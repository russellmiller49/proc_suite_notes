import json
import re
import datetime
from openpyxl import load_workbook, Workbook

# =============================================================================
# INPUTS
# =============================================================================
NOTE_ID = "note_001"
SOURCE_FILE = "note_001.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_001 SOURCE_FILE: note_001.txt 
INDICATION FOR OPERATION:  [REDACTED] is a 71 year old-year-old female who presents with multiple lung nodule, borderline lymphadenopathy, and left adrenal mass.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient or surrogate in detail.
Patient or surrogate indicated a wish to proceed with surgery and informed consent was signed.
 
CONSENT: Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient or surrogate.
The patient or surrogate read and signed the provided consent form / provided consent over the phone.
If consent was done via televideo, then the consent was witnessed by an assisting medical professional.
PREOPERATIVE DIAGNOSIS: R91.8 Other nonspecific abnormal finding of lung field.; lymphadenopathy;
left adrenal mass
POSTOPERATIVE DIAGNOSIS:  R91.8 Other nonspecific abnormal finding of lung field.; lymphadenopathy;
left adrenal mass
 
PROCEDURE:  
31899 Unlisted Procedure (Trach Change with Mature Tract or Procedure NOS)
31645 Therapeutic aspiration initial episode
31622 Dx bronchoscope/cell washing          
31623 Dx bronchoscope/brushing    
31624 Dx bronchoscope/lavage (BAL)    
31628 TBBX single lobe     
31629 TBNA single lobe   
31626 Fiducial marker placements, single or multiple     
31627 Navigational Bronchoscopy (computer assisted)
77012 Radiology / radiologic guidance for CT guided needle placement (CIOS)
76377 3D rendering with interpretation and reporting of CT, US, Tomo modality (ION Planning Station)
31654 Radial 
EBUS for peripheral lesion
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
31653 EBUS sampling 3 or more nodes
76982 Ultrasound Elastography, First Target Lesion
76983 Ultrasound Elastography, Additional Targets 
76983 Ultrasound Elastography, Additional Target 2
43238 EGD and EUS guided biopsies
 
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)
 
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure (22 MODIFIER):
This patient required the following at 3 different/discrete sites, although all 3 sites were within the same lobe they were in 
distinct/different locations within the lobe:  robotic navigational bronchoscopy to the different sites, brushings, radial EBUS, CT-guided needle placement, transbronchial needle aspiration, transbronchial cryo biopsies.
The patient required the following at 2 different/discrete sites:  bronchoalveolar lavage.
The patient required ultrasound elastrography evaluation of 7 different/discrete sites (6 lymph node sites and 1 adrenal mass site), with the images/interpretation used for intraprocedural decision-making;
this also extended the time necessary for the EBUS procedure and extended the time necessary for the EUS-B procedure.
This resulted in >150% increased work due to Increased intensity, Time, Technical difficulty of procedure, and Physical and mental effort required.
Apply to:      
31623 Dx bronchoscope/brushing    
31624 Dx bronchoscope/lavage (BAL)    
31628 TBBX single lobe     
31629 TBNA single lobe   
31627 Navigational Bronchoscopy (computer assisted)
77012 Radiology / radiologic guidance for CT guided needle placement (CIOS)
31654 Radial EBUS for peripheral lesion
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
31653 EBUS sampling 3 or more nodes
76982 Ultrasound Elastography, First Target Lesion
76983 Ultrasound Elastography, Additional Targets 
76983 Ultrasound Elastography, Additional Target 2
43238 EGD and EUS guided biopsies

ANESTHESIA: General Anesthesia
 
MONITORING: 
Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT(S): 
Flexible Therapeutic Bronchoscope
Linear EBUS 
Radial EBUS
Ion Robotic Bronchoscope
 
PROCEDURE IN DETAIL:
A timeout was performed (confirming the patient's name, procedure type, and procedure location).
After the successful induction of anesthesia, anesthesia placed an ETT.
Ventilation Parameters:
Mode	RR	TV	PEEP	FiO2	Flow Rate	Pmean
VCV	12	400	12	60	10	15
 
The Flexible Therapeutic Bronchoscope was advanced for airway examination.
Endobronchial topical lidocaine applied to the main carina, right carina 1, and left carina 2.
 
Initial Airway Examination Findings:
Trachea: Distal 1/3 normal.
Main Carina: Sharp
Right Lung Proximal Airways: Normal anatomic branching to the first subsegmental level with normal variant with 4-airway take-off at RUL.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Left Lung Proximal Airways:  Mildly tortuous LMSB with rotation of LC2.
Evidence of LUL lobectomy - stump in excellent condition without lesions.
LLL and LB7/8-10 segmental airways otherwise normal to the first subsegmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Mucosa: Normal.
Secretions: Moderate, thin, and clear  All secretions were suctioned to clear (therapeutic aspiration).
Successful therapeutic aspiration was performed to clean out the distal trachea, right mainstem bronchus, right upper lobe, bronchus intermedius, right middle lobe, right lower lobe, left mainstem bronchus, left upper lobe, left lower lobe from mucus.
The Flexible Therapeutic Bronchoscope was removed and the robotic-assisted navigational platform was set-up and the Ion robotic bronchoscope catheter advanced.
Right upper lobe apical segment nodule #1 - peripheral subsolid nodule with increasing density on recent imaging (labeled as RUL #1):
 
CT Chest scan was placed on separate planning station to generate 3D rendering of the pathway to target.
The navigational plan was reviewed and verified.  This was then loaded into robotic bronchoscopy platform.
Robotic navigation bronchoscopy was performed with Intuitive Ion platform.  Partial registration was used.
The Ion robotic catheter was used to engage the Apical Segment of RUL (RB1).
Target lesion is about 0.8 cm in diameter.   Under navigational guidance the Ion robotic catheter was advanced to 1.0 cm away from the planned target.
Cone Beam CT was performed: 3-D reconstructions were performed on an independent workstation.
Cios Spin system was used for evaluation of nodule location.  Standard dose spin was performed to acquire 3D volume.
This was passed on to Ion platform system for 3D reconstruction and nodule location.
Using the newly acquired nodule location, the Ion robotic system was adjusted to the new targeted location.
I personally interpreted the cone beam CT and 3-D reconstruction.
 
Robotic external working catheter was adjusted as noted above.
Transbronchial needle was advanced.  Cone Beam CT was repeated: 3-D reconstructions were performed on an independent workstation.
Cios Spin system was used for evaluation of nodule location.  Standard dose spin was performed to acquire 3D volume.
This was passed on to Ion platform system for 3D reconstruction and nodule location.
The newly acquired nodule location demonstrated the robotic catheter was in appropriate position. Cone beam CT confirmed tool-in-lesion.
I personally interpreted the cone beam CT and 3-D reconstruction.
Transbronchial needle aspiration was performed with 21G Needle and 23G Needle through the Ion extended working channel catheter.
Total 7 sample(s) were collected.  Samples sent for Cytology and Cultures.
Between needle aspiration biopsies, radial EBUS was performed in attempt to identify location of lesion, which showed Concentric pattern with the following features noted: Heterogeneous echogenicity.
Radial EBUS was performed in attempt to identify location of lesion, which showed Concentric pattern with the following features noted: Heterogeneous echogenicity.
Transbronchial cryobiopsy was performed with the 1.1 mm cryoprobe via the extended working channel catheter.
Freeze time of 4-6 seconds were used.  Total 8 sample(s) were collected.  Samples sent for Pathology and Tissue Cultures.
ROSE from ION procedure was noted to be:  Atypical cells, few inflammatory cells
 
Fiducial marker (0.8mm x 3mm soft tissue gold CIVCO) was loaded with bone wax and placed under fluoroscopy guidance.
Transbronchial brushing was performed with Protected cytology brush via the extended working channel catheter.  Total 1 sample(s) were collected.
Samples sent for Cytology.
 
Mini Bronchial alveolar lavage was performed via the extended working channel catheter.
Instilled 10 cc of NS, suction returned with 5 cc of BAL fluid.  Samples sent for Micro/Cultures.
Right upper lobe apical segment nodule #2 - more central subsolid nodule (labeled as RUL #2):
 
CT Chest scan was placed on separate planning station to generate 3D rendering of the pathway to target.
The navigational plan was reviewed and verified.  This was then loaded into robotic bronchoscopy platform.
Robotic navigation bronchoscopy was performed with Intuitive Ion platform.  Partial registration was repeated and used.
The Ion robotic catheter was used to engage the Apical Segment of RUL (RB1).
Target lesion is about 1.1 cm in diameter.   Under navigational guidance the Ion robotic catheter was advanced to 1.0 cm away from the planned target.
Cone Beam CT was performed: 3-D reconstructions were performed on an independent workstation.
Cios Spin system was used for evaluation of nodule location.  Standard dose spin was performed to acquire 3D volume.
This was passed on to Ion platform system for 3D reconstruction and nodule location.
Using the newly acquired nodule location, the Ion robotic system was adjusted to the new targeted location.
I personally interpreted the cone beam CT and 3-D reconstruction.
 
Robotic external working catheter was adjusted as noted above.
Transbronchial needle was advanced.  Cone Beam CT was repeated: 3-D reconstructions were performed on an independent workstation.
Cios Spin system was used for evaluation of nodule location.  Standard dose spin was performed to acquire 3D volume.
This was passed on to Ion platform system for 3D reconstruction and nodule location.
The newly acquired nodule location demonstrated the robotic catheter was in appropriate position. Cone beam CT confirmed tool-in-lesion.
I personally interpreted the cone beam CT and 3-D reconstruction.
Transbronchial needle aspiration was performed with 21G Needle and 23G Needle through the Ion extended working channel catheter.
Total 7 sample(s) were collected.  Samples sent for Cytology and Micro/Cultures.
Between needle aspiration biopsies, radial EBUS was performed in attempt to identify location of lesion, which showed Concentric pattern with the following features noted: Heterogeneous echogenicity.
Radial EBUS was performed in attempt to identify location of lesion, which showed Eccentric pattern with the following features noted: Heterogeneous echogenicity.
Transbronchial cryobiopsy was performed with the 1.1 mm cryoprobe via the extended working channel catheter.
Freeze time of 4-6 seconds were used.  Total 7 sample(s) were collected.  Samples sent for Pathology and Tissue Culture.
ROSE from ION procedure was noted to be:  Blood
 
Fiducial marker (0.8mm x 3mm soft tissue gold CIVCO) was loaded with bone wax and placed under fluoroscopy guidance.
Transbronchial brushing was performed with Protected cytology brush via the extended working channel catheter.  Total 1 sample(s) were collected.
Samples sent for Cytology.
 
Mini Bronchial alveolar lavage was performed via the extended working channel catheter.
Instilled 10 cc of NS, suction returned with 5 cc of NS.  Samples sent for Micro/Cultures.
Right upper lobe posterior segment (RB2) subpleural subsolid nodule (labeled as RUL #3):
 
CT Chest scan was placed on separate planning station to generate 3D rendering of the pathway to target.
The navigational plan was reviewed and verified.  This was then loaded into robotic bronchoscopy platform.
Robotic navigation bronchoscopy was performed with Intuitive Ion platform.  Partial registration was repeated and used.
The Ion robotic catheter was used to engage the Posterior Segment of RUL (RB2).
Target lesion is about 1.1 cm in diameter.   Under navigational guidance the Ion robotic catheter was advanced to 1.0 cm away from the planned target.
Cone Beam CT was performed: 3-D reconstructions were performed on an independent workstation.
Cios Spin system was used for evaluation of nodule location.  Standard dose spin was performed to acquire 3D volume.
This was passed on to Ion platform system for 3D reconstruction and nodule location.
Using the newly acquired nodule location, the Ion robotic system was adjusted to the new targeted location.
I personally interpreted the cone beam CT and 3-D reconstruction.
 
Robotic external working catheter was adjusted as noted above.
Transbronchial needle was advanced.  Cone Beam CT was repeated: 3-D reconstructions were performed on an independent workstation.
Cios Spin system was used for evaluation of nodule location.  High dose spin was performed to acquire 3D volume.
This was passed on to Ion platform system for 3D reconstruction and nodule location.
The newly acquired nodule location demonstrated the robotic catheter was in appropriate position. Cone beam CT confirmed tool-in-lesion.
I personally interpreted the cone beam CT and 3-D reconstruction.
Transbronchial needle aspiration was performed with 21G Needle and 23G Needle through the Ion extended working channel catheter.
Total 7 sample(s) were collected.  Samples sent for Cytology and Cultures.
Between needle aspiration biopsies, radial EBUS was performed in attempt to identify location of lesion, which showed Concentric pattern with the following features noted: Heterogeneous echogenicity.
Transbronchial cryobiopsy was performed with the 1.1 mm cryoprobe via the extended working channel catheter.
Freeze time of 4-5 seconds were used.  Total 8 sample(s) were collected.  Samples sent for Pathology and Tissue Cultures.
ROSE from ION procedure was noted to be:  Atypical cells - similar in morphology to site RUL #1.
Fiducial marker (0.8mm x 3mm soft tissue gold CIVCO) was loaded with bone wax and placed under fluoroscopy guidance.
Transbronchial brushing was performed with Protected cytology brush via the extended working channel catheter.  Total 1 sample(s) were collected.
Samples sent for Cytology.
 
Mini Bronchial alveolar lavage was performed via the extended working channel catheter.
Instilled 10 cc of NS, suction returned with 5 cc of BAL fluid.  Samples sent for Micro/Cultures.
The robotic extended working channel catheter was withdrawn.  Post-biopsy fluoroscopy images negative for pneumothorax.
 
The Flexible Therapeutic Bronchoscope was readvanced.
Bronchial alveolar lavage was performed at Apical Segment of RUL (RB1).
Instilled 60 cc of NS, suction returned with 15 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology. 
 
Bronchial alveolar lavage was performed at Posterior Segment of RUL (RB2).
Instilled 60 cc of NS, suction returned with 15 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology. 
 
Secretions, residual saline, and minimal blood suctioned to clear.
The Flexible Therapeutic Bronchoscope was withdrawn and the endobronchial ultrasound-capable (EBUS) bronchoscope was introduced with the following findings:
 
EBUS-Findings
Indications: Diagnostic and Staging
Technique:
All lymph node stations were assessed.
Only those 5 mm or greater in short axis were sampled, or if otherwise concerning appearing.
Lymph node sizing was performed by EBUS and sampling by transbronchial needle aspiration was performed using 25-gauge Needle and 21-gauge Needle.
Lymph Nodes/Sites Inspected: 
4R (lower paratracheal) node - Endobronchial ultrasound (EBUS) elastography mode was performed on this lymph node to assess its stiffness and tissue characteristics, and to identify different densities in this target lymph node.
Semi-qualitivative analysis of this lymph node demonstrated a Type 1 elastograpic pattern, predominantly soft (green/yellow), suggesting a reactive or benign process.
The elastography interpretation was separate from the typical interpretation of linear EBUS ultrasound images.
Along with the lymph node size, the elastographic pattern information/interpretation was used to make the decision to not perform transbronchial sampling of this lymph node.
This image was store and archived. 
4L (lower paratracheal) node - Endobronchial ultrasound (EBUS) elastography mode was performed on this lymph node to assess its stiffness and tissue characteristics, and to identify different densities in this target lymph node.
Semi-qualitivative analysis of this lymph node demonstrated a Type 1 elastograpic pattern, predominantly soft (green/yellow), suggesting a reactive or benign process.
The elastography interpretation was separate from the typical interpretation of linear EBUS ultrasound images.
Along with the lymph node size, the elastographic pattern information/interpretation was used to make the decision to not perform transbronchial sampling of this lymph node.
This image was store and archived. 
7 (subcarinal) node
10R lymph node
10L lymph node
11Rs lymph node
11Ri lymph node
11L lymph node
 
Overall EBUS TBNA ROSE Diagnosis: Not performed
 
EBUS Lymph Nodes Sampled:
Site 1: The 11L lymph node was < 10 mm on CT  and Hypermetabolic via PET-CT scan in 7/2025, but s/p surgical lymph node resection since that time.
The lymph node was photographed.
Elastography:  Endobronchial ultrasound (EBUS) elastography mode was performed to assess lymph node stiffness and tissue characteristics, and to identify different densities in this target lymph node.
Semi-qualitivative analysis of this lymph node demonstrated a Type 2 elastographic pattern with mixed soft (green/yellow) and stiff (blue) regions.
Given this heterogeneous and indeterminate appearance, TBNA was directed at representative areas to ensure comprehensive sampling and to minimize the risk of underdiagnosis.
The elastography interpretation was separate from the typical interpretation of linear EBUS ultrasound images.
Along with the lymph node size, the elastographic pattern interpretation was used to make the decision to perform transbronchial sampling of this lymph node and where to guide/location sampling within the lymph node.
This image was store and archived.  
Sampling:  The site was sampled.. 4 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Sent for cytology.
Preliminary ROSE Cytology was reported as not performed and without preliminary results. Final results are pending.
Site 2: The 7 (subcarinal) node was < 10 mm on CT  and Non-Hypermetabolic via PET-CT scan in 7/2025.
The lymph node was photographed.
Elastography:  Endobronchial ultrasound (EBUS) elastography mode was performed to assess lymph node stiffness and tissue characteristics, and to identify different densities in this target lymph node.
Semi-qualitivative analysis of this lymph node demonstrated a Type 1 elastograpic pattern, predominantly soft (green/yellow), suggesting a reactive or benign process.
Despite the benign appearance, TBNA was performed to confirm the absence of malignancy and to obtain cytology for diagnostic completeness.
The elastography interpretation was separate from the typical interpretation of linear EBUS ultrasound images.
Along with the lymph node size, the elastographic pattern interpretation was used to make the decision to perform transbronchial sampling of this lymph node and where to guide/location sampling within the lymph node.
This image was store and archived.  
Sampling:  The site was sampled.. 4 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Sent for cytology.
Preliminary ROSE Cytology was reported as not performed and without preliminary results. Final results are pending.
Site 3: The 11Rs lymph node was was < 10 mm on CT  and Non-Hypermetabolic via PET-CT scan in 7/2025.
The lymph node was photographed.  The photographed lymph node was <5mm in size;
however, during repeat assessment there was a larger lymph node >5mm in size that was identified, although a picture was not take.
Sampling was taken from both of these 11Rs lymph nodes.
Elastography:  Endobronchial ultrasound (EBUS) elastography mode was performed to assess lymph node stiffness and tissue characteristics, and to identify different densities in this target lymph node.
Semi-qualitivative analysis of this lymph node demonstrated a Type 1 elastograpic pattern, predominantly soft (green/yellow), suggesting a reactive or benign process.
Despite the benign appearance, TBNA was performed to confirm the absence of malignancy and to obtain cytology for diagnostic completeness.
The elastography interpretation was separate from the typical interpretation of linear EBUS ultrasound images.
Along with the lymph node size, the elastographic pattern interpretation was used to make the decision to perform transbronchial sampling of this lymph node and where to guide/location sampling within the lymph node.
This image was store and archived.  
Sampling:  The site was sampled.. 4 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Sent for cytology.
Preliminary ROSE Cytology was reported as not performed and without preliminary results. Final results are pending.
Site 4: The 11Ri lymph node was < 10 mm on CT  and Non-Hypermetabolic via PET-CT scan in 7/2025.
The lymph node was photographed.
Elastography:  Endobronchial ultrasound (EBUS) elastography mode was performed to assess lymph node stiffness and tissue characteristics, and to identify different densities in this target lymph node.
Semi-qualitivative analysis of this lymph node demonstrated a Type 1 elastograpic pattern, predominantly soft (green/yellow), suggesting a reactive or benign process.
Despite the benign appearance, TBNA was performed to confirm the absence of malignancy and to obtain cytology for diagnostic completeness.
The elastography interpretation was separate from the typical interpretation of linear EBUS ultrasound images.
Along with the lymph node size, the elastographic pattern interpretation was used to make the decision to perform transbronchial sampling of this lymph node and where to guide/location sampling within the lymph node.
This image was store and archived.  
Sampling:  The site was sampled.. 4 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Sent for cytology.
Preliminary ROSE Cytology was reported as not performed and without preliminary results. Final results are pending.
The EBUS bronchoscope was withdrawn and the Flexible Therapeutic Bronchoscope readvanced.  Minimal amount of blood was suctioned (therapeutic aspiration).
Inspection demonstrated no evidence of active bleeding.  Residual secretions were suctioned to cleared (therapeutic aspiration).
EGD findings:
EGD was performed to assess for esophageal or gastric in context of suspected new or recurrent cancer and to assess safety of advancing EBUS scope for EUS-B procedure.
Flexible therapeutic bronchoscope was advanced with the assistance of air insufflation, and allowing for dilation to the lumen for better intraluminal evaluation.
The esophagus was normal, GEJ normal, stomach was normal with mild amount of green bilious-appearing secretions, which were suctioned to clear.
The pyloric sphincter was assessed and normal.  The duodenum was not assessed.
Based on findings, no concerning lesions suggestion of esophageal or gastric malignancy and safe to proceed with EUS-B procedure.
Air was exsufflated while removing the scope.
 
EUS-B Findings
EUS-B was performed to assess concerning left adrenal mass in context of suspected new or recurrent cancer.
Indications: Diagnostic and Staging
Technique:
The left adrenal mass was assessed. It was >10mm in size, and was sampled.
The EUS-B scope was advanced with the assistance of air insufflation.
The left adrenal mass evaluation was performed via endoscopic-ultrasound with EBUS bronchoscope (EUS-B) and sampling by transesophageal/transgastric needle aspiration was performed using 25-gauge Needle.
Sites Inspected: 
Left adrenal mass
 
Overall EUS-B transesophageal/transgastric ROSE Diagnosis: "Lesional"
 
EUS-B Sites Sampled:
Site 1: The left adrenal mass was > 10 mm on CT  and Hypermetabolic via PET-CT scan in 7/2025.
The left adrenal mass was photographed.
Elastography:  Endobronchial ultrasound (EBUS) elastography mode was performed to assess lymph node stiffness and tissue characteristics, and to identify different densities in this target lymph node.
Semi-qualitivative analysis of this lymph node demonstrated a Type 2 elastographic pattern with mixed soft (green/yellow) and stiff (blue) regions.
Given this heterogeneous and indeterminate appearance, TBNA was directed at representative areas to ensure comprehensive sampling and to minimize the risk of underdiagnosis.
The elastography interpretation was separate from the typical interpretation of linear EBUS ultrasound images.
Along with the lymph node size, the elastographic pattern interpretation was used to make the decision to perform transbronchial sampling of this lymph node and where to guide/location sampling within the lymph node.
This image was store and archived.  
Sampling:  The site was sampled.. 4 endoscopic ultrasound guided transbronchial biopsies were performed with samples obtained.
Sent for cytology.  
Evidence of small hematoma forming around the adrenal mass on EUS-B evaluation.  Therefore, no further tissue obtained.
Preliminary ROSE Cytology was reported as adequate and "Lesional". Final results are pending.
The EUS-B scope was withdrawn and the Flexible Therapeutic Bronchoscope readvanced into esophagus and stomach with the assistance of air insufflation for repeat assessment.
No evidence of mucosal or intraluminal bleeding.  Air was exsufflated while removing the scope.
 
Bronchoscope was removed.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
ESTIMATED BLOOD LOSS:   Minimal
COMPLICATIONS:     None
 
SPECIMEN(S): 
-Right upper lobe apical segment nodule #1 - peripheral subsolid nodule with increasing density (labeled as RUL #1): TBNA (cytology and cultures), TBCBx (pathology and tissue culture), Brushing (cytology), Mini-BAL (cultures)
-Right upper lobe apical segment nodule #2 - more central subsolid nodule (labeled as RUL #2): TBNA (cytology and cultures), TBCBx (pathology and tissue culture), Brushing (cytology), Mini-BAL (cultures)
-Right upper lobe posterior segment (RB2) subpleural subsolid nodule (labeled as RUL #3): TBNA (cytology and cultures), TBCBx (pathology and tissue culture), Brushing (cytology), Mini-BAL (cultures)
-RUL RB1 (apical segment) BAL: 
 cell count, micro, and cytology
-RUL RB2 (posterior segment) BAL:  cell count, micro, and cytology
-EBUS TBNA of 11L lymph node, 7 (subcarinal) node, 11Rs lymph node, 11Ri lymph node:  cytology for each
-EUS-B transesopagheal/transgastric needle aspiration of left adrenal mass:  cytology
 
IMPRESSION/PLAN: [REDACTED] is a 71 year old-year-old female who presents for bronchoscopy for biopsy of multiple lung nodules, lymph node assessment, and left adrenal mass assessment.
-Obtain post-procedure CXR (ordered)
-Follow-up in IP clinic to review results in 1-2 weeks, as already scheduled (1/21/2026)."""

# =============================================================================
# DATA STRUCTURES
# =============================================================================

# Flags
FLAGS = {
    # Bronchoscopy (23)
    "diagnostic_bronchoscopy": 1,
    "bal": 1,
    "bronchial_wash": 0,
    "brushings": 1,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 1,
    "linear_ebus": 1,
    "radial_ebus": 1,
    "navigational_bronchoscopy": 1,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 1,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural (7)
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# Spans
spans_data = []

def create_span(span_text, label, normalized_value, event_id, section_type="PROCEDURE", context_prefix=None, match_index=None, schema_field=None):
    spans_data.append({
        "source_file": SOURCE_FILE,
        "note_id": NOTE_ID,
        "span_id": f"span_{len(spans_data) + 1}",
        "section_type": section_type,
        "context_prefix": context_prefix,
        "span_text": span_text,
        "match_index": match_index,
        "label": label,
        "normalized_value": normalized_value,
        "schema_field": schema_field,
        "event_id": event_id,
        "is_negated": "FALSE",
        "is_historical": "FALSE",
        "time_anchor": "",
        "reviewer": "Auto",
        "comments": "",
        "hydration_status": "needs_hydration"
    })

# Events List for V3
events_list = []

# =============================================================================
# EVENT LOGIC
# =============================================================================

# --- EVENT 1: Therapeutic Aspiration (General Airway) ---
e1 = "evt_01_asp"
create_span("Secretions: Moderate, thin, and clear  All secretions were suctioned to clear", "PROC_METHOD", "Therapeutic Aspiration", e1)
create_span("therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", e1, context_prefix="Successful")
create_span("clean out the distal trachea", "PROC_ACTION", "Airway Clearance", e1)

events_list.append({
    "id": e1, "type": "therapeutic_aspiration", "target": {"structure": "Airway"}, "method": "Suction",
    "outcomes": {}
})

# --- EVENT 2: RUL Apical Nodule 1 (Robotic) ---
e2 = "evt_02_rul1"
create_span("Right upper lobe apical segment nodule #1", "ANAT_LUNG_LOC", "RUL Apical Segment", e2)
create_span("Ion robotic bronchoscope", "DEV_INSTRUMENT", "Ion", e2)
create_span("Cone Beam CT", "PROC_METHOD", "CBCT", e2)
create_span("Cios Spin system", "DEV_INSTRUMENT", "Cios Spin", e2)
create_span("21G Needle", "DEV_NEEDLE", "21G", e2, context_prefix="RUL #1")
create_span("23G Needle", "DEV_NEEDLE", "23G", e2, context_prefix="RUL #1")
create_span("radial EBUS", "PROC_METHOD", "Radial EBUS", e2, context_prefix="RUL #1")
create_span("1.1 mm cryoprobe", "DEV_INSTRUMENT", "1.1mm Cryoprobe", e2, context_prefix="RUL #1")
create_span("Fiducial marker", "DEV_IMPLANT", "Fiducial", e2, context_prefix="RUL #1")
create_span("Protected cytology brush", "DEV_INSTRUMENT", "Cytology Brush", e2, context_prefix="RUL #1")
create_span("Mini Bronchial alveolar lavage", "PROC_METHOD", "Mini-BAL", e2, context_prefix="RUL #1")
create_span("Atypical cells", "OBS_ROSE", "Atypical cells", e2, context_prefix="RUL #1")

events_list.append({
    "id": e2, "type": "navigational_bronchoscopy",
    "target": {"location": {"lobe": "RUL", "segment": "Apical"}, "structure": "Lesion"},
    "lesion": {"type": "Nodule", "size_mm": 8},
    "method": "Robotic Navigational Bronchoscopy",
    "devices": ["Ion", "21G Needle", "23G Needle", "1.1mm Cryoprobe", "Fiducial", "Brush"],
    "specimens": ["Cytology", "Biopsy", "Microbiology"],
    "findings": ["Atypical cells"]
})

# --- EVENT 3: RUL Apical Nodule 2 (Robotic) ---
e3 = "evt_03_rul2"
create_span("Right upper lobe apical segment nodule #2", "ANAT_LUNG_LOC", "RUL Apical Segment", e3)
create_span("more central subsolid nodule", "OBS_LESION", "Subsolid Nodule", e3)
create_span("21G Needle", "DEV_NEEDLE", "21G", e3, context_prefix="RUL #2")
create_span("23G Needle", "DEV_NEEDLE", "23G", e3, context_prefix="RUL #2")
create_span("radial EBUS", "PROC_METHOD", "Radial EBUS", e3, context_prefix="RUL #2")
create_span("1.1 mm cryoprobe", "DEV_INSTRUMENT", "1.1mm Cryoprobe", e3, context_prefix="RUL #2")
create_span("Fiducial marker", "DEV_IMPLANT", "Fiducial", e3, context_prefix="RUL #2")
create_span("Protected cytology brush", "DEV_INSTRUMENT", "Cytology Brush", e3, context_prefix="RUL #2")
create_span("Mini Bronchial alveolar lavage", "PROC_METHOD", "Mini-BAL", e3, context_prefix="RUL #2")
create_span("Blood", "OBS_ROSE", "Blood", e3, context_prefix="ROSE from ION")

events_list.append({
    "id": e3, "type": "navigational_bronchoscopy",
    "target": {"location": {"lobe": "RUL", "segment": "Apical"}, "structure": "Lesion"},
    "lesion": {"type": "Nodule", "size_mm": 11},
    "method": "Robotic Navigational Bronchoscopy",
    "devices": ["Ion", "21G Needle", "23G Needle", "1.1mm Cryoprobe", "Fiducial", "Brush"],
    "specimens": ["Cytology", "Biopsy", "Microbiology"],
    "findings": ["Blood (ROSE)"]
})

# --- EVENT 4: RUL Posterior Nodule 3 (Robotic) ---
e4 = "evt_04_rul3"
create_span("Right upper lobe posterior segment (RB2) subpleural subsolid nodule", "ANAT_LUNG_LOC", "RUL Posterior Segment", e4)
create_span("21G Needle", "DEV_NEEDLE", "21G", e4, context_prefix="RUL #3")
create_span("23G Needle", "DEV_NEEDLE", "23G", e4, context_prefix="RUL #3")
create_span("radial EBUS", "PROC_METHOD", "Radial EBUS", e4, context_prefix="RUL #3")
create_span("1.1 mm cryoprobe", "DEV_INSTRUMENT", "1.1mm Cryoprobe", e4, context_prefix="RUL #3")
create_span("Fiducial marker", "DEV_IMPLANT", "Fiducial", e4, context_prefix="RUL #3")
create_span("Protected cytology brush", "DEV_INSTRUMENT", "Cytology Brush", e4, context_prefix="RUL #3")
create_span("Mini Bronchial alveolar lavage", "PROC_METHOD", "Mini-BAL", e4, context_prefix="RUL #3")
create_span("Atypical cells - similar in morphology to site RUL #1", "OBS_ROSE", "Atypical cells", e4)

events_list.append({
    "id": e4, "type": "navigational_bronchoscopy",
    "target": {"location": {"lobe": "RUL", "segment": "Posterior"}, "structure": "Lesion"},
    "lesion": {"type": "Nodule", "size_mm": 11},
    "method": "Robotic Navigational Bronchoscopy",
    "devices": ["Ion", "21G Needle", "23G Needle", "1.1mm Cryoprobe", "Fiducial", "Brush"],
    "specimens": ["Cytology", "Biopsy", "Microbiology"],
    "findings": ["Atypical cells"]
})

# --- EVENT 5: BAL RUL Apical ---
e5 = "evt_05_bal"
create_span("Bronchial alveolar lavage", "PROC_METHOD", "BAL", e5, context_prefix="Apical Segment of RUL")
create_span("Apical Segment of RUL", "ANAT_LUNG_LOC", "RUL Apical Segment", e5)
create_span("Instilled 60 cc", "MEAS_VOL", "60cc", e5, context_prefix="Apical Segment")

events_list.append({
    "id": e5, "type": "bal",
    "target": {"location": {"lobe": "RUL", "segment": "Apical"}},
    "method": "BAL",
    "measurements": {"instilled": "60 cc", "return": "15 cc"}
})

# --- EVENT 6: BAL RUL Posterior ---
e6 = "evt_06_bal"
create_span("Bronchial alveolar lavage", "PROC_METHOD", "BAL", e6, context_prefix="Posterior Segment of RUL")
create_span("Posterior Segment of RUL", "ANAT_LUNG_LOC", "RUL Posterior Segment", e6)

events_list.append({
    "id": e6, "type": "bal",
    "target": {"location": {"lobe": "RUL", "segment": "Posterior"}},
    "method": "BAL",
    "measurements": {"instilled": "60 cc", "return": "15 cc"}
})

# --- EVENT 7: EBUS 11L ---
e7 = "evt_07_ebus"
create_span("11L lymph node", "ANAT_LN_STATION", "11L", e7)
create_span("Endobronchial ultrasound (EBUS) elastography mode", "PROC_METHOD", "Elastography", e7, context_prefix="11L")
create_span("Type 2 elastographic pattern", "OBS_LESION", "Type 2 Elastography", e7, context_prefix="11L")
create_span("4 endobronchial ultrasound guided transbronchial biopsies", "PROC_ACTION", "TBNA", e7, context_prefix="11L")

events_list.append({
    "id": e7, "type": "linear_ebus",
    "target": {"station": "11L"},
    "method": "EBUS-TBNA",
    "specimens": ["Cytology"],
    "findings": ["Type 2 Elastography"]
})

# --- EVENT 8: EBUS 7 ---
e8 = "evt_08_ebus"
create_span("7 (subcarinal) node", "ANAT_LN_STATION", "7", e8)
create_span("Type 1 elastograpic pattern", "OBS_LESION", "Type 1 Elastography", e8, context_prefix="Site 2")
create_span("4 endobronchial ultrasound guided transbronchial biopsies", "PROC_ACTION", "TBNA", e8, context_prefix="Site 2")

events_list.append({
    "id": e8, "type": "linear_ebus",
    "target": {"station": "7"},
    "method": "EBUS-TBNA",
    "specimens": ["Cytology"],
    "findings": ["Type 1 Elastography"]
})

# --- EVENT 9: EBUS 11Rs ---
e9 = "evt_09_ebus"
create_span("11Rs lymph node", "ANAT_LN_STATION", "11Rs", e9)
create_span("Type 1 elastograpic pattern", "OBS_LESION", "Type 1 Elastography", e9, context_prefix="Site 3")
create_span("4 endobronchial ultrasound guided transbronchial biopsies", "PROC_ACTION", "TBNA", e9, context_prefix="Site 3")

events_list.append({
    "id": e9, "type": "linear_ebus",
    "target": {"station": "11Rs"},
    "method": "EBUS-TBNA",
    "specimens": ["Cytology"],
    "findings": ["Type 1 Elastography"]
})

# --- EVENT 10: EBUS 11Ri ---
e10 = "evt_10_ebus"
create_span("11Ri lymph node", "ANAT_LN_STATION", "11Ri", e10)
create_span("Type 1 elastograpic pattern", "OBS_LESION", "Type 1 Elastography", e10, context_prefix="Site 4")
create_span("4 endobronchial ultrasound guided transbronchial biopsies", "PROC_ACTION", "TBNA", e10, context_prefix="Site 4")

events_list.append({
    "id": e10, "type": "linear_ebus",
    "target": {"station": "11Ri"},
    "method": "EBUS-TBNA",
    "specimens": ["Cytology"],
    "findings": ["Type 1 Elastography"]
})

# --- EVENT 11: EUS-B Adrenal ---
e11 = "evt_11_eus"
create_span("left adrenal mass", "OBS_LESION", "Adrenal Mass", e11, section_type="EUS-B Findings")
create_span("EUS-B", "PROC_METHOD", "EUS-B", e11, section_type="EUS-B Findings")
create_span("25-gauge Needle", "DEV_NEEDLE", "25G", e11, section_type="EUS-B Findings")
create_span("endoscopic ultrasound guided transbronchial biopsies", "PROC_ACTION", "EUS-FNA", e11, section_type="EUS-B Findings")
create_span("small hematoma", "OUTCOME_COMPLICATION", "Hematoma", e11)
create_span("Lesional", "OBS_ROSE", "Lesional", e11)

events_list.append({
    "id": e11, "type": "eus",
    "target": {"structure": "Left Adrenal"},
    "method": "EUS-B FNA",
    "devices": ["25G Needle"],
    "specimens": ["Cytology"],
    "findings": ["Lesional", "Hematoma"]
})

# --- OUTCOMES ---
e_global = "evt_global"
create_span("No immediate complications", "OUTCOME_COMPLICATION", "None", e_global)

# =============================================================================
# HYDRATION LOGIC
# =============================================================================

def hydrate_spans(spans, note_text):
    hydrated = []
    for s in spans:
        entry = s.copy()
        text = s["span_text"]
        
        # Exact match count
        matches = [m.start() for m in re.finditer(re.escape(text), note_text)]
        count = len(matches)
        
        start_char = ""
        end_char = ""
        status = "ambiguous"
        
        if count == 1:
            start_char = matches[0]
            end_char = start_char + len(text)
            status = "hydrated_unique"
        elif count > 1:
            if s["context_prefix"]:
                # Window search (120 chars before)
                found = False
                for m_start in matches:
                    window_start = max(0, m_start - 120)
                    window_text = note_text[window_start:m_start]
                    if s["context_prefix"] in window_text:
                        start_char = m_start
                        end_char = m_start + len(text)
                        status = "hydrated_prefix_window"
                        found = True
                        break
                if not found and s["match_index"] is not None and s["match_index"] < count:
                    start_char = matches[s["match_index"]]
                    end_char = start_char + len(text)
                    status = "hydrated_match_index"
            elif s["match_index"] is not None and s["match_index"] < count:
                start_char = matches[s["match_index"]]
                end_char = start_char + len(text)
                status = "hydrated_match_index"
            else:
                status = f"ambiguous_count={count}"
        else:
            status = "not_found"

        entry["start_char"] = start_char
        entry["end_char"] = end_char
        entry["hydration_status"] = status
        hydrated.append(entry)
    return hydrated

hydrated_data = hydrate_spans(spans_data, NOTE_TEXT)

# =============================================================================
# EXCEL GENERATION
# =============================================================================

def generate_excel(template_path: str = TEMPLATE_PATH, output_path: str = OUTPUT_PATH) -> None:
    try:
        wb = load_workbook(template_path)
    except FileNotFoundError:
        wb = Workbook()

    # 1. Note_Text
    ws_text = wb["Note_Text"] if "Note_Text" in wb.sheetnames else wb.create_sheet("Note_Text")
    if ws_text.max_row == 1:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws_index = wb["Note_Index"] if "Note_Index" in wb.sheetnames else wb.create_sheet("Note_Index")
    headers_index = [
        "source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes",
        "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", "tbna_conventional",
        "linear_ebus", "radial_ebus", "navigational_bronchoscopy", "transbronchial_biopsy", "transbronchial_cryobiopsy",
        "therapeutic_aspiration", "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
        "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", "bronchial_thermoplasty",
        "whole_lung_lavage", "rigid_bronchoscopy", "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy",
        "pleurodesis", "pleural_biopsy", "fibrinolytic_therapy"
    ]
    if ws_index.max_row == 1:
        ws_index.append(headers_index)

    row_index = [SOURCE_FILE, NOTE_ID, "", "", "", "", "Complete", ""]
    for h in headers_index[8:]:
        row_index.append(FLAGS.get(h, 0))
    ws_index.append(row_index)

    # 3. Span_Annotations
    ws_span = wb["Span_Annotations"] if "Span_Annotations" in wb.sheetnames else wb.create_sheet("Span_Annotations")
    headers_span = [
        "source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", "match_index",
        "start_char", "end_char", "span_len", "label", "normalized_value", "schema_field", "event_id",
        "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"
    ]
    if ws_span.max_row == 1:
        ws_span.append(headers_span)

    for s in spans_data:
        ws_span.append([
            s["source_file"], s["note_id"], s["span_id"], s["section_type"], s["context_prefix"], s["span_text"], s["match_index"],
            "", "", f'=LEN(F{ws_span.max_row + 1})', s["label"], s["normalized_value"], s["schema_field"], s["event_id"],
            s["is_negated"], s["is_historical"], s["time_anchor"], s["reviewer"], s["comments"], s["hydration_status"]
        ])

    # 4. Span_Hydrated
    ws_hyd = wb["Span_Hydrated"] if "Span_Hydrated" in wb.sheetnames else wb.create_sheet("Span_Hydrated")
    if ws_hyd.max_row == 1:
        ws_hyd.append(headers_span)

    for h in hydrated_data:
        ws_hyd.append([
            h["source_file"], h["note_id"], h["span_id"], h["section_type"], h["context_prefix"], h["span_text"], h["match_index"],
            h["start_char"], h["end_char"], f'=LEN(F{ws_hyd.max_row + 1})', h["label"], h["normalized_value"], h["schema_field"], h["event_id"],
            h["is_negated"], h["is_historical"], h["time_anchor"], h["reviewer"], h["comments"], h["hydration_status"]
        ])

    # 5. Event_Log
    ws_event = wb["Event_Log"] if "Event_Log" in wb.sheetnames else wb.create_sheet("Event_Log")
    headers_event = [
        "source_file", "note_id", "event_id", "event_type", "method", "anatomy_target", "device", "needle_gauge",
        "stations", "counts", "measurements", "specimens", "findings", "is_historical", "reviewer", "comments",
        "device_size", "device_material", "outcome_airway_lumen_pre", "outcome_airway_lumen_post",
        "outcome_symptoms", "outcome_pleural", "outcome_complication"
    ]
    if ws_event.max_row == 1:
        ws_event.append(headers_event)

    # Map events_list to CSV structure
    for evt in events_list:
        ws_event.append([
            SOURCE_FILE, NOTE_ID, evt["id"], evt["type"], evt.get("method", ""),
            str(evt.get("target", "")), str(evt.get("devices", "")), "", "", "",
            str(evt.get("measurements", "")), str(evt.get("specimens", "")), str(evt.get("findings", "")),
            "FALSE", "Auto", "", "", "", "", "", "", "", "Hematoma" if "Hematoma" in evt.get("findings", []) else ""
        ])

    # 6. V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"] if "V3_Procedure_Events" in wb.sheetnames else wb.create_sheet("V3_Procedure_Events")
    headers_v3 = [
        "note_id", "event_id", "type", "target.anatomy_type", "target.location.lobe", "target.location.segment",
        "target.station", "lesion.type", "lesion.size_mm", "method", "devices_json", "measurements_json",
        "specimens_json", "findings_json", "evidence_quote", "stent.size", "stent.material_or_brand",
        "catheter.size_fr", "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post", "outcomes.symptoms",
        "outcomes.pleural", "outcomes.complications"
    ]
    if ws_v3.max_row == 1:
        ws_v3.append(headers_v3)

    for evt in events_list:
        tgt = evt.get("target", {})
        loc = tgt.get("location", {})
        ws_v3.append([
            NOTE_ID, evt["id"], evt["type"],
            tgt.get("structure", ""), loc.get("lobe", ""), loc.get("segment", ""), tgt.get("station", ""),
            evt.get("lesion", {}).get("type", ""), evt.get("lesion", {}).get("size_mm", ""),
            evt.get("method", ""), json.dumps(evt.get("devices", [])), json.dumps(evt.get("measurements", {})),
            json.dumps(evt.get("specimens", [])), json.dumps(evt.get("findings", [])), "",
            "", "", "", "", "", "", "", "Hematoma" if "Hematoma" in evt.get("findings", []) else ""
        ])

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"] if "V3_Registry_JSON" in wb.sheetnames else wb.create_sheet("V3_Registry_JSON")
    if ws_json.max_row == 1:
        ws_json.append(["schema_version", "note_id", "json_output"])

    final_json = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "procedures": events_list,
        "no_immediate_complications": True
    }
    ws_json.append(["v3.0", NOTE_ID, json.dumps(final_json, indent=2)])

    wb.save(output_path)
    print(f"Generated {output_path}")


if __name__ == "__main__":
    generate_excel()
