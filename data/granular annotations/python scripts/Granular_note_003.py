import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# =============================================================================
# INPUTS
# =============================================================================
NOTE_ID = "note_003"
SOURCE_FILE = "note_003.txt"
PROCEDURE_DATE = "" 
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_003 SOURCE_FILE: note_003.txt 
INDICATION FOR OPERATION:  [REDACTED]is a 23 year old-year-old female who presents with synovial cell sarcoma.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
 
CONSENT : Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient or surrogate.
The patient or surrogate read and signed the provided consent form / provided consent over the phone.
The consent was witnessed by an assisting medical professional.
 
PREOPERATIVE DIAGNOSIS: R91.8 Other nonspecific abnormal finding of lung field.
POSTOPERATIVE DIAGNOSIS:  R91.8 Other nonspecific abnormal finding of lung field.
PROCEDURE:  
31622 Dx bronchoscope/cell washing          
 
ANESTHESIA: 
99152 Moderate sedation: initial 15 minutes
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Disposable Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
 
Procedure performed under moderate sedation.
The following medications were provided:
Versed             2 mg
Fentanyl          75 mcg
 
Physician/patient face-to-face anesthesia start time:   0959
 
Physician/patient face-to-face anesthesia stop time:   1008
 
Total moderate sedation time was 9 minutes.
Patient was monitored continuously one-to-one throughout the entire procedure by the attending physician while anesthesia was administered.
The airway was inspected and mucus was suctioned. Lidocaine was applied to the vocal cords and airway.
The mouth and oropharynx were normal appearing. 
 
The airway anatomy was normal appearing to the segmental level bilaterally.
The previous biopsy site was inspected appeared to have healed well.
 
The patient tolerated the procedure well.
There were no immediate complications.  At the conclusion of the operation, the patient was transported to the recovery room in stable condition.
SPECIMEN(S): 
None
 
IMPRESSION/PLAN: [REDACTED]is a 23 year old-year-old female who presents for bronchoscopy for airway inspection.
 
-Follow up as needed"""

# =============================================================================
# DATA DEFINITIONS
# =============================================================================

# Procedure Flags (0/1)
PROCEDURE_FLAGS = {
    # Bronchoscopy (23)
    "diagnostic_bronchoscopy": 1, # "Dx bronchoscope"
    "bal": 0,
    "bronchial_wash": 0, # Header says cell washing, but "Specimens: None" and text only mentions mucus suction.
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0, # "mucus was suctioned" is usually incidental to inspection unless stated as therapeutic intent.
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural (7)
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# Spans (Anchor-First)
# Structure: (span_text, context_prefix, match_index_fallback, label, normalized_value, schema_field, event_id, is_historical, is_negated)
RAW_SPANS = [
    ("Dx bronchoscope", "PROCEDURE: \n31622", 0, "PROC_METHOD", "Bronchoscopy", "method", "ev1", False, False),
    ("Disposable Bronchoscope", "INSTRUMENT : \n", 0, "DEV_INSTRUMENT", "Bronchoscope", "device", "ev1", False, False),
    ("airway inspection", "Initial", 0, "PROC_ACTION", "Inspection", "method", "ev1", False, False),
    ("mucus was suctioned", "inspected and", 0, "PROC_ACTION", "Suction", "method", "ev1", False, False),
    ("The airway", "", 0, "ANAT_AIRWAY", "Airway", "target.anatomy_type", "ev1", False, False),
    
    # Findings
    ("normal appearing", "oropharynx were", 0, "OBS_LESION", "Normal", "findings", "ev1", False, False),
    ("airway anatomy was normal", "The", 0, "OBS_LESION", "Normal", "findings", "ev1", False, False),
    
    # Historical / Previous
    ("previous biopsy site", "The", 0, "ANAT_LUNG_LOC", "Biopsy Site", "target.location", "ev2", True, False),
    ("healed well", "have", 0, "OBS_LESION", "Healed", "findings", "ev2", True, False),
    
    # Outcomes
    ("no immediate complications", "There were", 0, "OUTCOME_COMPLICATION", "None", "outcomes.complications", "ev3", False, True)
]

# Event Definitions for V3
V3_EVENTS = [
    {
        "event_id": "ev1",
        "type": "diagnostic_bronchoscopy",
        "method": "Bronchoscopy",
        "target": {"anatomy_type": "Airway"},
        "devices_json": ["Bronchoscope"],
        "findings_json": ["Normal", "Mucus suctioned"],
        "evidence_quote": "Dx bronchoscope... airway anatomy was normal"
    },
    {
        "event_id": "ev2",
        "type": "historical_finding",
        "target": {"anatomy_type": "Biopsy Site"},
        "findings_json": ["Healed well"],
        "evidence_quote": "previous biopsy site... healed well",
        "is_historical": True
    },
    {
        "event_id": "ev3",
        "type": "outcome",
        "outcomes": {"complications": "None"},
        "evidence_quote": "There were no immediate complications"
    }
]

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def hydrate_span(note_text, span_text, context_prefix=None, match_index=None):
    """
    Finds start_char and end_char for a span.
    Logic:
    1. If count == 1, use it.
    2. If context_prefix provided, search for span where prefix is in preceding 120 chars.
    3. If match_index provided, use that specific instance.
    4. Else Ambiguous.
    """
    matches = [m for m in re.finditer(re.escape(span_text), note_text)]
    
    if not matches:
        return None, None, "not_found"
    
    # Logic 1: Unique
    if len(matches) == 1:
        return matches[0].start(), matches[0].end(), "hydrated_unique"
    
    # Logic 2: Context Prefix
    if context_prefix:
        # clean prefix for loose matching (optional, but strict here)
        for m in matches:
            start = m.start()
            # Look back 120 chars
            window_start = max(0, start - 120)
            preceding_text = note_text[window_start:start]
            if context_prefix in preceding_text:
                return m.start(), m.end(), "hydrated_prefix_window"
    
    # Logic 3: Match Index
    if match_index is not None and 0 <= match_index < len(matches):
        return matches[match_index].start(), matches[match_index].end(), "hydrated_match_index"
        
    return None, None, f"ambiguous_count={len(matches)}"

# =============================================================================
# MAIN SCRIPT
# =============================================================================

def generate_workbook():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        # Create a new workbook if template is missing (fallback for isolated execution)
        wb = openpyxl.Workbook()
        # Create necessary sheets
        for sheet in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            if sheet not in wb.sheetnames:
                wb.create_sheet(sheet)
        # Remove default
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # -------------------------------------------------------------------------
    # 1. Note_Text
    # -------------------------------------------------------------------------
    ws_text = wb["Note_Text"]
    # Check headers
    if ws_text.max_row == 1 and ws_text.cell(1,1).value is None:
        ws_text.append(["note_id", "source_file", "note_text"])
    
    # Append data
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # -------------------------------------------------------------------------
    # 2. Note_Index
    # -------------------------------------------------------------------------
    ws_index = wb["Note_Index"]
    
    # Define Flag Order (must match template exactly)
    flag_order = [
        "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
        "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy",
        "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration",
        "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
        "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation",
        "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
        "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis",
        "pleural_biopsy", "fibrinolytic_therapy"
    ]
    
    # Construct row
    # Metadata cols: source_file, note_id, encounter_id, procedure_date, site, reviewer, status, free_text_notes
    meta_row = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Draft", ""]
    flag_values = [PROCEDURE_FLAGS.get(f, 0) for f in flag_order]
    
    ws_index.append(meta_row + flag_values)

    # -------------------------------------------------------------------------
    # 3. Span_Annotations (Anchor-First) & 4. Span_Hydrated
    # -------------------------------------------------------------------------
    ws_anno = wb["Span_Annotations"]
    ws_hydra = wb["Span_Hydrated"]
    
    # Headers if empty
    headers = [
        "source_file", "note_id", "span_id", "section_type",
        "context_prefix", "span_text", "match_index",
        "start_char", "end_char", "span_len",
        "label", "normalized_value", "schema_field", "event_id",
        "is_negated", "is_historical", "time_anchor", "reviewer", "comments",
        "hydration_status"
    ]
    
    if ws_anno.max_row == 1:
        ws_anno.append(headers)
    if ws_hydra.max_row == 1:
        ws_hydra.append(headers)
        
    for i, span_data in enumerate(RAW_SPANS):
        # Unpack
        text, ctx, midx, label, norm, schema, evid, is_hist, is_neg = span_data
        span_id = f"{NOTE_ID}_s{i+1:03d}"
        
        # Hydrate
        start, end, status = hydrate_span(NOTE_TEXT, text, ctx, midx)
        span_len = len(text)
        
        # Row Base
        row_base = [
            SOURCE_FILE, NOTE_ID, span_id, "", # section_type blank
            ctx, text, midx if midx is not None else "",
        ]
        
        # Anno Row (Blanks for offsets)
        row_anno = row_base + [
            "", "", f"=LEN(F{ws_anno.max_row+1})", # Formula for len
            label, norm, schema, evid,
            is_neg, is_hist, "", "", "", # time_anchor, reviewer, comments
            "needs_hydration"
        ]
        ws_anno.append(row_anno)
        
        # Hydra Row (Filled offsets)
        row_hydra = row_base + [
            start if start is not None else "", 
            end if end is not None else "", 
            span_len,
            label, norm, schema, evid,
            is_neg, is_hist, "", "", "",
            status
        ]
        ws_hydra.append(row_hydra)

    # -------------------------------------------------------------------------
    # 5. Event_Log
    # -------------------------------------------------------------------------
    ws_event = wb["Event_Log"]
    if ws_event.max_row == 1:
        ws_event.append([
            "source_file", "note_id", "event_id", "event_type", "method",
            "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements",
            "specimens", "findings", "is_historical", "reviewer", "comments",
            "device_size", "device_material",
            "outcome_airway_lumen_pre", "outcome_airway_lumen_post",
            "outcome_symptoms", "outcome_pleural", "outcome_complication"
        ])
    
    for ev in V3_EVENTS:
        row = [
            SOURCE_FILE, NOTE_ID, ev["event_id"], ev.get("type",""), ev.get("method",""),
            ev.get("target",{}).get("anatomy_type",""), 
            ", ".join(ev.get("devices_json", [])),
            "", "", "", "", # needle, stations, counts, measurements
            "", # specimens
            ", ".join(ev.get("findings_json", [])),
            ev.get("is_historical", False), "", "", # reviewer, comments
            "", "", # dev size, mat
            "", "", "", "", # outcomes lumen/symp/pleural
            ev.get("outcomes", {}).get("complications", "")
        ]
        ws_event.append(row)

    # -------------------------------------------------------------------------
    # 6. V3_Procedure_Events
    # -------------------------------------------------------------------------
    ws_v3 = wb["V3_Procedure_Events"]
    if ws_v3.max_row == 1:
        ws_v3.append([
            "note_id", "event_id", "type",
            "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station",
            "lesion.type", "lesion.size_mm",
            "method", "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote",
            "stent.size", "stent.material_or_brand", "catheter.size_fr",
            "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post",
            "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"
        ])
        
    for ev in V3_EVENTS:
        row = [
            NOTE_ID, ev["event_id"], ev.get("type",""),
            ev.get("target",{}).get("anatomy_type",""), "","", "", # target details
            "", "", # lesion
            ev.get("method",""), 
            json.dumps(ev.get("devices_json",[])),
            "[]", "[]", 
            json.dumps(ev.get("findings_json",[])),
            ev.get("evidence_quote",""),
            "", "", "", # stent/cath
            "", "", "", "", # outcome lumen/symp/pleural
            ev.get("outcomes", {}).get("complications", "")
        ]
        ws_v3.append(row)

    # -------------------------------------------------------------------------
    # 7. V3_Registry_JSON
    # -------------------------------------------------------------------------
    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 1:
        ws_json.append(["schema_version", "note_id", "json_object"])
        
    registry_obj = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "procedures": V3_EVENTS,
        "no_immediate_complications": True # derived from text
    }
    
    ws_json.append(["v3.0", NOTE_ID, json.dumps(registry_obj, indent=2)])

    # -------------------------------------------------------------------------
    # Save
    # -------------------------------------------------------------------------
    wb.save(OUTPUT_PATH)
    print(f"Successfully generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()