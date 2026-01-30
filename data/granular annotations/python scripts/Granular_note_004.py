import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
from datetime import datetime

# ==========================================
# INPUTS
# ==========================================
NOTE_ID = "note_004"
SOURCE_FILE = "note_004.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_004 SOURCE_FILE: note_004.txt INDICATION FOR OPERATION:  [REDACTED]is a 71 year old-year-old female who presents with abnormal CT chest.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
 
CONSENT : Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient or surrogate.
The patient or surrogate read and signed the provided consent form / provided consent over the phone.
The consent was witnessed by an assisting medical professional.
 
PREOPERATIVE DIAGNOSIS: R91.8 Other nonspecific abnormal finding of lung field.
POSTOPERATIVE DIAGNOSIS:  R91.8 Other nonspecific abnormal finding of lung field.
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31624 Dx bronchoscope/lavage (BAL)    
32555 - pt with pleural fluid on CT and unclear source of fevers
 
50 Bilateral Procedures (Procedure done on both sides of the body) and 73 Like 53, but the procedure was not started
 
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required a bilateral procedure today when this procedure would typically be unilateral.
Apply to: 31624 Dx bronchoscope/lavage (BAL)    .
BAL performed in the RUL and LUL due to neutropenic fever and concern for infectious etiology or two synchronous processes
 
 
ANESTHESIA: 
99152 Moderate sedation: initial 15 minutes
 
Procedure performed under moderate sedation.
The following medications were provided:
Versed             2 mg
Fentanyl          75 mcg
 
Physician/patient face-to-face anesthesia start time:   15:40
 
Physician/patient face-to-face anesthesia stop time:   16:04
 
Total moderate sedation time was 24 minutes.
Patient was monitored continuously one-to-one throughout the entire procedure by the attending physician while anesthesia was administered 
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Disposable Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
 
A timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: supine
 
Initial Airway Inspection Findings:
 
The bronchoscope was introduced into the mouth and advanced to the level of the vocal cords.
A total of 6 mL of lidocaine was applied to the vocal cords.
The bronchoscope was advanced to through the vocal cords into the trachea.
The airways were examined to the subsegmental level bilaterally. 
 
The RML was fishmouthed. The mucosa appeared normal.
Successful therapeutic aspiration was performed to clean out the LB1/2 and LB3, RB1 and 2 from mucus plug.
Bronchial alveolar lavage was performed at Posterior Segment of RUL (RB2).
Instilled 80 cc of NS, suction returned with 15 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
 
Bronchial alveolar lavage was performed at Apical-Posterior Segment of LUL (LB1/2).
Instilled 60 cc of NS, suction returned with 15 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
 
 
The patient tolerated the procedure well.  There were no immediate complications.
The bronchoscope was removed
 
SPECIMEN(S): 
BAL RUL
BAL LUL
 
IMPRESSION/PLAN: [REDACTED]is a 71 year old-year-old female who presents for bronchoscopy for abnormal CT and to obtain specimens for cultures to help identify source of fever in neutropenic patient.
[ ] f/u culture"""

# ==========================================
# CONFIGURATION & DEFINITIONS
# ==========================================

# Procedure Flags (0/1)
PROCEDURE_FLAGS = {
    'diagnostic_bronchoscopy': 1,
    'bal': 1,
    'bronchial_wash': 0,
    'brushings': 0,
    'endobronchial_biopsy': 0,
    'tbna_conventional': 0,
    'linear_ebus': 0,
    'radial_ebus': 0,
    'navigational_bronchoscopy': 0,
    'transbronchial_biopsy': 0,
    'transbronchial_cryobiopsy': 0,
    'therapeutic_aspiration': 1,
    'foreign_body_removal': 0,
    'airway_dilation': 0,
    'airway_stent': 0,
    'thermal_ablation': 0,
    'tumor_debulking_non_thermal': 0,
    'cryotherapy': 0,
    'blvr': 0,
    'peripheral_ablation': 0,
    'bronchial_thermoplasty': 0,
    'whole_lung_lavage': 0,
    'rigid_bronchoscopy': 0,
    'thoracentesis': 0,
    'chest_tube': 0,
    'ipc': 0,
    'medical_thoracoscopy': 0,
    'pleurodesis': 0,
    'pleural_biopsy': 0,
    'fibrinolytic_therapy': 0
}

# Spans Definition
# Structure: (text, label, normalized_value, event_id, context_prefix (optional), match_index (optional), schema_field (optional))
SPANS = [
    # Event 1: Inspection / RML Finding
    ("RML", "ANAT_LUNG_LOC", "RML", "evt1", "The ", 0, "target.location.lobe"),
    ("fishmouthed", "OBS_LESION", "fishmouthed", "evt1", "RML was ", 0, "findings.observation"),
    
    # Event 2: Therapeutic Aspiration
    ("therapeutic aspiration", "PROC_METHOD", "therapeutic_aspiration", "evt2", "Successful ", 0, "method"),
    ("LB1/2", "ANAT_AIRWAY", "LB1/2", "evt2", None, 0, "target.anatomy_type"),
    ("LB3", "ANAT_AIRWAY", "LB3", "evt2", None, 0, "target.anatomy_type"),
    ("RB1", "ANAT_AIRWAY", "RB1", "evt2", None, 0, "target.anatomy_type"),
    ("RB2", "ANAT_AIRWAY", "RB2", "evt2", "and ", 0, "target.anatomy_type"),
    ("mucus plug", "OBS_LESION", "mucus plug", "evt2", None, 0, "findings.observation"),

    # Event 3: BAL RUL
    ("Bronchial alveolar lavage", "PROC_METHOD", "bal", "evt3", None, 0, "method"),
    ("Posterior Segment of RUL", "ANAT_LUNG_LOC", "RUL Posterior Segment", "evt3", None, 0, "target.location.segment"),
    ("RB2", "ANAT_AIRWAY", "RB2", "evt3", "(", 1, "target.anatomy_type"), # Second RB2 in text
    ("80 cc", "MEAS_VOL", "80", "evt3", "Instilled ", 0, "measurements.volume_instilled"),
    ("15 cc", "MEAS_VOL", "15", "evt3", "returned with ", 0, "measurements.volume_return"),
    ("Cell Count", "PROC_ACTION", "specimen_collected", "evt3", "sent for ", 0, "specimens.type"),
    ("Microbiology", "PROC_ACTION", "specimen_collected", "evt3", None, 0, "specimens.type"),
    ("Cytology", "PROC_ACTION", "specimen_collected", "evt3", None, 0, "specimens.type"),

    # Event 4: BAL LUL
    ("Bronchial alveolar lavage", "PROC_METHOD", "bal", "evt4", None, 1, "method"),
    ("Apical-Posterior Segment of LUL", "ANAT_LUNG_LOC", "LUL Apical-Posterior Segment", "evt4", None, 0, "target.location.segment"),
    ("LB1/2", "ANAT_AIRWAY", "LB1/2", "evt4", "(", 1, "target.anatomy_type"), # Second LB1/2 in text
    ("60 cc", "MEAS_VOL", "60", "evt4", "Instilled ", 0, "measurements.volume_instilled"),
    ("15 cc", "MEAS_VOL", "15", "evt4", "returned with ", 1, "measurements.volume_return"),
    ("Cell Count", "PROC_ACTION", "specimen_collected", "evt4", "sent for ", 1, "specimens.type"),
    ("Microbiology", "PROC_ACTION", "specimen_collected", "evt4", None, 1, "specimens.type"),
    ("Cytology", "PROC_ACTION", "specimen_collected", "evt4", None, 1, "specimens.type"),

    # Event 5: Outcome
    ("no immediate complications", "OUTCOME_COMPLICATION", "none", "evt5", None, 0, "outcomes.complications"),
]

# Event Definitions for Registry V3
EVENTS_V3 = [
    {
        "event_id": "evt1",
        "type": "inspection",
        "target": {"anatomy_type": "airway", "location": {"lobe": "RML"}},
        "findings": ["fishmouthed"]
    },
    {
        "event_id": "evt2",
        "type": "therapeutic_aspiration",
        "method": "therapeutic_aspiration",
        "target": {"anatomy_type": "airway", "location": {"segment": "LB1/2, LB3, RB1, RB2"}},
        "findings": ["mucus plug"]
    },
    {
        "event_id": "evt3",
        "type": "bal",
        "method": "bal",
        "target": {"anatomy_type": "airway", "location": {"segment": "RUL Posterior Segment"}},
        "measurements": {"instilled": "80 cc", "return": "15 cc"},
        "specimens": ["Cell Count", "Microbiology", "Cytology"]
    },
    {
        "event_id": "evt4",
        "type": "bal",
        "method": "bal",
        "target": {"anatomy_type": "airway", "location": {"segment": "LUL Apical-Posterior Segment"}},
        "measurements": {"instilled": "60 cc", "return": "15 cc"},
        "specimens": ["Cell Count", "Microbiology", "Cytology"]
    },
    {
        "event_id": "evt5",
        "type": "outcome",
        "outcomes": {"complications": "none"}
    }
]

# ==========================================
# HELPER FUNCTIONS
# ==========================================

def hydrate_span(text, span_text, context_prefix=None, match_index=None):
    """
    Finds start/end char offsets for span_text in text.
    Priorities:
    1. context_prefix (search preceding 120 chars)
    2. match_index (0-based index of occurrence)
    3. unique occurrence
    """
    matches = [m for m in re.finditer(re.escape(span_text), text)]
    
    if not matches:
        return None, None, "not_found"

    # Strategy 1: Context Prefix
    if context_prefix:
        for m in matches:
            start = m.start()
            search_window = text[max(0, start-120):start]
            if context_prefix in search_window:
                return m.start(), m.end(), "hydrated_prefix_window"
    
    # Strategy 2: Match Index
    if match_index is not None:
        if 0 <= match_index < len(matches):
            m = matches[match_index]
            return m.start(), m.end(), "hydrated_match_index"
        else:
            return None, None, f"index_out_of_bounds_count_{len(matches)}"

    # Strategy 3: Unique
    if len(matches) == 1:
        m = matches[0]
        return m.start(), m.end(), "hydrated_unique"

    # Strategy 4: Ambiguous
    return None, None, f"ambiguous_count_{len(matches)}"

def create_worksheets_if_missing(wb):
    required = {
        'Note_Text': ['note_id', 'source_file', 'note_text'],
        'Note_Index': ['source_file', 'note_id', 'encounter_id', 'procedure_date', 'site', 'reviewer', 'status', 'free_text_notes'] + list(PROCEDURE_FLAGS.keys()),
        'Span_Annotations': ['source_file', 'note_id', 'span_id', 'section_type', 'context_prefix', 'span_text', 'match_index', 'start_char', 'end_char', 'span_len', 'label', 'normalized_value', 'schema_field', 'event_id', 'is_negated', 'is_historical', 'time_anchor', 'reviewer', 'comments', 'hydration_status'],
        'Span_Hydrated': ['source_file', 'note_id', 'span_id', 'section_type', 'context_prefix', 'span_text', 'match_index', 'start_char', 'end_char', 'span_len', 'label', 'normalized_value', 'schema_field', 'event_id', 'is_negated', 'is_historical', 'time_anchor', 'reviewer', 'comments', 'hydration_status'],
        'Event_Log': ['source_file', 'note_id', 'event_id', 'event_type', 'method', 'anatomy_target', 'device', 'needle_gauge', 'stations', 'counts', 'measurements', 'specimens', 'findings', 'is_historical', 'reviewer', 'comments', 'device_size', 'device_material', 'outcome_airway_lumen_pre', 'outcome_airway_lumen_post', 'outcome_symptoms', 'outcome_pleural', 'outcome_complication'],
        'V3_Procedure_Events': ['note_id', 'event_id', 'type', 'target.anatomy_type', 'target.location.lobe', 'target.location.segment', 'target.station', 'lesion.type', 'lesion.size_mm', 'method', 'devices_json', 'measurements_json', 'specimens_json', 'findings_json', 'evidence_quote', 'stent.size', 'stent.material_or_brand', 'catheter.size_fr', 'outcomes.airway.lumen_pre', 'outcomes.airway.lumen_post', 'outcomes.symptoms', 'outcomes.pleural', 'outcomes.complications'],
        'V3_Registry_JSON': ['schema_version', 'note_id', 'procedures_json', 'no_immediate_complications']
    }
    
    for sheet_name, headers in required.items():
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)
        else:
            # Check if headers match, if not empty, append
            ws = wb[sheet_name]
            if ws.max_row == 0:
                ws.append(headers)

# ==========================================
# EXECUTION
# ==========================================

def main():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        print(f"Template not found: {TEMPLATE_PATH}. Creating new.")
        wb = openpyxl.Workbook()
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

    create_worksheets_if_missing(wb)

    # 1. Note_Text
    ws_text = wb['Note_Text']
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws_index = wb['Note_Index']
    row_data = [SOURCE_FILE, NOTE_ID, "", "", "", "", "Pending", ""]
    # Append flags
    for k in PROCEDURE_FLAGS:
        row_data.append(PROCEDURE_FLAGS[k])
    ws_index.append(row_data)

    # 3 & 4. Span_Annotations and Span_Hydrated
    ws_raw = wb['Span_Annotations']
    ws_hydra = wb['Span_Hydrated']

    span_id_counter = 1
    
    for s_data in SPANS:
        # Unpack
        text = s_data[0]
        label = s_data[1]
        norm = s_data[2]
        evt_id = s_data[3]
        
        ctx_prefix = s_data[4] if len(s_data) > 4 else None
        m_idx = s_data[5] if len(s_data) > 5 else None
        schema_field = s_data[6] if len(s_data) > 6 else None

        # Calculate Hydration
        start, end, status = hydrate_span(NOTE_TEXT, text, ctx_prefix, m_idx)
        span_len = len(text)
        
        # Determine span_id
        span_id = f"{NOTE_ID}_s{span_id_counter:03d}"
        span_id_counter += 1

        # Common data
        # 'source_file', 'note_id', 'span_id', 'section_type', 'context_prefix', 'span_text', 'match_index', 'start_char', 'end_char', 'span_len', 'label', 'normalized_value', 'schema_field', 'event_id', 'is_negated', 'is_historical', 'time_anchor', 'reviewer', 'comments', 'hydration_status'
        
        row_raw = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure", ctx_prefix, text, m_idx, 
            None, None, f"=LEN(F{ws_raw.max_row + 1})", # Formula for length
            label, norm, schema_field, evt_id, 
            False, False, False, "", "", "needs_hydration"
        ]
        
        row_hydra = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure", ctx_prefix, text, m_idx, 
            start, end, span_len,
            label, norm, schema_field, evt_id, 
            False, False, False, "", "", status
        ]

        ws_raw.append(row_raw)
        ws_hydra.append(row_hydra)

    # 5. Event_Log (Flat)
    ws_log = wb['Event_Log']
    # Summarize based on definitions
    for evt in EVENTS_V3:
        # Map V3 dict to flat log columns
        e_type = evt.get("type", "")
        method = evt.get("method", "")
        
        # Anatomy string
        anat_target = ""
        if "target" in evt:
            loc = evt["target"].get("location", {})
            parts = []
            if "lobe" in loc: parts.append(loc["lobe"])
            if "segment" in loc: parts.append(loc["segment"])
            anat_target = ", ".join(parts)
        
        findings = ", ".join(evt.get("findings", []))
        measurements = json.dumps(evt.get("measurements", {})) if "measurements" in evt else ""
        specimens = ", ".join(evt.get("specimens", []))
        outcomes_comp = evt.get("outcomes", {}).get("complications", "")

        row_log = [
            SOURCE_FILE, NOTE_ID, evt["event_id"], e_type, method,
            anat_target, "", "", "", "", measurements, specimens, findings, 
            False, "", "", "", "", 
            "", "", "", "", outcomes_comp
        ]
        ws_log.append(row_log)

    # 6. V3_Procedure_Events
    ws_v3 = wb['V3_Procedure_Events']
    for evt in EVENTS_V3:
        # Extract nested
        target = evt.get("target", {})
        loc = target.get("location", {})
        
        row_v3 = [
            NOTE_ID, evt["event_id"], evt.get("type"),
            target.get("anatomy_type"), loc.get("lobe"), loc.get("segment"), target.get("station"),
            "", "", # Lesion type, size
            evt.get("method"),
            "", # devices_json
            json.dumps(evt.get("measurements")) if "measurements" in evt else "",
            json.dumps(evt.get("specimens")) if "specimens" in evt else "",
            json.dumps(evt.get("findings")) if "findings" in evt else "",
            "", # evidence quote
            "", "", "", # Stent/Cath
            "", "", "", "", evt.get("outcomes", {}).get("complications", "") # Outcomes
        ]
        ws_v3.append(row_v3)

    # 7. V3_Registry_JSON
    ws_json = wb['V3_Registry_JSON']
    
    registry_obj = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": EVENTS_V3,
        "no_immediate_complications": True
    }
    
    ws_json.append(["3.0", NOTE_ID, json.dumps(EVENTS_V3, indent=2), True])

    # Save
    print(f"Saving to {OUTPUT_PATH}...")
    wb.save(OUTPUT_PATH)
    print("Done.")

if __name__ == "__main__":
    main()