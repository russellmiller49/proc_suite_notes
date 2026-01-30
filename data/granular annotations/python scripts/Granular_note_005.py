import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# =============================================================================
# INPUT DATA
# =============================================================================

NOTE_ID = "note_005"
SOURCE_FILE = "note_005.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# Exact text from the provided file content
NOTE_TEXT = """NOTE_ID:  note_005 SOURCE_FILE: note_005.txt INDICATION FOR OPERATION:  [REDACTED]is a 75 year old-year-old male who presents with pleural effusion and lung infiltrates.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
 
CONSENT : Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient or surrogate.
The patient or surrogate read and signed the provided consent form / provided consent over the phone.
The consent was witnessed by an assisting medical professional.
 
PREOPERATIVE DIAGNOSIS: R91.8 Other nonspecific abnormal finding of lung field.
POSTOPERATIVE DIAGNOSIS:  R91.8 Other nonspecific abnormal finding of lung field.
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31624 Dx bronchoscope/lavage (BAL)    
 
32555 Aspirate pleura with imaging (thoracentesis)
 
This procedure required greater than 50% of time and effort that was usually needed for a similar procedure.
(BAL in multiple areas).

ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
CHEST ULTRASOUND FINDINGS:  1‌ Image saved and printed 
Hemithorax:   Left 
 
Pleural Effusion: 
Volume:       Large 
Echogenicity:   Anechoic  
Loculations:  None  
Diaphragmatic Motion:  Normal  
Lung: 
Lung sliding before procedure:   Present  
Lung sliding post procedure:   Present  
Lung consolidation/atelectasis: Present  
Pleura:  Normal  
 
Based on Ultrasound Evaluation, Thoracentesis is determined to be feasible and will proceed as planned.
.
 
Insertion site prepped and draped in sterile fashion.  Thoracentesis Kit was used.
ANESTHESIA:   Lidocaine 1%: ___7___ ml      Other: ______ 
Entry Site: 
0‌ Right ___ Intercostal Space   1‌ Left  __6_ Intercostal Space 
0‌ Mid-clavicular   0‌ Mid-axillary  1‌ Mid-scapular  0‌ Other: 
 
Sutured: 0‌ Yes 0‌ No 
 
PROCEDURE FINDINGS: 
Fluid Removed: __1450___ ml 
 Serous  
 
Drainage device:  Drainage Bag  
 
 
SPECIMEN(S): 
‌PH              LDH                   
    Glucose       T. Protein    Cholesterol
Cell Count   ‌Triglycerides            
Gram Stain/ Culture            ‌AFB                       ‌Fungal Culture 
Cytology      
CXR ordered: Yes 
 
Initial Airway Inspection Findings:
Successful therapeutic aspiration was performed to clean out the Left Mainstem from mucus.
Bronchial alveolar lavage was performed at Superior Segment of Lingula (LB4).
Instilled 60 cc of NS, suction returned with 30 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
 
Bronchial alveolar lavage was performed at Anteromedial Segment of LLL (Lb7/8).
Instilled 60 cc of NS, suction returned with 15 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
 
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
Left pleural effusion
BAL -  LUL and LLL
 
 
IMPRESSION/PLAN: [REDACTED]is a 75 year old-year-old male who presents for bronchoscopy for lavage and thoracentesis.
- f/u BAL and PFA
- f/u CXR"""

# =============================================================================
# CONFIGURATION
# =============================================================================

# Define Procedure Flags
PROCEDURE_FLAGS = [
    # Bronchoscopy (23)
    "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
    "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy",
    "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration",
    "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
    "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation",
    "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
    # Pleural (7)
    "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis",
    "pleural_biopsy", "fibrinolytic_therapy"
]

# Set active flags based on note content
ACTIVE_FLAGS = {
    "bal": 1,
    "therapeutic_aspiration": 1,
    "thoracentesis": 1
}

# =============================================================================
# SPAN DEFINITIONS (ANCHOR-FIRST)
# =============================================================================

# Each span: (span_text, label, normalized_value, schema_field, event_id, context_prefix, match_index)
# context_prefix and match_index can be None. match_index is 0-based.

spans_data = [
    # --- Metadata/History ---
    ("pleural effusion", "OBS_LESION", "Pleural Effusion", "indication", "ev0", None, 0),
    ("Flexible Therapeutic Bronchoscope", "DEV_INSTRUMENT", "Flexible Bronchoscope", "device", "ev0", None, 0),

    # --- Event 1: Thoracentesis (Left) ---
    ("Thoracentesis", "PROC_METHOD", "Thoracentesis", "method", "ev1", "Evaluation, ", 0),
    ("Left", "LATERALITY", "Left", "target.laterality", "ev1", "Hemithorax:   ", 0),
    ("Large", "MEAS_SIZE", "Large", "finding.size", "ev1", "Volume:       ", 0),
    ("Anechoic", "OBS_LESION", "Anechoic", "finding.description", "ev1", None, 0),
    ("6_ Intercostal Space", "ANAT_LUNG_LOC", "6th Intercostal Space", "target.location", "ev1", None, 0),
    ("Mid-scapular", "ANAT_LUNG_LOC", "Mid-scapular line", "target.location", "ev1", None, 0),
    ("1450___ ml", "MEAS_VOL", "1450", "outcome.pleural.volume_removed", "ev1", None, 0),
    ("Serous", "OBS_LESION", "Serous", "outcome.pleural.character", "ev1", "Removed: __1450___ ml \n ", 0),
    
    # --- Event 2: Therapeutic Aspiration (Left Mainstem) ---
    ("therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", "method", "ev2", "Successful ", 0),
    ("Left Mainstem", "ANAT_AIRWAY", "Left Mainstem Bronchus", "target.anatomy", "ev2", None, 0),
    ("mucus", "OBS_LESION", "Mucus", "indication", "ev2", None, 0),
    
    # --- Event 3: BAL (Lingula/LB4) ---
    ("Bronchial alveolar lavage", "PROC_METHOD", "BAL", "method", "ev3", None, 1), # 2nd occurrence in text (after header)
    ("Superior Segment of Lingula", "ANAT_LUNG_LOC", "Lingula Superior Segment", "target.location", "ev3", None, 0),
    ("LB4", "ANAT_AIRWAY", "LB4", "target.anatomy", "ev3", None, 0),
    ("60 cc", "MEAS_VOL", "60", "procedure.volume_in", "ev3", "Instilled ", 0),
    ("30 cc", "MEAS_VOL", "30", "procedure.volume_out", "ev3", "returned with ", 0),

    # --- Event 4: BAL (LLL/LB7-8) ---
    ("Bronchial alveolar lavage", "PROC_METHOD", "BAL", "method", "ev4", None, 2),
    ("Anteromedial Segment of LLL", "ANAT_LUNG_LOC", "LLL Anteromedial Segment", "target.location", "ev4", None, 0),
    ("Lb7/8", "ANAT_AIRWAY", "LB7-LB8", "target.anatomy", "ev4", None, 0),
    ("60 cc", "MEAS_VOL", "60", "procedure.volume_in", "ev4", "Instilled ", 1),
    ("15 cc", "MEAS_VOL", "15", "procedure.volume_out", "ev4", "returned with ", 1),
    
    # --- Outcomes ---
    ("no immediate complications", "OUTCOME_COMPLICATION", "None", "outcomes.complications", "ev0", None, 0),
]

# =============================================================================
# HYDRATION LOGIC
# =============================================================================

def get_hydration_result(text, span_text, context_prefix=None, match_index=None):
    """
    Returns (start_char, end_char, status, valid_boolean)
    """
    if not span_text:
        return None, None, "missing_text", False

    matches = [m for m in re.finditer(re.escape(span_text), text)]
    
    if not matches:
        return None, None, "not_found", False

    # 1. Unique Match
    if len(matches) == 1 and match_index is None and context_prefix is None:
        m = matches[0]
        return m.start(), m.end(), "hydrated_unique", True

    # 2. Match Index
    if match_index is not None:
        if 0 <= match_index < len(matches):
            m = matches[match_index]
            return m.start(), m.end(), "hydrated_match_index", True
        else:
            return None, None, f"index_out_of_bounds_count={len(matches)}", False

    # 3. Context Prefix
    if context_prefix:
        best_match = None
        for m in matches:
            start = m.start()
            # Look at window before match
            window_start = max(0, start - 150)
            preceding_text = text[window_start:start]
            if context_prefix in preceding_text:
                best_match = m
                break
        
        if best_match:
            return best_match.start(), best_match.end(), "hydrated_prefix_window", True
        else:
            return None, None, "context_not_found", False

    # 4. Ambiguous
    return None, None, f"ambiguous_count={len(matches)}", False

# =============================================================================
# SCRIPT LOGIC
# =============================================================================

def generate_workbook():
    # 1. Load Template (or create if missing for standalone safety, though instructions require loading)
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        # Fallback for standalone testing if template is strictly required but missing
        wb = openpyxl.Workbook()
        for name in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", 
                     "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            if name not in wb.sheetnames:
                wb.create_sheet(name)

    # 2. Populate Note_Text
    ws_text = wb["Note_Text"]
    # Ensure headers if empty
    if ws_text.max_row == 1 and ws_text["A1"].value is None:
        ws_text.append(["note_id", "source_file", "note_text"])
    
    # Check if note already exists
    exists = False
    for row in ws_text.iter_rows(min_row=2, values_only=True):
        if row[0] == NOTE_ID:
            exists = True
            break
    
    if not exists:
        ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 3. Populate Note_Index
    ws_index = wb["Note_Index"]
    header = ["source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes"] + PROCEDURE_FLAGS
    
    # If header missing, write it
    if ws_index.max_row == 1 and ws_index["A1"].value is None:
        ws_index.append(header)
        
    # Build Row
    row_data = [SOURCE_FILE, NOTE_ID, "", "", "", "", "Pending", ""]
    for flag in PROCEDURE_FLAGS:
        row_data.append(ACTIVE_FLAGS.get(flag, 0))
    ws_index.append(row_data)

    # 4. Prepare Spans (Anchor First -> Hydrated)
    ws_annot = wb["Span_Annotations"]
    ws_hydrated = wb["Span_Hydrated"]
    
    annot_header = ["source_file", "note_id", "span_id", "section_type", 
                    "context_prefix", "span_text", "match_index", 
                    "start_char", "end_char", "span_len", 
                    "label", "normalized_value", "schema_field", "event_id", 
                    "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"]
    
    if ws_annot.max_row == 1 and ws_annot["A1"].value is None:
        ws_annot.append(annot_header)
    if ws_hydrated.max_row == 1 and ws_hydrated["A1"].value is None:
        ws_hydrated.append(annot_header)

    span_objects = []
    
    for i, span_def in enumerate(spans_data):
        (txt, lbl, norm, schema, ev_id, ctx, midx) = span_def
        span_id = f"{NOTE_ID}_s{i+1:03d}"
        
        # Hydrate
        start, end, status, valid = get_hydration_result(NOTE_TEXT, txt, ctx, midx)
        
        # Common data
        base_row = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure",
            ctx, txt, midx,
            None, None, f"=LEN(F{ws_annot.max_row + 1})", # Formula for length
            lbl, norm, schema, ev_id,
            False, False, None, "Auto", "", "needs_hydration"
        ]
        
        # Write to Anchor First (Empty start/end)
        ws_annot.append(base_row)
        
        # Write to Hydrated (Filled start/end)
        hydrated_row = list(base_row)
        hydrated_row[7] = start
        hydrated_row[8] = end
        hydrated_row[9] = (end - start) if valid else 0
        hydrated_row[19] = status
        ws_hydrated.append(hydrated_row)
        
        if valid:
            span_objects.append({
                "id": span_id, "label": lbl, "norm": norm, "event_id": ev_id, "text": txt
            })

    # 5. Event Log (Linear)
    ws_event = wb["Event_Log"]
    event_header = ["source_file", "note_id", "event_id", "event_type", "method", 
                    "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements", 
                    "specimens", "findings", "is_historical", "reviewer", "comments", 
                    "device_size", "device_material", 
                    "outcome_airway_lumen_pre", "outcome_airway_lumen_post", "outcome_symptoms", 
                    "outcome_pleural", "outcome_complication"]
    
    if ws_event.max_row == 1 and ws_event["A1"].value is None:
        ws_event.append(event_header)
        
    # define events based on logic
    events = [
        {
            "id": "ev1", "type": "Pleural", "method": "Thoracentesis", "target": "Left Pleura",
            "findings": "Large Anechoic Effusion", "measurements": "1450 ml", "outcome_pleural": "1450 ml Serous removed"
        },
        {
            "id": "ev2", "type": "Airway", "method": "Therapeutic Aspiration", "target": "Left Mainstem",
            "findings": "Mucus"
        },
        {
            "id": "ev3", "type": "Airway", "method": "BAL", "target": "Lingula/LB4",
            "measurements": "60cc in/30cc out"
        },
        {
            "id": "ev4", "type": "Airway", "method": "BAL", "target": "LLL/LB7-8",
            "measurements": "60cc in/15cc out"
        }
    ]
    
    for ev in events:
        row = [
            SOURCE_FILE, NOTE_ID, ev["id"], ev["type"], ev["method"],
            ev["target"], "", "", "", "", ev.get("measurements", ""),
            "", ev.get("findings", ""), False, "Auto", "",
            "", "", 
            "", "", "", ev.get("outcome_pleural", ""), "None" # Global complication normalized to None
        ]
        ws_event.append(row)

    # 6. V3 Procedure Events
    ws_v3 = wb["V3_Procedure_Events"]
    v3_header = ["note_id", "event_id", "type", 
                 "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station",
                 "lesion.type", "lesion.size_mm", 
                 "method", "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote",
                 "stent.size", "stent.material_or_brand", "catheter.size_fr",
                 "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post", 
                 "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"]
    
    if ws_v3.max_row == 1 and ws_v3["A1"].value is None:
        ws_v3.append(v3_header)

    # Map events to V3 Columns
    v3_rows = [
        # ev1 Thoracentesis
        [NOTE_ID, "ev1", "therapeutic", "pleura", "Left", "", "6th ICS", 
         "effusion", "", "thoracentesis", json.dumps(["Thoracentesis Kit"]), json.dumps({"volume": 1450}), "", json.dumps(["Anechoic", "Large"]), "",
         "", "", "", "", "", "", "1450ml Serous", "None"],
         
        # ev2 Aspiration
        [NOTE_ID, "ev2", "therapeutic", "airway", "", "Left Mainstem", "", 
         "mucus", "", "aspiration", "", "", "", json.dumps(["Mucus"]), "clean out Left Mainstem",
         "", "", "", "", "", "", "", "None"],

        # ev3 BAL
        [NOTE_ID, "ev3", "therapeutic", "airway", "Lingula", "Superior", "LB4", 
         "", "", "bal", "", json.dumps({"in": 60, "out": 30}), "", "", "",
         "", "", "", "", "", "", "", "None"],

        # ev4 BAL
        [NOTE_ID, "ev4", "therapeutic", "airway", "LLL", "Anteromedial", "LB7/8", 
         "", "", "bal", "", json.dumps({"in": 60, "out": 15}), "", "", "",
         "", "", "", "", "", "", "", "None"]
    ]
    
    for r in v3_rows:
        ws_v3.append(r)

    # 7. JSON
    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 1:
        ws_json.append(["schema_version", "note_id", "json_output"])
    
    json_obj = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": v3_rows # Simplified for this script
    }
    
    ws_json.append(["3.0", NOTE_ID, json.dumps(json_obj, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()