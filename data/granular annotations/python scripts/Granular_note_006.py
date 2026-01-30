import openpyxl
from openpyxl.utils import get_column_letter
import datetime
import json
import re

# -------------------------------------------------------------------------
# CONSTANTS & INPUTS
# -------------------------------------------------------------------------
NOTE_ID = "note_006"
SOURCE_FILE = "note_006.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_006 SOURCE_FILE: note_006.txt DATE OF PROCEDURE: [REDACTED]

INDICATION FOR OPERATION:  [REDACTED]is a 84 year old-year-old female who presents with respiratory failure.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
 
CONSENT : Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient or surrogate.
The patient or surrogate read and signed the provided consent form / provided consent over the phone.
The consent was witnessed by an assisting medical professional.
 
PREOPERATIVE DIAGNOSIS: J96.90 Respiratory Failure
 
POSTOPERATIVE DIAGNOSIS:  J96.90 Respiratory Failure
 
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31600 Incision of windpipe (perc trach)
31615 Visualization of windpipe (Tracheobronchoscopy through established tracheostomy incision)
76536 Ultrasound of Neck
 
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Disposable Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
 
Neck Ultrasound was performed to evaluate for any abnormal vessel, mass, structures at the site of percutaneous tracheostomy.
There were no significant vessels/mass were noted overlying the tracheostomy site on examination from the laryngeal prominence to the sternal notch.
Successful therapeutic aspiration was performed to clean out the Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1) and Left Carina (LC2) from mucus and blood.
The bronchoscope was retracted into the ETT tube and the ET tube retracted into the subglottic space under direct visualization.
The inferior border of the cricoid along with the proximal tracheal rings were visualized.
Next, the anterior neck was prepped and draped in the usual sterile fashion.
Lidocaine 1% 3ml ml was injected into the anterior neck.
A 1 cm incision was made vertically with a #10 blade down through the subcutaneous tissue, just inferior to the cricoid cartilage.
The introducer needle and passed between the 1st and 2nd tracheal rings and into the trachea under direct visualization.
Next, a J-wire was passed through the catheter, also visualized with the bronchoscope.
The site was then dilated using the 14Fr introducing dilator passed over the wire.
The 14 Fr dilator was then removed from the guide wire and an 8 Fr guiding catheter placed over the guide wire until the safety ridge on the guiding catheter was at skin level.
The Portex dilator was placed over the guiding catheter until the positioning mark was visualized via the bronchoscope.
The Portex dilator was then removed leaving the guiding catheter and guide wire assembly in place, all under direct visualization bronchoscopically.
Finally a Portex 7.0 tracheostomy tube with appropriate dilator was introduced over the guiding catheter into the trachea under direct visualization.
The dilator, guiding catheter, and J-wire were then removed and the tracheostomy tube left in place.
Tracheobronchoscopy was performed with insertion of bronchoscope through the tracheostomy to perform airway clearance and confirm tracheostomy position.
The Endotracheal tube was then removed and the ventilator connected to the tracheostomy tube.
Surgicel was placed preemptively around the tracheostomy site to reduce bleeding.
A Lyofoam drain sponge was placed under the tracheostomy tube prior to suturing into place.
The patient tolerated the procedure well.  There were no complications. The staff physician was present throughout the entire procedure.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
 
 
IMPRESSION/PLAN: [REDACTED]is a 84 year old-year-old female who presents for bronchoscopy for tracheostomy.
- f/u in 7 days for tracheostomy change"""

# -------------------------------------------------------------------------
# REGISTRY CONFIGURATION
# -------------------------------------------------------------------------
# Procedure Flags (30)
FLAGS = {
    'diagnostic_bronchoscopy': 1, # Code 31615 present
    'bal': 0,
    'bronchial_wash': 0,
    'brushings': 0,
    'endobronchial_biopsy': 0,
    'tbna_conventional': 0,
    'linear_ebus': 0,
    'radial_ebus': 0,
    'navigational_bronchoscopy': 0,
    'transbronchial_biopsy': 0,
    'transbronchial_cryobiopsy': 0,
    'therapeutic_aspiration': 1, # Code 31645 present and described
    'foreign_body_removal': 0,
    'airway_dilation': 0, # Used for trach insertion, not stenosis
    'airway_stent': 0,
    'thermal_ablation': 0,
    'tumor_debulking_non_thermal': 0,
    'cryotherapy': 0,
    'blvr': 0,
    'peripheral_ablation': 0,
    'bronchial_thermoplasty': 0,
    'whole_lung_lavage': 0,
    'rigid_bronchoscopy': 0,
    'thoracentesis': 0,
    'chest_tube': 0,
    'ipc': 0,
    'medical_thoracoscopy': 0,
    'pleurodesis': 0,
    'pleural_biopsy': 0,
    'fibrinolytic_therapy': 0
}

# -------------------------------------------------------------------------
# SPAN EXTRACTION LOGIC
# -------------------------------------------------------------------------
# Structure: (text, label, norm_value, schema_field, event_id, context_prefix)
SPANS = [
    # Metadata
    ("J96.90", "OBS_LESION", "Respiratory Failure", "indication", "meta", "PREOPERATIVE DIAGNOSIS: "),
    ("Disposable Bronchoscope", "DEV_INSTRUMENT", "Disposable Bronchoscope", "device", "meta", "INSTRUMENT : \n"),
    
    # Event 1: Therapeutic Aspiration (31645)
    ("Therapeutic aspiration", "PROC_METHOD", "Therapeutic aspiration", "method", "evt1", "31645 "),
    ("clean out", "PROC_ACTION", "clean out", "action", "evt1", "performed to "),
    ("Right Mainstem", "ANAT_AIRWAY", "Right Mainstem Bronchus", "target.anatomy_type", "evt1", "clean out the "),
    ("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "target.anatomy_type", "evt1", "Right Mainstem, "),
    ("Left Mainstem", "ANAT_AIRWAY", "Left Mainstem Bronchus", "target.anatomy_type", "evt1", "Bronchus Intermedius , "),
    ("Carina", "ANAT_AIRWAY", "Carina", "target.anatomy_type", "evt1", "Left Mainstem, "),
    ("RUL Carina", "ANAT_AIRWAY", "RUL Carina", "target.anatomy_type", "evt1", "Carina, "),
    ("RML Carina", "ANAT_AIRWAY", "RML Carina", "target.anatomy_type", "evt1", "(RC1), "),
    ("LUL Lingula Carina", "ANAT_AIRWAY", "LUL Lingula Carina", "target.anatomy_type", "evt1", "(RC2), "),
    ("Left Carina", "ANAT_AIRWAY", "Left Carina", "target.anatomy_type", "evt1", "(Lc1) and "),
    ("mucus", "OBS_ROSE", "mucus", "findings", "evt1", "from "),
    ("blood", "OBS_ROSE", "blood", "findings", "evt1", "mucus and "),

    # Event 2: Percutaneous Tracheostomy (31600)
    ("Incision of windpipe", "PROC_METHOD", "Percutaneous Tracheostomy", "method", "evt2", "31600 "),
    ("proximal tracheal rings", "ANAT_AIRWAY", "Trachea", "target.anatomy_type", "evt2", "along with the "),
    ("1 cm incision", "PROC_ACTION", "Incision", "action", "evt2", "A "),
    ("introducer needle", "DEV_NEEDLE", "Introducer Needle", "device", "evt2", "The "),
    ("1st and 2nd tracheal rings", "ANAT_AIRWAY", "Tracheal Rings 1-2", "target.anatomy_type", "evt2", "between the "),
    ("J-wire", "DEV_INSTRUMENT", "Guide Wire", "device", "evt2", "Next, a "),
    ("14Fr", "DEV_CATHETER_SIZE", "14Fr", "catheter.size_fr", "evt2", "using the "),
    ("introducing dilator", "DEV_INSTRUMENT", "Introducing Dilator", "device", "evt2", "14Fr "),
    ("8 Fr guiding catheter", "DEV_INSTRUMENT", "Guiding Catheter", "device", "evt2", "wire and an "),
    ("Portex dilator", "DEV_INSTRUMENT", "Portex Dilator", "device", "evt2", "The "),
    ("Portex 7.0 tracheostomy tube", "DEV_CATHETER", "Tracheostomy Tube", "device", "evt2", "Finally a "),
    ("7.0", "DEV_CATHETER_SIZE", "7.0", "catheter.size_fr", "evt2", "Portex "),

    # Event 3: Tracheobronchoscopy / Clearance
    ("Tracheobronchoscopy", "PROC_METHOD", "Tracheobronchoscopy", "method", "evt3", "place.\n"),
    ("airway clearance", "PROC_ACTION", "Airway Clearance", "action", "evt3", "perform "),
    ("confirm tracheostomy position", "PROC_ACTION", "Confirm Position", "action", "evt3", "clearance and "),

    # Complications / Outcomes
    ("There were no immediate complications", "OUTCOME_COMPLICATION", "none", "outcomes.complications", "evt1", "procedure well.  "),
]

# -------------------------------------------------------------------------
# HELPER FUNCTIONS
# -------------------------------------------------------------------------
def create_workbook():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        # Create minimal sheets if template missing (fallback)
        for sheet in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            if sheet not in wb.sheetnames:
                wb.create_sheet(sheet)
    return wb

def get_offset(full_text, span_text, context_prefix=None, match_index=0):
    """
    Returns (start, end, status)
    status: hydrated_unique, hydrated_prefix_window, hydrated_match_index, ambiguous_count=N, not_found
    """
    if not span_text:
        return "", "", "missing_text"
    
    # 1. Exact count check
    count = full_text.count(span_text)
    if count == 0:
        return "", "", "not_found"
    
    if count == 1:
        start = full_text.find(span_text)
        return start, start + len(span_text), "hydrated_unique"
    
    # 2. Context prefix
    if context_prefix:
        # Find all occurrences
        starts = [m.start() for m in re.finditer(re.escape(span_text), full_text)]
        for s in starts:
            window_start = max(0, s - 120)
            window = full_text[window_start:s]
            if context_prefix in window:
                return s, s + len(span_text), "hydrated_prefix_window"
    
    # 3. Match index
    starts = [m.start() for m in re.finditer(re.escape(span_text), full_text)]
    if match_index < len(starts):
        s = starts[match_index]
        return s, s + len(span_text), "hydrated_match_index"
    
    return "", "", f"ambiguous_count={count}"

# -------------------------------------------------------------------------
# MAIN EXECUTION
# -------------------------------------------------------------------------
def main():
    wb = create_workbook()
    
    # 1. Note_Text
    ws = wb["Note_Text"]
    # Assuming header exists, append
    row = [NOTE_ID, SOURCE_FILE, NOTE_TEXT]
    ws.append(row)
    
    # 2. Note_Index
    ws = wb["Note_Index"]
    # Columns: source_file, note_id, encounter_id, procedure_date, site, reviewer, status, free_text_notes, [flags...]
    meta_row = [SOURCE_FILE, NOTE_ID, "", "", "", "", "Pending", ""]
    # Append flags in order
    flag_keys = list(FLAGS.keys()) # Ensure order matches template requirement if strict, but usually append works if headers align.
    # We'll just assume the template has headers in the standard order or we append values.
    # To be safe, let's construct the row with values matching keys
    flag_values = [FLAGS[k] for k in flag_keys]
    ws.append(meta_row + flag_values)
    
    # 3. Span_Annotations & 4. Span_Hydrated
    ws_anno = wb["Span_Annotations"]
    ws_hydra = wb["Span_Hydrated"]
    
    hydrated_rows = []
    
    for idx, span_data in enumerate(SPANS):
        # Unpack
        text, label, norm, schema, evt_id, ctx = span_data
        
        # Calculate hydration
        start, end, status = get_offset(NOTE_TEXT, text, ctx)
        length = len(text)
        
        # Annotation Row (Anchor First - no offsets)
        # Cols: source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, start, end, len, label, norm, schema, event_id, neg, hist, time, rev, comm, status
        span_id = f"span_{idx+1:03d}"
        
        anno_row = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure", ctx, text, "", "", "", f"=LEN(F{ws_anno.max_row+1})",
            label, norm, schema, evt_id, 0, 0, "", "", "", "needs_hydration"
        ]
        ws_anno.append(anno_row)
        
        # Hydrated Row
        hydra_row = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure", ctx, text, "", start, end, length,
            label, norm, schema, evt_id, 0, 0, "", "", "", status
        ]
        ws_hydra.append(hydra_row)
        
    # 5. Event_Log
    ws_event = wb["Event_Log"]
    # We define 3 events based on analysis
    # evt1: Therapeutic Aspiration
    # evt2: Percutaneous Tracheostomy
    # evt3: Tracheobronchoscopy
    
    events = [
        {
            "id": "evt1",
            "type": "Bronchoscopy",
            "method": "Therapeutic Aspiration",
            "anatomy": "Right Mainstem, Left Mainstem, Carina",
            "findings": "Mucus and blood",
            "outcomes": {"complication": "none"}
        },
        {
            "id": "evt2",
            "type": "Percutaneous Tracheostomy",
            "method": "Percutaneous Tracheostomy",
            "anatomy": "Trachea (1st-2nd rings)",
            "device": "Portex 7.0 Tracheostomy Tube",
            "device_size": "7.0",
            "outcomes": {"complication": "none"}
        },
        {
            "id": "evt3",
            "type": "Bronchoscopy",
            "method": "Tracheobronchoscopy",
            "anatomy": "Airway",
            "findings": "Confirmed position",
            "outcomes": {"complication": "none"}
        }
    ]
    
    for evt in events:
        # source, note, id, type, method, anat, dev, gauge, station, count, meas, spec, find, hist, rev, comm, dev_size, dev_mat, out_lumen_pre, out_lumen_post, out_symp, out_pleural, out_comp
        row = [
            SOURCE_FILE, NOTE_ID, evt["id"], evt["type"], evt["method"],
            evt.get("anatomy", ""), evt.get("device", ""), "", "", "", "", "", evt.get("findings", ""),
            0, "", "", evt.get("device_size", ""), "",
            "", "", "", "", evt.get("outcomes", {}).get("complication", "")
        ]
        ws_event.append(row)

    # 6. V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    # One row per event, structured
    for evt in events:
        row = [
            NOTE_ID, evt["id"], evt["type"],
            "Airway" if "Airway" in evt.get("anatomy","") else "Trachea", # target.anatomy
            "", "", "", # lobe, seg, station
            "", "", # lesion type, size
            evt["method"],
            json.dumps([evt.get("device", "")]), # devices
            "", "", # meas, spec
            json.dumps([evt.get("findings", "")]), # findings
            "", # quote
            "", "", # stent size, mat
            evt.get("device_size", "") if "Tracheostomy" in evt["method"] else "", # catheter size (using for trach tube size)
            "", "", "", "", evt.get("outcomes", {}).get("complication", "")
        ]
        ws_v3.append(row)

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": events
    }
    ws_json.append([json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    main()