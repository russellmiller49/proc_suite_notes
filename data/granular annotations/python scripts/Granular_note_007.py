import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os
import datetime

# =============================================================================
# INPUT DATA
# =============================================================================

NOTE_ID = "note_007"
SOURCE_FILE = "note_007.txt"
PROCEDURE_DATE = "2020-01-20" # Arbitrary date as placeholder or inferred
NOTE_TEXT = """NOTE_ID:  note_007 SOURCE_FILE: note_007.txt INDICATION FOR OPERATION:  [REDACTED]is a 68 year old-year-old male who presents with persistent airleak.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
 
CONSENT : Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient or surrogate.
The patient or surrogate read and signed the provided consent form / provided consent over the phone.
The consent was witnessed by an assisting medical professional.
 
PREOPERATIVE DIAGNOSIS: J93.82 Other airleaks
 
POSTOPERATIVE DIAGNOSIS:  J93.82 Other airleaks
 
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31624 Dx bronchoscope/lavage (BAL)    
31634 Balloon occlusion or placement of occlusive substance 
31635 Foreign body removal
31647 Bronchial valve insert initial lobe 
 
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)  31624 BAL done in multiple lobes.
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
Flexible Hybrid (Pedatric) Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
 
Successful therapeutic aspiration was performed to clean out the Right Mainstem, Bronchus Intermedius , and Left Mainstem from mucus and mucus plug.
Bronchial alveolar lavage was performed at Superior Segment of Lingula (LB4) and Inferior Segment of Lingula (LB5).
Instilled 60 cc of NS, suction returned with 15 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
 
Bronchial alveolar lavage was performed at Lateral-basal Segment of RLL (RB9).
Instilled 60 cc of NS, suction returned with 15 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
 
Serial occlusion with endobronchial blocker (ardnts 7Fr) and Fogarty balloon was done to isolate the airleak to be at the RLL (Lateral and Posterior subsegment).
Airleak was reproduced with inspiratory hold at 30 and suction on pleurovac on -20cmH20.
Tisseel 2cc was used to block off a subsegment of the RLL posterior branch.
Size 7 spiration valve was placed but noted to be too large for the airway (RB10).  This was subsequently removed.
Then Size 6 spiration valve was placed in RB9, in good position.
Then size 6 spiration valve was placed in RB10, noted to be in poor position, this was removed again and replaced with another size 6 spiration valve in a better angle.
Final: 
RB9 - size 6 spiration valve
RB10- size 6 spiration valve
 
With airleak significantly decreased.
See Dr. Thistlethwaite's note for VATS and pleurodesis. 
 
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
BAL (x2) 
 
IMPRESSION/PLAN: [REDACTED]is a 68 year old-year-old male who presents for bronchoscopy for BAL and valve placement.
- f/u in BAL results
- f/u in 6 weeks for valve removal"""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# =============================================================================
# EXTRACTION LOGIC (Hardcoded for this specific note)
# =============================================================================

# Procedure Flags (0/1)
# Based on header and narrative
# 31645 Therapeutic aspiration -> therapeutic_aspiration
# 31624 BAL -> bal
# 31634 Balloon occlusion -> (not a top level flag, but present)
# 31635 Foreign body removal -> foreign_body_removal (header + valve removal)
# 31647 Valve insert -> blvr
PROCEDURE_FLAGS = {
    'diagnostic_bronchoscopy': 0, # Usually 0 if therapeutic/BAL is the main code, though often concurrent
    'bal': 1,
    'bronchial_wash': 0,
    'brushings': 0,
    'endobronchial_biopsy': 0,
    'tbna_conventional': 0,
    'linear_ebus': 0,
    'radial_ebus': 0,
    'navigational_bronchoscopy': 0,
    'transbronchial_biopsy': 0,
    'transbronchial_cryobiopsy': 0,
    'therapeutic_aspiration': 1,
    'foreign_body_removal': 1,
    'airway_dilation': 0,
    'airway_stent': 0,
    'thermal_ablation': 0,
    'tumor_debulking_non_thermal': 0,
    'cryotherapy': 0,
    'blvr': 1,
    'peripheral_ablation': 0,
    'bronchial_thermoplasty': 0,
    'whole_lung_lavage': 0,
    'rigid_bronchoscopy': 0,
    'thoracentesis': 0,
    'chest_tube': 0,
    'ipc': 0,
    'medical_thoracoscopy': 0,
    'pleurodesis': 0, # External note
    'pleural_biopsy': 0,
    'fibrinolytic_therapy': 0
}

# Spans
# Format: (text, label, normalized_value, event_id_ref, context_prefix_or_None)
# event_id_ref is an index to the EVENTS list (1-based)

SPANS = [
    # Event 1: Therapeutic Aspiration
    ("therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", 1, "Successful "),
    ("Right Mainstem", "ANAT_AIRWAY", "RMS", 1, "clean out the "),
    ("Bronchus Intermedius", "ANAT_AIRWAY", "BI", 1, "Right Mainstem, "),
    ("Left Mainstem", "ANAT_AIRWAY", "LMS", 1, "Bronchus Intermedius , and "),
    
    # Event 2: BAL 1 (Lingula)
    ("Bronchial alveolar lavage", "PROC_METHOD", "BAL", 2, "10] "),
    ("Superior Segment of Lingula", "ANAT_LUNG_LOC", "Lingula Superior", 2, None),
    ("LB4", "ANAT_AIRWAY", "LB4", 2, "Lingula ("),
    ("Inferior Segment of Lingula", "ANAT_LUNG_LOC", "Lingula Inferior", 2, None),
    ("LB5", "ANAT_AIRWAY", "LB5", 2, "Lingula ("),
    ("Instilled 60 cc", "MEAS_VOL", "60cc", 2, "11] "),
    ("returned with 15 cc", "MEAS_VOL", "15cc", 2, "Instilled 60 cc of NS, suction "),
    
    # Event 3: BAL 2 (RLL)
    ("Bronchial alveolar lavage", "PROC_METHOD", "BAL", 3, "Cytology.\n \n "),
    ("Lateral-basal Segment of RLL", "ANAT_LUNG_LOC", "RLL Lateral Basal", 3, None),
    ("RB9", "ANAT_AIRWAY", "RB9", 3, "Lateral-basal Segment of RLL ("),
    ("Instilled 60 cc", "MEAS_VOL", "60cc", 3, "13] "),
    ("returned with 15 cc", "MEAS_VOL", "15cc", 3, "13] Instilled 60 cc of NS, suction "),

    # Event 4: Occlusion/Isolation
    ("endobronchial blocker", "DEV_INSTRUMENT", "Endobronchial Blocker", 4, None),
    ("ardnts 7Fr", "DEV_CATHETER_SIZE", "7Fr", 4, None),
    ("Fogarty balloon", "DEV_INSTRUMENT", "Fogarty Balloon", 4, None),
    ("RLL", "ANAT_LUNG_LOC", "RLL", 4, "airleak to be at the "),
    ("Lateral and Posterior subsegment", "ANAT_LUNG_LOC", "RLL Lateral and Posterior", 4, None),
    ("Airleak was reproduced", "OBS_ROSE", "Airleak Reproduced", 4, None),

    # Event 5: Tisseel
    ("Tisseel", "PROC_METHOD", "Sealant Injection", 5, None),
    ("2cc", "MEAS_VOL", "2cc", 5, "Tisseel "),
    ("RLL posterior branch", "ANAT_AIRWAY", "RLL Posterior", 5, "subsegment of the "),

    # Event 6: Valve Trial (RB10, Size 7) - Removed
    ("Size 7", "MEAS_SIZE", "Size 7", 6, "17] "),
    ("spiration valve", "DEV_VALVE", "Spiration Valve", 6, "Size 7 "),
    ("RB10", "ANAT_AIRWAY", "RB10", 6, "too large for the airway ("),
    ("removed", "PROC_ACTION", "Removed", 6, "This was subsequently "),

    # Event 7: Valve Place (RB9, Size 6)
    ("Size 6", "MEAS_SIZE", "Size 6", 7, "18] Then "),
    ("spiration valve", "DEV_VALVE", "Spiration Valve", 7, "Then Size 6 "),
    ("placed", "PROC_ACTION", "Placed", 7, "spiration valve was "),
    ("RB9", "ANAT_AIRWAY", "RB9", 7, "placed in "),

    # Event 8: Valve Place (RB10, Size 6)
    ("size 6", "MEAS_SIZE", "Size 6", 8, "19] Then "),
    ("spiration valve", "DEV_VALVE", "Spiration Valve", 8, "Then size 6 "),
    ("RB10", "ANAT_AIRWAY", "RB10", 8, "placed in "),
    ("replaced", "PROC_ACTION", "Replaced", 8, None),
    ("size 6", "MEAS_SIZE", "Size 6", 8, "replaced with another "),
    ("spiration valve", "DEV_VALVE", "Spiration Valve", 8, "replaced with another size 6 "),

    # Outcome
    ("airleak significantly decreased", "OUTCOME_PLEURAL", "Airleak Decreased", 8, None),
    ("no immediate complications", "OUTCOME_COMPLICATION", "None", 8, "There were ")
]

EVENTS = [
    {
        "id": 1,
        "type": "Therapeutic Aspiration",
        "anatomy": "RMS, BI, LMS",
        "method": "Aspiration",
        "findings": "Mucus removed"
    },
    {
        "id": 2,
        "type": "BAL",
        "anatomy": "Lingula (LB4, LB5)",
        "specimen": "Microbiology, Cytology",
        "measurements": "Instilled 60cc, Return 15cc"
    },
    {
        "id": 3,
        "type": "BAL",
        "anatomy": "RLL (RB9)",
        "specimen": "Microbiology, Cytology",
        "measurements": "Instilled 60cc, Return 15cc"
    },
    {
        "id": 4,
        "type": "Occlusion",
        "anatomy": "RLL",
        "device": "Endobronchial Blocker, Fogarty Balloon",
        "findings": "Airleak isolated"
    },
    {
        "id": 5,
        "type": "Instillation",
        "anatomy": "RLL Posterior",
        "method": "Tisseel",
        "measurements": "2cc"
    },
    {
        "id": 6,
        "type": "Valve Placement (Trial)",
        "anatomy": "RB10",
        "device": "Spiration Valve Size 7",
        "findings": "Too large, removed"
    },
    {
        "id": 7,
        "type": "Valve Placement",
        "anatomy": "RB9",
        "device": "Spiration Valve Size 6",
        "findings": "Good position"
    },
    {
        "id": 8,
        "type": "Valve Placement",
        "anatomy": "RB10",
        "device": "Spiration Valve Size 6",
        "findings": "Replaced for better angle",
        "outcome_pleural": "Airleak significantly decreased",
        "outcome_complication": "None"
    }
]

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def hydrate_span(text, span_text, context_prefix=None):
    """
    Finds start_char and end_char for a span.
    Logic:
    1. If count == 1, use it.
    2. If context_prefix, search for span preceded by prefix in window.
    3. Else ambiguous (return first for now, or None).
    """
    matches = [m.start() for m in re.finditer(re.escape(span_text), text)]
    if not matches:
        return None, None, "not_found"
    
    if len(matches) == 1:
        return matches[0], matches[0] + len(span_text), "hydrated_unique"
    
    if context_prefix:
        # Search for prefix near matches
        for m_start in matches:
            # check preceding 150 chars
            window_start = max(0, m_start - 150)
            window = text[window_start:m_start]
            if context_prefix in window:
                return m_start, m_start + len(span_text), "hydrated_prefix_window"
    
    # Default to first match if still ambiguous, but mark it
    return matches[0], matches[0] + len(span_text), f"ambiguous_count={len(matches)}"

# =============================================================================
# MAIN GENERATION
# =============================================================================

def generate_workbook():
    # Load template
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        # Create blank if missing (fallback for environment without template)
        wb = openpyxl.Workbook()
        for s in ['Note_Text', 'Note_Index', 'Span_Annotations', 'Span_Hydrated', 'Event_Log', 'V3_Procedure_Events', 'V3_Registry_JSON']:
            wb.create_sheet(s)

    # 1. Note_Text
    ws = wb['Note_Text']
    if ws.max_row == 1:
        ws.append(["note_id", "source_file", "note_text"])
    ws.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws = wb['Note_Index']
    if ws.max_row == 1:
        headers = ["source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes"]
        headers += list(PROCEDURE_FLAGS.keys())
        ws.append(headers)
    
    row_data = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Completed", ""]
    row_data += [PROCEDURE_FLAGS[k] for k in PROCEDURE_FLAGS]
    ws.append(row_data)

    # 3. Span_Annotations (Anchor First)
    ws_anno = wb['Span_Annotations']
    if ws_anno.max_row == 1:
        ws_anno.append(["source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", "match_index", 
                        "start_char", "end_char", "span_len", "label", "normalized_value", "schema_field", "event_id", 
                        "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"])
    
    # 4. Span_Hydrated
    ws_hydra = wb['Span_Hydrated']
    if ws_hydra.max_row == 1:
        ws_hydra.append(["source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", "match_index", 
                         "start_char", "end_char", "span_len", "label", "normalized_value", "schema_field", "event_id", 
                         "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"])

    # Process Spans
    hydrated_rows = []
    
    for idx, (txt, label, norm, ev_id, ctx) in enumerate(SPANS):
        span_id = f"{NOTE_ID}_s{idx+1:02d}"
        
        # Hydrate
        start, end, status = hydrate_span(NOTE_TEXT, txt, ctx)
        
        # Common data
        row_base = [SOURCE_FILE, NOTE_ID, span_id, "Procedure", ctx, txt, 1, 
                    "", "", f"=LEN(F{ws_anno.max_row+1})", label, norm, "", f"event_{ev_id}", 
                    "FALSE", "FALSE", "", "Auto", "", "needs_hydration"]
        
        ws_anno.append(row_base)
        
        # Hydrated Row
        row_hyd = list(row_base)
        row_hyd[7] = start if start is not None else ""
        row_hyd[8] = end if end is not None else ""
        row_hyd[19] = status
        ws_hydra.append(row_hyd)

    # 5. Event_Log
    ws_log = wb['Event_Log']
    if ws_log.max_row == 1:
        ws_log.append(["source_file", "note_id", "event_id", "event_type", "method", "anatomy_target", "device", 
                       "needle_gauge", "stations", "counts", "measurements", "specimens", "findings", "is_historical", 
                       "reviewer", "comments", "device_size", "device_material", 
                       "outcome_airway_lumen_pre", "outcome_airway_lumen_post", "outcome_symptoms", "outcome_pleural", "outcome_complication"])

    for ev in EVENTS:
        ev_id = f"event_{ev['id']}"
        row = [SOURCE_FILE, NOTE_ID, ev_id, ev['type'], ev.get('method', ''), ev.get('anatomy', ''), ev.get('device', ''),
               "", "", "", ev.get('measurements', ''), ev.get('specimen', ''), ev.get('findings', ''), "FALSE", 
               "Auto", "", "", "", 
               "", "", "", ev.get('outcome_pleural', ''), ev.get('outcome_complication', '')]
        ws_log.append(row)

    # 6. V3_Procedure_Events
    ws_v3 = wb['V3_Procedure_Events']
    if ws_v3.max_row == 1:
        ws_v3.append(["note_id", "event_id", "type", "target.anatomy_type", "target.location.lobe", "target.location.segment", 
                      "target.station", "lesion.type", "lesion.size_mm", "method", "devices_json", "measurements_json", 
                      "specimens_json", "findings_json", "evidence_quote", "stent.size", "stent.material_or_brand", 
                      "catheter.size_fr", "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post", "outcomes.symptoms", 
                      "outcomes.pleural", "outcomes.complications"])
    
    for ev in EVENTS:
        # Simplified mapping for demonstration
        ws_v3.append([NOTE_ID, f"event_{ev['id']}", ev['type'], "Airway", "", "", "", "", "", ev.get('method', ''), 
                      json.dumps([ev.get('device')] if ev.get('device') else []), 
                      json.dumps([ev.get('measurements')] if ev.get('measurements') else []), 
                      json.dumps([ev.get('specimen')] if ev.get('specimen') else []), 
                      json.dumps([ev.get('findings')] if ev.get('findings') else []), 
                      "", "", "", "", "", "", "", ev.get('outcome_pleural', ''), ev.get('outcome_complication', '')])

    # 7. V3_Registry_JSON
    ws_json = wb['V3_Registry_JSON']
    if ws_json.max_row == 1:
        ws_json.append(["schema_version", "note_id", "json_output"])
    
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": EVENTS,
        "no_immediate_complications": True
    }
    ws_json.append(["3.0", NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()