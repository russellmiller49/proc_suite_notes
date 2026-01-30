import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import datetime

# ==========================================
# INPUT DATA
# ==========================================
NOTE_ID = "note_008"
SOURCE_FILE = "note_008.txt"
PROCEDURE_DATE = "2026-01-12" # Using current date as placeholder or empty if unknown
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_008 SOURCE_FILE: note_008.txt 
INDICATION FOR OPERATION:  [REDACTED]is a 29 year old-year-old male who presents with subglottic stenosis and tracheal stenosis.
CONSENT: Consent obtained before the procedure. The nature, indications, purpose, benefits, risks, potential complications, and alternatives to the procedure were discussed with the patient or surrogate decision-maker in detail.
Patient or surrogate decision-maker agreed to proceed with procedure.  The patient or surrogate decision-maker read and signed the provided consent form or provided consent over the phone.
If consent was provided over the phone, then the consent was witnessed by an assisting medical professional.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
 
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31630 Balloon dilation
31631 Dilate and tracheal stent placement   
31638 Revision of tracheal/bronchial stent     
31612 tracheal puncture
 
 
Mod 22 - Pt with tracheal reconstruction and subglottic stenosis that required emergent procedure for stent revision due to narrowing of the upper airway.
There increased technical skill or placing suture through stent and trachea due to tracheal flap.
This required increased effort and skill and resulted in 100% increased effort due to intensity, mental effort and 
 
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
 
PROCEDURE IN DETAIL:
A timeout was performed (confirming the patient's name, procedure type, and procedure location).
The Flexible Therapeutic Bronchoscope was advanced for airway examination.  Endobronchial topical lidocaine applied to the vocal cords, main carina, right carina 1, and left carina 2.
 
Initial Airway Examination Findings:
Larynx:  There are raised nodules on medial aspects of arytenoids, likely consistent with irritation.
Vocal Chords: Vocal cords are widely abducted.  Vocal cords were moving bilaterally.
A total of 4 mL of 2% lidocaine was instilled onto the vocal cords.
Subglottis: granulation tissue along the left portion and right portion that was soft but cause approximately 15% obstruction.
This extended approximately 1-2 cm. 
Trachea:  The proximal and mid trachea are abnormal.
Evidence of skin graft from the 9 o'clock to 3 o'clock position;
hair follicles and longer hair stubbles seen within the airway.
Numerous sutures visualized with area of pedunculated tissue around suture sites.  No evidence of dehiscence or necrosis seen.
Main Carina: Sharp
Right Lung Proximal Airways:  There is mild webbing of tissue at RUL take-off.
Left Lung Proximal Airways:  Normal anatomic branching to first subsegmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Mucosa: Trachea as described above.  More distal airways are normal.
Secretions: Moderate, thin, and clear.  All secretions were suctioned to clear (therapeutic aspiration).
Successful therapeutic aspiration was performed to clean out the trachea, right mainstem bronchus, right upper lobe, bronchus intermedius, right middle lobe, right lower lobe, left mainstem bronchus, left upper lobe, left lower lobe from mucus.
A jag wire was introduced into the iGel and advanced to the LLL.
The bronchoscope was removed and the jag wire left in place.
A bona stent 14 x 60 mm stent was deployed under direct visualization into the trachea.
Following deployment the stent was approximately 0.5 cm more distal than desired.
Stent revision was performed 0.5 cm proximally with forceps. Following revision the stent was in appropriate position.
A 12, 13.5, 15 Balloon was used to dilate and expand the stent distally, mid portion and proximally.
A 14 gauge angiocath was used to puncture the anterior tracheal wall more superiorly.
The stent in appropriate position the bronchoscope was used to guide a a 2.0 polyene suture through the anterior trachea through the angiocath.
Using forceps the suture was grasped and pull through the trachea and into the stent.
A second 14 gauge angiocath was used to puncture the anterior tracheal wall laterally.
Then a snare was introduced through the angiocatheter into the trachea.
The snare were used to bring the first suture through the angiocatheter.
Both sutures were pulled through the skin and then tied together at the skin.
The stent was secure across 3 struts in the trachea. 
 
The patient tolerated the procedure well.
There were no immediate complications.  At the conclusion of the operation, the patient had airway device removed.
ESTIMATED BLOOD LOSS:   None
COMPLICATIONS:    None
 
SPECIMEN(S): 
none
 
IMPRESSION/PLAN: [REDACTED]is a 29 year old-year-old male who presents for bronchoscopy for airway evaluation of tracheal stenosis.
[ ] f/u CXR
[ ] restart stent hygiene with albuterol 2.5 mg nebs and 3% hypertonic saline
[ ] add 1% lidocaine neb to assist with irritation of stent in subglottis
[ ] if able to tolerate the above consider ambulation
[ ] proximal end of the stent is <0.5 cm below the true vocal cords
[ ] should patient need to be intubated would use 6.0 ETT and keep cuff just below the cords to avoid dragging stent distally with ETT.
Stent is covered and could occlude airway
[ ] if there is concern for additional migration will consider addition of second suture"""

# ==========================================
# CONFIGURATION
# ==========================================

# Procedure Flags
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 0,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 0,
    "airway_dilation": 1,
    "airway_stent": 1,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# Anchors: List of dictionaries for Span_Annotations
# Fields: span_text, label, normalized_value, schema_field, context_prefix, section_type, event_id
ANCHORS = [
    # Event 1: Initial Findings (Larynx/Subglottis)
    {
        "span_text": "raised nodules",
        "label": "OBS_LESION",
        "normalized_value": "nodules",
        "section_type": "Initial Airway Examination Findings",
        "event_id": "evt_01"
    },
    {
        "span_text": "medial aspects of arytenoids",
        "label": "ANAT_AIRWAY",
        "normalized_value": "arytenoids",
        "section_type": "Initial Airway Examination Findings",
        "event_id": "evt_01"
    },
    {
        "span_text": "granulation tissue",
        "label": "OBS_LESION",
        "normalized_value": "granulation tissue",
        "section_type": "Initial Airway Examination Findings",
        "event_id": "evt_01"
    },
    {
        "span_text": "Subglottis",
        "label": "ANAT_AIRWAY",
        "normalized_value": "subglottis",
        "section_type": "Initial Airway Examination Findings",
        "event_id": "evt_01"
    },
    {
        "span_text": "15% obstruction",
        "label": "OUTCOME_AIRWAY_LUMEN_PRE",
        "normalized_value": "15",
        "section_type": "Initial Airway Examination Findings",
        "event_id": "evt_01"
    },
    {
        "span_text": "1-2 cm",
        "label": "MEAS_SIZE",
        "normalized_value": "1-2 cm",
        "section_type": "Initial Airway Examination Findings",
        "event_id": "evt_01"
    },

    # Event 2: Trachea Findings
    {
        "span_text": "Trachea",
        "label": "ANAT_AIRWAY",
        "normalized_value": "trachea",
        "context_prefix": "This extended approximately 1-2 cm.",
        "section_type": "Initial Airway Examination Findings",
        "event_id": "evt_02"
    },
    {
        "span_text": "skin graft",
        "label": "OBS_LESION",
        "normalized_value": "skin graft",
        "section_type": "Initial Airway Examination Findings",
        "event_id": "evt_02"
    },
    {
        "span_text": "hair follicles",
        "label": "OBS_LESION",
        "normalized_value": "hair follicles",
        "section_type": "Initial Airway Examination Findings",
        "event_id": "evt_02"
    },
    {
        "span_text": "Numerous sutures",
        "label": "OBS_LESION",
        "normalized_value": "sutures",
        "section_type": "Initial Airway Examination Findings",
        "event_id": "evt_02"
    },

    # Event 3: Therapeutic Aspiration
    {
        "span_text": "Therapeutic aspiration",
        "label": "PROC_METHOD",
        "normalized_value": "therapeutic_aspiration",
        "context_prefix": "Secretions: Moderate, thin, and clear.",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt_03"
    },
    {
        "span_text": "Secretions: Moderate, thin, and clear",
        "label": "OBS_LESION",
        "normalized_value": "secretions_moderate_thin_clear",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt_03"
    },
    {
        "span_text": "trachea",
        "label": "ANAT_AIRWAY",
        "normalized_value": "trachea",
        "context_prefix": "therapeutic aspiration was performed to clean out the",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt_03"
    },
    {
        "span_text": "right mainstem bronchus",
        "label": "ANAT_AIRWAY",
        "normalized_value": "RMB",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt_03"
    },
    {
        "span_text": "Successful therapeutic aspiration",
        "label": "PROC_ACTION",
        "normalized_value": "aspiration_complete",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt_03"
    },

    # Event 4: Stent Deployment
    {
        "span_text": "jag wire",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "guide_wire",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt_04"
    },
    {
        "span_text": "bona stent",
        "label": "DEV_STENT",
        "normalized_value": "Stent (Bona)",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt_04"
    },
    {
        "span_text": "14 x 60 mm",
        "label": "DEV_STENT_SIZE",
        "normalized_value": "14x60mm",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt_04"
    },
    {
        "span_text": "trachea",
        "label": "ANAT_AIRWAY",
        "normalized_value": "trachea",
        "context_prefix": "deployed under direct visualization into the",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt_04"
    },
    {
        "span_text": "deployed",
        "label": "PROC_ACTION",
        "normalized_value": "deployed",
        "context_prefix": "14 x 60 mm stent was",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt_04"
    },

    # Event 5: Stent Revision
    {
        "span_text": "Stent revision",
        "label": "PROC_METHOD",
        "normalized_value": "stent_revision",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt_05"
    },
    {
        "span_text": "forceps",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "forceps",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt_05"
    },
    {
        "span_text": "0.5 cm proximally",
        "label": "MEAS_SIZE",
        "normalized_value": "0.5cm",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt_05"
    },

    # Event 6: Balloon Dilation
    {
        "span_text": "12, 13.5, 15 Balloon",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "Balloon (CRE 12/13.5/15mm)",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt_06"
    },
    {
        "span_text": "dilate",
        "label": "PROC_ACTION",
        "normalized_value": "dilate",
        "context_prefix": "Balloon was used to",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt_06"
    },
    {
        "span_text": "expand the stent",
        "label": "PROC_ACTION",
        "normalized_value": "expand_stent",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt_06"
    },

    # Event 7: Suture Securement (Puncture)
    {
        "span_text": "14 gauge angiocath",
        "label": "DEV_NEEDLE",
        "normalized_value": "14G Angiocath",
        "context_prefix": "A",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt_07"
    },
    {
        "span_text": "puncture",
        "label": "PROC_ACTION",
        "normalized_value": "puncture",
        "context_prefix": "used to",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt_07"
    },
    {
        "span_text": "anterior tracheal wall",
        "label": "ANAT_AIRWAY",
        "normalized_value": "trachea_anterior",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt_07"
    },
    {
        "span_text": "2.0 polyene suture",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "Suture (2.0 Polyene)",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt_07"
    },
    {
        "span_text": "snare",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "snare",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt_07"
    },
    {
        "span_text": "secure",
        "label": "PROC_ACTION",
        "normalized_value": "secure",
        "context_prefix": "The stent was",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt_07"
    },
    
    # Outcomes
    {
        "span_text": "No immediate complications",
        "label": "OUTCOME_COMPLICATION",
        "normalized_value": "none",
        "section_type": "COMPLICATIONS",
        "event_id": "evt_08"
    }
]

# ==========================================
# HELPER FUNCTIONS
# ==========================================

def calculate_offsets(note_text, span_text, context_prefix=None):
    """
    Calculates start_char and end_char for a span.
    Priority:
    1. Exact single match.
    2. Context match (preceding 120 chars).
    3. Ambiguous -> Return None
    """
    matches = [m for m in re.finditer(re.escape(span_text), note_text)]
    
    if not matches:
        return None, None, "not_found"
    
    if len(matches) == 1:
        return matches[0].start(), matches[0].end(), "hydrated_unique"
    
    # Multiple matches, try context
    if context_prefix:
        for m in matches:
            start = m.start()
            # Look back 120 chars
            window_start = max(0, start - 120)
            preceding_text = note_text[window_start:start]
            if context_prefix in preceding_text:
                return start, m.end(), "hydrated_prefix_window"
    
    return None, None, f"ambiguous_count={len(matches)}"

def generate_excel_file():
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    
    # 1. Note_Text
    ws_text = wb["Note_Text"]
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # 2. Note_Index
    ws_index = wb["Note_Index"]
    # Header row assumed present. Append data.
    row_data = [
        SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Success", ""
    ]
    # Append flags in order of keys
    for k, v in PROCEDURE_FLAGS.items():
        row_data.append(v)
    ws_index.append(row_data)
    
    # 3. Span_Annotations & 4. Span_Hydrated
    ws_anno = wb["Span_Annotations"]
    ws_hydra = wb["Span_Hydrated"]
    
    hydrated_rows = []
    
    for anchor in ANCHORS:
        span_text = anchor["span_text"]
        context_prefix = anchor.get("context_prefix", "")
        
        start, end, status = calculate_offsets(NOTE_TEXT, span_text, context_prefix)
        span_len = len(span_text)
        
        # Base row for Anno (offsets blank)
        row_anno = [
            SOURCE_FILE, NOTE_ID, "", anchor["section_type"],
            context_prefix, span_text, "", # match_index
            "", "", f"=LEN(F{ws_anno.max_row + 1})", # start, end, len formula
            anchor["label"], anchor["normalized_value"], "", anchor["event_id"],
            "FALSE", "FALSE", "", "", "", "needs_hydration"
        ]
        ws_anno.append(row_anno)
        
        # Hydrated row
        row_hydra = [
            SOURCE_FILE, NOTE_ID, "", anchor["section_type"],
            context_prefix, span_text, "",
            start if start is not None else "",
            end if end is not None else "",
            span_len,
            anchor["label"], anchor["normalized_value"], "", anchor["event_id"],
            "FALSE", "FALSE", "", "", "", status
        ]
        ws_hydra.append(row_hydra)
        hydrated_rows.append(anchor)

    # 5. Event_Log
    ws_event = wb["Event_Log"]
    
    # Construct distinct events
    events = {}
    for anchor in ANCHORS:
        eid = anchor["event_id"]
        if eid not in events:
            events[eid] = {
                "type": "therapeutic", # Default
                "method": [],
                "anatomy": [],
                "device": [],
                "finding": [],
                "outcome_compl": [],
                "outcome_lumen_pre": [],
                "outcome_lumen_post": []
            }
        
        lbl = anchor["label"]
        val = anchor["normalized_value"]
        
        if lbl == "PROC_METHOD": events[eid]["method"].append(val)
        if lbl == "ANAT_AIRWAY": events[eid]["anatomy"].append(val)
        if "DEV_" in lbl: events[eid]["device"].append(val)
        if "OBS_" in lbl: events[eid]["finding"].append(val)
        if lbl == "OUTCOME_COMPLICATION": events[eid]["outcome_compl"].append(val)
        if lbl == "OUTCOME_AIRWAY_LUMEN_PRE": events[eid]["outcome_lumen_pre"].append(val)
    
    # Write events
    for eid, data in events.items():
        # Heuristics for event type/method
        event_type = "Procedure"
        method_str = ", ".join(set(data["method"]))
        if "aspiration" in method_str: method_str = "Therapeutic Aspiration"
        elif "revision" in method_str: method_str = "Stent Revision"
        
        row_event = [
            SOURCE_FILE, NOTE_ID, eid, event_type, method_str,
            ", ".join(set(data["anatomy"])),
            ", ".join(set(data["device"])),
            "", "", "", "", # gauge, stations, counts, measurements
            "", ", ".join(set(data["finding"])), # specimens, findings
            "FALSE", "", "", # historical, reviewer, comments
            "", "", # device size, material (simplified)
            ", ".join(set(data["outcome_lumen_pre"])),
            "", "", "", # lumen_post, symptoms, pleural
            ", ".join(set(data["outcome_compl"]))
        ]
        ws_event.append(row_event)

    # 6. V3_Procedure_Events & 7. V3_Registry_JSON
    ws_v3 = wb["V3_Procedure_Events"]
    ws_json = wb["V3_Registry_JSON"]
    
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": any(a["label"] == "OUTCOME_COMPLICATION" and a["normalized_value"] == "none" for a in ANCHORS),
        "procedures": []
    }
    
    # Simplified V3 Mapping
    for eid, data in events.items():
        proc_entry = {
            "event_id": eid,
            "method": list(set(data["method"])),
            "devices": list(set(data["device"])),
            "anatomy": list(set(data["anatomy"])),
            "findings": list(set(data["finding"]))
        }
        registry_data["procedures"].append(proc_entry)
        
        ws_v3.append([
            NOTE_ID, eid, "Therapeutic",
            ", ".join(set(data["anatomy"])), "", "", "", # anatomy split
            "", "", # lesion type/size
            ", ".join(set(data["method"])),
            json.dumps(list(set(data["device"]))),
            "", "", # measurements, specimens
            json.dumps(list(set(data["finding"]))),
            "", "", "", "", # quote, stent size, material, cath size
            ", ".join(set(data["outcome_lumen_pre"])),
            "", "", "", ", ".join(set(data["outcome_compl"]))
        ])
        
    ws_json.cell(row=1, column=1, value=json.dumps(registry_data, indent=2))

    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_excel_file()