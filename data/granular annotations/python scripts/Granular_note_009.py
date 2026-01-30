import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
from datetime import datetime
import os

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_009"
SOURCE_FILE = "note_009.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_009 SOURCE_FILE: note_009.txt INDICATION FOR OPERATION:  [REDACTED]is a 29 year old-year-old male who presents with airway stenosis.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
 
CONSENT : Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient or surrogate.
The patient or surrogate read and signed the provided consent form / provided consent over the phone.
The consent was witnessed by an assisting medical professional.
 
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
 
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
 
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31624 Dx bronchoscope/lavage (BAL)    
31625 Endobronchial Biopsy(s)
31640 Bronchoscopy with excision 
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
31635 Foreign body removal
 
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)
 
IP [REDACTED] CODE MOD DETAILS: 
Unusual 
Procedure:
This patient required a significant airway management and endobronchial hair removal.
This resulted in >40% increased work due to Increased intensity, Technical difficulty of procedure, and Physical and mental effort required.
Apply to: 31625 Endobronchial Biopsy(s)
31640 Bronchoscopy with excision . 
 
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
 
Airway noted to be stable with significant airway pressure.
Stent was noted to be migrated distally. 
 
Successful therapeutic aspiration was performed to clean out the Trachea (Proximal 1/3), Trachea (Middle 1/3), Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , and Left Mainstem from mucus.
Bronchial alveolar lavage was performed at Lateral Segment of RML (RB4) and Medial Segment of RML (RB5).
Instilled 20 cc of NS, suction returned with 10 cc of NS.  Samples sent for Cell Count and Microbiology (Cultures/Viral/Fungal).
Foreign body (stent) was noted in the distal trachea.  BonaStent (14x60) was removed with forceps.
Significant amount of hair was noted and was causing endobronchial obstruction at the level of graft leading to mucus trapping.
These were mechanically removed with forceps (excision of lesion due to obstruction).
There were significant granulation tissue that was noted and APC was used to ablate the granulation issue to reduce obstruction.
Endobronchial obstruction at Subglottic was treated with the following modalities:
 
Modality	Tools	Setting/Mode	Duration	Results
Electrocautery	 	 	 	 
APC	1.5mm 	Pulse 20 effect 2, flow 0.3	2-3 sec 	ablated
Laser	 	 	 	 
CoreCath	 	 	 	 
Cryoprobe	 	 	 	 
 
Prior to treatment, affected airway was note to be 60% patent.
After treatment, the airway was 90% patent. 
 
 
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
Bonastent 14x60
BAL RML
 
IMPRESSION/PLAN: [REDACTED]is a 29 year old-year-old male who presents for bronchoscopy for airway stenosis.
- critical airway watch for overnight 
- consider CPAP overnight if get into respiratory failure
- f/u in am with CXR"""

# -------------------------------------------------------------------------
# GENERATOR CLASS
# -------------------------------------------------------------------------

class Phase0Generator:
    def __init__(self, note_text, note_id, source_file):
        self.note_text = note_text
        self.note_id = note_id
        self.source_file = source_file
        self.wb = None
        self.spans = []
        self.events = []
        self.procedure_flags = {}

    def load_template(self, template_path):
        try:
            self.wb = openpyxl.load_workbook(template_path)
        except FileNotFoundError:
            # Create a blank workbook if template doesn't exist (fallback)
            self.wb = openpyxl.Workbook()
            for sheet in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
                self.wb.create_sheet(sheet)

    def set_procedure_flags(self):
        # Default all to 0
        flags = {
            "diagnostic_bronchoscopy": 0, "bal": 0, "bronchial_wash": 0, "brushings": 0,
            "endobronchial_biopsy": 0, "tbna_conventional": 0, "linear_ebus": 0, "radial_ebus": 0,
            "navigational_bronchoscopy": 0, "transbronchial_biopsy": 0, "transbronchial_cryobiopsy": 0,
            "therapeutic_aspiration": 0, "foreign_body_removal": 0, "airway_dilation": 0,
            "airway_stent": 0, "thermal_ablation": 0, "tumor_debulking_non_thermal": 0,
            "cryotherapy": 0, "blvr": 0, "peripheral_ablation": 0, "bronchial_thermoplasty": 0,
            "whole_lung_lavage": 0, "rigid_bronchoscopy": 0,
            "thoracentesis": 0, "chest_tube": 0, "ipc": 0, "medical_thoracoscopy": 0,
            "pleurodesis": 0, "pleural_biopsy": 0, "fibrinolytic_therapy": 0
        }
        
        # Logic based on Note Content
        txt_lower = self.note_text.lower()
        
        # Diagnostic
        if "bronchoscopy" in txt_lower:
            flags["diagnostic_bronchoscopy"] = 1
        
        # BAL
        if "bronchial alveolar lavage" in txt_lower or "bal" in txt_lower:
            flags["bal"] = 1
            
        # Therapeutic Aspiration
        if "therapeutic aspiration" in txt_lower:
            flags["therapeutic_aspiration"] = 1
            
        # Foreign Body Removal (Stent)
        if "foreign body" in txt_lower or "removed with forceps" in txt_lower:
            flags["foreign_body_removal"] = 1
            
        # Thermal Ablation (APC)
        if "apc" in txt_lower and "ablated" in txt_lower:
            flags["thermal_ablation"] = 1
            
        # Tumor Debulking (Non-thermal) - Hair/Granulation with forceps
        if "mechanically removed with forceps" in txt_lower:
            flags["tumor_debulking_non_thermal"] = 1
            
        # No placement of stent, only removal
        # No biopsy in text detail (codes exist but text describes excision/removal of hair/stent)
        
        self.procedure_flags = flags

    def extract_spans_and_events(self):
        # Helper to add span
        def add_span(text, label, norm_val, field, ev_id, ctx=None):
            if text not in self.note_text:
                return
            self.spans.append({
                "text": text,
                "label": label,
                "normalized_value": norm_val,
                "schema_field": field,
                "event_id": ev_id,
                "context_prefix": ctx
            })

        # --- Event 1: Therapeutic Aspiration ---
        ev1 = "evt_01_asp"
        add_span("Successful therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", "method", ev1)
        # Locations
        locs = [
            ("Trachea (Proximal 1/3)", "Trachea"),
            ("Trachea (Middle 1/3)", "Trachea"),
            ("Trachea (Distal 1/3)", "Trachea"),
            ("Right Mainstem", "R Mainstem"),
            ("Bronchus Intermedius", "Bronchus Intermedius"),
            ("Left Mainstem", "L Mainstem")
        ]
        for t, n in locs:
            add_span(t, "ANAT_AIRWAY", n, "target.anatomy", ev1)
        
        self.events.append({
            "event_id": ev1,
            "type": "Therapeutic Aspiration",
            "method": "Suction",
            "anatomy": "Trachea, RMS, BI, LMS"
        })

        # --- Event 2: BAL ---
        ev2 = "evt_02_bal"
        add_span("Bronchial alveolar lavage", "PROC_METHOD", "BAL", "method", ev2)
        add_span("Lateral Segment of RML (RB4)", "ANAT_LUNG_LOC", "RML Lateral", "target.location", ev2)
        add_span("Medial Segment of RML (RB5)", "ANAT_LUNG_LOC", "RML Medial", "target.location", ev2)
        add_span("Instilled 20 cc", "MEAS_VOL", "20", "measurements.instilled", ev2)
        add_span("suction returned with 10 cc", "MEAS_VOL", "10", "measurements.return", ev2)
        
        self.events.append({
            "event_id": ev2,
            "type": "BAL",
            "anatomy": "RML"
        })

        # --- Event 3: Foreign Body Removal (Stent) ---
        ev3 = "evt_03_fbr"
        add_span("Foreign body (stent)", "OBS_LESION", "Foreign Body (Stent)", "lesion.type", ev3)
        add_span("Stent was noted to be migrated distally", "OBS_LESION", "Stent Migration", "lesion.type", ev3)
        add_span("distal trachea", "ANAT_AIRWAY", "Trachea Distal", "target.anatomy", ev3, "Foreign body (stent) was noted in the ")
        add_span("BonaStent", "DEV_STENT_MATERIAL", "BonaStent", "stent.brand", ev3)
        add_span("14x60", "DEV_STENT_SIZE", "14x60", "stent.size", ev3, "BonaStent (")
        add_span("removed with forceps", "PROC_ACTION", "Removal", "method", ev3, "BonaStent (14x60) was ")
        
        self.events.append({
            "event_id": ev3,
            "type": "Foreign Body Removal",
            "method": "Forceps",
            "device": "BonaStent 14x60"
        })

        # --- Event 4: Hair Removal (Debulking) ---
        ev4 = "evt_04_debulk"
        add_span("Significant amount of hair", "OBS_LESION", "Hair", "lesion.type", ev4)
        add_span("endobronchial obstruction", "OBS_LESION", "Obstruction", "lesion.type", ev4)
        add_span("mechanically removed with forceps", "PROC_ACTION", "Mechanical Removal", "method", ev4)
        
        self.events.append({
            "event_id": ev4,
            "type": "Tumor Debulking (Non-thermal)",
            "method": "Forceps",
            "finding": "Hair obstruction"
        })

        # --- Event 5: APC Ablation ---
        ev5 = "evt_05_apc"
        add_span("significant granulation tissue", "OBS_LESION", "Granulation Tissue", "lesion.type", ev5)
        add_span("Subglottic", "ANAT_AIRWAY", "Subglottic", "target.anatomy", ev5)
        add_span("APC", "PROC_METHOD", "APC", "method", ev5)
        add_span("ablated", "PROC_ACTION", "Ablation", "method", ev5, "sec \t") # using unique context from table
        
        # Outcomes (Global/Airway)
        add_span("60% patent", "OUTCOME_AIRWAY_LUMEN_PRE", "60% patent", "outcomes.airway.lumen_pre", ev5)
        add_span("90% patent", "OUTCOME_AIRWAY_LUMEN_POST", "90% patent", "outcomes.airway.lumen_post", ev5)
        add_span("no immediate complications", "OUTCOME_COMPLICATION", "None", "outcomes.complications", ev5)

        self.events.append({
            "event_id": ev5,
            "type": "Thermal Ablation",
            "method": "APC",
            "anatomy": "Subglottic",
            "outcome_pre": "60% patent",
            "outcome_post": "90% patent"
        })

    def hydrate_spans(self):
        # Calculate start/end char offsets
        hydrated = []
        for s in self.spans:
            text = s["text"]
            count = self.note_text.count(text)
            start = -1
            status = "ambiguous"

            if count == 1:
                start = self.note_text.find(text)
                status = "hydrated_unique"
            elif count > 1 and s.get("context_prefix"):
                # Use context
                prefix = s["context_prefix"]
                # Find all occurrences
                starts = [m.start() for m in re.finditer(re.escape(text), self.note_text)]
                for st in starts:
                    # Look back 150 chars
                    window = self.note_text[max(0, st-150):st]
                    if prefix in window:
                        start = st
                        status = "hydrated_prefix"
                        break
            
            # Fallback for known duplicates without explicit context in this simplified script:
            # If still -1 and count > 0, pick first (simplification for Phase 0 workbook generation)
            if start == -1 and count > 0:
                 start = self.note_text.find(text)
                 status = "hydrated_fallback_first"

            if start != -1:
                end = start + len(text)
                hydrated.append({
                    **s,
                    "start_char": start,
                    "end_char": end,
                    "hydration_status": status
                })
            else:
                hydrated.append({
                    **s,
                    "start_char": "",
                    "end_char": "",
                    "hydration_status": f"ambiguous_count_{count}"
                })
        self.hydrated_spans = hydrated

    def write_data(self):
        # 1. Note_Text
        ws = self.wb["Note_Text"]
        if ws.max_row == 1:
             ws.append(["note_id", "source_file", "note_text"])
        ws.append([self.note_id, self.source_file, self.note_text])

        # 2. Note_Index
        ws = self.wb["Note_Index"]
        headers = ["source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes"]
        flag_keys = list(self.procedure_flags.keys())
        full_headers = headers + flag_keys
        
        # Check if headers exist, if not write them
        if ws.max_row == 1 and ws.cell(1,1).value != "source_file":
            ws.append(full_headers)
        
        row = [self.source_file, self.note_id, "", PROCEDURE_DATE, "", "", "Pending", ""]
        row += [self.procedure_flags[k] for k in flag_keys]
        ws.append(row)

        # 3. Span_Annotations
        ws = self.wb["Span_Annotations"]
        span_headers = ["source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", 
                        "match_index", "start_char", "end_char", "span_len", "label", "normalized_value", 
                        "schema_field", "event_id", "is_negated", "is_historical", "time_anchor", 
                        "reviewer", "comments", "hydration_status"]
        if ws.max_row == 1:
            ws.append(span_headers)
        
        span_id_counter = 1
        for s in self.spans:
            ws.append([
                self.source_file, self.note_id, f"span_{span_id_counter:03d}", "Procedure",
                s.get("context_prefix", ""), s["text"], "", "", "", f'=LEN(F{ws.max_row+1})',
                s["label"], s["normalized_value"], s["schema_field"], s["event_id"],
                "FALSE", "FALSE", "", "", "", "needs_hydration"
            ])
            span_id_counter += 1

        # 4. Span_Hydrated
        ws = self.wb["Span_Hydrated"]
        if ws.max_row == 1:
            ws.append(span_headers)
        
        span_id_counter = 1
        for s in self.hydrated_spans:
            ws.append([
                self.source_file, self.note_id, f"span_{span_id_counter:03d}", "Procedure",
                s.get("context_prefix", ""), s["text"], "", s["start_char"], s["end_char"], len(s["text"]),
                s["label"], s["normalized_value"], s["schema_field"], s["event_id"],
                "FALSE", "FALSE", "", "", "", s["hydration_status"]
            ])
            span_id_counter += 1

        # 5. Event_Log
        ws = self.wb["Event_Log"]
        ev_headers = ["source_file", "note_id", "event_id", "event_type", "method", "anatomy_target", "device", 
                      "needle_gauge", "stations", "counts", "measurements", "specimens", "findings", "is_historical", 
                      "reviewer", "comments", "device_size", "device_material", "outcome_airway_lumen_pre", 
                      "outcome_airway_lumen_post", "outcome_symptoms", "outcome_pleural", "outcome_complication"]
        if ws.max_row == 1:
            ws.append(ev_headers)
        
        for ev in self.events:
            # Map simple dict to columns
            row = [
                self.source_file, self.note_id, ev.get("event_id"), ev.get("type"), ev.get("method"), 
                ev.get("anatomy"), ev.get("device"), "", "", "", "", "", ev.get("finding"), "FALSE",
                "", "", ev.get("device_size"), "", ev.get("outcome_pre"), ev.get("outcome_post"), "", "", ""
            ]
            ws.append(row)

        # 6. V3_Procedure_Events
        ws = self.wb["V3_Procedure_Events"]
        v3_headers = ["note_id", "event_id", "type", "target.anatomy_type", "target.location.lobe", 
                      "target.location.segment", "target.station", "lesion.type", "lesion.size_mm", "method", 
                      "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote", 
                      "stent.size", "stent.material_or_brand", "catheter.size_fr", "outcomes.airway.lumen_pre", 
                      "outcomes.airway.lumen_post", "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"]
        if ws.max_row == 1:
            ws.append(v3_headers)

        for ev in self.events:
            # Simplified V3 mapping
            ws.append([
                self.note_id, ev.get("event_id"), ev.get("type"), "Airway", "", "", "", 
                "", "", ev.get("method"), "", "", "", "", "", 
                "", "", "", ev.get("outcome_pre"), ev.get("outcome_post"), "", "", ""
            ])

        # 7. V3_Registry_JSON
        ws = self.wb["V3_Registry_JSON"]
        if ws.max_row == 1:
            ws.append(["schema_version", "note_id", "json_output", "no_immediate_complications"])
        
        json_obj = {
            "schema_version": "3.0",
            "note_id": self.note_id,
            "procedures": self.events,
            "no_immediate_complications": True
        }
        ws.append(["3.0", self.note_id, json.dumps(json_obj, indent=2), True])

    def save(self):
        self.wb.save(OUTPUT_PATH)
        print(f"Generated {OUTPUT_PATH}")

# -------------------------------------------------------------------------
# EXECUTION
# -------------------------------------------------------------------------
if __name__ == "__main__":
    # If file exists in directory (uploaded), read it
    if os.path.exists(SOURCE_FILE):
        try:
            with open(SOURCE_FILE, 'r', encoding='utf-8') as f:
                content = f.read()
                # Basic sanity check to ensure it's not empty
                if len(content) > 100:
                    NOTE_TEXT = content
        except Exception as e:
            print(f"Could not read {SOURCE_FILE}, using embedded text. Error: {e}")

    gen = Phase0Generator(NOTE_TEXT, NOTE_ID, SOURCE_FILE)
    gen.load_template(TEMPLATE_PATH)
    gen.set_procedure_flags()
    gen.extract_spans_and_events()
    gen.hydrate_spans()
    gen.write_data()
    gen.save()