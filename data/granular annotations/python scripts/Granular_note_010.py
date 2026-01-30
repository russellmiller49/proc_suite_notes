import os
import json
import datetime
from openpyxl import load_workbook, Workbook

# -----------------------------------------------------------------------------
# 1. INPUT DATA
# -----------------------------------------------------------------------------
NOTE_ID = "note_010"
SOURCE_FILE = "note_010.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_010 SOURCE_FILE: note_010.txt INDICATION FOR OPERATION:  [REDACTED]is a 33 year old-year-old female who presents with hemoptysis.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
CONSENT : Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient or surrogate.
The patient or surrogate read and signed the provided consent form / provided consent over the phone.
The consent was witnessed by an assisting medical professional.
 
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
 
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
 
PROCEDURE:  
31646 Therapeutic aspiration subsequent episodes
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
 
ANESTHESIA: 
99152 Moderate sedation: initial 15 minutes
Procedure performed under moderate sedation.
The following medications were provided:
Versed             8 mg
Fentanyl          100 mcg
 
Physician/patient face-to-face anesthesia start time:   1450
 
Physician/patient face-to-face anesthesia stop time:   1510
 
Total moderate sedation time was 20 minutes.
Patient was monitored continuously one-to-one throughout the entire procedure by the attending physician while anesthesia was administered.
Sedation was administered by RN. 
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Hybrid (Pedatric) Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
 
PATIENT POSITION: Supine
 
Initial Airway Inspection Findings:
The tracheostomy tube is in good position.
Pharynx: Not assessed due to bronchoscopy introduction through tracheostomy tube.
Larynx: Not assessed due to bronchoscopy introduction through tracheostomy tube.
Vocal Cords: Not assessed due to bronchoscopy introduction through tracheostomy tube.
Trachea: Distal 1/3 normal.
Main Carina: Sharp
Right Lung Proximal Airways: Normal anatomic branching to segmental level.  Thin rusty secretions, therapeutic aspiration was done.
Cavity noted in the RB2 subsegment. Cavity was entered. 
Left Lung Proximal Airways: Endobronchial balloon at ligula, LUL with organized clot obscuring evaluation of distal airways.
Otherwise, normal anatomic branching to segmental level.  No evidence of mass, lesions, bleeding or other endobronchial pathology.
7Fr Ardnt blocker was noted in middle of LMS. Balloon deflated. 
Mucosa: Normal.
Secretions: Minimal, thin, and clear.
Successful therapeutic aspiration was performed to clean out the Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus, blood, and blood clots.
Uniblocker (5Fr) balloon occlusion was performed at the Left Carina (LC2) with 5 Fr Uniblocker and secured in place with the Uniblocker ventilator adaptor.
-Placed at the tape level at dark blue securement device
            -Confirmed that 3cc of air was required to inflate the balloon to fully occlude the 
LUL
The endobronchial blocker balloon was intermittently inflated in the LUL to assist with clot removal.
Endobronchial clot at LUL Lingula Carina (Lc1) and Left Carina (LC2) was treated with the following modalities:
 
Modality	Tools	Setting/Mode	Duration	Results
Cryoprobe	1.1mm Cryoprobe	N/A	5-10 second freezes	Excellent clot removal
 
After organized clot was removed from the LUL orifice.
When block balloon was deflated, no blood was noted coming from LB1/2 or LB3.  The blocker was left deflated.
The patient tolerated the procedure well.  There were no immediate complications.
SPECIMEN(S): 
 
IMPRESSION/PLAN: [REDACTED]is a 33 year old-year-old female who presents for bronchoscopy for evaluation of hemoptysis, cryotherapy (cryoprobe) were used to evacuate the LUL airways of organized clot.
5 Fr Uniblocker was placed and left deflated at the LUL.
The patient tolerated the procedure well and there were no immediate complications.
 
- prn bronchoscopy as needed
- decannulation as appropriate."""

# -----------------------------------------------------------------------------
# 2. DEFINITIONS & CONFIG
# -----------------------------------------------------------------------------

# PROCEDURE FLAGS (30 total)
PROC_FLAGS = {
    # Bronchoscopy (23)
    "diagnostic_bronchoscopy": 1, # Inspection performed
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1, # "Therapeutic aspiration subsequent episodes"
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 1, # "Cryotherapy (cryoprobe) were used"
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural (7)
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# -----------------------------------------------------------------------------
# 3. SPAN EXTRACTION LOGIC
# -----------------------------------------------------------------------------

# Helper to build span objects
def make_span(text, label, norm_val, field, event_id, ctx_pre=None, match_idx=None):
    return {
        "text": text,
        "label": label,
        "norm_val": norm_val,
        "field": field,
        "event_id": event_id,
        "ctx_pre": ctx_pre,
        "match_idx": match_idx
    }

raw_spans = []

# --- Event 1: General/Initial Inspection & Aspiration ---
ev1 = "ev1" # Inspection/Aspiration
raw_spans.append(make_span("hemoptysis", "INDICATION", "hemoptysis", "indication", ev1))
raw_spans.append(make_span("Flexible Hybrid (Pedatric) Bronchoscope", "DEV_INSTRUMENT", "Hybrid Pediatric Bronchoscope", "device", ev1))
raw_spans.append(make_span("tracheostomy tube", "ANAT_AIRWAY", "tracheostomy", "access_route", ev1))

# Findings
raw_spans.append(make_span("Thin rusty secretions", "OBS_LESION", "secretions", "finding", ev1))
raw_spans.append(make_span("therapeutic aspiration was done", "PROC_METHOD", "therapeutic aspiration", "procedure", ev1, ctx_pre="Thin rusty secretions, "))
raw_spans.append(make_span("Cavity noted in the RB2 subsegment", "OBS_LESION", "cavity", "finding", ev1))
raw_spans.append(make_span("RB2 subsegment", "ANAT_AIRWAY", "RB2", "anatomy", ev1))

# Left Lung initial finding
raw_spans.append(make_span("organized clot", "OBS_LESION", "clot", "finding", ev1, ctx_pre="LUL with "))
raw_spans.append(make_span("obscuring evaluation of distal airways", "OBS_LESION", "obscuring view", "finding", ev1))

# Historical/Existing Blocker
raw_spans.append(make_span("7Fr Ardnt blocker", "DEV_CATHETER", "7Fr Arndt Blocker", "device", ev1))
raw_spans.append(make_span("middle of LMS", "ANAT_AIRWAY", "Left Mainstem", "anatomy", ev1))

# Extensive Aspiration Event (could be ev2)
ev2 = "ev2" # Therapeutic Aspiration (Detailed)
raw_spans.append(make_span("Successful therapeutic aspiration", "PROC_METHOD", "therapeutic aspiration", "procedure", ev2))
raw_spans.append(make_span("clean out the Trachea", "PROC_ACTION", "clean out", "action", ev2))
# Anatomy list
raw_spans.append(make_span("Right Mainstem", "ANAT_AIRWAY", "RMS", "anatomy", ev2))
raw_spans.append(make_span("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "anatomy", ev2))
raw_spans.append(make_span("Left Mainstem", "ANAT_AIRWAY", "LMS", "anatomy", ev2))
raw_spans.append(make_span("Carina", "ANAT_AIRWAY", "Carina", "anatomy", ev2, ctx_pre="Left Mainstem, "))
raw_spans.append(make_span("RUL Carina (RC1)", "ANAT_AIRWAY", "RUL Carina", "anatomy", ev2))
raw_spans.append(make_span("RML Carina (RC2)", "ANAT_AIRWAY", "RML Carina", "anatomy", ev2))
raw_spans.append(make_span("LUL Lingula Carina (Lc1)", "ANAT_AIRWAY", "Lingula Carina", "anatomy", ev2))
raw_spans.append(make_span("Left Carina (LC2)", "ANAT_AIRWAY", "Left Carina", "anatomy", ev2))
raw_spans.append(make_span("mucus, blood, and blood clots", "OBS_LESION", "secretions/clots", "finding", ev2))

# --- Event 3: Uniblocker Placement ---
ev3 = "ev3"
raw_spans.append(make_span("Uniblocker (5Fr) balloon occlusion", "PROC_METHOD", "balloon occlusion", "procedure", ev3))
raw_spans.append(make_span("5 Fr Uniblocker", "DEV_CATHETER", "5Fr Uniblocker", "device", ev3))
raw_spans.append(make_span("Left Carina (LC2)", "ANAT_AIRWAY", "Left Carina", "anatomy", ev3, ctx_pre="performed at the "))
raw_spans.append(make_span("Uniblocker ventilator adaptor", "DEV_INSTRUMENT", "ventilator adaptor", "device", ev3))
raw_spans.append(make_span("3cc", "MEAS_VOL", "3cc", "volume", ev3))
raw_spans.append(make_span("fully occlude the \nLUL", "PROC_ACTION", "occlude", "action", ev3))

# --- Event 4: Cryotherapy ---
ev4 = "ev4"
raw_spans.append(make_span("Endobronchial clot", "OBS_LESION", "clot", "finding", ev4))
raw_spans.append(make_span("LUL Lingula Carina (Lc1)", "ANAT_AIRWAY", "Lingula Carina", "anatomy", ev4, ctx_pre="clot at "))
raw_spans.append(make_span("Left Carina (LC2)", "ANAT_AIRWAY", "Left Carina", "anatomy", ev4, ctx_pre="and "))

raw_spans.append(make_span("Cryoprobe", "PROC_METHOD", "Cryotherapy", "procedure", ev4, ctx_pre="Modality\tTools\t"))
raw_spans.append(make_span("1.1mm Cryoprobe", "DEV_INSTRUMENT", "1.1mm Cryoprobe", "device", ev4))
raw_spans.append(make_span("5-10 second freezes", "PROC_ACTION", "freeze 5-10s", "action", ev4))
raw_spans.append(make_span("Excellent clot removal", "OUTCOME_AIRWAY_LUMEN_POST", "clot removed", "outcome", ev4))

# Final checks
raw_spans.append(make_span("no blood was noted coming from LB1/2 or LB3", "OBS_LESION", "no bleeding", "finding", ev4))
raw_spans.append(make_span("The patient tolerated the procedure well", "OUTCOME_SYMPTOMS", "tolerated well", "outcome", ev4))
raw_spans.append(make_span("no immediate complications", "OUTCOME_COMPLICATION", "none", "outcome", ev4))

# Impression/Plan
raw_spans.append(make_span("cryotherapy (cryoprobe) were used", "PROC_METHOD", "cryotherapy", "procedure", ev4))
raw_spans.append(make_span("evacuate the LUL airways", "PROC_ACTION", "evacuate airways", "action", ev4))


# -----------------------------------------------------------------------------
# 4. HYDRATION ENGINE
# -----------------------------------------------------------------------------
def hydrate_spans(text, spans):
    hydrated = []
    text_lower = text.lower()
    
    for s in spans:
        stext = s["text"]
        
        # 1. Exact count check
        count = text.count(stext)
        start = -1
        status = "ambiguous"
        
        if count == 1:
            start = text.find(stext)
            status = "hydrated_unique"
        
        # 2. Context match
        elif count > 1 and s["ctx_pre"]:
            ctx = s["ctx_pre"]
            # scan all occurrences
            curr = 0
            candidates = []
            while True:
                idx = text.find(stext, curr)
                if idx == -1: break
                # look back 120 chars
                window = text[max(0, idx-120):idx]
                if ctx in window:
                    candidates.append(idx)
                curr = idx + 1
            
            if len(candidates) == 1:
                start = candidates[0]
                status = "hydrated_prefix_window"
            elif len(candidates) > 1:
                # Fallback to match_index if provided, else take first
                if s["match_idx"] is not None and s["match_idx"] < len(candidates):
                    start = candidates[s["match_idx"]]
                    status = "hydrated_match_index_window"
                else:
                    status = f"ambiguous_count={len(candidates)}_ctx"

        # 3. Match Index (without context or failed context)
        elif count > 1 and s["match_idx"] is not None:
            # find all
            curr = 0
            locs = []
            while True:
                idx = text.find(stext, curr)
                if idx == -1: break
                locs.append(idx)
                curr = idx + 1
            if s["match_idx"] < len(locs):
                start = locs[s["match_idx"]]
                status = "hydrated_match_index"
        
        # 4. Fallback: First occurrence if no other info
        elif count > 1:
            start = text.find(stext)
            status = f"ambiguous_count={count}_default_first"
        
        # Pack result
        row = s.copy()
        if start != -1:
            row["start"] = start
            row["end"] = start + len(stext)
            row["status"] = status
        else:
            row["start"] = ""
            row["end"] = ""
            row["status"] = "not_found"
            
        hydrated.append(row)
        
    return hydrated

hydrated_data = hydrate_spans(NOTE_TEXT, raw_spans)

# -----------------------------------------------------------------------------
# 5. EXCEL GENERATION
# -----------------------------------------------------------------------------
def generate_excel(template_path: str = TEMPLATE_PATH, output_path: str = OUTPUT_PATH) -> None:
    try:
        wb = load_workbook(template_path)
    except Exception:
        wb = Workbook()  # fallback

    # --- Sheet: Note_Text ---
    if "Note_Text" not in wb.sheetnames:
        wb.create_sheet("Note_Text")
    ws_text = wb["Note_Text"]
    if ws_text.max_row == 1:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # --- Sheet: Note_Index ---
    if "Note_Index" not in wb.sheetnames:
        wb.create_sheet("Note_Index")
    ws_idx = wb["Note_Index"]
    # Header usually exists, append data
    idx_row = [
        SOURCE_FILE, NOTE_ID, "", "", "", "", "Ambiguous", ""
    ]
    # Append flag values in correct order
    flag_order = [
        "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
        "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy",
        "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration",
        "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
        "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation",
        "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
        "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis",
        "pleural_biopsy", "fibrinolytic_therapy"
    ]
    for f in flag_order:
        idx_row.append(PROC_FLAGS.get(f, 0))

    ws_idx.append(idx_row)

    # --- Sheet: Span_Annotations ---
    if "Span_Annotations" not in wb.sheetnames:
        wb.create_sheet("Span_Annotations")
    ws_span = wb["Span_Annotations"]

    for i, h in enumerate(hydrated_data):
        ws_span.append([
            SOURCE_FILE, NOTE_ID, f"span_{i+1}", "",
            h.get("ctx_pre", ""), h["text"], h.get("match_idx", ""),
            "", "", f"=LEN(F{ws_span.max_row+1})",
            h["label"], h["norm_val"], h["field"], h["event_id"],
            0, 0, "", "Model", "", "needs_hydration"
        ])

    # --- Sheet: Span_Hydrated ---
    if "Span_Hydrated" not in wb.sheetnames:
        wb.create_sheet("Span_Hydrated")
    ws_hyd = wb["Span_Hydrated"]
    for i, h in enumerate(hydrated_data):
        ws_hyd.append([
            SOURCE_FILE, NOTE_ID, f"span_{i+1}", "",
            h.get("ctx_pre", ""), h["text"], h.get("match_idx", ""),
            h["start"], h["end"], len(h["text"]),
            h["label"], h["norm_val"], h["field"], h["event_id"],
            0, 0, "", "Model", "", h["status"]
        ])

    # --- Sheet: Event_Log ---
    if "Event_Log" not in wb.sheetnames:
        wb.create_sheet("Event_Log")
    ws_evt = wb["Event_Log"]

    events = [
        {
            "id": "ev1",
            "type": "Bronchoscopy",
            "method": "Inspection",
            "anatomy": "Trachea, RB2",
            "findings": "Hemoptysis, RB2 Cavity, Rusty secretions",
            "outcome_complication": "None"
        },
        {
            "id": "ev2",
            "type": "Therapeutic Aspiration",
            "method": "Aspiration",
            "anatomy": "RUL, RML, LUL, Lingula",
            "findings": "Mucus, blood, clots",
            "outcome_complication": "None"
        },
        {
            "id": "ev3",
            "type": "Bronchial Blocker",
            "method": "Balloon Occlusion",
            "anatomy": "LUL",
            "device": "5Fr Uniblocker",
            "outcome_complication": "None"
        },
        {
            "id": "ev4",
            "type": "Cryotherapy",
            "method": "Cryoprobe",
            "anatomy": "LUL Lingula",
            "device": "1.1mm Cryoprobe",
            "findings": "Organized clot",
            "outcome_lumen_post": "Excellent clot removal",
            "outcome_complication": "None"
        }
    ]

    for e in events:
        ws_evt.append([
            SOURCE_FILE, NOTE_ID, e["id"], e["type"], e["method"],
            e.get("anatomy", ""), e.get("device", ""), "", "", "", "", "",
            e.get("findings", ""), 0, "Model", "", "", "",
            "", e.get("outcome_lumen_post", ""), "", "", e.get("outcome_complication", "")
        ])

    # --- Sheet: V3_Procedure_Events ---
    if "V3_Procedure_Events" not in wb.sheetnames:
        wb.create_sheet("V3_Procedure_Events")
    ws_v3 = wb["V3_Procedure_Events"]

    ws_v3.append([
        NOTE_ID, "ev4", "Cryotherapy", "Airway", "LUL", "Lingula", "",
        "Clot", "", "Cryoprobe", json.dumps(["1.1mm Cryoprobe"]), "", "", json.dumps(["Organized clot"]),
        "cryotherapy (cryoprobe) were used to evacuate the LUL airways", "", "", "",
        "", "Excellent clot removal", "Tolerated well", "", "None"
    ])

    # --- Sheet: V3_Registry_JSON ---
    if "V3_Registry_JSON" not in wb.sheetnames:
        wb.create_sheet("V3_Registry_JSON")
    ws_json = wb["V3_Registry_JSON"]

    reg_json = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": [
            {
                "event_id": "ev4",
                "type": "Cryotherapy",
                "method": "Cryoprobe",
                "target": {"anatomy": "LUL"},
                "outcomes": {"complications": "None", "success": True}
            }
        ]
    }
    ws_json.append([json.dumps(reg_json, indent=2)])

    wb.save(output_path)


if __name__ == "__main__":
    generate_excel()
