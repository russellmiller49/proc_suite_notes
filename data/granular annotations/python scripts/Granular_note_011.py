import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os
import sys

# -------------------------------------------------------------------------
# INPUT DATA & CONFIGURATION
# -------------------------------------------------------------------------
NOTE_ID = "note_011"
SOURCE_FILE = "note_011.txt"
PROCEDURE_DATE = ""
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# Full text from the provided file content
NOTE_TEXT = """NOTE_ID:  note_011 SOURCE_FILE: note_011.txt INDICATION FOR OPERATION:  [REDACTED]is a 33 year old-year-old female who presents with hemoptysis.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
 
CONSENT : Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient or surrogate.
The patient or surrogate read and signed the provided consent form / provided consent over the phone.
The consent was witnessed by an assisting medical professional.
 
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
 
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
 
PROCEDURE:  
31646 Therapeutic aspiration subsequent episodes
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
31634 Balloon occlusion or placement of occlusive substance 
31624 Dx bronchoscope/lavage (BAL)    
31635 Foreign body removal
 

22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)
 
IP [REDACTED] CODE MOD 
DETAILS: 
Unusual Procedure:
This patient required cryotherapy for removal of organized clot throughout the left-sided airways.
This resulted in >100% increased work due to Time, Technical difficulty of procedure, and Physical and mental effort required.
Apply to: 31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy).
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Diagnostic Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
 
PATIENT POSITION: Supine
 
Initial Airway Inspection Findings:
The tracheostomy tube is in good position.
Pharynx: Not assessed due to bronchoscopy introduction through tracheostomy tube.
Larynx: Not assessed due to bronchoscopy introduction through tracheostomy tube.
Vocal Cords: Not assessed due to bronchoscopy introduction through tracheostomy tube.
Trachea: Distal 1/3 normal.
Main Carina: Sharp
Right Lung Proximal Airways: Normal anatomic branching to segmental level.  Thin rusty secretions, therapeutic aspiration was done.
Cavity noted in the RB2 subsegment. 
Left Lung Proximal Airways: Endobronchial balloon at ligula, LUL with organized clot obscuring evaluation of distal airways.
Otherwise, normal anatomic branching to segmental level.  No evidence of mass, lesions, bleeding or other endobronchial pathology.
7Fr Ardnt blocker was noted in middle of LMS. Balloon deflated. 
Mucosa: Normal.
Secretions: Minimal, thin, and clear.
Successful therapeutic aspiration was performed to clean out the Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus, blood, and blood clots.
Blocker (foreign body) was removed under direct visualization without any additional bleeding.
New Uniblocker (5Fr) balloon occlusion was performed at the Left Carina (LC2) with 5 Fr Uniblocker and secured in place with the Uniblocker ventilator adaptor.
-Placed at the tape level at dark blue securement device
            -Confirmed that 3cc of air was required to inflate the balloon to fully occlude the 
LUL
The endobronchial blocker balloon was left inflated in the LUL.
With balloon inflated, the following was performed. 
 
Endobronchial clot at LUL Lingula Carina (Lc1) and Left Carina (LC2) was treated with the following modalities:
 
Modality	Tools	Setting/Mode	Duration	Results
Cryoprobe	1.7mm Cryoprobe	N/A	5-10 second freezes	Excellent clot removal
 
After organized clot was removed from the LUL orifice.
When block balloon was deflated, additional fresh blood was noted coming from LB1/2.
The blocker was reinflated with 3cc of air with cessation of bleeding.
Bronchial alveolar lavage was performed at Posterior-Basal Segment of LLL (LB10).
Instilled 60 cc of NS, suction returned with 15 cc of NS.  Samples sent for Microbiology (Cultures/Viral/Fungal).
The patient tolerated the procedure well.  There were no immediate complications.
SPECIMEN(S): 
BAL 
 
IMPRESSION/PLAN: [REDACTED]is a 33 year old-year-old female who presents for bronchoscopy for evaluation of hemoptysis, cryotherapy (cryoprobe) were used to evacuate the LUL airways of organized clot.
5 Fr Uniblocker was placed and left inflated at the LUL.
The patient tolerated the procedure well and there were no immediate complications.
--Post procedure CXR 
--keep blocker up till IR guided BAE 
--repeat bronch [REDACTED] for bleeding evaluation and blocker take down"""

# Procedure Flags (30)
# Rules: 1 if explicitly performed/mentioned as done in the note.
FLAGS = {
    # Bronchoscopy (23)
    'diagnostic_bronchoscopy': 1, # "Dx bronchoscope"
    'bal': 1, # "Bronchial alveolar lavage was performed"
    'bronchial_wash': 0,
    'brushings': 0,
    'endobronchial_biopsy': 0,
    'tbna_conventional': 0,
    'linear_ebus': 0,
    'radial_ebus': 0,
    'navigational_bronchoscopy': 0,
    'transbronchial_biopsy': 0,
    'transbronchial_cryobiopsy': 0, # Cryotherapy was for clot removal, not biopsy
    'therapeutic_aspiration': 1, # "Therapeutic aspiration was done"
    'foreign_body_removal': 1, # "Blocker (foreign body) was removed"
    'airway_dilation': 0,
    'airway_stent': 0,
    'thermal_ablation': 0,
    'tumor_debulking_non_thermal': 0,
    'cryotherapy': 1, # "cryotherapy for removal of organized clot"
    'blvr': 0,
    'peripheral_ablation': 0,
    'bronchial_thermoplasty': 0,
    'whole_lung_lavage': 0,
    'rigid_bronchoscopy': 0, # Flexible used
    
    # Pleural (7)
    'thoracentesis': 0,
    'chest_tube': 0,
    'ipc': 0,
    'medical_thoracoscopy': 0,
    'pleurodesis': 0,
    'pleural_biopsy': 0,
    'fibrinolytic_therapy': 0
}

# -------------------------------------------------------------------------
# EXTRACTION LOGIC: SPANS & EVENTS
# -------------------------------------------------------------------------

# Spans Structure: (text, label, normalized_value, schema_field, event_id, [optional_context_prefix], [optional_match_index])
# Event IDs:
# E01: Initial Inspection & Aspiration (Right Lung)
# E02: Removal of Existing Blocker (Foreign Body)
# E03: Broad Therapeutic Aspiration (Clot/Mucus cleanup)
# E04: Placement of New Uniblocker
# E05: Cryotherapy (Clot Removal)
# E06: Re-inflation for Bleeding Control
# E07: BAL
# E08: Outcomes/Complications

SPANS_DATA = [
    # Metadata / Indication
    ("hemoptysis", "OBS_LESION", "hemoptysis", "indication", "E00"),
    ("tracheostomy tube", "ANAT_AIRWAY", "tracheostomy", "access_route", "E00"),
    
    # E01: Right Lung Inspection/Asp
    ("Cavity noted in the RB2 subsegment", "OBS_LESION", "cavity", "finding", "E01"),
    ("RB2", "ANAT_LUNG_LOC", "RUL posterior segment", "target_location", "E01"),
    ("Thin rusty secretions", "OBS_LESION", "rusty secretions", "finding", "E01"),
    ("therapeutic aspiration was done", "PROC_METHOD", "therapeutic_aspiration", "method", "E01"),
    
    # E02: Removal of Blocker (Foreign Body)
    ("7Fr Ardnt blocker", "DEV_CATHETER", "Arndt Blocker", "device", "E02"),
    ("LMS", "ANAT_AIRWAY", "Left Mainstem", "target_location", "E02"),
    ("Blocker (foreign body) was removed", "PROC_METHOD", "foreign_body_removal", "method", "E02"),
    
    # E03: Broad Aspiration
    ("Successful therapeutic aspiration", "PROC_METHOD", "therapeutic_aspiration", "method", "E03"),
    ("clean out the Trachea", "ANAT_AIRWAY", "Trachea", "target", "E03"),
    ("Right Mainstem", "ANAT_AIRWAY", "Right Mainstem", "target", "E03"),
    ("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "target", "E03"),
    ("Left Mainstem", "ANAT_AIRWAY", "Left Mainstem", "target", "E03", "Intermedius , "), 
    ("Carina", "ANAT_AIRWAY", "Carina", "target", "E03", "Left Mainstem, "), # Disambiguate from other carinas
    ("RUL Carina (RC1)", "ANAT_AIRWAY", "RUL Carina", "target", "E03"),
    ("RML Carina (RC2)", "ANAT_AIRWAY", "RML Carina", "target", "E03"),
    ("LUL Lingula Carina (Lc1)", "ANAT_AIRWAY", "Lingula Carina", "target", "E03"),
    ("Left Carina (LC2)", "ANAT_AIRWAY", "Left Carina", "target", "E03"),
    ("mucus, blood, and blood clots", "OBS_LESION", "blood_clots", "target_lesion", "E03"),

    # E04: New Blocker Placement
    ("New Uniblocker (5Fr)", "DEV_CATHETER", "Uniblocker", "device", "E04"),
    ("balloon occlusion", "PROC_METHOD", "balloon_occlusion", "method", "E04", "New Uniblocker (5Fr) "),
    ("Left Carina (LC2)", "ANAT_AIRWAY", "Left Carina", "target_location", "E04", "performed at the "),
    ("5Fr", "DEV_CATHETER_SIZE", "5Fr", "device_size", "E04"),
    ("3cc of air", "MEAS_VOL", "3cc", "inflation_volume", "E04", "Confirmed that "),
    ("fully occlude the \nLUL", "ANAT_LUNG_LOC", "LUL", "target_location", "E04"),

    # E05: Cryotherapy
    ("Endobronchial clot", "OBS_LESION", "clot", "target_lesion", "E05"),
    ("LUL Lingula Carina (Lc1)", "ANAT_AIRWAY", "Lingula Carina", "target_location", "E05", "Endobronchial clot at "),
    ("Left Carina (LC2)", "ANAT_AIRWAY", "Left Carina", "target_location", "E05", "and "),
    ("Cryoprobe", "PROC_METHOD", "cryotherapy", "method", "E05"),
    ("1.7mm Cryoprobe", "DEV_INSTRUMENT", "Cryoprobe 1.7mm", "device", "E05"),
    ("5-10 second freezes", "MEAS_TIME", "5-10s", "duration", "E05"),
    ("Excellent clot removal", "OUTCOME_AIRWAY_LUMEN_POST", "clot_removed", "outcome_success", "E05"),
    
    # E06: Re-Bleed & Re-Inflation
    ("block balloon was deflated", "PROC_ACTION", "deflation", "action", "E06"),
    ("additional fresh blood", "OBS_LESION", "fresh_blood", "finding", "E06"),
    ("LB1/2", "ANAT_AIRWAY", "LB1+LB2", "source_location", "E06"),
    ("blocker was reinflated", "PROC_ACTION", "inflation", "action", "E06"),
    ("cessation of bleeding", "OUTCOME_SYMPTOMS", "bleeding_stopped", "outcome_success", "E06"),

    # E07: BAL
    ("Bronchial alveolar lavage", "PROC_METHOD", "bal", "method", "E07"),
    ("Posterior-Basal Segment of LLL (LB10)", "ANAT_LUNG_LOC", "LLL posterior basal segment", "target_location", "E07"),
    ("Instilled 60 cc", "MEAS_VOL", "60ml", "instilled_vol", "E07"),
    ("returned with 15 cc", "MEAS_VOL", "15ml", "return_vol", "E07"),
    ("Microbiology (Cultures/Viral/Fungal)", "PROC_ACTION", "specimen_sent", "action", "E07"),

    # E08: Complications
    ("no immediate complications", "OUTCOME_COMPLICATION", "none", "outcome", "E08"),
    ("patient tolerated the procedure well", "OUTCOME_SYMPTOMS", "tolerated_well", "outcome", "E08"),
]

# -------------------------------------------------------------------------
# UTILS
# -------------------------------------------------------------------------

def create_worksheet_if_missing(wb, sheet_name, headers):
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
    else:
        ws = wb[sheet_name]
        # Check if headers exist, if not, append them
        if ws.max_row == 0:
            ws.append(headers)
    return wb[sheet_name]

def safe_find_offsets(text, span_text, context_prefix=None, match_index=0):
    """
    Finds start/end offsets for span_text. 
    Returns (start, end, hydration_status).
    """
    if not span_text:
        return "", "", "missing_text"
    
    # Escape special regex chars
    safe_span = re.escape(span_text)
    
    # 1. Exact count check
    matches = list(re.finditer(safe_span, text))
    count = len(matches)
    
    if count == 0:
        # Fallback: try case-insensitive (but prefer exact case)
        matches = list(re.finditer(safe_span, text, re.IGNORECASE))
        count = len(matches)
        if count == 0:
            return "", "", "not_found"

    # 2. Context prefix logic
    if context_prefix:
        safe_ctx = re.escape(context_prefix)
        # Look for context within 150 chars before
        best_match = None
        for m in matches:
            start_search = max(0, m.start() - 150)
            preceding_text = text[start_search:m.start()]
            # Normalize newlines for regex matching just in case
            if re.search(safe_ctx, preceding_text, re.IGNORECASE | re.DOTALL):
                best_match = m
                break
        
        if best_match:
            return best_match.start(), best_match.end(), "hydrated_prefix_window"
        
    # 3. Unique match
    if count == 1:
        m = matches[0]
        return m.start(), m.end(), "hydrated_unique"
    
    # 4. Match index
    if match_index < count:
        m = matches[match_index]
        return m.start(), m.end(), "hydrated_match_index"
    
    return "", "", f"ambiguous_count={count}"

# -------------------------------------------------------------------------
# SCRIPT EXECUTION
# -------------------------------------------------------------------------

def main():
    # Handle File Input if script is run in an environment where it can read the directory
    # (Here we just use the embedded constant for robustness)
    content = NOTE_TEXT

    # Initialize Workbook
    if os.path.exists(TEMPLATE_PATH):
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    else:
        # Create a new workbook if template doesn't exist (fallback)
        wb = openpyxl.Workbook()

    # 1. Note_Text
    ws_text = create_worksheet_if_missing(wb, "Note_Text", ["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, content])

    # 2. Note_Index
    headers_index = [
        "source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes"
    ] + list(FLAGS.keys())
    ws_index = create_worksheet_if_missing(wb, "Note_Index", headers_index)
    
    row_data = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "extraction_done", ""]
    row_data += [FLAGS[k] for k in FLAGS]
    ws_index.append(row_data)

    # 3. Span_Annotations & 4. Span_Hydrated
    headers_span = [
        "source_file", "note_id", "span_id", "section_type",
        "context_prefix", "span_text", "match_index",
        "start_char", "end_char", "span_len",
        "label", "normalized_value", "schema_field", "event_id",
        "is_negated", "is_historical", "time_anchor", "reviewer", "comments",
        "hydration_status"
    ]
    ws_span = create_worksheet_if_missing(wb, "Span_Annotations", headers_span)
    ws_hydrated = create_worksheet_if_missing(wb, "Span_Hydrated", headers_span)

    span_id_counter = 1
    
    for item in SPANS_DATA:
        # Default defaults
        ctx = ""
        midx = 0
        
        # Unpack based on length
        if len(item) == 5:
            txt, lbl, norm, field, evt = item
        elif len(item) == 6:
            txt, lbl, norm, field, evt, ctx = item
        elif len(item) == 7:
            txt, lbl, norm, field, evt, ctx, midx = item
        else:
            continue

        # Calculate offsets
        start, end, status = safe_find_offsets(content, txt, ctx, midx)
        span_len_val = (end - start) if isinstance(end, int) else 0
        span_id = f"{NOTE_ID}_s{span_id_counter:03d}"
        
        # Anchor Row (Empty offsets)
        row_anchor = [
            SOURCE_FILE, NOTE_ID, span_id, "", # section blank
            ctx, txt, midx,
            "", "", f"=LEN(F{ws_span.max_row+1})",
            lbl, norm, field, evt,
            0, 0, "", "model", "",
            "needs_hydration"
        ]
        ws_span.append(row_anchor)
        
        # Hydrated Row
        row_hydrated = [
            SOURCE_FILE, NOTE_ID, span_id, "",
            ctx, txt, midx,
            start, end, span_len_val,
            lbl, norm, field, evt,
            0, 0, "", "model", "",
            status
        ]
        ws_hydrated.append(row_hydrated)
        
        span_id_counter += 1

    # 5. Event_Log
    headers_event = [
        "source_file", "note_id", "event_id", "event_type", "method",
        "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements",
        "specimens", "findings", "is_historical", "reviewer", "comments",
        "device_size", "device_material",
        "outcome_airway_lumen_pre", "outcome_airway_lumen_post",
        "outcome_symptoms", "outcome_pleural", "outcome_complication"
    ]
    ws_event = create_worksheet_if_missing(wb, "Event_Log", headers_event)
    
    # Event Data Construction
    events_list = [
        {"id": "E01", "type": "therapeutic", "method": "therapeutic_aspiration", "target": "RUL posterior segment", "findings": "cavity, rusty secretions"},
        {"id": "E02", "type": "therapeutic", "method": "foreign_body_removal", "device": "Arndt Blocker 7Fr", "target": "Left Mainstem"},
        {"id": "E03", "type": "therapeutic", "method": "therapeutic_aspiration", "target": "Multiple Airways", "findings": "mucus, blood, clots"},
        {"id": "E04", "type": "therapeutic", "method": "balloon_occlusion", "device": "Uniblocker 5Fr", "measurements": "3cc", "target": "LUL", "device_size": "5Fr"},
        {"id": "E05", "type": "therapeutic", "method": "cryotherapy", "device": "Cryoprobe 1.7mm", "target": "LUL", "outcome_lumen_post": "clot_removed"},
        {"id": "E06", "type": "therapeutic", "method": "balloon_occlusion", "action": "reinflation", "target": "LB1+LB2", "outcome_symptoms": "bleeding_stopped"},
        {"id": "E07", "type": "diagnostic", "method": "bal", "target": "LLL posterior basal", "measurements": "Instilled 60cc, Return 15cc", "specimens": "microbiology"},
        {"id": "E08", "type": "outcome", "method": "n/a", "outcome_complication": "none", "outcome_symptoms": "tolerated_well"}
    ]
    
    for ev in events_list:
        row = [
            SOURCE_FILE, NOTE_ID, ev["id"], ev.get("type", ""), ev.get("method", ""),
            ev.get("target", ""), ev.get("device", ""), "", "", "", ev.get("measurements", ""),
            ev.get("specimens", ""), ev.get("findings", ""), 0, "model", "",
            ev.get("device_size", ""), "", # material empty
            "", ev.get("outcome_lumen_post", ""),
            ev.get("outcome_symptoms", ""), "", ev.get("outcome_complication", "")
        ]
        ws_event.append(row)

    # 6. V3_Procedure_Events
    headers_v3 = [
        "note_id", "event_id", "type",
        "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station",
        "lesion.type", "lesion.size_mm",
        "method", "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote",
        "stent.size", "stent.material_or_brand", "catheter.size_fr",
        "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post",
        "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"
    ]
    ws_v3 = create_worksheet_if_missing(wb, "V3_Procedure_Events", headers_v3)
    
    # Populate V3 using mapped events for richer JSON detail
    # E05 Cryotherapy
    ws_v3.append([
        NOTE_ID, "E05", "therapeutic",
        "airway", "LUL", "Lingula", "",
        "clot", "",
        "cryotherapy", json.dumps(["Cryoprobe 1.7mm"]), json.dumps({"duration": "5-10s"}), "", json.dumps(["Excellent clot removal"]), "Excellent clot removal",
        "", "", "",
        "", "clot_removed",
        "", "", ""
    ])
    
    # E07 BAL
    ws_v3.append([
        NOTE_ID, "E07", "diagnostic",
        "airway", "LLL", "Posterior Basal", "",
        "", "",
        "bal", "", json.dumps({"instilled": "60cc", "return": "15cc"}), json.dumps(["microbiology"]), "", "Samples sent for Microbiology",
        "", "", "",
        "", "",
        "", "", ""
    ])

    # 7. V3_Registry_JSON
    ws_json = create_worksheet_if_missing(wb, "V3_Registry_JSON", ["schema_version", "note_id", "json_output"])
    
    registry_data = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "procedures": events_list,
        "no_immediate_complications": True
    }
    
    ws_json.append(["3.0", NOTE_ID, json.dumps(registry_data, indent=2)])

    # Final Save
    wb.save(OUTPUT_PATH)
    print(f"Successfully generated {OUTPUT_PATH}")

if __name__ == "__main__":
    main()