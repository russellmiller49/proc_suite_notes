import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_012"
SOURCE_FILE = "note_012.txt"
PROCEDURE_DATE = ""  # Not explicitly in header; inferred context suggests Jan 2026 but unclear specific date.
NOTE_TEXT = """NOTE_ID:  note_012 SOURCE_FILE: note_012.txt INDICATION FOR OPERATION:  [REDACTED]is a 33 year old-year-old female who presents with aspergilloma and hemoptysis.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
 
CONSENT : Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient or surrogate.
The patient or surrogate read and signed the provided consent form / provided consent over the phone.
The consent was witnessed by an assisting medical professional.
 
PREOPERATIVE DIAGNOSIS: R91.8 Other nonspecific abnormal finding of lung field.
POSTOPERATIVE DIAGNOSIS:  R91.8 Other nonspecific abnormal finding of lung field.
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31573 Therapeutic injection(s) [eg, chemotherapy denervation agent or corticosteroid, injected percutaneous, transoral, or via endoscope channel]
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
31899NFD BRONCHOSCOPY W/ APPLICATION OF TRANEXAMIC ACID
 
 
ANESTHESIA: 
99152 Moderate sedation: initial 15 minutes
99153 Moderate sedation: each additional 15 minutes 
 
Procedure performed under moderate sedation.
The following medications were provided:
Etomidate                    20 mg
Rocuronium                 26 mg
Propofol gtt                  40 mcg/kg/min
Fentanyl gtt                 150 mcg/hr
 
Physician/patient face-to-face anesthesia start time:   1500
 
Physician/patient face-to-face anesthesia stop time:   1550
 
Total 
moderate sedation time was 50 minutes.  
 
Patient was monitored continuously one-to-one throughout the entire procedure by the attending physician while anesthesia was administered.
Sedation was administered by ICU RN. 
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Disposable Bronchoscope
 
ESTIMATED BLOOD LOSS:   Minimum
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
 
The tracheostomy tube is in good position.
Pharynx: Not assessed due to bronchoscopy introduction through tracheostomy tube.
Larynx: Not assessed due to bronchoscopy introduction through tracheostomy tube.
Vocal Cords: Not assessed due to bronchoscopy introduction through tracheostomy tube.
Trachea: Distal 1/3 normal.
Main Carina: Sharp
Right Lung Proximal Airways: Normal anatomic branching to segmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Left Lung Proximal Airways: Normal anatomic branching to segmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology. Uniblocker in place (deflated) at LMSB
Mucosa: Normal.
Secretions: Mild bloody secretions with left greater than right
 
Successful therapeutic aspiration was performed to clean out the Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus, blood, and blood clots.
Amphotericin 50mg in 20cc sterile water was instilled into the LUL with 10cc into the apico-posterior (LB1/2) subsegment and 10cc into the anterior (LB3) subsegment

Post-instillation, patient had mild ooze of fresh blood.
TXA total 1000mg was applied directly into the LUL through the bronchoscope.
Formalized clot was then extracted with cryoprobe to relieve airway obstruction.
Endobronchial clot at LUL Lingula Carina (Lc1) and Left Carina (LC2) was treated with the following modalities:
 
Modality	Tools	Setting/Mode	Duration	Results
Cryoprobe	1.1mm Cryoprobe	N/A	5-10 second freezes	Excellent clot removal
 
 
The patient tolerated the procedure well.
There were no immediate complications.  
SPECIMEN(S): 
none
 
IMPRESSION/PLAN: [REDACTED]is a 33 year old-year-old female who presents for bronchoscopy for for instillation of Amphotericin.
Patient tolerated the procedure well and there were no immediate complications.
--Continued care per primary team
--May consider repeat instillation of Amphotericin on 1/5/2026 and possible clot extraction"""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# -------------------------------------------------------------------------
# CONFIGURATION
# -------------------------------------------------------------------------

# Procedure Flags (30 total)
PROC_FLAGS = [
    "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
    "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy",
    "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration",
    "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
    "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation",
    "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
    "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy",
    "pleurodesis", "pleural_biopsy", "fibrinolytic_therapy"
]

# -------------------------------------------------------------------------
# LOGIC & PARSING
# -------------------------------------------------------------------------

def create_workbook():
    # 1. Load Template
    if not os.path.exists(TEMPLATE_PATH):
        # Fallback: create a blank workbook if template missing
        wb = openpyxl.Workbook()
        wb.remove(wb.active) # Remove default sheet
        for sheet_name in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            wb.create_sheet(sheet_name)
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    
    # 2. Populate Note_Text
    ws_text = wb["Note_Text"]
    if ws_text.max_row == 1 and ws_text.cell(1,1).value is None:
        ws_text.append(["note_id", "source_file", "note_text"])
    
    exists = False
    for row in ws_text.iter_rows(min_row=2, values_only=True):
        if row[0] == NOTE_ID:
            exists = True
            break
    
    if not exists:
        ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 3. Analyze Procedure Flags
    flags = {k: 0 for k in PROC_FLAGS}
    text_lower = NOTE_TEXT.lower()

    # Flag logic
    if "therapeutic aspiration" in text_lower:
        flags["therapeutic_aspiration"] = 1
    if "bronchoscopy" in text_lower:
        flags["diagnostic_bronchoscopy"] = 1
    if "cryoprobe" in text_lower or "cryotherapy" in text_lower:
        flags["cryotherapy"] = 1
    
    # 4. Populate Note_Index
    ws_index = wb["Note_Index"]
    headers_index = ["source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes"] + PROC_FLAGS
    if ws_index.max_row == 1 and ws_index.cell(1,1).value is None:
        ws_index.append(headers_index)
    
    row_data = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Not Started", ""]
    row_data.extend([flags[f] for f in PROC_FLAGS])
    ws_index.append(row_data)

    # 5. Generate Spans (Anchor-First)
    spans = []
    
    def add_span(text_match, label, norm_val, field, event_id, context=None, match_idx=0):
        if not text_match: return
        spans.append({
            "span_text": text_match,
            "label": label,
            "normalized_value": norm_val,
            "schema_field": field,
            "event_id": event_id,
            "context_prefix": context,
            "match_index": match_idx
        })

    # Event 1: General Inspection
    ev1 = "ev1"
    add_span("Disposable Bronchoscope", "DEV_INSTRUMENT", "Disposable Bronchoscope", "instrument", ev1)
    add_span("tracheostomy tube", "ANAT_AIRWAY", "Tracheostomy", "anatomy_entry", ev1)
    add_span("Trachea", "ANAT_AIRWAY", "Trachea", "anatomy_inspected", ev1, context="Vocal Cords: Not assessed due")
    add_span("Main Carina", "ANAT_AIRWAY", "Carina", "anatomy_inspected", ev1)
    add_span("Right Lung Proximal Airways", "ANAT_AIRWAY", "Right Lung", "anatomy_inspected", ev1)
    add_span("Left Lung Proximal Airways", "ANAT_AIRWAY", "Left Lung", "anatomy_inspected", ev1)
    add_span("Uniblocker", "DEV_INSTRUMENT", "Bronchial Blocker", "device", ev1)
    add_span("LMSB", "ANAT_AIRWAY", "LMSB", "target_location", ev1)

    # Event 2: Therapeutic Aspiration
    ev2 = "ev2"
    add_span("therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", "method", ev2, context="Successful")
    add_span("Trachea (Distal 1/3)", "ANAT_AIRWAY", "Trachea Distal", "target_location", ev2)
    add_span("Right Mainstem", "ANAT_AIRWAY", "RMS", "target_location", ev2)
    add_span("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "target_location", ev2)
    add_span("Left Mainstem", "ANAT_AIRWAY", "LMS", "target_location", ev2)
    add_span("Carina", "ANAT_AIRWAY", "Carina", "target_location", ev2, context="Left Mainstem, ")
    add_span("RUL Carina (RC1)", "ANAT_AIRWAY", "RUL Carina", "target_location", ev2)
    add_span("RML Carina (RC2)", "ANAT_AIRWAY", "RML Carina", "target_location", ev2)
    add_span("LUL Lingula Carina (Lc1)", "ANAT_AIRWAY", "Lingula Carina", "target_location", ev2, context="RML Carina (RC2), ")
    add_span("Left Carina (LC2)", "ANAT_AIRWAY", "Left Carina LC2", "target_location", ev2, context="Lingula Carina (Lc1), and ")
    add_span("mucus", "OBS_LESION", "Mucus", "finding", ev2)
    add_span("blood", "OBS_LESION", "Blood", "finding", ev2, context="mucus, ")
    add_span("blood clots", "OBS_LESION", "Blood Clot", "finding", ev2)

    # Event 3: Amphotericin Instillation
    ev3 = "ev3"
    add_span("Amphotericin", "PROC_METHOD", "Drug Instillation", "method", ev3)
    add_span("50mg", "MEAS_COUNT", "50mg", "dosage", ev3)
    add_span("20cc", "MEAS_VOL", "20cc", "volume", ev3)
    add_span("LUL", "ANAT_AIRWAY", "LUL", "target_location", ev3, context="instilled into the ")
    add_span("10cc", "MEAS_VOL", "10cc", "volume", ev3, context="subsegment and ")
    add_span("apico-posterior (LB1/2) subsegment", "ANAT_AIRWAY", "LB1+LB2", "target_location", ev3)
    add_span("10cc", "MEAS_VOL", "10cc", "volume", ev3, context="anterior (LB3)")
    add_span("anterior (LB3) subsegment", "ANAT_AIRWAY", "LB3", "target_location", ev3)

    # Event 4: TXA
    ev4 = "ev4"
    add_span("TXA", "PROC_METHOD", "Drug Instillation", "method", ev4)
    add_span("1000mg", "MEAS_COUNT", "1000mg", "dosage", ev4)
    add_span("LUL", "ANAT_AIRWAY", "LUL", "target_location", ev4, context="directly into the ")
    
    # Event 5: Cryotherapy (Extraction)
    ev5 = "ev5"
    add_span("extracted with cryoprobe", "PROC_METHOD", "Cryotherapy", "method", ev5)
    add_span("Formalized clot", "OBS_LESION", "Blood Clot", "target_lesion", ev5)
    add_span("LUL Lingula Carina (Lc1)", "ANAT_AIRWAY", "Lingula Carina", "target_location", ev5, context="Endobronchial clot at ")
    add_span("Left Carina (LC2)", "ANAT_AIRWAY", "Left Carina LC2", "target_location", ev5, context="and ")
    add_span("1.1mm Cryoprobe", "DEV_INSTRUMENT", "1.1mm Cryoprobe", "device", ev5)
    add_span("5-10 second freezes", "PROC_ACTION", "5-10s freeze", "duration", ev5)
    add_span("Excellent clot removal", "OBS_ROSE", "Success", "outcome_text", ev5)

    # Outcomes
    add_span("no immediate complications", "OUTCOME_COMPLICATION", "None", "outcome_complication", "ev_outcome", context="Patient tolerated")

    # 6. Hydrate Spans
    ws_anno = wb["Span_Annotations"]
    ws_hydra = wb["Span_Hydrated"]
    
    headers_span = [
        "source_file", "note_id", "span_id", "section_type",
        "context_prefix", "span_text", "match_index",
        "start_char", "end_char", "span_len",
        "label", "normalized_value", "schema_field", "event_id",
        "is_negated", "is_historical", "time_anchor", "reviewer", "comments",
        "hydration_status"
    ]
    
    if ws_anno.max_row == 1:
        ws_anno.append(headers_span)
        ws_hydra.append(headers_span)
    
    for idx, s in enumerate(spans):
        span_id = f"{NOTE_ID}_s{idx+1:03d}"
        txt = s["span_text"]
        start, end = -1, -1
        status = "needs_hydration"
        
        # Hydration
        if NOTE_TEXT.count(txt) == 1:
            start = NOTE_TEXT.find(txt)
            status = "hydrated_unique"
        elif s["context_prefix"]:
            occurrences = [m.start() for m in re.finditer(re.escape(txt), NOTE_TEXT)]
            best_occ = -1
            for occ in occurrences:
                window_start = max(0, occ - 120)
                window_text = NOTE_TEXT[window_start:occ]
                if s["context_prefix"] in window_text:
                    best_occ = occ
                    break
            if best_occ != -1:
                start = best_occ
                status = "hydrated_context"
            else:
                status = "ambiguous_context_fail"
        elif s.get("match_index") is not None:
             occurrences = [m.start() for m in re.finditer(re.escape(txt), NOTE_TEXT)]
             m_idx = s["match_index"]
             if m_idx < len(occurrences):
                 start = occurrences[m_idx]
                 status = "hydrated_index"
             else:
                 status = "ambiguous_index_out_of_bounds"
        else:
            status = "ambiguous_no_context"
            
        if start != -1:
            end = start + len(txt)
            
        row_anno = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure",
            s["context_prefix"], txt, s.get("match_index", ""),
            "", "", f"=LEN(F{ws_anno.max_row+1})",
            s["label"], s["normalized_value"], s["schema_field"], s["event_id"],
            "FALSE", "FALSE", "", "", "", "needs_hydration"
        ]
        ws_anno.append(row_anno)
        
        row_hydra = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure",
            s["context_prefix"], txt, s.get("match_index", ""),
            start if start != -1 else "", end if end != -1 else "", len(txt),
            s["label"], s["normalized_value"], s["schema_field"], s["event_id"],
            "FALSE", "FALSE", "", "", "", status
        ]
        ws_hydra.append(row_hydra)

    # 7. Event Log
    ws_event = wb["Event_Log"]
    headers_event = [
        "source_file", "note_id", "event_id", "event_type", "method",
        "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements",
        "specimens", "findings", "is_historical", "reviewer", "comments",
        "device_size", "device_material",
        "outcome_airway_lumen_pre", "outcome_airway_lumen_post",
        "outcome_symptoms", "outcome_pleural", "outcome_complication"
    ]
    if ws_event.max_row == 1:
        ws_event.append(headers_event)
        
    events_summary = {
        "ev1": ["Bronchoscopy", "Diagnostic", "Trachea/Carina/Lungs", "Disposable Bronchoscope", "", "", "", "", "Normal mucosa/No mass", "FALSE", "", "", "", "", "", "", "", "", "None"],
        "ev2": ["Bronchoscopy", "Therapeutic Aspiration", "Trachea/RMS/LMS/Carinas", "", "", "", "", "", "Mucus/Blood/Clots", "FALSE", "", "", "", "", "", "", "", "", "None"],
        "ev3": ["Bronchoscopy", "Drug Instillation", "LUL (LB1-3)", "", "", "", "Amphotericin 50mg", "", "Mild ooze", "FALSE", "", "", "", "", "", "", "", "", "None"],
        "ev4": ["Bronchoscopy", "Drug Instillation", "LUL", "", "", "", "TXA 1000mg", "", "", "FALSE", "", "", "", "", "", "", "", "", "None"],
        "ev5": ["Bronchoscopy", "Cryotherapy", "LUL Lingula/Left Carina", "1.1mm Cryoprobe", "", "", "", "", "Clot removal", "FALSE", "", "", "", "", "", "", "", "", "None"]
    }
    
    for eid, data in events_summary.items():
        row = [SOURCE_FILE, NOTE_ID, eid] + data
        while len(row) < len(headers_event):
            row.append("")
        ws_event.append(row)

    # 8. V3 Procedure Events
    ws_v3 = wb["V3_Procedure_Events"]
    headers_v3 = [
        "note_id", "event_id", "type",
        "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station",
        "lesion.type", "lesion.size_mm",
        "method", "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote",
        "stent.size", "stent.material_or_brand", "catheter.size_fr",
        "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post",
        "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"
    ]
    if ws_v3.max_row == 1:
        ws_v3.append(headers_v3)

    # JSON Registry
    registry_data = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "procedures": [],
        "no_immediate_complications": True
    }
    
    # Event 2
    v3_ev2 = [
        NOTE_ID, "ev2", "Therapeutic Bronchoscopy",
        "Airway", "Multiple", "", "",
        "Secretions", "",
        "Therapeutic Aspiration", "[]", "[]", "[]", '["Mucus", "Blood", "Clots"]', "Successful therapeutic aspiration...",
        "", "", "",
        "", "", "", "", "None"
    ]
    ws_v3.append(v3_ev2)
    registry_data["procedures"].append({"event_id": "ev2", "type": "Therapeutic Aspiration"})

    # Event 3
    v3_ev3 = [
        NOTE_ID, "ev3", "Therapeutic Bronchoscopy",
        "Airway", "LUL", "LB1/2, LB3", "",
        "", "",
        "Instillation", "[]", '{"drug": "Amphotericin", "dose": "50mg"}', "[]", "[]", "Amphotericin 50mg... instilled",
        "", "", "",
        "", "", "", "", "None"
    ]
    ws_v3.append(v3_ev3)
    registry_data["procedures"].append({"event_id": "ev3", "type": "Instillation"})
    
    # Event 5
    v3_ev5 = [
        NOTE_ID, "ev5", "Therapeutic Bronchoscopy",
        "Airway", "LUL", "Lingula", "",
        "Clot", "",
        "Cryotherapy", '["1.1mm Cryoprobe"]', "[]", "[]", '["Excellent clot removal"]', "Formalized clot was then extracted...",
        "", "", "",
        "", "", "", "", "None"
    ]
    ws_v3.append(v3_ev5)
    registry_data["procedures"].append({"event_id": "ev5", "type": "Cryotherapy"})

    # 9. Write JSON
    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 1:
        ws_json.append(["json_output"])
    ws_json.append([json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    create_workbook()