import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_013"
SOURCE_FILE = "note_013.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_013 SOURCE_FILE: note_013.txt INDICATION FOR OPERATION:  [REDACTED]is a 33 year old-year-old female who presents with respiratory failure and aspergilloma.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
 
CONSENT : Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient or surrogate.
The patient or surrogate read and signed the provided consent form / provided consent over the phone.
The consent was witnessed by an assisting medical professional.
 
PREOPERATIVE DIAGNOSIS: J96.90 Respiratory Failure;
Aspergilloma
 
POSTOPERATIVE DIAGNOSIS:  J96.90 Respiratory Failure; Aspergilloma
 
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31573 Therapeutic injection(s) [eg, chemotherapy denervation agent or corticosteroid, injected percutaneous, transoral, or via endoscope channel]
 
ANESTHESIA: 
99152 Moderate sedation: initial 15 minutes
99153 Moderate sedation: each additional 15 minutes 
 
Procedure performed under moderate sedation.
The following medications were provided:
Etomidate                    20 mg
Rocuronium                 50 mg
Dexmedetomidine gtt1 mcg/kg/hr
Propofol gtt                  40 mcg/kg/min
Fentanyl gtt                 150 mcg/hr
 
Physician/patient face-to-face anesthesia start time:   1439
 
Physician/patient face-to-face anesthesia stop time:   
1602
 
Total moderate sedation time was 83 minutes.  
 
Patient was monitored continuously one-to-one throughout the entire procedure by the attending physician while anesthesia was administered.
Sedation was administered by ICU RN. 
 
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Disposable Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: Supine
 
Initial Airway Inspection Findings:
 
The tracheostomy tube is in good position.
Pharynx: Not assessed due to bronchoscopy introduction through tracheostomy tube.
Larynx: Not assessed due to bronchoscopy introduction through tracheostomy tube.
Vocal Cords: Not assessed due to bronchoscopy introduction through tracheostomy tube.
Trachea: Distal 1/3 normal.
Main Carina: Sharp
Right Lung Proximal Airways: Normal anatomic branching to segmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Left Lung Proximal Airways: Normal anatomic branching to segmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology. Uniblocker in place (deflated) at LMSB
Mucosa: Normal.
Secretions: Moderate bloody secretions with left greater than right
 
Successful therapeutic aspiration was performed to clean out the Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus, blood, and blood clots.
Amphotericin 50mg in 20cc sterile water was instilled into the LUL with 10cc into the apico-posterior (LB1/2) subsegment and 10cc into the anterior (LB3) subsegment
 
The patient tolerated the procedure well.
There were no immediate complications.  At the conclusion of the operation, the patient was in stable condition.
SPECIMEN(S): 
--None
 
IMPRESSION/PLAN: [REDACTED]is a 33 year old-year-old female who presents for bronchoscopy for instillation of Amphotericin.
Patient tolerated the procedure well and there were no immediate complications.
--Continued care per primary team
--Repeat instillation of Amphotericin on [REDACTED]"""

# -------------------------------------------------------------------------
# REGISTRY SCHEMA & FLAGS
# -------------------------------------------------------------------------
PROCEDURE_FLAGS = [
    # Bronchoscopy (23)
    "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
    "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy",
    "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration",
    "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
    "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation",
    "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
    # Pleural (7)
    "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis",
    "pleural_biopsy", "fibrinolytic_therapy"
]

# Set active flags for this note
ACTIVE_FLAGS = {
    "therapeutic_aspiration": 1
    # Note: "Therapeutic injection" is performed but no direct flag exists for "instillation" 
    # other than generic buckets. We stick to the strict list.
}

# -------------------------------------------------------------------------
# SPAN DEFINITIONS (ANCHOR-FIRST)
# -------------------------------------------------------------------------
# Helper to define a span easily
def mk_span(text, label, norm=None, context=None, match_idx=0, section="PROCEDURE", event_id=None):
    return {
        "span_text": text,
        "label": label,
        "normalized_value": norm if norm else text,
        "context_prefix": context,
        "match_index": match_idx,
        "section_type": section,
        "event_id": event_id
    }

# Defining spans based on the note content
SPANS = [
    # --- Metadata / Diagnosis ---
    mk_span("Respiratory Failure", "OBS_LESION", "Respiratory Failure", section="PREOPERATIVE DIAGNOSIS"),
    mk_span("Aspergilloma", "OBS_LESION", "Aspergilloma", section="PREOPERATIVE DIAGNOSIS"),
    
    # --- Procedure Header ---
    mk_span("Therapeutic aspiration initial episode", "PROC_METHOD", "Therapeutic aspiration", section="PROCEDURE"),
    mk_span("Therapeutic injection(s)", "PROC_METHOD", "Therapeutic injection", section="PROCEDURE"),
    
    # --- Instruments ---
    mk_span("Disposable Bronchoscope", "DEV_INSTRUMENT", "Disposable Bronchoscope", section="INSTRUMENT"),
    
    # --- Airway Inspection (Event 1) ---
    mk_span("tracheostomy tube", "DEV_CATHETER", "Tracheostomy Tube", section="Initial Airway Inspection Findings", event_id="evt_01"),
    mk_span("Trachea: Distal 1/3 normal", "OBS_ROSE", "Normal", section="Initial Airway Inspection Findings", event_id="evt_01"),
    mk_span("Main Carina: Sharp", "OBS_ROSE", "Normal", section="Initial Airway Inspection Findings", event_id="evt_01"),
    mk_span("Right Lung Proximal Airways: Normal", "OBS_ROSE", "Normal", section="Initial Airway Inspection Findings", event_id="evt_01"),
    mk_span("Left Lung Proximal Airways: Normal", "OBS_ROSE", "Normal", section="Initial Airway Inspection Findings", event_id="evt_01"),
    mk_span("No evidence of mass", "OBS_LESION", "No mass", context="", section="Initial Airway Inspection Findings", event_id="evt_01"),
    mk_span("Uniblocker", "DEV_CATHETER", "Uniblocker", section="Initial Airway Inspection Findings", event_id="evt_01"),
    mk_span("LMSB", "ANAT_AIRWAY", "Left Mainstem Bronchus", section="Initial Airway Inspection Findings", event_id="evt_01"),
    
    # --- Therapeutic Aspiration (Event 2) ---
    mk_span("Successful therapeutic aspiration", "PROC_METHOD", "Therapeutic aspiration", section="PROCEDURE IN DETAIL", event_id="evt_02"),
    mk_span("clean out", "PROC_ACTION", "Aspiration", section="PROCEDURE IN DETAIL", event_id="evt_02"),
    mk_span("Trachea (Distal 1/3)", "ANAT_AIRWAY", "Trachea Distal 1/3", section="PROCEDURE IN DETAIL", event_id="evt_02"),
    mk_span("Right Mainstem", "ANAT_AIRWAY", "Right Mainstem", section="PROCEDURE IN DETAIL", event_id="evt_02"),
    mk_span("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", section="PROCEDURE IN DETAIL", event_id="evt_02"),
    mk_span("Left Mainstem", "ANAT_AIRWAY", "Left Mainstem", section="PROCEDURE IN DETAIL", event_id="evt_02"),
    mk_span("Carina", "ANAT_AIRWAY", "Carina", context="Left Mainstem, ", section="PROCEDURE IN DETAIL", event_id="evt_02"),
    mk_span("RUL Carina (RC1)", "ANAT_AIRWAY", "RUL Carina", section="PROCEDURE IN DETAIL", event_id="evt_02"),
    mk_span("RML Carina (RC2)", "ANAT_AIRWAY", "RML Carina", section="PROCEDURE IN DETAIL", event_id="evt_02"),
    mk_span("LUL Lingula Carina (Lc1)", "ANAT_AIRWAY", "LUL Lingula Carina", section="PROCEDURE IN DETAIL", event_id="evt_02"),
    mk_span("Left Carina (LC2)", "ANAT_AIRWAY", "Left Carina", section="PROCEDURE IN DETAIL", event_id="evt_02"),
    mk_span("mucus, blood, and blood clots", "OBS_LESION", "Secretions/Clots", section="PROCEDURE IN DETAIL", event_id="evt_02"),

    # --- Therapeutic Instillation (Event 3) ---
    mk_span("Amphotericin", "PROC_METHOD", "Instillation Agent", context="", section="PROCEDURE IN DETAIL", event_id="evt_03"),
    mk_span("instilled", "PROC_ACTION", "Instillation", section="PROCEDURE IN DETAIL", event_id="evt_03"),
    mk_span("LUL", "ANAT_AIRWAY", "LUL", context="instilled into the ", section="PROCEDURE IN DETAIL", event_id="evt_03"),
    mk_span("10cc", "MEAS_VOL", "10cc", context="with ", section="PROCEDURE IN DETAIL", event_id="evt_03"),
    mk_span("apico-posterior (LB1/2)", "ANAT_AIRWAY", "LUL Apico-posterior (LB1/2)", section="PROCEDURE IN DETAIL", event_id="evt_03"),
    mk_span("10cc", "MEAS_VOL", "10cc", context="subsegment and ", section="PROCEDURE IN DETAIL", event_id="evt_03"),
    mk_span("anterior (LB3)", "ANAT_AIRWAY", "LUL Anterior (LB3)", section="PROCEDURE IN DETAIL", event_id="evt_03"),

    # --- Outcomes ---
    mk_span("tolerated the procedure well", "OUTCOME_SYMPTOMS", "Tolerated well", section="PROCEDURE IN DETAIL"),
    mk_span("No immediate complications", "OUTCOME_COMPLICATION", "None", section="PROCEDURE IN DETAIL"),
    mk_span("stable condition", "OUTCOME_SYMPTOMS", "Stable", section="PROCEDURE IN DETAIL"),
]

# -------------------------------------------------------------------------
# EVENT DEFINITIONS
# -------------------------------------------------------------------------
EVENTS = [
    {
        "event_id": "evt_01",
        "type": "Diagnostic",
        "method": "Inspection",
        "anatomy_target": "Airways",
        "findings": "Normal/Sharp/No mass",
        "comments": "Initial inspection via tracheostomy"
    },
    {
        "event_id": "evt_02",
        "type": "Therapeutic",
        "method": "Therapeutic Aspiration",
        "anatomy_target": "Trachea, RMS, BI, LMS, Carina, RUL, RML, LUL",
        "findings": "Mucus, blood, clots removed",
        "comments": "Extensive clearing of airways"
    },
    {
        "event_id": "evt_03",
        "type": "Therapeutic",
        "method": "Instillation",
        "anatomy_target": "LUL (LB1/2, LB3)",
        "findings": "Amphotericin instilled",
        "measurements": "20cc total",
        "comments": "Amphotericin 50mg in 20cc"
    }
]

# -------------------------------------------------------------------------
# SCRIPT LOGIC
# -------------------------------------------------------------------------

def create_workbook():
    # 1. Load Template
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        # Fallback to create if template not found (for standalone testing)
        wb = openpyxl.Workbook()
        for s in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            wb.create_sheet(s)
            
    # 2. Populate Note_Text
    ws_text = wb["Note_Text"]
    # Check headers
    if ws_text.max_row == 1 and ws_text.cell(1, 1).value is None:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 3. Populate Note_Index
    ws_index = wb["Note_Index"]
    # Headers logic usually pre-filled in template, but we append row
    # Columns: source_file, note_id, encounter_id, procedure_date, site, reviewer, status, free_text_notes, [flags...]
    
    # Construct Flag Row
    flag_values = [ACTIVE_FLAGS.get(f, 0) for f in PROCEDURE_FLAGS]
    
    row_data = [
        SOURCE_FILE, NOTE_ID, "", "", "", "", "Pending", "Therapeutic aspiration and instillation of Amphotericin"
    ] + flag_values
    
    ws_index.append(row_data)

    # 4. Process Spans & Hydration
    ws_anno = wb["Span_Annotations"]
    ws_hydra = wb["Span_Hydrated"]
    
    # Headers if empty
    if ws_anno.max_row == 1 and ws_anno.cell(1, 1).value is None:
        headers = ["source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", 
                   "match_index", "start_char", "end_char", "span_len", "label", "normalized_value", 
                   "schema_field", "event_id", "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"]
        ws_anno.append(headers)
        ws_hydra.append(headers)

    span_counter = 1
    
    for s in SPANS:
        s_text = s["span_text"]
        
        # Hydration Logic
        start = -1
        end = -1
        status = "ambiguous"
        
        # 1. Exact count check
        count = NOTE_TEXT.count(s_text)
        
        if count == 1:
            start = NOTE_TEXT.find(s_text)
            status = "hydrated_unique"
        elif count > 1:
            # 2. Context Prefix
            if s.get("context_prefix"):
                # Find all starts
                all_starts = [m.start() for m in re.finditer(re.escape(s_text), NOTE_TEXT)]
                found = False
                for c_start in all_starts:
                    # check window before
                    window_start = max(0, c_start - 120)
                    window_text = NOTE_TEXT[window_start:c_start]
                    if s["context_prefix"] in window_text:
                        start = c_start
                        status = "hydrated_prefix_window"
                        found = True
                        break
                if not found:
                    # Fallback to match index if provided
                    m_idx = s.get("match_index", 0)
                    if m_idx < len(all_starts):
                        start = all_starts[m_idx]
                        status = "hydrated_match_index_fallback"
            else:
                # 3. Match Index
                m_idx = s.get("match_index", 0)
                all_starts = [m.start() for m in re.finditer(re.escape(s_text), NOTE_TEXT)]
                if m_idx < len(all_starts):
                    start = all_starts[m_idx]
                    status = "hydrated_match_index"
                else:
                    status = f"ambiguous_count={count}"
        else:
            status = "not_found"

        if start != -1:
            end = start + len(s_text)
            
        span_id = f"{NOTE_ID}_s{span_counter:03d}"
        span_counter += 1
        
        # Write to Span_Annotations (No offsets)
        row_anno = [
            SOURCE_FILE, NOTE_ID, span_id, s.get("section_type", ""), s.get("context_prefix", ""),
            s_text, s.get("match_index", ""), "", "", f"=LEN(F{ws_anno.max_row+1})",
            s["label"], s["normalized_value"], "", s.get("event_id", ""),
            "", "", "", "", "", "needs_hydration"
        ]
        ws_anno.append(row_anno)
        
        # Write to Span_Hydrated (With offsets)
        row_hydra = [
            SOURCE_FILE, NOTE_ID, span_id, s.get("section_type", ""), s.get("context_prefix", ""),
            s_text, s.get("match_index", ""), start if start!=-1 else "", end if end!=-1 else "", len(s_text),
            s["label"], s["normalized_value"], "", s.get("event_id", ""),
            "", "", "", "", "", status
        ]
        ws_hydra.append(row_hydra)

    # 5. Populate Event_Log
    ws_event = wb["Event_Log"]
    for e in EVENTS:
        # Mapping dict to columns
        # source_file, note_id, event_id, event_type, method, anatomy_target, device, needle_gauge, stations...
        row = [
            SOURCE_FILE, NOTE_ID, e["event_id"], e.get("type"), e.get("method"),
            e.get("anatomy_target"), "", "", "", "", e.get("measurements"),
            "", e.get("findings"), "", "", e.get("comments"),
            "", "", # device size/material
            "", "", "", "", "" # outcomes
        ]
        ws_event.append(row)
        
    # 6. Populate V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    for e in EVENTS:
        row = [
            NOTE_ID, e["event_id"], e.get("type"),
            "Airway", "", "", "", # anatomy breakdown
            "", "", # lesion
            e.get("method"), "", "", "", e.get("findings"), "", # devices, measurements json etc
            "", "", "", # stent/cath
            "", "", "", "", "None" if e["event_id"] == "evt_03" else "" # outcomes (simplification)
        ]
        ws_v3.append(row)

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    json_obj = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": EVENTS
    }
    ws_json.append([json.dumps(json_obj, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    create_workbook()