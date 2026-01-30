import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_014"
SOURCE_FILE = "note_014.txt"
PROCEDURE_DATE = "2026-01-12" # inferred from context or left blank if not in text, using prompt context date
# However, standard practice for these scripts is to leave blank if not explicit in note text.
# The note mentions [REDACTED] dates. I will leave blank.

NOTE_TEXT = """NOTE_ID:  note_014 SOURCE_FILE: note_014.txt INDICATION FOR OPERATION:  [REDACTED]is a 33 year old-year-old female who presents with hemoptysis.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
 
CONSENT : Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient or surrogate.
The patient or surrogate read and signed the provided consent form / provided consent over the phone.
The consent was witnessed by an assisting medical professional.
 
PREOPERATIVE DIAGNOSIS: Hemoptysis, airway obstruction, aspergilloma
 
POSTOPERATIVE DIAGNOSIS:  Hemoptysis, airway obstruction, aspergilloma
 
PROCEDURE:  
31646 Therapeutic aspiration subsequent episodes
31573 Therapeutic injection(s) [eg, chemotherapy denervation agent or corticosteroid, injected percutaneous, transoral, or via endoscope channel]
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
31899NFD BRONCHOSCOPY W/ APPLICATION OF TRANEXAMIC ACID
31634 Balloon occlusion or placement of occlusive substance 
 
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)
 
IP 
[REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required cryotherapy for removal of organized clot throughout the left-sided airways.
This resulted in >100% increased work due to Time, Technical difficulty of procedure, and Physical and mental effort required.
Apply to: 31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy).
ANESTHESIA: 
99152 Moderate sedation: initial 15 minutes
99153 Moderate sedation: each additional 15 minutes 
 
Procedure performed under moderate sedation.
The following medications were provided:
Versed gtt                    2 mg/hr
Versed boluses            10 mg total
Fentanyl gtt                 150 mcg/hr
Fentanyl boluses         125 mcg total
Propofol                       60 mg/kg/min
Dexmedatomidine    
   1 mcg/kg/hr
Cisatracurium              7 mg/hr
 
 
Physician/patient face-to-face anesthesia start time:   1156
 
Physician/patient face-to-face anesthesia stop time:   1400
 
Total moderate sedation time was 124 minutes.
Patient was monitored continuously one-to-one throughout the entire procedure by the attending physician while anesthesia was administered.
Sedation was administered by ICU RN. 
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Disposable Bronchoscope
 
ESTIMATED BLOOD LOSS:   Minimum
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: Supine
 
Initial Airway Inspection Findings:
The tracheostomy tube is in good position.
Pharynx: Not assessed due to bronchoscopy introduction through tracheostomy tube.
Larynx: Not assessed due to bronchoscopy introduction through tracheostomy tube.
Vocal Cords: Not assessed due to bronchoscopy introduction through tracheostomy tube.
Trachea: Distal 1/3 normal.
Main Carina: Sharp
Right Lung Proximal Airways: Normal anatomic branching to segmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Left Lung Proximal Airways: Endobronchial balloon at ligula, LUL and LB6 with organized clot obscuring evaluation of distal airways.
Otherwise, normal anatomic branching to segmental level.  No evidence of mass, lesions, bleeding or other endobronchial pathology.
Mucosa: Normal.
Secretions: Minimal, thin, and clear.  
 
Successful therapeutic aspiration was performed to clean out the Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus, blood, and blood clots.
Endobronchial clot at LUL Lingula Carina (Lc1) and Left Carina (LC2) and LB6 was treated with the following modalities:
 
Modality	Tools	Setting/Mode	Duration	Results
Cryoprobe	1.1mm Cryoprobe	N/A	5-10 second freezes	Excellent clot removal
 
After organized clot was removed from the LUL, Amphotericin 10mg in 10cc sterile water was instilled into the LUL proper.
The bronchoscope was wedged into the LUL to prevent backflow for 5 minutes.
TXA 400mg was applied directly into the LUL through the bronchoscope and balloon occlusion was performed at the Left Carina (LC2) with 5 Fr Uniblocker and secured in place with the Uniblocker ventilator adaptor.
-Placed at 27.5cm at dark blue securement device
            -Confirmed that 2.5cc of air was required to inflate the balloon to fully occlude the 
LUL
The endobronchial blocker balloon was left inflated in the LUL.
CLOT REMOVED FROM LUL
 
 
UNIBLOCKER IN PLACE AT LUL AT CONCLUSION OF PROCEDURE
 
 
The patient tolerated the procedure well.
There were no immediate complications.  At the conclusion of the operation, the patient was in stable condition.
SPECIMEN(S): 
--None
 
IMPRESSION/PLAN: [REDACTED]is a 33 year old-year-old female who presents for bronchoscopy for evaluation of hemoptysis and instillation of antifungal medication (Amphotericin).
A combination of cold saline and cryotherapy (cryoprobe) were used to evacuate the LUL and LB6 airways of organized clot.
Amphotericin was instilled into the LUL and a 5 Fr Uniblocker was placed and left inflated at the LUL.
The patient tolerated the procedure well and there were no immediate complications.
--Post procedure CXR
--Uniblocker placed at LUL orifice, inflated fully with 2.5cc of air, blocker distance at 27.5cm
--Please leave blocker in place for 2 hours post procedure (deflate at 1400 on [REDACTED])
--Plan for repeat bronchoscopy and likely repeat instillation of Amphotericin on [REDACTED]"""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# -------------------------------------------------------------------------
# CONFIGURATION & SCHEMA
# -------------------------------------------------------------------------
# 30 procedure flags
PROCEDURE_FLAGS = [
    "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
    "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy", "transbronchial_biopsy",
    "transbronchial_cryobiopsy", "therapeutic_aspiration", "foreign_body_removal", "airway_dilation",
    "airway_stent", "thermal_ablation", "tumor_debulking_non_thermal", "cryotherapy", "blvr",
    "peripheral_ablation", "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
    "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis", "pleural_biopsy",
    "fibrinolytic_therapy"
]

# Set active flags for this note
ACTIVE_FLAGS = {
    "therapeutic_aspiration": 1,
    "cryotherapy": 1,
    # "balloon_occlusion" is not a standard flag in this schema, so we stick to the closest valid flags
    # which are active.
}

# -------------------------------------------------------------------------
# EXTRACTION DATA (Spans & Events)
# -------------------------------------------------------------------------

# SPANS
# Fields: span_text, label, normalized_value, context_prefix, section_type, event_id
raw_spans = [
    # Event 1: Therapeutic Aspiration
    {
        "span_text": "Therapeutic aspiration",
        "label": "PROC_METHOD",
        "normalized_value": "therapeutic_aspiration",
        "section_type": "PROCEDURE",
        "event_id": "evt1"
    },
    {
        "span_text": "successful therapeutic aspiration",
        "label": "PROC_METHOD",
        "normalized_value": "therapeutic_aspiration",
        "context_prefix": "Secretions: Minimal, thin, and clear.  \n \n",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt1"
    },
    {
        "span_text": "Trachea (Distal 1/3)",
        "label": "ANAT_AIRWAY",
        "normalized_value": "trachea",
        "context_prefix": "clean out the ",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt1"
    },
    {
        "span_text": "Right Mainstem",
        "label": "ANAT_AIRWAY",
        "normalized_value": "rms",
        "context_prefix": "Trachea (Distal 1/3), ",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt1"
    },
    {
        "span_text": "Bronchus Intermedius",
        "label": "ANAT_AIRWAY",
        "normalized_value": "bi",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt1"
    },
    {
        "span_text": "Left Mainstem",
        "label": "ANAT_AIRWAY",
        "normalized_value": "lms",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt1"
    },
    {
        "span_text": "Carina",
        "label": "ANAT_AIRWAY",
        "normalized_value": "carina",
        "context_prefix": "Left Mainstem, ",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt1"
    },
    {
        "span_text": "RUL Carina (RC1)",
        "label": "ANAT_AIRWAY",
        "normalized_value": "rul_carina",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt1"
    },
    {
        "span_text": "RML Carina (RC2)",
        "label": "ANAT_AIRWAY",
        "normalized_value": "rml_carina",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt1"
    },
    {
        "span_text": "LUL Lingula Carina (Lc1)",
        "label": "ANAT_AIRWAY",
        "normalized_value": "lul_carina",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt1"
    },
    {
        "span_text": "Left Carina (LC2)",
        "label": "ANAT_AIRWAY",
        "normalized_value": "lul_carina_2",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt1"
    },
    {
        "span_text": "mucus, blood, and blood clots",
        "label": "OBS_LESION",
        "normalized_value": "secretions",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt1"
    },

    # Event 2: Cryotherapy
    {
        "span_text": "Endobronchial clot",
        "label": "OBS_LESION",
        "normalized_value": "clot",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt2"
    },
    {
        "span_text": "LUL Lingula Carina (Lc1)",
        "label": "ANAT_AIRWAY",
        "normalized_value": "lul_carina",
        "context_prefix": "Endobronchial clot at ",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt2"
    },
    {
        "span_text": "Left Carina (LC2)",
        "label": "ANAT_AIRWAY",
        "normalized_value": "lul_carina_2",
        "context_prefix": "LUL Lingula Carina (Lc1) and ",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt2"
    },
    {
        "span_text": "LB6",
        "label": "ANAT_AIRWAY",
        "normalized_value": "lb6",
        "context_prefix": "and Left Carina (LC2) and ",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt2"
    },
    {
        "span_text": "Cryoprobe",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "cryoprobe",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt2"
    },
    {
        "span_text": "1.1mm",
        "label": "MEAS_SIZE",
        "normalized_value": "1.1",
        "context_prefix": "Cryoprobe\t",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt2"
    },
    {
        "span_text": "Cryotherapy",
        "label": "PROC_METHOD",
        "normalized_value": "cryotherapy",
        "context_prefix": "This patient required ",
        "section_type": "CODE MOD DETAILS",
        "event_id": "evt2"
    },
    {
        "span_text": "Excellent clot removal",
        "label": "PROC_ACTION",
        "normalized_value": "clot_removal",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt2"
    },

    # Event 3: Amphotericin
    {
        "span_text": "Amphotericin",
        "label": "PROC_METHOD",
        "normalized_value": "instillation_amphotericin",
        "context_prefix": "LUL, ",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt3"
    },
    {
        "span_text": "10cc",
        "label": "MEAS_VOL",
        "normalized_value": "10",
        "context_prefix": "Amphotericin 10mg in ",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt3"
    },
    {
        "span_text": "LUL proper",
        "label": "ANAT_AIRWAY",
        "normalized_value": "lul",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt3"
    },

    # Event 4: TXA
    {
        "span_text": "TXA",
        "label": "PROC_METHOD",
        "normalized_value": "instillation_txa",
        "context_prefix": "minutes.\n",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt4"
    },
    {
        "span_text": "LUL",
        "label": "ANAT_AIRWAY",
        "normalized_value": "lul",
        "context_prefix": "directly into the ",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt4"
    },

    # Event 5: Uniblocker
    {
        "span_text": "balloon occlusion",
        "label": "PROC_METHOD",
        "normalized_value": "balloon_occlusion",
        "context_prefix": "bronchoscope and ",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt5"
    },
    {
        "span_text": "Left Carina (LC2)",
        "label": "ANAT_AIRWAY",
        "normalized_value": "lul_carina_2",
        "context_prefix": "occlusion was performed at the ",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt5"
    },
    {
        "span_text": "5 Fr",
        "label": "DEV_CATHETER_SIZE",
        "normalized_value": "5fr",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt5"
    },
    {
        "span_text": "Uniblocker",
        "label": "DEV_CATHETER",
        "normalized_value": "uniblocker",
        "context_prefix": "5 Fr ",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt5"
    },
    {
        "span_text": "2.5cc",
        "label": "MEAS_VOL",
        "normalized_value": "2.5",
        "context_prefix": "Confirmed that ",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt5"
    },
    
    # Outcomes
    {
        "span_text": "no immediate complications",
        "label": "OUTCOME_COMPLICATION",
        "normalized_value": "none",
        "section_type": "PROCEDURE IN DETAIL",
        "event_id": "evt_outcome"
    },
]

# EVENTS
# event_id mapped to V3 structure
v3_events = [
    {
        "event_id": "evt1",
        "type": "therapeutic_aspiration",
        "method": "suction",
        "target_stations": ["Trachea", "RMS", "LMS", "BI", "Carina", "RUL", "RML", "LUL", "Lingula"],
        "findings_json": ["mucus", "blood", "clots"],
        "evidence_quote": "Successful therapeutic aspiration was performed to clean out... from mucus, blood, and blood clots."
    },
    {
        "event_id": "evt2",
        "type": "cryotherapy",
        "method": "cryoprobe",
        "target_stations": ["LUL", "Lingula", "LB6"],
        "lesion_type": "clot",
        "devices_json": {"name": "Cryoprobe", "size": "1.1mm"},
        "evidence_quote": "Endobronchial clot at LUL Lingula Carina (Lc1) and Left Carina (LC2) and LB6 was treated with... Cryoprobe... Excellent clot removal"
    },
    {
        "event_id": "evt3",
        "type": "therapeutic_instillation",
        "method": "instillation",
        "target_stations": ["LUL"],
        "devices_json": {"drug": "Amphotericin", "dose": "10mg", "vol": "10cc"},
        "evidence_quote": "Amphotericin 10mg in 10cc sterile water was instilled into the LUL proper."
    },
    {
        "event_id": "evt4",
        "type": "therapeutic_instillation",
        "method": "instillation",
        "target_stations": ["LUL"],
        "devices_json": {"drug": "TXA", "dose": "400mg"},
        "evidence_quote": "TXA 400mg was applied directly into the LUL"
    },
    {
        "event_id": "evt5",
        "type": "balloon_occlusion",
        "method": "blocker",
        "target_stations": ["LUL", "LC2"],
        "devices_json": {"name": "Uniblocker", "size": "5 Fr", "fill": "2.5cc air"},
        "evidence_quote": "balloon occlusion was performed at the Left Carina (LC2) with 5 Fr Uniblocker... inflated to fully occlude the LUL"
    }
]

# -------------------------------------------------------------------------
# HELPER FUNCTIONS
# -------------------------------------------------------------------------

def hydrate_spans(text, spans):
    """
    Computes start_char, end_char based on span_text and context/indices.
    """
    hydrated = []
    
    # 1. First pass: count total occurrences of each span_text in text
    occurrence_map = {}
    for s in spans:
        st = s["span_text"]
        if st not in occurrence_map:
            occurrence_map[st] = [m.start() for m in re.finditer(re.escape(st), text)]
    
    # 2. Assign offsets
    for s in spans:
        st = s["span_text"]
        all_starts = occurrence_map.get(st, [])
        
        start = None
        status = "ambiguous"
        
        # Strategy A: Unique
        if len(all_starts) == 1:
            start = all_starts[0]
            status = "hydrated_unique"
        
        # Strategy B: Context Prefix
        elif len(all_starts) > 1 and s.get("context_prefix"):
            prefix = s["context_prefix"]
            # look for prefix in window before each occurrence
            valid_starts = []
            for candidate in all_starts:
                # check 120 chars before
                window_start = max(0, candidate - 120)
                window = text[window_start:candidate]
                if prefix in window:
                    valid_starts.append(candidate)
            
            if len(valid_starts) == 1:
                start = valid_starts[0]
                status = "hydrated_prefix_window"
            elif len(valid_starts) > 1:
                # If multiple match prefix (rare), take first or handle via match_index if we had it
                start = valid_starts[0] 
                status = "hydrated_prefix_ambiguous"
        
        # Strategy C: First available (fallback if no context)
        elif len(all_starts) > 0:
            start = all_starts[0]
            status = "hydrated_fallback_first"
        
        if start is not None:
            end = start + len(st)
            s_out = s.copy()
            s_out["start_char"] = start
            s_out["end_char"] = end
            s_out["hydration_status"] = status
            s_out["span_len"] = len(st)
            hydrated.append(s_out)
        else:
            # Append even if not found, with empty offsets
            s_out = s.copy()
            s_out["hydration_status"] = "missing_in_text"
            hydrated.append(s_out)
            
    return hydrated

def create_workbook():
    # Load template
    if not os.path.exists(TEMPLATE_PATH):
        # Fallback: create a blank workbook if template doesn't exist (for safety in standalone run)
        wb = openpyxl.Workbook()
        wb.create_sheet("Note_Text")
        wb.create_sheet("Note_Index")
        wb.create_sheet("Span_Annotations")
        wb.create_sheet("Span_Hydrated")
        wb.create_sheet("Event_Log")
        wb.create_sheet("V3_Procedure_Events")
        wb.create_sheet("V3_Registry_JSON")
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # 1. Note_Text
    ws_text = wb["Note_Text"]
    if ws_text.max_row == 1:
        # Headers should be there, if not, assume template has them. 
        # Append row
        ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    else:
        # Check if exists
        found = False
        for row in ws_text.iter_rows(min_row=2, values_only=True):
            if row[0] == NOTE_ID:
                found = True
                break
        if not found:
            ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws_index = wb["Note_Index"]
    # Columns: source_file, note_id, encounter_id, procedure_date, site, reviewer, status, free_text_notes, [flags...]
    # Create row
    row_data = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Done", ""]
    # Append flags
    for flag in PROCEDURE_FLAGS:
        row_data.append(ACTIVE_FLAGS.get(flag, 0))
    ws_index.append(row_data)

    # 3. Span_Annotations (Anchor First)
    ws_span = wb["Span_Annotations"]
    # Columns: source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, start, end, len, label, norm, schema, event_id, neg, hist, time, rev, comm, status
    
    # We need to hydrate first to get the correct valid spans for the next sheet, 
    # but strictly for Span_Annotations we leave start/end blank.
    
    # However, to be helpful, I will prepare the list and just write blanks.
    
    span_id_counter = 1
    
    final_spans_hydrated = hydrate_spans(NOTE_TEXT, raw_spans)
    
    for s in final_spans_hydrated:
        s_id = f"{NOTE_ID}_span_{span_id_counter:03d}"
        span_id_counter += 1
        
        row = [
            SOURCE_FILE, NOTE_ID, s_id, s.get("section_type", ""),
            s.get("context_prefix", ""), s["span_text"], "", # match_index left blank
            "", "", f'=LEN(F{ws_span.max_row + 1})', # start/end blank, len formula
            s["label"], s.get("normalized_value", ""), "", # schema_field blank
            s.get("event_id", ""), 0, 0, "", "", "", "needs_hydration"
        ]
        ws_span.append(row)

    # 4. Span_Hydrated
    ws_hyd = wb["Span_Hydrated"]
    # Same structure but with offsets filled
    span_id_counter = 1
    for s in final_spans_hydrated:
        s_id = f"{NOTE_ID}_span_{span_id_counter:03d}"
        span_id_counter += 1
        
        row = [
            SOURCE_FILE, NOTE_ID, s_id, s.get("section_type", ""),
            s.get("context_prefix", ""), s["span_text"], "",
            s.get("start_char", ""), s.get("end_char", ""), s.get("span_len", ""),
            s["label"], s.get("normalized_value", ""), "",
            s.get("event_id", ""), 0, 0, "", "", "", s.get("hydration_status", "error")
        ]
        ws_hyd.append(row)

    # 5. Event_Log
    ws_event = wb["Event_Log"]
    # Columns: source_file, note_id, event_id, event_type, method, anatomy_target, device, needle_gauge, stations, counts, measurements, specimens, findings, historical, reviewer, comments, dev_size, dev_mat, out_air_pre, out_air_post, out_symp, out_pleural, out_comp
    
    for ev in v3_events:
        # Flatten simple fields
        stations_str = ", ".join(ev.get("target_stations", []))
        dev_json = ev.get("devices_json", {})
        dev_str = dev_json.get("name", "")
        dev_size = dev_json.get("size", "")
        
        row = [
            SOURCE_FILE, NOTE_ID, ev["event_id"], ev["type"], ev["method"],
            "", dev_str, "", stations_str, "", "", "", json.dumps(ev.get("findings_json", [])), 0, "", "",
            dev_size, "", "", "", "", "", "none" if ev.get("event_id") == "evt_outcome" else ""
        ]
        ws_event.append(row)

    # 6. V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    # note_id, event_id, type, target.anat_type, lobe, segment, station, lesion.type, size, method, dev_json, meas_json, spec_json, find_json, quote...
    
    for ev in v3_events:
        row = [
            NOTE_ID, ev["event_id"], ev["type"], 
            "airway", "", "", ", ".join(ev.get("target_stations", [])), # simple flattening
            ev.get("lesion_type", ""), "", 
            ev["method"], json.dumps(ev.get("devices_json", {})), "", "", 
            json.dumps(ev.get("findings_json", [])), ev.get("evidence_quote", ""),
            "", "", "", "", "", "", "", ""
        ]
        ws_v3.append(row)

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    # Structure: schema_version, note_id, full_json
    
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": v3_events,
        "no_immediate_complications": True
    }
    
    ws_json.append(["3.0", NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    create_workbook()