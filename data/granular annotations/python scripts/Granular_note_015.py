import openpyxl
from openpyxl.utils import get_column_letter
import datetime
import re
import json

# =============================================================================
# INPUT DATA
# =============================================================================

NOTE_ID = "note_015"
SOURCE_FILE = "note_015.txt"
PROCEDURE_DATE = "2026-01-12" # Placeholder based on context or current date
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_015 SOURCE_FILE: note_015.txt INDICATION FOR OPERATION:  [REDACTED]is a 33 year old-year-old female who presents with massive hemoptysis, endobronchial clot causing acute airway obstruction and worsening of acute respiratory failure.
CONSENT: This procedure was emergent.  Serial consent was previously obtained before the procedure.
The nature, indications, purpose, benefits, risks, potential complications, and alternatives to the procedure were discussed with the patient or surrogate decision-maker in detail.
Patient or surrogate decision-maker previously agreed to proceed with procedure.
The patient or surrogate decision-maker previously read and signed the provided consent form or provided consent over the phone.
If consent was provided over the phone, then the consent was witnessed by an assisting medical professional.
PREOPERATIVE DIAGNOSIS: Massive hemoptysis, acute airway obstruction
POSTOPERATIVE DIAGNOSIS:  Same
 
PROCEDURE:  
31646 Therapeutic aspiration subsequent episodes
31622 Dx bronchoscope/cell washing          
31899NFD BRONCHOSCOPY W/ APPLICATION OF TRANEXAMIC ACID
31634 Balloon occlusion or placement of occlusive substance 
31635 Foreign body removal
 
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)
 
IP [REDACTED] CODE MOD DETAILS: Unusual Procedure (22 MODIFIER):
This patient was exceptionally critically ill due to acute respiratory failure and requiring max ventilator and ECMO support.
Patient required extensive therapeutic aspiration and removal of clot using forceps, which required much more time, mental/physical effort, and intensity.
This resulted in >100% increased work due to Increased intensity, Time, Technical difficulty of procedure, Severity of patient's condition, and Physical and mental effort required.
Apply to: 
31646 Therapeutic aspiration subsequent episodes
 
ANESTHESIA: 
99152 Moderate sedation: initial 15 minutes
99153 Moderate sedation: each additional 15 minutes 
99153 Moderate sedation: each additional 15 minutes 
Procedure performed under moderate sedation.
I took over the management of intraprocedural moderate sedation at 19:15PM.
Prior to this, it was managed by the primary ICU attending, Dr. Omar Mesarwi.
The following medications were provided:
Propofol gtt at 60 mcg/kg/min.
Precedex gtt at 1.3 mcg/kg/hr.
Fentanyl          IV 250 mcg
Dilaudid           IV 1mg
Physician/patient face-to-face anesthesia start time:   19:15 PM
Physician/patient face-to-face anesthesia stop time:   20:00 PM
Total moderate sedation time was 45 minutes.
Patient was monitored continuously one-to-one throughout the entire procedure by the attending physician while anesthesia was administered.
Sedation was administered by Gesicki, Killian Matthew, RN.  
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
Disposable Bronchoscope
 
PROCEDURE IN DETAIL:
This was an emergency procedure due to rapid clinical decompensation with associated massive hemoptysis with endobronchial clot causing acute airway obstruction.
Sedation administered and adjusted.  Tracheostomy tube (ID 7.5mm) already in place.
 
The Flexible Therapeutic Bronchoscope was advanced for airway examination.
Endobronchial topical lidocaine applied to the vocal cords, main carina, right carina 1, and left carina 2.
 
Initial Airway Examination Findings:
Trachea: Clot extending from the LMSB into the distal trachea without obstruction of the RMSB.
Trachea otherwise normal in anatomy.
Main Carina: Sharp
Right Lung Proximal Airways:  Evidence of yellow hard debris - likely pieces of fungal ball - noted in RB1.
Otherwise normal anatomic branching to the first subsegmental level.  No evidence of mass, lesions, bleeding or other endobronchial pathology.
Left Lung Proximal Airways: Large clot completely obstructing the LMSB - extending to the main carina.
LUL and LLL also completely obstructed by clot.  After extraction of clot, the LMSB and LLL were otherwise normal in anatomy without endobronchial lesions.
Incomplete assessment of LUL, as clot was not fully extracted from this area.
Mucosa: Mildly friable.
Secretions: Thin, blood-tinged.
All secretions were suctioned to clear (therapeutic aspiration). 
 
Successful therapeutic aspiration was performed to clean out the trachea, right mainstem bronchus, right upper lobe, bronchus intermedius, right middle lobe, right lower lobe, left mainstem bronchus, left upper lobe, left lower lobe from blood clot and mucus.
Endobronchial obstruction due to debris/clot was mechanically excised and removed using bland alligator forceps.
Foreign body removal of large endobronchial clot burden, which is not native to the airway, was performed.
The was performed using the below modalities:
-Numerous cold saline flushing to dislodge clot and control any distal bleeding.
-Therapeutic aspiration utilizing Neptune suction.
-Bland alligator forceps
 
Able to remove all clot from the LMSB and majority of LLL and LB7/8-10.
This included a large cast of clot that was dislodged and removed from LLL truncus basalis and basilar segments.
Some partially obstructing clot remained in the LLL superior segment (LB6), which could not be extracted despite extensive time.
View from mid LMSB with patent LLL bronchus and partially patent LUL bronchus.
LLL superior segment with residual partially obstructing clot:
 
The LUL remained largely filled with clot, although there may be some partial ventilation occurring here based on bubbling/tidaling of clot seen.
The  LUL Aspergilloma is the suspsected source of bleed.
Therefore, LUL clot was intentionally left in place to allow to organize and prevent further bleeding from the LUL.
Clot remaining in LUL. 
 
 
There was no active bleeding or pooling blood seen.
In hopes of preventing recurrence of active bleeding, the following endobronchial medications were instilled:
Endobronchial epinephrine 0.4mg was instilled into the LUL - targeting LB1/2 and LB3.
Endobronchial tranexamic acid 1000mg was instilled into the LUL - targeting LB1/2 and LB3.
A Uniblocker 9Fr was the only bronchial blocker available.  The Uniblocker ventilator adaptor was attached to the tracheostomy tube and the the 9Fr bronchial balloon blocker balloon was guided into the distal LMSB under direct visualization (using parallel Slim disposable scope via the tracheostomy tube).
-The bronchial blocker balloon was inflated to fully occlude the left mainstem bronchus.
-Placed at 25-26cm at the dark blue securement device.  Bronchial blocker secured tightly in position using this securement device.
-Confirmed that 8cc of air was needed to adequately inflate the balloon to fully occlude the LMSB.
The endobronchial blocker balloon was then deflated and left in the distal LMSB in case emergently needed for recurrent massive hemoptysis.
Residual secretions, saline, and minimal blood suctioned to clear.  Bronchoscope removed.
 
The patient tolerated the procedure well.
There were no immediate complications.  At the conclusion of the operation, the patient's respiratory status was improved.
She will remain in ICU for ongoing critical care management.
 
ESTIMATED BLOOD LOSS:   Minimal active bleeding.
Large amount of organized clot removed.
COMPLICATIONS:    None
 
SPECIMEN(S): 
None
 
IMPRESSION/PLAN: [REDACTED]is a 33 year old-year-old female who presents for bronchoscopy for therapeutic aspiration and foreign body removal of clot.
-Nebulized TXA.
-Hold all AC including heparin gtt for now.
-Keep well sedated to limit coughing.
-Continue broad-spectrum antimicrobials.
-Will tentatively plan for IP to repeat bronchoscopy tomorrow with therapeutic aspiration/cryoextraction of clot.
 
-I remain on-call for IP overnight.
Please do not hesitate to call me if I can provide any additional assistance."""

# =============================================================================
# EXTRACTION LOGIC
# =============================================================================

# Procedure Flags (0/1)
PROCEDURE_FLAGS = {
    'diagnostic_bronchoscopy': 1, # Code 31622 listed
    'bal': 0,
    'bronchial_wash': 1, # Code 31622 "cell washing"
    'brushings': 0,
    'endobronchial_biopsy': 0,
    'tbna_conventional': 0,
    'linear_ebus': 0,
    'radial_ebus': 0,
    'navigational_bronchoscopy': 0,
    'transbronchial_biopsy': 0,
    'transbronchial_cryobiopsy': 0,
    'therapeutic_aspiration': 1, # Explicitly stated multiple times
    'foreign_body_removal': 1, # Explicitly stated "Foreign body removal of clot"
    'airway_dilation': 0,
    'airway_stent': 0,
    'thermal_ablation': 0,
    'tumor_debulking_non_thermal': 0,
    'cryotherapy': 0,
    'blvr': 0,
    'peripheral_ablation': 0,
    'bronchial_thermoplasty': 0,
    'whole_lung_lavage': 0,
    'rigid_bronchoscopy': 0, # "Flexible Therapeutic Bronchoscope" used
    'thoracentesis': 0,
    'chest_tube': 0,
    'ipc': 0,
    'medical_thoracoscopy': 0,
    'pleurodesis': 0,
    'pleural_biopsy': 0,
    'fibrinolytic_therapy': 0
}

# Spans to Extract
# List of dicts: span_text, label, normalized_value (optional), context_prefix (optional)
SPANS_DATA = [
    # Devices
    {"span_text": "Flexible Therapeutic Bronchoscope", "label": "DEV_INSTRUMENT", "normalized_value": "Flexible Therapeutic Bronchoscope"},
    {"span_text": "Disposable Bronchoscope", "label": "DEV_INSTRUMENT", "normalized_value": "Disposable Bronchoscope"},
    {"span_text": "Tracheostomy tube", "label": "DEV_CATHETER", "normalized_value": "Tracheostomy tube", "context_prefix": "Sedation administered and adjusted.  "},
    {"span_text": "bland alligator forceps", "label": "DEV_INSTRUMENT", "normalized_value": "Alligator Forceps"},
    {"span_text": "Neptune suction", "label": "DEV_INSTRUMENT", "normalized_value": "Neptune Suction"},
    {"span_text": "Uniblocker 9Fr", "label": "DEV_CATHETER", "normalized_value": "Uniblocker 9Fr"},
    {"span_text": "bronchial blocker balloon", "label": "DEV_CATHETER", "normalized_value": "Bronchial Blocker"},
    
    # Anatomy
    {"span_text": "Trachea", "label": "ANAT_AIRWAY", "normalized_value": "Trachea", "context_prefix": "Initial Airway Examination Findings:\n"},
    {"span_text": "LMSB", "label": "ANAT_AIRWAY", "normalized_value": "LMSB", "context_prefix": "Clot extending from the "},
    {"span_text": "RMSB", "label": "ANAT_AIRWAY", "normalized_value": "RMSB", "context_prefix": "without obstruction of the "},
    {"span_text": "RB1", "label": "ANAT_AIRWAY", "normalized_value": "RB1", "context_prefix": "fungal ball - noted in "},
    {"span_text": "right mainstem bronchus", "label": "ANAT_AIRWAY", "normalized_value": "RMSB"},
    {"span_text": "right upper lobe", "label": "ANAT_LUNG_LOC", "normalized_value": "RUL"},
    {"span_text": "bronchus intermedius", "label": "ANAT_AIRWAY", "normalized_value": "Bronchus Intermedius"},
    {"span_text": "right middle lobe", "label": "ANAT_LUNG_LOC", "normalized_value": "RML"},
    {"span_text": "right lower lobe", "label": "ANAT_LUNG_LOC", "normalized_value": "RLL"},
    {"span_text": "left mainstem bronchus", "label": "ANAT_AIRWAY", "normalized_value": "LMSB"},
    {"span_text": "left upper lobe", "label": "ANAT_LUNG_LOC", "normalized_value": "LUL"},
    {"span_text": "left lower lobe", "label": "ANAT_LUNG_LOC", "normalized_value": "LLL"},
    {"span_text": "LB7/8-10", "label": "ANAT_AIRWAY", "normalized_value": "LB7-10"},
    {"span_text": "LLL truncus basalis", "label": "ANAT_AIRWAY", "normalized_value": "LLL truncus basalis"},
    {"span_text": "LLL superior segment", "label": "ANAT_AIRWAY", "normalized_value": "LB6"},
    {"span_text": "LB1/2", "label": "ANAT_AIRWAY", "normalized_value": "LB1/2"},
    {"span_text": "LB3", "label": "ANAT_AIRWAY", "normalized_value": "LB3"},

    # Observations/Pathology
    {"span_text": "Massive hemoptysis", "label": "OBS_LESION", "normalized_value": "Hemoptysis", "context_prefix": "presents with "},
    {"span_text": "endobronchial clot", "label": "OBS_LESION", "normalized_value": "Clot", "context_prefix": "presents with massive hemoptysis, "},
    {"span_text": "acute airway obstruction", "label": "OBS_LESION", "normalized_value": "Airway Obstruction", "context_prefix": "endobronchial clot causing "},
    {"span_text": "yellow hard debris", "label": "OBS_LESION", "normalized_value": "Debris"},
    {"span_text": "fungal ball", "label": "OBS_LESION", "normalized_value": "Fungal Ball"},
    {"span_text": "Aspergilloma", "label": "OBS_LESION", "normalized_value": "Aspergilloma"},
    
    # Procedures
    {"span_text": "Therapeutic aspiration", "label": "PROC_METHOD", "normalized_value": "Therapeutic Aspiration", "context_prefix": "31646 "},
    {"span_text": "Foreign body removal", "label": "PROC_METHOD", "normalized_value": "Foreign Body Removal", "context_prefix": "31635 "},
    {"span_text": "Balloon occlusion", "label": "PROC_METHOD", "normalized_value": "Balloon Occlusion", "context_prefix": "31634 "},
    {"span_text": "removal of clot", "label": "PROC_ACTION", "normalized_value": "Removal", "context_prefix": "extensive therapeutic aspiration and "},
    {"span_text": "mechanically excised", "label": "PROC_ACTION", "normalized_value": "Excision"},
    {"span_text": "cold saline flushing", "label": "PROC_METHOD", "normalized_value": "Saline Flush"},
    {"span_text": "Endobronchial epinephrine", "label": "PROC_METHOD", "normalized_value": "Instillation - Epinephrine"},
    {"span_text": "Endobronchial tranexamic acid", "label": "PROC_METHOD", "normalized_value": "Instillation - TXA"},
    
    # Measurements
    {"span_text": "8cc", "label": "MEAS_VOL", "normalized_value": "8cc", "context_prefix": "Confirmed that "},
    {"span_text": "7.5mm", "label": "MEAS_AIRWAY_DIAM", "normalized_value": "7.5mm", "context_prefix": "Tracheostomy tube (ID "}, # Strictly this is device size, but label set says MEAS_AIRWAY_DIAM? Or DEV_CATHETER_SIZE. Let's use DEV_CATHETER_SIZE in V3, but here label set. Prompt says DEV_CATHETER_SIZE exists.
    
    # Correcting 7.5mm label to DEV_CATHETER_SIZE if available in allowed list. 
    # Allowed: DEV_VALVE, DEV_STENT, DEV_CATHETER, DEV_NEEDLE, DEV_INSTRUMENT, DEV_STENT_MATERIAL, DEV_STENT_SIZE, DEV_CATHETER_SIZE
    # Re-doing 7.5mm
    {"span_text": "7.5mm", "label": "DEV_CATHETER_SIZE", "normalized_value": "7.5mm", "context_prefix": "Tracheostomy tube (ID "},

    # Outcomes
    {"span_text": "patient's respiratory status was improved", "label": "OUTCOME_SYMPTOMS", "normalized_value": "Improved"},
    {"span_text": "No immediate complications", "label": "OUTCOME_COMPLICATION", "normalized_value": "None", "is_negated": True},
    {"span_text": "minimal blood", "label": "OBS_LESION", "normalized_value": "Blood", "context_prefix": "Residual secretions, saline, and "},
]

# Events mapping
# Event 1: Therapeutic Aspiration (General)
# Event 2: Foreign Body Removal (Clot)
# Event 3: Instillation
# Event 4: Balloon Occlusion

EVENT_LOG_DATA = [
    {
        "event_id": "evt_01",
        "event_type": "Therapeutic Aspiration",
        "method": "Therapeutic aspiration",
        "anatomy_target": "Trachea, RMSB, RUL, BI, RML, RLL, LMSB, LUL, LLL",
        "device": "Neptune suction",
        "findings": "Blood clot and mucus cleared",
        "outcome_symptoms": "Improved"
    },
    {
        "event_id": "evt_02",
        "event_type": "Foreign Body Removal",
        "method": "Forceps Removal",
        "anatomy_target": "LMSB, LLL",
        "device": "Bland alligator forceps",
        "findings": "Large clot, fungal ball debris",
        "outcome_symptoms": "Improved"
    },
    {
        "event_id": "evt_03",
        "event_type": "Instillation",
        "method": "Instillation",
        "anatomy_target": "LUL (LB1/2, LB3)",
        "findings": "Targeted bleeding source",
        "comments": "Epinephrine and TXA instilled"
    },
    {
        "event_id": "evt_04",
        "event_type": "Balloon Occlusion",
        "method": "Balloon Occlusion",
        "anatomy_target": "Distal LMSB",
        "device": "Uniblocker 9Fr",
        "measurements": "8cc air",
        "findings": "Occluded LMSB, then deflated"
    }
]

V3_EVENTS_DATA = [
    {
        "note_id": NOTE_ID,
        "event_id": "evt_01",
        "type": "Therapeutic Aspiration",
        "target": {"anatomy_type": "Airway", "location": {"lobe": "Multiple"}},
        "method": "Therapeutic aspiration",
        "devices_json": json.dumps(["Neptune suction"]),
        "findings_json": json.dumps(["Blood clot", "Mucus"]),
        "outcomes": {"symptoms": "Improved"}
    },
    {
        "note_id": NOTE_ID,
        "event_id": "evt_02",
        "type": "Foreign Body Removal",
        "target": {"anatomy_type": "Airway", "location": {"lobe": "Left Lower Lobe"}},
        "method": "Forceps",
        "devices_json": json.dumps(["Bland alligator forceps"]),
        "findings_json": json.dumps(["Large clot", "Fungal debris"]),
        "outcomes": {"symptoms": "Improved"}
    },
    {
        "note_id": NOTE_ID,
        "event_id": "evt_03",
        "type": "Instillation",
        "target": {"anatomy_type": "Airway", "location": {"lobe": "Left Upper Lobe"}},
        "method": "Instillation",
        "findings_json": json.dumps(["Epinephrine", "TXA"]),
        "outcomes": {}
    },
    {
        "note_id": NOTE_ID,
        "event_id": "evt_04",
        "type": "Balloon Occlusion",
        "target": {"anatomy_type": "Airway", "location": {"lobe": "Left Mainstem Bronchus"}},
        "method": "Balloon Occlusion",
        "devices_json": json.dumps(["Uniblocker 9Fr"]),
        "measurements_json": json.dumps({"inflation_volume": "8cc"}),
        "outcomes": {}
    }
]

# =============================================================================
# PROCESSING FUNCTIONS
# =============================================================================

def hydrate_spans(note_text, spans):
    """Calculates start_char and end_char for spans based on heuristics."""
    hydrated = []
    
    for span in spans:
        text = span['span_text']
        row = span.copy()
        
        # Default: missing
        row['start_char'] = ""
        row['end_char'] = ""
        row['hydration_status'] = "needs_hydration"
        
        # 1. Exact match count == 1
        count = note_text.count(text)
        
        if count == 1:
            start = note_text.find(text)
            row['start_char'] = start
            row['end_char'] = start + len(text)
            row['hydration_status'] = "hydrated_unique"
            
        elif count > 1 and row.get('context_prefix'):
            # Context match
            prefix = row['context_prefix']
            # Find all occurrences
            starts = [m.start() for m in re.finditer(re.escape(text), note_text)]
            found = False
            for s in starts:
                # Look back 120 chars
                window_start = max(0, s - 120)
                window = note_text[window_start:s]
                if prefix in window:
                    row['start_char'] = s
                    row['end_char'] = s + len(text)
                    row['hydration_status'] = "hydrated_prefix_window"
                    found = True
                    break
            if not found:
                 row['hydration_status'] = f"ambiguous_count={count}_context_not_found"
        
        elif count > 1:
            # Fallback: take first (or logic could be improved if match_index provided)
            # For this script, we default to first if no context
            start = note_text.find(text)
            row['start_char'] = start
            row['end_char'] = start + len(text)
            row['hydration_status'] = f"hydrated_first_of_{count}" # Simplified
            
        else:
             row['hydration_status'] = "not_found"
        
        # Calculate length formula
        if row['start_char'] != "":
            row['span_len'] = f"=J{len(hydrated)+2}-I{len(hydrated)+2}" # Placeholder formula logic
        else:
            row['span_len'] = ""
            
        hydrated.append(row)
    return hydrated

def create_workbook():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        # Create a basic workbook if template missing (fallback)
        wb = openpyxl.Workbook()
        
    # 1. Note_Text
    if "Note_Text" not in wb.sheetnames: wb.create_sheet("Note_Text")
    ws = wb["Note_Text"]
    if ws.max_row == 1: ws.append(["note_id", "source_file", "note_text"])
    ws.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # 2. Note_Index
    if "Note_Index" not in wb.sheetnames: wb.create_sheet("Note_Index")
    ws = wb["Note_Index"]
    # Headers should exist in template, but assuming we append
    # Metadata columns + 30 flags
    row = [
        SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Success", ""
    ]
    # Append flags in order
    flag_keys = [
        'diagnostic_bronchoscopy', 'bal', 'bronchial_wash', 'brushings', 'endobronchial_biopsy', 
        'tbna_conventional', 'linear_ebus', 'radial_ebus', 'navigational_bronchoscopy', 
        'transbronchial_biopsy', 'transbronchial_cryobiopsy', 'therapeutic_aspiration', 
        'foreign_body_removal', 'airway_dilation', 'airway_stent', 'thermal_ablation', 
        'tumor_debulking_non_thermal', 'cryotherapy', 'blvr', 'peripheral_ablation', 
        'bronchial_thermoplasty', 'whole_lung_lavage', 'rigid_bronchoscopy',
        'thoracentesis', 'chest_tube', 'ipc', 'medical_thoracoscopy', 'pleurodesis', 
        'pleural_biopsy', 'fibrinolytic_therapy'
    ]
    for k in flag_keys:
        row.append(PROCEDURE_FLAGS.get(k, 0))
    ws.append(row)
    
    # 3. Span_Annotations
    if "Span_Annotations" not in wb.sheetnames: wb.create_sheet("Span_Annotations")
    ws_span = wb["Span_Annotations"]
    # Header: source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, start_char, end_char, span_len, label, normalized_value, schema_field, event_id, is_negated, is_historical, time_anchor, reviewer, comments, hydration_status
    
    # 4. Span_Hydrated
    if "Span_Hydrated" not in wb.sheetnames: wb.create_sheet("Span_Hydrated")
    ws_hydra = wb["Span_Hydrated"]
    
    hydrated_data = hydrate_spans(NOTE_TEXT, SPANS_DATA)
    
    for idx, item in enumerate(hydrated_data):
        span_id = f"span_{idx+1:03d}"
        
        # Span_Annotations Row (No offsets)
        row_anno = [
            SOURCE_FILE, NOTE_ID, span_id, "", item.get("context_prefix", ""), item["span_text"], "",
            "", "", "", # start, end, len (blank in Anno)
            item["label"], item.get("normalized_value", ""), "", "",
            item.get("is_negated", False), False, "", "", "", "needs_hydration"
        ]
        ws_span.append(row_anno)
        
        # Span_Hydrated Row (With offsets)
        row_hydra = [
            SOURCE_FILE, NOTE_ID, span_id, "", item.get("context_prefix", ""), item["span_text"], "",
            item["start_char"], item["end_char"], item.get("span_len"),
            item["label"], item.get("normalized_value", ""), "", "",
            item.get("is_negated", False), False, "", "", "", item["hydration_status"]
        ]
        ws_hydra.append(row_hydra)

    # 5. Event_Log
    if "Event_Log" not in wb.sheetnames: wb.create_sheet("Event_Log")
    ws_event = wb["Event_Log"]
    for evt in EVENT_LOG_DATA:
        row = [
            SOURCE_FILE, NOTE_ID, evt["event_id"], evt["event_type"], evt.get("method", ""),
            evt.get("anatomy_target", ""), evt.get("device", ""), "", "", "", evt.get("measurements", ""),
            "", evt.get("findings", ""), False, "", evt.get("comments", ""),
            "", "", "", "", evt.get("outcome_symptoms", ""), "", ""
        ]
        ws_event.append(row)

    # 6. V3_Procedure_Events
    if "V3_Procedure_Events" not in wb.sheetnames: wb.create_sheet("V3_Procedure_Events")
    ws_v3 = wb["V3_Procedure_Events"]
    for evt in V3_EVENTS_DATA:
        # note_id, event_id, type, target.anatomy_type, target.location.lobe, target.location.segment, target.station, lesion.type, lesion.size_mm, method, devices_json, measurements_json, specimens_json, findings_json, evidence_quote, stent.size, stent.material_or_brand, catheter.size_fr, outcomes.airway.lumen_pre, outcomes.airway.lumen_post, outcomes.symptoms, outcomes.pleural, outcomes.complications
        
        target = evt.get("target", {})
        loc = target.get("location", {})
        outcomes = evt.get("outcomes", {})
        
        row = [
            evt["note_id"], evt["event_id"], evt["type"],
            target.get("anatomy_type", ""), loc.get("lobe", ""), loc.get("segment", ""), target.get("station", ""),
            evt.get("lesion", {}).get("type", ""), evt.get("lesion", {}).get("size_mm", ""),
            evt.get("method", ""),
            evt.get("devices_json", "[]"),
            evt.get("measurements_json", "{}"),
            evt.get("specimens_json", "[]"),
            evt.get("findings_json", "[]"),
            "", # evidence_quote
            "", "", "", # stent/cath
            outcomes.get("airway", {}).get("lumen_pre", ""),
            outcomes.get("airway", {}).get("lumen_post", ""),
            outcomes.get("symptoms", ""),
            outcomes.get("pleural", ""),
            outcomes.get("complications", "")
        ]
        ws_v3.append(row)

    # 7. V3_Registry_JSON
    if "V3_Registry_JSON" not in wb.sheetnames: wb.create_sheet("V3_Registry_JSON")
    ws_json = wb["V3_Registry_JSON"]
    
    json_obj = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "procedures": V3_EVENTS_DATA,
        "no_immediate_complications": True # Based on note text
    }
    ws_json.append([json.dumps(json_obj, indent=2)])

    wb.save(OUTPUT_PATH)
    print(f"Workbook saved to {OUTPUT_PATH}")

if __name__ == "__main__":
    create_workbook()