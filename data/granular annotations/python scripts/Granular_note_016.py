import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# ==========================================
# 1. INPUTS & CONFIGURATION
# ==========================================
NOTE_ID = "note_016"
SOURCE_FILE = "note_016.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_016 SOURCE_FILE: note_016.txt INDICATION FOR OPERATION:  [REDACTED]is a 33 year old-year-old female who presents with acute respiratory failure.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy and Tracheostomy were discussed with the patient's surrogate (mother, Yvette Sese) in detail.
Patient unable to participate in discussion due to sedation and intubation.
Patient's surrogate indicated a wish to proceed with surgery and informed consent was signed. 
 
CONSENT : Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient or surrogate.
The patient or surrogate read and signed the provided consent form / provided consent over the phone.
The consent was witnessed by an assisting medical professional.
 
PREOPERATIVE DIAGNOSIS: J96.90 Respiratory Failure
 
POSTOPERATIVE DIAGNOSIS:  J96.90 Respiratory Failure
 
PROCEDURE:  
31645 Therapeutic aspiration initial episode
76536 Ultrasound of Neck
 
 
ANESTHESIA: 
99152 Moderate sedation: initial 15 minutes
99153 Moderate sedation: each additional 15 minutes 
 
Procedure performed under moderate sedation.
The following medications were provided:
Versed             4mg
Dilaudid          6mg
Etomidate20mg
Rocuronium     63mg x 2
Propofol60mcg/kg/min  
 
Physician/patient face-to-face anesthesia start time:   1543
            
Physician/patient face-to-face anesthesia stop time:   1822
 
Total moderate sedation time was 159 minutes.
Patient was monitored continuously one-to-one throughout the entire procedure by the attending physician while anesthesia was administered.
Sedation was administered by ICU RN. 
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
 
ESTIMATED BLOOD LOSS:   Minimum
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: Supine
 
Neck Ultrasound was performed to evaluate for any abnormal vessel, mass, structures at the site of percutaneous tracheostomy.
There were no significant vessels/mass were noted overlying the tracheostomy site on examination from the laryngeal prominence to the sternal notch.
High-riding innominate artery was noted just superior to the sternal notch.
Initial Airway Inspection Findings:
 
The endotracheal tube is in good position.
Pharynx: Not assessed due to bronchoscopy introduction through ETT.
Larynx: Not assessed due to bronchoscopy introduction through ETT.
Vocal Cords: Not assessed due to bronchoscopy introduction through ETT.
Trachea: Distal 1/3 normal.
Main Carina: Sharp
Right Lung Proximal Airways: Normal anatomic branching to segmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Left Lung Proximal Airways: Normal anatomic branching to segmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Mucosa: Normal.
Secretions: Copious thick and thin, light yellow secretions/mucus throughout  
 
Successful therapeutic aspiration was performed to clean out the Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient's endotracheal tube was removed and the patient was in stable condition.
SPECIMEN(S): 
--None
 
IMPRESSION/PLAN: [REDACTED]is a 33 year old-year-old female who presents for bronchoscopy and tracheostomy placement.
See separate documentation from Dr Hudson for modified open tracheostomy. Patient tolerated the procedure well and there were no complications.
--Post-procedure CXR
--Anticipate suture removal in ~7 days 
--Anticipate trach exchange in ~10 days"""

# ==========================================
# 2. DEFINITIONS & SCHEMA
# ==========================================

ALLOWED_LABELS = {
    "ANAT_LN_STATION", "ANAT_LUNG_LOC", "ANAT_AIRWAY", "ANAT_PLEURA", "LATERALITY",
    "DEV_VALVE", "DEV_STENT", "DEV_CATHETER", "DEV_NEEDLE", "DEV_INSTRUMENT",
    "DEV_STENT_MATERIAL", "DEV_STENT_SIZE", "DEV_CATHETER_SIZE",
    "MEAS_SIZE", "MEAS_VOL", "MEAS_PRESS", "MEAS_COUNT", "MEAS_AIRWAY_DIAM", "MEAS_PLEURAL_DRAIN",
    "PROC_METHOD", "PROC_ACTION", "OBS_ROSE", "OBS_LESION",
    "CTX_HISTORICAL", "CTX_TIME",
    "OUTCOME_AIRWAY_LUMEN_PRE", "OUTCOME_AIRWAY_LUMEN_POST", "OUTCOME_SYMPTOMS",
    "OUTCOME_PLEURAL", "OUTCOME_COMPLICATION"
}

PROCEDURE_FLAGS_MAP = {
    # Bronchoscopy (23)
    "diagnostic_bronchoscopy": False,
    "bal": False,
    "bronchial_wash": False,
    "brushings": False,
    "endobronchial_biopsy": False,
    "tbna_conventional": False,
    "linear_ebus": False,
    "radial_ebus": False,
    "navigational_bronchoscopy": False,
    "transbronchial_biopsy": False,
    "transbronchial_cryobiopsy": False,
    "therapeutic_aspiration": True, # Matches "31645 Therapeutic aspiration"
    "foreign_body_removal": False,
    "airway_dilation": False,
    "airway_stent": False,
    "thermal_ablation": False,
    "tumor_debulking_non_thermal": False,
    "cryotherapy": False,
    "blvr": False,
    "peripheral_ablation": False,
    "bronchial_thermoplasty": False,
    "whole_lung_lavage": False,
    "rigid_bronchoscopy": False,
    # Pleural (7)
    "thoracentesis": False,
    "chest_tube": False,
    "ipc": False,
    "medical_thoracoscopy": False,
    "pleurodesis": False,
    "pleural_biopsy": False,
    "fibrinolytic_therapy": False
}

# ==========================================
# 3. EXTRACTION LOGIC
# ==========================================

class SpanExtractor:
    def __init__(self, text):
        self.text = text
        self.spans = []
        self.events = []
        self.current_event_id = None

    def add_span(self, text_match, label, normalized=None, context_prefix=None, section="PROCEDURE"):
        if not text_match:
            return
        
        # Validation
        if label not in ALLOWED_LABELS:
            raise ValueError(f"Invalid label: {label}")
        
        # Calculate match index or uniqueness
        count = self.text.count(text_match)
        match_index = 1 # Default to first if not specified, hydration will resolve
        
        if count > 1 and not context_prefix:
            # Simple heuristic: find first occurrence after last span if possible, 
            # but for this script we stick to explicit unique strings or first match
            pass

        self.spans.append({
            "span_text": text_match,
            "label": label,
            "normalized_value": normalized if normalized else text_match,
            "context_prefix": context_prefix,
            "section_type": section,
            "event_id": self.current_event_id,
            "match_index": match_index
        })

    def start_event(self, event_id):
        self.current_event_id = event_id

def extract_data(text):
    extractor = SpanExtractor(text)
    
    # -- EVENT 1: Ultrasound Neck --
    extractor.start_event("evt_01")
    extractor.add_span("Ultrasound of Neck", "PROC_METHOD", "ultrasound", section="PROCEDURE_HEADER")
    extractor.add_span("tracheostomy site", "ANAT_AIRWAY", "trachea", context_prefix="overlying the ")
    extractor.add_span("High-riding innominate artery", "OBS_LESION", "vascular_anomaly", section="PROCEDURE_DETAIL")
    
    # -- EVENT 2: Inspection --
    extractor.start_event("evt_02")
    extractor.add_span("endotracheal tube", "DEV_CATHETER", "ETT", section="PROCEDURE_DETAIL")
    extractor.add_span("Trachea: Distal 1/3", "ANAT_AIRWAY", "trachea_distal", section="PROCEDURE_DETAIL")
    extractor.add_span("Main Carina", "ANAT_AIRWAY", "main_carina", section="PROCEDURE_DETAIL")
    extractor.add_span("Right Lung Proximal Airways", "ANAT_AIRWAY", "R_proximal", section="PROCEDURE_DETAIL")
    extractor.add_span("Left Lung Proximal Airways", "ANAT_AIRWAY", "L_proximal", section="PROCEDURE_DETAIL")
    extractor.add_span("No evidence of mass, lesions, bleeding", "OBS_LESION", "none", section="PROCEDURE_DETAIL")
    extractor.add_span("Copious thick and thin, light yellow secretions/mucus", "OBS_LESION", "secretions", section="PROCEDURE_DETAIL")

    # -- EVENT 3: Therapeutic Aspiration --
    extractor.start_event("evt_03")
    extractor.add_span("Therapeutic aspiration", "PROC_METHOD", "therapeutic_aspiration", context_prefix="Successful ")
    extractor.add_span("clean out", "PROC_ACTION", "aspiration", section="PROCEDURE_DETAIL")
    
    # Anatomical targets for aspiration
    locs = [
        ("Trachea (Distal 1/3)", "ANAT_AIRWAY", "trachea_distal", "clean out the "),
        ("Right Mainstem", "ANAT_AIRWAY", "RMS", "Distal 1/3), "),
        ("Bronchus Intermedius", "ANAT_AIRWAY", "BI", "Right Mainstem, "),
        ("Left Mainstem", "ANAT_AIRWAY", "LMS", "Bronchus Intermedius , "),
        ("Carina", "ANAT_AIRWAY", "carina", "Left Mainstem, "),
        ("RUL Carina (RC1)", "ANAT_AIRWAY", "RUL_carina", "Carina, "),
        ("RML Carina (RC2)", "ANAT_AIRWAY", "RML_carina", "RUL Carina (RC1), "),
        ("LUL Lingula Carina (Lc1)", "ANAT_AIRWAY", "LUL_lingula", "RML Carina (RC2), "),
        ("Left Carina (LC2)", "ANAT_AIRWAY", "L_carina", "LUL Lingula Carina (Lc1), ")
    ]
    
    for txt, lbl, norm, ctx in locs:
        extractor.add_span(txt, lbl, norm, context_prefix=ctx, section="PROCEDURE_DETAIL")
        
    extractor.add_span("mucus", "OBS_LESION", "mucus", context_prefix="(LC2) from ", section="PROCEDURE_DETAIL")
    
    # Outcomes
    extractor.add_span("tolerated the procedure well", "OUTCOME_SYMPTOMS", "tolerated_well", section="PROCEDURE_DETAIL")
    extractor.add_span("No immediate complications", "OUTCOME_COMPLICATION", "none", section="PROCEDURE_DETAIL")
    
    # Global / Final
    extractor.start_event("evt_global")
    extractor.add_span("31645 Therapeutic aspiration initial episode", "PROC_METHOD", "therapeutic_aspiration_code", section="PROCEDURE_HEADER")
    extractor.add_span("Flexible Therapeutic Bronchoscope", "DEV_INSTRUMENT", "bronchoscope", section="INSTRUMENT")

    return extractor.spans

# ==========================================
# 4. EXCEL GENERATION HELPERS
# ==========================================

def setup_workbook(template_path):
    if not os.path.exists(template_path):
        # Create a dummy workbook if template missing (fallback for valid code)
        wb = openpyxl.Workbook()
        wb.create_sheet("Note_Text")
        wb.create_sheet("Note_Index")
        wb.create_sheet("Span_Annotations")
        wb.create_sheet("Span_Hydrated")
        wb.create_sheet("Event_Log")
        wb.create_sheet("V3_Procedure_Events")
        wb.create_sheet("V3_Registry_JSON")
        return wb
    return openpyxl.load_workbook(template_path)

def hydrate_offsets(spans, note_text):
    hydrated = []
    for s in spans:
        row = s.copy()
        text = s['span_text']
        prefix = s.get('context_prefix')
        match_idx = s.get('match_index', 1)
        
        start = -1
        end = -1
        status = "ambiguous"

        # Strategy 1: Context Prefix
        if prefix and prefix in note_text and text in note_text:
            # Find all occurrences of prefix
            prefix_iter = re.finditer(re.escape(prefix), note_text)
            for m in prefix_iter:
                # Check if text follows within reasonable window (e.g. 0-5 chars)
                search_start = m.end()
                window = note_text[search_start : search_start + len(text) + 50]
                if text in window:
                    local_idx = window.find(text)
                    start = search_start + local_idx
                    end = start + len(text)
                    status = "hydrated_prefix"
                    break
        
        # Strategy 2: Unique
        if start == -1 and note_text.count(text) == 1:
            start = note_text.find(text)
            end = start + len(text)
            status = "hydrated_unique"
            
        # Strategy 3: Match Index
        if start == -1 and note_text.count(text) >= match_idx:
            # Find nth match
            curr = 0
            idx = -1
            while curr < match_idx:
                idx = note_text.find(text, idx + 1)
                curr += 1
            start = idx
            end = start + len(text)
            status = "hydrated_index"

        row['start_char'] = start if start != -1 else ""
        row['end_char'] = end if end != -1 else ""
        row['hydration_status'] = status
        row['span_len'] = len(text)
        hydrated.append(row)
    return hydrated

# ==========================================
# 5. MAIN EXECUTION
# ==========================================

def main():
    wb = setup_workbook(TEMPLATE_PATH)
    
    # -----------------------
    # 1. Note_Text
    # -----------------------
    ws = wb["Note_Text"]
    # Assuming header exists, append
    ws.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # -----------------------
    # 2. Note_Index
    # -----------------------
    ws = wb["Note_Index"]
    row_meta = [
        SOURCE_FILE, NOTE_ID, "", "", "", "", "Complete", ""
    ]
    # Add flags
    flag_values = [1 if PROCEDURE_FLAGS_MAP.get(k, False) else 0 for k in PROCEDURE_FLAGS_MAP]
    ws.append(row_meta + flag_values)

    # -----------------------
    # 3. Span_Annotations
    # -----------------------
    raw_spans = extract_data(NOTE_TEXT)
    ws_anno = wb["Span_Annotations"]
    
    # Headers: source, note, span_id, section, prefix, text, match_idx, start, end, len, label, norm, field, event, neg, hist, time, rev, comm, hyd_stat
    span_id_counter = 1
    for s in raw_spans:
        ws_anno.append([
            SOURCE_FILE, NOTE_ID, f"span_{span_id_counter:03d}", s['section_type'],
            s.get('context_prefix', ""), s['span_text'], s.get('match_index', 1),
            "", "", f"=LEN(F{ws_anno.max_row})", # Formula for len
            s['label'], s['normalized_value'], "", s['event_id'],
            "FALSE", "FALSE", "", "", "", "needs_hydration"
        ])
        s['span_id'] = f"span_{span_id_counter:03d}"
        span_id_counter += 1

    # -----------------------
    # 4. Span_Hydrated
    # -----------------------
    ws_hyd = wb["Span_Hydrated"]
    hydrated_data = hydrate_offsets(raw_spans, NOTE_TEXT)
    
    for h in hydrated_data:
        ws_hyd.append([
            SOURCE_FILE, NOTE_ID, h['span_id'], h['section_type'],
            h.get('context_prefix', ""), h['span_text'], h.get('match_index', 1),
            h['start_char'], h['end_char'], h['span_len'],
            h['label'], h['normalized_value'], "", h['event_id'],
            "FALSE", "FALSE", "", "", "", h['hydration_status']
        ])

    # -----------------------
    # 5. Event_Log
    # -----------------------
    ws_event = wb["Event_Log"]
    # Group by event_id
    events_map = {}
    for h in hydrated_data:
        eid = h['event_id']
        if eid not in events_map:
            events_map[eid] = {
                "type": "diagnostic", "method": [], "anatomy": [], "device": [], 
                "findings": [], "outcomes": [], "complications": []
            }
        
        lbl = h['label']
        val = h['normalized_value']
        
        if lbl == "PROC_METHOD":
            events_map[eid]["method"].append(val)
        elif lbl == "ANAT_AIRWAY":
            events_map[eid]["anatomy"].append(val)
        elif lbl == "DEV_INSTRUMENT" or lbl == "DEV_CATHETER":
            events_map[eid]["device"].append(val)
        elif lbl == "OBS_LESION":
            events_map[eid]["findings"].append(val)
        elif lbl == "OUTCOME_COMPLICATION":
            events_map[eid]["complications"].append(val)

    for eid, data in events_map.items():
        # Heuristics for event type
        etype = "diagnostic"
        if "therapeutic_aspiration" in data["method"]:
            etype = "therapeutic"
        
        ws_event.append([
            SOURCE_FILE, NOTE_ID, eid, etype, 
            ", ".join(set(data["method"])),
            ", ".join(set(data["anatomy"])),
            ", ".join(set(data["device"])),
            "", "", "", "", # gauge, station, counts, meas
            "", # specimens
            ", ".join(set(data["findings"])),
            "FALSE", "", "",
            "", "", # dev size/mat
            "", "", "", "", # outcomes airway/pleural
            ", ".join(set(data["complications"]))
        ])

    # -----------------------
    # 6. V3_Procedure_Events & JSON
    # -----------------------
    ws_v3 = wb["V3_Procedure_Events"]
    ws_json = wb["V3_Registry_JSON"]
    
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": [],
        "no_immediate_complications": True # Derived from note text "No immediate complications"
    }
    
    for eid, data in events_map.items():
        # Build V3 row
        row = [
            NOTE_ID, eid, data.get("type", "diagnostic"),
            "Airway", # target.anatomy_type
            "", "", "", # lobe, segment, station
            "", "", # lesion type, size
            json.dumps(data["method"]),
            json.dumps(data["device"]),
            "[]", "[]", # meas, specimens
            json.dumps(data["findings"]),
            "", # quote
            "", "", "", # stent props
            "", "", "", "", # outcomes
            json.dumps(data["complications"])
        ]
        ws_v3.append(row)
        
        # Build JSON object
        proc_obj = {
            "event_id": eid,
            "type": "Therapeutic" if "therapeutic" in str(data["method"]) else "Diagnostic",
            "method": data["method"],
            "anatomy": data["anatomy"],
            "findings": data["findings"]
        }
        registry_data["procedures"].append(proc_obj)

    ws_json.cell(row=1, column=1, value=json.dumps(registry_data, indent=2))

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    main()