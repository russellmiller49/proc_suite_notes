import openpyxl
from openpyxl.utils import get_column_letter
import os
import json
import re

# ==========================================
# INPUT DATA
# ==========================================
NOTE_ID = "note_017"
SOURCE_FILE = "note_017.txt"
PROCEDURE_DATE = "2026-01-12" # Placeholder/Current based on prompt context if not in text
NOTE_TEXT = """NOTE_ID:  note_017 SOURCE_FILE: note_017.txt INDICATION FOR OPERATION:  [REDACTED]is a 68 year old-year-old female who presents with bronchial stenosis.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31624 Dx bronchoscope/lavage (BAL)    
31625 Endobronchial Biopsy(s)
31654 Radial EBUS for peripheral lesion
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
31635 Foreign body removal
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required a extensive mechanical excision of endobronchial tissue to salvage 
the airway as well as two separate stent placement given abarrent anatomy.
This resulted in >80% increased work due to Technical difficulty of procedure and Physical and mental effort required.
Apply to: 31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy).
Pt with complex airway with significant stenosis of the LMSB and complete occlusion of the LLL bronchus that required multiple attempts and 
 
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Disposable Bronchoscope
ESTIMATED BLOOD LOSS:   Minimum
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: 
Initial Airway Inspection Findings:
An iGel was placed by anesthesia after adequate sedation.
Successful therapeutic aspiration was performed to clean out the Trachea (Middle 1/3), Trachea (Distal 1/3), Left Mainstem, Carina, LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus and mucus plug.
The stent was partially occluded with thick yellow,green mucus
The LUL bronchus was approximately 3-4 mm and the LLL bronchus could not be visualized.
Saline was instilled without visualization of the LLL.  
Endobronchial tumor was noted and required extensive excision and mechanical debridement.
(Alligator forceps) and 1.7 cryoprobe with 30 second freeze cycles to vascular occlusion and then cryotherapy for further debulking of the LUL ostium.
It appeared that the LMSB stent migrated distally and appeared to be covering the LMSB
Using forceps the proximal end of the microtech stent was grasped and then removed on en bloc with the bronchoscope.
The foreign body removal was difficult due to significant inflammation.
After removal the LMSB mucosa was ragged, irregular and there was dynamic collapse.
The LLL bronchus could not be visualized and the LC2 was difficult to delineate along the inferior portion.
The area of firm and unable to use multiple aliquots the LLL airway could not be identified.
Due to complex anatomy a jag wire was placed in the Lingula and the radial EBUS was utilized to identify vasculature and airways but the LLL was still unable to be identified.
Bronchial alveolar lavage was performed at Superior Segment of Lingula (LB4).
Instilled 40 cc of NS, suction returned with 15 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
 
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
- LUL BAL - cell count, culture and cytology
- LMSB EBBx - pathology 
IMPRESSION/PLAN: [REDACTED]is a 68 year old-year-old female who presents for bronchoscopy for bronchial stenosis.
Very odd presentation of cicatrization and benign stenosis of the LMSB without evidence of residual malignancy.
Very challenging anatomy of the airway and stenosis. Continues to have robust inflammatory response and no additional stents were placed.
- admit overnight
- start abx to treat MSSA, add prednisone 20 mg and obtain CT chest with contrast
- NPO after midnight for bronchoscopy [REDACTED]"""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# ==========================================
# CONFIGURATION & MAPPINGS
# ==========================================
PROCEDURE_FLAGS_MAP = {
    "diagnostic_bronchoscopy": ["Dx bronchoscope", "diagnostic bronchoscopy"],
    "bal": ["lavage (BAL)", "Bronchial alveolar lavage", "BAL - cell count"],
    "bronchial_wash": ["bronchial wash"],
    "brushings": ["brushings"],
    "endobronchial_biopsy": ["Endobronchial Biopsy", "EBBx"],
    "tbna_conventional": ["tbna", "transbronchial needle aspiration"],
    "linear_ebus": ["linear ebus", "ebus-tbna"],
    "radial_ebus": ["Radial EBUS"],
    "navigational_bronchoscopy": ["navigational", "veran", "superdimension", "ion"],
    "transbronchial_biopsy": ["transbronchial biopsy", "tbbx"],
    "transbronchial_cryobiopsy": ["transbronchial cryobiopsy", "cryobiopsy"],
    "therapeutic_aspiration": ["Therapeutic aspiration", "therapeutic aspiration"],
    "foreign_body_removal": ["Foreign body removal", "stent... removed", "stent was grasped and then removed"],
    "airway_dilation": ["balloon dilation", "dilated", "dilation"],
    "airway_stent": ["stent placement", "stent deployment", "stent removal", "stent was... removed", "additional stents"],
    "thermal_ablation": ["laser", "apc", "electrocautery", "argon plasma"],
    "tumor_debulking_non_thermal": ["mechanical excision", "mechanical debridement", "microdebrider"],
    "cryotherapy": ["cryotherapy", "cryoprobe"],
    "blvr": ["valve"],
    "peripheral_ablation": ["ablation", "microwave"],
    "bronchial_thermoplasty": ["thermoplasty"],
    "whole_lung_lavage": ["whole lung lavage"],
    "rigid_bronchoscopy": ["rigid bronchoscope", "rigid bronchoscopy"],
    "thoracentesis": ["thoracentesis"],
    "chest_tube": ["chest tube", "pigtail"],
    "ipc": ["indwelling pleural catheter", "ipc", "pleurx"],
    "medical_thoracoscopy": ["pleuroscopy", "medical thoracoscopy"],
    "pleurodesis": ["pleurodesis", "talc"],
    "pleural_biopsy": ["pleural biopsy"],
    "fibrinolytic_therapy": ["tpa", "dnase"]
}

# ==========================================
# PROCESSING LOGIC
# ==========================================

def clean_text(text):
    return text.replace('\r', '').strip()

def get_procedure_flags(text):
    text_lower = text.lower()
    flags = {}
    for flag, keywords in PROCEDURE_FLAGS_MAP.items():
        val = 0
        for kw in keywords:
            if kw in text_lower:
                val = 1
                break
        flags[flag] = val
    
    # Specific logic for this note based on context
    # Note mentions "31641 Destruction of tumor OR relief of stenosis" and "mechanical excision" -> tumor_debulking_non_thermal
    if "mechanical excision" in text_lower or "mechanical debridement" in text_lower:
        flags["tumor_debulking_non_thermal"] = 1
        
    return flags

def generate_spans(text):
    spans = []
    
    # Helper to add span
    def add_span(text_match, label, norm_val, schema_field=None, event_id=None, context_prefix=None, is_negated=False):
        if not text_match: return
        spans.append({
            "span_text": text_match,
            "label": label,
            "normalized_value": norm_val,
            "schema_field": schema_field,
            "event_id": event_id,
            "context_prefix": context_prefix,
            "is_negated": is_negated
        })

    # --- Event 1: Therapeutic Aspiration ---
    # "Successful therapeutic aspiration was performed to clean out the Trachea..."
    add_span("Therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", "method", "evt1")
    add_span("Trachea (Middle 1/3)", "ANAT_AIRWAY", "Trachea", "target.anatomy_type", "evt1")
    add_span("Trachea (Distal 1/3)", "ANAT_AIRWAY", "Trachea", "target.anatomy_type", "evt1")
    add_span("Left Mainstem", "ANAT_AIRWAY", "LMSB", "target.anatomy_type", "evt1")
    add_span("Carina", "ANAT_AIRWAY", "Carina", "target.anatomy_type", "evt1")
    add_span("LUL Lingula Carina (Lc1)", "ANAT_AIRWAY", "LUL", "target.anatomy_type", "evt1")
    add_span("Left Carina (LC2)", "ANAT_AIRWAY", "LUL", "target.anatomy_type", "evt1")
    add_span("mucus plug", "OBS_LESION", "Mucus Plug", "findings_json", "evt1")

    # --- Event 2: Stent Removal (Foreign Body) ---
    # "proximal end of the microtech stent was grasped and then removed"
    # "Foreign body removal"
    add_span("Foreign body removal", "PROC_METHOD", "Foreign Body Removal", "method", "evt2")
    add_span("microtech stent", "DEV_STENT", "Stent", "device", "evt2")
    add_span("forceps", "DEV_INSTRUMENT", "Forceps", "devices_json", "evt2", context_prefix="Using ")
    add_span("removed on en bloc", "PROC_ACTION", "Removed", "findings_json", "evt2")
    add_span("LMSB", "ANAT_AIRWAY", "LMSB", "target.anatomy_type", "evt2", context_prefix="covering the ")
    
    # --- Event 3: Tumor Debulking & Cryotherapy ---
    # "Endobronchial tumor was noted and required extensive excision and mechanical debridement."
    # "1.7 cryoprobe... cryotherapy"
    add_span("Endobronchial tumor", "OBS_LESION", "Tumor", "lesion.type", "evt3")
    add_span("extensive excision", "PROC_METHOD", "Mechanical Debulking", "method", "evt3")
    add_span("mechanical debridement", "PROC_METHOD", "Mechanical Debulking", "method", "evt3")
    add_span("Alligator forceps", "DEV_INSTRUMENT", "Alligator Forceps", "devices_json", "evt3")
    add_span("1.7 cryoprobe", "DEV_INSTRUMENT", "Cryoprobe", "devices_json", "evt3")
    add_span("cryotherapy", "PROC_METHOD", "Cryotherapy", "method", "evt3", context_prefix="and then ")
    add_span("LUL ostium", "ANAT_AIRWAY", "LUL", "target.anatomy_type", "evt3")

    # --- Event 4: Radial EBUS ---
    # "radial EBUS was utilized to identify vasculature and airways"
    # "Lingula"
    add_span("radial EBUS", "PROC_METHOD", "Radial EBUS", "method", "evt4")
    add_span("Lingula", "ANAT_LUNG_LOC", "Lingula", "target.location.lobe", "evt4", context_prefix="placed in the ")

    # --- Event 5: BAL ---
    # "Bronchial alveolar lavage was performed at Superior Segment of Lingula (LB4)."
    add_span("Bronchial alveolar lavage", "PROC_METHOD", "BAL", "method", "evt5")
    add_span("Superior Segment of Lingula (LB4)", "ANAT_LUNG_LOC", "Lingula Superior Segment", "target.location.segment", "evt5")
    add_span("Instilled 40 cc", "MEAS_VOL", "40", "measurements_json", "evt5")
    add_span("returned with 15 cc", "MEAS_VOL", "15", "measurements_json", "evt5")
    add_span("LUL BAL", "OBS_ROSE", "BAL", "specimens_json", "evt5")

    # --- Event 6: Endobronchial Biopsy ---
    # "LMSB EBBx - pathology"
    add_span("LMSB", "ANAT_AIRWAY", "LMSB", "target.anatomy_type", "evt6", context_prefix="- ")
    add_span("EBBx", "PROC_METHOD", "Endobronchial Biopsy", "method", "evt6")
    
    # --- Outcomes / General ---
    add_span("No immediate complications", "OUTCOME_COMPLICATION", "None", "outcomes.complications", "evt99")

    return spans

def hydrate_spans(spans, text):
    hydrated = []
    text_lower = text.lower()
    
    for s in spans:
        st = s["span_text"]
        row = s.copy()
        row["hydration_status"] = "ambiguous"
        row["start_char"] = ""
        row["end_char"] = ""
        
        matches = [m for m in re.finditer(re.escape(st), text)]
        
        if len(matches) == 1:
            m = matches[0]
            row["start_char"] = m.start()
            row["end_char"] = m.end()
            row["hydration_status"] = "hydrated_unique"
        elif len(matches) > 1:
            if s.get("context_prefix"):
                # Try context prefix
                found = False
                for m in matches:
                    start_look = max(0, m.start() - 120)
                    prefix_window = text[start_look:m.start()]
                    if s["context_prefix"] in prefix_window:
                        row["start_char"] = m.start()
                        row["end_char"] = m.end()
                        row["hydration_status"] = "hydrated_prefix_window"
                        found = True
                        break
                if not found:
                    # Fallback: first occurence if not specific logic
                    m = matches[0]
                    row["start_char"] = m.start()
                    row["end_char"] = m.end()
                    row["hydration_status"] = "hydrated_fallback_first"
            else:
                # Default to first
                m = matches[0]
                row["start_char"] = m.start()
                row["end_char"] = m.end()
                row["hydration_status"] = "hydrated_match_index_0"
        
        row["span_len"] = len(st)
        hydrated.append(row)
        
    return hydrated

def create_event_log(spans):
    events = {}
    
    for s in spans:
        eid = s.get("event_id")
        if not eid: continue
        
        if eid not in events:
            events[eid] = {
                "source_file": SOURCE_FILE,
                "note_id": NOTE_ID,
                "event_id": eid,
                "event_type": "",
                "method": "",
                "anatomy_target": "",
                "device": "",
                "findings": [],
                "specimens": [],
                "measurements": []
            }
        
        label = s["label"]
        val = s["normalized_value"]
        
        if label == "PROC_METHOD":
            events[eid]["method"] = val
            events[eid]["event_type"] = val # Proxy
        elif label.startswith("ANAT"):
            events[eid]["anatomy_target"] = val
        elif label.startswith("DEV"):
            events[eid]["device"] = val
        elif label == "OBS_LESION":
            events[eid]["findings"].append(val)
        elif label == "MEAS_VOL":
            events[eid]["measurements"].append(f"{val}cc")
        elif label == "OBS_ROSE":
            events[eid]["specimens"].append(val)

    # Convert to list
    rows = []
    for eid, data in events.items():
        data["findings"] = ", ".join(data["findings"])
        data["specimens"] = ", ".join(data["specimens"])
        data["measurements"] = ", ".join(data["measurements"])
        rows.append(data)
    return rows

def create_registry_json(spans, events_list):
    # Simplified JSON structure
    out = {
        "schema_version": "v3",
        "note_id": NOTE_ID,
        "procedures": events_list,
        "no_immediate_complications": True
    }
    return json.dumps(out, indent=2)

# ==========================================
# EXCEL GENERATION
# ==========================================

def main():
    if not os.path.exists(TEMPLATE_PATH):
        # Create dummy template if missing (for standalone run capability, though instruction says it exists)
        wb = openpyxl.Workbook()
        wb.create_sheet("Note_Text")
        wb.create_sheet("Note_Index")
        wb.create_sheet("Span_Annotations")
        wb.create_sheet("Span_Hydrated")
        wb.create_sheet("Event_Log")
        wb.create_sheet("V3_Procedure_Events")
        wb.create_sheet("V3_Registry_JSON")
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # 1. Note_Text
    ws = wb["Note_Text"]
    if ws.max_row == 1:
        ws.append(["note_id", "source_file", "note_text"])
    ws.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws = wb["Note_Index"]
    flags = get_procedure_flags(NOTE_TEXT)
    
    # Headers expected (meta + 30 flags)
    # Assuming template has headers, we append a row matching them.
    # We'll construct the row based on known flag order + metadata
    row = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Success", ""]
    
    flag_keys = [
        "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", 
        "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy", 
        "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration", 
        "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation", 
        "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", 
        "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
        "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis", 
        "pleural_biopsy", "fibrinolytic_therapy"
    ]
    
    for k in flag_keys:
        row.append(flags.get(k, 0))
    
    ws.append(row)

    # 3. Span_Annotations & 4. Span_Hydrated
    spans = generate_spans(NOTE_TEXT)
    hydrated_spans = hydrate_spans(spans, NOTE_TEXT)
    
    ws_ann = wb["Span_Annotations"]
    ws_hyd = wb["Span_Hydrated"]
    
    # Header: source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, start_char, end_char, span_len, label, normalized_value, schema_field, event_id, is_negated, is_historical, time_anchor, reviewer, comments, hydration_status
    
    for idx, s in enumerate(hydrated_spans):
        span_id = f"{NOTE_ID}_s{idx+1:03d}"
        
        # Annotation Row (Clean)
        ann_row = [
            SOURCE_FILE, NOTE_ID, span_id, "PROCEDURE",
            s.get("context_prefix", ""), s["span_text"], "", 
            "", "", s["span_len"], # Start/End blank
            s["label"], s["normalized_value"], s["schema_field"], s["event_id"],
            s["is_negated"], False, "", "", "", "needs_hydration"
        ]
        ws_ann.append(ann_row)
        
        # Hydrated Row
        hyd_row = [
            SOURCE_FILE, NOTE_ID, span_id, "PROCEDURE",
            s.get("context_prefix", ""), s["span_text"], "", 
            s["start_char"], s["end_char"], s["span_len"],
            s["label"], s["normalized_value"], s["schema_field"], s["event_id"],
            s["is_negated"], False, "", "", "", s["hydration_status"]
        ]
        ws_hyd.append(hyd_row)

    # 5. Event_Log
    events = create_event_log(spans)
    ws_evt = wb["Event_Log"]
    for e in events:
        # Map dict to columns roughly
        r = [
            SOURCE_FILE, NOTE_ID, e["event_id"], e["event_type"], e["method"],
            e["anatomy_target"], e["device"], "", "", "", e["measurements"],
            e["specimens"], e["findings"], False, "", "",
            "", "", "", "", "", "", "" # Outcomes blank for now in simple log
        ]
        ws_evt.append(r)

    # 6. V3_Procedure_Events (Simplified mapping)
    ws_v3 = wb["V3_Procedure_Events"]
    for e in events:
        r = [
            NOTE_ID, e["event_id"], e["event_type"],
            e["anatomy_target"], "", "", "", # Target breakdown
            "", "", # Lesion
            e["method"], "", e["measurements"], e["specimens"], e["findings"], "",
            "", "", "", # Stent/Cath
            "", "", "", "", "" # Outcomes
        ]
        ws_v3.append(r)

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    json_str = create_registry_json(spans, events)
    ws_json.append([json_str])

    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    main()