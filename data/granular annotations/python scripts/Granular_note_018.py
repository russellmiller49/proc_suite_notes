import openpyxl
from openpyxl.utils import get_column_letter
import json
import datetime
import re
import os

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_018"
SOURCE_FILE = "note_018.txt"
PROCEDURE_DATE = "2026-01-12" # Placeholder/Current
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_018 SOURCE_FILE: note_018.txt INDICATION FOR OPERATION:  [REDACTED]is a 60 year old-year-old female who presents with airway stenosis.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
 
CONSENT : Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient or surrogate.
The patient or surrogate read and signed the provided consent form / provided consent over the phone.
The consent was witnessed by an assisting medical professional.
 
PREOPERATIVE DIAGNOSIS: R91.8 Other nonspecific abnormal finding of lung field.
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
 
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31624 Dx bronchoscope/lavage (BAL)    
31631 Dilate and tracheal stent placement   
31640 Bronchoscopy with excision 
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
 
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
Flexible Hybrid (Pedatric) Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
 
Significant narrowing at the cord level, therapeutic scope was not able to be passed.
Surgical Planning Measurements:
 	 
Distance from vocal folds to stenosis	 0 mm
Location of stenosis relative to cricoid 	At cricoid
Distance from top of stenosis to bottom 	105 mm
Distance of bottom of stenosis to carina	90 mm
Approximation of most narrow portion	6 mm
 
Endobronchial obstruction at Subglottic was treated with the following modalities:
 
Modality	Tools	Setting/Mode	Duration	Results
Electrocautery	Knife (needle)	Endo cut I 2 2 1	4secs burst	Radial cut
APC	 	 	 	 
Laser	 	 	 	 
CoreCath	 	 	 	 
Cryoprobe	 	 	 	 
 
Prior to treatment, affected airway was note to be 50% patent.
After treatment, the airway was 50% patent. 
 
Balloon dilation was performed at Trachea (Proximal 1/3).
8/9/10 Elation balloon was used to perform dilation to 10 mm at the Trachea (Proximal 1/3).
Total 3 inflations with dilation time of 60 seconds each.
 
Endobronchial excision of tissue was performed at Trachea (Proximal 1/3).
Lesion was successfully removed.  Samples sent for Microbiology (Cultures/Viral/Fungal) and Cytology.
Sheet of tissue completed peeled off with forceps in a circular fashion.
Bronchial alveolar lavage was performed at Lateral Segment of RML (RB4) and Medial Segment of RML (RB5).
Instilled 60 cc of NS, suction returned with 20 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
 
Successful therapeutic aspiration was performed to clean out the Trachea (Proximal 1/3), Trachea (Middle 1/3), Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , and Left Mainstem from mucus and blood.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
EBBX , BAL
 
IMPRESSION/PLAN: [REDACTED]is a 60 year old-year-old female who presents for bronchoscopy for subglottic stenosis.
- repeat bronchoscopy in 4-6 weeks 
- f/u in clinic"""

# -------------------------------------------------------------------------
# CONFIGURATION
# -------------------------------------------------------------------------

PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1, # BAL
    "bal": 1,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0, # Excision is debulking here
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 0,
    "airway_dilation": 1,
    "airway_stent": 0,
    "thermal_ablation": 1, # Electrocautery
    "tumor_debulking_non_thermal": 1, # Forceps excision
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# -------------------------------------------------------------------------
# SPAN DATA
# -------------------------------------------------------------------------

# Define spans for logic. 
# Fields: span_text, label, normalized_value, schema_field, event_id, context_prefix
SPANS = [
    # Event 1: Electrocautery (Subglottic)
    {
        "span_text": "Subglottic",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Subglottic",
        "event_id": "ev1",
        "context_prefix": "Endobronchial obstruction at "
    },
    {
        "span_text": "Electrocautery",
        "label": "PROC_METHOD",
        "normalized_value": "Electrocautery",
        "event_id": "ev1",
        "context_prefix": "Results\n"
    },
    {
        "span_text": "Knife (needle)",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "Knife (needle)",
        "event_id": "ev1",
        "context_prefix": "Electrocautery\t"
    },
    {
        "span_text": "50% patent",
        "label": "OUTCOME_AIRWAY_LUMEN_PRE",
        "normalized_value": "50",
        "event_id": "ev1",
        "context_prefix": "Prior to treatment, affected airway was note to be "
    },
    {
        "span_text": "50% patent",
        "label": "OUTCOME_AIRWAY_LUMEN_POST",
        "normalized_value": "50",
        "event_id": "ev1",
        "context_prefix": "After treatment, the airway was "
    },
    
    # Event 2: Balloon Dilation
    {
        "span_text": "Balloon dilation",
        "label": "PROC_METHOD",
        "normalized_value": "Balloon dilation",
        "event_id": "ev2",
        "context_prefix": "patent. \n\n"
    },
    {
        "span_text": "Trachea (Proximal 1/3)",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Trachea (Proximal 1/3)",
        "event_id": "ev2",
        "context_prefix": "Balloon dilation was performed at "
    },
    {
        "span_text": "8/9/10 Elation balloon",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "Elation balloon 8-9-10mm",
        "event_id": "ev2",
        "context_prefix": ""
    },
    {
        "span_text": "dilation to 10 mm",
        "label": "MEAS_AIRWAY_DIAM",
        "normalized_value": "10",
        "event_id": "ev2",
        "context_prefix": "perform "
    },
    {
        "span_text": "Total 3 inflations",
        "label": "MEAS_COUNT",
        "normalized_value": "3",
        "event_id": "ev2",
        "context_prefix": ""
    },

    # Event 3: Excision (Forceps)
    {
        "span_text": "Endobronchial excision",
        "label": "PROC_METHOD",
        "normalized_value": "Endobronchial excision",
        "event_id": "ev3",
        "context_prefix": "each.\n\n"
    },
    {
        "span_text": "Trachea (Proximal 1/3)",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Trachea (Proximal 1/3)",
        "event_id": "ev3",
        "context_prefix": "Endobronchial excision of tissue was performed at "
    },
    {
        "span_text": "forceps",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "forceps",
        "event_id": "ev3",
        "context_prefix": "peeled off with "
    },
    {
        "span_text": "Lesion was successfully removed",
        "label": "OUTCOME_SYMPTOMS", # Using symptoms or general outcome as proxy for success if no specific lumen %
        "normalized_value": "successfully removed",
        "event_id": "ev3",
        "context_prefix": ""
    },

    # Event 4: BAL
    {
        "span_text": "Bronchial alveolar lavage",
        "label": "PROC_METHOD",
        "normalized_value": "Bronchial alveolar lavage",
        "event_id": "ev4",
        "context_prefix": ""
    },
    {
        "span_text": "Lateral Segment of RML",
        "label": "ANAT_LUNG_LOC",
        "normalized_value": "RML Lateral",
        "event_id": "ev4",
        "context_prefix": "performed at "
    },
    {
        "span_text": "Medial Segment of RML",
        "label": "ANAT_LUNG_LOC",
        "normalized_value": "RML Medial",
        "event_id": "ev4",
        "context_prefix": "(RB4) and "
    },
    
    # Event 5: Therapeutic Aspiration
    {
        "span_text": "Therapeutic aspiration",
        "label": "PROC_METHOD",
        "normalized_value": "Therapeutic aspiration",
        "event_id": "ev5",
        "context_prefix": "Successful "
    },
    {
        "span_text": "clean out the Trachea (Proximal 1/3), Trachea (Middle 1/3), Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , and Left Mainstem",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Trachea, RMS, BI, LMS",
        "event_id": "ev5",
        "context_prefix": "performed to "
    },

    # General / Outcome
    {
        "span_text": "no immediate complications",
        "label": "OUTCOME_COMPLICATION",
        "normalized_value": "none",
        "event_id": "ev_glob",
        "context_prefix": "There were "
    }
]

# -------------------------------------------------------------------------
# HELPER FUNCTIONS
# -------------------------------------------------------------------------

def locate_spans(note_text, span_list):
    """
    Locates spans in text, computing start/end/match_index.
    Returns list of dicts with populated offsets.
    """
    results = []
    
    for item in span_list:
        text = item["span_text"]
        prefix = item.get("context_prefix", "")
        
        # 1. Find all occurrences
        matches = [m.start() for m in re.finditer(re.escape(text), note_text)]
        
        selected_start = -1
        match_index = 0
        hydration_status = "ambiguous"
        
        if len(matches) == 1:
            selected_start = matches[0]
            match_index = 0
            hydration_status = "hydrated_unique"
        elif len(matches) > 1:
            # Try prefix match
            found = False
            if prefix:
                for idx, start in enumerate(matches):
                    # Check preceding 120 chars
                    search_window = note_text[max(0, start - 120):start]
                    if prefix in search_window:
                        selected_start = start
                        match_index = idx
                        hydration_status = "hydrated_prefix_window"
                        found = True
                        break
            
            if not found:
                # Default to first if no prefix or not found (fallback)
                selected_start = matches[0]
                match_index = 0
                hydration_status = "ambiguous_took_first"
        
        if selected_start != -1:
            item["start_char"] = selected_start
            item["end_char"] = selected_start + len(text)
            item["match_index"] = match_index
            item["hydration_status"] = hydration_status
            results.append(item)
        else:
            # Not found
            item["start_char"] = ""
            item["end_char"] = ""
            item["hydration_status"] = "missing"
            results.append(item)
            
    return results

def create_workbook():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        # Create a new workbook with required sheets if template not found (Safety net)
        wb = openpyxl.Workbook()
        for sheet_name in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
    return wb

def write_note_text(wb, text):
    ws = wb["Note_Text"]
    # Check if header exists, if not write it
    if ws.max_row == 1 and ws.cell(1,1).value is None:
        ws.append(["note_id", "source_file", "note_text"])
    
    ws.append([NOTE_ID, SOURCE_FILE, text])

def write_note_index(wb, flags):
    ws = wb["Note_Index"]
    # Header check
    headers = [
        "source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes",
        "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", "tbna_conventional",
        "linear_ebus", "radial_ebus", "navigational_bronchoscopy", "transbronchial_biopsy", "transbronchial_cryobiopsy",
        "therapeutic_aspiration", "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
        "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", "bronchial_thermoplasty",
        "whole_lung_lavage", "rigid_bronchoscopy", "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy",
        "pleurodesis", "pleural_biopsy", "fibrinolytic_therapy"
    ]
    
    if ws.max_row == 1 and ws.cell(1,1).value is None:
        ws.append(headers)
        
    row = [
        SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "ready", ""
    ]
    # Add flags in order
    flag_keys = headers[8:]
    for k in flag_keys:
        row.append(flags.get(k, 0))
    
    ws.append(row)

def write_spans_anchor(wb, spans):
    ws = wb["Span_Annotations"]
    headers = [
        "source_file", "note_id", "span_id", "section_type",
        "context_prefix", "span_text", "match_index",
        "start_char", "end_char", "span_len",
        "label", "normalized_value", "schema_field", "event_id",
        "is_negated", "is_historical", "time_anchor", "reviewer", "comments",
        "hydration_status"
    ]
    if ws.max_row == 1 and ws.cell(1,1).value is None:
        ws.append(headers)

    for idx, s in enumerate(spans):
        # Anchor first: start/end are blank
        row = [
            SOURCE_FILE, NOTE_ID, f"span_{idx+1}", "Procedure",
            s.get("context_prefix", ""), s["span_text"], s.get("match_index", 0),
            "", "", f'=LEN(F{ws.max_row + 1})', # Formula for len
            s["label"], s.get("normalized_value", ""), "", s.get("event_id", ""),
            "FALSE", "FALSE", "", "", "",
            "needs_hydration"
        ]
        ws.append(row)

def write_spans_hydrated(wb, spans):
    ws = wb["Span_Hydrated"]
    headers = [
        "source_file", "note_id", "span_id", "section_type",
        "context_prefix", "span_text", "match_index",
        "start_char", "end_char", "span_len",
        "label", "normalized_value", "schema_field", "event_id",
        "is_negated", "is_historical", "time_anchor", "reviewer", "comments",
        "hydration_status"
    ]
    if ws.max_row == 1 and ws.cell(1,1).value is None:
        ws.append(headers)

    for idx, s in enumerate(spans):
        row = [
            SOURCE_FILE, NOTE_ID, f"span_{idx+1}", "Procedure",
            s.get("context_prefix", ""), s["span_text"], s.get("match_index", 0),
            s.get("start_char", ""), s.get("end_char", ""), len(s["span_text"]),
            s["label"], s.get("normalized_value", ""), "", s.get("event_id", ""),
            "FALSE", "FALSE", "", "", "",
            s.get("hydration_status", "missing")
        ]
        ws.append(row)

def write_event_log(wb):
    ws = wb["Event_Log"]
    headers = [
        "source_file", "note_id", "event_id", "event_type", "method",
        "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements",
        "specimens", "findings", "is_historical", "reviewer", "comments",
        "device_size", "device_material",
        "outcome_airway_lumen_pre", "outcome_airway_lumen_post",
        "outcome_symptoms", "outcome_pleural", "outcome_complication"
    ]
    if ws.max_row == 1 and ws.cell(1,1).value is None:
        ws.append(headers)
    
    # Manually constructing event rows based on knowledge
    events = [
        {
            "event_id": "ev1", "event_type": "Therapeutic", "method": "Electrocautery", 
            "anatomy_target": "Subglottic", "device": "Knife (needle)",
            "outcome_airway_lumen_pre": "50", "outcome_airway_lumen_post": "50",
            "findings": "Significant narrowing at cord level"
        },
        {
            "event_id": "ev2", "event_type": "Therapeutic", "method": "Balloon dilation",
            "anatomy_target": "Trachea (Proximal 1/3)", "device": "Elation balloon 8/9/10",
            "measurements": "10 mm", "counts": "3 inflations"
        },
        {
            "event_id": "ev3", "event_type": "Therapeutic", "method": "Endobronchial excision",
            "anatomy_target": "Trachea (Proximal 1/3)", "device": "Forceps",
            "outcome_symptoms": "Lesion successfully removed",
            "specimens": "Microbiology, Cytology"
        },
        {
            "event_id": "ev4", "event_type": "Diagnostic", "method": "BAL",
            "anatomy_target": "RML Lateral/Medial",
            "specimens": "Cell Count, Micro, Cytology"
        },
        {
            "event_id": "ev5", "event_type": "Therapeutic", "method": "Therapeutic aspiration",
            "anatomy_target": "Trachea, RMS, BI, LMS",
            "findings": "Mucus and blood cleaned out"
        },
        {
            "event_id": "ev_glob", "event_type": "Global",
            "outcome_complication": "none"
        }
    ]

    for ev in events:
        row = [
            SOURCE_FILE, NOTE_ID, ev["event_id"], ev.get("event_type", ""), ev.get("method", ""),
            ev.get("anatomy_target", ""), ev.get("device", ""), "", "", ev.get("counts", ""), ev.get("measurements", ""),
            ev.get("specimens", ""), ev.get("findings", ""), "FALSE", "", "",
            "", "", # size/material
            ev.get("outcome_airway_lumen_pre", ""), ev.get("outcome_airway_lumen_post", ""),
            ev.get("outcome_symptoms", ""), "", ev.get("outcome_complication", "")
        ]
        ws.append(row)

def write_v3_procedure_events(wb):
    ws = wb["V3_Procedure_Events"]
    headers = [
        "note_id", "event_id", "type",
        "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station",
        "lesion.type", "lesion.size_mm",
        "method", "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote",
        "stent.size", "stent.material_or_brand", "catheter.size_fr",
        "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post",
        "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"
    ]
    if ws.max_row == 1 and ws.cell(1,1).value is None:
        ws.append(headers)

    # Simplified V3 mapping
    # Event 1: Electrocautery
    row1 = [
        NOTE_ID, "ev1", "therapeutic",
        "airway", "", "Subglottic", "",
        "stenosis", "",
        "Electrocautery", json.dumps(["Knife (needle)"]), "", "", json.dumps(["Significant narrowing"]), "",
        "", "", "",
        "50", "50",
        "", "", ""
    ]
    ws.append(row1)

    # Event 2: Balloon Dilation
    row2 = [
        NOTE_ID, "ev2", "therapeutic",
        "airway", "Trachea", "Proximal 1/3", "",
        "", "",
        "Balloon dilation", json.dumps(["Elation balloon"]), json.dumps({"target_diameter": 10, "unit": "mm"}), "", "", "",
        "", "", "",
        "", "",
        "", "", ""
    ]
    ws.append(row2)

    # Event 3: Excision
    row3 = [
        NOTE_ID, "ev3", "therapeutic",
        "airway", "Trachea", "Proximal 1/3", "",
        "lesion", "",
        "Excision", json.dumps(["Forceps"]), "", json.dumps(["Microbiology", "Cytology"]), "", "",
        "", "", "",
        "", "",
        "Lesion successfully removed", "", ""
    ]
    ws.append(row3)

    # Event 4: BAL
    row4 = [
        NOTE_ID, "ev4", "diagnostic",
        "airway", "RML", "Lateral/Medial", "",
        "", "",
        "BAL", "", "", json.dumps(["Cell Count", "Micro", "Cyto"]), "", "",
        "", "", "",
        "", "",
        "", "", ""
    ]
    ws.append(row4)

    # Event 5: Aspiration
    row5 = [
        NOTE_ID, "ev5", "therapeutic",
        "airway", "Trachea/Bronchi", "", "",
        "", "",
        "Therapeutic aspiration", "", "", "", json.dumps(["Mucus", "Blood"]), "",
        "", "", "",
        "", "",
        "", "", ""
    ]
    ws.append(row5)

    # Global Outcome
    row_glob = [
        NOTE_ID, "ev_glob", "outcome",
        "", "", "", "",
        "", "",
        "", "", "", "", "", "",
        "", "", "",
        "", "",
        "", "", "none"
    ]
    ws.append(row_glob)

def write_registry_json(wb):
    ws = wb["V3_Registry_JSON"]
    if ws.max_row == 1 and ws.cell(1,1).value is None:
        ws.append(["json_output"])
    
    # Construct a valid JSON object
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": [
            {
                "event_id": "ev1",
                "type": "therapeutic",
                "method": "Electrocautery",
                "target": {"anatomy_type": "airway", "location": "Subglottic"},
                "outcomes": {"airway": {"lumen_pre": 50, "lumen_post": 50}}
            },
            {
                "event_id": "ev2",
                "type": "therapeutic",
                "method": "Balloon dilation",
                "target": {"anatomy_type": "airway", "location": "Trachea (Proximal 1/3)"},
                "measurements": {"diameter_mm": 10}
            },
            {
                "event_id": "ev3",
                "type": "therapeutic",
                "method": "Excision",
                "target": {"anatomy_type": "airway", "location": "Trachea (Proximal 1/3)"},
                "outcomes": {"symptoms": "Lesion successfully removed"}
            },
            {
                "event_id": "ev4",
                "type": "diagnostic",
                "method": "BAL",
                "target": {"anatomy_type": "airway", "location": "RML"}
            }
        ],
        "no_immediate_complications": True
    }
    
    ws.append([json.dumps(registry_data, indent=2)])

# -------------------------------------------------------------------------
# MAIN EXECUTION
# -------------------------------------------------------------------------

def main():
    # 1. Hydrate spans
    hydrated_spans = locate_spans(NOTE_TEXT, SPANS)
    
    # 2. Create workbook
    wb = create_workbook()
    
    # 3. Populate sheets
    write_note_text(wb, NOTE_TEXT)
    write_note_index(wb, PROCEDURE_FLAGS)
    write_spans_anchor(wb, hydrated_spans)
    write_spans_hydrated(wb, hydrated_spans)
    write_event_log(wb)
    write_v3_procedure_events(wb)
    write_registry_json(wb)
    
    # 4. Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    main()