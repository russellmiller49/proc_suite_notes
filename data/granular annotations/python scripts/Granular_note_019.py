import re
import json
import datetime
from openpyxl import load_workbook, Workbook

# ==========================================
# INPUT DATA
# ==========================================
NOTE_ID = "note_019"
SOURCE_FILE = "note_019.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_019 SOURCE_FILE: note_019.txt INDICATION FOR OPERATION:  [REDACTED]is a 87 year old-year-old male who presents with Complicated Effusion.
The nature, purpose, risks, benefits and alternatives to Chest Ultrasound and Instillation of agents for fibrinolysis were discussed with the patient in detail.
Patient indicated a wish to proceed with procedure and informed consent was signed.
 
CONSENT : Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient or surrogate. The patient or surrogate provided consent.
PREOPERATIVE DIAGNOSIS:  Complicated Effusion
POSTOPERATIVE DIAGNOSIS: Same as preoperative diagnosis - see above.
PROCEDURE:  
76604 Ultrasound, chest (includes mediastinum), real time with image documentation
32561 Instillation(s), via chest tube/catheter, agent for fibrinolysis (eg, fibrinolytic agent for break up of multiloculated effusion);
initial day
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
PROCEDURE IN DETAIL:
 
PATIENT POSITION:  Sitting
 
LEFT CHEST ULTRASOUND PROCEDURE:
FINDINGS:
Hemithorax:  Left
Pleural Effusion: 
Volume:  Minimal
Echogenicity:  Anechoic
Loculations:  None
Diaphragmatic Motion:  Normal
Lung: 
Lung sliding before procedure:  Present
Lung sliding post procedure:  Present
Lung consolidation/atelectasis:  Absent
Pleura:  Normal
Image(s) saved and uploaded:  yes
 
RIGHT CHEST ULTRASOUND PROCEDURE:
FINDINGS:
Hemithorax:  Right
Pleural Effusion: 
Volume:  Large
Echogenicity:  Anechoic
Loculations:  Thick
Diaphragmatic Motion:  Diminished
Lung: 
Pleura:  Thick
Image(s) saved and uploaded:  yes
 
Date of chest tube insertion: 9/24/25
 
Side: RIGHT
1‌  10 mg tPA / 5 mg Dnasedose #:__1__ 
            0‌ 
 ___mg tPA                              dose #:____ 
0‌  Other medication:  
 
COMPLICATIONS:
1‌None 0‌Bleeding-EBL: ___ ml 0‌Pneumothorax 0‌Re- Expansion Pulmonary Edema 
0‌Other: 
 
IMPRESSION/PLAN: [REDACTED]is a 87 year old-year-old male who presents for Chest Ultrasound and Instillation of agents for fibrinolysis.
The patient tolerated the procedure well.  There were no immediate complications.
DISPOSITION: Home
 
-1hr dwell of lytics followed by drainage with vacutainer- 400cc bloody output drained
- pleural clinic in 10-14days for repeat lytics"""

# ==========================================
# CONFIGURATION
# ==========================================
# Flags for Note_Index
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 0, "bal": 0, "bronchial_wash": 0, "brushings": 0,
    "endobronchial_biopsy": 0, "tbna_conventional": 0, "linear_ebus": 0, "radial_ebus": 0,
    "navigational_bronchoscopy": 0, "transbronchial_biopsy": 0, "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0, "foreign_body_removal": 0, "airway_dilation": 0,
    "airway_stent": 0, "thermal_ablation": 0, "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0, "blvr": 0, "peripheral_ablation": 0, "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0, "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0, "chest_tube": 0, "ipc": 0, "medical_thoracoscopy": 0,
    "pleurodesis": 0, "pleural_biopsy": 0, "fibrinolytic_therapy": 1
}

# -------------------------------------------------------------------------
# SPAN DATA (Anchor-First)
# Structure: (span_text, label, normalized_value, context_prefix, event_id, section)
# -------------------------------------------------------------------------
# Event 1: Left Ultrasound
# Event 2: Right Ultrasound
# Event 3: Fibrinolysis
RAW_SPANS = [
    # Event 1: Left US
    ("Left", "LATERALITY", "Left", "Hemithorax:  ", "ev1", "LEFT CHEST ULTRASOUND PROCEDURE"),
    ("Minimal", "OBS_LESION", "Minimal", "Volume:  ", "ev1", "LEFT CHEST ULTRASOUND PROCEDURE"),
    ("Anechoic", "OBS_LESION", "Anechoic", "Echogenicity:  ", "ev1", "LEFT CHEST ULTRASOUND PROCEDURE"),
    ("None", "OBS_LESION", "No Loculations", "Loculations:  ", "ev1", "LEFT CHEST ULTRASOUND PROCEDURE"),
    
    # Event 2: Right US
    ("Right", "LATERALITY", "Right", "Hemithorax:  ", "ev2", "RIGHT CHEST ULTRASOUND PROCEDURE"),
    ("Large", "OBS_LESION", "Large", "Volume:  ", "ev2", "RIGHT CHEST ULTRASOUND PROCEDURE"),
    ("Anechoic", "OBS_LESION", "Anechoic", "Echogenicity:  ", "ev2", "RIGHT CHEST ULTRASOUND PROCEDURE"),
    ("Thick", "OBS_LESION", "Thick Loculations", "Loculations:  ", "ev2", "RIGHT CHEST ULTRASOUND PROCEDURE"),
    ("Diminished", "OBS_LESION", "Diminished Diaphragmatic Motion", "Diaphragmatic Motion:  ", "ev2", "RIGHT CHEST ULTRASOUND PROCEDURE"),
    ("Thick", "OBS_LESION", "Thick Pleura", "Pleura:  ", "ev2", "RIGHT CHEST ULTRASOUND PROCEDURE"),

    # Event 3: Fibrinolysis (Right)
    ("Instillation of agents for fibrinolysis", "PROC_METHOD", "Fibrinolysis", "alternatives to Chest Ultrasound and ", "ev3", "HEADER"),
    ("Instillation(s), via chest tube/catheter", "PROC_METHOD", "Fibrinolysis via Chest Tube", "32561 ", "ev3", "PROCEDURE"),
    ("9/24/25", "CTX_TIME", "2025-09-24", "Date of chest tube insertion: ", "ev3", "HISTORY"),
    ("chest tube insertion", "CTX_HISTORICAL", "Chest Tube Insertion", "Date of ", "ev3", "HISTORY"),
    ("RIGHT", "LATERALITY", "Right", "Side: ", "ev3", "PROCEDURE_DETAIL"),
    ("10 mg tPA", "PROC_METHOD", "tPA 10mg", "1‌  ", "ev3", "PROCEDURE_DETAIL"),
    ("5 mg Dnase", "PROC_METHOD", "Dnase 5mg", "/ ", "ev3", "PROCEDURE_DETAIL"),
    ("400cc", "MEAS_PLEURAL_DRAIN", "400", "drainage with vacutainer- ", "ev3", "DISPOSITION"),
    ("bloody output", "OUTCOME_PLEURAL", "bloody output", "400cc ", "ev3", "DISPOSITION"),
    ("None", "OUTCOME_COMPLICATION", "None", "1‌", "ev3", "COMPLICATIONS"),
    ("no immediate complications", "OUTCOME_COMPLICATION", "None", "There were ", "ev3", "IMPRESSION/PLAN")
]

# ==========================================
# HYDRATION LOGIC
# ==========================================
def find_offsets(text, span_text, context_prefix=None):
    """
    Returns (start, end, hydration_status)
    Status: unique, prefix_match, ambiguous_count=N, not_found
    """
    matches = [m for m in re.finditer(re.escape(span_text), text)]
    
    if not matches:
        return "", "", "not_found"
    
    if len(matches) == 1:
        return matches[0].start(), matches[0].end(), "hydrated_unique"
    
    # Context match
    if context_prefix:
        # Check preceding 120 chars for context
        best_match = None
        for m in matches:
            window_start = max(0, m.start() - 120)
            preceding_text = text[window_start:m.start()]
            if context_prefix in preceding_text:
                best_match = m
                break
        if best_match:
            return best_match.start(), best_match.end(), "hydrated_prefix_window"

    # If simple logic fails, fallback to first occurrence if no better signal, but mark ambiguous
    # For Phase 0 script, we'll implement a simple match index finder if specified in list, 
    # but since we only have list, we default to first or specific known distinct contexts.
    
    # Fallback: if we have multiple matches and context didn't resolve, 
    # check if the span list provided a specific intent (not implemented in simple tuple).
    # We will return the first one but mark ambiguous.
    return matches[0].start(), matches[0].end(), f"ambiguous_count={len(matches)}"

# ==========================================
# MAIN EXECUTION
# ==========================================
def generate_workbook():
    try:
        wb = load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        # Create a skeleton if template missing (fallback, though instructions say template required)
        wb = Workbook()
        for name in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            if name not in wb.sheetnames:
                wb.create_sheet(name)

    # 1. Note_Text
    ws_text = wb["Note_Text"]
    if ws_text.max_row == 1 and ws_text["A1"].value is None:
         ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws_index = wb["Note_Index"]
    # Columns: source_file, note_id, encounter_id, procedure_date, site, reviewer, status, free_text_notes, [flags...]
    # Map flags to list
    flag_keys = list(PROCEDURE_FLAGS.keys())
    flag_values = [PROCEDURE_FLAGS[k] for k in flag_keys]
    
    row_meta = [SOURCE_FILE, NOTE_ID, "", "", "", "", "Pending", ""]
    ws_index.append(row_meta + flag_values)

    # 3. Span_Annotations & 4. Span_Hydrated
    ws_span = wb["Span_Annotations"]
    ws_hydrated = wb["Span_Hydrated"]
    
    # Headers check (assuming template has them, but for safety in logic)
    # Just appending rows
    
    hydrated_rows = []
    
    for span_data in RAW_SPANS:
        text_val, label, norm, ctx, ev_id, section = span_data
        
        # Calculate offsets
        start, end, status = find_offsets(NOTE_TEXT, text_val, ctx)
        span_len = len(text_val)
        
        # Common fields
        # source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, 
        # start, end, len, label, normalized, schema, event_id, negated, historical, time, reviewer, comments, hydration
        
        # Determine specialized flags
        is_hist = "1" if label == "CTX_HISTORICAL" else "0"
        is_neg = "0" # Simple default
        
        # Schema field mapping (approximate for display)
        schema_map = {
            "LATERALITY": "target.location",
            "OBS_LESION": "findings",
            "PROC_METHOD": "method",
            "MEAS_PLEURAL_DRAIN": "outcomes.pleural_drain",
            "OUTCOME_PLEURAL": "outcomes.pleural",
            "OUTCOME_COMPLICATION": "outcomes.complications",
            "CTX_TIME": "metadata.date",
            "CTX_HISTORICAL": "metadata.history"
        }
        schema_field = schema_map.get(label, "other")

        # Row for Span_Annotations (blank start/end)
        row_anno = [
            SOURCE_FILE, NOTE_ID, f"{ev_id}_{text_val[:10]}", section,
            ctx, text_val, "", # match_index
            "", "", span_len, # start, end blank
            label, norm, schema_field, ev_id,
            is_neg, is_hist, "", "", "", "needs_hydration"
        ]
        ws_span.append(row_anno)
        
        # Row for Span_Hydrated
        row_hyd = [
            SOURCE_FILE, NOTE_ID, f"{ev_id}_{text_val[:10]}", section,
            ctx, text_val, "", 
            start, end, span_len,
            label, norm, schema_field, ev_id,
            is_neg, is_hist, "", "", "", status
        ]
        ws_hydrated.append(row_hyd)
        
        hydrated_rows.append({
            "event_id": ev_id,
            "label": label,
            "norm": norm,
            "text": text_val
        })

    # 5. Event_Log (Granular)
    ws_event = wb["Event_Log"]
    # Group by event_id
    events_map = {}
    for r in hydrated_rows:
        eid = r["event_id"]
        if eid not in events_map:
            events_map[eid] = {
                "type": "Procedure" if "ev3" in eid else "Diagnostic",
                "method": [],
                "findings": [],
                "outcome_pleural": [],
                "outcome_complication": [],
                "measurements": [],
                "lateral": []
            }
        
        if r["label"] == "PROC_METHOD":
            events_map[eid]["method"].append(r["norm"])
        elif r["label"] == "OBS_LESION":
            events_map[eid]["findings"].append(r["norm"])
        elif r["label"] == "LATERALITY":
            events_map[eid]["lateral"].append(r["norm"])
        elif r["label"] == "OUTCOME_PLEURAL":
            events_map[eid]["outcome_pleural"].append(r["norm"])
        elif r["label"] == "OUTCOME_COMPLICATION":
            events_map[eid]["outcome_complication"].append(r["norm"])
        elif r["label"] == "MEAS_PLEURAL_DRAIN":
            events_map[eid]["measurements"].append(f"Drain: {r['norm']}")

    for eid, data in events_map.items():
        # Flatten
        lat = data["lateral"][0] if data["lateral"] else ""
        method = ", ".join(data["method"])
        findings = ", ".join(data["findings"])
        out_pleural = ", ".join(data["outcome_pleural"])
        out_comp = ", ".join(data["outcome_complication"])
        meas = ", ".join(data["measurements"])
        
        etype = "Ultrasound" if "ev1" in eid or "ev2" in eid else "Fibrinolysis"
        
        # source_file, note_id, event_id, event_type, method, anatomy, device...
        # Map to columns by index or just append assuming order
        # Columns: source_file, note_id, event_id, event_type, method, anatomy_target, device, needle_gauge, stations, counts, measurements, specimens, findings, is_historical, reviewer, comments, device_size, device_material, outcome_airway_lumen_pre, outcome_airway_lumen_post, outcome_symptoms, outcome_pleural, outcome_complication
        
        row_log = [
            SOURCE_FILE, NOTE_ID, eid, etype, method,
            lat, "", "", "", "", meas,
            "", findings, "0", "", "",
            "", "", # device size/mat
            "", "", "", out_pleural, out_comp
        ]
        ws_event.append(row_log)

    # 6. V3_Procedure_Events & 7. V3_Registry_JSON
    # (Simplified for this generator, just outputting JSON structure and mapping basic rows)
    ws_v3 = wb["V3_Procedure_Events"]
    
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True, # Based on text
        "procedures": []
    }
    
    for eid, data in events_map.items():
        # Create V3 entry
        proc_obj = {
            "event_id": eid,
            "type": "Ultrasound" if "ev1" in eid or "ev2" in eid else "Fibrinolysis",
            "target": {
                "anatomy_type": "Hemithorax",
                "location": {"laterality": data["lateral"][0] if data["lateral"] else "Unspecified"}
            },
            "method": data["method"],
            "outcomes": {
                "pleural": data["outcome_pleural"],
                "complications": data["outcome_complication"]
            }
        }
        registry_data["procedures"].append(proc_obj)
        
        # Append to V3 sheet (sparse)
        # note_id, event_id, type, target.anatomy_type, target.location.lobe, target.location.segment, target.station...
        ws_v3.append([
            NOTE_ID, eid, proc_obj["type"], "Hemithorax", 
            proc_obj["target"]["location"].get("laterality", ""), "", "",
            "", "", # lesion
            str(data["method"]), "", str(data["measurements"]), "", str(data["findings"]), "",
            "", "", "", # stent/cath
            "", "", "", str(data["outcome_pleural"]), str(data["outcome_complication"])
        ])

    ws_json = wb["V3_Registry_JSON"]
    ws_json.append([json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()