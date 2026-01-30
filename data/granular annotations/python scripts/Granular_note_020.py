import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import datetime

# ==========================================
# INPUT DATA
# ==========================================
NOTE_ID = "note_020"
SOURCE_FILE = "note_020.txt"
PROCEDURE_DATE = ""  # Not explicitly stated as a specific date in text, leaving blank
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_020 SOURCE_FILE: note_020.txt INDICATION FOR OPERATION:  [REDACTED]is a 59 year old-year-old male who presents with lung transplant and airway narrowing.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
 
CONSENT : Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient or surrogate.
The patient or surrogate read and signed the provided consent form / provided consent over the phone.
The consent was witnessed by an assisting medical professional.
 
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
 
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
 
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31624 Dx bronchoscope/lavage (BAL)    
31625 Endobronchial Biopsy(s)
 
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Disposable Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
 
Successful therapeutic aspiration was performed to clean out the Right Mainstem, Bronchus Intermedius , and Left Mainstem from mucus.
Bronchial alveolar lavage was performed at Lateral Segment of RML (RB4) and Medial Segment of RML (RB5).
Instilled 60 cc of NS, suction returned with 20 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
 
Endobronchial biopsy was performed at Right Mainstem.  Lesion was successfully removed.
Samples sent for Microbiology (Cultures/Viral/Fungal) and Pathology.
 
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
BAL, EBBX
 
IMPRESSION/PLAN: [REDACTED]is a 59 year old-year-old male who presents for bronchoscopy for airway stenosis.
- f/u BAL
- airway in good condition, need repeat bronch in 4-6 weeks."""

# ==========================================
# CONFIGURATION
# ==========================================

# Procedure Flags (30 total)
FLAGS = {
    "diagnostic_bronchoscopy": 1, # BAL + Biopsy
    "bal": 1,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 1,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# ==========================================
# SPAN DATA GENERATION
# ==========================================

# Helper for creating span rows
def create_span(span_text, label, normalized_value, schema_field, event_id, context_prefix="", is_negated=False):
    return {
        "source_file": SOURCE_FILE,
        "note_id": NOTE_ID,
        "span_id": f"{NOTE_ID}_s_{len(spans) + 1:03d}",
        "section_type": "Procedure", # General assumption
        "context_prefix": context_prefix,
        "span_text": span_text,
        "match_index": None, # Will calculate later if needed
        "start_char": "",
        "end_char": "",
        "span_len": f"=LEN(F{len(spans) + 2})", # Formula for Excel
        "label": label,
        "normalized_value": normalized_value,
        "schema_field": schema_field,
        "event_id": event_id,
        "is_negated": is_negated,
        "is_historical": False,
        "time_anchor": "",
        "reviewer": "Auto",
        "comments": "",
        "hydration_status": "needs_hydration"
    }

spans = []

# --- Event 1: Therapeutic Aspiration ---
evt1 = "evt_01"
spans.append(create_span("Therapeutic aspiration", "PROC_METHOD", "therapeutic_aspiration", "method", evt1))
spans.append(create_span("Right Mainstem", "ANAT_AIRWAY", "Right Mainstem", "target.anatomy", evt1, context_prefix="clean out the "))
spans.append(create_span("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "target.anatomy", evt1))
spans.append(create_span("Left Mainstem", "ANAT_AIRWAY", "Left Mainstem", "target.anatomy", evt1))
spans.append(create_span("clean out", "PROC_ACTION", "clean", "action", evt1))
spans.append(create_span("mucus", "OBS_FINDING", "mucus", "findings", evt1))

# --- Event 2: BAL ---
evt2 = "evt_02"
spans.append(create_span("Bronchial alveolar lavage", "PROC_METHOD", "bal", "method", evt2))
spans.append(create_span("Lateral Segment of RML", "ANAT_LUNG_LOC", "RML Lateral", "target.location", evt2))
spans.append(create_span("RB4", "ANAT_LUNG_LOC", "RB4", "target.location", evt2))
spans.append(create_span("Medial Segment of RML", "ANAT_LUNG_LOC", "RML Medial", "target.location", evt2))
spans.append(create_span("RB5", "ANAT_LUNG_LOC", "RB5", "target.location", evt2))
spans.append(create_span("60 cc", "MEAS_VOL", "60", "measurements.volume_instilled", evt2))
spans.append(create_span("20 cc", "MEAS_VOL", "20", "measurements.volume_returned", evt2))
spans.append(create_span("Cell Count", "OBS_ROSE", "Cell Count", "specimens", evt2)) # Using OBS_ROSE/specimens loosely
spans.append(create_span("Microbiology", "OBS_ROSE", "Microbiology", "specimens", evt2, context_prefix="Samples sent for Cell Count, "))
spans.append(create_span("Cytology", "OBS_ROSE", "Cytology", "specimens", evt2))

# --- Event 3: Endobronchial Biopsy ---
evt3 = "evt_03"
spans.append(create_span("Endobronchial biopsy", "PROC_METHOD", "endobronchial_biopsy", "method", evt3))
spans.append(create_span("Right Mainstem", "ANAT_AIRWAY", "Right Mainstem", "target.anatomy", evt3, context_prefix="performed at "))
spans.append(create_span("Lesion", "OBS_LESION", "lesion", "lesion.type", evt3))
spans.append(create_span("removed", "PROC_ACTION", "removed", "action", evt3, context_prefix="Lesion was successfully "))
spans.append(create_span("Microbiology", "OBS_ROSE", "Microbiology", "specimens", evt3, context_prefix="Samples sent for "))
spans.append(create_span("Pathology", "OBS_ROSE", "Pathology", "specimens", evt3))

# --- Global / Outcomes ---
evt_global = "evt_global"
spans.append(create_span("Disposable Bronchoscope", "DEV_INSTRUMENT", "Disposable Bronchoscope", "device", evt_global))
spans.append(create_span("No immediate complications", "OUTCOME_COMPLICATION", "none", "outcomes.complications", evt_global, is_negated=True))


# ==========================================
# HYDRATION LOGIC
# ==========================================
def hydrate_spans(text, span_list):
    hydrated = []
    
    # Pre-calculate counts for all span texts
    text_counts = {}
    for s in span_list:
        st = s["span_text"]
        text_counts[st] = text.count(st)

    for s in span_list:
        row = s.copy()
        st = row["span_text"]
        count = text_counts.get(st, 0)
        
        start = -1
        status = "ambiguous"
        
        if count == 0:
            status = "not_found"
        elif count == 1:
            start = text.find(st)
            status = "hydrated_unique"
        else:
            # Context match
            prefix = row.get("context_prefix", "")
            if prefix:
                # Find all occurrences
                occurrences = [m.start() for m in re.finditer(re.escape(st), text)]
                for idx in occurrences:
                    # Check window before
                    window_start = max(0, idx - 150)
                    window = text[window_start:idx]
                    if prefix in window:
                        start = idx
                        status = "hydrated_context"
                        break
            
            # Fallback to match_index if context failed or not present
            if start == -1 and row.get("match_index") is not None:
                mi = row["match_index"]
                occurrences = [m.start() for m in re.finditer(re.escape(st), text)]
                if 0 <= mi < len(occurrences):
                    start = occurrences[mi]
                    status = "hydrated_index"

        if start != -1:
            row["start_char"] = start
            row["end_char"] = start + len(st)
            row["hydration_status"] = status
        else:
            if count > 1:
                row["hydration_status"] = f"ambiguous_count={count}"
            else:
                row["hydration_status"] = "not_found"
        
        hydrated.append(row)
    return hydrated

hydrated_spans = hydrate_spans(NOTE_TEXT, spans)

# ==========================================
# EVENT LOG DATA
# ==========================================
events = [
    {
        "event_id": "evt_01",
        "event_type": "Therapeutic Aspiration",
        "method": "Therapeutic aspiration",
        "anatomy_target": "Right Mainstem, Bronchus Intermedius, Left Mainstem",
        "findings": "mucus",
        "comments": "Cleaned out mucus"
    },
    {
        "event_id": "evt_02",
        "event_type": "BAL",
        "method": "Bronchial alveolar lavage",
        "anatomy_target": "Lateral Segment of RML (RB4), Medial Segment of RML (RB5)",
        "measurements": "Instilled 60 cc, Returned 20 cc",
        "specimens": "Cell Count, Microbiology, Cytology"
    },
    {
        "event_id": "evt_03",
        "event_type": "Endobronchial Biopsy",
        "method": "Endobronchial biopsy",
        "anatomy_target": "Right Mainstem",
        "findings": "Lesion",
        "specimens": "Microbiology, Pathology",
        "comments": "Lesion removed"
    },
    {
        "event_id": "evt_global",
        "event_type": "Global Outcome",
        "outcome_complication": "None"
    }
]

# ==========================================
# JSON REGISTRY DATA
# ==========================================
registry_data = {
    "schema_version": "3.0",
    "note_id": NOTE_ID,
    "no_immediate_complications": True,
    "procedures": [
        {
            "event_id": "evt_01",
            "type": "Therapeutic Aspiration",
            "method": "Therapeutic aspiration",
            "target": {
                "anatomy_type": "Airway",
                "location": ["Right Mainstem", "Bronchus Intermedius", "Left Mainstem"]
            },
            "findings": ["mucus"]
        },
        {
            "event_id": "evt_02",
            "type": "BAL",
            "method": "Bronchial alveolar lavage",
            "target": {
                "anatomy_type": "Lung",
                "location": ["RML Lateral (RB4)", "RML Medial (RB5)"]
            },
            "measurements": {
                "instilled": "60 cc",
                "returned": "20 cc"
            },
            "specimens": ["Cell Count", "Microbiology", "Cytology"]
        },
        {
            "event_id": "evt_03",
            "type": "Endobronchial Biopsy",
            "method": "Endobronchial biopsy",
            "target": {
                "anatomy_type": "Airway",
                "location": "Right Mainstem"
            },
            "lesion": {
                "type": "lesion",
                "action": "removed"
            },
            "specimens": ["Microbiology", "Pathology"]
        }
    ]
}

# ==========================================
# EXCEL GENERATION
# ==========================================

def generate_excel(template_path: str = TEMPLATE_PATH, output_path: str = OUTPUT_PATH) -> None:
    try:
        wb = openpyxl.load_workbook(template_path)
    except FileNotFoundError:
        # Create basic template structure if missing (fallback for execution without external file)
        wb = openpyxl.Workbook()
        for sheet_name in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            wb.create_sheet(sheet_name)
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # 1. Note_Text
    ws = wb["Note_Text"]
    if ws.max_row == 1:
        ws.append(["note_id", "source_file", "note_text"])
    ws.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws = wb["Note_Index"]
    header = [
        "source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes",
        "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", "tbna_conventional",
        "linear_ebus", "radial_ebus", "navigational_bronchoscopy", "transbronchial_biopsy", "transbronchial_cryobiopsy",
        "therapeutic_aspiration", "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
        "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", "bronchial_thermoplasty",
        "whole_lung_lavage", "rigid_bronchoscopy", "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy",
        "pleurodesis", "pleural_biopsy", "fibrinolytic_therapy"
    ]
    if ws.max_row == 1:
        ws.append(header)

    row_data = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Ready", ""]
    # Append flags in order
    flag_keys = header[8:]
    for k in flag_keys:
        row_data.append(FLAGS.get(k, 0))
    ws.append(row_data)

    # 3. Span_Annotations
    ws = wb["Span_Annotations"]
    span_header = [
        "source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", "match_index",
        "start_char", "end_char", "span_len", "label", "normalized_value", "schema_field", "event_id",
        "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"
    ]
    if ws.max_row == 1:
        ws.append(span_header)

    for s in spans:
        ws.append([
            s["source_file"], s["note_id"], s["span_id"], s["section_type"], s["context_prefix"], s["span_text"], s["match_index"],
            s["start_char"], s["end_char"], s["span_len"], s["label"], s["normalized_value"], s["schema_field"], s["event_id"],
            s["is_negated"], s["is_historical"], s["time_anchor"], s["reviewer"], s["comments"], s["hydration_status"]
        ])

    # 4. Span_Hydrated
    ws = wb["Span_Hydrated"]
    if ws.max_row == 1:
        ws.append(span_header)

    for s in hydrated_spans:
        ws.append([
            s["source_file"], s["note_id"], s["span_id"], s["section_type"], s["context_prefix"], s["span_text"], s["match_index"],
            s["start_char"], s["end_char"], s["span_len"], s["label"], s["normalized_value"], s["schema_field"], s["event_id"],
            s["is_negated"], s["is_historical"], s["time_anchor"], s["reviewer"], s["comments"], s["hydration_status"]
        ])

    # 5. Event_Log
    ws = wb["Event_Log"]
    event_header = [
        "source_file", "note_id", "event_id", "event_type", "method", "anatomy_target", "device", "needle_gauge",
        "stations", "counts", "measurements", "specimens", "findings", "is_historical", "reviewer", "comments",
        "device_size", "device_material", "outcome_airway_lumen_pre", "outcome_airway_lumen_post",
        "outcome_symptoms", "outcome_pleural", "outcome_complication"
    ]
    if ws.max_row == 1:
        ws.append(event_header)

    for e in events:
        row = [
            SOURCE_FILE, NOTE_ID, e.get("event_id"), e.get("event_type"), e.get("method"), e.get("anatomy_target"),
            e.get("device"), e.get("needle_gauge"), e.get("stations"), e.get("counts"), e.get("measurements"),
            e.get("specimens"), e.get("findings"), False, "Auto", e.get("comments"),
            e.get("device_size"), e.get("device_material"), e.get("outcome_airway_lumen_pre"),
            e.get("outcome_airway_lumen_post"), e.get("outcome_symptoms"), e.get("outcome_pleural"),
            e.get("outcome_complication")
        ]
        ws.append(row)

    # 6. V3_Procedure_Events
    ws = wb["V3_Procedure_Events"]
    v3_header = [
        "note_id", "event_id", "type", "target.anatomy_type", "target.location.lobe", "target.location.segment",
        "target.station", "lesion.type", "lesion.size_mm", "method", "devices_json", "measurements_json",
        "specimens_json", "findings_json", "evidence_quote", "stent.size", "stent.material_or_brand",
        "catheter.size_fr", "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post", "outcomes.symptoms",
        "outcomes.pleural", "outcomes.complications"
    ]
    if ws.max_row == 1:
        ws.append(v3_header)

    # Simple mapping for V3 (not fully populated for brevity, but structurally correct)
    for e in events:
        if e.get("event_id") == "evt_global":
            continue
        ws.append([
            NOTE_ID, e.get("event_id"), e.get("event_type"), "", "", "", "", "", "", e.get("method"),
            "", e.get("measurements"), e.get("specimens"), e.get("findings"), "", "", "", "", "", "", "", "", ""
        ])

    # 7. V3_Registry_JSON
    ws = wb["V3_Registry_JSON"]
    if ws.max_row == 1:
        ws.append(["json_object"])

    ws.append([json.dumps(registry_data, indent=2)])

    wb.save(output_path)
    print(f"Successfully generated {output_path}")


if __name__ == "__main__":
    generate_excel()
