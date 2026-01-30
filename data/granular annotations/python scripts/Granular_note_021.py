import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import re
import json
import os
from datetime import datetime

# =============================================================================
# 1. INPUT DATA
# =============================================================================

NOTE_ID = "note_021"
SOURCE_FILE = "note_021.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_021 SOURCE_FILE: note_021.txt INDICATION FOR OPERATION:  [REDACTED]is a 68 year old-year-old male who presents with persistent airleak.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
 
CONSENT : Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient or surrogate.
The patient or surrogate read and signed the provided consent form / provided consent over the phone.
The consent was witnessed by an assisting medical professional.
 
PREOPERATIVE DIAGNOSIS: J93.82 Other airleaks
 
POSTOPERATIVE DIAGNOSIS:  J93.82 Other airleaks
 
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31624 Dx bronchoscope/lavage (BAL)    
31634 Balloon occlusion or placement of occlusive substance 
31635 Foreign body removal
31647 Bronchial valve insert initial lobe 
 
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)  31624 BAL done in multiple lobes.
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
Flexible Hybrid (Pedatric) Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
 
Successful therapeutic aspiration was performed to clean out the Right Mainstem, Bronchus Intermedius , and Left Mainstem from mucus and mucus plug.
Bronchial alveolar lavage was performed at Superior Segment of Lingula (LB4) and Inferior Segment of Lingula (LB5).
Instilled 60 cc of NS, suction returned with 15 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
 
Bronchial alveolar lavage was performed at Lateral-basal Segment of RLL (RB9).
Instilled 60 cc of NS, suction returned with 15 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
 
Serial occlusion with endobronchial blocker (ardnts 7Fr) and Fogarty balloon was done to isolate the airleak to be at the RLL (Lateral and Posterior subsegment).
Airleak was reproduced with inspiratory hold at 30 and suction on pleurovac on -20cmH20.
Tisseel 2cc was used to block off a subsegment of the RLL posterior branch.
Size 7 spiration valve was placed but noted to be too large for the airway (RB10).  This was subsequently removed.
Then Size 6 spiration valve was placed in RB9, in good position.
Then size 6 spiration valve was placed in RB10, noted to be in poor position, this was removed again and replaced with another size 6 spiration valve in a better angle.
Final: 
RB9 - size 6 spiration valve
RB10- size 6 spiration valve
 
With airleak significantly decreased.
See Dr. Thistlethwaite's note for VATS and pleurodesis. 
 
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
BAL (x2) 
 
IMPRESSION/PLAN: [REDACTED]is a 68 year old-year-old male who presents for bronchoscopy for BAL and valve placement.
- f/u in BAL results
- f/u in 6 weeks for valve removal"""

# =============================================================================
# 2. DEFINITIONS & UTILS
# =============================================================================

PROCEDURE_FLAGS = [
    # Bronchoscopy (23)
    "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
    "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy",
    "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration",
    "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
    "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation",
    "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
    # Pleural (7)
    "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis",
    "pleural_biopsy", "fibrinolytic_therapy"
]

def clean_text(text):
    return text.strip().replace('\r', '')

def get_exact_substring(full_text, substring, context_prefix=None, instance_index=0):
    """
    Finds a substring in full_text.
    Returns (start, end) or (None, None).
    Logic:
    1. If context_prefix is provided, search near that prefix.
    2. Else, use instance_index (0-based) to find the nth occurrence.
    """
    if not substring or substring not in full_text:
        return None, None

    if context_prefix:
        # Scan for prefix
        p_len = len(context_prefix)
        start_search = 0
        while True:
            p_idx = full_text.find(context_prefix, start_search)
            if p_idx == -1:
                break
            # Look for substring within a window after prefix (e.g. 150 chars)
            window_size = 200
            search_region = full_text[p_idx:p_idx+window_size]
            sub_idx = search_region.find(substring)
            if sub_idx != -1:
                final_start = p_idx + sub_idx
                return final_start, final_start + len(substring)
            start_search = p_idx + 1
        # Fallback if prefix logic fails but string exists: use index 0
        idx = full_text.find(substring)
        return idx, idx + len(substring)

    else:
        # Use instance index
        current_idx = -1
        for _ in range(instance_index + 1):
            current_idx = full_text.find(substring, current_idx + 1)
            if current_idx == -1:
                return None, None
        return current_idx, current_idx + len(substring)

class WorkbookGenerator:
    def __init__(self):
        self.wb = None
        self.note_text = clean_text(NOTE_TEXT)
        self.spans = []
        self.events = []
        self.proc_flags = {k: 0 for k in PROCEDURE_FLAGS}
        
    def load_template(self):
        if not os.path.exists(TEMPLATE_PATH):
            # Create a dummy workbook if template is missing (for safety in constrained envs)
            self.wb = openpyxl.Workbook()
            for sheet in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", 
                          "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
                self.wb.create_sheet(sheet)
        else:
            self.wb = openpyxl.load_workbook(TEMPLATE_PATH)

    def extract_data(self):
        # 1. Flags
        # Based on HEADER codes and detailed text
        self.proc_flags['diagnostic_bronchoscopy'] = 1 # 31624
        self.proc_flags['bal'] = 1 # 31624 + text
        self.proc_flags['therapeutic_aspiration'] = 1 # 31645 + text
        self.proc_flags['foreign_body_removal'] = 1 # 31635 in header. Though text is sparse, we respect the billing code presence.
        self.proc_flags['blvr'] = 1 # 31647 Valve placement
        self.proc_flags['pleurodesis'] = 0 # Specifically refers to another note

        # 2. Events & Spans
        
        # --- Event 1: Therapeutic Aspiration ---
        evt_asp = "evt_01"
        self.events.append({
            "event_id": evt_asp,
            "type": "therapeutic_aspiration",
            "method": "therapeutic aspiration",
            "target_anatomy": "Right Mainstem, Bronchus Intermedius , and Left Mainstem",
            "findings": "mucus and mucus plug"
        })
        self.add_span(evt_asp, "PROC_METHOD", "therapeutic aspiration", "Successful ")
        self.add_span(evt_asp, "ANAT_AIRWAY", "Right Mainstem", "clean out the ")
        self.add_span(evt_asp, "ANAT_AIRWAY", "Bronchus Intermedius", "Right Mainstem, ")
        self.add_span(evt_asp, "ANAT_AIRWAY", "Left Mainstem", "Intermedius , and ")
        self.add_span(evt_asp, "OBS_LESION", "mucus", "from ")
        self.add_span(evt_asp, "OBS_LESION", "mucus plug", "and ")

        # --- Event 2: BAL Lingula ---
        evt_bal1 = "evt_02"
        self.events.append({
            "event_id": evt_bal1,
            "type": "bal",
            "method": "Bronchial alveolar lavage",
            "target_anatomy": "Superior Segment of Lingula (LB4) and Inferior Segment of Lingula (LB5)",
            "measurements_json": {"instilled": "60 cc", "return": "15 cc"}
        })
        self.add_span(evt_bal1, "PROC_METHOD", "Bronchial alveolar lavage", "plug.\n")
        self.add_span(evt_bal1, "ANAT_LUNG_LOC", "Superior Segment of Lingula", "performed at ")
        self.add_span(evt_bal1, "ANAT_AIRWAY", "LB4", "Lingula (")
        self.add_span(evt_bal1, "ANAT_LUNG_LOC", "Inferior Segment of Lingula", "and ")
        self.add_span(evt_bal1, "ANAT_AIRWAY", "LB5", "Lingula (")
        self.add_span(evt_bal1, "MEAS_VOL", "60 cc", "Instilled ")
        self.add_span(evt_bal1, "MEAS_VOL", "15 cc", "returned with ")

        # --- Event 3: BAL RLL ---
        evt_bal2 = "evt_03"
        self.events.append({
            "event_id": evt_bal2,
            "type": "bal",
            "method": "Bronchial alveolar lavage",
            "target_anatomy": "Lateral-basal Segment of RLL (RB9)",
            "measurements_json": {"instilled": "60 cc", "return": "15 cc"}
        })
        self.add_span(evt_bal2, "PROC_METHOD", "Bronchial alveolar lavage", "Cytology.\n \n")
        self.add_span(evt_bal2, "ANAT_LUNG_LOC", "Lateral-basal Segment of RLL", "performed at ")
        self.add_span(evt_bal2, "ANAT_AIRWAY", "RB9", "(", match_index=0) # Need careful matching: (RB9)
        self.add_span(evt_bal2, "MEAS_VOL", "60 cc", "Instilled ", match_index=1)
        self.add_span(evt_bal2, "MEAS_VOL", "15 cc", "returned with ", match_index=1)

        # --- Event 4: Occlusion / Isolation ---
        evt_occ = "evt_04"
        self.events.append({
            "event_id": evt_occ,
            "type": "occlusion",
            "method": "Serial occlusion",
            "devices": ["endobronchial blocker", "Fogarty balloon"],
            "target_anatomy": "RLL (Lateral and Posterior subsegment)",
            "findings": "Airleak was reproduced"
        })
        self.add_span(evt_occ, "PROC_METHOD", "Serial occlusion", "Cytology.\n \n")
        self.add_span(evt_occ, "DEV_CATHETER", "endobronchial blocker", "with ")
        self.add_span(evt_occ, "DEV_CATHETER_SIZE", "7Fr", "ardnts ")
        self.add_span(evt_occ, "DEV_CATHETER", "Fogarty balloon", "and ")
        self.add_span(evt_occ, "ANAT_LUNG_LOC", "RLL", "airleak to be at the ")
        self.add_span(evt_occ, "ANAT_LUNG_LOC", "Lateral and Posterior subsegment", "RLL (")
        self.add_span(evt_occ, "OBS_ROSE", "Airleak was reproduced", "subsegment).\n")

        # --- Event 5: Tisseel (Sealant) ---
        evt_seal = "evt_05"
        self.events.append({
            "event_id": evt_seal,
            "type": "sealant_injection",
            "method": "Tisseel 2cc was used",
            "target_anatomy": "RLL posterior branch",
            "measurements_json": {"volume": "2cc"}
        })
        self.add_span(evt_seal, "PROC_METHOD", "Tisseel", "-20cmH20.\n")
        self.add_span(evt_seal, "MEAS_VOL", "2cc", "Tisseel ")
        self.add_span(evt_seal, "ANAT_AIRWAY", "RLL posterior branch", "subsegment of the ")

        # --- Event 6: Valve Attempt 1 (RB10, Size 7) ---
        evt_v1 = "evt_06"
        self.events.append({
            "event_id": evt_v1,
            "type": "blvr",
            "method": "valve was placed",
            "target_anatomy": "RB10",
            "device": "spiration valve",
            "size": "Size 7",
            "outcome": "removed"
        })
        self.add_span(evt_v1, "DEV_VALVE", "spiration valve", "Size 7 ")
        self.add_span(evt_v1, "MEAS_SIZE", "Size 7", "branch.\n")
        self.add_span(evt_v1, "PROC_METHOD", "placed", "valve was ")
        self.add_span(evt_v1, "ANAT_AIRWAY", "RB10", "airway (")
        self.add_span(evt_v1, "OBS_LESION", "too large for the airway", "noted to be ")
        self.add_span(evt_v1, "PROC_ACTION", "removed", "subsequently ")

        # --- Event 7: Valve Placement 1 (RB9, Size 6) ---
        evt_v2 = "evt_07"
        self.events.append({
            "event_id": evt_v2,
            "type": "blvr",
            "method": "valve was placed",
            "target_anatomy": "RB9",
            "device": "spiration valve",
            "size": "Size 6"
        })
        self.add_span(evt_v2, "DEV_VALVE", "spiration valve", "Size 6 ", match_index=0)
        self.add_span(evt_v2, "MEAS_SIZE", "Size 6", "Then ")
        self.add_span(evt_v2, "PROC_METHOD", "placed", "valve was ", match_index=1)
        self.add_span(evt_v2, "ANAT_AIRWAY", "RB9", "placed in ")
        self.add_span(evt_v2, "OBS_LESION", "good position", ", in ")

        # --- Event 8: Valve Attempt 2 (RB10, Size 6) ---
        evt_v3 = "evt_08"
        self.events.append({
            "event_id": evt_v3,
            "type": "blvr",
            "method": "valve was placed",
            "target_anatomy": "RB10",
            "device": "spiration valve",
            "size": "size 6",
            "outcome": "removed"
        })
        self.add_span(evt_v3, "DEV_VALVE", "spiration valve", "size 6 ", match_index=1)
        self.add_span(evt_v3, "MEAS_SIZE", "size 6", "Then ")
        self.add_span(evt_v3, "PROC_METHOD", "placed", "valve was ", match_index=2)
        self.add_span(evt_v3, "ANAT_AIRWAY", "RB10", "placed in ", match_index=1)
        self.add_span(evt_v3, "OBS_LESION", "poor position", "to be in ")
        self.add_span(evt_v3, "PROC_ACTION", "removed", "this was ")

        # --- Event 9: Valve Placement 2 (RB10, Size 6) ---
        evt_v4 = "evt_09"
        self.events.append({
            "event_id": evt_v4,
            "type": "blvr",
            "method": "replaced",
            "target_anatomy": "RB10",
            "device": "spiration valve",
            "size": "size 6"
        })
        self.add_span(evt_v4, "PROC_METHOD", "replaced", "again and ")
        self.add_span(evt_v4, "MEAS_SIZE", "size 6", "another ")
        self.add_span(evt_v4, "DEV_VALVE", "spiration valve", "size 6 ", match_index=2)
        self.add_span(evt_v4, "OBS_LESION", "better angle", "in a ")

        # Global Outcomes
        self.add_span("global", "OUTCOME_COMPLICATION", "no immediate complications", "were ")
        self.add_span("global", "OUTCOME_SYMPTOMS", "airleak significantly decreased", "With ")

    def add_span(self, event_id, label, span_text, context_prefix=None, match_index=None):
        self.spans.append({
            "event_id": event_id,
            "label": label,
            "span_text": span_text,
            "context_prefix": context_prefix,
            "match_index": match_index
        })

    def write_excel(self):
        self.load_template()
        
        # 1. Note_Text
        ws = self.wb["Note_Text"]
        ws.append([NOTE_ID, SOURCE_FILE, self.note_text])

        # 2. Note_Index
        ws = self.wb["Note_Index"]
        row_idx = [
            SOURCE_FILE, NOTE_ID, "", "", "", "", "Pending", ""
        ]
        # Append flags
        for flag in PROCEDURE_FLAGS:
            row_idx.append(self.proc_flags.get(flag, 0))
        ws.append(row_idx)

        # 3. Span_Annotations
        ws_span = self.wb["Span_Annotations"]
        ws_hyd = self.wb["Span_Hydrated"]
        
        # Header check for Span_Annotations (assuming template has them, but logic requires appending)
        
        span_id_counter = 1
        
        for s in self.spans:
            # Anchor Logic
            span_text = s["span_text"]
            ctx = s["context_prefix"]
            midx = s.get("match_index", 0) if s.get("match_index") is not None else 0
            
            # Hydration
            start, end = get_exact_substring(self.note_text, span_text, ctx, midx)
            if start is None:
                hyd_status = "ambiguous"
                start_val, end_val = "", ""
            else:
                hyd_status = "hydrated"
                start_val, end_val = start, end

            # Common Row
            # Columns: source, note_id, span_id, section, prefix, text, match_idx, start, end, len, label, norm, schema, event, neg, hist, time, rev, comm, hyd_status
            
            row_base = [
                SOURCE_FILE, NOTE_ID, f"span_{span_id_counter:03d}", "Procedure",
                ctx, span_text, midx
            ]
            
            # Annotation Sheet (No offsets)
            row_anno = row_base + ["", "", f"=LEN(F{ws_span.max_row+1})", s["label"], "", "", s["event_id"], 
                                   "FALSE", "FALSE", "", "", "", "needs_hydration"]
            ws_span.append(row_anno)
            
            # Hydrated Sheet (With offsets)
            row_hyd = row_base + [start_val, end_val, len(span_text) if start_val!="" else "", s["label"], "", "", s["event_id"], 
                                  "FALSE", "FALSE", "", "", "", hyd_status]
            ws_hyd.append(row_hyd)
            
            span_id_counter += 1

        # 4. Event_Log
        ws_log = self.wb["Event_Log"]
        for e in self.events:
            row = [
                SOURCE_FILE, NOTE_ID, e["event_id"], e["type"], e.get("method", ""),
                e.get("target_anatomy", ""), e.get("device", ""), "", "", "",
                json.dumps(e.get("measurements_json", {})),
                "", e.get("findings", ""), "FALSE", "", "",
                e.get("size", ""), "", "", "", "", "", e.get("outcome", "")
            ]
            ws_log.append(row)

        # 5. V3_Procedure_Events
        ws_v3 = self.wb["V3_Procedure_Events"]
        for e in self.events:
            # Map loosely to V3 columns
            # note_id, event_id, type, target.anat, target.lobe, target.seg, target.stat, lesion.type, lesion.size, method, dev_json, meas_json, spec_json, find_json
            row = [
                NOTE_ID, e["event_id"], e["type"],
                e.get("target_anatomy", ""), "", "", "", # target breakdown left basic
                "", "", # lesion
                e.get("method", ""),
                json.dumps({"device": e.get("device"), "size": e.get("size")}) if e.get("device") else "",
                json.dumps(e.get("measurements_json", {})),
                "",
                json.dumps({"findings": e.get("findings")}) if e.get("findings") else "",
                "", # evidence
                e.get("size", "") if "valve" in str(e.get("device", "")).lower() else "",
                "", # stent mat
                "7Fr" if "blocker" in str(e.get("devices", "")) else "",
                "", "", "", "", e.get("outcome", "")
            ]
            ws_v3.append(row)

        # 6. JSON
        ws_json = self.wb["V3_Registry_JSON"]
        registry_data = {
            "schema_version": "3.0",
            "note_id": NOTE_ID,
            "procedures": self.events,
            "no_immediate_complications": True
        }
        ws_json.append([json.dumps(registry_data, indent=2)])

        # Save
        self.wb.save(OUTPUT_PATH)
        print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    gen = WorkbookGenerator()
    gen.extract_data()
    gen.write_excel()