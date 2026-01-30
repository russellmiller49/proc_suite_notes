import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_022"
SOURCE_FILE = "note_022.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_022 SOURCE_FILE: note_022.txt PROCEDURE: Flexible bronchoscopy with BAL and TBBX under moderate sedation with fluoroscopic guidance 
INDICATION : Lung Transplant
CONSENT : Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient.
The patient read and signed the provided consent form. The consent was witnessed by an assisting medical proffesional.
MONITORING : Continuous telemetry, BP and oxygen saturation monitored
INSTRUMENT : Olympus Video Bronchoscope
ANESTHESIA : Moderate IV sedation with: fentanyl 50 mcg and Versed 2.5 mg;
Local anesthesia with: Lidocaine 2% Solution ~20mL intratracheal. Fluoro time 28s
Sedation time was 24 min during this procedure
 
Time out: Immediately prior to procedure a "time out" was called to verify the correct patient, procedure, equipment, support staff and site/side  marked as required.
SUMMARY : Initially a time-out was performed prior to the procedure, confirming the patient's name, procedure type and procedure location.
After the patient was properly positioned and sedated and topical anesthesia applied, the bronchoscope was introduced through the mouth with O2 being administered at all times.
This was done without difficulty. The bronchoscope was passed by the carina, which was examined for sharpness, position and texture.
The bronchial orifices were systematically identified, evaluated and suctioned free of secretions and close attention was paid to color, texture, positions, size and patency.
Mild distortion of airways noted bilaterally. 
 
FINDINGS
Pharynx: Normal 
Larynx: Normal 
Vocal Chords: Normal 
Trachea: Normal 
Carina: Sharp
 
Bronchial Tree
Right Lung: Normal anastomosis, no strictures, ischemia or black eschar
 
Left Lung: Normal anastomosis, no strictures, ischemia or black eschar
 
No evidence of mass, lesions, bleeding or other intra-bronchial pathology.
Mucosa: normal
Secretions: none. Suctioned to clear (therapeutic aspiration). 
 
Studies/samples: 
1. BAL 1: 60 cc instilled in the RML and 20 cc returned
1. BAL 1: 60 cc instilled in the LLL and 20 cc returned
2. Biopsies: 10 TBBX attempted and 10 samples obtained from the RLL, lateral basal segment 
- EBL: 5ml
Complications: None
 
 
IMPRESSION: 
Flexible bronchoscopy with BAL and TBBX under moderate sedation
 
Post procedure, no dyspnea, chest pain or changes to his voice
Exam: CTA over the area of the lavage and biopsy, no voice change or crepitus noted
Samples sent for microbiology and pathology"""

# -------------------------------------------------------------------------
# CONFIGURATION & DEFINITIONS
# -------------------------------------------------------------------------
PROCEDURE_FLAGS = {
    # Bronchoscopy (23)
    'diagnostic_bronchoscopy': 1,
    'bal': 1,
    'bronchial_wash': 0,
    'brushings': 0,
    'endobronchial_biopsy': 0,
    'tbna_conventional': 0,
    'linear_ebus': 0,
    'radial_ebus': 0,
    'navigational_bronchoscopy': 0,
    'transbronchial_biopsy': 1,
    'transbronchial_cryobiopsy': 0,
    'therapeutic_aspiration': 1,
    'foreign_body_removal': 0,
    'airway_dilation': 0,
    'airway_stent': 0,
    'thermal_ablation': 0,
    'tumor_debulking_non_thermal': 0,
    'cryotherapy': 0,
    'blvr': 0,
    'peripheral_ablation': 0,
    'bronchial_thermoplasty': 0,
    'whole_lung_lavage': 0,
    'rigid_bronchoscopy': 0,
    # Pleural (7)
    'thoracentesis': 0,
    'chest_tube': 0,
    'ipc': 0,
    'medical_thoracoscopy': 0,
    'pleurodesis': 0,
    'pleural_biopsy': 0,
    'fibrinolytic_therapy': 0
}

# Define Spans
# Structure: (text, label, normalized_value, event_id, context_prefix)
raw_spans = [
    # Metadata/Devices
    ("Olympus Video Bronchoscope", "DEV_INSTRUMENT", "Olympus Video Bronchoscope", "evt0", None),
    ("fentanyl 50 mcg", "PROC_METHOD", "Sedation-Fentanyl", "evt0", None),
    ("Versed 2.5 mg", "PROC_METHOD", "Sedation-Versed", "evt0", None),
    ("Lidocaine 2% Solution", "PROC_METHOD", "Anesthesia-Local", "evt0", None),
    ("Lung Transplant", "CTX_HISTORICAL", "Lung Transplant", "evt0", None),
    
    # Inspection / General Findings
    ("Mild distortion of airways", "OBS_LESION", "Airway Distortion", "evt0", None),
    ("Carina: Sharp", "OBS_FINDING", "Carina Sharp", "evt0", None),
    
    # Therapeutic Aspiration (General/Secretions)
    ("suctioned free of secretions", "PROC_ACTION", "Suction", "evt1", None),
    ("therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", "evt1", "Suctioned to clear"),

    # Right Lung Inspection
    ("Right Lung", "ANAT_LUNG_LOC", "Right Lung", "evt_inspect_R", "Bronchial Tree"),
    ("Normal anastomosis", "OBS_FINDING", "Normal Anastomosis", "evt_inspect_R", "Right Lung:"),
    ("no strictures", "OBS_FINDING", "No Strictures", "evt_inspect_R", "Right Lung:"),
    
    # Left Lung Inspection
    ("Left Lung", "ANAT_LUNG_LOC", "Left Lung", "evt_inspect_L", "is or black eschar"),
    ("Normal anastomosis", "OBS_FINDING", "Normal Anastomosis", "evt_inspect_L", "Left Lung:"),
    ("no strictures", "OBS_FINDING", "No Strictures", "evt_inspect_L", "Left Lung:"),

    # BAL 1 (RML)
    ("BAL", "PROC_METHOD", "BAL", "evt_bal_1", "1."),
    ("60 cc", "MEAS_VOL", "60ml", "evt_bal_1", "BAL 1:"),
    ("RML", "ANAT_LUNG_LOC", "RML", "evt_bal_1", "instilled in the"),
    ("20 cc", "MEAS_VOL", "20ml", "evt_bal_1", "RML and"),
    
    # BAL 2 (LLL) - Note text says "1. BAL 1:" again for the second line
    ("BAL", "PROC_METHOD", "BAL", "evt_bal_2", "RML and 20 cc returned"), # Context from prev line
    ("60 cc", "MEAS_VOL", "60ml", "evt_bal_2", "instilled in the LLL"),
    ("LLL", "ANAT_LUNG_LOC", "LLL", "evt_bal_2", "instilled in the"),
    ("20 cc", "MEAS_VOL", "20ml", "evt_bal_2", "LLL and"),

    # TBBX (RLL)
    ("TBBX", "PROC_METHOD", "Transbronchial Biopsy", "evt_tbbx", "2. Biopsies: 10"),
    ("10", "MEAS_COUNT", "10", "evt_tbbx", "2. Biopsies:"),
    ("RLL, lateral basal segment", "ANAT_LUNG_LOC", "RLL LB", "evt_tbbx", "obtained from the"),
    ("10 samples", "MEAS_COUNT", "10", "evt_tbbx", "attempted and"),

    # Outcomes
    ("None", "OUTCOME_COMPLICATION", "None", "evt0", "Complications:"),
    ("no dyspnea", "OUTCOME_SYMPTOMS", "None", "evt0", "Post procedure,"),
    ("no voice change", "OUTCOME_SYMPTOMS", "None", "evt0", "biopsy,"),
]

# -------------------------------------------------------------------------
# HELPER FUNCTIONS
# -------------------------------------------------------------------------

def hydrate_spans(text, span_list):
    hydrated = []
    
    for idx, (span_text, label, norm_val, evt_id, ctx) in enumerate(span_list):
        row = {
            "source_file": SOURCE_FILE,
            "note_id": NOTE_ID,
            "span_id": f"span_{idx+1:03d}",
            "section_type": "Procedure",
            "context_prefix": ctx,
            "span_text": span_text,
            "match_index": None,
            "start_char": None,
            "end_char": None,
            "span_len": len(span_text),
            "label": label,
            "normalized_value": norm_val,
            "schema_field": "",
            "event_id": evt_id,
            "is_negated": False,
            "is_historical": False,
            "time_anchor": "",
            "reviewer": "Auto",
            "comments": "",
            "hydration_status": "unresolved"
        }
        
        # Hydration Logic
        count = text.count(span_text)
        
        if count == 0:
            row["hydration_status"] = "not_found"
            hydrated.append(row)
            continue
            
        start_idx = -1
        
        if count == 1:
            start_idx = text.find(span_text)
            row["hydration_status"] = "hydrated_unique"
        elif ctx:
            # Context search
            matches = [m.start() for m in re.finditer(re.escape(span_text), text)]
            found = False
            for m_start in matches:
                # look back 150 chars
                window_start = max(0, m_start - 150)
                window = text[window_start:m_start]
                if ctx in window:
                    start_idx = m_start
                    row["hydration_status"] = "hydrated_context"
                    found = True
                    break
            if not found:
                 # Fallback to first if context fails (or manual check)
                 row["hydration_status"] = f"ambiguous_ctx_fail_count={count}"
        else:
             row["hydration_status"] = f"ambiguous_count={count}"
             
        if start_idx != -1:
            row["start_char"] = start_idx
            row["end_char"] = start_idx + len(span_text)
            
        hydrated.append(row)
        
    return hydrated

def generate_events(hydrated_spans):
    events = {}
    
    for row in hydrated_spans:
        eid = row['event_id']
        if eid not in events:
            events[eid] = {
                "source_file": SOURCE_FILE,
                "note_id": NOTE_ID,
                "event_id": eid,
                "event_type": "Procedure", # default
                "method": [],
                "anatomy_target": [],
                "device": [],
                "counts": [],
                "measurements": [],
                "findings": [],
                "outcomes": {}
            }
        
        lbl = row['label']
        val = row['normalized_value']
        
        if lbl == "PROC_METHOD":
            events[eid]['method'].append(val)
        elif lbl == "PROC_ACTION":
            events[eid]['method'].append(val)
        elif lbl.startswith("ANAT_"):
            events[eid]['anatomy_target'].append(val)
        elif lbl.startswith("DEV_"):
            events[eid]['device'].append(val)
        elif lbl == "MEAS_COUNT":
            events[eid]['counts'].append(val)
        elif lbl.startswith("MEAS_") and lbl != "MEAS_COUNT":
            events[eid]['measurements'].append(f"{val} ({row['span_text']})")
        elif lbl.startswith("OBS_"):
            events[eid]['findings'].append(val)
        elif lbl.startswith("OUTCOME_"):
            events[eid]['outcomes'][lbl] = val

    # Format for sheet
    event_rows = []
    for eid, data in events.items():
        # Clean up lists
        data['method'] = list(set(data['method']))
        data['anatomy_target'] = list(set(data['anatomy_target']))
        
        out_row = [
            SOURCE_FILE,
            NOTE_ID,
            eid,
            data['event_type'],
            ", ".join(data['method']),
            ", ".join(data['anatomy_target']),
            ", ".join(data['device']),
            "", # gauge
            "", # stations
            ", ".join(data['counts']),
            ", ".join(data['measurements']),
            "", # specimens
            ", ".join(data['findings']),
            False, # historical
            "Auto",
            "",
            "", # size
            "", # material
            data['outcomes'].get('OUTCOME_AIRWAY_LUMEN_PRE', ""),
            data['outcomes'].get('OUTCOME_AIRWAY_LUMEN_POST', ""),
            data['outcomes'].get('OUTCOME_SYMPTOMS', ""),
            data['outcomes'].get('OUTCOME_PLEURAL', ""),
            data['outcomes'].get('OUTCOME_COMPLICATION', "")
        ]
        event_rows.append(out_row)
        
    return event_rows, events

def main():
    # 1. Load Template
    if not os.path.exists(TEMPLATE_PATH):
        # Create dummy if missing for script validity
        wb = openpyxl.Workbook()
        wb.create_sheet("Note_Text")
        wb.create_sheet("Note_Index")
        wb.create_sheet("Span_Annotations")
        wb.create_sheet("Span_Hydrated")
        wb.create_sheet("Event_Log")
        wb.create_sheet("V3_Procedure_Events")
        wb.create_sheet("V3_Registry_JSON")
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # 2. Populate Note_Text
    ws_text = wb["Note_Text"]
    if ws_text.max_row == 1:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 3. Populate Note_Index
    ws_index = wb["Note_Index"]
    # Headers usually exist. If empty, write them (simplified check)
    if ws_index.max_row == 0:
        headers = ["source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes"] + list(PROCEDURE_FLAGS.keys())
        ws_index.append(headers)
    
    # Construct row
    index_row = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Auto-Generated", ""]
    for k in PROCEDURE_FLAGS.keys():
        index_row.append(PROCEDURE_FLAGS[k])
    ws_index.append(index_row)

    # 4. Hydrate Spans
    hydrated_data = hydrate_spans(NOTE_TEXT, raw_spans)

    # 5. Write Span_Annotations (Anchor First - no offsets)
    ws_anno = wb["Span_Annotations"]
    if ws_anno.max_row == 0:
        ws_anno.append(["source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", "match_index", "start_char", "end_char", "span_len", "label", "normalized_value", "schema_field", "event_id", "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"])

    for row in hydrated_data:
        ws_anno.append([
            row["source_file"], row["note_id"], row["span_id"], row["section_type"],
            row["context_prefix"], row["span_text"], row["match_index"],
            "", "", f'=LEN(INDIRECT("F"&ROW()))', # Formula for length
            row["label"], row["normalized_value"], row["schema_field"], row["event_id"],
            row["is_negated"], row["is_historical"], row["time_anchor"],
            row["reviewer"], row["comments"], "needs_hydration"
        ])

    # 6. Write Span_Hydrated
    ws_hydra = wb["Span_Hydrated"]
    if ws_hydra.max_row == 0:
        # Same headers
        ws_hydra.append(["source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", "match_index", "start_char", "end_char", "span_len", "label", "normalized_value", "schema_field", "event_id", "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"])

    for row in hydrated_data:
        ws_hydra.append([
            row["source_file"], row["note_id"], row["span_id"], row["section_type"],
            row["context_prefix"], row["span_text"], row["match_index"],
            row["start_char"], row["end_char"], len(row["span_text"]),
            row["label"], row["normalized_value"], row["schema_field"], row["event_id"],
            row["is_negated"], row["is_historical"], row["time_anchor"],
            row["reviewer"], row["comments"], row["hydration_status"]
        ])

    # 7. Generate and Write Event Log
    event_rows, event_dict = generate_events(hydrated_data)
    ws_events = wb["Event_Log"]
    if ws_events.max_row == 0:
        ws_events.append(["source_file", "note_id", "event_id", "event_type", "method", "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements", "specimens", "findings", "is_historical", "reviewer", "comments", "device_size", "device_material", "outcome_airway_lumen_pre", "outcome_airway_lumen_post", "outcome_symptoms", "outcome_pleural", "outcome_complication"])
    
    for er in event_rows:
        ws_events.append(er)

    # 8. V3 Procedure Events (Flat mapping)
    ws_v3 = wb["V3_Procedure_Events"]
    if ws_v3.max_row == 0:
        ws_v3.append(["note_id", "event_id", "type", "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station", "lesion.type", "lesion.size_mm", "method", "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote", "stent.size", "stent.material_or_brand", "catheter.size_fr", "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post", "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"])

    for eid, data in event_dict.items():
        # Basic mapping logic for demo
        anatomy = ", ".join(data['anatomy_target'])
        method = ", ".join(data['method'])
        ws_v3.append([
            NOTE_ID, eid, "Procedure", "Airway" if "Lung" in anatomy or "RML" in anatomy else "Unknown",
            anatomy, "", "", "", "", method, 
            json.dumps(data['device']), json.dumps(data['measurements']), 
            "", json.dumps(data['findings']), "", "", "", "",
            data['outcomes'].get('OUTCOME_AIRWAY_LUMEN_PRE', ""),
            data['outcomes'].get('OUTCOME_AIRWAY_LUMEN_POST', ""),
            data['outcomes'].get('OUTCOME_SYMPTOMS', ""),
            data['outcomes'].get('OUTCOME_PLEURAL', ""),
            data['outcomes'].get('OUTCOME_COMPLICATION', "")
        ])

    # 9. JSON Generation
    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 0:
        ws_json.append(["schema_version", "note_id", "procedures_json", "no_immediate_complications"])
    
    registry_obj = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "events": event_dict,
        "no_immediate_complications": "None" in str(event_dict.values()) # simplistic check
    }
    
    ws_json.append(["3.0", NOTE_ID, json.dumps(registry_obj, indent=2), True])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    main()