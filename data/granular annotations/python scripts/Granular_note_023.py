import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import json
import re
import os
from datetime import datetime

# ==========================================
# 1. INPUT DATA & CONFIGURATION
# ==========================================

NOTE_ID = "note_023"
SOURCE_FILE = "note_023.txt"
PROCEDURE_DATE = "2026-01-12" # Placeholder
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_023 SOURCE_FILE: note_023.txt INDICATION FOR OPERATION:  [REDACTED]is a 74 year old-year-old female who presents with hemoptysis after bronchoscopy.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: hemoptysis
POSTOPERATIVE DIAGNOSIS:  hemoptysis
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31622 Dx bronchoscope/cell washing          
ANESTHESIA: 
Continuous sedation
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Disposable Bronchoscope
ESTIMATED BLOOD LOSS:   Minimum
COMPLICATIONS:    None
PROCEDURE IN DETAIL: A timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: supine
Initial Airway Inspection Findings:
The bronchoscope was introduced into the ETT.
The ETT is 2 cm above the main carina. The bronchial blocker was in the RLL but was mostly deflated without evidence of spill over into the other segments.
Successful therapeutic aspiration was performed to clean out the LLL and LMSB from mucus.
The balloon was was fully deflated and then removed. Then a new 7 F Ardnt bronchial blocker was replaced through the existing ETT.
The blocker was positioned in the distal BI. 3 mL of air was introduced into the blocker and resulted in occlusion of the BI without occlusion of the RUL.
During exchange of the bronchial blocker there was no increase in the blood seen in the RLL lateral segment.
This was no aspirated due to no obvious formed clot.
The bronchoscope was removed. 
 
The patient tolerated the procedure well.
There were no immediate complications.  
SPECIMEN(S): 
none
IMPRESSION/PLAN: [REDACTED]is a 74 year old-year-old female who presents for bronchoscopy for hemoptysis.
[ ] keep blocker in place at 40 cm at the connective device
[ ] ETT at 20 cm at the front tooth
[ ] stop paralysis once this is not effective stop fentanyl and repeat bronchoscopy at 1500 today
[ ] family updated"""

# ==========================================
# 2. DEFINITIONS & SCHEMA
# ==========================================

PROCEDURE_FLAGS = [
    "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
    "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy",
    "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration",
    "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
    "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation",
    "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
    "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis",
    "pleural_biopsy", "fibrinolytic_therapy"
]

# Set flags based on note content
FLAGS_VALUES = {flag: 0 for flag in PROCEDURE_FLAGS}
FLAGS_VALUES["diagnostic_bronchoscopy"] = 1 # CPT 31622 + inspection
FLAGS_VALUES["therapeutic_aspiration"] = 1 # CPT 31645 + "Successful therapeutic aspiration"
# bronchial_wash is 0 despite CPT code because SPECIMEN(S): none and intent was cleaning (aspiration)

# ==========================================
# 3. EXTRACTION LOGIC (SPANS & EVENTS)
# ==========================================

# Raw Spans Data
# Fields: span_text, label, normalized_value, context_prefix (optional)
raw_spans = [
    # Metadata/Indication
    {"text": "hemoptysis", "label": "OBS_ROSE", "norm": "hemoptysis", "context": "presents with"},
    
    # Procedure 1: Diagnostic / Inspection
    {"text": "Disposable Bronchoscope", "label": "DEV_INSTRUMENT", "norm": "bronchoscope"},
    {"text": "supine", "label": "CTX_TIME", "norm": "supine"}, # Using closest cat
    {"text": "ETT", "label": "DEV_CATHETER", "norm": "ETT", "context": "introduced into the"},
    {"text": "main carina", "label": "ANAT_AIRWAY", "norm": "main carina"},
    {"text": "bronchial blocker", "label": "DEV_CATHETER", "norm": "bronchial blocker", "context": "The"},
    {"text": "RLL", "label": "ANAT_LUNG_LOC", "norm": "RLL", "context": "blocker was in the"},
    
    # Procedure 2: Therapeutic Aspiration
    {"text": "Therapeutic aspiration", "label": "PROC_METHOD", "norm": "therapeutic_aspiration", "context": "Successful"},
    {"text": "clean out", "label": "PROC_ACTION", "norm": "clean out"},
    {"text": "LLL", "label": "ANAT_LUNG_LOC", "norm": "LLL"},
    {"text": "LMSB", "label": "ANAT_AIRWAY", "norm": "LMSB"},
    {"text": "mucus", "label": "OBS_LESION", "norm": "mucus"}, # Using OBS_LESION as catch-all for findings like mucus/clot
    
    # Procedure 3: Blocker Exchange/Placement
    {"text": "removed", "label": "PROC_ACTION", "norm": "remove", "context": "deflated and then"},
    {"text": "7 F", "label": "DEV_CATHETER_SIZE", "norm": "7Fr"},
    {"text": "Ardnt", "label": "DEV_CATHETER", "norm": "Arndt"},
    {"text": "bronchial blocker", "label": "DEV_CATHETER", "norm": "bronchial blocker", "context": "Ardnt"},
    {"text": "replaced", "label": "PROC_ACTION", "norm": "replace"},
    {"text": "distal BI", "label": "ANAT_AIRWAY", "norm": "Bronchus Intermedius"},
    {"text": "3 mL", "label": "MEAS_VOL", "norm": "3 mL"},
    {"text": "occlusion", "label": "OBS_LESION", "norm": "occlusion", "context": "resulted in"},
    
    # Outcome
    {"text": "No immediate complications", "label": "OUTCOME_COMPLICATION", "norm": "none"},
    {"text": "tolerated the procedure well", "label": "OUTCOME_SYMPTOMS", "norm": "tolerated well"},
]

# Map spans to events
# Event 1: Diagnostic/Inspection
# Event 2: Therapeutic Aspiration
# Event 3: Bronchial Blocker Placement
event_map = {
    "Disposable Bronchoscope": "evt1",
    "ETT": "evt1",
    "main carina": "evt1",
    "bronchial blocker": "evt3", # Context dependent, but let's map generally for this simplistic script or separate
    "Therapeutic aspiration": "evt2",
    "clean out": "evt2",
    "LLL": "evt2",
    "LMSB": "evt2",
    "mucus": "evt2",
    "removed": "evt3",
    "7 F": "evt3",
    "Ardnt": "evt3",
    "replaced": "evt3",
    "distal BI": "evt3",
    "3 mL": "evt3",
    "occlusion": "evt3",
    "No immediate complications": "evt_outcome",
    "tolerated the procedure well": "evt_outcome"
}

# Process spans into structured list
processed_spans = []
for i, item in enumerate(raw_spans):
    span_obj = {
        "source_file": SOURCE_FILE,
        "note_id": NOTE_ID,
        "span_id": f"span_{i+1:03d}",
        "section_type": "Procedure", # Simplified
        "context_prefix": item.get("context", ""),
        "span_text": item["text"],
        "label": item["label"],
        "normalized_value": item["norm"],
        "schema_field": item["label"], # Simplified mapping
        "event_id": event_map.get(item["text"], "evt0"),
        "is_negated": False,
        "is_historical": False,
        "hydration_status": "needs_hydration"
    }
    processed_spans.append(span_obj)

# Hydration Logic
hydrated_spans = []
for span in processed_spans:
    text = span["span_text"]
    prefix = span["context_prefix"]
    
    start_char = -1
    end_char = -1
    method = "not_found"
    
    # Strategy 1: Exact unique match
    if NOTE_TEXT.count(text) == 1:
        start_char = NOTE_TEXT.find(text)
        method = "hydrated_unique"
    
    # Strategy 2: Context prefix
    elif prefix and prefix in NOTE_TEXT:
        # Find all occurrences of text
        indices = [m.start() for m in re.finditer(re.escape(text), NOTE_TEXT)]
        best_idx = -1
        for idx in indices:
            # Check preceding 120 chars
            window_start = max(0, idx - 120)
            window = NOTE_TEXT[window_start:idx]
            if prefix in window:
                best_idx = idx
                break
        if best_idx != -1:
            start_char = best_idx
            method = "hydrated_prefix_window"
            
    # Strategy 3: Fallback first match (if ambiguous and no prefix)
    if start_char == -1:
        start_char = NOTE_TEXT.find(text)
        method = "ambiguous_first_match" if start_char != -1 else "not_found"

    if start_char != -1:
        end_char = start_char + len(text)
        
    h_span = span.copy()
    h_span["start_char"] = start_char
    h_span["end_char"] = end_char
    h_span["span_len"] = len(text)
    h_span["hydration_status"] = method
    hydrated_spans.append(h_span)


# Events Data
events_data = [
    {
        "event_id": "evt1",
        "type": "diagnostic_bronchoscopy",
        "method": "Inspection",
        "target_anatomy": "Trachea", # ETT/Carina
        "devices": "Disposable Bronchoscope",
        "findings": "ETT 2cm above carina",
        "outcomes": ""
    },
    {
        "event_id": "evt2",
        "type": "therapeutic_aspiration",
        "method": "Aspiration",
        "target_anatomy": "LMSB, LLL",
        "devices": "Bronchoscope",
        "findings": "Mucus",
        "outcomes": "Successful cleaning"
    },
    {
        "event_id": "evt3",
        "type": "bronchial_blocker",
        "method": "Placement",
        "target_anatomy": "Bronchus Intermedius (distal BI)",
        "devices": "7 Fr Arndt Bronchial Blocker",
        "measurements": "3 mL air",
        "findings": "Occlusion achieved",
        "outcomes": ""
    }
]

registry_json = {
    "schema_version": "3.0",
    "note_id": NOTE_ID,
    "no_immediate_complications": True,
    "procedures": events_data
}

# ==========================================
# 4. EXCEL GENERATION
# ==========================================

class Phase0Workbook:
    def __init__(self, template_path, output_path):
        self.output_path = output_path
        # Create a new workbook in memory since we don't have the physical template
        # In a real scenario, we would load: wb = openpyxl.load_workbook(template_path)
        self.wb = openpyxl.Workbook()
        self.setup_sheets()

    def setup_sheets(self):
        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
        
        # Define required sheets and headers
        self.sheets = {
            "Note_Text": ["note_id", "source_file", "note_text"],
            "Note_Index": ["source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes"] + PROCEDURE_FLAGS,
            "Span_Annotations": ["source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", "match_index", "start_char", "end_char", "span_len", "label", "normalized_value", "schema_field", "event_id", "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"],
            "Span_Hydrated": ["source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", "match_index", "start_char", "end_char", "span_len", "label", "normalized_value", "schema_field", "event_id", "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"],
            "Event_Log": ["source_file", "note_id", "event_id", "event_type", "method", "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements", "specimens", "findings", "is_historical", "reviewer", "comments", "device_size", "device_material", "outcome_airway_lumen_pre", "outcome_airway_lumen_post", "outcome_symptoms", "outcome_pleural", "outcome_complication"],
            "V3_Procedure_Events": ["note_id", "event_id", "type", "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station", "lesion.type", "lesion.size_mm", "method", "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote", "stent.size", "stent.material_or_brand", "catheter.size_fr", "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post", "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"],
            "V3_Registry_JSON": ["json_object"]
        }

        for name, headers in self.sheets.items():
            ws = self.wb.create_sheet(name)
            ws.append(headers)
            # Style headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    def write_note_text(self):
        ws = self.wb["Note_Text"]
        ws.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    def write_note_index(self):
        ws = self.wb["Note_Index"]
        row = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Analysis Complete", ""]
        # Append flags
        for flag in PROCEDURE_FLAGS:
            row.append(FLAGS_VALUES[flag])
        ws.append(row)

    def write_spans(self, spans_list, sheet_name):
        ws = self.wb[sheet_name]
        for span in spans_list:
            row = [
                span["source_file"], span["note_id"], span["span_id"], span["section_type"],
                span["context_prefix"], span["span_text"], "", # match_index
                span.get("start_char", ""), span.get("end_char", ""), span.get("span_len", f'=LEN(F{ws.max_row+1})'),
                span["label"], span["normalized_value"], span["schema_field"], span["event_id"],
                span["is_negated"], span["is_historical"], "", "", "", # time/reviewer/comments
                span["hydration_status"]
            ]
            ws.append(row)

    def write_event_log(self):
        ws = self.wb["Event_Log"]
        # Simplified event log writing based on events_data
        for evt in events_data:
            row = [
                SOURCE_FILE, NOTE_ID, evt["event_id"], evt["type"], evt["method"],
                evt["target_anatomy"], evt["devices"], "", "", "", evt.get("measurements", ""),
                "", evt.get("findings", ""), False, "", "",
                "", "", "", "", "", "", "" # Extra granular cols left blank for now
            ]
            ws.append(row)

    def write_v3_json(self):
        ws = self.wb["V3_Registry_JSON"]
        json_str = json.dumps(registry_json, indent=2)
        ws.append([json_str])
        # Resize column
        ws.column_dimensions['A'].width = 100
        ws['A2'].alignment = Alignment(wrap_text=True)

    def save(self):
        self.wb.save(self.output_path)
        print(f"Workbook saved to {self.output_path}")

# ==========================================
# 5. EXECUTION
# ==========================================

def main():
    try:
        # Check if file exists to load content (Simulated here with hardcoded text)
        if os.path.exists(SOURCE_FILE):
            with open(SOURCE_FILE, 'r', encoding='utf-8') as f:
                content = f.read()
                # In real usage, update NOTE_TEXT here
                # global NOTE_TEXT
                # NOTE_TEXT = content
        
        workbook = Phase0Workbook(TEMPLATE_PATH, OUTPUT_PATH)
        workbook.write_note_text()
        workbook.write_note_index()
        workbook.write_spans(processed_spans, "Span_Annotations")
        workbook.write_spans(hydrated_spans, "Span_Hydrated")
        workbook.write_event_log()
        workbook.write_v3_json()
        workbook.save()
        
    except Exception as e:
        print(f"Error generating workbook: {e}")

if __name__ == "__main__":
    main()