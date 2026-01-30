import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# =============================================================================
# INPUT CONFIGURATION
# =============================================================================
NOTE_ID = "note_024"
SOURCE_FILE = "note_024.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_024 SOURCE_FILE: note_024.txt PROCEDURE: Flexible bronchoscopy with BAL and TBBX under moderate sedation
INDICATION : Lung Transplant 
CONSENT : Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient.
The patient read and signed the provided consent form. The consent was witnessed by an assisting medical proffesional.
MONITORING : Continuous telemetry, BP and oxygen saturation monitored
INSTRUMENT : Olympus Video Bronchoscope
ANESTHESIA : Moderate IV sedation with: fentanyl 50 mcg and Versed 2 mg;
Local anesthesia with: Lidocaine 2% Solution ~8ml via adomizer
Sedation time was 15 min during this procedure
 
 
Time out: Immediately prior to procedure a "time out" was called to verify the correct patient, procedure, equipment, support staff and site/side  marked as required.
SUMMARY : Initially a time-out was performed prior to the procedure, confirming the patient's name, procedure type and procedure location.
After the patient was properly positioned and sedated and topical anesthesia applied, the bronchoscope was introduced through the mouth with O2 being administered at all times.
This was done without difficulty. The bronchoscope was passed by the carina, which was examined for sharpness, position and texture.
The bronchial orifices were systematically identified, evaluated and suctioned free of secretions and close attention was paid to color, texture, positions, size and patency.
FINDINGS
Pharynx: Normal 
Larynx: Normal 
Vocal Chords: Normal 
Trachea: Normal
Carina: Sharp
Bronchial Tree
Right Lung: Normal anastamosis, no strictures, ischemia or black eschar
Left Lung: Normal anastamosis, no strictures, ischemia or black eschar
 
No evidence of mass, lesions, bleeding or other intra-bronchial pathology.
Mucosa: normal
Secretions: none
 
Studies/samples: 
1. BAL 1: 50 cc instilled in the LLL and 20 cc returned
2. BAL 2: 50 cc instilled in the RLL and 25 cc returned
2. Biopsies: 11 TBBX attempted and 11 samples obtained from the RLL
- EBL: none
Complications: None
 
 
IMPRESSION: 
Flexible bronchoscopy with BAL and TBBX under moderate sedation
Post procedure, no dyspnea, chest pain or changes to his voice
Exam: CTA over the area of the lavage and biopsy, no voice change or crepitus noted
Samples sent for microbiology and pathology
Lavage from both sides included cytology order to make sure there is no recurrent CA"""

# =============================================================================
# DATA EXTRACTION (DETERMINISTIC)
# =============================================================================

# 1. Procedure Flags
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 1,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 1,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 2. Spans (Anchor-First Definition)
# Structure: (span_text, label, normalized_value, event_id, context_prefix or None)
SPANS_DATA = [
    # Event 1: General Inspection / Bronchoscopy
    ("Flexible bronchoscopy", "PROC_METHOD", "Flexible Bronchoscopy", "evt_01", None),
    ("Olympus Video Bronchoscope", "DEV_INSTRUMENT", "Olympus Video Bronchoscope", "evt_01", None),
    ("Right Lung: Normal anastamosis", "OBS_ROSE", "Normal anastomosis", "evt_01", None),
    ("Left Lung: Normal anastamosis", "OBS_ROSE", "Normal anastomosis", "evt_01", None),

    # Event 2: BAL 1 (LLL)
    ("BAL 1", "PROC_METHOD", "BAL", "evt_02", None),
    ("LLL", "ANAT_LUNG_LOC", "LLL", "evt_02", "instilled in the "),
    ("50 cc", "MEAS_VOL", "50", "evt_02", "BAL 1: "),
    ("20 cc", "MEAS_VOL", "20", "evt_02", "LLL and "),

    # Event 3: BAL 2 (RLL)
    ("BAL 2", "PROC_METHOD", "BAL", "evt_03", None),
    ("RLL", "ANAT_LUNG_LOC", "RLL", "evt_03", "instilled in the "),
    ("50 cc", "MEAS_VOL", "50", "evt_03", "BAL 2: "),
    ("25 cc", "MEAS_VOL", "25", "evt_03", "RLL and "),

    # Event 4: TBBX (RLL)
    ("TBBX", "PROC_METHOD", "Transbronchial Biopsy", "evt_04", "Biopsies: 11 "),
    ("11", "MEAS_COUNT", "11", "evt_04", "Biopsies: "),
    ("RLL", "ANAT_LUNG_LOC", "RLL", "evt_04", "obtained from the "),
    
    # Global Outcomes / Context
    ("Complications: None", "OUTCOME_COMPLICATION", "None", "global", None),
    ("no dyspnea", "OUTCOME_SYMPTOMS", "No dyspnea", "global", "Post procedure, "),
    ("chest pain", "OUTCOME_SYMPTOMS", "No chest pain", "global", "no dyspnea, "),
    ("no voice change", "OUTCOME_SYMPTOMS", "No voice change", "global", None),
]

# =============================================================================
# LOGIC & UTILITIES
# =============================================================================

def create_workbook():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        # Create a fallback if template missing (minimal structure)
        wb = openpyxl.Workbook()
        for sheet_name in ["Note_Text", "Note_Index", "Span_Annotations", 
                           "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
    return wb

def hydrate_offsets(text, spans):
    hydrated_spans = []
    
    for item in spans:
        span_text, label, norm_val, event_id, context = item
        
        # Defaults
        start = -1
        end = -1
        status = "unknown"
        
        # 1. Exact unique match
        if text.count(span_text) == 1:
            start = text.find(span_text)
            status = "hydrated_unique"
        
        # 2. Context Match
        elif context and text.count(span_text) > 1:
            # Find all occurrences
            occurrences = [m.start() for m in re.finditer(re.escape(span_text), text)]
            found = False
            for occ in occurrences:
                # Look at preceding 120 chars
                window_start = max(0, occ - 120)
                preceding_text = text[window_start:occ]
                if context in preceding_text:
                    start = occ
                    status = "hydrated_context"
                    found = True
                    break
            if not found:
                status = "ambiguous_context_not_found"
        
        # 3. Fallback (First Occurrence if no context provided but multiple exist)
        elif text.count(span_text) > 1 and not context:
            start = text.find(span_text)
            status = "hydrated_first_occurrence_ambiguous"
            
        else:
            status = "not_found"

        if start != -1:
            end = start + len(span_text)
            
        hydrated_spans.append({
            "span_text": span_text,
            "label": label,
            "normalized_value": norm_val,
            "event_id": event_id,
            "context_prefix": context,
            "start_char": start if start != -1 else None,
            "end_char": end if end != -1 else None,
            "hydration_status": status
        })
        
    return hydrated_spans

def generate_v3_structure(hydrated_spans):
    # Group by event
    events = {}
    
    # Pre-process outcomes from global
    global_complication = None
    global_symptoms = []
    
    for s in hydrated_spans:
        if s['event_id'] == 'global':
            if s['label'] == 'OUTCOME_COMPLICATION':
                global_complication = s['normalized_value']
            if s['label'] == 'OUTCOME_SYMPTOMS':
                global_symptoms.append(s['normalized_value'])
                
    for s in hydrated_spans:
        eid = s['event_id']
        if eid == 'global': continue
        
        if eid not in events:
            events[eid] = {
                "type": "Diagnostic" if "bronchoscopy" in s['span_text'].lower() else s['normalized_value'],
                "method": [],
                "target": {"location": {}, "station": None, "anatomy_type": "Airway"},
                "devices": [],
                "measurements": [],
                "findings": [],
                "outcomes": {
                    "complications": global_complication,
                    "symptoms": ", ".join(global_symptoms) if global_symptoms else None
                }
            }
            
        evt = events[eid]
        lbl = s['label']
        val = s['normalized_value']
        
        if lbl == "PROC_METHOD":
            if val not in evt['method']: evt['method'].append(val)
            # Update type to specific if currently generic
            if evt['type'] == "Diagnostic" and val != "Flexible Bronchoscopy":
                evt['type'] = val
                
        elif lbl == "ANAT_LUNG_LOC":
            if "LL" in val or "UL" in val or "ML" in val or "Lingula" in val:
                evt['target']['location']['lobe'] = val
            else:
                evt['target']['location']['segment'] = val
                
        elif lbl == "MEAS_VOL":
            evt['measurements'].append({"type": "volume", "value": val, "unit": "cc"})
        elif lbl == "MEAS_COUNT":
            evt['measurements'].append({"type": "count", "value": val, "unit": "samples"})
            
        elif lbl == "OBS_ROSE":
            evt['findings'].append(val)
            
        elif lbl == "DEV_INSTRUMENT":
            evt['devices'].append({"type": "bronchoscope", "name": val})

    # Flatten for V3 sheet
    v3_rows = []
    for eid, data in events.items():
        row = {
            "note_id": NOTE_ID,
            "event_id": eid,
            "type": data['type'],
            "target.location.lobe": data['target']['location'].get('lobe'),
            "target.location.segment": data['target']['location'].get('segment'),
            "method": ", ".join(data['method']),
            "measurements_json": json.dumps(data['measurements']),
            "findings_json": json.dumps(data['findings']),
            "outcomes.complications": data['outcomes']['complications'],
            "outcomes.symptoms": data['outcomes']['symptoms']
        }
        v3_rows.append(row)
        
    return v3_rows, events

# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    wb = create_workbook()
    
    # 1. Populate Note_Text
    ws_text = wb["Note_Text"]
    # Check headers
    if ws_text.max_row == 1 and ws_text.cell(1,1).value is None:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # 2. Populate Note_Index
    ws_index = wb["Note_Index"]
    # Headers logic omitted for brevity, assuming template exists or raw append
    # Construct row
    meta = [SOURCE_FILE, NOTE_ID, "", "", "", "", "Success", ""]
    flags = [PROCEDURE_FLAGS[k] for k in PROCEDURE_FLAGS]
    ws_index.append(meta + flags)
    
    # 3. Hydrate Spans
    hydrated_data = hydrate_offsets(NOTE_TEXT, SPANS_DATA)
    
    # 4. Populate Span Sheets
    ws_span_anno = wb["Span_Annotations"]
    ws_span_hydra = wb["Span_Hydrated"]
    
    # Headers for reference
    headers = ["source_file", "note_id", "span_id", "section_type", 
               "context_prefix", "span_text", "match_index", 
               "start_char", "end_char", "span_len", 
               "label", "normalized_value", "schema_field", "event_id", 
               "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"]

    if ws_span_anno.max_row == 1 and ws_span_anno.cell(1,1).value is None:
        ws_span_anno.append(headers)
    if ws_span_hydra.max_row == 1 and ws_span_hydra.cell(1,1).value is None:
        ws_span_hydra.append(headers)

    for idx, h in enumerate(hydrated_data):
        span_id = f"{NOTE_ID}_s{idx:03d}"
        
        # Common data
        row_base = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure",
            h['context_prefix'], h['span_text'], None,
            None, None, len(h['span_text']),
            h['label'], h['normalized_value'], None, h['event_id'],
            False, False, None, "Auto", None
        ]
        
        # Annotation Row (No offsets)
        row_anno = row_base + ["needs_hydration"]
        ws_span_anno.append(row_anno)
        
        # Hydrated Row (With offsets)
        row_hydra = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure",
            h['context_prefix'], h['span_text'], None,
            h['start_char'], h['end_char'], len(h['span_text']),
            h['label'], h['normalized_value'], None, h['event_id'],
            False, False, None, "Auto", None, h['hydration_status']
        ]
        ws_span_hydra.append(row_hydra)

    # 5. Populate V3_Procedure_Events & JSON
    v3_rows, v3_tree = generate_v3_structure(hydrated_data)
    ws_v3 = wb["V3_Procedure_Events"]
    
    # Ensure headers if new
    v3_headers = ["note_id", "event_id", "type", "target.anatomy_type", "target.location.lobe", 
                  "target.location.segment", "target.station", "lesion.type", "lesion.size_mm", 
                  "method", "devices_json", "measurements_json", "specimens_json", "findings_json", 
                  "evidence_quote", "stent.size", "stent.material_or_brand", "catheter.size_fr", 
                  "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post", "outcomes.symptoms", 
                  "outcomes.pleural", "outcomes.complications"]
    
    if ws_v3.max_row == 1 and ws_v3.cell(1,1).value is None:
        ws_v3.append(v3_headers)
        
    for row in v3_rows:
        # Map dict to ordered list matching headers
        # Defaults to None if key missing
        xl_row = [
            row.get('note_id'), row.get('event_id'), row.get('type'), "Airway", row.get('target.location.lobe'),
            row.get('target.location.segment'), None, None, None,
            row.get('method'), None, row.get('measurements_json'), None, row.get('findings_json'),
            None, None, None, None,
            None, None, row.get('outcomes.symptoms'),
            None, row.get('outcomes.complications')
        ]
        ws_v3.append(xl_row)

    # 6. Event Log (Simplified flattened view)
    ws_event = wb["Event_Log"]
    if ws_event.max_row == 1 and ws_event.cell(1,1).value is None:
        ws_event.append(["source_file", "note_id", "event_id", "event_type", "method", "anatomy_target", 
                         "measurements", "findings", "outcome_symptoms", "outcome_complication"])
                         
    for row in v3_rows:
        ws_event.append([
            SOURCE_FILE, NOTE_ID, row.get('event_id'), row.get('type'), row.get('method'), 
            row.get('target.location.lobe') or row.get('target.location.segment'),
            row.get('measurements_json'), row.get('findings_json'),
            row.get('outcomes.symptoms'), row.get('outcomes.complications')
        ])

    # 7. V3 Registry JSON
    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 1 and ws_json.cell(1,1).value is None:
        ws_json.append(["schema_version", "note_id", "json_output"])
    
    final_json = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "procedures": v3_tree,
        "no_immediate_complications": True if "None" in str(v3_rows[0].get('outcomes.complications')) else False
    }
    
    ws_json.append(["v3.0", NOTE_ID, json.dumps(final_json, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    main()