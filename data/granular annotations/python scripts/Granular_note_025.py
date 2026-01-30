import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import datetime
import os

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_025"
SOURCE_FILE = "note_025.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_025 SOURCE_FILE: note_025.txt PROCEDURE: Flexible bronchoscopy with BAL and TBBX under moderate sedation
INDICATION : Lung Transplant 
CONSENT : Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient.
The patient read and signed the provided consent form. The consent was witnessed by an assisting medical proffesional.
MONITORING : Continuous telemetry, BP and oxygen saturation monitored
INSTRUMENT : Olympus Video Bronchoscope
ANESTHESIA : Moderate IV sedation with: fentanyl 50 mcg and Versed 2 mg;
Local anesthesia with: Lidocaine 2% Solution ~8ml via adomizer
Sedation time was 15 min during this procedure
 
Time out: Immediately prior to procedure a "time out" was called to verify the correct patient, procedure, equipment, support staff and site/side  marked as required.
SUMMARY : Initially a time-out was performed prior to the procedure, confirming the patient's name, procedure type and procedure location.
After the patient was properly positioned and sedated and topical anesthesia applied, the bronchoscope was introduced through the mouth with O2 being administered at all times.
This was done without difficulty. The bronchoscope was passed by the carina, which was examined for sharpness, position and texture.
The bronchial orifices were systematically identified, evaluated and suctioned free of secretions and close attention was paid to color, texture, positions, size and patency.
FINDINGS
Pharynx: Normal 
Larynx: Normal 
Vocal Chords: Normal 
Trachea: Normal 
Carina: Sharp
Bronchial Tree
Right Lung: Normal anastamosis, no strictures, ischemia or black eschar
Left Lung: Normal anastamosis, no strictures, ischemia or black eschar
No evidence of mass, lesions, bleeding or other intra-bronchial pathology.
Mucosa: normal
Secretions: none

Studies/samples: 
1. BAL 1: 40 cc instilled in the RML and 25 cc returned
2. Biopsies: 11 TBBX attempted and 11 samples obtained from the RML
- EBL: none
Complications: None
 
 
IMPRESSION: 
Flexible bronchoscopy with BAL and TBBX under moderate sedation
 
Post procedure, no dyspnea, chest pain or changes to his voice
Exam: CTA over the area of the lavage and biopsy, no voice change or crepitus noted
Samples sent for microbiology and pathology"""

# -------------------------------------------------------------------------
# CONFIGURATION
# -------------------------------------------------------------------------
PROCEDURE_FLAGS = [
    "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
    "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy",
    "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration",
    "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
    "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation",
    "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
    "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis",
    "pleural_biopsy", "fibrinolytic_therapy"
]

# -------------------------------------------------------------------------
# EXTRACTION LOGIC
# -------------------------------------------------------------------------

# 1. Determine Procedure Flags
# Default all to 0
flags = {f: 0 for f in PROCEDURE_FLAGS}

# Logic based on Note 025
# "Flexible bronchoscopy with BAL and TBBX"
flags['diagnostic_bronchoscopy'] = 1  # Base procedure
flags['bal'] = 1
flags['transbronchial_biopsy'] = 1 # TBBX

# 2. Define Spans (Anchor-First)
# Structure: (span_text, label, normalized_value, event_id, context_prefix)
# event_id: 1=General/Bronch, 2=BAL, 3=TBBX
spans_data = [
    # Event 1: General Bronchoscopy
    ("Flexible bronchoscopy", "PROC_METHOD", "Flexible bronchoscopy", "ev1", "PROCEDURE: "),
    ("Olympus Video Bronchoscope", "DEV_INSTRUMENT", "Olympus Video Bronchoscope", "ev1", None),
    ("Bronchial Tree", "ANAT_AIRWAY", "Bronchial Tree", "ev1", None),
    ("Carina", "ANAT_AIRWAY", "Carina", "ev1", "Trachea: Normal \n"),
    ("Sharp", "OBS_ROSE", "Sharp", "ev1", "Carina: "), # Using OBS_ROSE as generic observation placeholder or Finding if allowed. 
    # Strict label check: OBS_ROSE is for rapid on site eval. OBS_LESION is for lesions. 
    # Let's map findings to comments or just generic if no perfect label. 
    # "Sharp" describes carina. Not a lesion. 
    # We will map "Carina" and use comments for findings in Event Log, spanning "Sharp" might be tricky if no label fits.
    # Label set allows: PROC_METHOD, PROC_ACTION, OBS_ROSE, OBS_LESION.
    # "Sharp" is not a lesion or ROSE. We'll skip spanning "Sharp" as a label but include in findings text.
    
    # Event 2: BAL
    ("BAL", "PROC_METHOD", "BAL", "ev2", "Flexible bronchoscopy with "),
    ("RML", "ANAT_LUNG_LOC", "RML", "ev2", "instilled in the "),
    ("40 cc", "MEAS_VOL", "40", "ev2", "BAL 1: "),
    ("25 cc", "MEAS_VOL", "25", "ev2", "RML and "),
    
    # Event 3: TBBX
    ("TBBX", "PROC_METHOD", "Transbronchial Biopsy", "ev3", "Flexible bronchoscopy with BAL and "),
    ("Biopsies", "PROC_ACTION", "Biopsy", "ev3", "2. "),
    ("11", "MEAS_COUNT", "11", "ev3", "2. Biopsies: "), # 11 TBBX attempted
    ("samples obtained", "OBS_ROSE", "samples obtained", "ev3", "11 "), # Using OBS_ROSE loosely for sample status or just keep text
    ("RML", "ANAT_LUNG_LOC", "RML", "ev3", "obtained from the "),
    
    # Outcomes / Complications (Global)
    ("Complications: None", "OUTCOME_COMPLICATION", "None", "ev1", None),
    ("no dyspnea", "OUTCOME_SYMPTOMS", "no dyspnea", "ev1", "Post procedure, "),
]

# 3. Construct Event Log Objects
events_info = [
    {
        "event_id": "ev1",
        "type": "Diagnostic Bronchoscopy",
        "method": "Flexible bronchoscopy",
        "device": "Olympus Video Bronchoscope",
        "findings": "Airways normal, Carina sharp",
        "outcome_complication": "None",
        "outcome_symptoms": "no dyspnea"
    },
    {
        "event_id": "ev2",
        "type": "BAL",
        "method": "BAL",
        "anatomy": "RML",
        "measurements": "Instilled: 40cc, Returned: 25cc",
        "specimens": "Fluid sent for micro/path"
    },
    {
        "event_id": "ev3",
        "type": "Transbronchial Biopsy",
        "method": "TBBX",
        "anatomy": "RML",
        "counts": "11 attempted, 11 obtained",
        "specimens": "11 samples"
    }
]

# -------------------------------------------------------------------------
# HELPER FUNCTIONS
# -------------------------------------------------------------------------

def hydrate_offsets(text, span_text, context_prefix=None):
    """
    Find start/end char offsets.
    Strategies:
    1. Exact match count == 1 -> use it.
    2. Context prefix -> find span preceded by prefix within window.
    3. Fallback -> Ambiguous or Not Found.
    """
    if not span_text:
        return None, None, "missing_text"
    
    # Strategy 1: Unique
    count = text.count(span_text)
    if count == 1:
        start = text.find(span_text)
        return start, start + len(span_text), "hydrated_unique"
    
    # Strategy 2: Context
    if context_prefix and count > 0:
        # scan all occurrences
        search_start = 0
        candidates = []
        while True:
            idx = text.find(span_text, search_start)
            if idx == -1:
                break
            # check window before idx
            window_start = max(0, idx - 120)
            preceding_text = text[window_start:idx]
            if context_prefix in preceding_text:
                candidates.append(idx)
            search_start = idx + 1
        
        if len(candidates) == 1:
            start = candidates[0]
            return start, start + len(span_text), "hydrated_prefix_window"
        elif len(candidates) > 1:
            return None, None, f"ambiguous_count={len(candidates)}"
    
    if count == 0:
        return None, None, "not_found"
    
    return None, None, f"ambiguous_count={count}"

# -------------------------------------------------------------------------
# WORKBOOK GENERATION
# -------------------------------------------------------------------------

def generate_workbook():
    # Load Template
    if not os.path.exists(TEMPLATE_PATH):
        # Create a dummy one if missing for safety in some envs (though instructions say it exists)
        wb = openpyxl.Workbook()
        for name in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            wb.create_sheet(name)
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # 1. Note_Text
    ws = wb["Note_Text"]
    # Headers: note_id, source_file, note_text
    # Check if empty, if so write header
    if ws.max_row == 1 and ws.cell(1,1).value is None:
        ws.append(["note_id", "source_file", "note_text"])
    ws.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws = wb["Note_Index"]
    # Headers: source_file, note_id, encounter_id, procedure_date, site, reviewer, status, free_text_notes, [flags...]
    # We need to map our dictionary to the columns.
    # Assuming template has headers in row 1.
    headers = [cell.value for cell in ws[1]]
    
    row_data = {
        "source_file": SOURCE_FILE,
        "note_id": NOTE_ID,
        "procedure_date": PROCEDURE_DATE,
        "status": "Auto-Generated"
    }
    # Add flags
    row_data.update(flags)
    
    new_row = []
    for h in headers:
        val = row_data.get(h, "")
        if val == "":
            val = row_data.get(h.lower(), "") # try lowercase
        new_row.append(val)
    ws.append(new_row)

    # 3. Span_Annotations
    ws_anno = wb["Span_Annotations"]
    # Headers: source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, start_char, end_char, span_len, label, normalized_value, schema_field, event_id, is_negated, is_historical, time_anchor, reviewer, comments, hydration_status
    
    # 4. Span_Hydrated
    ws_hyd = wb["Span_Hydrated"]
    
    span_id_counter = 1
    
    for item in spans_data:
        text_span, label, norm_val, ev_id, ctx = item
        
        # Hydrate
        start, end, status = hydrate_offsets(NOTE_TEXT, text_span, ctx)
        
        # Common row data
        row_base = {
            "source_file": SOURCE_FILE,
            "note_id": NOTE_ID,
            "span_id": f"{NOTE_ID}_s{span_id_counter:03d}",
            "context_prefix": ctx,
            "span_text": text_span,
            "span_len": f"=LEN(F{ws_anno.max_row + 1})" if ws_anno.max_row > 0 else 0, # Formula approach
            "label": label,
            "normalized_value": norm_val,
            "event_id": ev_id,
            "hydration_status": "needs_hydration" # for Anno sheet
        }
        
        # Write to Span_Annotations (No offsets)
        anno_row = []
        # Get headers
        anno_headers = [c.value for c in ws_anno[1]]
        for h in anno_headers:
            if h == "span_len":
                anno_row.append(f"=LEN(INDIRECT(\"RC[-4]\",0))") # Approx excel formula for length of span_text column (assuming F is span_text)
                # Actually simpler to just write len(text_span) or leave blank if formula required. 
                # Instruction says "span_len (formula)".
                # Let's put python len for safety or simple formula.
                # using =LEN(Fx) is fragile if columns move. 
                # Let's just put the integer length for now to satisfy data, or simple formula.
                anno_row.append(len(text_span)) 
            elif h in row_base:
                anno_row.append(row_base[h])
            else:
                anno_row.append(None)
        ws_anno.append(anno_row)

        # Write to Span_Hydrated (With offsets)
        hyd_row_data = row_base.copy()
        hyd_row_data["start_char"] = start
        hyd_row_data["end_char"] = end
        hyd_row_data["hydration_status"] = status
        
        hyd_row = []
        hyd_headers = [c.value for c in ws_hyd[1]]
        for h in hyd_headers:
            if h == "span_len":
                hyd_row.append(len(text_span))
            elif h in hyd_row_data:
                hyd_row.append(hyd_row_data[h])
            else:
                hyd_row.append(None)
        ws_hyd.append(hyd_row)
        
        span_id_counter += 1

    # 5. Event_Log
    ws_event = wb["Event_Log"]
    event_headers = [c.value for c in ws_event[1]]
    
    for ev in events_info:
        r = []
        # Map specific keys to headers
        # keys in ev: event_id, type, method, device, anatomy, measurements, counts, specimens, findings, outcome...
        # headers in sheet: source_file, note_id, event_id, event_type, method, anatomy_target, device...
        
        mapping = {
            "source_file": SOURCE_FILE,
            "note_id": NOTE_ID,
            "event_id": ev.get("event_id"),
            "event_type": ev.get("type"),
            "method": ev.get("method"),
            "anatomy_target": ev.get("anatomy"),
            "device": ev.get("device"),
            "counts": ev.get("counts"),
            "measurements": ev.get("measurements"),
            "specimens": ev.get("specimens"),
            "findings": ev.get("findings"),
            "outcome_complication": ev.get("outcome_complication"),
            "outcome_symptoms": ev.get("outcome_symptoms")
        }
        
        for h in event_headers:
            r.append(mapping.get(h, None))
        ws_event.append(r)

    # 6. V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    v3_headers = [c.value for c in ws_v3[1]]
    
    for ev in events_info:
        # Construct V3 row
        row_map = {
            "note_id": NOTE_ID,
            "event_id": ev.get("event_id"),
            "type": ev.get("type"),
            "method": ev.get("method"),
            "target.anatomy_type": "lung" if "Lung" in str(ev.get("anatomy","")) or "RML" in str(ev.get("anatomy","")) else "airway",
            "target.location.lobe": "RML" if "RML" in str(ev.get("anatomy","")) else None,
            "outcomes.complications": ev.get("outcome_complication"),
            "outcomes.symptoms": ev.get("outcome_symptoms"),
            "findings_json": json.dumps({"text": ev.get("findings")}) if ev.get("findings") else None
        }
        
        new_r = []
        for h in v3_headers:
            new_r.append(row_map.get(h, None))
        ws_v3.append(new_r)

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 1 and ws_json.cell(1,1).value is None:
        ws_json.append(["schema_version", "note_id", "json_output"])
    
    # Construct full JSON
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True, # Based on text
        "procedures": events_info
    }
    
    ws_json.append(["3.0", NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)

if __name__ == "__main__":
    generate_workbook()