import re
import datetime
import json
import os
from openpyxl import load_workbook, Workbook

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_026"
SOURCE_FILE = "note_026.txt"
PROCEDURE_DATE = "2026-01-12" # Using current date as placeholder or blank if preferred
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_026 SOURCE_FILE: note_026.txt INDICATION FOR OPERATION:  [REDACTED]is a 30 year old-year-old male who presents with Complicated Effusion.
The nature, purpose, risks, benefits and alternatives to Chest Ultrasound were discussed with the patient in detail.
Patient indicated a wish to proceed with procedure and informed consent was signed.
 
CONSENT : Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient or surrogate.
The patient or surrogate read and signed the provided consent form / provided consent over the phone.
The consent was witnessed by an assisting medical professional.
 
PREOPERATIVE DIAGNOSIS:  Complicated Effusion
POSTOPERATIVE DIAGNOSIS: Same as preoperative diagnosis - see above.
PROCEDURE:  
76604 Ultrasound, chest (includes mediastinum), real time with image documentation
 
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
PROCEDURE IN DETAIL:
 
 
PATIENT POSITION: 
0‌ Supine  1‌ Sitting   
0‌ Lateral Decubitus:  0‌ Right 0‌ Left 
 
CHEST ULTRASOUND FINDINGS:  1‌ Image saved and up loaded to PACS
Hemithorax:   0‌ Right  1‌ Left 
 
Pleural Effusion: 
Volume:       0‌ None  1‌ Minimal  0‌ Small  0‌ Moderate  0‌ Large 
Echogenicity:   1‌ Anechoic  0‌ Hypoechoic  0‌ Isoechoic  0‌ Hyperechoic 
Loculations:  0‌ None  1‌Thin  0‌ Thick 
Diaphragmatic Motion:  1‌ Normal  0‌ Diminished  0‌ Absent  

Lung: 
Lung sliding before procedure:   1‌ Present  0‌ Absent 
Lung sliding post procedure:   0‌ Present  0‌ Absent 
Lung consolidation/atelectasis: 0‌ Present  0‌  Absent 
Pleura:  1‌ Normal  0‌ Thick  0‌ Nodular 
 
Atelectasis improved and posterior collection trace compared to previous day.
No drainable fluid.
 
COMPLICATIONS:
1‌None 0‌Bleeding-EBL: ___ ml 0‌Pneumothorax 0‌Re- Expansion Pulmonary Edema 
0‌Other: 
 
IMPRESSION/PLAN: [REDACTED]is a 30 year old-year-old male who presents for Chest Ultrasound.
The patient tolerated the procedure well.  There were no immediate complications.  
 
 
DISPOSITION: Nursing Unit
D/c chest tube"""

# -------------------------------------------------------------------------
# CONFIGURATION
# -------------------------------------------------------------------------
# Procedure flags
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 0, "bal": 0, "bronchial_wash": 0, "brushings": 0,
    "endobronchial_biopsy": 0, "tbna_conventional": 0, "linear_ebus": 0, "radial_ebus": 0,
    "navigational_bronchoscopy": 0, "transbronchial_biopsy": 0, "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0, "foreign_body_removal": 0, "airway_dilation": 0,
    "airway_stent": 0, "thermal_ablation": 0, "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0, "blvr": 0, "peripheral_ablation": 0, "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0, "rigid_bronchoscopy": 0,
    "thoracentesis": 0, "chest_tube": 0, "ipc": 0, "medical_thoracoscopy": 0,
    "pleurodesis": 0, "pleural_biopsy": 0, "fibrinolytic_therapy": 0
}

# Although "Chest Ultrasound" isn't a flag, we process the note. 
# "D/c chest tube" is a plan, not necessarily a procedure performed *during* this note's procedure block (Ultrasound).
# We will keep flags as 0 but capture the data in spans/events.

# -------------------------------------------------------------------------
# EXTRACTION LOGIC
# -------------------------------------------------------------------------
spans_data = []

def add_span(text_match, label, norm_val, field, event_id, context_prefix=None, is_negated=False):
    spans_data.append({
        "span_text": text_match,
        "label": label,
        "normalized_value": norm_val,
        "schema_field": field,
        "event_id": event_id,
        "context_prefix": context_prefix,
        "is_negated": is_negated
    })

# Event 1: Chest Ultrasound (Diagnostic)
ev1 = "evt_01"

# Indication
add_span("Complicated Effusion", "OBS_LESION", "complicated effusion", "indication", ev1)

# Laterality / Anatomy
add_span("Left", "LATERALITY", "left", "target.location.laterality", ev1, context_prefix="Hemithorax:   0‌ Right  1‌ ")

# Findings - Effusion
add_span("Pleural Effusion", "ANAT_PLEURA", "pleural effusion", "target.anatomy", ev1)
add_span("Minimal", "MEAS_VOL", "minimal", "findings.effusion_volume", ev1, context_prefix="Volume:       0‌ None  1‌ ")
add_span("Anechoic", "OBS_ROSE", "anechoic", "findings.echogenicity", ev1)
add_span("Thin", "OBS_LESION", "thin loculations", "findings.loculations", ev1, context_prefix="Loculations:  0‌ None  1‌")
add_span("Normal", "OBS_LESION", "normal diaphragmatic motion", "findings.diaphragm", ev1, context_prefix="Diaphragmatic Motion:  1‌ ")

# Findings - Lung/Pleura
add_span("Lung sliding before procedure:   1‌ Present", "OBS_LESION", "lung sliding present", "findings.lung_sliding", ev1)
add_span("Pleura:  1‌ Normal", "OBS_LESION", "pleura normal", "findings.pleura", ev1)

# Specific Findings Text
add_span("Atelectasis improved", "OBS_LESION", "atelectasis improved", "findings.other", ev1)
add_span("posterior collection trace", "OBS_LESION", "posterior collection trace", "findings.other", ev1)
add_span("No drainable fluid", "OBS_LESION", "no drainable fluid", "findings.fluid", ev1)

# Procedure Name
add_span("Ultrasound, chest", "PROC_METHOD", "chest ultrasound", "procedure_name", ev1)

# Complications
add_span("no immediate complications", "OUTCOME_COMPLICATION", "none", "outcomes.complications", ev1)

# Disposition (Event 2 - Plan/Action)
ev2 = "evt_02"
add_span("D/c chest tube", "PROC_ACTION", "discontinue chest tube", "plan.action", ev2)

# -------------------------------------------------------------------------
# WORKBOOK GENERATION
# -------------------------------------------------------------------------
def generate_excel(template_path: str = TEMPLATE_PATH, output_path: str = OUTPUT_PATH) -> None:
    try:
        wb = load_workbook(template_path)
    except FileNotFoundError:
        # Create blank if template missing (fallback for demo)
        wb = Workbook()
        for name in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            wb.create_sheet(name)

    # 1. Note_Text
    ws_text = wb["Note_Text"]
    if ws_text.max_row == 1:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws_index = wb["Note_Index"]
    row_meta = [
        SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Success", ""
    ]
    # Append flags
    flag_values = [PROCEDURE_FLAGS[k] for k in PROCEDURE_FLAGS]
    ws_index.append(row_meta + flag_values)

    # 3. Span_Annotations
    ws_span = wb["Span_Annotations"]
    span_rows = []
    for idx, s in enumerate(spans_data):
        span_id = f"span_{idx+1:03d}"
        row = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure",
            s.get("context_prefix", ""), s["span_text"], "", "", "", f'=LEN(F{ws_span.max_row + 1 + idx})',
            s["label"], s["normalized_value"], s["schema_field"], s["event_id"],
            s["is_negated"], False, "", "", "", "needs_hydration"
        ]
        span_rows.append(row)
        ws_span.append(row)

    # 4. Span_Hydrated
    ws_hyd = wb["Span_Hydrated"]

    def find_offsets(text, span, prefix=None):
        count = text.count(span)
        if count == 0:
            return None, None, "not_found"
        if count == 1:
            start = text.find(span)
            return start, start + len(span), "hydrated_unique"

        if prefix:
            starts = [m.start() for m in re.finditer(re.escape(span), text)]
            for st in starts:
                window = text[max(0, st-120):st]
                if prefix in window:
                    return st, st + len(span), "hydrated_prefix_window"

        return None, None, f"ambiguous_count={count}"

    for row_data in span_rows:
        span_text = row_data[5]
        prefix = row_data[4]

        start, end, status = find_offsets(NOTE_TEXT, span_text, prefix)

        new_row = list(row_data)
        new_row[7] = start if start is not None else ""
        new_row[8] = end if end is not None else ""
        new_row[19] = status
        ws_hyd.append(new_row)

    # 5. Event_Log
    ws_event = wb["Event_Log"]
    events = [
        {
            "id": "evt_01", "type": "Diagnostic", "method": "Ultrasound",
            "findings": "Left pleural effusion (minimal, anechoic, thin loculations), No drainable fluid",
            "outcome_complication": "none"
        },
        {
            "id": "evt_02", "type": "Plan", "method": "Discontinue Chest Tube",
            "findings": "D/c chest tube"
        }
    ]

    for e in events:
        row = [
            SOURCE_FILE, NOTE_ID, e["id"], e["type"], e["method"],
            "", "", "", "", "", "", "", e.get("findings", ""), False, "", "",
            "", "", "", "", "", "", e.get("outcome_complication", "")
        ]
        ws_event.append(row)

    # 6. V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    for e in events:
        row = [
            NOTE_ID, e["id"], e["type"],
            "", "", "", "",
            "", "",
            e["method"],
            "", "", "", json.dumps(e.get("findings", "")), "",
            "", "", "",
            "", "", "", "", e.get("outcome_complication", "")
        ]
        ws_v3.append(row)

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    json_obj = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": events
    }
    ws_json.append([json.dumps(json_obj, indent=2)])

    wb.save(output_path)


if __name__ == "__main__":
    generate_excel()
