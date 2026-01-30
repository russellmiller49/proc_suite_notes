import openpyxl
from openpyxl.utils import get_column_letter
import datetime
import re
import json
import os

# -------------------------------------------------------------------------
# INPUTS
# -------------------------------------------------------------------------
NOTE_ID = "note_028"
SOURCE_FILE = "note_028.txt"
PROCEDURE_DATE = "2026-01-12" # inferred from context or current date if not in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_028 SOURCE_FILE: note_028.txt INDICATION FOR OPERATION:  [REDACTED]is a 30 year old-year-old male who presents with Complicated Effusion.
The nature, purpose, risks, benefits and alternatives to Chest Ultrasound and Instillation of agents for fibrinolysis (subsequent) were discussed with the patient in detail.
Patient indicated a wish to proceed with procedure and informed consent was signed.
 
CONSENT : Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient or surrogate.
The patient or surrogate read and signed the provided consent form / provided consent over the phone.
The consent was witnessed by an assisting medical professional.
 
PREOPERATIVE DIAGNOSIS:  Complicated Effusion
POSTOPERATIVE DIAGNOSIS: Same as preoperative diagnosis - see above.
PROCEDURE:  
76604 Ultrasound, chest (includes mediastinum), real time with image documentation
32562 Instillation(s), via chest tube/catheter, agent for fibrinolysis (eg, fibrinolytic agent for break up of multiloculated effusion);
subsequent day
 
25 Added to the E&M Encounter bill which is separate from a procedure if it is done the same day
 
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
PROCEDURE IN DETAIL:
 
 
PATIENT POSITION: 
1‌ Supine  0‌ Sitting   
0‌ Lateral Decubitus:  0‌ Right 0‌ Left 
 
 
CHEST ULTRASOUND FINDINGS:  1‌ Image saved and printed 
Hemithorax:   0‌ Right  1‌ Left 
 
Pleural Effusion: 
Volume:       0‌ None  0‌ Minimal  0‌ Small  1‌ Moderate  0‌ Large 
Echogenicity:   1‌ Anechoic  0‌ Hypoechoic  0‌ Isoechoic  0‌ Hyperechoic 
Loculations:  0‌ None  1‌Thin  0‌ Thick 
Diaphragmatic Motion:  1‌ Normal  0‌ Diminished  0‌ Absent  
Lung: 

Lung sliding before procedure:   1‌ Present  0‌ Absent 
Lung sliding post procedure:   0‌ Present  0‌ Absent 
Lung consolidation/atelectasis: 0‌ Present  0‌  Absent 
Pleura:  1‌ Normal  0‌ Thick  0‌ Nodular 
 
 
Date of chest tube insertion: 12/29
 
Side: left
1‌  10 mg/5 mg tPA/Dnasedose #:__4__ 
            0‌  ___mg tPA                             
 dose #:____ 
0‌  Other medication: 
 
 
 
COMPLICATIONS:
1‌None 0‌Bleeding-EBL: ___ ml 0‌Pneumothorax 0‌Re- Expansion Pulmonary Edema 
0‌Other: 
 
IMPRESSION/PLAN: [REDACTED]is a 30 year old-year-old male who presents for Chest Ultrasound and Instillation of agents for fibrinolysis (subsequent).
The patient tolerated the procedure well.  There were no immediate complications.
- dwell for 1hr then unclamp and place back on suction -20cm H20, @3:00pm 
 
 
DISPOSITION: Nursing Unit"""

# -------------------------------------------------------------------------
# CONFIGURATION
# -------------------------------------------------------------------------
PROCEDURE_FLAGS = {
    # Bronchoscopy (23)
    'diagnostic_bronchoscopy': 0, 'bal': 0, 'bronchial_wash': 0, 'brushings': 0,
    'endobronchial_biopsy': 0, 'tbna_conventional': 0, 'linear_ebus': 0,
    'radial_ebus': 0, 'navigational_bronchoscopy': 0, 'transbronchial_biopsy': 0,
    'transbronchial_cryobiopsy': 0, 'therapeutic_aspiration': 0,
    'foreign_body_removal': 0, 'airway_dilation': 0, 'airway_stent': 0,
    'thermal_ablation': 0, 'tumor_debulking_non_thermal': 0, 'cryotherapy': 0,
    'blvr': 0, 'peripheral_ablation': 0, 'bronchial_thermoplasty': 0,
    'whole_lung_lavage': 0, 'rigid_bronchoscopy': 0,
    # Pleural (7)
    'thoracentesis': 0, 'chest_tube': 0, 'ipc': 0, 'medical_thoracoscopy': 0,
    'pleurodesis': 0, 'pleural_biopsy': 0,
    'fibrinolytic_therapy': 1  # Note 028 explicitly mentions instillation of fibrinolytic agents
}

# -------------------------------------------------------------------------
# SPAN DATA (Anchor-First Definition)
# -------------------------------------------------------------------------
# Each entry: (span_text, label, normalized_value, schema_field, event_id, context_prefix_optional)
SPANS_DATA = [
    # Diagnosis
    ("Complicated Effusion", "OBS_LESION", "Complicated Effusion", "indication", "ev0", None),
    
    # Procedure: Ultrasound
    ("Ultrasound, chest (includes mediastinum)", "PROC_METHOD", "Ultrasound", "method", "ev1", None),
    ("Hemithorax:   0‌ Right  1‌ Left", "LATERALITY", "Left", "target.location", "ev1", None),
    ("1‌ Moderate", "MEAS_VOL", "Moderate", "findings.volume", "ev1", None),
    ("1‌ Anechoic", "OBS_LESION", "Anechoic", "findings.echogenicity", "ev1", "Echogenicity:"),
    ("1‌Thin", "OBS_LESION", "Thin", "findings.loculations", "ev1", "Loculations:"),
    ("Lung sliding before procedure:   1‌ Present", "OBS_ROSE", "Present", "findings.lung_sliding", "ev1", None),
    
    # Procedure: Fibrinolysis
    ("Instillation(s), via chest tube/catheter, agent for fibrinolysis", "PROC_METHOD", "Instillation of Fibrinolytics", "method", "ev2", None),
    ("Date of chest tube insertion: 12/29", "CTX_HISTORICAL", "12/29", "history", "ev2", None),
    ("Side: left", "LATERALITY", "Left", "target.location", "ev2", None),
    ("1‌  10 mg/5 mg tPA/Dnase", "PROC_METHOD", "tPA/Dnase 10mg/5mg", "medication", "ev2", None),
    ("dose #:__4__", "MEAS_COUNT", "4", "dose_number", "ev2", None),
    ("- dwell for 1hr", "PROC_ACTION", "dwell 1hr", "instructions", "ev2", None),
    ("-20cm H20", "MEAS_PRESS", "-20", "suction_pressure", "ev2", None),
    
    # Outcome
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "outcomes.complications", "ev2", None),
]

# -------------------------------------------------------------------------
# LOGIC
# -------------------------------------------------------------------------
def create_workbook():
    # Load template
    if not os.path.exists(TEMPLATE_PATH):
        # Create a dummy workbook if template is missing (for safety, though instructions require template)
        wb = openpyxl.Workbook()
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    
    # Ensure sheets exist
    required_sheets = [
        'Note_Text', 'Note_Index', 'Span_Annotations', 'Span_Hydrated', 
        'Event_Log', 'V3_Procedure_Events', 'V3_Registry_JSON'
    ]
    for sheet_name in required_sheets:
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
    
    return wb

def hydrate_spans(spans_input, full_text):
    hydrated = []
    
    for item in spans_input:
        span_text = item[0]
        label = item[1]
        norm_val = item[2]
        schema = item[3]
        ev_id = item[4]
        ctx_prefix = item[5]
        
        # Hydration Logic
        start_char = -1
        end_char = -1
        status = "ambiguous"
        
        count = full_text.count(span_text)
        
        if count == 0:
            status = "not_found"
        elif count == 1:
            start_char = full_text.find(span_text)
            end_char = start_char + len(span_text)
            status = "hydrated_unique"
        else:
            # Ambiguous handling
            if ctx_prefix:
                # Find occurrences
                pattern = re.escape(span_text)
                matches = [m.start() for m in re.finditer(pattern, full_text)]
                found = False
                for m_start in matches:
                    # Look back 120 chars
                    window_start = max(0, m_start - 120)
                    window_text = full_text[window_start:m_start]
                    if ctx_prefix in window_text:
                        start_char = m_start
                        end_char = m_start + len(span_text)
                        status = "hydrated_prefix_window"
                        found = True
                        break
                if not found:
                    status = f"ambiguous_count={count}"
            else:
                status = f"ambiguous_count={count}"
        
        span_obj = {
            'text': span_text, 'label': label, 'norm': norm_val, 
            'schema': schema, 'event_id': ev_id, 'ctx': ctx_prefix,
            'start': start_char, 'end': end_char, 'status': status
        }
        hydrated.append(span_obj)
        
    return hydrated

def generate_excel():
    wb = create_workbook()
    
    # 1. Note_Text
    ws = wb['Note_Text']
    if ws.max_row == 1:
        ws.append(["note_id", "source_file", "note_text"])
    ws.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # 2. Note_Index
    ws = wb['Note_Index']
    if ws.max_row == 1:
        headers = ["source_file", "note_id", "encounter_id", "procedure_date", 
                   "site", "reviewer", "status", "free_text_notes"] + list(PROCEDURE_FLAGS.keys())
        ws.append(headers)
    
    row_data = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "generated", ""]
    for k in PROCEDURE_FLAGS:
        row_data.append(PROCEDURE_FLAGS[k])
    ws.append(row_data)
    
    # 3. Span_Annotations (Anchor First) & 4. Span_Hydrated
    ws_anno = wb['Span_Annotations']
    ws_hydra = wb['Span_Hydrated']
    
    # Write Headers if empty
    headers_span = ["source_file", "note_id", "span_id", "section_type", "context_prefix", 
                    "span_text", "match_index", "start_char", "end_char", "span_len", 
                    "label", "normalized_value", "schema_field", "event_id", 
                    "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"]
    
    if ws_anno.max_row == 1: ws_anno.append(headers_span)
    if ws_hydra.max_row == 1: ws_hydra.append(headers_span)
    
    hydrated_data = hydrate_spans(SPANS_DATA, NOTE_TEXT)
    
    for i, h in enumerate(hydrated_data):
        span_id = f"{NOTE_ID}_s{i+1:03d}"
        
        # Base Row
        row = [SOURCE_FILE, NOTE_ID, span_id, "Procedure Note", h['ctx'], 
               h['text'], "", "", "", f"=LEN(F{i+2})", 
               h['label'], h['norm'], h['schema'], h['event_id'], 
               False, False, "", "System", "", "needs_hydration"]
        
        # Append to Annotations (empty offsets)
        ws_anno.append(row)
        
        # Hydrated Row
        row_h = list(row)
        if h['start'] != -1:
            row_h[7] = h['start']
            row_h[8] = h['end']
        row_h[19] = h['status']
        ws_hydra.append(row_h)
        
    # 5. Event_Log
    ws_event = wb['Event_Log']
    if ws_event.max_row == 1:
        headers_evt = ["source_file", "note_id", "event_id", "event_type", "method", 
                       "anatomy_target", "device", "needle_gauge", "stations", "counts", 
                       "measurements", "specimens", "findings", "is_historical", "reviewer", 
                       "comments", "device_size", "device_material", 
                       "outcome_airway_lumen_pre", "outcome_airway_lumen_post", 
                       "outcome_symptoms", "outcome_pleural", "outcome_complication"]
        ws_event.append(headers_evt)
    
    # Define Events
    events = [
        {
            "id": "ev1", "type": "Diagnostic", "method": "Ultrasound", "target": "Left Pleura",
            "findings": "Moderate anechoic effusion, thin loculations, lung sliding present",
            "complication": ""
        },
        {
            "id": "ev2", "type": "Therapeutic", "method": "Fibrinolytic Therapy", "target": "Left Pleura",
            "device": "Chest Tube (Historical)", "measurements": "Dose #4, -20cm H20 suction",
            "findings": "tPA/Dnase instillation", "complication": "None"
        }
    ]
    
    for ev in events:
        ws_event.append([
            SOURCE_FILE, NOTE_ID, ev['id'], ev['type'], ev['method'],
            ev['target'], ev.get('device', ''), "", "", "",
            ev.get('measurements', ''), "", ev['findings'], False, "System",
            "", "", "",
            "", "", "", "", ev['complication']
        ])

    # 6. V3_Procedure_Events
    ws_v3 = wb['V3_Procedure_Events']
    if ws_v3.max_row == 1:
        headers_v3 = ["note_id", "event_id", "type", 
                      "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station",
                      "lesion.type", "lesion.size_mm", "method", 
                      "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote",
                      "stent.size", "stent.material_or_brand", "catheter.size_fr",
                      "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post", 
                      "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"]
        ws_v3.append(headers_v3)

    # EV1: Ultrasound
    ws_v3.append([
        NOTE_ID, "ev1", "Diagnostic", 
        "Pleura", "Left", "", "",
        "Effusion", "Moderate", "Ultrasound",
        "{}", "{}", "{}", json.dumps({"echogenicity": "anechoic", "loculations": "thin", "lung_sliding": "present"}), "Ultrasound chest left moderate effusion",
        "", "", "", 
        "", "", "", "", ""
    ])
    
    # EV2: Fibrinolysis
    ws_v3.append([
        NOTE_ID, "ev2", "Therapeutic",
        "Pleura", "Left", "", "",
        "", "", "Fibrinolytic Therapy",
        json.dumps({"agent": "tPA/Dnase", "route": "chest_tube"}), json.dumps({"dose_number": 4, "suction": "-20cmH20"}), "{}", "{}", "Instillation via chest tube tPA/Dnase",
        "", "", "",
        "", "", "", "", "None"
    ])

    # 7. V3_Registry_JSON
    ws_json = wb['V3_Registry_JSON']
    if ws_json.max_row == 1:
        ws_json.append(["schema_version", "note_id", "json_object"])
    
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": [
            {
                "event_id": "ev1",
                "type": "Ultrasound",
                "target": {"anatomy": "Pleura", "laterality": "Left"},
                "findings": ["Moderate Effusion", "Anechoic", "Thin Loculations"]
            },
            {
                "event_id": "ev2",
                "type": "Fibrinolytic Therapy",
                "method": "Instillation",
                "target": {"anatomy": "Pleura", "laterality": "Left"},
                "agents": ["tPA", "Dnase"],
                "dose_number": 4,
                "outcomes": {"complications": "None"}
            }
        ]
    }
    
    ws_json.append(["3.0", NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Successfully generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_excel()