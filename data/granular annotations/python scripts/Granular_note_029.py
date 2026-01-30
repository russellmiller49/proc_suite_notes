import re
import json
import datetime
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# ==========================================
# 1. INPUT DATA & CONFIGURATION
# ==========================================
NOTE_ID = "note_029"
SOURCE_FILE = "note_029.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# Full text from the provided file
NOTE_TEXT = """NOTE_ID:  note_029 SOURCE_FILE: note_029.txt INDICATION FOR OPERATION:  [REDACTED]is a 30 year old-year-old male who presents with Pleural Effusion.
The nature, purpose, risks, benefits and alternatives to Chest Ultrasound and Chest tube placement were discussed with the patient in detail.
Patient indicated a wish to proceed with procedure and informed consent was signed.
 
CONSENT : Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient or surrogate.
The patient or surrogate read and signed the provided consent form / provided consent over the phone.
The consent was witnessed by an assisting medical professional.
 
PREOPERATIVE DIAGNOSIS:  Pleural Effusion
POSTOPERATIVE DIAGNOSIS: Same as preoperative diagnosis - see above.
PROCEDURE:  
76604 Ultrasound, chest (includes mediastinum), real time with image documentation
32557 Insert catheter pleura with imaging (chest tube)
 
25 Added to the E&M Encounter bill which is separate from a procedure if it is done the same day
 MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
PROCEDURE IN DETAIL:
 
 
PATIENT POSITION: 
1‌ Supine  0‌ Sitting   
0‌ Lateral Decubitus:  0‌ Right 0‌ Left 
 
CHEST ULTRASOUND FINDINGS:  1‌ Image saved and printed 
Hemithorax:   0‌ Right  1‌ Left 
 
Pleural Effusion: 
Volume:       0‌ None  0‌ Minimal  0‌ Small  0‌ Moderate  1‌ Large 
Echogenicity:   1‌ Anechoic  0‌ Hypoechoic  0‌ Isoechoic  0‌ Hyperechoic 
Loculations:  0‌ None  1‌Thin  0‌ Thick 
Diaphragmatic Motion:  1‌ Normal  0‌ Diminished  0‌ Absent  
Lung: 
Lung 
sliding before procedure:   1‌ Present  0‌ Absent 
Lung sliding post procedure:   0‌ Present  0‌ Absent 
Lung consolidation/atelectasis: 0‌ Present  0‌  Absent 
Pleura:  1‌ Normal  0‌ Thick  0‌ Nodular 
 
Insertion site prepped and draped in sterile fashion.
ANESTHESIA:   Lidocaine 1%: ___15___ ml      Other: ______ 
Entry Site: 
0‌ Right ___ Intercostal Space   1‌ Left  _4th__ Intercostal Space 
0‌ Mid-clavicular   1‌ Mid-axillary  0‌ Mid-scapular  0‌ Other: 
 
Size:  0‌ 6Fr  0‌ 8Fr   0‌ 12FR   1‌ 14Fr  0‌ 16Fr   0‌ 18Fr  0‌ 24Fr   0‌ 32 Fr   0‌ Other: 
Sutured: 1‌ Yes 0‌ No 
 
PROCEDURE FINDINGS: 
A  pigtail catheter was inserted using the Seldinger technique.
Entry into the pleural space was confirmed with the easy removal of minimal serous appearing pleural fluid and air bubbles.
A guidewire was inserted using the introducer needle pointed in the apical posterior direction. The introducer needle was then removed.
A dilator was then inserted over the wire with a twisting motion in order to form a tract for catheter insertion.
The dilator was removed and the pigtail catheter (with trochar) was advanced over the guidewire.
The catheter was inserted into the chest until all catheter holes were well within the chest.
The guidewire and trochar were then removed.  The tube was then attached to the collection drain apparatus and secured in place with suture and covered.
Fluid Removed: __300___ ml 
1‌ Serous  0‌ Serosanguinous 0‌ Bloody  0‌ Chylous 0‌ Other: 
 
Pleural pressures:   If not measured, please check here 1‌
Opening:    cmH20   
500 ml:       cmH20   
1000ml:      cmH20       
1500ml:      cmH20   
2000ml       cmH20.
Drainage device:   1‌ Pleurovac    0‌ Drainage Bag  0‌ Heimlich Valve  0‌ Pneumostat  0‌ Other: 
Suction: 1‌ No 0‌Yes, - ___ cmH20 
 
 
SPECIMEN(S): 
0‌None           1‌PH               1‌ LDH                        1‌Glucose       1‌T.
Protein    1‌Cholesterol
1‌Cell Count   0‌ ADA             0‌Triglycerides            0‌Amylase 
1‌Gram Stain/ Culture            1‌AFB                         1‌Fungal Culture 
1‌Cytology      0‌Flow Cytometry                 
              0‌Other: 
 
CXR ordered: 1‌ Yes 0‌ No 
 
 
 
COMPLICATIONS:
1‌None 0‌Bleeding-EBL: ___ ml 0‌Pneumothorax 0‌Re- Expansion Pulmonary Edema 
0‌Other: 
 
IMPRESSION/PLAN: [REDACTED]is a 30 year old-year-old male who presents for Chest Ultrasound and Chest tube placement on the LEFT.
The patient tolerated the procedure well.  There were no immediate complications.
- f/u post-op CXR
- f/u pleural fluid studies 
- keep chest tube to waterseal
- small bore chest tube flushing q8h as per orders
- daily CXR while chest tube is in place 
 
 
DISPOSITION: Nursing Unit"""

# ==========================================
# 2. DEFINITIONS & SCHEMA
# ==========================================

# Procedure Flags (Binary 0/1)
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 0, "bal": 0, "bronchial_wash": 0, "brushings": 0,
    "endobronchial_biopsy": 0, "tbna_conventional": 0, "linear_ebus": 0,
    "radial_ebus": 0, "navigational_bronchoscopy": 0, "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0, "therapeutic_aspiration": 0,
    "foreign_body_removal": 0, "airway_dilation": 0, "airway_stent": 0,
    "thermal_ablation": 0, "tumor_debulking_non_thermal": 0, "cryotherapy": 0,
    "blvr": 0, "peripheral_ablation": 0, "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0, "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,  # Part of chest tube, not standalone
    "chest_tube": 1,
    "ipc": 0, "medical_thoracoscopy": 0, "pleurodesis": 0,
    "pleural_biopsy": 0, "fibrinolytic_therapy": 0
}

# Annotated Spans (Anchor-First)
# Structure: (span_text, label, normalized_value, schema_field, event_id, context_prefix)
RAW_SPANS = [
    # Event 1: Chest Ultrasound
    ("Ultrasound, chest", "PROC_METHOD", "Ultrasound", "method", "event_1", None),
    ("Left", "LATERALITY", "Left", "target.location.side", "event_1", "Hemithorax:   0‌ Right  1‌ "),
    ("Large", "MEAS_SIZE", "Large", "lesion.size_mm", "event_1", "0‌ Moderate  1‌ "),
    ("Anechoic", "OBS_LESION", "Anechoic", "findings", "event_1", None),
    ("Thin", "OBS_LESION", "Thin Loculations", "findings", "event_1", "Loculations:  0‌ None  1‌"),
    
    # Event 2: Chest Tube Placement
    ("Insert catheter pleura", "PROC_METHOD", "Chest Tube Insertion", "method", "event_2", None),
    ("Left", "LATERALITY", "Left", "target.location.side", "event_2", "Right ___ Intercostal Space   1‌ "),
    ("4th Intercostal Space", "ANAT_LUNG_LOC", "4th Intercostal Space", "target.location.segment", "event_2", None),
    ("Mid-axillary", "ANAT_LUNG_LOC", "Mid-axillary", "target.location.segment", "event_2", None),
    ("14Fr", "DEV_CATHETER_SIZE", "14Fr", "catheter.size_fr", "event_2", None),
    ("pigtail catheter", "DEV_CATHETER", "Pigtail", "device", "event_2", "The dilator was removed and the "),
    ("Seldinger technique", "PROC_METHOD", "Seldinger technique", "method", "event_2", None),
    ("300", "MEAS_VOL", "300", "outcomes.pleural", "event_2", "Fluid Removed: __"),
    ("Serous", "OBS_ROSE", "Serous", "specimens", "event_2", "1‌ "),
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "outcomes.complications", "event_2", None),
]

# Event Metadata
EVENTS = [
    {
        "event_id": "event_1",
        "event_type": "Diagnostic",
        "method": "Ultrasound",
        "anatomy_target": "Pleura",
        "findings": "Large Anechoic Effusion",
    },
    {
        "event_id": "event_2",
        "event_type": "Therapeutic",
        "method": "Chest Tube",
        "anatomy_target": "Pleural Space",
        "device": "Pigtail Catheter",
        "device_size": "14Fr",
        "outcome_pleural": "300ml drained",
        "outcome_complication": "None"
    }
]

# ==========================================
# 3. HELPER FUNCTIONS
# ==========================================

def get_start_end(text, span, context_prefix=None):
    """
    Finds start/end indices of a span in text.
    Uses context_prefix to disambiguate if provided.
    """
    if context_prefix:
        # Search for prefix + span, but with loose whitespace handling
        # Escape special regex chars in prefix and span
        pattern_str = re.escape(context_prefix) + r"\s*" + re.escape(span)
        pattern = re.compile(pattern_str, re.IGNORECASE | re.DOTALL)
        match = pattern.search(text)
        if match:
            # The span is at the end of the match
            full_match_text = match.group(0)
            # Find where the span actually starts within the match
            span_start_in_match = full_match_text.rfind(span)
            start_idx = match.start() + span_start_in_match
            end_idx = start_idx + len(span)
            return start_idx, end_idx
            
    # Fallback: Count occurrences
    count = text.count(span)
    if count == 1:
        start = text.find(span)
        return start, start + len(span)
    
    # If ambiguous and no context provided, return -1 (will flag as needs_hydration)
    return -1, -1

def create_worksheets(wb):
    """Creates specific worksheets if they don't exist."""
    required_sheets = [
        "Note_Text", "Note_Index", "Span_Annotations", 
        "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"
    ]
    
    # Remove default sheet if empty
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
        
    for sheet_name in required_sheets:
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)

def generate_workbook():
    try:
        wb = load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        wb = Workbook()
        create_worksheets(wb)

    # 1. Populate Note_Text
    ws_text = wb["Note_Text"]
    if ws_text.max_row == 1:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Populate Note_Index
    ws_index = wb["Note_Index"]
    # Headers usually exist, if not we assume template structure. 
    # Appending metadata row
    row_meta = [
        SOURCE_FILE, NOTE_ID, "", "", "", "", "Ready", ""
    ]
    # Append flags
    flag_values = [PROCEDURE_FLAGS.get(k, 0) for k in PROCEDURE_FLAGS.keys()]
    ws_index.append(row_meta + flag_values)

    # 3. Populate Span_Annotations (Anchor First - No Offsets)
    ws_anno = wb["Span_Annotations"]
    if ws_anno.max_row == 1:
        ws_anno.append([
            "source_file", "note_id", "span_id", "section_type",
            "context_prefix", "span_text", "match_index",
            "start_char", "end_char", "span_len",
            "label", "normalized_value", "schema_field", "event_id",
            "is_negated", "is_historical", "time_anchor", "reviewer", "comments",
            "hydration_status"
        ])

    # 4. Populate Span_Hydrated (Calculated Offsets)
    ws_hydrated = wb["Span_Hydrated"]
    if ws_hydrated.max_row == 1:
        ws_hydrated.append([
            "source_file", "note_id", "span_id", "section_type",
            "context_prefix", "span_text", "match_index",
            "start_char", "end_char", "span_len",
            "label", "normalized_value", "schema_field", "event_id",
            "is_negated", "is_historical", "time_anchor", "reviewer", "comments",
            "hydration_status"
        ])

    span_counter = 1
    hydrated_rows = []

    for item in RAW_SPANS:
        span_text, label, norm_val, schema_field, event_id, context_prefix = item
        
        # Calculate Offsets
        start, end = get_start_end(NOTE_TEXT, span_text, context_prefix)
        
        hydration_status = "hydrated"
        if start == -1:
            hydration_status = "ambiguous"
            start = ""
            end = ""
        
        span_id = f"{NOTE_ID}_s{span_counter:02d}"
        
        # Row for Span_Annotations (Empty start/end)
        anno_row = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure",
            context_prefix, span_text, "",
            "", "", f"=LEN(F{ws_anno.max_row+1})", # Excel Formula for len
            label, norm_val, schema_field, event_id,
            "FALSE", "FALSE", "", "", "",
            "needs_hydration"
        ]
        ws_anno.append(anno_row)

        # Row for Span_Hydrated (Filled start/end)
        hydrated_row = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure",
            context_prefix, span_text, "",
            start, end, len(span_text) if isinstance(start, int) else "",
            label, norm_val, schema_field, event_id,
            "FALSE", "FALSE", "", "", "",
            hydration_status
        ]
        ws_hydrated.append(hydrated_row)
        hydrated_rows.append(hydrated_row)
        
        span_counter += 1

    # 5. Populate Event_Log
    ws_event = wb["Event_Log"]
    if ws_event.max_row == 1:
        headers = [
            "source_file", "note_id", "event_id", "event_type", "method",
            "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements",
            "specimens", "findings", "is_historical", "reviewer", "comments",
            "device_size", "device_material",
            "outcome_airway_lumen_pre", "outcome_airway_lumen_post",
            "outcome_symptoms", "outcome_pleural", "outcome_complication"
        ]
        ws_event.append(headers)

    for ev in EVENTS:
        row = [
            SOURCE_FILE, NOTE_ID, ev["event_id"], ev["event_type"], ev.get("method", ""),
            ev.get("anatomy_target", ""), ev.get("device", ""), "", "", "", "",
            "", ev.get("findings", ""), "FALSE", "", "",
            ev.get("device_size", ""), ev.get("device_material", ""),
            "", "", # Airway Lumen
            "", # Symptoms
            ev.get("outcome_pleural", ""),
            ev.get("outcome_complication", "")
        ]
        ws_event.append(row)

    # 6. Populate V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    if ws_v3.max_row == 1:
        headers = [
            "note_id", "event_id", "type",
            "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station",
            "lesion.type", "lesion.size_mm",
            "method", "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote",
            "stent.size", "stent.material_or_brand", "catheter.size_fr",
            "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post",
            "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"
        ]
        ws_v3.append(headers)
    
    # Simple map for V3 sheet
    for ev in EVENTS:
        v3_row = [
            NOTE_ID, ev["event_id"], ev["event_type"],
            ev.get("anatomy_target", ""), "", "", "",
            "", "", # lesion
            ev.get("method", ""), "[]", "[]", "[]", "[]", "",
            "", "", ev.get("device_size", ""), # stent/catheter
            "", "", "", ev.get("outcome_pleural", ""), ev.get("outcome_complication", "")
        ]
        ws_v3.append(v3_row)

    # 7. Generate V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 1:
        ws_json.append(["json_output"])
    
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": EVENTS,
        "no_immediate_complications": True # Derived from note content
    }
    
    json_str = json.dumps(registry_data, indent=2)
    ws_json.append([json_str])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()