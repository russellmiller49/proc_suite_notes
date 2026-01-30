import openpyxl
from openpyxl.utils import get_column_letter
import os
import datetime
import re

# -------------------------------------------------------------------------
# CONSTANTS & INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_030"
SOURCE_FILE = "note_030.txt"
PROCEDURE_DATE = "2026-01-12" # inferred or current
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_030 SOURCE_FILE: note_030.txt INDICATION FOR OPERATION:  [REDACTED]is a 65 year old-year-old male who presents with tracheostomy change.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
 
CONSENT : Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient or surrogate.
The patient or surrogate read and signed the provided consent form / provided consent over the phone.
The consent was witnessed by an assisting medical professional.
 
PREOPERATIVE DIAGNOSIS: J96.90 Respiratory Failure
 
POSTOPERATIVE DIAGNOSIS:  J96.90 Respiratory Failure
 
PROCEDURE:  
31899 Unlisted Procedure (Trach Change with Mature Tract or Procedure NOS)
31645 Therapeutic aspiration initial episode
31899NFJ TRACHEOSTOMY CHANGE AFTER ESTABLISHMENT OF FISTULA TRACT;
WITHOUT AN E/M SERVICE OR OTHER ENDOSCOPY PROCEDURE
31615 Visualization of windpipe (Tracheobronchoscopy through established tracheostomy incision)
 
ANESTHESIA: 
99152 Moderate sedation: initial 15 minutes
99153 Moderate sedation: each additional 15 minutes 
Procedure performed under moderate sedation.
The following medications were provided:
Versed             4 mg
Fentanyl          100 mcg
 
Physician/patient face-to-face anesthesia start time:   11:50
 
Physician/patient face-to-face anesthesia stop time:   12:20
 
Total moderate sedation time was 30 minutes.
Patient was monitored continuously one-to-one throughout the entire procedure by the attending physician while anesthesia was administered.
Sedation was administered by RN. 
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Disposable Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
 
Successful therapeutic aspiration was performed to clean out the Right Mainstem, Bronchus Intermedius , and Left Mainstem from mucus.
Upper airway was suctioned and cleared.  Endotracheal suctioning performed.  The cuff was deflated and the tracheostomy tube was easily removed.
The stoma appeared widely patent and no granulation tissue was immediately visualized.
The new tracheostomy tube was then placed with obturator in place.
The obturator was removed, inner cannula was placed and the cuff inflated.
Percutaneous tracheostomy was changed from:
Portex cuffed Trach ISO/ID size 8.0mm
 
To 
Portex cuffed Trach ISO/ID size 8.0mm
 
without issue.
Tracheobronchoscopy was performed with insertion of bronchoscope through the tracheostomy to perform airway clearance and confirm tracheostomy position.
The patient tolerated the procedure well.  There were no immediate complications.
SPECIMEN(S): 
 
 
IMPRESSION/PLAN: [REDACTED]is a 65 year old-year-old male who presents for bronchoscopy for tracheostomy change.
- f/u as outpatient for tracheostomy management"""

# -------------------------------------------------------------------------
# CONFIGURATION
# -------------------------------------------------------------------------
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0,
}

# Spans to extract
# Fields: span_text, label, normalized_value, schema_field, event_id, context_prefix, match_index
SPANS = [
    # Procedure Methods
    ("Tracheobronchoscopy", "PROC_METHOD", "bronchoscopy", "method", "evt_1", None, 0),
    ("Therapeutic aspiration", "PROC_METHOD", "therapeutic_aspiration", "method", "evt_2", None, 0),
    ("Moderate sedation", "PROC_METHOD", "moderate_sedation", "method", "evt_1", None, 0),
    
    # Anatomy
    ("Right Mainstem", "ANAT_AIRWAY", "RMS", "target.anatomy_type", "evt_2", None, 0),
    ("Bronchus Intermedius", "ANAT_AIRWAY", "BI", "target.anatomy_type", "evt_2", None, 0),
    ("Left Mainstem", "ANAT_AIRWAY", "LMS", "target.anatomy_type", "evt_2", None, 0),
    ("stoma", "ANAT_AIRWAY", "stoma", "target.anatomy_type", "evt_3", None, 0),
    
    # Devices
    ("Disposable Bronchoscope", "DEV_INSTRUMENT", "bronchoscope", "devices_json", "evt_1", None, 0),
    ("tracheostomy tube", "DEV_CATHETER", "tracheostomy_tube", "devices_json", "evt_3", "The cuff was deflated and the", 0), # Removed tube
    ("Portex", "DEV_CATHETER", "Portex", "devices_json", "evt_3", "Percutaneous tracheostomy was changed from", 0), # Old
    ("8.0mm", "DEV_CATHETER_SIZE", "8.0mm", "catheter.size_fr", "evt_3", "Portex cuffed Trach ISO/ID size", 0), # Old size
    
    ("tracheostomy tube", "DEV_CATHETER", "tracheostomy_tube", "devices_json", "evt_4", "The new", 0), # New tube
    ("Portex", "DEV_CATHETER", "Portex", "devices_json", "evt_4", "To \nPortex", 0), # New
    ("8.0mm", "DEV_CATHETER_SIZE", "8.0mm", "catheter.size_fr", "evt_4", "To \nPortex cuffed Trach ISO/ID size", 0), # New size
    
    # Actions/Findings
    ("clean out", "PROC_ACTION", "suction", "findings_json", "evt_2", None, 0),
    ("removed", "PROC_ACTION", "removed", "findings_json", "evt_3", "tracheostomy tube was easily", 0),
    ("placed", "PROC_ACTION", "placed", "findings_json", "evt_4", "tracheostomy tube was then", 0),
    ("granulation tissue", "OBS_FINDING", "granulation_tissue", "findings_json", "evt_3", None, 0),
    
    # Outcomes
    ("No immediate complications", "OUTCOME_COMPLICATION", "none", "outcomes.complications", "evt_1", None, 0),
    ("tolerated the procedure well", "OUTCOME_SYMPTOMS", "tolerated_well", "outcomes.symptoms", "evt_1", None, 0),
]

EVENTS = [
    {
        "event_id": "evt_1",
        "type": "diagnostic_bronchoscopy",
        "method": "bronchoscopy",
        "findings": "visualized windpipe",
        "outcomes": {
            "symptoms": "tolerated_well",
            "complications": "none"
        }
    },
    {
        "event_id": "evt_2",
        "type": "therapeutic_aspiration",
        "method": "therapeutic_aspiration",
        "target_anatomy": ["RMS", "BI", "LMS"],
        "action": "clean out mucus"
    },
    {
        "event_id": "evt_3",
        "type": "device_removal",
        "method": "tracheostomy_change",
        "device": "Portex Trach 8.0mm",
        "action": "removed"
    },
    {
        "event_id": "evt_4",
        "type": "device_placement",
        "method": "tracheostomy_change",
        "device": "Portex Trach 8.0mm",
        "action": "placed"
    }
]

# -------------------------------------------------------------------------
# HELPER FUNCTIONS
# -------------------------------------------------------------------------
def get_hydration_data(note_text, span_text, context_prefix=None, match_index=None):
    """
    Finds start_char and end_char for a span.
    Returns: (start, end, status_string)
    """
    if not span_text:
        return None, None, "missing_text"
        
    matches = [m for m in re.finditer(re.escape(span_text), note_text)]
    
    if not matches:
        return None, None, "not_found"
    
    # 1. Unique match
    if len(matches) == 1:
        m = matches[0]
        return m.start(), m.end(), "hydrated_unique"
        
    # 2. Context prefix
    if context_prefix:
        # Search for prefix near matches
        best_match = None
        for m in matches:
            # Check preceding 120 chars
            window_start = max(0, m.start() - 120)
            preceding_text = note_text[window_start:m.start()]
            if context_prefix in preceding_text:
                best_match = m
                break
        if best_match:
            return best_match.start(), best_match.end(), "hydrated_prefix_window"
            
    # 3. Match Index
    if match_index is not None and 0 <= match_index < len(matches):
        m = matches[match_index]
        return m.start(), m.end(), "hydrated_match_index"
        
    return None, None, f"ambiguous_count={len(matches)}"

def create_workbook():
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
        
    return wb

def ensure_sheet(wb, name, headers):
    if name not in wb.sheetnames:
        ws = wb.create_sheet(name)
        ws.append(headers)
    else:
        ws = wb[name]
    return ws

# -------------------------------------------------------------------------
# MAIN EXECUTION
# -------------------------------------------------------------------------
def main():
    wb = openpyxl.load_workbook(TEMPLATE_PATH) if os.path.exists(TEMPLATE_PATH) else create_workbook()
    
    # 1. Note_Text
    ws_text = ensure_sheet(wb, "Note_Text", ["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # 2. Note_Index
    ws_index = ensure_sheet(wb, "Note_Index", [
        "source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes"
    ] + list(PROCEDURE_FLAGS.keys()))
    
    idx_row = [
        SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "pending", ""
    ] + [PROCEDURE_FLAGS[k] for k in PROCEDURE_FLAGS]
    ws_index.append(idx_row)
    
    # 3. Span_Annotations (Anchor First)
    ws_spans = ensure_sheet(wb, "Span_Annotations", [
        "source_file", "note_id", "span_id", "section_type",
        "context_prefix", "span_text", "match_index",
        "start_char", "end_char", "span_len",
        "label", "normalized_value", "schema_field", "event_id",
        "is_negated", "is_historical", "time_anchor", "reviewer", "comments",
        "hydration_status"
    ])
    
    # 4. Span_Hydrated
    ws_hydrated = ensure_sheet(wb, "Span_Hydrated", [
        "source_file", "note_id", "span_id", "section_type",
        "context_prefix", "span_text", "match_index",
        "start_char", "end_char", "span_len",
        "label", "normalized_value", "schema_field", "event_id",
        "is_negated", "is_historical", "time_anchor", "reviewer", "comments",
        "hydration_status"
    ])
    
    span_id_counter = 1
    
    for span_def in SPANS:
        text, label, norm, field, evt, ctx, midx = span_def
        
        # Calculate hydration
        start, end, status = get_hydration_data(NOTE_TEXT, text, ctx, midx)
        span_len = len(text) if text else 0
        
        row_base = [
            SOURCE_FILE, NOTE_ID, f"span_{span_id_counter:03d}", "Procedure",
            ctx, text, midx
        ]
        
        # Write to Anchor First (blank offsets)
        row_anchor = row_base + [
            "", "", f"=LEN(F{ws_spans.max_row+1})", # formula for len
            label, norm, field, evt,
            "FALSE", "FALSE", "", "", "",
            "needs_hydration"
        ]
        ws_spans.append(row_anchor)
        
        # Write to Hydrated (filled offsets)
        row_hyd = row_base + [
            start, end, span_len,
            label, norm, field, evt,
            "FALSE", "FALSE", "", "", "",
            status
        ]
        ws_hydrated.append(row_hyd)
        
        span_id_counter += 1
        
    # 5. Event_Log
    ws_events = ensure_sheet(wb, "Event_Log", [
        "source_file", "note_id", "event_id", "event_type", "method",
        "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements",
        "specimens", "findings", "is_historical", "reviewer", "comments",
        "device_size", "device_material",
        "outcome_airway_lumen_pre", "outcome_airway_lumen_post",
        "outcome_symptoms", "outcome_pleural", "outcome_complication"
    ])
    
    for evt in EVENTS:
        # Defaults
        out_sympt = evt.get("outcomes", {}).get("symptoms", "")
        out_comp = evt.get("outcomes", {}).get("complications", "")
        
        row = [
            SOURCE_FILE, NOTE_ID, evt["event_id"], evt["type"], evt.get("method", ""),
            str(evt.get("target_anatomy", "")), evt.get("device", ""), "", "", "", "",
            "", evt.get("action", "") or evt.get("findings", ""), "FALSE", "", "",
            "", "", # size/material
            "", "", # lumen
            out_sympt, "", out_comp
        ]
        ws_events.append(row)

    # 6. V3_Procedure_Events
    ws_v3 = ensure_sheet(wb, "V3_Procedure_Events", [
        "note_id", "event_id", "type",
        "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station",
        "lesion.type", "lesion.size_mm",
        "method", "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote",
        "stent.size", "stent.material_or_brand", "catheter.size_fr",
        "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post",
        "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"
    ])
    
    for evt in EVENTS:
        out_sympt = evt.get("outcomes", {}).get("symptoms", "")
        out_comp = evt.get("outcomes", {}).get("complications", "")
        
        row_v3 = [
            NOTE_ID, evt["event_id"], evt["type"],
            str(evt.get("target_anatomy", "")), "", "", "",
            "", "",
            evt.get("method", ""), 
            f'{{"device": "{evt.get("device", "")}"}}' if evt.get("device") else "", 
            "", "", 
            f'{{"findings": "{evt.get("action", "")}"}}', "",
            "", "", "", # specific props
            "", "",
            out_sympt, "", out_comp
        ]
        ws_v3.append(row_v3)

    # 7. V3_Registry_JSON
    ws_json = ensure_sheet(wb, "V3_Registry_JSON", ["schema_version", "note_id", "json_output"])
    ws_json.append(["1.0", NOTE_ID, "{}"]) # Placeholder

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    main()