import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os
from datetime import datetime

# =============================================================================
# INPUT DATA
# =============================================================================

NOTE_ID = "note_031"
SOURCE_FILE = "note_031.txt"
PROCEDURE_DATE = "2026-01-12" # Placeholder based on current context if needed, or blank
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_031 SOURCE_FILE: note_031.txt INDICATION FOR OPERATION:  [REDACTED]is a 65 year old-year-old male who presents with respiratory failure.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
 
CONSENT : Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient or surrogate.
The patient or surrogate read and signed the provided consent form / provided consent over the phone.
The consent was witnessed by an assisting medical professional.
 
PREOPERATIVE DIAGNOSIS: J96.90 Respiratory Failure
 
POSTOPERATIVE DIAGNOSIS:  J96.90 Respiratory Failure
 
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31624 Dx bronchoscope/lavage (BAL)    
31600 Incision of windpipe (perc trach)
76536 Ultrasound of Neck
 
 
ANESTHESIA: 
99152 Moderate sedation: initial 15 minutes
99153 Moderate sedation: each additional 15 minutes 
 
Procedure performed under moderate sedation.
The following medications were provided:
Etomidate30mg
Rocuronium80mg
Propofol gtt50 mcg/kg/min
Fentanyl gtt150 mcg/hr
 
Physician/patient face-to-face anesthesia start time:   1326
 
Physician/patient face-to-face anesthesia stop time:   1520
 
Total moderate sedation time was 114 minutes.
Patient was monitored continuously one-to-one throughout the entire procedure by the attending physician while anesthesia was administered.
Sedation was administered by ICU RN. 
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Diagnostic Bronchoscope
 
ESTIMATED BLOOD LOSS:   Minimum
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: Supine
 
Initial Airway Inspection Findings:
The endotracheal tube is in good position.
Pharynx: Not assessed due to bronchoscopy introduction through ETT.
Larynx: Not assessed due to bronchoscopy introduction through ETT.
Vocal Cords: Not assessed due to bronchoscopy introduction through ETT.
Trachea: Distal 1/3 normal.
Main Carina: Sharp
Right Lung Proximal Airways: Normal anatomic branching to segmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Left Lung Proximal Airways: Normal anatomic branching to segmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Mucosa: Normal.
Secretions: Minimal, thin, and clear.
Successful therapeutic aspiration was performed to clean out the Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus.
Bronchial alveolar lavage was performed at Superior Segment of Lingula (LB4) and Inferior Segment of Lingula (LB5).
Instilled 60 cc of NS, suction returned with 15 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
 
Neck Ultrasound was performed to evaluate for any abnormal vessel, mass, structures at the site of percutaneous tracheostomy.
There were no significant vessels/mass were noted overlying the tracheostomy site on examination from the laryngeal prominence to the sternal notch.
The bronchoscope was retracted into the ETT tube and the ET tube retracted into the subglottic space under direct visualization.
The inferior border of the cricoid along with the proximal tracheal rings were visualized.
Next, the anterior neck was prepped and draped in the usual sterile fashion.
Lidocaine 1% 3ml ml was injected into the anterior neck.
A 1 cm incision was made horizontally with a #10 blade down through the subcutaneous tissue, just inferior to the cricoid cartilage.
Suture (3-0 prolene) was placed around the incision site to create a Rummel tourniquet.
The introducer needle was then passed between the 1st and 2nd tracheal rings and into the trachea under direct visualization.
Next, a J-wire was passed through the catheter, also visualized with the bronchoscope.
The site was then dilated using the 14Fr introducing dilator passed over the wire.
The 14 Fr dilator was then removed from the guide wire and an 8 Fr guiding catheter placed over the guide wire until the safety ridge on the guiding catheter was at skin level.
The tissue dilator was placed over the guiding catheter until the positioning mark was visualized via the bronchoscope.
The tissue dilator was then removed leaving the guiding catheter and guide wire assembly in place, all under direct visualization bronchoscopically.
Finally a Shiley 8.0 cuffed tracheostomy tube with appropriate dilator was introduced over the guiding catheter into the trachea under direct visualization.
The dilator, guiding catheter, and J-wire were then removed and the tracheostomy tube left in place.
This was confirmed to be in good position bronchoscopically.  The Endotracheal tube was then removed and the ventilator connected to the tracheostomy tube.
Surgicel was placed preemptively around the tracheostomy site to reduce bleeding.
A Lyofoam drain sponge was placed under the tracheostomy tube prior to suturing into place.
The patient tolerated the procedure well.  There were no complications. The staff physician was present throughout the entire procedure.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was in stable condition.
SPECIMEN(S): 
--Lingula BAL (cell count, micro, cyto)
 
IMPRESSION/PLAN: [REDACTED]is a 65 year old-year-old male who presents for bronchoscopy and tracheostomy tube placement.
Portex 8.0mm cuffed tracheostomy tube and Rummel tourniquet placed. Patient tolerated the procedure well and there were no immediate complications.
--Post procedure CXR
--Anticipate suture removal in 7 days 
--Anticipate trach change in 10 days"""

# =============================================================================
# CONFIGURATION & FLAGGING
# =============================================================================

# Flags 0/1
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 1,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 0,
    "airway_dilation": 0, # Note: Dilation used for trach insertion, not therapeutic airway dilation
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# =============================================================================
# SPAN DATA DEFINITION
# =============================================================================

# Structure: (text, label, norm_value, context_prefix, event_id)
# Note: event_id linkage:
# evt01: Inspection/General
# evt02: Therapeutic Aspiration
# evt03: BAL
# evt04: Percutaneous Tracheostomy
# evt05: Outcomes

RAW_SPANS = [
    # Header/Dx
    ("J96.90", "OBS_FINDING", "Respiratory Failure", "DIAGNOSIS:", None),
    
    # Event 1: Inspection (No specific pathology found, but anatomy checked)
    ("Trachea: Distal 1/3 normal", "OBS_FINDING", "Normal", None, "evt01"),
    ("Right Lung Proximal Airways: Normal", "OBS_FINDING", "Normal", None, "evt01"),
    ("Left Lung Proximal Airways: Normal", "OBS_FINDING", "Normal", None, "evt01"),
    ("Secretions: Minimal, thin, and clear", "OBS_FINDING", "Secretions", None, "evt01"),

    # Event 2: Therapeutic Aspiration
    ("Therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", "31645", "evt02"),
    ("clean out", "PROC_ACTION", "Suction", "was performed to", "evt02"),
    ("Trachea (Distal 1/3)", "ANAT_AIRWAY", "Trachea", "clean out the", "evt02"),
    ("Right Mainstem", "ANAT_AIRWAY", "RMS", "Distal 1/3),", "evt02"),
    ("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "Right Mainstem,", "evt02"),
    ("Left Mainstem", "ANAT_AIRWAY", "LMS", "Bronchus Intermedius ,", "evt02"),
    ("Carina", "ANAT_AIRWAY", "Carina", "Left Mainstem,", "evt02"),
    ("RUL Carina (RC1)", "ANAT_AIRWAY", "RUL Carina", "Carina,", "evt02"),
    ("mucus", "OBS_FINDING", "Mucus", "LC2) from", "evt02"),

    # Event 3: BAL
    ("Bronchial alveolar lavage", "PROC_METHOD", "BAL", "18]", "evt03"),
    ("Superior Segment of Lingula (LB4)", "ANAT_LUNG_LOC", "Lingula Superior", "performed at", "evt03"),
    ("Inferior Segment of Lingula (LB5)", "ANAT_LUNG_LOC", "Lingula Inferior", "and", "evt03"),
    ("Instilled 60 cc", "MEAS_VOL", "60cc", None, "evt03"),
    ("returned with 15 cc", "MEAS_VOL", "15cc", "suction", "evt03"),
    ("Cell Count", "PROC_ACTION", "Specimen Analysis", "sent for", "evt03"),
    ("Microbiology", "PROC_ACTION", "Specimen Analysis", "Cell Count,", "evt03"),
    ("Cytology", "PROC_ACTION", "Specimen Analysis", "Fungal), and", "evt03"),

    # Event 4: Tracheostomy
    ("Incision of windpipe (perc trach)", "PROC_METHOD", "Percutaneous Tracheostomy", "31600", "evt04"),
    ("Neck Ultrasound", "PROC_METHOD", "Ultrasound", "Cytology.\n \n", "evt04"),
    ("no significant vessels/mass", "OBS_FINDING", "Normal", "There were", "evt04"),
    ("Incision", "PROC_ACTION", "Incision", "A 1 cm", "evt04"),
    ("inferior to the cricoid cartilage", "ANAT_AIRWAY", "Cricoid", "tissue, just", "evt04"),
    ("Rummel tourniquet", "DEV_INSTRUMENT", "Rummel tourniquet", "create a", "evt04"),
    ("introducer needle", "DEV_NEEDLE", "Introducer Needle", "The", "evt04"),
    ("1st and 2nd tracheal rings", "ANAT_AIRWAY", "Tracheal Rings 1-2", "between the", "evt04"),
    ("J-wire", "DEV_INSTRUMENT", "J-wire", "Next, a", "evt04"),
    ("14Fr introducing dilator", "DEV_INSTRUMENT", "14Fr Dilator", "using the", "evt04"),
    ("8 Fr guiding catheter", "DEV_CATHETER", "8Fr Guiding Catheter", "wire and an", "evt04"),
    ("tissue dilator", "DEV_INSTRUMENT", "Tissue Dilator", "The", "evt04"),
    ("Shiley", "DEV_CATHETER", "Tracheostomy Tube", "Finally a", "evt04"),
    ("8.0", "DEV_CATHETER_SIZE", "8.0", "Shiley", "evt04"),
    ("cuffed tracheostomy tube", "DEV_CATHETER", "Tracheostomy Tube", "Shiley 8.0", "evt04"),
    ("Portex 8.0mm cuffed tracheostomy tube", "DEV_CATHETER", "Tracheostomy Tube", "IMPRESSION/PLAN", "evt04"),
    
    # Outcomes
    ("tolerated the procedure well", "OUTCOME_SYMPTOMS", "Tolerated Well", "The patient", "evt05"),
    ("no complications", "OUTCOME_COMPLICATION", "None", "There were", "evt05"),
    ("stable condition", "OBS_FINDING", "Stable", "patient was in", "evt05")
]

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def clean_text(text):
    return text.replace('\r', '').strip()

def find_offsets(full_text, span_text, context_prefix=None, match_index=None):
    """
    Find start_char, end_char based on hydration logic.
    Returns: (start, end, status_msg)
    """
    if not span_text:
        return None, None, "missing_text"
    
    # 1. Exact match count
    matches = [m.start() for m in re.finditer(re.escape(span_text), full_text)]
    count = len(matches)
    
    if count == 0:
        return None, None, "not_found"
    
    if count == 1:
        return matches[0], matches[0] + len(span_text), "hydrated_unique"
    
    # 2. Context Prefix
    if context_prefix:
        # Look for occurrence where prefix is in preceding 150 chars
        for m_start in matches:
            window_start = max(0, m_start - 150)
            preceding_text = full_text[window_start:m_start]
            if context_prefix in preceding_text:
                return m_start, m_start + len(span_text), "hydrated_prefix_window"
    
    # 3. Match Index
    if match_index is not None and 0 <= match_index < count:
        chosen = matches[match_index]
        return chosen, chosen + len(span_text), "hydrated_match_index"
    
    # Fallback if no context worked but multiple exist
    return None, None, f"ambiguous_count={count}"

# =============================================================================
# WORKBOOK GENERATION
# =============================================================================

def create_workbook():
    # 1. Load Template
    if not os.path.exists(TEMPLATE_PATH):
        # Fallback for demonstration if template doesn't exist, create simple one
        wb = openpyxl.Workbook()
        ws_flags = wb.create_sheet("Note_Index")
        ws_flags.append(["source_file", "note_id", "procedure_date"] + list(PROCEDURE_FLAGS.keys()))
        ws_text = wb.create_sheet("Note_Text")
        ws_text.append(["note_id", "source_file", "note_text"])
        ws_spans = wb.create_sheet("Span_Annotations")
        ws_spans.append(["source_file","note_id","span_id","section_type","context_prefix","span_text","match_index","start_char","end_char","span_len","label","normalized_value","schema_field","event_id","is_negated","is_historical","time_anchor","reviewer","comments","hydration_status"])
        ws_hydrated = wb.create_sheet("Span_Hydrated")
        ws_hydrated.append(["source_file","note_id","span_id","section_type","context_prefix","span_text","match_index","start_char","end_char","span_len","label","normalized_value","schema_field","event_id","is_negated","is_historical","time_anchor","reviewer","comments","hydration_status"])
        wb.create_sheet("Event_Log")
        wb.create_sheet("V3_Procedure_Events")
        wb.create_sheet("V3_Registry_JSON")
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # 2. Populate Note_Text
    ws_text = wb["Note_Text"]
    # Check if empty, if so add header
    if ws_text.max_row == 1 and ws_text.cell(1,1).value is None:
        ws_text.append(["note_id", "source_file", "note_text"])
    
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 3. Populate Note_Index (Flags)
    ws_index = wb["Note_Index"]
    header = [cell.value for cell in ws_index[1]]
    
    row_data = {
        "source_file": SOURCE_FILE,
        "note_id": NOTE_ID,
        "procedure_date": PROCEDURE_DATE,
        "status": "Auto-Generated"
    }
    # Add flags
    row_data.update(PROCEDURE_FLAGS)
    
    new_row = []
    for col in header:
        new_row.append(row_data.get(col, ""))
    ws_index.append(new_row)

    # 4. Process Spans & Hydration
    ws_spans = wb["Span_Annotations"]
    ws_hydrated = wb["Span_Hydrated"]
    
    # Load events for later grouping
    events_map = {} # event_id -> list of dicts

    for i, (span_txt, label, norm, ctx, evt_id) in enumerate(RAW_SPANS):
        span_id = f"span_{i+1:03d}"
        
        # Hydrate
        start, end, status = find_offsets(NOTE_TEXT, span_txt, context_prefix=ctx)
        span_len = len(span_txt) if span_txt else 0
        
        # Common data
        row_dict = {
            "source_file": SOURCE_FILE,
            "note_id": NOTE_ID,
            "span_id": span_id,
            "span_text": span_txt,
            "context_prefix": ctx,
            "label": label,
            "normalized_value": norm,
            "event_id": evt_id,
            "start_char": start,
            "end_char": end,
            "span_len": span_len,
            "hydration_status": status
        }
        
        # Write to Span_Annotations (Blank offsets)
        ws_spans.append([
            SOURCE_FILE, NOTE_ID, span_id, "", ctx, span_txt, "", "", "", span_len,
            label, norm, "", evt_id, "FALSE", "FALSE", "", "Auto", "", "needs_hydration"
        ])

        # Write to Span_Hydrated (Filled offsets)
        ws_hydrated.append([
            SOURCE_FILE, NOTE_ID, span_id, "", ctx, span_txt, "", start if start is not None else "", end if end is not None else "", span_len,
            label, norm, "", evt_id, "FALSE", "FALSE", "", "Auto", "", status
        ])
        
        # Collect for Event Log
        if evt_id:
            if evt_id not in events_map:
                events_map[evt_id] = []
            events_map[evt_id].append(row_dict)

    # 5. Populate Event_Log & V3_Procedure_Events
    ws_events = wb["Event_Log"]
    ws_v3 = wb["V3_Procedure_Events"]
    
    # Ensure headers exist if creating new
    if ws_events.max_row == 1:
        # Add basic header if missing
        ws_events.append(["source_file", "note_id", "event_id", "event_type", "method", "anatomy", "device", "outcomes"])

    v3_registry_list = []

    for evt_id, spans in events_map.items():
        # Aggregate data for this event
        method = next((s['normalized_value'] for s in spans if s['label'] == 'PROC_METHOD'), "")
        action = next((s['normalized_value'] for s in spans if s['label'] == 'PROC_ACTION'), "")
        anatomy = [s['normalized_value'] for s in spans if s['label'] in ('ANAT_AIRWAY', 'ANAT_LUNG_LOC', 'ANAT_PLEURA')]
        devices = [s['normalized_value'] for s in spans if s['label'] in ('DEV_INSTRUMENT', 'DEV_CATHETER', 'DEV_NEEDLE', 'DEV_STENT')]
        outcomes = [s['normalized_value'] for s in spans if s['label'] in ('OUTCOME_SYMPTOMS', 'OUTCOME_COMPLICATION')]
        measurements = [f"{s['span_text']} ({s['normalized_value']})" for s in spans if 'MEAS' in s['label']]
        
        # Write flat event log
        ws_events.append([
            SOURCE_FILE, NOTE_ID, evt_id, 
            "Procedure", 
            method or action, 
            ", ".join(anatomy), 
            ", ".join(devices), 
            ", ".join(outcomes)
        ])
        
        # Build V3 Object
        v3_obj = {
            "event_id": evt_id,
            "type": method,
            "method": method,
            "target": {"anatomy": anatomy},
            "devices": devices,
            "outcomes": outcomes,
            "measurements": measurements
        }
        
        ws_v3.append([NOTE_ID, evt_id, method, str(anatomy), "", "", "", "", "", str(devices), str(measurements), "", "", "", "", "", "", "", "", "", ""])
        v3_registry_list.append(v3_obj)

    # 6. JSON Export
    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 1:
        ws_json.append(["schema_version", "note_id", "json_output"])
    
    final_json = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True, # Derived from note content
        "procedures": v3_registry_list
    }
    
    ws_json.append(["3.0", NOTE_ID, json.dumps(final_json, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    create_workbook()