import os
import re
import json
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# --- INPUT CONFIGURATION ---
NOTE_ID = "note_033"
SOURCE_FILE = "note_033.txt"
PROCEDURE_DATE = "2025-12-15" # inferred from chest tube date context or left blank if uncertain, using context date
NOTE_TEXT = """NOTE_ID:  note_033 SOURCE_FILE: note_033.txt  INDICATION FOR OPERATION:  [REDACTED]is a 47 year old-year-old male who presents with Pleural Effusion and Complicated Effusion.
The nature, purpose, risks, benefits and alternatives to Instillation of agents for fibrinolysis (initial) were discussed with the patient in detail.
Patient indicated a wish to proceed with procedure and informed consent was signed.
PREOPERATIVE DIAGNOSIS:  Pleural Effusion
POSTOPERATIVE DIAGNOSIS: Same as preoperative diagnosis - see above.
PROCEDURE:  
32561 Instillation(s), via chest tube/catheter, agent for fibrinolysis (eg, fibrinolytic agent for break up of multiloculated effusion);
initial day
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
PROCEDURE IN DETAIL:
 
PATIENT POSITION: 
0‌ Supine  0‌ Sitting   
0‌ Lateral Decubitus:  0‌ Right 0‌ Left 
 
Date of chest tube insertion: 12/15/25
 
Side: right
1‌  10 mg/5 mg tPA/Dnasedose #:_1___ 
            0‌  ___mg tPA                              dose #:____ 
0‌  Other medication:  
 
COMPLICATIONS:
0‌None 0‌Bleeding-EBL: ___ ml 0‌Pneumothorax 0‌Re- Expansion Pulmonary Edema 
0‌Other: 
 
IMPRESSION/PLAN: 
[REDACTED]is a 47 year old-year-old male who presents for Instillation of agents for fibrinolysis (initial).
The patient tolerated the procedure well.  There were no immediate complications.  
 
 
DISPOSITION: Home"""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# --- DEFINITIONS & SCHEMA ---

PROCEDURE_FLAGS = {
    'diagnostic_bronchoscopy': 0, 'bal': 0, 'bronchial_wash': 0, 'brushings': 0, 
    'endobronchial_biopsy': 0, 'tbna_conventional': 0, 'linear_ebus': 0, 'radial_ebus': 0, 
    'navigational_bronchoscopy': 0, 'transbronchial_biopsy': 0, 'transbronchial_cryobiopsy': 0, 
    'therapeutic_aspiration': 0, 'foreign_body_removal': 0, 'airway_dilation': 0, 
    'airway_stent': 0, 'thermal_ablation': 0, 'tumor_debulking_non_thermal': 0, 
    'cryotherapy': 0, 'blvr': 0, 'peripheral_ablation': 0, 'bronchial_thermoplasty': 0, 
    'whole_lung_lavage': 0, 'rigid_bronchoscopy': 0, 
    'thoracentesis': 0, 'chest_tube': 0, 'ipc': 0, 'medical_thoracoscopy': 0, 
    'pleurodesis': 0, 'pleural_biopsy': 0, 'fibrinolytic_therapy': 1
}

# Anchor-first Spans: text, label, norm_val, schema_field, event_id
# Note: Normalized values guide the registry mapping.
SPANS_DATA = [
    # Diagnosis/Indication
    ("Pleural Effusion", "OBS_LESION", "Pleural Effusion", "lesion.type", "E1"),
    ("Complicated Effusion", "OBS_LESION", "Complicated Effusion", "lesion.type", "E1"),
    
    # Procedure Method
    ("Instillation of agents for fibrinolysis", "PROC_METHOD", "fibrinolytic_therapy", "method", "E1"),
    ("via chest tube", "DEV_CATHETER", "chest_tube", "device", "E1"),
    
    # Anatomy
    ("Side: right", "LATERALITY", "Right", "target.location.lobe", "E1"),
    
    # Agents (Fibrinolytics)
    ("10 mg", "MEAS_VOL", "10mg", "measurements_json", "E1"),
    ("tPA", "DEV_INSTRUMENT", "tPA", "devices_json", "E1"),
    ("5 mg", "MEAS_VOL", "5mg", "measurements_json", "E1"),
    ("Dnase", "DEV_INSTRUMENT", "DNase", "devices_json", "E1"),
    
    # Outcomes
    ("tolerated the procedure well", "OUTCOME_SYMPTOMS", "tolerated well", "outcomes.symptoms", "E1"),
    ("no immediate complications", "OUTCOME_COMPLICATION", "none", "outcomes.complications", "E1"),
]

# --- GENERATOR CLASS ---

class Phase0Generator:
    def __init__(self, note_text, note_id, source_file, proc_date, flags, spans_data, template_path, output_path):
        self.note_text = note_text
        self.note_id = note_id
        self.source_file = source_file
        self.proc_date = proc_date
        self.flags = flags
        self.spans_data = spans_data
        self.template_path = template_path
        self.output_path = output_path
        self.wb = None

    def load_template(self):
        if not os.path.exists(self.template_path):
            # Create a dummy workbook if template missing (fallback for valid script execution)
            self.wb = Workbook()
            self.wb.create_sheet("Note_Text")
            self.wb.create_sheet("Note_Index")
            self.wb.create_sheet("Span_Annotations")
            self.wb.create_sheet("Span_Hydrated")
            self.wb.create_sheet("Event_Log")
            self.wb.create_sheet("V3_Procedure_Events")
            self.wb.create_sheet("V3_Registry_JSON")
        else:
            self.wb = load_workbook(self.template_path)

    def populate_note_text(self):
        ws = self.wb["Note_Text"]
        # Assuming headers: note_id, source_file, note_text
        ws.append([self.note_id, self.source_file, self.note_text])

    def populate_note_index(self):
        ws = self.wb["Note_Index"]
        # Headers: source_file, note_id, encounter_id, procedure_date, site, reviewer, status, free_text_notes, [flags...]
        row = [
            self.source_file, self.note_id, "", self.proc_date, "", "", "pre-filled", ""
        ]
        # Append flags in order
        flag_keys = list(self.flags.keys()) # Ensure order matches strict list if needed, but dict preservation is standard in 3.7+
        # Explicit order as per prompt reqs to ensure column alignment if template has pre-defined headers
        ordered_keys = [
            'diagnostic_bronchoscopy', 'bal', 'bronchial_wash', 'brushings', 'endobronchial_biopsy', 
            'tbna_conventional', 'linear_ebus', 'radial_ebus', 'navigational_bronchoscopy', 
            'transbronchial_biopsy', 'transbronchial_cryobiopsy', 'therapeutic_aspiration', 
            'foreign_body_removal', 'airway_dilation', 'airway_stent', 'thermal_ablation', 
            'tumor_debulking_non_thermal', 'cryotherapy', 'blvr', 'peripheral_ablation', 
            'bronchial_thermoplasty', 'whole_lung_lavage', 'rigid_bronchoscopy',
            'thoracentesis', 'chest_tube', 'ipc', 'medical_thoracoscopy', 'pleurodesis', 
            'pleural_biopsy', 'fibrinolytic_therapy'
        ]
        for k in ordered_keys:
            row.append(self.flags.get(k, 0))
        ws.append(row)

    def hydrate_spans(self):
        annot_ws = self.wb["Span_Annotations"]
        hydrated_ws = self.wb["Span_Hydrated"]
        
        # Write headers if empty (simple check)
        if annot_ws.max_row == 1:
            headers = ["source_file", "note_id", "span_id", "section_type", "context_prefix", 
                       "span_text", "match_index", "start_char", "end_char", "span_len", 
                       "label", "normalized_value", "schema_field", "event_id", 
                       "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"]
            annot_ws.append(headers)
            hydrated_ws.append(headers)

        span_id_counter = 1
        
        for sp in self.spans_data:
            text, label, norm, schema_field, event_id = sp
            
            # Logic for start/end
            start_char = ""
            end_char = ""
            hydration_status = "needs_hydration"
            match_index = 0
            
            # Simple hydration logic
            count = self.note_text.count(text)
            
            final_start = None
            final_end = None
            
            if count == 1:
                final_start = self.note_text.find(text)
                final_end = final_start + len(text)
                hydration_status = "hydrated_unique"
            elif count > 1:
                # Default to first for this script unless context provided
                final_start = self.note_text.find(text)
                final_end = final_start + len(text)
                hydration_status = f"hydrated_first_of_{count}" # Simplified
            else:
                hydration_status = "missing"
            
            # Write to Annotations (Anchor First - no offsets)
            row_annot = [
                self.source_file, self.note_id, f"span_{span_id_counter}", "Body", "", 
                text, "", "", "", f"=LEN(F{annot_ws.max_row+1})", 
                label, norm, schema_field, event_id, 
                "FALSE", "FALSE", "", "Auto", "", "needs_hydration"
            ]
            annot_ws.append(row_annot)
            
            # Write to Hydrated
            if final_start is not None:
                row_hyd = [
                    self.source_file, self.note_id, f"span_{span_id_counter}", "Body", "", 
                    text, 1, final_start, final_end, len(text), 
                    label, norm, schema_field, event_id, 
                    "FALSE", "FALSE", "", "Auto", "", hydration_status
                ]
                hydrated_ws.append(row_hyd)
            
            span_id_counter += 1

    def populate_event_log(self):
        ws = self.wb["Event_Log"]
        # Basic event row for the fibrinolytic therapy
        # columns: source_file, note_id, event_id, event_type, method, anatomy_target, device, needle_gauge, stations, counts, measurements, specimens, findings, is_historical, reviewer, comments, device_size, device_material, outcome_airway_lumen_pre, outcome_airway_lumen_post, outcome_symptoms, outcome_pleural, outcome_complication
        
        row = [
            self.source_file, self.note_id, "E1", "fibrinolytic_therapy", "instillation",
            "Right Pleura", "Chest Tube, tPA, DNase", "", "", "", "10mg, 5mg",
            "", "Effusion", "FALSE", "Auto", "Instillation via existing chest tube",
            "", "", # device size/mat
            "", "", # lumen
            "tolerated well", # symptoms
            "", # pleural outcome
            "none" # complication
        ]
        ws.append(row)

    def populate_v3_events(self):
        ws = self.wb["V3_Procedure_Events"]
        # note_id, event_id, type, target.anatomy_type, target.location.lobe, target.location.segment, target.station, lesion.type, lesion.size_mm, method, devices_json, measurements_json, specimens_json, findings_json, evidence_quote, stent.size, stent.material_or_brand, catheter.size_fr, outcomes.airway.lumen_pre, outcomes.airway.lumen_post, outcomes.symptoms, outcomes.pleural, outcomes.complications
        
        devices = json.dumps(["Chest Tube", "tPA", "DNase"])
        measurements = json.dumps({"tPA": "10mg", "DNase": "5mg"})
        
        row = [
            self.note_id, "E1", "fibrinolytic_therapy", 
            "pleura", "right", "", "", 
            "complicated effusion", "", 
            "instillation", devices, measurements, "", "", "",
            "", "", "", # stent/cath
            "", "", # lumen
            "tolerated well", "", "none"
        ]
        ws.append(row)

    def generate_json(self):
        ws = self.wb["V3_Registry_JSON"]
        
        data = {
            "schema_version": "v3.0",
            "note_id": self.note_id,
            "no_immediate_complications": True,
            "procedures": [
                {
                    "event_id": "E1",
                    "type": "fibrinolytic_therapy",
                    "target": {
                        "anatomy_type": "pleura",
                        "location": {"lobe": "right"}
                    },
                    "method": "instillation",
                    "agents": [
                        {"name": "tPA", "dose": "10mg"},
                        {"name": "DNase", "dose": "5mg"}
                    ],
                    "outcomes": {
                        "symptoms": "tolerated well",
                        "complications": "none"
                    }
                }
            ]
        }
        
        ws.append([json.dumps(data, indent=2)])

    def run(self):
        self.load_template()
        self.populate_note_text()
        self.populate_note_index()
        self.hydrate_spans()
        self.populate_event_log()
        self.populate_v3_events()
        self.generate_json()
        self.wb.save(self.output_path)
        print(f"Generated {self.output_path}")

# --- EXECUTION ---
if __name__ == "__main__":
    generator = Phase0Generator(
        NOTE_TEXT, NOTE_ID, SOURCE_FILE, PROCEDURE_DATE, 
        PROCEDURE_FLAGS, SPANS_DATA, TEMPLATE_PATH, OUTPUT_PATH
    )
    generator.run()