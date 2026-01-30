import json
import re
import os
from openpyxl import Workbook, load_workbook

# ==========================================
# 1. INPUTS
# ==========================================
NOTE_ID = "note_034"
SOURCE_FILE = "note_034.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_034 SOURCE_FILE: note_034.txt INDICATION FOR OPERATION:  [REDACTED]is a 47 year old-year-old male who presents with Pleural Effusion and Complicated Effusion.
The nature, purpose, risks, benefits and alternatives to Instillation of agents for fibrinolysis (initial) were discussed with the patient in detail.
Patient indicated a wish to proceed with procedure and informed consent was signed.
PREOPERATIVE DIAGNOSIS:  Pleural Effusion
POSTOPERATIVE DIAGNOSIS: Same as preoperative diagnosis - see above.
PROCEDURE:  
32561 Instillation(s), via chest tube/catheter, agent for fibrinolysis (eg, fibrinolytic agent for break up of multiloculated effusion);
initial day
 
PROCEDURE IN DETAIL:
 
PATIENT POSITION: 
0‌‌ Supine  0‌‌ Sitting   
0‌‌ Lateral Decubitus:  0‌‌ Right 0‌‌ Left 
 
Date of chest tube insertion: 12/15/25
 
Side: right
1‌‌  10 mg/5 mg tPA/Dnase dose #:_2___ 
            0‌‌  ___mg tPA                              dose #:____ 
0‌‌  Other medication:  
 
COMPLICATIONS:
0‌‌None 0‌‌Bleeding-EBL: ___ ml 0‌‌Pneumothorax 0‌‌Re- Expansion Pulmonary Edema 
0‌‌Other: 

 
IMPRESSION/PLAN: [REDACTED]is a 47 year old-year-old male who presents for Instillation of agents for fibrinolysis (initial).
The patient tolerated the procedure well.  There were no immediate complications.  
 
 
DISPOSITION: Home"""

# ==========================================
# 2. DATA STRUCTURES & LOGIC
# ==========================================

class Span:
    def __init__(self, text, label, context_prefix=None, normalized_value=None, schema_field=None, event_id=None, is_negated=False, is_historical=False):
        self.text = text
        self.label = label
        self.context_prefix = context_prefix
        self.normalized_value = normalized_value
        self.schema_field = schema_field
        self.event_id = event_id
        self.is_negated = is_negated
        self.is_historical = is_historical
        self.start_char = None
        self.end_char = None
        self.hydration_status = "needs_hydration"

    def hydrate(self, full_text):
        if not self.text:
            self.hydration_status = "error_empty_text"
            return

        # 1. Exact match count
        count = full_text.count(self.text)
        
        if count == 1:
            self.start_char = full_text.find(self.text)
            self.end_char = self.start_char + len(self.text)
            self.hydration_status = "hydrated_unique"
        elif count > 1:
            if self.context_prefix:
                # Find all starts
                starts = [m.start() for m in re.finditer(re.escape(self.text), full_text)]
                best_start = -1
                
                # Check preceding window for context
                for s in starts:
                    window_start = max(0, s - 120)
                    preceding_text = full_text[window_start:s]
                    if self.context_prefix in preceding_text:
                        best_start = s
                        break
                
                if best_start != -1:
                    self.start_char = best_start
                    self.end_char = best_start + len(self.text)
                    self.hydration_status = "hydrated_prefix_window"
                else:
                    self.hydration_status = f"ambiguous_prefix_not_found_count={count}"
            else:
                self.hydration_status = f"ambiguous_no_context_count={count}"
        else:
            self.hydration_status = "error_not_found"

def normalize_text(text):
    return text.strip().replace('\r\n', ' ').replace('\n', ' ')

# ==========================================
# 3. EXTRACTION LOGIC
# ==========================================

spans = []
events = []

# --- Event 1: Fibrinolytic Instillation ---
event_id = "event_01"

# Anatomy/Laterality
spans.append(Span("right", "LATERALITY", context_prefix="Side:", normalized_value="Right", event_id=event_id))

# Indication/Diagnosis
spans.append(Span("Pleural Effusion", "OBS_LESION", context_prefix="presents with", normalized_value="Pleural effusion", event_id=event_id))
spans.append(Span("Complicated Effusion", "OBS_LESION", context_prefix="Pleural Effusion and", normalized_value="Complicated effusion", event_id=event_id))

# Procedure
spans.append(Span("Instillation of agents for fibrinolysis", "PROC_METHOD", context_prefix="alternatives to", normalized_value="Fibrinolytic therapy", event_id=event_id))
spans.append(Span("chest tube", "DEV_CATHETER", context_prefix="via", normalized_value="Chest Tube", event_id=event_id))

# Agents & Dosage
spans.append(Span("10 mg", "MEAS_VOL", context_prefix="1", normalized_value="10mg", event_id=event_id, schema_field="measurements"))
spans.append(Span("5 mg", "MEAS_VOL", context_prefix="10 mg/", normalized_value="5mg", event_id=event_id, schema_field="measurements"))
spans.append(Span("tPA", "PROC_ACTION", context_prefix="5 mg", normalized_value="tPA", event_id=event_id))
spans.append(Span("Dnase", "PROC_ACTION", context_prefix="tPA/", normalized_value="Dnase", event_id=event_id))

# Dose Count
spans.append(Span("2", "MEAS_COUNT", context_prefix="dose #:", normalized_value="2", event_id=event_id))

# Outcome
spans.append(Span("No immediate complications", "OUTCOME_COMPLICATION", context_prefix="well.", normalized_value="None", event_id=event_id))

# Historical Context (Chest Tube Insertion Date)
spans.append(Span("12/15/25", "CTX_TIME", context_prefix="insertion:", normalized_value="2025-12-15", is_historical=True))


# --- Event Construction ---
evt = {
    "note_id": NOTE_ID,
    "event_id": event_id,
    "type": "Fibrinolytic Therapy",
    "target": {
        "anatomy_type": "Pleura",
        "location": {"lobe": "Right", "segment": ""}
    },
    "method": "Instillation via Chest Tube",
    "devices_json": json.dumps(["Chest Tube (Existing)"]),
    "measurements_json": json.dumps({"tPA": "10mg", "Dnase": "5mg", "Dose #": "2"}),
    "findings_json": json.dumps(["Pleural Effusion", "Complicated Effusion"]),
    "outcomes": {
        "complications": "None"
    }
}
events.append(evt)

# ==========================================
# 4. PROCEDURE FLAGS
# ==========================================

# Default all to 0
proc_flags = {
    "diagnostic_bronchoscopy": 0, "bal": 0, "bronchial_wash": 0, "brushings": 0, "endobronchial_biopsy": 0,
    "tbna_conventional": 0, "linear_ebus": 0, "radial_ebus": 0, "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0, "transbronchial_cryobiopsy": 0, "therapeutic_aspiration": 0,
    "foreign_body_removal": 0, "airway_dilation": 0, "airway_stent": 0, "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0, "cryotherapy": 0, "blvr": 0, "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0, "whole_lung_lavage": 0, "rigid_bronchoscopy": 0,
    "thoracentesis": 0, "chest_tube": 0, "ipc": 0, "medical_thoracoscopy": 0,
    "pleurodesis": 0, "pleural_biopsy": 0, "fibrinolytic_therapy": 0
}

# Set detected flags
# Note: Chest tube is the route (existing), not the procedure performed. The procedure is fibrinolysis.
proc_flags["fibrinolytic_therapy"] = 1

# ==========================================
# 5. HYDRATION & WORKBOOK GENERATION
# ==========================================

def generate_excel(template_path: str = TEMPLATE_PATH, output_path: str = OUTPUT_PATH) -> None:
    try:
        wb = load_workbook(template_path)
    except FileNotFoundError:
        wb = Workbook()  # Fallback if template missing (should be provided)

    # --- Sheet 1: Note_Text ---
    ws_text = wb["Note_Text"] if "Note_Text" in wb.sheetnames else wb.create_sheet("Note_Text")
    if ws_text.max_row == 1:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # --- Sheet 2: Note_Index ---
    ws_index = wb["Note_Index"] if "Note_Index" in wb.sheetnames else wb.create_sheet("Note_Index")
    row_meta = [SOURCE_FILE, NOTE_ID, "", "", "", "", "Pending", ""]
    flag_order = [
        "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
        "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy",
        "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration",
        "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
        "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation",
        "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
        "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis",
        "pleural_biopsy", "fibrinolytic_therapy"
    ]
    row_flags = [proc_flags[f] for f in flag_order]
    ws_index.append(row_meta + row_flags)

    # --- Sheet 3 & 4: Spans ---
    ws_span_anno = wb["Span_Annotations"] if "Span_Annotations" in wb.sheetnames else wb.create_sheet("Span_Annotations")
    ws_span_hydra = wb["Span_Hydrated"] if "Span_Hydrated" in wb.sheetnames else wb.create_sheet("Span_Hydrated")

    headers_span = [
        "source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text",
        "match_index", "start_char", "end_char", "span_len", "label", "normalized_value",
        "schema_field", "event_id", "is_negated", "is_historical", "time_anchor", "reviewer",
        "comments", "hydration_status"
    ]
    if ws_span_anno.max_row == 1:
        ws_span_anno.append(headers_span)
    if ws_span_hydra.max_row == 1:
        ws_span_hydra.append(headers_span)

    for idx, span in enumerate(spans):
        span.hydrate(NOTE_TEXT)

        row_base = [
            SOURCE_FILE, NOTE_ID, f"span_{idx+1:03d}", "", span.context_prefix, span.text,
            "", "", "", f"=LEN(F{idx+2})", span.label, span.normalized_value,
            span.schema_field, span.event_id, span.is_negated, span.is_historical, "", "", "",
            "",  # hydration_status
        ]

        row_anno = list(row_base)
        row_anno[19] = "needs_hydration"
        ws_span_anno.append(row_anno)

        row_hydra = list(row_base)
        row_hydra[7] = span.start_char
        row_hydra[8] = span.end_char
        row_hydra[19] = span.hydration_status
        ws_span_hydra.append(row_hydra)

    # --- Sheet 5: Event_Log ---
    ws_event_log = wb["Event_Log"] if "Event_Log" in wb.sheetnames else wb.create_sheet("Event_Log")
    if ws_event_log.max_row == 1:
        ws_event_log.append([
            "source_file", "note_id", "event_id", "event_type", "method", "anatomy_target",
            "device", "needle_gauge", "stations", "counts", "measurements", "specimens", "findings",
            "is_historical", "reviewer", "comments", "device_size", "device_material",
            "outcome_airway_lumen_pre", "outcome_airway_lumen_post", "outcome_symptoms",
            "outcome_pleural", "outcome_complication"
        ])

    for e in events:
        ws_event_log.append([
            SOURCE_FILE, NOTE_ID, e["event_id"], e["type"], e["method"], e["target"]["location"]["lobe"],
            "Chest Tube", "", "", "", e["measurements_json"], "", e["findings_json"],
            False, "", "", "", "", "", "", "", "", e["outcomes"]["complications"]
        ])

    # --- Sheet 6: V3_Procedure_Events ---
    ws_v3 = wb["V3_Procedure_Events"] if "V3_Procedure_Events" in wb.sheetnames else wb.create_sheet("V3_Procedure_Events")
    if ws_v3.max_row == 1:
        ws_v3.append([
            "note_id", "event_id", "type", "target.anatomy_type", "target.location.lobe",
            "target.location.segment", "target.station", "lesion.type", "lesion.size_mm",
            "method", "devices_json", "measurements_json", "specimens_json", "findings_json",
            "evidence_quote", "stent.size", "stent.material_or_brand", "catheter.size_fr",
            "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post", "outcomes.symptoms",
            "outcomes.pleural", "outcomes.complications"
        ])

    for e in events:
        ws_v3.append([
            NOTE_ID, e["event_id"], e["type"], e["target"]["anatomy_type"], e["target"]["location"]["lobe"],
            "", "", "Effusion", "", e["method"], e["devices_json"], e["measurements_json"],
            "", e["findings_json"], "", "", "", "", "", "", "", "", e["outcomes"]["complications"]
        ])

    # --- Sheet 7: V3_Registry_JSON ---
    ws_json = wb["V3_Registry_JSON"] if "V3_Registry_JSON" in wb.sheetnames else wb.create_sheet("V3_Registry_JSON")
    if ws_json.max_row == 1:
        ws_json.append(["schema_version", "note_id", "json_output"])

    final_json = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": events
    }
    ws_json.append(["v3.0", NOTE_ID, json.dumps(final_json, indent=2)])

    wb.save(output_path)
    print(f"Generated {output_path}")


if __name__ == "__main__":
    generate_excel()
