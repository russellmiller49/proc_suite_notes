import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# ==============================================================================
# 1. INPUTS
# ==============================================================================
NOTE_ID = "note_035"
SOURCE_FILE = "note_035.txt"
PROCEDURE_DATE = ""  # Not explicitly stated as current date, historical is 12/15/25
NOTE_TEXT = """NOTE_ID:  note_035 SOURCE_FILE: note_035.txt INDICATION FOR OPERATION:  [REDACTED]is a 47 year old-year-old male who presents with Complicated Effusion.
The nature, purpose, risks, benefits and alternatives to Instillation of agents for fibrinolysis (subsequent) were discussed with the patient in detail.
Patient indicated a wish to proceed with procedure and informed consent was signed.
 
CONSENT : Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient or surrogate.
The patient or surrogate read and signed the provided consent form / provided consent over the phone.
The consent was witnessed by an assisting medical professional.
 
PREOPERATIVE DIAGNOSIS:  Complicated Effusion
POSTOPERATIVE DIAGNOSIS: Same as preoperative diagnosis - see above.
PROCEDURE:  
32562 Instillation(s), via chest tube/catheter, agent for fibrinolysis (eg, fibrinolytic agent for break up of multiloculated effusion);
subsequent day
was present throughout the entire procedure.
 
PROCEDURE IN DETAIL:
 
PATIENT POSITION: 
0‌ Supine  0‌ Sitting   
0‌ Lateral Decubitus:  0‌ Right 0‌ Left 
 
Date of chest tube insertion: 12/15/25 
 
Side: right
1‌  10 mg/5 mg tPA/Dnasedose #:__3__ 
            0‌  ___mg tPA                              dose #:____ 
0‌  Other medication:  
 
COMPLICATIONS:
0‌None 0‌Bleeding-EBL: ___ ml 
0‌Pneumothorax 0‌Re- Expansion Pulmonary Edema 
0‌Other: 
 
IMPRESSION/PLAN: [REDACTED]is a 47 year old-year-old male who presents for Instillation of agents for fibrinolysis (subsequent).
The patient tolerated the procedure well.  There were no immediate complications."""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# ==============================================================================
# 2. DEFINITIONS & CONFIG
# ==============================================================================

# Procedure Flags (30)
PROCEDURE_FLAGS = [
    "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
    "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy",
    "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration",
    "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
    "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation",
    "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
    "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis",
    "pleural_biopsy", "fibrinolytic_therapy"
]

# Spans to Extract
# Format: (span_text, label, normalized_value, schema_field, event_id_ref, context_prefix_opt)
SPANS_DATA = [
    ("Complicated Effusion", "OBS_LESION", "complicated effusion", "indication", "ev1", None),
    ("Instillation of agents for fibrinolysis", "PROC_METHOD", "fibrinolytic instillation", "procedure", "ev1", "The nature, purpose, risks, benefits and alternatives to "),
    ("Side: right", "LATERALITY", "Right", "target.location", "ev1", None),
    ("tPA/Dnase", "PROC_METHOD", "tPA/Dnase", "medication", "ev1", None),
    ("dose #:__3__", "MEAS_COUNT", "3", "dose_number", "ev1", None),
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "outcomes.complications", "ev1", None),
    ("tolerated the procedure well", "OUTCOME_SYMPTOMS", "tolerated well", "outcomes.symptoms", "ev1", None)
]

# Event Log Data
EVENT_DATA = [
    {
        "event_id": "ev1",
        "event_type": "Fibrinolytic Therapy",
        "method": "Instillation of fibrinolytic agents",
        "anatomy_target": "Right Pleural Space",
        "device": "tPA/Dnase",
        "findings": "Complicated Effusion",
        "outcome_complication": "None",
        "outcome_symptoms": "Tolerated well"
    }
]

# ==============================================================================
# 3. UTILITY FUNCTIONS
# ==============================================================================

def clean_text(text):
    return text.replace('\r\n', '\n').strip()

def get_offsets(text, span_text, context_prefix=None, match_index=None):
    """
    Returns (start_char, end_char, hydration_status)
    """
    if not span_text:
        return None, None, "missing_text"
    
    # 1. Exact match count
    count = text.count(span_text)
    
    if count == 0:
        return None, None, "not_found"
    
    if count == 1:
        start = text.find(span_text)
        return start, start + len(span_text), "hydrated_unique"
    
    # 2. Context prefix
    if context_prefix:
        # Search for prefix + some window + span
        # Simple approach: find all occurrences, check preceding text
        indices = [m.start() for m in re.finditer(re.escape(span_text), text)]
        for idx in indices:
            # Look back up to 150 chars
            window_start = max(0, idx - 150)
            preceding = text[window_start:idx]
            if context_prefix in preceding:
                return idx, idx + len(span_text), "hydrated_prefix_window"
        
        # Fallback if prefix not found near (shouldn't happen if data correct)
        return None, None, "ambiguous_prefix_fail"

    # 3. Match index (0-based)
    if match_index is not None and 0 <= match_index < count:
        indices = [m.start() for m in re.finditer(re.escape(span_text), text)]
        start = indices[match_index]
        return start, start + len(span_text), "hydrated_match_index"

    return None, None, f"ambiguous_count={count}"

# ==============================================================================
# 4. MAIN GENERATION LOGIC
# ==============================================================================

def generate_workbook():
    # Load Template
    if not os.path.exists(TEMPLATE_PATH):
        # Create a dummy workbook if template missing (for robustness in some envs, though instructions say it must exist)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # ---------------------------------------------------------
    # Sheet 1: Note_Text
    # ---------------------------------------------------------
    ws_text = wb["Note_Text"] if "Note_Text" in wb.sheetnames else wb.create_sheet("Note_Text")
    if ws_text.max_row == 1 and ws_text["A1"].value is None:
        ws_text.append(["note_id", "source_file", "note_text"])
    
    # Check if note already exists to avoid dupes (simple check)
    exists = False
    for row in ws_text.iter_rows(min_row=2, values_only=True):
        if row[0] == NOTE_ID:
            exists = True
            break
    
    if not exists:
        ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # ---------------------------------------------------------
    # Sheet 2: Note_Index (Flags)
    # ---------------------------------------------------------
    ws_index = wb["Note_Index"] if "Note_Index" in wb.sheetnames else wb.create_sheet("Note_Index")
    # Headers: metadata + 30 flags
    headers = ["source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes"] + PROCEDURE_FLAGS
    
    if ws_index.max_row == 1 and ws_index["A1"].value != headers[0]:
        ws_index.append(headers)
    
    # Determine Flag Values
    # In this note: Fibrinolytic Therapy = 1. Others 0.
    # Chest tube is mentioned but clearly historical ("Date of chest tube insertion: 12/15/25").
    # Current procedure is Instillation.
    
    flag_values = [0] * 30
    try:
        idx = PROCEDURE_FLAGS.index("fibrinolytic_therapy")
        flag_values[idx] = 1
    except ValueError:
        pass

    row_data = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "auto_generated", ""] + flag_values
    ws_index.append(row_data)

    # ---------------------------------------------------------
    # Sheet 3 & 4: Span_Annotations & Span_Hydrated
    # ---------------------------------------------------------
    ws_span = wb["Span_Annotations"] if "Span_Annotations" in wb.sheetnames else wb.create_sheet("Span_Annotations")
    ws_hydrated = wb["Span_Hydrated"] if "Span_Hydrated" in wb.sheetnames else wb.create_sheet("Span_Hydrated")

    span_headers = [
        "source_file", "note_id", "span_id", "section_type",
        "context_prefix", "span_text", "match_index",
        "start_char", "end_char", "span_len",
        "label", "normalized_value", "schema_field", "event_id",
        "is_negated", "is_historical", "time_anchor", "reviewer", "comments",
        "hydration_status"
    ]

    if ws_span.max_row == 1: ws_span.append(span_headers)
    if ws_hydrated.max_row == 1: ws_hydrated.append(span_headers)

    span_id_counter = 1
    
    for item in SPANS_DATA:
        text_val, label, norm_val, field, ev_id, ctx = item
        
        # Calculate hydration
        start, end, status = get_offsets(NOTE_TEXT, text_val, context_prefix=ctx)
        length = len(text_val) if text_val else 0
        
        # Row for Span_Annotations (Blank start/end)
        row_annot = [
            SOURCE_FILE, NOTE_ID, f"span_{span_id_counter:03d}", "Procedure",
            ctx, text_val, "", # match_index left blank unless strictly needed
            "", "", f"=LEN(F{ws_span.max_row+1})", # formula for length
            label, norm_val, field, ev_id,
            False, False, "", "auto", "",
            "needs_hydration"
        ]
        ws_span.append(row_annot)

        # Row for Span_Hydrated (Filled start/end)
        row_hyd = [
            SOURCE_FILE, NOTE_ID, f"span_{span_id_counter:03d}", "Procedure",
            ctx, text_val, "",
            start if start is not None else "", 
            end if end is not None else "",
            length,
            label, norm_val, field, ev_id,
            False, False, "", "auto", "",
            status
        ]
        ws_hydrated.append(row_hyd)
        
        span_id_counter += 1

    # ---------------------------------------------------------
    # Sheet 5: Event_Log
    # ---------------------------------------------------------
    ws_event = wb["Event_Log"] if "Event_Log" in wb.sheetnames else wb.create_sheet("Event_Log")
    event_headers = [
        "source_file", "note_id", "event_id", "event_type", "method",
        "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements",
        "specimens", "findings", "is_historical", "reviewer", "comments",
        "device_size", "device_material",
        "outcome_airway_lumen_pre", "outcome_airway_lumen_post",
        "outcome_symptoms", "outcome_pleural", "outcome_complication"
    ]
    if ws_event.max_row == 1: ws_event.append(event_headers)

    for ev in EVENT_DATA:
        row = [
            SOURCE_FILE, NOTE_ID, ev.get("event_id"), ev.get("event_type"), ev.get("method"),
            ev.get("anatomy_target"), ev.get("device"), "", "", "", "",
            "", ev.get("findings"), False, "auto", "",
            "", "", # size, material
            "", "", # airway outcomes
            ev.get("outcome_symptoms"), "", ev.get("outcome_complication")
        ]
        ws_event.append(row)

    # ---------------------------------------------------------
    # Sheet 6: V3_Procedure_Events
    # ---------------------------------------------------------
    ws_v3 = wb["V3_Procedure_Events"] if "V3_Procedure_Events" in wb.sheetnames else wb.create_sheet("V3_Procedure_Events")
    v3_headers = [
        "note_id", "event_id", "type",
        "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station",
        "lesion.type", "lesion.size_mm",
        "method", "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote",
        "stent.size", "stent.material_or_brand", "catheter.size_fr",
        "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post",
        "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"
    ]
    if ws_v3.max_row == 1: ws_v3.append(v3_headers)

    for ev in EVENT_DATA:
        # Construct JSONs
        devices = [{"name": ev.get("device")}] if ev.get("device") else []
        findings = [{"name": ev.get("findings")}] if ev.get("findings") else []
        
        row = [
            NOTE_ID, ev.get("event_id"), ev.get("event_type"),
            "Pleura", "", "", "", # anatomy
            "Effusion", "", # lesion
            ev.get("method"), json.dumps(devices), "[]", "[]", json.dumps(findings), "",
            "", "", "", # stent/cath
            "", "", # lumen
            ev.get("outcome_symptoms"), "", ev.get("outcome_complication")
        ]
        ws_v3.append(row)

    # ---------------------------------------------------------
    # Sheet 7: V3_Registry_JSON
    # ---------------------------------------------------------
    ws_json = wb["V3_Registry_JSON"] if "V3_Registry_JSON" in wb.sheetnames else wb.create_sheet("V3_Registry_JSON")
    if ws_json.max_row == 1: ws_json.append(["schema_version", "note_id", "json_object"])
    
    registry_obj = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": [
            {
                "event_id": ev.get("event_id"),
                "type": ev.get("event_type"),
                "method": ev.get("method"),
                "target": {"structure": "Pleura", "laterality": "Right"},
                "outcomes": {
                    "symptoms": ev.get("outcome_symptoms"),
                    "complications": ev.get("outcome_complication")
                }
            } for ev in EVENT_DATA
        ],
        "no_immediate_complications": True
    }
    
    ws_json.append(["3.0", NOTE_ID, json.dumps(registry_obj, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()