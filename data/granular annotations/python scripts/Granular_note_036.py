import re
import json
import os
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

# =============================================================================
# INPUT DATA & CONFIGURATION
# =============================================================================

NOTE_ID = "note_036"
SOURCE_FILE = "note_036.txt"
PROCEDURE_DATE = "2026-01-12" # inferred from context or left blank if unknown, using context date
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# Exact text from the source provided
NOTE_TEXT = """NOTE_ID:  note_036 SOURCE_FILE: note_036.txt INDICATION: Pleural Effusion.  

PREOPERATIVE DIAGNOSIS:  Pleural Effusion
POSTOPERATIVE DIAGNOSIS: Same as preoperative diagnosis - see above.
PROCEDURE:  
32552 Removal of indwelling tunneled pleural catheter with cuff
PROCEDURE IN DETAIL:
PATIENT POSITION: 
0‌ Supine  0‌ Sitting   
1‌ Lateral Decubitus:  0‌ Right 1‌ Left 
 
Date of TPC insertion: [REDACTED]

Side: 1‌ Right  0‌ Left 
 
Reason for TPC removal: 
0‌  No output due to auto-pleurodesis 
0‌  No output due to successful talc pleurodesis 
0‌  No output due to blocked TPC 
0‌  Patient request due to discomfort 
1‌  Patient request and partial dislodgement
 
Catheter Site:       1‌  Clean/dry     0‌ Erythema 
   0‌  Tenderness with palpation    0‌  Drainage 
Anesthesia:0‌ Yes 1‌ No 
Lidocaine 1%: ______ ml   Other: ______ 
Sutured: 0‌ Yes 1‌ No 
 
Clean dry dressing applied and should remain in place for a minimum of 72 hours.
Ok to shower with occlusive dressing.  
 
Complications:  
0‌ Yes, if so what: 
1‌  No 
 
Antibiotics: 
0‌ Yes, if so why: 
1‌  No 
 
 
COMPLICATIONS:
1‌None 0‌Bleeding-EBL: ___ ml 0‌Pneumothorax 0‌Re- Expansion Pulmonary Edema 
0‌Other: 
 
IMPRESSION/PLAN: [REDACTED]is a 47 year old-year-old male who presents for Tunnel Pleural Catheter Removal.
The patient tolerated the procedure well.  There were no immediate complications.
-Follow up in Interventional Pulmonary Pleural clinic
-Follow up CXR
 
DISPOSITION: Nursing Unit"""

# =============================================================================
# FLAGGING LOGIC
# =============================================================================

# Default all to 0
PROCEDURE_FLAGS = {
    # Bronchoscopy (23)
    "diagnostic_bronchoscopy": 0, "bal": 0, "bronchial_wash": 0, "brushings": 0,
    "endobronchial_biopsy": 0, "tbna_conventional": 0, "linear_ebus": 0, "radial_ebus": 0,
    "navigational_bronchoscopy": 0, "transbronchial_biopsy": 0, "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0, "foreign_body_removal": 0, "airway_dilation": 0,
    "airway_stent": 0, "thermal_ablation": 0, "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0, "blvr": 0, "peripheral_ablation": 0, "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0, "rigid_bronchoscopy": 0,
    # Pleural (7)
    "thoracentesis": 0, "chest_tube": 0, "ipc": 0, "medical_thoracoscopy": 0,
    "pleurodesis": 0, "pleural_biopsy": 0, "fibrinolytic_therapy": 0
}

# Logic for note_036:
# Procedure is "Removal of indwelling tunneled pleural catheter".
# While technically a removal, we map this to the IPC domain.
PROCEDURE_FLAGS["ipc"] = 1

# =============================================================================
# SPAN ANNOTATIONS (ANCHOR-FIRST)
# =============================================================================

# Spans definition
# Fields: text, label, normalized, section, context_prefix (optional), match_index (optional), event_id (optional)
SPANS_DATA = [
    {
        "text": "Pleural Effusion",
        "label": "OBS_LESION",
        "normalized": "Pleural Effusion",
        "section": "INDICATION",
        "event_id": "ev1",
        "match_index": 0 # First occurrence in Indication
    },
    {
        "text": "Removal of indwelling tunneled pleural catheter with cuff",
        "label": "PROC_METHOD",
        "normalized": "IPC Removal",
        "section": "PROCEDURE",
        "event_id": "ev1"
    },
    {
        "text": "Right",
        "label": "LATERALITY",
        "normalized": "Right",
        "section": "PROCEDURE IN DETAIL",
        "context_prefix": "Side: 1‌ ", # Using context to distinguish from other 'Right'
        "event_id": "ev1"
    },
    {
        "text": "partial dislodgement",
        "label": "OBS_LESION",
        "normalized": "Dislodgement",
        "section": "PROCEDURE IN DETAIL",
        "event_id": "ev1",
        "comments": "Reason for removal"
    },
    {
        "text": "Clean/dry",
        "label": "OBS_LESION",
        "normalized": "clean/dry",
        "section": "PROCEDURE IN DETAIL",
        "event_id": "ev1",
        "comments": "Catheter Site condition"
    },
    {
        "text": "Clean dry dressing applied",
        "label": "PROC_ACTION",
        "normalized": "Dressing applied",
        "section": "PROCEDURE IN DETAIL",
        "event_id": "ev1"
    },
    {
        "text": "No immediate complications",
        "label": "OUTCOME_COMPLICATION",
        "normalized": "None",
        "section": "IMPRESSION/PLAN",
        "event_id": "ev1"
    },
    {
        "text": "Nursing Unit",
        "label": "PROC_ACTION",
        "normalized": "Disposition: Nursing Unit",
        "section": "DISPOSITION",
        "event_id": "ev1"
    }
]

# =============================================================================
# EVENTS & V3 MAPPING
# =============================================================================

EVENTS_DATA = [
    {
        "event_id": "ev1",
        "type": "pleural_procedure",
        "method": "IPC Removal",
        "target": {
            "anatomy_type": "Pleura",
            "location": {"lobe": "Right"} # Mapping Laterality to general location
        },
        "lesion": {
            "type": "Pleural Effusion"
        },
        "findings": {
            "site_condition": "clean/dry",
            "reason": "partial dislodgement"
        },
        "outcomes": {
            "complications": "None",
            "pleural": "Stable"
        }
    }
]

# =============================================================================
# CORE HELPER FUNCTIONS
# =============================================================================

def get_substring_offset(full_text, substring, context_prefix=None, match_index=0):
    """
    Finds the start/end indices of a substring.
    Strategies:
    1. If context_prefix provided: find occurrences where prefix appears in preceding 120 chars.
    2. If match_index provided: pick the Nth occurrence (0-based).
    3. Default: first occurrence.
    """
    if not substring:
        return None, None, "empty_string"
    
    # Strategy 1: Context Prefix
    if context_prefix:
        # iterate all occurrences
        start = 0
        matches = []
        while True:
            idx = full_text.find(substring, start)
            if idx == -1:
                break
            # Check context
            window_start = max(0, idx - 120)
            preceding_text = full_text[window_start:idx]
            if context_prefix in preceding_text:
                matches.append(idx)
            start = idx + 1
        
        if matches:
            # If match_index is specified relative to context matches, use it, else 0
            selected_idx = matches[match_index] if match_index < len(matches) else matches[0]
            return selected_idx, selected_idx + len(substring), "hydrated_prefix_window"
        else:
            return None, None, "prefix_not_found"

    # Strategy 2: Match Index (global)
    occurrences = [m.start() for m in re.finditer(re.escape(substring), full_text)]
    if not occurrences:
        return None, None, "not_found"
    
    if match_index < len(occurrences):
        return occurrences[match_index], occurrences[match_index] + len(substring), "hydrated_match_index"
    
    return None, None, f"index_out_of_bounds_count_{len(occurrences)}"

def create_or_load_workbook():
    if os.path.exists(TEMPLATE_PATH):
        return openpyxl.load_workbook(TEMPLATE_PATH)
    else:
        # Fallback if template missing (should not happen per prompt instructions)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        return wb

def ensure_sheet_headers(ws, headers):
    if ws.max_row == 0 or (ws.max_row == 1 and ws.cell(1,1).value is None):
        ws.append(headers)

# =============================================================================
# MAIN PROCESSING
# =============================================================================

def main():
    wb = create_or_load_workbook()

    # 1. Note_Text Sheet
    if "Note_Text" not in wb.sheetnames:
        wb.create_sheet("Note_Text")
    ws_text = wb["Note_Text"]
    ensure_sheet_headers(ws_text, ["note_id", "source_file", "note_text"])
    # Check if note already exists to avoid duplicates
    existing_ids = [row[0] for row in ws_text.iter_rows(min_row=2, values_only=True)]
    if NOTE_ID not in existing_ids:
        ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index Sheet
    if "Note_Index" not in wb.sheetnames:
        wb.create_sheet("Note_Index")
    ws_index = wb["Note_Index"]
    
    # Headers for Note_Index
    flag_keys = list(PROCEDURE_FLAGS.keys())
    index_headers = [
        "source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes"
    ] + flag_keys
    ensure_sheet_headers(ws_index, index_headers)
    
    # Row Data
    index_row = [
        SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "pending", ""
    ] + [PROCEDURE_FLAGS[k] for k in flag_keys]
    ws_index.append(index_row)

    # 3. Span_Annotations (Anchor First)
    if "Span_Annotations" not in wb.sheetnames:
        wb.create_sheet("Span_Annotations")
    ws_span_base = wb["Span_Annotations"]
    
    span_headers = [
        "source_file", "note_id", "span_id", "section_type",
        "context_prefix", "span_text", "match_index",
        "start_char", "end_char", "span_len",
        "label", "normalized_value", "schema_field", "event_id",
        "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"
    ]
    ensure_sheet_headers(ws_span_base, span_headers)

    # 4. Span_Hydrated
    if "Span_Hydrated" not in wb.sheetnames:
        wb.create_sheet("Span_Hydrated")
    ws_span_hyd = wb["Span_Hydrated"]
    ensure_sheet_headers(ws_span_hyd, span_headers)

    # Process Spans
    hydrated_rows = []
    
    for idx, span_def in enumerate(SPANS_DATA):
        span_id = f"{NOTE_ID}_s{idx+1:02d}"
        
        text = span_def.get("text", "")
        ctx = span_def.get("context_prefix")
        midx = span_def.get("match_index", 0)
        
        # Calculate hydration
        start, end, status = get_substring_offset(NOTE_TEXT, text, context_prefix=ctx, match_index=midx)
        span_len = len(text)
        
        # Base Row (No offsets)
        base_row = [
            SOURCE_FILE, NOTE_ID, span_id, span_def.get("section", ""),
            ctx, text, midx,
            None, None, f"=LEN(F{ws_span_base.max_row + 1})", # Formula for length
            span_def.get("label", ""), span_def.get("normalized", ""), "", span_def.get("event_id", ""),
            0, 0, "", "", span_def.get("comments", ""), "needs_hydration"
        ]
        ws_span_base.append(base_row)
        
        # Hydrated Row
        hyd_row = [
            SOURCE_FILE, NOTE_ID, span_id, span_def.get("section", ""),
            ctx, text, midx,
            start, end, span_len,
            span_def.get("label", ""), span_def.get("normalized", ""), "", span_def.get("event_id", ""),
            0, 0, "", "", span_def.get("comments", ""), status
        ]
        ws_span_hyd.append(hyd_row)

    # 5. Event_Log
    if "Event_Log" not in wb.sheetnames:
        wb.create_sheet("Event_Log")
    ws_events = wb["Event_Log"]
    
    event_log_headers = [
        "source_file", "note_id", "event_id", "event_type", "method",
        "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements",
        "specimens", "findings", "is_historical", "reviewer", "comments",
        "device_size", "device_material",
        "outcome_airway_lumen_pre", "outcome_airway_lumen_post",
        "outcome_symptoms", "outcome_pleural", "outcome_complication"
    ]
    ensure_sheet_headers(ws_events, event_log_headers)

    for ev in EVENTS_DATA:
        row = [
            SOURCE_FILE, NOTE_ID, ev["event_id"], ev["type"], ev["method"],
            json.dumps(ev["target"]), "", "", "", "", "",
            "", json.dumps(ev["findings"]), 0, "", "",
            "", "", # device size/mat
            "", "", # airway lumen
            "", ev["outcomes"].get("pleural", ""), ev["outcomes"].get("complications", "")
        ]
        ws_events.append(row)

    # 6. V3_Procedure_Events
    if "V3_Procedure_Events" not in wb.sheetnames:
        wb.create_sheet("V3_Procedure_Events")
    ws_v3 = wb["V3_Procedure_Events"]
    
    v3_headers = [
        "note_id", "event_id", "type",
        "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station",
        "lesion.type", "lesion.size_mm",
        "method", "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote",
        "stent.size", "stent.material_or_brand", "catheter.size_fr",
        "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post",
        "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"
    ]
    ensure_sheet_headers(ws_v3, v3_headers)

    for ev in EVENTS_DATA:
        tgt = ev["target"]
        loc = tgt.get("location", {})
        
        row = [
            NOTE_ID, ev["event_id"], ev["type"],
            tgt.get("anatomy_type", ""), loc.get("lobe", ""), loc.get("segment", ""), "",
            ev["lesion"].get("type", ""), "",
            ev["method"], "[]", "[]", "[]", json.dumps(ev["findings"]), "",
            "", "", "", # stent/cath
            "", "", # lumen
            "", ev["outcomes"].get("pleural", ""), ev["outcomes"].get("complications", "")
        ]
        ws_v3.append(row)

    # 7. V3_Registry_JSON
    if "V3_Registry_JSON" not in wb.sheetnames:
        wb.create_sheet("V3_Registry_JSON")
    ws_json = wb["V3_Registry_JSON"]
    
    # Construct final JSON structure
    registry_obj = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": EVENTS_DATA,
        "no_immediate_complications": True # derived from note text
    }
    
    ws_json.append(["schema_version", "note_id", "json_content"])
    ws_json.append(["3.0", NOTE_ID, json.dumps(registry_obj, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Successfully generated {OUTPUT_PATH}")

if __name__ == "__main__":
    main()