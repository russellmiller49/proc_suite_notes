import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# =============================================================================
# INPUT DATA & CONFIGURATION
# =============================================================================

NOTE_ID = "note_037"
SOURCE_FILE = "note_037.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_037 SOURCE_FILE: note_037.txt INDICATION FOR OPERATION:  [REDACTED]is a 76 year old-year-old female who presents with airway stenosis.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
 
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
 
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31640 Bronchoscopy with excision 
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
 
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)
 
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required multiple modalities for debulking and to treat bronchial stenosis.
This resulted in >100% increased work due to Increased intensity, Time, Technical difficulty of procedure, and Physical and mental effort required.
Apply to: 31640 Bronchoscopy with excision 
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy).
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
Flexible Hybrid (Pedatric) Bronchoscope
 
ESTIMATED BLOOD LOSS:   Moderate
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: Supine
 
Initial Airway Inspection Findings:
 
The laryngeal mask airway is in good position.
Pharynx: Not assessed due to bronchoscopy introduction through LMA.
Larynx: Normal.
Vocal Cords: Tissue/web at anterior commissure 
Trachea: Extrinsic compression from tumor at distal trachea, right side as well as some extruding tumor
Main Carina: Sharp
Right Lung Proximal Airways: Obstructed airways at RUL and beyond the proximal bronchus intermedius secondary to extrinsic compression and tumor in-growth 
Left Lung Proximal Airways: Normal anatomic branching to segmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Mucosa: Erythematous and Friable.
Secretions: Minimal, thin, and clear mucus.
Blood at the bronchus intermedius and distal right-sided airways   
 
Successful therapeutic aspiration was performed to clean out the Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus.
TISSUE AT VOCAL CORDS/ANTERIOR COMMISSURE 
 
 
EXTRINSIC COMPRESSION / INGROWING TUMOR AT DISTAL TRACHEA 
 
 
BRONCHUS INTERMEDIUS
 
 
Endobronchial tumor was noted and excised with mechanical debridement using forceps.
Bleeding/oozing was noted from the airways so decision was made to intubate patient with endotracheal tube.
Anesthesia placed 8.0mm ETT without issue.
 
Endobronchial obstruction at Bronchus Intermedius  was treated with the following modalities:
 
Modality	Tools	Setting/Mode	Duration	Results
Mechanical	Forceps	N/A	N/A	Tissue/tumor debulking
APC	2.3mm Straightfire probe	Forced, effect 3	 	Tissue/tumor debulking and hemostasis
Cryoprobe	2.4mm Cryoprobe	 	3-10 second applications	Tissue/tumor debulking and hemostasis
 
Prior to treatment, affected airway was note to be 5% patent.
After treatment, the airway was 15% patent. 
 
Bleeding/oozing was treated with cold saline, TXA (total 1000mg), and epinephrine (total 1000mg).
BRONCHUS INTERMEDIUS AT CONCLUSION
 
 
DISTAL RIGHT-SIDED AIRWAYS (RIGHT LOWER LOBE) AT CONCLUSION
 
 
The patient tolerated the procedure well.
There were no immediate complications.  At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
--None
 
IMPRESSION/PLAN: [REDACTED]is a 76 year old-year-old female who presents for bronchoscopy for evaluation of airway stenosis.
Patient was noted to have extensive endobronchial tumor involvement at the right-sided airways.
This was treated with multiple modalities including forceps, APC, and cryotherapy.
At the conclusion of the case patent distal right-sided airways were identified.
--Post procedure CXR
--Continued care per primary team
--If patient has scant hemoptysis, would treat with TXA 500mg NEB Q8h"""

# =============================================================================
# DATA STRUCTURES & LOGIC
# =============================================================================

def create_span(span_text, label, normalized_value, event_id, context_prefix=None, schema_field=None):
    return {
        "source_file": SOURCE_FILE,
        "note_id": NOTE_ID,
        "span_id": f"{NOTE_ID}_span_{len(spans) + 1:03d}",
        "section_type": "Procedure",
        "context_prefix": context_prefix if context_prefix else "",
        "span_text": span_text,
        "match_index": 0, # Default to 0, logic handles specific indexing if needed
        "start_char": "",
        "end_char": "",
        "span_len": f'=LEN(F{len(spans) + 2})',
        "label": label,
        "normalized_value": normalized_value,
        "schema_field": schema_field if schema_field else "",
        "event_id": event_id,
        "is_negated": "FALSE",
        "is_historical": "FALSE",
        "time_anchor": "",
        "reviewer": "AI_Model",
        "comments": "",
        "hydration_status": "needs_hydration"
    }

spans = []

# --- Event IDs ---
EVT_DX = "evt_diagnosis"
EVT_GEN = "evt_general"
EVT_ASP = "evt_aspiration"
EVT_MECH = "evt_forceps"
EVT_APC = "evt_apc"
EVT_CRYO = "evt_cryo"
EVT_INTUB = "evt_intubation"
EVT_OUT = "evt_outcome"
EVT_DRUGS = "evt_drugs"

# --- 1. Diagnosis/Indications ---
spans.append(create_span("airway stenosis", "OBS_LESION", "stenosis", EVT_DX, "presents with "))
spans.append(create_span("J98.09", "OBS_LESION", "Other diseases of bronchus", EVT_DX, "PREOPERATIVE DIAGNOSIS: "))

# --- 2. General / Anesthesia ---
spans.append(create_span("General Anesthesia", "PROC_METHOD", "General Anesthesia", EVT_GEN))
spans.append(create_span("Flexible Therapeutic Bronchoscope", "DEV_INSTRUMENT", "Flexible Therapeutic Bronchoscope", EVT_GEN))
spans.append(create_span("Supine", "PROC_METHOD", "Supine", EVT_GEN))
spans.append(create_span("laryngeal mask airway", "DEV_INSTRUMENT", "LMA", EVT_GEN))

# --- 3. Findings (Pre-Op) ---
spans.append(create_span("Tissue/web", "OBS_LESION", "web", EVT_DX, "Vocal Cords: "))
spans.append(create_span("anterior commissure", "ANAT_AIRWAY", "Larynx", EVT_DX))
spans.append(create_span("Extrinsic compression", "OBS_LESION", "extrinsic_compression", EVT_DX, "Trachea: "))
spans.append(create_span("distal trachea", "ANAT_AIRWAY", "Trachea", EVT_DX))
spans.append(create_span("tumor in-growth", "OBS_LESION", "tumor", EVT_DX))
spans.append(create_span("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", EVT_DX, "proximal "))

# --- 4. Therapeutic Aspiration ---
spans.append(create_span("Therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", EVT_ASP, "31645 "))
spans.append(create_span("Trachea (Distal 1/3)", "ANAT_AIRWAY", "Trachea", EVT_ASP, "clean out the "))
spans.append(create_span("Right Mainstem", "ANAT_AIRWAY", "RMS", EVT_ASP))
spans.append(create_span("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", EVT_ASP, "Right Mainstem, "))
spans.append(create_span("mucus", "OBS_LESION", "mucus", EVT_ASP))

# --- 5. Mechanical Debulking (Forceps) ---
spans.append(create_span("mechanical debridement", "PROC_ACTION", "debridement", EVT_MECH))
spans.append(create_span("forceps", "DEV_INSTRUMENT", "Forceps", EVT_MECH, "using "))
spans.append(create_span("Forceps", "DEV_INSTRUMENT", "Forceps", EVT_MECH, "Mechanical\t"))
spans.append(create_span("Tissue/tumor debulking", "PROC_ACTION", "debulking", EVT_MECH, "N/A\t"))

# --- 6. APC (Thermal) ---
spans.append(create_span("APC", "PROC_METHOD", "APC", EVT_APC, "Mechanical\t"))
spans.append(create_span("2.3mm Straightfire probe", "DEV_INSTRUMENT", "2.3mm Straightfire probe", EVT_APC))
spans.append(create_span("hemostasis", "PROC_ACTION", "hemostasis", EVT_APC, "debulking and "))

# --- 7. Cryotherapy ---
spans.append(create_span("Cryoprobe", "PROC_METHOD", "Cryotherapy", EVT_CRYO, "APC\t")) # Table logic
spans.append(create_span("2.4mm Cryoprobe", "DEV_INSTRUMENT", "2.4mm Cryoprobe", EVT_CRYO))
spans.append(create_span("3-10 second", "MEAS_COUNT", "3-10s", EVT_CRYO))

# --- 8. Intubation (Mid-procedure) ---
spans.append(create_span("endotracheal tube", "DEV_INSTRUMENT", "ETT", EVT_INTUB))
spans.append(create_span("8.0mm", "MEAS_SIZE", "8.0mm", EVT_INTUB, "placed "))

# --- 9. Outcomes (Lumen) ---
# "Prior to treatment, affected airway was note to be 5% patent."
spans.append(create_span("5% patent", "OUTCOME_AIRWAY_LUMEN_PRE", "5", EVT_OUT, "note to be "))
# "After treatment, the airway was 15% patent."
spans.append(create_span("15% patent", "OUTCOME_AIRWAY_LUMEN_POST", "15", EVT_OUT, "airway was "))

# --- 10. Complications / Drugs ---
spans.append(create_span("No immediate complications", "OUTCOME_COMPLICATION", "None", EVT_OUT))
spans.append(create_span("TXA", "DEV_INSTRUMENT", "Tranexamic Acid", EVT_DRUGS, "saline, "))
spans.append(create_span("epinephrine", "DEV_INSTRUMENT", "Epinephrine", EVT_DRUGS))

# =============================================================================
# HYDRATION LOGIC
# =============================================================================

def get_snippet(text, start, length=120):
    s = max(0, start - length)
    return text[s:start]

def hydrate_spans(full_text, span_list):
    # Sort by length desc to handle overlapping substrings if needed, but standard logic usually ok
    # We maintain order for the sheet though.
    
    for span in span_list:
        stext = span['span_text']
        prefix = span['context_prefix']
        
        matches = [m.start() for m in re.finditer(re.escape(stext), full_text)]
        
        if not matches:
            span['hydration_status'] = "not_found"
            continue
            
        found_start = -1
        
        if len(matches) == 1:
            found_start = matches[0]
            span['hydration_status'] = "hydrated_unique"
        elif prefix:
            # Look for prefix in preceding window
            candidates = []
            for m_start in matches:
                snippet = get_snippet(full_text, m_start)
                if prefix in snippet:
                    candidates.append(m_start)
            if len(candidates) >= 1:
                # If multiple, take first or use match_index logic if we implemented strict indexing
                found_start = candidates[0] 
                span['hydration_status'] = "hydrated_prefix"
            else:
                span['hydration_status'] = "ambiguous_prefix_not_found"
        else:
            # Ambiguous with no prefix
            # Simple fallback: take the first one or leave empty? 
            # Per instructions: "Else: ambiguous."
            # We will default to first for "best guess" but mark ambiguous
            found_start = matches[0]
            span['hydration_status'] = f"ambiguous_count={len(matches)}"

        if found_start != -1:
            span['start_char'] = found_start
            span['end_char'] = found_start + len(stext)

hydrate_spans(NOTE_TEXT, spans)

# =============================================================================
# EXCEL GENERATION
# =============================================================================

def generate_excel(template_path: str = TEMPLATE_PATH, output_path: str = OUTPUT_PATH) -> None:
    try:
        wb = openpyxl.load_workbook(template_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # 1. Note_Text Sheet
    if "Note_Text" not in wb.sheetnames:
        wb.create_sheet("Note_Text")
    ws_text = wb["Note_Text"]
    if ws_text.max_row == 1:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index Sheet (Procedure Flags)
    if "Note_Index" not in wb.sheetnames:
        wb.create_sheet("Note_Index")
    ws_index = wb["Note_Index"]

    flags = {
        "diagnostic_bronchoscopy": 1,
        "bal": 0,
        "bronchial_wash": 0,
        "brushings": 0,
        "endobronchial_biopsy": 0,
        "tbna_conventional": 0,
        "linear_ebus": 0,
        "radial_ebus": 0,
        "navigational_bronchoscopy": 0,
        "transbronchial_biopsy": 0,
        "transbronchial_cryobiopsy": 0,
        "therapeutic_aspiration": 1,
        "foreign_body_removal": 0,
        "airway_dilation": 0,
        "airway_stent": 0,
        "thermal_ablation": 1,
        "tumor_debulking_non_thermal": 1,
        "cryotherapy": 1,
        "blvr": 0,
        "peripheral_ablation": 0,
        "bronchial_thermoplasty": 0,
        "whole_lung_lavage": 0,
        "rigid_bronchoscopy": 0,
        "thoracentesis": 0,
        "chest_tube": 0,
        "ipc": 0,
        "medical_thoracoscopy": 0,
        "pleurodesis": 0,
        "pleural_biopsy": 0,
        "fibrinolytic_therapy": 0
    }

    row_meta = [
        SOURCE_FILE, NOTE_ID, "", "", "", "", "Complete", ""
    ]
    row_flags = list(flags.values())
    ws_index.append(row_meta + row_flags)

    # 3. Span_Annotations
    if "Span_Annotations" not in wb.sheetnames:
        wb.create_sheet("Span_Annotations")
    ws_anno = wb["Span_Annotations"]
    if ws_anno.max_row == 0:
        ws_anno.append(["source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text",
                        "match_index", "start_char", "end_char", "span_len", "label", "normalized_value",
                        "schema_field", "event_id", "is_negated", "is_historical", "time_anchor",
                        "reviewer", "comments", "hydration_status"])

    for s in spans:
        ws_anno.append([
            s["source_file"], s["note_id"], s["span_id"], s["section_type"], s["context_prefix"], s["span_text"],
            s["match_index"], "", "", s["span_len"], s["label"], s["normalized_value"],
            s["schema_field"], s["event_id"], s["is_negated"], s["is_historical"], s["time_anchor"],
            s["reviewer"], s["comments"], s["hydration_status"]
        ])

    # 4. Span_Hydrated
    if "Span_Hydrated" not in wb.sheetnames:
        wb.create_sheet("Span_Hydrated")
    ws_hydra = wb["Span_Hydrated"]
    if ws_hydra.max_row == 0:
        ws_hydra.append(["source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text",
                         "match_index", "start_char", "end_char", "span_len", "label", "normalized_value",
                         "schema_field", "event_id", "is_negated", "is_historical", "time_anchor",
                         "reviewer", "comments", "hydration_status"])

    for s in spans:
        ws_hydra.append([
            s["source_file"], s["note_id"], s["span_id"], s["section_type"], s["context_prefix"], s["span_text"],
            s["match_index"], s["start_char"], s["end_char"], s["span_len"], s["label"], s["normalized_value"],
            s["schema_field"], s["event_id"], s["is_negated"], s["is_historical"], s["time_anchor"],
            s["reviewer"], s["comments"], s["hydration_status"]
        ])

    # 5. Event_Log (Flat)
    if "Event_Log" not in wb.sheetnames:
        wb.create_sheet("Event_Log")
    ws_event = wb["Event_Log"]
    headers_event = ["source_file", "note_id", "event_id", "event_type", "method", "anatomy_target", "device",
                     "needle_gauge", "stations", "counts", "measurements", "specimens", "findings", "is_historical",
                     "reviewer", "comments", "device_size", "device_material",
                     "outcome_airway_lumen_pre", "outcome_airway_lumen_post", "outcome_symptoms", "outcome_pleural", "outcome_complication"]
    if ws_event.max_row == 0:
        ws_event.append(headers_event)

    events_data = [
        {
            "id": EVT_DX, "type": "Assessment", "method": "N/A", "anatomy": "Airway, Trachea",
            "findings": "Stenosis, Tumor, Extrinsic Compression", "outcome_complication": "None"
        },
        {
            "id": EVT_ASP, "type": "Therapeutic", "method": "Aspiration", "anatomy": "Trachea, RMS, BI",
            "findings": "Mucus"
        },
        {
            "id": EVT_MECH, "type": "Therapeutic", "method": "Mechanical Debulking", "anatomy": "Bronchus Intermedius",
            "device": "Forceps"
        },
        {
            "id": EVT_APC, "type": "Therapeutic", "method": "APC", "anatomy": "Bronchus Intermedius",
            "device": "2.3mm Straightfire probe"
        },
        {
            "id": EVT_CRYO, "type": "Therapeutic", "method": "Cryotherapy", "anatomy": "Bronchus Intermedius",
            "device": "2.4mm Cryoprobe", "measurements": "3-10s"
        },
        {
            "id": EVT_OUT, "type": "Outcome", "method": "N/A", "anatomy": "Right Sided Airways",
            "outcome_airway_lumen_pre": "5%", "outcome_airway_lumen_post": "15%", "outcome_complication": "None"
        }
    ]

    for e in events_data:
        row = [
            SOURCE_FILE, NOTE_ID, e["id"], e["type"], e.get("method", ""), e.get("anatomy", ""), e.get("device", ""),
            "", "", "", e.get("measurements", ""), "", e.get("findings", ""), "FALSE", "AI", "",
            "", "",
            e.get("outcome_airway_lumen_pre", ""), e.get("outcome_airway_lumen_post", ""), "", "", e.get("outcome_complication", "")
        ]
        ws_event.append(row)

    # 6. V3_Procedure_Events (Nested)
    if "V3_Procedure_Events" not in wb.sheetnames:
        wb.create_sheet("V3_Procedure_Events")
    ws_v3 = wb["V3_Procedure_Events"]
    headers_v3 = ["note_id", "event_id", "type", "target.anatomy_type", "target.location.lobe",
                  "target.location.segment", "target.station", "lesion.type", "lesion.size_mm",
                  "method", "devices_json", "measurements_json", "specimens_json", "findings_json",
                  "evidence_quote", "stent.size", "stent.material_or_brand", "catheter.size_fr",
                  "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post", "outcomes.symptoms",
                  "outcomes.pleural", "outcomes.complications"]
    if ws_v3.max_row == 0:
        ws_v3.append(headers_v3)

    v3_rows = [
        {"eid": EVT_ASP, "type": "therapeutic_aspiration", "anat": "airway", "lobe": "right_main", "method": "suction", "dev": [], "find": ["mucus"]},
        {"eid": EVT_MECH, "type": "tumor_debulking_non_thermal", "anat": "airway", "lobe": "bronchus_intermedius", "method": "forceps", "dev": [{"name": "Forceps"}], "find": ["tumor"]},
        {"eid": EVT_APC, "type": "thermal_ablation", "anat": "airway", "lobe": "bronchus_intermedius", "method": "apc", "dev": [{"name": "2.3mm Straightfire probe"}], "find": ["hemostasis"]},
        {"eid": EVT_CRYO, "type": "cryotherapy", "anat": "airway", "lobe": "bronchus_intermedius", "method": "cryoprobe", "dev": [{"name": "2.4mm Cryoprobe"}], "meas": [{"name": "duration", "value": "3-10s"}]},
        {"eid": EVT_OUT, "type": "outcome", "anat": "airway", "lobe": "right_lower_lobe", "lumen_pre": "5", "lumen_post": "15", "comp": "None"},
    ]

    for v in v3_rows:
        row = [
            NOTE_ID, v["eid"], v["type"], v["anat"], v.get("lobe", ""), "", "", "", "",
            v.get("method", ""),
            json.dumps(v.get("dev", [])),
            json.dumps(v.get("meas", [])),
            "[]",
            json.dumps(v.get("find", [])),
            "", "", "", "",
            v.get("lumen_pre", ""), v.get("lumen_post", ""), "", "", v.get("comp", "")
        ]
        ws_v3.append(row)

    # 7. JSON Generation
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": v3_rows
    }

    if "V3_Registry_JSON" not in wb.sheetnames:
        wb.create_sheet("V3_Registry_JSON")
    ws_json = wb["V3_Registry_JSON"]
    ws_json.append([json.dumps(registry_data, indent=2)])

    wb.save(output_path)
    print(f"Generated {output_path}")


if __name__ == "__main__":
    generate_excel()
