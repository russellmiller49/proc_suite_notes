import openpyxl
from openpyxl.utils import get_column_letter
import json
import os
import re
from datetime import datetime

# =============================================================================
# INPUT CONFIGURATION
# =============================================================================
NOTE_ID = "note_039"
SOURCE_FILE = "note_039.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_039 SOURCE_FILE: note_039.txt INDICATION FOR OPERATION:  is a 41 year old-year-old female who presents with endobronchial obstruction.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
 
CONSENT : Obtained before the procedure.
Its indications and potential complications and alternatives were discussed with the patient or surrogate.
The patient or surrogate read and signed the provided consent form / provided consent over the phone.
The consent was witnessed by an assisting medical professional.
 
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
 
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
 
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31624 Dx bronchoscope/lavage (BAL)    
31625 Endobronchial Biopsy(s)
31640 Bronchoscopy with excision 
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
 
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)
 
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required 
tumor debulking and relief of endobronchial obstruction via multiple modalities.
This resulted in >100% increased work due to Increased intensity, Time, Technical difficulty of procedure, and Physical and mental effort required.
Apply to: 31640 Bronchoscopy with excision 
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy).
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Rigid Bronchoscope
Flexible Therapeutic Bronchoscope
 
ESTIMATED BLOOD LOSS:   Minimum
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: Supine
 
After induction of muscle relaxants, tooth or gum protector was placed.
The black rigid barrel was introduced through the mouth and advanced in the midline while keeping the alignment with the axis of the trachea and minimizing pressure to the teeth.
The vocal cords were identified and the rigid bronchoscope was advanced carefully while minimizing contact with them.
Once the rigid bronchoscope was positioned at the mid-trachea, jet ventilation was initiated and chest wall movement was confirmed.
Initial Airway Inspection Findings:
Vocal Cords: Normal without mass/lesions
Trachea: Normal.
Main Carina: Sharp
Right Lung Proximal Airways: Right mainstem bronchus with endobronchial mass
Left Lung Proximal Airways: Normal anatomic branching to segmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Mucosa: Normal.
Secretions: Moderate thick and thin light yellow mucus/secretions.
RIGHT MAINSTEM BRONCHUS ENDOBRONCHIAL TUMOR
 
 
Successful therapeutic aspiration was performed to clean out the Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus.
Electrocautery snare was utilized to ensnare the endobronchial tumor and remove the brunt of the endbronchial lesion from its stalk emanating from the RUL carina (RC1).
Once removed from its stalk, the cryoprobe was used to freeze/grasp the endobronchial mass and remove it en bloc with the flexible bronchoscope through the rigid bronchoscope.
RESIDUAL TUMOR AFTER SNARE
 
 
Next, mini microwave catheter was used to treat the residual tumor. Two applications/treatments were administered.
A combination of cryotherapy and APC were used to remove tumor from the right upper lobe and debulk.
In sum, the endobronchial obstruction at Right Mainstem and RUL Carina (RC1) was treated with the following modalities:
 
Modality	Tools	Setting/Mode	Duration	Results
Electrocautery	2cm Snare	EndoCut Q	3-5 second applications	Good tumor destruction/debulking
APC	1.5mm Straightfire probe	Forced, Flow 0.5 LPM, 40W	2-3 second applications	Good tumor destruction/debulking
Microwave	Mini antenna	1 kJ, 62⁰C
 
2.5 kJ, 80⁰C	1min 20sec
 
3min	Good tumor destruction/debulking
Cryoprobe	2.4mm Cryoprobe	N/A	15-30 second freezes	Good tumor destruction/debulking
 
Prior to treatment, affected airway was note to be 1% patent.
After treatment, the airway was 100% patent. 
 
Bronchial alveolar lavage was performed at Superior Segment of Lingula (LB4) and Inferior Segment of Lingula (LB5).
Instilled 60 cc of NS, suction returned with 20 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
 
 
VIEW FROM RMSB AT THE CONCLUSION OF PROCEDURE
 
 
VIEW OF RIGHT UPPER LOBE AT THE CONCLUSION OF PROCEDURE
 
 
The patient tolerated the procedure well.
There were no immediate complications.  At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
--Right mainstem endobronchial tumor (path)
--LUL BAL (cell count, cyto, micro)
 
IMPRESSION/PLAN: [REDACTED]is a 41 year old-year-old female who presents for bronchoscopy for evaluation and management of right mainstem endobronchial tumor.
The tumor was debulked using a combination of mechanical debridement, cryotherapy, APC, electrocautery, and microwave ablation.
Total patency was restored to the right mainstem bronchus at the conclusion of the procedure.
The patient tolerated the procedure well and there were no immediate complications.
--Post procedure CXR
--Follow up path and BAL results
--Recommend treatment for post-obstructive PNA if not already initiated
--Repeat bronchoscopy with Interventional Pulmonology in 4 weeks"""

# =============================================================================
# DATA EXTRACTION
# =============================================================================

# 1. Procedure Flags
PROC_FLAGS = {
    "diagnostic_bronchoscopy": 1, # Code 31624/31625
    "bal": 1,                     # Code 31624
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 1,    # Code 31625
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,  # Code 31645
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 1,        # Code 31641 (APC/Microwave)
    "tumor_debulking_non_thermal": 1, # Code 31640 (Snare/Mechanical)
    "cryotherapy": 1,             # Cryotherapy mentioned
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 1,      # Mentioned explicitly
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0,
}

# 2. Events & Spans
events_list = [
    {
        "event_id": "evt_01",
        "type": "Rigid Bronchoscopy",
        "method": "rigid_bronchoscopy",
        "findings": "Right mainstem endobronchial mass",
        "devices": ["Rigid Bronchoscope"],
        "anatomy": "trachea"
    },
    {
        "event_id": "evt_02",
        "type": "Therapeutic Aspiration",
        "method": "therapeutic_aspiration",
        "anatomy": "Right Mainstem, Bronchus Intermedius, Left Mainstem, Carina",
        "findings": "Moderate thick and thin light yellow mucus/secretions"
    },
    {
        "event_id": "evt_03",
        "type": "Tumor Debulking (Snare/Electro)",
        "method": "tumor_debulking_non_thermal",
        "anatomy": "RUL carina (RC1)",
        "devices": ["Electrocautery snare", "2cm Snare"],
        "action": "ensnare the endobronchial tumor"
    },
    {
        "event_id": "evt_04",
        "type": "Tumor Debulking (Cryo)",
        "method": "cryotherapy",
        "anatomy": "Right mainstem",
        "devices": ["Cryoprobe", "2.4mm Cryoprobe"],
        "action": "freeze/grasp"
    },
    {
        "event_id": "evt_05",
        "type": "Tumor Debulking (Microwave)",
        "method": "thermal_ablation",
        "anatomy": "Right mainstem",
        "devices": ["mini microwave catheter", "Mini antenna"],
        "action": "treat the residual tumor"
    },
    {
        "event_id": "evt_06",
        "type": "Tumor Debulking (APC)",
        "method": "thermal_ablation",
        "anatomy": "right upper lobe",
        "devices": ["APC", "1.5mm Straightfire probe"],
        "action": "debulk"
    },
    {
        "event_id": "evt_07",
        "type": "Outcome Assessment",
        "method": "inspection",
        "anatomy": "Right Mainstem",
        "outcomes": {
            "lumen_pre": "1%",
            "lumen_post": "100%",
            "complication": "None"
        }
    },
    {
        "event_id": "evt_08",
        "type": "BAL",
        "method": "bal",
        "anatomy": "Superior Segment of Lingula (LB4)",
        "specimens": ["LUL BAL"]
    }
]

spans_data = [
    # Metadata / Instruments
    ("Rigid Bronchoscope", "DEV_INSTRUMENT", "Rigid Bronchoscope", "evt_01"),
    ("Flexible Therapeutic Bronchoscope", "DEV_INSTRUMENT", "Flexible Bronchoscope", "evt_01"),
    ("endobronchial obstruction", "OBS_LESION", "endobronchial obstruction", "evt_01"),
    
    # Inspection / Anatomy
    ("Right mainstem bronchus with endobronchial mass", "OBS_LESION", "endobronchial mass", "evt_01"),
    ("Moderate thick and thin light yellow mucus", "OBS_LESION", "mucus", "evt_02"),
    
    # Therapeutic Aspiration
    ("Therapeutic aspiration", "PROC_METHOD", "therapeutic_aspiration", "evt_02"),
    ("Right Mainstem", "ANAT_AIRWAY", "RMS", "evt_02"),
    ("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "evt_02"),
    ("Left Mainstem", "ANAT_AIRWAY", "LMS", "evt_02"),
    ("LUL Lingula Carina (Lc1)", "ANAT_AIRWAY", "LUL Lingula Carina", "evt_02"),
    
    # Tumor - Snare
    ("Electrocautery snare", "DEV_INSTRUMENT", "Electrocautery snare", "evt_03"),
    ("ensnare the endobronchial tumor", "PROC_ACTION", "ensnare", "evt_03"),
    ("RUL carina (RC1)", "ANAT_AIRWAY", "RUL Carina", "evt_03"),
    ("2cm Snare", "DEV_INSTRUMENT", "2cm Snare", "evt_03"),
    ("EndoCut Q", "MEAS_OTHER", "EndoCut Q", "evt_03"),
    
    # Tumor - Cryo
    ("cryoprobe", "DEV_INSTRUMENT", "cryoprobe", "evt_04"),
    ("freeze/grasp", "PROC_ACTION", "freeze_grasp", "evt_04"),
    ("2.4mm Cryoprobe", "DEV_INSTRUMENT", "2.4mm Cryoprobe", "evt_04"),
    
    # Tumor - Microwave
    ("mini microwave catheter", "DEV_INSTRUMENT", "mini microwave catheter", "evt_05"),
    ("Microwave", "PROC_METHOD", "thermal_ablation", "evt_05"),
    ("Mini antenna", "DEV_INSTRUMENT", "Mini antenna", "evt_05"),
    ("1 kJ, 62⁰C", "MEAS_OTHER", "1 kJ, 62C", "evt_05"),
    
    # Tumor - APC
    ("APC", "PROC_METHOD", "thermal_ablation", "evt_06"),
    ("1.5mm Straightfire probe", "DEV_INSTRUMENT", "1.5mm Straightfire probe", "evt_06"),
    ("40W", "MEAS_OTHER", "40W", "evt_06"),
    
    # Outcomes
    ("affected airway was note to be 1% patent", "OUTCOME_AIRWAY_LUMEN_PRE", "1", "evt_07"),
    ("airway was 100% patent", "OUTCOME_AIRWAY_LUMEN_POST", "100", "evt_07"),
    
    # BAL
    ("Bronchial alveolar lavage", "PROC_METHOD", "bal", "evt_08"),
    ("Superior Segment of Lingula (LB4)", "ANAT_LUNG_LOC", "LUL_Lingula_Superior", "evt_08"),
    ("Inferior Segment of Lingula (LB5)", "ANAT_LUNG_LOC", "LUL_Lingula_Inferior", "evt_08"),
    ("Instilled 60 cc", "MEAS_VOL", "60", "evt_08"),
    ("suction returned with 20 cc", "MEAS_VOL", "20", "evt_08"),
    
    # Outcome Complications
    ("There were no immediate complications", "OUTCOME_COMPLICATION", "none", "evt_07"),
]

# =============================================================================
# LOGIC
# =============================================================================

def hydrate_span(full_text, span_text, context_prefix=None, match_index=None):
    """
    Finds start_char and end_char for a span.
    Logic:
    1. If context_prefix provided, search near it.
    2. If match_index provided, pick that occurrence (0-based).
    3. If unique, pick it.
    4. Else, ambiguous.
    """
    matches = [m.start() for m in re.finditer(re.escape(span_text), full_text)]
    
    if not matches:
        return None, None, "not_found"
    
    if len(matches) == 1:
        start = matches[0]
        return start, start + len(span_text), "hydrated_unique"
        
    if context_prefix:
        # Search for context prefix
        c_matches = [m.start() for m in re.finditer(re.escape(context_prefix), full_text)]
        for c_start in c_matches:
            # Look for span within reasonable window after context
            window_end = min(len(full_text), c_start + len(context_prefix) + 150)
            chunk = full_text[c_start:window_end]
            s_in_chunk = chunk.find(span_text)
            if s_in_chunk != -1:
                abs_start = c_start + s_in_chunk
                return abs_start, abs_start + len(span_text), "hydrated_prefix_window"

    if match_index is not None and match_index < len(matches):
        start = matches[match_index]
        return start, start + len(span_text), "hydrated_match_index"

    # Fallback: Ambiguous, return first but mark it
    start = matches[0]
    return start, start + len(span_text), f"ambiguous_count={len(matches)}"

def generate_workbook():
    # 1. Load Template
    if not os.path.exists(TEMPLATE_PATH):
        # Create a dummy one for standalone execution if template missing (though instructions say it exists)
        wb = openpyxl.Workbook()
        wb.create_sheet("Note_Text")
        wb.create_sheet("Note_Index")
        wb.create_sheet("Span_Annotations")
        wb.create_sheet("Span_Hydrated")
        wb.create_sheet("Event_Log")
        wb.create_sheet("V3_Procedure_Events")
        wb.create_sheet("V3_Registry_JSON")
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # 2. Populate Note_Text
    ws_text = wb["Note_Text"]
    # Check headers
    if ws_text.max_row == 1 and ws_text.cell(1,1).value is None:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 3. Populate Note_Index
    ws_index = wb["Note_Index"]
    # Columns: source_file, note_id, encounter_id, procedure_date, site, reviewer, status, free_text_notes, [flags...]
    flag_keys = list(PROC_FLAGS.keys())
    row_data = [
        SOURCE_FILE, NOTE_ID, "", "", "", "", "Pending", ""
    ] + [PROC_FLAGS[k] for k in flag_keys]
    ws_index.append(row_data)

    # 4. Process Spans -> Annotations & Hydrated
    ws_anno = wb["Span_Annotations"]
    ws_hydra = wb["Span_Hydrated"]
    
    # Ensure headers if empty
    anno_headers = ["source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", "match_index", "start_char", "end_char", "span_len", "label", "normalized_value", "schema_field", "event_id", "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"]
    if ws_anno.max_row == 0 or ws_anno.cell(1,1).value is None:
        ws_anno.append(anno_headers)
        ws_hydra.append(anno_headers)

    span_counter = 1
    for txt, label, norm, evt_id in spans_data:
        # Context/Match Index logic
        ctx = None
        midx = None
        
        # Simple disambiguation for common terms
        if txt == "APC" and evt_id == "evt_06":
            # Just pick first occurrence after 'Modality' table or in table
            midx = 0 
        if txt == "Right Mainstem" and evt_id == "evt_02":
            # "Successful therapeutic aspiration ... Right Mainstem"
            ctx = "Successful therapeutic aspiration"
        
        start, end, status = hydrate_span(NOTE_TEXT, txt, ctx, midx)
        
        row_base = [
            SOURCE_FILE, NOTE_ID, f"span_{span_counter:03d}", "procedure_body", 
            ctx, txt, midx, 
            "", "", f"=LEN(INDIRECT(\"RC[-4]\",0))", # Formulas for annotation sheet
            label, norm, "", evt_id, 
            0, 0, "", "", "", "needs_hydration"
        ]
        ws_anno.append(row_base)
        
        row_hydra = list(row_base)
        row_hydra[7] = start
        row_hydra[8] = end
        row_hydra[9] = end - start if start is not None else 0
        row_hydra[19] = status
        ws_hydra.append(row_hydra)
        
        span_counter += 1

    # 5. Populate Event_Log
    ws_event = wb["Event_Log"]
    if ws_event.max_row == 0 or ws_event.cell(1,1).value is None:
        ws_event.append(["source_file", "note_id", "event_id", "event_type", "method", "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements", "specimens", "findings", "is_historical", "reviewer", "comments", "device_size", "device_material", "outcome_airway_lumen_pre", "outcome_airway_lumen_post", "outcome_symptoms", "outcome_pleural", "outcome_complication"])

    for evt in events_list:
        devs = ", ".join(evt.get("devices", []))
        anat = evt.get("anatomy", "")
        method = evt.get("method", "")
        outcomes = evt.get("outcomes", {})
        
        ws_event.append([
            SOURCE_FILE, NOTE_ID, evt["event_id"], evt["type"], method,
            anat, devs, "", "", "", "", 
            ", ".join(evt.get("specimens", [])),
            evt.get("findings", ""),
            0, "", "", "", "",
            outcomes.get("lumen_pre", ""),
            outcomes.get("lumen_post", ""),
            "", "", outcomes.get("complication", "")
        ])

    # 6. Populate V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    v3_headers = ["note_id", "event_id", "type", "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station", "lesion.type", "lesion.size_mm", "method", "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote", "stent.size", "stent.material_or_brand", "catheter.size_fr", "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post", "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"]
    if ws_v3.max_row == 0 or ws_v3.cell(1,1).value is None:
        ws_v3.append(v3_headers)

    registry_events = []
    
    for evt in events_list:
        # Construct V3 Row
        v3_row = ["" for _ in range(len(v3_headers))]
        v3_row[0] = NOTE_ID
        v3_row[1] = evt["event_id"]
        
        # Mapping Logic
        evt_type = evt["type"]
        method = evt["method"]
        
        v3_row[2] = method # type/method column
        
        # Anatomy logic (simple heuristics)
        anat_str = evt.get("anatomy", "").lower()
        if "right mainstem" in anat_str:
            v3_row[3] = "airway"
            v3_row[4] = "RUL" # RMS often mapped to RUL or just Right Mainstem in specific fields
        elif "lingula" in anat_str:
            v3_row[3] = "airway"
            v3_row[4] = "LUL"
            v3_row[5] = "lingula"
        
        v3_row[9] = method
        v3_row[10] = json.dumps(evt.get("devices", []))
        
        # Outcomes
        outcomes = evt.get("outcomes", {})
        v3_row[18] = outcomes.get("lumen_pre", "")
        v3_row[19] = outcomes.get("lumen_post", "")
        v3_row[22] = outcomes.get("complication", "")

        ws_v3.append(v3_row)
        
        # JSON Construction
        registry_events.append({
            "event_id": evt["event_id"],
            "method": method,
            "target": {"anatomy": evt.get("anatomy")},
            "devices": evt.get("devices", []),
            "outcomes": outcomes
        })

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 0:
        ws_json.append(["schema_version", "note_id", "json_object"])
    
    final_json = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": registry_events,
        "no_immediate_complications": True # derived from text
    }
    
    ws_json.append(["3.0", NOTE_ID, json.dumps(final_json, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Successfully generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()