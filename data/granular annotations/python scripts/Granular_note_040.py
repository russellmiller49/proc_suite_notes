import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_040"
SOURCE_FILE = "note_040.txt"
PROCEDURE_DATE = "2026-01-12" # inferred from context or left generic
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_040 SOURCE_FILE: note_040.txt INDICATION FOR OPERATION:  [REDACTED]is a 68 year old-year-old female who presents with Recurrent Effusion.
The nature, purpose, risks, benefits and alternatives to Chest Ultrasound and Thoracentesis were discussed with the patient in detail.
Patient indicated a wish to proceed with procedure and informed consent was signed.
PREOPERATIVE DIAGNOSIS:  Recurrent Effusion
POSTOPERATIVE DIAGNOSIS: Recurrent Effusion
 
PROCEDURE:  
76604 Ultrasound, chest (includes mediastinum), real time with image documentation
32555 Aspirate pleura with imaging (thoracentesis)
 
50 Bilateral Procedures (Procedure done on both sides of the body) (bilateral chest ultrasound)
 
Local ONLY
PROCEDURE IN DETAIL:
 
PATIENT POSITION: 
0‌ Supine  1‌ Sitting   
0‌ Lateral Decubitus:  0‌ Right 0‌ Left 
 
LEFT CHEST ULTRASOUND FINDINGS:  1‌ Image saved and uploaded to patient's medical record
Hemithorax:   0‌ Right  1‌ Left 
 
Pleural Effusion: 
Volume:       0‌ None  0‌ Minimal  1‌ Small 
 0‌ Moderate  0‌ Large 
Echogenicity:   1‌ Anechoic  0‌ Hypoechoic  0‌ Isoechoic  0‌ Hyperechoic 
Loculations:  1‌ None  0‌Thin  0‌ Thick 
Diaphragmatic Motion:  1‌ Normal  0‌ Diminished  0‌ Absent  
Lung: 
Lung sliding before procedure:   1‌ Present  0‌ Absent 
Lung sliding post procedure:   1‌ Present  0‌ Absent 
Lung consolidation/atelectasis: 0‌ Present  1‌  Absent 
Pleura:  1‌ Normal  0‌ Thick  0‌ Nodular 
 
 
 
RIGHT CHEST ULTRASOUND FINDINGS:  1‌ Image saved and uploaded to patient's medical record 
Hemithorax: 
  1‌ Right  0‌ Left 
 
Pleural Effusion: 
Volume:       0‌ None  0‌ Minimal  0‌ Small  0‌ Moderate  1‌ Large 
Echogenicity:   1‌ Anechoic  0‌ Hypoechoic  0‌ Isoechoic  0‌ Hyperechoic 
Loculations:  1‌ None  0‌Thin  0‌ Thick 
Diaphragmatic Motion:  1‌ Normal  0‌ Diminished  0‌ Absent  
Lung: 
Lung sliding before procedure:   1‌ Present  0‌ Absent 
Lung sliding post procedure:   1‌ Present  0‌ Absent 
Lung consolidation/atelectasis: 1‌ Present  0‌  Absent 
Pleura:  1‌ 
Normal  0‌ Thick  0‌ Nodular 
 
 
 
Based on Ultrasound Evaluation, Thoracentesis is determined to be feasible and will proceed as planned.
.
 
Insertion site prepped and draped in sterile fashion.  Thoracentesis Kit was used.
ANESTHESIA:   Lidocaine 1%: 10 ml      Other: ______ 
Entry Site: 
1‌ Right 8th Intercostal Space   0‌ Left  ___ Intercostal Space 
0‌ Mid-clavicular   1‌ Posterior-axillary  0‌ Mid-scapular  0‌ Other: 
 
Sutured: 0‌ Yes 1‌ No 
 
PROCEDURE FINDINGS: 
Fluid Removed: 900 mL, stopped due to symptoms (coughing, complaint of dry throat)
1‌ Serous  0‌ Serosanguinous 0‌ Bloody  0‌ Chylous 0‌ Other: 
 
Drainage device:   1‌ Pleurovac    0‌ Drainage Bag  0‌ Heimlich Valve  0‌ Pneumostat  0‌ Other: 
Suction: 1‌ No 
0‌Yes, - ___ cmH20 
 
RIGHT CHEST ULTRASOUND S/P THORACENTESIS
 
 
SPECIMEN(S): 
0‌None           1‌PH               1‌ LDH                        1‌Glucose       1‌T.
Protein    1‌Cholesterol
1‌Cell Count   1‌ ADA             0‌Triglycerides            0‌Amylase 
1‌Gram Stain/ Culture            1‌AFB                         1‌Fungal Culture 
1‌Cytology      0‌Flow Cytometry                 
              0‌Other: 
 
CXR ordered: 1‌ Yes 0‌ No 
 
 
COMPLICATIONS:
1‌None 0‌Bleeding-EBL: ___ ml 0‌Pneumothorax 0‌Re- Expansion Pulmonary Edema 
0‌Other: 
 
IMPRESSION/PLAN: [REDACTED]is a 68 year old-year-old female who presents for Chest Ultrasound and Thoracentesis.
The patient tolerated the procedure well.  There were no immediate complications.
--Post procedure CXR
--Follow up pleural studies
--Continued care per primary team
 
 
DISPOSITION: Nursing Unit"""

# -------------------------------------------------------------------------
# CONFIGURATION
# -------------------------------------------------------------------------
# 30 procedure flags
PROCEDURE_FLAGS = [
    "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
    "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy", 
    "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration", 
    "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation", 
    "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", 
    "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
    "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis", 
    "pleural_biopsy", "fibrinolytic_therapy"
]

# Set specific flags for this note
ACTIVE_FLAGS = {
    "thoracentesis": 1
}

# -------------------------------------------------------------------------
# DATA EXTRACTION LOGIC
# -------------------------------------------------------------------------

# Define Spans
# Structure: (text, label, normalized_value, schema_field, event_id, context_prefix)
# event_1 = Left Ultrasound
# event_2 = Right Ultrasound
# event_3 = Right Thoracentesis

raw_spans = [
    # Metadata/Diagnosis
    ("Recurrent Effusion", "OBS_LESION", "Pleural Effusion", "lesion.type", "event_0", "DIAGNOSIS:  "),

    # Event 1: Left Ultrasound
    ("Left", "LATERALITY", "Left", "target.location.side", "event_1", "Hemithorax:   0‌ Right  1‌ "),
    ("Small", "OBS_LESION", "Small", "lesion.size_mm", "event_1", "0‌ Minimal  1‌ "),
    ("Anechoic", "OBS_LESION", "Anechoic", "lesion.type", "event_1", "Echogenicity:   1‌ "),
    
    # Event 2: Right Ultrasound
    ("Right", "LATERALITY", "Right", "target.location.side", "event_2", "Hemithorax: \n  1‌ "),
    ("Large", "OBS_LESION", "Large", "lesion.size_mm", "event_2", "0‌ Moderate  1‌ "),
    ("Anechoic", "OBS_LESION", "Anechoic", "lesion.type", "event_2", "Echogenicity:   1‌ "),
    ("Consolidation/atelectasis: 1‌ Present", "OBS_LESION", "Consolidation/Atelectasis", "lesion.type", "event_2", "Lung "),

    # Event 3: Thoracentesis (Therapeutic)
    ("Thoracentesis", "PROC_METHOD", "Thoracentesis", "method", "event_3", "Evaluation, "),
    ("Right", "LATERALITY", "Right", "target.location.side", "event_3", "Entry Site: \n1‌ "),
    ("8th Intercostal Space", "ANAT_LUNG_LOC", "8th Intercostal Space", "target.location.segment", "event_3", "Right "),
    ("Posterior-axillary", "ANAT_LUNG_LOC", "Posterior-axillary", "target.location.segment", "event_3", "0‌ Mid-clavicular   1‌ "),
    ("900 mL", "MEAS_VOL", "900", "measurements.volume", "event_3", "Fluid Removed: "),
    ("Serous", "OBS_LESION", "Serous", "findings.appearance", "event_3", "1‌ "),
    ("Pleurovac", "DEV_DEVICE", "Pleurovac", "device.brand", "event_3", "Drainage device:   1‌ "),
    ("PH", "PROC_ACTION", "Specimen Sent", "specimens.type", "event_3", "0‌None           1‌"),
    ("LDH", "PROC_ACTION", "Specimen Sent", "specimens.type", "event_3", "1‌PH               1‌ "),
    ("Glucose", "PROC_ACTION", "Specimen Sent", "specimens.type", "event_3", "1‌ LDH                        1‌"),
    ("Protein", "PROC_ACTION", "Specimen Sent", "specimens.type", "event_3", "T.\n"),
    ("Cholesterol", "PROC_ACTION", "Specimen Sent", "specimens.type", "event_3", "Protein    1‌"),
    ("Cell Count", "PROC_ACTION", "Specimen Sent", "specimens.type", "event_3", "1‌"),
    ("ADA", "PROC_ACTION", "Specimen Sent", "specimens.type", "event_3", "1‌Cell Count   1‌ "),
    ("Gram Stain/ Culture", "PROC_ACTION", "Specimen Sent", "specimens.type", "event_3", "0‌Amylase \n1‌"),
    ("AFB", "PROC_ACTION", "Specimen Sent", "specimens.type", "event_3", "Culture            1‌"),
    ("Fungal Culture", "PROC_ACTION", "Specimen Sent", "specimens.type", "event_3", "1‌AFB                         1‌"),
    ("Cytology", "PROC_ACTION", "Specimen Sent", "specimens.type", "event_3", "1‌"),

    # Outcomes
    ("coughing, complaint of dry throat", "OUTCOME_SYMPTOMS", "Cough/Dry throat", "outcomes.symptoms", "event_3", "stopped due to symptoms ("),
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "outcomes.complications", "event_3", "well.  There were "),
]

# -------------------------------------------------------------------------
# HELPER FUNCTIONS
# -------------------------------------------------------------------------

def create_workbook():
    if not os.path.exists(TEMPLATE_PATH):
        # Create a dummy template if strict mode requirements fail (just for safety in this snippet)
        wb = openpyxl.Workbook()
        for name in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            wb.create_sheet(name)
        return wb
    else:
        return openpyxl.load_workbook(TEMPLATE_PATH)

def get_hydration(text, span_text, context_prefix=None):
    """
    Returns (start_char, end_char, hydration_status)
    """
    if not span_text:
        return "", "", "missing_text"
    
    # Strategy 1: Context prefix
    if context_prefix:
        # We search for context prefix, then look for span_text shortly after
        # Normalize spaces for regex search
        pattern = re.escape(context_prefix) + r"[\s\S]{0,150}" + re.escape(span_text)
        match = re.search(pattern, text)
        if match:
            # The span is at the end of the match
            full_match_str = match.group(0)
            span_start = match.start() + full_match_str.rfind(span_text)
            span_end = span_start + len(span_text)
            return span_start, span_end, "hydrated_prefix_window"

    # Strategy 2: Exact count = 1
    count = text.count(span_text)
    if count == 1:
        start = text.find(span_text)
        return start, start + len(span_text), "hydrated_unique"
    
    # Strategy 3: Fallback (Ambiguous)
    if count > 1:
        return "", "", f"ambiguous_count={count}"
    
    return "", "", "not_found"

def format_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    return cell

# -------------------------------------------------------------------------
# GENERATION
# -------------------------------------------------------------------------

class WorkbookGenerator:
    def __init__(self):
        self.wb = create_workbook()
        self.hydrated_spans = []

    def populate_note_text(self):
        ws = self.wb["Note_Text"]
        # Determine next empty row
        row = ws.max_row + 1
        if ws.max_row == 1 and ws.cell(1,1).value is None: row = 1 # Handle empty sheet
        
        ws.cell(row, 1, NOTE_ID)
        ws.cell(row, 2, SOURCE_FILE)
        ws.cell(row, 3, NOTE_TEXT)

    def populate_note_index(self):
        ws = self.wb["Note_Index"]
        row = ws.max_row + 1
        
        # Meta columns
        ws.cell(row, 1, SOURCE_FILE)
        ws.cell(row, 2, NOTE_ID)
        ws.cell(row, 4, PROCEDURE_DATE)
        ws.cell(row, 7, "extracted") # Status
        
        # Flags start at col 9 usually, but let's map headers dynamically if possible or hardcode per spec
        # Assuming template headers match the list order in PROCEDURE_FLAGS
        # We will write to columns 9 through 38 (30 flags)
        
        # Find start of flags. Assuming they start at column 9 based on typical templates
        col_start = 9
        for i, flag in enumerate(PROCEDURE_FLAGS):
            val = ACTIVE_FLAGS.get(flag, 0)
            ws.cell(row, col_start + i, val)

    def populate_spans(self):
        ws_anno = self.wb["Span_Annotations"]
        ws_hydra = self.wb["Span_Hydrated"]
        
        row_anno = ws_anno.max_row + 1
        row_hydra = ws_hydra.max_row + 1
        
        for span in raw_spans:
            text, label, norm, field, event, context = span
            
            # Anchor Logic
            start, end, status = get_hydration(NOTE_TEXT, text, context)
            
            # Common Data
            data = [
                SOURCE_FILE, NOTE_ID, f"span_{row_anno}", "procedure_note",
                context, text, "", # match_index left blank
                "", "", # start/end blank for Annotation
                f"=LEN(F{row_anno})", # formula
                label, norm, field, event,
                "FALSE", "FALSE", "", "system", "", # is_negated, is_hist, time, reviewer, comments
                "needs_hydration"
            ]
            
            # Write to Annotation
            for i, val in enumerate(data, 1):
                ws_anno.cell(row_anno, i, val)
                
            # Write to Hydrated
            hydra_data = list(data)
            hydra_data[7] = start
            hydra_data[8] = end
            hydra_data[19] = status
            
            for i, val in enumerate(hydra_data, 1):
                ws_hydra.cell(row_hydra, i, val)
            
            # Store for downstream
            self.hydrated_spans.append({
                "event_id": event,
                "label": label,
                "norm": norm,
                "text": text,
                "field": field
            })
            
            row_anno += 1
            row_hydra += 1

    def populate_event_log(self):
        ws = self.wb["Event_Log"]
        row = ws.max_row + 1
        
        # Group spans by event
        events = {}
        for s in self.hydrated_spans:
            eid = s["event_id"]
            if eid not in events: events[eid] = []
            events[eid].append(s)
            
        # Define Event Objects
        # Event 1: Left Ultrasound
        ws.cell(row, 1, SOURCE_FILE)
        ws.cell(row, 2, NOTE_ID)
        ws.cell(row, 3, "event_1")
        ws.cell(row, 4, "Diagnostic Ultrasound")
        ws.cell(row, 5, "Ultrasound") # method
        ws.cell(row, 6, "Left Pleura") # anatomy
        ws.cell(row, 13, "Small effusion") # findings
        row += 1
        
        # Event 2: Right Ultrasound
        ws.cell(row, 1, SOURCE_FILE)
        ws.cell(row, 2, NOTE_ID)
        ws.cell(row, 3, "event_2")
        ws.cell(row, 4, "Diagnostic Ultrasound")
        ws.cell(row, 5, "Ultrasound") # method
        ws.cell(row, 6, "Right Pleura") # anatomy
        ws.cell(row, 13, "Large effusion, Consolidation") # findings
        row += 1
        
        # Event 3: Thoracentesis
        ws.cell(row, 1, SOURCE_FILE)
        ws.cell(row, 2, NOTE_ID)
        ws.cell(row, 3, "event_3")
        ws.cell(row, 4, "Therapeutic Thoracentesis")
        ws.cell(row, 5, "Thoracentesis") # method
        ws.cell(row, 6, "Right Pleura (8th ICS)") # anatomy
        ws.cell(row, 7, "Pleurovac") # device
        ws.cell(row, 10, "900 mL") # measurements
        ws.cell(row, 11, "PH, LDH, Glucose, Protein, Cholesterol, Cell Count, ADA, Gram Stain, AFB, Fungal, Cytology") # specimens
        ws.cell(row, 13, "Serous fluid") # findings
        ws.cell(row, 21, "Coughing, dry throat") # symptoms
        ws.cell(row, 23, "None") # complications
        row += 1

    def populate_v3_json(self):
        ws = self.wb["V3_Registry_JSON"]
        row = ws.max_row + 1
        
        # Construct JSON Structure
        data = {
            "schema_version": "v3.0",
            "note_id": NOTE_ID,
            "no_immediate_complications": True,
            "procedures": [
                {
                    "event_id": "event_1",
                    "type": "Diagnostic Ultrasound",
                    "method": "Ultrasound",
                    "target": {"anatomy_type": "Pleura", "location": {"side": "Left"}},
                    "lesion": {"type": "Pleural Effusion", "size_mm": "Small"}
                },
                {
                    "event_id": "event_2",
                    "type": "Diagnostic Ultrasound",
                    "method": "Ultrasound",
                    "target": {"anatomy_type": "Pleura", "location": {"side": "Right"}},
                    "lesion": [{"type": "Pleural Effusion", "size_mm": "Large"}, {"type": "Consolidation/Atelectasis"}]
                },
                {
                    "event_id": "event_3",
                    "type": "Therapeutic Thoracentesis",
                    "method": "Thoracentesis",
                    "target": {"anatomy_type": "Pleura", "location": {"side": "Right", "segment": "8th Intercostal Space, Posterior-axillary"}},
                    "device": {"brand": "Pleurovac"},
                    "measurements": {"volume": 900, "unit": "mL"},
                    "findings": {"appearance": "Serous"},
                    "specimens": ["PH", "LDH", "Glucose", "Protein", "Cholesterol", "Cell Count", "ADA", "Gram Stain/Culture", "AFB", "Fungal Culture", "Cytology"],
                    "outcomes": {
                        "symptoms": "Cough/Dry throat",
                        "complications": "None"
                    }
                }
            ]
        }
        
        ws.cell(row, 1, json.dumps(data, indent=2))

    def save(self):
        self.wb.save(OUTPUT_PATH)
        print(f"Saved to {OUTPUT_PATH}")

    def generate(self):
        self.populate_note_text()
        self.populate_note_index()
        self.populate_spans()
        self.populate_event_log()
        self.populate_v3_json()
        self.save()

# -------------------------------------------------------------------------
# EXECUTION
# -------------------------------------------------------------------------
if __name__ == "__main__":
    generator = WorkbookGenerator()
    generator.generate()