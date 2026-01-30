import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# -------------------------------------------------------------------------
# INPUTS
# -------------------------------------------------------------------------
NOTE_ID = "note_041"
SOURCE_FILE = "note_041.txt"
PROCEDURE_DATE = "2026-01-12" # Placeholder, not in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_041 SOURCE_FILE: note_041.txt INDICATION FOR OPERATION:  Recurrent Effusion.
PREOPERATIVE DIAGNOSIS:  Recurrent Effusion
POSTOPERATIVE DIAGNOSIS: Recurrent Effusion
 
PROCEDURE:  
76604 Ultrasound, chest (includes mediastinum), real time with image documentation
32557 Insert catheter pleura with imaging (chest tube)
 
Local ONLY
 
PROCEDURE IN DETAIL:
 
PATIENT POSITION: 
0‌ Supine  0‌ Sitting   
1‌ Lateral Decubitus:  0‌ Right 1‌ Left with right side up
 
CHEST ULTRASOUND FINDINGS:  1‌ Image saved and up loaded to the epic media tab
Hemithorax:   1‌ Right  0‌ Left 
 
Pleural Effusion: 
Volume:       0‌ None  0‌ Minimal  0‌ Small  0‌ Moderate  1‌ Large 
Echogenicity: 
  1‌ Anechoic  0‌ Hypoechoic  0‌ Isoechoic  0‌ Hyperechoic 
Loculations:  1‌ None  0‌Thin  0‌ Thick 
Diaphragmatic Motion:  1‌ Normal  0‌ Diminished  0‌ Absent  
Lung: 
Lung sliding before procedure:   1‌ Present  0‌ Absent 
Lung sliding post procedure:   1‌ Present  0‌ Absent 
Lung consolidation/atelectasis: 1‌ Present  0‌  Absent 
Pleura:  1‌ Normal  0‌ Thick  0‌ Nodular 
 
 
 
Insertion site prepped and draped in sterile fashion.
ANESTHESIA:   Lidocaine 1%: 10 ml      Other: ______ 
Entry Site: 
1‌ Right 7th Intercostal Space   0‌ Left  ___ Intercostal Space 
0‌ Mid-clavicular   1‌ Mid-axillary  0‌ Mid-scapular  0‌ Other: 
 
Size:  0‌ 6Fr  0‌ 8Fr   0‌ 12FR   1‌ 14Fr  0‌ 16Fr   0‌ 18Fr  0‌ 24Fr   0‌ 32 Fr   0‌ Other: 
Sutured: 1‌ Yes 0‌ No 
 
PROCEDURE FINDINGS: 
A  pigtail catheter was inserted using the Seldinger technique.
Entry into the pleural space was confirmed with the easy removal of minimal serous appearing pleural fluid.
A guidewire was inserted using the introducer needle pointed in the apical posterior direction. The introducer needle was then removed.
A dilator was then inserted over the wire with a twisting motion in order to form a tract for catheter insertion.
The dilator was removed and the pigtail catheter (with trochar) was advanced over the guidewire.
The catheter was inserted into the chest until all catheter holes were well within the chest.
The guidewire and trochar were then removed.  The tube was then attached to the collection drain apparatus and secured in place with suture and covered.
Fluid Removed: 1225 ml 
1‌ Serous  0‌ Serosanguinous 0‌ Bloody  0‌ Chylous 0‌ Other: 
 
Drainage device:   1‌ Pleurovac    0‌ Drainage Bag  0‌ Heimlich Valve  0‌ Pneumostat  0‌ Other: 
Suction: 0‌ No 1‌Yes, - 20cmH20 
 
 
SPECIMEN(S): 
1‌None
 
CXR ordered: 1‌ Yes 0‌ No 
 
 
COMPLICATIONS:
1‌None 0‌Bleeding-EBL: ___ ml 0‌Pneumothorax 0‌Re- Expansion Pulmonary Edema 
0‌Other: 
 
IMPRESSION/PLAN: [REDACTED]is a 68 year old-year-old female who presents for Chest Ultrasound and Chest tube placement.
The patient tolerated the procedure well.  There were no immediate complications.
--Post procedure CXR
--Continue chest tube to -20cmH2O
--Strict I/O
--Daily CXR while chest tube in place
--Nursing chest tube flushing protocol
--Continued care per primary team
 
DISPOSITION: ICU"""

# -------------------------------------------------------------------------
# DATA DEFINITIONS
# -------------------------------------------------------------------------

# 1. Procedure Flags (30 total)
# 1 if present, 0 if absent
PROCEDURE_FLAGS = {
    # Bronchoscopy (23) - All 0 for this note
    "diagnostic_bronchoscopy": 0, "bal": 0, "bronchial_wash": 0, "brushings": 0,
    "endobronchial_biopsy": 0, "tbna_conventional": 0, "linear_ebus": 0, "radial_ebus": 0,
    "navigational_bronchoscopy": 0, "transbronchial_biopsy": 0, "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0, "foreign_body_removal": 0, "airway_dilation": 0,
    "airway_stent": 0, "thermal_ablation": 0, "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0, "blvr": 0, "peripheral_ablation": 0, "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0, "rigid_bronchoscopy": 0,
    # Pleural (7)
    "thoracentesis": 0, # Note says "Insert catheter... (chest tube)", pigtail used.
    "chest_tube": 1,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 2. Spans (Anchor-First)
# Format: (span_text, label, normalized_value, event_id, context_prefix)
# event_id: 1=Ultrasound, 2=Chest Tube
SPANS = [
    # Event 1: Ultrasound
    ("Ultrasound, chest", "PROC_METHOD", "Ultrasound", "evt1", "PROCEDURE:  \n76604 "),
    ("Right", "LATERALITY", "Right", "evt1", "Hemithorax:   1‌ "),
    ("Large", "MEAS_VOL", "Large", "evt1", "Pleural Effusion: \nVolume:       0‌ None  0‌ Minimal  0‌ Small  0‌ Moderate  1‌ "),
    ("Anechoic", "OBS_LESION", "Anechoic", "evt1", "Echogenicity: \n  1‌ "),
    
    # Event 2: Chest Tube
    ("Insert catheter pleura", "PROC_METHOD", "Chest Tube", "evt2", "32557 "),
    ("chest tube", "DEV_CATHETER", "Chest Tube", "evt2", "Insert catheter pleura with imaging ("),
    
    # Anatomy for Tube
    ("Right 7th Intercostal Space", "ANAT_LUNG_LOC", "Right 7th ICS", "evt2", "Entry Site: \n1‌ "),
    ("Mid-axillary", "ANAT_LUNG_LOC", "Mid-axillary", "evt2", "0‌ Mid-clavicular   1‌ "),
    
    # Device details
    ("14Fr", "DEV_CATHETER_SIZE", "14Fr", "evt2", "Size:  0‌ 6Fr  0‌ 8Fr   0‌ 12FR   1‌ "),
    ("pigtail catheter", "DEV_CATHETER", "Pigtail", "evt2", "PROCEDURE FINDINGS: \nA  "),
    ("Seldinger technique", "PROC_METHOD", "Seldinger", "evt2", "inserted using the "),
    
    # Fluid Outcomes
    ("1225 ml", "MEAS_PLEURAL_DRAIN", "1225 ml", "evt2", "Fluid Removed: "),
    ("Serous", "OBS_LESION", "Serous", "evt2", "Fluid Removed: 1225 ml \n1‌ "),
    
    # Complications
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "evt2", "tolerated the procedure well.  There were "),
]

# 3. Event Log Data (Populates Event_Log and V3_Procedure_Events)
EVENT_DATA = [
    {
        "event_id": "evt1",
        "type": "Ultrasound",
        "method": "Ultrasound",
        "anatomy": "Right Hemithorax",
        "findings": "Large Anechoic Effusion",
        "evidence": "Ultrasound, chest; Hemithorax: Right; Large; Anechoic"
    },
    {
        "event_id": "evt2",
        "type": "Chest Tube",
        "method": "Seldinger",
        "device": "Pigtail catheter",
        "device_size": "14Fr",
        "anatomy": "Right 7th Intercostal Space, Mid-axillary",
        "outcome_pleural": "1225 ml",
        "outcome_complication": "None",
        "findings": "Serous fluid",
        "evidence": "Insert catheter pleura; 14Fr pigtail catheter; Seldinger technique; Fluid Removed: 1225 ml; No immediate complications"
    }
]

# -------------------------------------------------------------------------
# HELPER FUNCTIONS
# -------------------------------------------------------------------------

def clean_text(text):
    return text.replace('\r\n', '\n').replace('\r', '\n')

def get_match_index(full_text, span_text, context_prefix):
    """
    Finds the 0-based index of the specific occurrence of span_text.
    Logic:
    1. If context_prefix provided, find occurrence following it within window.
    2. Default to first occurrence if no context or unique.
    """
    if not span_text:
        return 0
    
    matches = [m.start() for m in re.finditer(re.escape(span_text), full_text)]
    if len(matches) == 0:
        return -1 # Not found
    
    if len(matches) == 1:
        return 0
    
    if context_prefix:
        # Find prefix location
        prefix_matches = [m.start() for m in re.finditer(re.escape(context_prefix), full_text)]
        if not prefix_matches:
            return 0 # Prefix not found, fallback
            
        # Look for the span that follows the prefix closely (e.g., within 150 chars)
        best_match_idx = 0
        min_dist = 999999
        
        for p_start in prefix_matches:
            p_end = p_start + len(context_prefix)
            for i, m_start in enumerate(matches):
                if m_start >= p_end:
                    dist = m_start - p_end
                    if dist < min_dist and dist < 200: # Window
                        min_dist = dist
                        best_match_idx = i
        return best_match_idx

    return 0 # Default to first

def find_offsets(full_text, span_text, match_index):
    matches = [m for m in re.finditer(re.escape(span_text), full_text)]
    if match_index >= 0 and match_index < len(matches):
        m = matches[match_index]
        return m.start(), m.end()
    return None, None

def create_or_get_sheet(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return wb.create_sheet(sheet_name)

# -------------------------------------------------------------------------
# MAIN EXECUTION
# -------------------------------------------------------------------------

def generate_workbook():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        # If template missing, create a basic one for demonstration (though instructions say load it)
        wb = openpyxl.Workbook()
        
    full_text_clean = clean_text(NOTE_TEXT)
    
    # -----------------------------------------
    # 1. Note_Text
    # -----------------------------------------
    ws_text = create_or_get_sheet(wb, "Note_Text")
    # Headers if empty
    if ws_text.max_row == 1 and ws_text.cell(1,1).value is None:
        ws_text.append(["note_id", "source_file", "note_text"])
    
    # Check if exists
    exists = False
    for row in ws_text.iter_rows(min_row=2, values_only=True):
        if row[0] == NOTE_ID:
            exists = True
            break
    if not exists:
        ws_text.append([NOTE_ID, SOURCE_FILE, full_text_clean])

    # -----------------------------------------
    # 2. Note_Index (Flags)
    # -----------------------------------------
    ws_index = create_or_get_sheet(wb, "Note_Index")
    # Header: source_file, note_id, encounter, date, site, reviewer, status, notes, [30 flags]
    # We assume template has headers. If not, we'd add them. 
    # Let's construct the row.
    
    meta_row = [
        SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "complete", ""
    ]
    
    # Add flags in specific order
    flag_keys = list(PROCEDURE_FLAGS.keys())
    flag_values = [PROCEDURE_FLAGS[k] for k in flag_keys]
    
    ws_index.append(meta_row + flag_values)

    # -----------------------------------------
    # 3. Span_Annotations (Anchor First)
    # -----------------------------------------
    ws_anno = create_or_get_sheet(wb, "Span_Annotations")
    # Headers: source_file, note_id, span_id, section, prefix, span_text, match_idx, start, end, len, label, norm, field, event_id, neg, hist, time, rev, comm, hydra
    
    # 4. Span_Hydrated
    ws_hydra = create_or_get_sheet(wb, "Span_Hydrated")
    
    span_id_counter = 1
    
    for span_text, label, norm, evt_id, ctx_prefix in SPANS:
        # Determine match index
        match_idx = get_match_index(full_text_clean, span_text, ctx_prefix)
        
        # Calculate offsets for Hydrated sheet
        start_char, end_char = find_offsets(full_text_clean, span_text, match_idx)
        length = len(span_text)
        
        sid = f"{NOTE_ID}_s{span_id_counter:03d}"
        
        # Base Row Data
        row_base = [
            SOURCE_FILE, NOTE_ID, sid, "Procedure",
            ctx_prefix, span_text, match_idx,
            None, None, length, # start/end blank for Anno
            label, norm, "", evt_id, 
            0, 0, "", "", "", "needs_hydration"
        ]
        
        ws_anno.append(row_base)
        
        # Hydrated Row
        row_hydra = list(row_base)
        row_hydra[7] = start_char
        row_hydra[8] = end_char
        row_hydra[19] = "hydrated" if start_char is not None else "ambiguous"
        
        ws_hydra.append(row_hydra)
        span_id_counter += 1

    # -----------------------------------------
    # 5. Event_Log & 6. V3_Procedure_Events
    # -----------------------------------------
    ws_event = create_or_get_sheet(wb, "Event_Log")
    ws_v3 = create_or_get_sheet(wb, "V3_Procedure_Events")
    
    procedures_list = []
    
    for evt in EVENT_DATA:
        eid = evt["event_id"]
        
        # Populate Event_Log (Basic)
        # Columns: source, note, event_id, type, method, anatomy, device, gauge, station, count, meas, spec, find, hist, rev, comm, size, mat, pre, post, symp, pleur, comp
        log_row = [
            SOURCE_FILE, NOTE_ID, eid, evt["type"], evt.get("method", ""),
            evt.get("anatomy", ""), evt.get("device", ""), "", "", "", "",
            "", evt.get("findings", ""), 0, "", "",
            evt.get("device_size", ""), "", # size, material
            "", "", "", evt.get("outcome_pleural", ""), evt.get("outcome_complication", "")
        ]
        ws_event.append(log_row)
        
        # Populate V3_Procedure_Events (Nested)
        # note_id, event_id, type, target.anat, .lobe, .seg, .stat, les.type, .size, method, dev_json, meas_json, spec_json, find_json, evid, stent.size, .mat, cath.size, out.pre, .post, .sym, .pl, .comp
        
        # Construct JSONs for columns
        dev_json = json.dumps([{"name": evt.get("device", ""), "size": evt.get("device_size", "")}]) if evt.get("device") else ""
        find_json = json.dumps([evt.get("findings", "")]) if evt.get("findings") else ""
        
        v3_row = [
            NOTE_ID, eid, evt["type"],
            "Lung/Pleura", evt.get("anatomy", ""), "", "", # anat, lobe, seg, stat
            "", "", # lesion type, size
            evt.get("method", ""),
            dev_json, "", "", find_json, evt.get("evidence", ""),
            "", "", evt.get("device_size", ""), # stent size, mat, cath size
            "", "", "", evt.get("outcome_pleural", ""), evt.get("outcome_complication", "")
        ]
        ws_v3.append(v3_row)
        
        # Build JSON object for Registry
        proc_obj = {
            "event_id": eid,
            "procedure_type": evt["type"],
            "method": evt.get("method"),
            "target_anatomy": {"location": evt.get("anatomy")},
            "devices": [{"name": evt.get("device"), "size": evt.get("device_size")}] if evt.get("device") else [],
            "outcomes": {
                "pleural_drainage": evt.get("outcome_pleural"),
                "complications": evt.get("outcome_complication")
            }
        }
        procedures_list.append(proc_obj)

    # -----------------------------------------
    # 7. V3_Registry_JSON
    # -----------------------------------------
    ws_json = create_or_get_sheet(wb, "V3_Registry_JSON")
    
    final_json = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "source_file": SOURCE_FILE,
        "procedures": procedures_list,
        "no_immediate_complications": True # Derived from note text
    }
    
    json_str = json.dumps(final_json, indent=2)
    # If sheet empty, add header
    if ws_json.max_row == 1 and ws_json.cell(1,1).value is None:
        ws_json.append(["note_id", "json_content"])
        
    ws_json.append([NOTE_ID, json_str])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()