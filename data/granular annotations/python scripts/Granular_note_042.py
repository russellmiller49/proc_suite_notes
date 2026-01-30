import openpyxl
from openpyxl.utils import get_column_letter
import re
import os
import json
from datetime import datetime

# ==========================================
# INPUT DATA
# ==========================================
NOTE_ID = "note_042"
SOURCE_FILE = "note_042.txt"
PROCEDURE_DATE = "2025-12-22" # Inferred from "Anticipate suture removal in 7 days (~12/29/2025)" -> 29-7 = 22. Also "12/22/2025" mentioned for PEG usage.
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_042 SOURCE_FILE: note_042.txt INDICATION FOR OPERATION:  68 year old-year-old female with respiratory failure.
PREOPERATIVE DIAGNOSIS: J96.90 Respiratory Failure
 
POSTOPERATIVE DIAGNOSIS:  J96.90 Respiratory Failure
 
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31624 Dx bronchoscope/lavage (BAL)    
31600 Incision of windpipe (perc trach)
43246 Esophagogastroduodenoscopy, flexible, transoral;
with directed placement of percutaneous gastrostomy tube
76536 Ultrasound of Neck
 
ANESTHESIA: 
99152 Moderate sedation: initial 15 minutes
99153 Moderate sedation: each additional 15 minutes 
 
Procedure performed under moderate sedation.
The following medications were provided:
Versed             4 mg
Fentanyl          100 mcg
Etomidate20mg
Rocuronium43mg
 
Physician/patient face-to-face anesthesia start time:   1513
 
Physician/patient face-to-face anesthesia stop time:   1700
 
Total moderate sedation time was 107 minutes.
Patient was monitored continuously one-to-one throughout the entire procedure by the attending physician while anesthesia was administered.
Sedation was administered by ICU RN. 
 
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
 
ESTIMATED BLOOD LOSS:   Minimum
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: Supine
 
Initial Airway Inspection Findings:
 
The endotracheal tube is in good position.
Pharynx: Not assessed due to bronchoscopy introduction through ETT.
Larynx: Not assessed due to bronchoscopy introduction through ETT.
Vocal Cords: Not assessed due to bronchoscopy introduction through ETT.
Trachea: Distal 1/3 normal.
Main Carina: Sharp
Right Lung Proximal Airways: Normal anatomic branching to segmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Left Lung Proximal Airways: Normal anatomic branching to segmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Mucosa: Normal.
Secretions: Copious thick and thin light-yellow mucus/secretions throughout.
Successful therapeutic aspiration was performed to clean out the Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus.
Bronchial alveolar lavage was performed at Lateral Segment of RML (RB4) and Medial Segment of RML (RB5).
Instilled 40 cc of NS, suction returned with 25 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
 
===========================================
 
NECK ULTRASOUND OF PROPOSED TRACHEOSTOMY INSERTION SITE
 
 
The bronchoscope was retracted into the ETT tube and the ET tube retracted into the subglottic space under direct visualization.
The inferior border of the cricoid along with the proximal tracheal rings were visualized.
Next, the anterior neck was prepped and draped in the usual sterile fashion.
Lidocaine 1% 3ml ml was injected into the anterior neck.
A 1 cm incision was made horizontallywith a #10 blade down through the subcutaneous tissue, just inferior to the cricoid cartilage.
The introducer needle and passed between the 1st and 2nd tracheal rings and into the trachea under direct visualization.
Next, a J-wire was passed through the catheter, also visualized with the bronchoscope.
The site was then dilated using the 14Fr introducing dilator passed over the wire.
The 14 Fr dilator was then removed from the guide wire and an 8 Fr guiding catheter placed over the guide wire until the safety ridge on the guiding catheter was at skin level.
The tissue dilator was placed over the guiding catheter until the positioning mark was visualized via the bronchoscope.
The tissue dilator was then removed leaving the guiding catheter and guide wire assembly in place, all under direct visualization bronchoscopically.
Finally a Portex 7.0mm cuffed tracheostomy tube with appropriate dilator was introduced over the guiding catheter into the trachea under direct visualization.
The dilator, guiding catheter, and J-wire were then removed and the tracheostomy tube left in place.
This was confirmed to be in good position bronchoscopically.  The Endotracheal tube was then removed and the ventilator connected to the tracheostomy tube.
Surgicel was placed preemptively around the tracheostomy site to reduce bleeding.
A Lyofoam drain sponge was placed under the tracheostomy tube prior to suturing into place.
The patient tolerated the procedure well.  There were no complications. The staff physician was present throughout the entire procedure.
===========================================
 
Under sterile prep and draping, the abdomen was evaluated and 2 cm below the costal margin on the left, the likely point of insertion was identified.
Ultrasound was used and hepatic tissue was identified just under the standard/traditional PEG insertion site.
TRADITIONAL INSERTION SITE
 
 
As such, a new more lateral insertion site was chosen and no hepatic tissue was identified on ultrasound.
PROPOSED INSERTION SITE
 
 
The scope was introduced through the mouth without difficulty and with continuous insufflation, the stomach was reached within 10 seconds.
Then, the point of digital pressure was identified in the stomach and transillumination was accomplished successfully and without any doubts.
At this point, a 1cm incision was carried out in the skin and a 14ga angiocath was introduced.
Using modified Seldinger technique, the angiocath was advanced and a wire was introduced.
Using pull through technique, the guide wire was pulled through the mouth and linked to the 20Fr PEG catheter in usual fashion.
Thereafter, the wire was pulled through the abdominal wall and the PEG tube positioned correctly with the bumper at 1.5 cm.
The remaining air was suctioned out and complete apposition of the stomach and esophagus was seen. There were no complications.
The total procedural time was 20 minutes.
 
INSTRUCTIONS:
PEG can be used for medications 6 hrs post procedure (@2200)
PEG (or NG tube) can be used for feeds 6 hrs post procedure (@2200).
No enteral feeding prior.
OK to restart systemic anticoagulation @2200, if needed.
 
The patient tolerated the procedure well.
There were no immediate complications.  At the conclusion of the operation, the patient was in stable condition.
SPECIMEN(S): 
--RML BAL (cell count, cyto, micro)
 
IMPRESSION/PLAN: [REDACTED]is a 68 year old-year-old female who presents for tracheostomy and PEG tube placement.
--Post-procedure CXR
--Anticipate suture removal in 7 days (~12/29/2025)
--Anticipate trach change in 10 days (~1/1/2026)
--PEG can be used for medications and tube feeds at 2200 on 12/22/2025
--If required to place gauze underneath the skin bumper, please use only 1 thin layer and change as needed, as this can lead to tension on the gastric cuff and result in complication.
--Please call the Interventional pulmonary fellow on call should there be any issues with the PEG."""

# ==========================================
# 1. FLAG EXTRACTION
# ==========================================
def extract_flags(text):
    text_lower = text.lower()
    flags = {
        'diagnostic_bronchoscopy': 0, 'bal': 0, 'bronchial_wash': 0, 'brushings': 0,
        'endobronchial_biopsy': 0, 'tbna_conventional': 0, 'linear_ebus': 0, 'radial_ebus': 0,
        'navigational_bronchoscopy': 0, 'transbronchial_biopsy': 0, 'transbronchial_cryobiopsy': 0,
        'therapeutic_aspiration': 0, 'foreign_body_removal': 0, 'airway_dilation': 0, 'airway_stent': 0,
        'thermal_ablation': 0, 'tumor_debulking_non_thermal': 0, 'cryotherapy': 0, 'blvr': 0,
        'peripheral_ablation': 0, 'bronchial_thermoplasty': 0, 'whole_lung_lavage': 0, 'rigid_bronchoscopy': 0,
        'thoracentesis': 0, 'chest_tube': 0, 'ipc': 0, 'medical_thoracoscopy': 0, 'pleurodesis': 0,
        'pleural_biopsy': 0, 'fibrinolytic_therapy': 0
    }
    
    # Logic based on Note content
    if "therapeutic aspiration" in text_lower or "31645" in text:
        flags['therapeutic_aspiration'] = 1
    
    if "bal" in text_lower or "bronchial alveolar lavage" in text_lower or "31624" in text:
        flags['bal'] = 1
        
    # Tracheostomy logic (perc trach) - Often involves dilation, but usually mapped to specific trach flag if available.
    # Registry flags provided don't have "tracheostomy". 
    # Will stick to explicit flags found in text.
    
    return flags

# ==========================================
# 2. SPAN EXTRACTION (Anchor-First)
# ==========================================
def extract_spans(text):
    spans = []
    
    def add_span(text_match, label, norm_val, field, event_id, context_prefix="", is_negated=False):
        if not text_match: return
        spans.append({
            'span_text': text_match,
            'label': label,
            'normalized_value': norm_val,
            'schema_field': field,
            'event_id': event_id,
            'context_prefix': context_prefix,
            'is_negated': is_negated
        })

    # Event 1: Therapeutic Aspiration
    e1_id = "event_01"
    add_span("Therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", "procedure.method", e1_id)
    add_span("Copious thick and thin light-yellow mucus/secretions", "OBS_LESION", "mucus", "findings", e1_id)
    add_span("Successful therapeutic aspiration was performed", "PROC_METHOD", "Therapeutic Aspiration", "procedure.method", e1_id)
    
    # Locations for Aspiration
    locs = [
        ("Trachea (Distal 1/3)", "Trachea"),
        ("Right Mainstem", "RMS"),
        ("Bronchus Intermedius", "Bronchus Intermedius"),
        ("Left Mainstem", "LMS"),
        ("Carina", "Carina"),
        ("RUL Carina", "RUL Carina"),
        ("RML Carina", "RML Carina"),
        ("LUL Lingula Carina", "LUL Lingula"),
        ("Left Carina", "Left Carina") # Usually means LMS/LUL junction or secondary carina
    ]
    for txt, norm in locs:
        add_span(txt, "ANAT_AIRWAY", norm, "target.anatomy", e1_id)
    
    add_span("mucus", "OBS_LESION", "mucus", "findings", e1_id, context_prefix="and Left Carina (LC2) from ")

    # Event 2: BAL
    e2_id = "event_02"
    add_span("Bronchial alveolar lavage was performed", "PROC_METHOD", "BAL", "procedure.method", e2_id)
    add_span("Lateral Segment of RML", "ANAT_LUNG_LOC", "RML Lateral", "target.location", e2_id)
    add_span("Medial Segment of RML", "ANAT_LUNG_LOC", "RML Medial", "target.location", e2_id)
    add_span("Instilled 40 cc", "MEAS_VOL", "40", "measurements.instilled", e2_id)
    add_span("returned with 25 cc", "MEAS_VOL", "25", "measurements.return", e2_id)
    add_span("Cell Count", "PROC_ACTION", "Cell Count", "specimens", e2_id)
    add_span("Microbiology", "PROC_ACTION", "Microbiology", "specimens", e2_id)
    add_span("Cytology", "PROC_ACTION", "Cytology", "specimens", e2_id)

    # Event 3: Tracheostomy (Percutaneous)
    e3_id = "event_03"
    add_span("Incision of windpipe", "PROC_METHOD", "Tracheostomy", "procedure.method", e3_id)
    add_span("perc trach", "PROC_METHOD", "Percutaneous Tracheostomy", "procedure.method", e3_id)
    add_span("NECK ULTRASOUND", "PROC_METHOD", "Ultrasound Guidance", "procedure.method", e3_id)
    
    # Devices / Steps
    add_span("1 cm incision", "PROC_ACTION", "Incision", "procedure.action", e3_id)
    add_span("introducer needle", "DEV_NEEDLE", "Introducer Needle", "device.name", e3_id)
    add_span("J-wire", "DEV_INSTRUMENT", "J-wire", "device.name", e3_id)
    add_span("14Fr introducing dilator", "DEV_INSTRUMENT", "Dilator", "device.name", e3_id)
    add_span("14Fr", "DEV_CATHETER_SIZE", "14Fr", "device.size", e3_id, context_prefix="dilated using the ")
    add_span("Portex 7.0mm cuffed tracheostomy tube", "DEV_CATHETER", "Tracheostomy Tube", "device.name", e3_id)
    add_span("7.0mm", "DEV_CATHETER_SIZE", "7.0mm", "device.size", e3_id, context_prefix="Portex ")
    add_span("Surgicel", "DEV_INSTRUMENT", "Surgicel", "device.name", e3_id)
    add_span("Lyofoam drain sponge", "DEV_INSTRUMENT", "Lyofoam drain sponge", "device.name", e3_id)
    
    # Locations
    add_span("trachea", "ANAT_AIRWAY", "Trachea", "target.anatomy", e3_id, context_prefix="catheter into the ")
    
    # Event 4: PEG
    e4_id = "event_04"
    add_span("placement of percutaneous gastrostomy tube", "PROC_METHOD", "PEG Placement", "procedure.method", e4_id)
    add_span("20Fr PEG catheter", "DEV_CATHETER", "PEG Catheter", "device.name", e4_id)
    add_span("20Fr", "DEV_CATHETER_SIZE", "20Fr", "device.size", e4_id, context_prefix="linked to the ")

    # Outcomes (Global or per event - usually global in note, mapped to main events)
    add_span("No immediate complications", "OUTCOME_COMPLICATION", "None", "outcomes.complication", e1_id, is_negated=True)
    add_span("No immediate complications", "OUTCOME_COMPLICATION", "None", "outcomes.complication", e2_id, is_negated=True)
    add_span("No immediate complications", "OUTCOME_COMPLICATION", "None", "outcomes.complication", e3_id, is_negated=True)

    return spans

# ==========================================
# 3. HYDRATION & PROCESSING
# ==========================================
def hydrate_spans(spans, full_text):
    hydrated = []
    for s in spans:
        txt = s['span_text']
        prefix = s.get('context_prefix', '')
        
        start = -1
        count = full_text.count(txt)
        
        if count == 1:
            start = full_text.find(txt)
        elif count > 1 and prefix:
            # Find all occurrences
            occurrences = [m.start() for m in re.finditer(re.escape(txt), full_text)]
            # Check prefix
            for occ in occurrences:
                # Look back 120 chars
                window_start = max(0, occ - 120)
                window = full_text[window_start:occ]
                if prefix in window:
                    start = occ
                    break
            # Fallback: if matches not unique by prefix, pick first or ambiguous
            if start == -1 and occurrences:
                start = occurrences[0] # Fallback to first
        elif count > 1:
            # Ambiguous without prefix, take first for now or mark ambiguous
            start = full_text.find(txt)
            
        if start != -1:
            end = start + len(txt)
            s['start_char'] = start
            s['end_char'] = end
            s['hydration_status'] = 'hydrated'
        else:
            s['start_char'] = ''
            s['end_char'] = ''
            s['hydration_status'] = 'missing'
            
        hydrated.append(s)
    return hydrated

def generate_excel():
    # Load template
    if not os.path.exists(TEMPLATE_PATH):
        # Create a dummy one if missing (for standalone testing validity), 
        # but in production this should fail or use existing.
        # We assume it exists as per instructions.
        wb = openpyxl.Workbook()
        wb.create_sheet("Note_Text")
        wb.create_sheet("Note_Index")
        wb.create_sheet("Span_Annotations")
        wb.create_sheet("Span_Hydrated")
        wb.create_sheet("Event_Log")
        wb.create_sheet("V3_Procedure_Events")
        wb.create_sheet("V3_Registry_JSON")
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # 1. Note_Text
    ws_text = wb["Note_Text"]
    # Check headers
    if ws_text.max_row == 1 and ws_text.cell(1,1).value is None:
         ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws_index = wb["Note_Index"]
    # Headers: source_file, note_id, encounter_id, procedure_date, site, reviewer, status, free_text_notes, [flags...]
    # We need to map flags to columns. Assuming columns 9+ are flags.
    # Let's just append a row with dictionary mapping if headers exist, or creating them.
    # For simplicity in this script, we assume the template has correct headers.
    # We will read headers to map flags.
    headers = [c.value for c in ws_index[1]]
    row_data = [None] * len(headers)
    
    # Metadata
    meta_map = {
        "source_file": SOURCE_FILE,
        "note_id": NOTE_ID,
        "procedure_date": PROCEDURE_DATE,
        "status": "Auto-Generated"
    }
    
    flags = extract_flags(NOTE_TEXT)
    
    for i, h in enumerate(headers):
        if h in meta_map:
            row_data[i] = meta_map[h]
        elif h in flags:
            row_data[i] = flags[h]
        elif h == "reviewer":
            row_data[i] = "Phase0_Bot"
    
    # Ensure mandatory 0s for missing flags in dict
    for f in flags:
        if f in headers:
            idx = headers.index(f)
            if row_data[idx] is None:
                row_data[idx] = flags[f]
                
    ws_index.append(row_data)

    # 3. Span_Annotations
    ws_span = wb["Span_Annotations"]
    raw_spans = extract_spans(NOTE_TEXT)
    
    span_headers = ["source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", "match_index", "start_char", "end_char", "span_len", "label", "normalized_value", "schema_field", "event_id", "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"]
    
    # Write Header if empty
    if ws_span.max_row == 1 and ws_span.cell(1,1).value is None:
        ws_span.append(span_headers)
        
    for i, s in enumerate(raw_spans):
        row = [
            SOURCE_FILE, NOTE_ID, f"span_{i+1:03d}", "Procedure", s.get('context_prefix'), s['span_text'],
            None, None, None, len(s['span_text']),
            s['label'], s['normalized_value'], s['schema_field'], s['event_id'],
            s['is_negated'], False, None, "Phase0_Bot", None, "needs_hydration"
        ]
        ws_span.append(row)

    # 4. Span_Hydrated
    ws_hydrated = wb["Span_Hydrated"]
    if ws_hydrated.max_row == 1 and ws_hydrated.cell(1,1).value is None:
        ws_hydrated.append(span_headers)
        
    hydrated_spans_list = hydrate_spans(raw_spans, NOTE_TEXT)
    for i, s in enumerate(hydrated_spans_list):
        row = [
            SOURCE_FILE, NOTE_ID, f"span_{i+1:03d}", "Procedure", s.get('context_prefix'), s['span_text'],
            None, s['start_char'], s['end_char'], len(s['span_text']),
            s['label'], s['normalized_value'], s['schema_field'], s['event_id'],
            s['is_negated'], False, None, "Phase0_Bot", None, s['hydration_status']
        ]
        ws_hydrated.append(row)

    # 5. Event_Log
    ws_event = wb["Event_Log"]
    if ws_event.max_row == 1 and ws_event.cell(1,1).value is None:
        ws_event.append(["source_file", "note_id", "event_id", "event_type", "method", "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements", "specimens", "findings", "is_historical", "reviewer", "comments", "device_size", "device_material", "outcome_airway_lumen_pre", "outcome_airway_lumen_post", "outcome_symptoms", "outcome_pleural", "outcome_complication"])

    # Define events manually based on extraction
    events_data = {
        "event_01": {
            "type": "Therapeutic Aspiration", "method": "Therapeutic Aspiration",
            "anatomy": "Trachea, RMS, BI, LMS, Carina, RUL, RML, LUL", "findings": "Mucus",
            "complication": "None"
        },
        "event_02": {
            "type": "BAL", "method": "BAL", "anatomy": "RML (Lateral, Medial)",
            "measurements": "Instilled: 40cc, Returned: 25cc", "specimens": "Cell Count, Micro, Cyto",
            "complication": "None"
        },
        "event_03": {
            "type": "Tracheostomy", "method": "Percutaneous Tracheostomy", "anatomy": "Trachea",
            "device": "Tracheostomy Tube, J-wire, Dilator", "device_size": "7.0mm, 14Fr",
            "complication": "None"
        },
        "event_04": {
            "type": "PEG", "method": "PEG Placement", "anatomy": "Stomach/Abdomen",
            "device": "PEG Catheter", "device_size": "20Fr"
        }
    }
    
    for eid, dat in events_data.items():
        ws_event.append([
            SOURCE_FILE, NOTE_ID, eid, dat.get("type"), dat.get("method"),
            dat.get("anatomy"), dat.get("device"), None, None, None,
            dat.get("measurements"), dat.get("specimens"), dat.get("findings"),
            False, "Phase0_Bot", None, dat.get("device_size"), None,
            None, None, None, None, dat.get("complication")
        ])

    # 6. V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    if ws_v3.max_row == 1 and ws_v3.cell(1,1).value is None:
        ws_v3.append(["note_id", "event_id", "type", "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station", "lesion.type", "lesion.size_mm", "method", "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote", "stent.size", "stent.material_or_brand", "catheter.size_fr", "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post", "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"])

    # Map events to V3 rows
    # Event 1
    ws_v3.append([NOTE_ID, "event_01", "Therapeutic Aspiration", "Airway", "Multi", None, None, None, None, "Therapeutic Aspiration", None, None, None, json.dumps(["Mucus"]), None, None, None, None, None, None, None, None, "None"])
    # Event 2
    ws_v3.append([NOTE_ID, "event_02", "BAL", "Airway", "RML", "Lateral/Medial", None, None, None, "BAL", None, json.dumps({"instilled": 40, "return": 25}), json.dumps(["Cell Count", "Micro", "Cyto"]), None, None, None, None, None, None, None, None, None, "None"])
    # Event 3
    ws_v3.append([NOTE_ID, "event_03", "Tracheostomy", "Airway", "Trachea", None, None, None, None, "Percutaneous Tracheostomy", json.dumps(["Trach Tube 7.0mm", "Dilator 14Fr"]), None, None, None, None, None, None, "7.0mm", None, None, None, None, "None"])
    # Event 4
    ws_v3.append([NOTE_ID, "event_04", "PEG", "GI", None, None, None, None, None, "PEG Placement", json.dumps(["PEG Catheter 20Fr"]), None, None, None, None, None, None, "20Fr", None, None, None, None, None])

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 1:
        ws_json.append(["json_output"])
        
    registry_data = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "procedures": [events_data[k] for k in events_data],
        "no_immediate_complications": True
    }
    ws_json.append([json.dumps(registry_data, indent=2)])

    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_excel()