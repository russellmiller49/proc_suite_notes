import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# -------------------------------------------------------------------------
# 1. IMPORTS & CONSTANTS
# -------------------------------------------------------------------------
NOTE_ID = "note_043"
SOURCE_FILE = "note_043.txt"
PROCEDURE_DATE = ""  # Not explicitly defined for the current service date in header, only tube insertion date.
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_043 SOURCE_FILE: note_043.txt INDICATION FOR OPERATION:  [REDACTED]is a 67 year old-year-old male who presents with Complicated Effusion.
The nature, purpose, risks, benefits and alternatives to Chest Ultrasound and Instillation of agents for fibrinolysis (subsequent) were discussed with the patient in detail.
Patient indicated a wish to proceed with procedure and informed consent was signed.
PREOPERATIVE DIAGNOSIS:  Complicated Effusion
POSTOPERATIVE DIAGNOSIS: Same as preoperative diagnosis - see above.
PROCEDURE:  
76604 Ultrasound, chest (includes mediastinum), real time with image documentation
32562 Instillation(s), via chest tube/catheter, agent for fibrinolysis (eg, fibrinolytic agent for break up of multiloculated effusion);
subsequent day
 
PROCEDURE IN DETAIL:
 
PATIENT POSITION: 
0‌ Supine  1‌ Sitting   
0‌ Lateral Decubitus:  0‌ Right 0‌ Left 
 
CHEST ULTRASOUND FINDINGS:  1‌ Image saved and uploaded to patient's medical record 
Hemithorax:   0‌ Right  1‌ Left 
 
Pleural Effusion: 
Volume:       0‌ None  0‌ Minimal  1‌ Small  0‌ Moderate  0‌ Large 
Echogenicity:   1‌ Anechoic  0‌ Hypoechoic  0‌ Isoechoic  0‌ Hyperechoic 
Loculations:  0‌ None  1‌Thin  1‌ Thick 
Diaphragmatic Motion:  1‌ Normal  0‌ Diminished  
0‌ Absent  
Lung: 
Lung sliding before procedure:   0‌ Present  1‌ Absent 
Lung sliding post procedure:   0‌ Present  1‌ Absent 
Lung consolidation/atelectasis: 1‌ Present  0‌  Absent 
Pleura:  0‌ Normal  1‌ Thick  0‌ Nodular 
 
 
 
Date of chest tube insertion: 12/22/2025
 
Side: left
1‌  5 mg/5 mg tPA/Dnasedose #: 2
            0‌  ___mg tPA                        
      dose #:____ 
0‌  Other medication: 
 
 
COMPLICATIONS:
1‌None 0‌Bleeding-EBL: ___ ml 0‌Pneumothorax 0‌Re- Expansion Pulmonary Edema 
0‌Other: 
 
IMPRESSION/PLAN: [REDACTED]is a 67 year old-year-old male who presents for Chest Ultrasound and Instillation of agents for fibrinolysis (subsequent).
The patient tolerated the procedure well.  There were no immediate complications.
--Unclamp chest tube in 1 hour
--Continue strict I/O
--Continue daily CXR while chest tube in place
--Continue nursing chest tube flushing protocol
 
DISPOSITION: Nursing Unit"""

# -------------------------------------------------------------------------
# 2. CONFIGURATION (PROCEDURE FLAGS)
# -------------------------------------------------------------------------
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 0,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0, # Existing tube used, not inserted today
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 1,
}

# -------------------------------------------------------------------------
# 3. DATA DEFINITION (SPANS)
# -------------------------------------------------------------------------
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("Ultrasound, chest", "PROC_METHOD", "Ultrasound", "PROCEDURE:  \n76604 ", "evt_01"),
    ("Instillation", "PROC_METHOD", "Instillation", "32562 ", "evt_02"),
    ("fibrinolysis", "PROC_METHOD", "Fibrinolysis", "agent for ", "evt_02"),
    ("Left", "LATERALITY", "Left", "Hemithorax:   0‌ Right  1‌ ", "evt_01"),
    ("Small", "MEAS_VOL", "Small", "Volume:       0‌ None  0‌ Minimal  1‌ ", "evt_01"),
    ("Anechoic", "OBS_LESION", "Anechoic", "Echogenicity:   1‌ ", "evt_01"),
    ("Thin", "OBS_LESION", "Loculations Thin", "Loculations:  0‌ None  1‌", "evt_01"),
    ("Thick", "OBS_LESION", "Loculations Thick", "Thin  1‌ ", "evt_01"),
    ("chest tube", "DEV_CATHETER", "Chest Tube", "via ", "evt_02"),
    ("left", "LATERALITY", "Left", "Side: ", "evt_02"),
    ("tPA/Dnase", "PROC_METHOD", "tPA/DNase", "5 mg/5 mg ", "evt_02"),
    ("None", "OUTCOME_COMPLICATION", "None", "COMPLICATIONS:\n1‌", "evt_02"),
    ("tolerated the procedure well", "OUTCOME_SYMPTOMS", "Tolerated well", "The patient ", "evt_02"),
]

# -------------------------------------------------------------------------
# 4. EVENT DEFINITIONS
# -------------------------------------------------------------------------
EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Diagnostic",
        "method": ["Ultrasound"],
        "anatomy": ["Pleura"],
        "laterality": "Left",
        "devices": [],
        "samples": [],
        "findings": ["Small Effusion", "Anechoic", "Loculations"],
        "outcomes": []
    },
    {
        "event_id": "evt_02",
        "procedure_type": "Therapeutic",
        "method": ["Fibrinolysis", "Instillation"],
        "anatomy": ["Pleura"],
        "laterality": "Left",
        "devices": ["Chest Tube"],
        "medications": ["tPA", "DNase"],
        "samples": [],
        "findings": [],
        "outcomes": ["Tolerated well", "No complications"]
    }
]

# -------------------------------------------------------------------------
# 5. HELPER FUNCTIONS
# -------------------------------------------------------------------------
def clean_text(text):
    if not text:
        return ""
    return text.strip().replace('\r', '')

def hydrate_span(text, span_text, context_prefix):
    # Normalize line endings in the main text
    text_normalized = text.replace('\r\n', '\n').replace('\r', '\n')
    
    # Escape regex characters in context and span
    escaped_context = re.escape(context_prefix)
    escaped_span = re.escape(span_text)
    
    # Handle the 'hidden' zero-width characters or check-mark placeholders often found in these notes
    # The note uses "1\u200c" (digit one + zero width non-joiner) sometimes. 
    # We allow whitespace or special chars between context and span in the regex.
    pattern = f"{escaped_context}\s*{escaped_span}"
    
    match = re.search(pattern, text_normalized, re.DOTALL)
    
    if match:
        # Calculate full text start index
        full_start = match.start()
        # The span starts after the context
        # We need to find the length of the matched context part roughly
        # Because we used \s* in the middle, we can't just add len(context).
        # Let's search for the span within the match object's group.
        matched_string = match.group(0)
        span_start_in_match = matched_string.rfind(span_text)
        
        start_char = full_start + span_start_in_match
        end_char = start_char + len(span_text)
        return start_char, end_char
    else:
        # Fallback: simple search if context missing or slightly off
        start_char = text_normalized.find(span_text)
        if start_char != -1:
            return start_char, start_char + len(span_text)
        return "", ""

# -------------------------------------------------------------------------
# 6. WORKBOOK GENERATION FUNCTION
# -------------------------------------------------------------------------
def generate_workbook():
    # Create or load workbook
    if os.path.exists(TEMPLATE_PATH):
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    else:
        wb = openpyxl.Workbook()
        # Create default sheets if template doesn't exist
        for sheet_name in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
        # Remove default 'Sheet'
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # -------------------------------------------------------
    # Sheet 1: Note_Text
    # -------------------------------------------------------
    ws_text = wb["Note_Text"]
    # Headers if empty
    if ws_text.max_row == 1 and ws_text["A1"].value is None:
        ws_text.append(["note_id", "source_file", "note_text"])
    
    # Write Data
    # Check if note already exists to avoid dupes (basic check)
    exists = False
    for row in ws_text.iter_rows(min_row=2, values_only=True):
        if row[0] == NOTE_ID:
            exists = True
            break
    if not exists:
        ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # -------------------------------------------------------
    # Sheet 2: Note_Index
    # -------------------------------------------------------
    ws_index = wb["Note_Index"]
    # Headers
    headers_index = ["note_id", "source_file", "procedure_date"] + list(PROCEDURE_FLAGS.keys())
    if ws_index.max_row == 1 and ws_index["A1"].value is None:
        ws_index.append(headers_index)
    
    # Write Row
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE] + list(PROCEDURE_FLAGS.values())
    ws_index.append(row_data)

    # -------------------------------------------------------
    # Sheet 3: Span_Annotations
    # -------------------------------------------------------
    ws_anno = wb["Span_Annotations"]
    headers_anno = ["note_id", "span_text", "label", "normalized_value", "context_prefix", "start_char", "end_char", "event_id"]
    if ws_anno.max_row == 1 and ws_anno["A1"].value is None:
        ws_anno.append(headers_anno)
    
    for span in SPANS:
        # span = (text, label, norm, context, evt)
        # Leave start/end blank for this sheet as per instructions (or pre-fill 0 if preferred, keeping blank for manual review)
        ws_anno.append([NOTE_ID, span[0], span[1], span[2], span[3], "", "", span[4]])

    # -------------------------------------------------------
    # Sheet 4: Span_Hydrated
    # -------------------------------------------------------
    ws_hydra = wb["Span_Hydrated"]
    if ws_hydra.max_row == 1 and ws_hydra["A1"].value is None:
        ws_hydra.append(headers_anno) # Same headers
    
    for span in SPANS:
        start, end = hydrate_span(NOTE_TEXT, span[0], span[3])
        ws_hydra.append([NOTE_ID, span[0], span[1], span[2], span[3], start, end, span[4]])

    # -------------------------------------------------------
    # Sheet 5: Event_Log (Flattened)
    # -------------------------------------------------------
    ws_event = wb["Event_Log"]
    headers_event = ["note_id", "event_id", "category", "value"]
    if ws_event.max_row == 1 and ws_event["A1"].value is None:
        ws_event.append(headers_event)
    
    for evt in EVENTS:
        eid = evt["event_id"]
        # Write methods
        for m in evt.get("method", []):
            ws_event.append([NOTE_ID, eid, "Method", m])
        # Write anatomy
        for a in evt.get("anatomy", []):
            ws_event.append([NOTE_ID, eid, "Anatomy", a])
        # Write findings
        for f in evt.get("findings", []):
            ws_event.append([NOTE_ID, eid, "Finding", f])
        # Devices
        for d in evt.get("devices", []):
            ws_event.append([NOTE_ID, eid, "Device", d])

    # -------------------------------------------------------
    # Sheet 6: V3_Procedure_Events
    # -------------------------------------------------------
    ws_v3 = wb["V3_Procedure_Events"]
    headers_v3 = ["note_id", "event_id", "procedure_type", "method", "anatomy", "laterality", "devices", "samples", "findings", "outcomes"]
    if ws_v3.max_row == 1 and ws_v3["A1"].value is None:
        ws_v3.append(headers_v3)
    
    for evt in EVENTS:
        row = [
            NOTE_ID,
            evt["event_id"],
            evt.get("procedure_type", ""),
            json.dumps(evt.get("method", [])),
            json.dumps(evt.get("anatomy", [])),
            evt.get("laterality", ""),
            json.dumps(evt.get("devices", [])),
            json.dumps(evt.get("samples", [])),
            json.dumps(evt.get("findings", [])),
            json.dumps(evt.get("outcomes", []))
        ]
        ws_v3.append(row)

    # -------------------------------------------------------
    # Sheet 7: V3_Registry_JSON
    # -------------------------------------------------------
    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 1 and ws_json["A1"].value is None:
        ws_json.append(["note_id", "json_data"])
    
    registry_data = {
        "note_id": NOTE_ID,
        "procedure_date": PROCEDURE_DATE,
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS
    }
    
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Workbook generated successfully: {OUTPUT_PATH}")

# -------------------------------------------------------------------------
# 7. EXECUTION BLOCK
# -------------------------------------------------------------------------
if __name__ == "__main__":
    generate_workbook()