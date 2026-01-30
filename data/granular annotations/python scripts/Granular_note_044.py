import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# =============================================================================
# INPUT DATA
# =============================================================================

NOTE_ID = "note_044"
SOURCE_FILE = "note_044.txt"
PROCEDURE_DATE = "2026-01-12" # Placeholder based on current context if strictly needed, else blank
NOTE_TEXT = """NOTE_ID:  note_044 SOURCE_FILE: note_044.txt INDICATION FOR OPERATION:  [REDACTED]is a 67 year old-year-old male who presents with Complicated Effusion.
PREOPERATIVE DIAGNOSIS:  Complicated Effusion
POSTOPERATIVE DIAGNOSIS: Complicated Effusion
 
PROCEDURE:  
76604 Ultrasound, chest (includes mediastinum), real time with image documentation
32557 Insert catheter pleura with imaging (chest tube)
32561 Instillation(s), via chest tube/catheter, agent for fibrinolysis (eg, fibrinolytic agent for break up of multiloculated effusion);
initial day
 
Local ONLY
 
PROCEDURE IN DETAIL:
 
 
PATIENT POSITION: 
0‌ Supine  1‌ Sitting   
0‌ Lateral Decubitus:  0‌ Right 0‌ Left 
 
CHEST ULTRASOUND FINDINGS:  1‌ Image saved and uploaded to patient's medical record
Hemithorax:   0‌ Right  1‌ Left 
 
Pleural Effusion: 
Volume:       0‌ None  0‌ Minimal  0‌ Small  0‌ Moderate  1‌ Large 
Echogenicity:   1‌ Anechoic  0‌ Hypoechoic  0‌ Isoechoic  0‌ Hyperechoic 
Loculations:  0‌ None  1‌Thin  1‌ Thick 
Diaphragmatic Motion:  1‌ Normal  0‌ 
Diminished  0‌ Absent  
Lung: 
Lung sliding before procedure:   0‌ Present  1‌ Absent 
Lung sliding post procedure:   0‌ Present  1‌ Absent 
Lung consolidation/atelectasis: 1‌ Present  0‌  Absent 
Pleura:  0‌ Normal  1‌ Thick  0‌ Nodular 
 
 
 
  
Insertion site prepped and draped in sterile fashion.
ANESTHESIA:   Lidocaine 1%: 15 ml      Other: ______ 
Entry Site: 
0‌ Right ___ Intercostal Space   1‌ Left  7th Intercostal Space 
0‌ Mid-clavicular   1‌ Post-axillary  0‌ Mid-scapular  0‌ Other: 
 
Size:  0‌ 6Fr  0‌ 8Fr   0‌ 12FR   1‌ 14Fr  0‌ 16Fr   0‌ 18Fr  0‌ 24Fr   0‌ 32 Fr   0‌ Other: 
Sutured: 1‌ Yes 0‌ No 
 
PROCEDURE FINDINGS: 
A  pigtail catheter was inserted using the Seldinger technique.
Entry into the pleural space was confirmed with the easy removal of minimal serous appearing pleural fluid and air bubbles.
A guidewire was inserted using the introducer needle pointed in the apical posterior direction. The introducer needle was then removed.
A dilator was then inserted over the wire with a twisting motion in order to form a tract for catheter insertion.
The dilator was removed and the pigtail catheter (with trochar) was advanced over the guidewire.
The catheter was inserted into the chest until all catheter holes were well within the chest.
The guidewire and trochar were then removed.  The tube was then attached to the collection drain apparatus and secured in place with suture and covered.
Fluid Removed: 150 ml 
0‌ Serous  1‌ Serosanguinous 0‌ Bloody  0‌ Chylous 0‌ Other: 
 
Pleural pressures:   If not measured, please check here 1‌

Drainage device:   1‌ Pleurovac    0‌ Drainage Bag  0‌ Heimlich Valve  0‌ Pneumostat  0‌ Other: 
Suction: 0‌ No 1‌Yes, -20cmH20 
 
 
SPECIMEN(S): 
0‌None           0‌PH               1‌ LDH                
        1‌Glucose       1‌T.
Protein    1‌Cholesterol
1‌Cell Count   0‌ ADA             0‌Triglycerides            0‌Amylase 
1‌Gram Stain/ Culture            1‌AFB                         1‌Fungal Culture 
1‌Cytology      0‌Flow Cytometry                 
              0‌Other: 
 
CXR ordered: 1‌ Yes 0‌ No 
 
===============================================
 
FIBRINOLYTIC THERAPY
 
Date of chest tube insertion: [REDACTED]
Side: left
1‌  5 mg/5 mg tPA/Dnasedose #:1 
            0‌  ___mg tPA                              dose #:____ 
0‌  Other medication: 
 
 
COMPLICATIONS:
1‌None 0‌Bleeding-EBL: ___ ml 0‌Pneumothorax 0‌Re- Expansion Pulmonary 
Edema 
0‌Other: 
 
IMPRESSION/PLAN: [REDACTED]is a 67 year old-year-old male who presents for Chest Ultrasound, Chest tube placement, and Instillation of agents for fibrinolysis (initial).
The patient tolerated the procedure well.  There were no immediate complications.
--Post-procedure CXR
--Unclamp chest tube in 1 hour
--Strict I/O
--Daily CXR while chest tube in place
--Nursing chest tube flushing protocol
 
DISPOSITION: Nursing Unit"""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# =============================================================================
# CONFIGURATION
# =============================================================================

PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 0, "bal": 0, "bronchial_wash": 0, "brushings": 0,
    "endobronchial_biopsy": 0, "tbna_conventional": 0, "linear_ebus": 0, "radial_ebus": 0,
    "navigational_bronchoscopy": 0, "transbronchial_biopsy": 0, "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0, "foreign_body_removal": 0, "airway_dilation": 0,
    "airway_stent": 0, "thermal_ablation": 0, "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0, "blvr": 0, "peripheral_ablation": 0, "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0, "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0, "chest_tube": 1, "ipc": 0, "medical_thoracoscopy": 0,
    "pleurodesis": 0, "pleural_biopsy": 0, "fibrinolytic_therapy": 1
}

# =============================================================================
# SPAN DATA
# =============================================================================

# Helper to structure spans: (text, label, normalized, prefix_context, field)
SPANS_DATA = [
    # Event 1: Chest Ultrasound
    ("Ultrasound, chest", "PROC_METHOD", "Chest Ultrasound", None, "procedure"),
    ("Left", "LATERALITY", "Left", "Hemithorax:   0‌ Right  1‌ ", "target.laterality"),
    ("Large", "OBS_LESION", "Large", "0‌ Moderate  1‌ ", "finding.size"),
    ("Thick", "OBS_LESION", "Thick", "Loculations:  0‌ None  1‌Thin  1‌ ", "finding.morphology"),
    ("Absent", "OBS_LESION", "Absent", "Lung sliding before procedure:   0‌ Present  1‌ ", "finding.sliding"),
    ("Thick", "OBS_LESION", "Thick", "Pleura:  0‌ Normal  1‌ ", "finding.pleura"),

    # Event 2: Chest Tube
    ("Left", "LATERALITY", "Left", "Right ___ Intercostal Space   1‌ ", "target.laterality"),
    ("7th Intercostal Space", "ANAT_LUNG_LOC", "7th Intercostal Space", "1‌ Left  ", "target.location"),
    ("Post-axillary", "ANAT_LUNG_LOC", "Post-axillary", "0‌ Mid-clavicular   1‌ ", "target.location"),
    ("14Fr", "DEV_CATHETER_SIZE", "14Fr", "0‌ 12FR   1‌ ", "device.size"),
    ("Pigtail catheter", "DEV_CATHETER", "Pigtail catheter", "A  ", "device.type"),
    ("Seldinger technique", "PROC_METHOD", "Seldinger technique", "inserted using the ", "procedure.method"),
    ("150 ml", "MEAS_PLEURAL_DRAIN", "150 ml", "Fluid Removed: ", "measurement.volume"),
    ("Serosanguinous", "OBS_LESION", "Serosanguinous", "0‌ Serous  1‌ ", "finding.appearance"), # Note: Prompt text shows checkmarks 1 Serosanguinous
    
    # Event 3: Fibrinolysis
    ("Instillation of agents for fibrinolysis", "PROC_METHOD", "Fibrinolysis", "Chest tube placement, and ", "procedure.method"),
    ("tPA/Dnase", "DEV_INSTRUMENT", "tPA/Dnase", "5 mg/5 mg ", "device.agent"),
    
    # Outcomes
    ("None", "OUTCOME_COMPLICATION", "None", "COMPLICATIONS:\n1‌", "outcome.complication"),
    ("tolerated the procedure well", "OUTCOME_SYMPTOMS", "tolerated well", None, "outcome.symptoms")
]

# Note: In the text `0‌ Serous  1‌ Serosanguinous`, 1 indicates selection. I used Serosanguinous in the list above.

# =============================================================================
# HYDRATION LOGIC
# =============================================================================

def find_span_offsets(full_text, span_text, context_prefix=None, match_index=0):
    """
    Locates start_char and end_char for a span.
    Logic:
    1. Find all exact matches of span_text.
    2. If context_prefix is provided, filter matches where prefix appears in preceding 120 chars.
    3. If multiple remain (or no context provided), use match_index.
    """
    matches = [m.start() for m in re.finditer(re.escape(span_text), full_text)]
    
    if not matches:
        return None, None, "not_found"
    
    selected_start = None
    hydration_status = "ambiguous"
    
    if len(matches) == 1:
        selected_start = matches[0]
        hydration_status = "hydrated_unique"
    elif context_prefix:
        # Filter by context
        valid_matches = []
        for start in matches:
            # Look back up to 120 chars
            window_start = max(0, start - 120)
            preceding_text = full_text[window_start:start]
            # loose check for prefix in window (ignoring some whitespace diffs)
            if context_prefix.strip() in preceding_text:
                valid_matches.append(start)
        
        if valid_matches:
            # If multiple valid matches with context, use match_index (default 0)
            if match_index < len(valid_matches):
                selected_start = valid_matches[match_index]
                hydration_status = "hydrated_prefix_window"
            else:
                selected_start = valid_matches[0] # Fallback
                hydration_status = "hydrated_fallback"
        else:
            # Context not found, fall back to simple index
            if match_index < len(matches):
                selected_start = matches[match_index]
                hydration_status = "hydrated_match_index_no_context"
    else:
        if match_index < len(matches):
            selected_start = matches[match_index]
            hydration_status = "hydrated_match_index"
        else:
            selected_start = matches[0]
            hydration_status = "hydrated_overflow"

    if selected_start is not None:
        return selected_start, selected_start + len(span_text), hydration_status
    
    return None, None, "error"

# =============================================================================
# WORKBOOK GENERATION
# =============================================================================

def generate_workbook():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        # Create a blank one for demonstration if template missing (in real run, must exist)
        wb = openpyxl.Workbook()
        for name in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            wb.create_sheet(name)

    # 1. Note_Text
    ws_text = wb["Note_Text"]
    # Assuming header exists, append row
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws_index = wb["Note_Index"]
    # Construct row: metadata + 30 flags
    idx_row = [
        SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Success", ""
    ]
    # Add flags in order
    flag_keys = [
        "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", 
        "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy", 
        "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration", 
        "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation", 
        "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", 
        "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
        "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis", 
        "pleural_biopsy", "fibrinolytic_therapy"
    ]
    for k in flag_keys:
        idx_row.append(PROCEDURE_FLAGS.get(k, 0))
    ws_index.append(idx_row)

    # 3. Span_Annotations & 4. Span_Hydrated
    ws_span = wb["Span_Annotations"]
    ws_hydrated = wb["Span_Hydrated"]
    
    # Define Event IDs for mapping
    # E1: Ultrasound, E2: Tube, E3: Fibrinolysis
    # Spans 0-5 -> E1
    # Spans 6-13 -> E2
    # Spans 14-15 -> E3
    # Spans 16-17 -> E2 (Outcomes usually mapped to main therapeutic event, Chest Tube)
    
    events_map = (
        ["E1"] * 6 +   # US
        ["E2"] * 8 +   # Tube
        ["E3"] * 2 +   # Fibrinolysis
        ["E2"] * 2     # Outcomes
    )

    span_id_counter = 1
    
    # Process spans
    for i, (text, label, norm, prefix, field) in enumerate(SPANS_DATA):
        span_id = f"{NOTE_ID}_s{span_id_counter:03d}"
        event_id = events_map[i]
        
        # Calculate offsets
        start, end, status = find_span_offsets(NOTE_TEXT, text, prefix)
        
        # Row for Span_Annotations (No offsets)
        row_anno = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure", prefix, text, 
            1 if prefix else 0, # Match index proxy
            "", "", len(text), # Start/End blank
            label, norm, field, event_id,
            0, 0, 0, "Auto", "", "needs_hydration"
        ]
        ws_span.append(row_anno)
        
        # Row for Span_Hydrated (With offsets)
        row_hyd = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure", prefix, text,
            1 if prefix else 0,
            start if start is not None else "",
            end if end is not None else "",
            len(text),
            label, norm, field, event_id,
            0, 0, 0, "Auto", "", status
        ]
        ws_hydrated.append(row_hyd)
        
        span_id_counter += 1

    # 5. Event_Log
    ws_event = wb["Event_Log"]
    # E1: Chest Ultrasound
    ws_event.append([
        SOURCE_FILE, NOTE_ID, "E1", "Chest Ultrasound", "Ultrasound", 
        "Left Hemithorax", "Ultrasound", "", "", "", "Large", "", "Loculations, Thick Pleura", 
        0, "Auto", "Diagnostic", "", "", "", "", "", "", "None"
    ])
    # E2: Chest Tube
    ws_event.append([
        SOURCE_FILE, NOTE_ID, "E2", "Chest Tube", "Chest tube placement", 
        "Left 7th Intercostal", "Pigtail catheter", "14Fr", "", "", "150 ml", 
        "", "Serosanguinous Fluid", 
        0, "Auto", "Therapeutic", "14Fr", "Pigtail", "", "", "Tolerated well", "Serosanguinous", "None"
    ])
    # E3: Fibrinolytic Therapy
    ws_event.append([
        SOURCE_FILE, NOTE_ID, "E3", "Fibrinolytic Therapy", "Instillation", 
        "Pleural Space", "tPA/Dnase", "", "", "", "", "", "", 
        0, "Auto", "", "", "", "", "", "", "", ""
    ])

    # 6. V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    # E1
    ws_v3.append([
        NOTE_ID, "E1", "Diagnostic", "Pleura", "Left", "", "", 
        "Effusion", "Large", "Ultrasound", "{}", "{}", "{}", '{"findings": ["Thick", "Loculated"]}', "",
        "", "", "", "", "", "", "", ""
    ])
    # E2
    ws_v3.append([
        NOTE_ID, "E2", "Therapeutic", "Pleura", "Left", "", "", 
        "", "", "Chest Tube", '{"device": "Pigtail", "size": "14Fr"}', '{"drainage": "150 ml"}', "{}", "{}", "Entry into pleural space confirmed",
        "", "", "14Fr", "", "", "Tolerated well", "Serosanguinous", "None"
    ])
    # E3
    ws_v3.append([
        NOTE_ID, "E3", "Therapeutic", "Pleura", "Left", "", "",
        "", "", "Fibrinolysis", '{"agent": "tPA/Dnase"}', "{}", "{}", "{}", "",
        "", "", "", "", "", "", "", ""
    ])

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": [
            {
                "event_id": "E1",
                "type": "Diagnostic",
                "method": "Chest Ultrasound",
                "target": {"anatomy": "Pleura", "laterality": "Left"},
                "findings": ["Large Effusion", "Thick Loculations"]
            },
            {
                "event_id": "E2",
                "type": "Therapeutic",
                "method": "Chest Tube",
                "device": {"type": "Pigtail catheter", "size": "14Fr"},
                "target": {"anatomy": "Pleura", "laterality": "Left", "location": "7th Intercostal Space"},
                "outcomes": {"complications": "None", "pleural_fluid": "150 ml Serosanguinous"}
            },
            {
                "event_id": "E3",
                "type": "Therapeutic",
                "method": "Fibrinolysis",
                "device": {"agent": "tPA/Dnase"}
            }
        ]
    }
    
    ws_json.append([json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Workbook saved to {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()