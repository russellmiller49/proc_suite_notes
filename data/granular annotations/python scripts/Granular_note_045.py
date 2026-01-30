import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_045"
SOURCE_FILE = "note_045.txt"
PROCEDURE_DATE = "2026-01-12" # Using current date as placeholder or extracted if available
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_045 SOURCE_FILE: note_045.txt INDICATION FOR OPERATION:  [REDACTED]is a 67 year old-year-old male who presents with Pleural Effusion.
The nature, purpose, risks, benefits and alternatives to Chest Ultrasound and Thoracentesis were discussed with the patient in detail.
Patient indicated a wish to proceed with procedure and informed consent was signed.
PREOPERATIVE DIAGNOSIS:  Pleural Effusion
POSTOPERATIVE DIAGNOSIS: Same as preoperative diagnosis - see above.
PROCEDURE:  
76604 Ultrasound, chest (includes mediastinum), real time with image documentation
32555 Aspirate pleura with imaging (thoracentesis)
 
25 Added to the E&M Encounter bill which is separate from a procedure if it is done the same day
 
Local ONLY
 
PROCEDURE IN DETAIL: 
PATIENT POSITION: 
0‌ Supine  1‌ Sitting   
0‌ Lateral Decubitus:  0‌ Right 0‌ Left 
 
CHEST ULTRASOUND FINDINGS:  1‌ Image saved and printed 
Hemithorax:   0‌ Right  1‌ Left 
 
Pleural Effusion: 
Volume:       0‌ None  0‌ Minimal  0‌ Small  0‌ Moderate  1‌ 
Large 
Echogenicity:   1‌ Anechoic  0‌ Hypoechoic  0‌ Isoechoic  1‌ Hyperechoic 
Loculations:  0‌ None  1‌Thin  1‌ Thick 
Diaphragmatic Motion:  0‌ Normal  1‌ Diminished  0‌ Absent  
Lung: 
Lung sliding before procedure:   0‌ Present  0‌ Absent 
Lung sliding post procedure:   0‌ Present  0‌ Absent 
Lung consolidation/atelectasis: 0‌ Present  0‌  Absent 
Pleura:  0‌ Normal  1‌ Thick  0‌ Nodular 
 
Based on Ultrasound Evaluation, Thoracentesis is determined to be feasible and will proceed as planned.
.
 
Insertion site prepped and draped in sterile fashion.  Thoracentesis Kit was used.
ANESTHESIA:   Lidocaine 1%: ___10___ ml      Other: ______ 
Entry Site: 
0‌ Right ___ Intercostal Space   1‌ Left  __6th_ Intercostal Space 
0‌ Mid-clavicular   0‌ Mid-axillary  1‌ Mid-scapular  0‌ Other: 
 
Sutured: 0‌ Yes 1‌ No 
 
PROCEDURE FINDINGS: 
Fluid Removed: __150___ ml 
0‌ Serous  1‌ Serosanguinous 0‌ Bloody  0‌ Chylous 0‌ Other: 
 
Drainage device:   0‌ Pleurovac    1‌ Drainage Bag  0‌ Heimlich Valve  0‌ Pneumostat  0‌ Other: 
Suction: 0‌ No 0‌Yes, - ___ cmH20 
 
 
SPECIMEN(S): 
0‌None 
          0‌PH               1‌ LDH                        1‌Glucose       1‌T.
Protein    1‌Cholesterol
1‌Cell Count   0‌ ADA             0‌Triglycerides            0‌Amylase 
1‌Gram Stain/ Culture            1‌AFB                         1‌Fungal Culture 
1‌Cytology      0‌Flow Cytometry                 
              0‌Other: 
 
CXR ordered: 1‌ Yes 0‌ No 
 
 
COMPLICATIONS:
1‌None 0‌Bleeding-EBL: ___ ml 0‌Pneumothorax 0‌Re- Expansion Pulmonary Edema 
0‌Other: 
 
IMPRESSION/PLAN: [REDACTED]is a 67 year old-year-old male who presents for Chest Ultrasound and Thoracentesis.
The patient tolerated the procedure well.  There were no immediate complications.
- f/u CXR
- f/u studies 
 
 
DISPOSITION: Nursing Unit"""

# -------------------------------------------------------------------------
# CONFIGURATION
# -------------------------------------------------------------------------
FLAGS_ORDER = [
    'diagnostic_bronchoscopy', 'bal', 'bronchial_wash', 'brushings', 'endobronchial_biopsy',
    'tbna_conventional', 'linear_ebus', 'radial_ebus', 'navigational_bronchoscopy',
    'transbronchial_biopsy', 'transbronchial_cryobiopsy', 'therapeutic_aspiration',
    'foreign_body_removal', 'airway_dilation', 'airway_stent', 'thermal_ablation',
    'tumor_debulking_non_thermal', 'cryotherapy', 'blvr', 'peripheral_ablation',
    'bronchial_thermoplasty', 'whole_lung_lavage', 'rigid_bronchoscopy',
    'thoracentesis', 'chest_tube', 'ipc', 'medical_thoracoscopy', 'pleurodesis',
    'pleural_biopsy', 'fibrinolytic_therapy'
]

# Set active flags for this note
ACTIVE_FLAGS = {'thoracentesis'}

# -------------------------------------------------------------------------
# PARSING & EXTRACTION
# -------------------------------------------------------------------------

# Helper to find span offsets
def find_offsets(text, span_text, context_prefix=None, match_index=0):
    """
    Returns (start_char, end_char, method_used)
    method_used: 'unique', 'prefix_window', 'match_index', 'ambiguous'
    """
    if not span_text:
        return None, None, "empty"
    
    # 1. Exact count check
    count = text.count(span_text)
    if count == 0:
        return None, None, "not_found"
    
    if count == 1:
        start = text.find(span_text)
        return start, start + len(span_text), "hydrated_unique"
    
    # 2. Context prefix
    if context_prefix:
        # Scan all occurrences, check 120 chars before
        indices = [m.start() for m in re.finditer(re.escape(span_text), text)]
        for idx in indices:
            window_start = max(0, idx - 120)
            pre_window = text[window_start:idx]
            if context_prefix in pre_window:
                return idx, idx + len(span_text), "hydrated_prefix_window"
    
    # 3. Match index
    indices = [m.start() for m in re.finditer(re.escape(span_text), text)]
    if match_index < len(indices):
        idx = indices[match_index]
        return idx, idx + len(span_text), "hydrated_match_index"
    
    return None, None, f"ambiguous_count={count}"

# Define Spans
# Format: (span_text, label, normalized_value, schema_field, event_id, context_prefix, match_index)
spans_data = [
    # Diagnosis
    ("Pleural Effusion", "OBS_LESION", "Pleural Effusion", "lesion.type", "ev1", "DIAGNOSIS:", 0),
    
    # Ultrasound Findings (Mapped to Thoracentesis Event Context or Findings)
    ("Left", "LATERALITY", "Left", "target.location.laterality", "ev1", "Hemithorax:", 0),
    ("Large", "OBS_LESION", "Large", "lesion.size_qualitative", "ev1", "Volume:", 0),
    ("Anechoic", "OBS_LESION", "Anechoic", "lesion.characteristic", "ev1", "Echogenicity:", 0),
    ("Thick", "OBS_LESION", "Thick", "target.pleura_characteristic", "ev1", "Pleura:", 0),
    ("Diminished", "OBS_LESION", "Diminished", "target.diaphragm_motion", "ev1", "Diaphragmatic Motion:", 0),
    
    # Procedure Details
    ("Thoracentesis", "PROC_METHOD", "Thoracentesis", "method", "ev1", "Based on Ultrasound Evaluation,", 0),
    ("Left", "LATERALITY", "Left", "target.location.laterality", "ev1", "Entry Site:", 0),
    ("6th Intercostal Space", "ANAT_LUNG_LOC", "6th Intercostal Space", "target.location.segment", "ev1", "Left", 0),
    ("Mid-scapular", "ANAT_LUNG_LOC", "Mid-scapular line", "target.location.detail", "ev1", "Mid-axillary", 0),
    
    # Medication
    ("Lidocaine 1%", "PROC_MEDICATION", "Lidocaine 1%", "medication", "ev1", "ANESTHESIA:", 0),
    
    # Findings/Action
    ("150", "MEAS_VOL", "150", "measurements.volume_removed", "ev1", "Fluid Removed:", 0),
    ("Serosanguinous", "OBS_LESION", "Serosanguinous", "lesion.appearance", "ev1", "Fluid Removed:", 0),
    
    # Specimens
    ("LDH", "PROC_ACTION", "LDH", "specimens.type", "ev1", "SPECIMEN(S):", 0),
    ("Glucose", "PROC_ACTION", "Glucose", "specimens.type", "ev1", "LDH", 0),
    ("T. Protein", "PROC_ACTION", "Total Protein", "specimens.type", "ev1", "Glucose", 0),
    ("Cholesterol", "PROC_ACTION", "Cholesterol", "specimens.type", "ev1", "Protein", 0),
    ("Cell Count", "PROC_ACTION", "Cell Count", "specimens.type", "ev1", "Cholesterol", 0),
    ("Gram Stain/ Culture", "PROC_ACTION", "Gram Stain/Culture", "specimens.type", "ev1", "Amylase", 0),
    ("AFB", "PROC_ACTION", "AFB", "specimens.type", "ev1", "Culture", 0),
    ("Fungal Culture", "PROC_ACTION", "Fungal Culture", "specimens.type", "ev1", "AFB", 0),
    ("Cytology", "PROC_ACTION", "Cytology", "specimens.type", "ev1", "Fungal Culture", 0),
    
    # Outcomes
    ("None", "OUTCOME_COMPLICATION", "None", "outcomes.complications", "ev1", "COMPLICATIONS:", 0),
    ("tolerated the procedure well", "OUTCOME_SYMPTOMS", "tolerated well", "outcomes.symptoms", "ev1", None, 0),
]

# Event Data
event_map = {
    "ev1": {
        "type": "Thoracentesis",
        "method": "Thoracentesis",
        "anatomy": "Left Pleural Space",
        "findings": ["Pleural Effusion", "Large", "Anechoic", "Serosanguinous"],
        "measurements": {"volume_removed": "150 ml"},
        "specimens": ["LDH", "Glucose", "T. Protein", "Cholesterol", "Cell Count", "Gram Stain/Culture", "AFB", "Fungal Culture", "Cytology"],
        "complications": "None"
    }
}

# -------------------------------------------------------------------------
# WORKBOOK GENERATION
# -------------------------------------------------------------------------

def generate_workbook():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        print(f"Error: Template {TEMPLATE_PATH} not found.")
        return

    # 1. Note_Text
    ws_text = wb["Note_Text"]
    # Check if header exists, if not write it (assuming template has it, but safe to append)
    # Template usually has headers. We append data.
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws_index = wb["Note_Index"]
    row_data = [
        SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Extraction", ""
    ]
    # Add flags
    for flag in FLAGS_ORDER:
        row_data.append(1 if flag in ACTIVE_FLAGS else 0)
    ws_index.append(row_data)

    # 3. Span_Annotations (Anchor First) & 4. Span_Hydrated
    ws_annot = wb["Span_Annotations"]
    ws_hydra = wb["Span_Hydrated"]
    
    # Columns for Span_Annotations: 
    # source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, 
    # start_char, end_char, span_len, label, normalized_value, schema_field, event_id, 
    # is_negated, is_historical, time_anchor, reviewer, comments, hydration_status
    
    span_id_counter = 1
    
    for sp_txt, label, norm, schema, ev_id, ctx, midx in spans_data:
        span_id = f"{NOTE_ID}_s{span_id_counter:03d}"
        span_id_counter += 1
        
        # Calculate hydration
        start, end, status = find_offsets(NOTE_TEXT, sp_txt, ctx, midx)
        length = len(sp_txt) if sp_txt else 0
        
        # Base row
        row_base = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure", ctx, sp_txt, midx
        ]
        
        # Annot row (Blank offsets)
        row_annot = row_base + [
            None, None, f"=LEN(F{ws_annot.max_row + 1})", 
            label, norm, schema, ev_id, 
            0, 0, 0, "", "", "needs_hydration"
        ]
        ws_annot.append(row_annot)
        
        # Hydra row (Filled offsets)
        row_hydra = row_base + [
            start, end, length,
            label, norm, schema, ev_id,
            0, 0, 0, "", "", status
        ]
        ws_hydra.append(row_hydra)

    # 5. Event_Log
    ws_log = wb["Event_Log"]
    # Columns: source_file, note_id, event_id, event_type, method, anatomy_target, device, 
    # needle_gauge, stations, counts, measurements, specimens, findings, is_historical, reviewer, comments, 
    # device_size, device_material, outcome_airway_lumen_pre, outcome_airway_lumen_post, 
    # outcome_symptoms, outcome_pleural, outcome_complication
    
    # Event 1
    ev1 = event_map["ev1"]
    row_ev1 = [
        SOURCE_FILE, NOTE_ID, "ev1", ev1["type"], ev1["method"],
        ev1["anatomy"], "", "", "", "", 
        json.dumps(ev1["measurements"]), json.dumps(ev1["specimens"]), json.dumps(ev1["findings"]),
        0, "System", "",
        "", "", "", "",
        "Tolerated well", "", ev1["complications"]
    ]
    ws_log.append(row_ev1)

    # 6. V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    # note_id, event_id, type, target.anatomy_type, target.location.lobe, target.location.segment, 
    # target.station, lesion.type, lesion.size_mm, method, devices_json, measurements_json, 
    # specimens_json, findings_json, evidence_quote, stent.size, stent.material_or_brand, 
    # catheter.size_fr, outcomes.airway.lumen_pre, outcomes.airway.lumen_post, outcomes.symptoms, 
    # outcomes.pleural, outcomes.complications
    
    row_v3_1 = [
        NOTE_ID, "ev1", ev1["type"], "Pleura", "Left", "6th Intercostal",
        "", "Pleural Effusion", "Large", ev1["method"], "[]", json.dumps(ev1["measurements"]),
        json.dumps(ev1["specimens"]), json.dumps(ev1["findings"]), "", "", "",
        "", "", "", "Tolerated well", "", ev1["complications"]
    ]
    ws_v3.append(row_v3_1)

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": [
            {
                "event_id": "ev1",
                "type": ev1["type"],
                "method": ev1["method"],
                "target": {"location": {"laterality": "Left", "segment": "6th Intercostal"}},
                "measurements": ev1["measurements"],
                "specimens": ev1["specimens"],
                "outcomes": {"complications": ev1["complications"], "symptoms": "Tolerated well"}
            }
        ],
        "no_immediate_complications": True
    }
    ws_json.append([json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()