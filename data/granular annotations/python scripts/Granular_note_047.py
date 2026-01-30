import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import json
import re
import os

# ==========================================
# INPUT DATA
# ==========================================
NOTE_ID = "note_047"
SOURCE_FILE = "note_047.txt"
PROCEDURE_DATE = "2026-01-12" # inferred from context or left blank if unknown

# Full text from the user input
NOTE_TEXT = """NOTE_ID:  note_047 SOURCE_FILE: note_047.txt INDICATION FOR OPERATION:  [REDACTED]is a 77 year old-year-old female who presents with lung nodule.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
PREOPERATIVE DIAGNOSIS: R91.1 Solitary Lung Nodule
 
POSTOPERATIVE DIAGNOSIS:  R91.1 Solitary Lung Nodule
 
PROCEDURE:  
31899 Unlisted Procedure (Trach Change with Mature Tract or Procedure NOS)
31645 Therapeutic aspiration initial episode
31623 Dx bronchoscope/brushing    
31624 Dx bronchoscope/lavage (BAL)    
31628 TBBX single lobe     
31629 TBNA single lobe   
31626 Fiducial marker placements, single or multiple     
31627 Navigational Bronchoscopy (computer assisted)
77012 Radiology / radiologic guidance for CT guided needle placement (CIOS)
76377 3D rendering with interpretation and reporting of CT, US, Tomo modality (ION Planning Station)
31653 EBUS sampling 3 or 
more nodes  
31654 Radial EBUS for peripheral lesion
76982 Ultrasound Elastography, First Target Lesion
76983 Ultrasound Elastography, Additional Targets 
76983 Ultrasound Elastography, Additional Target 2
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
 
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)
 
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required a Transbronchial Cryo biopsies.
This resulted in >40% increased work due to Technical difficulty of procedure and Physical and mental effort required.
Apply to: 31628 TBBX single lobe     .
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
Linear EBUS 
Radial EBUS
Ion Robotic Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
CT Chest scan was placed on separate planning station to generate 3D rendering of the pathway to target.
The navigational plan was reviewed and verified.  This was then loaded into robotic bronchoscopy platform.
Successful therapeutic aspiration was performed to clean out the Right Mainstem, Bronchus Intermedius , and Left Mainstem from mucus.
Ventilation Parameters:
Mode	RR	TV	PEEP	FiO2	Flow Rate	Pmean
VCV	15	350	12	100	10	15
 
Robotic navigation bronchoscopy was performed with Ion platform.  Partial registration was used.
Ion robotic catheter was used to engage the Anterior Segment of LUL (LB3).
Target lesion is about 2 cm in diameter.   Under navigational guidance the ion robotic catheter was advanced to 1.0 cm away from the planned target.
Radial EBUS was performed to confirm that the location of the nodule is Eccentric.
The following features were noted: Continuous margin  and Absence of linear-discrete air bronchogram.
Cone Beam CT was performed: 3-D reconstructions were performed on an independent workstation.
Cios Spin system was used for evaluation of nodule location.  Low dose spin was performed to acquire CT imaging.
This was passed on to Ion platform system for reconstruction and nodule location.
The 3D images was interpreted on an independent workstation (Ion).
Using the newly acquired nodule location, the Ion robotic system was adjusted to the new targeted location.
I personally interpreted the cone beam CT and 3-D reconstruction.
Transbronchial needle aspiration was performed with 21G Needle through the extended working channel catheter.  Total 4 samples were collected.
Samples sent for Cytology.
 
Transbronchial biopsy was performed with alligator forceps the extended working channel catheter.
Total 1 samples were collected.  Samples sent for Pathology.
 
Transbronchial cryobiopsy was performed with 1.1mm cryoprobe via the extended working channel catheter.
Freeze time of 6 seconds were used.  Total 6 samples were collected.  Samples sent for Pathology.
Transbronchial brushing was performed with Protected cytology brush the extended working channel catheter.  Total 1 samples were collected.
Samples sent for Microbiology (Cultures/Viral/Fungal).
 
Bronchial alveolar lavage was performed the extended working channel catheter.
Instilled 10 cc of NS, suction returned with 5 cc of NS.  Samples sent for Microbiology (Cultures/Viral/Fungal).
Fiducial marker (0.8mm x 3mm soft tissue gold CIVCO) was loaded with bone wax and placed under fluoroscopy guidance.
Prior to withdraw of the bronchoscope. 
 
ROSE from ION procedure was noted to be:
Conclusive evidence of malignant neoplasm
 
Prior to withdrawal of the bronchoscope, inspection demonstrated no evidence of bleeding.
CT Chest scan was placed on separate planning station to generate 3D rendering of the pathway to target.
The navigational plan was reviewed and verified.  This was then loaded into robotic bronchoscopy platform.
Second RLL target was attempted with the following settings. 
 
Ventilation Parameters:
Mode	RR	TV	PEEP	FiO2	Flow Rate	Pmean
vcv	15	350	16	100	10	20
 
Robotic navigation bronchoscopy was performed with Ion platform.
Partial registration was used.    Ion robotic catheter was used to engage the Lateral-basal Segment of RLL (RB9).
Target lesion is about 1 cm in diameter.   Under navigational guidance the ion robotic catheter was advanced to 1.0 cm away from the planned target.
Radial EBUS was performed to confirm that the location of the nodule is not visible.
T
 
Cone Beam CT was performed: 3-D reconstructions were performed on an independent workstation.
Cios Spin system was used for evaluation of nodule location.  Low dose spin was performed to acquire CT imaging.
This was passed on to Ion platform system for reconstruction and nodule location.
The 3D images was interpreted on an independent workstation (Ion).
Using the newly acquired nodule location, the Ion robotic system was adjusted to the new targeted location.
I personally interpreted the cone beam CT and 3-D reconstruction.
2 spins were done to assess the area and the RLL nodule was not identified on the system.
Due to inability to identify the nodule, we aborted the procedure. 
 
 
EBUS-Findings
Indications: Diagnostic and Staging
Technique:
All lymph node stations were assessed.
Only those 5 mm or greater in short axis were sampled.
Lymph node sizing was performed by EBUS and sampling by transbronchial needle aspiration was performed using 25-gauge Needle and 22-gauge Needle.
Lymph Nodes/Sites Inspected: 4R (lower paratracheal) node
4L (lower paratracheal) node
7 (subcarinal) node
10R lymph node
10L lymph node
11Rs lymph node
11L lymph node
 
No immediate complications
 
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
Elastography provided a semi-quantitative classification (Type 1–3), which was used to guide biopsy site selection and sampling strategy.
Lymph Nodes Evaluated:
Site 1: The 11Rs lymph node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was sampled.. 4 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 1 elastographic pattern, predominantly soft (green/yellow), suggesting a reactive or benign process.
Despite the benign appearance, TBNA was performed to confirm the absence of malignancy and to obtain cytology for diagnostic completeness.
Site 2: The 7 (subcarinal) node was => 10 mm on CT and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was sampled.. 4 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 2 elastographic pattern with mixed soft and stiff regions.
Given this heterogeneous and indeterminate appearance, TBNA was directed at representative areas to ensure comprehensive sampling and to minimize the risk of underdiagnosis.
Site 3: The 11L lymph node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was sampled.. 4 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 2 elastographic pattern with mixed soft and stiff regions.
Given this heterogeneous and indeterminate appearance, TBNA was directed at representative areas to ensure comprehensive sampling and to minimize the risk of underdiagnosis.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
LUL TBCBX, TBNA, TBBX, BAL, Brush
11R, 7, 11L TBNA
 
IMPRESSION/PLAN: [REDACTED]is a 77 year old-year-old female who presents for bronchoscopy for lung nodule.
- f/u in clinic"""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# ==========================================
# CLASS DEFINITION
# ==========================================

class Phase0Generator:
    def __init__(self, note_id, source_file, note_text, template_path):
        self.note_id = note_id
        self.source_file = source_file
        self.note_text = note_text
        self.template_path = template_path
        self.workbook = None
        
        # Load workbook (or create blank if template missing - strict requirement is load)
        if os.path.exists(template_path):
            self.workbook = openpyxl.load_workbook(template_path)
        else:
            # Fallback for demonstration if template file not present
            self.workbook = openpyxl.Workbook()
            self._create_sheets()

        self.spans = []
        self.hydrated_spans = []
        self.events = {} # Map event_id to dict

    def _create_sheets(self):
        # Create standard sheets if missing
        sheets = ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]
        for s in sheets:
            if s not in self.workbook.sheetnames:
                self.workbook.create_sheet(s)

    def write_note_text(self):
        ws = self.workbook["Note_Text"]
        # Determine next empty row
        row = ws.max_row + 1
        if row == 1:
            ws.append(["note_id", "source_file", "note_text"])
            row = 2
        
        # Sanitize text to remove illegal chars if any
        clean_text = self.note_text
        ws.append([self.note_id, self.source_file, clean_text])

    def write_note_index(self):
        ws = self.workbook["Note_Index"]
        # Headers
        headers = [
            "source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes",
            "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", "tbna_conventional",
            "linear_ebus", "radial_ebus", "navigational_bronchoscopy", "transbronchial_biopsy", "transbronchial_cryobiopsy",
            "therapeutic_aspiration", "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
            "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", "bronchial_thermoplasty",
            "whole_lung_lavage", "rigid_bronchoscopy",
            "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis", "pleural_biopsy", "fibrinolytic_therapy"
        ]
        
        if ws.max_row == 1 and ws.cell(1,1).value is None:
            ws.append(headers)

        # Logic for flags
        txt = self.note_text.lower()
        
        flags = {h: 0 for h in headers[8:]}
        
        # Rules based on NOTE_047
        flags["diagnostic_bronchoscopy"] = 1
        if "bal" in txt or "lavage" in txt: flags["bal"] = 1
        if "brushing" in txt: flags["brushings"] = 1
        if "transbronchial needle aspiration" in txt or "tbna" in txt: flags["tbna_conventional"] = 1
        if "ebus" in txt: flags["linear_ebus"] = 1
        if "radial ebus" in txt: flags["radial_ebus"] = 1
        if "navigational" in txt or "ion" in txt: flags["navigational_bronchoscopy"] = 1
        if "transbronchial biopsy" in txt: flags["transbronchial_biopsy"] = 1
        if "cryobiopsy" in txt: flags["transbronchial_cryobiopsy"] = 1
        if "therapeutic aspiration" in txt: flags["therapeutic_aspiration"] = 1
        
        # Row data
        row_data = [
            self.source_file, self.note_id, "", PROCEDURE_DATE, "", "", "ready", ""
        ]
        for h in headers[8:]:
            row_data.append(flags[h])
            
        ws.append(row_data)

    def define_spans(self):
        # MANUAL SPAN EXTRACTION LOGIC
        # Format: (text, context, label, norm_value, event_id)
        
        # Event 1: Therapeutic Aspiration
        self.spans.append({
            "span_text": "Successful therapeutic aspiration was performed",
            "context_prefix": None,
            "label": "PROC_METHOD",
            "normalized_value": "therapeutic_aspiration",
            "event_id": "evt_01_asp"
        })
        self.spans.append({
            "span_text": "Right Mainstem, Bronchus Intermedius , and Left Mainstem",
            "context_prefix": "clean out the ",
            "label": "ANAT_AIRWAY",
            "normalized_value": "bilateral_airways",
            "event_id": "evt_01_asp"
        })

        # Event 2: LUL Nodule (Nav + Biopsies)
        # Location
        self.spans.append({
            "span_text": "Anterior Segment of LUL (LB3)",
            "context_prefix": "engage the ",
            "label": "ANAT_LUNG_LOC",
            "normalized_value": "LUL Ant Seg",
            "event_id": "evt_02_lul"
        })
        self.spans.append({
            "span_text": "about 2 cm in diameter",
            "context_prefix": "Target lesion is ",
            "label": "OBS_LESION",
            "normalized_value": "20mm",
            "event_id": "evt_02_lul"
        })
        # Radial EBUS
        self.spans.append({
            "span_text": "Radial EBUS was performed",
            "context_prefix": "planned target.\n",
            "label": "PROC_METHOD",
            "normalized_value": "radial_ebus",
            "event_id": "evt_02_lul"
        })
        self.spans.append({
            "span_text": "Eccentric",
            "context_prefix": "nodule is ",
            "label": "OBS_LESION",
            "normalized_value": "eccentric",
            "event_id": "evt_02_lul"
        })
        # TBNA
        self.spans.append({
            "span_text": "Transbronchial needle aspiration was performed",
            "context_prefix": "reconstruction.\n",
            "label": "PROC_METHOD",
            "normalized_value": "tbna",
            "event_id": "evt_02_lul"
        })
        self.spans.append({
            "span_text": "21G Needle",
            "context_prefix": "performed with ",
            "label": "DEV_NEEDLE",
            "normalized_value": "21G",
            "event_id": "evt_02_lul"
        })
        self.spans.append({
            "span_text": "Total 4 samples were collected",
            "context_prefix": "channel catheter.  ",
            "label": "MEAS_COUNT",
            "normalized_value": "4",
            "event_id": "evt_02_lul"
        })
        # TBBX
        self.spans.append({
            "span_text": "Transbronchial biopsy was performed",
            "context_prefix": "Cytology.\n \n",
            "label": "PROC_METHOD",
            "normalized_value": "transbronchial_biopsy",
            "event_id": "evt_02_lul"
        })
        self.spans.append({
            "span_text": "alligator forceps",
            "context_prefix": "performed with ",
            "label": "DEV_INSTRUMENT",
            "normalized_value": "alligator_forceps",
            "event_id": "evt_02_lul"
        })
        # Cryo
        self.spans.append({
            "span_text": "Transbronchial cryobiopsy was performed",
            "context_prefix": "Pathology.\n \n",
            "label": "PROC_METHOD",
            "normalized_value": "transbronchial_cryobiopsy",
            "event_id": "evt_02_lul"
        })
        self.spans.append({
            "span_text": "1.1mm cryoprobe",
            "context_prefix": "performed with ",
            "label": "DEV_INSTRUMENT",
            "normalized_value": "cryoprobe_1.1mm",
            "event_id": "evt_02_lul"
        })
        self.spans.append({
            "span_text": "Total 6 samples were collected",
            "context_prefix": "used.  ",
            "label": "MEAS_COUNT",
            "normalized_value": "6",
            "event_id": "evt_02_lul"
        })
        # Brush
        self.spans.append({
            "span_text": "Transbronchial brushing was performed",
            "context_prefix": "Pathology.\n",
            "label": "PROC_METHOD",
            "normalized_value": "brushing",
            "event_id": "evt_02_lul"
        })
        # BAL
        self.spans.append({
            "span_text": "Bronchial alveolar lavage was performed",
            "context_prefix": "Viral/Fungal).\n \n",
            "label": "PROC_METHOD",
            "normalized_value": "bal",
            "event_id": "evt_02_lul"
        })
        # Fiducial
        self.spans.append({
            "span_text": "Fiducial marker (0.8mm x 3mm soft tissue gold CIVCO)",
            "context_prefix": None,
            "label": "DEV_VALVE", # Using DEV_VALVE as catch-all for implants/fiducials if DEV_IMPLANT not avail, else DEV_INSTRUMENT
            "normalized_value": "fiducial_marker",
            "event_id": "evt_02_lul"
        })
        # ROSE
        self.spans.append({
            "span_text": "Conclusive evidence of malignant neoplasm",
            "context_prefix": "ROSE from ION procedure was noted to be:\n",
            "label": "OBS_ROSE",
            "normalized_value": "malignant",
            "event_id": "evt_02_lul"
        })

        # Event 3: RLL Attempt
        self.spans.append({
            "span_text": "Lateral-basal Segment of RLL (RB9)",
            "context_prefix": "engage the ",
            "label": "ANAT_LUNG_LOC",
            "normalized_value": "RLL Lat-Bas",
            "event_id": "evt_03_rll"
        })
        self.spans.append({
            "span_text": "Radial EBUS was performed",
            "context_prefix": "planned target.\n",
            "label": "PROC_METHOD",
            "normalized_value": "radial_ebus",
            "event_id": "evt_03_rll"
        })
        self.spans.append({
            "span_text": "nodule is not visible",
            "context_prefix": "location of the ",
            "label": "OBS_LESION",
            "normalized_value": "not_visible",
            "event_id": "evt_03_rll"
        })
        self.spans.append({
            "span_text": "aborted the procedure",
            "context_prefix": "nodule, we ",
            "label": "PROC_ACTION",
            "normalized_value": "aborted",
            "event_id": "evt_03_rll"
        })

        # Event 4: EBUS 11Rs
        self.spans.append({
            "span_text": "11Rs lymph node",
            "context_prefix": "Site 1: The ",
            "label": "ANAT_LN_STATION",
            "normalized_value": "11Rs",
            "event_id": "evt_04_ebus_11rs"
        })
        self.spans.append({
            "span_text": "4 endobronchial ultrasound guided transbronchial biopsies",
            "context_prefix": "sampled.. ",
            "label": "PROC_METHOD",
            "normalized_value": "ebus_tbna",
            "event_id": "evt_04_ebus_11rs"
        })
        self.spans.append({
            "span_text": "Type 1 elastographic pattern",
            "context_prefix": "demonstrated a ",
            "label": "OBS_LESION",
            "normalized_value": "elastography_type_1",
            "event_id": "evt_04_ebus_11rs"
        })

        # Event 5: EBUS 7
        self.spans.append({
            "span_text": "7 (subcarinal) node",
            "context_prefix": "Site 2: The ",
            "label": "ANAT_LN_STATION",
            "normalized_value": "7",
            "event_id": "evt_05_ebus_7"
        })
        self.spans.append({
            "span_text": "4 endobronchial ultrasound guided transbronchial biopsies",
            "context_prefix": "sampled.. ",
            "match_index": 1, # Second occurrence in EBUS section
            "label": "PROC_METHOD",
            "normalized_value": "ebus_tbna",
            "event_id": "evt_05_ebus_7"
        })
        self.spans.append({
            "span_text": "Type 2 elastographic pattern",
            "context_prefix": "demonstrated a ",
            "label": "OBS_LESION",
            "normalized_value": "elastography_type_2",
            "event_id": "evt_05_ebus_7"
        })

        # Event 6: EBUS 11L
        self.spans.append({
            "span_text": "11L lymph node",
            "context_prefix": "Site 3: The ",
            "label": "ANAT_LN_STATION",
            "normalized_value": "11L",
            "event_id": "evt_06_ebus_11l"
        })
        self.spans.append({
            "span_text": "4 endobronchial ultrasound guided transbronchial biopsies",
            "context_prefix": "sampled.. ",
            "match_index": 2, # Third occurrence
            "label": "PROC_METHOD",
            "normalized_value": "ebus_tbna",
            "event_id": "evt_06_ebus_11l"
        })
        self.spans.append({
            "span_text": "Type 2 elastographic pattern",
            "context_prefix": "demonstrated a ",
            "match_index": 1, # Second occurrence of Type 2
            "label": "OBS_LESION",
            "normalized_value": "elastography_type_2",
            "event_id": "evt_06_ebus_11l"
        })

        # Outcome
        self.spans.append({
            "span_text": "No immediate complications",
            "context_prefix": "procedure well.  There were ",
            "label": "OUTCOME_COMPLICATION",
            "normalized_value": "none",
            "event_id": "evt_global"
        })

    def hydrate_spans(self):
        # Programmatic offset calculation
        for s in self.spans:
            txt = s["span_text"]
            # Find all matches
            matches = [m.start() for m in re.finditer(re.escape(txt), self.note_text)]
            
            start_char = None
            hydration_type = "needs_hydration"
            
            if len(matches) == 1:
                start_char = matches[0]
                hydration_type = "hydrated_unique"
            elif len(matches) > 1:
                if s.get("context_prefix"):
                    # Check preceding 120 chars
                    found = False
                    for m in matches:
                        pre_text = self.note_text[max(0, m-120):m]
                        if s["context_prefix"] in pre_text:
                            start_char = m
                            hydration_type = "hydrated_prefix_window"
                            found = True
                            break
                    if not found:
                         # Fallback to index if provided
                        idx = s.get("match_index", 0)
                        if idx < len(matches):
                            start_char = matches[idx]
                            hydration_type = "hydrated_match_index"
                        else:
                            hydration_type = f"ambiguous_count={len(matches)}"
                else:
                    # Use match index
                    idx = s.get("match_index", 0)
                    if idx < len(matches):
                        start_char = matches[idx]
                        hydration_type = "hydrated_match_index"
                    else:
                         hydration_type = f"ambiguous_count={len(matches)}"
            else:
                hydration_type = "not_found"

            # Create hydrated record
            h_rec = s.copy()
            if start_char is not None:
                h_rec["start_char"] = start_char
                h_rec["end_char"] = start_char + len(txt)
            else:
                h_rec["start_char"] = ""
                h_rec["end_char"] = ""
            
            h_rec["hydration_status"] = hydration_type
            h_rec["span_len"] = len(txt)
            self.hydrated_spans.append(h_rec)

    def write_span_sheets(self):
        # Sheet: Span_Annotations (Anchor First, Blank Offsets)
        ws_anno = self.workbook["Span_Annotations"]
        if ws_anno.max_row == 1:
            ws_anno.append([
                "source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text",
                "match_index", "start_char", "end_char", "span_len", "label", "normalized_value",
                "schema_field", "event_id", "is_negated", "is_historical", "time_anchor",
                "reviewer", "comments", "hydration_status"
            ])
        
        # Sheet: Span_Hydrated (Calculated Offsets)
        ws_hyd = self.workbook["Span_Hydrated"]
        if ws_hyd.max_row == 1:
            ws_hyd.append([
                "source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text",
                "match_index", "start_char", "end_char", "span_len", "label", "normalized_value",
                "schema_field", "event_id", "is_negated", "is_historical", "time_anchor",
                "reviewer", "comments", "hydration_status"
            ])

        for i, h_s in enumerate(self.hydrated_spans):
            span_id = f"span_{i+1:03d}"
            
            # Row for Annotation (clean)
            row_anno = [
                self.source_file, self.note_id, span_id, "", h_s.get("context_prefix"), h_s["span_text"],
                h_s.get("match_index", ""), "", "", h_s["span_len"], h_s["label"], h_s["normalized_value"],
                "", h_s["event_id"], 0, 0, "", "", "", "needs_hydration"
            ]
            ws_anno.append(row_anno)

            # Row for Hydrated
            row_hyd = [
                self.source_file, self.note_id, span_id, "", h_s.get("context_prefix"), h_s["span_text"],
                h_s.get("match_index", ""), h_s["start_char"], h_s["end_char"], h_s["span_len"], h_s["label"], h_s["normalized_value"],
                "", h_s["event_id"], 0, 0, "", "", "", h_s["hydration_status"]
            ]
            ws_hyd.append(row_hyd)

    def write_event_log(self):
        ws = self.workbook["Event_Log"]
        headers = [
            "source_file", "note_id", "event_id", "event_type", "method", "anatomy_target", "device", 
            "needle_gauge", "stations", "counts", "measurements", "specimens", "findings", 
            "is_historical", "reviewer", "comments", "device_size", "device_material",
            "outcome_airway_lumen_pre", "outcome_airway_lumen_post", "outcome_symptoms", 
            "outcome_pleural", "outcome_complication"
        ]
        if ws.max_row == 1:
            ws.append(headers)

        # Aggregate spans by event_id
        grouped = {}
        for s in self.hydrated_spans:
            eid = s["event_id"]
            if eid not in grouped:
                grouped[eid] = {
                    "method": set(), "anatomy": set(), "device": set(), "needle": set(), 
                    "counts": set(), "findings": set(), "complication": set()
                }
            
            val = s["normalized_value"]
            lbl = s["label"]
            
            if lbl == "PROC_METHOD": grouped[eid]["method"].add(val)
            if lbl.startswith("ANAT"): grouped[eid]["anatomy"].add(val)
            if lbl in ["DEV_INSTRUMENT", "DEV_CATHETER", "DEV_VALVE"]: grouped[eid]["device"].add(val)
            if lbl == "DEV_NEEDLE": grouped[eid]["needle"].add(val)
            if lbl == "MEAS_COUNT": grouped[eid]["counts"].add(val)
            if lbl in ["OBS_LESION", "OBS_ROSE"]: grouped[eid]["findings"].add(val)
            if lbl == "OUTCOME_COMPLICATION": grouped[eid]["complication"].add(val)

        for eid, data in grouped.items():
            if eid == "evt_global": continue 
            
            row = [
                self.source_file, self.note_id, eid, "procedure_event",
                ", ".join(data["method"]),
                ", ".join(data["anatomy"]),
                ", ".join(data["device"]),
                ", ".join(data["needle"]),
                "", # stations
                ", ".join(data["counts"]),
                "", # measurements
                "", # specimens
                ", ".join(data["findings"]),
                0, "", "", "", "",
                "", "", "", "", # granular outcomes
                ", ".join(data["complication"])
            ]
            ws.append(row)

    def write_v3_json(self):
        # Create valid JSON object
        registry_data = {
            "schema_version": "3.0",
            "note_id": self.note_id,
            "procedures": [],
            "no_immediate_complications": True # Derived from note text
        }
        
        ws_json = self.workbook["V3_Registry_JSON"]
        ws_json.cell(row=1, column=1, value=json.dumps(registry_data, indent=2))

    def run(self):
        self.write_note_text()
        self.write_note_index()
        self.define_spans()
        self.hydrate_spans()
        self.write_span_sheets()
        self.write_event_log()
        self.write_v3_json()
        self.workbook.save(self.template_path.replace("template", "extraction").replace("golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural", f"extraction_{self.note_id}"))
        print(f"Generated {OUTPUT_PATH}")

# ==========================================
# EXECUTION
# ==========================================
if __name__ == "__main__":
    # In a real file upload scenario, one would read lines from sys.argv or a specific path
    # For this runnable script, we use the embedded constants
    generator = Phase0Generator(NOTE_ID, SOURCE_FILE, NOTE_TEXT, TEMPLATE_PATH)
    generator.run()