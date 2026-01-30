import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os
from datetime import datetime

# =============================================================================
# 1. INPUT DATA & CONFIGURATION
# =============================================================================

NOTE_ID = "note_048"
SOURCE_FILE = "note_048.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_048 SOURCE_FILE: note_048.txt INDICATION FOR OPERATION:  [REDACTED]is a 65 year old-year-old male who presents with respiratory failure.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
PREOPERATIVE DIAGNOSIS: J96.90 Respiratory Failure
 
POSTOPERATIVE DIAGNOSIS:  J96.90 Respiratory Failure
 
PROCEDURE:  
31899 Unlisted Procedure (Trach Change with Mature Tract or Procedure NOS)
31645 Therapeutic aspiration initial episode
 
 
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
Initial Airway Inspection Findings:
After adequate sedation an iGel airway was placed by anesthesia and ventilation continued via the tracheostomy tube.
There was a large amount of thick, foul smelling secretions at the level of the glottis. These were therapeutically aspirated.
There was moderate edema and the true vocal cords were unable to be visualized.
There is a small shelf of tissue along the 12 oclock to 2 oclock position when the tracheostomy tube enters.
This does not cause significant obstruction. The cuff of the tracheostomy tube deflated the bronchoscope can easily pass the tracheostomy tube.
The bronchoscope was removed and introduced into the existing tracheostomy tube.
Successful therapeutic aspiration was performed to clean out the Supraglottic, Vocal Cord, Subglottic, Trachea (Proximal 1/3), Trachea (Middle 1/3), Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus and mucus plug.
Under direct visualization from the glottis the cuff was deflated and the tracheostomy tube was easily removed.
The stoma appeared widely patent and some granulation tissue was immediately visualized.
The new tracheostomy tube was then placed with obturator in place.
The obturator was removed, inner cannula was placed and the cuff inflated under direct visualization which confirmed occlusion of the entire trachea.
Percutaneous tracheostomy was changed from:
Portex cuffed Trach ISO/ID size 8.0mm suctionaid 
To 
Portex cuffed Trach ISO/ID size 8.0mm suctionaid 
 
without issue.
Pictures and videos were taken from the tracheostomy tube and from the mouth.
 
The patient tolerated the procedure well.
There were no immediate complications.  At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
 
 
IMPRESSION/PLAN: [REDACTED]is a 65 year old-year-old male who presents for bronchoscopy for respiratory failure.
-Follow up tracheostomy change in ~3 months
- DME order for Portex 8.0 inner cannulas"""

# =============================================================================
# 2. DATA EXTRACTION (Simulation)
# =============================================================================

# Procedure Flags
PROCEDURE_FLAGS = {
    'diagnostic_bronchoscopy': 0, # Primary intent was therapeutic/trach change
    'bal': 0,
    'bronchial_wash': 0,
    'brushings': 0,
    'endobronchial_biopsy': 0,
    'tbna_conventional': 0,
    'linear_ebus': 0,
    'radial_ebus': 0,
    'navigational_bronchoscopy': 0,
    'transbronchial_biopsy': 0,
    'transbronchial_cryobiopsy': 0,
    'therapeutic_aspiration': 1, # Explicitly listed and performed
    'foreign_body_removal': 0,
    'airway_dilation': 0,
    'airway_stent': 0,
    'thermal_ablation': 0,
    'tumor_debulking_non_thermal': 0,
    'cryotherapy': 0,
    'blvr': 0,
    'peripheral_ablation': 0,
    'bronchial_thermoplasty': 0,
    'whole_lung_lavage': 0,
    'rigid_bronchoscopy': 0,
    'thoracentesis': 0,
    'chest_tube': 0,
    'ipc': 0,
    'medical_thoracoscopy': 0,
    'pleurodesis': 0,
    'pleural_biopsy': 0,
    'fibrinolytic_therapy': 0
}

# Spans (Anchor-First)
# Structure: (text, label, normalized, field, context_prefix or None)
RAW_SPANS = [
    ("respiratory failure", "CTX_INDICATION", "Respiratory failure", "indication", "presents with "),
    ("Flexible Therapeutic Bronchoscope", "DEV_INSTRUMENT", "Flexible Bronchoscope", "instrument", None),
    ("iGel airway", "DEV_INSTRUMENT", "iGel airway", "instrument", None),
    ("thick, foul smelling secretions", "OBS_LESION", "Thick foul smelling secretions", "finding", None),
    ("glottis", "ANAT_AIRWAY", "Glottis", "anatomy", "level of the "),
    ("therapeutically aspirated", "PROC_METHOD", "Therapeutic Aspiration", "method", "These were "),
    ("moderate edema", "OBS_LESION", "Moderate edema", "finding", None),
    ("small shelf of tissue", "OBS_LESION", "Tissue shelf", "finding", None),
    ("tracheostomy tube", "DEV_CATHETER", "Tracheostomy tube", "device", "enters. \nThis does not cause significant obstruction. The cuff of the "),
    ("therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", "method", "Successful "),
    ("Supraglottic", "ANAT_AIRWAY", "Supraglottic", "anatomy", "clean out the "),
    ("Vocal Cord", "ANAT_AIRWAY", "Vocal Cords", "anatomy", "Supraglottic, "),
    ("Subglottic", "ANAT_AIRWAY", "Subglottic", "anatomy", "Vocal Cord, "),
    ("Trachea (Proximal 1/3)", "ANAT_AIRWAY", "Trachea Proximal", "anatomy", None),
    ("Trachea (Middle 1/3)", "ANAT_AIRWAY", "Trachea Middle", "anatomy", None),
    ("Trachea (Distal 1/3)", "ANAT_AIRWAY", "Trachea Distal", "anatomy", None),
    ("Right Mainstem", "ANAT_AIRWAY", "RMS", "anatomy", None),
    ("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "anatomy", None),
    ("Left Mainstem", "ANAT_AIRWAY", "LMS", "anatomy", None),
    ("Carina", "ANAT_AIRWAY", "Carina", "anatomy", "Left Mainstem, "),
    ("RUL Carina", "ANAT_AIRWAY", "RUL Carina", "anatomy", None),
    ("RML Carina", "ANAT_AIRWAY", "RML Carina", "anatomy", None),
    ("LUL Lingula Carina", "ANAT_AIRWAY", "LUL Lingula Carina", "anatomy", None),
    ("Left Carina", "ANAT_AIRWAY", "Left Carina", "anatomy", None),
    ("mucus and mucus plug", "OBS_LESION", "Mucus plug", "finding", None),
    ("granulation tissue", "OBS_LESION", "Granulation tissue", "finding", "some "),
    ("Portex cuffed Trach ISO/ID size 8.0mm suctionaid", "DEV_CATHETER", "Portex Trach 8.0mm", "device", "changed from:\n"),
    ("Portex cuffed Trach ISO/ID size 8.0mm suctionaid", "DEV_CATHETER", "Portex Trach 8.0mm", "device", "To \n"),
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "complication", None)
]

# Events Logic
EVENTS = [
    {
        "event_id": "evt_01",
        "type": "Therapeutic Aspiration",
        "method": "therapeutic_aspiration",
        "anatomy": "Tracheobronchial Tree",
        "findings": ["Thick foul smelling secretions", "Mucus plug"],
        "outcomes": {
            "airway_lumen_post": "Patent" # Inferred from "widely patent" stoma and successful aspiration
        }
    },
    {
        "event_id": "evt_02",
        "type": "Tracheostomy Change",
        "method": "Tracheostomy Change",
        "device": "Portex Trach 8.0mm",
        "findings": ["Granulation tissue", "Tissue shelf"],
        "outcomes": {
            "complication": "None"
        }
    }
]

# =============================================================================
# 3. HELPER FUNCTIONS
# =============================================================================

def find_offset(text, span_text, context_prefix=None, match_index=0):
    """
    Finds start/end char offsets.
    Priority:
    1. Exact single match.
    2. Context prefix match (within 120 chars).
    3. Match index (if explicitly provided and valid).
    """
    if not span_text:
        return None, None, "missing_text"
    
    # Escape regex special chars in span_text
    escaped_span = re.escape(span_text)
    
    # Find all matches
    matches = [m for m in re.finditer(escaped_span, text)]
    
    if not matches:
        return None, None, "not_found"
    
    # Logic 1: Unique match
    if len(matches) == 1:
        m = matches[0]
        return m.start(), m.end(), "hydrated_unique"
    
    # Logic 2: Context prefix
    if context_prefix:
        # Normalize spaces in context
        clean_ctx = " ".join(context_prefix.split())
        for m in matches:
            # Look back up to 120 chars
            start_search = max(0, m.start() - 120)
            preceding_text = text[start_search:m.start()]
            # Normalize preceding for loose matching
            clean_preceding = " ".join(preceding_text.split())
            if clean_ctx in clean_preceding:
                return m.start(), m.end(), "hydrated_prefix_window"
    
    # Logic 3: Match Index
    if 0 <= match_index < len(matches):
        m = matches[match_index]
        return m.start(), m.end(), "hydrated_match_index"
    
    return None, None, f"ambiguous_count={len(matches)}"

def create_workbook():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        # Create a basic workbook if template missing (fallback, though template required)
        wb = openpyxl.Workbook()
        for sheet in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            if sheet not in wb.sheetnames:
                wb.create_sheet(sheet)
    return wb

# =============================================================================
# 4. MAIN GENERATION LOGIC
# =============================================================================

def main():
    wb = create_workbook()
    
    # ---------------------------------------------------------
    # Sheet 1: Note_Text
    # ---------------------------------------------------------
    ws_text = wb["Note_Text"]
    # Assuming headers exist: note_id, source_file, note_text
    # Append data
    next_row = ws_text.max_row + 1
    ws_text.cell(row=next_row, column=1, value=NOTE_ID)
    ws_text.cell(row=next_row, column=2, value=SOURCE_FILE)
    ws_text.cell(row=next_row, column=3, value=NOTE_TEXT)

    # ---------------------------------------------------------
    # Sheet 2: Note_Index
    # ---------------------------------------------------------
    ws_index = wb["Note_Index"]
    # Row 2 (assuming row 1 is header)
    r_idx = ws_index.max_row + 1
    
    # Metadata columns (A-H)
    metadata = [
        SOURCE_FILE, NOTE_ID, "", "", "", "", "success", ""
    ]
    for col, val in enumerate(metadata, 1):
        ws_index.cell(row=r_idx, column=col, value=val)
    
    # Flags columns (starts at I=9)
    # Order matches the keys in PROCEDURE_FLAGS
    flag_keys = [
        'diagnostic_bronchoscopy', 'bal', 'bronchial_wash', 'brushings', 'endobronchial_biopsy', 
        'tbna_conventional', 'linear_ebus', 'radial_ebus', 'navigational_bronchoscopy', 
        'transbronchial_biopsy', 'transbronchial_cryobiopsy', 'therapeutic_aspiration', 
        'foreign_body_removal', 'airway_dilation', 'airway_stent', 'thermal_ablation', 
        'tumor_debulking_non_thermal', 'cryotherapy', 'blvr', 'peripheral_ablation', 
        'bronchial_thermoplasty', 'whole_lung_lavage', 'rigid_bronchoscopy', 
        'thoracentesis', 'chest_tube', 'ipc', 'medical_thoracoscopy', 'pleurodesis', 
        'pleural_biopsy', 'fibrinolytic_therapy'
    ]
    
    for i, key in enumerate(flag_keys):
        ws_index.cell(row=r_idx, column=9+i, value=PROCEDURE_FLAGS.get(key, 0))

    # ---------------------------------------------------------
    # Sheet 3 & 4: Span_Annotations & Span_Hydrated
    # ---------------------------------------------------------
    ws_anno = wb["Span_Annotations"]
    ws_hydra = wb["Span_Hydrated"]
    
    # Headers match logic, just starting filling from row 2
    row_num = ws_anno.max_row + 1
    
    # Span ID counter
    span_counter = 1
    
    # Process extracted spans
    # RAW_SPANS tuple: (text, label, normalized, field, context_prefix)
    
    hydrated_rows = []
    
    for txt, label, norm, field, ctx in RAW_SPANS:
        # Generate Span ID
        span_id = f"{NOTE_ID}_s{span_counter:02d}"
        
        # Hydrate
        start, end, status = find_offset(NOTE_TEXT, txt, context_prefix=ctx)
        
        # Common data
        row_data = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure Note", # section type dummy
            ctx, txt, None, # match_index left blank unless we used logic 3 explicitly (here we used ctx mostly)
            None, None, f"=LEN(F{row_num})", # start/end blank for Anno, formula for len
            label, norm, field, "", # event_id blank
            0, 0, 0, "gpt-4-extracted", "", "needs_hydration"
        ]
        
        # Write to Span_Annotations
        for c, val in enumerate(row_data, 1):
            ws_anno.cell(row=row_num, column=c, value=val)
            
        # Prepare Hydrated Row
        h_row_data = list(row_data)
        h_row_data[7] = start if start is not None else "" # Start Char
        h_row_data[8] = end if end is not None else ""     # End Char
        h_row_data[19] = status
        
        # Write to Span_Hydrated
        # Hydrated sheet usually mirrors structure but has start/end filled
        # Assuming identical columns except start/end at indices 7,8 (H, I)
        # Re-mapping based on template structure:
        # A:Source, B:Note, C:SpanID, D:Section, E:Ctx, F:Text, G:MatchIdx, H:Start, I:End, J:Len...
        
        hyd_row_idx = ws_hydra.max_row + 1
        for c, val in enumerate(h_row_data, 1):
            ws_hydra.cell(row=hyd_row_idx, column=c, value=val)
        
        span_counter += 1
        row_num += 1

    # ---------------------------------------------------------
    # Sheet 5: Event_Log
    # ---------------------------------------------------------
    ws_event = wb["Event_Log"]
    ev_row = ws_event.max_row + 1
    
    for evt in EVENTS:
        # Mapping dict keys to columns loosely
        # Columns: source, note_id, event_id, event_type, method, anatomy, device...
        ws_event.cell(row=ev_row, column=1, value=SOURCE_FILE)
        ws_event.cell(row=ev_row, column=2, value=NOTE_ID)
        ws_event.cell(row=ev_row, column=3, value=evt["event_id"])
        ws_event.cell(row=ev_row, column=4, value=evt["type"])
        ws_event.cell(row=ev_row, column=5, value=evt["method"])
        ws_event.cell(row=ev_row, column=6, value=evt.get("anatomy", ""))
        ws_event.cell(row=ev_row, column=7, value=evt.get("device", ""))
        
        # Findings to string
        findings_str = ", ".join(evt.get("findings", []))
        ws_event.cell(row=ev_row, column=13, value=findings_str)
        
        # Outcomes
        outcomes = evt.get("outcomes", {})
        ws_event.cell(row=ev_row, column=20, value=outcomes.get("airway_lumen_post", ""))
        ws_event.cell(row=ev_row, column=23, value=outcomes.get("complication", ""))
        
        ev_row += 1

    # ---------------------------------------------------------
    # Sheet 6: V3_Procedure_Events
    # ---------------------------------------------------------
    ws_v3 = wb["V3_Procedure_Events"]
    v3_row = ws_v3.max_row + 1
    
    for evt in EVENTS:
        ws_v3.cell(row=v3_row, column=1, value=NOTE_ID)
        ws_v3.cell(row=v3_row, column=2, value=evt["event_id"])
        ws_v3.cell(row=v3_row, column=3, value=evt["type"])
        
        # Anatomy split logic (simplified for script)
        if "anatomy" in evt:
            ws_v3.cell(row=v3_row, column=4, value="Airway") # target.anatomy_type
        
        ws_v3.cell(row=v3_row, column=10, value=evt["method"])
        
        # Outcomes
        outcomes = evt.get("outcomes", {})
        if "airway_lumen_post" in outcomes:
            ws_v3.cell(row=v3_row, column=19, value=outcomes["airway_lumen_post"])
        if "complication" in outcomes:
            ws_v3.cell(row=v3_row, column=22, value=outcomes["complication"])
            
        v3_row += 1

    # ---------------------------------------------------------
    # Sheet 7: V3_Registry_JSON
    # ---------------------------------------------------------
    ws_json = wb["V3_Registry_JSON"]
    
    registry_data = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "procedures": EVENTS,
        "no_immediate_complications": True
    }
    
    json_str = json.dumps(registry_data, indent=2)
    ws_json.cell(row=1, column=1, value=json_str)

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    main()