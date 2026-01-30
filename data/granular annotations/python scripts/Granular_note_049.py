import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import json
import re
import datetime
import os

# ==============================================================================
# 1. INPUTS & CONFIGURATION
# ==============================================================================
NOTE_ID = "note_049"
SOURCE_FILE = "note_049.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# Full text from the provided source
NOTE_TEXT = """NOTE_ID:  note_049 SOURCE_FILE: note_049.txt INDICATION FOR OPERATION:  [REDACTED]is a 22 year old-year-old female who presents with airway stenosis.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
 
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
 
PROCEDURE:  
31646 Therapeutic aspiration subsequent episodes
31624 Dx bronchoscope/lavage (BAL)    
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Disposable Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
 
The laryngeal mask airway is in good position.
Pharynx: Not assessed due to bronchoscopy introduction through LMA.
Larynx: Normal.
Vocal Cords: Normal without mass/lesions
Trachea: Mildly tortuous
Main Carina: Sharp
Right Lung Proximal Airways: Normal anatomic branching to segmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Left Lung Proximal Airways: The distal left mainstem bronchus was noted to be slightly stenotic (~8mm).
Otherwise, there was normal anatomic branching to segmental level in the LLL, noting historic obliteration of LUL.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Mucosa: Normal.
Secretions: Minimal, thin, and clear.
DISTAL LEFT MAINSTEM BRONCHUS
 ]
Successful therapeutic aspiration was performed to clean out the Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus.
On PEEP 0, the patient's airways were noted to have slight malacia <50% at the proximal and distal left mainstem bronchus.
MALACIA AT PROXIMAL LMSB

MALACIA AT DISTAL LMSB

DISTAL LMSB WITH AND WITHOUT PEEP (left and right, respectively)
 
Bronchial alveolar lavage was performed at Anteromedial Segment of LLL (Lb7/8), Lateral-basal Segment of LLL (LB9), and Posterior-Basal Segment of LLL (LB10).
Instilled 60 cc of NS, suction returned with 20 cc of NS.  Samples sent for Cell Count and Microbiology (Cultures/Viral/Fungal).
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
--LLL BAL (cell count, micro)
 
IMPRESSION/PLAN: [REDACTED]is a 22 year old-year-old female who presents for bronchoscopy for evaluation of airway stenosis.
The patient tolerated the procedure well and there were no immediate complications.
--Follow up BAL results
--Repeat bronchoscopy in 4 weeks for possible dilation, stenting"""

# ==============================================================================
# 2. PROCEDURE FLAGS & DEFINITIONS
# ==============================================================================

# 1 = Performed, 0 = Not Performed
PROCEDURE_FLAGS = {
    # Bronchoscopy (23)
    "diagnostic_bronchoscopy": 1,
    "bal": 1,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 0,
    "airway_dilation": 0, # Planned for future, not done here
    "airway_stent": 0, # Planned for future
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    
    # Pleural (7)
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# ==============================================================================
# 3. SPAN DEFINITIONS (ANCHOR-FIRST)
# ==============================================================================
# Structure: (span_text, label, normalized_value, event_id, context_prefix, match_index)
# context_prefix and match_index can be None.

RAW_SPANS = [
    # Event 1: Diagnostic Inspection / Initial Findings
    ("Dx bronchoscope/lavage", "PROC_METHOD", "Diagnostic Bronchoscopy", "evt_dx", None, None),
    ("distal left mainstem bronchus", "ANAT_AIRWAY", "Left Mainstem Bronchus (Distal)", "evt_dx", None, None),
    ("slightly stenotic (~8mm)", "OBS_LESION", "Stenosis (8mm)", "evt_dx", None, None),
    ("slight malacia <50%", "OBS_LESION", "Malacia (<50%)", "evt_dx", None, None),
    ("proximal and distal left mainstem bronchus", "ANAT_AIRWAY", "Left Mainstem Bronchus", "evt_dx", None, None),
    ("Disposable Bronchoscope", "DEV_INSTRUMENT", "Disposable Bronchoscope", "evt_dx", None, None),

    # Event 2: Therapeutic Aspiration
    ("Therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", "evt_asp", None, None),
    ("clean out... from mucus", "PROC_ACTION", "Mucus Removal", "evt_asp", None, None),
    
    # Locations for aspiration
    ("Trachea (Distal 1/3)", "ANAT_AIRWAY", "Trachea (Distal 1/3)", "evt_asp", None, None),
    ("Right Mainstem", "ANAT_AIRWAY", "Right Mainstem Bronchus", "evt_asp", None, None),
    ("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "evt_asp", None, None),
    ("Left Mainstem", "ANAT_AIRWAY", "Left Mainstem Bronchus", "evt_asp", "Bronchus Intermedius , ", None),
    ("Carina", "ANAT_AIRWAY", "Main Carina", "evt_asp", "Left Mainstem, ", None), # Context: "Left Mainstem, Carina"
    ("RUL Carina (RC1)", "ANAT_AIRWAY", "RUL Carina", "evt_asp", None, None),
    ("RML Carina (RC2)", "ANAT_AIRWAY", "RML Carina", "evt_asp", None, None),
    ("LUL Lingula Carina (Lc1)", "ANAT_AIRWAY", "Lingula Carina", "evt_asp", None, None),
    ("Left Carina (LC2)", "ANAT_AIRWAY", "Left Upper Lobe Carina", "evt_asp", None, None),

    # Event 3: BAL
    ("Bronchial alveolar lavage", "PROC_METHOD", "BAL", "evt_bal", None, None),
    ("Anteromedial Segment of LLL (Lb7/8)", "ANAT_LUNG_LOC", "LLL Anteromedial Segment", "evt_bal", None, None),
    ("Lateral-basal Segment of LLL (LB9)", "ANAT_LUNG_LOC", "LLL Lateral-basal Segment", "evt_bal", None, None),
    ("Posterior-Basal Segment of LLL (LB10)", "ANAT_LUNG_LOC", "LLL Posterior-basal Segment", "evt_bal", None, None),
    ("Instilled 60 cc", "MEAS_VOL", "60 ml", "evt_bal", None, None),
    ("returned with 20 cc", "MEAS_VOL", "20 ml", "evt_bal", None, None),
    ("Samples sent for Cell Count", "PROC_ACTION", "Sample for Cell Count", "evt_bal", None, None),

    # Outcomes/Complications
    ("no immediate complications", "OUTCOME_COMPLICATION", "None", "evt_global", "tolerated the procedure well.  There were ", None),
    ("tolerated the procedure well", "OUTCOME_SYMPTOMS", "Tolerated Well", "evt_global", "The patient ", 0), # Match index 0 (first occurrence)
]

# ==============================================================================
# 4. CLASSES & HELPERS
# ==============================================================================

class Span:
    def __init__(self, span_text, label, normalized, event_id, context=None, match_index=None):
        self.span_text = span_text
        self.label = label
        self.normalized = normalized
        self.event_id = event_id
        self.context = context
        self.match_index = match_index
        # Calculated
        self.start_char = None
        self.end_char = None
        self.hydration_status = "pending"

def find_offsets(text, span_text, context_prefix=None, match_index=None):
    """
    Hydrates offsets based on search logic:
    1. If unique -> return.
    2. If context_prefix -> search window before.
    3. If match_index -> return specific occurrence.
    4. Else -> ambiguous.
    """
    escaped_text = re.escape(span_text)
    # Allow loose matching for whitespace
    pattern = escaped_text.replace(r'\ ', r'\s+')
    matches = list(re.finditer(pattern, text, re.IGNORECASE))
    
    if not matches:
        return None, None, "not_found"

    # Case 1: Unique
    if len(matches) == 1:
        m = matches[0]
        return m.start(), m.end(), "hydrated_unique"

    # Case 2: Context Prefix
    if context_prefix:
        # Check 120 chars before each match
        for m in matches:
            start = m.start()
            window_start = max(0, start - 120)
            window_text = text[window_start:start]
            # Simple check: is context string inside the window?
            # Normalize whitespace for check
            if "..." in context_prefix:
                # Fuzzy context not implemented fully, strip "..."
                clean_ctx = context_prefix.replace("...", "")
            else:
                clean_ctx = context_prefix
            
            if clean_ctx.lower() in window_text.lower():
                return m.start(), m.end(), "hydrated_prefix_window"

    # Case 3: Match Index
    if match_index is not None and 0 <= match_index < len(matches):
        m = matches[match_index]
        return m.start(), m.end(), "hydrated_match_index"

    return None, None, f"ambiguous_count={len(matches)}"

def create_workbook():
    if os.path.exists(TEMPLATE_PATH):
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    else:
        # Fallback if template missing (should not happen per prompt, but for safety)
        wb = openpyxl.Workbook()
    return wb

def get_or_create_sheet(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return wb.create_sheet(sheet_name)

# ==============================================================================
# 5. MAIN GENERATION LOGIC
# ==============================================================================

def main():
    print(f"Generating {OUTPUT_PATH}...")
    
    # 1. Hydrate Spans
    spans = []
    for raw in RAW_SPANS:
        s = Span(*raw)
        start, end, status = find_offsets(NOTE_TEXT, s.span_text, s.context, s.match_index)
        s.start_char = start
        s.end_char = end
        s.hydration_status = status
        spans.append(s)

    # 2. Load Workbook
    wb = create_workbook()
    
    # 3. Populate Note_Text
    ws_text = get_or_create_sheet(wb, "Note_Text")
    # Headers usually: note_id, source_file, note_text
    if ws_text.max_row == 1:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 4. Populate Note_Index
    ws_index = get_or_create_sheet(wb, "Note_Index")
    # Expected Headers: source_file, note_id, encounter_id, procedure_date, site, reviewer, status, free_text_notes, [flags...]
    # We will assume row 1 is headers. We append to row 2.
    
    # Construct row data
    index_row = [
        SOURCE_FILE, NOTE_ID, "", "", "", "", "Pending", ""
    ]
    # Add flags in specific order (assumed based on prompt listing)
    flag_keys = list(PROCEDURE_FLAGS.keys())
    for key in flag_keys:
        index_row.append(PROCEDURE_FLAGS[key])
    
    ws_index.append(index_row)

    # 5. Populate Span_Annotations (Anchor First)
    ws_anno = get_or_create_sheet(wb, "Span_Annotations")
    # Header: source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, start_char, end_char, span_len, label, normalized_value, schema_field, event_id, is_negated, is_historical, time_anchor, reviewer, comments, hydration_status
    
    # Write Header if empty
    headers = ["source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", "match_index", "start_char", "end_char", "span_len", "label", "normalized_value", "schema_field", "event_id", "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"]
    if ws_anno.max_row == 0 or ws_anno.cell(1,1).value is None:
        ws_anno.append(headers)

    for i, s in enumerate(spans):
        row = [
            SOURCE_FILE, NOTE_ID, f"span_{i+1:03d}", "Procedure Note", 
            s.context, s.span_text, s.match_index, 
            "", "", f"=LEN(F{i+2})", # start/end blank, formula for len
            s.label, s.normalized, "", s.event_id,
            "FALSE", "FALSE", "", "", "", "needs_hydration"
        ]
        ws_anno.append(row)

    # 6. Populate Span_Hydrated
    ws_hydro = get_or_create_sheet(wb, "Span_Hydrated")
    if ws_hydro.max_row == 0 or ws_hydro.cell(1,1).value is None:
        ws_hydro.append(headers)

    for i, s in enumerate(spans):
        span_len = (s.end_char - s.start_char) if s.start_char is not None else 0
        row = [
            SOURCE_FILE, NOTE_ID, f"span_{i+1:03d}", "Procedure Note", 
            s.context, s.span_text, s.match_index, 
            s.start_char, s.end_char, span_len,
            s.label, s.normalized, "", s.event_id,
            "FALSE", "FALSE", "", "", "", s.hydration_status
        ]
        ws_hydro.append(row)

    # 7. Populate Event_Log
    ws_event = get_or_create_sheet(wb, "Event_Log")
    event_headers = [
        "source_file", "note_id", "event_id", "event_type", "method", 
        "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements", 
        "specimens", "findings", "is_historical", "reviewer", "comments",
        "device_size", "device_material", 
        "outcome_airway_lumen_pre", "outcome_airway_lumen_post", 
        "outcome_symptoms", "outcome_pleural", "outcome_complication"
    ]
    if ws_event.max_row == 0 or ws_event.cell(1,1).value is None:
        ws_event.append(event_headers)

    # Aggregating data for events
    events_data = {
        "evt_dx": {
            "type": "Diagnostic Bronchoscopy", "method": "Bronchoscopy", "anatomy": [], "findings": ["Stenosis", "Malacia"], 
            "outcomes": {}
        },
        "evt_asp": {
            "type": "Therapeutic", "method": "Therapeutic Aspiration", 
            "anatomy": ["Trachea", "Right Mainstem", "Bronchus Intermedius", "Left Mainstem", "Carina", "RUL", "RML", "Lingula", "LUL"],
            "findings": ["Mucus"], "outcomes": {}
        },
        "evt_bal": {
            "type": "Diagnostic", "method": "BAL", 
            "anatomy": ["LLL Anteromedial", "LLL Lateral-basal", "LLL Posterior-basal"],
            "measurements": ["Instilled 60ml", "Return 20ml"], "specimens": ["Cell Count"],
            "outcomes": {}
        },
        "evt_global": {
            "type": "Outcome", "method": "", "anatomy": [], "outcomes": {
                "complication": "None", "symptoms": "Tolerated Well"
            }
        }
    }

    for eid, data in events_data.items():
        row = [
            SOURCE_FILE, NOTE_ID, eid, data.get("type"), data.get("method"),
            ", ".join(data.get("anatomy", [])), "", "", "", "", 
            ", ".join(data.get("measurements", [])),
            ", ".join(data.get("specimens", [])),
            ", ".join(data.get("findings", [])),
            "FALSE", "", "", "", "",
            "", "", # lumen pre/post
            data["outcomes"].get("symptoms", ""),
            "",
            data["outcomes"].get("complication", "")
        ]
        ws_event.append(row)

    # 8. Populate V3_Procedure_Events
    ws_v3 = get_or_create_sheet(wb, "V3_Procedure_Events")
    v3_headers = [
        "note_id", "event_id", "type", 
        "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station",
        "lesion.type", "lesion.size_mm",
        "method", "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote",
        "stent.size", "stent.material_or_brand", "catheter.size_fr",
        "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post",
        "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"
    ]
    if ws_v3.max_row == 0 or ws_v3.cell(1,1).value is None:
        ws_v3.append(v3_headers)
    
    # Simple mapping for V3 (best effort based on complexity)
    # Event: Therapeutic Aspiration
    ws_v3.append([
        NOTE_ID, "evt_asp", "Therapeutic",
        "Airway", "Multilobar", "Multiple", "",
        "", "",
        "Therapeutic Aspiration", "[]", "[]", "[]", '["Mucus"]', "clean out... from mucus",
        "", "", "",
        "", "", "Tolerated Well", "", "None"
    ])
    # Event: BAL
    ws_v3.append([
        NOTE_ID, "evt_bal", "Diagnostic",
        "Lung", "LLL", "Segments 7,8,9,10", "",
        "", "",
        "BAL", "[]", '["Instilled 60ml", "Return 20ml"]', '["Cell Count", "Micro"]', "[]", "Bronchial alveolar lavage...",
        "", "", "",
        "", "", "", "", ""
    ])

    # 9. V3_Registry_JSON
    ws_json = get_or_create_sheet(wb, "V3_Registry_JSON")
    registry_obj = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": [
            {"event_id": "evt_asp", "method": "Therapeutic Aspiration", "sites": ["Trachea", "LMSB", "RMSB", "BI"]},
            {"event_id": "evt_bal", "method": "BAL", "sites": ["LLL"]}
        ]
    }
    ws_json.cell(1, 1).value = json.dumps(registry_obj, indent=2)

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Done. Saved to {OUTPUT_PATH}")

if __name__ == "__main__":
    main()