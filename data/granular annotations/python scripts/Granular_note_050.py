import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_050"
SOURCE_FILE = "note_050.txt"
PROCEDURE_DATE = "2026-01-12"  # Using current date as placeholder or blank if preferred
NOTE_TEXT = """NOTE_ID:  note_050 SOURCE_FILE: note_050.txt INDICATION FOR OPERATION:  [REDACTED]is a 22 year old-year-old female who presents with airway stenosis.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
 
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
 
PROCEDURE:  
31646 Therapeutic aspiration subsequent episodes
31899NFN BRONCHOSCOPY WITH ENDOBRONCHIAL ULTRASOUND (EBUS) OF MEDIASTINAL AND/OR HILAR LYMPH NODES WITHOUT BIOPSY
31630 Balloon dilation
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
 
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
Linear EBUS 
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
 
The laryngeal mask airway is in good position.
Pharynx: Not assessed due to bronchoscopy introduction through LMA.
Larynx: Normal.
Vocal Cords: Normal without mass/lesions
Trachea: Mildly tortuous
Main Carina: Sharp
Right Lung Proximal Airways: Normal anatomic branching to segmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Left Lung Proximal Airways: The distal left mainstem bronchus was noted to be slightly stenotic (~8mm).
Otherwise, there was normal anatomic branching to segmental level in the LLL, noting historic obliteration of LUL.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Mucosa: Normal.
Secretions: Minimal, thin, and clear.
Successful therapeutic aspiration was performed to clean out the Right Mainstem, Bronchus Intermedius , and Left Mainstem from mucus.
LMS stenosis was dilated with  10/11/12 Elation balloon was used to perform dilation to 10 mm at the Left Mainstem.
Total 1 inflations with dilation time of 60 seconds each.
 
LUL was completely fused.
Linear ebus was used to identify the area where the prior airway was.
Needle knife was used to cut into the airway.  Serial needle 25G, 22G and 19G was used to reenforce the pathway.
Mustang balloon was used to engage the airway but was only able to pass into the airway for 1/4 of the balloon unable to dilate open the airway.
This was stopped given the inability to completely engage the LUL passage. 
 
The patient tolerated the procedure well.
There were no immediate complications.  At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
 
 
IMPRESSION/PLAN: [REDACTED]is a 22 year old-year-old female who presents for bronchoscopy for evaluation of airway stenosis.
The patient tolerated the procedure well and there were no immediate complications.
 
--repeat bronch in 8 weeks"""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# -------------------------------------------------------------------------
# CONFIGURATION
# -------------------------------------------------------------------------
# Procedure Flags Logic
# Map specific keywords or phrases to the 30 flags.
PROC_FLAGS_MAP = {
    "diagnostic_bronchoscopy": True, # Implied by inspection
    "bal": False,
    "bronchial_wash": False,
    "brushings": False,
    "endobronchial_biopsy": False,
    "tbna_conventional": False,
    "linear_ebus": True,  # "Linear EBUS"
    "radial_ebus": False,
    "navigational_bronchoscopy": False,
    "transbronchial_biopsy": False,
    "transbronchial_cryobiopsy": False,
    "therapeutic_aspiration": True, # "Therapeutic aspiration"
    "foreign_body_removal": False,
    "airway_dilation": True, # "Balloon dilation"
    "airway_stent": False,
    "thermal_ablation": False,
    "tumor_debulking_non_thermal": False,
    "cryotherapy": False,
    "blvr": False,
    "peripheral_ablation": False,
    "bronchial_thermoplasty": False,
    "whole_lung_lavage": False,
    "rigid_bronchoscopy": False,
    "thoracentesis": False,
    "chest_tube": False,
    "ipc": False,
    "medical_thoracoscopy": False,
    "pleurodesis": False,
    "pleural_biopsy": False,
    "fibrinolytic_therapy": False
}

# -------------------------------------------------------------------------
# DATA EXTRACTION LOGIC
# -------------------------------------------------------------------------

def extract_spans_and_events(text):
    spans = []
    events = []
    
    # --- Event 1: Inspection & Diagnosis (LMS Stenosis) ---
    ev1_id = "evt_01"
    spans.append({
        "span_text": "distal left mainstem bronchus",
        "label": "ANAT_AIRWAY",
        "norm": "Left Mainstem Bronchus",
        "event_id": ev1_id,
        "context": "Left Lung Proximal Airways: The"
    })
    spans.append({
        "span_text": "stenotic",
        "label": "OBS_LESION",
        "norm": "Stenosis",
        "event_id": ev1_id,
        "context": "noted to be slightly"
    })
    spans.append({
        "span_text": "~8mm",
        "label": "OUTCOME_AIRWAY_LUMEN_PRE",
        "norm": "8mm",
        "event_id": ev1_id,
        "context": "slightly stenotic ("
    })
    
    # --- Event 2: Therapeutic Aspiration ---
    ev2_id = "evt_02"
    spans.append({
        "span_text": "Therapeutic aspiration",
        "label": "PROC_METHOD",
        "norm": "Therapeutic Aspiration",
        "event_id": ev2_id,
        "context": "Successful"
    })
    spans.append({
        "span_text": "Right Mainstem",
        "label": "ANAT_AIRWAY",
        "norm": "Right Mainstem Bronchus",
        "event_id": ev2_id,
        "context": "clean out the"
    })
    spans.append({
        "span_text": "Bronchus Intermedius",
        "label": "ANAT_AIRWAY",
        "norm": "Bronchus Intermedius",
        "event_id": ev2_id,
        "context": "Right Mainstem,"
    })
    spans.append({
        "span_text": "Left Mainstem",
        "label": "ANAT_AIRWAY",
        "norm": "Left Mainstem Bronchus",
        "event_id": ev2_id,
        "context": "Bronchus Intermedius , and"
    })
    
    events.append({
        "event_id": ev2_id,
        "type": "Therapeutic Aspiration",
        "method": "Suction",
        "anatomy": "RMS, BI, LMS",
        "outcome_symptoms": "N/A"
    })

    # --- Event 3: Dilation of LMS ---
    ev3_id = "evt_03"
    spans.append({
        "span_text": "LMS",
        "label": "ANAT_AIRWAY",
        "norm": "Left Mainstem Bronchus",
        "event_id": ev3_id,
        "context": "clean out" # This is tricky context for "LMS" at start of sentence? No, "LMS stenosis was dilated"
    })
    spans.append({
        "span_text": "stenosis",
        "label": "OBS_LESION",
        "norm": "Stenosis",
        "event_id": ev3_id,
        "context": "LMS"
    })
    spans.append({
        "span_text": "dilated",
        "label": "PROC_ACTION",
        "norm": "Dilation",
        "event_id": ev3_id,
        "context": "was"
    })
    spans.append({
        "span_text": "10/11/12 Elation balloon",
        "label": "DEV_INSTRUMENT",
        "norm": "Elation Balloon 10-11-12mm",
        "event_id": ev3_id,
        "context": "with"
    })
    spans.append({
        "span_text": "10 mm",
        "label": "OUTCOME_AIRWAY_LUMEN_POST",
        "norm": "10mm",
        "event_id": ev3_id,
        "context": "dilation to"
    })
    
    events.append({
        "event_id": ev3_id,
        "type": "Airway Dilation",
        "method": "Balloon Dilation",
        "anatomy": "Left Mainstem Bronchus",
        "device": "Elation Balloon",
        "outcome_lumen_pre": "8mm", # inferred from inspection
        "outcome_lumen_post": "10mm"
    })

    # --- Event 4: Recanalization Attempt LUL ---
    ev4_id = "evt_04"
    spans.append({
        "span_text": "LUL",
        "label": "ANAT_AIRWAY",
        "norm": "Left Upper Lobe",
        "event_id": ev4_id,
        "context": "Total 1 inflations" # Context is previous sentence ending
    })
    spans.append({
        "span_text": "completely fused",
        "label": "OBS_LESION",
        "norm": "Complete Obstruction",
        "event_id": ev4_id,
        "context": "LUL was"
    })
    spans.append({
        "span_text": "Linear ebus",
        "label": "DEV_INSTRUMENT",
        "norm": "Linear EBUS Scope",
        "event_id": ev4_id,
        "context": "fused."
    })
    spans.append({
        "span_text": "Needle knife",
        "label": "DEV_INSTRUMENT",
        "norm": "Needle Knife",
        "event_id": ev4_id,
        "context": "airway was."
    })
    spans.append({
        "span_text": "cut",
        "label": "PROC_ACTION",
        "norm": "Cut",
        "event_id": ev4_id,
        "context": "used to"
    })
    spans.append({
        "span_text": "25G",
        "label": "DEV_NEEDLE",
        "norm": "25G",
        "event_id": ev4_id,
        "context": "Serial needle"
    })
    spans.append({
        "span_text": "22G",
        "label": "DEV_NEEDLE",
        "norm": "22G",
        "event_id": ev4_id,
        "context": "25G,"
    })
    spans.append({
        "span_text": "19G",
        "label": "DEV_NEEDLE",
        "norm": "19G",
        "event_id": ev4_id,
        "context": "and"
    })
    spans.append({
        "span_text": "Mustang balloon",
        "label": "DEV_INSTRUMENT",
        "norm": "Mustang Balloon",
        "event_id": ev4_id,
        "context": "pathway."
    })
    spans.append({
        "span_text": "unable to dilate open",
        "label": "OUTCOME_AIRWAY_LUMEN_POST", # Failure implies no change or still obstructed
        "norm": "Failure",
        "event_id": ev4_id,
        "context": "balloon"
    })

    events.append({
        "event_id": ev4_id,
        "type": "Recanalization/Dilation",
        "method": "Needle Knife / Balloon",
        "anatomy": "Left Upper Lobe",
        "device": "Needle Knife, Needles, Mustang Balloon",
        "findings": "Completely fused",
        "outcome_lumen_post": "Unable to dilate"
    })

    # --- Event 5: Global Outcome ---
    ev5_id = "evt_05"
    spans.append({
        "span_text": "No immediate complications",
        "label": "OUTCOME_COMPLICATION",
        "norm": "None",
        "event_id": ev5_id,
        "context": "procedure well."
    })

    events.append({
        "event_id": ev5_id,
        "type": "Global Outcome",
        "outcome_complication": "None"
    })

    return spans, events

# -------------------------------------------------------------------------
# HYDRATION LOGIC
# -------------------------------------------------------------------------

def hydrate_offsets(text, spans):
    """
    Computes start_char and end_char for each span based on logic:
    1. Exact unique match.
    2. Context prefix match (120 chars).
    3. Match index fallback.
    """
    hydrated_spans = []
    
    for s in spans:
        target = s['span_text']
        matches = [m.start() for m in re.finditer(re.escape(target), text)]
        
        start_char = None
        status = "ambiguous"
        
        if len(matches) == 1:
            start_char = matches[0]
            status = "hydrated_unique"
        elif len(matches) > 1:
            # Try context
            if 'context' in s and s['context']:
                ctx = s['context']
                # Search for context + target proximity
                best_m = None
                for m in matches:
                    # Look back 120 chars
                    window_start = max(0, m - 120)
                    window_text = text[window_start:m]
                    if ctx in window_text:
                        best_m = m
                        break
                if best_m is not None:
                    start_char = best_m
                    status = "hydrated_prefix_window"
                else:
                    # Fallback to order if we assume sequential extraction or add a match_index manually
                    # For this script, we'll default to first match if context fails, but flag it
                    start_char = matches[0] 
                    status = f"ambiguous_count={len(matches)}_default_first"
            else:
                start_char = matches[0]
                status = f"ambiguous_count={len(matches)}_default_first"
        else:
            status = "not_found"
            
        if start_char is not None:
            s['start_char'] = start_char
            s['end_char'] = start_char + len(target)
            s['hydration_status'] = status
        else:
            s['start_char'] = ""
            s['end_char'] = ""
            s['hydration_status'] = status
            
        hydrated_spans.append(s)
        
    return hydrated_spans

# -------------------------------------------------------------------------
# MAIN EXECUTION
# -------------------------------------------------------------------------

def create_workbook():
    # 1. Prepare Data
    raw_spans, raw_events = extract_spans_and_events(NOTE_TEXT)
    hydrated_spans = hydrate_offsets(NOTE_TEXT, raw_spans)
    
    # 2. Load Template
    if not os.path.exists(TEMPLATE_PATH):
        print(f"Template not found at {TEMPLATE_PATH}. Please ensure file exists.")
        return

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    
    # 3. Populate Sheets
    
    # --- Sheet: Note_Text ---
    if "Note_Text" not in wb.sheetnames:
        wb.create_sheet("Note_Text")
    ws_text = wb["Note_Text"]
    # Append row: note_id, source_file, note_text
    # Assuming header exists, find next empty row
    row_idx = ws_text.max_row + 1
    ws_text.cell(row=row_idx, column=1, value=NOTE_ID)
    ws_text.cell(row=row_idx, column=2, value=SOURCE_FILE)
    ws_text.cell(row=row_idx, column=3, value=NOTE_TEXT)

    # --- Sheet: Note_Index ---
    if "Note_Index" not in wb.sheetnames:
        wb.create_sheet("Note_Index")
    ws_idx = wb["Note_Index"]
    row_idx = ws_idx.max_row + 1
    
    # Metadata columns (A-H)
    metadata = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Success", ""]
    for col, val in enumerate(metadata, 1):
        ws_idx.cell(row=row_idx, column=col, value=val)
        
    # Procedure Flags (Starting column I, assuming template order matches prompt list)
    # The prompt lists 30 flags. We map them to columns I onwards.
    # We must ensure the order matches the template usually, but here we just append in order of the prompt list.
    flag_keys = list(PROC_FLAGS_MAP.keys())
    for i, key in enumerate(flag_keys):
        val = 1 if PROC_FLAGS_MAP[key] else 0
        ws_idx.cell(row=row_idx, column=9 + i, value=val)

    # --- Sheet: Span_Annotations ---
    if "Span_Annotations" not in wb.sheetnames:
        wb.create_sheet("Span_Annotations")
    ws_span = wb["Span_Annotations"]
    
    # Columns: source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, start_char, end_char, span_len, label, normalized_value, schema_field, event_id, is_negated, is_historical, time_anchor, reviewer, comments, hydration_status
    
    row_num = ws_span.max_row + 1
    for i, s in enumerate(raw_spans):
        ws_span.cell(row=row_num, column=1, value=SOURCE_FILE)
        ws_span.cell(row=row_num, column=2, value=NOTE_ID)
        ws_span.cell(row=row_num, column=3, value=f"span_{i+1:03d}")
        ws_span.cell(row=row_num, column=4, value="Procedure") # section placeholder
        ws_span.cell(row=row_num, column=5, value=s.get('context', ''))
        ws_span.cell(row=row_num, column=6, value=s['span_text'])
        ws_span.cell(row=row_num, column=7, value="") # match_index
        ws_span.cell(row=row_num, column=8, value="") # start_char BLANK
        ws_span.cell(row=row_num, column=9, value="") # end_char BLANK
        ws_span.cell(row=row_num, column=10, value=f"=LEN(F{row_num})")
        ws_span.cell(row=row_num, column=11, value=s['label'])
        ws_span.cell(row=row_num, column=12, value=s.get('norm', ''))
        ws_span.cell(row=row_num, column=13, value="")
        ws_span.cell(row=row_num, column=14, value=s['event_id'])
        ws_span.cell(row=row_num, column=15, value="FALSE")
        ws_span.cell(row=row_num, column=16, value="FALSE")
        ws_span.cell(row=row_num, column=17, value="")
        ws_span.cell(row=row_num, column=18, value="Auto")
        ws_span.cell(row=row_num, column=19, value="")
        ws_span.cell(row=row_num, column=20, value="needs_hydration")
        row_num += 1

    # --- Sheet: Span_Hydrated ---
    if "Span_Hydrated" not in wb.sheetnames:
        wb.create_sheet("Span_Hydrated")
    ws_hyd = wb["Span_Hydrated"]
    
    row_num = ws_hyd.max_row + 1
    for i, s in enumerate(hydrated_spans):
        ws_hyd.cell(row=row_num, column=1, value=SOURCE_FILE)
        ws_hyd.cell(row=row_num, column=2, value=NOTE_ID)
        ws_hyd.cell(row=row_num, column=3, value=f"span_{i+1:03d}")
        ws_hyd.cell(row=row_num, column=4, value="Procedure")
        ws_hyd.cell(row=row_num, column=5, value=s.get('context', ''))
        ws_hyd.cell(row=row_num, column=6, value=s['span_text'])
        ws_hyd.cell(row=row_num, column=7, value="")
        ws_hyd.cell(row=row_num, column=8, value=s['start_char'])
        ws_hyd.cell(row=row_num, column=9, value=s['end_char'])
        ws_hyd.cell(row=row_num, column=10, value=len(s['span_text']))
        ws_hyd.cell(row=row_num, column=11, value=s['label'])
        ws_hyd.cell(row=row_num, column=12, value=s.get('norm', ''))
        ws_hyd.cell(row=row_num, column=13, value="")
        ws_hyd.cell(row=row_num, column=14, value=s['event_id'])
        ws_hyd.cell(row=row_num, column=15, value="FALSE")
        ws_hyd.cell(row=row_num, column=16, value="FALSE")
        ws_hyd.cell(row=row_num, column=17, value="")
        ws_hyd.cell(row=row_num, column=18, value="Auto")
        ws_hyd.cell(row=row_num, column=19, value="")
        ws_hyd.cell(row=row_num, column=20, value=s['hydration_status'])
        row_num += 1

    # --- Sheet: Event_Log ---
    if "Event_Log" not in wb.sheetnames:
        wb.create_sheet("Event_Log")
    ws_evt = wb["Event_Log"]
    
    # Columns: source_file, note_id, event_id, event_type, method, anatomy_target, device, needle_gauge, stations, counts, measurements, specimens, findings, is_historical, reviewer, comments, device_size, device_material, outcome_airway_lumen_pre, outcome_airway_lumen_post, outcome_symptoms, outcome_pleural, outcome_complication
    
    row_num = ws_evt.max_row + 1
    for e in raw_events:
        ws_evt.cell(row=row_num, column=1, value=SOURCE_FILE)
        ws_evt.cell(row=row_num, column=2, value=NOTE_ID)
        ws_evt.cell(row=row_num, column=3, value=e['event_id'])
        ws_evt.cell(row=row_num, column=4, value=e.get('type', ''))
        ws_evt.cell(row=row_num, column=5, value=e.get('method', ''))
        ws_evt.cell(row=row_num, column=6, value=e.get('anatomy', ''))
        ws_evt.cell(row=row_num, column=7, value=e.get('device', ''))
        ws_evt.cell(row=row_num, column=13, value=e.get('findings', ''))
        ws_evt.cell(row=row_num, column=19, value=e.get('outcome_lumen_pre', ''))
        ws_evt.cell(row=row_num, column=20, value=e.get('outcome_lumen_post', ''))
        ws_evt.cell(row=row_num, column=23, value=e.get('outcome_complication', ''))
        row_num += 1

    # --- Sheet: V3_Procedure_Events ---
    if "V3_Procedure_Events" not in wb.sheetnames:
        wb.create_sheet("V3_Procedure_Events")
    ws_v3 = wb["V3_Procedure_Events"]
    
    row_num = ws_v3.max_row + 1
    for e in raw_events:
        ws_v3.cell(row=row_num, column=1, value=NOTE_ID)
        ws_v3.cell(row=row_num, column=2, value=e['event_id'])
        ws_v3.cell(row=row_num, column=3, value=e.get('type', ''))
        # Mapping simple fields for V3 demonstration
        if "anatomy" in e:
            ws_v3.cell(row=row_num, column=4, value="Airway") # target.anatomy_type
            ws_v3.cell(row=row_num, column=5, value=e['anatomy']) # Lobe/loc
        ws_v3.cell(row=row_num, column=10, value=e.get('method', ''))
        ws_v3.cell(row=row_num, column=19, value=e.get('outcome_lumen_pre', ''))
        ws_v3.cell(row=row_num, column=20, value=e.get('outcome_lumen_post', ''))
        ws_v3.cell(row=row_num, column=22, value=e.get('outcome_complication', ''))
        row_num += 1

    # --- Sheet: V3_Registry_JSON ---
    if "V3_Registry_JSON" not in wb.sheetnames:
        wb.create_sheet("V3_Registry_JSON")
    ws_json = wb["V3_Registry_JSON"]
    
    json_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": raw_events,
        "no_immediate_complications": True
    }
    
    ws_json.cell(row=ws_json.max_row+1, column=1, value=json.dumps(json_data, indent=2))

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated workbook at {OUTPUT_PATH}")

if __name__ == "__main__":
    create_workbook()