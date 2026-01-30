import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os
from datetime import datetime

# =============================================================================
# INPUTS
# =============================================================================
NOTE_ID = "note_051"
SOURCE_FILE = "note_051.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_051 SOURCE_FILE: note_051.txt INDICATION FOR OPERATION:  [REDACTED]is a 27 year old-year-old female who presents with lung infiltrates.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
 
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
 
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31624 Dx bronchoscope/lavage (BAL)    
 
ANESTHESIA: 
General Anesthesia
 
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
 
Successful therapeutic aspiration was performed to clean out the Right Mainstem, Bronchus Intermedius , and Left Mainstem from mucus and mucus plug.
Bronchial alveolar lavage was performed at Lateral Segment of RML (RB4) and Medial Segment of RML (RB5).
Instilled 60 cc of NS, suction returned with 20 cc of NS.  Samples sent for Microbiology (Cultures/Viral/Fungal).
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
RML BAL
 
IMPRESSION/PLAN: [REDACTED]is a 27 year old-year-old female who presents for bronchoscopy for lung infiltrates.
- f/u in clinic"""

# =============================================================================
# CONFIGURATION
# =============================================================================

PROCEDURE_FLAGS = [
    "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
    "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy",
    "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration",
    "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
    "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation",
    "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
    "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis",
    "pleural_biopsy", "fibrinolytic_therapy"
]

# Set active flags based on note content
ACTIVE_FLAGS = {
    "diagnostic_bronchoscopy": 1, # Implied by BAL/31624
    "bal": 1,
    "therapeutic_aspiration": 1,
}

# Initialize all flags to 0
NOTE_FLAGS = {flag: ACTIVE_FLAGS.get(flag, 0) for flag in PROCEDURE_FLAGS}

# =============================================================================
# SPAN EXTRACTION
# =============================================================================

# Helper to define spans
def s(text, label, norm=None, field=None, event="1", prefix=None):
    return {
        "span_text": text,
        "label": label,
        "normalized_value": norm,
        "schema_field": field,
        "event_id": event,
        "context_prefix": prefix
    }

SPANS = [
    # Event 1: Therapeutic Aspiration
    s("Therapeutic aspiration", "PROC_METHOD", "therapeutic_aspiration", "method", "1"),
    s("clean out", "PROC_ACTION", "clean", "action", "1"),
    s("Right Mainstem", "ANAT_AIRWAY", "RMS", "target.anatomy", "1"),
    s("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "target.anatomy", "1"),
    s("Left Mainstem", "ANAT_AIRWAY", "LMS", "target.anatomy", "1"),
    s("mucus and mucus plug", "OBS_LESION", "mucus plug", "findings", "1"),
    
    # Event 2: BAL
    s("Bronchial alveolar lavage", "PROC_METHOD", "bal", "method", "2"),
    s("Lateral Segment of RML", "ANAT_LUNG_LOC", "RML Lateral", "target.anatomy", "2"),
    s("RB4", "ANAT_AIRWAY", "RB4", "target.anatomy", "2"),
    s("Medial Segment of RML", "ANAT_LUNG_LOC", "RML Medial", "target.anatomy", "2"),
    s("RB5", "ANAT_AIRWAY", "RB5", "target.anatomy", "2"),
    s("60 cc", "MEAS_VOL", "60", "measurements.instilled_vol", "2", prefix="Instilled "),
    s("20 cc", "MEAS_VOL", "20", "measurements.return_vol", "2", prefix="suction returned with "),
    s("Samples sent for Microbiology", "PROC_ACTION", "microbiology", "specimens", "2"),
    
    # Outcomes / Global
    s("No immediate complications", "OUTCOME_COMPLICATION", "none", "outcomes.complications", "global"),
]

# =============================================================================
# HYDRATION LOGIC
# =============================================================================

def hydrate_spans(text, spans):
    hydrated = []
    for sp in spans:
        item = sp.copy()
        target = item["span_text"]
        prefix = item.get("context_prefix")
        
        # Default status
        item["start_char"] = ""
        item["end_char"] = ""
        item["hydration_status"] = "ambiguous"
        
        matches = [m.start() for m in re.finditer(re.escape(target), text)]
        
        if not matches:
            item["hydration_status"] = "not_found"
            hydrated.append(item)
            continue
            
        selected_index = None
        
        # 1. Unique match
        if len(matches) == 1:
            selected_index = matches[0]
            item["hydration_status"] = "hydrated_unique"
            
        # 2. Context prefix
        elif prefix:
            for m_idx in matches:
                # Look back 120 chars
                start_search = max(0, m_idx - 120)
                context_window = text[start_search:m_idx]
                if prefix in context_window:
                    selected_index = m_idx
                    item["hydration_status"] = "hydrated_prefix_window"
                    break
        
        # Fallback: if ambiguous and no resolution found
        if selected_index is None:
            # For this script, we default to the first occurrence if ambiguous to ensure data flows,
            # but mark it. Real systems might need manual review.
            selected_index = matches[0]
            item["hydration_status"] = f"ambiguous_count={len(matches)}"

        if selected_index is not None:
            item["start_char"] = selected_index
            item["end_char"] = selected_index + len(target)
            item["span_len"] = len(target)
            
        hydrated.append(item)
    return hydrated

# =============================================================================
# EXCEL GENERATION
# =============================================================================

def create_workbook():
    # 1. Load Template
    if not os.path.exists(TEMPLATE_PATH):
        # Create a dummy workbook if template missing (for robustness in standalone envs)
        wb = openpyxl.Workbook()
        wb.create_sheet("Note_Text")
        wb.create_sheet("Note_Index")
        wb.create_sheet("Span_Annotations")
        wb.create_sheet("Span_Hydrated")
        wb.create_sheet("Event_Log")
        wb.create_sheet("V3_Procedure_Events")
        wb.create_sheet("V3_Registry_JSON")
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # 2. Note_Text
    ws = wb["Note_Text"]
    # Check headers
    if ws.max_row == 1 and ws.cell(1,1).value is None:
        ws.append(["note_id", "source_file", "note_text"])
    ws.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 3. Note_Index
    ws = wb["Note_Index"]
    # Headers: source_file, note_id, encounter_id, procedure_date, site, reviewer, status, free_text_notes, [flags...]
    row_meta = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Auto-Generated", ""]
    row_flags = [NOTE_FLAGS[f] for f in PROCEDURE_FLAGS]
    ws.append(row_meta + row_flags)

    # 4. Span_Annotations (Anchor First)
    ws = wb["Span_Annotations"]
    # Columns: source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, start_char, end_char, span_len, label, normalized_value, schema_field, event_id, is_negated, is_historical, time_anchor, reviewer, comments, hydration_status
    
    annot_rows = []
    for i, sp in enumerate(SPANS):
        row = [
            SOURCE_FILE, NOTE_ID, f"span_{i+1:03d}", "", 
            sp.get("context_prefix", ""), sp["span_text"], "", 
            "", "", f'=LEN(F{ws.max_row+1+i})', # Formula for len
            sp["label"], sp["normalized_value"], sp["schema_field"], sp["event_id"],
            "", "", "", "Auto", "", "needs_hydration"
        ]
        ws.append(row)

    # 5. Span_Hydrated
    ws = wb["Span_Hydrated"]
    hydrated_data = hydrate_spans(NOTE_TEXT, SPANS)
    
    # Columns mirror Span_Annotations but with filled offsets
    # source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, start_char, end_char, span_len, label, normalized_value, schema_field, event_id, is_negated, is_historical, time_anchor, reviewer, comments, hydration_status
    
    for i, h in enumerate(hydrated_data):
        row = [
            SOURCE_FILE, NOTE_ID, f"span_{i+1:03d}", "",
            h.get("context_prefix", ""), h["span_text"], "",
            h["start_char"], h["end_char"], h.get("span_len", ""),
            h["label"], h["normalized_value"], h["schema_field"], h["event_id"],
            "", "", "", "Auto", "", h["hydration_status"]
        ]
        ws.append(row)

    # 6. Event_Log (Best Effort)
    ws = wb["Event_Log"]
    # Columns: source_file, note_id, event_id, event_type, method, anatomy_target, device, needle_gauge, stations, counts, measurements, specimens, findings, is_historical, reviewer, comments, device_size, device_material, outcome_airway_lumen_pre, outcome_airway_lumen_post, outcome_symptoms, outcome_pleural, outcome_complication
    
    # Group spans by event
    events = {}
    for h in hydrated_data:
        eid = h["event_id"]
        if eid not in events:
            events[eid] = {
                "type": "", "method": [], "anatomy": [], "findings": [], 
                "measurements": [], "specimens": [], "complications": []
            }
        
        if h["label"] == "PROC_METHOD":
            events[eid]["type"] = h["normalized_value"]
            events[eid]["method"].append(h["span_text"])
        elif h["label"] in ["ANAT_AIRWAY", "ANAT_LUNG_LOC"]:
            events[eid]["anatomy"].append(h["normalized_value"])
        elif h["label"] == "OBS_LESION":
            events[eid]["findings"].append(h["normalized_value"])
        elif h["label"] == "MEAS_VOL":
            events[eid]["measurements"].append(f"{h['normalized_value']}cc")
        elif h["label"] == "PROC_ACTION" and h["schema_field"] == "specimens":
            events[eid]["specimens"].append(h["normalized_value"])
        elif h["label"] == "OUTCOME_COMPLICATION":
            events[eid]["complications"].append(h["normalized_value"])

    # Write Event 1 (Therapeutic Aspiration)
    if "1" in events:
        e1 = events["1"]
        ws.append([
            SOURCE_FILE, NOTE_ID, "1", "therapeutic_aspiration", ",".join(e1["method"]),
            ",".join(e1["anatomy"]), "", "", "", "", "", "", ",".join(e1["findings"]),
            "FALSE", "Auto", "", "", "", "", "", "", "", ""
        ])
        
    # Write Event 2 (BAL)
    if "2" in events:
        e2 = events["2"]
        ws.append([
            SOURCE_FILE, NOTE_ID, "2", "bal", ",".join(e2["method"]),
            ",".join(e2["anatomy"]), "", "", "", "", ",".join(e2["measurements"]), 
            ",".join(e2["specimens"]), "", "FALSE", "Auto", "", "", "", "", "", "", "", ""
        ])

    # 7. V3_Procedure_Events & JSON
    ws_v3 = wb["V3_Procedure_Events"]
    ws_json = wb["V3_Registry_JSON"]
    
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": [],
        "no_immediate_complications": True # derived from global outcome
    }

    # Event 1
    v3_row_1 = [
        NOTE_ID, "1", "therapeutic_aspiration",
        "airway", "", "", "", # target: type, lobe, segment, station
        "mucus_plug", "", # lesion type, size
        "aspiration", "[]", "{}", "[]", json.dumps(["mucus"]), "", # method, dev, meas, spec, find, quote
        "", "", "", # stent/cath
        "", "", "", "", "none" # outcomes
    ]
    ws_v3.append(v3_row_1)
    
    proc_1 = {
        "event_id": "1",
        "type": "therapeutic_aspiration",
        "target": [{"anatomy_type": "airway", "location": "RMS"}, {"anatomy_type": "airway", "location": "LMS"}, {"anatomy_type": "airway", "location": "BI"}],
        "findings": ["mucus"],
        "method": "aspiration"
    }
    registry_data["procedures"].append(proc_1)

    # Event 2
    v3_row_2 = [
        NOTE_ID, "2", "bal",
        "lung", "RML", "lateral, medial", "", 
        "", "",
        "lavage", "[]", json.dumps({"instilled": 60, "return": 20}), json.dumps(["microbiology"]), "[]", "",
        "", "", "",
        "", "", "", "", "none"
    ]
    ws_v3.append(v3_row_2)

    proc_2 = {
        "event_id": "2",
        "type": "bal",
        "target": [{"anatomy_type": "lung", "lobe": "RML", "segment": "lateral"}, {"anatomy_type": "lung", "lobe": "RML", "segment": "medial"}],
        "measurements": {"instilled_vol": 60, "return_vol": 20},
        "specimens": ["microbiology"]
    }
    registry_data["procedures"].append(proc_2)

    # Write JSON
    ws_json.append([json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    create_workbook()