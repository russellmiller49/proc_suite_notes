import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# -------------------------------------------------------------------------
# INPUTS
# -------------------------------------------------------------------------
NOTE_ID = "note_052"
SOURCE_FILE = "note_052.txt"
PROCEDURE_DATE = ""
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_052 SOURCE_FILE: note_052.txt INDICATION FOR OPERATION:  [REDACTED]is a 82 year old-year-old female who presents with lymphadenopathy.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
PREOPERATIVE DIAGNOSIS: R59.0 Localized enlarged lymph nodes
POSTOPERATIVE DIAGNOSIS:  R59.0 Localized enlarged lymph nodes
 
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31653 EBUS sampling 3 or more nodes  
76982 Ultrasound Elastography, First Target Lesion
76983 Ultrasound Elastography, Additional Targets 
76983 Ultrasound Elastography, Additional Target 2
 
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
Linear EBUS 
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
 
Successful therapeutic aspiration was performed to clean out the Right Mainstem, Bronchus Intermedius , and Left Mainstem from mucus.
EBUS-Findings
Indications: Diagnostic
Technique:
All lymph node stations were assessed. Only those 5 mm or greater in short axis were sampled.
Lymph node sizing was performed by EBUS and sampling by transbronchial needle aspiration was performed using 25-gauge Needle and 22-gauge Needle.
Lymph Nodes/Sites Inspected: 4R (lower paratracheal) node
11Rs lymph node
11L lymph node
 
Overall ROSE Diagnosis: Suggestive of benign-appearing lymphoid tissue
 
No immediate complications
 
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
Elastography provided a semi-quantitative classification (Type 1–3), which was used to guide biopsy site selection and sampling strategy.
Lymph Nodes Evaluated:
Site 1: The 11L lymph node was => 10 mm on CT and Hypermetabolic via PET-CT scan.
The lymph node was photographed. The site was sampled.. 4 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 2 elastographic pattern with mixed soft and stiff regions.
Given this heterogeneous and indeterminate appearance, TBNA was directed at representative areas to ensure comprehensive sampling and to minimize the risk of underdiagnosis.
Preliminary ROSE Cytology was reported as adequate and suggestive of anthracotic pigments. Final results are pending.
Site 2: The 4R (lower paratracheal) node was => 10 mm on CT and Hypermetabolic via PET-CT scan.
The lymph node was photographed. The site was sampled.. 4 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 2 elastographic pattern with mixed soft and stiff regions.
Given this heterogeneous and indeterminate appearance, TBNA was directed at representative areas to ensure comprehensive sampling and to minimize the risk of underdiagnosis.
Preliminary ROSE Cytology was reported as adequate and suggestive of anthracotic pigments. Final results are pending.
Site 3: The 11Rs lymph node was => 10 mm on CT and Hypermetabolic via PET-CT scan.
The lymph node was photographed. The site was sampled.. 4 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 2 elastographic pattern with mixed soft and stiff regions.
Given this heterogeneous and indeterminate appearance, TBNA was directed at representative areas to ensure comprehensive sampling and to minimize the risk of underdiagnosis.
Preliminary ROSE Cytology was reported as adequate and suggestive of anthracotic pigments. Final results are pending.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
11L, 11Rs, 4R TBNA
 
IMPRESSION/PLAN: [REDACTED]is a 82 year old-year-old female who presents for bronchoscopy for lymphadenopathy.
- f/u final path
- f/u in clinic"""

# -------------------------------------------------------------------------
# DATA STRUCTURES
# -------------------------------------------------------------------------

# Flag logic
FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,  # General bronch indicated
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,      # Explicit mention and use
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1, # Explicitly stated CPT and in text
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# -------------------------------------------------------------------------
# SPAN DEFINITIONS (ANCHOR-FIRST)
# -------------------------------------------------------------------------
# Each tuple: (span_text, label, normalized_value, event_id, context_prefix or match_index)
# context_prefix: string to search before. match_index: integer 0-based index if ambiguous.

RAW_SPANS = [
    # Global Devices/Instruments
    ("Flexible Therapeutic Bronchoscope", "DEV_INSTRUMENT", "flexible_bronchoscope", "ev0", None),
    ("Linear EBUS", "DEV_INSTRUMENT", "linear_ebus", "ev0", None),
    ("25-gauge Needle", "DEV_NEEDLE", "25G", "ev0", None),
    ("22-gauge Needle", "DEV_NEEDLE", "22G", "ev0", None),

    # Event 1: Therapeutic Aspiration
    ("Successful therapeutic aspiration", "PROC_METHOD", "therapeutic_aspiration", "ev1", None),
    ("Right Mainstem", "ANAT_AIRWAY", "RMS", "ev1", "clean out the"),
    ("Bronchus Intermedius", "ANAT_AIRWAY", "BI", "ev1", None),
    ("Left Mainstem", "ANAT_AIRWAY", "LMS", "ev1", "clean out the"),
    ("mucus", "OBS_LESION", "mucus", "ev1", "Left Mainstem from"),

    # Event 2: Station 11L (Site 1)
    ("11L lymph node", "ANAT_LN_STATION", "11L", "ev2", "Site 1: The"),
    ("=> 10 mm", "MEAS_SIZE", "10mm", "ev2", "Site 1:"),
    ("Hypermetabolic", "OBS_LESION", "hypermetabolic", "ev2", "Site 1:"),
    ("4", "MEAS_COUNT", "4", "ev2", "Site 1:"), # "4 endobronchial ultrasound..."
    ("endobronchial ultrasound guided transbronchial biopsies", "PROC_METHOD", "linear_ebus_tbna", "ev2", "Site 1:"),
    ("Type 2 elastographic pattern", "OBS_LESION", "elastography_type_2", "ev2", "Site 1:"),
    ("mixed soft and stiff regions", "OBS_LESION", "heterogeneous_stiffness", "ev2", "Site 1:"),
    ("anthracotic pigments", "OBS_ROSE", "anthracosis", "ev2", "Site 1:"),

    # Event 3: Station 4R (Site 2)
    ("4R (lower paratracheal) node", "ANAT_LN_STATION", "4R", "ev3", "Site 2: The"),
    ("=> 10 mm", "MEAS_SIZE", "10mm", "ev3", "Site 2:"),
    ("Hypermetabolic", "OBS_LESION", "hypermetabolic", "ev3", "Site 2:"),
    ("4", "MEAS_COUNT", "4", "ev3", "Site 2:"),
    ("endobronchial ultrasound guided transbronchial biopsies", "PROC_METHOD", "linear_ebus_tbna", "ev3", "Site 2:"),
    ("Type 2 elastographic pattern", "OBS_LESION", "elastography_type_2", "ev3", "Site 2:"),
    ("mixed soft and stiff regions", "OBS_LESION", "heterogeneous_stiffness", "ev3", "Site 2:"),
    ("anthracotic pigments", "OBS_ROSE", "anthracosis", "ev3", "Site 2:"),

    # Event 4: Station 11Rs (Site 3)
    ("11Rs lymph node", "ANAT_LN_STATION", "11Rs", "ev4", "Site 3: The"),
    ("=> 10 mm", "MEAS_SIZE", "10mm", "ev4", "Site 3:"),
    ("Hypermetabolic", "OBS_LESION", "hypermetabolic", "ev4", "Site 3:"),
    ("4", "MEAS_COUNT", "4", "ev4", "Site 3:"),
    ("endobronchial ultrasound guided transbronchial biopsies", "PROC_METHOD", "linear_ebus_tbna", "ev4", "Site 3:"),
    ("Type 2 elastographic pattern", "OBS_LESION", "elastography_type_2", "ev4", "Site 3:"),
    ("mixed soft and stiff regions", "OBS_LESION", "heterogeneous_stiffness", "ev4", "Site 3:"),
    ("anthracotic pigments", "OBS_ROSE", "anthracosis", "ev4", "Site 3:"),

    # Outcomes
    ("No immediate complications", "OUTCOME_COMPLICATION", "none", "ev99", "procedure well.  There were")
]

# -------------------------------------------------------------------------
# SCRIPT LOGIC
# -------------------------------------------------------------------------

def create_workbook():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        # Create a basic workbook structure if template is missing (fallback)
        wb = openpyxl.Workbook()
        for sheet in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            if sheet not in wb.sheetnames:
                wb.create_sheet(sheet)
    return wb

def hydrate_spans(note_text, raw_spans):
    hydrated = []
    
    for idx, (span_text, label, norm, event_id, context) in enumerate(raw_spans):
        start_char = -1
        end_char = -1
        hydration_status = "ambiguous"
        
        # Strategy 1: Unique exact match
        count = note_text.count(span_text)
        if count == 1:
            start_char = note_text.find(span_text)
            hydration_status = "hydrated_unique"
        
        # Strategy 2: Context Prefix
        elif context and isinstance(context, str):
            matches = [m.start() for m in re.finditer(re.escape(span_text), note_text)]
            for m_start in matches:
                # Look at preceding 120 chars
                window_start = max(0, m_start - 120)
                preceding = note_text[window_start:m_start]
                if context in preceding:
                    start_char = m_start
                    hydration_status = "hydrated_prefix_window"
                    break
        
        # Strategy 3: Match Index (if context is integer)
        elif isinstance(context, int) and count > context:
            matches = [m.start() for m in re.finditer(re.escape(span_text), note_text)]
            if len(matches) > context:
                start_char = matches[context]
                hydration_status = "hydrated_match_index"

        # Strategy 4: Fallback to first if ambiguous but exists (Annotator Correction needed)
        if start_char == -1 and count > 0:
             start_char = note_text.find(span_text)
             hydration_status = f"ambiguous_count={count}_default_first"

        if start_char != -1:
            end_char = start_char + len(span_text)

        hydrated.append({
            "span_id": f"s{idx+1}",
            "span_text": span_text,
            "label": label,
            "normalized_value": norm,
            "event_id": event_id,
            "start_char": start_char if start_char != -1 else "",
            "end_char": end_char if end_char != -1 else "",
            "hydration_status": hydration_status,
            "context": context
        })
    return hydrated

def main():
    wb = create_workbook()
    
    # 1. Note_Text
    ws = wb["Note_Text"]
    if ws.max_row == 1: # Header check
        ws.append(["note_id", "source_file", "note_text"])
    ws.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws = wb["Note_Index"]
    # Headers implied by template; strict mapping
    # Assuming standard headers: source_file, note_id, ..., [flags]
    # We append a row matching the template columns conceptually
    row_data = [
        SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "final", "Parsed by script"
    ]
    # Append flag values in specific order
    for k, v in FLAGS.items():
        row_data.append(v)
    ws.append(row_data)

    # 3. Span_Annotations (Anchor First - no offsets)
    ws_ann = wb["Span_Annotations"]
    # Header: source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, start, end, len, label, norm, schema, event, neg, hist, time, rev, comm, status
    
    hydrated_data = hydrate_spans(NOTE_TEXT, RAW_SPANS)
    
    for h in hydrated_data:
        ctx_prefix = h['context'] if isinstance(h['context'], str) else ""
        match_idx = h['context'] if isinstance(h['context'], int) else ""
        
        ws_ann.append([
            SOURCE_FILE, NOTE_ID, h['span_id'], "Procedure", ctx_prefix, h['span_text'], match_idx,
            "", "", f"=LEN(F{ws_ann.max_row+1})", # Formula for len
            h['label'], h['normalized_value'], "", h['event_id'],
            "FALSE", "FALSE", "", "Script", "", "needs_hydration"
        ])

    # 4. Span_Hydrated
    ws_hyd = wb["Span_Hydrated"]
    for h in hydrated_data:
        ctx_prefix = h['context'] if isinstance(h['context'], str) else ""
        match_idx = h['context'] if isinstance(h['context'], int) else ""
        
        ws_hyd.append([
            SOURCE_FILE, NOTE_ID, h['span_id'], "Procedure", ctx_prefix, h['span_text'], match_idx,
            h['start_char'], h['end_char'], len(h['span_text']),
            h['label'], h['normalized_value'], "", h['event_id'],
            "FALSE", "FALSE", "", "Script", "", h['hydration_status']
        ])

    # 5. Event_Log (Simplified Logic)
    ws_evt = wb["Event_Log"]
    # Define events manually based on logical grouping
    events = [
        {
            "id": "ev1", "type": "therapeutic_aspiration", "method": "Aspiration", 
            "anatomy": "RMS, BI, LMS", "findings": "Mucus"
        },
        {
            "id": "ev2", "type": "linear_ebus_tbna", "method": "EBUS-TBNA",
            "anatomy": "11L", "device": "Linear EBUS, 25G/22G", "counts": "4 passes",
            "findings": "Hypermetabolic, Type 2 elastography, Anthracosis"
        },
        {
            "id": "ev3", "type": "linear_ebus_tbna", "method": "EBUS-TBNA",
            "anatomy": "4R", "device": "Linear EBUS, 25G/22G", "counts": "4 passes",
            "findings": "Hypermetabolic, Type 2 elastography, Anthracosis"
        },
        {
            "id": "ev4", "type": "linear_ebus_tbna", "method": "EBUS-TBNA",
            "anatomy": "11Rs", "device": "Linear EBUS, 25G/22G", "counts": "4 passes",
            "findings": "Hypermetabolic, Type 2 elastography, Anthracosis"
        }
    ]
    
    for e in events:
        ws_evt.append([
            SOURCE_FILE, NOTE_ID, e["id"], e["type"], e["method"], e.get("anatomy", ""),
            e.get("device", ""), "", "", e.get("counts", ""), "", "", 
            e.get("findings", ""), "FALSE", "Script", ""
        ])

    # 6. V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    # Map events to V3 schema columns
    
    # Event 1: Therapeutic Aspiration
    ws_v3.append([
        NOTE_ID, "ev1", "Therapeutic Aspiration", "Airway", "", "", "", # Target
        "", "", "Suction/Aspiration", # Method
        "", "", "", '{"finding": "mucus"}', "Successful therapeutic aspiration... clean out... from mucus",
        "", "", "", "", "", "", "", "" # Outcomes
    ])

    # Event 2: EBUS 11L
    ws_v3.append([
        NOTE_ID, "ev2", "EBUS-TBNA", "Lymph Node", "", "", "11L",
        "Lymphadenopathy", "10", "TBNA", 
        '{"needle": ["25G", "22G"]}', '{"passes": 4}', "", 
        '{"rose": "Anthracotic pigments", "elastography": "Type 2", "metabolic": "Hypermetabolic"}',
        "Site 1: The 11L lymph node...",
        "", "", "", "", "", "", "", ""
    ])

    # Event 3: EBUS 4R
    ws_v3.append([
        NOTE_ID, "ev3", "EBUS-TBNA", "Lymph Node", "", "", "4R",
        "Lymphadenopathy", "10", "TBNA", 
        '{"needle": ["25G", "22G"]}', '{"passes": 4}', "", 
        '{"rose": "Anthracotic pigments", "elastography": "Type 2", "metabolic": "Hypermetabolic"}',
        "Site 2: The 4R...",
        "", "", "", "", "", "", "", ""
    ])

    # Event 4: EBUS 11Rs
    ws_v3.append([
        NOTE_ID, "ev4", "EBUS-TBNA", "Lymph Node", "", "", "11Rs",
        "Lymphadenopathy", "10", "TBNA", 
        '{"needle": ["25G", "22G"]}', '{"passes": 4}', "", 
        '{"rose": "Anthracotic pigments", "elastography": "Type 2", "metabolic": "Hypermetabolic"}',
        "Site 3: The 11Rs...",
        "", "", "", "", "", "", "", ""
    ])

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": events
    }
    ws_json.append([json.dumps(registry_data, indent=2)])

    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    main()