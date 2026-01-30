import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_053"
SOURCE_FILE = "note_053.txt"
PROCEDURE_DATE = "2026-01-12" # inferred from context or left blank if not in text, using prompt context date for file creation but specific note date is not explicit in header. Leaving blank or using context.
NOTE_TEXT = """NOTE_ID:  note_053 SOURCE_FILE: note_053.txt Pre-Op Diagnosis: Respiratory failure, on VV ECMO, CTEPH, acute PE, RV failure, reintubation
 
Post-Op Diagnosis: Same
 
Procedure: Tracheostomy, modified open/percutaneous

Brief Clinical Note: 
49F w/polycythemia vera with JAK2 mutation and prior thrombogenic events p/w SOB/CP found to have group 4 CTEPH w/ massive PEs w/RV strain/thrombus 12/1 (Loma Linda) s/p failed/incomplete IR thrombectomy c/b Lt CVA s/p neuroIR thrombectomy 12/3 further c/b CV decompensation s/p Lt fem VA ECMO w/DPC 12/4, failed wean 12/10 so converted to VV ECMO.
Extubated 12/11. S/p oxygenator exchange 12/12. Re-intubated 12/14 2/2 secretions. 12/15 bronch, AF RVR.
New LLE DVT on bival 12/16 though was ECMO cannula there.
From her Lt MCA CVA has Rt hemiplegia 
 
In need of tracheostomy to assist with ventilator weaning, secretion management, and ongoing potential rehabilitation.
Primary team, ICU team, and family in agreement. Ventilator settings, labs, and body habitus adequate for bedside procedure.
Has slightly high riding innominate artery (though not classic) as well as a large bridging neck vein.
VV ECMO/DVT/PE anticoagulation held and will restart shortly thereafter. Consent signed.
Details of Procedure Performed:  
 
The patient remained in the ICU with a previously placed single-lumen endotracheal tube.
The patient was positioned supine with a padded shoulder roll in place, the head of bed elevated, and the neck moderately extended.
A surgical brief was performed and the level of inspired fraction of oxygen was reduced to 40%.
Patient remained on VV ECMO support. After adequate anesthesia, landmarks were palpated including the thyroid notch, cricoid, and suprasternal notch.
Neck ultrasound was performed to confirm vascular anatomy. Exposure was excellent.
The anterior neck was prepped and draped in the standard fashion. 
 
A surgical time out was performed.
5mL of 1% lidocaine was injected 2-3 cm above the sternal notch.
It was made higher due to the known partially high-riding innominate.
A 3cm horizontal skin incision was made with a #15 blade knife in a natural, relaxed skin tension line midway between the cricoid cartilage and the sternal notch.
Bovie cautery was used to dissect through the platysma and down to the strap muscles, with effort made to avoid the anterior jugular veins.
There were two vertical engorged veins along the midline raphe with a bridging vein between then.
The bridging vein was dissected, suture ligated, and divided. This allowed visualization of the midline raphe.
The strap muscles were separated vertically at the midline with Bovie cautery. The strap muscles are retracted laterally.
Using a combination of blunt and careful cautery midline dissection, the anterior tracheal wall was freed of soft tissue.
The innominate artery was identified and protected. There was good soft tissue between the innominate and the trachea which was not disturbed.
The trachea was palpated to find the 2nd and 3rd tracheal rings.
There was adequate distance between the innominate artery and the planned insertion site.
Under bronchoscopic guidance, the endotracheal tube cuff was deflated and was pulled back below the level of the vocal cords.
A finder needle was placed through the anterior trachea between the 2nd and 3rd cartilaginous rings.
Once confirmed by bronchoscopy, a seldinger technique was used to serially dilate the tracheotomy over a wire and to place a # 7 cuffed shiley tracheostomy.
The tracheostomy cuff was inflated. Cross table ventilation was performed and the tracheostomy placement was confirmed by bronchoscopy through the tracheostomy as well as CO2 return on the anesthesia monitor.
The tracheostomy inner cannula was then placed. Surgicel was placed in the wound bed though hemostasis was great but plan to resume anticoagulation very shortly.
The tracheostomy was secured with loose 2-0 prolene sutures to the skin and tracheostomy straps to maintain airway but prevent skin pressure injuries.
The endotracheal tube was then withdrawn completely. The tracheostomy obturator was provided to the bedside care team.
The patient remained in the ICU in critical condition. There were no complications.
At the end of the procedure, I contacted her husband Carl Pinell at 626-991-6311 
 
EBL: 1ml
 
Fluids: per anesthesia record
 
Specimens: none
 
Implants: none
 
Drains: none
 
Complications: none immediately apparent
 
Disposition: stable, ventilated, remained in ICU, on VV ECMO
 
Plan: 
Okay to resume bival in 1 hour
Trach ties to stay for 7-10 days
No need for routine trach exchange within first 14 day if not having issues
Do not over inflate the cuff"""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# -------------------------------------------------------------------------
# CONFIGURATION
# -------------------------------------------------------------------------
LABELS = {
    "ANATOMY": ["ANAT_LN_STATION", "ANAT_LUNG_LOC", "ANAT_AIRWAY", "ANAT_PLEURA", "LATERALITY"],
    "DEVICES": ["DEV_VALVE", "DEV_STENT", "DEV_CATHETER", "DEV_NEEDLE", "DEV_INSTRUMENT", "DEV_STENT_MATERIAL", "DEV_STENT_SIZE", "DEV_CATHETER_SIZE"],
    "MEASUREMENTS": ["MEAS_SIZE", "MEAS_VOL", "MEAS_PRESS", "MEAS_COUNT", "MEAS_AIRWAY_DIAM", "MEAS_PLEURAL_DRAIN"],
    "PROCEDURE": ["PROC_METHOD", "PROC_ACTION", "OBS_ROSE", "OBS_LESION"],
    "CONTEXT": ["CTX_HISTORICAL", "CTX_TIME"],
    "OUTCOME": ["OUTCOME_AIRWAY_LUMEN_PRE", "OUTCOME_AIRWAY_LUMEN_POST", "OUTCOME_SYMPTOMS", "OUTCOME_PLEURAL", "OUTCOME_COMPLICATION"]
}

PROCEDURE_FLAGS = [
    # Bronchoscopy (23)
    "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", 
    "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy", 
    "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration", 
    "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation", 
    "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", 
    "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
    # Pleural (7)
    "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis", 
    "pleural_biopsy", "fibrinolytic_therapy"
]

# -------------------------------------------------------------------------
# PARSING LOGIC
# -------------------------------------------------------------------------

def create_span(text, span_text, label, normalized_value, schema_field, event_id, 
                context_prefix=None, is_negated=False, is_historical=False):
    """
    Creates a span dictionary. 
    """
    return {
        "source_file": SOURCE_FILE,
        "note_id": NOTE_ID,
        "span_id": None, # Generated later
        "section_type": "Procedure", # General assumption
        "context_prefix": context_prefix,
        "span_text": span_text,
        "match_index": None, # Calculated later during hydration
        "start_char": None,
        "end_char": None,
        "span_len": f'=LEN(INDIRECT("RC[-4]",0))',
        "label": label,
        "normalized_value": normalized_value,
        "schema_field": schema_field,
        "event_id": event_id,
        "is_negated": is_negated,
        "is_historical": is_historical,
        "time_anchor": None,
        "reviewer": "Auto",
        "comments": None,
        "hydration_status": "needs_hydration"
    }

def hydrate_spans(spans, full_text):
    """
    Calculates start_char, end_char, and match_index logic.
    """
    hydrated_rows = []
    
    # Track occurrence counts for match_index logic
    text_occurrences = {} # span_text -> list of start_offsets
    
    for s in spans:
        st = s["span_text"]
        if st not in text_occurrences:
            # Find all occurrences
            offsets = [m.start() for m in re.finditer(re.escape(st), full_text)]
            text_occurrences[st] = offsets
    
    # Assign specific occurrences
    # We maintain a consumed set to try to avoid overlapping identical spans if possible, 
    # but strictly following the prompt's logic:
    # 1. If count==1, use it.
    # 2. If context, use context.
    # 3. Else ambiguous/match_index.
    
    # We need to process sequentially or group by span_text to assign match_index correctly if manual
    
    for s in spans:
        st = s["span_text"]
        candidates = text_occurrences.get(st, [])
        
        start = None
        end = None
        h_status = "ambiguous"
        
        if len(candidates) == 1:
            start = candidates[0]
            h_status = "hydrated_unique"
        elif len(candidates) > 1:
            # Try context
            if s["context_prefix"]:
                ctx = s["context_prefix"]
                best_c = None
                for c in candidates:
                    # check preceding 120 chars
                    window_start = max(0, c - 120)
                    window_text = full_text[window_start:c]
                    if ctx in window_text:
                        best_c = c
                        break
                if best_c is not None:
                    start = best_c
                    h_status = "hydrated_prefix_window"
            
            # If still none, check if match_index is manually provided (not in this flow),
            # or just default to first available (or mark ambiguous)
            if start is None:
                # Naive: take the first one that hasn't been heavily used? 
                # For this script, we'll just take the first one to ensure validity, 
                # but mark as ambiguous count
                start = candidates[0]
                h_status = f"ambiguous_count={len(candidates)}"
        
        else:
            h_status = "not_found"
        
        if start is not None:
            end = start + len(st)
        
        # Create hydrated row
        row = s.copy()
        row["start_char"] = start
        row["end_char"] = end
        row["hydration_status"] = h_status
        hydrated_rows.append(row)

    return hydrated_rows

def generate_phase0():
    # 1. Create Flags
    flags = {f: 0 for f in PROCEDURE_FLAGS}
    
    # Logic for this specific note
    # "Under bronchoscopic guidance" -> diagnostic_bronchoscopy
    if "bronchoscopic guidance" in NOTE_TEXT or "confirmed by bronchoscopy" in NOTE_TEXT:
        flags["diagnostic_bronchoscopy"] = 1
    
    # "serially dilate the tracheotomy" -> airway_dilation
    if "dilate the tracheotomy" in NOTE_TEXT:
        flags["airway_dilation"] = 1

    # 2. Extract Spans
    spans = []
    
    # Event 1: Tracheostomy (Main)
    e1 = "evt_trach_1"
    
    spans.append(create_span(NOTE_TEXT, "Tracheostomy, modified open/percutaneous", "PROC_METHOD", "percutaneous tracheostomy", "method", e1))
    spans.append(create_span(NOTE_TEXT, "Respiratory failure", "PROC_ACTION", "indication", "indication", e1))
    spans.append(create_span(NOTE_TEXT, "VV ECMO", "CTX_HISTORICAL", "VV ECMO", "history", e1))
    
    # Anatomy
    spans.append(create_span(NOTE_TEXT, "cricoid", "ANAT_AIRWAY", "cricoid cartilage", "target.anatomy_type", e1))
    spans.append(create_span(NOTE_TEXT, "sternal notch", "ANAT_AIRWAY", "sternal notch", "landmark", e1, context_prefix="above the"))
    spans.append(create_span(NOTE_TEXT, "anterior jugular veins", "ANAT_AIRWAY", "anterior jugular vein", "landmark", e1))
    spans.append(create_span(NOTE_TEXT, "trachea", "ANAT_AIRWAY", "trachea", "target.anatomy_type", e1, context_prefix="soft tissue between"))
    spans.append(create_span(NOTE_TEXT, "2nd and 3rd tracheal rings", "ANAT_AIRWAY", "tracheal rings 2-3", "target.location.segment", e1))
    spans.append(create_span(NOTE_TEXT, "innominate artery", "ANAT_AIRWAY", "innominate artery", "landmark", e1, context_prefix="slightly high riding"))

    # Procedure steps / Methods
    spans.append(create_span(NOTE_TEXT, "Neck ultrasound", "PROC_METHOD", "ultrasound guidance", "method", e1))
    spans.append(create_span(NOTE_TEXT, "incision", "PROC_ACTION", "incision", "method", e1))
    spans.append(create_span(NOTE_TEXT, "dissect", "PROC_ACTION", "dissection", "method", e1))
    spans.append(create_span(NOTE_TEXT, "finder needle", "DEV_NEEDLE", "finder needle", "device", e1))
    spans.append(create_span(NOTE_TEXT, "seldinger technique", "PROC_METHOD", "seldinger technique", "method", e1))
    spans.append(create_span(NOTE_TEXT, "serially dilate", "PROC_ACTION", "dilation", "method", e1))
    spans.append(create_span(NOTE_TEXT, "tracheostomy cuff", "DEV_CATHETER", "tracheostomy tube cuff", "device", e1))
    
    # Main Device
    spans.append(create_span(NOTE_TEXT, "# 7", "MEAS_SIZE", "7", "stent.size", e1))
    spans.append(create_span(NOTE_TEXT, "cuffed shiley tracheostomy", "DEV_CATHETER", "tracheostomy tube", "device", e1))

    # Event 2: Bronchoscopy Guidance
    e2 = "evt_bronch_1"
    spans.append(create_span(NOTE_TEXT, "bronchoscopic guidance", "PROC_METHOD", "bronchoscopy", "method", e2))
    spans.append(create_span(NOTE_TEXT, "confirmed by bronchoscopy", "PROC_ACTION", "confirmation", "method", e2))
    spans.append(create_span(NOTE_TEXT, "endotracheal tube", "DEV_CATHETER", "endotracheal tube", "device", e2, context_prefix="single-lumen"))

    # Outcomes
    spans.append(create_span(NOTE_TEXT, "no complications", "OUTCOME_COMPLICATION", "none", "outcomes.complications", e1, context_prefix="There were"))
    
    # 3. Hydrate
    hydrated_spans = hydrate_spans(spans, NOTE_TEXT)
    
    # 4. Generate Events
    events_log = [
        {
            "event_id": e1,
            "event_type": "Therapeutic Airway",
            "method": "Percutaneous Tracheostomy",
            "anatomy_target": "Trachea",
            "device": "Shiley Tracheostomy Tube #7",
            "outcome_complication": "None"
        },
        {
            "event_id": e2,
            "event_type": "Diagnostic Bronchoscopy",
            "method": "Bronchoscopy",
            "findings": "Used for guidance"
        }
    ]

    # 5. V3 JSON
    v3_events = [
        {
            "note_id": NOTE_ID,
            "event_id": e1,
            "type": "Therapeutic Airway",
            "target": {"anatomy_type": "Trachea", "station": "2nd-3rd rings"},
            "method": "Percutaneous Tracheostomy",
            "devices_json": json.dumps(["Shiley Tracheostomy Tube #7"]),
            "stent": {"size": "7"},
            "outcomes": {"complications": "None"}
        },
        {
            "note_id": NOTE_ID,
            "event_id": e2,
            "type": "Diagnostic Bronchoscopy",
            "method": "Bronchoscopy",
            "evidence_quote": "Under bronchoscopic guidance"
        }
    ]
    
    registry_json = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": v3_events
    }

    # -------------------------------------------------------------------------
    # EXCEL GENERATION
    # -------------------------------------------------------------------------
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        # Create sheets if missing logic implied, but for this script assuming template exists or raw creation
        for sheet_name in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)

    # 1. Note_Text
    ws = wb["Note_Text"]
    if ws.max_row == 1:
        ws.append(["note_id", "source_file", "note_text"])
    ws.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws = wb["Note_Index"]
    if ws.max_row == 1:
        headers = ["source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes"] + PROCEDURE_FLAGS
        ws.append(headers)
    
    row_data = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Pre-annotated", ""]
    row_data.extend([flags[f] for f in PROCEDURE_FLAGS])
    ws.append(row_data)

    # 3. Span_Annotations
    ws = wb["Span_Annotations"]
    if ws.max_row == 1:
        headers = ["source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", "match_index", 
                   "start_char", "end_char", "span_len", "label", "normalized_value", "schema_field", "event_id", 
                   "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"]
        ws.append(headers)
    
    for idx, s in enumerate(spans):
        s["span_id"] = f"{NOTE_ID}_span_{idx+1:03d}"
        row = [
            s["source_file"], s["note_id"], s["span_id"], s["section_type"], s["context_prefix"], s["span_text"],
            s["match_index"], "", "", s["span_len"], s["label"], s["normalized_value"], s["schema_field"],
            s["event_id"], s["is_negated"], s["is_historical"], s["time_anchor"], s["reviewer"], s["comments"], "needs_hydration"
        ]
        ws.append(row)

    # 4. Span_Hydrated
    ws = wb["Span_Hydrated"]
    if ws.max_row == 1:
        # Same headers
        ws.append(headers)
    
    for idx, s in enumerate(hydrated_spans):
        s["span_id"] = f"{NOTE_ID}_span_{idx+1:03d}"
        row = [
            s["source_file"], s["note_id"], s["span_id"], s["section_type"], s["context_prefix"], s["span_text"],
            s["match_index"], s["start_char"], s["end_char"], s["span_len"], s["label"], s["normalized_value"], s["schema_field"],
            s["event_id"], s["is_negated"], s["is_historical"], s["time_anchor"], s["reviewer"], s["comments"], s["hydration_status"]
        ]
        ws.append(row)

    # 5. Event_Log
    ws = wb["Event_Log"]
    if ws.max_row == 1:
        headers = ["source_file", "note_id", "event_id", "event_type", "method", "anatomy_target", "device", "needle_gauge", 
                   "stations", "counts", "measurements", "specimens", "findings", "is_historical", "reviewer", "comments",
                   "device_size", "device_material", "outcome_airway_lumen_pre", "outcome_airway_lumen_post", 
                   "outcome_symptoms", "outcome_pleural", "outcome_complication"]
        ws.append(headers)
    
    for e in events_log:
        row = [
            SOURCE_FILE, NOTE_ID, e["event_id"], e["event_type"], e.get("method"), e.get("anatomy_target"), e.get("device"),
            "", "", "", "", "", e.get("findings"), False, "Auto", "",
            "", "", "", "", "", "", e.get("outcome_complication")
        ]
        ws.append(row)

    # 6. V3_Procedure_Events
    ws = wb["V3_Procedure_Events"]
    if ws.max_row == 1:
        headers = ["note_id", "event_id", "type", "target.anatomy_type", "target.location.lobe", "target.location.segment", 
                   "target.station", "lesion.type", "lesion.size_mm", "method", "devices_json", "measurements_json", 
                   "specimens_json", "findings_json", "evidence_quote", "stent.size", "stent.material_or_brand", 
                   "catheter.size_fr", "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post", "outcomes.symptoms", 
                   "outcomes.pleural", "outcomes.complications"]
        ws.append(headers)
    
    for e in v3_events:
        tgt = e.get("target", {})
        out = e.get("outcomes", {})
        row = [
            NOTE_ID, e["event_id"], e["type"], tgt.get("anatomy_type"), tgt.get("location", {}).get("lobe"),
            tgt.get("location", {}).get("segment"), tgt.get("station"), "", "", e.get("method"), e.get("devices_json"),
            "", "", "", e.get("evidence_quote"), e.get("stent", {}).get("size"), "", "",
            out.get("airway", {}).get("lumen_pre"), out.get("airway", {}).get("lumen_post"), 
            out.get("symptoms"), out.get("pleural"), out.get("complications")
        ]
        ws.append(row)

    # 7. V3_Registry_JSON
    ws = wb["V3_Registry_JSON"]
    if ws.max_row == 1:
        ws.append(["json_object"])
    
    ws.append([json.dumps(registry_json, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_phase0()