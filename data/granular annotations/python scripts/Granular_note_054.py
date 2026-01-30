import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os
from datetime import datetime

# =============================================================================
# INPUT DATA & CONFIGURATION
# =============================================================================

NOTE_ID = "note_054"
SOURCE_FILE = "note_054.txt"
PROCEDURE_DATE = "2026-01-12" # Placeholder/Inferred
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_054 SOURCE_FILE: note_054.txt INDICATION FOR OPERATION:  [REDACTED]is a 49 year old-year-old female who presents with respiratory failure.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
PREOPERATIVE DIAGNOSIS: J96.90 Respiratory Failure
 
POSTOPERATIVE DIAGNOSIS:  J96.90 Respiratory Failure
 
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31624 Dx bronchoscope/lavage (BAL)    
31502 Tracheotomy tube change prior to establishment of fistula tract
 
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)
 
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required wound packing at tracheostomy stoma.
This resulted in >50% increased work due to Time, Technical difficulty of procedure, and Physical and mental effort required.
Apply to: 31502 Tracheotomy tube change prior to establishment of fistula tract.
ANESTHESIA: 
99152 Moderate sedation: initial 15 minutes
99153 Moderate sedation: each additional 15 minutes 
 
Procedure performed under moderate sedation.
The following medications were provided:
Versed             5 mg (total)
Fentanyl          150 mcg (total)
Propofol gtt60mcg/kg/min
 
Physician/patient face-to-face anesthesia start time:   1338
 
Physician/patient face-to-face anesthesia stop time:   1510
 
Total moderate sedation time was 92 minutes.
Patient was monitored continuously one-to-one throughout the entire procedure by the attending physician while anesthesia was administered.
Sedation was administered by ICU RN. 
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Diagnostic Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: Supine
 
Initial Airway Inspection Findings:
The tracheostomy tube is in good position.
Pharynx: Normal.
Larynx: Normal.
Vocal Cords: Normal without mass/lesions
Trachea: Normal.
Main Carina: Sharp
Right Lung Proximal Airways: Normal anatomic branching to segmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Left Lung Proximal Airways: Normal anatomic branching to segmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Mucosa: Normal.
Secretions: Minimal, thin, and clear.
TRACHEOSTOMY TUBE FROM ABOVE
 
Successful therapeutic aspiration was performed to clean out the Subglottic, Trachea (Proximal 1/3), Trachea (Middle 1/3), Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus.
Bronchial alveolar lavage was performed at Lateral Segment of RML (RB4) and Medial Segment of RML (RB5).
Instilled 40 cc of NS, suction returned with 15 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
 
Upper airway was suctioned and cleared.  Endotracheal suctioning performed.
The cuff was deflated and the tracheostomy tube was easily removed.
The stoma appeared widely patent however an erosion was visualized at the 4 o'clock (inferolateral) position.
STOMA EROSION
 
The new tracheostomy tube was then placed with obturator in place.
The obturator was removed, inner cannula was placed and the cuff inflated.
Percutaneous tracheostomy was changed from:
Shiley cuffed Trach ISO/ID size 7.0mm
 
To Portex cuffed Trach ISO/ID size 7.0mm without issue.
During the procedure Cardiothoracic Surgery (Dr Hudson) was consulted as they had placed the original tracheostomy tube.
Under their guidance and expertise, the erosion was packed with iodoform gauze.
 
The patient tolerated the procedure well.
There were no immediate complications.  At the conclusion of the operation, the patient was in stable condition.
SPECIMEN(S): 
--RML BAL (cell count, micro, cyto)
 
IMPRESSION/PLAN: [REDACTED]is a 49 year old-year-old female who presents for tracheostomy tube exchange and bronchoscopy.
The patient underwent bronchoscopy and RML BAL. No airway bleeding, oozing, or significant secretions/mucus was appreciated.
Tracheostomy tube exchange was performed with removal of of the patient's Shiley 7mm cuffed tube and placement of a Portex 7mm cuffed tube.
On inspection of the patient's stoma an erosion was noted at the inferolateral aspect.
Cardiothoracic Surgery was consulted and they recommended packing with iodoform gauze.
--Portex 7mm cuffed tracheostomy tube placed
--Stoma erosion packed with iodoform gauze
--Recommend Wound Care evaluation
--Follow up RML BAL results
--Continued care per primary team"""

# =============================================================================
# DOMAIN KNOWLEDGE & MAPPING
# =============================================================================

# Record-level flags
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 1,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# Annotated Spans (Anchor-first)
# Fields: span_text, context_prefix, label, normalized_value, schema_field, event_id
SPANS_DATA = [
    {
        "span_text": "Therapeutic aspiration",
        "context_prefix": "31645",
        "label": "PROC_METHOD",
        "normalized_value": "therapeutic_aspiration",
        "event_id": "evt_asp_01"
    },
    {
        "span_text": "Bronchial alveolar lavage",
        "context_prefix": "31624 Dx bronchoscope/lavage",
        "label": "PROC_METHOD",
        "normalized_value": "bal",
        "event_id": "evt_bal_01"
    },
    {
        "span_text": "Tracheotomy tube change",
        "context_prefix": "31502",
        "label": "PROC_METHOD",
        "normalized_value": "tracheostomy_tube_change",
        "event_id": "evt_trach_01"
    },
    # Therapeutic Aspiration Anatomy
    {
        "span_text": "Subglottic",
        "context_prefix": "clean out the",
        "label": "ANAT_AIRWAY",
        "normalized_value": "subglottic",
        "event_id": "evt_asp_01"
    },
    {
        "span_text": "Trachea (Proximal 1/3)",
        "context_prefix": "Subglottic,",
        "label": "ANAT_AIRWAY",
        "normalized_value": "trachea_proximal",
        "event_id": "evt_asp_01"
    },
    {
        "span_text": "Trachea (Middle 1/3)",
        "context_prefix": "Proximal 1/3),",
        "label": "ANAT_AIRWAY",
        "normalized_value": "trachea_middle",
        "event_id": "evt_asp_01"
    },
    {
        "span_text": "Trachea (Distal 1/3)",
        "context_prefix": "Middle 1/3),",
        "label": "ANAT_AIRWAY",
        "normalized_value": "trachea_distal",
        "event_id": "evt_asp_01"
    },
    {
        "span_text": "Right Mainstem",
        "context_prefix": "Distal 1/3),",
        "label": "ANAT_AIRWAY",
        "normalized_value": "rm_bronchus",
        "event_id": "evt_asp_01"
    },
    {
        "span_text": "Bronchus Intermedius",
        "context_prefix": "Right Mainstem,",
        "label": "ANAT_AIRWAY",
        "normalized_value": "bronchus_intermedius",
        "event_id": "evt_asp_01"
    },
    {
        "span_text": "Left Mainstem",
        "context_prefix": "Bronchus Intermedius ,",
        "label": "ANAT_AIRWAY",
        "normalized_value": "lm_bronchus",
        "event_id": "evt_asp_01"
    },
    {
        "span_text": "Carina",
        "context_prefix": "Left Mainstem,",
        "label": "ANAT_AIRWAY",
        "normalized_value": "main_carina",
        "event_id": "evt_asp_01"
    },
    # BAL Details
    {
        "span_text": "Lateral Segment of RML",
        "context_prefix": "performed at",
        "label": "ANAT_LUNG_LOC",
        "normalized_value": "RML_lateral",
        "event_id": "evt_bal_01"
    },
    {
        "span_text": "Medial Segment of RML",
        "context_prefix": "and",
        "label": "ANAT_LUNG_LOC",
        "normalized_value": "RML_medial",
        "event_id": "evt_bal_01"
    },
    {
        "span_text": "40 cc",
        "context_prefix": "Instilled",
        "label": "MEAS_VOL",
        "normalized_value": "40",
        "event_id": "evt_bal_01"
    },
    {
        "span_text": "15 cc",
        "context_prefix": "suction returned with",
        "label": "MEAS_VOL",
        "normalized_value": "15",
        "event_id": "evt_bal_01"
    },
    # Trach Change Details
    {
        "span_text": "Shiley cuffed Trach",
        "context_prefix": "from:\n",
        "label": "DEV_CATHETER",
        "normalized_value": "shiley_cuffed_trach",
        "event_id": "evt_trach_01"
    },
    {
        "span_text": "7.0mm",
        "context_prefix": "Trach ISO/ID size",
        "label": "DEV_CATHETER_SIZE",
        "normalized_value": "7.0",
        "event_id": "evt_trach_01"
    },
    {
        "span_text": "Portex cuffed Trach",
        "context_prefix": "To",
        "label": "DEV_CATHETER",
        "normalized_value": "portex_cuffed_trach",
        "event_id": "evt_trach_01"
    },
    {
        "span_text": "7.0mm",
        "context_prefix": "To Portex cuffed Trach ISO/ID size",
        "label": "DEV_CATHETER_SIZE",
        "normalized_value": "7.0",
        "event_id": "evt_trach_01"
    },
    {
        "span_text": "erosion",
        "context_prefix": "however an",
        "label": "OBS_LESION",
        "normalized_value": "erosion",
        "event_id": "evt_trach_01"
    },
    {
        "span_text": "packed with iodoform gauze",
        "context_prefix": "erosion was",
        "label": "PROC_ACTION",
        "normalized_value": "packing",
        "event_id": "evt_stoma_01"
    },
    # Outcomes
    {
        "span_text": "No immediate complications",
        "context_prefix": "There were",
        "label": "OUTCOME_COMPLICATION",
        "normalized_value": "none",
        "event_id": "evt_glob_01"
    }
]

# Events Definitions for V3 sheets
EVENTS_INFO = [
    {
        "event_id": "evt_asp_01",
        "type": "therapeutic_aspiration",
        "method": "Aspiration",
        "anatomy": "Trachea, Mainstems, Lobar",
        "findings": "Mucus cleaned out"
    },
    {
        "event_id": "evt_bal_01",
        "type": "bal",
        "method": "Lavage",
        "anatomy": "RML (Lateral, Medial)",
        "measurements": "Instilled: 40cc, Return: 15cc"
    },
    {
        "event_id": "evt_trach_01",
        "type": "tracheostomy_tube_change", # Custom type for logic
        "method": "Tube Exchange",
        "anatomy": "Trachea",
        "device": "Shiley 7.0mm -> Portex 7.0mm",
        "findings": "Stoma erosion at 4 o'clock"
    },
    {
        "event_id": "evt_stoma_01",
        "type": "wound_care",
        "method": "Packing",
        "anatomy": "Tracheostomy Stoma",
        "device": "Iodoform Gauze"
    }
]

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def clean_text(text):
    """Normalize whitespace."""
    return re.sub(r'\s+', ' ', text).strip()

def hydrate_spans(text, spans):
    """
    Compute start_char and end_char for each span based on logic:
    1. Exact unique match.
    2. Context prefix match (preceding 120 chars).
    3. Match index (fallback).
    """
    hydrated = []
    
    for s in spans:
        span_text = s["span_text"]
        ctx = s.get("context_prefix", "")
        
        # Default state
        start = -1
        end = -1
        status = "ambiguous"
        
        # Find all occurrences
        matches = [m.start() for m in re.finditer(re.escape(span_text), text)]
        
        if len(matches) == 0:
            status = "not_found"
        elif len(matches) == 1:
            start = matches[0]
            status = "hydrated_unique"
        else:
            # Try context
            if ctx:
                # Normalize context for search (remove extra whitespace)
                norm_ctx = clean_text(ctx)
                found_ctx = False
                for m_start in matches:
                    # Look back 120 chars
                    window_start = max(0, m_start - 120)
                    window_text = clean_text(text[window_start:m_start])
                    if norm_ctx in window_text:
                        start = m_start
                        status = "hydrated_context"
                        found_ctx = True
                        break
                
                if not found_ctx:
                    # Fallback: if context looks like a direct prefix in source, try strict check
                    # Logic: just pick the first one if context fails? No, keep ambiguous.
                    # But for this script, we can try to be smart about ordinality if provided
                    pass
            
            # If still ambiguous and we have hardcoded knowledge (rare in auto, but valid here)
            # We skip match_index logic for simplicity unless strictly needed.
            if start == -1:
                 # Last resort: if we have multiple, pick first? No, unsafe.
                 # For the purpose of this script, we leave it as ambiguous if context fails.
                 status = f"ambiguous_count_{len(matches)}"
                 start = matches[0] # Default to first for visualization, but flag it
        
        if start != -1:
            end = start + len(span_text)
        
        row = s.copy()
        row["start_char"] = start
        row["end_char"] = end
        row["span_len"] = len(span_text)
        row["hydration_status"] = status
        hydrated.append(row)
        
    return hydrated

def generate_v3_json(note_id, events, spans):
    """Generate the V3 Registry JSON object."""
    # Group spans by event
    event_spans = {}
    for s in spans:
        eid = s.get("event_id")
        if eid:
            if eid not in event_spans:
                event_spans[eid] = []
            event_spans[eid].append(s)

    procedures = []
    
    for evt in events:
        eid = evt["event_id"]
        espans = event_spans.get(eid, [])
        
        # Build base object
        obj = {
            "event_id": eid,
            "type": evt["type"],
            "method": evt.get("method"),
            "target": {
                "anatomy_type": "airway", # default
                "location": {},
                "station": None
            },
            "devices": [],
            "measurements": [],
            "specimens": [],
            "findings": []
        }
        
        # Populate from spans
        for sp in espans:
            lbl = sp["label"]
            val = sp["normalized_value"]
            
            if lbl == "ANAT_LUNG_LOC":
                obj["target"]["location"]["lobe"] = val
            elif lbl == "MEAS_VOL":
                obj["measurements"].append({"type": "volume", "value": val, "unit": "cc"})
            elif lbl.startswith("DEV_"):
                obj["devices"].append({"type": lbl, "value": val})
            elif lbl == "OBS_LESION":
                obj["findings"].append(val)
        
        procedures.append(obj)
        
    return json.dumps({
        "schema_version": "3.0",
        "note_id": note_id,
        "procedures": procedures,
        "no_immediate_complications": True # derived from note
    }, indent=2)

# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    # 1. Load Template
    if not os.path.exists(TEMPLATE_PATH):
        # Create a dummy one if missing (for standalone testing validity, though prompt implies it exists)
        wb = openpyxl.Workbook()
        for name in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            wb.create_sheet(name)
        # Remove default
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # 2. Populate Note_Text
    ws_text = wb["Note_Text"]
    # Assuming headers: note_id, source_file, note_text
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 3. Populate Note_Index (Flags)
    ws_index = wb["Note_Index"]
    # Row construction: Meta fields + 30 flags
    meta_row = [
        SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Success", "Generated by Phase0 Script"
    ]
    
    # Flags keys strictly ordered as per prompt list? 
    # The prompt lists them in groups. We'll map them carefully.
    flag_keys = [
        "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", 
        "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy", 
        "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration", 
        "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation", 
        "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", 
        "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
        "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis", 
        "pleural_biopsy", "fibrinolytic_therapy"
    ]
    
    flags_values = [PROCEDURE_FLAGS.get(k, 0) for k in flag_keys]
    ws_index.append(meta_row + flags_values)

    # 4. Process Spans (Hydration)
    hydrated_spans = hydrate_spans(NOTE_TEXT, SPANS_DATA)

    # 5. Populate Span_Annotations (Anchor First - No offsets)
    ws_anno = wb["Span_Annotations"]
    # Columns: source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, start, end, len, label, norm_val, field, event_id, neg, hist, time, rev, comm, hyd_status
    
    for i, s in enumerate(hydrated_spans):
        span_id = f"{NOTE_ID}_span_{i+1:03d}"
        row = [
            SOURCE_FILE, NOTE_ID, span_id, "", s.get("context_prefix"), s["span_text"], "",
            "", "", f"=LEN(F{i+2})", # Formula for len
            s["label"], s["normalized_value"], "", s["event_id"],
            0, 0, "", "", "", "needs_hydration"
        ]
        ws_anno.append(row)

    # 6. Populate Span_Hydrated
    ws_hyd = wb["Span_Hydrated"]
    for i, s in enumerate(hydrated_spans):
        span_id = f"{NOTE_ID}_span_{i+1:03d}"
        row = [
            SOURCE_FILE, NOTE_ID, span_id, "", s.get("context_prefix"), s["span_text"], "",
            s["start_char"], s["end_char"], s["span_len"],
            s["label"], s["normalized_value"], "", s["event_id"],
            0, 0, "", "", "", s["hydration_status"]
        ]
        ws_hyd.append(row)

    # 7. Populate Event_Log
    ws_event = wb["Event_Log"]
    for evt in EVENTS_INFO:
        # source, note, id, type, method, anat, dev, gauge, station, count, meas, spec, find, hist, rev, comm, ...
        row = [
            SOURCE_FILE, NOTE_ID, evt["event_id"], evt["type"], evt.get("method", ""),
            evt.get("anatomy", ""), evt.get("device", ""), "", "", "", 
            evt.get("measurements", ""), "", evt.get("findings", ""), 
            0, "", "", "", "", "", "", "", "", "none"
        ]
        ws_event.append(row)

    # 8. Populate V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    for evt in EVENTS_INFO:
        # Simplification for script generation
        row = [
            NOTE_ID, evt["event_id"], evt["type"], 
            "", "", "", "", # Target breakdown
            "", "", # Lesion
            evt.get("method"), "", evt.get("measurements", ""), "", evt.get("findings", ""), "",
            "", "", "", "", "", "", "", ""
        ]
        ws_v3.append(row)

    # 9. Populate V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    json_str = generate_v3_json(NOTE_ID, EVENTS_INFO, hydrated_spans)
    ws_json.append(["3.0", NOTE_ID, json_str, 1])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Successfully generated {OUTPUT_PATH}")

if __name__ == "__main__":
    main()