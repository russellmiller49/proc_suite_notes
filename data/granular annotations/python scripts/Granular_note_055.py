import os
import re
import json
import datetime
import openpyxl
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION & INPUTS
# =============================================================================
NOTE_ID = "note_055"
SOURCE_FILE = "note_055.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# Full text content from the provided file
NOTE_TEXT = """NOTE_ID:  note_055 SOURCE_FILE: note_055.txt INDICATION FOR OPERATION:  [REDACTED]is a 67 year old-year-old male who presents with lung nodule.
PREOPERATIVE DIAGNOSIS: R91.8 Other nonspecific abnormal finding of lung field.
POSTOPERATIVE DIAGNOSIS:  R91.8 Other nonspecific abnormal finding of lung field.
PROCEDURE:  
31899 Unlisted Procedure (Trach Change with Mature Tract or Procedure NOS)
31645 Therapeutic aspiration initial episode
31624 Dx bronchoscope/lavage (BAL)    
31628 TBBX single lobe     
31632 TBBX additional lobes  
31629 TBNA single lobe   
31633 TBNA additional lobes   
31626 Fiducial marker placements, single or multiple     
31627 Navigational Bronchoscopy (computer assisted)
77012 Radiology / radiologic guidance for CT guided needle placement (CIOS)
76377 3D rendering with interpretation and reporting of CT, US, Tomo modality (ION Planning Station)
31652 EBUS sampling 1 or 2 nodes
31654 Radial EBUS for peripheral lesion
76982 Ultrasound 
Elastography, First Target Lesion
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
 
 
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required a Transbronchial Cryo biopsies, Robotic Navigation to more than one site, Multiple Bronchoalveolar lavages in different locations, Brushings at multiple locations, and Radial EBUS performed at multiple locations.
This resulted in >40% increased work due to Increased intensity, Technical difficulty of procedure, and Physical and mental effort required.
Apply to: 31628 TBBX single lobe     
31632 TBBX additional lobes  
31629 TBNA single lobe   
31633 TBNA additional lobes   
31627 Navigational Bronchoscopy (computer assisted).
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
Linear EBUS 
Radial EBUS
Ion Robotic Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
 
CT Chest scan was placed on separate planning station to generate 3D rendering of the pathway to target.
The navigational plan was reviewed and verified.  This was then loaded into robotic bronchoscopy platform.
Successful therapeutic aspiration was performed to clean out the Right Mainstem, Bronchus Intermedius , and Left Mainstem from mucus.
Ventilation Parameters:
Mode	RR	TV	PEEP	FiO2	Flow Rate	Pmean
vcv	15	350	16	100	10	20
 
RLL:
Robotic navigation bronchoscopy was performed with Ion platform.  Partial registration was used.
Ion robotic catheter was used to engage the Superior Segment of RLL (RB6).
Target lesion is about 0.5 cm in diameter.   Under navigational guidance the ion robotic catheter was advanced to 1.0 cm away from the planned target.
Radial EBUS was performed to confirm that the location of the nodule is Eccentric.
The following features were noted: Continuous margin  and Absence of linear-discrete air bronchogram.
Cone Beam CT was performed: 3-D reconstructions were performed on an independent workstation.
Cios Spin system was used for evaluation of nodule location.  Low dose spin was performed to acquire CT imaging.
This was passed on to Ion platform system for reconstruction and nodule location.
The 3D images was interpreted on an independent workstation (Ion).
Using the newly acquired nodule location, the Ion robotic system was adjusted to the new targeted location.
I personally interpreted the cone beam CT and 3-D reconstruction.
Transbronchial needle aspiration was performed with 21G Needle through the extended working channel catheter.  Total 4 samples were collected.
Samples sent for Cytology.
 
Transbronchial cryobiopsy was performed with 1.1mm cryoprobe via the extended working channel catheter.
Freeze time of 6 seconds were used.  Total 6 samples were collected.  Samples sent for Pathology.
Transbronchial brushing was performed with Protected cytology brush the extended working channel catheter.  Total 1 samples were collected.
Samples sent for Microbiology (Cultures/Viral/Fungal).
 
Bronchial alveolar lavage was performed the extended working channel catheter.
Instilled 10 cc of NS, suction returned with 5 cc of NS.  Samples sent for Microbiology (Cultures/Viral/Fungal).
Fiducial marker (0.8mm x 3mm soft tissue gold CIVCO) was loaded with bone wax and placed under fluoroscopy guidance.
Prior to withdraw of the bronchoscope. 
 
ROSE from ION procedure was noted to be:
Conclusive evidence of malignant neoplasm
 
Prior to withdrawal of the bronchoscope, inspection demonstrated no evidence of bleeding.
LLL:
 
Robotic navigation bronchoscopy was performed with Ion platform.  Partial registration was used.
Ion robotic catheter was used to engage the Posterior-Basal Segment of LLL (LB10).
Target lesion is about 0.5 cm in diameter.   Under navigational guidance the ion robotic catheter was advanced to 1.0 cm away from the planned target.
Radial EBUS was performed to confirm that the location of the nodule is Eccentric.
The following features were noted: Continuous margin .
 
Cone Beam CT was performed: 3-D reconstructions were performed on an independent workstation.
Cios Spin system was used for evaluation of nodule location.  Low dose spin was performed to acquire CT imaging.
This was passed on to Ion platform system for reconstruction and nodule location.
The 3D images was interpreted on an independent workstation (Ion).
Using the newly acquired nodule location, the Ion robotic system was adjusted to the new targeted location.
I personally interpreted the cone beam CT and 3-D reconstruction.
Transbronchial needle aspiration was performed with 21G Needle through the extended working channel catheter.  Total 6 samples were collected.
Samples sent for Cytology.
 
Transbronchial cryobiopsy was performed with 1.1mm cryoprobe via the extended working channel catheter.
Freeze time of 6 seconds were used.  Total 6 samples were collected.  Samples sent for Pathology.
Transbronchial brushing was performed with Protected cytology brush the extended working channel catheter.  Total 1 samples were collected.
Samples sent for Microbiology (Cultures/Viral/Fungal).
 
Bronchial alveolar lavage was performed the extended working channel catheter.
Instilled 10 cc of NS, suction returned with 5 cc of NS.  Samples sent for Microbiology (Cultures/Viral/Fungal).
Fiducial marker (0.8mm x 3mm soft tissue gold CIVCO) was loaded with bone wax and placed under fluoroscopy guidance.
Prior to withdraw of the bronchoscope. 
 
ROSE from ION procedure was noted to be:
Conclusive evidence of malignant neoplasm
 
Prior to withdrawal of the bronchoscope, inspection demonstrated no evidence of bleeding.
EBUS-Findings
Indications: Diagnostic
Technique:
All lymph node stations were assessed. Only those 5 mm or greater in short axis were sampled.
Lymph node sizing was performed by EBUS and sampling by transbronchial needle aspiration was performed using 22-gauge Needle.
Lymph Nodes/Sites Inspected: 7 (subcarinal) node
 
No immediate complications
 
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
Elastography provided a semi-quantitative classification (Type 1–3), which was used to guide biopsy site selection and sampling strategy.
Lymph Nodes Evaluated:
Site 1: The 7 (subcarinal) node was => 10 mm on CT and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was sampled.. 4 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 2 elastographic pattern with mixed soft and stiff regions.
Given this heterogeneous and indeterminate appearance, TBNA was directed at representative areas to ensure comprehensive sampling and to minimize the risk of underdiagnosis.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
7 TBNA
RLL TBCBX, TBNA, Brush, BAL
LLL TBCBX, TBNA, Brush, BAL
 
IMPRESSION/PLAN: [REDACTED]is a 67 year old-year-old male who presents for bronchoscopy for lung nodules.
- f/u in clinic"""

# =============================================================================
# DATA EXTRACTION LOGIC
# =============================================================================

def generate_extraction_data():
    """
    Returns a dictionary containing all extraction data:
    - flags (dict)
    - spans (list of dicts)
    - events (list of dicts)
    - v3_data (dict)
    """
    
    # 1. Procedure Flags
    flags = {
        # Bronchoscopy
        "diagnostic_bronchoscopy": 1,
        "bal": 1,
        "bronchial_wash": 0,
        "brushings": 1,
        "endobronchial_biopsy": 0,
        "tbna_conventional": 1,
        "linear_ebus": 1,
        "radial_ebus": 1,
        "navigational_bronchoscopy": 1,
        "transbronchial_biopsy": 1, # CPT 31628 mentioned + Cryo TBBX
        "transbronchial_cryobiopsy": 1,
        "therapeutic_aspiration": 1,
        "foreign_body_removal": 0,
        "airway_dilation": 0,
        "airway_stent": 0,
        "thermal_ablation": 0,
        "tumor_debulking_non_thermal": 0,
        "cryotherapy": 0, # Diagnostic cryo flagged above
        "blvr": 0,
        "peripheral_ablation": 0,
        "bronchial_thermoplasty": 0,
        "whole_lung_lavage": 0,
        "rigid_bronchoscopy": 0,
        # Pleural
        "thoracentesis": 0,
        "chest_tube": 0,
        "ipc": 0,
        "medical_thoracoscopy": 0,
        "pleurodesis": 0,
        "pleural_biopsy": 0,
        "fibrinolytic_therapy": 0,
    }

    # 2. Spans (Anchor First)
    spans = []
    
    def add_span(text, label, field, value=None, match_idx=0, context=None):
        if value is None:
            value = text
        spans.append({
            "span_text": text,
            "label": label,
            "schema_field": field,
            "normalized_value": value,
            "match_index": match_idx,
            "context_prefix": context
        })

    # Event 1: Therapeutic Aspiration
    add_span("Successful therapeutic aspiration", "PROC_METHOD", "method", "therapeutic_aspiration", 0)
    add_span("Right Mainstem", "ANAT_AIRWAY", "anatomy", "Right Mainstem", 0)
    add_span("Bronchus Intermedius", "ANAT_AIRWAY", "anatomy", "Bronchus Intermedius", 0)
    add_span("Left Mainstem", "ANAT_AIRWAY", "anatomy", "Left Mainstem", 0)
    
    # Event 2: RLL
    # Note: "Radial EBUS" appears in header (idx0), RLL (idx1), LLL (idx2). 
    # But checking text:
    # 1. Source 9: "Radial EBUS" (Instrument list)
    # 2. Source 16: "Radial EBUS was performed" (RLL)
    # 3. Source 35: "Radial EBUS was performed" (LLL)
    # We want indices 1 and 2 for the procedures.
    
    # RLL Anchors
    add_span("RLL", "ANAT_LUNG_LOC", "target.location.lobe", "RLL", 0, "Superior Segment of")
    add_span("Superior Segment", "ANAT_LUNG_LOC", "target.location.segment", "Superior", 0)
    add_span("RB6", "ANAT_LUNG_LOC", "target.location.segment", "RB6", 0)
    add_span("0.5 cm", "MEAS_SIZE", "lesion.size_mm", "5", 0, "Target lesion is about")
    add_span("Radial EBUS", "PROC_METHOD", "method", "Radial EBUS", 1, "navigational guidance") # Index 1 (Index 0 is in instrument list)
    add_span("Eccentric", "OBS_LESION", "lesion.type", "Eccentric", 0)
    add_span("Transbronchial needle aspiration", "PROC_METHOD", "method", "TBNA", 0, "I personally interpreted") # idx 0 is RLL
    add_span("21G Needle", "DEV_NEEDLE", "device", "21G", 0)
    add_span("Transbronchial cryobiopsy", "PROC_METHOD", "method", "Cryobiopsy", 1) # Idx 0 is in 'Unusual Procedure' header
    add_span("1.1mm cryoprobe", "DEV_INSTRUMENT", "device", "1.1mm cryoprobe", 0)
    add_span("Transbronchial brushing", "PROC_METHOD", "method", "Brushing", 0)
    add_span("Protected cytology brush", "DEV_INSTRUMENT", "device", "cytology brush", 0)
    add_span("Bronchial alveolar lavage", "PROC_METHOD", "method", "BAL", 0)
    add_span("Fiducial marker", "DEV_INSTRUMENT", "device", "Fiducial", 1) # Idx 0 is CPT code list, Idx 1 is RLL
    add_span("malignant neoplasm", "OBS_ROSE", "findings", "Malignant", 0)

    # Event 3: LLL
    add_span("LLL", "ANAT_LUNG_LOC", "target.location.lobe", "LLL", 0, "Posterior-Basal Segment of")
    add_span("Posterior-Basal Segment", "ANAT_LUNG_LOC", "target.location.segment", "Posterior-Basal", 0)
    add_span("LB10", "ANAT_LUNG_LOC", "target.location.segment", "LB10", 0)
    add_span("0.5 cm", "MEAS_SIZE", "lesion.size_mm", "5", 1) # 2nd occurrence
    add_span("Radial EBUS", "PROC_METHOD", "method", "Radial EBUS", 2) # 3rd occurrence
    add_span("Eccentric", "OBS_LESION", "lesion.type", "Eccentric", 1)
    add_span("Transbronchial needle aspiration", "PROC_METHOD", "method", "TBNA", 1) # idx 1 is LLL
    add_span("21G Needle", "DEV_NEEDLE", "device", "21G", 1)
    add_span("Transbronchial cryobiopsy", "PROC_METHOD", "method", "Cryobiopsy", 2) # Idx 2 is LLL
    add_span("1.1mm cryoprobe", "DEV_INSTRUMENT", "device", "1.1mm cryoprobe", 1)
    add_span("Transbronchial brushing", "PROC_METHOD", "method", "Brushing", 1)
    add_span("Protected cytology brush", "DEV_INSTRUMENT", "device", "cytology brush", 1)
    add_span("Bronchial alveolar lavage", "PROC_METHOD", "method", "BAL", 1)
    add_span("Fiducial marker", "DEV_INSTRUMENT", "device", "Fiducial", 2) # Idx 2 is LLL
    add_span("malignant neoplasm", "OBS_ROSE", "findings", "Malignant", 1)

    # Event 4: EBUS
    add_span("7 (subcarinal) node", "ANAT_LN_STATION", "target.station", "7", 0)
    add_span("22-gauge Needle", "DEV_NEEDLE", "device", "22G", 0)
    add_span("Elastography", "PROC_METHOD", "method", "Elastography", 0) # Index 0 (source 5 header?) Let's check.
    # "Elastography, First Target Lesion" (Header? No, Source 5).
    # "Endobronchial ultrasound (EBUS) elastography" (Source 52).
    # Let's use "Type 2" as the anchor for finding.
    add_span("Type 2", "OBS_LESION", "findings", "Type 2 Elastography", 0)
    add_span("No immediate complications", "OUTCOME_COMPLICATION", "outcomes.complications", "None", 1) # 2nd occurrence (EBUS section)

    # 3. Events Table
    events = []
    
    # Event 1
    events.append({
        "event_id": "EVT_01",
        "event_type": "Therapeutic",
        "method": "Therapeutic aspiration",
        "anatomy_target": "Right Mainstem, Bronchus Intermedius, Left Mainstem",
        "comments": "Mucus clearance"
    })
    
    # Event 2
    events.append({
        "event_id": "EVT_02",
        "event_type": "Diagnostic",
        "method": "Navigational Bronchoscopy",
        "anatomy_target": "RLL Superior Segment (RB6)",
        "device": "Ion Catheter, Radial EBUS, 21G Needle, 1.1mm Cryoprobe, Brush, Fiducial",
        "findings": "0.5 cm, Eccentric, Malignant",
        "specimens": "TBNA x4, Cryo x6, Brush x1, BAL"
    })

    # Event 3
    events.append({
        "event_id": "EVT_03",
        "event_type": "Diagnostic",
        "method": "Navigational Bronchoscopy",
        "anatomy_target": "LLL Posterior-Basal Segment (LB10)",
        "device": "Ion Catheter, Radial EBUS, 21G Needle, 1.1mm Cryoprobe, Brush, Fiducial",
        "findings": "0.5 cm, Eccentric, Malignant",
        "specimens": "TBNA x6, Cryo x6, Brush x1, BAL"
    })

    # Event 4
    events.append({
        "event_id": "EVT_04",
        "event_type": "Diagnostic",
        "method": "Linear EBUS",
        "anatomy_target": "Station 7",
        "device": "22G Needle",
        "findings": "Type 2 Elastography",
        "specimens": "TBNA x4"
    })

    # 4. V3 JSON Construction
    v3_data = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": [
            {
                "event_id": "EVT_01",
                "type": "Therapeutic",
                "method": "Therapeutic aspiration",
                "target": {
                    "anatomy_type": "Airway",
                    "location": {
                        "airway_regions": ["Right Mainstem", "Bronchus Intermedius", "Left Mainstem"]
                    }
                }
            },
            {
                "event_id": "EVT_02",
                "type": "Diagnostic",
                "method": "Navigational Bronchoscopy",
                "target": {
                    "anatomy_type": "Lung",
                    "location": {
                        "lobe": "RLL",
                        "segment": "Superior (RB6)"
                    }
                },
                "lesion": {
                    "size_mm": 5,
                    "type": "Eccentric"
                },
                "devices_json": ["Ion Catheter", "Radial EBUS", "21G Needle", "1.1mm Cryoprobe", "Cytology Brush", "Fiducial"],
                "specimens_json": ["TBNA", "Cryobiopsy", "Brush", "BAL"],
                "findings_json": ["Malignant (ROSE)"]
            },
            {
                "event_id": "EVT_03",
                "type": "Diagnostic",
                "method": "Navigational Bronchoscopy",
                "target": {
                    "anatomy_type": "Lung",
                    "location": {
                        "lobe": "LLL",
                        "segment": "Posterior-Basal (LB10)"
                    }
                },
                "lesion": {
                    "size_mm": 5,
                    "type": "Eccentric"
                },
                "devices_json": ["Ion Catheter", "Radial EBUS", "21G Needle", "1.1mm Cryoprobe", "Cytology Brush", "Fiducial"],
                "specimens_json": ["TBNA", "Cryobiopsy", "Brush", "BAL"],
                "findings_json": ["Malignant (ROSE)"]
            },
            {
                "event_id": "EVT_04",
                "type": "Diagnostic",
                "method": "Linear EBUS",
                "target": {
                    "anatomy_type": "Lymph Node",
                    "station": "7"
                },
                "devices_json": ["22G Needle"],
                "findings_json": ["Type 2 Elastography"]
            }
        ]
    }
    
    return flags, spans, events, v3_data

# =============================================================================
# WORKBOOK GENERATION
# =============================================================================

def process_template():
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Template not found at {TEMPLATE_PATH}")

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    flags, spans, events, v3_data = generate_extraction_data()

    # 1. Note_Text
    if "Note_Text" not in wb.sheetnames:
        wb.create_sheet("Note_Text")
    ws_text = wb["Note_Text"]
    if ws_text.max_row == 1:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    if "Note_Index" not in wb.sheetnames:
        wb.create_sheet("Note_Index")
    ws_index = wb["Note_Index"]
    
    # Headers logic (simplified for script brevity, assuming template headers exist)
    # We append the metadata + flags
    flag_keys = [
        "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", 
        "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy", 
        "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration", 
        "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation", 
        "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", 
        "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
        "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis", 
        "pleural_biopsy", "fibrinolytic_therapy"
    ]
    
    row_data = [
        SOURCE_FILE, NOTE_ID, "", "", "", "", "Pending", ""
    ] + [flags[k] for k in flag_keys]
    
    ws_index.append(row_data)

    # 3. Span_Annotations & 4. Span_Hydrated
    if "Span_Annotations" not in wb.sheetnames:
        wb.create_sheet("Span_Annotations")
    ws_span = wb["Span_Annotations"]
    
    if "Span_Hydrated" not in wb.sheetnames:
        wb.create_sheet("Span_Hydrated")
    ws_hydra = wb["Span_Hydrated"]
    
    # Define Span Headers (Standard)
    span_headers = [
        "source_file", "note_id", "span_id", "section_type", "context_prefix",
        "span_text", "match_index", "start_char", "end_char", "span_len",
        "label", "normalized_value", "schema_field", "event_id", "is_negated",
        "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"
    ]
    
    # Write headers if empty
    if ws_span.max_row == 1 and ws_span.cell(1,1).value is None:
        ws_span.append(span_headers)
    if ws_hydra.max_row == 1 and ws_hydra.cell(1,1).value is None:
        ws_hydra.append(span_headers)

    span_id_counter = 1
    
    for s in spans:
        # Hydration Logic
        span_text = s["span_text"]
        match_idx = s["match_index"]
        context = s.get("context_prefix", "")
        
        start = -1
        end = -1
        hydration_status = "unhydrated"
        
        # Find all occurrences
        occurrences = [m.start() for m in re.finditer(re.escape(span_text), NOTE_TEXT)]
        
        if len(occurrences) == 0:
            hydration_status = "not_found"
        elif len(occurrences) == 1:
            start = occurrences[0]
            hydration_status = "hydrated_unique"
        else:
            # Multiple occurrences
            if context:
                # Try context window
                found_with_context = False
                for occ in occurrences:
                    window_start = max(0, occ - 120)
                    preceding_text = NOTE_TEXT[window_start:occ]
                    if context in preceding_text:
                        start = occ
                        hydration_status = "hydrated_context"
                        found_with_context = True
                        break
                if not found_with_context:
                    # Fallback to index if provided
                    if match_idx < len(occurrences):
                        start = occurrences[match_idx]
                        hydration_status = "hydrated_index_fallback"
                    else:
                        hydration_status = "ambiguous_index_out_of_bounds"
            else:
                # Use index
                if match_idx < len(occurrences):
                    start = occurrences[match_idx]
                    hydration_status = "hydrated_index"
                else:
                    hydration_status = "ambiguous_index_out_of_bounds"

        if start != -1:
            end = start + len(span_text)

        # Base Row
        row = [
            SOURCE_FILE, NOTE_ID, f"SPAN_{span_id_counter:03d}", "", s.get("context_prefix"),
            span_text, match_idx, "", "", f"=LEN(F{ws_span.max_row + 1})",
            s["label"], s["normalized_value"], s["schema_field"], "", 
            "FALSE", "FALSE", "", "", "", "needs_hydration"
        ]
        ws_span.append(row)
        
        # Hydrated Row
        row_h = list(row)
        row_h[7] = start if start != -1 else ""
        row_h[8] = end if end != -1 else ""
        row_h[9] = len(span_text)
        row_h[19] = hydration_status
        ws_hydra.append(row_h)
        
        span_id_counter += 1

    # 5. Event_Log
    if "Event_Log" not in wb.sheetnames:
        wb.create_sheet("Event_Log")
    ws_event = wb["Event_Log"]
    
    # Headers
    event_headers = [
        "source_file", "note_id", "event_id", "event_type", "method", "anatomy_target",
        "device", "needle_gauge", "stations", "counts", "measurements", "specimens",
        "findings", "is_historical", "reviewer", "comments", "device_size", "device_material",
        "outcome_airway_lumen_pre", "outcome_airway_lumen_post", "outcome_symptoms", 
        "outcome_pleural", "outcome_complication"
    ]
    if ws_event.max_row == 1 and ws_event.cell(1,1).value is None:
        ws_event.append(event_headers)

    for e in events:
        row = [
            SOURCE_FILE, NOTE_ID, e["event_id"], e["event_type"], e["method"], e["anatomy_target"],
            e.get("device", ""), "", "", "", "", e.get("specimens", ""),
            e.get("findings", ""), "FALSE", "", e.get("comments", ""), "", "",
            "", "", "", "", ""
        ]
        ws_event.append(row)

    # 6. V3_Procedure_Events
    if "V3_Procedure_Events" not in wb.sheetnames:
        wb.create_sheet("V3_Procedure_Events")
    ws_v3 = wb["V3_Procedure_Events"]
    
    v3_headers = [
        "note_id", "event_id", "type", 
        "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station",
        "lesion.type", "lesion.size_mm", "method", 
        "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote",
        "stent.size", "stent.material_or_brand", "catheter.size_fr",
        "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post", 
        "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"
    ]
    if ws_v3.max_row == 1 and ws_v3.cell(1,1).value is None:
        ws_v3.append(v3_headers)

    for p in v3_data["procedures"]:
        t = p.get("target", {})
        l = t.get("location", {})
        les = p.get("lesion", {})
        
        row = [
            NOTE_ID, p["event_id"], p["type"],
            t.get("anatomy_type"), l.get("lobe"), l.get("segment"), t.get("station"),
            les.get("type"), les.get("size_mm"), p["method"],
            json.dumps(p.get("devices_json", [])), 
            json.dumps(p.get("measurements_json", [])),
            json.dumps(p.get("specimens_json", [])),
            json.dumps(p.get("findings_json", [])),
            "", "", "", "", "", "", "", "", ""
        ]
        ws_v3.append(row)

    # 7. V3_Registry_JSON
    if "V3_Registry_JSON" not in wb.sheetnames:
        wb.create_sheet("V3_Registry_JSON")
    ws_json = wb["V3_Registry_JSON"]
    ws_json.append([json.dumps(v3_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    # Check for text file override
    if os.path.exists(SOURCE_FILE):
        try:
            with open(SOURCE_FILE, 'r', encoding='utf-8') as f:
                content = f.read()
                if content.strip():
                    NOTE_TEXT = content
                    print(f"Loaded text from {SOURCE_FILE}")
        except Exception as e:
            print(f"Error reading file, using embedded text: {e}")
            
    process_template()