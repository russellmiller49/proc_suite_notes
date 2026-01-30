import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os
from datetime import datetime

# =============================================================================
# INPUT DATA
# =============================================================================

NOTE_ID = "note_056"
SOURCE_FILE = "note_056.txt"
PROCEDURE_DATE = "2026-01-12" # inferred from context or left blank if unknown, using context date
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_056 SOURCE_FILE: note_056.txt INDICATION FOR OPERATION:  [REDACTED]is a 91 year old-year-old female who presents with airway obstruction.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
 
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
 
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31624 Dx bronchoscope/lavage (BAL)    
31899NFN BRONCHOSCOPY WITH ENDOBRONCHIAL ULTRASOUND (EBUS) OF MEDIASTINAL AND/OR HILAR LYMPH NODES WITHOUT BIOPSY
31630 Balloon dilation
31631 Dilate and tracheal stent placement   
43200 Esophagoscopy, flexible, transoral;
diagnostic, including collection of specimen(s) by brushing or washing, when performed
 
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)
 
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required a complex tracheal stent placement.
This resulted in >100% increased work due to Increased intensity, Time, Technical difficulty of procedure, and Physical and mental effort required.
Apply to: 31631 Dilate and tracheal stent placement   .
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Rigid Bronchoscope
Flexible Therapeutic Bronchoscope
Linear EBUS 
 
ESTIMATED BLOOD LOSS:   Minimum
 
COMPLICATIONS:    Injury to posterior distal trachea
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: Supine
 
Initial Airway Inspection Findings:
  
The laryngeal mask airway is in good position.
Pharynx: Not assessed due to bronchoscopy introduction through LMA.
Larynx: Mild edema and erythema
Vocal Cords: Normal without mass/lesions
Trachea: Normal.
Main Carina: Sharp
Right Lung Proximal Airways: Normal anatomic branching to segmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Left Lung Proximal Airways: Normal anatomic branching to segmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Mucosa: Friable.
Secretions: Copious thick and thin, clear to tan secretions/mucus throughout with left > right.
ERYTHEMATOUS AND EDEMATOUS SUPRAGLOTTIC SPACE
 
 
Successful therapeutic aspiration was performed to clean out the Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus.
Bronchial alveolar lavage was performed at Anterior Segment of LUL (LB3).
Instilled 40 cc of NS, suction returned with 20 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
 
After initial airway inspection and BAL the therapeutic bronchoscope was removed and the linear EBUS scope introduced via the LMA.
The patient's trachea was inspected and particular attention was paid to the subglottic posterior trachea where suspicious tissue was identified on cross-sectional imaging.
No suspicious lesion was identified and no samples were taken.
The linear EBUS was removed and repeat inspeciton with the therapeutic scope revealed a full thickness tear in the airway, presumably from the EBUS scope.
POSTERIOR MEMBRANE DEFECT
Advanced Gastroenterology service was consulted intraoperatively and came to the OR for consultation.
The defect was observed from within the trachea and the bronchoscope was introduced into the defect where no esophageal lumen was identified, only soft tissue c/w mediastinum.
The scope was subsequently introduced into the esophagus to evaluate for esophageal injury.
With Gastroenterology present, the length of the esophagus was examined and no esophageal defect/injury was identified.
Decision was then made to place a tracheal stent across the defect and forgo esophageal stenting.
After induction of muscle relaxants, tooth or gum protector was placed.
The black long rigid barrel was introduced through the mouth and advanced in the midline while keeping the alignment with the axis of the trachea and minimizing pressure to the teeth.
The vocal cords were identified and the rigid bronchoscope was advanced carefully while minimizing contact with them.
Once the rigid bronchoscope was positioned at the mid-trachea, jet ventilation (14 DP, 80 F, 100% FiO2) was initiated and chest wall movement was confirmed.
Intubation was easy. 
 
 
Measurements were taken of the posterior membrane defect as follows:
 
 	 
Distance from vocal folds to membrane defect	 8 cm
Distance from top of defect to bottom 	2.5 cm
Distance of bottom of defect to main carina	3 cm
 
The black rigid bronchoscope was removed and the patient was reintubated with the yellow rigid bronchoscope in the standard fashion as described above.
Intubation was easy. 
 
A silicone Y-stent (16mm x 13mm x 13mm) was initially modified to have tracheal portion cover the defect with short bronchial limbs for migration prevention.
However, significant difficulty was encountered advancing the stent into the deployer, eventually causing a mild tear of a bronchial limb.
Thus, the stent was modified to a tubular tracheal stent measuring 16mmx 45mm.
The modified tracheal stent was loaded on to the rigid stent loader and placed in the trachea.
The rigid forceps were then used to grasp the silicone stent and position it squarely across the tracheal defect.
Balloon dilation was performed at Trachea (Distal 1/3) within the silicone stent to seat it in place.
12/13.5/15 Elation balloon was used to perform dilation to 15 mm at the Trachea (Distal 1/3) (within the silicone stent).
Total 1 inflations with dilation time of 30 seconds each.
SILICONE STENT AT THE CONCLUSION OF THE PROCEDURE 
The patient tolerated the procedure well.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
--LUL BAL (cell count, micro, cyto)
 
IMPRESSION/PLAN: [REDACTED]is a 91 year old-year-old female who presents for bronchoscopy for evaluation of airway occlusion.
No occlusion was identified on bronchoscopic evaluation. Copious mucus/secretions were suctioned from throughout the airways and it is surmised that the occlusion on prior cross-sectional imaging was from mucus impaction.
The patient was examined with endobronchial ultrasound but no suspicious lesions were identified and no biopsies were performed.
On re-inspection the patient had a posterior membrane defect at the distal trachea.
A silicone tracheal stent was placed via rigid bronchoscopy to span the defect.
--Starte stent Hydration Regimen TID in the following order (this has been ordered): 
1) Albuterol nebs
2) Hypertonic saline (3%) nebs
3) Flutter valve
--Guaifenesin 1200mg PO BID
--Please engage case management to obtain nebulizer and medications (as above) for patient prior to discharge as patient will require this regimen as long as he has stent in place
--Follow up BAL results
--Anticipate bronchoscopic re-evaluation of stent in appx 2 weeks"""

# =============================================================================
# DATA STRUCTURES & PARSING
# =============================================================================

# 1. Procedure Flags (30 flags)
# Based on analysis of the note:
# - Diagnostic bronch: Yes (initial scope mentioned)
# - BAL: Yes
# - EBUS: Yes (Linear)
# - Therapeutic Aspiration: Yes
# - Airway Dilation: Yes (Balloon)
# - Airway Stent: Yes (Silicone)
# - Rigid Bronch: Yes
# - All others: No/Not mentioned
PROCEDURE_FLAGS = {
    # Bronchoscopy (23)
    "diagnostic_bronchoscopy": 1,
    "bal": 1,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 0,
    "airway_dilation": 1,
    "airway_stent": 1,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 1,
    # Pleural (7)
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 2. Spans (Anchor-First)
# Format: (span_text, label, normalized_value, schema_field, event_id, context_prefix)
spans_data = [
    # Event 1: Therapeutic Aspiration
    ("therapeutic aspiration", "PROC_METHOD", "therapeutic aspiration", "method", "ev1", "Successful "),
    ("clean out the Trachea", "PROC_ACTION", "clean airway", "findings_json", "ev1", None),
    ("Copious thick and thin, clear to tan secretions", "OBS_ROSE", "mucus secretions", "findings_json", "ev1", None),
    
    # Event 2: BAL
    ("Bronchial alveolar lavage", "PROC_METHOD", "BAL", "method", "ev2", None),
    ("Anterior Segment of LUL (LB3)", "ANAT_LUNG_LOC", "LUL Anterior Segment (LB3)", "target.location.segment", "ev2", None),
    ("Instilled 40 cc", "MEAS_VOL", "40 cc", "measurements_json", "ev2", None),
    ("returned with 20 cc", "MEAS_VOL", "20 cc", "measurements_json", "ev2", None),
    
    # Event 3: EBUS & Complication
    ("linear EBUS scope", "PROC_METHOD", "Linear EBUS", "method", "ev3", None),
    ("subglottic posterior trachea", "ANAT_AIRWAY", "Trachea (Posterior)", "target.anatomy_type", "ev3", None),
    ("No suspicious lesion was identified", "OBS_LESION", "No lesion", "findings_json", "ev3", None),
    
    # The complication is attributed to EBUS
    ("full thickness tear in the airway", "OUTCOME_COMPLICATION", "airway tear", "outcomes.complications", "ev3", None),
    ("presumably from the EBUS scope", "CTX_HISTORICAL", "cause of injury", "evidence_quote", "ev3", None),
    ("Injury to posterior distal trachea", "OUTCOME_COMPLICATION", "tracheal injury", "outcomes.complications", "ev3", None),
    
    # Complication measurements
    ("Distance from top of defect to bottom 	2.5 cm", "MEAS_SIZE", "2.5 cm", "measurements_json", "ev3", None),
    
    # Event 4: Esophagoscopy (Diagnostic Rule-out)
    ("introduced into the esophagus", "PROC_METHOD", "Esophagoscopy", "method", "ev4", None),
    ("no esophageal defect/injury", "OBS_LESION", "No injury", "findings_json", "ev4", None),
    
    # Event 5: Rigid Bronchoscopy & Stent Placement
    ("rigid bronchoscope", "PROC_METHOD", "Rigid Bronchoscopy", "method", "ev5", "identified and the "),
    ("black long rigid barrel", "DEV_INSTRUMENT", "Rigid Bronchoscope (Black)", "devices_json", "ev5", None),
    ("yellow rigid bronchoscope", "DEV_INSTRUMENT", "Rigid Bronchoscope (Yellow)", "devices_json", "ev5", None),
    
    # Stent (Successful one)
    ("tubular tracheal stent", "DEV_STENT", "Tubular Stent", "devices_json", "ev5", None),
    ("16mmx 45mm", "DEV_STENT_SIZE", "16x45mm", "stent.size", "ev5", None),
    ("silicone stent", "DEV_STENT_MATERIAL", "Silicone", "stent.material_or_brand", "ev5", "grasp the "),
    ("position it squarely across the tracheal defect", "PROC_ACTION", "stent placement", "findings_json", "ev5", None),
    
    # Event 6: Balloon Dilation
    ("Balloon dilation", "PROC_METHOD", "Balloon Dilation", "method", "ev6", None),
    ("Trachea (Distal 1/3)", "ANAT_AIRWAY", "Trachea (Distal 1/3)", "target.anatomy_type", "ev6", "dilation was performed at "),
    ("12/13.5/15 Elation balloon", "DEV_INSTRUMENT", "Elation Balloon (12-15mm)", "devices_json", "ev6", None),
    ("dilation to 15 mm", "MEAS_AIRWAY_DIAM", "15 mm", "outcomes.airway.lumen_post", "ev6", None),
    ("Total 1 inflations", "MEAS_COUNT", "1", "measurements_json", "ev6", None),
    
    # General Outcomes
    ("patient tolerated the procedure well", "OUTCOME_SYMPTOMS", "Tolerated well", "outcomes.symptoms", "ev5", None),
]

# 3. Events Definition for Log & V3
events_info = {
    "ev1": {
        "type": "Therapeutic Aspiration",
        "anatomy": "Trachea/Bronchi",
        "findings": "Copious secretions, cleared",
        "outcome_symptoms": "Improvement implied (cleared)"
    },
    "ev2": {
        "type": "BAL",
        "anatomy": "LUL Anterior Segment",
        "specimens": "Cell count, Micro, Cyto",
        "measurements": "Instilled 40cc, Return 20cc"
    },
    "ev3": {
        "type": "EBUS",
        "anatomy": "Trachea (Posterior)",
        "findings": "No lesion seen, but caused Tear",
        "outcome_complication": "Full thickness airway tear"
    },
    "ev4": {
        "type": "Esophagoscopy",
        "anatomy": "Esophagus",
        "findings": "No injury identified"
    },
    "ev5": {
        "type": "Airway Stent Placement",
        "method": "Rigid Bronchoscopy",
        "device": "Silicone Tubular Stent",
        "size": "16x45mm",
        "anatomy": "Trachea",
        "outcome_symptoms": "Tolerated well"
    },
    "ev6": {
        "type": "Balloon Dilation",
        "anatomy": "Trachea (Distal 1/3)",
        "device": "Elation Balloon",
        "measurements": "Dilated to 15mm"
    }
}

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def clean_text(text):
    if not text: return ""
    return re.sub(r'[\r\n]+', ' ', text).strip()

def find_offset(full_text, span_text, context_prefix=None, match_index=0):
    """
    Returns (start_char, end_char, message)
    Logic:
    1. If context_prefix provided, find occurrence preceded by prefix.
    2. Else if unique, return it.
    3. Else use match_index.
    """
    if not span_text in full_text:
        return None, None, "not_found"
    
    # Escape regex special chars in span_text
    escaped_span = re.escape(span_text)
    
    # 1. Context Match
    if context_prefix:
        # Look for prefix within ~120 chars before span
        pattern = re.compile(re.escape(context_prefix) + r".{0,120}?" + escaped_span, re.DOTALL)
        match = pattern.search(full_text)
        if match:
            # The span is at the end of the match
            start = match.end() - len(span_text)
            end = match.end()
            return start, end, "hydrated_prefix_window"
        else:
            return None, None, "context_miss"

    # 2. All matches
    matches = [m for m in re.finditer(escaped_span, full_text)]
    
    if len(matches) == 1:
        return matches[0].start(), matches[0].end(), "hydrated_unique"
    
    if len(matches) > 1:
        if match_index < len(matches):
            return matches[match_index].start(), matches[match_index].end(), "hydrated_match_index"
        else:
            return None, None, f"ambiguous_count={len(matches)}"
            
    return None, None, "error"

# =============================================================================
# MAIN GENERATION
# =============================================================================

def generate_workbook():
    # 1. Load Template
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        # Create blank if template missing (fallback)
        wb = openpyxl.Workbook()
        for sheet in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            wb.create_sheet(sheet)
        # Remove default
        if "Sheet" in wb.sheetnames: del wb["Sheet"]

    # 2. Note_Text
    ws_text = wb["Note_Text"]
    # Ensure headers if empty
    if ws_text.max_row == 1 and ws_text.cell(1,1).value is None:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 3. Note_Index
    ws_index = wb["Note_Index"]
    # Row 1 is headers usually. Row 2 is data.
    # We append a new row.
    # Metadata cols: source_file, note_id, encounter_id, procedure_date, site, reviewer, status, free_text_notes
    meta_row = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Success", ""]
    
    # Flags keys order must match the template or list defined above.
    # Assuming template matches the keys in PROCEDURE_FLAGS dictionary order.
    # We will write keys in a fixed order for safety.
    flag_keys = list(PROCEDURE_FLAGS.keys())
    flag_values = [PROCEDURE_FLAGS[k] for k in flag_keys]
    
    ws_index.append(meta_row + flag_values)

    # 4. Span_Annotations (Anchor First) & Span_Hydrated
    ws_anno = wb["Span_Annotations"]
    ws_hydra = wb["Span_Hydrated"]
    
    # Headers check
    headers = ["source_file", "note_id", "span_id", "section_type", 
               "context_prefix", "span_text", "match_index", 
               "start_char", "end_char", "span_len", 
               "label", "normalized_value", "schema_field", "event_id", 
               "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"]
    
    if ws_anno.max_row == 1 and ws_anno.cell(1,1).value is None:
        ws_anno.append(headers)
        ws_hydra.append(headers)

    span_counter = 1
    
    for item in spans_data:
        text, label, norm, schema, ev_id, ctx = item
        
        span_id = f"{NOTE_ID}_s{span_counter:02d}"
        
        # Hydration
        start, end, status = find_offset(NOTE_TEXT, text, context_prefix=ctx)
        length = len(text)
        
        # Row for Annotations (Blank offsets)
        row_anno = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure",
            ctx, text, 0,
            "", "", f"=LEN(F{ws_anno.max_row+1})", # Formula for len
            label, norm, schema, ev_id,
            False, False, "", "", "", "needs_hydration"
        ]
        ws_anno.append(row_anno)
        
        # Row for Hydrated
        row_hydra = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure",
            ctx, text, 0,
            start, end, length,
            label, norm, schema, ev_id,
            False, False, "", "", "", status
        ]
        ws_hydra.append(row_hydra)
        
        span_counter += 1

    # 5. Event_Log
    ws_log = wb["Event_Log"]
    # Columns: source_file, note_id, event_id, event_type, method, anatomy_target, device, needle_gauge, stations, counts, measurements, specimens, findings, is_historical, reviewer, comments, device_size, device_material, outcome_airway_lumen_pre, outcome_airway_lumen_post, outcome_symptoms, outcome_pleural, outcome_complication
    
    if ws_log.max_row == 1 and ws_log.cell(1,1).value is None:
         ws_log.append(["source_file", "note_id", "event_id", "event_type", "method", "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements", "specimens", "findings", "is_historical", "reviewer", "comments", "device_size", "device_material", "outcome_airway_lumen_pre", "outcome_airway_lumen_post", "outcome_symptoms", "outcome_pleural", "outcome_complication"])

    for ev_id, info in events_info.items():
        row = [
            SOURCE_FILE, NOTE_ID, ev_id,
            info.get("type", ""),
            info.get("method", ""),
            info.get("anatomy", ""),
            info.get("device", ""),
            "", "", "", # needle, stations, counts
            info.get("measurements", ""),
            info.get("specimens", ""),
            info.get("findings", ""),
            False, "", "",
            info.get("size", ""), # device_size
            info.get("material", ""), # device_material
            "", # lumen_pre
            "", # lumen_post
            info.get("outcome_symptoms", ""),
            "", # outcome_pleural
            info.get("outcome_complication", "")
        ]
        ws_log.append(row)

    # 6. V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    # Header check omitted for brevity, assuming template
    
    for ev_id, info in events_info.items():
        # Mapping dict to V3 columns structure
        # note_id, event_id, type, target.anatomy, ...
        row = [
            NOTE_ID, ev_id, info.get("type"),
            "Airway" if "Trachea" in info.get("anatomy", "") else "Lung", # simplified
            "", "", "", # lobe/seg/station
            "", "", # lesion type/size
            info.get("method", ""),
            json.dumps({"device": info.get("device")}) if info.get("device") else "",
            json.dumps({"measurements": info.get("measurements")}) if info.get("measurements") else "",
            json.dumps({"specimens": info.get("specimens")}) if info.get("specimens") else "",
            json.dumps({"findings": info.get("findings")}) if info.get("findings") else "",
            "", # quote
            info.get("size", ""), # stent size
            info.get("material", ""), # stent material
            "", # cath size
            "", "", # lumens
            info.get("outcome_symptoms", ""),
            "", # pleural
            info.get("outcome_complication", "")
        ]
        ws_v3.append(row)

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 1:
        ws_json.append(["schema_version", "note_id", "json_output"])
    
    # Construct JSON object
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": [],
        "no_immediate_complications": False # Set to false because of the tear
    }
    
    for ev_id, info in events_info.items():
        event_obj = {
            "event_id": ev_id,
            "type": info.get("type"),
            "target": {"anatomy": info.get("anatomy")},
            "method": info.get("method"),
            "outcomes": {
                "complications": info.get("outcome_complication"),
                "symptoms": info.get("outcome_symptoms")
            }
        }
        registry_data["procedures"].append(event_obj)
        
    ws_json.append(["3.0", NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()