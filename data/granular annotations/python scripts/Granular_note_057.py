import openpyxl
from openpyxl.utils import get_column_letter
import json
import datetime
import re
import os

# =============================================================================
# INPUT DATA & CONFIGURATION
# =============================================================================
NOTE_ID = "note_057"
SOURCE_FILE = "note_057.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_057 SOURCE_FILE: note_057.txt INDICATION FOR OPERATION:  [REDACTED]is a 68 year old-year-old female who presents with airway stenosis.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
PROCEDURE:  
31646 Therapeutic aspiration subsequent episodes
31622 Dx bronchoscope/cell washing          
31624 Dx bronchoscope/lavage (BAL)    
31628 TBBX single lobe     
31632 TBBX additional lobes  
31627 Navigational Bronchoscopy (computer assisted)
77012 Radiology / radiologic guidance for CT guided needle placement (CIOS)
76377 3D rendering with interpretation and reporting of CT, US, Tomo modality (ION Planning Station)
31654 Radial EBUS for peripheral lesion
31630 Balloon dilation
31641 Destruction of tumor OR relief of 
stenosis by any method other than excision (eg. laser therapy, cryotherapy)
31635 Foreign body removal
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required a extensive ablation and debulking with excision along the left mainstem and left lower lobe with multiple dilations and multiple episodes of excision between other therapeutic modalities.
This resulted in >90% increased work due to Increased intensity, Time, Technical difficulty of procedure, and Physical and mental effort required.
Apply to: 31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy).
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
Radial EBUS
Ion Robotic Bronchoscope
Disposable Bronchoscope
ESTIMATED BLOOD LOSS:   Moderate
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
Initial Airway Inspection Findings:
The airway was inspected.
The trachea and right lung were intact and appropriate to the segmental level.
The left mainstem was inflamed and the left upper lobe was visualized, but the left lower lobe was again not visualized.
CT Chest scan was placed on separate planning station to generate 3D rendering of the pathway to target.
The navigational plan was reviewed and verified.  This was then loaded into robotic bronchoscopy platform.
Ventilation Parameters:
Mode	RR	TV	PEEP	FiO2	Flow Rate	Pmean
VCV	 	300	12	100	10	15
 
Robotic navigation bronchoscopy was performed with Ion platform.  Partial registration was used.
Ion robotic catheter was used to engage the left lower lobe presumed oriface.
Target lesion is about 1 cm in diameter.   Under navigational guidance the ion robotic catheter was advanced to 1.0 cm away from the planned target.
Radial EBUS was performed to confirm that the location of the target is not well visuzlized.
The following features were noted: soft tissue.
Cone Beam CT was performed: 3-D reconstructions were performed on an independent workstation.
Cios Spin system was used for evaluation of target location.  Low dose spin was performed to acquire CT imaging.
This was passed on to Ion platform system for reconstruction and nodule location.
The 3D images was interpreted on an independent workstation (Ion).
Using the newly acquired nodule location, the Ion robotic system was adjusted to the new targeted location.
I personally interpreted the cone beam CT and 3-D reconstruction.
The target airway was visualized and targets were placed ensure optimal trajectory of the tools to open the obstructed airway.
Transbronchial needle aspiration was performed with 19G Needle through the extended working channel catheter.  Total 2 samples were collected.
Samples sent for Microbiology (Cultures/Viral/Fungal) and Cytology.
Transbronchial biopsy was performed with precisor the extended working channel catheter.
Total 2 samples were collected.  Samples sent for Microbiology (Cultures/Viral/Fungal) and Pathology.
The robotic bronchoscope was advanced until airway lumen was visualized with the vision probe.
The catheter was then drawn back and the robotic bronchoscope was disconnected.
The therapeutic bronchoscope was then used to clear the airway. The left lung was irrigated with saline.
Granulation tissue and inflamed tissue was excised with the precisor, pulmonary, and large forceps.
Balloon dilation was performed at left lower lobe orifice.  8/9/10 Elation balloon was used to perform dilation to 10 mm at the left lower lobe orifice.
Total 1 inflations with dilation time of 60 seconds each.
Balloon dilation was performed at left lower lobe orifice.
10/11/12 Elation balloon was used to perform dilation to 12 mm at the left lower lobe orifice.
Total 3 inflations with dilation time of 60 seconds each.
The left upper lobe and the left lower lobe were better visualized.
Serial irrigation with iced saline was performed with excision of granulation tissue and blood clot/foreign body.
The 2.5 cm microwave catheter was used to ablate the inflamed tissue for three minutes(3.5KJ and 67 degrees Celsius reached), then a pause for further and extensive excision, followed by one minute (1.3 KJ and 49 degrees Celsius reached)
The patient tolerated the procedure well.
There were no immediate complications.  At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
•	Left lower lobe transbronchial needle aspiration, transbronchial forceps biopsies
•	Left mainstem and left lower lobe biopsies
•	Left lower lobe bronchoalveolar lavage
IMPRESSION/PLAN: [REDACTED]is a 68 year old-year-old female who presents for bronchoscopy for airway stenosis.
-Follow up CXR
-Follow up bronchoscopic lab work
-Plan for follow up bronchoscopy in ~1 week"""

# =============================================================================
# STATIC DATA: FLAGS & SPANS
# =============================================================================

# Procedure Flags (30 items)
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 1,  # "31624 Dx bronchoscope/lavage (BAL)"
    "bronchial_wash": 1, # "31622 Dx bronchoscope/cell washing"
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 1, # "19G Needle" - usually conventional or robotic TBNA. Note describes "Ion robotic catheter" + TBNA. Mapping as conv for this schema if robotic specific flag absent, or strictly 1 if needle used.
    "linear_ebus": 0,
    "radial_ebus": 1, # "31654 Radial EBUS"
    "navigational_bronchoscopy": 1, # "Ion"
    "transbronchial_biopsy": 1, # "31628 TBBX"
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1, # "31646 Therapeutic aspiration"
    "foreign_body_removal": 0, # Listed in CPTs but text says "excision of... blood clot/foreign body" in context of cleaning airway. Often incidental. Let's keep 0 unless specific retrieval device mentioned for non-biological FB. Note says "excision...".
    "airway_dilation": 1, # "31630 Balloon dilation"
    "airway_stent": 0,
    "thermal_ablation": 1, # "microwave catheter"
    "tumor_debulking_non_thermal": 1, # "excised with the precisor... forceps"
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# Anchor-First Spans
# Structure: (context_prefix, span_text, label, normalized_value, event_id)
# Note: context_prefix is used for hydration if span_text is not unique.
RAW_SPANS = [
    # Diagnosis
    ("", "airway stenosis", "OBS_LESION", "Airway Stenosis", "evt_dx"),

    # Event 1: Navigation / Imaging
    ("Robotic navigation bronchoscopy was performed with ", "Ion platform", "PROC_METHOD", "Ion", "evt01"),
    ("robotic catheter was used to engage the ", "left lower lobe", "ANAT_LUNG_LOC", "Left Lower Lobe", "evt01"),
    ("", "Radial EBUS", "PROC_METHOD", "Radial EBUS", "evt01"),
    ("", "Cios Spin system", "DEV_INSTRUMENT", "Cios Spin", "evt01"),
    ("", "Cone Beam CT", "PROC_METHOD", "Cone Beam CT", "evt01"),

    # Event 2: TBNA LLL
    ("", "19G Needle", "DEV_NEEDLE", "19G", "evt02"),
    ("aspiration was performed with 19G Needle through the ", "extended working channel catheter", "DEV_CATHETER", "Extended Working Channel", "evt02"),
    ("Needle through the extended working channel catheter.  ", "Total 2 samples", "MEAS_COUNT", "2", "evt02"),

    # Event 3: TBBX LLL
    ("Transbronchial biopsy was performed with ", "precisor", "DEV_INSTRUMENT", "Precisor", "evt03"),
    ("Transbronchial biopsy was performed with precisor the ", "extended working channel catheter", "DEV_CATHETER", "Extended Working Channel", "evt03"),
    ("Total 2 samples were collected.  Samples sent for ", "Total 2 samples", "MEAS_COUNT", "2", "evt03"),

    # Event 4: Mechanical Debulking
    ("therapeutic bronchoscope was then used to clear the ", "airway", "ANAT_AIRWAY", "Airway", "evt04"),
    ("", "pulmonary, and large forceps", "DEV_INSTRUMENT", "Forceps", "evt04"),
    ("Granulation tissue and inflamed tissue was ", "excised", "PROC_ACTION", "Excision", "evt04"),
    ("", "Granulation tissue", "OBS_LESION", "Granulation Tissue", "evt04"),

    # Event 5: Balloon Dilation 1
    ("Balloon dilation was performed at ", "left lower lobe orifice", "ANAT_AIRWAY", "Left Lower Lobe Orifice", "evt05"),
    ("", "8/9/10 Elation balloon", "DEV_INSTRUMENT", "Elation Balloon", "evt05"),
    ("Elation balloon was used to perform dilation to ", "10 mm", "MEAS_AIRWAY_DIAM", "10 mm", "evt05"),
    ("Total 1 inflations with dilation time of ", "60 seconds", "MEAS_TIME", "60 s", "evt05"),

    # Event 6: Balloon Dilation 2
    ("Balloon dilation was performed at ", "left lower lobe orifice", "ANAT_AIRWAY", "Left Lower Lobe Orifice", "evt06"),
    ("", "10/11/12 Elation balloon", "DEV_INSTRUMENT", "Elation Balloon", "evt06"),
    ("Elation balloon was used to perform dilation to ", "12 mm", "MEAS_AIRWAY_DIAM", "12 mm", "evt06"),
    ("Total 3 inflations with dilation time of ", "60 seconds", "MEAS_TIME", "60 s", "evt06"),

    # Event 7: Thermal Ablation (Microwave)
    ("", "2.5 cm microwave catheter", "DEV_INSTRUMENT", "Microwave Catheter", "evt07"),
    ("used to ", "ablate", "PROC_ACTION", "Ablation", "evt07"),
    ("ablate the ", "inflamed tissue", "OBS_LESION", "Inflamed Tissue", "evt07"),
    ("inflamed tissue for ", "three minutes", "MEAS_TIME", "3 min", "evt07"),

    # Outcomes
    ("", "No immediate complications", "OUTCOME_COMPLICATION", "None", "evt_outcome"),
    ("The ", "left upper lobe and the left lower lobe were better visualized", "OUTCOME_AIRWAY_LUMEN_POST", "Better Visualized", "evt_outcome"),
]

# Event Log Data
EVENTS = [
    {
        "event_id": "evt01",
        "event_type": "navigational_bronchoscopy",
        "method": "Ion, Radial EBUS, Cone Beam CT",
        "anatomy_target": "Left Lower Lobe",
        "device": "Ion Catheter",
        "findings": "Target lesion 1cm, soft tissue confirmed",
        "comments": "Localization sequence"
    },
    {
        "event_id": "evt02",
        "event_type": "tbna_conventional",
        "method": "Transbronchial Needle Aspiration",
        "anatomy_target": "Left Lower Lobe",
        "device": "19G Needle",
        "counts": "2 samples",
        "specimens": "Microbiology, Cytology"
    },
    {
        "event_id": "evt03",
        "event_type": "transbronchial_biopsy",
        "method": "Biopsy",
        "anatomy_target": "Left Lower Lobe",
        "device": "Precisor",
        "counts": "2 samples",
        "specimens": "Microbiology, Pathology"
    },
    {
        "event_id": "evt04",
        "event_type": "tumor_debulking_non_thermal",
        "method": "Excision",
        "anatomy_target": "Left Mainstem / Left Lower Lobe",
        "device": "Forceps (Pulmonary, Large), Precisor",
        "findings": "Granulation and inflamed tissue",
        "comments": "Mechanical debulking"
    },
    {
        "event_id": "evt05",
        "event_type": "airway_dilation",
        "method": "Balloon Dilation",
        "anatomy_target": "Left Lower Lobe Orifice",
        "device": "8/9/10 Elation Balloon",
        "outcome_airway_lumen_post": "10 mm",
        "measurements": "60 sec duration"
    },
    {
        "event_id": "evt06",
        "event_type": "airway_dilation",
        "method": "Balloon Dilation",
        "anatomy_target": "Left Lower Lobe Orifice",
        "device": "10/11/12 Elation Balloon",
        "outcome_airway_lumen_post": "12 mm",
        "measurements": "60 sec duration x 3"
    },
    {
        "event_id": "evt07",
        "event_type": "thermal_ablation",
        "method": "Microwave Ablation",
        "anatomy_target": "Inflamed tissue (LMS/LLL)",
        "device": "2.5 cm microwave catheter",
        "measurements": "3 min at 3.5KJ, 1 min at 1.3KJ",
        "comments": "Ablation of inflamed tissue"
    },
    {
        "event_id": "evt_outcome",
        "event_type": "outcome",
        "outcome_complication": "None",
        "outcome_symptoms": "Better visualized",
        "comments": "Patient stable"
    }
]

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def count_spans(note_text, span_text):
    return note_text.count(span_text)

def hydrate_spans(raw_spans, note_text):
    hydrated = []
    
    for ctx, txt, lbl, norm, evt in raw_spans:
        count = count_spans(note_text, txt)
        start = -1
        end = -1
        status = "ambiguous"
        
        if count == 0:
            status = "not_found"
        elif count == 1:
            start = note_text.find(txt)
            end = start + len(txt)
            status = "hydrated_unique"
        else:
            # Context match
            if ctx:
                search_start = 0
                found = False
                for _ in range(count):
                    idx = note_text.find(txt, search_start)
                    if idx == -1: break
                    # Look back 120 chars for context
                    window = note_text[max(0, idx-120):idx]
                    if ctx in window:
                        start = idx
                        end = idx + len(txt)
                        status = "hydrated_context"
                        found = True
                        break
                    search_start = idx + 1
                if not found:
                    status = f"ambiguous_count={count}"
            else:
                status = f"ambiguous_count={count}"

        row = {
            "source_file": SOURCE_FILE,
            "note_id": NOTE_ID,
            "span_id": f"{NOTE_ID}_{len(hydrated)+1:03}",
            "section_type": "Procedure",
            "context_prefix": ctx,
            "span_text": txt,
            "match_index": 1 if count == 1 else None, # Simplified for now
            "start_char": start if start != -1 else None,
            "end_char": end if end != -1 else None,
            "span_len": len(txt),
            "label": lbl,
            "normalized_value": norm,
            "schema_field": "",
            "event_id": evt,
            "is_negated": False,
            "is_historical": False,
            "time_anchor": "",
            "reviewer": "Auto",
            "comments": "",
            "hydration_status": status
        }
        hydrated.append(row)
    return hydrated

def generate_v3_structure(events, spans):
    v3_rows = []
    for ev in events:
        if ev["event_type"] == "outcome": continue
        
        # Build JSON components
        target = {
            "anatomy_type": "Bronchial" if "Lower Lobe" in ev.get("anatomy_target", "") else "Unknown",
            "location": {"lobe": ev.get("anatomy_target", ""), "segment": ""},
            "station": ""
        }
        
        # Outcomes mapping
        outcome_airway = {}
        if "outcome_airway_lumen_post" in ev:
            outcome_airway["lumen_post"] = ev["outcome_airway_lumen_post"]
            
        row = {
            "note_id": NOTE_ID,
            "event_id": ev["event_id"],
            "type": ev["event_type"],
            "target.anatomy_type": target["anatomy_type"],
            "target.location.lobe": target["location"]["lobe"],
            "target.location.segment": "",
            "target.station": "",
            "lesion.type": "Stenosis/Tumor",
            "lesion.size_mm": "10" if "1 cm" in ev.get("findings","") else "",
            "method": ev.get("method", ""),
            "devices_json": json.dumps([ev.get("device", "")]),
            "measurements_json": json.dumps([ev.get("measurements", "")]),
            "specimens_json": json.dumps([ev.get("specimens", "")]),
            "findings_json": json.dumps([ev.get("findings", "")]),
            "evidence_quote": "",
            "stent.size": "",
            "stent.material_or_brand": "",
            "catheter.size_fr": "19G" if "19G" in ev.get("device", "") else "",
            "outcomes.airway.lumen_pre": "",
            "outcomes.airway.lumen_post": ev.get("outcome_airway_lumen_post", ""),
            "outcomes.symptoms": "",
            "outcomes.pleural": "",
            "outcomes.complications": ""
        }
        v3_rows.append(row)
    return v3_rows

# =============================================================================
# MAIN EXECUTION
# =============================================================================

def run():
    print(f"Loading template: {TEMPLATE_PATH}")
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        print("ERROR: Template file not found. Please ensure the template is in the directory.")
        return

    # 1. Note_Text
    ws_text = wb["Note_Text"]
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws_index = wb["Note_Index"]
    index_row = [
        SOURCE_FILE, NOTE_ID, "", "", "", "", "Pending", ""
    ]
    # Append flags in order of schema
    flag_keys = [
        "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
        "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy",
        "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration",
        "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
        "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation",
        "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
        "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis",
        "pleural_biopsy", "fibrinolytic_therapy"
    ]
    for k in flag_keys:
        index_row.append(PROCEDURE_FLAGS.get(k, 0))
    ws_index.append(index_row)

    # Hydrate Spans
    hydrated_spans = hydrate_spans(RAW_SPANS, NOTE_TEXT)

    # 3. Span_Annotations (Anchor First)
    ws_anno = wb["Span_Annotations"]
    for s in hydrated_spans:
        ws_anno.append([
            s["source_file"], s["note_id"], s["span_id"], s["section_type"],
            s["context_prefix"], s["span_text"], s["match_index"],
            "", "", f"=LEN(F{ws_anno.max_row+1})", # Formula for len
            s["label"], s["normalized_value"], s["schema_field"], s["event_id"],
            s["is_negated"], s["is_historical"], s["time_anchor"], s["reviewer"], s["comments"],
            "needs_hydration"
        ])

    # 4. Span_Hydrated
    if "Span_Hydrated" not in wb.sheetnames:
        ws_hyd = wb.create_sheet("Span_Hydrated")
        # Copy header from anno
        for col in range(1, ws_anno.max_column + 1):
            ws_hyd.cell(row=1, column=col, value=ws_anno.cell(row=1, column=col).value)
    else:
        ws_hyd = wb["Span_Hydrated"]

    for s in hydrated_spans:
        ws_hyd.append([
            s["source_file"], s["note_id"], s["span_id"], s["section_type"],
            s["context_prefix"], s["span_text"], s["match_index"],
            s["start_char"], s["end_char"], s["span_len"],
            s["label"], s["normalized_value"], s["schema_field"], s["event_id"],
            s["is_negated"], s["is_historical"], s["time_anchor"], s["reviewer"], s["comments"],
            s["hydration_status"]
        ])

    # 5. Event_Log
    ws_event = wb["Event_Log"]
    for e in EVENTS:
        ws_event.append([
            SOURCE_FILE, NOTE_ID, e.get("event_id"), e.get("event_type"), e.get("method"),
            e.get("anatomy_target"), e.get("device"), "", "", e.get("counts"), e.get("measurements"),
            e.get("specimens"), e.get("findings"), False, "Auto", e.get("comments"),
            "", "",
            "", e.get("outcome_airway_lumen_post"), # Pre/Post
            e.get("outcome_symptoms"), e.get("outcome_pleural"), e.get("outcome_complication")
        ])

    # 6. V3_Procedure_Events
    v3_data = generate_v3_structure(EVENTS, hydrated_spans)
    ws_v3 = wb["V3_Procedure_Events"]
    for r in v3_data:
        ws_v3.append([
            r["note_id"], r["event_id"], r["type"],
            r["target.anatomy_type"], r["target.location.lobe"], r["target.location.segment"], r["target.station"],
            r["lesion.type"], r["lesion.size_mm"],
            r["method"], r["devices_json"], r["measurements_json"], r["specimens_json"], r["findings_json"], r["evidence_quote"],
            r["stent.size"], r["stent.material_or_brand"], r["catheter.size_fr"],
            r["outcomes.airway.lumen_pre"], r["outcomes.airway.lumen_post"],
            r["outcomes.symptoms"], r["outcomes.pleural"], r["outcomes.complications"]
        ])

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    registry_obj = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": v3_data,
        "no_immediate_complications": True
    }
    ws_json.append([json.dumps(registry_obj, indent=2)])

    print(f"Saving output to: {OUTPUT_PATH}")
    wb.save(OUTPUT_PATH)
    print("Done.")

if __name__ == "__main__":
    run()