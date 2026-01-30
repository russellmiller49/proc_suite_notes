import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# -------------------------------------------------------------------------
# CONSTANTS & INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_059"
SOURCE_FILE = "note_059.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_059 SOURCE_FILE: note_059.txt INDICATION FOR OPERATION:  [REDACTED]is a 58 year old-year-old female who presents with lung nodule.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
PREOPERATIVE DIAGNOSIS: R91.1 Solitary Lung Nodule
 
POSTOPERATIVE DIAGNOSIS:  R91.1 Solitary Lung Nodule
 
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31623 Dx bronchoscope/brushing    
31624 Dx bronchoscope/lavage (BAL)    
31628 TBBX single lobe     
31629 TBNA single lobe   
31626 Fiducial marker placements, single or multiple     
31627 Navigational Bronchoscopy (computer assisted)
77012 Radiology / radiologic guidance for CT guided needle placement (CIOS)
76377 3D rendering with interpretation and reporting of CT, US, Tomo modality (ION Planning Station)
31653 EBUS sampling 3 or more nodes  
31654 Radial EBUS for peripheral lesion
76982 Ultrasound 
Elastography, First Target Lesion
76983 Ultrasound Elastography, Additional Targets 
76983 Ultrasound Elastography, Additional Target 2
 
 
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
Linear EBUS 
Radial EBUS
Ion Robotic Bronchoscope
Disposable Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
 
CT Chest scan was placed on separate planning station to generate 3D rendering of the pathway to target.
The navigational plan was reviewed and verified.  This was then loaded into robotic bronchoscopy platform.
Successful therapeutic aspiration was performed to clean out the Right Mainstem, Bronchus Intermedius , and Left Mainstem from mucus.
Ventilation Parameters:
Mode	RR	TV	PEEP	FiO2	Flow Rate	Pmean
vcv	14	375	15	100	10	20
 
Robotic navigation bronchoscopy was performed with Ion platform.  Partial registration was used.
Ion robotic catheter was used to engage the Posterior-basal Segment of RLL (RB10).
Target lesion is about 1.5 cm in diameter.   Under navigational guidance the ion robotic catheter was advanced to 1.0 cm away from the planned target.
Radial EBUS was performed to confirm that the location of the nodule is Concentric.
The following features were noted: Continuous margin .
 
Cone Beam CT was performed: 3-D reconstructions were performed on an independent workstation.
Cios Spin system was used for evaluation of nodule location.  Low dose spin was performed to acquire CT imaging.
This was passed on to Ion platform system for reconstruction and nodule location.
The 3D images was interpreted on an independent workstation (Ion).
Using the newly acquired nodule location, the Ion robotic system was adjusted to the new targeted location.
I personally interpreted the cone beam CT and 3-D reconstruction.
Transbronchial needle aspiration was performed with 21G Needle through the extended working channel catheter.  Total 6 samples were collected.
Samples sent for Cytology.
 
Transbronchial biopsy was performed with alligator forceps the extended working channel catheter.
Total 1 samples were collected.  Samples sent for Pathology.
 
Transbronchial cryobiopsy was performed with 1.1mm cryoprobe via the extended working channel catheter.
Freeze time of 6 seconds were used.  Total 6 samples were collected.  Samples sent for Pathology.
Transbronchial brushing was performed with Protected cytology brush the extended working channel catheter.  Total 1 samples were collected.
Samples sent for Microbiology (Cultures/Viral/Fungal).
 
Bronchial alveolar lavage was performed the extended working channel catheter.
Instilled 60 cc of NS, suction returned with 15 cc of NS.  Samples sent for Microbiology (Cultures/Viral/Fungal).
Fiducial marker (0.8mm x 3mm soft tissue gold CIVCO) was loaded with bone wax and placed under fluoroscopy guidance.
Prior to withdraw of the bronchoscope. 
 
ROSE from ION procedure was noted to be:
Cells present causing suspicion of malignant neoplasm
 
Prior to withdrawal of the bronchoscope, inspection demonstrated no evidence of bleeding.
EBUS-Findings
Indications: Diagnostic and Diagnostic and Staging
Technique:
All lymph node stations were assessed.
Only those 5 mm or greater in short axis were sampled.
Lymph node sizing was performed by EBUS and sampling by transbronchial needle aspiration was performed using 25-gauge Needle and 22-gauge Needle.
Lymph Nodes/Sites Inspected: 4R (lower paratracheal) node
4L (lower paratracheal) node
7 (subcarinal) node
10R lymph node
10L lymph node
11Rs lymph node
11Ri lymph node
11L lymph node
 
No immediate complications
 
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
Elastography provided a semi-quantitative classification (Type 1–3), which was used to guide biopsy site selection and sampling strategy.
Lymph Nodes Evaluated:
Site 1: The 11L lymph node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was sampled.. 4 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 2 elastographic pattern with mixed soft and stiff regions.
Given this heterogeneous and indeterminate appearance, TBNA was directed at representative areas to ensure comprehensive sampling and to minimize the risk of underdiagnosis.
Site 2: The 7 (subcarinal) node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was sampled.. 4 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 1 elastographic pattern, predominantly soft (green/yellow), suggesting a reactive or benign process.
Despite the benign appearance, TBNA was performed to confirm the absence of malignancy and to obtain cytology for diagnostic completeness.
Site 3: The 11Rs lymph node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was sampled.. 4 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 2 elastographic pattern with mixed soft and stiff regions.
Given this heterogeneous and indeterminate appearance, TBNA was directed at representative areas to ensure comprehensive sampling and to minimize the risk of underdiagnosis.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
RLL TBCBX, TBNA, Brush, BAL
TBNA 11R, 11L, 7
 
IMPRESSION/PLAN: [REDACTED]is a 58 year old-year-old female who presents for bronchoscopy for lung nodule.
- f/u in clinic
- f/u cxr"""

# -------------------------------------------------------------------------
# SPAN DATA (ANCHOR-FIRST)
# -------------------------------------------------------------------------
# Format: (span_text, label, normalized_value, schema_field, event_id, context_prefix_opt)
SPANS_DATA = [
    # --- Event 1: Therapeutic Aspiration ---
    ("Successful therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", "method", "evt_1", None),
    ("Right Mainstem", "ANAT_AIRWAY", "Right Mainstem Bronchus", "target.anatomy_type", "evt_1", None),
    ("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "target.anatomy_type", "evt_1", None),
    ("Left Mainstem", "ANAT_AIRWAY", "Left Mainstem Bronchus", "target.anatomy_type", "evt_1", None),
    ("mucus", "OBS_LESION", "Mucus", "findings", "evt_1", None),

    # --- Event 2: RLL Navigation/Imaging (Shared Context) ---
    ("Ion robotic catheter", "PROC_METHOD", "Robotic Bronchoscopy", "method", "evt_2", None),
    ("Posterior-basal Segment of RLL", "ANAT_LUNG_LOC", "RLL Posterior Basal", "target.location", "evt_2", None),
    ("RB10", "ANAT_LUNG_LOC", "RB10", "target.location.segment", "evt_2", None),
    ("Radial EBUS", "PROC_METHOD", "Radial EBUS", "method", "evt_2", None),
    ("Concentric", "OBS_LESION", "Concentric", "lesion.view", "evt_2", None),
    ("1.5 cm", "MEAS_SIZE", "1.5", "lesion.size_mm", "evt_2", "Target lesion is about "),
    ("Cios Spin system", "PROC_METHOD", "Cone Beam CT", "method", "evt_2", None),

    # --- Event 3: RLL TBNA ---
    ("Transbronchial needle aspiration", "PROC_METHOD", "TBNA", "method", "evt_3", "21G Needle through the extended"),
    ("21G Needle", "DEV_NEEDLE", "21G", "device", "evt_3", None),
    ("Total 6 samples", "MEAS_COUNT", "6", "specimens.count", "evt_3", "21G Needle through the extended"),

    # --- Event 4: RLL TBBX ---
    ("Transbronchial biopsy", "PROC_METHOD", "Transbronchial Biopsy", "method", "evt_4", "alligator forceps"),
    ("alligator forceps", "DEV_INSTRUMENT", "Alligator Forceps", "device", "evt_4", None),
    ("Total 1 samples", "MEAS_COUNT", "1", "specimens.count", "evt_4", "Transbronchial biopsy was performed with alligator forceps"),

    # --- Event 5: RLL Cryobiopsy ---
    ("Transbronchial cryobiopsy", "PROC_METHOD", "Transbronchial Cryobiopsy", "method", "evt_5", None),
    ("1.1mm cryoprobe", "DEV_INSTRUMENT", "1.1mm Cryoprobe", "device", "evt_5", None),
    ("Freeze time of 6 seconds", "PROC_ACTION", "Freeze time 6s", "procedure.settings", "evt_5", None),
    ("Total 6 samples", "MEAS_COUNT", "6", "specimens.count", "evt_5", "Freeze time of 6 seconds were used."),

    # --- Event 6: RLL Brush ---
    ("Transbronchial brushing", "PROC_METHOD", "Bronchial Brushing", "method", "evt_6", None),
    ("Protected cytology brush", "DEV_INSTRUMENT", "Protected Cytology Brush", "device", "evt_6", None),
    ("Total 1 samples", "MEAS_COUNT", "1", "specimens.count", "evt_6", "Transbronchial brushing was performed"),

    # --- Event 7: RLL BAL ---
    ("Bronchial alveolar lavage", "PROC_METHOD", "BAL", "method", "evt_7", None),
    ("Instilled 60 cc", "MEAS_VOL", "60", "procedure.instilled_cc", "evt_7", None),
    ("returned with 15 cc", "MEAS_VOL", "15", "procedure.return_cc", "evt_7", None),

    # --- Event 8: Fiducial ---
    ("Fiducial marker", "PROC_METHOD", "Fiducial Placement", "method", "evt_8", None),
    ("0.8mm x 3mm soft tissue gold CIVCO", "DEV_INSTRUMENT", "CIVCO Gold Fiducial 0.8x3mm", "device", "evt_8", None),

    # --- ROSE (Associated with RLL) ---
    ("Cells present causing suspicion of malignant neoplasm", "OBS_ROSE", "Suspicious for malignancy", "findings", "evt_3", None),

    # --- EBUS General ---
    ("Linear EBUS", "PROC_METHOD", "Linear EBUS", "method", "evt_9", "Flexible Therapeutic Bronchoscope"),
    ("All lymph node stations were assessed", "PROC_ACTION", "Systematic EBUS Assessment", "action", "evt_9", None),

    # --- Event 9: Station 11L ---
    ("11L lymph node", "ANAT_LN_STATION", "11L", "target.station", "evt_9", "Site 1: The "),
    ("Type 2 elastographic pattern", "OBS_LESION", "Type 2 (Mixed)", "findings.elastography", "evt_9", "Site 1: The "),
    ("4 endobronchial ultrasound guided transbronchial biopsies", "PROC_METHOD", "EBUS-TBNA", "method", "evt_9", "Site 1: The "),

    # --- Event 10: Station 7 ---
    ("7 (subcarinal) node", "ANAT_LN_STATION", "7", "target.station", "evt_10", "Site 2: The "),
    ("Type 1 elastographic pattern", "OBS_LESION", "Type 1 (Benign/Soft)", "findings.elastography", "evt_10", "Site 2: The "),
    ("4 endobronchial ultrasound guided transbronchial biopsies", "PROC_METHOD", "EBUS-TBNA", "method", "evt_10", "Site 2: The "),
    ("benign appearance", "OBS_LESION", "Benign Appearance", "findings", "evt_10", None),

    # --- Event 11: Station 11Rs ---
    ("11Rs lymph node", "ANAT_LN_STATION", "11Rs", "target.station", "evt_11", "Site 3: The "),
    ("Type 2 elastographic pattern", "OBS_LESION", "Type 2 (Mixed)", "findings.elastography", "evt_11", "Site 3: The "),
    ("4 endobronchial ultrasound guided transbronchial biopsies", "PROC_METHOD", "EBUS-TBNA", "method", "evt_11", "Site 3: The "),

    # --- Outcomes ---
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "outcomes.complications", "evt_0", "well.  There were "),
]

# -------------------------------------------------------------------------
# FLAG CONFIGURATION
# -------------------------------------------------------------------------
# 1 = Present, 0 = Absent
PROCEDURE_FLAGS = {
    # Bronchoscopy (23)
    "diagnostic_bronchoscopy": 1,
    "bal": 1,
    "bronchial_wash": 0,
    "brushings": 1,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 1,
    "linear_ebus": 1,
    "radial_ebus": 1,
    "navigational_bronchoscopy": 1,
    "transbronchial_biopsy": 1,
    "transbronchial_cryobiopsy": 1,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural (7)
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# -------------------------------------------------------------------------
# HELPER FUNCTIONS
# -------------------------------------------------------------------------
def hydrate_spans(text, spans):
    """
    Computes start_char, end_char based on text count and context.
    Returns list of enriched span dicts.
    """
    hydrated = []
    
    for item in spans:
        txt, label, norm, schema, evt, ctx = item
        
        row = {
            "source_file": SOURCE_FILE,
            "note_id": NOTE_ID,
            "span_id": f"span_{len(hydrated)+1:03d}",
            "section_type": "Procedure",
            "context_prefix": ctx if ctx else "",
            "span_text": txt,
            "match_index": "", # logic below
            "start_char": "",
            "end_char": "",
            "span_len": len(txt),
            "label": label,
            "normalized_value": norm,
            "schema_field": schema,
            "event_id": evt,
            "is_negated": "FALSE",
            "is_historical": "FALSE",
            "time_anchor": "",
            "reviewer": "Auto",
            "comments": "",
            "hydration_status": "ambiguous"
        }

        # Logic to find offsets
        matches = [m.start() for m in re.finditer(re.escape(txt), text)]
        
        if len(matches) == 1:
            row['start_char'] = matches[0]
            row['end_char'] = matches[0] + len(txt)
            row['hydration_status'] = "hydrated_unique"
            row['match_index'] = 1
        elif len(matches) > 1:
            # Context match
            found = False
            if ctx:
                # Look for context in window before each match
                for idx, m_start in enumerate(matches):
                    window_start = max(0, m_start - 120)
                    window_text = text[window_start:m_start]
                    if ctx in window_text:
                        row['start_char'] = m_start
                        row['end_char'] = m_start + len(txt)
                        row['hydration_status'] = "hydrated_prefix_window"
                        row['match_index'] = idx + 1
                        found = True
                        break
            
            if not found:
                # Default to first if not specific
                row['start_char'] = matches[0]
                row['end_char'] = matches[0] + len(txt)
                row['hydration_status'] = "hydrated_fallback_first"
                row['match_index'] = 1
        else:
            row['hydration_status'] = "missing_text"

        hydrated.append(row)
    
    return hydrated

def generate_v3_data(hydrated_rows):
    """
    Aggregates hydrated rows into V3 Event structures and JSON.
    """
    events = {}
    
    for row in hydrated_rows:
        eid = row['event_id']
        if eid == "evt_0": continue # General outcomes
        
        if eid not in events:
            events[eid] = {
                "note_id": NOTE_ID,
                "event_id": eid,
                "type": "Procedure",
                "target": {"anatomy": [], "location": []},
                "lesion": {},
                "method": [],
                "devices": [],
                "measurements": [],
                "specimens": [],
                "findings": [],
                "outcomes": {}
            }
        
        evt = events[eid]
        lbl = row['label']
        val = row['normalized_value']
        
        if "ANAT" in lbl:
            if "STATION" in lbl: evt["target"]["station"] = val
            elif "LUNG" in lbl: evt["target"]["location"].append(val)
            else: evt["target"]["anatomy"].append(val)
        elif "PROC_METHOD" in lbl:
            evt["method"].append(val)
        elif "DEV" in lbl:
            evt["devices"].append(val)
        elif "MEAS" in lbl:
            evt["measurements"].append(f"{row['schema_field']}: {val}")
        elif "OBS" in lbl:
            evt["findings"].append(val)
        
    # JSON Construction
    json_out = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "procedures": list(events.values()),
        "no_immediate_complications": True
    }
    
    return events, json_out

# -------------------------------------------------------------------------
# EXCEL GENERATION
# -------------------------------------------------------------------------
def create_workbook():
    if not os.path.exists(TEMPLATE_PATH):
        # Create a dummy template if missing (for standalone testing)
        wb = openpyxl.Workbook()
        wb.create_sheet("Note_Text")
        wb.create_sheet("Note_Index")
        wb.create_sheet("Span_Annotations")
        wb.create_sheet("Span_Hydrated")
        wb.create_sheet("Event_Log")
        wb.create_sheet("V3_Procedure_Events")
        wb.create_sheet("V3_Registry_JSON")
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # 1. Note_Text
    ws = wb["Note_Text"]
    if ws.max_row == 1:
        ws.append(["note_id", "source_file", "note_text"])
    ws.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws = wb["Note_Index"]
    header = [
        "source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes",
        "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", "tbna_conventional",
        "linear_ebus", "radial_ebus", "navigational_bronchoscopy", "transbronchial_biopsy", "transbronchial_cryobiopsy",
        "therapeutic_aspiration", "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
        "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", "bronchial_thermoplasty",
        "whole_lung_lavage", "rigid_bronchoscopy",
        "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis", "pleural_biopsy", "fibrinolytic_therapy"
    ]
    
    # Ensure header
    if ws.max_row == 1:
        # Clear and write correct header if needed, or just append
        pass 

    row_data = [SOURCE_FILE, NOTE_ID, "", "", "", "", "Ready", ""]
    # Append flags in order
    flag_keys = header[8:]
    for key in flag_keys:
        row_data.append(PROCEDURE_FLAGS.get(key, 0))
    ws.append(row_data)

    # 3. Span_Annotations & 4. Span_Hydrated
    hydrated_data = hydrate_spans(NOTE_TEXT, SPANS_DATA)
    
    ws_anno = wb["Span_Annotations"]
    ws_hydra = wb["Span_Hydrated"]
    
    # Headers logic omitted for brevity, assuming template has them.
    # We append data mapped to columns.
    
    columns = [
        "source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", 
        "match_index", "start_char", "end_char", "span_len", "label", "normalized_value", 
        "schema_field", "event_id", "is_negated", "is_historical", "time_anchor", "reviewer", 
        "comments", "hydration_status"
    ]

    for row in hydrated_data:
        # Annotation Sheet (No offsets)
        anno_row = [row[k] for k in columns]
        anno_row[7] = "" # start_char
        anno_row[8] = "" # end_char
        anno_row[19] = "needs_hydration"
        ws_anno.append(anno_row)

        # Hydrated Sheet (With offsets)
        hydra_row = [row[k] for k in columns]
        ws_hydra.append(hydra_row)

    # 5. Event_Log
    ws_log = wb["Event_Log"]
    events_map, json_obj = generate_v3_data(hydrated_data)
    
    for eid, data in events_map.items():
        log_row = [
            SOURCE_FILE, NOTE_ID, eid, data["type"], 
            ", ".join(data["method"]), 
            str(data["target"]), 
            ", ".join(data["devices"]), 
            "", # needle gauge logic omitted for brevity
            data["target"].get("station", ""),
            "", "", # counts/measurements
            str(data["specimens"]), 
            str(data["findings"]),
            "FALSE", "Auto", "", "", "", "", "", "", "", ""
        ]
        ws_log.append(log_row)

    # 6. V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    for eid, data in events_map.items():
        v3_row = [
            NOTE_ID, eid, data["type"],
            ", ".join(data["target"].get("anatomy", [])),
            ", ".join(data["target"].get("location", [])),
            "", # segment
            data["target"].get("station", ""),
            "", # lesion type
            "", # size
            ", ".join(data["method"]),
            json.dumps(data["devices"]),
            json.dumps(data["measurements"]),
            json.dumps(data["specimens"]),
            json.dumps(data["findings"]),
            "", "", "", "", "", "", "", "", ""
        ]
        ws_v3.append(v3_row)

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    ws_json.append(["v3.0", NOTE_ID, json.dumps(json_obj, indent=2), "TRUE"])

    wb.save(OUTPUT_PATH)
    print(f"Successfully generated {OUTPUT_PATH}")

if __name__ == "__main__":
    create_workbook()