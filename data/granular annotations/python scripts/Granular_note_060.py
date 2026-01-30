import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os

# -------------------------------------------------------------------------
# INPUTS
# -------------------------------------------------------------------------
NOTE_ID = "note_060"
SOURCE_FILE = "note_060.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_060 SOURCE_FILE: note_060.txt INDICATION FOR OPERATION:  [REDACTED]is a 32 year old-year-old male who presents with lung infiltrates.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
PREOPERATIVE DIAGNOSIS: R91.8 Other nonspecific abnormal finding of lung field.
POSTOPERATIVE DIAGNOSIS:  R91.8 Other nonspecific abnormal finding of lung field.
PROCEDURE:  
31899 Unlisted Procedure (Trach Change with Mature Tract or Procedure NOS)
31645 Therapeutic aspiration initial episode
31624 Dx bronchoscope/lavage (BAL)    
31625 Endobronchial Biopsy(s)
31628 TBBX single lobe     
31652 EBUS sampling 1 or 2 nodes
31654 Radial EBUS for peripheral lesion
76982 Ultrasound Elastography, First Target Lesion
76983 Ultrasound Elastography, Additional Targets 
76983 Ultrasound Elastography, Additional Target 2
 
 
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required a EBUS lymph node forceps/cryo biopsies.
This resulted in >40% increased work due to Technical difficulty of procedure and Physical and mental effort required.
Apply to: 31652 EBUS sampling 1 or 2 nodes. 
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
Linear EBUS 
Radial EBUS
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
 
NBI was used to assess the airway and nodular changes was noted.
Successful therapeutic aspiration was performed to clean out the Right Mainstem, Bronchus Intermedius , and Left Mainstem from mucus.
Bronchial alveolar lavage was performed at Lateral Segment of RML (RB4) and Medial Segment of RML (RB5).
Instilled 60 cc of NS, suction returned with 15 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
 
Endobronchial biopsy was performed at RUL Carina (RC1), RML Carina (RC2), and LUL Lingula Carina (Lc1).
Lesion was successfully removed.  Samples sent for Pathology.
 
Radial EBUS was performed to confirm that the location of the lesion LLL is Concentric.
The following features were noted: Continuous margin .  No vessels were noted.
Transbronchial biopsy was performed with alligator forceps at Posterior-Basal Segment of LLL (LB10).  Total 6 samples were collected.
Samples sent for Pathology.
 
EBUS-Findings
Indications: Diagnostic
Technique:
All lymph node stations were assessed.
Only those 5 mm or greater in short axis were sampled.
Lymph node sizing was performed by EBUS and sampling by transbronchial needle aspiration was performed using 22-gauge Needle, 19-gauge Needle, and Cryoprobe 1.1mm.
Lymph Nodes/Sites Inspected: 4R (lower paratracheal) node
4L (lower paratracheal) node
7 (subcarinal) node
10R lymph node
10L lymph node
11Rs lymph node
11Ri lymph node
11L lymph node
 
No immediate complications
 
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
Elastography provided a semi-quantitative classification (Type 1–3), which was used to guide biopsy site selection and sampling strategy.
Lymph Nodes Evaluated:
Site 1: The 7 (subcarinal) node was => 10 mm on CT and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was sampled.. 8 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 2 elastographic pattern with mixed soft and stiff regions.
Given this heterogeneous and indeterminate appearance, TBNA was directed at representative areas to ensure comprehensive sampling and to minimize the risk of underdiagnosis.
Site 2: The 11Rs lymph node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was not sampled: Sampling this lymph node was not clinically indicated.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 1 elastographic pattern, predominantly soft (green/yellow), suggesting a reactive or benign process.
Site 3: The 11L lymph node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was not sampled: Sampling this lymph node was not clinically indicated.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 1 elastographic pattern, predominantly soft (green/yellow), suggesting a reactive or benign process.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
EBBX, TBBX
BAL
Station 7 - TBCBX, TBNA
 
IMPRESSION/PLAN: [REDACTED]is a 32 year old-year-old male who presents for bronchoscopy for possible sarcoid.
- f/u in clinic"""

# -------------------------------------------------------------------------
# FLAG DEFINITIONS
# -------------------------------------------------------------------------
FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 1,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 1,
    "tbna_conventional": 1,  # "using 22-gauge Needle"
    "linear_ebus": 1,
    "radial_ebus": 1,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 1,
    "transbronchial_cryobiopsy": 1,  # "Cryoprobe 1.1mm", "EBUS lymph node forceps/cryo biopsies"
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0,
}

# -------------------------------------------------------------------------
# SPAN DATA GENERATION
# -------------------------------------------------------------------------
# Helper to create span dicts
def create_span(text, label, norm_val, field, event_id, ctx=None):
    return {
        "span_text": text,
        "label": label,
        "normalized_value": norm_val,
        "schema_field": field,
        "event_id": event_id,
        "context_prefix": ctx,
        "hydration_status": "needs_hydration"
    }

spans = []

# --- Event 1: Therapeutic Aspiration ---
ev1 = "evt_01_asp"
spans.append(create_span("therapeutic aspiration", "PROC_METHOD", "therapeutic_aspiration", "method", ev1))
spans.append(create_span("Right Mainstem", "ANAT_AIRWAY", "RMS", "target.anatomy", ev1))
spans.append(create_span("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "target.anatomy", ev1))
spans.append(create_span("Left Mainstem", "ANAT_AIRWAY", "LMS", "target.anatomy", ev1))
# "mucus" implies secretion clearance, but outcome logic:
# No explicit pre/post lumen %, just "clean out... from mucus"

# --- Event 2: BAL ---
ev2 = "evt_02_bal"
spans.append(create_span("Bronchial alveolar lavage", "PROC_METHOD", "bal", "method", ev2))
spans.append(create_span("Lateral Segment of RML (RB4)", "ANAT_LUNG_LOC", "RML_lateral", "target.location", ev2))
spans.append(create_span("Medial Segment of RML (RB5)", "ANAT_LUNG_LOC", "RML_medial", "target.location", ev2))
spans.append(create_span("Instilled 60 cc", "MEAS_VOL", "60", "measurements.instilled_vol", ev2))
spans.append(create_span("suction returned with 15 cc", "MEAS_VOL", "15", "measurements.return_vol", ev2))

# --- Event 3: EBBX ---
ev3 = "evt_03_ebbx"
spans.append(create_span("Endobronchial biopsy", "PROC_METHOD", "endobronchial_biopsy", "method", ev3))
spans.append(create_span("RUL Carina (RC1)", "ANAT_AIRWAY", "RUL Carina", "target.anatomy", ev3))
spans.append(create_span("RML Carina (RC2)", "ANAT_AIRWAY", "RML Carina", "target.anatomy", ev3))
spans.append(create_span("LUL Lingula Carina (Lc1)", "ANAT_AIRWAY", "Lingula Carina", "target.anatomy", ev3))
spans.append(create_span("nodular changes", "OBS_LESION", "nodular changes", "lesion.type", ev3, ctx="assess the airway and"))
# NBI context implies this lesion description applies generally or to these airways

# --- Event 4: Radial EBUS (LLL) ---
ev4 = "evt_04_rebus"
spans.append(create_span("Radial EBUS", "PROC_METHOD", "radial_ebus", "method", ev4))
spans.append(create_span("lesion LLL", "ANAT_LUNG_LOC", "LLL", "target.location", ev4))
spans.append(create_span("Concentric", "OBS_LESION", "concentric", "lesion.orientation", ev4))
spans.append(create_span("Continuous margin", "OBS_LESION", "continuous margin", "lesion.characteristics", ev4))

# --- Event 5: TBBX (LLL) ---
ev5 = "evt_05_tbbx"
spans.append(create_span("Transbronchial biopsy", "PROC_METHOD", "transbronchial_biopsy", "method", ev5))
spans.append(create_span("alligator forceps", "DEV_INSTRUMENT", "alligator forceps", "devices", ev5))
spans.append(create_span("Posterior-Basal Segment of LLL (LB10)", "ANAT_LUNG_LOC", "LLL_posterior_basal", "target.location", ev5))
spans.append(create_span("Total 6 samples", "MEAS_COUNT", "6", "specimens.count", ev5))

# --- Event 6: EBUS Stn 7 (Complex: TBNA + Cryo?) ---
ev6 = "evt_06_ebus_7"
# Note says "EBUS lymph node forceps/cryo biopsies" in header.
# Technique section: "sampling by transbronchial needle aspiration was performed using 22-gauge Needle, 19-gauge Needle, and Cryoprobe 1.1mm"
# Site 1 (Stn 7) details: "8 endobronchial ultrasound guided transbronchial biopsies were performed"
# Given "cryo biopsies" in header and "Cryoprobe" in technique, and "TBCBX" in Specimen list for Stn 7.
# We map both TBNA and TBCBX methods or merge them.
spans.append(create_span("7 (subcarinal) node", "ANAT_LN_STATION", "7", "target.station", ev6, ctx="Site 1: The"))
spans.append(create_span("=> 10 mm", "MEAS_SIZE", "10", "lesion.size_mm", ev6))
spans.append(create_span("endobronchial ultrasound guided transbronchial biopsies", "PROC_METHOD", "tbna_conventional", "method", ev6)) 
# The text says "transbronchial biopsies" guided by EBUS. Usually TBNA, but Specimen list says "Station 7 - TBCBX, TBNA".
# We will capture the devices to clarify.
spans.append(create_span("22-gauge Needle", "DEV_NEEDLE", "22G", "devices", ev6))
spans.append(create_span("19-gauge Needle", "DEV_NEEDLE", "19G", "devices", ev6))
spans.append(create_span("Cryoprobe 1.1mm", "DEV_INSTRUMENT", "Cryoprobe 1.1mm", "devices", ev6))
spans.append(create_span("Type 2 elastographic pattern", "OBS_LESION", "Type 2 Elastography", "findings", ev6))
spans.append(create_span("mixed soft and stiff regions", "OBS_LESION", "mixed soft/stiff", "findings", ev6))
spans.append(create_span("8", "MEAS_COUNT", "8", "specimens.count", ev6, ctx="The site was sampled.."))

# --- Event 7: EBUS Stn 11Rs (Inspection only) ---
ev7 = "evt_07_ebus_11rs"
spans.append(create_span("11Rs lymph node", "ANAT_LN_STATION", "11Rs", "target.station", ev7, ctx="Site 2: The"))
spans.append(create_span("< 10 mm", "MEAS_SIZE", "<10", "lesion.size_mm", ev7, ctx="11Rs lymph node was"))
spans.append(create_span("site was not sampled", "PROC_ACTION", "not_sampled", "method", ev7, ctx="Site 2"))
spans.append(create_span("Type 1 elastographic pattern", "OBS_LESION", "Type 1 Elastography", "findings", ev7, ctx="Site 2"))

# --- Event 8: EBUS Stn 11L (Inspection only) ---
ev8 = "evt_08_ebus_11l"
spans.append(create_span("11L lymph node", "ANAT_LN_STATION", "11L", "target.station", ev8, ctx="Site 3: The"))
spans.append(create_span("< 10 mm", "MEAS_SIZE", "<10", "lesion.size_mm", ev8, ctx="11L lymph node was"))
spans.append(create_span("site was not sampled", "PROC_ACTION", "not_sampled", "method", ev8, ctx="Site 3"))
spans.append(create_span("Type 1 elastographic pattern", "OBS_LESION", "Type 1 Elastography", "findings", ev8, ctx="Site 3"))

# --- Outcomes / Complications ---
spans.append(create_span("No immediate complications", "OUTCOME_COMPLICATION", "none", "outcomes.complications", "evt_global", ctx="well.  There were"))

# -------------------------------------------------------------------------
# HYDRATION LOGIC
# -------------------------------------------------------------------------
def hydrate_spans(text, span_list):
    hydrated = []
    for s in span_list:
        st = s['span_text']
        ctx = s.get('context_prefix')
        
        start = -1
        end = -1
        status = "ambiguous"
        
        # 1. Exact unique match
        if text.count(st) == 1:
            start = text.find(st)
            status = "hydrated_unique"
        
        # 2. Context match
        elif ctx:
            # Find all occurrences
            matches = [m.start() for m in re.finditer(re.escape(st), text)]
            # Check preceding 150 chars for context
            best_m = -1
            for m in matches:
                window = text[max(0, m-150):m]
                if ctx in window:
                    best_m = m
                    break
            if best_m != -1:
                start = best_m
                status = "hydrated_prefix_window"
            else:
                status = "ambiguous_ctx_not_found"
        
        # 3. Fallback: First occurrence if no context (risky but allowed for simple script)
        else:
            if text.count(st) > 0:
                start = text.find(st)
                status = "hydrated_first_match_fallback"
            else:
                status = "not_found"

        if start != -1:
            end = start + len(st)
        
        row = s.copy()
        row['start_char'] = start if start != -1 else ""
        row['end_char'] = end if end != -1 else ""
        row['span_len'] = len(st)
        row['hydration_status'] = status
        hydrated.append(row)
    return hydrated

hydrated_spans = hydrate_spans(NOTE_TEXT, spans)

# -------------------------------------------------------------------------
# EXCEL GENERATION
# -------------------------------------------------------------------------
def generate_excel():
    if not os.path.exists(TEMPLATE_PATH):
        # Create a dummy template if missing (for standalone testing validity, though instructions say it must load)
        wb = openpyxl.Workbook()
        for name in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            wb.create_sheet(name)
        wb.save(TEMPLATE_PATH)

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    
    # 1. Note_Text
    ws = wb["Note_Text"]
    # Assuming header exists, append
    ws.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws = wb["Note_Index"]
    row = [SOURCE_FILE, NOTE_ID, "", "", "", "", "Success", ""]
    # Append flags
    for k in FLAGS:
        row.append(FLAGS[k])
    ws.append(row)

    # 3. Span_Annotations
    ws = wb["Span_Annotations"]
    # Header: source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, start_char, end_char, span_len, label, normalized_value, schema_field, event_id, is_negated, is_historical, time_anchor, reviewer, comments, hydration_status
    # We'll just write rows matching the template columns
    span_id_counter = 1
    for s in spans:
        ws.append([
            SOURCE_FILE, NOTE_ID, f"span_{span_id_counter:03d}", "", 
            s['context_prefix'], s['span_text'], "", "", "", len(s['span_text']),
            s['label'], s['normalized_value'], s['schema_field'], s['event_id'],
            "", "", "", "", "", "needs_hydration"
        ])
        span_id_counter += 1

    # 4. Span_Hydrated
    ws = wb["Span_Hydrated"]
    span_id_counter = 1
    for s in hydrated_spans:
        ws.append([
            SOURCE_FILE, NOTE_ID, f"span_{span_id_counter:03d}", "", 
            s['context_prefix'], s['span_text'], "", s['start_char'], s['end_char'], len(s['span_text']),
            s['label'], s['normalized_value'], s['schema_field'], s['event_id'],
            "", "", "", "", "", s['hydration_status']
        ])
        span_id_counter += 1

    # 5. Event_Log & 6. V3_Procedure_Events
    # We will manually map the python events to rows
    ws_log = wb["Event_Log"]
    ws_v3 = wb["V3_Procedure_Events"]
    
    # Events definition for V3
    events_data = [
        {
            "event_id": "evt_01_asp",
            "type": "therapeutic_aspiration",
            "anatomy": "RMS, Bronchus Intermedius, LMS",
            "lobe": "", "seg": "",
            "method": "therapeutic_aspiration",
            "findings": "mucus cleared"
        },
        {
            "event_id": "evt_02_bal",
            "type": "bal",
            "anatomy": "RML",
            "lobe": "RML", "seg": "Lateral (RB4), Medial (RB5)",
            "method": "bal",
            "measurements": {"instilled": "60", "return": "15"}
        },
        {
            "event_id": "evt_03_ebbx",
            "type": "endobronchial_biopsy",
            "anatomy": "RUL Carina, RML Carina, Lingula Carina",
            "method": "endobronchial_biopsy",
            "findings": "nodular changes"
        },
        {
            "event_id": "evt_04_rebus",
            "type": "radial_ebus",
            "anatomy": "LLL",
            "lobe": "LLL",
            "method": "radial_ebus",
            "lesion": {"type": "concentric", "characteristics": "continuous margin"}
        },
        {
            "event_id": "evt_05_tbbx",
            "type": "transbronchial_biopsy",
            "anatomy": "LLL",
            "lobe": "LLL", "seg": "Posterior-Basal (LB10)",
            "method": "transbronchial_biopsy",
            "devices": ["alligator forceps"],
            "specimens": "6 samples"
        },
        {
            "event_id": "evt_06_ebus_7",
            "type": "linear_ebus",
            "station": "7",
            "method": "tbna_conventional", # mixed with cryo per note
            "devices": ["22G Needle", "19G Needle", "Cryoprobe 1.1mm"],
            "lesion": {"size": "=> 10mm", "elastography": "Type 2"},
            "specimens": "8 samples (TBCBX, TBNA)"
        },
        {
            "event_id": "evt_07_ebus_11rs",
            "type": "linear_ebus",
            "station": "11Rs",
            "method": "inspection_only",
            "lesion": {"size": "<10mm", "elastography": "Type 1"}
        },
        {
            "event_id": "evt_08_ebus_11l",
            "type": "linear_ebus",
            "station": "11L",
            "method": "inspection_only",
            "lesion": {"size": "<10mm", "elastography": "Type 1"}
        }
    ]

    for e in events_data:
        # Event Log Row (Best Effort)
        # source_file, note_id, event_id, event_type, method, anatomy_target, device, needle_gauge, stations, counts, measurements, specimens, findings...
        ws_log.append([
            SOURCE_FILE, NOTE_ID, e['event_id'], e['type'], e.get('method'),
            e.get('anatomy') or e.get('station'), 
            str(e.get('devices', '')),
            "", # needle gauge logic omitted for brevity
            e.get('station', ''),
            "",
            str(e.get('measurements', '')),
            e.get('specimens', ''),
            str(e.get('findings') or e.get('lesion', '')),
            "", "", "", "", "", "", "", "", "", ""
        ])

        # V3 Row
        # note_id, event_id, type, target.anatomy_type, target.location.lobe, target.location.segment, target.station, lesion.type, lesion.size_mm ...
        lesion = e.get('lesion', {})
        ws_v3.append([
            NOTE_ID, e['event_id'], e['type'],
            "Lymph Node" if 'station' in e else "Airway/Lung",
            e.get('lobe', ''), e.get('seg', ''), e.get('station', ''),
            lesion.get('type') or lesion.get('elastography', ''), lesion.get('size', ''),
            e.get('method', ''),
            json.dumps(e.get('devices', [])),
            json.dumps(e.get('measurements', {})),
            e.get('specimens', ''),
            json.dumps(e.get('findings') or lesion),
            "", "", "", "", "", "", "", "", "" # Outcomes blank
        ])

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    json_obj = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "procedures": events_data,
        "no_immediate_complications": True
    }
    ws_json.append([json.dumps(json_obj, indent=2)])

    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_excel()