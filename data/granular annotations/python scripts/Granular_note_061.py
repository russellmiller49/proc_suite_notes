import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os
import datetime

# =============================================================================
# INPUT DATA & CONFIGURATION
# =============================================================================

NOTE_ID = "note_061"
SOURCE_FILE = "note_061.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# Full text from the provided source, cleaned of citation tags for processing
NOTE_TEXT = """NOTE_ID:  note_061 SOURCE_FILE: note_061.txt INDICATION FOR OPERATION:  [REDACTED]is a 68 year old-year-old male who presents with stent evaluation.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
 
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
 
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
ANESTHESIA: 
General Anesthesia
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
 
Successful therapeutic aspiration was performed to clean out the Right Mainstem, Bronchus Intermedius , and Left Mainstem from mucus and mucus plug.
Endobronchial obstruction at Left Mainstem was treated with the following modalities:
 
Modality	Tools	Setting/Mode	Duration	Results
Electrocautery	 	 	 	 
APC	1.5mm	Straight.
Pulse, 2	10sec	ablated
Laser	 	 	 	 
CoreCath	 	 	 	 
Cryoprobe	 	 	 	 
 
Tissue was ablated.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
 
 
IMPRESSION/PLAN: [REDACTED]is a 68 year old-year-old male who presents for bronchoscopy for stent evaluation.
- repeat bronch in 6-8 weeks."""

# =============================================================================
# LABELING DATA (ANCHOR-FIRST)
# =============================================================================

# Procedure Flags (30 standard flags)
# Based on note: Therapeutic Aspiration (Yes), Thermal Ablation (APC - Yes).
# Stent Evaluation is listed, but no stent placement/removal codes/actions described, just inspection/cleaning.
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 0,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,  # "31645 Therapeutic aspiration"
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 1,        # "APC... ablated"
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,             # Listed in code title, but APC used in table.
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# Spans to extract
# Fields: span_text, label, normalized_value, schema_field, event_id, context_prefix (optional)
SPANS = [
    # Header/Context
    {
        "span_text": "stent evaluation",
        "label": "CTX_HISTORICAL",
        "normalized_value": "stent evaluation",
        "schema_field": "indication",
        "event_id": "evt0",
        "comments": "Indication"
    },
    {
        "span_text": "Flexible Therapeutic Bronchoscope",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "Flexible Therapeutic Bronchoscope",
        "schema_field": "instrument",
        "event_id": "evt0"
    },
    
    # Event 1: Therapeutic Aspiration
    {
        "span_text": "Successful therapeutic aspiration",
        "label": "PROC_METHOD",
        "normalized_value": "Therapeutic Aspiration",
        "schema_field": "method",
        "event_id": "evt1"
    },
    {
        "span_text": "Right Mainstem",
        "label": "ANAT_AIRWAY",
        "normalized_value": "RMS",
        "schema_field": "target.anatomy",
        "event_id": "evt1"
    },
    {
        "span_text": "Bronchus Intermedius",
        "label": "ANAT_AIRWAY",
        "normalized_value": "BI",
        "schema_field": "target.anatomy",
        "event_id": "evt1"
    },
    {
        "span_text": "Left Mainstem",
        "label": "ANAT_AIRWAY",
        "normalized_value": "LMS",
        "schema_field": "target.anatomy",
        "event_id": "evt1",
        "context_prefix": "clean out the Right Mainstem, Bronchus Intermedius , and "
    },
    {
        "span_text": "mucus and mucus plug",
        "label": "OBS_LESION",
        "normalized_value": "mucus plug",
        "schema_field": "finding",
        "event_id": "evt1"
    },

    # Event 2: Thermal Ablation (APC)
    {
        "span_text": "Endobronchial obstruction",
        "label": "OBS_LESION",
        "normalized_value": "obstruction",
        "schema_field": "lesion.type",
        "event_id": "evt2"
    },
    {
        "span_text": "Left Mainstem",
        "label": "ANAT_AIRWAY",
        "normalized_value": "LMS",
        "schema_field": "target.anatomy",
        "event_id": "evt2",
        "context_prefix": "Endobronchial obstruction at "
    },
    {
        "span_text": "APC",
        "label": "PROC_METHOD",
        "normalized_value": "APC",
        "schema_field": "method",
        "event_id": "evt2"
    },
    {
        "span_text": "1.5mm",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "1.5mm Probe",
        "schema_field": "device.size",
        "event_id": "evt2"
    },
    {
        "span_text": "ablated",
        "label": "PROC_ACTION",
        "normalized_value": "ablated",
        "schema_field": "action",
        "event_id": "evt2",
        "context_prefix": "Pulse, 2\t10sec\t"
    },

    # Event 3: Outcomes
    {
        "span_text": "No immediate complications",
        "label": "OUTCOME_COMPLICATION",
        "normalized_value": "None",
        "schema_field": "outcomes.complications",
        "event_id": "evt3"
    }
]

# Event Definitions for V3 Logging
EVENTS_METADATA = [
    {
        "event_id": "evt1",
        "type": "Therapeutic Aspiration",
        "method": "Therapeutic Aspiration",
        "target_anatomy": "RMS, BI, LMS",
        "findings": "Mucus/Plug"
    },
    {
        "event_id": "evt2",
        "type": "Thermal Ablation",
        "method": "APC",
        "target_anatomy": "LMS",
        "findings": "Endobronchial obstruction",
        "device": "1.5mm APC Probe"
    },
    {
        "event_id": "evt3",
        "type": "Outcome",
        "outcome_complication": "None"
    }
]

# =============================================================================
# PROCESSING LOGIC
# =============================================================================

def create_workbook_if_needed():
    if not os.path.exists(TEMPLATE_PATH):
        # Create a dummy template if not found (strictly for script validity)
        wb = openpyxl.Workbook()
        wb.create_sheet("Note_Text")
        wb.create_sheet("Note_Index")
        wb.create_sheet("Span_Annotations")
        wb.create_sheet("Span_Hydrated")
        wb.create_sheet("Event_Log")
        wb.create_sheet("V3_Procedure_Events")
        wb.create_sheet("V3_Registry_JSON")
        wb.save(TEMPLATE_PATH)
    return openpyxl.load_workbook(TEMPLATE_PATH)

def get_offset_info(note_text, span_text, context_prefix=None, match_index=None):
    """
    Calculates start_char, end_char, and hydration_status.
    """
    matches = [m for m in re.finditer(re.escape(span_text), note_text)]
    
    if not matches:
        return "", "", "not_found"
    
    selected_match = None
    status = "ambiguous"

    if len(matches) == 1:
        selected_match = matches[0]
        status = "hydrated_unique"
    elif context_prefix:
        # Look for prefix in window before match
        for m in matches:
            start = m.start()
            # check 120 chars before
            window_start = max(0, start - 120)
            preceding_text = note_text[window_start:start]
            if context_prefix in preceding_text:
                selected_match = m
                status = "hydrated_prefix_window"
                break
    elif match_index is not None and match_index < len(matches):
        selected_match = matches[match_index]
        status = "hydrated_match_index"
    
    if selected_match:
        return selected_match.start(), selected_match.end(), status
    
    return "", "", f"ambiguous_count={len(matches)}"

def run_extraction():
    wb = create_workbook_if_needed()
    
    # 1. Populate Note_Text
    ws_text = wb["Note_Text"]
    # Check if headers exist, if empty add them
    if ws_text.max_row == 1 and ws_text["A1"].value is None:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Populate Note_Index
    ws_index = wb["Note_Index"]
    # Append Metadata + Flags
    row_data = [
        SOURCE_FILE, NOTE_ID, "", "", "", "", "extracted", "" # Metadata cols
    ]
    # Add flags in order
    flag_keys = list(PROCEDURE_FLAGS.keys())
    for k in flag_keys:
        row_data.append(PROCEDURE_FLAGS[k])
    ws_index.append(row_data)

    # 3. Populate Span_Annotations & 4. Span_Hydrated
    ws_anno = wb["Span_Annotations"]
    ws_hydra = wb["Span_Hydrated"]
    
    # Headers for reference (assuming template has them, but mapping by index)
    # COLS: source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, start, end, len, label, norm_val, schema, event_id, neg, hist, time, rev, comm, status
    
    span_id_counter = 1
    
    for span in SPANS:
        s_text = span["span_text"]
        c_prefix = span.get("context_prefix", "")
        
        # Calculate offsets
        start, end, status = get_offset_info(NOTE_TEXT, s_text, context_prefix=c_prefix)
        length = len(s_text)
        
        # Base row data
        row_base = [
            SOURCE_FILE,
            NOTE_ID,
            f"span_{span_id_counter:03d}",
            "", # section type
            c_prefix,
            s_text,
            "", # match_index
        ]
        
        # Annotation Row (offsets blank)
        anno_row = row_base + ["", "", f"=LEN(F{ws_anno.max_row+1})"] + [
            span["label"],
            span.get("normalized_value", ""),
            span.get("schema_field", ""),
            span.get("event_id", ""),
            0, # is_negated
            0, # is_historical
            "", # time_anchor
            "Auto", # reviewer
            span.get("comments", ""),
            "needs_hydration"
        ]
        ws_anno.append(anno_row)

        # Hydrated Row (offsets filled)
        hydra_row = row_base + [start, end, length] + [
            span["label"],
            span.get("normalized_value", ""),
            span.get("schema_field", ""),
            span.get("event_id", ""),
            0,
            0,
            "",
            "Auto",
            span.get("comments", ""),
            status
        ]
        ws_hydra.append(hydra_row)
        
        span_id_counter += 1

    # 5. Populate Event_Log
    ws_log = wb["Event_Log"]
    for evt in EVENTS_METADATA:
        # Map fields to Event_Log columns
        # source, note, event_id, type, method, anatomy, device, gauge, stations, counts, meas, spec, find, hist, rev, comm, size, mat, out_air_pre, out_air_post, out_sym, out_pl, out_comp
        row = [
            SOURCE_FILE,
            NOTE_ID,
            evt["event_id"],
            evt.get("type", ""),
            evt.get("method", ""),
            evt.get("target_anatomy", ""),
            evt.get("device", ""),
            "", # gauge
            "", # stations
            "", # counts
            "", # measurements
            "", # specimens
            evt.get("findings", ""),
            0, # historical
            "Auto", # reviewer
            "", # comments
            "", # device size (granular)
            "", # device mat
            "", # out lumen pre
            "", # out lumen post
            "", # out sym
            "", # out pleural
            evt.get("outcome_complication", "")
        ]
        ws_log.append(row)

    # 6. Populate V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    for evt in EVENTS_METADATA:
        row = [
            NOTE_ID,
            evt["event_id"],
            evt.get("type", ""),
            "", # target.type
            "", # lobe
            "", # segment
            "", # station
            "", # lesion.type
            "", # lesion.size
            evt.get("method", ""),
            json.dumps({"device": evt.get("device")}) if evt.get("device") else "",
            "", # meas json
            "", # spec json
            json.dumps({"findings": evt.get("findings")}) if evt.get("findings") else "",
            "", # quote
            "", # stent size
            "", # stent mat
            "", # cath size
            "", # air lumen pre
            "", # air lumen post
            "", # sym
            "", # pleural
            evt.get("outcome_complication", "")
        ]
        ws_v3.append(row)

    # 7. Generate V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": EVENTS_METADATA,
        "no_immediate_complications": True
    }
    
    json_str = json.dumps(registry_data, indent=2)
    ws_json.append([json_str])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Extraction saved to {OUTPUT_PATH}")

if __name__ == "__main__":
    run_extraction()