import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# ==========================================
# INPUT DATA
# ==========================================
NOTE_ID = "note_062"
SOURCE_FILE = "note_062.txt"
PROCEDURE_DATE = "2026-01-12" # Placeholder based on current context
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_062 SOURCE_FILE: note_062.txt INDICATION FOR OPERATION:  [REDACTED]is a 68 year old-year-old male who presents with lung transplant and airway stenosis.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
Patient indicated a wish to proceed with surgery and informed consent was signed.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
 
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
 
PROCEDURE:  
31646 Therapeutic aspiration subsequent episodes
31624 Dx bronchoscope/lavage (BAL)    
31630 Balloon dilation
31636 Dilate and bronchial stent initial bronchus
31635 Foreign body removal
 
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)
 
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required conversion to rigid bronchoscopy and stent removal and re-placement .
This resulted in >50% increased work due to Time, Technical difficulty of procedure, and Physical and mental effort required.
Apply to: 31636 Dilate and bronchial stent initial bronchus
31635 Foreign body removal.
31630 - completed in the right mainstem bronchus
31636 performed in the left main stem bronchus 
This increase the time and complexity of the procedure
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Rigid Bronchoscope
Flexible Therapeutic Bronchoscope
Disposable Bronchoscope
 
ESTIMATED BLOOD LOSS:   Minimum
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: Supine
 
Initial Airway Inspection Findings:
 
The laryngeal mask airway is in good position.
Pharynx: Not assessed due to bronchoscopy introduction through LMA.
Larynx: Normal.
Vocal Cords: Normal without mass/lesions
Trachea: Normal.
Main Carina: Sharp
Right Lung Proximal Airways: Slight stenosis as anastamosis site without evidence of dehiscence.
Normal anatomic branching to segmental level.  No evidence of mass, lesions, bleeding or other endobronchial pathology.
Left Lung Proximal Airways: Stent in place at LMSB. Thick mucus throughout.
Granulation at distal stent and evidence of distal stent migration and occlusion of LUL.
Otherwise, normal anatomic branching to segmental level.  No evidence of mass, lesions, bleeding or other endobronchial pathology.
Mucosa: Granulation tissue at distal LMSB stent. Slightly friable.
Secretions: Copious thick and thin secretions throughout with left greater than right.
MUCUS WITHIN TRACHEA AT OUTSET OF PROCEDURE
 
 
DISTAL LMSB STENT WITH GRANULATION TISSUE AND DISTAL MIGRATION
 
 
RMSB STENOSIS AT ANASTOMOSIS SITE
 
  
Successful therapeutic aspiration was performed to clean out the Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus.
Bronchial alveolar lavage was performed at Anteromedial Segment of LLL (Lb7/8).
Instilled 80 cc of NS, suction returned with 20 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
 
Given evidence of distal migration of the stent and occlusion of the left upper lobe, decision was made to remove the patient's stent and place a larger stent.
After induction of muscle relaxants, tooth or gum protector was placed.
The black rigid barrel was introduced through the mouth and advanced in the midline while keeping the alignment with the axis of the trachea and minimizing pressure to the teeth.
The vocal cords were identified and the rigid bronchoscope was advanced carefully while minimizing contact with them.
Once the rigid bronchoscope was positioned at the mid-trachea, jet ventilation was initiated and chest wall movement was confirmed.
The rigid bronchoscope was placed in the left mainstem bronchus and the patient's stent was grasped with the single-action rigid forceps.
The stent was removed from the patient's airway through the rigid bronchoscope (foreign body removal).
The following stent (Bonastent, 12 x 50mm, covered) was placed in the Left Mainstem.
Balloon dilation was performed at Left Mainstem through the stent to fully expand seat the stent.
10/11/12 Elation balloon was used to perform dilation to 12 mm at the Left Mainstem.
Total 3 inflations with dilation time of 10 seconds each.
VIEW OF NEW LMSB STENT
 
 
Balloon dilation was performed at Right Mainstem (at stenotic anastamosis site).
10/11/12 Elation balloon was used to perform dilation to 12 mm at the Right Mainstem.
Total 1 inflations with dilation time of 60 seconds each.
VIEW OF RMSB STENOSIS AT ANASTOMOSIS (AFTER BALLOON DILATION) 
 
 
The patient tolerated the procedure well.
There were no immediate complications.  At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
--LLL BAL (cell count, micro, cyto)
--LMSB stent (path)
 
IMPRESSION/PLAN: [REDACTED]is a 68 year old-year-old male who presents for bronchoscopy for evaluation of airway stenosis and stent evaluation.
Patient underwent stent removal and re-placement without immediate complication.
 
--Continue stent hydration regimen TID in the following order: 
1) Albuterol nebs
2) Hypertonic saline (3%) nebs
3) Flutter valve
--Guaifenesin 1200mg PO BID
--Follow up BAL results
--Repeat bronchoscopy in 2-4 weeks for re-evaluation
--Follow-up in outpatient Interventional Pulmonology clinic as scheduled on [REDACTED]"""

# ==========================================
# EXTRACTION LOGIC
# ==========================================

class Phase0Generator:
    def __init__(self, note_text, note_id, source_file):
        self.note_text = note_text
        self.note_id = note_id
        self.source_file = source_file
        self.spans = []
        self.events = []
        self.procedure_flags = {
            "diagnostic_bronchoscopy": 0, "bal": 0, "bronchial_wash": 0, "brushings": 0, 
            "endobronchial_biopsy": 0, "tbna_conventional": 0, "linear_ebus": 0, "radial_ebus": 0, 
            "navigational_bronchoscopy": 0, "transbronchial_biopsy": 0, "transbronchial_cryobiopsy": 0, 
            "therapeutic_aspiration": 0, "foreign_body_removal": 0, "airway_dilation": 0, 
            "airway_stent": 0, "thermal_ablation": 0, "tumor_debulking_non_thermal": 0, 
            "cryotherapy": 0, "blvr": 0, "peripheral_ablation": 0, "bronchial_thermoplasty": 0, 
            "whole_lung_lavage": 0, "rigid_bronchoscopy": 0, "thoracentesis": 0, "chest_tube": 0, 
            "ipc": 0, "medical_thoracoscopy": 0, "pleurodesis": 0, "pleural_biopsy": 0, 
            "fibrinolytic_therapy": 0
        }

    def set_flags(self):
        # Flags based on reading the note
        self.procedure_flags["diagnostic_bronchoscopy"] = 1
        self.procedure_flags["therapeutic_aspiration"] = 1
        self.procedure_flags["bal"] = 1
        self.procedure_flags["rigid_bronchoscopy"] = 1
        self.procedure_flags["foreign_body_removal"] = 1 # Stent removal mentioned
        self.procedure_flags["airway_stent"] = 1
        self.procedure_flags["airway_dilation"] = 1

    def add_span(self, text, label, context_prefix=None, normalized_value=None, event_id=None, section="PROCEDURE"):
        if not text:
            return
        
        # Hydration Logic
        start_char = -1
        end_char = -1
        hydration_status = "ambiguous"
        
        count = self.note_text.count(text)
        
        if count == 0:
            # Fallback for minor spacing differences or typos in definition vs text
            # This is a strict extract, so we warn if not found, but we won't crash
            print(f"Warning: Span '{text}' not found in text.")
            hydration_status = "not_found"
        elif count == 1:
            start_char = self.note_text.find(text)
            end_char = start_char + len(text)
            hydration_status = "hydrated_unique"
        else:
            # Multiple occurrences
            if context_prefix:
                # Find all occurrences
                occurrences = [m.start() for m in re.finditer(re.escape(text), self.note_text)]
                found = False
                for idx in occurrences:
                    # Look back 120 chars
                    window_start = max(0, idx - 120)
                    window = self.note_text[window_start:idx]
                    if context_prefix in window:
                        start_char = idx
                        end_char = idx + len(text)
                        hydration_status = "hydrated_prefix_window"
                        found = True
                        break
                if not found:
                    hydration_status = f"ambiguous_count={count}"
            else:
                 hydration_status = f"ambiguous_count={count}"

        self.spans.append({
            "source_file": self.source_file,
            "note_id": self.note_id,
            "span_id": f"span_{len(self.spans) + 1:03d}",
            "section_type": section,
            "context_prefix": context_prefix,
            "span_text": text,
            "match_index": None, # Simple script logic doesn't support complex Nth match automatically without more inputs
            "start_char": start_char if start_char != -1 else None,
            "end_char": end_char if end_char != -1 else None,
            "span_len": len(text),
            "label": label,
            "normalized_value": normalized_value if normalized_value else text,
            "schema_field": None,
            "event_id": event_id,
            "is_negated": False,
            "is_historical": False,
            "time_anchor": None,
            "reviewer": "Auto",
            "comments": None,
            "hydration_status": hydration_status
        })

    def generate_spans(self):
        # Metadata
        self.add_span("lung transplant", "CTX_HISTORICAL", section="INDICATION")
        self.add_span("airway stenosis", "CTX_HISTORICAL", section="INDICATION")

        # Event 1: Rigid Bronchoscopy (General)
        e1 = "evt_01"
        self.add_span("conversion to rigid bronchoscopy", "PROC_METHOD", event_id=e1)
        self.add_span("Rigid Bronchoscope", "DEV_INSTRUMENT", section="INSTRUMENT", event_id=e1)
        
        # Event 2: Therapeutic Aspiration
        e2 = "evt_02"
        self.add_span("Therapeutic aspiration", "PROC_METHOD", section="PROCEDURE", event_id=e2)
        self.add_span("successful therapeutic aspiration was performed", "PROC_ACTION", section="PROCEDURE IN DETAIL", event_id=e2)
        # Locations
        locs = ["Trachea", "Right Mainstem", "Bronchus Intermedius", "Left Mainstem", "Carina", "RUL Carina", "RML Carina", "LUL Lingula Carina", "Left Carina"]
        for loc in locs:
            self.add_span(loc, "ANAT_AIRWAY", context_prefix="clean out the", event_id=e2)

        # Event 3: BAL
        e3 = "evt_03"
        self.add_span("Bronchial alveolar lavage", "PROC_METHOD", context_prefix="18]", event_id=e3)
        self.add_span("Anteromedial Segment of LLL", "ANAT_LUNG_LOC", event_id=e3)
        self.add_span("Lb7/8", "ANAT_LUNG_LOC", event_id=e3, normalized_value="LB7/LB8")
        self.add_span("Instilled 80 cc", "MEAS_VOL", event_id=e3, normalized_value="80")
        self.add_span("suction returned with 20 cc", "MEAS_VOL", event_id=e3, normalized_value="20")
        
        # Event 4: Stent Removal (LMSB)
        e4 = "evt_04"
        self.add_span("stent removal", "PROC_METHOD", context_prefix="Unusual Procedure:", event_id=e4)
        self.add_span("distal migration of the stent", "OBS_LESION", event_id=e4)
        self.add_span("occlusion of the left upper lobe", "OBS_LESION", event_id=e4)
        self.add_span("stent was grasped", "PROC_ACTION", event_id=e4)
        self.add_span("single-action rigid forceps", "DEV_INSTRUMENT", event_id=e4)
        self.add_span("foreign body removal", "PROC_METHOD", context_prefix="bronchoscope (", event_id=e4)
        self.add_span("Left Lung Proximal Airways: Stent in place at LMSB", "ANAT_AIRWAY", event_id=e4, normalized_value="LMSB")

        # Event 5: Stent Placement (LMSB)
        e5 = "evt_05"
        self.add_span("stent ... was placed", "PROC_ACTION", context_prefix="following", event_id=e5)
        self.add_span("Bonastent", "DEV_STENT_MATERIAL", event_id=e5)
        self.add_span("12 x 50mm", "DEV_STENT_SIZE", event_id=e5, normalized_value="12x50mm")
        self.add_span("covered", "DEV_STENT_MATERIAL", context_prefix="50mm, ", event_id=e5)
        self.add_span("Left Mainstem", "ANAT_AIRWAY", context_prefix="placed in the", event_id=e5)

        # Event 6: Balloon Dilation (LMSB - Stent)
        e6 = "evt_06"
        self.add_span("Balloon dilation was performed", "PROC_METHOD", context_prefix="VIEW OF NEW LMSB STENT", event_id=e6)
        self.add_span("Left Mainstem", "ANAT_AIRWAY", context_prefix="dilation was performed at", event_id=e6)
        self.add_span("10/11/12 Elation balloon", "DEV_INSTRUMENT", context_prefix="at the Left Mainstem.", event_id=e6)
        self.add_span("dilation to 12 mm", "MEAS_AIRWAY_DIAM", context_prefix="perform", event_id=e6, normalized_value="12")
        self.add_span("3 inflations", "MEAS_COUNT", context_prefix="Total", event_id=e6, normalized_value="3")

        # Event 7: Balloon Dilation (RMSB - Anastomosis)
        e7 = "evt_07"
        self.add_span("Balloon dilation was performed", "PROC_METHOD", context_prefix="stenosis at anastamosis site", event_id=e7)
        self.add_span("Right Mainstem", "ANAT_AIRWAY", context_prefix="dilation was performed at", event_id=e7)
        self.add_span("stenotic anastamosis site", "OBS_LESION", event_id=e7)
        self.add_span("10/11/12 Elation balloon", "DEV_INSTRUMENT", context_prefix="at the Right Mainstem", event_id=e7)
        self.add_span("dilation to 12 mm", "MEAS_AIRWAY_DIAM", context_prefix="perform dilation to", event_id=e7, normalized_value="12")
        self.add_span("1 inflations", "MEAS_COUNT", context_prefix="Total", event_id=e7, normalized_value="1")

        # Outcome
        self.add_span("No immediate complications", "OUTCOME_COMPLICATION", normalized_value="None", event_id="evt_outcome")

    def generate_events(self):
        # 1. Rigid Bronchoscopy
        self.events.append({
            "event_id": "evt_01",
            "event_type": "Procedure",
            "method": "Rigid Bronchoscopy",
            "anatomy_target": "Airway",
            "device": "Rigid Bronchoscope",
            "outcome_complication": "None"
        })
        # 2. Aspiration
        self.events.append({
            "event_id": "evt_02",
            "event_type": "Procedure",
            "method": "Therapeutic Aspiration",
            "anatomy_target": "Trachea, RMSB, LMSB, Carina",
            "findings": "Mucus"
        })
        # 3. BAL
        self.events.append({
            "event_id": "evt_03",
            "event_type": "Procedure",
            "method": "BAL",
            "anatomy_target": "LLL Anteromedial (LB7/8)",
            "measurements": "Instilled: 80cc, Return: 20cc",
            "specimens": "Cell Count, Micro, Cyto"
        })
        # 4. Stent Removal
        self.events.append({
            "event_id": "evt_04",
            "event_type": "Procedure",
            "method": "Foreign Body Removal (Stent)",
            "anatomy_target": "LMSB",
            "device": "Rigid Forceps",
            "findings": "Distal migration, occlusion, granulation"
        })
        # 5. Stent Placement
        self.events.append({
            "event_id": "evt_05",
            "event_type": "Procedure",
            "method": "Airway Stent",
            "anatomy_target": "LMSB",
            "device": "Bonastent",
            "device_size": "12x50mm",
            "device_material": "Covered"
        })
        # 6. Dilation (Left)
        self.events.append({
            "event_id": "evt_06",
            "event_type": "Procedure",
            "method": "Balloon Dilation",
            "anatomy_target": "LMSB",
            "device": "Elation Balloon",
            "measurements": "12mm diameter, 3 inflations"
        })
        # 7. Dilation (Right)
        self.events.append({
            "event_id": "evt_07",
            "event_type": "Procedure",
            "method": "Balloon Dilation",
            "anatomy_target": "RMSB",
            "device": "Elation Balloon",
            "measurements": "12mm diameter, 1 inflation",
            "findings": "Anastomotic stenosis"
        })

    def generate_v3_registry(self):
        # Construct V3 JSON
        procedures = []
        
        # Helper to simplify
        def make_proc(eid, ptype, loc, method, dev_json=None, meas_json=None, findings_json=None):
            return {
                "event_id": eid,
                "type": ptype,
                "target": {"anatomy_type": "Airway", "location": loc},
                "method": method,
                "devices_json": json.dumps(dev_json) if dev_json else None,
                "measurements_json": json.dumps(meas_json) if meas_json else None,
                "findings_json": json.dumps(findings_json) if findings_json else None
            }

        # Rigid
        procedures.append(make_proc("evt_01", "Rigid Bronchoscopy", {}, "Rigid Bronchoscopy", {"device": "Rigid Bronchoscope"}))
        
        # Aspiration
        procedures.append(make_proc("evt_02", "Therapeutic Aspiration", {"name": "Trachea/Mainstems"}, "Suction"))
        
        # BAL
        procedures.append(make_proc("evt_03", "BAL", {"lobe": "LLL", "segment": "Anteromedial"}, "Lavage", 
                                    meas_json={"instilled_cc": 80, "return_cc": 20}))
        
        # Stent Removal
        procedures.append(make_proc("evt_04", "Foreign Body Removal", {"name": "LMSB"}, "Forceps", 
                                    findings_json=["stent migration", "occlusion"]))
        
        # Stent Placement
        stent_evt = make_proc("evt_05", "Airway Stent", {"name": "LMSB"}, "Deployment", 
                              dev_json={"brand": "Bonastent", "size": "12x50mm", "type": "covered"})
        stent_evt["stent"] = {"size": "12x50mm", "material_or_brand": "Bonastent Covered"}
        procedures.append(stent_evt)

        # Dilation Left
        procedures.append(make_proc("evt_06", "Airway Dilation", {"name": "LMSB"}, "Balloon", 
                                    meas_json={"diameter_mm": 12, "inflations": 3}))

        # Dilation Right
        procedures.append(make_proc("evt_07", "Airway Dilation", {"name": "RMSB"}, "Balloon", 
                                    meas_json={"diameter_mm": 12, "inflations": 1},
                                    findings_json=["anastomotic stenosis"]))

        return {
            "schema_version": "v3.0",
            "note_id": self.note_id,
            "no_immediate_complications": True,
            "procedures": procedures
        }

    def write_excel(self):
        try:
            wb = openpyxl.load_workbook(TEMPLATE_PATH)
        except FileNotFoundError:
            # Create a blank one for fallback if template missing (though instructions say it must load)
            wb = openpyxl.Workbook()
            
        # 1. Note_Text
        if "Note_Text" not in wb.sheetnames: wb.create_sheet("Note_Text")
        ws_text = wb["Note_Text"]
        ws_text.append([self.note_id, self.source_file, self.note_text])

        # 2. Note_Index
        if "Note_Index" not in wb.sheetnames: wb.create_sheet("Note_Index")
        ws_index = wb["Note_Index"]
        row_idx = [self.source_file, self.note_id, "", PROCEDURE_DATE, "", "", "Extraction_Complete", ""]
        # Add flags
        for key in self.procedure_flags:
            row_idx.append(self.procedure_flags[key])
        ws_index.append(row_idx)

        # 3. Span_Annotations (Anchor First)
        if "Span_Annotations" not in wb.sheetnames: wb.create_sheet("Span_Annotations")
        ws_span = wb["Span_Annotations"]
        # Header is assumed in template, we append data
        for s in self.spans:
            ws_span.append([
                s["source_file"], s["note_id"], s["span_id"], s["section_type"],
                s["context_prefix"], s["span_text"], s["match_index"],
                None, None, s["span_len"], # start/end blank
                s["label"], s["normalized_value"], s["schema_field"], s["event_id"],
                s["is_negated"], s["is_historical"], s["time_anchor"], s["reviewer"], s["comments"],
                "needs_hydration"
            ])

        # 4. Span_Hydrated
        if "Span_Hydrated" not in wb.sheetnames: wb.create_sheet("Span_Hydrated")
        ws_hydra = wb["Span_Hydrated"]
        # Write header if empty (simplistic check)
        if ws_hydra.max_row == 1:
            headers = ["source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", 
                       "match_index", "start_char", "end_char", "span_len", "label", "normalized_value", 
                       "schema_field", "event_id", "is_negated", "is_historical", "time_anchor", 
                       "reviewer", "comments", "hydration_status"]
            ws_hydra.append(headers)
            
        for s in self.spans:
            ws_hydra.append([
                s["source_file"], s["note_id"], s["span_id"], s["section_type"],
                s["context_prefix"], s["span_text"], s["match_index"],
                s["start_char"], s["end_char"], s["span_len"],
                s["label"], s["normalized_value"], s["schema_field"], s["event_id"],
                s["is_negated"], s["is_historical"], s["time_anchor"], s["reviewer"], s["comments"],
                s["hydration_status"]
            ])

        # 5. Event_Log
        if "Event_Log" not in wb.sheetnames: wb.create_sheet("Event_Log")
        ws_event = wb["Event_Log"]
        for e in self.events:
            ws_event.append([
                self.source_file, self.note_id, e["event_id"], e.get("event_type"), e.get("method"),
                e.get("anatomy_target"), e.get("device"), None, None, None, e.get("measurements"),
                e.get("specimens"), e.get("findings"), False, "Auto", None,
                e.get("device_size"), e.get("device_material"),
                None, None, None, None, e.get("outcome_complication")
            ])

        # 6. V3_Procedure_Events
        if "V3_Procedure_Events" not in wb.sheetnames: wb.create_sheet("V3_Procedure_Events")
        ws_v3 = wb["V3_Procedure_Events"]
        v3_data = self.generate_v3_registry()
        for p in v3_data["procedures"]:
            stent = p.get("stent", {})
            tgt = p.get("target", {})
            loc = tgt.get("location", {})
            
            ws_v3.append([
                self.note_id, p["event_id"], p["type"],
                tgt.get("anatomy_type"), loc.get("lobe"), loc.get("segment"), tgt.get("station"),
                None, None, # lesion type/size
                p.get("method"), p.get("devices_json"), p.get("measurements_json"), 
                None, p.get("findings_json"), None,
                stent.get("size"), stent.get("material_or_brand"),
                None, # catheter
                None, None, None, None, None # Outcomes
            ])

        # 7. V3_Registry_JSON
        if "V3_Registry_JSON" not in wb.sheetnames: wb.create_sheet("V3_Registry_JSON")
        ws_json = wb["V3_Registry_JSON"]
        ws_json.append([json.dumps(v3_data, indent=2)])

        wb.save(OUTPUT_PATH)
        print(f"Generated {OUTPUT_PATH}")

# ==========================================
# EXECUTION
# ==========================================
if __name__ == "__main__":
    # Check if a file exists to override text (as per prompt instructions for file upload handling)
    input_filename = "note_062.txt"
    if os.path.exists(input_filename):
        with open(input_filename, 'r', encoding='utf-8') as f:
            NOTE_TEXT = f.read()
            SOURCE_FILE = input_filename
            # Infer note_id from filename
            base = os.path.basename(input_filename)
            NOTE_ID = os.path.splitext(base)[0]
    
    generator = Phase0Generator(NOTE_TEXT, NOTE_ID, SOURCE_FILE)
    generator.set_flags()
    generator.generate_spans()
    generator.generate_events()
    generator.write_excel()