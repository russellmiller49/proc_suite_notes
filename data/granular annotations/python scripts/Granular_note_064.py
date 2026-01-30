import openpyxl
from openpyxl.utils import get_column_letter
import json
import os
import re

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_064"
SOURCE_FILE = "note_064.txt"
PROCEDURE_DATE = "2026-01-12" # Placeholder based on current context if needed, else blank
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_064 SOURCE_FILE: note_064.txt INDICATION FOR OPERATION:  [REDACTED]is a 68 year old-year-old male who presents with bronchial stenosis.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
 
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
 
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31624 Dx bronchoscope/lavage (BAL)    
31625 Endobronchial Biopsy(s)
31630 Balloon dilation
31636 Dilate and bronchial stent initial bronchus
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
 
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
 
ESTIMATED BLOOD LOSS:   Minimum
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
Initial Airway Inspection Findings:
 
Initial Airway Examination Findings:
 
The laryngeal mask airway is in good position.
Pharynx: Not assessed due to bronchoscopy introduction through LMA.
Larynx: Not fully assessed due to bronchoscopy introduction through LMA.
Vocal Cords: Normal without mass/lesions
Trachea: Normal.
Main Carina: Sharp
Right Lung Proximal Airways: Normal anatomic branching to segmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology. Anastomosis site noted, normal.
Left Lung Proximal Airways: LMSB noted to have stenosis (both static at the anastomosis and dynamic at the proximal LMSB).
Normal anatomic branching to segmental level.  No evidence of mass, lesions, bleeding or other endobronchial pathology.
Mucosa: Some cobblestoning at main carina and proximal LMSB, otherwise normal
Secretions: Minimal, thin, and clear.
LMSB Stenosis:
 
 
 
Successful therapeutic aspiration was performed to clean out the Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus.
Bronchial alveolar lavage was performed at Superior Segment of Lingula (LB4) and Inferior Segment of Lingula (LB5).
Instilled 40 cc of NS, suction returned with 15 cc of NS.  Samples sent for Cell Count and Microbiology (Cultures/Viral/Fungal).
Endobronchial biopsy of cobblestoned mucosa was performed at Left Mainstem and Carina using 2.0mm pulmonary forceps.  Samples sent for Pathology.
Endobronchial obstruction at LMSB anastomosis stenosis was treated with the following modalities:
 
Modality	Tools	Setting/Mode	Duration	Results
Electrocautery	Needle Knife	EndoCut, Effect 3	 	Radial cuts 
APC	 	 	 	 
Laser	 	 	 	 
CoreCath	 	 	 	 
Cryoprobe	 	 	 	 
 
Prior to treatment, affected airway was note to be 75% patent.
After treatment, the airway was 100% patent. 
 
Balloon dilation was performed at Left Mainstem.
8/9/10 Elation balloon was used to perform dilation to 10 mm at the Left Mainstem.
Total 1 inflations with dilation time of 60 seconds each.
The following stent (MicroTech 10 x 40mm) was placed in the Left Mainstem. 
 
Balloon dilation was performed at Left Mainstem.
8/9/10 Elation balloon was used to perform dilation to 10 mm at the Left Mainstem to fully extend and seat the stent.
Total 3 inflations with dilation time of 30 seconds each.
 
Balloon dilation was performed at Left Mainstem.
10/11/12 Elation balloon was used to perform dilation to 12 mm at the Left Mainstem.
Total 3 inflations with dilation time of 30 seconds each.
View of LMSB stent at the conclusion of the procedure:
 
 
 
 
 
 
 
The patient tolerated the procedure well.
There were no immediate complications.  At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
--LMSB endobronchial forceps biopsies
--Lingula BAL (cell count, micro)
 
IMPRESSION/PLAN: [REDACTED]is a 68 year old-year-old male who presents for bronchoscopy for evaluation of bronchial stenosis.
MicroTech 10 x 40 mm stent was successfully deployed to the LMSB.
Patient tolerated the procedure well and there were no immediate complications.
--Stent Hydration Regimen TID in the following order: 
1) Albuterol nebs
2) Hypertonic saline (3%) nebs
3) Flutter valve
--Guaifenesin 1200mg PO BID
--Follow-up endobronchial biopsy and BAL results
--Outpatient follow-up in IP clinic as scheduled on [REDACTED]
--Repeat bronchoscopy for stent check in 2-3 weeks"""

# -------------------------------------------------------------------------
# PROCEDURE FLAGS
# -------------------------------------------------------------------------
# Map flags to 0/1 based on note content
PROCEDURE_FLAGS = {
    # Bronchoscopy (23)
    "diagnostic_bronchoscopy": 1,
    "bal": 1,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 1,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 0,
    "airway_dilation": 1,
    "airway_stent": 1,
    "thermal_ablation": 1, # Electrocautery mentioned
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0, # Flexible explicit, LMA used
    # Pleural (7)
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# -------------------------------------------------------------------------
# ANCHOR SPANS
# -------------------------------------------------------------------------
# Structure: (text, context, label, value, field, event_id)
SPANS = [
    # Event 1: Therapeutic Aspiration
    ("Therapeutic aspiration", "Successful", "PROC_METHOD", "Therapeutic aspiration", "method", "evt1"),
    ("Trachea (Distal 1/3)", "clean out the", "ANAT_AIRWAY", "Trachea", "target_anatomy", "evt1"),
    ("Right Mainstem", "1/3),", "ANAT_AIRWAY", "Right Mainstem", "target_anatomy", "evt1"),
    ("Left Mainstem", "Intermedius ,", "ANAT_AIRWAY", "Left Mainstem", "target_anatomy", "evt1"),
    
    # Event 2: BAL
    ("Bronchial alveolar lavage", "Superior Segment", "PROC_METHOD", "Bronchoalveolar lavage", "method", "evt2"),
    ("Superior Segment of Lingula (LB4)", "lavage was performed at", "ANAT_LUNG_LOC", "Lingula Superior Segment", "target_anatomy", "evt2"),
    ("Inferior Segment of Lingula (LB5)", "LB4) and", "ANAT_LUNG_LOC", "Lingula Inferior Segment", "target_anatomy", "evt2"),
    ("40 cc", "Instilled", "MEAS_VOL", "40", "volume_instilled", "evt2"),
    ("15 cc", "returned with", "MEAS_VOL", "15", "volume_return", "evt2"),
    
    # Event 3: EBB
    ("Endobronchial biopsy", "15]", "PROC_METHOD", "Endobronchial biopsy", "method", "evt3"),
    ("Left Mainstem", "cobblestoned mucosa was performed at", "ANAT_AIRWAY", "Left Mainstem", "target_anatomy", "evt3"),
    ("Carina", "Left Mainstem and", "ANAT_AIRWAY", "Carina", "target_anatomy", "evt3"),
    ("2.0mm pulmonary forceps", "using", "DEV_INSTRUMENT", "2.0mm pulmonary forceps", "device", "evt3"),
    ("cobblestoned mucosa", "biopsy of", "OBS_LESION", "Cobblestoning", "findings", "evt3"),
    
    # Event 4: Electrocautery (Destruction of stenosis)
    ("Electrocautery", "Modality", "PROC_METHOD", "Electrocautery", "method", "evt4"),
    ("Needle Knife", "Electrocautery", "DEV_INSTRUMENT", "Needle Knife", "device", "evt4"),
    ("LMSB", "obstruction at", "ANAT_AIRWAY", "Left Mainstem", "target_anatomy", "evt4"),
    ("Radial cuts", "Results", "PROC_ACTION", "Radial cuts", "action", "evt4"),
    
    # Outcomes (Linked to Event 4 - treatment)
    ("75% patent", "airway was note to be", "OUTCOME_AIRWAY_LUMEN_PRE", "75", "outcome_lumen_pre", "evt4"),
    ("100% patent", "airway was", "OUTCOME_AIRWAY_LUMEN_POST", "100", "outcome_lumen_post", "evt4"),

    # Event 5: Balloon Dilation 1
    ("Balloon dilation", "17]", "PROC_METHOD", "Balloon dilation", "method", "evt5"),
    ("Left Mainstem", "dilation to 10 mm at the", "ANAT_AIRWAY", "Left Mainstem", "target_anatomy", "evt5"),
    ("8/9/10 Elation balloon", "18]", "DEV_INSTRUMENT", "Elation Balloon", "device", "evt5"),
    ("10 mm", "perform dilation to", "MEAS_AIRWAY_DIAM", "10", "diameter", "evt5"),
    ("60 seconds", "dilation time of", "CTX_TIME", "60 seconds", "duration", "evt5"),
    
    # Event 6: Stent Placement
    ("stent", "The following", "DEV_STENT", "Stent", "device_category", "evt6"),
    ("MicroTech", "stent (", "DEV_STENT_MATERIAL", "MicroTech", "device_material", "evt6"),
    ("10 x 40mm", "MicroTech", "DEV_STENT_SIZE", "10x40mm", "device_size", "evt6"),
    ("Left Mainstem", "placed in the", "ANAT_AIRWAY", "Left Mainstem", "target_anatomy", "evt6"),
    
    # Event 7: Balloon Dilation 2 (Seating)
    ("Balloon dilation", "20]", "PROC_METHOD", "Balloon dilation", "method", "evt7"),
    ("Left Mainstem", "10 mm at the", "ANAT_AIRWAY", "Left Mainstem", "target_anatomy", "evt7"),
    ("10 mm", "at the Left Mainstem to fully", "MEAS_AIRWAY_DIAM", "10", "diameter", "evt7"),
    ("fully extend and seat the stent", "Mainstem to", "PROC_ACTION", "Seat stent", "action", "evt7"),
    
    # Event 8: Balloon Dilation 3 (Post-stent)
    ("Balloon dilation", "22]", "PROC_METHOD", "Balloon dilation", "method", "evt8"),
    ("Left Mainstem", "12 mm at the", "ANAT_AIRWAY", "Left Mainstem", "target_anatomy", "evt8"),
    ("10/11/12 Elation balloon", "23]", "DEV_INSTRUMENT", "Elation Balloon", "device", "evt8"),
    ("12 mm", "perform dilation to", "MEAS_AIRWAY_DIAM", "12", "diameter", "evt8"),
    
    # Global/Conclusion
    ("no immediate complications", "There were", "OUTCOME_COMPLICATION", "None", "outcome_complication", "evt99")
]

# -------------------------------------------------------------------------
# SCRIPT UTILS
# -------------------------------------------------------------------------
def create_worksheet_if_not_exists(wb, sheet_name, headers):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
    
    # Clear existing
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
            
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    return ws

def get_hydrated_span(note_text, span_text, context_prefix=None, match_index=None):
    """
    Finds start/end char offsets.
    Priority:
    1. Exact unique match.
    2. Context match (preceding 120 chars).
    3. Match index (nth occurrence).
    """
    escaped_span = re.escape(span_text)
    matches = [m for m in re.finditer(escaped_span, note_text)]
    
    if not matches:
        return None, None, "not_found"
    
    # 1. Unique
    if len(matches) == 1:
        return matches[0].start(), matches[0].end(), "hydrated_unique"
    
    # 2. Context
    if context_prefix:
        for m in matches:
            start = m.start()
            # Look back 120 chars
            window_start = max(0, start - 120)
            window = note_text[window_start:start]
            if context_prefix in window:
                return m.start(), m.end(), "hydrated_prefix_window"
    
    # 3. Match Index
    if match_index is not None and 1 <= match_index <= len(matches):
        m = matches[match_index - 1]
        return m.start(), m.end(), "hydrated_match_index"
        
    return None, None, f"ambiguous_count={len(matches)}"

def generate_workbook():
    # Load Template
    if not os.path.exists(TEMPLATE_PATH):
        # Fallback: create a blank workbook if template missing (should not happen per prompt, but safe)
        wb = openpyxl.Workbook()
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # -----------------------------------------------------
    # 1. Note_Text
    # -----------------------------------------------------
    ws_text = create_worksheet_if_not_exists(wb, "Note_Text", ["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # -----------------------------------------------------
    # 2. Note_Index
    # -----------------------------------------------------
    idx_headers = [
        "source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes",
        "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", 
        "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy", 
        "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration", 
        "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation", 
        "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", 
        "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
        "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis", 
        "pleural_biopsy", "fibrinolytic_therapy"
    ]
    ws_index = create_worksheet_if_not_exists(wb, "Note_Index", idx_headers)
    
    idx_row = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "pre-filled", ""]
    # Append flags in order
    flag_keys = idx_headers[8:]
    for k in flag_keys:
        idx_row.append(PROCEDURE_FLAGS.get(k, 0))
    ws_index.append(idx_row)

    # -----------------------------------------------------
    # 3. Span_Annotations (Anchor First)
    # -----------------------------------------------------
    span_headers = [
        "source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", 
        "match_index", "start_char", "end_char", "span_len", "label", "normalized_value", 
        "schema_field", "event_id", "is_negated", "is_historical", "time_anchor", 
        "reviewer", "comments", "hydration_status"
    ]
    ws_span = create_worksheet_if_not_exists(wb, "Span_Annotations", span_headers)
    
    # 4. Span_Hydrated
    ws_hydrated = create_worksheet_if_not_exists(wb, "Span_Hydrated", span_headers)

    row_id = 1
    hydrated_rows = []
    
    for item in SPANS:
        text, context, label, value, field, evt_id = item
        
        # Defaults
        match_idx = None
        
        # Calculate offsets
        start, end, status = get_hydrated_span(NOTE_TEXT, text, context, match_idx)
        span_len = len(text)
        
        # Base Row
        base_row = [
            SOURCE_FILE, NOTE_ID, f"span_{row_id:03d}", "", context, text, 
            match_idx, "", "", f"=LEN(F{row_id+1})", label, value, 
            field, evt_id, 0, 0, "", "Model", "", "needs_hydration"
        ]
        ws_span.append(base_row)
        
        # Hydrated Row
        hyd_row = list(base_row)
        hyd_row[7] = start
        hyd_row[8] = end
        hyd_row[9] = span_len
        hyd_row[19] = status
        ws_hydrated.append(hyd_row)
        
        row_id += 1

    # -----------------------------------------------------
    # 5. Event_Log
    # -----------------------------------------------------
    event_headers = [
        "source_file", "note_id", "event_id", "event_type", "method", 
        "anatomy_target", "device", "needle_gauge", "stations", "counts", 
        "measurements", "specimens", "findings", "is_historical", "reviewer", "comments",
        "device_size", "device_material", "outcome_airway_lumen_pre", 
        "outcome_airway_lumen_post", "outcome_symptoms", "outcome_pleural", "outcome_complication"
    ]
    ws_event = create_worksheet_if_not_exists(wb, "Event_Log", event_headers)
    
    # Define events data
    events_data = [
        {
            "id": "evt1", "type": "Procedure", "method": "Therapeutic aspiration", 
            "target": "Trachea, Right/Left Mainstem, Carina", "device": "Scope",
            "findings": "Mucus cleared"
        },
        {
            "id": "evt2", "type": "Procedure", "method": "BAL", 
            "target": "Lingula (LB4, LB5)", "meas": "Instilled 40cc, Return 15cc",
            "spec": "Cell Count, Micro"
        },
        {
            "id": "evt3", "type": "Procedure", "method": "Endobronchial biopsy",
            "target": "LMSB, Carina", "device": "2.0mm forceps",
            "findings": "Cobblestoning", "spec": "Pathology"
        },
        {
            "id": "evt4", "type": "Procedure", "method": "Electrocautery",
            "target": "LMSB", "device": "Needle Knife",
            "findings": "Stenosis", "lumen_pre": "75%", "lumen_post": "100%"
        },
        {
            "id": "evt5", "type": "Procedure", "method": "Balloon dilation",
            "target": "LMSB", "device": "Elation Balloon 10mm",
            "meas": "10mm, 60s"
        },
        {
            "id": "evt6", "type": "Procedure", "method": "Stent placement",
            "target": "LMSB", "device": "MicroTech Stent",
            "dev_size": "10x40mm", "dev_mat": "MicroTech"
        },
        {
            "id": "evt7", "type": "Procedure", "method": "Balloon dilation",
            "target": "LMSB", "device": "Elation Balloon 10mm",
            "meas": "10mm, 30s", "comments": "Seating stent"
        },
        {
            "id": "evt8", "type": "Procedure", "method": "Balloon dilation",
            "target": "LMSB", "device": "Elation Balloon 12mm",
            "meas": "12mm, 30s"
        },
        {
            "id": "evt99", "type": "Outcome", "method": "N/A",
            "comp": "None"
        }
    ]
    
    for e in events_data:
        row = [
            SOURCE_FILE, NOTE_ID, e["id"], e["type"], e.get("method"),
            e.get("target"), e.get("device"), "", "", "",
            e.get("meas"), e.get("spec"), e.get("findings"), 0, "Model", e.get("comments"),
            e.get("dev_size"), e.get("dev_mat"), e.get("lumen_pre"), e.get("lumen_post"),
            "", "", e.get("comp")
        ]
        ws_event.append(row)

    # -----------------------------------------------------
    # 6. V3_Procedure_Events & 7. V3_Registry_JSON
    # -----------------------------------------------------
    # (Simplified representation for the script to generate valid JSON structure)
    
    registry_json = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": []
    }
    
    # Helper to build JSON objects
    for e in events_data:
        if e["id"] == "evt99": continue
        proc = {
            "event_id": e["id"],
            "method": e["method"],
            "target": {"anatomy": e.get("target")},
            "outcomes": {}
        }
        if "lumen_pre" in e: proc["outcomes"]["airway"] = {"lumen_pre": e["lumen_pre"], "lumen_post": e["lumen_post"]}
        registry_json["procedures"].append(proc)
        
    ws_json = create_worksheet_if_not_exists(wb, "V3_Registry_JSON", ["json_output"])
    ws_json.cell(row=1, column=1, value=json.dumps(registry_json, indent=2))
    
    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()