import re
import datetime
import json
from openpyxl import Workbook, load_workbook

# ==========================================
# 1. INPUTS
# ==========================================
NOTE_ID = "note_065"
SOURCE_FILE = "note_065.txt"
PROCEDURE_DATE = ""  # Not specified in text
# Full text content from the provided file
NOTE_TEXT = """NOTE_ID:  note_065 SOURCE_FILE: note_065.txt INDICATION FOR OPERATION:  [REDACTED] is a 84 year old-year-old female who presents with lung mass.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: R91.8 Other nonspecific abnormal finding of lung field.
POSTOPERATIVE DIAGNOSIS:  R91.8 Other nonspecific abnormal finding of lung field.
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31652 EBUS sampling 1 or 2 nodes
 
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Linear EBUS 
Disposable Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
 
Successful therapeutic aspiration was performed to clean out the Right Mainstem, Bronchus Intermedius , and Left Mainstem from mucus.
EBUS-Findings
Indications: Diagnostic
Technique:
All lymph node stations were assessed. Only those 5 mm or greater in short axis were sampled.
Lymph node sizing was performed by EBUS and sampling by transbronchial needle aspiration was performed using 25-gauge Needle and 22-gauge Needle.
Lymph Nodes/Sites Inspected: 4R (lower paratracheal) node
11Rs lymph node
 
Overall ROSE Diagnosis: Positive for malignancy
 
No immediate complications
 
 
Lymph Nodes Evaluated:
Site 1: The 11Rs lymph node was => 10 mm on CT and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was sampled.. 6 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Preliminary ROSE Cytology was reported as not adequate and suggestive of blood and bronchial cells. Final results are pending.
Site 2: The 4R (lower paratracheal) node was => 10 mm on CT and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was sampled.. 8 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Preliminary ROSE Cytology was reported as adequate and suggestive of Positive for malignancy. Final results are pending.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
TBNA and TBNB station 11R and 4R
 
IMPRESSION/PLAN: [REDACTED]is a 84 year old-year-old female who presents for bronchoscopy for lung mass.
- f/u pathology"""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# ==========================================
# 2. PARSING & EXTRACTION LOGIC
# ==========================================

# --- 2a. Procedure Flags ---
# Default all to 0
proc_flags = {
    'diagnostic_bronchoscopy': 0, 'bal': 0, 'bronchial_wash': 0, 'brushings': 0,
    'endobronchial_biopsy': 0, 'tbna_conventional': 0, 'linear_ebus': 0, 'radial_ebus': 0,
    'navigational_bronchoscopy': 0, 'transbronchial_biopsy': 0, 'transbronchial_cryobiopsy': 0,
    'therapeutic_aspiration': 0, 'foreign_body_removal': 0, 'airway_dilation': 0,
    'airway_stent': 0, 'thermal_ablation': 0, 'tumor_debulking_non_thermal': 0,
    'cryotherapy': 0, 'blvr': 0, 'peripheral_ablation': 0, 'bronchial_thermoplasty': 0,
    'whole_lung_lavage': 0, 'rigid_bronchoscopy': 0,
    'thoracentesis': 0, 'chest_tube': 0, 'ipc': 0, 'medical_thoracoscopy': 0,
    'pleurodesis': 0, 'pleural_biopsy': 0, 'fibrinolytic_therapy': 0
}

# Logic based on text analysis
txt_lower = NOTE_TEXT.lower()

if "therapeutic aspiration" in txt_lower:
    proc_flags['therapeutic_aspiration'] = 1

if "linear ebus" in txt_lower or "ebus" in txt_lower:
    proc_flags['linear_ebus'] = 1

if "diagnostic" in txt_lower and "bronchoscopy" in txt_lower:
    proc_flags['diagnostic_bronchoscopy'] = 1

# Note mentions "transbronchial biopsies" and "TBNB" in specimens
if "transbronchial biopsies" in txt_lower or "tbnb" in txt_lower:
    proc_flags['transbronchial_biopsy'] = 1
    
# "transbronchial needle aspiration" is explicitly mentioned, but usually subsumed by EBUS if EBUS is present.
# However, registry often wants 'tbna_conventional' if non-guided, or just 'linear_ebus' if guided.
# We will leave tbna_conventional 0 as it is EBUS-guided.

# --- 2b. Span Extraction (Anchor-First) ---
spans = []

def add_span(text, label, normalized_val, schema_field, event_id, context_prefix="", is_negated=False):
    if not text:
        return
    spans.append({
        "span_text": text,
        "label": label,
        "normalized_value": normalized_val,
        "schema_field": schema_field,
        "event_id": event_id,
        "context_prefix": context_prefix,
        "is_negated": is_negated,
        "hydration_status": "needs_hydration"
    })

# Event 1: Therapeutic Aspiration
evt_asp = "evt_01"
add_span("Therapeutic aspiration", "PROC_METHOD", "Therapeutic aspiration", "method", evt_asp)
add_span("Right Mainstem", "ANAT_AIRWAY", "Right Mainstem", "target.anatomy", evt_asp)
add_span("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "target.anatomy", evt_asp)
add_span("Left Mainstem", "ANAT_AIRWAY", "Left Mainstem", "target.anatomy", evt_asp)
add_span("mucus", "OBS_LESION", "mucus", "findings", evt_asp)

# Event 2: EBUS General / Setup (Needles)
# We can link needles to specific sites or a general setup event. 
# Text says: "using 25-gauge Needle and 22-gauge Needle" generally.
# We will create a general EBUS event for the method/device if not strictly per node. 
# However, V3 usually wants events per station. We'll add needles to both station events since we don't know which was which.
# Or better: Text implies both were available/used.
evt_ebus_gen = "evt_ebus_gen" # Conceptual, but let's stick to station-based events
evt_11r = "evt_11r"
evt_4r = "evt_4r"

# 11Rs Event
add_span("11Rs lymph node", "ANAT_LN_STATION", "11Rs", "target.station", evt_11r, context_prefix="Site 1: The ")
add_span("=> 10 mm", "MEAS_SIZE", ">=10mm", "lesion.size_mm", evt_11r, context_prefix="11Rs lymph node was ")
add_span("endobronchial ultrasound guided transbronchial biopsies", "PROC_METHOD", "EBUS-TBNB", "method", evt_11r, context_prefix="Site 1")
add_span("6", "MEAS_COUNT", "6", "specimens.count", evt_11r, context_prefix="The site was sampled.. ")
add_span("not adequate", "OBS_ROSE", "not adequate", "findings", evt_11r, context_prefix="Preliminary ROSE Cytology was reported as ")

# 4R Event
add_span("4R (lower paratracheal) node", "ANAT_LN_STATION", "4R", "target.station", evt_4r, context_prefix="Site 2: The ")
add_span("=> 10 mm", "MEAS_SIZE", ">=10mm", "lesion.size_mm", evt_4r, context_prefix="4R (lower paratracheal) node was ")
add_span("endobronchial ultrasound guided transbronchial biopsies", "PROC_METHOD", "EBUS-TBNB", "method", evt_4r, context_prefix="Site 2")
add_span("8", "MEAS_COUNT", "8", "specimens.count", evt_4r, context_prefix="The site was sampled.. ")
add_span("Positive for malignancy", "OBS_ROSE", "positive", "findings", evt_4r, context_prefix="suggestive of ")

# Needles (General Section) - Associate with both or create a device event?
# Let's add them to both for completeness, as "using 25... and 22..." implies availability for the procedure.
# Using match_index for uniqueness since they appear once in the "Technique" section.
add_span("25-gauge Needle", "DEV_NEEDLE", "25G", "devices", evt_11r)
add_span("22-gauge Needle", "DEV_NEEDLE", "22G", "devices", evt_11r)
add_span("25-gauge Needle", "DEV_NEEDLE", "25G", "devices", evt_4r)
add_span("22-gauge Needle", "DEV_NEEDLE", "22G", "devices", evt_4r)

# Specimens
add_span("TBNA and TBNB", "PROC_METHOD", "TBNA and TBNB", "method", evt_11r, context_prefix="SPECIMEN(S): \n")
add_span("TBNA and TBNB", "PROC_METHOD", "TBNA and TBNB", "method", evt_4r, context_prefix="SPECIMEN(S): \n")

# Complications / Outcomes
evt_outcome = "evt_outcome"
add_span("No immediate complications", "OUTCOME_COMPLICATION", "None", "outcomes.complications", evt_outcome, context_prefix="procedure well.  There were ")

# Diagnosis
add_span("Positive for malignancy", "OBS_ROSE", "Positive", "findings", evt_outcome, context_prefix="Overall ROSE Diagnosis: ")

# --- 2c. Hydration Logic ---
def get_snippet(full_text, start, length=120):
    s = max(0, start - length)
    return full_text[s:start]

hydrated_spans = []
for s in spans:
    txt_val = s["span_text"]
    if not txt_val: 
        continue
    
    # 1. Exact count check
    count = NOTE_TEXT.count(txt_val)
    start_char = -1
    end_char = -1
    method = "ambiguous"

    if count == 1:
        start_char = NOTE_TEXT.find(txt_val)
        method = "hydrated_unique"
    elif count > 1:
        # Context match
        if s["context_prefix"]:
            # Find all occurrences
            occurrences = [m.start() for m in re.finditer(re.escape(txt_val), NOTE_TEXT)]
            best_match = -1
            for occ in occurrences:
                snippet = get_snippet(NOTE_TEXT, occ)
                # Loose check: is context mostly in snippet?
                # Clean strings for comparison
                ctx_clean = s["context_prefix"].strip()[:20] # take last 20 chars of context
                if ctx_clean in snippet:
                    best_match = occ
                    break
            if best_match != -1:
                start_char = best_match
                method = "hydrated_prefix_window"
            else:
                method = f"ambiguous_count={count}"
        else:
             method = f"ambiguous_count={count}"
    else:
        method = "not_found"

    if start_char != -1:
        end_char = start_char + len(txt_val)
    
    # Update span row
    row = s.copy()
    row["start_char"] = start_char if start_char != -1 else ""
    row["end_char"] = end_char if end_char != -1 else ""
    row["span_len"] = len(txt_val)
    row["hydration_status"] = method
    hydrated_spans.append(row)

def generate_excel(template_path: str = TEMPLATE_PATH, output_path: str = OUTPUT_PATH) -> None:
    try:
        wb = load_workbook(template_path)
    except FileNotFoundError:
        wb = Workbook()  # Fallback if template missing (shouldn't happen per instructions)

    def get_sheet(name):
        if name in wb.sheetnames:
            return wb[name]
        return wb.create_sheet(name)

    # 3.1 Note_Text
    ws_text = get_sheet("Note_Text")
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 3.2 Note_Index
    ws_index = get_sheet("Note_Index")
    meta_row = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Todo", ""]
    flag_order = [
        'diagnostic_bronchoscopy', 'bal', 'bronchial_wash', 'brushings', 'endobronchial_biopsy',
        'tbna_conventional', 'linear_ebus', 'radial_ebus', 'navigational_bronchoscopy',
        'transbronchial_biopsy', 'transbronchial_cryobiopsy', 'therapeutic_aspiration',
        'foreign_body_removal', 'airway_dilation', 'airway_stent', 'thermal_ablation',
        'tumor_debulking_non_thermal', 'cryotherapy', 'blvr', 'peripheral_ablation',
        'bronchial_thermoplasty', 'whole_lung_lavage', 'rigid_bronchoscopy',
        'thoracentesis', 'chest_tube', 'ipc', 'medical_thoracoscopy', 'pleurodesis',
        'pleural_biopsy', 'fibrinolytic_therapy'
    ]
    flag_values = [proc_flags[f] for f in flag_order]
    ws_index.append(meta_row + flag_values)

    # 3.3 Span_Annotations & Span_Hydrated
    ws_span_def = get_sheet("Span_Annotations")
    ws_span_hyd = get_sheet("Span_Hydrated")

    span_id_counter = 1
    for s in hydrated_spans:
        sid = f"{NOTE_ID}_s{span_id_counter:03d}"
        span_id_counter += 1

        row_data = [
            SOURCE_FILE, NOTE_ID, sid, "", s.get("context_prefix", ""), s["span_text"], "",
            s["start_char"], s["end_char"], s["span_len"],
            s["label"], s["normalized_value"], s["schema_field"], s["event_id"],
            s["is_negated"], False, "", "", "", s["hydration_status"]
        ]

        ws_span_hyd.append(row_data)
        row_def = list(row_data)
        row_def[7] = ""
        row_def[8] = ""
        ws_span_def.append(row_def)

    # 3.4 Event_Log
    ws_event = get_sheet("Event_Log")

    events_map = {}
    for s in hydrated_spans:
        eid = s["event_id"]
        if eid not in events_map:
            events_map[eid] = {
                "method": set(), "anatomy": set(), "devices": set(), "measurements": set(),
                "findings": set(), "stations": set(), "specimens": set(), "needle": set(),
                "complication": set()
            }

        lbl = s["label"]
        val = s["normalized_value"]

        if lbl == "PROC_METHOD":
            events_map[eid]["method"].add(val)
        if lbl == "ANAT_AIRWAY":
            events_map[eid]["anatomy"].add(val)
        if lbl == "ANAT_LN_STATION":
            events_map[eid]["stations"].add(val)
        if lbl == "DEV_NEEDLE":
            events_map[eid]["needle"].add(val)
        if lbl in ["MEAS_SIZE", "MEAS_COUNT"]:
            events_map[eid]["measurements"].add(f"{lbl}:{val}")
        if lbl in ["OBS_ROSE", "OBS_LESION"]:
            events_map[eid]["findings"].add(val)
        if lbl == "OUTCOME_COMPLICATION":
            events_map[eid]["complication"].add(val)

    for eid, data in events_map.items():
        etype = "Procedure"
        if "outcome" in eid:
            etype = "Outcome"

        row = [
            SOURCE_FILE, NOTE_ID, eid, etype,
            ", ".join(data["method"]),
            ", ".join(data["anatomy"]),
            ", ".join(data["devices"]),
            ", ".join(data["needle"]),
            ", ".join(data["stations"]),
            "",
            ", ".join(data["measurements"]),
            ", ".join(data["specimens"]),
            ", ".join(data["findings"]),
            False, "", "", "", "", "", "", "", "",
            ", ".join(data["complication"])
        ]
        ws_event.append(row)

    # 3.5 V3_Procedure_Events & JSON
    ws_v3 = get_sheet("V3_Procedure_Events")
    registry_procedures = []

    p1 = {
        "event_id": evt_asp,
        "type": "Therapeutic Aspiration",
        "target": {"anatomy_type": "Airway", "location": {"name": "Right Mainstem, Bronchus Intermedius, Left Mainstem"}},
        "method": "Therapeutic aspiration",
        "findings": [{"description": "mucus"}]
    }
    registry_procedures.append(p1)

    p2 = {
        "event_id": evt_11r,
        "type": "EBUS-TBNA",
        "target": {"anatomy_type": "Lymph Node", "station": "11Rs"},
        "lesion": {"size_mm": ">=10"},
        "method": "EBUS-TBNB",
        "devices": [{"type": "Needle", "size": "25G"}, {"type": "Needle", "size": "22G"}],
        "specimens": [{"count": 6, "type": "Biopsy"}],
        "findings": [{"description": "ROSE: not adequate"}]
    }
    registry_procedures.append(p2)

    p3 = {
        "event_id": evt_4r,
        "type": "EBUS-TBNA",
        "target": {"anatomy_type": "Lymph Node", "station": "4R"},
        "lesion": {"size_mm": ">=10"},
        "method": "EBUS-TBNB",
        "devices": [{"type": "Needle", "size": "25G"}, {"type": "Needle", "size": "22G"}],
        "specimens": [{"count": 8, "type": "Biopsy"}],
        "findings": [{"description": "ROSE: Positive"}]
    }
    registry_procedures.append(p3)

    for p in registry_procedures:
        ws_v3.append([
            NOTE_ID, p["event_id"], p["type"],
            "LN" if "station" in p["target"] else "Airway",
            "", "",
            p["target"].get("station", ""),
            "",
            p.get("lesion", {}).get("size_mm", ""),
            p.get("method", ""),
            json.dumps(p.get("devices", [])),
            "",
            json.dumps(p.get("specimens", [])),
            json.dumps(p.get("findings", [])),
            "",
            "", "", "",
            "", "", "", "", ""
        ])

    ws_json = get_sheet("V3_Registry_JSON")
    json_obj = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": registry_procedures
    }
    ws_json.append([json.dumps(json_obj, indent=2)])

    wb.save(output_path)
    print(f"Generated {output_path}")


if __name__ == "__main__":
    generate_excel()
