import openpyxl
from openpyxl.utils import get_column_letter
import json
import os
import re

# =============================================================================
# 1. INPUT DATA (Hardcoded for Note_066)
# =============================================================================

NOTE_ID = "note_066"
SOURCE_FILE = "note_066.txt"
PROCEDURE_DATE = "2026-01-12" # inferred from current date in prompt or left blank. Prompt says "Current time is... 2026". Procedure date is usually date of service. Leaving blank is safer if not explicit.
# However, for completeness, I will leave it blank in the metadata as per instructions "procedure_date... (blank) unless explicit".
PROCEDURE_DATE_META = "" 

# Full text from the prompt
NOTE_TEXT = """NOTE_ID:  note_066 SOURCE_FILE: note_066.txt INDICATION FOR OPERATION:  [REDACTED] is a 77 year old-year-old female who presents with lung nodule.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: R91.8 Other nonspecific abnormal finding of lung field.
POSTOPERATIVE DIAGNOSIS:  R91.8 Other nonspecific abnormal finding of lung field.
PROCEDURE:  
31899 Unlisted Procedure (Trach Change with Mature Tract or Procedure NOS)
31645 Therapeutic aspiration initial episode
31623 Dx bronchoscope/brushing    
31624 Dx bronchoscope/lavage (BAL)    
31628 TBBX single lobe     
31632 TBBX additional lobes  
31629 TBNA single lobe   
31633 TBNA additional lobes   
31626 Fiducial marker placements, single or multiple     
31627 Navigational Bronchoscopy (computer assisted)
77012 Radiology / radiologic guidance for CT guided needle placement (CIOS)
76377 3D rendering with interpretation and reporting of CT, US, Tomo modality (ION Planning Station)
31899NFN BRONCHOSCOPY WITH ENDOBRONCHIAL ULTRASOUND (EBUS) OF 
MEDIASTINAL AND/OR HILAR LYMPH NODES WITHOUT BIOPSY
31654 Radial EBUS for peripheral lesion
 
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)
 
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required a Transbronchial Cryo biopsies, Robotic Navigation to more than one site, Multiple Bronchoalveolar lavages in different locations, Brushings at multiple locations, and Radial EBUS performed at multiple locations.
This resulted in >40% increased work due to Technical difficulty of procedure and Physical and mental effort required.
Apply to: 31623 Dx bronchoscope/brushing    
31624 Dx bronchoscope/lavage (BAL)    
31628 TBBX single lobe     
31629 TBNA single lobe   
31654 Radial EBUS for peripheral lesion.
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Linear EBUS 
Radial EBUS
Ion Robotic Bronchoscope
Disposable Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
 
CT Chest scan was placed on separate planning station to generate 3D rendering of the pathway to target.
The navigational plan was reviewed and verified.  This was then loaded into robotic bronchoscopy platform.
Successful therapeutic aspiration was performed to clean out the Right Mainstem, Bronchus Intermedius , and Left Mainstem from mucus.
Ventilation Parameters:
Mode	RR	TV	PEEP	FiO2	Flow Rate	Pmean
vcv	14	300	12	100	10	15
 
RUL Target:
Robotic navigation bronchoscopy was performed with Ion platform.  Partial registration was used.
Ion robotic catheter was used to engage the Anterior Segment of RUL (RB3).
Target lesion is about 1 cm in diameter.   Under navigational guidance the ion robotic catheter was advanced to 1.0 cm away from the planned target.
Radial EBUS was performed to confirm that the location of the nodule is Eccentric.
The following features were noted: Continuous margin  and Absence of linear-discrete air bronchogram.
Cone Beam CT was performed: 3-D reconstructions were performed on an independent workstation.
Cios Spin system was used for evaluation of nodule location.  Low dose spin was performed to acquire CT imaging.
This was passed on to Ion platform system for reconstruction and nodule location.
The 3D images was interpreted on an independent workstation (Ion).
Using the newly acquired nodule location, the Ion robotic system was adjusted to the new targeted location.
I personally interpreted the cone beam CT and 3-D reconstruction.
Transbronchial needle aspiration was performed with 21G Needle through the extended working channel catheter.  Total 4 samples were collected.
Samples sent for Cytology.
 
Transbronchial biopsy was performed with alligator forceps the extended working channel catheter.
Total 1 samples were collected.  Samples sent for Pathology.
 
Transbronchial cryobiopsy was performed with 1.1mm cryoprobe via the extended working channel catheter.
Freeze time of 6 seconds were used.  Total 6 samples were collected.  Samples sent for Pathology.
Transbronchial brushing was performed with Protected cytology brush the extended working channel catheter.  Total 1 samples were collected.
Samples sent for Microbiology (Cultures/Viral/Fungal).
 
Bronchial alveolar lavage was performed the extended working channel catheter.
Instilled 20 cc of NS, suction returned with 10 cc of NS.  Samples sent for Microbiology (Cultures/Viral/Fungal).
Fiducial marker (0.8mm x 3mm soft tissue gold CIVCO) was loaded with bone wax and placed under fluoroscopy guidance.
Prior to withdraw of the bronchoscope. 
 
ROSE from ION procedure was noted to be:
No evidence of a malignant neoplasm
 
RLL Target:
 
Robotic navigation bronchoscopy was performed with Ion platform.
Partial registration was used.    Ion robotic catheter was used to engage the Superior Segment of RLL (RB6).
Target lesion is about 1 cm in diameter.   Under navigational guidance the ion robotic catheter was advanced to 1.0 cm away from the planned target.
Radial EBUS was performed to confirm that the location of the nodule is Concentric.
The following features were noted: Continuous margin .
 
Cone Beam CT was performed: 3-D reconstructions were performed on an independent workstation.
Cios Spin system was used for evaluation of nodule location.  Low dose spin was performed to acquire CT imaging.
This was passed on to Ion platform system for reconstruction and nodule location.
The 3D images was interpreted on an independent workstation (Ion).
Using the newly acquired nodule location, the Ion robotic system was adjusted to the new targeted location.
I personally interpreted the cone beam CT and 3-D reconstruction.
Transbronchial needle aspiration was performed with 21G Needle through the extended working channel catheter.  Total 4 samples were collected.
Samples sent for Cytology.
 
Transbronchial cryobiopsy was performed with 1.1mm cryoprobe via the extended working channel catheter.
Freeze time of 6 seconds were used.  Total 6 samples were collected.  Samples sent for Pathology.
Transbronchial brushing was performed with Protected cytology brush the extended working channel catheter.  Total 1 samples were collected.
Samples sent for Microbiology (Cultures/Viral/Fungal).
 
Bronchial alveolar lavage was performed the extended working channel catheter.
Instilled 40 cc of NS, suction returned with 15 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
 
Fiducial marker (0.8mm x 3mm soft tissue gold CIVCO) was loaded with bone wax and placed under fluoroscopy guidance.
Prior to withdraw of the bronchoscope. 
 
ROSE from ION procedure was noted to be:
No evidence of a malignant neoplasm
 
Prior to withdrawal of the bronchoscope, inspection demonstrated no evidence of bleeding.
LUL Target:
 
Robotic navigation bronchoscopy was performed with Ion platform.  Partial registration was used.
Ion robotic catheter was used to engage the Apical-Posterior Segment of LUL (LB1/2).
Target lesion is about 1 cm in diameter.   Under navigational guidance the ion robotic catheter was advanced to 1.0 cm away from the planned target.
Radial EBUS was performed to confirm that the location of the nodule is Eccentric.
The following features were noted: Continuous margin  and Absence of linear-discrete air bronchogram.
Cone Beam CT was performed: 3-D reconstructions were performed on an independent workstation.
Cios Spin system was used for evaluation of nodule location.  Low dose spin was performed to acquire CT imaging.
This was passed on to Ion platform system for reconstruction and nodule location.
The 3D images was interpreted on an independent workstation (Ion).
Using the newly acquired nodule location, the Ion robotic system was adjusted to the new targeted location.
I personally interpreted the cone beam CT and 3-D reconstruction.
Transbronchial needle aspiration was performed with 21G Needle through the extended working channel catheter.  Total 4 samples were collected.
Samples sent for Cytology.
 
 
Transbronchial cryobiopsy was performed with 1.1mm cryoprobe via the extended working channel catheter.
Freeze time of 6 seconds were used.  Total 6 samples were collected.  Samples sent for Pathology.
Transbronchial brushing was performed with Protected cytology brush the extended working channel catheter.  Total 1 samples were collected.
Samples sent for Microbiology (Cultures/Viral/Fungal).
 
Bronchial alveolar lavage was performed the extended working channel catheter.
Instilled 20 cc of NS, suction returned with 5 cc of NS.  Samples sent for Microbiology (Cultures/Viral/Fungal).
Fiducial marker (0.8mm x 3mm soft tissue gold CIVCO) was loaded with bone wax and placed under fluoroscopy guidance.
Prior to withdraw of the bronchoscope. 
 
ROSE from ION procedure was noted to be:
No evidence of a malignant neoplasm
 
Prior to withdrawal of the bronchoscope, inspection demonstrated no evidence of bleeding.
EBUS-Findings
Indications: Diagnostic
Technique:
All lymph node stations were assessed. Only those 5 mm or greater in short axis were sampled.
Lymph node sizing was performed by EBUS no sampling was done as none met biopsy criteria.
Lymph Nodes/Sites Inspected: 4R (lower paratracheal) node
4L (lower paratracheal) node
7 (subcarinal) node
10R lymph node
10L lymph node
11Rs lymph node
11Ri lymph node
11L lymph node
 
No immediate complications
 
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
Elastography provided a semi-quantitative classification (Type 1–3), which was used to guide biopsy site selection and sampling strategy.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
RUL TBNA, TBCBX, Brush, BAL, TBBX
RLL TBNA, TBCBX, Brush, BAL
LUL TBNA, TBCBX, Brush, BAL
 
IMPRESSION/PLAN: [REDACTED] is a 77 year old-year-old female who presents for bronchoscopy for lung nodules.
- f/u in clinic for results"""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# Check if file exists locally and override NOTE_TEXT if so (per strict file handling rule)
if os.path.exists(SOURCE_FILE):
    with open(SOURCE_FILE, "r", encoding="utf-8") as f:
        NOTE_TEXT = f.read()

# =============================================================================
# 2. DEFINITIONS & FLAGGING LOGIC
# =============================================================================

# Flags 30
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 1,
    "bronchial_wash": 0,
    "brushings": 1,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 1, # TBNA performed on nodules
    "linear_ebus": 1, # Performed for staging/sizing
    "radial_ebus": 1,
    "navigational_bronchoscopy": 1,
    "transbronchial_biopsy": 1,
    "transbronchial_cryobiopsy": 1,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# =============================================================================
# 3. SPAN EXTRACTION CONFIGURATION
# =============================================================================

# Define spans manually based on analysis of NOTE_TEXT.
# Fields: span_text, context_prefix (optional), match_index (optional, 0-based), label, normalized_value, event_id
# event_id map:
# 1: Therapeutic Aspiration
# 2: RUL Intervention
# 3: RLL Intervention
# 4: LUL Intervention
# 5: EBUS Staging (No biopsy)

SPANS = [
    # --- Event 1: Therapeutic Aspiration ---
    {
        "span_text": "Successful therapeutic aspiration",
        "label": "PROC_METHOD",
        "normalized_value": "Therapeutic aspiration",
        "event_id": "1",
        "context_prefix": None
    },
    {
        "span_text": "Right Mainstem",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Right mainstem bronchus",
        "event_id": "1",
        "context_prefix": "Successful therapeutic aspiration was performed to clean out the "
    },
    {
        "span_text": "Bronchus Intermedius",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Bronchus intermedius",
        "event_id": "1",
        "context_prefix": "Successful therapeutic aspiration was performed to clean out the Right Mainstem, "
    },
    {
        "span_text": "Left Mainstem",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Left mainstem bronchus",
        "event_id": "1",
        "context_prefix": "Bronchus Intermedius , and "
    },
    
    # --- Event 2: RUL Target (RB3) ---
    {
        "span_text": "Robotic navigation bronchoscopy",
        "label": "PROC_METHOD",
        "normalized_value": "Navigational bronchoscopy",
        "event_id": "2",
        "context_prefix": "RUL Target:\n"
    },
    {
        "span_text": "Anterior Segment of RUL (RB3)",
        "label": "ANAT_LUNG_LOC",
        "normalized_value": "RUL Anterior Segment",
        "event_id": "2",
        "context_prefix": "Ion robotic catheter was used to engage the "
    },
    {
        "span_text": "about 1 cm",
        "label": "MEAS_SIZE",
        "normalized_value": "10",
        "event_id": "2",
        "context_prefix": "RUL Target:\nRobotic navigation bronchoscopy was performed with Ion platform" # Broad context
    },
    {
        "span_text": "Eccentric",
        "label": "OBS_LESION",
        "normalized_value": "eccentric",
        "event_id": "2",
        "context_prefix": "Radial EBUS was performed to confirm that the location of the nodule is "
    },
    {
        "span_text": "Transbronchial needle aspiration",
        "label": "PROC_METHOD",
        "normalized_value": "TBNA",
        "event_id": "2",
        "context_prefix": "I personally interpreted the cone beam CT and 3-D reconstruction.\n" # RUL specific context
    },
    {
        "span_text": "21G Needle",
        "label": "DEV_NEEDLE",
        "normalized_value": "21G",
        "event_id": "2",
        "context_prefix": "Transbronchial needle aspiration was performed with "
    },
    {
        "span_text": "Transbronchial biopsy",
        "label": "PROC_METHOD",
        "normalized_value": "Transbronchial biopsy",
        "event_id": "2",
        "context_prefix": "Samples sent for Cytology.\n \n"
    },
    {
        "span_text": "alligator forceps",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "Alligator forceps",
        "event_id": "2",
        "context_prefix": "Transbronchial biopsy was performed with "
    },
    {
        "span_text": "Transbronchial cryobiopsy",
        "label": "PROC_METHOD",
        "normalized_value": "Transbronchial cryobiopsy",
        "event_id": "2",
        "context_prefix": "Samples sent for Pathology.\n \n"
    },
    {
        "span_text": "1.1mm cryoprobe",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "1.1mm Cryoprobe",
        "event_id": "2",
        "context_prefix": "Transbronchial cryobiopsy was performed with "
    },
    {
        "span_text": "Transbronchial brushing",
        "label": "PROC_METHOD",
        "normalized_value": "Bronchial brushing",
        "event_id": "2",
        "context_prefix": "Samples sent for Pathology.\n"
    },
    {
        "span_text": "Protected cytology brush",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "Protected cytology brush",
        "event_id": "2",
        "context_prefix": "Transbronchial brushing was performed with "
    },
    {
        "span_text": "Bronchial alveolar lavage",
        "label": "PROC_METHOD",
        "normalized_value": "BAL",
        "event_id": "2",
        "context_prefix": "Samples sent for Microbiology (Cultures/Viral/Fungal).\n \n"
    },
    {
        "span_text": "Fiducial marker",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "Fiducial marker",
        "event_id": "2",
        "context_prefix": "Samples sent for Microbiology (Cultures/Viral/Fungal).\n"
    },
    {
        "span_text": "No evidence of a malignant neoplasm",
        "label": "OBS_ROSE",
        "normalized_value": "Benign",
        "event_id": "2",
        "context_prefix": "ROSE from ION procedure was noted to be:\n"
    },

    # --- Event 3: RLL Target (RB6) ---
    {
        "span_text": "Robotic navigation bronchoscopy",
        "label": "PROC_METHOD",
        "normalized_value": "Navigational bronchoscopy",
        "event_id": "3",
        "context_prefix": "RLL Target:\n \n"
    },
    {
        "span_text": "Superior Segment of RLL (RB6)",
        "label": "ANAT_LUNG_LOC",
        "normalized_value": "RLL Superior Segment",
        "event_id": "3",
        "context_prefix": "Ion robotic catheter was used to engage the "
    },
    {
        "span_text": "Concentric",
        "label": "OBS_LESION",
        "normalized_value": "concentric",
        "event_id": "3",
        "context_prefix": "Radial EBUS was performed to confirm that the location of the nodule is "
    },
    {
        "span_text": "Transbronchial needle aspiration",
        "label": "PROC_METHOD",
        "normalized_value": "TBNA",
        "event_id": "3",
        "match_index": 1 # Second occurrence
    },
    {
        "span_text": "21G Needle",
        "label": "DEV_NEEDLE",
        "normalized_value": "21G",
        "event_id": "3",
        "match_index": 1
    },
    {
        "span_text": "Transbronchial cryobiopsy",
        "label": "PROC_METHOD",
        "normalized_value": "Transbronchial cryobiopsy",
        "event_id": "3",
        "match_index": 1
    },
    {
        "span_text": "1.1mm cryoprobe",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "1.1mm Cryoprobe",
        "event_id": "3",
        "match_index": 1
    },
    {
        "span_text": "Transbronchial brushing",
        "label": "PROC_METHOD",
        "normalized_value": "Bronchial brushing",
        "event_id": "3",
        "match_index": 1
    },
    {
        "span_text": "Protected cytology brush",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "Protected cytology brush",
        "event_id": "3",
        "match_index": 1
    },
    {
        "span_text": "Bronchial alveolar lavage",
        "label": "PROC_METHOD",
        "normalized_value": "BAL",
        "event_id": "3",
        "match_index": 1
    },
    {
        "span_text": "Fiducial marker",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "Fiducial marker",
        "event_id": "3",
        "match_index": 1
    },
    {
        "span_text": "No evidence of a malignant neoplasm",
        "label": "OBS_ROSE",
        "normalized_value": "Benign",
        "event_id": "3",
        "match_index": 1
    },

    # --- Event 4: LUL Target (LB1/2) ---
    {
        "span_text": "Robotic navigation bronchoscopy",
        "label": "PROC_METHOD",
        "normalized_value": "Navigational bronchoscopy",
        "event_id": "4",
        "context_prefix": "LUL Target:\n \n"
    },
    {
        "span_text": "Apical-Posterior Segment of LUL (LB1/2)",
        "label": "ANAT_LUNG_LOC",
        "normalized_value": "LUL Apical-Posterior Segment",
        "event_id": "4",
        "context_prefix": "Ion robotic catheter was used to engage the "
    },
    {
        "span_text": "Eccentric",
        "label": "OBS_LESION",
        "normalized_value": "eccentric",
        "event_id": "4",
        "context_prefix": "Radial EBUS was performed to confirm that the location of the nodule is ",
        "match_index": 1 # Second "Eccentric" (First was RUL)
    },
    {
        "span_text": "Transbronchial needle aspiration",
        "label": "PROC_METHOD",
        "normalized_value": "TBNA",
        "event_id": "4",
        "match_index": 2
    },
    {
        "span_text": "21G Needle",
        "label": "DEV_NEEDLE",
        "normalized_value": "21G",
        "event_id": "4",
        "match_index": 2
    },
    {
        "span_text": "Transbronchial cryobiopsy",
        "label": "PROC_METHOD",
        "normalized_value": "Transbronchial cryobiopsy",
        "event_id": "4",
        "match_index": 2
    },
    {
        "span_text": "1.1mm cryoprobe",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "1.1mm Cryoprobe",
        "event_id": "4",
        "match_index": 2
    },
    {
        "span_text": "Transbronchial brushing",
        "label": "PROC_METHOD",
        "normalized_value": "Bronchial brushing",
        "event_id": "4",
        "match_index": 2
    },
    {
        "span_text": "Protected cytology brush",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "Protected cytology brush",
        "event_id": "4",
        "match_index": 2
    },
    {
        "span_text": "Bronchial alveolar lavage",
        "label": "PROC_METHOD",
        "normalized_value": "BAL",
        "event_id": "4",
        "match_index": 2
    },
    {
        "span_text": "Fiducial marker",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "Fiducial marker",
        "event_id": "4",
        "match_index": 2
    },
    {
        "span_text": "No evidence of a malignant neoplasm",
        "label": "OBS_ROSE",
        "normalized_value": "Benign",
        "event_id": "4",
        "match_index": 2
    },

    # --- Event 5: EBUS Staging (Findings) ---
    {
        "span_text": "Lymph node sizing was performed by EBUS",
        "label": "PROC_METHOD",
        "normalized_value": "Linear EBUS",
        "event_id": "5",
        "context_prefix": "Technique:\nAll lymph node stations were assessed. Only those 5 mm or greater in short axis were sampled.\n"
    },
    {
        "span_text": "no sampling was done",
        "label": "PROC_ACTION",
        "normalized_value": "No biopsy",
        "event_id": "5",
        "context_prefix": "Lymph node sizing was performed by EBUS "
    },
    {
        "span_text": "4R (lower paratracheal) node",
        "label": "ANAT_LN_STATION",
        "normalized_value": "4R",
        "event_id": "5",
        "context_prefix": "Lymph Nodes/Sites Inspected: "
    },
    {
        "span_text": "4L (lower paratracheal) node",
        "label": "ANAT_LN_STATION",
        "normalized_value": "4L",
        "event_id": "5",
        "context_prefix": "4R (lower paratracheal) node\n"
    },
    {
        "span_text": "7 (subcarinal) node",
        "label": "ANAT_LN_STATION",
        "normalized_value": "7",
        "event_id": "5",
        "context_prefix": "4L (lower paratracheal) node\n"
    },
    {
        "span_text": "10R lymph node",
        "label": "ANAT_LN_STATION",
        "normalized_value": "10R",
        "event_id": "5",
        "context_prefix": "7 (subcarinal) node\n"
    },
    {
        "span_text": "10L lymph node",
        "label": "ANAT_LN_STATION",
        "normalized_value": "10L",
        "event_id": "5",
        "context_prefix": "10R lymph node\n"
    },
    {
        "span_text": "11Rs lymph node",
        "label": "ANAT_LN_STATION",
        "normalized_value": "11Rs",
        "event_id": "5",
        "context_prefix": "10L lymph node\n"
    },
    {
        "span_text": "11Ri lymph node",
        "label": "ANAT_LN_STATION",
        "normalized_value": "11Ri",
        "event_id": "5",
        "context_prefix": "11Rs lymph node\n"
    },
    {
        "span_text": "11L lymph node",
        "label": "ANAT_LN_STATION",
        "normalized_value": "11L",
        "event_id": "5",
        "context_prefix": "11Ri lymph node\n"
    },
    {
        "span_text": "No immediate complications",
        "label": "OUTCOME_COMPLICATION",
        "normalized_value": "None",
        "event_id": "5",
        "context_prefix": "11L lymph node\n \n"
    }
]

# =============================================================================
# 4. HELPER FUNCTIONS
# =============================================================================

def find_offset(full_text, span_text, context_prefix=None, match_index=None):
    """
    Finds the start/end char indices of span_text in full_text.
    Resolution logic:
    1. If context_prefix provided: find occurrence preceded by context (approx 120 chars).
    2. If match_index provided: find Nth occurrence.
    3. If unique: use it.
    4. Else: Ambiguous.
    """
    if not span_text or span_text not in full_text:
        return None, None, "not_found"
    
    # Get all occurrences
    occurrences = [m.start() for m in re.finditer(re.escape(span_text), full_text)]
    
    if len(occurrences) == 1:
        start = occurrences[0]
        return start, start + len(span_text), "hydrated_unique"
    
    # Context matching
    if context_prefix:
        # Normalize context for robust matching (collapse whitespace)
        norm_context = " ".join(context_prefix.split())
        
        for start in occurrences:
            # Check preceding 150 chars (slightly expanded for safety)
            preceding_window = full_text[max(0, start - 150):start]
            norm_window = " ".join(preceding_window.split())
            
            # Simple substring check in normalized window
            if norm_context in norm_window:
                return start, start + len(span_text), "hydrated_prefix_window"
            
            # Fallback: check if context is literally immediately preceding (less likely due to newlines)
            if context_prefix in preceding_window:
                 return start, start + len(span_text), "hydrated_prefix_window"

    # Match index
    if match_index is not None and 0 <= match_index < len(occurrences):
        start = occurrences[match_index]
        return start, start + len(span_text), "hydrated_match_index"

    return None, None, f"ambiguous_count={len(occurrences)}"

# =============================================================================
# 5. WORKBOOK GENERATION
# =============================================================================

def generate_workbook():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        # Create a blank one if template missing (fallback)
        wb = openpyxl.Workbook()

    # --- 1. Note_Text ---
    if "Note_Text" not in wb.sheetnames:
        wb.create_sheet("Note_Text")
    ws_text = wb["Note_Text"]
    if ws_text.max_row == 1:
         ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # --- 2. Note_Index ---
    if "Note_Index" not in wb.sheetnames:
        wb.create_sheet("Note_Index")
    ws_index = wb["Note_Index"]
    
    # Metadata row
    metadata_headers = ["source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes"]
    # Check if headers exist, if not add them (assuming new sheet)
    if ws_index.max_row == 1 and ws_index.cell(1,1).value != "source_file":
         ws_index.append(metadata_headers + list(PROCEDURE_FLAGS.keys()))
    
    # Append data
    flag_values = [PROCEDURE_FLAGS[key] for key in PROCEDURE_FLAGS]
    ws_index.append([SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE_META, "", "", "", ""] + flag_values)

    # --- 3. Span_Annotations (Anchor First) ---
    if "Span_Annotations" not in wb.sheetnames:
        wb.create_sheet("Span_Annotations")
    ws_span = wb["Span_Annotations"]
    
    headers_span = [
        "source_file", "note_id", "span_id", "section_type",
        "context_prefix", "span_text", "match_index",
        "start_char", "end_char", "span_len",
        "label", "normalized_value", "schema_field", "event_id",
        "is_negated", "is_historical", "time_anchor", "reviewer", "comments",
        "hydration_status"
    ]
    if ws_span.max_row == 1 and ws_span.cell(1,1).value != "source_file":
        ws_span.append(headers_span)

    # --- 4. Span_Hydrated ---
    if "Span_Hydrated" not in wb.sheetnames:
        wb.create_sheet("Span_Hydrated")
    ws_hydrated = wb["Span_Hydrated"]
    if ws_hydrated.max_row == 1 and ws_hydrated.cell(1,1).value != "source_file":
        ws_hydrated.append(headers_span)

    # Process spans
    hydrated_rows = []
    
    for idx, span_def in enumerate(SPANS):
        span_id = f"span_{idx+1}"
        span_text = span_def["span_text"]
        context = span_def.get("context_prefix", "")
        match_idx = span_def.get("match_index", None)
        label = span_def["label"]
        norm_val = span_def["normalized_value"]
        event_id = span_def["event_id"]
        
        # Calculate offsets
        start, end, status = find_offset(NOTE_TEXT, span_text, context, match_idx)
        
        # Row for Span_Annotations (Blank offsets)
        row_anchor = [
            SOURCE_FILE, NOTE_ID, span_id, "", # section blank
            context, span_text, match_idx if match_idx is not None else "",
            "", "", f'=LEN(F{ws_span.max_row+1})', # Formula for length
            label, norm_val, "", event_id,
            "FALSE", "FALSE", "", "", "",
            "needs_hydration"
        ]
        ws_span.append(row_anchor)
        
        # Row for Span_Hydrated (Filled offsets)
        row_hydrated = [
            SOURCE_FILE, NOTE_ID, span_id, "",
            context, span_text, match_idx if match_idx is not None else "",
            start if start is not None else "", end if end is not None else "", (end - start) if start is not None else "",
            label, norm_val, "", event_id,
            "FALSE", "FALSE", "", "", "",
            status
        ]
        ws_hydrated.append(row_hydrated)

    # --- 5. Event_Log ---
    if "Event_Log" not in wb.sheetnames:
        wb.create_sheet("Event_Log")
    ws_event = wb["Event_Log"]
    event_headers = [
        "source_file", "note_id", "event_id", "event_type", "method",
        "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements",
        "specimens", "findings", "is_historical", "reviewer", "comments",
        "device_size", "device_material",
        "outcome_airway_lumen_pre", "outcome_airway_lumen_post",
        "outcome_symptoms", "outcome_pleural", "outcome_complication"
    ]
    if ws_event.max_row == 1:
        ws_event.append(event_headers)

    # Define Event Data
    events_data = {
        "1": {
            "type": "Therapeutic",
            "method": "Therapeutic aspiration",
            "anatomy": "Right Mainstem, Bronchus Intermedius, Left Mainstem",
            "findings": "Cleaned out mucus",
            "specimens": ""
        },
        "2": {
            "type": "Diagnostic/Interventional",
            "method": "Navigational bronchoscopy, TBNA, TBBX, Cryo, Brush, BAL",
            "anatomy": "RUL Anterior Segment (RB3)",
            "device": "21G Needle, Alligator forceps, 1.1mm cryoprobe, Cytology brush, Fiducial marker",
            "findings": "Nodule eccentric, no linear air bronchogram, ROSE Benign",
            "specimens": "Cytology, Pathology, Microbiology"
        },
        "3": {
            "type": "Diagnostic/Interventional",
            "method": "Navigational bronchoscopy, TBNA, Cryo, Brush, BAL",
            "anatomy": "RLL Superior Segment (RB6)",
            "device": "21G Needle, 1.1mm cryoprobe, Cytology brush, Fiducial marker",
            "findings": "Nodule concentric, ROSE Benign",
            "specimens": "Cytology, Pathology, Microbiology"
        },
        "4": {
            "type": "Diagnostic/Interventional",
            "method": "Navigational bronchoscopy, TBNA, Cryo, Brush, BAL",
            "anatomy": "LUL Apical-Posterior Segment (LB1/2)",
            "device": "21G Needle, 1.1mm cryoprobe, Cytology brush, Fiducial marker",
            "findings": "Nodule eccentric, ROSE Benign",
            "specimens": "Cytology, Pathology, Microbiology"
        },
        "5": {
            "type": "Staging",
            "method": "Linear EBUS",
            "anatomy": "Mediastinal/Hilar LNs",
            "stations": "4R, 4L, 7, 10R, 10L, 11Rs, 11Ri, 11L",
            "findings": "Sizing only, no biopsy",
            "complication": "None"
        }
    }

    for eid, data in events_data.items():
        ws_event.append([
            SOURCE_FILE, NOTE_ID, eid, data.get("type"), data.get("method"),
            data.get("anatomy"), data.get("device"), data.get("needle_gauge", ""), data.get("stations", ""), "", "",
            data.get("specimens", ""), data.get("findings", ""), "FALSE", "", "",
            "", "", "", "", "", "", data.get("complication", "")
        ])

    # --- 6. V3_Procedure_Events ---
    if "V3_Procedure_Events" not in wb.sheetnames:
        wb.create_sheet("V3_Procedure_Events")
    ws_v3 = wb["V3_Procedure_Events"]
    v3_headers = [
        "note_id", "event_id", "type",
        "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station",
        "lesion.type", "lesion.size_mm",
        "method", "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote",
        "stent.size", "stent.material_or_brand", "catheter.size_fr",
        "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post",
        "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"
    ]
    if ws_v3.max_row == 1:
        ws_v3.append(v3_headers)

    # Populate V3 (Best effort mapping)
    # Event 1
    ws_v3.append([NOTE_ID, "1", "therapeutic_aspiration", "airway", "", "Mainstem/Bronchus Intermedius", "", "", "", "aspiration", "", "", "", "", "Successful therapeutic aspiration was performed...", "", "", "", "", "", "", "", ""])
    
    # Event 2
    ws_v3.append([NOTE_ID, "2", "navigational_bronchoscopy", "lung", "RUL", "Anterior Segment", "", "nodule", "10", "TBNA/TBBX/Cryo/BAL/Brush", json.dumps(["21G Needle", "Alligator forceps", "1.1mm Cryoprobe", "Cytology brush", "Fiducial"]), "", json.dumps(["Cytology", "Pathology", "Microbiology"]), json.dumps(["Eccentric", "ROSE Benign"]), "RUL Target... Ion robotic catheter", "", "", "", "", "", "", "", ""])

    # Event 3
    ws_v3.append([NOTE_ID, "3", "navigational_bronchoscopy", "lung", "RLL", "Superior Segment", "", "nodule", "10", "TBNA/Cryo/BAL/Brush", json.dumps(["21G Needle", "1.1mm Cryoprobe", "Cytology brush", "Fiducial"]), "", json.dumps(["Cytology", "Pathology", "Microbiology"]), json.dumps(["Concentric", "ROSE Benign"]), "RLL Target... Ion robotic catheter", "", "", "", "", "", "", "", ""])

    # Event 4
    ws_v3.append([NOTE_ID, "4", "navigational_bronchoscopy", "lung", "LUL", "Apical-Posterior Segment", "", "nodule", "10", "TBNA/Cryo/BAL/Brush", json.dumps(["21G Needle", "1.1mm Cryoprobe", "Cytology brush", "Fiducial"]), "", json.dumps(["Cytology", "Pathology", "Microbiology"]), json.dumps(["Eccentric", "ROSE Benign"]), "LUL Target... Ion robotic catheter", "", "", "", "", "", "", "", ""])

    # Event 5
    ws_v3.append([NOTE_ID, "5", "ebus_staging", "lymph_node", "", "", "4R,4L,7,10R,10L,11Rs,11Ri,11L", "", "", "linear_ebus_sizing", "", "", "", "no_biopsy", "Lymph node sizing was performed by EBUS...", "", "", "", "", "", "", "", "None"])

    # --- 7. V3_Registry_JSON ---
    if "V3_Registry_JSON" not in wb.sheetnames:
        wb.create_sheet("V3_Registry_JSON")
    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 1:
        ws_json.append(["schema_version", "note_id", "json_output"])
    
    registry_json = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "procedures": [
            {"event_id": "1", "type": "therapeutic_aspiration"},
            {"event_id": "2", "type": "navigational_bronchoscopy", "target": "RUL Anterior"},
            {"event_id": "3", "type": "navigational_bronchoscopy", "target": "RLL Superior"},
            {"event_id": "4", "type": "navigational_bronchoscopy", "target": "LUL Apical-Posterior"},
            {"event_id": "5", "type": "linear_ebus", "action": "staging_only"}
        ],
        "no_immediate_complications": True
    }
    ws_json.append(["v3.0", NOTE_ID, json.dumps(registry_json, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()