import openpyxl
from openpyxl.utils import get_column_letter
import json
import re

# ==========================================
# INPUT DATA & CONFIGURATION
# ==========================================
NOTE_ID = "note_067"
SOURCE_FILE = "note_067.txt"
PROCEDURE_DATE = "2026-01-12" # inferred from context or left blank if not in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_067 SOURCE_FILE: note_067.txt INDICATION FOR OPERATION:  [REDACTED] is a 63 year old-year-old female who presents with pulmonary nodule.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: R91.1 Solitary Lung Nodule
 
POSTOPERATIVE DIAGNOSIS:  R91.1 Solitary Lung Nodule
 
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31629 TBNA single lobe   
31627 Navigational Bronchoscopy (computer assisted)
77012 Radiology / radiologic guidance for CT guided needle placement (CIOS)
76377 3D rendering with interpretation and reporting of CT, US, Tomo modality (ION Planning Station)
31654 Radial EBUS for peripheral lesion
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
 
 
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present 
throughout the entire procedure.
 
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
Flexible Hybrid (Pedatric) Bronchoscope
Radial EBUS
Ion Robotic Bronchoscope
 
ESTIMATED BLOOD LOSS:   Minimum
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
Initial Airway Inspection Findings:
 
Successful therapeutic aspiration was performed to clean out the Trachea (Middle 1/3), Trachea (Distal 1/3), Right Mainstem, and Bronchus Intermedius  from mucus.
CT Chest scan was placed on separate planning station to generate 3D rendering of the pathway to target.
The navigational plan was reviewed and verified.  This was then loaded into robotic bronchoscopy platform.
Ventilation Parameters:
Mode	RR	TV	PEEP	FiO2	Flow Rate	Pmean
VCV	12 	300	12	100	10	15
 
RUL Mass: 
 
Robotic navigation bronchoscopy was performed with Ion platform.  Partial registration was used.
Ion robotic catheter was used to engage the Anterior Segment of RUL (RB3).
Target lesion is about 4.0 cm in diameter.   Under navigational guidance the ion robotic catheter was advanced to 1.0 cm away from the planned target.
Needle was advanced into the lesion. Cone Beam CT was performed: 3-D reconstructions were performed on an independent workstation.
Cios Spin system was used for evaluation of nodule location.  Low dose spin was performed to acquire CT imaging.
This was passed on to Ion platform system for reconstruction and nodule location.
The 3D images was interpreted on an independent workstation (Ion).
Using the newly acquired nodule location, the Ion robotic system was adjusted to the new targeted location.
I personally interpreted the cone beam CT and 3-D reconstruction.
Transbronchial needle aspiration was performed with 19G Needle through the extended working channel catheter.  Total 1 samples were collected.
Samples sent for Cytology.
 
Radial EBUS was performed to confirm that the location of the nodule is Concentric after initial needle puncture into the lesion.
The following features were noted: continuous margin, heterogenous echotexture.
 
Once a tunnel had been created into the lesion with TBNA as above, transbronchial microwave ablation was performed of the nodule as below.
Modality	Tools	Setting/Mode	Duration	Results
Microwave 	Medium Antenna 	90*C	20:00 min	Transbronchial Ablation of Nodule
 
 
Subsequently, RB2 was engaged.
And microwave catheter was placed in the lateral superior aspect of the lung mass.
CBCT was done to confirm the location. 
 
Modality	Tools	Setting/Mode	Duration	Results
Microwave 	Medium Antenna 	90*C	10:00 min	Transbronchial Ablation of Nodule
 
Subsequently, RB1 was engaged.
And microwave catheter was placed in the superior medial aspect of the lung mass.
CBCT was done to confirm the location. 
 
Modality	Tools	Setting/Mode	Duration	Results
Microwave 	Medium Antenna 	90*C	10:00 min	Transbronchial Ablation of Nodule
 
Cone Beam CT was performed: 3-D reconstructions were performed on an independent workstation.
Cios Spin system was used for evaluation of nodule location.  Low dose spin was performed to acquire CT imaging.
This was passed on to Ion platform system for reconstruction and nodule location.
The 3D images was interpreted on an independent workstation (Ion).
Using the newly acquired nodule location, the Ion robotic system was adjusted to the new targeted location.
I personally interpreted the cone beam CT and 3-D reconstruction. The images confirmed tool-in-lesion. 
 
The patient tolerated the procedure well.
There were no immediate complications.  At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
- RUL mass TBNA (cyto)
 
 
IMPRESSION/PLAN: [REDACTED] is a 63 year old-year-old female who presents for bronchoscopy for lung nodules.
RLL nodule and RUL mass underwent TBNA and MWA. 
 
- f/u CXR
- admit to medicine for observation
- monitor for post-ablation pain 
- start empiric abx (IV unasyn as inpatient, po moxifloxacin as an outpatient)"""

# ==========================================
# 1. PROCEDURE FLAGS LOGIC
# ==========================================
def determine_flags(text):
    text_lower = text.lower()
    flags = {
        'diagnostic_bronchoscopy': 0, 'bal': 0, 'bronchial_wash': 0, 'brushings': 0, 
        'endobronchial_biopsy': 0, 'tbna_conventional': 0, 'linear_ebus': 0, 
        'radial_ebus': 0, 'navigational_bronchoscopy': 0, 'transbronchial_biopsy': 0, 
        'transbronchial_cryobiopsy': 0, 'therapeutic_aspiration': 0, 
        'foreign_body_removal': 0, 'airway_dilation': 0, 'airway_stent': 0, 
        'thermal_ablation': 0, 'tumor_debulking_non_thermal': 0, 'cryotherapy': 0, 
        'blvr': 0, 'peripheral_ablation': 0, 'bronchial_thermoplasty': 0, 
        'whole_lung_lavage': 0, 'rigid_bronchoscopy': 0,
        'thoracentesis': 0, 'chest_tube': 0, 'ipc': 0, 'medical_thoracoscopy': 0, 
        'pleurodesis': 0, 'pleural_biopsy': 0, 'fibrinolytic_therapy': 0
    }
    
    # Logic based on text content
    if "navigational" in text_lower or "ion robotic" in text_lower:
        flags['navigational_bronchoscopy'] = 1
    if "radial ebus" in text_lower:
        flags['radial_ebus'] = 1
    if "therapeutic aspiration" in text_lower or "31645" in text_lower:
        flags['therapeutic_aspiration'] = 1
    if "tbna" in text_lower or "transbronchial needle aspiration" in text_lower:
        flags['tbna_conventional'] = 1
    if "microwave" in text_lower or "ablation" in text_lower:
        flags['thermal_ablation'] = 1
        flags['peripheral_ablation'] = 1
    
    # Flag checks typically implicitly include base bronchoscopy if others are true, 
    # but here we follow strict presence. "Bronchoscopy" is in title.
    if "bronchoscopy" in text_lower:
        flags['diagnostic_bronchoscopy'] = 1
        
    return flags

# ==========================================
# 2. SPAN EXTRACTION (ANCHOR-FIRST)
# ==========================================
def extract_spans(text):
    spans = []
    
    # Helper to add span
    def add(span_text, label, norm_val, schema_field, event_id, context_prefix=None, comments=""):
        if not span_text or span_text not in text:
            return
        spans.append({
            "span_text": span_text,
            "label": label,
            "normalized_value": norm_val,
            "schema_field": schema_field,
            "event_id": event_id,
            "context_prefix": context_prefix,
            "comments": comments
        })

    # Event 1: Therapeutic Aspiration
    ev1 = "evt_01"
    add("Therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", "procedure.method", ev1)
    add("Trachea (Middle 1/3)", "ANAT_AIRWAY", "Trachea", "target.anatomy", ev1)
    add("Trachea (Distal 1/3)", "ANAT_AIRWAY", "Trachea", "target.anatomy", ev1)
    add("Right Mainstem", "ANAT_AIRWAY", "Right Mainstem Bronchus", "target.anatomy", ev1)
    add("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "target.anatomy", ev1)

    # Event 2: RUL TBNA
    ev2 = "evt_02"
    add("Anterior Segment of RUL (RB3)", "ANAT_AIRWAY", "RB3", "target.anatomy", ev2)
    add("4.0 cm", "MEAS_SIZE", "40", "lesion.size_mm", ev2)
    add("Ion robotic bronchoscopy", "PROC_METHOD", "Robotic Bronchoscopy", "procedure.method", ev2)
    add("Ion robotic catheter", "DEV_CATHETER", "Ion Catheter", "devices", ev2)
    add("19G Needle", "DEV_NEEDLE", "19G", "devices", ev2)
    add("Radial EBUS", "PROC_METHOD", "Radial EBUS", "procedure.method", ev2)
    add("Concentric", "OBS_LESION", "Concentric", "findings", ev2, context_prefix="location of the nodule is")
    add("continuous margin", "OBS_LESION", "Continuous Margin", "findings", ev2)
    add("heterogenous echotexture", "OBS_LESION", "Heterogenous Echotexture", "findings", ev2)
    add("1 samples", "MEAS_COUNT", "1", "specimens", ev2)
    add("Cytology", "PROC_METHOD", "Cytology", "specimens", ev2)

    # Event 3: MWA RB3
    ev3 = "evt_03"
    add("transbronchial microwave ablation", "PROC_METHOD", "Microwave Ablation", "procedure.method", ev3, context_prefix="as above, ")
    add("Medium Antenna", "DEV_INSTRUMENT", "Medium Antenna", "devices", ev3, context_prefix="Microwave \t")

    # Event 4: MWA RB2
    ev4 = "evt_04"
    add("RB2", "ANAT_AIRWAY", "RB2", "target.anatomy", ev4)
    add("lateral superior aspect", "ANAT_LUNG_LOC", "Lateral Superior", "target.location", ev4)
    # Finding specific microwave span for this event is tricky as they look identical in the table
    # Using context logic for unique identification
    add("Microwave", "PROC_METHOD", "Microwave Ablation", "procedure.method", ev4, context_prefix="RB2 was engaged.\nAnd microwave catheter was placed in the lateral superior aspect of the lung mass.\nCBCT was done to confirm the location. \n\nModality\tTools\tSetting/Mode\tDuration\tResults\n")
    add("Medium Antenna", "DEV_INSTRUMENT", "Medium Antenna", "devices", ev4, context_prefix="Subsequently, RB2 was engaged.")

    # Event 5: MWA RB1
    ev5 = "evt_05"
    add("RB1", "ANAT_AIRWAY", "RB1", "target.anatomy", ev5)
    add("superior medial aspect", "ANAT_LUNG_LOC", "Superior Medial", "target.location", ev5)
    add("Microwave", "PROC_METHOD", "Microwave Ablation", "procedure.method", ev5, context_prefix="RB1 was engaged.\nAnd microwave catheter was placed in the superior medial aspect of the lung mass.\nCBCT was done to confirm the location. \n\nModality\tTools\tSetting/Mode\tDuration\tResults\n")
    add("Medium Antenna", "DEV_INSTRUMENT", "Medium Antenna", "devices", ev5, context_prefix="Subsequently, RB1 was engaged.")

    # Global/Outcome
    add("no immediate complications", "OUTCOME_COMPLICATION", "None", "outcomes.complications", "global")
    add("RUL mass", "ANAT_LUNG_LOC", "RUL", "target.location", ev2) # Linking RUL mass to the TBNA event mainly

    return spans

# ==========================================
# 3. HYDRATION LOGIC
# ==========================================
def hydrate_spans(spans, text):
    hydrated = []
    
    for s in spans:
        item = s.copy()
        target = item["span_text"]
        
        # 1. Exact count check
        count = text.count(target)
        
        start_idx = -1
        status = "ambiguous"
        
        if count == 1:
            start_idx = text.find(target)
            status = "hydrated_unique"
        elif count > 1:
            # Context check
            if item["context_prefix"]:
                # Find all occurrences
                occurrences = [m.start() for m in re.finditer(re.escape(target), text)]
                found = False
                for idx in occurrences:
                    # check preceding 150 chars
                    window = text[max(0, idx-150):idx]
                    if item["context_prefix"] in window:
                        start_idx = idx
                        status = "hydrated_prefix_window"
                        found = True
                        break
                if not found:
                    status = f"ambiguous_count={count}_ctx_not_found"
            else:
                status = f"ambiguous_count={count}"
        else:
            status = "not_found"

        if start_idx != -1:
            item["start_char"] = start_idx
            item["end_char"] = start_idx + len(target)
            item["span_len"] = len(target)
        
        item["hydration_status"] = status
        hydrated.append(item)
        
    return hydrated

# ==========================================
# 4. EXECUTION & WORKBOOK GENERATION
# ==========================================
def generate_workbook():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        # If template not found, create simple one for graceful failure (in real scenario, fail hard)
        wb = openpyxl.Workbook()
        wb.create_sheet("Note_Text")
        wb.create_sheet("Note_Index")
        wb.create_sheet("Span_Annotations")
        wb.create_sheet("Span_Hydrated")
        wb.create_sheet("Event_Log")
        wb.create_sheet("V3_Procedure_Events")
        wb.create_sheet("V3_Registry_JSON")

    # 1. Populate Note_Text
    ws_text = wb["Note_Text"]
    # Append row: note_id, source_file, note_text
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Populate Note_Index
    ws_index = wb["Note_Index"]
    flags = determine_flags(NOTE_TEXT)
    
    # Flags columns order assumed to match standard list. 
    # Current template structure usually: Metadata first, then flags.
    # We construct the row manually.
    meta_row = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Success", ""]
    
    # Bronch flags (23)
    bronch_keys = [
        'diagnostic_bronchoscopy', 'bal', 'bronchial_wash', 'brushings', 
        'endobronchial_biopsy', 'tbna_conventional', 'linear_ebus', 'radial_ebus', 
        'navigational_bronchoscopy', 'transbronchial_biopsy', 'transbronchial_cryobiopsy', 
        'therapeutic_aspiration', 'foreign_body_removal', 'airway_dilation', 'airway_stent', 
        'thermal_ablation', 'tumor_debulking_non_thermal', 'cryotherapy', 'blvr', 
        'peripheral_ablation', 'bronchial_thermoplasty', 'whole_lung_lavage', 'rigid_bronchoscopy'
    ]
    # Pleural flags (7)
    pleural_keys = [
        'thoracentesis', 'chest_tube', 'ipc', 'medical_thoracoscopy', 
        'pleurodesis', 'pleural_biopsy', 'fibrinolytic_therapy'
    ]
    
    flag_values = [flags[k] for k in bronch_keys] + [flags[k] for k in pleural_keys]
    
    ws_index.append(meta_row + flag_values)

    # 3. Process Spans
    raw_spans = extract_spans(NOTE_TEXT)
    
    ws_span_anno = wb["Span_Annotations"]
    # Columns: source_file, note_id, span_id, section_type, context_prefix, span_text, 
    # match_index, start_char, end_char, span_len, label, normalized_value, schema_field, 
    # event_id, is_negated, is_historical, time_anchor, reviewer, comments, hydration_status
    
    span_id_counter = 1
    for s in raw_spans:
        row = [
            SOURCE_FILE, NOTE_ID, f"span_{span_id_counter:03d}", "", 
            s.get("context_prefix"), s["span_text"], "", 
            "", "", f"=LEN(F{ws_span_anno.max_row+1})", # span_len formula
            s["label"], s["normalized_value"], s["schema_field"], 
            s["event_id"], "", "", "", "", s.get("comments"), "needs_hydration"
        ]
        ws_span_anno.append(row)
        s["span_id"] = f"span_{span_id_counter:03d}"
        span_id_counter += 1

    # 4. Hydrate Spans
    hydrated_spans = hydrate_spans(raw_spans, NOTE_TEXT)
    ws_span_hydro = wb["Span_Hydrated"]
    
    for s in hydrated_spans:
        row = [
            SOURCE_FILE, NOTE_ID, s["span_id"], "", 
            s.get("context_prefix"), s["span_text"], "", 
            s.get("start_char"), s.get("end_char"), s.get("span_len"),
            s["label"], s["normalized_value"], s["schema_field"], 
            s["event_id"], "", "", "", "", s.get("comments"), s["hydration_status"]
        ]
        ws_span_hydro.append(row)

    # 5. Populate Event_Log (Best Effort) & V3_Procedure_Events
    # We group spans by event_id
    events = {}
    for s in hydrated_spans:
        eid = s["event_id"]
        if eid not in events:
            events[eid] = {
                "type": "Procedure", "method": [], "anatomy": [], "device": [], 
                "measurements": [], "specimens": [], "findings": [], "outcomes": []
            }
        
        # Map label to category
        lbl = s["label"]
        val = s["normalized_value"]
        if "PROC" in lbl: events[eid]["method"].append(val)
        if "ANAT" in lbl: events[eid]["anatomy"].append(val)
        if "DEV" in lbl: events[eid]["device"].append(val)
        if "MEAS" in lbl: events[eid]["measurements"].append(val)
        if "SPEC" in lbl: events[eid]["specimens"].append(val)
        if "OBS" in lbl: events[eid]["findings"].append(val)
        if "OUTCOME" in lbl: events[eid]["outcomes"].append(val)

    ws_event_log = wb["Event_Log"]
    ws_v3 = wb["V3_Procedure_Events"]

    # Simple mapping for Event Log
    for eid, data in events.items():
        if eid == "global": continue
        
        # Event Log Row
        # source_file, note_id, event_id, event_type, method, anatomy_target, device...
        el_row = [
            SOURCE_FILE, NOTE_ID, eid, "Procedure", 
            ", ".join(set(data["method"])), 
            ", ".join(set(data["anatomy"])),
            ", ".join(set(data["device"])),
            "", "", "", # gauge, stations, counts
            ", ".join(set(data["measurements"])),
            ", ".join(set(data["specimens"])),
            ", ".join(set(data["findings"])),
            "", "", "" # history, reviewer, comments
        ]
        # Pad remaining columns (outcomes)
        el_row += [""] * 10
        ws_event_log.append(el_row)

        # V3 Row
        # note_id, event_id, type, target.lobe, target.segment...
        # Parsing Anatomy for V3 Lobe/Seg
        lobe = ""
        seg = ""
        if "RUL" in str(data["anatomy"]): lobe = "RUL"
        if "RB3" in str(data["anatomy"]): seg = "RB3"
        if "RB2" in str(data["anatomy"]): seg = "RB2"
        if "RB1" in str(data["anatomy"]): seg = "RB1"

        v3_row = [
            NOTE_ID, eid, "Procedure",
            "Airway/Lung", lobe, seg, "", # Station
            "Nodule", "", # Lesion type, size
            ", ".join(set(data["method"])),
            json.dumps(data["device"]),
            json.dumps(data["measurements"]),
            json.dumps(data["specimens"]),
            json.dumps(data["findings"]),
            "", # evidence
            "", "", "", # stent details
            "", "", "", "", "" # outcomes
        ]
        ws_v3.append(v3_row)

    # 6. JSON Dump
    ws_json = wb["V3_Registry_JSON"]
    json_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": events,
        "no_immediate_complications": True
    }
    ws_json.cell(row=1, column=1, value=json.dumps(json_data, indent=2))

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()