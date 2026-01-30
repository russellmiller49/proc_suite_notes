import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import json
import datetime
import re
import sys

# =============================================================================
# CONSTANTS & CONFIGURATION
# =============================================================================

NOTE_ID = "note_068"
SOURCE_FILE = "note_068.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# Reconstructed text from the provided source content (stripping [source] tags)
NOTE_TEXT = """NOTE_ID:  note_068 SOURCE_FILE: note_068.txt INDICATION FOR OPERATION:  [REDACTED] is a 65 year old-year-old male who presents with lung mass.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
 
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
 
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
 
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
 
Stent in RML orifice in good position.
Successful therapeutic aspiration was performed to clean out the Right Mainstem, Bronchus Intermedius , and Left Mainstem from mucus.
Endobronchial obstruction at Trachea (Distal 1/3), Right Mainstem, and Left Mainstem was treated with the following modalities:
 
Modality\tTools\tSetting/Mode\tDuration\tResults
Electrocautery\t \t \t \t 
APC\t2.3 mm\tstraight\tPulse effect 4\tablated
Laser\t \t \t \t 
CoreCath\t \t \t \t 
Cryoprobe\t \t \t \t 
 
Prior to treatment, affected airway was note to be 90% patent.
After treatment, the airway was 100% patent. 
 
 
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
none
 
IMPRESSION/PLAN: [REDACTED] is a 65 year old-year-old male who presents for bronchoscopy for stent evaluation.
- repeat bronchoscopy in 4-6 weeks
- f/u in clinic with consideration to perform PDT"""

# =============================================================================
# DATA DEFINITIONS
# =============================================================================

# Record-level flags (0/1)
PROCEDURE_FLAGS = {
    # Bronchoscopy (23)
    "diagnostic_bronchoscopy": 1, # Implied by flexible bronch + inspection
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1, # Explicit 31645
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0, # Existing stent mentioned, no placement/removal
    "thermal_ablation": 1, # APC used
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural (7)
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# Anchor-First Spans
# Structure: (text, label, normalized, schema_field, event_id, context_prefix, is_negated, is_historical)
SPANS_DATA = [
    # Metadata/Diagnosis
    ("J98.09 Other diseases of bronchus", "OBS_LESION", "Other diseases of bronchus", "lesion.type", "evt_dx", "PREOPERATIVE DIAGNOSIS: ", False, False),
    ("Flexible Therapeutic Bronchoscope", "DEV_INSTRUMENT", "Flexible Bronchoscope", "device.type", "evt_proc_gen", "INSTRUMENT : \n", False, False),
    
    # Event 1: Stent Inspection
    ("Stent", "DEV_STENT", "Airway Stent", "stent.type", "evt_stent_inspect", "Initial Airway Inspection Findings:\n \n", False, True),
    ("RML orifice", "ANAT_AIRWAY", "RML", "target.anatomy_type", "evt_stent_inspect", "Stent in ", False, True),
    ("in good position", "OBS_LESION", "Good Position", "findings.description", "evt_stent_inspect", "RML orifice ", False, True),
    
    # Event 2: Therapeutic Aspiration
    ("therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", "method", "evt_asp", "Successful ", False, False),
    ("Right Mainstem", "ANAT_AIRWAY", "RMS", "target.anatomy_type", "evt_asp", "clean out the ", False, False),
    ("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "target.anatomy_type", "evt_asp", "Right Mainstem, ", False, False),
    ("Left Mainstem", "ANAT_AIRWAY", "LMS", "target.anatomy_type", "evt_asp", "Bronchus Intermedius , and ", False, False),
    ("mucus", "OBS_LESION", "Mucus", "findings.description", "evt_asp", "Left Mainstem from ", False, False),
    
    # Event 3: Thermal Ablation (APC)
    ("Endobronchial obstruction", "OBS_LESION", "Obstruction", "findings.description", "evt_apc", "\n", False, False),
    ("Trachea (Distal 1/3)", "ANAT_AIRWAY", "Trachea", "target.anatomy_type", "evt_apc", "Endobronchial obstruction at ", False, False),
    ("Right Mainstem", "ANAT_AIRWAY", "RMS", "target.anatomy_type", "evt_apc", "Trachea (Distal 1/3), ", False, False),
    ("Left Mainstem", "ANAT_AIRWAY", "LMS", "target.anatomy_type", "evt_apc", "Right Mainstem, and ", False, False),
    ("treated", "PROC_ACTION", "Treated", "procedure.action", "evt_apc", "was ", False, False),
    ("APC", "PROC_METHOD", "APC", "method", "evt_apc", "Electrocautery\t \t \t \t \n", False, False),
    ("2.3 mm", "DEV_CATHETER_SIZE", "2.3 mm", "catheter.size", "evt_apc", "APC\t", False, False),
    ("ablated", "PROC_ACTION", "Ablation", "procedure.action", "evt_apc", "Pulse effect 4\t", False, False),
    ("90% patent", "OUTCOME_AIRWAY_LUMEN_PRE", "90% patent", "outcomes.airway.lumen_pre", "evt_apc", "note to be ", False, False),
    ("100% patent", "OUTCOME_AIRWAY_LUMEN_POST", "100% patent", "outcomes.airway.lumen_post", "evt_apc", "airway was ", False, False),
    
    # Event 4: Complications
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "outcomes.complications", "evt_comp", "well.  There were ", True, False),
]

# Event Log Data
EVENT_LOG_ENTRIES = [
    {
        "event_id": "evt_stent_inspect",
        "event_type": "Inspection",
        "anatomy_target": "RML",
        "device": "Stent",
        "findings": "Stent in good position",
        "is_historical": "Yes"
    },
    {
        "event_id": "evt_asp",
        "event_type": "Therapeutic Aspiration",
        "method": "Suction",
        "anatomy_target": "RMS, Bronchus Intermedius, LMS",
        "findings": "Mucus cleaned out"
    },
    {
        "event_id": "evt_apc",
        "event_type": "Thermal Ablation",
        "method": "APC",
        "device": "APC Probe",
        "device_size": "2.3 mm",
        "anatomy_target": "Trachea (Distal 1/3), RMS, LMS",
        "outcome_airway_lumen_pre": "90% patent",
        "outcome_airway_lumen_post": "100% patent"
    },
    {
        "event_id": "evt_comp",
        "event_type": "Complication",
        "outcome_complication": "No immediate complications"
    }
]

# V3 Procedure Events
V3_EVENTS = [
    {
        "note_id": NOTE_ID,
        "event_id": "evt_stent_inspect",
        "type": "Inspection",
        "target": {"anatomy_type": "RML", "location": {"lobe": "RML"}},
        "stent": {"status": "existing"},
        "findings_json": ["Stent in good position"],
        "evidence_quote": "Stent in RML orifice in good position"
    },
    {
        "note_id": NOTE_ID,
        "event_id": "evt_asp",
        "type": "Therapeutic Aspiration",
        "method": "Suction",
        "target": {"anatomy_type": "RMS, BI, LMS"},
        "findings_json": ["Mucus"],
        "evidence_quote": "Successful therapeutic aspiration was performed to clean out the Right Mainstem, Bronchus Intermedius , and Left Mainstem from mucus."
    },
    {
        "note_id": NOTE_ID,
        "event_id": "evt_apc",
        "type": "Thermal Ablation",
        "method": "APC",
        "target": {"anatomy_type": "Trachea, RMS, LMS"},
        "devices_json": [{"type": "APC Probe", "size": "2.3 mm"}],
        "outcomes": {
            "airway": {
                "lumen_pre": "90% patent",
                "lumen_post": "100% patent"
            }
        },
        "evidence_quote": "Endobronchial obstruction... treated with... APC 2.3 mm... ablated"
    },
    {
        "note_id": NOTE_ID,
        "event_id": "evt_comp",
        "type": "Outcome",
        "outcomes": {
            "complications": "None"
        },
        "evidence_quote": "There were no immediate complications."
    }
]

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def clean_text(text):
    return text.strip()

def normalize_value(val):
    if not val: return ""
    return str(val).strip()

def hydrate_offsets(text, span_text, context_prefix=None, match_index=None):
    """
    Calculates start_char, end_char based on text, span, context, or index.
    """
    if not span_text or span_text not in text:
        return None, None, "not_found"
    
    # 1. Context Match
    if context_prefix:
        # Find all occurrences of span
        starts = [m.start() for m in re.finditer(re.escape(span_text), text)]
        best_start = -1
        
        for start in starts:
            # Look at preceding window
            window_start = max(0, start - 150)
            preceding_text = text[window_start:start]
            if context_prefix in preceding_text:
                best_start = start
                break
        
        if best_start != -1:
            return best_start, best_start + len(span_text), "hydrated_prefix_window"
            
    # 2. Unique Match
    count = text.count(span_text)
    if count == 1:
        start = text.find(span_text)
        return start, start + len(span_text), "hydrated_unique"
    
    # 3. Match Index
    if match_index is not None and isinstance(match_index, int):
        starts = [m.start() for m in re.finditer(re.escape(span_text), text)]
        if 0 <= match_index < len(starts):
            s = starts[match_index]
            return s, s + len(span_text), "hydrated_match_index"
    
    # Default: Ambiguous
    return None, None, f"ambiguous_count={count}"

# =============================================================================
# MAIN LOGIC
# =============================================================================

def main():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        print(f"Error: Template {TEMPLATE_PATH} not found.")
        sys.exit(1)

    # 1. Populate Note_Text
    # -------------------------------------------------------------------------
    if "Note_Text" not in wb.sheetnames:
        wb.create_sheet("Note_Text")
    ws_text = wb["Note_Text"]
    # Assuming header exists: note_id, source_file, note_text
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Populate Note_Index
    # -------------------------------------------------------------------------
    if "Note_Index" not in wb.sheetnames:
        wb.create_sheet("Note_Index")
    ws_index = wb["Note_Index"]
    
    # Headers for reference (not writing them, assuming template has them)
    # source_file, note_id, encounter_id, procedure_date, site, reviewer, status, free_text_notes, [flags...]
    
    index_row = [
        SOURCE_FILE, NOTE_ID, "", "", "", "", "Pending", ""
    ]
    
    # Append flags in order
    flag_keys = list(PROCEDURE_FLAGS.keys())
    for k in flag_keys:
        index_row.append(PROCEDURE_FLAGS[k])
        
    ws_index.append(index_row)

    # 3. Populate Span_Annotations (Anchor First) & 4. Span_Hydrated
    # -------------------------------------------------------------------------
    if "Span_Annotations" not in wb.sheetnames:
        wb.create_sheet("Span_Annotations")
    ws_anno = wb["Span_Annotations"]
    
    if "Span_Hydrated" not in wb.sheetnames:
        wb.create_sheet("Span_Hydrated")
    ws_hydra = wb["Span_Hydrated"]
    
    # Write spans
    span_id_counter = 1
    
    for item in SPANS_DATA:
        # Unpack tuple
        # (text, label, normalized, schema_field, event_id, context_prefix, is_negated, is_historical)
        span_text = item[0]
        label = item[1]
        norm_val = item[2]
        schema_f = item[3]
        evt_id = item[4]
        ctx_pre = item[5]
        is_neg = item[6]
        is_hist = item[7]
        
        # Determine unique span ID
        sid = f"{NOTE_ID}_s{span_id_counter:03d}"
        span_id_counter += 1
        
        # Anchor Row (offsets blank)
        # source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, 
        # start_char, end_char, span_len, label, normalized_value, schema_field, event_id, 
        # is_negated, is_historical, time_anchor, reviewer, comments, hydration_status
        
        # Calculate len formula
        len_formula = f'=LEN(F{ws_anno.max_row + 1})' # Assuming span_text is col F (index 6)
        
        row_anno = [
            SOURCE_FILE, NOTE_ID, sid, "", ctx_pre, span_text, "", 
            "", "", len_formula, label, norm_val, schema_f, evt_id,
            is_neg, is_hist, "", "", "", "needs_hydration"
        ]
        ws_anno.append(row_anno)
        
        # Hydrated Row
        start, end, status = hydrate_offsets(NOTE_TEXT, span_text, context_prefix=ctx_pre)
        
        row_hydra = [
            SOURCE_FILE, NOTE_ID, sid, "", ctx_pre, span_text, "", 
            start, end, (end - start) if start is not None else "", label, norm_val, schema_f, evt_id,
            is_neg, is_hist, "", "", "", status
        ]
        ws_hydra.append(row_hydra)

    # 5. Populate Event_Log
    # -------------------------------------------------------------------------
    if "Event_Log" not in wb.sheetnames:
        wb.create_sheet("Event_Log")
    ws_event = wb["Event_Log"]
    
    for evt in EVENT_LOG_ENTRIES:
        # source_file, note_id, event_id, event_type, method, anatomy_target, device, needle_gauge, 
        # stations, counts, measurements, specimens, findings, is_historical, reviewer, comments,
        # device_size, device_material, outcome_airway_lumen_pre, outcome_airway_lumen_post, 
        # outcome_symptoms, outcome_pleural, outcome_complication
        
        row_evt = [
            SOURCE_FILE, NOTE_ID, evt.get("event_id", ""), evt.get("event_type", ""), evt.get("method", ""),
            evt.get("anatomy_target", ""), evt.get("device", ""), "",
            "", "", "", "", evt.get("findings", ""), evt.get("is_historical", False), "", "",
            evt.get("device_size", ""), evt.get("device_material", ""),
            evt.get("outcome_airway_lumen_pre", ""), evt.get("outcome_airway_lumen_post", ""),
            evt.get("outcome_symptoms", ""), evt.get("outcome_pleural", ""), evt.get("outcome_complication", "")
        ]
        ws_event.append(row_evt)

    # 6. Populate V3_Procedure_Events
    # -------------------------------------------------------------------------
    if "V3_Procedure_Events" not in wb.sheetnames:
        wb.create_sheet("V3_Procedure_Events")
    ws_v3 = wb["V3_Procedure_Events"]
    
    for v3e in V3_EVENTS:
        # Flatten target
        tgt = v3e.get("target", {})
        loc = tgt.get("location", {})
        
        # Flatten outputs
        out = v3e.get("outcomes", {})
        out_aw = out.get("airway", {})
        
        row_v3 = [
            NOTE_ID, v3e.get("event_id"), v3e.get("type"),
            tgt.get("anatomy_type"), loc.get("lobe"), loc.get("segment"), tgt.get("station"),
            "", "", # lesion type/size
            v3e.get("method"), 
            json.dumps(v3e.get("devices_json", [])),
            "", # measurements
            "", # specimens
            json.dumps(v3e.get("findings_json", [])),
            v3e.get("evidence_quote"),
            v3e.get("stent", {}).get("size", ""), v3e.get("stent", {}).get("material_or_brand", ""),
            "", # catheter size
            out_aw.get("lumen_pre", ""), out_aw.get("lumen_post", ""),
            out.get("symptoms", ""), out.get("pleural", ""), out.get("complications", "")
        ]
        ws_v3.append(row_v3)

    # 7. Populate V3_Registry_JSON
    # -------------------------------------------------------------------------
    if "V3_Registry_JSON" not in wb.sheetnames:
        wb.create_sheet("V3_Registry_JSON")
    ws_json = wb["V3_Registry_JSON"]
    
    registry_obj = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": V3_EVENTS,
        "no_immediate_complications": True
    }
    
    ws_json.cell(row=1, column=1, value=json.dumps(registry_obj, indent=2))

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    main()