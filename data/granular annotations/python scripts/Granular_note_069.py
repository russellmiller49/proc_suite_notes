import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import datetime
import os

# -------------------------------------------------------------------------
# INPUTS
# -------------------------------------------------------------------------
NOTE_ID = "note_069"
SOURCE_FILE = "note_069.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_069 SOURCE_FILE: note_069.txt INDICATION FOR OPERATION:  [REDACTED] is a 64 year old-year-old male who presents with bronchial stenosis related to endobronchial tumor.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
 
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31622 Dx bronchoscope/cell washing          
31624 Dx bronchoscope/lavage (BAL)    
31625 Endobronchial Biopsy(s)
31630 Balloon dilation - right lower lobe bronchus (no stent placed)
31636 Dilate and bronchial stent initial bronchus - right middle lobe bronchus (dilation and stent placement)
31640 Bronchoscopy with excision 
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
 
XS Separate Structure, a service 
that is distinct because it was performed on a separate organ/structure
AND
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's 
 
XS Separate Structure:
CPT 31625 was performed at the trachea, right-sided airways (right mainstem bronchus, right upper lobe bronchus, bronchus intermedius, right middle lobe bronchus, and right lower lobe bronchus) and left-sided airways (left mainstem bronchus).
The trachea, right-sided airways, and left-sided airways are separate structures - thus these should be treated as distinct services x3.
CPT 31640 was performed at the trachea, right-sided airways (right mainstem bronchus, right upper lobe bronchus, bronchus intermedius, right middle lobe bronchus, and right lower lobe bronchus) and left-sided airways (left mainstem bronchus).
The trachea, right-sided airways, and left-sided airways are separate structures - thus these should be treated as distinct services x3.
CPT 31641 was performed at the trachea, right-sided airways (right mainstem bronchus, right upper lobe bronchus, bronchus intermedius, right middle lobe bronchus, and right lower lobe bronchus) and left-sided airways (left mainstem bronchus).
The trachea, right-sided airways, and left-sided airways are separate structures - thus these should be treated as distinct services x3.
AND
 
Unusual Procedure (22 MODIFIER):
This patient required extensive endobronchial biopsies of endobronchial tumor in/at the trachea (middle and distal trachea) and main carina, and extensive endobronchial biopsies of endobronchial tumor in the right-sided airways (right mainstem bronchus, right upper lobe bronchus, bronchus intermedius, right middle lobe bronchus, and right lower lobe bronchus).
This patient required extensive mechanical excision of obstructing tumors in/at the trachea (middle and distal trachea) and main carina, and extensive mechanical excision of obstructing tumors in the right-sided airways (right mainstem bronchus, right upper lobe bronchus, bronchus intermedius, right middle lobe bronchus, and right lower lobe bronchus).
This patient required extensive tumor ablation/destruction of obstructing tumors in/at the trachea (middle and distal trachea) and main carina, and extensive tumor ablation/destruction of obstructing tumors in the right-sided airways (right mainstem bronchus, right upper lobe bronchus, bronchus intermedius, right middle lobe bronchus, and right lower lobe bronchus).
This resulted in >150% increased work due to Increased intensity, Time, Technical difficulty of procedure, Severity of patient's condition, and Physical and mental effort required.
Apply to: 
31625 Endobronchial Biopsy(s)
31640 Bronchoscopy with excision 
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
 
 
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
Flexible Hybrid (Pedatric) Bronchoscope
 
PROCEDURE IN DETAIL:
A timeout was performed (confirming the patient's name, procedure type, and procedure location).
Sedation initiated.  LMA placed.
 
The Flexible Therapeutic Bronchoscope was advanced for airway examination.
Endobronchial topical lidocaine applied to the vocal cords, main carina, right carina 1, and left carina 2.
 
Initial Airway Inspection Findings:
Initial Airway Examination Findings:
Pharynx: Not assessed due to bronchoscopy introduction through LMA.
Larynx:  No lesions.  
Vocal Chords: Normal without mass/lesions.  Appropriate abduction and adduction of vocal cords.
Extensive partially obstructing vascular endobronchial tumor and overlying white necrotic debris involving the mid-trachea, distal trachea, main carina, right mainstem bronchus, right upper lobe bronchus, bronchus intermedius, right middle lobe bronchus, and right lower lobe bronchus, and left mainstem bronchus.
Most severe at the distal trachea, right mainstem bronchus and bronchus intermedius.
Extrinsic compression also contributing in the bronchus intermedius, right middle lobe, and right lower lboe.
The RML bronchus was highly stenotic and completely obstructed by tumor, debris, and mucus.
Unable to traverse with therapeutic bronchoscope initially.  Moderate stenosis of RLL RB6 and RLL truncus basalis due to tumor and extrinsic compression.
Mucosa: Normal except at sites of endobronchial tumor.
Secretions: Moderate, thin, and clear  All secretions were suctioned to clear (therapeutic aspiration).
Successful therapeutic aspiration was performed to clean out the trachea, right mainstem bronchus, right upper lobe, bronchus intermedius, right middle lobe, right lower lobe, left mainstem bronchus, left upper lobe, left lower lobe from mucus and overlying necrotic debris.
Endobronchial biopsy was performed at tumor in/at the trachea, main carina, right mainstem bronchus, right upper lobe, bronchus intermedius, right middle lobe, right lower lobe, and left mainstem bronchus.
This tumor was combined into a single sample labeled "trachea" endobronchial biopsy.  Lesion was successfully removed.  Samples sent for Pathology.
Obstructing endobronchial tumor was noted in/at the trachea, main carina, right mainstem bronchus, right upper lobe, bronchus intermedius, right middle lobe, right lower lobe, and left mainstem bronchus.
The obstructing tumor at the trachea, main carina, right mainstem bronchus, right upper lobe, bronchus intermedius, right middle lobe, right lower lobe, and left mainstem bronchus was excised with mechanical debridement using bland alligator forceps.
Obstructing endobronchial tumor at trachea, main carina, right mainstem bronchus, right upper lobe, bronchus intermedius, right middle lobe, right lower lobe, and left mainstem bronchus was treated with the following modalities:
Modality	Tools	Setting/Mode	Duration	Results
APC	1.5mm Pulmonary axial 'straight-fire' probe	forcedAPC, 0.5 LPM, Max Watts 30-40.
1-5 second bursts	Extensive tumor ablation and tumor coagulation in the trachea, main carina, right mainstem bronchus, right upper lobe, bronchus intermedius, right middle lobe, right lower lobe, and left mainstem bronchus.
Achieved good tumor ablation and hemostasis.
 
After tumor excision/ablation able to traverse right middle lobe bronchus with a Hybrid bronchoscope and saline hydrodissection - confirmed RB4 (RML lateral segment) was patent and that RB5 (RML medial segment) appeared viable, although filled with mucus and debris.
Balloon dilation was performed at right middle lobe bronchus stenosis.
6/7/8 Elation balloon was used to perform dilation to 7 mm at the right middle lobe bronchus stenosis.
Total 2 inflations with dilation time of 60 and 30 seconds each.
Afterward, able to traverse with therapeutic bronchoscope with moderate pressure.
Debris and mucos 
 
Balloon dilation was performed at right lower lobe bronchus take-off.
6/7/8 Elation balloon was used to perform dilation to 8 mm at the right lower lobe bronchus take-off.
Total 1 inflations with dilation time of 60 seconds each.
Bronchial alveolar lavage was performed at right middle lobe bronchus (Lateral Segment of RML (RB4) and Medial Segment of RML (RB5)).
Instilled 60 cc of NS, suction returned with 15 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
 
The following stent (AEROMini 6x15mm) was placed in the right middle lobe bronchus.
Seated very well, and maintained patency of both RB4 and RB5 segments.
Prior to treatment, the trachea was note to be 80% patent.  After treatment, the airway was 95% patent.
Prior to treatment, the right-sided airways (except the RML bronchus) was note to be 60% patent.
After treatment, the airway was 90% patent.  
Prior to treatment, the RML bronchus was note to be 1% patent.
After treatment, the airway was 80% patent.  
Prior to treatment, the left mainstem bronchus was note to be 80% patent.
After treatment, the airway was 95% patent.
 
Residual secretions, saline, and blood were suctioned to clear.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient's LMA was removed in the operating room and transported to the recovery room in stable condition.
ESTIMATED BLOOD LOSS:   Minimal
COMPLICATIONS:    None
 
SPECIMEN(S): 
-Endobronchial biopsy (samples from trachea, main carina, right mainstem bronchus, right upper lobe bronchus, bronchus intermedius, right middle lobe bronchus, and right lower lobe bronchus, and left mainstem bronchus combined into single sample titled "trachea"):  Pathology
-RML BAL:  cell count, micro/cultures, cytology
 
IMPRESSION/PLAN: [REDACTED] is a 64 year old-year-old male who presents for bronchoscopy for endobronchial tumor excision/ablation, airway dilation, and right middle lobe stent placement.
-Extensive endobronchial tumor excision and ablation in the trachea, main carina, right mainstem bronchus, right upper lobe bronchus, bronchus intermedius, right middle lobe bronchus, and right lower lobe bronchus, and left mainstem bronchus.
-Stent placement in RML bronchus (AEROMini 6x15mm)
-F/u post-op CXR
-F/u EBBx and BAL results.
-Stent hydration/hygiene protocol:
-QID albuterol nebulizer, hypertonic saline (3%, 4mL) nebulizer, and flutter valve regimen
-Mucinex 1200mg  BID for mucolysis/clearance
-This is very important to avoid stent obstruction and should be continued as long as patient has airway stent in place.
-Please engage case management to obtain nebulizer device and medications (as above) for patient prior to discharge as patient will require this regimen as long as he has stent in place.
-Will need repeat bronchoscopy with airway evaluation, stent check, and likely additional tumor excision/ablation in ~3-4 weeks.
Consideration of rigid bronchoscopy with micro-debrider given extensive endobronchial tumor involvement."""

# -------------------------------------------------------------------------
# PROCEDURE FLAGS
# -------------------------------------------------------------------------
FLAGS = {
    # Bronchoscopy
    'diagnostic_bronchoscopy': 1,
    'bal': 1,
    'bronchial_wash': 0,
    'brushings': 0,
    'endobronchial_biopsy': 1,
    'tbna_conventional': 0,
    'linear_ebus': 0,
    'radial_ebus': 0,
    'navigational_bronchoscopy': 0,
    'transbronchial_biopsy': 0,
    'transbronchial_cryobiopsy': 0,
    'therapeutic_aspiration': 1,
    'foreign_body_removal': 0,
    'airway_dilation': 1,
    'airway_stent': 1,
    'thermal_ablation': 1,
    'tumor_debulking_non_thermal': 1,
    'cryotherapy': 0,
    'blvr': 0,
    'peripheral_ablation': 0,
    'bronchial_thermoplasty': 0,
    'whole_lung_lavage': 0,
    'rigid_bronchoscopy': 0,
    
    # Pleural
    'thoracentesis': 0,
    'chest_tube': 0,
    'ipc': 0,
    'medical_thoracoscopy': 0,
    'pleurodesis': 0,
    'pleural_biopsy': 0,
    'fibrinolytic_therapy': 0
}

# -------------------------------------------------------------------------
# SPAN DATA
# -------------------------------------------------------------------------
# Helper to build spans
# Fields: span_text, label, normalized_value, event_id, context_prefix, section_type
spans_data = []

def add_span(text, label, value, event_id, context="", section="PROCEDURE IN DETAIL"):
    spans_data.append({
        "span_text": text,
        "label": label,
        "normalized_value": value,
        "event_id": event_id,
        "context_prefix": context,
        "section_type": section
    })

# Event 1: Setup & Inspection
add_span("Flexible Therapeutic Bronchoscope", "DEV_INSTRUMENT", "Flexible Therapeutic Bronchoscope", "ev1")
add_span("LMA", "DEV_INSTRUMENT", "LMA", "ev1")
add_span("Extensive partially obstructing vascular endobronchial tumor", "OBS_LESION", "endobronchial tumor", "ev1")
add_span("mid-trachea", "ANAT_AIRWAY", "Trachea", "ev1", "involving the")
add_span("distal trachea", "ANAT_AIRWAY", "Distal Trachea", "ev1", "mid-trachea,")
add_span("main carina", "ANAT_AIRWAY", "Main Carina", "ev1", "distal trachea,")
add_span("right mainstem bronchus", "ANAT_AIRWAY", "RMB", "ev1", "main carina,")
add_span("right upper lobe bronchus", "ANAT_AIRWAY", "RUL", "ev1", "right mainstem bronchus,")
add_span("bronchus intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "ev1", "upper lobe bronchus,")
add_span("right middle lobe bronchus", "ANAT_AIRWAY", "RML", "ev1", "bronchus intermedius,")
add_span("right lower lobe bronchus", "ANAT_AIRWAY", "RLL", "ev1", "middle lobe bronchus, and")
add_span("left mainstem bronchus", "ANAT_AIRWAY", "LMB", "ev1", "lower lobe bronchus, and")

add_span("RML bronchus", "ANAT_AIRWAY", "RML", "ev1", "The")
add_span("completely obstructed", "OBS_LESION", "Obstruction", "ev1")
add_span("RLL RB6", "ANAT_AIRWAY", "RLL Superior Segment", "ev1")
add_span("RLL truncus basalis", "ANAT_AIRWAY", "RLL Basal Segments", "ev1")
add_span("Moderate stenosis", "OBS_LESION", "Stenosis", "ev1")

# Event 2: Therapeutic Aspiration
add_span("therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", "ev2", "clear (")
add_span("clean out", "PROC_ACTION", "Clean", "ev2", "performed to")
add_span("necrotic debris", "OBS_LESION", "Necrotic Debris", "ev2", "overlying")

# Event 3: Endobronchial Biopsy
add_span("Endobronchial biopsy", "PROC_METHOD", "Endobronchial Biopsy", "ev3")
add_span("trachea", "ANAT_AIRWAY", "Trachea", "ev3", "in/at the")
add_span("single sample", "SPECIMEN", "Biopsy Sample", "ev3")

# Event 4: Mechanical Debridement
add_span("excised with mechanical debridement", "PROC_METHOD", "Mechanical Debridement", "ev4")
add_span("bland alligator forceps", "DEV_INSTRUMENT", "Alligator Forceps", "ev4")
add_span("Obstructing endobronchial tumor", "OBS_LESION", "Endobronchial Tumor", "ev4", "forceps.\n")

# Event 5: Thermal Ablation (APC)
add_span("APC", "PROC_METHOD", "APC", "ev5")
add_span("1.5mm Pulmonary axial 'straight-fire' probe", "DEV_INSTRUMENT", "APC Probe", "ev5")
add_span("Extensive tumor ablation", "PROC_ACTION", "Ablation", "ev5")
add_span("tumor coagulation", "PROC_ACTION", "Coagulation", "ev5")

# Event 6: Balloon Dilation RML
add_span("Balloon dilation", "PROC_METHOD", "Balloon Dilation", "ev6", "debris.\n")
add_span("right middle lobe bronchus stenosis", "ANAT_AIRWAY", "RML", "ev6")
add_span("6/7/8 Elation balloon", "DEV_INSTRUMENT", "Elation Balloon", "ev6", "stenosis.\n")
add_span("dilation to 7 mm", "PROC_ACTION", "Dilation", "ev6")
add_span("2 inflations", "MEAS_COUNT", "2", "ev6")

# Event 7: Balloon Dilation RLL
add_span("Balloon dilation", "PROC_METHOD", "Balloon Dilation", "ev7", "mucos \n \n")
add_span("right lower lobe bronchus take-off", "ANAT_AIRWAY", "RLL", "ev7")
add_span("6/7/8 Elation balloon", "DEV_INSTRUMENT", "Elation Balloon", "ev7", "take-off.\n")
add_span("dilation to 8 mm", "PROC_ACTION", "Dilation", "ev7")
add_span("1 inflations", "MEAS_COUNT", "1", "ev7")

# Event 8: BAL RML
add_span("Bronchial alveolar lavage", "PROC_METHOD", "BAL", "ev8")
add_span("right middle lobe bronchus", "ANAT_AIRWAY", "RML", "ev8", "performed at")
add_span("RB4", "ANAT_AIRWAY", "RML Lateral Segment", "ev8", "RML (")
add_span("RB5", "ANAT_AIRWAY", "RML Medial Segment", "ev8", "RML (")
add_span("15 cc", "MEAS_VOL", "15 cc", "ev8")

# Event 9: Stent RML
add_span("stent", "DEV_STENT", "Stent", "ev9", "following")
add_span("AEROMini", "DEV_STENT_MATERIAL", "AEROMini", "ev9")
add_span("6x15mm", "DEV_STENT_SIZE", "6x15mm", "ev9")
add_span("right middle lobe bronchus", "ANAT_AIRWAY", "RML", "ev9", "placed in the")
add_span("maintained patency", "PROC_ACTION", "Patency Check", "ev9")

# Event 10: Outcomes
add_span("trachea", "ANAT_AIRWAY", "Trachea", "ev10", "treatment, the")
add_span("80% patent", "OUTCOME_AIRWAY_LUMEN_PRE", "80%", "ev10", "note to be")
add_span("95% patent", "OUTCOME_AIRWAY_LUMEN_POST", "95%", "ev10", "airway was")

add_span("right-sided airways", "ANAT_AIRWAY", "Right Airways", "ev10")
add_span("60% patent", "OUTCOME_AIRWAY_LUMEN_PRE", "60%", "ev10", "note to be")
add_span("90% patent", "OUTCOME_AIRWAY_LUMEN_POST", "90%", "ev10", "airway was")

add_span("RML bronchus", "ANAT_AIRWAY", "RML", "ev10", "treatment, the")
add_span("1% patent", "OUTCOME_AIRWAY_LUMEN_PRE", "1%", "ev10", "note to be")
add_span("80% patent", "OUTCOME_AIRWAY_LUMEN_POST", "80%", "ev10", "airway was")

add_span("left mainstem bronchus", "ANAT_AIRWAY", "LMB", "ev10", "treatment, the")
add_span("80% patent", "OUTCOME_AIRWAY_LUMEN_PRE", "80%", "ev10", "note to be")
add_span("95% patent", "OUTCOME_AIRWAY_LUMEN_POST", "95%", "ev10", "airway was")

add_span("No immediate complications", "OUTCOME_COMPLICATION", "None", "ev10")


# -------------------------------------------------------------------------
# EVENT LOG GENERATION
# -------------------------------------------------------------------------
# Structure: event_id -> dict
event_registry = {
    "ev1": {
        "event_type": "Diagnostic",
        "method": "Inspection",
        "anatomy_target": "Trachea, RMB, RUL, BI, RML, RLL, LMB",
        "findings": "Extensive obstructing tumor, stenosis",
        "device": "Flexible Therapeutic Bronchoscope"
    },
    "ev2": {
        "event_type": "Therapeutic",
        "method": "Therapeutic Aspiration",
        "anatomy_target": "Trachea, Bilateral Airways",
        "findings": "Mucus and necrotic debris",
        "device": "Suction"
    },
    "ev3": {
        "event_type": "Diagnostic",
        "method": "Endobronchial Biopsy",
        "anatomy_target": "Trachea, RMB, RUL, BI, RML, RLL, LMB",
        "specimens": "Single combined sample (Trachea)",
        "findings": "Tumor"
    },
    "ev4": {
        "event_type": "Therapeutic",
        "method": "Mechanical Debridement",
        "anatomy_target": "Trachea, RMB, RUL, BI, RML, RLL, LMB",
        "device": "Alligator Forceps"
    },
    "ev5": {
        "event_type": "Therapeutic",
        "method": "Thermal Ablation",
        "anatomy_target": "Trachea, RMB, RUL, BI, RML, RLL, LMB",
        "device": "APC Probe"
    },
    "ev6": {
        "event_type": "Therapeutic",
        "method": "Balloon Dilation",
        "anatomy_target": "RML",
        "device": "Elation Balloon (6/7/8mm)",
        "measurements": "Dilation to 7mm",
        "counts": "2 inflations"
    },
    "ev7": {
        "event_type": "Therapeutic",
        "method": "Balloon Dilation",
        "anatomy_target": "RLL",
        "device": "Elation Balloon (6/7/8mm)",
        "measurements": "Dilation to 8mm",
        "counts": "1 inflation"
    },
    "ev8": {
        "event_type": "Diagnostic",
        "method": "BAL",
        "anatomy_target": "RML (RB4, RB5)",
        "measurements": "Instilled 60cc, Returned 15cc",
        "specimens": "Cell count, Micro, Cytology"
    },
    "ev9": {
        "event_type": "Therapeutic",
        "method": "Airway Stent",
        "anatomy_target": "RML",
        "device": "AEROMini Stent",
        "device_size": "6x15mm",
        "device_material": "AEROMini"
    },
    "ev10": {
        "event_type": "Outcome",
        "method": "Assessment",
        "anatomy_target": "Airways",
        "outcome_airway_lumen_pre": "1% (RML)",
        "outcome_airway_lumen_post": "80% (RML)",
        "outcome_complication": "None"
    }
}

# -------------------------------------------------------------------------
# SCRIPT UTILS
# -------------------------------------------------------------------------
def clean_text(text):
    return text.replace('\r', '').replace('\t', ' ')

def find_offset(text, span_text, context_prefix, match_index):
    # If perfect match unique
    if text.count(span_text) == 1:
        return text.find(span_text), "hydrated_unique"
    
    # If context prefix
    if context_prefix:
        # Scan for context
        idx = 0
        while True:
            idx = text.find(span_text, idx)
            if idx == -1: break
            
            # Check previous ~120 chars
            window = text[max(0, idx-120):idx]
            if context_prefix in window:
                return idx, "hydrated_prefix_window"
            idx += 1
            
    # Fallback to match index
    if match_index > 0:
        current = 0
        for _ in range(match_index):
            loc = text.find(span_text, current)
            if loc == -1: return -1, "ambiguous_index_fail"
            current = loc + 1
        return loc, "hydrated_match_index"

    return -1, "ambiguous"

def generate_excel():
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    
    # 1. Note_Text
    ws = wb["Note_Text"]
    ws.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # 2. Note_Index
    ws = wb["Note_Index"]
    row = [
        SOURCE_FILE, NOTE_ID, "", "", "", "", "Success", ""
    ]
    # Append flags in order
    for k in FLAGS:
        row.append(FLAGS[k])
    ws.append(row)
    
    # 3. Span_Annotations & 4. Span_Hydrated
    ws_anno = wb["Span_Annotations"]
    ws_hydra = wb["Span_Hydrated"]
    
    # Sort spans by occurrence in text to approximate flow, though not strictly required
    # We will just process list order.
    
    cleaned_note = clean_text(NOTE_TEXT)
    
    for span in spans_data:
        # Calculate hydration
        start, status = find_offset(cleaned_note, span["span_text"], span.get("context_prefix"), 0)
        end = start + len(span["span_text"]) if start != -1 else -1
        span_len = len(span["span_text"])
        
        # Base Row
        row_base = [
            SOURCE_FILE, NOTE_ID, "", span.get("section_type", "PROCEDURE IN DETAIL"),
            span.get("context_prefix", ""), span["span_text"], 1 # Match index default 1
        ]
        
        # Anno Row (Blanks for offsets)
        row_anno = row_base + [
            "", "", f"=LEN(F{ws_anno.max_row+1})",
            span["label"], span["normalized_value"], "", span["event_id"],
            "", "", "", "", "", "needs_hydration"
        ]
        ws_anno.append(row_anno)
        
        # Hydra Row (Filled offsets)
        row_hydra = row_base + [
            start if start!=-1 else "", end if end!=-1 else "", span_len,
            span["label"], span["normalized_value"], "", span["event_id"],
            "", "", "", "", "", status
        ]
        ws_hydra.append(row_hydra)

    # 5. Event_Log
    ws_event = wb["Event_Log"]
    for eid, data in event_registry.items():
        row = [
            SOURCE_FILE, NOTE_ID, eid, data.get("event_type"), data.get("method"),
            data.get("anatomy_target"), data.get("device"), "", "", data.get("counts"),
            data.get("measurements"), data.get("specimens"), data.get("findings"),
            "", "", "",
            data.get("device_size"), data.get("device_material"),
            data.get("outcome_airway_lumen_pre"), data.get("outcome_airway_lumen_post"),
            data.get("outcome_symptoms"), data.get("outcome_pleural"),
            data.get("outcome_complication")
        ]
        ws_event.append(row)

    # 6. V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    
    # Map events to V3 structure
    # Common outcomes (applied to therapeutic events or separate)
    # We have explicit events.
    
    for eid, data in event_registry.items():
        # Defaults
        loc_lobe = ""
        if "RML" in str(data.get("anatomy_target")): loc_lobe = "RML"
        elif "RLL" in str(data.get("anatomy_target")): loc_lobe = "RLL"
        
        # Construct row
        row = [
            NOTE_ID, eid, data.get("event_type"),
            "Airway", loc_lobe, "", "", # Target
            "", "", # Lesion
            data.get("method"), "", "", "", "", "", # Method/Devices/Meas/Spec/Findings/Quote
            data.get("device_size"), data.get("device_material"), "", # Stent/Cath
            data.get("outcome_airway_lumen_pre"), data.get("outcome_airway_lumen_post"), # Outcomes
            "", "", data.get("outcome_complication")
        ]
        ws_v3.append(row)

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    json_obj = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "procedures": list(event_registry.values()),
        "no_immediate_complications": True
    }
    ws_json.cell(row=1, column=1, value=json.dumps(json_obj, indent=2))

    wb.save(OUTPUT_PATH)

if __name__ == "__main__":
    generate_excel()