import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import json
import re
import os

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_070"
SOURCE_FILE = "note_070.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_070 SOURCE_FILE: note_070.txt INDICATION FOR OPERATION:  [REDACTED] is a 51 year old-year-old female who presents with lung nodule.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: R91.1 Solitary Lung Nodule
 
POSTOPERATIVE DIAGNOSIS:  R91.1 Solitary Lung Nodule
 
PROCEDURE:  
31899 Unlisted Procedure (Trach Change with Mature Tract or Procedure NOS)
31645 Therapeutic aspiration initial episode
31623 Dx bronchoscope/brushing    
31624 Dx bronchoscope/lavage (BAL)    
31628 TBBX single lobe     
31629 TBNA single lobe   
31626 Fiducial marker placements, single or multiple     
31627 Navigational Bronchoscopy (computer assisted)
77012 Radiology / radiologic guidance for CT guided needle placement (CIOS)
76377 3D rendering with interpretation and reporting of CT, US, Tomo modality (ION Planning Station)
31652 EBUS sampling 1 or 
2 nodes
31654 Radial EBUS for peripheral lesion
76982 Ultrasound Elastography, First Target Lesion
76983 Ultrasound Elastography, Additional Targets 
76983 Ultrasound Elastography, Additional Target 2
31630 Balloon dilation
 
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)
 
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required a Transbronchial Cryo biopsies.
This resulted in >40% increased work due to Technical difficulty of procedure and Physical and mental effort required.
Apply to: 31628 TBBX single lobe. 
 
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Linear EBUS 
Radial EBUS
Ion Robotic Bronchoscope
Disposable Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
 
CT Chest scan was placed on separate planning station to generate 3D rendering of the pathway to target.
The navigational plan was reviewed and verified.  This was then loaded into robotic bronchoscopy platform.
Successful therapeutic aspiration was performed to clean out the Right Mainstem, Bronchus Intermedius , and Left Mainstem from mucus.
Ventilation Parameters:
Mode	RR	TV	PEEP	FiO2	Flow Rate	Pmean
vcv	14	400	12	100	10	15
 
Robotic navigation bronchoscopy was performed with Ion platform.  Partial registration was used.
Ion robotic catheter was used to engage the Anterior Segment of RUL (RB3).
Target lesion is about 1 cm in diameter.   
 
Significant difficulties were encountered during the navigation due to a small carina.
5mmx20mm mustang balloon was used to dilate the distal airway under direct fluoro-guidance with Omnipaque 240 as inflation agent.
After dilation, the robotic bronchoscope was easily navigated to the distal nodule under navigational guidance the ion robotic catheter was advanced to 1.0 cm away from the planned target.
Radial EBUS was performed to confirm that the location of the nodule is Eccentric.
The following features were noted: Continuous margin .
 
Cone Beam CT was performed: 3-D reconstructions were performed on an independent workstation.
Cios Spin system was used for evaluation of nodule location.  Low dose spin was performed to acquire CT imaging.
This was passed on to Ion platform system for reconstruction and nodule location.
The 3D images was interpreted on an independent workstation (Ion).
Using the newly acquired nodule location, the Ion robotic system was adjusted to the new targeted location.
I personally interpreted the cone beam CT and 3-D reconstruction.
Transbronchial needle aspiration was performed with 21G Needle through the extended working channel catheter.  Total 6 samples were collected.
Samples sent for Cytology.
 
Transbronchial biopsy was performed with alligator forceps the extended working channel catheter.
Total 2 samples were collected.  Samples sent for Pathology.
 
Transbronchial cryobiopsy was performed with 1.1mm cryoprobe via the extended working channel catheter.
Freeze time of 6 seconds were used.  Total 6 samples were collected.  Samples sent for Pathology.
Transbronchial brushing was performed with Protected cytology brush the extended working channel catheter.  Total 1 samples were collected.
Samples sent for Microbiology (Cultures/Viral/Fungal).
 
Bronchial alveolar lavage was performed the extended working channel catheter.
Instilled 40 cc of NS, suction returned with 15 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
 
Fiducial marker (0.8mm x 3mm soft tissue gold CIVCO) was loaded with bone wax and placed under fluoroscopy guidance.
Prior to withdraw of the bronchoscope. 
 
ROSE from ION procedure was noted to be:
Conclusive evidence of malignant neoplasm
 
Prior to withdrawal of the bronchoscope, inspection demonstrated no evidence of bleeding.
EBUS-Findings
Indications: Diagnostic and Staging
Technique:
All lymph node stations were assessed. Only those 5 mm or greater in short axis were sampled.
Lymph node sizing was performed by EBUS and sampling by transbronchial needle aspiration was performed using 22-gauge Needle.
Lymph Nodes/Sites Inspected: 4R (lower paratracheal) node
4L (lower paratracheal) node
7 (subcarinal) node
11Rs lymph node
11L lymph node
 
No immediate complications
 
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
Elastography provided a semi-quantitative classification (Type 1â€“3), which was used to guide biopsy site selection and sampling strategy.
Lymph Nodes Evaluated:
Site 1: The 11L lymph node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was not sampled: Sampling this lymph node was not clinically indicated.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 1 elastographic pattern, predominantly soft (green/yellow), suggesting a reactive or benign process.
Site 2: The 7 (subcarinal) node was => 10 mm on CT and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was sampled.. 4 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 2 elastographic pattern with mixed soft and stiff regions.
Given this heterogeneous and indeterminate appearance, TBNA was directed at representative areas to ensure comprehensive sampling and to minimize the risk of underdiagnosis.
Site 3: The 4R (lower paratracheal) node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was not sampled: Sampling this lymph node was not clinically indicated.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 1 elastographic pattern, predominantly soft (green/yellow), suggesting a reactive or benign process.
Site 4: The 11Rs lymph node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was not sampled: Sampling this lymph node was not clinically indicated.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 1 elastographic pattern, predominantly soft (green/yellow), suggesting a reactive or benign process.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
TBCBX, TBBX, TBNA, Brush, BAL RUL lung nodule
TBNA station 7 
 
IMPRESSION/PLAN: [REDACTED] is a 51 year old-year-old female who presents for bronchoscopy for lung nodule.
- f/u results in clinic
- f/u cxr"""

# -------------------------------------------------------------------------
# UTILS
# -------------------------------------------------------------------------

def clean_text(text):
    return text.strip().replace('\r', '')

def count_occurrences(text, sub):
    return text.count(sub)

def find_contextual_start(full_text, substring, context_prefix, match_index):
    # If context prefix provided, use it
    if context_prefix:
        pattern = re.escape(context_prefix) + r".{0,120}?" + re.escape(substring)
        matches = list(re.finditer(pattern, full_text, re.DOTALL))
        if matches:
            # We want the start of the substring, not the prefix
            m = matches[0]
            # find the substring inside the match
            sub_start = m.group(0).find(substring)
            return m.start() + sub_start
    
    # If match_index provided
    if match_index is not None:
        matches = [m.start() for m in re.finditer(re.escape(substring), full_text)]
        if 0 <= match_index < len(matches):
            return matches[match_index]
    
    # Fallback: simple first occurrence
    return full_text.find(substring)

# -------------------------------------------------------------------------
# EXTRACTION LOGIC
# -------------------------------------------------------------------------

def extract_flags(text):
    flags = {
        'diagnostic_bronchoscopy': 1,
        'bal': 0, 'bronchial_wash': 0, 'brushings': 0, 'endobronchial_biopsy': 0,
        'tbna_conventional': 0, 'linear_ebus': 0, 'radial_ebus': 0,
        'navigational_bronchoscopy': 0, 'transbronchial_biopsy': 0,
        'transbronchial_cryobiopsy': 0, 'therapeutic_aspiration': 0,
        'foreign_body_removal': 0, 'airway_dilation': 0, 'airway_stent': 0,
        'thermal_ablation': 0, 'tumor_debulking_non_thermal': 0,
        'cryotherapy': 0, 'blvr': 0, 'peripheral_ablation': 0,
        'bronchial_thermoplasty': 0, 'whole_lung_lavage': 0, 'rigid_bronchoscopy': 0,
        'thoracentesis': 0, 'chest_tube': 0, 'ipc': 0,
        'medical_thoracoscopy': 0, 'pleurodesis': 0, 'pleural_biopsy': 0,
        'fibrinolytic_therapy': 0
    }
    
    text_lower = text.lower()
    
    if 'bal' in text_lower or 'lavage' in text_lower:
        flags['bal'] = 1
    if 'brush' in text_lower:
        flags['brushings'] = 1
    if 'tbna' in text_lower or 'transbronchial needle' in text_lower:
        flags['tbna_conventional'] = 1
    if 'linear ebus' in text_lower or 'ebus' in text_lower:
        flags['linear_ebus'] = 1
    if 'radial ebus' in text_lower:
        flags['radial_ebus'] = 1
    if 'navigational' in text_lower or 'ion' in text_lower:
        flags['navigational_bronchoscopy'] = 1
    if 'transbronchial biopsy' in text_lower or 'tbbx' in text_lower:
        flags['transbronchial_biopsy'] = 1
    if 'cryobiopsy' in text_lower or 'cryo biopsies' in text_lower:
        flags['transbronchial_cryobiopsy'] = 1
    if 'therapeutic aspiration' in text_lower:
        flags['therapeutic_aspiration'] = 1
    if 'dilation' in text_lower or 'balloon' in text_lower:
        flags['airway_dilation'] = 1
        
    return flags

def generate_spans_and_events(text):
    spans = []
    events = []
    
    # Helper to add span
    def add_span(text_segment, label, norm_val, field, event_id, prefix=None, match_idx=0):
        if text_segment not in text:
            return
        spans.append({
            'span_text': text_segment,
            'label': label,
            'normalized_value': norm_val,
            'schema_field': field,
            'event_id': event_id,
            'context_prefix': prefix,
            'match_index': match_idx
        })

    # --- EVENT 1: Therapeutic Aspiration ---
    evt_asp = "evt_aspiration"
    add_span("Therapeutic aspiration", "PROC_METHOD", "Therapeutic aspiration", "method", evt_asp)
    add_span("Right Mainstem", "ANAT_AIRWAY", "Right Mainstem", "target.anatomy_type", evt_asp)
    add_span("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "target.anatomy_type", evt_asp)
    add_span("Left Mainstem", "ANAT_AIRWAY", "Left Mainstem", "target.anatomy_type", evt_asp)
    add_span("clean out", "PROC_ACTION", "clean out", "action", evt_asp)
    add_span("mucus", "OBS_ROSE", "mucus", "findings", evt_asp)
    
    events.append({
        "event_id": evt_asp,
        "type": "Therapeutic Aspiration",
        "method": "Suction",
        "anatomy_target": "RMS, BI, LMS",
        "findings": "mucus"
    })

    # --- EVENT 2: Navigation & Nodule Target ---
    evt_nav = "evt_nav_nodule"
    add_span("Robotic navigation bronchoscopy", "PROC_METHOD", "Robotic Bronchoscopy", "method", evt_nav)
    add_span("Ion", "DEV_INSTRUMENT", "Ion", "device", evt_nav)
    add_span("Anterior Segment of RUL", "ANAT_LUNG_LOC", "RUL Anterior Segment", "target.location.segment", evt_nav)
    add_span("RB3", "ANAT_LUNG_LOC", "RB3", "target.location.segment", evt_nav)
    add_span("1 cm", "MEAS_SIZE", "10", "lesion.size_mm", evt_nav, prefix="Target lesion is about ")
    add_span("Radial EBUS", "PROC_METHOD", "Radial EBUS", "method", evt_nav)
    add_span("Eccentric", "OBS_LESION", "Eccentric", "lesion.type", evt_nav)
    add_span("Continuous margin", "OBS_LESION", "Continuous margin", "lesion.type", evt_nav)
    add_span("Cios Spin", "DEV_INSTRUMENT", "Cios Spin", "device", evt_nav)
    add_span("Cone Beam CT", "PROC_METHOD", "Cone Beam CT", "method", evt_nav)
    
    events.append({
        "event_id": evt_nav,
        "type": "Navigation/Imaging",
        "method": "Ion, Radial EBUS, Cone Beam CT",
        "anatomy_target": "RUL Anterior Segment (RB3)",
        "findings": "Eccentric nodule, 1cm"
    })

    # --- EVENT 3: Airway Dilation ---
    evt_dil = "evt_dilation"
    add_span("small carina", "ANAT_AIRWAY", "Carina", "target.anatomy_type", evt_dil)
    add_span("mustang balloon", "DEV_INSTRUMENT", "Mustang Balloon", "device", evt_dil)
    add_span("5mmx20mm", "DEV_CATHETER_SIZE", "5mmx20mm", "device_size", evt_dil)
    add_span("dilate", "PROC_ACTION", "Dilation", "method", evt_dil, prefix="used to ")
    
    events.append({
        "event_id": evt_dil,
        "type": "Airway Dilation",
        "method": "Balloon",
        "device": "Mustang Balloon",
        "device_size": "5mmx20mm",
        "anatomy_target": "Carina (distal airway)"
    })

    # --- EVENT 4: TBNA Nodule ---
    evt_tbna = "evt_tbna_nodule"
    add_span("Transbronchial needle aspiration", "PROC_METHOD", "TBNA", "method", evt_tbna, prefix="reconstruction.\n")
    add_span("21G Needle", "DEV_NEEDLE", "21G", "device", evt_tbna)
    add_span("6 samples", "MEAS_COUNT", "6", "specimens.count", evt_tbna, prefix="Total ")
    
    events.append({
        "event_id": evt_tbna,
        "type": "TBNA",
        "method": "TBNA",
        "anatomy_target": "RUL Nodule",
        "needle_gauge": "21G",
        "counts": "6 samples"
    })

    # --- EVENT 5: TBBX Nodule ---
    evt_tbbx = "evt_tbbx_nodule"
    add_span("Transbronchial biopsy", "PROC_METHOD", "Transbronchial biopsy", "method", evt_tbbx)
    add_span("alligator forceps", "DEV_INSTRUMENT", "Alligator Forceps", "device", evt_tbbx)
    add_span("2 samples", "MEAS_COUNT", "2", "specimens.count", evt_tbbx, prefix="Total ")
    
    events.append({
        "event_id": evt_tbbx,
        "type": "TBBX",
        "method": "Forceps",
        "anatomy_target": "RUL Nodule",
        "counts": "2 samples"
    })

    # --- EVENT 6: Cryobiopsy Nodule ---
    evt_cryo = "evt_cryo_nodule"
    add_span("Transbronchial cryobiopsy", "PROC_METHOD", "Cryobiopsy", "method", evt_cryo)
    add_span("1.1mm cryoprobe", "DEV_INSTRUMENT", "Cryoprobe 1.1mm", "device", evt_cryo)
    add_span("Freeze time of 6 seconds", "PROC_METHOD", "6s Freeze", "method", evt_cryo)
    add_span("6 samples", "MEAS_COUNT", "6", "specimens.count", evt_cryo, prefix="Total ")
    
    events.append({
        "event_id": evt_cryo,
        "type": "Cryobiopsy",
        "method": "Cryoprobe",
        "anatomy_target": "RUL Nodule",
        "counts": "6 samples"
    })

    # --- EVENT 7: Brushing Nodule ---
    evt_brush = "evt_brush_nodule"
    add_span("Transbronchial brushing", "PROC_METHOD", "Brushing", "method", evt_brush)
    add_span("Protected cytology brush", "DEV_INSTRUMENT", "Protected Cytology Brush", "device", evt_brush)
    add_span("1 samples", "MEAS_COUNT", "1", "specimens.count", evt_brush, prefix="Total ")
    
    events.append({
        "event_id": evt_brush,
        "type": "Brushing",
        "method": "Brush",
        "anatomy_target": "RUL Nodule",
        "counts": "1 samples"
    })

    # --- EVENT 8: BAL Nodule ---
    evt_bal = "evt_bal_nodule"
    add_span("Bronchial alveolar lavage", "PROC_METHOD", "BAL", "method", evt_bal)
    add_span("40 cc", "MEAS_VOL", "40", "measurements.instilled", evt_bal)
    add_span("15 cc", "MEAS_VOL", "15", "measurements.returned", evt_bal)
    
    events.append({
        "event_id": evt_bal,
        "type": "BAL",
        "method": "Lavage",
        "anatomy_target": "RUL Nodule",
        "measurements": "40cc in, 15cc out"
    })

    # --- EVENT 9: Fiducial ---
    evt_fid = "evt_fiducial"
    add_span("Fiducial marker", "PROC_METHOD", "Fiducial Placement", "method", evt_fid)
    add_span("0.8mm x 3mm", "DEV_DEVICE_SIZE", "0.8mm x 3mm", "device_size", evt_fid)
    add_span("gold CIVCO", "DEV_INSTRUMENT", "Gold CIVCO", "device", evt_fid)
    
    events.append({
        "event_id": evt_fid,
        "type": "Fiducial Placement",
        "method": "Placement",
        "device": "Gold CIVCO Marker",
        "anatomy_target": "RUL Nodule"
    })
    
    # --- ROSE Findings ---
    add_span("Malignant neoplasm", "OBS_ROSE", "Malignant neoplasm", "findings", "evt_nav")

    # --- EVENT 10: EBUS Station 11L ---
    evt_11l = "evt_ebus_11l"
    add_span("11L", "ANAT_LN_STATION", "11L", "target.station", evt_11l, prefix="Site 1: The ")
    add_span("< 10 mm", "MEAS_SIZE", "<10", "lesion.size_mm", evt_11l, prefix="lymph node was ")
    add_span("not sampled", "PROC_ACTION", "Not Sampled", "method", evt_11l, prefix="The site was ")
    add_span("Type 1", "OBS_LESION", "Type 1", "lesion.type", evt_11l, prefix="demonstrated a ")
    
    events.append({
        "event_id": evt_11l,
        "type": "EBUS Inspection",
        "anatomy_target": "11L",
        "findings": "<10mm, Type 1",
        "method": "Inspection only"
    })

    # --- EVENT 11: EBUS Station 7 ---
    evt_7 = "evt_ebus_7"
    add_span("7 (subcarinal)", "ANAT_LN_STATION", "7", "target.station", evt_7, prefix="Site 2: The ")
    add_span("=> 10 mm", "MEAS_SIZE", ">=10", "lesion.size_mm", evt_7, prefix="node was ")
    add_span("sampled", "PROC_ACTION", "Sampled", "method", evt_7, prefix="The site was ")
    add_span("4", "MEAS_COUNT", "4", "specimens.count", evt_7, prefix="sampled.. ")
    add_span("transbronchial biopsies", "PROC_METHOD", "TBNA", "method", evt_7, prefix="ultrasound guided ")
    add_span("Type 2", "OBS_LESION", "Type 2", "lesion.type", evt_7, prefix="demonstrated a ")
    add_span("22-gauge Needle", "DEV_NEEDLE", "22G", "device", evt_7, prefix="performed using ")

    events.append({
        "event_id": evt_7,
        "type": "EBUS TBNA",
        "anatomy_target": "7",
        "findings": ">=10mm, Type 2",
        "method": "TBNA",
        "counts": "4 samples",
        "needle_gauge": "22G"
    })

    # --- EVENT 12: EBUS Station 4R ---
    evt_4r = "evt_ebus_4r"
    add_span("4R", "ANAT_LN_STATION", "4R", "target.station", evt_4r, prefix="Site 3: The ")
    add_span("< 10 mm", "MEAS_SIZE", "<10", "lesion.size_mm", evt_4r, prefix="node was ", match_idx=1) # Need to handle multiple matches carefully
    add_span("not sampled", "PROC_ACTION", "Not Sampled", "method", evt_4r, prefix="The site was ", match_idx=1)
    
    events.append({
        "event_id": evt_4r,
        "type": "EBUS Inspection",
        "anatomy_target": "4R",
        "findings": "<10mm, Type 1",
        "method": "Inspection only"
    })
    
    # --- EVENT 13: EBUS Station 11Rs ---
    evt_11rs = "evt_ebus_11rs"
    add_span("11Rs", "ANAT_LN_STATION", "11Rs", "target.station", evt_11rs, prefix="Site 4: The ")
    
    events.append({
        "event_id": evt_11rs,
        "type": "EBUS Inspection",
        "anatomy_target": "11Rs",
        "findings": "<10mm, Type 1",
        "method": "Inspection only"
    })

    # --- OUTCOMES ---
    add_span("No immediate complications", "OUTCOME_COMPLICATION", "None", "outcomes.complications", "global", prefix="well.  There were ")

    return spans, events

# -------------------------------------------------------------------------
# EXECUTION
# -------------------------------------------------------------------------

def create_workbook():
    # 1. Load Template
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Template not found: {TEMPLATE_PATH}")
    
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    
    # 2. Extract Data
    flags = extract_flags(NOTE_TEXT)
    spans, event_log_data = generate_spans_and_events(NOTE_TEXT)
    
    # 3. Populate Note_Text
    ws_text = wb["Note_Text"]
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # 4. Populate Note_Index
    ws_index = wb["Note_Index"]
    row_meta = [
        SOURCE_FILE, NOTE_ID, "", "", "", "", "Draft", ""
    ]
    # Map flags to Note_Index columns (skipping metadata cols)
    # The template has specific columns for these flags.
    # We assume the order matches the list below as per instructions.
    flag_keys = [
        'diagnostic_bronchoscopy', 'bal', 'bronchial_wash', 'brushings', 'endobronchial_biopsy',
        'tbna_conventional', 'linear_ebus', 'radial_ebus', 'navigational_bronchoscopy', 'transbronchial_biopsy',
        'transbronchial_cryobiopsy', 'therapeutic_aspiration', 'foreign_body_removal', 'airway_dilation', 'airway_stent',
        'thermal_ablation', 'tumor_debulking_non_thermal', 'cryotherapy', 'blvr', 'peripheral_ablation',
        'bronchial_thermoplasty', 'whole_lung_lavage', 'rigid_bronchoscopy',
        'thoracentesis', 'chest_tube', 'ipc', 'medical_thoracoscopy', 'pleurodesis', 'pleural_biopsy', 'fibrinolytic_therapy'
    ]
    
    flag_values = [flags.get(k, 0) for k in flag_keys]
    ws_index.append(row_meta + flag_values)
    
    # 5. Populate Span_Annotations
    ws_span = wb["Span_Annotations"]
    # Columns: source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, start_char, end_char, span_len, label, normalized_value, schema_field, event_id, is_negated, is_historical, time_anchor, reviewer, comments, hydration_status
    
    hydrated_rows = []
    
    for i, s in enumerate(spans):
        span_id = f"{NOTE_ID}_s{i+1:03d}"
        
        # Calculate hydration
        start = find_contextual_start(NOTE_TEXT, s['span_text'], s['context_prefix'], s.get('match_index'))
        end = start + len(s['span_text']) if start != -1 else -1
        status = "hydrated" if start != -1 else "missing"
        
        # Base row
        row = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure",
            s['context_prefix'], s['span_text'], s.get('match_index', 0),
            None, None, f"=LEN(F{ws_span.max_row+1})", # Formula for length
            s['label'], s['normalized_value'], s['schema_field'], s['event_id'],
            0, 0, "", "", "", "needs_hydration"
        ]
        ws_span.append(row)
        
        # Hydrated row
        h_row = list(row)
        h_row[7] = start
        h_row[8] = end
        h_row[9] = len(s['span_text'])
        h_row[19] = status
        hydrated_rows.append(h_row)
        
    # 6. Populate Span_Hydrated
    ws_hyd = wb["Span_Hydrated"]
    for hr in hydrated_rows:
        ws_hyd.append(hr)
        
    # 7. Populate Event_Log
    ws_event = wb["Event_Log"]
    for e in event_log_data:
        # source_file, note_id, event_id, event_type, method, anatomy_target, device, needle_gauge, stations, counts, measurements, specimens, findings, is_historical, reviewer, comments, device_size, device_material, outcome_airway_lumen_pre, outcome_airway_lumen_post, outcome_symptoms, outcome_pleural, outcome_complication
        row = [
            SOURCE_FILE, NOTE_ID, e['event_id'], e['type'], e.get('method', ''),
            e.get('anatomy_target', ''), e.get('device', ''), e.get('needle_gauge', ''),
            '', e.get('counts', ''), e.get('measurements', ''), '',
            e.get('findings', ''), 0, "", "",
            e.get('device_size', ''), '',
            '', '', '', '', ''
        ]
        ws_event.append(row)
        
    # 8. Populate V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    for e in event_log_data:
        # note_id, event_id, type, target.anatomy_type, target.location.lobe, target.location.segment, target.station, lesion.type, lesion.size_mm, method, devices_json, measurements_json, specimens_json, findings_json, evidence_quote, stent.size, stent.material_or_brand, catheter.size_fr, outcomes.airway.lumen_pre, outcomes.airway.lumen_post, outcomes.symptoms, outcomes.pleural, outcomes.complications
        
        # Minimal mapping for V3
        row = [
            NOTE_ID, e['event_id'], e['type'],
            e.get('anatomy_target', ''), '', '', '',
            '', '', e.get('method', ''),
            json.dumps({"device": e.get('device')}) if e.get('device') else "",
            "", "", "", "",
            "", "", "",
            "", "", "", "", ""
        ]
        ws_v3.append(row)
        
    # 9. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": event_log_data,
        "no_immediate_complications": True
    }
    ws_json.append([json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    create_workbook()