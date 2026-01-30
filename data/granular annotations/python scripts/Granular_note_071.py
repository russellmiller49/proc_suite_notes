import openpyxl
from openpyxl.utils import get_column_letter
import json
import re

# =============================================================================
# 1. CONSTANTS & INPUTS
# =============================================================================
NOTE_ID = "note_071"
SOURCE_FILE = "note_071.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_071 SOURCE_FILE: note_071.txt INDICATION FOR OPERATION:  [REDACTED] is a 85 year old-year-old female who presents with airway stenosis.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
 
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
 
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31624 Dx bronchoscope/lavage (BAL)    
31625 Endobronchial Biopsy(s)
31630 Balloon dilation
 
 
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Disposable Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
Successful therapeutic aspiration was performed to clean out the Right Mainstem, Bronchus Intermedius , and Left Mainstem from mucus.
Endobronchial biopsy was performed at Bronchus Intermedius .  Two areas were biopsied with NBI guidance.
Samples sent for Microbiology (Cultures/Viral/Fungal) and Pathology.
 
Stents are in good position.
The (iCAST 7x16) was noted in the right lower airway.  (iCAST 5x16) was noted in the right middle lobe airway.
Bronchial alveolar lavage was performed at RLL.  Instilled 40 cc of NS, suction returned with 15 cc of NS.
Samples sent for Microbiology (Cultures/Viral/Fungal).
 
Balloon dilation was performed at RB2 orifice.
5mm mustang balloon was used to perform dilation to 5mm at the RB2.
Total 3 inflations with dilation time of 60 seconds each.
 
The patient tolerated the procedure well.
There were no immediate complications.  At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
EBBX
BAL RLL
 
IMPRESSION/PLAN: [REDACTED] is a 85 year old-year-old female who presents for bronchoscopy for airway narrowing.
- f/u in clinic
- f/u in 12 weeks for bronch evaluation of stent"""

# =============================================================================
# 2. FLAGS & METADATA
# =============================================================================
# Initialize all flags to 0
procedure_flags = {
    # Bronchoscopy (23)
    "diagnostic_bronchoscopy": 0, "bal": 0, "bronchial_wash": 0, "brushings": 0,
    "endobronchial_biopsy": 0, "tbna_conventional": 0, "linear_ebus": 0, "radial_ebus": 0,
    "navigational_bronchoscopy": 0, "transbronchial_biopsy": 0, "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0, "foreign_body_removal": 0, "airway_dilation": 0,
    "airway_stent": 0, "thermal_ablation": 0, "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0, "blvr": 0, "peripheral_ablation": 0, "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0, "rigid_bronchoscopy": 0,
    # Pleural (7)
    "thoracentesis": 0, "chest_tube": 0, "ipc": 0, "medical_thoracoscopy": 0,
    "pleurodesis": 0, "pleural_biopsy": 0, "fibrinolytic_therapy": 0
}

# Set flags based on note content
procedure_flags["diagnostic_bronchoscopy"] = 1 # CPT 31624 listed
procedure_flags["bal"] = 1 # CPT 31624 and text "Bronchial alveolar lavage"
procedure_flags["endobronchial_biopsy"] = 1 # CPT 31625 and text "Endobronchial biopsy"
procedure_flags["therapeutic_aspiration"] = 1 # CPT 31645 and text "Therapeutic aspiration"
procedure_flags["airway_dilation"] = 1 # CPT 31630 and text "Balloon dilation"

# =============================================================================
# 3. SPAN GENERATION (Anchor-First)
# =============================================================================
raw_spans = []

def add_span(text, label, normalized="", schema="", event_id="", prefix="", is_hist=False, is_neg=False, comments=""):
    raw_spans.append({
        "span_text": text,
        "label": label,
        "normalized_value": normalized,
        "schema_field": schema,
        "event_id": event_id,
        "context_prefix": prefix,
        "is_historical": is_hist,
        "is_negated": is_neg,
        "comments": comments
    })

# --- Event 1: Therapeutic Aspiration ---
e1 = "evt01" # Aspiration
add_span("Therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", "method", e1)
add_span("clean out", "PROC_ACTION", "Clean/Clear", "action", e1)
add_span("Right Mainstem", "ANAT_AIRWAY", "Right Mainstem Bronchus", "target.anatomy", e1)
add_span("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "target.anatomy", e1, prefix="Right Mainstem, ")
add_span("Left Mainstem", "ANAT_AIRWAY", "Left Mainstem Bronchus", "target.anatomy", e1)
add_span("mucus", "OBS_LESION", "Mucus", "findings", e1)

# --- Event 2: Endobronchial Biopsy ---
e2 = "evt02" # EBBX
add_span("Endobronchial biopsy", "PROC_METHOD", "Endobronchial Biopsy", "method", e2)
add_span("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "target.anatomy", e2, prefix="Endobronchial biopsy was performed at ")
add_span("Two areas", "MEAS_COUNT", "2", "counts", e2)
add_span("NBI guidance", "PROC_METHOD", "NBI Guidance", "method", e2)

# --- Event 3: Existing Stents (Findings) ---
e3 = "evt03" # Findings (Stents)
add_span("Stents", "DEV_STENT", "Stent", "findings", e3, is_hist=True)
add_span("in good position", "OBS_LESION", "Good Position", "findings", e3, is_hist=True)
add_span("iCAST 7x16", "DEV_STENT", "iCast 7x16", "device", e3, is_hist=True)
add_span("right lower airway", "ANAT_AIRWAY", "Right Lower Lobe", "target.anatomy", e3)
add_span("iCAST 5x16", "DEV_STENT", "iCast 5x16", "device", e3, is_hist=True)
add_span("right middle lobe airway", "ANAT_AIRWAY", "Right Middle Lobe", "target.anatomy", e3)

# --- Event 4: BAL ---
e4 = "evt04" # BAL
add_span("Bronchial alveolar lavage", "PROC_METHOD", "Bronchoalveolar Lavage", "method", e4)
add_span("RLL", "ANAT_LUNG_LOC", "RLL", "target.anatomy", e4)
add_span("40 cc", "MEAS_VOL", "40", "measurements", e4)
add_span("15 cc", "MEAS_VOL", "15", "measurements", e4)

# --- Event 5: Balloon Dilation ---
e5 = "evt05" # Dilation
add_span("Balloon dilation", "PROC_METHOD", "Balloon Dilation", "method", e5)
add_span("RB2 orifice", "ANAT_AIRWAY", "RB2", "target.anatomy", e5)
add_span("5mm mustang balloon", "DEV_INSTRUMENT", "Mustang Balloon 5mm", "device", e5)
add_span("dilation to 5mm", "MEAS_AIRWAY_DIAM", "5", "outcomes.airway.lumen_post", e5)
add_span("3 inflations", "MEAS_COUNT", "3", "counts", e5)
add_span("60 seconds", "MEAS_COUNT", "60", "counts", e5, comments="duration per inflation")

# --- Outcomes ---
e_out = "evt_out"
add_span("No immediate complications", "OUTCOME_COMPLICATION", "None", "outcomes.complications", e_out, is_neg=True)

# =============================================================================
# 4. HYDRATION & OFFSET CALCULATION
# =============================================================================
hydrated_spans = []

def get_offsets(text, span_text, context_prefix):
    # Strategy 1: Exact unique match
    count = text.count(span_text)
    if count == 1:
        start = text.find(span_text)
        return start, start + len(span_text), "hydrated_unique"
    
    # Strategy 2: Context prefix
    if context_prefix and count > 0:
        # Scan all occurrences
        start_idx = 0
        candidates = []
        while True:
            idx = text.find(span_text, start_idx)
            if idx == -1:
                break
            # Check prefix in preceding 150 chars
            window = text[max(0, idx - 150):idx]
            if context_prefix in window:
                candidates.append(idx)
            start_idx = idx + 1
        
        if len(candidates) == 1:
            return candidates[0], candidates[0] + len(span_text), "hydrated_prefix_window"
        if len(candidates) > 1:
            # Fallback to first context match if multiple (risky but deterministic)
            return candidates[0], candidates[0] + len(span_text), "hydrated_prefix_window_multi"

    # Strategy 3: Ambiguous
    return None, None, f"ambiguous_count={count}"

span_id_counter = 1
for s in raw_spans:
    start, end, status = get_offsets(NOTE_TEXT, s["span_text"], s["context_prefix"])
    
    row = {
        "source_file": SOURCE_FILE,
        "note_id": NOTE_ID,
        "span_id": f"span_{span_id_counter:03d}",
        "section_type": "Procedure", # Simplified
        "context_prefix": s["context_prefix"],
        "span_text": s["span_text"],
        "match_index": "", # Not strictly used here, relying on code logic
        "start_char": start if start is not None else "",
        "end_char": end if end is not None else "",
        "span_len": len(s["span_text"]),
        "label": s["label"],
        "normalized_value": s["normalized_value"],
        "schema_field": s["schema_field"],
        "event_id": s["event_id"],
        "is_negated": s["is_negated"],
        "is_historical": s["is_historical"],
        "time_anchor": "",
        "reviewer": "Auto",
        "comments": s["comments"],
        "hydration_status": status
    }
    hydrated_spans.append(row)
    span_id_counter += 1

# =============================================================================
# 5. GENERATE EVENTS & REGISTRY JSON
# =============================================================================
# V3 Event Mapping
v3_events = []

# Event 1: Aspiration
v3_events.append({
    "note_id": NOTE_ID, "event_id": "evt01", "type": "Therapeutic Aspiration",
    "target": {"anatomy_type": "Airway", "location": {"lobe": "", "segment": ""}, "station": "RMS, BI, LMS"},
    "lesion": {"type": "Mucus", "size_mm": ""},
    "method": "Therapeutic Aspiration",
    "devices_json": json.dumps([]),
    "measurements_json": json.dumps([]),
    "specimens_json": json.dumps([]),
    "findings_json": json.dumps(["Cleaned out mucus"]),
    "evidence_quote": "Successful therapeutic aspiration was performed to clean out the Right Mainstem, Bronchus Intermedius , and Left Mainstem from mucus.",
    "stent": {"size": "", "material_or_brand": ""},
    "catheter": {"size_fr": ""},
    "outcomes": {"airway": {"lumen_pre": "", "lumen_post": ""}, "symptoms": "", "pleural": "", "complications": ""}
})

# Event 2: Biopsy
v3_events.append({
    "note_id": NOTE_ID, "event_id": "evt02", "type": "Endobronchial Biopsy",
    "target": {"anatomy_type": "Airway", "location": {"lobe": "", "segment": ""}, "station": "BI"},
    "lesion": {"type": "", "size_mm": ""},
    "method": "Endobronchial Biopsy",
    "devices_json": json.dumps(["NBI"]),
    "measurements_json": json.dumps({"count": 2}),
    "specimens_json": json.dumps(["EBBX"]),
    "findings_json": json.dumps([]),
    "evidence_quote": "Endobronchial biopsy was performed at Bronchus Intermedius .  Two areas were biopsied with NBI guidance.",
    "stent": {"size": "", "material_or_brand": ""},
    "catheter": {"size_fr": ""},
    "outcomes": {"airway": {"lumen_pre": "", "lumen_post": ""}, "symptoms": "", "pleural": "", "complications": ""}
})

# Event 3: Stent Findings (Historical)
v3_events.append({
    "note_id": NOTE_ID, "event_id": "evt03", "type": "Inspection",
    "target": {"anatomy_type": "Airway", "location": {"lobe": "RLL/RML", "segment": ""}, "station": ""},
    "lesion": {"type": "", "size_mm": ""},
    "method": "Inspection",
    "devices_json": json.dumps(["iCAST 7x16", "iCAST 5x16"]),
    "measurements_json": json.dumps([]),
    "specimens_json": json.dumps([]),
    "findings_json": json.dumps(["Stents in good position"]),
    "evidence_quote": "Stents are in good position. The (iCAST 7x16) was noted in the right lower airway.  (iCAST 5x16) was noted in the right middle lobe airway.",
    "stent": {"size": "7x16, 5x16", "material_or_brand": "iCAST"},
    "catheter": {"size_fr": ""},
    "outcomes": {"airway": {"lumen_pre": "", "lumen_post": ""}, "symptoms": "", "pleural": "", "complications": ""}
})

# Event 4: BAL
v3_events.append({
    "note_id": NOTE_ID, "event_id": "evt04", "type": "BAL",
    "target": {"anatomy_type": "Lung", "location": {"lobe": "RLL", "segment": ""}, "station": ""},
    "lesion": {"type": "", "size_mm": ""},
    "method": "BAL",
    "devices_json": json.dumps([]),
    "measurements_json": json.dumps({"instilled": "40 cc", "return": "15 cc"}),
    "specimens_json": json.dumps(["BAL RLL"]),
    "findings_json": json.dumps([]),
    "evidence_quote": "Bronchial alveolar lavage was performed at RLL.",
    "stent": {"size": "", "material_or_brand": ""},
    "catheter": {"size_fr": ""},
    "outcomes": {"airway": {"lumen_pre": "", "lumen_post": ""}, "symptoms": "", "pleural": "", "complications": ""}
})

# Event 5: Balloon Dilation
v3_events.append({
    "note_id": NOTE_ID, "event_id": "evt05", "type": "Airway Dilation",
    "target": {"anatomy_type": "Airway", "location": {"lobe": "RUL", "segment": "RB2"}, "station": ""},
    "lesion": {"type": "Stenosis", "size_mm": ""},
    "method": "Balloon Dilation",
    "devices_json": json.dumps(["5mm mustang balloon"]),
    "measurements_json": json.dumps({"inflations": 3, "duration": "60 sec"}),
    "specimens_json": json.dumps([]),
    "findings_json": json.dumps([]),
    "evidence_quote": "Balloon dilation was performed at RB2 orifice... 5mm mustang balloon was used to perform dilation to 5mm",
    "stent": {"size": "", "material_or_brand": ""},
    "catheter": {"size_fr": ""},
    "outcomes": {"airway": {"lumen_pre": "", "lumen_post": "5mm"}, "symptoms": "", "pleural": "", "complications": "None"}
})

registry_json = {
    "schema_version": "v3.0",
    "note_id": NOTE_ID,
    "procedures": v3_events,
    "no_immediate_complications": True
}

# =============================================================================
# 6. EXCEL GENERATION
# =============================================================================
def generate_excel(template_path: str = TEMPLATE_PATH, output_path: str = OUTPUT_PATH) -> None:
    try:
        wb = openpyxl.load_workbook(template_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()  # Fallback for standalone run

    def append_row(sheet_name, row_data, headers=None):
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            if headers:
                ws.append(headers)
        else:
            ws = wb[sheet_name]
        ws.append(row_data)

    append_row("Note_Text", [NOTE_ID, SOURCE_FILE, NOTE_TEXT], ["note_id", "source_file", "note_text"])

    flag_keys = list(procedure_flags.keys())
    index_headers = ["source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes"] + flag_keys
    index_row = [SOURCE_FILE, NOTE_ID, "", "", "", "", "Ready", ""] + [procedure_flags[k] for k in flag_keys]
    append_row("Note_Index", index_row, index_headers)

    span_headers = ["source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text",
                    "match_index", "start_char", "end_char", "span_len", "label", "normalized_value",
                    "schema_field", "event_id", "is_negated", "is_historical", "time_anchor",
                    "reviewer", "comments", "hydration_status"]

    for s in hydrated_spans:
        row = [s["source_file"], s["note_id"], s["span_id"], s["section_type"], s["context_prefix"], s["span_text"],
               s["match_index"], "", "", s["span_len"], s["label"], s["normalized_value"],
               s["schema_field"], s["event_id"], s["is_negated"], s["is_historical"], s["time_anchor"],
               s["reviewer"], s["comments"], "needs_hydration"]
        append_row("Span_Annotations", row, span_headers)

    for s in hydrated_spans:
        row = [s["source_file"], s["note_id"], s["span_id"], s["section_type"], s["context_prefix"], s["span_text"],
               s["match_index"], s["start_char"], s["end_char"], s["span_len"], s["label"], s["normalized_value"],
               s["schema_field"], s["event_id"], s["is_negated"], s["is_historical"], s["time_anchor"],
               s["reviewer"], s["comments"], s["hydration_status"]]
        append_row("Span_Hydrated", row, span_headers)

    event_log_headers = ["source_file", "note_id", "event_id", "event_type", "method", "anatomy_target", "device",
                         "needle_gauge", "stations", "counts", "measurements", "specimens", "findings",
                         "is_historical", "reviewer", "comments", "device_size", "device_material",
                         "outcome_airway_lumen_pre", "outcome_airway_lumen_post", "outcome_symptoms",
                         "outcome_pleural", "outcome_complication"]

    for evt in v3_events:
        row = [
            SOURCE_FILE, NOTE_ID, evt["event_id"], evt["type"], evt["method"],
            f"{evt['target']['anatomy_type']} {evt['target']['location']} {evt['target']['station']}",
            evt["devices_json"], "", "", "", evt["measurements_json"], evt["specimens_json"], evt["findings_json"],
            False, "Auto", "", evt["stent"]["size"], evt["stent"]["material_or_brand"],
            evt["outcomes"]["airway"]["lumen_pre"], evt["outcomes"]["airway"]["lumen_post"],
            evt["outcomes"]["symptoms"], evt["outcomes"]["pleural"], evt["outcomes"]["complications"]
        ]
        append_row("Event_Log", row, event_log_headers)

    v3_headers = ["note_id", "event_id", "type", "target.anatomy_type", "target.location.lobe",
                  "target.location.segment", "target.station", "lesion.type", "lesion.size_mm",
                  "method", "devices_json", "measurements_json", "specimens_json", "findings_json",
                  "evidence_quote", "stent.size", "stent.material_or_brand", "catheter.size_fr",
                  "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post", "outcomes.symptoms",
                  "outcomes.pleural", "outcomes.complications"]

    for evt in v3_events:
        row = [
            evt["note_id"], evt["event_id"], evt["type"], evt["target"]["anatomy_type"],
            evt["target"]["location"]["lobe"], evt["target"]["location"]["segment"], evt["target"]["station"],
            evt["lesion"]["type"], evt["lesion"]["size_mm"], evt["method"], evt["devices_json"],
            evt["measurements_json"], evt["specimens_json"], evt["findings_json"], evt["evidence_quote"],
            evt["stent"]["size"], evt["stent"]["material_or_brand"], evt["catheter"]["size_fr"],
            evt["outcomes"]["airway"]["lumen_pre"], evt["outcomes"]["airway"]["lumen_post"],
            evt["outcomes"]["symptoms"], evt["outcomes"]["pleural"], evt["outcomes"]["complications"]
        ]
        append_row("V3_Procedure_Events", row, v3_headers)

    if "V3_Registry_JSON" not in wb.sheetnames:
        wb.create_sheet("V3_Registry_JSON")
    ws_json = wb["V3_Registry_JSON"]
    ws_json.append(["note_id", "json_content"])
    ws_json.append([NOTE_ID, json.dumps(registry_json, indent=2)])

    wb.save(output_path)
    print(f"Generated {output_path}")


if __name__ == "__main__":
    generate_excel()
