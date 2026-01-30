import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import datetime
import re
import json
import sys

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_072"
SOURCE_FILE = "note_072.txt"
PROCEDURE_DATE = "2026-01-12" # Using current context date as placeholder or blank if preferred
NOTE_TEXT = """NOTE_ID:  note_072 SOURCE_FILE: note_072.txt INDICATION FOR OPERATION:  [REDACTED] is a 49 year old-year-old adult who presents with RPP.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
 
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
 
1899 Unlisted Procedure (Trach Change with Mature Tract or Procedure NOS)
31645 Therapeutic aspiration initial episode
31625 Endobronchial Biopsy(s)
31630 Balloon dilation
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
 
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required) and XS Separate Structure, a service that is distinct because it was performed on a separate organ/structure
 
IP [REDACTED] 
CODE MOD DETAILS: 
Unusual Procedure (22 MODIFIER):
This patient required multiple forms of ablative therapy (soft coag and cryotherapy) to multiple different structures, including glottis, infraglottis, subglottis, and tracheal structures.
This resulted in >50% increased work due to Time, Technical difficulty of procedure, Severity of patient's condition, and Physical and mental effort required.
Apply to: 
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
 
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Rigid Tracheoscope
Flexible Therapeutic Bronchoscope
 
PROCEDURE IN DETAIL:
A timeout was performed (confirming the patient's name, procedure type, and procedure location).
Sedation initiated.  LMA placed.
 
The Flexible Therapeutic Bronchoscope was advanced for airway examination.
Endobronchial topical lidocaine applied to the vocal cords and larynx.
 
Initial Airway Inspection Findings:
Vocal cords with appropriate adduction/abduction.
Large patch of papillomas at anterior commissure of larynx and right vocal cord.
Proximal trachea with focal area of spiraling tissue creating a focal area of mild stenosis - about 20% stenosis prior to treatment.
Able to easily traverse area of mild stenosis with therapeutic bronchoscope without issue.
Assessment of right airways revealed normal anatomic branching to first subsegmental level.
Otherwise no evidence of mass, lesions, bleeding or other endobronchial pathology.
Assessment of left airways revealed normal anatomic branching to first subsegmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology  
 
VC papilloma are treated with combination of cryospray, electro-cautery (PreciSect/softcoag), and mechanically debulking.
Endobronchial papillomatous tumors at proximal trachea, subglottic region, infraglottic region, left vocal fold, and anterior commissure were treated with the following modalities.
The base of all patch sites were treated with cryotherapy.
Modality	Tools	Setting/Mode	Duration	Results
Cryospray	2.3 mm catheter	Normal flow, Freeze	5 sec burst	Tumor frozen, 5 sites 3-5 treatment each
 
 
Endobronchial tumor and coagulated debris was noted and excised with mechanical debridement using alligator forceps.
Send to path and microbiology.
 
Rigid tracheoscope was used to dilate the proximal tracheal obstruction.  Mechanically coring through the tumor.
Black non-ventilating scope was used to perform dilation to 12 mm at the Trachea (Proximal 1/3).
Total 2 inflations with dilation time of 300 seconds each.
Cryospray was done at the tumor site proximal trachea for 2 different sites.
Total 3-5 treatments with 5 seconds of freeze at Normal flow was completed for both sides.
Prior to treatment, proximal trachea airway was note to be 40% patent.  After treatment, the airway was 90% patent.
Throughout the procedure, careful attention was paid to suction any secretions and blood to prevent it falling into the distal trachea or more distal airways.
Residual secretions and blood were suctioned to clear.  Video imaging obtained and saved.  Bronchoscope removed.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient had airway device removed in the operating room and transported to the recovery room in stable condition.
ESTIMATED BLOOD LOSS:   Trivial
COMPLICATIONS:    None
 
SPECIMEN(S): 
Trachea biopsy:  Pathology
 
 
IMPRESSION/PLAN: [REDACTED] is a 49 year old-year-old adult who presents for bronchoscopy for blation of papillomatous tumors in the context of recurrent respiratory papillomatosis.
-Recommend repeat bronchoscopy in 4-6 weeks to reassess recurrence of papillomas and consideration of additional treatment."""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# -------------------------------------------------------------------------
# CONFIGURATION
# -------------------------------------------------------------------------

PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 1,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 0,
    "airway_dilation": 1,
    "airway_stent": 0,
    "thermal_ablation": 1,
    "tumor_debulking_non_thermal": 1,
    "cryotherapy": 1,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 1,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0,
}

# -------------------------------------------------------------------------
# SPAN DATA GENERATION
# -------------------------------------------------------------------------

def create_spans():
    """
    Returns a list of dictionaries representing spans.
    Each dict must have:
    span_text, context_prefix, section_type, label, normalized_value, schema_field, event_id, etc.
    """
    spans = []

    def add_span(text, label, norm, field, event_id, section="PROCEDURE", prefix="", is_hist=False, is_neg=False, comments=""):
        spans.append({
            "span_text": text,
            "context_prefix": prefix,
            "section_type": section,
            "label": label,
            "normalized_value": norm,
            "schema_field": field,
            "event_id": event_id,
            "is_historical": is_hist,
            "is_negated": is_neg,
            "comments": comments,
            "reviewer": "AI_EXTRACTOR",
            "match_index": 0 # Default, logic handles collision later
        })

    # --- Header / Diagnosis ---
    add_span("RPP", "OBS_LESION", "Recurrent Respiratory Papillomatosis", "indication", "evt0", section="HEADER")
    add_span("J98.09", "CTX_HISTORICAL", "J98.09", "icd_code", "evt0", section="HEADER")
    add_span("Therapeutic aspiration", "PROC_METHOD", "therapeutic_aspiration", "procedure_type", "evt0", section="HEADER")
    add_span("Endobronchial Biopsy", "PROC_METHOD", "endobronchial_biopsy", "procedure_type", "evt0", section="HEADER")
    add_span("Balloon dilation", "PROC_METHOD", "balloon_dilation", "procedure_type", "evt0", section="HEADER")
    add_span("Destruction of tumor OR relief of stenosis", "PROC_METHOD", "destruction_of_tumor", "procedure_type", "evt0", section="HEADER")

    # --- Anesthesia / Setup ---
    add_span("General Anesthesia", "PROC_METHOD", "General Anesthesia", "anesthesia_type", "evt1", section="ANESTHESIA")
    add_span("LMA placed", "PROC_ACTION", "LMA placement", "airway_management", "evt1", section="PROCEDURE")
    
    # --- Instruments ---
    add_span("Rigid Tracheoscope", "DEV_INSTRUMENT", "Rigid Tracheoscope", "device_type", "evt2", section="INSTRUMENT")
    add_span("Flexible Therapeutic Bronchoscopy", "DEV_INSTRUMENT", "Flexible Bronchoscope", "device_type", "evt2", section="INSTRUMENT")

    # --- Initial Inspection (Flexible) ---
    add_span("Flexible Therapeutic Bronchoscopy was advanced", "PROC_ACTION", "Bronchoscopy", "method", "evt2", section="PROCEDURE")
    add_span("Large patch of papillomas", "OBS_LESION", "Papilloma", "lesion_type", "evt2", section="PROCEDURE")
    add_span("anterior commissure of larynx", "ANAT_AIRWAY", "Anterior Commissure", "anatomy_location", "evt2", section="PROCEDURE")
    add_span("right vocal cord", "ANAT_AIRWAY", "Right Vocal Cord", "anatomy_location", "evt2", section="PROCEDURE", prefix="commissure of larynx and ")

    # Trachea inspection
    add_span("Proximal trachea", "ANAT_AIRWAY", "Proximal Trachea", "anatomy_location", "evt2", section="PROCEDURE", prefix="patch of papillomas at ") # Need prefix to distinguish
    add_span("mild stenosis", "OBS_LESION", "Stenosis", "lesion_type", "evt2", section="PROCEDURE", prefix="focal area of ")
    add_span("about 20% stenosis prior to treatment", "OUTCOME_AIRWAY_LUMEN_PRE", "80%", "airway_lumen_pre", "evt2", section="PROCEDURE")

    # Normal Anatomy Negations
    add_span("right airways", "ANAT_AIRWAY", "Right Airways", "anatomy_location", "evt3", section="PROCEDURE", prefix="Assessment of ")
    add_span("no evidence of mass, lesions, bleeding", "OBS_LESION", "None", "findings", "evt3", section="PROCEDURE", is_neg=True, prefix="Otherwise ")
    add_span("left airways", "ANAT_AIRWAY", "Left Airways", "anatomy_location", "evt4", section="PROCEDURE", prefix="Assessment of ")
    add_span("No evidence of mass, lesions, bleeding", "OBS_LESION", "None", "findings", "evt4", section="PROCEDURE", is_neg=True, prefix="subsegmental level.\n")

    # --- Ablative Therapy (Complex) ---
    add_span("cryospray", "PROC_METHOD", "Cryospray", "method", "evt5", section="PROCEDURE", prefix="combination of ")
    add_span("electro-cautery (PreciSect/softcoag)", "PROC_METHOD", "Electro-cautery", "method", "evt5", section="PROCEDURE")
    add_span("mechanically debulking", "PROC_METHOD", "Mechanical Debulking", "method", "evt5", section="PROCEDURE")
    
    # Target locations for ablation
    add_span("Endobronchial papillomatous tumors", "OBS_LESION", "Papilloma", "lesion_type", "evt5", section="PROCEDURE")
    add_span("proximal trachea", "ANAT_AIRWAY", "Proximal Trachea", "anatomy_location", "evt5", section="PROCEDURE", prefix="papillomatous tumors at ")
    add_span("subglottic region", "ANAT_AIRWAY", "Subglottis", "anatomy_location", "evt5", section="PROCEDURE")
    add_span("infraglottic region", "ANAT_AIRWAY", "Infraglottis", "anatomy_location", "evt5", section="PROCEDURE")
    add_span("left vocal fold", "ANAT_AIRWAY", "Left Vocal Cord", "anatomy_location", "evt5", section="PROCEDURE")
    add_span("anterior commissure", "ANAT_AIRWAY", "Anterior Commissure", "anatomy_location", "evt5", section="PROCEDURE", prefix="left vocal fold, and ")

    # Specifics of Cryo
    add_span("2.3 mm catheter", "DEV_CATHETER_SIZE", "2.3mm", "device_size", "evt5", section="PROCEDURE")
    add_span("5 sec burst", "MEAS_TIME", "5 sec", "duration", "evt5", section="PROCEDURE")
    add_span("5 sites", "MEAS_COUNT", "5", "count", "evt5", section="PROCEDURE")

    # Mechanical excision
    add_span("mechanical debridement", "PROC_METHOD", "Mechanical Debridement", "method", "evt5", section="PROCEDURE", prefix="excised with ")
    add_span("alligator forceps", "DEV_INSTRUMENT", "Alligator Forceps", "device_type", "evt5", section="PROCEDURE")
    add_span("Send to path", "PROC_ACTION", "Pathology", "specimen_handling", "evt5", section="PROCEDURE")

    # --- Rigid / Dilation ---
    add_span("Rigid tracheoscope", "DEV_INSTRUMENT", "Rigid Tracheoscope", "device_type", "evt6", section="PROCEDURE", prefix="microbiology.\n\n")
    add_span("dilate the proximal tracheal obstruction", "PROC_ACTION", "Dilation", "method", "evt6", section="PROCEDURE")
    add_span("proximal tracheal obstruction", "OBS_LESION", "Obstruction", "lesion_type", "evt6", section="PROCEDURE", prefix="dilate the ")
    add_span("Mechanically coring", "PROC_METHOD", "Coring", "method", "evt6", section="PROCEDURE")
    
    # Dilation specifics
    add_span("Black non-ventilating scope", "DEV_INSTRUMENT", "Rigid Scope", "device_type", "evt6", section="PROCEDURE")
    add_span("dilation to 12 mm", "MEAS_AIRWAY_DIAM", "12 mm", "target_diameter", "evt6", section="PROCEDURE")
    add_span("Trachea (Proximal 1/3)", "ANAT_AIRWAY", "Proximal Trachea", "anatomy_location", "evt6", section="PROCEDURE", prefix="12 mm at the ")
    add_span("Total 2 inflations", "MEAS_COUNT", "2", "count", "evt6", section="PROCEDURE")
    add_span("300 seconds each", "MEAS_TIME", "300 sec", "duration", "evt6", section="PROCEDURE")

    # --- More Cryo (Proximal Trachea specific) ---
    add_span("Cryospray was done", "PROC_METHOD", "Cryospray", "method", "evt7", section="PROCEDURE", prefix="seconds each.\n")
    add_span("proximal trachea", "ANAT_AIRWAY", "Proximal Trachea", "anatomy_location", "evt7", section="PROCEDURE", prefix="tumor site ")
    add_span("2 different sites", "MEAS_COUNT", "2", "count", "evt7", section="PROCEDURE")
    add_span("Total 3-5 treatments", "MEAS_COUNT", "3-5", "count", "evt7", section="PROCEDURE")

    # --- Outcomes ---
    # "Prior to treatment, proximal trachea airway was note to be 40% patent." 
    # NOTE: Earlier it said 20% stenosis (80% patent). Now 40% patent. Capturing both, assigning this one to the rigid event context.
    add_span("40% patent", "OUTCOME_AIRWAY_LUMEN_PRE", "40%", "airway_lumen_pre", "evt6", section="PROCEDURE") 
    add_span("After treatment, the airway was 90% patent", "OUTCOME_AIRWAY_LUMEN_POST", "90%", "airway_lumen_post", "evt6", section="PROCEDURE")
    
    add_span("suction any secretions", "PROC_ACTION", "Suction", "method", "evt8", section="PROCEDURE")
    add_span("Trachea biopsy", "PROC_ACTION", "Biopsy", "method", "evt9", section="SPECIMEN")
    
    # Complications
    add_span("no immediate complications", "OUTCOME_COMPLICATION", "None", "complication", "evt99", section="PROCEDURE", is_neg=True)
    add_span("COMPLICATIONS:    None", "OUTCOME_COMPLICATION", "None", "complication", "evt99", section="COMPLICATIONS", is_neg=True)

    return spans

# -------------------------------------------------------------------------
# EVENT LOG DATA GENERATION
# -------------------------------------------------------------------------

def create_events():
    """
    Returns list of event dicts.
    """
    events = []
    
    # Event 1: Anesthesia
    events.append({
        "event_id": "evt1",
        "event_type": "Anesthesia",
        "method": "General Anesthesia",
        "comments": "LMA placed"
    })
    
    # Event 2: Diagnostic Bronch (Flexible)
    events.append({
        "event_id": "evt2",
        "event_type": "Diagnostic Bronchoscopy",
        "method": "Flexible Bronchoscopy",
        "anatomy_target": "Larynx, Trachea",
        "findings": "Papillomas at anterior commissure, vocal cord, proximal trachea (20% stenosis)",
        "outcome_airway_lumen_pre": "80%" # inferred from 20% stenosis
    })
    
    # Event 5: Ablation (Multi-modal)
    events.append({
        "event_id": "evt5",
        "event_type": "Therapeutic Bronchoscopy",
        "method": "Cryospray, Electro-cautery, Mechanical Debulking",
        "anatomy_target": "Proximal Trachea, Subglottis, Vocal Folds",
        "device": "2.3 mm catheter, Alligator forceps",
        "counts": "5 sites",
        "specimens": "Endobronchial tumor/debris sent to path"
    })

    # Event 6: Rigid Dilation
    events.append({
        "event_id": "evt6",
        "event_type": "Airway Dilation",
        "method": "Rigid Dilation (Coring)",
        "anatomy_target": "Proximal Trachea",
        "device": "Rigid Tracheoscope, Black non-ventilating scope",
        "measurements": "Dilated to 12mm, 2 inflations @ 300sec",
        "outcome_airway_lumen_pre": "40%",
        "outcome_airway_lumen_post": "90%"
    })

    # Event 7: Additional Cryo
    events.append({
        "event_id": "evt7",
        "event_type": "Cryotherapy",
        "method": "Cryospray",
        "anatomy_target": "Proximal Trachea",
        "counts": "2 sites, 3-5 treatments"
    })

    # Event 99: Complications
    events.append({
        "event_id": "evt99",
        "event_type": "Outcome",
        "outcome_complication": "None"
    })

    return events

# -------------------------------------------------------------------------
# HELPER CLASSES & LOGIC
# -------------------------------------------------------------------------

class TextHydrator:
    def __init__(self, full_text):
        self.full_text = full_text.replace('\r\n', '\n') # Normalize line endings

    def hydrate_span(self, span_data):
        target = span_data['span_text']
        prefix = span_data.get('context_prefix', '')
        match_idx = span_data.get('match_index', 0)
        
        # 1. Unique Match
        count = self.full_text.count(target)
        if count == 1:
            start = self.full_text.find(target)
            return start, start + len(target), "hydrated_unique"
        
        # 2. Context Match (Window 120 chars)
        if prefix:
            # Find all occurrences
            starts = [m.start() for m in re.finditer(re.escape(target), self.full_text)]
            best_start = -1
            for s in starts:
                # Look back 120 chars
                window_start = max(0, s - 120)
                window = self.full_text[window_start:s]
                if prefix in window:
                    best_start = s
                    break
            
            if best_start != -1:
                return best_start, best_start + len(target), "hydrated_prefix_window"

        # 3. Match Index Fallback
        if count > 1:
            starts = [m.start() for m in re.finditer(re.escape(target), self.full_text)]
            if 0 <= match_idx < len(starts):
                s = starts[match_idx]
                return s, s + len(target), "hydrated_match_index"

        return None, None, f"ambiguous_count={count}"

class Phase0WorkbookGenerator:
    def __init__(self, template_path, output_path):
        self.template_path = template_path
        self.output_path = output_path
        self.wb = None

    def load_template(self):
        try:
            self.wb = openpyxl.load_workbook(self.template_path)
        except FileNotFoundError:
            # Create fresh if missing (fallback for standalone runs)
            self.wb = openpyxl.Workbook()
            # Create standard sheets
            for name in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
                if name not in self.wb.sheetnames:
                    self.wb.create_sheet(name)

    def write_note_text(self):
        ws = self.wb["Note_Text"]
        # Ensure headers
        if ws.max_row == 1 and ws.cell(1, 1).value is None:
            ws.append(["note_id", "source_file", "note_text"])
        
        ws.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    def write_note_index(self):
        ws = self.wb["Note_Index"]
        # Headers based on requirements
        headers = ["source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes"]
        flag_keys = list(PROCEDURE_FLAGS.keys())
        headers.extend(flag_keys)
        
        if ws.max_row == 1 and ws.cell(1, 1).value is None:
            ws.append(headers)
        
        row_data = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Check", ""]
        row_data.extend([PROCEDURE_FLAGS[k] for k in flag_keys])
        ws.append(row_data)

    def write_spans(self):
        spans = create_spans()
        hydrator = TextHydrator(NOTE_TEXT)
        
        ws_anno = self.wb["Span_Annotations"]
        ws_hydr = self.wb["Span_Hydrated"]
        
        # Headers
        headers = ["source_file", "note_id", "span_id", "section_type", 
                   "context_prefix", "span_text", "match_index", 
                   "start_char", "end_char", "span_len", 
                   "label", "normalized_value", "schema_field", "event_id", 
                   "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"]
        
        if ws_anno.max_row == 1 and ws_anno.cell(1, 1).value is None:
            ws_anno.append(headers)
        if ws_hydr.max_row == 1 and ws_hydr.cell(1, 1).value is None:
            ws_hydr.append(headers)

        for idx, s in enumerate(spans):
            span_id = f"{NOTE_ID}_s{idx+1:03d}"
            
            # Annotation Row (Anchor First)
            anno_row = [
                SOURCE_FILE, NOTE_ID, span_id, s['section_type'],
                s['context_prefix'], s['span_text'], s.get('match_index', 0),
                "", "", f"=LEN(F{ws_anno.max_row+1})", # Formulas not evaluated but written
                s['label'], s['normalized_value'], s['schema_field'], s['event_id'],
                s['is_negated'], s['is_historical'], False, s['reviewer'], s['comments'], "needs_hydration"
            ]
            ws_anno.append(anno_row)
            
            # Hydration
            start, end, status = hydrator.hydrate_span(s)
            hydr_row = list(anno_row)
            hydr_row[7] = start
            hydr_row[8] = end
            hydr_row[9] = (end - start) if start is not None else 0
            hydr_row[19] = status
            ws_hydr.append(hydr_row)

    def write_events(self):
        ws = self.wb["Event_Log"]
        headers = ["source_file", "note_id", "event_id", "event_type", "method",
                   "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements",
                   "specimens", "findings", "is_historical", "reviewer", "comments",
                   "device_size", "device_material",
                   "outcome_airway_lumen_pre", "outcome_airway_lumen_post",
                   "outcome_symptoms", "outcome_pleural", "outcome_complication"]
        
        if ws.max_row == 1 and ws.cell(1, 1).value is None:
            ws.append(headers)
            
        events = create_events()
        for e in events:
            row = [
                SOURCE_FILE, NOTE_ID, e.get('event_id'), e.get('event_type'), e.get('method'),
                e.get('anatomy_target'), e.get('device'), e.get('needle_gauge'), e.get('stations'), e.get('counts'), e.get('measurements'),
                e.get('specimens'), e.get('findings'), False, "AI", e.get('comments'),
                e.get('device_size'), e.get('device_material'),
                e.get('outcome_airway_lumen_pre'), e.get('outcome_airway_lumen_post'),
                e.get('outcome_symptoms'), e.get('outcome_pleural'), e.get('outcome_complication')
            ]
            ws.append(row)

    def write_v3_registry(self):
        # Placeholder for V3 JSON and Events logic
        # For this exercise, we essentially dump the event log into JSON structure
        ws_json = self.wb["V3_Registry_JSON"]
        if ws_json.max_row == 1:
            ws_json.append(["schema_version", "note_id", "json_output"])
        
        registry_data = {
            "schema_version": "3.0",
            "note_id": NOTE_ID,
            "no_immediate_complications": True, # Based on extraction
            "procedures": create_events()
        }
        
        json_str = json.dumps(registry_data, indent=2)
        ws_json.append(["3.0", NOTE_ID, json_str])

    def save(self):
        self.wb.save(self.output_path)
        print(f"Generated {self.output_path}")

# -------------------------------------------------------------------------
# EXECUTION
# -------------------------------------------------------------------------
if __name__ == "__main__":
    generator = Phase0WorkbookGenerator(TEMPLATE_PATH, OUTPUT_PATH)
    generator.load_template()
    generator.write_note_text()
    generator.write_note_index()
    generator.write_spans()
    generator.write_events()
    generator.write_v3_registry()
    generator.save()