import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_073"
SOURCE_FILE = "note_073.txt"
PROCEDURE_DATE = "2026-01-12" # Using current date as placeholder or extraction if available, note implies recent.
NOTE_TEXT = """NOTE_ID:  note_073 SOURCE_FILE: note_073.txt INDICATION FOR OPERATION:  [REDACTED]is a 29 year old-year-old male who presents with tracheal stenosis.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
 
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
 
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31623 Dx bronchoscope/brushing    
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
 
 
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
 
ESTIMATED BLOOD LOSS:   Minimum
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
 
The laryngeal mask airway is in good position.
Pharynx: Not assessed due to bronchoscopy introduction through LMA.
Larynx: Not fully assessed due to bronchoscopy introduction through LMA.
Vocal Cords: Normal without mass/lesions
Trachea: Tracheal stent in place with slight granulation tissue and the anterior and posterior aspect of the proximal stent.
Mild thick yellow mucus within the stent
Main Carina: Sharp
Right Lung Proximal Airways: Normal anatomic branching to segmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Left Lung Proximal Airways: Normal anatomic branching to segmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Mucosa: Normal.
Secretions: As described above.
GRANULATION TISSUE AT THE PROXIMAL STENT
 
 
 
 
Successful therapeutic aspiration was performed to clean out the Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus.
The 2.4mm cryoprobe was used to treat the anterior granulation tissue with 30 second freeze-thaw cycles (total of 4 cycles).
The granulation tissue at the posterior trachea was treated with the 2.4mm cryoprobe in which tissue was frozen to the probe and removed en bloc with the bronchoscope.
In sum, the granulation tissue at Subglottic was treated with the following modalities:
 
Modality	Tools	Setting/Mode	Duration	Results
Cryoprobe	2.4mm Cryoprobe	N/A	30 second freeze/thaw cycles	Excellent tissue destruction
 
Mild oozing/bleeding was noted and treated with cold saline.
Lastly, mucus/secretions were removed from within the stent using a combination of saline flushes and cytology brush.
AREAS OF GRANULATION  TISSUE TREATMENT AT THE CONCLUSION OF THE PROCEDURE
 
 
STENT AT THE CONCLUSION OF THE PROCEDURE
 
 
 
The patient tolerated the procedure well.
There were no immediate complications.  At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
--Tracheal tumor/granulation tissue removed with cryoprobe
 
IMPRESSION/PLAN: [REDACTED]is a 29 year old-year-old male who presents for bronchoscopy for evaluation of tracheal stenosis.
Granulation tissue at the anterior and posterior aspects of the proximal tracheal stent were treated with cryoprobe with excellent tissue debulking.
--Post procedure CXR
--Follow up path results
--Patient to follow up in outpatient IP clinic to go over results and next steps"""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# -------------------------------------------------------------------------
# CONFIGURATION
# -------------------------------------------------------------------------

PROCEDURE_FLAGS = {
    # Bronchoscopy (23)
    "diagnostic_bronchoscopy": 1, # "Dx bronchoscope"
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 1, # "Dx bronchoscope/brushing", "cytology brush"
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1, # "Therapeutic aspiration initial episode"
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0, # Stent present, but not placed/removed
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0, 
    "cryotherapy": 1, # "cryotherapy", "cryoprobe"
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural (7)
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# -------------------------------------------------------------------------
# SPAN GENERATION
# -------------------------------------------------------------------------

# Helper to build spans compactly
def create_span(text, label, norm_val, field, event_id, context_prefix="", is_hist=False, comments=""):
    return {
        "source_file": SOURCE_FILE,
        "note_id": NOTE_ID,
        "span_id": f"{NOTE_ID}_s{len(SPANS) + 1:03d}",
        "section_type": "Procedure", # Simplified
        "context_prefix": context_prefix,
        "span_text": text,
        "match_index": 0, # Default to first if not specified
        "start_char": None,
        "end_char": None,
        "span_len": len(text),
        "label": label,
        "normalized_value": norm_val,
        "schema_field": field,
        "event_id": event_id,
        "is_negated": False,
        "is_historical": is_hist,
        "time_anchor": "",
        "reviewer": "Auto",
        "comments": comments,
        "hydration_status": "needs_hydration"
    }

SPANS = []

# Event 1: Therapeutic Aspiration (General airway clearance)
ev1 = "evt_01"
SPANS.append(create_span("Therapeutic aspiration", "PROC_METHOD", "Therapeutic aspiration", "method", ev1))
SPANS.append(create_span("clean out", "PROC_ACTION", "clean out", "action", ev1))
SPANS.append(create_span("Trachea (Distal 1/3)", "ANAT_AIRWAY", "Trachea", "target.anatomy_type", ev1))
SPANS.append(create_span("Right Mainstem", "ANAT_AIRWAY", "RMS", "target.anatomy_type", ev1))
SPANS.append(create_span("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "target.anatomy_type", ev1))
SPANS.append(create_span("Left Mainstem", "ANAT_AIRWAY", "LMS", "target.anatomy_type", ev1))
SPANS.append(create_span("Carina", "ANAT_AIRWAY", "Carina", "target.anatomy_type", ev1, context_prefix="Left Mainstem, "))

# Event 2: Cryotherapy (Anterior Granulation)
ev2 = "evt_02"
SPANS.append(create_span("2.4mm cryoprobe", "DEV_INSTRUMENT", "Cryoprobe 2.4mm", "devices_json", ev2))
SPANS.append(create_span("anterior granulation tissue", "OBS_LESION", "Granulation tissue (Anterior)", "lesion.type", ev2))
SPANS.append(create_span("30 second", "MEAS_COUNT", "30", "measurements_json", ev2)) # Duration
SPANS.append(create_span("4 cycles", "MEAS_COUNT", "4", "measurements_json", ev2))
SPANS.append(create_span("Excellent tissue destruction", "OBS_ROSE", "Excellent tissue destruction", "findings_json", ev2))

# Event 3: Cryotherapy (Posterior Granulation)
ev3 = "evt_03"
SPANS.append(create_span("granulation tissue at the posterior trachea", "OBS_LESION", "Granulation tissue (Posterior)", "lesion.type", ev3))
SPANS.append(create_span("removed en bloc", "PROC_ACTION", "removed en bloc", "action", ev3))
SPANS.append(create_span("Excellent tissue debulking", "OBS_ROSE", "Excellent tissue debulking", "findings_json", ev3))

# Event 4: Stent cleaning (Brush)
ev4 = "evt_04"
SPANS.append(create_span("cytology brush", "DEV_INSTRUMENT", "Cytology brush", "devices_json", ev4))
SPANS.append(create_span("within the stent", "ANAT_AIRWAY", "Stent lumen", "target.anatomy_type", ev4))
SPANS.append(create_span("mucus/secretions were removed", "PROC_ACTION", "Mucus removal", "action", ev4))

# General/Historical
SPANS.append(create_span("tracheal stenosis", "OBS_LESION", "tracheal stenosis", "lesion.type", "evt_hist_01"))
SPANS.append(create_span("Tracheal stent", "DEV_STENT", "Tracheal stent", "stent.material_or_brand", "evt_hist_02", is_hist=True))
SPANS.append(create_span("No immediate complications", "OUTCOME_COMPLICATION", "none", "outcomes.complications", "global"))

# -------------------------------------------------------------------------
# HYDRATION LOGIC
# -------------------------------------------------------------------------

def hydrate_spans(spans, full_text):
    hydrated = []
    for s in spans:
        target = s["span_text"]
        prefix = s.get("context_prefix", "")
        # Explicit match index logic if needed, but simple strategies first
        
        matches = list(re.finditer(re.escape(target), full_text))
        
        selected_match = None
        status = "ambiguous"
        
        if len(matches) == 1:
            selected_match = matches[0]
            status = "hydrated_unique"
        elif len(matches) > 1:
            if prefix:
                # Look for prefix in window before
                for m in matches:
                    start = m.start()
                    window_start = max(0, start - 150)
                    pre_text = full_text[window_start:start]
                    if prefix in pre_text:
                        selected_match = m
                        status = "hydrated_prefix_window"
                        break
            
            if not selected_match:
                # Fallback to first if not specified or found
                selected_match = matches[0]
                status = f"hydrated_match_index_0_of_{len(matches)}"

        if selected_match:
            s_copy = s.copy()
            s_copy["start_char"] = selected_match.start()
            s_copy["end_char"] = selected_match.end()
            s_copy["hydration_status"] = status
            hydrated.append(s_copy)
        else:
            s_copy = s.copy()
            s_copy["hydration_status"] = "not_found"
            hydrated.append(s_copy)
            
    return hydrated

HYDRATED_SPANS = hydrate_spans(SPANS, NOTE_TEXT)

# -------------------------------------------------------------------------
# EVENT LOG GENERATION
# -------------------------------------------------------------------------

EVENTS_FLAT = [
    {
        "event_id": "evt_01",
        "event_type": "Procedure",
        "method": "Therapeutic aspiration",
        "anatomy_target": "Trachea, RMS, LMS, Carina",
        "findings": "Mucus cleaned out"
    },
    {
        "event_id": "evt_02",
        "event_type": "Procedure",
        "method": "Cryotherapy",
        "device": "2.4mm Cryoprobe",
        "anatomy_target": "Anterior Granulation Tissue",
        "findings": "Excellent tissue destruction"
    },
    {
        "event_id": "evt_03",
        "event_type": "Procedure",
        "method": "Cryotherapy",
        "anatomy_target": "Posterior Granulation Tissue",
        "findings": "Excellent tissue debulking, removed en bloc"
    },
    {
        "event_id": "evt_04",
        "event_type": "Procedure",
        "method": "Brushing",
        "device": "Cytology brush",
        "anatomy_target": "Stent Lumen",
        "findings": "Mucus removed"
    }
]

# -------------------------------------------------------------------------
# EXCEL GENERATION
# -------------------------------------------------------------------------

def generate_excel():
    if not os.path.exists(TEMPLATE_PATH):
        # Create a dummy template if missing (for strict standalone compliance, though logic assumes it exists)
        wb = openpyxl.Workbook()
        wb.create_sheet("Note_Text")
        wb.create_sheet("Note_Index")
        wb.create_sheet("Span_Annotations")
        wb.create_sheet("Span_Hydrated")
        wb.create_sheet("Event_Log")
        wb.create_sheet("V3_Procedure_Events")
        wb.create_sheet("V3_Registry_JSON")
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # 1. Note_Text
    ws = wb["Note_Text"]
    if ws.max_row == 1:
        ws.append(["note_id", "source_file", "note_text"])
    ws.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws = wb["Note_Index"]
    header = [
        "source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes",
        "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", 
        "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy", 
        "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration", 
        "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation", 
        "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", 
        "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
        "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis", 
        "pleural_biopsy", "fibrinolytic_therapy"
    ]
    
    # If header missing/empty, write it
    if ws.max_row < 1 or (ws.max_row == 1 and ws.cell(1,1).value is None):
        ws.append(header)
        
    row_data = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "Auto", "Pending", ""]
    # Append flags in order
    flag_keys = header[8:]
    for k in flag_keys:
        row_data.append(PROCEDURE_FLAGS.get(k, 0))
    ws.append(row_data)

    # 3. Span_Annotations
    ws = wb["Span_Annotations"]
    headers_span = [
        "source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", 
        "match_index", "start_char", "end_char", "span_len", "label", "normalized_value", 
        "schema_field", "event_id", "is_negated", "is_historical", "time_anchor", "reviewer", 
        "comments", "hydration_status"
    ]
    if ws.max_row == 0: ws.append(headers_span)
    
    for s in SPANS:
        ws.append([
            s["source_file"], s["note_id"], s["span_id"], s["section_type"], s["context_prefix"], 
            s["span_text"], s["match_index"], "", "", f"=LEN(F{ws.max_row+1})", s["label"], 
            s["normalized_value"], s["schema_field"], s["event_id"], s["is_negated"], 
            s["is_historical"], s["time_anchor"], s["reviewer"], s["comments"], "needs_hydration"
        ])

    # 4. Span_Hydrated
    ws = wb["Span_Hydrated"]
    if ws.max_row == 0: ws.append(headers_span)
    
    for s in HYDRATED_SPANS:
        ws.append([
            s["source_file"], s["note_id"], s["span_id"], s["section_type"], s["context_prefix"], 
            s["span_text"], s["match_index"], s["start_char"], s["end_char"], s["span_len"], 
            s["label"], s["normalized_value"], s["schema_field"], s["event_id"], s["is_negated"], 
            s["is_historical"], s["time_anchor"], s["reviewer"], s["comments"], s["hydration_status"]
        ])

    # 5. Event_Log
    ws = wb["Event_Log"]
    headers_event = [
        "source_file", "note_id", "event_id", "event_type", "method", "anatomy_target", "device", 
        "needle_gauge", "stations", "counts", "measurements", "specimens", "findings", 
        "is_historical", "reviewer", "comments", "device_size", "device_material", 
        "outcome_airway_lumen_pre", "outcome_airway_lumen_post", "outcome_symptoms", 
        "outcome_pleural", "outcome_complication"
    ]
    if ws.max_row == 0: ws.append(headers_event)
    
    for e in EVENTS_FLAT:
        row = [
            SOURCE_FILE, NOTE_ID, e.get("event_id"), e.get("event_type"), e.get("method"), 
            e.get("anatomy_target"), e.get("device"), "", "", "", "", "", e.get("findings"), 
            False, "Auto", "", "", "", "", "", "", "", "none" if "complications" in e.get("findings", "").lower() else ""
        ]
        ws.append(row)

    # 6. V3_Procedure_Events
    ws = wb["V3_Procedure_Events"]
    headers_v3 = [
        "note_id", "event_id", "type", "target.anatomy_type", "target.location.lobe", 
        "target.location.segment", "target.station", "lesion.type", "lesion.size_mm", "method", 
        "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote", 
        "stent.size", "stent.material_or_brand", "catheter.size_fr", "outcomes.airway.lumen_pre", 
        "outcomes.airway.lumen_post", "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"
    ]
    if ws.max_row == 0: ws.append(headers_v3)

    # Simplified mapping for V3
    for e in EVENTS_FLAT:
        ws.append([
            NOTE_ID, e.get("event_id"), "Procedure", e.get("anatomy_target"), "", "", "", 
            "Granulation" if "Granulation" in e.get("anatomy_target", "") else "", "", 
            e.get("method"), e.get("device"), "", "", e.get("findings"), "", "", "", "", "", "", "", "", ""
        ])

    # 7. V3_Registry_JSON
    ws = wb["V3_Registry_JSON"]
    if ws.max_row == 0: ws.append(["schema_version", "note_id", "json_content", "no_immediate_complications"])
    
    json_obj = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": EVENTS_FLAT,
        "no_immediate_complications": True
    }
    ws.append(["3.0", NOTE_ID, json.dumps(json_obj, indent=2), True])

    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_excel()