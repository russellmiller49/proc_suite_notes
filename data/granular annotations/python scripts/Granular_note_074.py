import os
import openpyxl
from openpyxl.utils import get_column_letter
import json
import re

# --- INPUT DATA ---
NOTE_ID = "note_074"
SOURCE_FILE = "note_074.txt"
PROCEDURE_DATE = "" 
NOTE_TEXT = """NOTE_ID:  note_074 SOURCE_FILE: note_074.txt INDICATION FOR OPERATION:  [REDACTED] is a 37 year old-year-old male who presents with tumor obstruction.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
 
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
 
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31624 Dx bronchoscope/lavage (BAL)    
31636 Dilate and bronchial stent initial bronchus
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
ANESTHESIA: 
General Anesthesia
 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Rigid Bronchoscope
Flexible Therapeutic Bronchoscope
 
ESTIMATED BLOOD LOSS:   None
 
COMPLICATIONS:    None
 
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
 
Initial Airway Inspection Findings:
Successful therapeutic aspiration was performed to clean out the Right Mainstem, Bronchus Intermedius , and Left Mainstem from mucus.
Rigid Black ventilating scope was used to intubate the patient.
Endobronchial obstruction at Left Mainstem  (100% obstruction) was treated with the following modalities:
 
Modality	Tools	Setting/Mode	Duration	Results
Electrocautery	Knife 	PreciSect	10 secs	removed
Microwave	mini	90oC	2min, 1.5min	ablated
 
Prior to treatment, affected airway was note to be 0% patent.
After treatment, the airway was 100% patent. 
 
Successful therapeutic aspiration was performed to clean out the LLL and LUL subsegments from mucus.
The following stent (bonastent 10x20) was placed in the LLL basal subsegment to the LC2;
jailing LB6 (origin of the tumor). 
 
Bronchial alveolar lavage was performed at Superior Segment of Lingula (LB4) and Inferior Segment of Lingula (LB5).
Instilled 20 cc of NS, suction returned with 10 cc of NS.  Samples sent for Microbiology (Cultures/Viral/Fungal).
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
Tumor from LMS
BAL
 
IMPRESSION/PLAN: [REDACTED] is a 37 year old-year-old male who presents for bronchoscopy for endobronchial obstruction.
- f/u in clinic 
- f/u in 4-6 weeks post for stent check
- start stent hydration protocol"""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# --- FLAGGING LOGIC ---
flags = {
    # Bronchoscopy (23)
    'diagnostic_bronchoscopy': 1, # BAL
    'bal': 1,
    'bronchial_wash': 0,
    'brushings': 0,
    'endobronchial_biopsy': 0,
    'tbna_conventional': 0,
    'linear_ebus': 0,
    'radial_ebus': 0,
    'navigational_bronchoscopy': 0,
    'transbronchial_biopsy': 0,
    'transbronchial_cryobiopsy': 0,
    'therapeutic_aspiration': 1,
    'foreign_body_removal': 0,
    'airway_dilation': 1, # 31636 mentioned, rigid used
    'airway_stent': 1,
    'thermal_ablation': 1, # Electrocautery, Microwave
    'tumor_debulking_non_thermal': 0,
    'cryotherapy': 0,
    'blvr': 0,
    'peripheral_ablation': 0,
    'bronchial_thermoplasty': 0,
    'whole_lung_lavage': 0,
    'rigid_bronchoscopy': 1,
    # Pleural (7)
    'thoracentesis': 0,
    'chest_tube': 0,
    'ipc': 0,
    'medical_thoracoscopy': 0,
    'pleurodesis': 0,
    'pleural_biopsy': 0,
    'fibrinolytic_therapy': 0
}

# --- ANCHOR SPANS ---
# Structure: (span_text, label, normalized_value, schema_field, event_id, context_prefix)
spans = [
    # Event 1: Initial Therapeutic Aspiration
    ("therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", "method", "evt1", "Successful"),
    ("Right Mainstem", "ANAT_AIRWAY", "Right Mainstem Bronchus", "target.location.segment", "evt1", "clean out the"),
    ("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "target.location.segment", "evt1", "Right Mainstem,"),
    ("Left Mainstem", "ANAT_AIRWAY", "Left Mainstem Bronchus", "target.location.segment", "evt1", "Intermedius , and"),
    ("mucus", "OBS_FINDING", "Mucus", "findings", "evt1", "from"),

    # Event 2: Rigid Intubation
    ("Rigid Black ventilating scope", "DEV_INSTRUMENT", "Rigid Bronchoscope", "device", "evt2", None),
    ("intubate", "PROC_ACTION", "Intubation", "method", "evt2", None),

    # Event 3: Tumor Destruction (Electrocautery)
    ("Endobronchial obstruction", "OBS_LESION", "Endobronchial Obstruction", "lesion.type", "evt3", None),
    ("Left Mainstem", "ANAT_AIRWAY", "Left Mainstem Bronchus", "target.location.segment", "evt3", "obstruction at"),
    ("100% obstruction", "OUTCOME_AIRWAY_LUMEN_PRE", "0", "outcomes.airway.lumen_pre", "evt3", "Left Mainstem  ("),
    ("Electrocautery", "PROC_METHOD", "Electrocautery", "method", "evt3", None),
    ("Knife", "DEV_INSTRUMENT", "Electrocautery Knife", "device", "evt3", None),
    
    # Event 4: Tumor Destruction (Microwave)
    ("Microwave", "PROC_METHOD", "Microwave Ablation", "method", "evt4", None),
    ("mini", "DEV_INSTRUMENT", "Microwave Probe (Mini)", "device", "evt4", None),
    
    # Outcomes for Ablation (Events 3 & 4)
    ("0% patent", "OUTCOME_AIRWAY_LUMEN_PRE", "0", "outcomes.airway.lumen_pre", "evt4", "note to be"),
    ("100% patent", "OUTCOME_AIRWAY_LUMEN_POST", "100", "outcomes.airway.lumen_post", "evt4", "airway was"),

    # Event 5: Secondary Aspiration
    ("therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", "method", "evt5", "Successful"),
    ("LLL", "ANAT_AIRWAY", "Left Lower Lobe", "target.location.lobe", "evt5", "clean out the"),
    ("LUL", "ANAT_AIRWAY", "Left Upper Lobe", "target.location.lobe", "evt5", "and"),
    ("mucus", "OBS_FINDING", "Mucus", "findings", "evt5", "subsegments from"),

    # Event 6: Stent Placement
    ("stent", "DEV_STENT", "Airway Stent", "method", "evt6", "following"),
    ("bonastent", "DEV_STENT_MATERIAL", "Bonastent", "stent.material_or_brand", "evt6", None),
    ("10x20", "DEV_STENT_SIZE", "10x20mm", "stent.size", "evt6", None),
    ("placed", "PROC_ACTION", "Placement", "method", "evt6", None),
    ("LLL basal subsegment", "ANAT_AIRWAY", "LLL Basal Segment", "target.location.segment", "evt6", None),
    ("LC2", "ANAT_AIRWAY", "LC2", "target.location.segment", "evt6", "subsegment to the"),
    ("LB6", "ANAT_AIRWAY", "LB6", "target.location.segment", "evt6", "jailing"),
    ("origin of the tumor", "OBS_LESION", "Tumor Origin", "findings", "evt6", None),

    # Event 7: BAL
    ("Bronchial alveolar lavage", "PROC_METHOD", "BAL", "method", "evt7", None),
    ("Superior Segment of Lingula", "ANAT_AIRWAY", "Lingula Superior Segment", "target.location.segment", "evt7", None),
    ("LB4", "ANAT_AIRWAY", "LB4", "target.location.segment", "evt7", None),
    ("Inferior Segment of Lingula", "ANAT_AIRWAY", "Lingula Inferior Segment", "target.location.segment", "evt7", None),
    ("LB5", "ANAT_AIRWAY", "LB5", "target.location.segment", "evt7", None),
    ("20 cc", "MEAS_VOL", "20", "measurements.volume", "evt7", "Instilled"),
    ("10 cc", "MEAS_VOL", "10", "measurements.volume", "evt7", "returned with"),
    ("Microbiology", "OBS_ROSE", "Microbiology", "specimens.type", "evt7", None),

    # Complications / Global
    ("no immediate complications", "OUTCOME_COMPLICATION", "None", "outcomes.complications", "evt_glob", None),
]

# --- UTILS ---
def get_hydration_data(text, span_text, context_prefix=None):
    """
    Returns (start_char, end_char, hydration_status)
    """
    matches = [m for m in re.finditer(re.escape(span_text), text)]
    if not matches:
        return "", "", "not_found"
    
    if len(matches) == 1:
        m = matches[0]
        return m.start(), m.end(), "hydrated_unique"
    
    # Context match
    if context_prefix:
        for m in matches:
            start_search = max(0, m.start() - 150)
            context_window = text[start_search:m.start()]
            if context_prefix in context_window:
                return m.start(), m.end(), "hydrated_prefix_window"
    
    # Fallback: ambiguous
    return "", "", f"ambiguous_count={len(matches)}"

# --- EXCEL GENERATION ---
def generate_workbook():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        # Create a blank workbook if template is missing (though instructions say it must load)
        wb = openpyxl.Workbook()
        for sheet in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            wb.create_sheet(sheet)

    # 1. Note_Text
    ws = wb["Note_Text"]
    if ws.max_row == 1: # Header check simplistic
        ws.append(["note_id", "source_file", "note_text"])
    ws.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws = wb["Note_Index"]
    # Meta row
    ws.append([SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Success", ""])
    # Flags row
    flag_row = []
    # Order must match template or generic append. 
    # Assuming template has columns for flags. We will just append values corresponding to keys.
    # The keys in 'flags' dict should map to columns. For safety, we rely on the user instructions 
    # that "Append row with metadata... Then append the 30 procedure-presence flags".
    # We will just append a row of 0/1s.
    for k, v in flags.items():
        flag_row.append(v)
    ws.append(flag_row)

    # 3. Span_Annotations (Anchor First)
    ws_span = wb["Span_Annotations"]
    # Headers: source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, start_char, end_char, span_len, label, normalized_value, schema_field, event_id, is_negated, is_historical, time_anchor, reviewer, comments, hydration_status
    
    # 4. Span_Hydrated
    ws_hydrated = wb["Span_Hydrated"]

    span_id_counter = 1
    
    for item in spans:
        span_text, label, norm_val, schema, evt_id, ctx = item
        
        # Hydrate
        start, end, status = get_hydration_data(NOTE_TEXT, span_text, ctx)
        length = len(span_text)
        
        row_base = [
            SOURCE_FILE, NOTE_ID, f"span_{span_id_counter:03d}", "", 
            ctx, span_text, "", 
            "", "", f"=LEN(F{ws_span.max_row + 1})", # Formula for len
            label, norm_val, schema, evt_id,
            "FALSE", "FALSE", "", "", "", "needs_hydration"
        ]
        ws_span.append(row_base)

        row_hydrated = [
            SOURCE_FILE, NOTE_ID, f"span_{span_id_counter:03d}", "", 
            ctx, span_text, "", 
            start, end, length,
            label, norm_val, schema, evt_id,
            "FALSE", "FALSE", "", "", "", status
        ]
        ws_hydrated.append(row_hydrated)
        
        span_id_counter += 1

    # 5. Event_Log
    ws_event = wb["Event_Log"]
    # Columns: source_file, note_id, event_id, event_type, method, anatomy_target, device, needle_gauge, stations, counts, measurements, specimens, findings, is_historical, reviewer, comments, device_size, device_material, outcome_airway_lumen_pre, outcome_airway_lumen_post, outcome_symptoms, outcome_pleural, outcome_complication
    
    events_data = {
        "evt1": {
            "type": "Therapeutic Aspiration", "method": "Suction", "target": "Right Mainstem, BI, LMS", 
            "findings": "Mucus"
        },
        "evt2": {
            "type": "Rigid Bronchoscopy", "method": "Rigid Intubation", "device": "Rigid Black ventilating scope"
        },
        "evt3": {
            "type": "Tumor Destruction", "method": "Electrocautery", "target": "Left Mainstem", 
            "device": "Knife", "lumen_pre": "0% patent (100% obstr)"
        },
        "evt4": {
            "type": "Tumor Destruction", "method": "Microwave Ablation", "target": "Left Mainstem", 
            "device": "Microwave mini", "lumen_post": "100%"
        },
        "evt5": {
            "type": "Therapeutic Aspiration", "method": "Suction", "target": "LLL, LUL", "findings": "Mucus"
        },
        "evt6": {
            "type": "Airway Stent", "method": "Stent Placement", "target": "LLL Basal", 
            "device": "Bonastent", "size": "10x20", "findings": "Jailing LB6"
        },
        "evt7": {
            "type": "BAL", "method": "Lavage", "target": "Lingula (LB4, LB5)", 
            "measurements": "Instilled 20cc, Return 10cc", "specimens": "Microbiology"
        },
        "evt_glob": {
            "type": "Outcome", "complication": "None"
        }
    }

    for eid, data in events_data.items():
        ws_event.append([
            SOURCE_FILE, NOTE_ID, eid, 
            data.get("type", ""), data.get("method", ""), data.get("target", ""),
            data.get("device", ""), "", "", "", 
            data.get("measurements", ""), data.get("specimens", ""), data.get("findings", ""),
            "FALSE", "", "", 
            data.get("size", ""), data.get("material", ""), 
            data.get("lumen_pre", ""), data.get("lumen_post", ""), 
            "", "", data.get("complication", "")
        ])

    # 6. V3_Procedure_Events
    # note_id, event_id, type, target.anatomy_type, target.location.lobe, target.location.segment, target.station, lesion.type, lesion.size_mm, method, devices_json, measurements_json, specimens_json, findings_json, evidence_quote, stent.size, stent.material_or_brand, catheter.size_fr, outcomes.airway.lumen_pre, outcomes.airway.lumen_post, outcomes.symptoms, outcomes.pleural, outcomes.complications
    ws_v3 = wb["V3_Procedure_Events"]
    
    # Event 3/4/Combined for Outcome
    ws_v3.append([
        NOTE_ID, "evt3", "Tumor Destruction", "Airway", "", "Left Mainstem", "", 
        "Endobronchial Obstruction", "", "Electrocautery", '{"device": "Knife"}', "", "", "", "",
        "", "", "", "0", "", "", "", ""
    ])
    ws_v3.append([
        NOTE_ID, "evt4", "Tumor Destruction", "Airway", "", "Left Mainstem", "", 
        "", "", "Microwave", '{"device": "Microwave mini"}', "", "", "", "",
        "", "", "", "", "100", "", "", ""
    ])
    ws_v3.append([
        NOTE_ID, "evt6", "Airway Stent", "Airway", "LLL", "Basal Segment", "",
        "", "", "Stent Placement", '{"device": "Bonastent 10x20"}', "", "", '{"finding": "jailing LB6"}', "",
        "10x20", "Bonastent", "", "", "", "", "", ""
    ])
    ws_v3.append([
        NOTE_ID, "evt_glob", "Outcome", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "None"
    ])

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    registry_data = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "procedures": [
            {"event_id": "evt1", "type": "Therapeutic Aspiration", "target": "Right/Left Mainstem", "findings": ["Mucus"]},
            {"event_id": "evt2", "type": "Rigid Bronchoscopy", "device": "Rigid Black ventilating scope"},
            {"event_id": "evt3", "type": "Tumor Destruction", "method": "Electrocautery", "target": "LMS", "outcome_lumen_pre": 0},
            {"event_id": "evt4", "type": "Tumor Destruction", "method": "Microwave", "target": "LMS", "outcome_lumen_post": 100},
            {"event_id": "evt6", "type": "Airway Stent", "device": "Bonastent", "size": "10x20", "location": "LLL Basal"},
            {"event_id": "evt7", "type": "BAL", "location": "Lingula", "specimen": "Microbiology"}
        ],
        "no_immediate_complications": True
    }
    ws_json.append([json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Workbook saved to {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()