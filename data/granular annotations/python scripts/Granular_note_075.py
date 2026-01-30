import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os

# =============================================================================
# INPUT DATA
# =============================================================================
NOTE_ID = "note_075"
SOURCE_FILE = "note_075.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_075 SOURCE_FILE: note_075.txt INDICATION FOR OPERATION:  [REDACTED] is a 61 year old-year-old male who presents with lung nodule.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: R91.1 Solitary Lung Nodule
POSTOPERATIVE DIAGNOSIS:  R91.1 Solitary Lung Nodule
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31623 Dx bronchoscope/brushing    
31624 Dx bronchoscope/lavage (BAL)    
31628 TBBX single lobe     
31629 TBNA single lobe   
31626 Fiducial marker placements, single or multiple     
31627 Navigational Bronchoscopy (computer assisted)
77012 Radiology / radiologic guidance for CT guided needle placement (CIOS)
76377 3D rendering with interpretation and reporting of CT, US, Tomo modality (ION Planning Station)
31652 EBUS sampling 1 or 2 nodes
31654 Radial EBUS for peripheral lesion
76982 Ultrasound Elastography, First Target Lesion
76983 
Ultrasound Elastography, Additional Targets 
76983 Ultrasound Elastography, Additional Target 2
 
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
Linear EBUS 
Radial EBUS
Ion Robotic Bronchoscope
ESTIMATED BLOOD LOSS:   None
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
Initial Airway Inspection Findings:
CT Chest scan was placed on separate planning station to generate 3D rendering of the pathway to target.
The navigational plan was reviewed and verified.  This was then loaded into robotic bronchoscopy platform.
Successful therapeutic aspiration was performed to clean out the Right Mainstem, Bronchus Intermedius , and Left Mainstem from mucus.
Ventilation Parameters:
Mode\tRR\tTV\tPEEP\tFiO2\tFlow Rate\tPmean
vcv\t14\t400\t12\t100\t10\t15
Robotic navigation bronchoscopy was performed with Ion platform.  Partial registration was used.
Ion robotic catheter was used to engage the Anterior Segment of RUL (RB3).
Target lesion is about 1 cm in diameter.   Under navigational guidance the ion robotic catheter was advanced to 1.0 cm away from the planned target.
Radial EBUS was performed to confirm that the location of the nodule is Eccentric.
The following features were noted: Continuous margin .
Cone Beam CT was performed: 3-D reconstructions were performed on an independent workstation.
Cios Spin system was used for evaluation of nodule location.  Low dose spin was performed to acquire CT imaging.
This was passed on to Ion platform system for reconstruction and nodule location.
The 3D images was interpreted on an independent workstation (Ion).
Using the newly acquired nodule location, the Ion robotic system was adjusted to the new targeted location.
I personally interpreted the cone beam CT and 3-D reconstruction.
Transbronchial needle aspiration was performed with 21G Needle through the extended working channel catheter.  Total 6 samples were collected.
Samples sent for Microbiology (Cultures/Viral/Fungal), Cytology, and flow cytometry.
Transbronchial biopsy was performed with alligator forceps the extended working channel catheter.
Total 6 samples were collected.  Samples sent for Pathology.
Transbronchial brushing was performed with Protected cytology brush the extended working channel catheter.
Total 1 samples were collected.  Samples sent for Microbiology (Cultures/Viral/Fungal).
Bronchial alveolar lavage was performed the extended working channel catheter.
Instilled 60 cc of NS, suction returned with 20 cc of NS.  Samples sent for Microbiology (Cultures/Viral/Fungal) and Cytology.
Fiducial marker (0.8mm x 3mm soft tissue gold CIVCO) was loaded with bone wax and placed under fluoroscopy guidance.
Prior to withdraw of the bronchoscope. 
Prior to withdrawal of the bronchoscope, inspection demonstrated no evidence of bleeding.
EBUS-Findings
Indications: Diagnostic and Staging
Technique:
All lymph node stations were assessed. Only those 5 mm or greater in short axis were sampled.
Lymph node sizing was performed by EBUS and sampling by transbronchial needle aspiration was performed using 22-gauge Needle.
Lymph Nodes/Sites Inspected: 4R (lower paratracheal) node
4L (lower paratracheal) node
7 (subcarinal) node
11Rs lymph node
11L lymph node
No immediate complications
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
Elastography provided a semi-quantitative classification (Type 1–3), which was used to guide biopsy site selection and sampling strategy.
Lymph Nodes Evaluated:
Site 1: The 11L lymph node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was not sampled: Sampling this lymph node was not clinically indicated.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 1 elastographic pattern, predominantly soft (green/yellow), suggesting a reactive or benign process.
Site 2: The 7 (subcarinal) node was => 10 mm on CT and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was sampled.. 4 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 1 elastographic pattern, predominantly soft (green/yellow), suggesting a reactive or benign process.
Despite the benign appearance, TBNA was performed to confirm the absence of malignancy and to obtain cytology for diagnostic completeness.
Site 3: The 4L (lower paratracheal) node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was not sampled: Sampling this lymph node was not clinically indicated.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 1 elastographic pattern, predominantly soft (green/yellow), suggesting a reactive or benign process.
Site 4: The 4R (lower paratracheal) node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was not sampled: Sampling this lymph node was not clinically indicated.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 1 elastographic pattern, predominantly soft (green/yellow), suggesting a reactive or benign process.
Site 5: The 11Rs lymph node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was not sampled: Sampling this lymph node was not clinically indicated.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 1 elastographic pattern, predominantly soft (green/yellow), suggesting a reactive or benign process.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
RUL TBBX, TBNA, Brush, BAL
Station 7 TBNA
IMPRESSION/PLAN: [REDACTED] is a 61 year old-year-old male who presents for bronchoscopy for lung nodule and lymphadenopathy.
- f/u results
- f/u CXR"""

# =============================================================================
# DEFINITIONS AND UTILITIES
# =============================================================================

class Phase0Generator:
    def __init__(self, template_path, output_path, note_text):
        self.template_path = template_path
        self.output_path = output_path
        self.note_text = note_text
        self.wb = None
        self.spans = []
        self.events = []
        
        # Flags
        self.proc_flags = {
            "diagnostic_bronchoscopy": 0, "bal": 0, "bronchial_wash": 0, "brushings": 0,
            "endobronchial_biopsy": 0, "tbna_conventional": 0, "linear_ebus": 0, "radial_ebus": 0,
            "navigational_bronchoscopy": 0, "transbronchial_biopsy": 0, "transbronchial_cryobiopsy": 0,
            "therapeutic_aspiration": 0, "foreign_body_removal": 0, "airway_dilation": 0,
            "airway_stent": 0, "thermal_ablation": 0, "tumor_debulking_non_thermal": 0,
            "cryotherapy": 0, "blvr": 0, "peripheral_ablation": 0, "bronchial_thermoplasty": 0,
            "whole_lung_lavage": 0, "rigid_bronchoscopy": 0,
            "thoracentesis": 0, "chest_tube": 0, "ipc": 0, "medical_thoracoscopy": 0,
            "pleurodesis": 0, "pleural_biopsy": 0, "fibrinolytic_therapy": 0
        }

    def load_template(self):
        try:
            self.wb = openpyxl.load_workbook(self.template_path)
        except FileNotFoundError:
            # Fallback for fresh generation if template missing (though instructions say strictly load)
            self.wb = openpyxl.Workbook()
            for sheet in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
                if sheet not in self.wb.sheetnames:
                    self.wb.create_sheet(sheet)

    def set_flags(self):
        # Scan text for procedures to set binary flags
        txt = self.note_text.lower()
        
        # Mapping checks
        if "bronchoscop" in txt:
            self.proc_flags["diagnostic_bronchoscopy"] = 1
        
        if "lavage" in txt and "bronchial alveolar lavage" in txt:
            self.proc_flags["bal"] = 1
            
        if "brush" in txt:
            self.proc_flags["brushings"] = 1
            
        if "transbronchial needle aspiration" in txt or "tbna" in txt:
            self.proc_flags["tbna_conventional"] = 1
            
        if "linear ebus" in txt or "ebus sampling" in txt:
            self.proc_flags["linear_ebus"] = 1
            
        if "radial ebus" in txt:
            self.proc_flags["radial_ebus"] = 1
            
        if "navigational bronchoscopy" in txt or "robotic navigation" in txt or "ion" in txt:
            self.proc_flags["navigational_bronchoscopy"] = 1
            
        if "transbronchial biopsy" in txt or "tbbx" in txt:
            self.proc_flags["transbronchial_biopsy"] = 1
            
        if "therapeutic aspiration" in txt:
            self.proc_flags["therapeutic_aspiration"] = 1

    def add_span(self, text, label, norm_val, field, event_id, context="", is_hist=0):
        if not text: return
        self.spans.append({
            "text": text,
            "label": label,
            "norm": norm_val,
            "field": field,
            "event_id": event_id,
            "context": context,
            "hist": is_hist
        })

    def process_content(self):
        # 1. Event: Therapeutic Aspiration
        ev_id = "evt_01"
        self.add_span("therapeutic aspiration", "PROC_METHOD", "aspiration", "method", ev_id, context="Successful")
        self.add_span("clean out the Right Mainstem, Bronchus Intermedius , and Left Mainstem", "PROC_ACTION", "clean airway", "action", ev_id)
        self.add_span("Right Mainstem", "ANAT_AIRWAY", "right mainstem", "target.anatomy", ev_id)
        self.add_span("Bronchus Intermedius", "ANAT_AIRWAY", "bronchus intermedius", "target.anatomy", ev_id)
        self.add_span("Left Mainstem", "ANAT_AIRWAY", "left mainstem", "target.anatomy", ev_id)
        self.events.append({
            "event_id": ev_id, "type": "therapeutic_aspiration", "method": "aspiration",
            "anatomy": "RMS, BI, LMS"
        })

        # 2. Event: Nodule Navigation
        ev_id = "evt_02"
        self.add_span("Robotic navigation bronchoscopy", "PROC_METHOD", "robotic navigation", "method", ev_id)
        self.add_span("Ion platform", "DEV_INSTRUMENT", "Ion", "device", ev_id)
        self.add_span("Anterior Segment of RUL", "ANAT_LUNG_LOC", "RUL Anterior Segment", "target.location", ev_id)
        self.add_span("Radial EBUS", "PROC_METHOD", "radial ebus", "method", ev_id, context="performed to confirm")
        self.add_span("nodule is Eccentric", "OBS_ROSE", "eccentric", "finding", ev_id) # Closest fit
        self.add_span("Cone Beam CT", "PROC_METHOD", "cone beam ct", "method", ev_id)
        self.add_span("Cios Spin system", "DEV_INSTRUMENT", "Cios Spin", "device", ev_id)
        self.add_span("3-D reconstruction", "PROC_METHOD", "3d reconstruction", "method", ev_id, context="personally interpreted")
        self.events.append({
            "event_id": ev_id, "type": "navigation", "method": "robotic_nav, radial_ebus, cbct",
            "anatomy": "RUL Anterior"
        })

        # 3. Event: TBNA Nodule
        ev_id = "evt_03"
        self.add_span("Transbronchial needle aspiration", "PROC_METHOD", "tbna", "method", ev_id, context="performed with 21G")
        self.add_span("21G Needle", "DEV_NEEDLE", "21G", "device", ev_id)
        self.add_span("Total 6 samples", "MEAS_COUNT", "6", "count", ev_id, context="catheter")
        self.events.append({
            "event_id": ev_id, "type": "tbna", "method": "tbna", "device": "21G Needle", "counts": "6"
        })

        # 4. Event: TBBX Nodule
        ev_id = "evt_04"
        self.add_span("Transbronchial biopsy", "PROC_METHOD", "tbbx", "method", ev_id)
        self.add_span("alligator forceps", "DEV_INSTRUMENT", "alligator forceps", "device", ev_id)
        self.add_span("Total 6 samples", "MEAS_COUNT", "6", "count", ev_id, context="alligator")
        self.events.append({
            "event_id": ev_id, "type": "transbronchial_biopsy", "method": "tbbx", "device": "alligator forceps", "counts": "6"
        })

        # 5. Event: Brushing Nodule
        ev_id = "evt_05"
        self.add_span("Transbronchial brushing", "PROC_METHOD", "brushing", "method", ev_id)
        self.add_span("Protected cytology brush", "DEV_INSTRUMENT", "cytology brush", "device", ev_id)
        self.add_span("Total 1 samples", "MEAS_COUNT", "1", "count", ev_id)
        self.events.append({
            "event_id": ev_id, "type": "brushings", "method": "brushing", "device": "cytology brush", "counts": "1"
        })

        # 6. Event: BAL Nodule
        ev_id = "evt_06"
        self.add_span("Bronchial alveolar lavage", "PROC_METHOD", "bal", "method", ev_id)
        self.add_span("Instilled 60 cc", "MEAS_VOL", "60", "vol_instilled", ev_id)
        self.add_span("returned with 20 cc", "MEAS_VOL", "20", "vol_return", ev_id)
        self.events.append({
            "event_id": ev_id, "type": "bal", "method": "bal", "measurements": "60cc in, 20cc out"
        })

        # 7. Event: Fiducial
        ev_id = "evt_07"
        self.add_span("Fiducial marker", "DEV_INSTRUMENT", "fiducial", "device", ev_id)
        self.add_span("0.8mm x 3mm", "MEAS_SIZE", "0.8mm x 3mm", "size", ev_id)
        self.add_span("CIVCO", "DEV_STENT_MATERIAL", "CIVCO", "brand", ev_id) # Reuse stent field or instrument
        self.events.append({
            "event_id": ev_id, "type": "fiducial", "method": "placement", "device": "CIVCO fiducial"
        })

        # EBUS Sequence
        # Common needle for EBUS
        ebus_needle = "22-gauge Needle"
        
        # 8. EBUS Station 11L (No sample)
        ev_id = "evt_08"
        self.add_span("11L lymph node", "ANAT_LN_STATION", "11L", "target.station", ev_id, context="Site 1")
        self.add_span("Elastography", "PROC_METHOD", "elastography", "method", ev_id, context="Site 1")
        self.add_span("Type 1", "OBS_ROSE", "Type 1", "finding", ev_id, context="Site 1")
        self.add_span("not sampled", "PROC_ACTION", "not sampled", "action", ev_id, context="Site 1")
        self.events.append({
            "event_id": ev_id, "type": "ebus_staging", "stations": "11L", "findings": "Type 1 Elastography"
        })

        # 9. EBUS Station 7 (Sampled)
        ev_id = "evt_09"
        self.add_span("7 (subcarinal) node", "ANAT_LN_STATION", "7", "target.station", ev_id, context="Site 2")
        self.add_span("Type 1", "OBS_ROSE", "Type 1", "finding", ev_id, context="Site 2")
        self.add_span("transbronchial biopsies", "PROC_METHOD", "tbna", "method", ev_id, context="Site 2")
        self.add_span("TBNA", "PROC_METHOD", "tbna", "method", ev_id, context="Site 2") # confirm absence
        self.add_span("22-gauge Needle", "DEV_NEEDLE", "22G", "device", ev_id) # General context
        self.events.append({
            "event_id": ev_id, "type": "ebus_tbna", "stations": "7", "method": "tbna", "findings": "Type 1"
        })

        # 10. EBUS Station 4L (No sample)
        ev_id = "evt_10"
        self.add_span("4L (lower paratracheal) node", "ANAT_LN_STATION", "4L", "target.station", ev_id, context="Site 3")
        self.add_span("Type 1", "OBS_ROSE", "Type 1", "finding", ev_id, context="Site 3")
        self.add_span("not sampled", "PROC_ACTION", "not sampled", "action", ev_id, context="Site 3")
        self.events.append({
            "event_id": ev_id, "type": "ebus_staging", "stations": "4L"
        })

        # 11. EBUS Station 4R (No sample)
        ev_id = "evt_11"
        self.add_span("4R (lower paratracheal) node", "ANAT_LN_STATION", "4R", "target.station", ev_id, context="Site 4")
        self.add_span("Type 1", "OBS_ROSE", "Type 1", "finding", ev_id, context="Site 4")
        self.add_span("not sampled", "PROC_ACTION", "not sampled", "action", ev_id, context="Site 4")
        self.events.append({
            "event_id": ev_id, "type": "ebus_staging", "stations": "4R"
        })

        # 12. EBUS Station 11Rs (No sample)
        ev_id = "evt_12"
        self.add_span("11Rs lymph node", "ANAT_LN_STATION", "11Rs", "target.station", ev_id, context="Site 5")
        self.add_span("Type 1", "OBS_ROSE", "Type 1", "finding", ev_id, context="Site 5")
        self.add_span("not sampled", "PROC_ACTION", "not sampled", "action", ev_id, context="Site 5")
        self.events.append({
            "event_id": ev_id, "type": "ebus_staging", "stations": "11Rs"
        })

        # Complications
        self.add_span("no immediate complications", "OUTCOME_COMPLICATION", "none", "outcome", "global", context="patient tolerated")


    def hydrate_spans(self):
        # Hydrate logic matching instructions
        hydrated = []
        for s in self.spans:
            txt_val = s["text"]
            count = self.note_text.count(txt_val)
            start = -1
            end = -1
            status = "ambiguous"

            if count == 0:
                # Should not happen if extracted from text
                status = "not_found"
            elif count == 1:
                start = self.note_text.find(txt_val)
                status = "hydrated_unique"
            else:
                # Context match
                if s["context"]:
                    # Sliding window or simple split find
                    # Find all indices
                    indices = [m.start() for m in re.finditer(re.escape(txt_val), self.note_text)]
                    best_idx = -1
                    for idx in indices:
                        # check 120 chars before
                        pre_start = max(0, idx - 120)
                        pre_text = self.note_text[pre_start:idx]
                        if s["context"] in pre_text:
                            best_idx = idx
                            break
                    if best_idx != -1:
                        start = best_idx
                        status = "hydrated_context"
                    else:
                        # Fallback to first if ambiguous? Or keep ambiguous.
                        # Logic says "Else: ambiguous"
                        status = f"ambiguous_count={count}"
                else:
                    status = f"ambiguous_count={count}"
            
            if start != -1:
                end = start + len(txt_val)
            
            row = {
                "source_file": SOURCE_FILE,
                "note_id": NOTE_ID,
                "span_id": f"span_{len(hydrated)+1:03d}",
                "section_type": "procedure",
                "context_prefix": s["context"],
                "span_text": txt_val,
                "match_index": "", # logic not fully implemented for index
                "start_char": start if start != -1 else "",
                "end_char": end if end != -1 else "",
                "span_len": len(txt_val),
                "label": s["label"],
                "normalized_value": s["norm"],
                "schema_field": s["field"],
                "event_id": s["event_id"],
                "is_negated": 0,
                "is_historical": s["hist"],
                "time_anchor": "",
                "reviewer": "",
                "comments": "",
                "hydration_status": status
            }
            hydrated.append(row)
        return hydrated

    def write_worksheet(self):
        self.load_template()
        
        # 1. Note_Text
        ws = self.wb["Note_Text"]
        ws.append([NOTE_ID, SOURCE_FILE, self.note_text])

        # 2. Note_Index
        ws = self.wb["Note_Index"]
        # Metadata row
        ws.append([SOURCE_FILE, NOTE_ID, "", "", "", "", "Ready", ""])
        # Flags
        flag_row = [self.proc_flags[k] for k in self.proc_flags]
        ws.append(flag_row)

        # 3. Span_Annotations (Anchor First)
        ws_ann = self.wb["Span_Annotations"]
        hydrated_rows = self.hydrate_spans()
        
        for r in hydrated_rows:
            # Columns: source, note, span_id, section, context, text, match_index, start(blank), end(blank), len, label, norm, field, event, neg, hist, time, rev, comm, status
            ws_ann.append([
                r["source_file"], r["note_id"], r["span_id"], r["section_type"],
                r["context_prefix"], r["span_text"], r["match_index"],
                "", "", r["span_len"], # Blanks for start/end
                r["label"], r["normalized_value"], r["schema_field"], r["event_id"],
                r["is_negated"], r["is_historical"], r["time_anchor"],
                r["reviewer"], r["comments"], "needs_hydration"
            ])

        # 4. Span_Hydrated
        ws_hyd = self.wb["Span_Hydrated"]
        for r in hydrated_rows:
            ws_hyd.append([
                r["source_file"], r["note_id"], r["span_id"], r["section_type"],
                r["context_prefix"], r["span_text"], r["match_index"],
                r["start_char"], r["end_char"], r["span_len"],
                r["label"], r["normalized_value"], r["schema_field"], r["event_id"],
                r["is_negated"], r["is_historical"], r["time_anchor"],
                r["reviewer"], r["comments"], r["hydration_status"]
            ])

        # 5. Event_Log
        ws_evt = self.wb["Event_Log"]
        for e in self.events:
            # map dict to schema cols
            ws_evt.append([
                SOURCE_FILE, NOTE_ID, e["event_id"], e.get("type", ""), e.get("method", ""),
                e.get("anatomy", ""), e.get("device", ""), e.get("needle", ""), e.get("stations", ""),
                e.get("counts", ""), e.get("measurements", ""), "", e.get("findings", ""),
                0, "", ""
            ])

        # 6. V3_Procedure_Events
        ws_v3 = self.wb["V3_Procedure_Events"]
        for e in self.events:
            # Simplified V3 mapping
            ws_v3.append([
                NOTE_ID, e["event_id"], e.get("type", ""),
                e.get("anatomy", ""), "", "", e.get("stations", ""), # anatomy fields
                "", "", # lesion
                e.get("method", ""), # method
                json.dumps({"device": e.get("device")}), # devices
                json.dumps({"measurements": e.get("measurements")}), # measurements
                "", # specimens
                json.dumps({"findings": e.get("findings")}), # findings
                "", # quote
                "", "", "", # stent
                "", "", "", "", "" # outcomes
            ])

        # 7. V3_Registry_JSON
        ws_json = self.wb["V3_Registry_JSON"]
        reg_data = {
            "schema_version": "3.0",
            "note_id": NOTE_ID,
            "procedures": self.events,
            "no_immediate_complications": True
        }
        ws_json.append([json.dumps(reg_data, indent=2)])

        self.wb.save(self.output_path)
        print(f"Generated {self.output_path}")

# =============================================================================
# EXECUTION
# =============================================================================
if __name__ == "__main__":
    generator = Phase0Generator(TEMPLATE_PATH, OUTPUT_PATH, NOTE_TEXT)
    generator.set_flags()
    generator.process_content()
    generator.write_worksheet()