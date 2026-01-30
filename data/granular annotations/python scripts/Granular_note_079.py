import openpyxl
from openpyxl.utils import get_column_letter
import re
import os
import json

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_079"
SOURCE_FILE = "note_079.txt"
PROCEDURE_DATE = "2020-01-20" # Placeholder based on instruction context or leave blank
NOTE_TEXT = """NOTE_ID:  note_079 SOURCE_FILE: note_079.txt INDICATION FOR OPERATION:  [REDACTED]is a 60 year old-year-old male who presents with respiratory failure and neck mass.
PREOPERATIVE DIAGNOSIS: J96.90 Respiratory Failure
POSTOPERATIVE DIAGNOSIS:  J96.90 Respiratory Failure
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31624 Dx bronchoscope/lavage (BAL)    
31653 EBUS sampling 3 or more nodes  
76981 Ultrasound Elastography, Parenchyma of Organ
76982 Ultrasound Elastography, First Target Lesion
76983 Ultrasound Elastography, Additional Targets 
76983 Ultrasound Elastography, Additional Target 2
31630 Balloon dilation
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient 
required a EBUS lymph node forceps/cryo biopsies and very challenging rigid bronchoscopy intubation due to deviated upper airway anatomy and tracheal stenosis.
This also required intra-operative ENT consultation for tracheostomy. This resulted in >100% increased work due to Time, Technical difficulty of procedure, and Physical and mental effort required.
Apply to: 31653 EBUS sampling 3 or more nodes  .
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Rigid Tracheoscope
Flexible Therapeutic Bronchoscope
Flexible Hybrid (Pedatric) Bronchoscope
Linear EBUS 
ESTIMATED BLOOD LOSS:   Minimum
COMPLICATIONS:    Injury to left vocal cord
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: Supine
Initial Airway Inspection Findings:
The endotracheal tube is in good position.
Pharynx: Not assessed due to bronchoscopy introduction through ETT.
Larynx: Not assessed due to bronchoscopy introduction through ETT.
Vocal Cords: Not assessed due to bronchoscopy introduction through ETT.
Trachea: Distal 1/3 normal.
Main Carina: Sharp
Right Lung Proximal Airways: Normal anatomic branching to segmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Left Lung Proximal Airways: Normal anatomic branching to segmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Mucosa: Normal.
Secretions: Large burden, thick, and clear bilaterally.
The endotracheal tube itself was noted to have flaking/peeling plastic the length of the tube.
ENDOTRACHEAL TUBE FLAKING / DEBRIS
 
Successful therapeutic aspiration was performed to clean out the Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus.
Bronchial alveolar lavage was performed at Lateral Segment of RML (RB4) and Medial Segment of RML (RB5).
Instilled 40 cc of NS, suction returned with 15 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
After induction of muscle relaxants, tooth or gum protector was placed.
The black rigid barrel tracheoscope was introduced through the mouth and advanced in the midline while keeping the alignment with the axis of the trachea and minimizing pressure to the teeth.
This was done side by side with the endotracheal tube in place.
The endotracheal tube was removed and there was moderate blood in the oropharynx despite suctioning.
Additionally, there was anterior and left displacement of the airway.
The glottic inlet could not be fully appreciated and patient began to desaturate so intubation was aborted and anesthesia re-recruited with BMV which was easy, though required oral airway and two hand seal on the mask.
The black barrel was exchanged for a red short barrel for the second intubation attempt.
Barrel was advanced into the mouth and into the larynx.
The vocal cords were only partially identified but the rigid bronchoscope was able to be advanced into the trachea.
Once the rigid bronchoscope was positioned at the mid-trachea, jet ventilation was initiated and chest wall movement was confirmed.
Intubation was very challenging due to above. 
No endobronchial or endotracheal tumor/disease was noted.
There was severe extrinsic compression from the posterior membrane at the proximal aspect of the trachea in the subglottic space.
This compression was noted to be displacing the trachea anteriorly.
EXTRINSIC COMPRESSION AT THE POSTERIOR TRACHEAL AS VIEWED THROUGH RIGID TRACHEOSCOPE
 
LEFT VOCAL CORD ERYTHEMA / INJURY
Stenosis Measurements:
Distance from vocal folds to stenosis	5 mm
Location of stenosis relative to cricoid 	Above cricoid (subglottic funnel)
Distance from top of stenosis to bottom 	5.5 cm
Distance of bottom of stenosis to carina	5.7 cm
EBUS-Findings
Indications: Diagnostic and Staging
Technique:
All lymph node stations were assessed.
Only those 5 mm or greater in short axis were sampled.
Lymph node sizing was performed by EBUS and sampling by transbronchial needle aspiration was performed using 25-gauge Needle and 22-gauge Needle.
Lymph Nodes/Sites Inspected: 4R (lower paratracheal) node
4L (lower paratracheal) node
7 (subcarinal) node
11Rs lymph node
11Ri lymph node
11L lymph node
Tracheal mass
Overall ROSE Diagnosis: Positive for malignancy
No immediate complications
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
Elastography provided a semi-quantitative classification (Type 1–3), which was used to guide biopsy site selection and sampling strategy.
Lymph Nodes Evaluated:
Site 1: The 11Rs lymph node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was sampled.. 4 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 3 elastographic pattern, predominantly stiff (blue), raising concern for malignancy.
Based on this finding, the lymph node was prioritized for TBNA with multiple passes obtained to maximize diagnostic yield.
Site 2: The 4R (lower paratracheal) node was => 10 mm on CT and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was sampled.. 5 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 3 elastographic pattern, predominantly stiff (blue), raising concern for malignancy.
Based on this finding, the lymph node was prioritized for TBNA with multiple passes obtained to maximize diagnostic yield.
Site 3: The 7 (subcarinal) node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was sampled.. 4 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 2 elastographic pattern with mixed soft and stiff regions.
Given this heterogeneous and indeterminate appearance, TBNA was directed at representative areas to ensure comprehensive sampling and to minimize the risk of underdiagnosis.
Site 4: The 4L (lower paratracheal) node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was sampled.. 5 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 2 elastographic pattern with mixed soft and stiff regions.
Given this heterogeneous and indeterminate appearance, TBNA was directed at representative areas to ensure comprehensive sampling and to minimize the risk of underdiagnosis.
Site 5: The 11L lymph node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was sampled.. 4 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 3 elastographic pattern, predominantly stiff (blue), raising concern for malignancy.
Based on this finding, the lymph node was prioritized for TBNA with multiple passes obtained to maximize diagnostic yield.
Site 6: The tracheal mass was => 10 mm on CT and Metabolic activity unknown or PET-CT scan unavailable.
The mass was photographed. The site was sampled.. 8 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Additionally, 5 endobronchial ultrasound guided transbronchial CRYObiopsies were performed with samples obtained.
Preliminary ROSE Cytology was reported as adequate and suggestive of Positive for malignancy. Final results are pending.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was in stable condition.
Given the patient's prolonged intubation at OSH (~2 weeks), distorted upper airway anatomy, tracheal stenosis, and possible vocal cord injury, decision was made to consult ENT intraoperatively for surgical tracheostomy placement under the same anesthesia event.
They reviewed the case and agreed with surgical tracheostomy placement.
At the conclusion of the case the rigid tracheoscope was extubated and the patient underwent fiberoptic intubation with a 7.0 ETT to facilitate ENT procedure.
SPECIMEN(S): 
--RML BAL (cell count, micro, cyto)
--Station 11Rs, 4R, 7, 4L, 11L, tracheal mass TBNA
--Tracheal mass cryobiopsies
 
IMPRESSION/PLAN: [REDACTED]is a 60 year old-year-old male who presents for bronchoscopy for evaluation of neck mass with extrinsic compression.
Patient tolerated the procedure well. The patient remained intubated and sedated and the conclusion of our portion of the procedure.
The care of the patient was turned over to the anesthesia and ENT teams for tracheostomy placement.
--Post procedure CXR
--Follow up BAL, TBNA, and cryobiopsy results
--Continued care per primary team"""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# -------------------------------------------------------------------------
# CONFIG / LOOKUPS
# -------------------------------------------------------------------------

PROCEDURE_FLAGS = [
    "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
    "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy",
    "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration",
    "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
    "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation",
    "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
    "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis",
    "pleural_biopsy", "fibrinolytic_therapy"
]

# Set flags based on note content
# Logic:
# - rigid_bronchoscopy: Explicitly mentioned "Rigid Tracheoscope"
# - therapeutic_aspiration: Explicitly mentioned "Successful therapeutic aspiration"
# - bal: Explicitly mentioned "Bronchial alveolar lavage"
# - linear_ebus: Explicitly mentioned "Linear EBUS"
# - tbna_conventional: Needle aspiration with EBUS
# - transbronchial_cryobiopsy: "transbronchial CRYObiopsies"
# - airway_dilation: Header 31630, but text "rigid bronchoscope was able to be advanced" suggests dilation by rigid scope. Will flag but focus events on rigid.
# - diagnostic_bronchoscopy: Yes

FLAGS_TO_SET = {
    "diagnostic_bronchoscopy": 1,
    "bal": 1,
    "tbna_conventional": 1,
    "linear_ebus": 1,
    "transbronchial_cryobiopsy": 1,
    "therapeutic_aspiration": 1,
    "rigid_bronchoscopy": 1,
    # "airway_dilation": 1 # Often implicit in rigid, but without explicit balloon/bougie narrative, we'll leave 0 or 1? Header has it. Let's set 1 to be safe given header, but no specific balloon event.
    "airway_dilation": 1 
}

# -------------------------------------------------------------------------
# SPAN DATA GENERATION
# -------------------------------------------------------------------------

# Helper to create span row
def create_span(span_text, label, normalized_value, schema_field, event_id, 
                context_prefix=None, is_negated=False, is_historical=False, comments=""):
    return {
        "span_text": span_text,
        "label": label,
        "normalized_value": normalized_value,
        "schema_field": schema_field,
        "event_id": event_id,
        "context_prefix": context_prefix,
        "is_negated": is_negated,
        "is_historical": is_historical,
        "comments": comments
    }

spans = []

# Event 1: Therapeutic Aspiration
ev1 = "evt_01_asp"
spans.append(create_span("therapeutic aspiration", "PROC_METHOD", "therapeutic aspiration", "method", ev1))
spans.append(create_span("clean out the Trachea", "PROC_ACTION", "clean out", "action", ev1))
spans.append(create_span("Trachea (Distal 1/3)", "ANAT_AIRWAY", "trachea", "target.anatomy", ev1))
spans.append(create_span("Right Mainstem", "ANAT_AIRWAY", "RMS", "target.anatomy", ev1))
spans.append(create_span("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "target.anatomy", ev1))
spans.append(create_span("Left Mainstem", "ANAT_AIRWAY", "LMS", "target.anatomy", ev1))
spans.append(create_span("Large burden, thick, and clear", "OBS_FINDING", "thick clear secretions", "findings", ev1, context_prefix="Secretions: "))

# Event 2: BAL
ev2 = "evt_02_bal"
spans.append(create_span("Bronchial alveolar lavage", "PROC_METHOD", "bal", "method", ev2))
spans.append(create_span("Lateral Segment of RML (RB4)", "ANAT_LUNG_LOC", "RML Lateral", "target.location", ev2))
spans.append(create_span("Medial Segment of RML (RB5)", "ANAT_LUNG_LOC", "RML Medial", "target.location", ev2))
spans.append(create_span("Instilled 40 cc", "MEAS_VOL", "40", "measurements.instilled", ev2))
spans.append(create_span("returned with 15 cc", "MEAS_VOL", "15", "measurements.returned", ev2))

# Event 3: Rigid Bronchoscopy / Intubation
ev3 = "evt_03_rigid"
spans.append(create_span("black rigid barrel tracheoscope", "DEV_INSTRUMENT", "rigid bronchoscope", "device", ev3))
spans.append(create_span("red short barrel", "DEV_INSTRUMENT", "rigid bronchoscope short", "device", ev3))
spans.append(create_span("advanced into the trachea", "PROC_ACTION", "advance", "action", ev3))
spans.append(create_span("severe extrinsic compression", "OBS_LESION", "extrinsic compression", "lesion.type", ev3))
spans.append(create_span("trachea", "ANAT_AIRWAY", "trachea", "target.anatomy", ev3, context_prefix="compression was noted to be displacing the "))

# Event 4: EBUS Station 11Rs
ev4 = "evt_04_ebus_11rs"
spans.append(create_span("11Rs lymph node", "ANAT_LN_STATION", "11Rs", "target.station", ev4, context_prefix="Site 1: The "))
spans.append(create_span("4 endobronchial ultrasound guided transbronchial biopsies", "PROC_METHOD", "ebus_tbna", "method", ev4, context_prefix="sampled.. "))
spans.append(create_span("Type 3 elastographic pattern", "OBS_FINDING", "elastography_type_3", "findings", ev4, context_prefix="demonstrated a "))
spans.append(create_span("predominantly stiff (blue)", "OBS_FINDING", "stiff", "findings", ev4))

# Event 5: EBUS Station 4R
ev5 = "evt_05_ebus_4r"
spans.append(create_span("4R (lower paratracheal) node", "ANAT_LN_STATION", "4R", "target.station", ev5, context_prefix="Site 2: The "))
spans.append(create_span("5 endobronchial ultrasound guided transbronchial biopsies", "PROC_METHOD", "ebus_tbna", "method", ev5, context_prefix="sampled.. "))
spans.append(create_span("Type 3 elastographic pattern", "OBS_FINDING", "elastography_type_3", "findings", ev5, context_prefix="target lymph node demonstrated a "))

# Event 6: EBUS Station 7
ev6 = "evt_06_ebus_7"
spans.append(create_span("7 (subcarinal) node", "ANAT_LN_STATION", "7", "target.station", ev6, context_prefix="Site 3: The "))
spans.append(create_span("4 endobronchial ultrasound guided transbronchial biopsies", "PROC_METHOD", "ebus_tbna", "method", ev6, context_prefix="sampled.. "))
spans.append(create_span("Type 2 elastographic pattern", "OBS_FINDING", "elastography_type_2", "findings", ev6, context_prefix="demonstrated a "))
spans.append(create_span("mixed soft and stiff regions", "OBS_FINDING", "mixed soft/stiff", "findings", ev6))

# Event 7: EBUS Station 4L
ev7 = "evt_07_ebus_4l"
spans.append(create_span("4L (lower paratracheal) node", "ANAT_LN_STATION", "4L", "target.station", ev7, context_prefix="Site 4: The "))
spans.append(create_span("5 endobronchial ultrasound guided transbronchial biopsies", "PROC_METHOD", "ebus_tbna", "method", ev7, context_prefix="sampled.. "))
spans.append(create_span("Type 2 elastographic pattern", "OBS_FINDING", "elastography_type_2", "findings", ev7, context_prefix="target lymph node demonstrated a "))

# Event 8: EBUS Station 11L
ev8 = "evt_08_ebus_11l"
spans.append(create_span("11L lymph node", "ANAT_LN_STATION", "11L", "target.station", ev8, context_prefix="Site 5: The "))
spans.append(create_span("4 endobronchial ultrasound guided transbronchial biopsies", "PROC_METHOD", "ebus_tbna", "method", ev8, context_prefix="sampled.. "))
spans.append(create_span("Type 3 elastographic pattern", "OBS_FINDING", "elastography_type_3", "findings", ev8, context_prefix="target lymph node demonstrated a "))

# Event 9: EBUS Tracheal Mass
ev9 = "evt_09_ebus_mass"
spans.append(create_span("tracheal mass", "OBS_LESION", "mass", "lesion.type", ev9, context_prefix="Site 6: The "))
spans.append(create_span("8 endobronchial ultrasound guided transbronchial biopsies", "PROC_METHOD", "ebus_tbna", "method", ev9))
spans.append(create_span("Positive for malignancy", "OBS_ROSE", "positive", "findings.rose", ev9, context_prefix="ROSE Cytology was reported as adequate and suggestive of "))

# Event 10: Cryobiopsy Tracheal Mass
ev10 = "evt_10_cryo"
spans.append(create_span("5 endobronchial ultrasound guided transbronchial CRYObiopsies", "PROC_METHOD", "cryobiopsy", "method", ev10))
spans.append(create_span("tracheal mass", "OBS_LESION", "mass", "lesion.type", ev10, context_prefix="Site 6: The "))

# Complications / Outcomes
ev_out = "evt_outcome"
spans.append(create_span("No immediate complications", "OUTCOME_COMPLICATION", "none", "outcomes.complications", ev_out))
spans.append(create_span("surgical tracheostomy placement", "PROC_METHOD", "tracheostomy", "method", "evt_11_trach"))
spans.append(create_span("Injury to left vocal cord", "OUTCOME_COMPLICATION", "vocal cord injury", "outcomes.complications", ev_out, context_prefix="COMPLICATIONS:    "))

# -------------------------------------------------------------------------
# PROCESSING
# -------------------------------------------------------------------------

def hydrate_offsets(text, span_list):
    hydrated = []
    for sp in span_list:
        stext = sp["span_text"]
        prefix = sp["context_prefix"]
        
        start = -1
        end = -1
        method = "ambiguous"
        
        # 1. Exact unique
        if text.count(stext) == 1:
            start = text.find(stext)
            method = "hydrated_unique"
        # 2. Context
        elif prefix and text.count(stext) > 1:
            # Find all occurrences
            occurrences = [m.start() for m in re.finditer(re.escape(stext), text)]
            best_occ = -1
            for occ in occurrences:
                # check window before
                window = text[max(0, occ - 120):occ]
                if prefix in window:
                    best_occ = occ
                    break
            if best_occ != -1:
                start = best_occ
                method = "hydrated_prefix_window"
            else:
                method = f"ambiguous_count={text.count(stext)}_prefix_not_found"
        else:
            method = f"ambiguous_count={text.count(stext)}"

        if start != -1:
            end = start + len(stext)
        
        row = sp.copy()
        row["start_char"] = start if start != -1 else ""
        row["end_char"] = end if end != -1 else ""
        row["hydration_status"] = method
        row["span_len"] = len(stext)
        hydrated.append(row)
    return hydrated

hydrated_spans = hydrate_offsets(NOTE_TEXT, spans)

# -------------------------------------------------------------------------
# WRITING TO EXCEL
# -------------------------------------------------------------------------

def generate_excel(template_path: str = TEMPLATE_PATH, output_path: str = OUTPUT_PATH) -> None:
    if os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
    else:
        wb = openpyxl.Workbook()

    ws_text = wb["Note_Text"] if "Note_Text" in wb.sheetnames else wb.create_sheet("Note_Text")
    if ws_text.max_row == 1:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    ws_index = wb["Note_Index"] if "Note_Index" in wb.sheetnames else wb.create_sheet("Note_Index")
    if ws_index.max_row == 1:
        headers = ["source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes"] + PROCEDURE_FLAGS
        ws_index.append(headers)

    flag_values = [FLAGS_TO_SET.get(f, 0) for f in PROCEDURE_FLAGS]
    ws_index.append([SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "pending", ""] + flag_values)

    ws_spans = wb["Span_Annotations"] if "Span_Annotations" in wb.sheetnames else wb.create_sheet("Span_Annotations")
    if ws_spans.max_row == 1:
        ws_spans.append(["source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", "match_index", "start_char", "end_char", "span_len", "label", "normalized_value", "schema_field", "event_id", "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"])

    for i, h in enumerate(hydrated_spans):
        ws_spans.append([
            SOURCE_FILE, NOTE_ID, f"span_{i+1:03d}", "", h["context_prefix"], h["span_text"], "", "", "", h["span_len"],
            h["label"], h["normalized_value"], h["schema_field"], h["event_id"],
            h["is_negated"], h["is_historical"], "", "", h["comments"], "needs_hydration"
        ])

    ws_hyd = wb["Span_Hydrated"] if "Span_Hydrated" in wb.sheetnames else wb.create_sheet("Span_Hydrated")
    if ws_hyd.max_row == 1:
        ws_hyd.append(["source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", "match_index", "start_char", "end_char", "span_len", "label", "normalized_value", "schema_field", "event_id", "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"])

    for i, h in enumerate(hydrated_spans):
        ws_hyd.append([
            SOURCE_FILE, NOTE_ID, f"span_{i+1:03d}", "", h["context_prefix"], h["span_text"], "", h["start_char"], h["end_char"], h["span_len"],
            h["label"], h["normalized_value"], h["schema_field"], h["event_id"],
            h["is_negated"], h["is_historical"], "", "", h["comments"], h["hydration_status"]
        ])

    ws_event = wb["Event_Log"] if "Event_Log" in wb.sheetnames else wb.create_sheet("Event_Log")
    if ws_event.max_row == 1:
        ws_event.append(["source_file", "note_id", "event_id", "event_type", "method", "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements", "specimens", "findings", "is_historical", "reviewer", "comments", "device_size", "device_material", "outcome_airway_lumen_pre", "outcome_airway_lumen_post", "outcome_symptoms", "outcome_pleural", "outcome_complication"])

    events_data = [
        (ev1, "therapeutic_aspiration", "aspiration", "Trachea/Mainstems", "", "", "", "", "thick clear secretions", "", "thick clear secretions", False),
        (ev2, "bal", "bal", "RML", "", "", "", "", "40cc/15cc", "cytology", "", False),
        (ev3, "rigid_bronchoscopy", "rigid_scope", "Trachea", "Rigid Barrel", "", "", "", "Extrinsic Compression", "", "Severe compression", False),
        (ev4, "ebus_tbna", "ebus", "11Rs", "", "22G/25G", "11Rs", "4 passes", "", "cyto", "Type 3 Elastography", False),
        (ev5, "ebus_tbna", "ebus", "4R", "", "22G/25G", "4R", "5 passes", "", "cyto", "Type 3 Elastography", False),
        (ev6, "ebus_tbna", "ebus", "7", "", "22G/25G", "7", "4 passes", "", "cyto", "Type 2 Elastography", False),
        (ev7, "ebus_tbna", "ebus", "4L", "", "22G/25G", "4L", "5 passes", "", "cyto", "Type 2 Elastography", False),
        (ev8, "ebus_tbna", "ebus", "11L", "", "22G/25G", "11L", "4 passes", "", "cyto", "Type 3 Elastography", False),
        (ev9, "ebus_tbna", "ebus", "Tracheal Mass", "", "22G/25G", "Mass", "8 passes", "", "cyto", "ROSE Positive", False),
        (ev10, "cryobiopsy", "cryobiopsy", "Tracheal Mass", "", "", "", "5 samples", "", "biopsy", "Mass", False),
        ("evt_11_trach", "other", "tracheostomy", "Trachea", "", "", "", "", "", "", "ENT Procedure", False),
    ]

    for ev in events_data:
        row = [SOURCE_FILE, NOTE_ID, ev[0], ev[1], ev[2], ev[3], ev[4], ev[5], ev[6], ev[7], ev[8], ev[9], ev[10], ev[11], "", "", "", "", "", "", "", "", ""]
        ws_event.append(row)

    ws_v3 = wb["V3_Procedure_Events"] if "V3_Procedure_Events" in wb.sheetnames else wb.create_sheet("V3_Procedure_Events")
    if ws_v3.max_row == 1:
        ws_v3.append(["note_id", "event_id", "type", "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station", "lesion.type", "lesion.size_mm", "method", "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote", "stent.size", "stent.material_or_brand", "catheter.size_fr", "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post", "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"])

    v3_rows = [
        [NOTE_ID, ev1, "therapeutic_aspiration", "airway", "R/L", "Mainstem", "", "", "", "aspiration", "", "", "", "thick secretions", "", "", "", "", "", "", "", "", ""],
        [NOTE_ID, ev2, "bal", "lung", "RML", "Lateral/Medial", "", "", "", "bal", "", '{"vol_instilled":40, "vol_returned":15}', "cytology", "", "", "", "", "", "", "", "", "", ""],
        [NOTE_ID, ev3, "rigid_bronchoscopy", "airway", "", "", "", "extrinsic_compression", "", "rigid_scope", '["rigid_barrel"]', "", "", "severe compression", "", "", "", "", "", "", "", "", ""],
        [NOTE_ID, ev4, "ebus_tbna", "ln_station", "", "", "11Rs", "", "", "tbna", '["needle"]', "", "cytology", "elastography_type_3", "", "", "", "", "", "", "", "", ""],
    ]
    for r in v3_rows:
        ws_v3.append(r)

    ws_json = wb["V3_Registry_JSON"] if "V3_Registry_JSON" in wb.sheetnames else wb.create_sheet("V3_Registry_JSON")
    if ws_json.max_row == 1:
        ws_json.append(["schema_version", "note_id", "json_output"])

    json_obj = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": []
    }
    ws_json.append(["v3.0", NOTE_ID, json.dumps(json_obj, indent=2)])

    wb.save(output_path)


if __name__ == "__main__":
    generate_excel()
