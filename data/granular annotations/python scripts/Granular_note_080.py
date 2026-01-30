import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import re
import os
import json
from datetime import datetime

# ==========================================
# INPUTS
# ==========================================
NOTE_ID = "note_080"
SOURCE_FILE = "note_080.txt"
PROCEDURE_DATE = "2026-01-12" # Placeholder, actual date not in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_080 SOURCE_FILE: note_080.txt INDICATION FOR OPERATION:  [REDACTED]is a 62 year old-year-old male who presents with lung nodules.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: R91.1 Solitary Lung Nodule
POSTOPERATIVE DIAGNOSIS:  R91.1 Solitary Lung Nodule
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31622 Dx bronchoscope/cell washing          
31623 Dx bronchoscope/brushing    
31624 Dx bronchoscope/lavage (BAL)    
31628 TBBX single lobe     
31632 TBBX additional lobes  
31629 TBNA single lobe   
31633 TBNA additional lobes   
31626 Fiducial marker placements, single or multiple     
31627 Navigational Bronchoscopy (computer assisted)
77012 Radiology / radiologic guidance for CT guided needle placement (CIOS)
76377 3D rendering with interpretation and reporting 
of CT, US, Tomo modality (ION Planning Station)
31899NFN BRONCHOSCOPY WITH ENDOBRONCHIAL ULTRASOUND (EBUS) OF MEDIASTINAL AND/OR HILAR LYMPH NODES WITHOUT BIOPSY
31654 Radial EBUS for peripheral lesion
76981 Ultrasound Elastography, Parenchyma of Organ
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required a Transbronchial Cryo biopsies, Robotic Navigation to more than one site, Multiple Bronchoalveolar lavages in different locations, Brushings at multiple locations, and Radial EBUS performed at multiple locations.
This resulted in >60% increased work due to Increased intensity, Time, Technical difficulty of procedure, and Physical and mental effort required.
Apply to: 31623 Dx bronchoscope/brushing    
31624 Dx bronchoscope/lavage (BAL)    
31628 TBBX single lobe     
31632 TBBX additional lobes  
31654 Radial EBUS for peripheral lesion.
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Linear EBUS 
Radial EBUS
Ion Robotic Bronchoscope
Disposable Bronchoscope
ESTIMATED BLOOD LOSS:   Minimum
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
Initial Airway Inspection Findings:
Normal appearing airway anatomy and mucosa bilaterally to the segmental level.
CT Chest scan was placed on separate planning station to generate 3D rendering of the pathway to target.
The navigational plan was reviewed and verified.  This was then loaded into robotic bronchoscopy platform.
Successful therapeutic aspiration was performed to clean out the Trachea (Middle 1/3), Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus.
Ventilation Parameters:
Mode	RR	TV	PEEP	FiO2	Flow Rate	Pmean
VCV	 	300	12	100	10	15
Robotic navigation bronchoscopy was performed with Ion platform.  Partial registration was used.
Ion robotic catheter was used to engage the Apical-Posterior Segment of LUL (LB1/2).
Target lesion is about 0.8 cm in diameter.   Under navigational guidance the ion robotic catheter was advanced to 1.0 cm away from the planned target.
Radial EBUS was performed to confirm that the location of the nodule is Eccentric.
The following features were noted: Continuous margin  and Absence of linear-discrete air bronchogram.
Cone Beam CT was performed: 3-D reconstructions were performed on an independent workstation.
Cios Spin system was used for evaluation of nodule location.  Low dose spin was performed to acquire CT imaging.
This was passed on to Ion platform system for reconstruction and nodule location.
The 3D images was interpreted on an independent workstation (Ion).
Using the newly acquired nodule location, the Ion robotic system was adjusted to the new targeted location.
I personally interpreted the cone beam CT and 3-D reconstruction.
Transbronchial needle aspiration was performed with 21G Needle through the extended working channel catheter.  Total 4 samples were collected.
Samples sent for Microbiology (Cultures/Viral/Fungal) and Cytology.
Transbronchial cryobiopsy was performed with 1.1mm cryoprobe via the extended working channel catheter.
Freeze time of 6 seconds were used.  Total 10 samples were collected.  Samples sent for Microbiology (Cultures/Viral/Fungal) and Pathology.
Fiducial marker was loaded with bone wax and placed under fluoroscopy guidance. Prior to withdraw of the bronchoscope.
Transbronchial brushing was performed with Protected cytology brush the extended working channel catheter.  Total 1 samples were collected.
Samples sent for Cytology.
Bronchial alveolar lavage was performed the extended working channel catheter.
Instilled 20 cc of NS, suction returned with 5 cc of NS.  Samples sent for Microbiology (Cultures/Viral/Fungal).
ROSE from ION procedure was noted to be:
No evidence of a malignant neoplasm
 
Right upper lobe 
Robotic navigation bronchoscopy was performed with Ion platform.
Partial registration was used.    Ion robotic catheter was used to engage the Apical Segment of RUL (RB1).
Target lesion is about 0.6 cm in diameter.   Under navigational guidance the ion robotic catheter was advanced to 1.0 cm away from the planned target.
Radial EBUS was performed to confirm that the location of the nodule is Eccentric.
The following features were noted: Continuous margin  and Absence of linear-discrete air bronchogram.
Cone Beam CT was performed: 3-D reconstructions were performed on an independent workstation.
Cios Spin system was used for evaluation of nodule location.  Low dose spin was performed to acquire CT imaging.
This was passed on to Ion platform system for reconstruction and nodule location.
The 3D images was interpreted on an independent workstation (Ion).
Using the newly acquired nodule location, the Ion robotic system was adjusted to the new targeted location.
I personally interpreted the cone beam CT and 3-D reconstruction.
Transbronchial needle aspiration was performed with 21G Needle through the extended working channel catheter.  Total 4 samples were collected.
Samples sent for Microbiology (Cultures/Viral/Fungal) and Cytology.
Transbronchial cryobiopsy was performed with 1.1mm cryoprobe via the extended working channel catheter.
Freeze time of 6 seconds were used.  Total 6 samples were collected.  Samples sent for Microbiology (Cultures/Viral/Fungal) and Pathology.
Transbronchial brushing was performed with Protected cytology brush the extended working channel catheter.  Total 1 samples were collected.
Samples sent for Cytology.
Bronchial alveolar lavage was performed the extended working channel catheter.
Instilled 20 cc of NS, suction returned with 5 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
ROSE from ION procedure was noted to be:
Atypical cells present but no evidence of malignant neoplasm
 
Successful therapeutic aspiration was performed to clean out the Trachea (Middle 1/3), Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus and blood.
Left upper lobe Bronchial alveolar lavage was performed at Apical-Posterior Segment of LUL (LB1/2), Anterior Segment of LUL (LB3), Superior Segment of Lingula (LB4), and Inferior Segment of Lingula (LB5).
Instilled 40 cc of NS, suction returned with 15 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
Right upper lobe Bronchial alveolar lavage was performed at Apical Segment of RUL (RB1), Posterior Segment of RUL (RB2), and Anterior Segment of RUL (RB3).
Instilled 40 cc of NS, suction returned with 15 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
EBUS-Findings
Indications: Diagnostic
Technique:
All lymph node stations were assessed.
Only those 5 mm or greater in short axis were sampled.
Lymph Nodes/Sites Inspected: 4R (lower paratracheal) node
4L (lower paratracheal) node
7 (subcarinal) node
11Rs lymph node
11Ri lymph node
11L lymph node
No immediate complications
Based upon the size and ultrasound appearance, biopsies were not taken.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
Elastography provided a semi-quantitative classification (Type 1–3), which was used to guide biopsy site selection and sampling strategy.
Lymph Nodes Evaluated:
Site 1: The 11L lymph node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was not sampled: Sampling this lymph node was not clinically indicated.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 2 elastographic pattern with mixed soft and stiff regions.
Given the ultrasound appearance and size, biopsies were not taken.
Site 2: The 4L (lower paratracheal) node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was not sampled: Sampling this lymph node was not clinically indicated.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 2 elastographic pattern with mixed soft and stiff regions.
Given the ultrasound appearance and size, biopsies were not taken.
Site 3: The 7 (subcarinal) node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was not sampled: Sampling this lymph node was not clinically indicated.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 2 elastographic pattern with mixed soft and stiff regions.
Given the ultrasound appearance and size, biopsies were not taken.
 
The patient tolerated the procedure well.
There were no immediate complications.  At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
•	Left upper lobe transbronchial needle aspiration, transbronchial cryobiopsy, transbronchial brush, sub segmental bronchoalveolar lavage, lobar bronchoalveolar lavage
•	Right upper lobe transbronchial needle aspiration, transbronchial cryobiopsy, transbronchial brush, sub segmental bronchoalveolar lavage, lobar bronchoalveolar lavage
IMPRESSION/PLAN: [REDACTED]is a 62 year old-year-old male who presents for bronchoscopy for lung nodules.
-Follow up bronchoscopic lab work
-Follow up CXR
"""

# ==========================================
# 1. SETUP & UTILS
# ==========================================
def create_workbook_if_missing():
    if not os.path.exists(TEMPLATE_PATH):
        wb = openpyxl.Workbook()
        # Create required sheets
        for sheet_name in [
            "Note_Text", "Note_Index", "Span_Annotations", 
            "Span_Hydrated", "Event_Log", 
            "V3_Procedure_Events", "V3_Registry_JSON"
        ]:
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
        # Remove default Sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        # Headers
        wb["Note_Text"].append(["note_id", "source_file", "note_text"])
        
        note_index_headers = [
            "source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes",
            "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", 
            "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy", 
            "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration", 
            "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation", 
            "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", 
            "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
            "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis", "pleural_biopsy", "fibrinolytic_therapy"
        ]
        wb["Note_Index"].append(note_index_headers)

        span_headers = [
            "source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", "match_index", 
            "start_char", "end_char", "span_len", "label", "normalized_value", "schema_field", "event_id", 
            "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"
        ]
        wb["Span_Annotations"].append(span_headers)
        wb["Span_Hydrated"].append(span_headers)
        
        wb.save(TEMPLATE_PATH)

# ==========================================
# 2. NOTE_INDEX (FLAGS)
# ==========================================
flags = {
    "diagnostic_bronchoscopy": 1,
    "bal": 1,
    "bronchial_wash": 1, # "cell washing" in codes
    "brushings": 1,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 1,
    "linear_ebus": 1,
    "radial_ebus": 1,
    "navigational_bronchoscopy": 1,
    "transbronchial_biopsy": 0, # Cryo preferred flag
    "transbronchial_cryobiopsy": 1,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# ==========================================
# 4. SPAN EXTRACTION (Anchor First)
# ==========================================
spans = []
span_id_counter = 1

def add_span(text, label, norm, field, event_id, context=None, section="Procedure"):
    global span_id_counter
    if text not in NOTE_TEXT:
        return
    
    # Context logic for hydration later
    match_index = 0 
    if NOTE_TEXT.count(text) > 1:
        # Simple heuristic: if we want the second occurrence, user must specify match_index
        # Here we rely on context_prefix if provided
        pass

    spans.append({
        "source_file": SOURCE_FILE,
        "note_id": NOTE_ID,
        "span_id": f"{NOTE_ID}_s{span_id_counter:03d}",
        "section_type": section,
        "context_prefix": context, # 120 chars before
        "span_text": text,
        "match_index": 0, # Default to 0, hydration updates this
        "start_char": "",
        "end_char": "",
        "span_len": f"=LEN(F{len(spans)+2})",
        "label": label,
        "normalized_value": norm,
        "schema_field": field,
        "event_id": event_id,
        "is_negated": "FALSE",
        "is_historical": "FALSE",
        "time_anchor": "",
        "reviewer": "Auto",
        "comments": "",
        "hydration_status": "needs_hydration"
    })
    span_id_counter += 1

# --- Event 1: Therapeutic Aspiration ---
evt_asp = "evt_01_asp"
add_span("Successful therapeutic aspiration was performed", "PROC_METHOD", "therapeutic aspiration", "method", evt_asp)
targets_asp = [
    "Trachea (Middle 1/3)", "Trachea (Distal 1/3)", "Right Mainstem", "Bronchus Intermedius", 
    "Left Mainstem", "Carina", "RUL Carina (RC1)", "RML Carina (RC2)", 
    "LUL Lingula Carina (Lc1)", "Left Carina (LC2)"
]
for t in targets_asp:
    add_span(t, "ANAT_AIRWAY", t, "target.anatomy_type", evt_asp, context="Successful therapeutic aspiration was performed to clean out the")

# --- Event 2: LUL Nodule (Apical-Posterior LB1/2) ---
evt_lul = "evt_02_lul"
# Nav
add_span("Robotic navigation bronchoscopy", "PROC_METHOD", "Navigational Bronchoscopy", "method", evt_lul, context="Ventilation Parameters")
add_span("Ion robotic catheter", "DEV_CATHETER", "Ion Robotic Catheter", "device", evt_lul, context="Partial registration was used.")
add_span("Apical-Posterior Segment of LUL (LB1/2)", "ANAT_LUNG_LOC", "LUL Apical-Posterior", "target.location.segment", evt_lul)
add_span("0.8 cm", "MEAS_SIZE", "8", "lesion.size_mm", evt_lul, context="Target lesion is about")
# Radial EBUS
add_span("Radial EBUS", "PROC_METHOD", "Radial EBUS", "method", evt_lul, context="navigational guidance the ion robotic catheter")
# TBNA
add_span("Transbronchial needle aspiration", "PROC_METHOD", "TBNA", "method", evt_lul, context="3-D reconstruction.")
add_span("21G Needle", "DEV_NEEDLE", "21G", "device", evt_lul, context="Transbronchial needle aspiration was performed with")
add_span("Total 4 samples were collected", "MEAS_COUNT", "4", "specimens_json", evt_lul, context="21G Needle through the extended working channel catheter")
# Cryo
add_span("Transbronchial cryobiopsy", "PROC_METHOD", "Transbronchial Cryobiopsy", "method", evt_lul, context="Samples sent for Microbiology (Cultures/Viral/Fungal) and Cytology.")
add_span("1.1mm cryoprobe", "DEV_INSTRUMENT", "1.1mm cryoprobe", "device", evt_lul, context="Transbronchial cryobiopsy was performed with")
add_span("Freeze time of 6 seconds", "MEAS_TIME", "6s", "measurements_json", evt_lul, context="Transbronchial cryobiopsy was performed with 1.1mm cryoprobe")
add_span("Total 10 samples were collected", "MEAS_COUNT", "10", "specimens_json", evt_lul, context="Freeze time of 6 seconds were used")
# Fiducial
add_span("Fiducial marker", "DEV_INSTRUMENT", "Fiducial Marker", "device", evt_lul, context="Samples sent for Microbiology (Cultures/Viral/Fungal) and Pathology.")
# Brush
add_span("Transbronchial brushing", "PROC_METHOD", "Transbronchial Brushing", "method", evt_lul, context="Prior to withdraw of the bronchoscope.")
add_span("Protected cytology brush", "DEV_INSTRUMENT", "Protected cytology brush", "device", evt_lul, context="Transbronchial brushing was performed with")
# BAL (Nodule specific)
add_span("Bronchial alveolar lavage", "PROC_METHOD", "BAL", "method", evt_lul, context="Samples sent for Cytology.")
add_span("Instilled 20 cc of NS", "MEAS_VOL", "20", "measurements_json", evt_lul, context="Bronchial alveolar lavage was performed the extended working channel catheter.")
# ROSE
add_span("No evidence of a malignant neoplasm", "OBS_ROSE", "negative", "findings_json", evt_lul, context="ROSE from ION procedure was noted to be:")

# --- Event 3: RUL Nodule (Apical RB1) ---
evt_rul = "evt_03_rul"
# Nav
add_span("Robotic navigation bronchoscopy", "PROC_METHOD", "Navigational Bronchoscopy", "method", evt_rul, context="Right upper lobe")
add_span("Apical Segment of RUL (RB1)", "ANAT_LUNG_LOC", "RUL Apical", "target.location.segment", evt_rul)
add_span("0.6 cm", "MEAS_SIZE", "6", "lesion.size_mm", evt_rul, context="Target lesion is about")
# Radial
add_span("Radial EBUS", "PROC_METHOD", "Radial EBUS", "method", evt_rul, context="advanced to 1.0 cm away from the planned target.")
# TBNA
add_span("Transbronchial needle aspiration", "PROC_METHOD", "TBNA", "method", evt_rul, context="I personally interpreted the cone beam CT and 3-D reconstruction.")
add_span("Total 4 samples were collected", "MEAS_COUNT", "4", "specimens_json", evt_rul, context="21G Needle through the extended working channel catheter.  Total 4")
# Cryo
add_span("Transbronchial cryobiopsy", "PROC_METHOD", "Transbronchial Cryobiopsy", "method", evt_rul, context="Samples sent for Microbiology (Cultures/Viral/Fungal) and Cytology.")
add_span("Total 6 samples were collected", "MEAS_COUNT", "6", "specimens_json", evt_rul, context="Freeze time of 6 seconds were used.")
# Brush
add_span("Transbronchial brushing", "PROC_METHOD", "Transbronchial Brushing", "method", evt_rul, context="Microbiology (Cultures/Viral/Fungal) and Pathology.")
# BAL (Nodule)
add_span("Bronchial alveolar lavage", "PROC_METHOD", "BAL", "method", evt_rul, context="Samples sent for Cytology.")
add_span("Instilled 20 cc of NS", "MEAS_VOL", "20", "measurements_json", evt_rul, context="Instilled 20 cc of NS, suction returned with 5 cc of NS.")
# ROSE
add_span("Atypical cells present but no evidence of malignant neoplasm", "OBS_ROSE", "atypical", "findings_json", evt_rul, context="ROSE from ION procedure was noted to be:")

# --- Event 4: LUL Lobar BAL ---
evt_bal_lul = "evt_04_bal_lul"
add_span("Left upper lobe Bronchial alveolar lavage", "PROC_METHOD", "BAL", "method", evt_bal_lul)
add_span("Apical-Posterior Segment of LUL (LB1/2)", "ANAT_LUNG_LOC", "LUL Apical-Posterior", "target.location.segment", evt_bal_lul, context="Left upper lobe Bronchial alveolar lavage was performed at")
add_span("Instilled 40 cc of NS", "MEAS_VOL", "40", "measurements_json", evt_bal_lul, context="Inferior Segment of Lingula (LB5).")

# --- Event 5: RUL Lobar BAL ---
evt_bal_rul = "evt_05_bal_rul"
add_span("Right upper lobe Bronchial alveolar lavage", "PROC_METHOD", "BAL", "method", evt_bal_rul)
add_span("Apical Segment of RUL (RB1)", "ANAT_LUNG_LOC", "RUL Apical", "target.location.segment", evt_bal_rul, context="Right upper lobe Bronchial alveolar lavage was performed at")
add_span("Instilled 40 cc of NS", "MEAS_VOL", "40", "measurements_json", evt_bal_rul, context="Anterior Segment of RUL (RB3).")

# --- Event 6: EBUS Inspection ---
evt_ebus = "evt_06_ebus"
add_span("All lymph node stations were assessed", "PROC_METHOD", "EBUS Inspection", "method", evt_ebus)
add_span("11L lymph node", "ANAT_LN_STATION", "11L", "target.station", evt_ebus, context="Site 1: The")
add_span("site was not sampled", "PROC_ACTION", "not sampled", "findings_json", evt_ebus, context="The lymph node was photographed. The")
add_span("4L (lower paratracheal) node", "ANAT_LN_STATION", "4L", "target.station", evt_ebus, context="Site 2: The")
add_span("7 (subcarinal) node", "ANAT_LN_STATION", "7", "target.station", evt_ebus, context="Site 3: The")

# --- Outcome ---
evt_outcome = "evt_07_outcome"
add_span("No immediate complications", "OUTCOME_COMPLICATION", "none", "outcomes.complications", evt_outcome, context="The patient tolerated the procedure well.")

# ==========================================
# 5. SPAN HYDRATION
# ==========================================
def generate_excel(template_path: str = TEMPLATE_PATH, output_path: str = OUTPUT_PATH) -> None:
    create_workbook_if_missing()
    wb = openpyxl.load_workbook(template_path)

    # Note_Text
    if "Note_Text" not in wb.sheetnames:
        wb.create_sheet("Note_Text")
    ws_text = wb["Note_Text"]
    if ws_text.max_row == 1:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # Note_Index
    if "Note_Index" not in wb.sheetnames:
        wb.create_sheet("Note_Index")
    ws_index = wb["Note_Index"]
    if ws_index.max_row == 1:
        # Minimal fallback header; preferred is to run with the official template.
        note_index_headers = [
            "source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes",
            "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
            "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy",
            "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration",
            "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
            "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation",
            "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
            "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis",
            "pleural_biopsy", "fibrinolytic_therapy"
        ]
        ws_index.append(note_index_headers)

    row_data = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Success", ""]
    header_row = [cell.value for cell in ws_index[1]]
    for h in header_row[8:]:
        row_data.append(flags.get(h, 0))
    ws_index.append(row_data)

    # Spans
    if "Span_Annotations" not in wb.sheetnames:
        wb.create_sheet("Span_Annotations")
    if "Span_Hydrated" not in wb.sheetnames:
        wb.create_sheet("Span_Hydrated")
    ws_annot = wb["Span_Annotations"]
    ws_hydrated = wb["Span_Hydrated"]

    for s in spans:
        ws_annot.append([
            s["source_file"], s["note_id"], s["span_id"], s["section_type"], s["context_prefix"],
            s["span_text"], s["match_index"], "", "", s["span_len"],
            s["label"], s["normalized_value"], s["schema_field"], s["event_id"],
            s["is_negated"], s["is_historical"], s["time_anchor"], s["reviewer"], s["comments"], "needs_hydration"
        ])

        text = s["span_text"]
        ctx = s["context_prefix"] or ""
        count = NOTE_TEXT.count(text)
        start = -1

        if count == 1:
            start = NOTE_TEXT.find(text)
            status = "hydrated_unique"
        elif count > 1 and ctx:
            matches = [m.start() for m in re.finditer(re.escape(text), NOTE_TEXT)]
            found = False
            for m in matches:
                window_start = max(0, m - 120)
                window = NOTE_TEXT[window_start:m]
                if ctx in window:
                    start = m
                    status = "hydrated_prefix_window"
                    found = True
                    break
            if not found:
                status = f"ambiguous_count={count}"
        else:
            status = f"ambiguous_count={count}"

        if start != -1:
            s["start_char"] = start
            s["end_char"] = start + len(text)
        s["hydration_status"] = status

        ws_hydrated.append([
            s["source_file"], s["note_id"], s["span_id"], s["section_type"], s["context_prefix"],
            s["span_text"], s["match_index"], s["start_char"], s["end_char"], s["span_len"],
            s["label"], s["normalized_value"], s["schema_field"], s["event_id"],
            s["is_negated"], s["is_historical"], s["time_anchor"], s["reviewer"], s["comments"], s["hydration_status"]
        ])

    # Event_Log
    if "Event_Log" not in wb.sheetnames:
        wb.create_sheet("Event_Log")
    ws_event = wb["Event_Log"]

    events_data = [
        {"id": evt_asp, "type": "Procedure", "method": "Therapeutic Aspiration", "target": "Trachea/Mainstems/Carinas", "device": "", "findings": "Cleared mucus"},
        {"id": evt_lul, "type": "Procedure", "method": "Navigational Bronchoscopy/TBNA/Cryo/BAL", "target": "LUL Apical-Posterior", "device": "Ion, 21G Needle, 1.1mm Cryoprobe, Fiducial", "measurements": "0.8cm lesion, 6s freeze", "specimens": "TBNA x4, Cryo x10, Brush x1, BAL", "findings": "ROSE Negative"},
        {"id": evt_rul, "type": "Procedure", "method": "Navigational Bronchoscopy/TBNA/Cryo/BAL", "target": "RUL Apical", "device": "Ion, 21G Needle, 1.1mm Cryoprobe", "measurements": "0.6cm lesion, 6s freeze", "specimens": "TBNA x4, Cryo x6, Brush x1, BAL", "findings": "ROSE Atypical"},
        {"id": evt_bal_lul, "type": "Procedure", "method": "Lobar BAL", "target": "LUL Segments", "measurements": "40cc instilled"},
        {"id": evt_bal_rul, "type": "Procedure", "method": "Lobar BAL", "target": "RUL Segments", "measurements": "40cc instilled"},
        {"id": evt_ebus, "type": "Procedure", "method": "EBUS Inspection", "stations": "4R,4L,7,11", "findings": "Not sampled (<10mm)"},
        {"id": evt_outcome, "type": "Outcome", "outcome_complication": "None"},
    ]

    for e in events_data:
        row = [
            SOURCE_FILE, NOTE_ID, e.get("id"), e.get("type"), e.get("method"), e.get("target"),
            e.get("device"), "", e.get("stations"), "", e.get("measurements"), e.get("specimens"),
            e.get("findings"), "FALSE", "Auto", "", "", "", "", "", "", "", e.get("outcome_complication", "")
        ]
        ws_event.append(row)

    # V3 JSON (append-only)
    if "V3_Registry_JSON" not in wb.sheetnames:
        wb.create_sheet("V3_Registry_JSON")
    ws_json = wb["V3_Registry_JSON"]
    registry_data = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": events_data
    }
    ws_json.append([json.dumps(registry_data, indent=2)])

    wb.save(output_path)


if __name__ == "__main__":
    generate_excel()
