import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import json
import os
import re

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_081"
SOURCE_FILE = "note_081.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_081 SOURCE_FILE: note_081.txt INDICATION FOR OPERATION:  [REDACTED]is a 48 year old-year-old female who presents with dyspnea.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: R91.8 Other nonspecific abnormal finding of lung field.
POSTOPERATIVE DIAGNOSIS:  R91.8 Other nonspecific abnormal finding of lung field.
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31622 Dx bronchoscope/cell washing          
31625 Endobronchial Biopsy(s)
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Disposable Bronchoscope
ESTIMATED BLOOD LOSS:   None
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
Initial Airway Inspection Findings:
The airway was inspected. The vocal cords were normal appearing.
Lidocaine was applied to the vocal cords and the airway.
The airway anatomy was notable for scarring at the subglottic space, the previous site of stenosis.
There was minimal narrowing at this site. There remainder of the airway was normal appearing to the segmental level bilaterally.
Successful therapeutic aspiration was performed to clean out the Vocal Cord, Subglottic, Trachea (Proximal 1/3), Trachea (Middle 1/3), Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus.
Endobronchial biopsy was performed at Subglottic.  Lesion was successfully removed.  Samples sent for Pathology.
 
The patient tolerated the procedure well.
There were no immediate complications.  At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
Subglottic endobronchial biopsies
IMPRESSION/PLAN: [REDACTED]is a 48 year old-year-old female who presents for bronchoscopy for dyspnea.
-Follow up bronchoscopic lab work"""

# -------------------------------------------------------------------------
# PROCEDURE FLAGS
# -------------------------------------------------------------------------
# Explicitly mentioned: 
# 31645 Therapeutic aspiration -> therapeutic_aspiration
# 31622 Dx bronchoscope/cell washing -> diagnostic_bronchoscopy, bronchial_wash
# 31625 Endobronchial Biopsy(s) -> endobronchial_biopsy

FLAGS = {
    # Bronchoscopy (23)
    'diagnostic_bronchoscopy': 1,
    'bal': 0,
    'bronchial_wash': 1,
    'brushings': 0,
    'endobronchial_biopsy': 1,
    'tbna_conventional': 0,
    'linear_ebus': 0,
    'radial_ebus': 0,
    'navigational_bronchoscopy': 0,
    'transbronchial_biopsy': 0,
    'transbronchial_cryobiopsy': 0,
    'therapeutic_aspiration': 1,
    'foreign_body_removal': 0,
    'airway_dilation': 0,
    'airway_stent': 0,
    'thermal_ablation': 0,
    'tumor_debulking_non_thermal': 0,
    'cryotherapy': 0,
    'blvr': 0,
    'peripheral_ablation': 0,
    'bronchial_thermoplasty': 0,
    'whole_lung_lavage': 0,
    'rigid_bronchoscopy': 0,
    # Pleural (7)
    'thoracentesis': 0,
    'chest_tube': 0,
    'ipc': 0,
    'medical_thoracoscopy': 0,
    'pleurodesis': 0,
    'pleural_biopsy': 0,
    'fibrinolytic_therapy': 0
}

# -------------------------------------------------------------------------
# SPAN ANNOTATIONS
# -------------------------------------------------------------------------
# Schema: [span_text, label, normalized_value, event_id, context_prefix (optional)]
# Event IDs: E1 (Inspection/Dx), E2 (Therapeutic Aspiration), E3 (Biopsy), E0 (Global/Complications)

SPANS_DATA = [
    # Indication
    ("dyspnea", "OUTCOME_SYMPTOMS", "dyspnea", "E1", "presents with "),
    
    # Instruments
    ("Disposable Bronchoscope", "DEV_INSTRUMENT", "Disposable Bronchoscope", "E1", None),
    
    # Inspection / Findings
    ("scarring", "OBS_LESION", "scarring", "E1", "notable for "),
    ("subglottic space", "ANAT_AIRWAY", "Subglottic", "E1", "scarring at the "),
    ("stenosis", "OBS_LESION", "stenosis", "E1", "previous site of "),
    ("minimal narrowing", "OUTCOME_AIRWAY_LUMEN_PRE", "minimal narrowing", "E1", "There was "),
    ("at this site", "ANAT_AIRWAY", "Subglottic", "E1", "minimal narrowing "),
    
    # Therapeutic Aspiration (E2)
    ("Therapeutic aspiration", "PROC_METHOD", "Therapeutic aspiration", "E2", "31645 "),
    ("therapeutic aspiration", "PROC_METHOD", "Therapeutic aspiration", "E2", "Successful "),
    ("clean out", "PROC_ACTION", "clean out", "E2", None),
    
    # Locations for Aspiration (E2)
    ("Vocal Cord", "ANAT_AIRWAY", "Vocal Cords", "E2", "clean out the "),
    ("Subglottic", "ANAT_AIRWAY", "Subglottic", "E2", "Vocal Cord, "),
    ("Trachea (Proximal 1/3)", "ANAT_AIRWAY", "Trachea Proximal", "E2", None),
    ("Trachea (Middle 1/3)", "ANAT_AIRWAY", "Trachea Mid", "E2", None),
    ("Trachea (Distal 1/3)", "ANAT_AIRWAY", "Trachea Distal", "E2", None),
    ("Right Mainstem", "ANAT_AIRWAY", "RMS", "E2", None),
    ("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "E2", None),
    ("Left Mainstem", "ANAT_AIRWAY", "LMS", "E2", None),
    ("Carina", "ANAT_AIRWAY", "Carina", "E2", "Left Mainstem, "),
    ("RUL Carina", "ANAT_AIRWAY", "RUL Carina", "E2", None),
    ("RML Carina", "ANAT_AIRWAY", "RML Carina", "E2", None),
    ("LUL Lingula Carina", "ANAT_AIRWAY", "LUL Lingula Carina", "E2", None),
    ("Left Carina", "ANAT_AIRWAY", "LC2", "E2", "and "),
    ("mucus", "OBS_LESION", "mucus", "E2", None),
    
    # Biopsy (E3)
    ("Endobronchial biopsy", "PROC_METHOD", "Endobronchial biopsy", "E3", None),
    ("Subglottic", "ANAT_AIRWAY", "Subglottic", "E3", "biopsy was performed at "),
    ("Lesion", "OBS_LESION", "Lesion", "E3", "at Subglottic.  "),
    ("successfully removed", "PROC_ACTION", "Removed", "E3", "Lesion was "),
    ("Samples", "OBS_ROSE", "Samples", "E3", None),
    
    # Global / Outcome
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "E0", None),
]

# -------------------------------------------------------------------------
# HELPER FUNCTIONS
# -------------------------------------------------------------------------

def load_or_create_wb(path):
    if os.path.exists(path):
        return openpyxl.load_workbook(path)
    return openpyxl.Workbook()

def ensure_sheet(wb, name, headers):
    if name in wb.sheetnames:
        ws = wb[name]
        # Clear existing data except headers if strictly following "generate" paradigm, 
        # but here we usually append or overwrite. 
        # For safety in this "generator" task, we assume we are filling a fresh template.
    else:
        ws = wb.create_sheet(name)
        ws.append(headers)
    return ws

def hydrate_span(text, span_text, context_prefix=None, match_index=None):
    """
    Finds start/end chars for a span.
    Returns: (start, end, status)
    """
    if not span_text or span_text not in text:
        return None, None, "not_found"
    
    matches = [m.start() for m in re.finditer(re.escape(span_text), text)]
    
    if len(matches) == 1:
        start = matches[0]
        return start, start + len(span_text), "hydrated_unique"
    
    if context_prefix:
        # Look for context in window before occurrence
        for m_start in matches:
            window_start = max(0, m_start - 120)
            preceding_text = text[window_start:m_start]
            if context_prefix in preceding_text:
                return m_start, m_start + len(span_text), "hydrated_prefix_window"
    
    if match_index is not None and 0 <= match_index < len(matches):
        start = matches[match_index]
        return start, start + len(span_text), "hydrated_match_index"
        
    return None, None, f"ambiguous_count={len(matches)}"

def generate_excel(template_path: str = TEMPLATE_PATH, output_path: str = OUTPUT_PATH) -> None:
    wb = load_or_create_wb(template_path)

    # 1. Note_Text
    ws_text = ensure_sheet(wb, "Note_Text", ["note_id", "source_file", "note_text"])
    ws_text.delete_rows(2, ws_text.max_row)
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws_index = ensure_sheet(wb, "Note_Index", [
        "source_file", "note_id", "encounter_id", "procedure_date", "site",
        "reviewer", "status", "free_text_notes"
    ] + list(FLAGS.keys()))

    for row in ws_index.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    index_row = [SOURCE_FILE, NOTE_ID, "", "", "", "", "Pending", ""]
    index_row += [FLAGS[k] for k in FLAGS]
    ws_index.append(index_row)

    # 3. Span_Annotations (Anchor First)
    ws_span = ensure_sheet(wb, "Span_Annotations", [
        "source_file", "note_id", "span_id", "section_type", "context_prefix",
        "span_text", "match_index", "start_char", "end_char", "span_len",
        "label", "normalized_value", "schema_field", "event_id",
        "is_negated", "is_historical", "time_anchor", "reviewer", "comments",
        "hydration_status"
    ])
    ws_span.delete_rows(2, ws_span.max_row)

    # 4. Span_Hydrated
    ws_hydrated = ensure_sheet(wb, "Span_Hydrated", [
        "source_file", "note_id", "span_id", "section_type", "context_prefix",
        "span_text", "match_index", "start_char", "end_char", "span_len",
        "label", "normalized_value", "schema_field", "event_id",
        "is_negated", "is_historical", "time_anchor", "reviewer", "comments",
        "hydration_status"
    ])
    ws_hydrated.delete_rows(2, ws_hydrated.max_row)

    span_id_counter = 1
    hydrated_rows = []

    for item in SPANS_DATA:
        s_text, label, norm, ev_id, ctx = item

        span_id = f"{NOTE_ID}_S{span_id_counter:03d}"
        span_id_counter += 1

        row_raw = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure", ctx,
            s_text, None, None, None, f"=LEN(F{span_id_counter})",
            label, norm, "", ev_id,
            False, False, "", "", "", "needs_hydration"
        ]
        ws_span.append(row_raw)

        start, end, status = hydrate_span(NOTE_TEXT, s_text, context_prefix=ctx)

        row_hyd = list(row_raw)
        row_hyd[7] = start
        row_hyd[8] = end
        row_hyd[9] = len(s_text) if start is not None else 0
        row_hyd[19] = status

        ws_hydrated.append(row_hyd)
        hydrated_rows.append(row_hyd)

    # 5. Event_Log
    ws_event = ensure_sheet(wb, "Event_Log", [
        "source_file", "note_id", "event_id", "event_type", "method",
        "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements",
        "specimens", "findings", "is_historical", "reviewer", "comments",
        "device_size", "device_material",
        "outcome_airway_lumen_pre", "outcome_airway_lumen_post",
        "outcome_symptoms", "outcome_pleural", "outcome_complication"
    ])
    ws_event.delete_rows(2, ws_event.max_row)

    events = {
        "E1": {
            "type": "Diagnostic",
            "method": "Inspection",
            "anatomy": ["Subglottic"],
            "findings": ["scarring", "stenosis"],
            "outcomes": {"outcome_airway_lumen_pre": "minimal narrowing", "outcome_symptoms": "dyspnea"},
        },
        "E2": {"type": "Therapeutic", "method": "Therapeutic aspiration", "anatomy": ["Vocal Cords", "Subglottic", "Trachea Proximal", "Trachea Mid", "Trachea Distal", "RMS", "Bronchus Intermedius", "LMS", "Carina", "RUL Carina", "RML Carina", "LUL Lingula Carina", "LC2"], "findings": ["mucus"]},
        "E3": {"type": "Therapeutic", "method": "Endobronchial biopsy", "anatomy": ["Subglottic"], "findings": ["Lesion"], "specimens": ["Subglottic endobronchial biopsies"], "comments": "Lesion removed"},
        "E0": {"type": "Global", "outcomes": {"outcome_complication": "None"}},
    }

    for eid, data in events.items():
        row = [
            SOURCE_FILE, NOTE_ID, eid, data.get("type", ""), data.get("method", ""),
            ", ".join(data.get("anatomy", [])),
            "", "", "", "", "",
            ", ".join(data.get("specimens", [])),
            ", ".join(data.get("findings", [])),
            False, "", data.get("comments", ""),
            "", "",
            data.get("outcomes", {}).get("outcome_airway_lumen_pre", ""),
            data.get("outcomes", {}).get("outcome_airway_lumen_post", ""),
            data.get("outcomes", {}).get("outcome_symptoms", ""),
            data.get("outcomes", {}).get("outcome_pleural", ""),
            data.get("outcomes", {}).get("outcome_complication", "")
        ]
        ws_event.append(row)

    # 6. V3_Procedure_Events
    ws_v3 = ensure_sheet(wb, "V3_Procedure_Events", [
        "note_id", "event_id", "type",
        "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station",
        "lesion.type", "lesion.size_mm",
        "method", "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote",
        "stent.size", "stent.material_or_brand", "catheter.size_fr",
        "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post",
        "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"
    ])
    ws_v3.delete_rows(2, ws_v3.max_row)

    v3_rows = []
    v3_rows.append([NOTE_ID, "E1", "Diagnostic", "Airway", "Subglottic", "", "", "Scarring", "", "Inspection", "[]", "[]", "[]", json.dumps(["scarring", "stenosis"]), "scarring at the subglottic space", "", "", "", "minimal narrowing", "", "dyspnea", "", ""])
    v3_rows.append([NOTE_ID, "E2", "Therapeutic", "Airway", "Multi-site", "", "", "Mucus", "", "Therapeutic Aspiration", "[]", "[]", "[]", json.dumps(["mucus"]), "Successful therapeutic aspiration... clean out...", "", "", "", "", "", "", "", ""])
    v3_rows.append([NOTE_ID, "E3", "Therapeutic", "Airway", "Subglottic", "", "", "Lesion", "", "Endobronchial Biopsy", "[]", "[]", json.dumps(["Subglottic biopsy"]), json.dumps(["Lesion"]), "Lesion was successfully removed", "", "", "", "", "Patent (Implied)", "", "", ""])
    v3_rows.append([NOTE_ID, "E0", "Global", "", "", "", "", "", "", "", "[]", "[]", "[]", "[]", "No immediate complications", "", "", "", "", "", "", "", "None"])

    for r in v3_rows:
        ws_v3.append(r)

    # 7. V3_Registry_JSON
    ws_json = ensure_sheet(wb, "V3_Registry_JSON", ["schema_version", "note_id", "json_output"])
    ws_json.delete_rows(2, ws_json.max_row)

    registry_obj = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": [
            {"event_id": "E1", "type": "Diagnostic", "method": "Inspection", "target": {"anatomy": "Subglottic"}, "findings": ["Scarring", "Stenosis"], "outcomes": {"airway": {"lumen_pre": "minimal narrowing"}, "symptoms": "dyspnea"}},
            {"event_id": "E2", "type": "Therapeutic", "method": "Therapeutic Aspiration", "target": {"anatomy": "Multi-site Airway"}, "findings": ["Mucus"]},
            {"event_id": "E3", "type": "Therapeutic", "method": "Endobronchial Biopsy", "target": {"anatomy": "Subglottic"}, "findings": ["Lesion"], "action": "Removed"},
        ],
        "global_outcomes": {"complications": "None"}
    }
    ws_json.append(["3.0", NOTE_ID, json.dumps(registry_obj, indent=2)])

    wb.save(output_path)
    print(f"Generated {output_path}")


if __name__ == "__main__":
    generate_excel()
