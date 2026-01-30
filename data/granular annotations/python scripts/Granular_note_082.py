import openpyxl
from openpyxl.utils import get_column_letter
import json
import os
import re

# =============================================================================
# 1. INPUT DATA
# =============================================================================
NOTE_ID = "note_082"
SOURCE_FILE = "note_082.txt"
PROCEDURE_DATE = ""  # Not explicitly in text, leave blank
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_082 SOURCE_FILE: note_082.txt INDICATION FOR OPERATION:  [REDACTED]is a 68 year old-year-old female who presents with airway stenosis.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
PROCEDURE:  
31646 Therapeutic aspiration subsequent episodes
31622 Dx bronchoscope/cell washing          
31624 Dx bronchoscope/lavage (BAL)    
31628 TBBX single lobe     
31632 TBBX additional lobes  
31627 Navigational Bronchoscopy (computer assisted)
77012 Radiology / radiologic guidance for CT guided needle placement (CIOS)
76377 3D rendering with interpretation and reporting of CT, US, Tomo modality (ION Planning Station)
31654 Radial EBUS for peripheral lesion
31630 Balloon dilation
31641 Destruction of tumor OR relief of 
stenosis by any method other than excision (eg. laser therapy, cryotherapy)
31635 Foreign body removal
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required a extensive ablation and debulking with excision along the left mainstem and left lower lobe with multiple dilations and multiple episodes of excision between other therapeutic modalities.
This resulted in >90% increased work due to Increased intensity, Time, Technical difficulty of procedure, and Physical and mental effort required.
Apply to: 31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy).
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
Radial EBUS
Ion Robotic Bronchoscope
Disposable Bronchoscope
ESTIMATED BLOOD LOSS:   Moderate
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
Initial Airway Inspection Findings:
The airway was inspected.
The trachea and right lung were intact and appropriate to the segmental level.
The left mainstem was inflamed and the left upper lobe was visualized, but the left lower lobe was again not visualized.
CT Chest scan was placed on separate planning station to generate 3D rendering of the pathway to target.
The navigational plan was reviewed and verified.  This was then loaded into robotic bronchoscopy platform.
Ventilation Parameters:
Mode	RR	TV	PEEP	FiO2	Flow Rate	Pmean
VCV	 	300	12	100	10	15
 
Robotic navigation bronchoscopy was performed with Ion platform.  Partial registration was used.
Ion robotic catheter was used to engage the left lower lobe presumed oriface.
Target lesion is about 1 cm in diameter.   Under navigational guidance the ion robotic catheter was advanced to 1.0 cm away from the planned target.
Radial EBUS was performed to confirm that the location of the target is not well visuzlized.
The following features were noted: soft tissue.
Cone Beam CT was performed: 3-D reconstructions were performed on an independent workstation.
Cios Spin system was used for evaluation of target location.  Low dose spin was performed to acquire CT imaging.
This was passed on to Ion platform system for reconstruction and nodule location.
The 3D images was interpreted on an independent workstation (Ion).
Using the newly acquired nodule location, the Ion robotic system was adjusted to the new targeted location.
I personally interpreted the cone beam CT and 3-D reconstruction.
The target airway was visualized and targets were placed ensure optimal trajectory of the tools to open the obstructed airway.
Transbronchial needle aspiration was performed with 19G Needle through the extended working channel catheter.  Total 2 samples were collected.
Samples sent for Microbiology (Cultures/Viral/Fungal) and Cytology.
Transbronchial biopsy was performed with precisor the extended working channel catheter.
Total 2 samples were collected.  Samples sent for Microbiology (Cultures/Viral/Fungal) and Pathology.
The robotic bronchoscope was advanced until airway lumen was visualized with the vision probe.
The catheter was then drawn back and the robotic bronchoscope was disconnected.
The therapeutic bronchoscope was then used to clear the airway. The left lung was irrigated with saline.
Granulation tissue and inflamed tissue was excised with the precisor, pulmonary, and large forceps.
Balloon dilation was performed at left lower lobe orifice.  8/9/10 Elation balloon was used to perform dilation to 10 mm at the left lower lobe orifice.
Total 1 inflations with dilation time of 60 seconds each.
Balloon dilation was performed at left lower lobe orifice.
10/11/12 Elation balloon was used to perform dilation to 12 mm at the left lower lobe orifice.
Total 3 inflations with dilation time of 60 seconds each.
The left upper lobe and the left lower lobe were better visualized.
Serial irrigation with iced saline was performed with excision of granulation tissue and blood clot/foreign body.
The 2.5 cm microwave catheter was used to ablate the inflamed tissue for three minutes(3.5KJ and 67 degrees Celsius reached), then a pause for further and extensive excision, followed by one minute (1.3 KJ and 49 degrees Celsius reached)
The patient tolerated the procedure well.
There were no immediate complications.  At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
•	Left lower lobe transbronchial needle aspiration, transbronchial forceps biopsies
•	Left mainstem and left lower lobe biopsies
•	Left lower lobe bronchoalveolar lavage
IMPRESSION/PLAN: [REDACTED]is a 68 year old-year-old female who presents for bronchoscopy for airway stenosis.
-Follow up CXR
-Follow up bronchoscopic lab work
-Plan for follow up bronchoscopy in ~1 week"""

# =============================================================================
# 2. CONFIG & SCHEMA
# =============================================================================

# Flags for Index Sheet
PROCEDURE_FLAGS = [
    "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
    "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy",
    "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration",
    "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
    "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation",
    "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
    "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis",
    "pleural_biopsy", "fibrinolytic_therapy"
]

# =============================================================================
# 3. EXTRACTION LOGIC
# =============================================================================

def extract_flags(text):
    text_lower = text.lower()
    flags = {k: 0 for k in PROCEDURE_FLAGS}

    # Direct mapping based on note content
    if "31627" in text or "navigational" in text_lower or "ion platform" in text_lower:
        flags["navigational_bronchoscopy"] = 1
    if "31654" in text or "radial ebus" in text_lower:
        flags["radial_ebus"] = 1
    if "31624" in text or "bal" in text_lower or "lavage" in text_lower:
        flags["bal"] = 1
    if "31622" in text or "cell washing" in text_lower:
        flags["bronchial_wash"] = 1
    if "31630" in text or "dilation" in text_lower:
        flags["airway_dilation"] = 1
    if "31646" in text or "therapeutic aspiration" in text_lower:
        flags["therapeutic_aspiration"] = 1
    if "31628" in text or "tbbx" in text or "transbronchial biopsy" in text_lower:
        flags["transbronchial_biopsy"] = 1
    if "31641" in text or "ablate" in text_lower or "microwave catheter" in text_lower:
        flags["thermal_ablation"] = 1
    if "excision" in text_lower and ("forceps" in text_lower or "precisor" in text_lower):
        flags["tumor_debulking_non_thermal"] = 1
    if "31635" in text or "foreign body removal" in text_lower:
        flags["foreign_body_removal"] = 1
    
    # TBNA - Mentioned "Transbronchial needle aspiration... 19G Needle"
    if "transbronchial needle aspiration" in text_lower or "tbna" in text_lower:
        flags["tbna_conventional"] = 1

    # Base diagnostic bronchoscope always true if others present
    flags["diagnostic_bronchoscopy"] = 1
    
    return flags

def generate_spans(text):
    spans = []
    
    # Helper to add span
    def add_span(text_match, label, normalized, field, event_id, context_prefix=None, comments=""):
        if not text_match: return
        spans.append({
            "span_text": text_match,
            "label": label,
            "normalized_value": normalized,
            "schema_field": field,
            "event_id": event_id,
            "context_prefix": context_prefix,
            "comments": comments
        })

    # --- Event 1: Navigation ---
    add_span("Ion platform", "PROC_METHOD", "Navigational Bronchoscopy", "method", "evt1")
    add_span("robotic catheter", "DEV_CATHETER", "Ion Catheter", "devices", "evt1")
    add_span("left lower lobe", "ANAT_LUNG_LOC", "Left Lower Lobe", "target.location.lobe", "evt1", context_prefix="engage the")

    # --- Event 2: Radial EBUS ---
    add_span("Radial EBUS", "PROC_METHOD", "Radial EBUS", "method", "evt2", context_prefix="Therapeutic Bronchoscope\n")
    add_span("1.0 cm", "MEAS_SIZE", "1.0 cm", "lesion.size_mm", "evt2", context_prefix="advanced to")

    # --- Event 3: TBNA ---
    add_span("Transbronchial needle aspiration", "PROC_METHOD", "TBNA", "method", "evt3")
    add_span("19G Needle", "DEV_NEEDLE", "19G", "device", "evt3")
    add_span("2 samples", "MEAS_COUNT", "2", "counts", "evt3", context_prefix="Total")

    # --- Event 4: TBBX ---
    add_span("Transbronchial biopsy", "PROC_METHOD", "Transbronchial Biopsy", "method", "evt4")
    add_span("precisor", "DEV_INSTRUMENT", "Precisor", "devices", "evt4")
    add_span("extended working channel catheter", "DEV_CATHETER", "EWC Catheter", "devices", "evt4", context_prefix="precisor the")

    # --- Event 5: Debulking / Excision ---
    add_span("excised", "PROC_ACTION", "Excision", "method", "evt5", context_prefix="inflamed tissue was")
    add_span("large forceps", "DEV_INSTRUMENT", "Large Forceps", "devices", "evt5")
    add_span("Granulation tissue", "OBS_LESION", "Granulation Tissue", "findings", "evt5")

    # --- Event 6: Dilation 1 ---
    add_span("Balloon dilation", "PROC_METHOD", "Balloon Dilation", "method", "evt6", context_prefix="performed at left lower")
    add_span("8/9/10 Elation balloon", "DEV_INSTRUMENT", "Elation Balloon 8-9-10mm", "devices", "evt6")
    add_span("dilation to 10 mm", "OUTCOME_AIRWAY_LUMEN_POST", "10 mm", "outcomes.airway.lumen_post", "evt6")
    add_span("left lower lobe orifice", "ANAT_AIRWAY", "Left Lower Lobe Orifice", "target.anatomy_type", "evt6")

    # --- Event 7: Dilation 2 ---
    add_span("10/11/12 Elation balloon", "DEV_INSTRUMENT", "Elation Balloon 10-11-12mm", "devices", "evt7")
    add_span("dilation to 12 mm", "OUTCOME_AIRWAY_LUMEN_POST", "12 mm", "outcomes.airway.lumen_post", "evt7")

    # --- Event 8: Ablation ---
    add_span("microwave catheter", "DEV_CATHETER", "Microwave Catheter", "devices", "evt8")
    add_span("2.5 cm", "DEV_CATHETER_SIZE", "2.5 cm", "catheter.size_fr", "evt8", context_prefix="The") # Length really, but maps best here
    add_span("ablate", "PROC_ACTION", "Ablation", "method", "evt8")
    add_span("3.5KJ", "MEAS_ENERGY", "3.5KJ", "measurements_json", "evt8")
    add_span("67 degrees Celsius", "MEAS_TEMP", "67 C", "measurements_json", "evt8")
    add_span("1.3 KJ", "MEAS_ENERGY", "1.3 KJ", "measurements_json", "evt8")

    # --- Event 9: Clearance/Complications/Global ---
    add_span("Therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", "method", "evt9")
    add_span("no immediate complications", "OUTCOME_COMPLICATION", "None", "outcomes.complications", "evt9")

    return spans

def hydrate_spans(spans, text):
    hydrated = []
    
    for s in spans:
        span_text = s["span_text"]
        
        # Hydration logic
        start = -1
        matches = [m.start() for m in re.finditer(re.escape(span_text), text)]
        
        status = "ambiguous"
        
        if len(matches) == 1:
            start = matches[0]
            status = "hydrated_unique"
        elif len(matches) > 1:
            # Context match
            if s.get("context_prefix"):
                prefix = s["context_prefix"]
                best_dist = 9999
                best_start = -1
                for m in matches:
                    # Look backwards 150 chars
                    window = text[max(0, m - 150):m]
                    if prefix in window:
                        start = m
                        status = "hydrated_prefix_window"
                        break
            
            # Fallback if context didn't help but matches exist
            if start == -1:
                # Default to first if specific enough or just mark ambiguous
                # For this script, we'll pick first to ensure valid offset if ambiguous
                start = matches[0] 
                status = "hydrated_fallback_first"

        if start != -1:
            end = start + len(span_text)
            s_copy = s.copy()
            s_copy["start_char"] = start
            s_copy["end_char"] = end
            s_copy["span_len"] = len(span_text)
            s_copy["hydration_status"] = status
            hydrated.append(s_copy)
            
    return hydrated

# =============================================================================
# 4. EXCEL GENERATION
# =============================================================================

def create_workbook():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except:
        wb = openpyxl.Workbook()
        # Create default sheets if template missing (fallback)
        for name in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", 
                     "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            if name not in wb.sheetnames:
                wb.create_sheet(name)

    # 1. Note_Text
    ws = wb["Note_Text"]
    if ws.max_row == 1:
        ws.append(["note_id", "source_file", "note_text"])
    ws.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws = wb["Note_Index"]
    flags = extract_flags(NOTE_TEXT)
    # Header check handled by template existence, else we append
    # Construct row: Meta cols then flags
    # Meta: source_file, note_id, encounter_id, procedure_date, site, reviewer, status, free_text
    meta_row = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Draft", ""]
    flag_vals = [flags.get(f, 0) for f in PROCEDURE_FLAGS]
    ws.append(meta_row + flag_vals)

    # 3. Span_Annotations
    ws = wb["Span_Annotations"]
    raw_spans = generate_spans(NOTE_TEXT)
    # source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, start, end, len, label, norm, field, event, neg, hist, time, rev, comm, hydra
    span_id_counter = 1
    for s in raw_spans:
        row = [
            SOURCE_FILE, NOTE_ID, f"span_{span_id_counter:03d}", "Procedure",
            s.get("context_prefix", ""), s["span_text"], "", "", "", "",
            s["label"], s["normalized_value"], s["schema_field"], s["event_id"],
            "FALSE", "FALSE", "", "", s.get("comments", ""), "needs_hydration"
        ]
        ws.append(row)
        s["span_id"] = f"span_{span_id_counter:03d}"
        span_id_counter += 1

    # 4. Span_Hydrated
    ws = wb["Span_Hydrated"]
    hydrated_spans = hydrate_spans(raw_spans, NOTE_TEXT)
    for s in hydrated_spans:
        row = [
            SOURCE_FILE, NOTE_ID, s["span_id"], "Procedure",
            s.get("context_prefix", ""), s["span_text"], "", s["start_char"], s["end_char"], s["span_len"],
            s["label"], s["normalized_value"], s["schema_field"], s["event_id"],
            "FALSE", "FALSE", "", "", s.get("comments", ""), s["hydration_status"]
        ]
        ws.append(row)

    # 5. Event_Log (Simplified for script)
    ws = wb["Event_Log"]
    # Group by event_id
    events = {}
    for s in raw_spans:
        eid = s["event_id"]
        if eid not in events: events[eid] = {"method": "", "device": "", "anatomy": "", "outcome_lumen_post": "", "outcome_compl": ""}
        if s["schema_field"] == "method": events[eid]["method"] = s["normalized_value"]
        if s["schema_field"] == "devices": events[eid]["device"] = s["normalized_value"]
        if s["schema_field"] == "target.anatomy_type" or s["schema_field"] == "target.location.lobe": events[eid]["anatomy"] = s["normalized_value"]
        if s["schema_field"] == "outcomes.airway.lumen_post": events[eid]["outcome_lumen_post"] = s["normalized_value"]
        if s["schema_field"] == "outcomes.complications": events[eid]["outcome_compl"] = s["normalized_value"]

    for eid, data in events.items():
        row = [
            SOURCE_FILE, NOTE_ID, eid, "Procedure", data["method"],
            data["anatomy"], data["device"], "", "", "", "",
            "", "", "", "", "",
            "", "", "", data["outcome_lumen_post"], "", "", data["outcome_compl"]
        ]
        ws.append(row)

    # 6. V3_Registry_JSON
    ws = wb["V3_Registry_JSON"]
    json_obj = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True if "None" in events.get("evt9", {}).get("outcome_compl", "") else False,
        "procedures": list(events.keys())
    }
    ws.append([json.dumps(json_obj, indent=2)])

    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    create_workbook()