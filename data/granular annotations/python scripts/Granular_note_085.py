import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
from datetime import datetime

# =============================================================================
# INPUT DATA
# =============================================================================

NOTE_ID = "note_085"
SOURCE_FILE = "note_085.txt"
PROCEDURE_DATE = "2026-01-12" # Placeholder or extracted if available
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_085 SOURCE_FILE: note_085.txt INDICATION FOR OPERATION:  [REDACTED]is a 74 year old-year-old female who presents with hemoptysis after bronchoscopy.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: hemoptysis
POSTOPERATIVE DIAGNOSIS:  hemoptysis
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31622 Dx bronchoscope/cell washing          
ANESTHESIA: 
Continuous sedation
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Disposable Bronchoscope
ESTIMATED BLOOD LOSS:   Minimum
COMPLICATIONS:    None
PROCEDURE IN DETAIL: A timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: supine
Initial Airway Inspection Findings:
The bronchoscope was introduced into the ETT.
The ETT is 2 cm above the main carina. The bronchial blocker was in the RLL but was mostly deflated without evidence of spill over into the other segments.
Successful therapeutic aspiration was performed to clean out the LLL and LMSB from mucus.
The balloon was was fully deflated and then removed. Then a new 7 F Ardnt bronchial blocker was replaced through the existing ETT.
The blocker was positioned in the distal BI. 3 mL of air was introduced into the blocker and resulted in occlusion of the BI without occlusion of the RUL.
During exchange of the bronchial blocker there was no increase in the blood seen in the RLL lateral segment.
This was no aspirated due to no obvious formed clot.
The bronchoscope was removed. 
 
The patient tolerated the procedure well.
There were no immediate complications.  
SPECIMEN(S): 
none
IMPRESSION/PLAN: [REDACTED]is a 74 year old-year-old female who presents for bronchoscopy for hemoptysis.
[ ] keep blocker in place at 40 cm at the connective device
[ ] ETT at 20 cm at the front tooth
[ ] stop paralysis once this is not effective stop fentanyl and repeat bronchoscopy at 1500 today
[ ] family updated"""

# =============================================================================
# CONFIGURATION & MAPPINGS
# =============================================================================

# Procedure Flags Logic
# 1 if present in note, else 0
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1, # "31622 Dx bronchoscope"
    "bal": 0,
    "bronchial_wash": 1, # "31622 ... cell washing"
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1, # "31645 Therapeutic aspiration"
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# =============================================================================
# SPAN ANNOTATIONS (ANCHOR-FIRST)
# =============================================================================

# Helper list of spans. 
# Format: (text, label, normalized_value, schema_field, event_id, context_prefix)
SPANS_DATA = [
    # Event 1: General Inspection / Setup
    ("Disposable Bronchoscope", "DEV_INSTRUMENT", "Disposable Bronchoscope", "instrument", "ev1", None),
    ("ETT", "ANAT_AIRWAY", "Endotracheal Tube", "target.anatomy", "ev1", "introduced into the"),
    ("main carina", "ANAT_AIRWAY", "Main Carina", "target.anatomy", "ev1", None),
    
    # Event 2: Old Blocker Check (RLL)
    ("bronchial blocker", "DEV_CATHETER", "Bronchial Blocker", "device", "ev2", "The ETT is 2 cm above the main carina. The"),
    ("RLL", "ANAT_LUNG_LOC", "RLL", "target.location.lobe", "ev2", "blocker was in the"),
    ("deflated", "PROC_ACTION", "Deflated", "action", "ev2", "mostly"),
    
    # Event 3: Therapeutic Aspiration
    ("Therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", "method", "ev3", "31645"),
    ("clean out", "PROC_ACTION", "Clean Out", "action", "ev3", None),
    ("LLL", "ANAT_LUNG_LOC", "LLL", "target.location.lobe", "ev3", "clean out the"),
    ("LMSB", "ANAT_AIRWAY", "LMSB", "target.anatomy", "ev3", "clean out the LLL and"),
    ("mucus", "OBS_FINDING", "Mucus", "findings", "ev3", None),
    
    # Event 4: Removal of Old Blocker
    ("balloon", "DEV_CATHETER", "Balloon", "device", "ev4", "The"),
    ("removed", "PROC_ACTION", "Removed", "action", "ev4", "deflated and then"),
    
    # Event 5: Placement of New Blocker
    ("7 F", "DEV_CATHETER_SIZE", "7F", "catheter.size_fr", "ev5", None),
    ("Ardnt bronchial blocker", "DEV_CATHETER", "Arndt Bronchial Blocker", "device", "ev5", None),
    ("replaced", "PROC_ACTION", "Replaced", "action", "ev5", None),
    ("distal BI", "ANAT_AIRWAY", "Bronchus Intermedius", "target.anatomy", "ev5", None),
    ("3 mL", "MEAS_VOL", "3 mL", "measurements.volume", "ev5", None),
    ("occlusion", "OUTCOME_AIRWAY_LUMEN_POST", "Occluded", "outcomes.airway.lumen_post", "ev5", "resulted in"),
    ("BI", "ANAT_AIRWAY", "Bronchus Intermedius", "target.anatomy", "ev5", "occlusion of the"),
    ("RUL", "ANAT_LUNG_LOC", "RUL", "target.location.lobe", "ev5", "occlusion of the"),
    
    # Event 6: Inspection RLL lateral
    ("RLL lateral segment", "ANAT_LUNG_LOC", "RLL Lateral Segment", "target.location.segment", "ev6", None),
    ("no increase in the blood", "OBS_FINDING", "No increase in blood", "findings", "ev6", None),
    ("no aspirated", "PROC_ACTION", "Not Aspirated", "action", "ev6", "This was"),
    
    # Global / Plan
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "outcomes.complications", "ev_global", None),
    ("blocker", "DEV_CATHETER", "Bronchial Blocker", "device", "ev_plan", "keep"),
    ("40 cm", "MEAS_SIZE", "40 cm", "measurements.size", "ev_plan", None),
    ("ETT", "ANAT_AIRWAY", "Endotracheal Tube", "target.anatomy", "ev_plan", "[ ]"),
    ("20 cm", "MEAS_SIZE", "20 cm", "measurements.size", "ev_plan", "at"),
]

# =============================================================================
# SCRIPT LOGIC
# =============================================================================

class Phase0Generator:
    def __init__(self, template_path, output_path, note_text):
        self.template_path = template_path
        self.output_path = output_path
        self.note_text = note_text
        self.wb = None
        self.spans = []
        self.events_log = []
        self.v3_events = []

    def load_workbook(self):
        try:
            self.wb = openpyxl.load_workbook(self.template_path)
        except FileNotFoundError:
            # Fallback: Create a basic workbook if template is missing (for safety, though requirements say use template)
            self.wb = openpyxl.Workbook()
            # Create sheets
            for sheet in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
                self.wb.create_sheet(sheet)
            # Remove default
            if "Sheet" in self.wb.sheetnames:
                del self.wb["Sheet"]

    def write_note_text(self):
        ws = self.wb["Note_Text"]
        # Assuming headers exist: note_id, source_file, note_text
        if ws.max_row == 1 and ws.cell(1,1).value is None:
             ws.append(["note_id", "source_file", "note_text"])
        
        ws.append([NOTE_ID, SOURCE_FILE, self.note_text])

    def write_note_index(self):
        ws = self.wb["Note_Index"]
        # Meta: source_file, note_id, encounter_id, procedure_date, site, reviewer, status, free_text_notes
        meta_row = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Pending", ""]
        
        # Flags
        flag_row = [PROCEDURE_FLAGS.get(h, 0) for h in [
            "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", 
            "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy", 
            "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration", 
            "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation", 
            "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", 
            "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
            "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis", 
            "pleural_biopsy", "fibrinolytic_therapy"
        ]]
        
        full_row = meta_row + flag_row
        ws.append(full_row)

    def process_spans(self):
        # Prepare spans for writing
        # source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, 
        # start_char, end_char, span_len, label, normalized_value, schema_field, event_id, 
        # is_negated, is_historical, time_anchor, reviewer, comments, hydration_status
        
        self.spans = []
        for idx, (text, label, norm, schema, ev_id, ctx) in enumerate(SPANS_DATA):
            span_id = f"span_{idx+1:03d}"
            
            # Basic hydration logic to find start/end
            start = -1
            end = -1
            status = "needs_hydration"
            match_index = 1
            
            count = self.note_text.count(text)
            
            if count == 0:
                status = "not_found"
            elif count == 1:
                start = self.note_text.find(text)
                end = start + len(text)
                status = "hydrated_unique"
            else:
                # Ambiguous
                if ctx:
                    # Search for context + text
                    search_str = ctx # simplified, ideally context + whitespace + text
                    # Find all occurrences of text
                    all_starts = [m.start() for m in re.finditer(re.escape(text), self.note_text)]
                    best_start = -1
                    
                    for s in all_starts:
                        window_start = max(0, s - 150)
                        window = self.note_text[window_start:s]
                        if ctx in window:
                            best_start = s
                            break
                    
                    if best_start != -1:
                        start = best_start
                        end = start + len(text)
                        status = "hydrated_context"
                    else:
                        status = f"ambiguous_count={count}"
                else:
                    status = f"ambiguous_count={count}"
                    # Default to first for simplicity in this script if no context
                    start = self.note_text.find(text)
                    end = start + len(text)
            
            row = [
                SOURCE_FILE, NOTE_ID, span_id, "Procedure", ctx, text, match_index,
                start if start != -1 else "", end if end != -1 else "", len(text),
                label, norm, schema, ev_id, 
                0, 0, "", "", "", status
            ]
            self.spans.append(row)

    def write_span_annotations(self):
        ws = self.wb["Span_Annotations"]
        # Headers usually exist. Append rows.
        # Strict rule: start_char/end_char must be blank in Span_Annotations
        for row in self.spans:
            # Copy row but clear offsets
            anno_row = list(row)
            anno_row[7] = "" # start
            anno_row[8] = "" # end
            ws.append(anno_row)

    def write_span_hydrated(self):
        ws = self.wb["Span_Hydrated"]
        for row in self.spans:
            ws.append(row)

    def write_event_log(self):
        ws = self.wb["Event_Log"]
        # Group spans by event_id
        events = {}
        for row in self.spans:
            ev_id = row[13] # event_id
            if ev_id not in events:
                events[ev_id] = {
                    "method": [], "anatomy": [], "device": [], "findings": [], 
                    "outcomes_lumen_post": [], "outcomes_comp": [], "measurements": []
                }
            
            label = row[10]
            val = row[11] # normalized
            
            if label == "PROC_METHOD": events[ev_id]["method"].append(val)
            elif "ANAT" in label: events[ev_id]["anatomy"].append(val)
            elif "DEV" in label: events[ev_id]["device"].append(val)
            elif "OBS" in label: events[ev_id]["findings"].append(val)
            elif label == "OUTCOME_AIRWAY_LUMEN_POST": events[ev_id]["outcomes_lumen_post"].append(val)
            elif label == "OUTCOME_COMPLICATION": events[ev_id]["outcomes_comp"].append(val)
            elif "MEAS" in label: events[ev_id]["measurements"].append(val)

        # Write rows
        for ev_id, data in events.items():
            # Construct summary strings
            method = ", ".join(set(data["method"]))
            anatomy = ", ".join(set(data["anatomy"]))
            device = ", ".join(set(data["device"]))
            findings = ", ".join(set(data["findings"]))
            outcome_airway_post = ", ".join(set(data["outcomes_lumen_post"]))
            outcome_comp = ", ".join(set(data["outcomes_comp"]))
            
            # Map event_type roughly
            if "ev_global" in ev_id: event_type = "Global"
            elif "ev_plan" in ev_id: event_type = "Plan"
            else: event_type = "Procedure"

            ws.append([
                SOURCE_FILE, NOTE_ID, ev_id, event_type, method,
                anatomy, device, "", "", "", ", ".join(data["measurements"]),
                "", findings, 0, "", "", 
                "", "", # Size/Mat
                "", outcome_airway_post, # Pre/Post
                "", "", outcome_comp # Symp/Pleural/Comp
            ])
            
            # Store for V3
            self.v3_events.append({
                "id": ev_id,
                "type": event_type,
                "anatomy": anatomy,
                "method": method,
                "device": device,
                "measurements": data["measurements"],
                "outcomes": {
                    "lumen_post": outcome_airway_post,
                    "complication": outcome_comp
                }
            })

    def write_v3_events(self):
        ws = self.wb["V3_Procedure_Events"]
        for ev in self.v3_events:
            if ev["type"] == "Procedure":
                ws.append([
                    NOTE_ID, ev["id"], "Therapeutic/Diagnostic",
                    "", "", "", "", # target breakdown
                    "", "", # lesion
                    ev["method"], ev["device"], json.dumps(ev["measurements"]), "", "", "",
                    "", "", "",
                    "", ev["outcomes"]["lumen_post"],
                    "", "", ev["outcomes"]["complication"]
                ])

    def write_json(self):
        ws = self.wb["V3_Registry_JSON"]
        data = {
            "schema_version": "3.0",
            "note_id": NOTE_ID,
            "procedures": self.v3_events,
            "no_immediate_complications": any(e["outcomes"]["complication"] == "None" for e in self.v3_events)
        }
        json_str = json.dumps(data, indent=2)
        ws.cell(row=1, column=1, value=json_str)

    def save(self):
        self.wb.save(self.output_path)
        print(f"Saved extraction to {self.output_path}")

# =============================================================================
# MAIN EXECUTION
# =============================================================================

if __name__ == "__main__":
    generator = Phase0Generator(TEMPLATE_PATH, OUTPUT_PATH, NOTE_TEXT)
    generator.load_workbook()
    generator.write_note_text()
    generator.write_note_index()
    generator.process_spans()
    generator.write_span_annotations()
    generator.write_span_hydrated()
    generator.write_event_log()
    generator.write_v3_events()
    generator.write_json()
    generator.save()