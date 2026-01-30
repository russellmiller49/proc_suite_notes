import sys
import os
import json
import re
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

# =============================================================================
# 1. CONSTANTS & INPUTS
# =============================================================================
NOTE_ID = "note_086"
SOURCE_FILE = "note_086.txt"
PROCEDURE_DATE = "2026-01-12" # Using context date as placeholder or blank if preferred
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_086 SOURCE_FILE: note_086.txt INDICATION FOR OPERATION:  [REDACTED] is a 57 year old-year-old male who presents with bronchial stenosis.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
PROCEDURE:  
31899 Unlisted Procedure (Trach Change with Mature Tract or Procedure NOS)
31645 Therapeutic aspiration initial episode
31624 Dx bronchoscope/lavage (BAL)    
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
PROCEDURE IN DETAIL:
A timeout was performed (confirming the patient's name, procedure type, and procedure location).
Sedation initiated.  LMA placed.
The Flexible Therapeutic Bronchoscope was advanced for airway examination.
Endobronchial topical lidocaine applied to the vocal cords, main carina, right carina 1, and left carina 2.
Initial Airway Examination Findings:
Larynx: Not fully assessed due to bronchoscopy introduction through LMA.
Vocal Chords: Normal without mass/lesions
Trachea: Mildy tortuous, otherwise normal.
Main Carina: Sharp
Right Lung Proximal Airways: Right Lung Proximal Airways: The right anastomosis site was intact with visible intact blue sutures at site;
mild stenosis of right anastomosis without dehiscence. The RUL and RB1-3 were normal to segmental level.
The bronchus intermedius and more distal lobes/airways are mildly rotated clockwise.
The RML bronchus with moderate stenosis - estimate ~5mm patent; unable to traverse with therapeutic bronchoscope.
The RML mucosa has an anthracotic appearance to it.  However, this is overall the best I have seen his RML look before an intervention;
no associated granulation tissue.  Able to visualize patent RB5/RB5.
Left Lung Proximal Airways: The left anastomosis site was intact with visible intact blue sutures at site;
mild stenosis of left anastomosis. The LUL bronchus and LB1-3 were normal to first subsegmental level.
The lingula is patent and LB4-5 are normal to first subsegmental level - able to traverse lingular bronchus with therapeutic bronchoscope.
LLL and LB6-10 have normal anatomic branching to segmental level.
No other evidence of mass, lesions, bleeding or other endobronchial pathology. 
Mucosa: Normal except as described above.
Secretions: Moderate, thin, and clear  All secretions were suctioned to clear (therapeutic aspiration).
Successful therapeutic aspiration was performed to clean out the trachea, right mainstem bronchus, right upper lobe, bronchus intermedius, right middle lobe, right lower lobe, left mainstem bronchus, left upper lobe, left lower lobe from mucus.
Bronchial alveolar lavage was performed at right middle lobe (Lateral Segment of RML (RB4) and Medial Segment of RML (RB5)).
Instilled 40 cc of NS, suction returned with 15 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
Residual secretions, saline suctioned to clear.  No bleeding.  Bronchoscope removed.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient's LMA was removed in the operating room and transported to the recovery room in stable condition.
ESTIMATED BLOOD LOSS:   None
COMPLICATIONS:    None
SPECIMEN(S): 
RML BAL - cell count, cultures/micro, cytology
IMPRESSION/PLAN: [REDACTED] is a 57 year old-year-old male who presents for bronchoscopy for airway evaluation and BAL.
-RML bronchus appears overall the best I have seen it at the beginning of the bronchoscopy;
no granulation tissue or signs of irritation.  Although unable to traverse with therapeutic bronchoscope, the RB4 and RB5 subsegments are grossly patent.
Therefore, to hopefully avoid undo inflammation, no dilation or other intervention was performed today.
-Repeat bronchoscopy in 2-3 weeks to confirm stability of RML bronchus.  Consideration of interventions depending on appearance of airway."""

# =============================================================================
# 2. DEFINITIONS: FIELDS, COLUMNS, FLAGS
# =============================================================================
PROCEDURE_FLAGS = [
    # Bronchoscopy (23)
    "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
    "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy",
    "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration",
    "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
    "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation",
    "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
    # Pleural (7)
    "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis",
    "pleural_biopsy", "fibrinolytic_therapy"
]

INDEX_COLS = ["source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes"] + PROCEDURE_FLAGS

SPAN_COLS = [
    "source_file", "note_id", "span_id", "section_type",
    "context_prefix", "span_text", "match_index",
    "start_char", "end_char", "span_len",
    "label", "normalized_value", "schema_field", "event_id",
    "is_negated", "is_historical", "time_anchor", "reviewer", "comments",
    "hydration_status"
]

EVENT_COLS = [
    "source_file", "note_id", "event_id", "event_type", "method",
    "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements",
    "specimens", "findings", "is_historical", "reviewer", "comments",
    "device_size", "device_material",
    "outcome_airway_lumen_pre", "outcome_airway_lumen_post",
    "outcome_symptoms", "outcome_pleural", "outcome_complication"
]

V3_COLS = [
    "note_id", "event_id", "type",
    "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station",
    "lesion.type", "lesion.size_mm",
    "method", "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote",
    "stent.size", "stent.material_or_brand", "catheter.size_fr",
    "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post",
    "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"
]

# =============================================================================
# 3. EXTRACTION LOGIC
# =============================================================================
def generate_data():
    
    # --- A. Procedure Flags ---
    # Scan text for positive assertions
    flags_dict = {f: 0 for f in PROCEDURE_FLAGS}
    
    # 1. Diagnostic Bronchoscopy (31624, "Flexible Therapeutic Bronchoscope")
    if "31624" in NOTE_TEXT or "Bronchoscope was advanced" in NOTE_TEXT:
        flags_dict["diagnostic_bronchoscopy"] = 1
        
    # 2. Therapeutic Aspiration (31645)
    if "31645" in NOTE_TEXT or "Therapeutic aspiration" in NOTE_TEXT:
        flags_dict["therapeutic_aspiration"] = 1
        
    # 3. BAL (31624, "lavage")
    if "BAL" in NOTE_TEXT or "Bronchial alveolar lavage" in NOTE_TEXT:
        flags_dict["bal"] = 1
        
    # Note: Text explicitly says "no dilation or other intervention was performed".
    
    # --- B. Span Extraction ---
    spans = []
    
    def add_span(text, label, norm_val=None, event="evt1", field=None, context=None):
        if not text: return
        # Simple exact search for anchor
        if text not in NOTE_TEXT:
            print(f"Warning: Span text '{text}' not found in note.")
            return
            
        spans.append({
            "span_text": text,
            "label": label,
            "normalized_value": norm_val,
            "event_id": event,
            "schema_field": field,
            "context_prefix": context
        })

    # EVENT 1: INSPECTION / DIAGNOSTIC
    evt1 = "evt1"
    add_span("Flexible Therapeutic Bronchoscope", "DEV_INSTRUMENT", "Flexible Bronchoscope", evt1, "device")
    add_span("airway examination", "PROC_METHOD", "Inspection", evt1, "method")
    add_span("mild stenosis of right anastomosis", "OBS_LESION", "Stenosis", evt1, "findings")
    add_span("right anastomosis", "ANAT_AIRWAY", "Right Mainstem Bronchus", evt1, "target_anatomy") # Approximated anatomy
    
    # RML findings
    add_span("RML bronchus", "ANAT_LUNG_LOC", "RML", evt1, "target_anatomy")
    add_span("moderate stenosis", "OBS_LESION", "Stenosis", evt1, "findings", context="RML bronchus with ")
    add_span("~5mm patent", "MEAS_AIRWAY_DIAM", "5mm", evt1, "measurement", context="estimate ")
    add_span("unable to traverse", "OBS_LESION", "Obstruction", evt1, "findings")
    add_span("anthracotic appearance", "OBS_LESION", "Anthracosis", evt1, "findings")
    add_span("no associated granulation tissue", "OBS_LESION", "Granulation Tissue", evt1, "findings", context="best I have seen his RML look before an intervention;\n")
    
    # Left side
    add_span("left anastomosis site", "ANAT_AIRWAY", "Left Mainstem Bronchus", evt1, "target_anatomy")
    add_span("mild stenosis of left anastomosis", "OBS_LESION", "Stenosis", evt1, "findings")
    
    # EVENT 2: THERAPEUTIC ASPIRATION
    evt2 = "evt2"
    add_span("Therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", evt2, "method", context="31645 ")
    add_span("Secretions: Moderate, thin, and clear", "OBS_LESION", "Secretions", evt2, "findings")
    add_span("suctioned to clear", "PROC_ACTION", "Suction", evt2, "action")
    # Anatomy list for aspiration
    for anat in ["trachea", "right mainstem bronchus", "right upper lobe", "bronchus intermedius", "right middle lobe", "right lower lobe", "left mainstem bronchus", "left upper lobe", "left lower lobe"]:
        add_span(anat, "ANAT_AIRWAY", anat.title(), evt2, "target_anatomy", context="clean out the ")

    # EVENT 3: BAL
    evt3 = "evt3"
    add_span("Bronchial alveolar lavage", "PROC_METHOD", "BAL", evt3, "method")
    add_span("Lateral Segment of RML (RB4)", "ANAT_LUNG_LOC", "RML Lateral Segment", evt3, "target_anatomy")
    add_span("Medial Segment of RML (RB5)", "ANAT_LUNG_LOC", "RML Medial Segment", evt3, "target_anatomy")
    add_span("Instilled 40 cc", "MEAS_VOL", "40cc", evt3, "measurement")
    add_span("returned with 15 cc", "MEAS_VOL", "15cc", evt3, "measurement")
    add_span("Cell Count", "PROC_ACTION", "Specimen", evt3, "specimen", context="Samples sent for ")
    add_span("Cytology", "PROC_ACTION", "Specimen", evt3, "specimen", context="(Cultures/Viral/Fungal), and ")

    # OUTCOMES / GLOBAL
    # No dilation performed
    add_span("no dilation", "PROC_METHOD", "None", "evt_neg", "is_negated", context="inflammation, ")
    
    # Outcomes
    add_span("No immediate complications", "OUTCOME_COMPLICATION", "None", "evt_outcome")
    add_span("patient tolerated the procedure well", "OUTCOME_SYMPTOMS", "Tolerated Well", "evt_outcome")

    # --- C. Hydration & Processing ---
    # This happens in the writer function where we calculate offsets
    
    return flags_dict, spans

# =============================================================================
# 4. EXCEL GENERATION UTILS
# =============================================================================
def get_or_create_sheet(wb, name, headers):
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(name)
        ws.append(headers)
    return ws

def hydrate_offsets(spans, full_text):
    hydrated = []
    
    for s in spans:
        txt = s["span_text"]
        if not txt: continue
        
        row = s.copy()
        
        # 1. Exact count
        count = full_text.count(txt)
        start = -1
        
        if count == 1:
            start = full_text.find(txt)
            row["hydration_status"] = "hydrated_unique"
            
        elif count > 1:
            # Context match
            if s.get("context_prefix"):
                ctx = s["context_prefix"]
                # Find all occurrences
                occurrences = [m.start() for m in re.finditer(re.escape(txt), full_text)]
                match_found = False
                for occ in occurrences:
                    # Check preceding 120 chars
                    window = full_text[max(0, occ-120):occ]
                    if ctx in window:
                        start = occ
                        row["hydration_status"] = "hydrated_prefix_window"
                        match_found = True
                        break
                if not match_found:
                    # Fallback to first if ambiguous context
                    start = full_text.find(txt)
                    row["hydration_status"] = f"ambiguous_count={count}_context_fail"
            else:
                # Default to first
                start = full_text.find(txt)
                row["hydration_status"] = f"ambiguous_count={count}"
        else:
            row["hydration_status"] = "missing"
            
        if start != -1:
            row["start_char"] = start
            row["end_char"] = start + len(txt)
            row["span_len"] = len(txt)
        
        hydrated.append(row)
        
    return hydrated

def write_workbook():
    if os.path.exists(TEMPLATE_PATH):
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    else:
        wb = openpyxl.Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    flags, raw_spans = generate_data()
    
    # 1. Note_Text
    ws_text = get_or_create_sheet(wb, "Note_Text", ["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # 2. Note_Index
    ws_idx = get_or_create_sheet(wb, "Note_Index", INDEX_COLS)
    idx_row = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Success", ""]
    # Append flags
    for f in PROCEDURE_FLAGS:
        idx_row.append(flags.get(f, 0))
    ws_idx.append(idx_row)
    
    # 3. Span_Annotations (Anchor First)
    ws_span = get_or_create_sheet(wb, "Span_Annotations", SPAN_COLS)
    
    span_id_counter = 1
    for s in raw_spans:
        row = [
            SOURCE_FILE, NOTE_ID, f"span_{span_id_counter:03d}", "PROCEDURE",
            s.get("context_prefix", ""), s["span_text"], s.get("match_index", ""),
            "", "", f'=LEN(F{ws_span.max_row+1})', # Formulas for len
            s["label"], s.get("normalized_value", ""), s.get("schema_field", ""), s.get("event_id", ""),
            s.get("is_negated", ""), "", "", "", "", "needs_hydration"
        ]
        ws_span.append(row)
        # Store ID for hydration
        s["span_id"] = f"span_{span_id_counter:03d}"
        span_id_counter += 1

    # 4. Span_Hydrated
    hydrated_data = hydrate_offsets(raw_spans, NOTE_TEXT)
    ws_hyd = get_or_create_sheet(wb, "Span_Hydrated", SPAN_COLS)
    for h in hydrated_data:
        row = [
            SOURCE_FILE, NOTE_ID, h["span_id"], "PROCEDURE",
            h.get("context_prefix", ""), h["span_text"], "",
            h.get("start_char", ""), h.get("end_char", ""), h.get("span_len", ""),
            h["label"], h.get("normalized_value", ""), h.get("schema_field", ""), h.get("event_id", ""),
            h.get("is_negated", ""), "", "", "", "", h.get("hydration_status", "")
        ]
        ws_hyd.append(row)

    # 5. Event_Log
    ws_event = get_or_create_sheet(wb, "Event_Log", EVENT_COLS)
    
    # Event 1: Inspection
    ws_event.append([
        SOURCE_FILE, NOTE_ID, "evt1", "Diagnostic Bronchoscopy", "Inspection",
        "RML Bronchus, Anastomoses", "Flexible Bronchoscope", "", "", "", "Stenosis (~5mm RML)",
        "", "Mild/Mod Stenosis, No Granulation", "", "", "",
        "", "", "5mm patent (RML)", "", "", "", ""
    ])
    
    # Event 2: Therapeutic Aspiration
    ws_event.append([
        SOURCE_FILE, NOTE_ID, "evt2", "Therapeutic Aspiration", "Aspiration",
        "Trachea, Bilateral Tree", "", "", "", "", "",
        "Secretions", "Moderate, thin secretions", "", "", "",
        "", "", "", "", "Tolerated well", "", "None"
    ])
    
    # Event 3: BAL
    ws_event.append([
        SOURCE_FILE, NOTE_ID, "evt3", "BAL", "Lavage",
        "RML (RB4, RB5)", "", "", "", "", "Instilled 40cc, Return 15cc",
        "Cell Count, Cytology, Micro", "", "", "", "",
        "", "", "", "", "", "", ""
    ])

    # 6. V3_Procedure_Events
    ws_v3 = get_or_create_sheet(wb, "V3_Procedure_Events", V3_COLS)
    
    # Evt1
    ws_v3.append([
        NOTE_ID, "evt1", "Diagnostic Bronchoscopy",
        "Airway", "RML", "Bronchus", "",
        "Stenosis", "",
        "Inspection", "", '{"stenosis_grade": "moderate"}', "", '{"finding": "anthracosis"}', "moderate stenosis - estimate ~5mm patent",
        "", "", "",
        "5mm", "",
        "", "", ""
    ])
    
    # Evt2
    ws_v3.append([
        NOTE_ID, "evt2", "Therapeutic Aspiration",
        "Airway", "Bilateral", "Tree", "",
        "Secretions", "",
        "Suction", "", "", "", '{"secretions": "moderate, thin"}', "All secretions were suctioned to clear",
        "", "", "",
        "", "",
        "Tolerated well", "", "None"
    ])

    # Evt3
    ws_v3.append([
        NOTE_ID, "evt3", "BAL",
        "Airway", "RML", "RB4/RB5", "",
        "", "",
        "Lavage", "", '{"instilled": "40cc", "return": "15cc"}', '["Cell Count", "Micro", "Cyto"]', "", "Instilled 40 cc of NS",
        "", "", "",
        "", "",
        "", "", ""
    ])

    # 7. V3_Registry_JSON
    ws_json = get_or_create_sheet(wb, "V3_Registry_JSON", ["schema_version", "note_id", "json_output"])
    
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": [
            {
                "event_id": "evt1",
                "type": "Diagnostic Bronchoscopy",
                "target": {"anatomy": "RML Bronchus"},
                "findings": ["Stenosis", "Anthracosis"],
                "measurements": {"lumen_pre": "5mm"}
            },
            {
                "event_id": "evt2",
                "type": "Therapeutic Aspiration",
                "target": {"anatomy": "Bilateral Airways"},
                "outcome": {"symptoms": "Tolerated well", "complications": "None"}
            },
            {
                "event_id": "evt3",
                "type": "BAL",
                "target": {"anatomy": "RML"},
                "specimens": ["Cell Count", "Microbiology", "Cytology"]
            }
        ]
    }
    
    ws_json.append(["3.0", NOTE_ID, json.dumps(registry_data, indent=2)])

    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    write_workbook()