import re
import json
import datetime
import os
from openpyxl import Workbook, load_workbook

# =============================================================================
# 1. INPUT DATA
# =============================================================================

NOTE_ID = "note_087"
SOURCE_FILE = "note_087.txt"
PROCEDURE_DATE = "2026-01-12" # inferred from context or left generic
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_087 SOURCE_FILE: note_087.txt INDICATION FOR OPERATION:  [REDACTED]is a 57 year old-year-old male who presents with bronchial stenosis.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31622 Dx bronchoscope/cell washing          
31624 Dx bronchoscope/lavage (BAL)    
31630 Balloon dilation
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required a Pt with transplant 
stenosis that required 4 balloon dilations and 6 applications of cryo .
This resulted in >50% increased work due to Time, Technical difficulty of procedure, and Physical and mental effort required.
Apply to: 31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy).
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
ESTIMATED BLOOD LOSS:   Minimum
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: supine
Initial Airway Inspection Findings:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
Initial Airway Inspection Findings:
The laryngeal mask airway is in good position. The vocal cords appear normal.
The subglottic space is normal. The trachea is of normal caliber. The carina is sharp.
The tracheobronchial tree was examined to at least the first subsegmental level.
Airway exam notable for stenosis of the RML (75% patent).
Diameter was 5-6 mm
Moderate clear secretions bilaterally, white thicker mucus eminating from RML.
Successful therapeutic aspiration was performed to clean out the Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, and RML Carina (RC2) from mucus and mucus plug.
Endobronchial obstruction at RML was treated with the following modalities:
Modality	Tools	Setting/Mode	Duration	Results
Electrocautery	 	 	 	 
APC	 	 	 	 
Laser	 	 	 	 
CoreCath	 	 	 	 
Cryoprobe	1.7mm probe	 	30sec freeze-thaw cycles;
total 6 applications	Ablation
 
Balloon dilation was performed at RML.  6/7/8 Elation balloon was used to perform dilation to 8 mm at the RML.
Total 2 inflations with dilation time of 60 seconds each.
Prior to treatment, affected airway was note to be 75% patent.  After treatment, the airway was 100% patent.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was successfully extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
BAL - RML
IMPRESSION/PLAN: [REDACTED]is a 57 year old-year-old male who presents for bronchoscopy for bronchial stenosis.
[ ] f/u results
[ ] f/u CXR
[ ] repeat bronchoscopy on [REDACTED]"""

# =============================================================================
# 2. DEFINITIONS & UTILS
# =============================================================================

PROCEDURE_FLAGS = [
    "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings",
    "endobronchial_biopsy", "tbna_conventional", "linear_ebus", "radial_ebus",
    "navigational_bronchoscopy", "transbronchial_biopsy", "transbronchial_cryobiopsy",
    "therapeutic_aspiration", "foreign_body_removal", "airway_dilation", "airway_stent",
    "thermal_ablation", "tumor_debulking_non_thermal", "cryotherapy", "blvr",
    "peripheral_ablation", "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
    "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis",
    "pleural_biopsy", "fibrinolytic_therapy"
]

LABELS = [
    "ANAT_LN_STATION", "ANAT_LUNG_LOC", "ANAT_AIRWAY", "ANAT_PLEURA", "LATERALITY",
    "DEV_VALVE", "DEV_STENT", "DEV_CATHETER", "DEV_NEEDLE", "DEV_INSTRUMENT",
    "DEV_STENT_MATERIAL", "DEV_STENT_SIZE", "DEV_CATHETER_SIZE",
    "MEAS_SIZE", "MEAS_VOL", "MEAS_PRESS", "MEAS_COUNT", "MEAS_AIRWAY_DIAM", "MEAS_PLEURAL_DRAIN",
    "PROC_METHOD", "PROC_ACTION", "OBS_ROSE", "OBS_LESION",
    "CTX_HISTORICAL", "CTX_TIME",
    "OUTCOME_AIRWAY_LUMEN_PRE", "OUTCOME_AIRWAY_LUMEN_POST",
    "OUTCOME_SYMPTOMS", "OUTCOME_PLEURAL", "OUTCOME_COMPLICATION"
]

def create_blank_workbook():
    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    return wb

def get_or_create_sheet(wb, title, headers):
    if title in wb.sheetnames:
        ws = wb[title]
    else:
        ws = wb.create_sheet(title)
        ws.append(headers)
    return ws

def hydrate_span(full_text, span_text, context_prefix=None, match_index=None):
    """
    Returns (start_char, end_char, hydration_status)
    """
    if not span_text:
        return None, None, "missing_text"
    
    # 1. Exact count check
    count = full_text.count(span_text)
    
    if count == 0:
        return None, None, "not_found"
    
    if count == 1:
        start = full_text.find(span_text)
        return start, start + len(span_text), "hydrated_unique"
    
    # 2. Context prefix
    if context_prefix:
        # Find all occurrences
        starts = [m.start() for m in re.finditer(re.escape(span_text), full_text)]
        best_start = -1
        
        for s in starts:
            # Look at preceding 120 chars
            window_start = max(0, s - 120)
            window_end = s
            preceding_text = full_text[window_start:window_end]
            if context_prefix in preceding_text:
                best_start = s
                break
        
        if best_start != -1:
            return best_start, best_start + len(span_text), "hydrated_prefix_window"

    # 3. Match Index
    if match_index is not None and isinstance(match_index, int):
        starts = [m.start() for m in re.finditer(re.escape(span_text), full_text)]
        if 0 <= match_index < len(starts):
            s = starts[match_index]
            return s, s + len(span_text), "hydrated_match_index"

    return None, None, f"ambiguous_count={count}"

# =============================================================================
# 3. EXTRACTION LOGIC
# =============================================================================

def extract_data(note_text):
    # Flags logic
    flags = {f: 0 for f in PROCEDURE_FLAGS}
    
    # Simple keyword mapping for flags based on this specific note content
    # Note: Logic is specialized for this note's explicit codes and text
    lower_text = note_text.lower()
    
    if "31645" in note_text or "therapeutic aspiration" in lower_text:
        flags["therapeutic_aspiration"] = 1
    if "31622" in note_text or "bronchial wash" in lower_text or "cell washing" in lower_text:
        flags["bronchial_wash"] = 1
    if "31624" in note_text or "bal" in lower_text:
        flags["bal"] = 1
    if "31630" in note_text or "balloon dilation" in lower_text:
        flags["airway_dilation"] = 1
    if "31641" in note_text or "cryotherapy" in lower_text or "cryo" in lower_text:
        flags["cryotherapy"] = 1
    
    # Note mentions "Dx bronchoscope" so generic flag
    if "dx bronchoscope" in lower_text:
        flags["diagnostic_bronchoscopy"] = 1

    # SPAN GENERATION
    spans = []
    
    # Helper to add span
    def add_span(text, label, norm_val, event_id=None, prefix=None, field=None):
        spans.append({
            "span_text": text,
            "label": label,
            "normalized_value": norm_val,
            "event_id": event_id,
            "context_prefix": prefix,
            "schema_field": field
        })

    # --- Event 1: Therapeutic Aspiration ---
    evt_asp = "evt_01"
    add_span("Therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", evt_asp, prefix="31645 ")
    add_span("clean out the Trachea", "PROC_ACTION", "Clean", evt_asp)
    add_span("Trachea (Distal 1/3)", "ANAT_AIRWAY", "Trachea", evt_asp)
    add_span("Right Mainstem", "ANAT_AIRWAY", "RMS", evt_asp, prefix="Distal 1/3), ")
    add_span("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", evt_asp)
    add_span("Left Mainstem", "ANAT_AIRWAY", "LMS", evt_asp)
    add_span("RML Carina (RC2)", "ANAT_AIRWAY", "RML", evt_asp)
    add_span("mucus and mucus plug", "OBS_LESION", "Mucus Plug", evt_asp)

    # --- Event 2: Cryotherapy ---
    evt_cryo = "evt_02"
    add_span("Destruction of tumor OR relief of stenosis", "PROC_ACTION", "Relief of stenosis", evt_cryo)
    add_span("cryotherapy", "PROC_METHOD", "Cryotherapy", evt_cryo, prefix="laser therapy, ")
    add_span("Cryoprobe", "DEV_INSTRUMENT", "Cryoprobe", evt_cryo)
    add_span("1.7mm probe", "DEV_INSTRUMENT", "Cryoprobe 1.7mm", evt_cryo)
    add_span("30sec freeze-thaw cycles", "PROC_METHOD", "Freeze-Thaw", evt_cryo)
    add_span("total 6 applications", "MEAS_COUNT", "6", evt_cryo)
    add_span("RML", "ANAT_AIRWAY", "RML", evt_cryo, prefix="Endobronchial obstruction at ")
    add_span("Ablation", "PROC_ACTION", "Ablation", evt_cryo)

    # --- Event 3: Balloon Dilation ---
    evt_bal_dil = "evt_03"
    add_span("Balloon dilation", "PROC_METHOD", "Balloon Dilation", evt_bal_dil, prefix="31630 ")
    add_span("RML", "ANAT_AIRWAY", "RML", evt_bal_dil, prefix="Balloon dilation was performed at ")
    add_span("6/7/8 Elation balloon", "DEV_INSTRUMENT", "Elation Balloon 6-7-8mm", evt_bal_dil)
    add_span("dilation to 8 mm", "MEAS_AIRWAY_DIAM", "8 mm", evt_bal_dil)
    add_span("Total 2 inflations", "MEAS_COUNT", "2", evt_bal_dil)
    add_span("dilation time of 60 seconds each", "CTX_TIME", "60 seconds", evt_bal_dil)

    # Outcomes associated with Dilation/Cryo (the therapeutic sequence)
    # The note says "Prior to treatment... 75% patent. After treatment... 100% patent."
    # Linking this to the primary therapeutic event (could be either, linking to dilation as it's the mechanical opening)
    add_span("75% patent", "OUTCOME_AIRWAY_LUMEN_PRE", "75%", evt_bal_dil, prefix="note to be ")
    add_span("100% patent", "OUTCOME_AIRWAY_LUMEN_POST", "100%", evt_bal_dil)

    # --- Event 4: BAL ---
    evt_bal = "evt_04"
    add_span("BAL", "PROC_METHOD", "BAL", evt_bal, prefix="31624 Dx bronchoscope/lavage (")
    add_span("BAL", "PROC_METHOD", "BAL", evt_bal, prefix="SPECIMEN(S):\n")
    add_span("RML", "ANAT_AIRWAY", "RML", evt_bal, prefix="BAL - ")

    # --- General Findings / Complications ---
    add_span("No immediate complications", "OUTCOME_COMPLICATION", "None", "evt_global")
    add_span("stenosis of the RML", "OBS_LESION", "Stenosis", "evt_global")
    add_span("Diameter was 5-6 mm", "MEAS_SIZE", "5-6 mm", "evt_global") # Anatomical finding

    return flags, spans

# =============================================================================
# 4. EXCEL GENERATION
# =============================================================================

def generate_excel():
    if os.path.exists(TEMPLATE_PATH):
        wb = load_workbook(TEMPLATE_PATH)
    else:
        wb = create_blank_workbook()

    flags, spans = extract_data(NOTE_TEXT)

    # 1. Note_Text
    ws_text = get_or_create_sheet(wb, "Note_Text", ["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    headers_index = ["source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes"] + PROCEDURE_FLAGS
    ws_index = get_or_create_sheet(wb, "Note_Index", headers_index)
    row_index = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "pending", ""]
    for f in PROCEDURE_FLAGS:
        row_index.append(flags.get(f, 0))
    ws_index.append(row_index)

    # 3. Span_Annotations
    headers_span = [
        "source_file", "note_id", "span_id", "section_type",
        "context_prefix", "span_text", "match_index",
        "start_char", "end_char", "span_len",
        "label", "normalized_value", "schema_field", "event_id",
        "is_negated", "is_historical", "time_anchor", "reviewer", "comments",
        "hydration_status"
    ]
    ws_anno = get_or_create_sheet(wb, "Span_Annotations", headers_span)
    
    # 4. Span_Hydrated
    ws_hydrated = get_or_create_sheet(wb, "Span_Hydrated", headers_span)

    # Process spans
    hydrated_rows = []
    
    for idx, s in enumerate(spans):
        span_id = f"span_{idx+1:03d}"
        
        # Calculate hydration
        start, end, status = hydrate_span(NOTE_TEXT, s["span_text"], s["context_prefix"])
        span_len = len(s["span_text"]) if s["span_text"] else 0
        
        # Base row for Anno (no offsets)
        row_anno = [
            SOURCE_FILE, NOTE_ID, span_id, "", 
            s["context_prefix"], s["span_text"], "",
            "", "", f"=LEN(F{ws_anno.max_row+1})", # Formula for len
            s["label"], s["normalized_value"], s["schema_field"], s["event_id"],
            "FALSE", "FALSE", "", "", "", "needs_hydration"
        ]
        ws_anno.append(row_anno)
        
        # Hydrated row
        row_hyd = [
            SOURCE_FILE, NOTE_ID, span_id, "",
            s["context_prefix"], s["span_text"], "",
            start if start is not None else "", end if end is not None else "", span_len,
            s["label"], s["normalized_value"], s["schema_field"], s["event_id"],
            "FALSE", "FALSE", "", "", "", status
        ]
        ws_hydrated.append(row_hyd)
        
        # Store for event log usage
        s["start"] = start

    # 5. Event_Log & 6. V3_Procedure_Events & 7. JSON
    # Group by event_id
    events = {}
    for s in spans:
        eid = s["event_id"]
        if not eid: continue
        if eid not in events:
            events[eid] = {
                "method": [], "anatomy": [], "device": [], "measurements": [], 
                "outcomes": {}, "specimens": [], "findings": []
            }
        
        if s["label"] == "PROC_METHOD": events[eid]["method"].append(s["normalized_value"])
        if s["label"].startswith("ANAT"): events[eid]["anatomy"].append(s["normalized_value"])
        if s["label"].startswith("DEV"): events[eid]["device"].append(s["normalized_value"])
        if s["label"].startswith("MEAS"): events[eid]["measurements"].append(f"{s['normalized_value']} ({s['span_text']})")
        
        # Outcomes
        if s["label"] == "OUTCOME_AIRWAY_LUMEN_PRE": events[eid]["outcomes"]["lumen_pre"] = s["normalized_value"]
        if s["label"] == "OUTCOME_AIRWAY_LUMEN_POST": events[eid]["outcomes"]["lumen_post"] = s["normalized_value"]
        if s["label"] == "OUTCOME_COMPLICATION": events[eid]["outcomes"]["complication"] = s["normalized_value"]

    # Write Event Log
    headers_log = [
        "source_file", "note_id", "event_id", "event_type", "method",
        "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements",
        "specimens", "findings", "is_historical", "reviewer", "comments",
        "device_size", "device_material",
        "outcome_airway_lumen_pre", "outcome_airway_lumen_post",
        "outcome_symptoms", "outcome_pleural", "outcome_complication"
    ]
    ws_log = get_or_create_sheet(wb, "Event_Log", headers_log)
    
    # Write V3 Events
    headers_v3 = [
        "note_id", "event_id", "type",
        "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station",
        "lesion.type", "lesion.size_mm",
        "method", "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote",
        "stent.size", "stent.material_or_brand", "catheter.size_fr",
        "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post",
        "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"
    ]
    ws_v3 = get_or_create_sheet(wb, "V3_Procedure_Events", headers_v3)

    json_events = []

    for eid, data in events.items():
        # Determine main method
        method_str = ", ".join(set(data["method"]))
        anatomy_str = ", ".join(set(data["anatomy"]))
        device_str = ", ".join(set(data["device"]))
        meas_str = ", ".join(data["measurements"])
        
        # Log Row
        row_log = [
            SOURCE_FILE, NOTE_ID, eid, "Therapeutic/Diagnostic", method_str,
            anatomy_str, device_str, "", "", "", meas_str,
            "", "", "FALSE", "", "",
            "", "", # size/material
            data["outcomes"].get("lumen_pre", ""),
            data["outcomes"].get("lumen_post", ""),
            "", "",
            data["outcomes"].get("complication", "")
        ]
        ws_log.append(row_log)
        
        # V3 Row
        row_v3 = [
            NOTE_ID, eid, "Procedure",
            "Airway", anatomy_str, "", "",
            "", "",
            method_str, json.dumps(data["device"]), json.dumps(data["measurements"]), "[]", "[]", "",
            "", "", "",
            data["outcomes"].get("lumen_pre", ""),
            data["outcomes"].get("lumen_post", ""),
            "", "",
            data["outcomes"].get("complication", "")
        ]
        ws_v3.append(row_v3)
        
        # JSON obj
        json_events.append({
            "event_id": eid,
            "method": method_str,
            "target": anatomy_str,
            "outcomes": data["outcomes"]
        })

    # 7. V3_Registry_JSON
    ws_json = get_or_create_sheet(wb, "V3_Registry_JSON", ["json_output"])
    final_json = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": json_events,
        "no_immediate_complications": True # derived from text
    }
    ws_json.append([json.dumps(final_json, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_excel()