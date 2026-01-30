import openpyxl
import re
import json
from openpyxl.utils import get_column_letter

# ========================================================================================
# 1. INPUT DATA (User-Provided Note)
# ========================================================================================

NOTE_ID = "note_088"
SOURCE_FILE = "note_088.txt"
PROCEDURE_DATE = ""  # Not available in text
NOTE_TEXT = """NOTE_ID:  note_088 SOURCE_FILE: note_088.txt INDICATION FOR OPERATION:  [REDACTED]is a 57 year old-year-old male who presents with bronchial stenosis.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31630 Balloon dilation
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
Flexible Hybrid (Pedatric) Bronchoscope
ESTIMATED BLOOD LOSS:   Minimum
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
Initial Airway Inspection Findings:
The laryngeal mask airway is in good position. The vocal cords appear normal.
The subglottic space is normal. The trachea is of normal caliber. The carina is sharp.
The tracheobronchial tree was examined to at least the first subsegmental level.
Airway exam notable for stenosis of the RML (25% patent). 
Moderate clear secretions bilaterally, white thicker mucus eminating from RML.
Successful therapeutic aspiration was performed to clean out the Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, and RML Carina (RC2) from mucus and mucus plug.
Endobronchial obstruction at RML was treated with the following modalities:
Modality	Tools	Setting/Mode	Duration	Results
Electrocautery	Needle Knife	EndoCut I, Effect 3, 3d/2i	 	Radial cuts
APC	 	 	 	 
Laser	 	 	 	 
CoreCath	 	 	 	 
Cryoprobe	1.7mm probe	 	30sec freeze-thaw cycles	Ablation
Balloon dilation was performed at RML.
6/7/8 Elation balloon was used to perform dilation to 6 mm at the RML.
Total 2 inflations with dilation time of 60 seconds each.
Balloon dilation was performed at RML.
6/7/8 Elation balloon was used to perform dilation to 8 mm at the RML.
Total 2 inflations with dilation time of 60 seconds each.
Prior to treatment, affected airway was note to be 25% patent.  After treatment, the airway was 100% patent.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
None
IMPRESSION/PLAN: [REDACTED]is a 57 year old-year-old male who presents for bronchoscopy for bronchial stenosis.
RML airway moderately stenosed again, repeat dilation and touch cryotherapy. 
- bronch with dilation in 2 weeks"""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"


# ========================================================================================
# 2. DEFINITIONS & CONFIG
# ========================================================================================

# Procedure flags (exact column names required by V3)
PROC_FLAGS = [
    # Bronchoscopy (23)
    'diagnostic_bronchoscopy', 'bal', 'bronchial_wash', 'brushings', 'endobronchial_biopsy',
    'tbna_conventional', 'linear_ebus', 'radial_ebus', 'navigational_bronchoscopy',
    'transbronchial_biopsy', 'transbronchial_cryobiopsy', 'therapeutic_aspiration',
    'foreign_body_removal', 'airway_dilation', 'airway_stent', 'thermal_ablation',
    'tumor_debulking_non_thermal', 'cryotherapy', 'blvr', 'peripheral_ablation',
    'bronchial_thermoplasty', 'whole_lung_lavage', 'rigid_bronchoscopy',
    # Pleural (7)
    'thoracentesis', 'chest_tube', 'ipc', 'medical_thoracoscopy', 'pleurodesis',
    'pleural_biopsy', 'fibrinolytic_therapy'
]

# Initialize flags map
flags_map = {f: 0 for f in PROC_FLAGS}

# ----------------------------------------------------------------------------------------
# 3. EXTRACTION LOGIC (Deterministic)
# ----------------------------------------------------------------------------------------

# Clean text for easier searching (keeping original for offsets)
clean_text = NOTE_TEXT.lower().replace('\n', ' ')

# --- Flag Detection ---
if "therapeutic aspiration" in clean_text:
    flags_map['therapeutic_aspiration'] = 1
if "balloon dilation" in clean_text or "dilation" in clean_text:
    flags_map['airway_dilation'] = 1
if "cryotherapy" in clean_text or "cryoprobe" in clean_text:
    flags_map['cryotherapy'] = 1
if "electrocautery" in clean_text or "needle knife" in clean_text:
    flags_map['thermal_ablation'] = 1 # Electrocautery falls under thermal

# Basic diagnostic?
if "bronchoscopy" in clean_text:
    flags_map['diagnostic_bronchoscopy'] = 1

# --- Span Extraction List ---
# Structure: (span_text, label, normalized_value, schema_field, event_id, context_prefix)

spans = []

# --- Event 1: Therapeutic Aspiration ---
# "Successful therapeutic aspiration was performed to clean out the Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, and RML Carina (RC2) from mucus and mucus plug."
evt1 = "evt_01"
# Spans
spans.append(("therapeutic aspiration", "PROC_METHOD", "therapeutic aspiration", "method", evt1, "Successful "))
spans.append(("clean out", "PROC_ACTION", "clean out", "action", evt1, "to "))
spans.append(("Trachea (Distal 1/3)", "ANAT_AIRWAY", "trachea distal", "target.anatomy", evt1, "out the "))
spans.append(("Right Mainstem", "ANAT_AIRWAY", "right mainstem bronchus", "target.anatomy", evt1, "Trachea (Distal 1/3), "))
spans.append(("Bronchus Intermedius", "ANAT_AIRWAY", "bronchus intermedius", "target.anatomy", evt1, "Right Mainstem, "))
spans.append(("Left Mainstem", "ANAT_AIRWAY", "left mainstem bronchus", "target.anatomy", evt1, "Bronchus Intermedius , "))
spans.append(("RML Carina (RC2)", "ANAT_AIRWAY", "rml carina", "target.anatomy", evt1, "Left Mainstem, and "))
spans.append(("mucus and mucus plug", "OBS_LESION", "mucus plug", "findings", evt1, "from "))

# --- Event 2: Electrocautery / Needle Knife (Thermal Ablation) ---
# "Endobronchial obstruction at RML was treated with the following modalities:"
# "Electrocautery	Needle Knife	EndoCut I, Effect 3, 3d/2i		Radial cuts"
evt2 = "evt_02"
spans.append(("Endobronchial obstruction", "OBS_LESION", "obstruction", "lesion.type", evt2, None))
spans.append(("RML", "ANAT_AIRWAY", "rml", "target.anatomy", evt2, "obstruction at "))
spans.append(("Electrocautery", "PROC_METHOD", "electrocautery", "method", evt2, "Results\n"))
spans.append(("Needle Knife", "DEV_INSTRUMENT", "needle knife", "devices", evt2, "Electrocautery\t"))
spans.append(("Radial cuts", "PROC_ACTION", "radial cuts", "action", evt2, "Effect 3, 3d/2i\t \t"))

# --- Event 3: Cryotherapy ---
# "Cryoprobe	1.7mm probe		30sec freeze-thaw cycles	Ablation"
evt3 = "evt_03"
spans.append(("Cryoprobe", "PROC_METHOD", "cryotherapy", "method", evt3, "CoreCath\t \t \t \t \n"))
spans.append(("1.7mm probe", "DEV_INSTRUMENT", "cryoprobe 1.7mm", "devices", evt3, "Cryoprobe\t"))
spans.append(("Ablation", "PROC_ACTION", "ablation", "action", evt3, "freeze-thaw cycles\t"))

# --- Event 4: Balloon Dilation 1 ---
# "Balloon dilation was performed at RML."
# "6/7/8 Elation balloon was used to perform dilation to 6 mm at the RML."
evt4 = "evt_04"
spans.append(("Balloon dilation", "PROC_METHOD", "balloon dilation", "method", evt4, "cycles	Ablation\n"))
spans.append(("RML", "ANAT_AIRWAY", "rml", "target.anatomy", evt4, "performed at "))
spans.append(("6/7/8 Elation balloon", "DEV_INSTRUMENT", "elation balloon", "devices", evt4, "at RML.\n"))
spans.append(("dilation", "PROC_ACTION", "dilation", "action", evt4, "used to perform "))
spans.append(("6 mm", "MEAS_SIZE", "6 mm", "measurements", evt4, "dilation to "))
# Context for "RML" in this sentence: "at the RML."
spans.append(("RML", "ANAT_AIRWAY", "rml", "target.anatomy", evt4, "at the "))

# --- Event 5: Balloon Dilation 2 ---
# "Balloon dilation was performed at RML." (Second occurrence)
# "6/7/8 Elation balloon was used to perform dilation to 8 mm at the RML."
evt5 = "evt_05"
# Context for second "Balloon dilation": after "60 seconds each.\n"
spans.append(("Balloon dilation", "PROC_METHOD", "balloon dilation", "method", evt5, "60 seconds each.\n"))
spans.append(("RML", "ANAT_AIRWAY", "rml", "target.anatomy", evt5, "performed at ")) # This is the second "performed at RML"
spans.append(("6/7/8 Elation balloon", "DEV_INSTRUMENT", "elation balloon", "devices", evt5, "at RML.\n")) # Second occurrence
spans.append(("dilation", "PROC_ACTION", "dilation", "action", evt5, "used to perform ")) # Second occurrence
spans.append(("8 mm", "MEAS_SIZE", "8 mm", "measurements", evt5, "dilation to "))
# Context for "RML" in this sentence: "at the RML." (Second occurrence)
spans.append(("RML", "ANAT_AIRWAY", "rml", "target.anatomy", evt5, "at the "))

# --- Outcomes ---
# "Prior to treatment, affected airway was note to be 25% patent."
# "After treatment, the airway was 100% patent."
# "There were no immediate complications."
evt_out = "evt_outcome"
spans.append(("25% patent", "OUTCOME_AIRWAY_LUMEN_PRE", "25%", "outcomes.airway.lumen_pre", evt_out, "to be "))
spans.append(("100% patent", "OUTCOME_AIRWAY_LUMEN_POST", "100%", "outcomes.airway.lumen_post", evt_out, "airway was "))
spans.append(("no immediate complications", "OUTCOME_COMPLICATION", "none", "outcomes.complications", evt_out, "There were "))

# --- Diagnosis / Indication ---
# "stenosis of the RML (25% patent)"
spans.append(("stenosis", "OBS_LESION", "stenosis", "lesion.type", "evt_dx", "notable for "))
spans.append(("RML", "ANAT_AIRWAY", "rml", "target.anatomy", "evt_dx", "of the "))
spans.append(("25% patent", "MEAS_AIRWAY_DIAM", "25%", "measurements", "evt_dx", "RML (")) # Using measurement here, outcome above

# ========================================================================================
# 4. EXCEL GENERATION LOGIC
# ========================================================================================

def generate_workbook():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        # Create blank if missing (fallback for valid script execution)
        wb = openpyxl.Workbook()
        for name in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            wb.create_sheet(name)

    # ------------------------------------------------------------------------------------
    # Sheet 1: Note_Text
    # ------------------------------------------------------------------------------------
    ws = wb["Note_Text"]
    if ws.max_row == 1:
        ws.append(["note_id", "source_file", "note_text"])
    ws.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # ------------------------------------------------------------------------------------
    # Sheet 2: Note_Index (Flags)
    # ------------------------------------------------------------------------------------
    ws = wb["Note_Index"]
    # Ensure headers
    headers = [
        "source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes"
    ] + PROC_FLAGS
    
    # Check if headers exist, if not write them
    if ws.max_row == 1:
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)
            
    # Write Data
    row_data = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "To Do", ""]
    for flag in PROC_FLAGS:
        row_data.append(flags_map[flag])
    ws.append(row_data)

    # ------------------------------------------------------------------------------------
    # Sheet 3: Span_Annotations (Anchor First)
    # ------------------------------------------------------------------------------------
    ws_span = wb["Span_Annotations"]
    span_headers = [
        "source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text",
        "match_index", "start_char", "end_char", "span_len", "label", "normalized_value",
        "schema_field", "event_id", "is_negated", "is_historical", "time_anchor",
        "reviewer", "comments", "hydration_status"
    ]
    if ws_span.max_row == 1:
        ws_span.append(span_headers)

    # ------------------------------------------------------------------------------------
    # Sheet 4: Span_Hydrated (Computed Offsets)
    # ------------------------------------------------------------------------------------
    ws_hyd = wb["Span_Hydrated"]
    if ws_hyd.max_row == 1:
        ws_hyd.append(span_headers)

    # Helper for hydration
    def find_offset(text, span, context):
        # 1. Exact match count
        count = text.count(span)
        if count == 0:
            return None, None, "missing"
        
        # 2. Context match
        if context:
            # Look for context + span
            # We scan for all occurrences of span
            starts = [m.start() for m in re.finditer(re.escape(span), text)]
            best_start = -1
            
            # Simple context check: Look in preceding 150 chars
            for s in starts:
                window_start = max(0, s - 150)
                preceding = text[window_start:s]
                if context in preceding:
                    best_start = s
                    break
            
            if best_start != -1:
                return best_start, best_start + len(span), "hydrated_prefix_window"
            
            # Fallback if context not found exactly but unique span
            if count == 1:
                idx = text.find(span)
                return idx, idx + len(span), "hydrated_unique_fallback"

        # 3. Unique match
        if count == 1:
            idx = text.find(span)
            return idx, idx + len(span), "hydrated_unique"
        
        return None, None, f"ambiguous_count={count}"

    # Write Spans
    for i, (txt, lbl, norm, field, evt, ctx) in enumerate(spans):
        span_id = f"{NOTE_ID}_s{i+1:03d}"
        
        # Hydrate
        start, end, status = find_offset(NOTE_TEXT, txt, ctx)
        
        # Common row
        # (leave start/end blank for Span_Annotations)
        row_base = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure", ctx, txt, 
            "", "", "", f"=LEN(F{ws_span.max_row + 1})", 
            lbl, norm, field, evt, 
            "FALSE", "FALSE", "", "", "", "needs_hydration"
        ]
        ws_span.append(row_base)
        
        # Hydrated row
        row_hyd = list(row_base)
        row_hyd[7] = start if start is not None else ""
        row_hyd[8] = end if end is not None else ""
        row_hyd[9] = end - start if (start is not None and end is not None) else ""
        row_hyd[19] = status
        ws_hyd.append(row_hyd)

    # ------------------------------------------------------------------------------------
    # Sheet 5: Event_Log
    # ------------------------------------------------------------------------------------
    ws_evt = wb["Event_Log"]
    evt_headers = [
        "source_file", "note_id", "event_id", "event_type", "method",
        "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements",
        "specimens", "findings", "is_historical", "reviewer", "comments",
        "device_size", "device_material",
        "outcome_airway_lumen_pre", "outcome_airway_lumen_post",
        "outcome_symptoms", "outcome_pleural", "outcome_complication"
    ]
    if ws_evt.max_row == 1:
        ws_evt.append(evt_headers)

    # Define events data
    # Event 1: Aspiration
    ws_evt.append([
        SOURCE_FILE, NOTE_ID, "evt_01", "Therapeutic Aspiration", "therapeutic aspiration",
        "trachea, mainstems, RML", "None", "", "", "", "",
        "", "mucus plug", "FALSE", "", "Cleaned multiple sites",
        "", "",
        "", "", "", "", ""
    ])
    # Event 2: Thermal Ablation (Electrocautery)
    ws_evt.append([
        SOURCE_FILE, NOTE_ID, "evt_02", "Thermal Ablation", "electrocautery",
        "RML", "Needle Knife", "", "", "", "",
        "", "obstruction", "FALSE", "", "Radial cuts",
        "", "",
        "", "", "", "", ""
    ])
    # Event 3: Cryotherapy
    ws_evt.append([
        SOURCE_FILE, NOTE_ID, "evt_03", "Cryotherapy", "cryotherapy",
        "RML", "Cryoprobe", "", "", "", "",
        "", "ablation", "FALSE", "", "Freeze-thaw cycles",
        "1.7mm", "",
        "", "", "", "", ""
    ])
    # Event 4: Balloon 1
    ws_evt.append([
        SOURCE_FILE, NOTE_ID, "evt_04", "Airway Dilation", "balloon dilation",
        "RML", "Elation balloon", "", "", "", "6 mm",
        "", "", "FALSE", "", "1st dilation",
        "6mm", "",
        "25%", "100%", "", "", "none"
    ])
    # Event 5: Balloon 2
    ws_evt.append([
        SOURCE_FILE, NOTE_ID, "evt_05", "Airway Dilation", "balloon dilation",
        "RML", "Elation balloon", "", "", "", "8 mm",
        "", "", "FALSE", "", "2nd dilation",
        "8mm", "",
        "25%", "100%", "", "", "none"
    ])

    # ------------------------------------------------------------------------------------
    # Sheet 6: V3_Procedure_Events
    # ------------------------------------------------------------------------------------
    ws_v3 = wb["V3_Procedure_Events"]
    v3_headers = [
        "note_id", "event_id", "type",
        "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station",
        "lesion.type", "lesion.size_mm",
        "method", "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote",
        "stent.size", "stent.material_or_brand", "catheter.size_fr",
        "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post",
        "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"
    ]
    if ws_v3.max_row == 1:
        ws_v3.append(v3_headers)

    # Mapped rows
    # Evt 1
    ws_v3.append([
        NOTE_ID, "evt_01", "Therapeutic Aspiration",
        "Airway", "RML", "", "",
        "mucus", "",
        "therapeutic aspiration", "[]", "[]", "[]", "['mucus plug']", "Successful therapeutic aspiration...",
        "", "", "",
        "", "", "", "", ""
    ])
    # Evt 2
    ws_v3.append([
        NOTE_ID, "evt_02", "Thermal Ablation",
        "Airway", "RML", "", "",
        "obstruction", "",
        "electrocautery", "['Needle Knife']", "[]", "[]", "['radial cuts']", "Endobronchial obstruction at RML was treated...",
        "", "", "",
        "", "", "", "", ""
    ])
    # Evt 3
    ws_v3.append([
        NOTE_ID, "evt_03", "Cryotherapy",
        "Airway", "RML", "", "",
        "", "",
        "cryotherapy", "['Cryoprobe 1.7mm']", "[]", "[]", "['ablation']", "Cryoprobe 1.7mm probe...",
        "", "", "",
        "", "", "", "", ""
    ])
    # Evt 4
    ws_v3.append([
        NOTE_ID, "evt_04", "Airway Dilation",
        "Airway", "RML", "", "",
        "", "",
        "balloon dilation", "['Elation balloon']", "['6 mm']", "[]", "[]", "Balloon dilation... 6 mm",
        "", "", "",
        "25%", "100%", "", "", "none"
    ])
     # Evt 5
    ws_v3.append([
        NOTE_ID, "evt_05", "Airway Dilation",
        "Airway", "RML", "", "",
        "", "",
        "balloon dilation", "['Elation balloon']", "['8 mm']", "[]", "[]", "Balloon dilation... 8 mm",
        "", "", "",
        "25%", "100%", "", "", "none"
    ])

    # ------------------------------------------------------------------------------------
    # Sheet 7: V3_Registry_JSON
    # ------------------------------------------------------------------------------------
    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 1:
        ws_json.append(["schema_version", "note_id", "json_object"])
    
    registry_data = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "procedures": [
            {"event_id": "evt_01", "type": "Therapeutic Aspiration", "target": {"location": "RML"}},
            {"event_id": "evt_02", "type": "Thermal Ablation", "method": "Electrocautery"},
            {"event_id": "evt_03", "type": "Cryotherapy"},
            {"event_id": "evt_04", "type": "Airway Dilation", "outcomes": {"airway": {"lumen_pre": "25%", "lumen_post": "100%"}}},
            {"event_id": "evt_05", "type": "Airway Dilation"}
        ],
        "no_immediate_complications": True
    }
    ws_json.append(["v3.0", NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()