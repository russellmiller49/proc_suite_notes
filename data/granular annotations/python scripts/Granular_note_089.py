import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_089"
SOURCE_FILE = "note_089.txt"
PROCEDURE_DATE = ""  # Not specified in text
NOTE_TEXT = """NOTE_ID:  note_089 SOURCE_FILE: note_089.txt INDICATION FOR OPERATION:  [REDACTED]is a 57 year old-year-old male who presents with airway narrowing.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
ESTIMATED BLOOD LOSS:   None
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: 
Initial Airway Inspection Findings:
Right middle lobe orifice was significantly narrowed.
Successful therapeutic aspiration was performed to clean out the Right Mainstem and Bronchus Intermedius  from mucus.
Endobronchial obstruction at RML orifice was treated with the following modalities:
Modality	Tools	Setting/Mode	Duration	Results
Electrocautery	knife	endoI 4, 4, 1	4 sec	Radial cuts
APC	 	 	 	 
Laser	 	 	 	 
CoreCath	 	 	 	 
Cryoprobe	 	 	 	 
Balloon dilation was performed at right middle lobe orifice.
6/7/8 Elation balloon was used to perform dilation to 8 mm at the right middle lobe orifice.
Total 3 inflations with dilation time of 60 seconds each.
Prior to treatment, affected airway was note to be 10% patent.  After treatment, the airway was 100% patent.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
none
IMPRESSION/PLAN: [REDACTED]is a 57 year old-year-old male who presents for bronchoscopy for airway narrowing.
- 2 weeks f/u bronch (order placed with level 2)"""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# -------------------------------------------------------------------------
# CONFIGURATION
# -------------------------------------------------------------------------
PROCEDURE_FLAGS = [
    "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
    "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy",
    "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration",
    "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
    "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation",
    "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
    "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis",
    "pleural_biopsy", "fibrinolytic_therapy"
]

# -------------------------------------------------------------------------
# GENERATOR CLASS
# -------------------------------------------------------------------------
class Phase0WorkbookGenerator:
    def __init__(self):
        self.wb = None
        self.note_text = NOTE_TEXT
        self.spans = []
        self.events = []
        self.flags = {k: 0 for k in PROCEDURE_FLAGS}
        
    def load_template(self):
        if not os.path.exists(TEMPLATE_PATH):
            # Create a dummy workbook if template is missing (for robust script execution in isolation)
            self.wb = openpyxl.Workbook()
            self.wb.remove(self.wb.active)
            for sheet_name in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
                self.wb.create_sheet(sheet_name)
        else:
            self.wb = openpyxl.load_workbook(TEMPLATE_PATH)

    def process_flags(self):
        # 1. Base Bronchoscopy
        self.flags['diagnostic_bronchoscopy'] = 1 # Implicit base
        
        # 2. Therapeutic Aspiration (Explicit)
        if "Therapeutic aspiration" in self.note_text or "therapeutic aspiration" in self.note_text:
            self.flags['therapeutic_aspiration'] = 1
            
        # 3. Airway Dilation (Explicit)
        if "Balloon dilation" in self.note_text:
            self.flags['airway_dilation'] = 1
            
        # 4. Thermal Ablation (Electrocautery)
        if "Electrocautery" in self.note_text:
            self.flags['thermal_ablation'] = 1
            
        # 5. Tumor Debulking (Non-thermal) - No explicit mechanical debulking mentioned outside of aspiration/cautery
        
        # 6. Cryotherapy - Mentioned in list but blank results, so 0.
        
        # 7. Rigid - Explicitly "Flexible Therapeutic Bronchoscope"
        
    def generate_spans_and_events(self):
        # Event 1: Therapeutic Aspiration
        e1_id = "evt_01"
        self.events.append({
            "event_id": e1_id,
            "type": "therapeutic_aspiration",
            "method": "aspiration",
            "anatomy": "Right Mainstem, Bronchus Intermedius",
            "outcome_complication": "none"
        })
        self.add_span("Therapeutic aspiration", "PROC_METHOD", "aspiration", e1_id, context="Successful")
        self.add_span("Right Mainstem", "ANAT_AIRWAY", "Right Mainstem", e1_id)
        self.add_span("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", e1_id)
        
        # Event 2: Electrocautery (Thermal Ablation)
        e2_id = "evt_02"
        self.events.append({
            "event_id": e2_id,
            "type": "thermal_ablation",
            "method": "electrocautery",
            "anatomy": "Right Middle Lobe Orifice",
            "device": "knife"
        })
        self.add_span("Electrocautery", "PROC_METHOD", "electrocautery", e2_id)
        self.add_span("knife", "DEV_INSTRUMENT", "electrocautery knife", e2_id)
        self.add_span("RML orifice", "ANAT_AIRWAY", "RML", e2_id, context="Endobronchial obstruction at")

        # Event 3: Balloon Dilation
        e3_id = "evt_03"
        self.events.append({
            "event_id": e3_id,
            "type": "airway_dilation",
            "method": "balloon_dilation",
            "anatomy": "Right Middle Lobe Orifice",
            "device": "6/7/8 Elation balloon",
            "outcome_lumen_pre": "10%",
            "outcome_lumen_post": "100%"
        })
        self.add_span("Balloon dilation", "PROC_METHOD", "balloon_dilation", e3_id)
        self.add_span("right middle lobe orifice", "ANAT_AIRWAY", "RML", e3_id, context="Balloon dilation was performed at")
        self.add_span("6/7/8 Elation balloon", "DEV_INSTRUMENT", "Elation balloon 6-7-8mm", e3_id)
        self.add_span("8 mm", "MEAS_AIRWAY_DIAM", "8 mm", e3_id, context="dilation to")
        self.add_span("3 inflations", "MEAS_COUNT", "3", e3_id)
        
        # Outcomes (Shared between RML interventions, linking to dilation primarily as it was the final step)
        self.add_span("10% patent", "OUTCOME_AIRWAY_LUMEN_PRE", "10%", e3_id)
        self.add_span("100% patent", "OUTCOME_AIRWAY_LUMEN_POST", "100%", e3_id)
        
        # Global Complications
        self.add_span("No immediate complications", "OUTCOME_COMPLICATION", "none", "global")

    def add_span(self, text, label, value, event_id, context=None):
        self.spans.append({
            "text": text,
            "label": label,
            "value": value,
            "event_id": event_id,
            "context": context,
            "hydration": "needs_hydration"
        })

    def hydrate_spans(self):
        hydrated = []
        for s in self.spans:
            txt = s["text"]
            count = self.note_text.count(txt)
            
            start_char = ""
            end_char = ""
            status = "ambiguous"
            
            if count == 1:
                start_char = self.note_text.find(txt)
                end_char = start_char + len(txt)
                status = "hydrated_unique"
            elif count > 1 and s["context"]:
                # Try to find context within a window before occurrence
                matches = [m.start() for m in re.finditer(re.escape(txt), self.note_text)]
                found = False
                for m_start in matches:
                    window_start = max(0, m_start - 120)
                    window_text = self.note_text[window_start:m_start]
                    if s["context"] in window_text:
                        start_char = m_start
                        end_char = m_start + len(txt)
                        status = "hydrated_context"
                        found = True
                        break
                if not found:
                    status = f"ambiguous_count={count}_ctx_fail"
            else:
                status = f"ambiguous_count={count}"
                
            row = s.copy()
            row["start_char"] = start_char
            row["end_char"] = end_char
            row["hydration_status"] = status
            row["len"] = len(txt)
            hydrated.append(row)
        return hydrated

    def write_worksheet(self):
        self.load_template()
        
        # 1. Note_Text
        ws = self.wb["Note_Text"]
        ws.append([NOTE_ID, SOURCE_FILE, self.note_text])
        
        # 2. Note_Index
        ws = self.wb["Note_Index"]
        row = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Success", ""]
        row += [self.flags[f] for f in PROCEDURE_FLAGS]
        ws.append(row)
        
        # 3. Span_Annotations (Anchor First)
        ws = self.wb["Span_Annotations"]
        # Header: source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, start_char, end_char, span_len, label, normalized_value, schema_field, event_id, is_negated, is_historical, time_anchor, reviewer, comments, hydration_status
        for i, s in enumerate(self.spans):
            ws.append([
                SOURCE_FILE, NOTE_ID, f"span_{i+1:03d}", "", s["context"] or "", s["text"], "", 
                "", "", f"=LEN(F{i+2})", s["label"], s["value"], "", s["event_id"], 
                "", "", "", "", "", s["hydration"]
            ])
            
        # 4. Span_Hydrated
        ws = self.wb["Span_Hydrated"]
        h_spans = self.hydrate_spans()
        for i, s in enumerate(h_spans):
            ws.append([
                SOURCE_FILE, NOTE_ID, f"span_{i+1:03d}", "", s["context"] or "", s["text"], "", 
                s["start_char"], s["end_char"], s["len"], s["label"], s["value"], "", s["event_id"], 
                "", "", "", "", "", s["hydration_status"]
            ])

        # 5. Event_Log
        ws = self.wb["Event_Log"]
        # Headers: source_file, note_id, event_id, event_type, method, anatomy_target, device, needle_gauge, stations, counts, measurements, specimens, findings, is_historical, reviewer, comments, device_size, device_material, outcome_airway_lumen_pre, outcome_airway_lumen_post, outcome_symptoms, outcome_pleural, outcome_complication
        for e in self.events:
            ws.append([
                SOURCE_FILE, NOTE_ID, e["event_id"], e["type"], e.get("method",""), 
                e.get("anatomy",""), e.get("device",""), "", "", "", "", "", "", "", "", "", 
                "", "", # Size/Mat
                e.get("outcome_lumen_pre",""), e.get("outcome_lumen_post",""), "", "", e.get("outcome_complication","")
            ])
            
        # 6. V3_Procedure_Events
        ws = self.wb["V3_Procedure_Events"]
        for e in self.events:
            # Simple mapping for demo
            ws.append([
                NOTE_ID, e["event_id"], e["type"], 
                "airway", "", "", "", # Anatomy columns
                "", "", # Lesion
                e.get("method",""), "{}", "{}", "{}", "{}", "", # JSONs + Quote
                "", "", "", # Stent/Cath
                e.get("outcome_lumen_pre",""), e.get("outcome_lumen_post",""), "", "", e.get("outcome_complication","")
            ])

        # 7. V3_Registry_JSON
        ws = self.wb["V3_Registry_JSON"]
        reg_data = {
            "schema_version": "v3.0",
            "note_id": NOTE_ID,
            "no_immediate_complications": True,
            "procedures": self.events
        }
        ws.append([json.dumps(reg_data, indent=2)])
        
        self.wb.save(OUTPUT_PATH)

# -------------------------------------------------------------------------
# EXECUTION
# -------------------------------------------------------------------------
if __name__ == "__main__":
    generator = Phase0WorkbookGenerator()
    generator.process_flags()
    generator.generate_spans_and_events()
    generator.write_worksheet()
    print(f"Workbook generated: {OUTPUT_PATH}")