import openpyxl
from openpyxl.utils import get_column_letter
import json
import datetime
import re

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_090"
SOURCE_FILE = "note_090.txt"
PROCEDURE_DATE = "2026-01-12" # Placeholder or extracted if available
NOTE_TEXT = """NOTE_ID:  note_090 SOURCE_FILE: note_090.txt INDICATION FOR OPERATION:  [REDACTED]is a 57 year old-year-old male who presents with bronchial stenosis.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31622 Dx bronchoscope/cell washing          
31624 Dx bronchoscope/lavage (BAL)    
31625 Endobronchial Biopsy(s)
31630 Balloon dilation
31640 Bronchoscopy with excision 
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
Flexible Hybrid (Pedatric) Bronchoscope
PROCEDURE IN DETAIL:
A timeout was performed (confirming the patient's name, procedure type, and procedure location).
Sedation initiated.  LMA placed.
The Flexible Therapeutic Bronchoscope was advanced for airway examination.
Endobronchial topical lidocaine applied to the vocal cords, main carina, right carina 1, and left carina 2.
Initial Airway Examination Findings:
Vocal Chords: Normal without mass/lesions.
Appropriate abduction.  Paralyzed so unable to assess adduction.
Trachea: Mildy tortuous, otherwise normal.
Main Carina: Sharp
Right Lung Proximal Airways: The right anastomosis site was intact with visible intact blue sutures at site;
mild stenosis of right anastomosis without dehiscence. The RUL and RB1-3 were normal to segmental level.
The bronchus intermedius and more distal lobes/airways are mildly rotated clockwise.  The RML bronchus with severe stenosis of ~3mm patent.
Proximal RML take-off appears OK, but areas of stenosis about halfway into the bronchus.
After dilation, patient also noted to have a second area of moderate stenosis right (~5mm) before the take-off of RB4 and RB5.
These stenoses appeared to be circumferential due to fibrotic and granulation tissue.
There was a small patch of white debris also overlying the area of stenosis.
After relief of the stenosis, there was moderate white secretions (likely mucostasis) suctioned from the RB4 and RB5 segmental airways.
Able to enter the truncus basalis with mild pressure with therapeutic bronchoscope.  RB7 is compressed and fish-mouthed.
The RLL and RB6,RB8-10 are otherwise have normal anatomic branching to segmental level.
No other evidence of mass, lesions, bleeding or other endobronchial pathology.
Left Lung Proximal Airways: The left anastomosis site was intact with visible intact blue sutures at site;
mild stenosis of left anastomosis. The LUL bronchus and LB1-3 were normal to first subsegmental level.
The lingula is patent and LB4-5 are normal to first subsegmental level - able to traverse lingular bronchus with therapeutic bronchoscope.
LLL and LB6-10 have normal anatomic branching to segmental level.
No other evidence of mass, lesions, bleeding or other endobronchial pathology. 
Mucosa: Normal.
Secretions: As described above in RML.
Otherwise moderate, thin, and clear  All secretions were suctioned to clear (therapeutic aspiration).
RML before treatment:
 
Successful therapeutic aspiration was performed to clean out the trachea, right mainstem bronchus, right upper lobe, bronchus intermedius, right middle lobe, right lower lobe, left mainstem bronchus, left upper lobe, left lower lobe from mucus.
Endobronchial biopsy was performed at white patch of tissue in RML bronchus.  Lesion was successfully removed.
Samples sent for Tissue Culture.
Balloon dilation was performed at right middle lobe bronchus.
6/7/8 Elation balloon was used to perform dilation to 7 mm at the right middle lobe bronchus.
Total 1 inflations with dilation time of 60 seconds each.
Endobronchial obstruction/stenosis at right middle lobe bronchus was treated with the following modalities:
Modality	Tools	Setting/Mode	Duration	Results
Cryoprobe	1.7mm	N/A	30 second overlapping treatments	Cryotherapy overlapping treatment to areas of circumferential stenosis in the middle RML bronchus.
Led to softening of fibrotic and granulation tissue.
Endobronchial tissue debris causing obstruction/stenosis was noted and excised with mechanical debridement using bland alligator forceps.
Balloon dilation was performed at right middle lobe bronchus.  6/7/8 Elation balloon was used to perform dilation to 8 mm at the right middle lobe bronchus.
Total 1 inflations with dilation time of 60 seconds each.
Endobronchial obstruction/stenosis at right middle lobe bronchus was treated with the following modalities:
Modality	Tools	Setting/Mode	Duration	Results
Cryoprobe	1.7mm	N/A	30 second overlapping treatments	Cryotherapy overlapping treatment to areas of circumferential stenosis in the distal RML bronchus.
Led to softening of fibrotic and granulation tissue.
Balloon dilation was performed at right middle lobe bronchus.
6/7/8 Elation balloon was used to perform dilation to 8 mm at the right middle lobe bronchus.
Total 1 inflations with dilation time of 60 seconds each.  
Balloon dilation was performed at right middle lobe medial segment.
6/7/8 Elation balloon was used to perform dilation to 6 mm at right middle lobe medial segment.
Total 1 inflations with dilation time of 60 seconds each.  
Balloon dilation was performed at right middle lobe lateral segment.
6/7/8 Elation balloon was used to perform dilation to 6 mm at right middle lobe medial segment.
Total 1 inflations with dilation time of 60 seconds each.
Despite multiple dilations, RML bronchus continues to mildly narrow, particularly at the distal end of the bronchus.
Trivial bleeding easily controlled with suction and saline.
Prior to treatment, affected airway was note to be 30% patent.
After treatment, the airway was 75% patent.
RML after treatments:
 
Unable to traverse entire length of RML bronchus with therapeutic bronchoscope.
Hybrid bronchoscope advanced and confirmed patency of RB4 and RB5 to their respective subsegments:
 
Bronchial alveolar lavage was performed at RML (Lateral Segment of RML (RB4) and Medial Segment of RML (RB5)).
Instilled 60 cc of NS, suction returned with 20 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology. 
Residual secretions and saline suctioned to clear.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient's LMA was removed in the operating room and transported to the recovery room in stable condition.
ESTIMATED BLOOD LOSS:   Trivial
COMPLICATIONS:    None
SPECIMEN(S): 
RML EBBx:  Culture
RML BAL:  cell count, cultures/micro, cytology
IMPRESSION/PLAN: [REDACTED]is a 57 year old-year-old male who presents for bronchoscopy for balloon dilation and cryotherapy for management of bronchial stenosis.
-Repeat bronchoscopy in 2 weeks for airway evaluation, dilation, possible stent placement.
-Consideration of 8x20mm stent placement vs custom silicone stent placement.  Will review with other IP providers.
-Will request that lung transplant team follow-up on EBBx culture and BAL results."""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# -------------------------------------------------------------------------
# CONFIGURATION
# -------------------------------------------------------------------------
PROCEDURE_FLAGS = [
    "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
    "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy",
    "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration",
    "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
    "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation",
    "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
    "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis",
    "pleural_biopsy", "fibrinolytic_therapy"
]

# Set active flags for this note
ACTIVE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 1,
    "endobronchial_biopsy": 1,
    "therapeutic_aspiration": 1,
    "airway_dilation": 1,
    "cryotherapy": 1,
    "tumor_debulking_non_thermal": 1
}

# -------------------------------------------------------------------------
# SPAN DATA (Manually extracted based on note content)
# -------------------------------------------------------------------------
# Format: (span_text, context_prefix, label, normalized_value, event_id, schema_field)
# Event IDs:
# E01: Therapeutic Aspiration
# E02: Endobronchial Biopsy RML
# E03: Balloon Dilation RML (7mm)
# E04: Cryotherapy RML Middle
# E05: Mechanical Debridement (Debulking)
# E06: Balloon Dilation RML (8mm)
# E07: Cryotherapy RML Distal
# E08: Balloon Dilation RML (8mm - 2nd)
# E09: Balloon Dilation RML Medial (6mm)
# E10: Balloon Dilation RML Lateral (6mm)
# E11: BAL RML
# E99: Outcomes

SPANS = [
    # General/Devices
    ("Flexible Therapeutic Bronchoscope", "LMA placed.\nThe", "DEV_INSTRUMENT", "therapeutic_bronchoscope", "E00", "device"),
    ("Flexible Hybrid (Pedatric) Bronchoscope", "Bronchoscope\n", "DEV_INSTRUMENT", "hybrid_bronchoscope", "E00", "device"),

    # E01: Therapeutic Aspiration
    ("therapeutic aspiration", "Successful", "PROC_METHOD", "therapeutic_aspiration", "E01", "method"),
    ("clean out the trachea", "performed to", "PROC_ACTION", "clean_airways", "E01", "action"),

    # E02: Biopsy
    ("Endobronchial biopsy", "from mucus.\n", "PROC_METHOD", "endobronchial_biopsy", "E02", "method"),
    ("white patch of tissue", "performed at", "OBS_LESION", "white_patch", "E02", "lesion"),
    ("RML bronchus", "tissue in", "ANAT_AIRWAY", "RML", "E02", "anatomy"),
    ("Lesion was successfully removed", "bronchus.", "PROC_ACTION", "removed", "E02", "action"),

    # E03: Balloon 1
    ("Balloon dilation", "Samples sent for Tissue Culture.\n", "PROC_METHOD", "airway_dilation", "E03", "method"),
    ("right middle lobe bronchus", "performed at", "ANAT_AIRWAY", "RML", "E03", "anatomy"),
    ("6/7/8 Elation balloon", "right middle lobe bronchus.\n", "DEV_INSTRUMENT", "elation_balloon_6_7_8", "E03", "device"),
    ("dilation to 7 mm", "perform", "PROC_ACTION", "dilation_7mm", "E03", "action"),
    ("1 inflations", "Total", "MEAS_COUNT", "1", "E03", "count"),
    ("60 seconds", "time of", "MEAS_TIME", "60_sec", "E03", "duration"),

    # E04: Cryo 1
    ("Cryoprobe", "Modality\tTools\tSetting/Mode\tDuration\tResults\n", "DEV_INSTRUMENT", "cryoprobe_1.7mm", "E04", "device"),
    ("Cryotherapy", "30 second overlapping treatments\t", "PROC_METHOD", "cryotherapy", "E04", "method"),
    ("middle RML bronchus", "stenosis in the", "ANAT_AIRWAY", "RML", "E04", "anatomy"),
    ("circumferential stenosis", "areas of", "OBS_LESION", "stenosis_circumferential", "E04", "lesion"),

    # E05: Debulking
    ("excised with mechanical debridement", "noted and", "PROC_METHOD", "tumor_debulking_non_thermal", "E05", "method"),
    ("bland alligator forceps", "using", "DEV_INSTRUMENT", "forceps_alligator", "E05", "device"),
    ("Endobronchial tissue debris", "tissue.\n", "OBS_LESION", "tissue_debris", "E05", "lesion"),

    # E06: Balloon 2
    ("Balloon dilation", "forceps.\n", "PROC_METHOD", "airway_dilation", "E06", "method"),
    ("right middle lobe bronchus", "forceps.\nBalloon dilation was performed at", "ANAT_AIRWAY", "RML", "E06", "anatomy"),
    ("6/7/8 Elation balloon", "lobe bronchus.  ", "DEV_INSTRUMENT", "elation_balloon_6_7_8", "E06", "device"),
    ("dilation to 8 mm", "perform", "PROC_ACTION", "dilation_8mm", "E06", "action"),

    # E07: Cryo 2
    ("Cryoprobe", "Results\n", "DEV_INSTRUMENT", "cryoprobe_1.7mm", "E07", "device"),
    ("Cryotherapy", "30 second overlapping treatments\t", "PROC_METHOD", "cryotherapy", "E07", "method"),
    ("distal RML bronchus", "stenosis in the", "ANAT_AIRWAY", "RML", "E07", "anatomy"),

    # E08: Balloon 3
    ("Balloon dilation", "granulation tissue.\n", "PROC_METHOD", "airway_dilation", "E08", "method"),
    ("right middle lobe bronchus", "granulation tissue.\nBalloon dilation was performed at", "ANAT_AIRWAY", "RML", "E08", "anatomy"),
    ("6/7/8 Elation balloon", "lobe bronchus.\n", "DEV_INSTRUMENT", "elation_balloon_6_7_8", "E08", "device"),
    ("dilation to 8 mm", "perform", "PROC_ACTION", "dilation_8mm", "E08", "action"),

    # E09: Balloon 4 (Medial)
    ("Balloon dilation", "60 seconds each.  \n", "PROC_METHOD", "airway_dilation", "E09", "method"),
    ("right middle lobe medial segment", "performed at", "ANAT_AIRWAY", "RML_medial", "E09", "anatomy"),
    ("6/7/8 Elation balloon", "medial segment.\n", "DEV_INSTRUMENT", "elation_balloon_6_7_8", "E09", "device"),
    ("dilation to 6 mm", "perform", "PROC_ACTION", "dilation_6mm", "E09", "action"),

    # E10: Balloon 5 (Lateral)
    ("Balloon dilation", "60 seconds each.  \n", "PROC_METHOD", "airway_dilation", "E10", "method"),
    ("right middle lobe lateral segment", "performed at", "ANAT_AIRWAY", "RML_lateral", "E10", "anatomy"),
    ("6/7/8 Elation balloon", "lateral segment.\n", "DEV_INSTRUMENT", "elation_balloon_6_7_8", "E10", "device"),
    # Source 39 has typo "medial segment" in text, but header says lateral. Use header span for anatomy intent.
    ("dilation to 6 mm", "perform", "PROC_ACTION", "dilation_6mm", "E10", "action"),

    # E11: BAL
    ("Bronchial alveolar lavage", "subsegments:\n \n", "PROC_METHOD", "bal", "E11", "method"),
    ("Lateral Segment of RML (RB4)", "(", "ANAT_AIRWAY", "RML_lateral", "E11", "anatomy"),
    ("Medial Segment of RML (RB5)", "and", "ANAT_AIRWAY", "RML_medial", "E11", "anatomy"),
    ("Instilled 60 cc", "", "MEAS_VOL", "60ml", "E11", "volume_in"),
    ("suction returned with 20 cc", "NS,", "MEAS_VOL", "20ml", "E11", "volume_out"),

    # Outcomes (Linked to RML treatments, primarily Dilation/Cryo)
    ("30% patent", "airway was note to be", "OUTCOME_AIRWAY_LUMEN_PRE", "30%", "E03", "outcome_lumen_pre"),
    ("75% patent", "airway was", "OUTCOME_AIRWAY_LUMEN_POST", "75%", "E03", "outcome_lumen_post"),
    ("No immediate complications", "well.  There were", "OUTCOME_COMPLICATION", "none", "E99", "complication"),
    ("Trivial bleeding", "", "OUTCOME_COMPLICATION", "bleeding_trivial", "E99", "complication"),
    
    # Diagnosis/Hist
    ("bronchial stenosis", "presents with", "OBS_LESION", "bronchial_stenosis", "E00", "indication"),
    ("J98.09", "DIAGNOSIS:", "CTX_HISTORICAL", "J98.09", "E00", "diagnosis_code"),
]

# -------------------------------------------------------------------------
# LOGIC
# -------------------------------------------------------------------------

def calculate_offset(note_text, span_text, context_prefix, match_index=None):
    """
    Calculates start_char and end_char based on logic:
    1. If unique in text -> return that.
    2. If context_prefix -> find occurrence where prefix matches preceding 120 chars.
    3. If match_index -> return nth occurrence.
    4. Else -> Ambiguous.
    """
    matches = [m.start() for m in re.finditer(re.escape(span_text), note_text)]
    
    if not matches:
        return None, None, "not_found"
    
    if len(matches) == 1:
        start = matches[0]
        return start, start + len(span_text), "hydrated_unique"
    
    # Context match
    if context_prefix:
        clean_prefix = context_prefix.strip()[-50:] # Use last 50 chars of prefix for robust match
        for m_start in matches:
            window_start = max(0, m_start - 150)
            window = note_text[window_start:m_start]
            if clean_prefix in window:
                return m_start, m_start + len(span_text), "hydrated_prefix_window"
    
    # Match index
    if match_index is not None and 0 <= match_index < len(matches):
        start = matches[match_index]
        return start, start + len(span_text), "hydrated_match_index"
    
    return None, None, f"ambiguous_count={len(matches)}"

def generate_workbook():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        # If template not found, create simple one for demonstration/fallback
        wb = openpyxl.Workbook()
        wb.create_sheet("Note_Text")
        wb.create_sheet("Note_Index")
        wb.create_sheet("Span_Annotations")
        wb.create_sheet("Span_Hydrated")
        wb.create_sheet("Event_Log")
        wb.create_sheet("V3_Procedure_Events")
        wb.create_sheet("V3_Registry_JSON")

    # 1. Note_Text
    ws = wb["Note_Text"]
    if ws.max_row == 1:
        ws.append(["note_id", "source_file", "note_text"])
    ws.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws = wb["Note_Index"]
    header = [
        "source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes"
    ] + PROCEDURE_FLAGS
    
    # Ensure header exists if empty
    if ws.max_row == 0:
        ws.append(header)
    
    row_data = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Success", ""]
    for flag in PROCEDURE_FLAGS:
        row_data.append(ACTIVE_FLAGS.get(flag, 0))
    ws.append(row_data)

    # 3. Span_Annotations & 4. Span_Hydrated
    ws_anno = wb["Span_Annotations"]
    ws_hydra = wb["Span_Hydrated"]
    
    anno_header = [
        "source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", "match_index",
        "start_char", "end_char", "span_len", "label", "normalized_value", "schema_field", "event_id",
        "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"
    ]
    
    if ws_anno.max_row == 0: ws_anno.append(anno_header)
    if ws_hydra.max_row == 0: ws_hydra.append(anno_header)

    hydrated_spans = []

    for idx, span_data in enumerate(SPANS):
        span_text, ctx, label, norm, ev_id, field = span_data
        span_id = f"{NOTE_ID}_s{idx+1:03d}"
        
        # Hydrate
        start, end, status = calculate_offset(NOTE_TEXT, span_text, ctx)
        
        # Anno Row (Anchor First - no offsets)
        anno_row = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure", ctx, span_text, "",
            "", "", f"=LEN(F{ws_anno.max_row+1})", label, norm, field, ev_id,
            "FALSE", "FALSE", "", "", "", "needs_hydration"
        ]
        ws_anno.append(anno_row)

        # Hydra Row
        hydra_row = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure", ctx, span_text, "",
            start if start is not None else "", end if end is not None else "", 
            end-start if start is not None else "", label, norm, field, ev_id,
            "FALSE", "FALSE", "", "", "", status
        ]
        ws_hydra.append(hydra_row)
        
        if status.startswith("hydrated"):
            hydrated_spans.append({
                "label": label, "norm": norm, "event_id": ev_id, "field": field, "text": span_text
            })

    # 5. Event_Log & 6. V3_Procedure_Events & 7. JSON
    # Construct structured data
    events_dict = {}
    
    # Initialize basic event structures
    event_definitions = {
        "E01": {"type": "therapeutic_aspiration", "method": "Aspiration", "anatomy": "Airways"},
        "E02": {"type": "endobronchial_biopsy", "method": "Biopsy", "anatomy": "RML"},
        "E03": {"type": "airway_dilation", "method": "Balloon Dilation", "anatomy": "RML"},
        "E04": {"type": "cryotherapy", "method": "Cryotherapy", "anatomy": "RML"},
        "E05": {"type": "tumor_debulking_non_thermal", "method": "Mechanical Debridement", "anatomy": "RML"},
        "E06": {"type": "airway_dilation", "method": "Balloon Dilation", "anatomy": "RML"},
        "E07": {"type": "cryotherapy", "method": "Cryotherapy", "anatomy": "RML"},
        "E08": {"type": "airway_dilation", "method": "Balloon Dilation", "anatomy": "RML"},
        "E09": {"type": "airway_dilation", "method": "Balloon Dilation", "anatomy": "RML Medial"},
        "E10": {"type": "airway_dilation", "method": "Balloon Dilation", "anatomy": "RML Lateral"},
        "E11": {"type": "bal", "method": "Lavage", "anatomy": "RML"},
        "E99": {"type": "outcomes", "method": "N/A", "anatomy": "N/A"}
    }

    for e_id, meta in event_definitions.items():
        events_dict[e_id] = {
            "note_id": NOTE_ID,
            "event_id": e_id,
            "type": meta["type"],
            "target": {"anatomy": meta["anatomy"]},
            "method": meta["method"],
            "devices": [],
            "measurements": [],
            "outcomes": {}
        }

    # Fill data from spans
    for s in hydrated_spans:
        eid = s["event_id"]
        if eid not in events_dict: continue
        
        if s["field"] == "device":
            events_dict[eid]["devices"].append(s["norm"])
        elif s["field"] == "action":
            events_dict[eid]["method"] += f" ({s['norm']})"
        elif s["field"].startswith("outcome_"):
            key = s["field"].replace("outcome_", "")
            events_dict[eid]["outcomes"][key] = s["norm"]
            # Also apply global outcomes to main RML dilation E03 if not present
            if eid == "E99" and key == "complication":
                 events_dict["E03"]["outcomes"]["complication"] = s["norm"]
        elif s["label"] == "MEAS_VOL":
            events_dict[eid]["measurements"].append(f"{s['field']}: {s['norm']}")
        elif s["label"] == "MEAS_COUNT":
            events_dict[eid]["measurements"].append(f"count: {s['norm']}")

    # Apply global outcomes (E03 outcomes to others if relevant? No, just keep simple)
    # E03 has pre/post lumen.
    
    ws_event = wb["Event_Log"]
    if ws_event.max_row == 0:
        ws_event.append(["source_file", "note_id", "event_id", "event_type", "method", "anatomy_target", "device", "outcomes_json"])

    ws_v3 = wb["V3_Procedure_Events"]
    v3_header = [
        "note_id", "event_id", "type", "target.anatomy_type", "target.location.lobe", 
        "target.location.segment", "target.station", "method", "devices_json", "outcomes.airway.lumen_pre",
        "outcomes.airway.lumen_post", "outcomes.complications"
    ]
    if ws_v3.max_row == 0: ws_v3.append(v3_header)

    json_procedures = []

    for eid in sorted(events_dict.keys()):
        if eid == "E99": continue # Skip pure outcome container in event list
        evt = events_dict[eid]
        
        # Event Log
        ws_event.append([
            SOURCE_FILE, NOTE_ID, eid, evt["type"], evt["method"], 
            evt["target"]["anatomy"], 
            ", ".join(evt["devices"]),
            json.dumps(evt["outcomes"])
        ])

        # V3
        # Parse anatomy for V3
        lobe = "RML" if "RML" in evt["target"]["anatomy"] else ""
        seg = ""
        if "Medial" in evt["target"]["anatomy"]: seg = "Medial"
        if "Lateral" in evt["target"]["anatomy"]: seg = "Lateral"
        
        ws_v3.append([
            NOTE_ID, eid, evt["type"], "Airway", lobe, seg, "", 
            evt["method"], json.dumps(evt["devices"]),
            evt["outcomes"].get("lumen_pre", ""),
            evt["outcomes"].get("lumen_post", ""),
            evt["outcomes"].get("complication", "")
        ])

        # JSON
        json_item = {
            "event_id": eid,
            "type": evt["type"],
            "target": {"lobe": lobe, "segment": seg},
            "method": evt["method"],
            "devices": evt["devices"],
            "outcomes": evt["outcomes"]
        }
        json_procedures.append(json_item)

    # 7. Registry JSON
    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 0: ws_json.append(["json_output"])
    
    final_json = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": json_procedures,
        "no_immediate_complications": True if events_dict["E99"]["outcomes"].get("complication") == "none" else False
    }
    ws_json.append([json.dumps(final_json, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()