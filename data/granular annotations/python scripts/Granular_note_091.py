import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_091"
SOURCE_FILE = "note_091.txt"
PROCEDURE_DATE = "2026-01-12" # Using current date as placeholder or blank if preferred
NOTE_TEXT = """NOTE_ID:  note_091 SOURCE_FILE: note_091.txt INDICATION FOR OPERATION:  [REDACTED]is a 57 year old-year-old male who presents with bronchial stenosis.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31624 Dx bronchoscope/lavage (BAL)    
31625 Endobronchial Biopsy(s)
31630 Balloon dilation
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Disposable Bronchoscope
ESTIMATED BLOOD LOSS:   Minimum
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
Initial Airway Inspection Findings:
The endotracheal tube is in good position.
The visualized portion of the trachea is of normal caliber. The carina is sharp.
The tracheobronchial tree was examined to at least the first subsegmental level.
Airway exam notable for mildly tortuous airways as before, intact anastomoses bilaterally.
The RML orifice was ~75% narrowed with combination of weblike stenosis and pedunculated granulation tissue.
Mild secretions bilaterally but with thick purulent secretions eminating from the RML.
Successful therapeutic aspiration was performed to clean out the Bronchus Intermedius  and RML Carina (RC2) from mucus and mucus plug.
Endobronchial biopsy was performed at RML.  Lesion was successfully removed.  Samples sent for Microbiology (Cultures/Viral/Fungal).
Balloon dilation was performed at RML.  Mustang 4mm balloon was used to perform dilation to 4 mm at the RML.
Total 1 inflations with dilation time of 60 seconds each.
Balloon dilation was performed at RML.
Mustang 6/7/8 balloon was used to perform dilation to 6 and 7 mm at the RML.
Total 2 inflations with dilation time of 60 seconds each.
Bronchial alveolar lavage was performed at RML.
Instilled 60 cc of NS, suction returned with 15 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
Endobronchial obstruction at RML was treated with the following modalities:
Modality	Tools	Setting/Mode	Duration	Results
Electrocautery	 	 	 	 
APC	 	 	 	 
Laser	 	 	 	 
CoreCath	 	 	 	 
Cryoprobe	1.7mm probe	 	30sec freeze thaw cycles	Ablation 
Prior to treatment, affected airway was note to be 25% patent.
After treatment, the airway was 100% patent. 
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
- RML EBBx (micro)
- RML BAL
IMPRESSION/PLAN: [REDACTED]is a 57 year old-year-old male who presents for bronchoscopy for bronchial stenosis of the RML.
Ongoing recurrence of stenosis, currently on stent holiday. Will need ongoing close vigilance of the airway due to tendency to fuse shut.
- f/u BAL results
- repeat bronch in 2 weeks for dilation"""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# -------------------------------------------------------------------------
# CONSTANTS & CONFIG
# -------------------------------------------------------------------------
PROCEDURE_FLAGS = [
    # Bronchoscopy (23)
    "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", 
    "endobronchial_biopsy", "tbna_conventional", "linear_ebus", "radial_ebus", 
    "navigational_bronchoscopy", "transbronchial_biopsy", "transbronchial_cryobiopsy", 
    "therapeutic_aspiration", "foreign_body_removal", "airway_dilation", 
    "airway_stent", "thermal_ablation", "tumor_debulking_non_thermal", 
    "cryotherapy", "blvr", "peripheral_ablation", "bronchial_thermoplasty", 
    "whole_lung_lavage", "rigid_bronchoscopy",
    # Pleural (7)
    "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", 
    "pleurodesis", "pleural_biopsy", "fibrinolytic_therapy"
]

# -------------------------------------------------------------------------
# HELPER FUNCTIONS
# -------------------------------------------------------------------------
def clean_text(text):
    return text.strip().replace('\r', '').replace('\t', ' ')

def get_or_create_sheet(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return wb.create_sheet(sheet_name)

def safe_find_index(full_text, span_text, start_search=0):
    try:
        return full_text.index(span_text, start_search)
    except ValueError:
        return -1

def hydrate_span(full_text, span_text, context_prefix=None, match_index=None):
    # Logic:
    # 1. Exact count == 1 -> use that
    # 2. Context prefix -> search window
    # 3. Match index -> nth occurrence
    
    count = full_text.count(span_text)
    if count == 0:
        return None, None, "not_found"
    
    if count == 1:
        start = full_text.find(span_text)
        return start, start + len(span_text), "hydrated_unique"
    
    # Ambiguous resolution
    if context_prefix:
        # Find all occurrences
        starts = [m.start() for m in re.finditer(re.escape(span_text), full_text)]
        for s in starts:
            # Check preceding 120 chars
            window_start = max(0, s - 120)
            window = full_text[window_start:s]
            if context_prefix in window:
                return s, s + len(span_text), "hydrated_prefix_window"
    
    if match_index is not None and match_index > 0:
        starts = [m.start() for m in re.finditer(re.escape(span_text), full_text)]
        if len(starts) >= match_index:
            s = starts[match_index - 1] # 1-based index in input
            return s, s + len(span_text), "hydrated_match_index"

    return None, None, f"ambiguous_count={count}"

# -------------------------------------------------------------------------
# GENERATION LOGIC
# -------------------------------------------------------------------------
def generate_workbook():
    # 1. Load Template
    if not os.path.exists(TEMPLATE_PATH):
        # Create a dummy workbook if template is missing (for robust standalone execution in some envs)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # 2. Sheet: Note_Text
    ws_text = get_or_create_sheet(wb, "Note_Text")
    if ws_text.max_row == 1:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 3. Sheet: Note_Index
    ws_index = get_or_create_sheet(wb, "Note_Index")
    if ws_index.max_row == 1:
        headers = ["source_file", "note_id", "encounter_id", "procedure_date", "site", 
                   "reviewer", "status", "free_text_notes"] + PROCEDURE_FLAGS
        ws_index.append(headers)
    
    # Calculate Flags
    flags = {f: 0 for f in PROCEDURE_FLAGS}
    
    # Logic for flags based on note content
    txt_lower = NOTE_TEXT.lower()
    
    if "therapeutic aspiration" in txt_lower or "31645" in txt_lower:
        flags["therapeutic_aspiration"] = 1
    if "endobronchial biopsy" in txt_lower or "31625" in txt_lower:
        flags["endobronchial_biopsy"] = 1
    if "bal" in txt_lower or "lavage" in txt_lower:
        flags["bal"] = 1
    if "balloon dilation" in txt_lower or "31630" in txt_lower:
        flags["airway_dilation"] = 1
    if "cryoprobe" in txt_lower or "cryotherapy" in txt_lower:
        flags["cryotherapy"] = 1
    
    # "Destruction of tumor OR relief of stenosis" (31641) + Cryoprobe ablation implies cryotherapy is the method
    # Diagnostic usually implied if therapeutic is done, but 31624 is explicitly listed
    if "31624" in txt_lower or "dx bronchoscope" in txt_lower:
        flags["diagnostic_bronchoscopy"] = 1

    row_data = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "pre-filled", ""]
    row_data += [flags[f] for f in PROCEDURE_FLAGS]
    ws_index.append(row_data)

    # 4. Spans & Events
    # We will construct a list of span dictionaries, then process them
    
    spans_data = []
    
    # Helper to add span
    def add_span(text, label, norm_val, event_id, field=None, prefix=None, index=None):
        spans_data.append({
            "span_text": text,
            "label": label,
            "normalized_value": norm_val,
            "event_id": event_id,
            "schema_field": field,
            "context_prefix": prefix,
            "match_index": index
        })

    # --- EVENT DEFINITIONS ---
    
    # Event 1: Therapeutic Aspiration
    # Successful therapeutic aspiration was performed to clean out the Bronchus Intermedius and RML Carina (RC2) from mucus and mucus plug.
    evt_asp = "evt_01_asp"
    add_span("Therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", evt_asp, "method")
    add_span("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", evt_asp, "target.anatomy_type")
    add_span("RML Carina", "ANAT_AIRWAY", "RML Carina", evt_asp, "target.anatomy_type")
    add_span("mucus", "OBS_ROSE", "Mucus", evt_asp, "findings_json")
    
    # Event 2: Endobronchial Biopsy
    # Endobronchial biopsy was performed at RML. Lesion was successfully removed.
    evt_bx = "evt_02_bx"
    add_span("Endobronchial biopsy", "PROC_METHOD", "Endobronchial Biopsy", evt_bx, "method")
    add_span("RML", "ANAT_LUNG_LOC", "RML", evt_bx, "target.location.lobe", prefix="Endobronchial biopsy was performed at")
    add_span("Lesion", "OBS_LESION", "Lesion", evt_bx, "lesion.type", prefix="Endobronchial biopsy was performed at RML.  ")
    add_span("removed", "PROC_ACTION", "Removed", evt_bx, "action")
    
    # Event 3: Balloon Dilation 1
    # Balloon dilation was performed at RML. Mustang 4mm balloon was used to perform dilation to 4 mm...
    # Total 1 inflations...
    evt_dil1 = "evt_03_dil"
    add_span("Balloon dilation", "PROC_METHOD", "Balloon Dilation", evt_dil1, "method", prefix="")
    add_span("RML", "ANAT_LUNG_LOC", "RML", evt_dil1, "target.location.lobe", prefix="Balloon dilation was performed at")
    add_span("Mustang", "DEV_INSTRUMENT", "Mustang Balloon", evt_dil1, "devices_json")
    add_span("4mm balloon", "DEV_INSTRUMENT", "4mm Balloon", evt_dil1, "devices_json")
    add_span("4 mm", "MEAS_AIRWAY_DIAM", "4 mm", evt_dil1, "measurements_json", prefix="used to perform dilation to")
    add_span("1 inflations", "MEAS_COUNT", "1", evt_dil1, "measurements_json")
    add_span("60 seconds", "MEAS_TIME", "60 s", evt_dil1, "measurements_json", prefix="dilation time of")

    # Event 4: Balloon Dilation 2
    # Mustang 6/7/8 balloon... dilation to 6 and 7 mm...
    # Total 2 inflations...
    evt_dil2 = "evt_04_dil"
    add_span("Mustang 6/7/8 balloon", "DEV_INSTRUMENT", "Mustang 6/7/8 Balloon", evt_dil2, "devices_json")
    add_span("6 and 7 mm", "MEAS_AIRWAY_DIAM", "6, 7 mm", evt_dil2, "measurements_json")
    add_span("2 inflations", "MEAS_COUNT", "2", evt_dil2, "measurements_json")
    add_span("60 seconds", "MEAS_TIME", "60 s", evt_dil2, "measurements_json", prefix="Total 2 inflations with dilation time of")

    # Event 5: BAL
    # Bronchial alveolar lavage was performed at RML.
    evt_bal = "evt_05_bal"
    add_span("Bronchial alveolar lavage", "PROC_METHOD", "BAL", evt_bal, "method")
    add_span("RML", "ANAT_LUNG_LOC", "RML", evt_bal, "target.location.lobe", prefix="Bronchial alveolar lavage was performed at")
    add_span("60 cc", "MEAS_VOL", "60 cc", evt_bal, "measurements_json", prefix="Instilled")
    add_span("15 cc", "MEAS_VOL", "15 cc", evt_bal, "measurements_json", prefix="suction returned with")

    # Event 6: Cryotherapy (Ablation)
    # Endobronchial obstruction at RML... Cryoprobe... Ablation
    evt_cryo = "evt_06_cryo"
    add_span("Endobronchial obstruction", "OBS_LESION", "Obstruction", evt_cryo, "lesion.type")
    add_span("RML", "ANAT_LUNG_LOC", "RML", evt_cryo, "target.location.lobe", prefix="Endobronchial obstruction at")
    add_span("Cryoprobe", "DEV_INSTRUMENT", "Cryoprobe", evt_cryo, "devices_json")
    add_span("1.7mm probe", "DEV_INSTRUMENT", "1.7mm Cryoprobe", evt_cryo, "devices_json")
    add_span("30sec freeze thaw cycles", "PROC_METHOD", "Freeze Thaw Cycles", evt_cryo, "method")
    add_span("Ablation", "PROC_METHOD", "Ablation", evt_cryo, "method")

    # Outcomes & Findings (Event 99 - Global/Context)
    evt_global = "evt_99_global"
    add_span("RML orifice", "ANAT_AIRWAY", "RML Orifice", evt_global, "findings_json")
    add_span("~75% narrowed", "OUTCOME_AIRWAY_LUMEN_PRE", "25% patent", evt_global, "outcomes.airway.lumen_pre") # 75% narrowed = 25% patent
    add_span("weblike stenosis", "OBS_LESION", "Weblike Stenosis", evt_global, "lesion.type")
    add_span("pedunculated granulation tissue", "OBS_LESION", "Granulation Tissue", evt_global, "lesion.type")
    
    # Explicit Pre/Post statements
    add_span("25% patent", "OUTCOME_AIRWAY_LUMEN_PRE", "25%", evt_global, "outcomes.airway.lumen_pre")
    add_span("100% patent", "OUTCOME_AIRWAY_LUMEN_POST", "100%", evt_global, "outcomes.airway.lumen_post")
    
    # Complications
    add_span("No immediate complications", "OUTCOME_COMPLICATION", "None", evt_global, "outcomes.complications")
    
    # Indication
    add_span("bronchial stenosis", "OBS_LESION", "Bronchial Stenosis", evt_global, "indication")

    # -------------------------------------------------------------------------
    # WRITE SPANS & HYDRATE
    # -------------------------------------------------------------------------
    ws_span = get_or_create_sheet(wb, "Span_Annotations")
    ws_hydrated = get_or_create_sheet(wb, "Span_Hydrated")

    if ws_span.max_row == 1:
        headers = ["source_file", "note_id", "span_id", "section_type",
                   "context_prefix", "span_text", "match_index",
                   "start_char", "end_char", "span_len",
                   "label", "normalized_value", "schema_field", "event_id",
                   "is_negated", "is_historical", "time_anchor", "reviewer", "comments",
                   "hydration_status"]
        ws_span.append(headers)
        ws_hydrated.append(headers)

    span_id_counter = 1
    
    for s in spans_data:
        # Generate ID
        sid = f"span_{span_id_counter:03d}"
        span_id_counter += 1
        
        # Hydrate
        start, end, status = hydrate_span(NOTE_TEXT, s["span_text"], s["context_prefix"], s["match_index"])
        
        # Row Base
        row = [
            SOURCE_FILE, NOTE_ID, sid, "procedure_note",
            s["context_prefix"], s["span_text"], s["match_index"],
            "", "", f"=LEN(F{ws_span.max_row+1})", # Formula for length
            s["label"], s["normalized_value"], s["schema_field"], s["event_id"],
            False, False, "procedure_date", "", "",
            "needs_hydration"
        ]
        
        ws_span.append(row)
        
        # Hydrated Row
        h_row = list(row)
        h_row[7] = start
        h_row[8] = end
        h_row[9] = len(s["span_text"]) if start is not None else 0
        h_row[19] = status
        ws_hydrated.append(h_row)

    # -------------------------------------------------------------------------
    # POPULATE EVENTS (Event_Log & V3_Procedure_Events)
    # -------------------------------------------------------------------------
    ws_event = get_or_create_sheet(wb, "Event_Log")
    ws_v3 = get_or_create_sheet(wb, "V3_Procedure_Events")
    
    if ws_event.max_row == 1:
        headers = ["source_file", "note_id", "event_id", "event_type", "method",
                   "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements",
                   "specimens", "findings", "is_historical", "reviewer", "comments",
                   "device_size", "device_material",
                   "outcome_airway_lumen_pre", "outcome_airway_lumen_post",
                   "outcome_symptoms", "outcome_pleural", "outcome_complication"]
        ws_event.append(headers)

    if ws_v3.max_row == 1:
        v3_headers = ["note_id", "event_id", "type",
                      "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station",
                      "lesion.type", "lesion.size_mm",
                      "method", "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote",
                      "stent.size", "stent.material_or_brand", "catheter.size_fr",
                      "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post",
                      "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"]
        ws_v3.append(v3_headers)

    # Define structured event objects for V3
    events_map = {
        "evt_01_asp": {
            "type": "therapeutic", "method": "Therapeutic Aspiration",
            "target.anatomy_type": "Bronchus Intermedius, RML Carina",
            "findings_json": ["mucus", "mucus plug"],
            "evidence_quote": "Successful therapeutic aspiration was performed to clean out the Bronchus Intermedius and RML Carina"
        },
        "evt_02_bx": {
            "type": "diagnostic", "method": "Endobronchial Biopsy",
            "target.location.lobe": "RML",
            "lesion.type": "Lesion",
            "specimens_json": ["Microbiology"],
            "findings_json": ["Lesion removed"],
            "evidence_quote": "Endobronchial biopsy was performed at RML. Lesion was successfully removed."
        },
        "evt_03_dil": {
            "type": "therapeutic", "method": "Balloon Dilation",
            "target.location.lobe": "RML",
            "devices_json": ["Mustang 4mm balloon"],
            "measurements_json": ["4 mm", "1 inflations", "60 seconds"],
            "outcomes.airway.lumen_post": "100%", # Inferred from summary
            "evidence_quote": "Mustang 4mm balloon was used to perform dilation to 4 mm"
        },
        "evt_04_dil": {
            "type": "therapeutic", "method": "Balloon Dilation",
            "target.location.lobe": "RML",
            "devices_json": ["Mustang 6/7/8 balloon"],
            "measurements_json": ["6 mm", "7 mm", "2 inflations", "60 seconds"],
            "outcomes.airway.lumen_post": "100%", # Final outcome
            "evidence_quote": "Mustang 6/7/8 balloon was used to perform dilation to 6 and 7 mm"
        },
        "evt_05_bal": {
            "type": "diagnostic", "method": "BAL",
            "target.location.lobe": "RML",
            "measurements_json": ["Instilled 60 cc", "Returned 15 cc"],
            "specimens_json": ["Cell Count", "Microbiology", "Cytology"],
            "evidence_quote": "Bronchial alveolar lavage was performed at RML."
        },
        "evt_06_cryo": {
            "type": "therapeutic", "method": "Cryotherapy Ablation",
            "target.location.lobe": "RML",
            "lesion.type": "Endobronchial obstruction",
            "devices_json": ["1.7mm Cryoprobe"],
            "measurements_json": ["30sec freeze thaw cycles"],
            "outcomes.airway.lumen_pre": "25%",
            "outcomes.airway.lumen_post": "100%",
            "evidence_quote": "Endobronchial obstruction at RML was treated with... Cryoprobe... Ablation"
        }
    }

    # Write to V3 Sheet
    for eid, e in events_map.items():
        row = [
            NOTE_ID, eid, e.get("type", ""),
            e.get("target.anatomy_type", ""), e.get("target.location.lobe", ""), "", "",
            e.get("lesion.type", ""), "",
            e.get("method", ""),
            json.dumps(e.get("devices_json", [])),
            json.dumps(e.get("measurements_json", [])),
            json.dumps(e.get("specimens_json", [])),
            json.dumps(e.get("findings_json", [])),
            e.get("evidence_quote", ""),
            "", "", "",
            e.get("outcomes.airway.lumen_pre", ""),
            e.get("outcomes.airway.lumen_post", ""),
            "", "", e.get("outcomes.complications", "")
        ]
        ws_v3.append(row)
        
        # Simplified Event Log
        log_row = [
            SOURCE_FILE, NOTE_ID, eid, e.get("type", ""), e.get("method", ""),
            f"{e.get('target.location.lobe','')} {e.get('target.anatomy_type','')}".strip(),
            json.dumps(e.get("devices_json", [])),
            "", "", "", json.dumps(e.get("measurements_json", [])),
            json.dumps(e.get("specimens_json", [])),
            json.dumps(e.get("findings_json", [])),
            False, "", "", "", "",
            e.get("outcomes.airway.lumen_pre", ""),
            e.get("outcomes.airway.lumen_post", ""),
            "", "", e.get("outcomes.complications", "")
        ]
        ws_event.append(log_row)

    # 5. Registry JSON
    ws_json = get_or_create_sheet(wb, "V3_Registry_JSON")
    if ws_json.max_row == 1:
        ws_json.append(["json_output"])
    
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": list(events_map.values()),
        "no_immediate_complications": True
    }
    ws_json.append([json.dumps(registry_data, indent=2)])

    # 6. Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()