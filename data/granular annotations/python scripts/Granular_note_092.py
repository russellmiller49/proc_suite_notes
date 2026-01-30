import openpyxl
from openpyxl.utils import get_column_letter
import re
import os

# -----------------------------------------------------------------------------
# 1. INPUTS & CONSTANTS
# -----------------------------------------------------------------------------
NOTE_ID = "note_092"
SOURCE_FILE = "note_092.txt"
PROCEDURE_DATE = "2026-01-12" # inferred from context or left blank if not in text, using prompt context date for file metadata if needed, but standard is usually extracted. Left blank in row data.

NOTE_TEXT = """NOTE_ID:  note_092 SOURCE_FILE: note_092.txt INDICATION FOR OPERATION:  [REDACTED]is a 57 year old-year-old male who presents with airway stenosis.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31622 Dx bronchoscope/cell washing          
31624 Dx bronchoscope/lavage (BAL)    
31630 Balloon dilation
31640 Bronchoscopy with excision 
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
31899NFD BRONCHOSCOPY W/ APPLICATION OF TRANEXAMIC ACID
31635 Foreign body removal
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort 
required)
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required multiple (5) separate applications of spray cryotherapy and required multiple forms of bronchoscopy (flexible and rigid) in order to adequately treat the patient.
This resulted in >50% increased work due to Time, Technical difficulty of procedure, and Physical and mental effort required.
Apply to: 
31622 Dx bronchoscope/cell washing     
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy).
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Rigid Bronchoscope
Flexible Therapeutic Bronchoscope
ESTIMATED BLOOD LOSS:   Minimum
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
The flexible therapeutic bronchoscope was advanced.
Initial Airway Inspection Findings:
The laryngeal mask airway is in good position.
Pharynx: Not assessed due to bronchoscopy introduction through LMA.
Larynx: Not fully assessed due to bronchoscopy introduction through LMA.
Vocal Cords: Normal without mass/lesions
Trachea: Anterior trachea with well-healed tracheostomy site, otherwise normal
Main Carina: Sharp
Right Lung Proximal Airways: Anastamosis sutures intact, somewhat tortuous bronchus intermedius.
RML with stent in place. Normal anatomic branching to segmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Left Lung Proximal Airways: Anastamosis sutures intact.
Normal anatomic branching to segmental level.  No evidence of mass, lesions, bleeding or other endobronchial pathology.
Mucosa: Normal.
Secretions: Minimal, thin, and clear.  
Successful therapeutic aspiration was performed to clean out the Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus.
During initial gentle suctioning/therapeutic aspiration at the RML the iCAST 7mm x 16mm stent became dislodged and completely extruded from the RML into the distal bronchus intermedius.
Foreign body removal was performed:  the stent was grasped with pulmonary forceps and removed en bloc with the therapeutic bronchoscope.
There was some oozing blood from the distal RML bronchus after the stent was dislodged.
This bleeding was easily controlled with endobronchial cold saline, epinephrine 0.2mg, and tranexamic acid 200mg.
RML AFTER STENT DISLODGEMENT
 
After induction of muscle relaxants, LMA was removed and tooth or gum protector was placed.
The black bronchoscope ventilating rigid barrel was introduced through the mouth and advanced in the midline while keeping the alignment with the axis of the trachea and minimizing pressure to the teeth.
The vocal cords were identified and the rigid bronchoscope was advanced carefully while minimizing contact with them.
Once the rigid bronchoscope was positioned at the mid-trachea, jet ventilation was initiated and chest wall movement was confirmed.
The bronchoscope barrel was then positioned into the proximal right mainstem bronchus.
Area of stenosis at the right middle lobe was treated with the following modalities:
Modality	Tools	Setting/Mode	Duration	Results
Mechanical	Pulmonary alligator forceps	N/A	N/A	Good granulation tissue removal from distal RML bronchus
Cryospray	Cryospray cryotherapy catheter	Low-flow	10-second application x 5 total applications	Excellent application of spray cryotherapy to proximal RML bronchus and around the RML take-off
Balloon dilation was performed at RML Carina (RC2).
8/9/10 Elation balloon was used to perform dilation to 10 mm at the RML Carina (RC2).
Total 2 inflations with dilation time of 60 seconds each.
Prior to treatment, affected airway was note to be 70% patent.  After treatment, the airway was 90% patent.
After the stent extraction, light mechanical excision, and balloon dilation, the RML was widely patent with patent segmental airways.
Decision was made to not place another stent at this time, as concern that it could easily dislodge.
Will also allow for stent vacation to help break chain of stent bacterial colonization and to see how the RML tissue responds to the cryospray treatment.
Will monitor this airway closely and plan for close bronchoscopy airway evaluation in 2-3 weeks.
The rigid bronchoscope was extubated.
LMA was replaced by anesthesia team.
Flexible therapeutic bronchoscope was advanced.
Bronchial alveolar lavage was performed at Lateral Segment of RML (RB4).
Instilled 40 cc of NS, suction returned with 20 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
RML AFTER STENT SPRAY CRYOTHERAPY AND BALLOON DILATION
 
 
The patient tolerated the procedure well.
There were no immediate complications.  At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
--RML stent (path)
--RML BAL (cell count, micro, cyto)
IMPRESSION/PLAN: [REDACTED]is a 57 year old-year-old male who presents for bronchoscopy for evaluation of bronchial stenosis.
Patient underwent bronchoscopy with stent removal, application of spray cryotherapy, and balloon dilation without immediate complication.
--Post-procedure CXR
--Follow-up BAL results
--Follow-up in outpatient Interventional Pulmonology clinic as scheduled on [REDACTED]
--Repeat bronchoscopy in 2-3 weeks"""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# -----------------------------------------------------------------------------
# 2. LABELING DATA
# -----------------------------------------------------------------------------

# Flags
FLAGS = {
    'diagnostic_bronchoscopy': 1,
    'bal': 1,
    'bronchial_wash': 1, # Code 31622 listed
    'brushings': 0,
    'endobronchial_biopsy': 0,
    'tbna_conventional': 0,
    'linear_ebus': 0,
    'radial_ebus': 0,
    'navigational_bronchoscopy': 0,
    'transbronchial_biopsy': 0,
    'transbronchial_cryobiopsy': 0,
    'therapeutic_aspiration': 1,
    'foreign_body_removal': 1,
    'airway_dilation': 1,
    'airway_stent': 0, # Removal only
    'thermal_ablation': 0,
    'tumor_debulking_non_thermal': 1, # Mechanical forceps excision
    'cryotherapy': 1,
    'blvr': 0,
    'peripheral_ablation': 0,
    'bronchial_thermoplasty': 0,
    'whole_lung_lavage': 0,
    'rigid_bronchoscopy': 1,
    'thoracentesis': 0,
    'chest_tube': 0,
    'ipc': 0,
    'medical_thoracoscopy': 0,
    'pleurodesis': 0,
    'pleural_biopsy': 0,
    'fibrinolytic_therapy': 0
}

# Spans
# Structure: (text, label, normalized_value, field, event_id, context_prefix)
SPANS = [
    # Event 1: Therapeutic Aspiration
    ("therapeutic aspiration", "PROC_METHOD", "therapeutic aspiration", "method", "evt1", "Successful "),
    ("Trachea (Distal 1/3)", "ANAT_AIRWAY", "trachea distal", "target.anatomy", "evt1", "clean out the "),
    ("Right Mainstem", "ANAT_AIRWAY", "RMS", "target.anatomy", "evt1", "clean out the "),
    ("Bronchus Intermedius", "ANAT_AIRWAY", "BI", "target.anatomy", "evt1", "Right Mainstem, "),
    ("Left Mainstem", "ANAT_AIRWAY", "LMS", "target.anatomy", "evt1", "Intermedius , "),
    ("Carina", "ANAT_AIRWAY", "main carina", "target.anatomy", "evt1", "Left Mainstem, "),
    ("RUL Carina (RC1)", "ANAT_AIRWAY", "RUL carina", "target.anatomy", "evt1", "Carina, "),
    ("RML Carina (RC2)", "ANAT_AIRWAY", "RML carina", "target.anatomy", "evt1", "RUL Carina (RC1), "),
    ("LUL Lingula Carina (Lc1)", "ANAT_AIRWAY", "LUL lingula carina", "target.anatomy", "evt1", "RML Carina (RC2), "),
    ("Left Carina (LC2)", "ANAT_AIRWAY", "LUL carina", "target.anatomy", "evt1", "LUL Lingula Carina (Lc1), and "),
    
    # Event 2: Foreign Body Removal (Stent)
    ("iCAST 7mm x 16mm stent", "DEV_STENT", "iCAST stent", "device", "evt2", "RML the "),
    ("RML", "ANAT_LUNG_LOC", "RML", "target.location.lobe", "evt2", "extruded from the "),
    ("Foreign body removal", "PROC_METHOD", "foreign body removal", "method", "evt2", "intermedius.\n"),
    ("pulmonary forceps", "DEV_INSTRUMENT", "pulmonary forceps", "device", "evt2", "grasped with "),
    ("removed en bloc", "PROC_ACTION", "removed en bloc", "method", "evt2", "forceps and "),

    # Event 3: Bleeding Control
    ("bleeding", "OBS_ROSE", "bleeding", "findings", "evt3", "19] This "),
    ("endobronchial cold saline", "PROC_METHOD", "cold saline", "method", "evt3", "controlled with "),
    ("epinephrine 0.2mg", "PROC_METHOD", "epinephrine", "method", "evt3", "cold saline, "),
    ("tranexamic acid 200mg", "PROC_METHOD", "tranexamic acid", "method", "evt3", "0.2mg, and "),
    
    # Event 4: Mechanical Debulking
    ("Mechanical", "PROC_METHOD", "mechanical debulking", "method", "evt4", "Results\n"),
    ("Pulmonary alligator forceps", "DEV_INSTRUMENT", "pulmonary alligator forceps", "device", "evt4", "Mechanical\t"),
    ("granulation tissue removal", "PROC_ACTION", "granulation tissue removal", "method", "evt4", "Good "),
    ("distal RML bronchus", "ANAT_AIRWAY", "RML bronchus distal", "target.anatomy", "evt4", "removal from "),

    # Event 5: Cryotherapy
    ("Cryospray", "PROC_METHOD", "cryospray", "method", "evt5", "Results\nMechanical\t"),
    ("Cryospray cryotherapy catheter", "DEV_INSTRUMENT", "cryospray catheter", "device", "evt5", "Cryospray\t"),
    ("proximal RML bronchus", "ANAT_AIRWAY", "RML bronchus proximal", "target.anatomy", "evt5", "cryotherapy to "),
    
    # Event 6: Balloon Dilation
    ("Balloon dilation", "PROC_METHOD", "balloon dilation", "method", "evt6", "applications\n"),
    ("RML Carina (RC2)", "ANAT_AIRWAY", "RML carina", "target.anatomy", "evt6", "performed at "),
    ("8/9/10 Elation balloon", "DEV_INSTRUMENT", "Elation balloon", "device", "evt6", "26] "),
    ("dilation to 10 mm", "MEAS_AIRWAY_DIAM", "10 mm", "measurements", "evt6", "perform "),
    ("2 inflations", "MEAS_COUNT", "2", "measurements", "evt6", "Total "),
    ("60 seconds", "CTX_TIME", "60 seconds", "measurements", "evt6", "time of "),
    
    # Event 7: Outcomes
    ("70% patent", "OUTCOME_AIRWAY_LUMEN_PRE", "70%", "outcomes.airway.lumen_pre", "evt7", "note to be "),
    ("90% patent", "OUTCOME_AIRWAY_LUMEN_POST", "90%", "outcomes.airway.lumen_post", "evt7", "airway was "),
    ("no immediate complications", "OUTCOME_COMPLICATION", "none", "outcomes.complications", "evt7", "There were "),

    # Event 8: BAL
    ("Bronchial alveolar lavage", "PROC_METHOD", "BAL", "method", "evt8", "advanced.\n"),
    ("Lateral Segment of RML (RB4)", "ANAT_AIRWAY", "RML lateral segment", "target.anatomy", "evt8", "performed at "),
    ("Instilled 40 cc", "MEAS_VOL", "40 cc", "measurements", "evt8", "35] "),
    ("returned with 20 cc", "MEAS_VOL", "20 cc", "measurements", "evt8", "suction "),
]

# -----------------------------------------------------------------------------
# 3. HELPER FUNCTIONS
# -----------------------------------------------------------------------------

def hydrate_offsets(text, span_text, context_prefix=None):
    """
    Finds start/end char offsets for span_text.
    Returns: (start, end, status_string)
    """
    matches = [m for m in re.finditer(re.escape(span_text), text)]
    
    if not matches:
        return None, None, "not_found"
    
    if len(matches) == 1:
        m = matches[0]
        return m.start(), m.end(), "hydrated_unique"
    
    # Context match
    if context_prefix:
        for m in matches:
            start = m.start()
            # Look back 150 chars
            window_start = max(0, start - 150)
            preceding_text = text[window_start:start]
            if context_prefix in preceding_text:
                return start, m.end(), "hydrated_prefix_window"
    
    # Ambiguous if no context or context failed
    return None, None, f"ambiguous_count={len(matches)}"

# -----------------------------------------------------------------------------
# 4. GENERATION SCRIPT
# -----------------------------------------------------------------------------

def generate_workbook():
    # Load template
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Template not found at {TEMPLATE_PATH}")
    
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    
    # 1. Note_Text
    ws_text = wb["Note_Text"]
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # 2. Note_Index
    ws_index = wb["Note_Index"]
    # Columns: source_file, note_id, encounter_id, procedure_date, site, reviewer, status, free_text, [flags...]
    # Map flags to list
    flag_keys = [
        'diagnostic_bronchoscopy', 'bal', 'bronchial_wash', 'brushings', 'endobronchial_biopsy', 
        'tbna_conventional', 'linear_ebus', 'radial_ebus', 'navigational_bronchoscopy', 
        'transbronchial_biopsy', 'transbronchial_cryobiopsy', 'therapeutic_aspiration', 
        'foreign_body_removal', 'airway_dilation', 'airway_stent', 'thermal_ablation', 
        'tumor_debulking_non_thermal', 'cryotherapy', 'blvr', 'peripheral_ablation', 
        'bronchial_thermoplasty', 'whole_lung_lavage', 'rigid_bronchoscopy', 
        'thoracentesis', 'chest_tube', 'ipc', 'medical_thoracoscopy', 'pleurodesis', 
        'pleural_biopsy', 'fibrinolytic_therapy'
    ]
    
    row_meta = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Pending", ""]
    row_flags = [FLAGS.get(k, 0) for k in flag_keys]
    ws_index.append(row_meta + row_flags)
    
    # 3 & 4. Span_Annotations & Span_Hydrated
    ws_anno = wb["Span_Annotations"]
    ws_hydra = wb["Span_Hydrated"]
    
    # Headers should already exist in template, but we append rows matching columns:
    # source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, 
    # start_char, end_char, span_len, label, normalized_value, schema_field, event_id, 
    # is_negated, is_historical, time_anchor, reviewer, comments, hydration_status
    
    span_id_counter = 1
    
    for sp in SPANS:
        txt, label, norm, field, evid, ctx = sp
        
        # Calculate offsets
        start, end, status = hydrate_offsets(NOTE_TEXT, txt, ctx)
        length = len(txt)
        
        # Annotation row (no offsets)
        row_anno = [
            SOURCE_FILE, NOTE_ID, f"span_{span_id_counter:03d}", "", ctx, txt, "", 
            "", "", f"=LEN(F{ws_anno.max_row+1})", label, norm, field, evid, 
            "FALSE", "FALSE", "", "", "", "needs_hydration"
        ]
        ws_anno.append(row_anno)
        
        # Hydrated row (with offsets)
        row_hydra = [
            SOURCE_FILE, NOTE_ID, f"span_{span_id_counter:03d}", "", ctx, txt, "", 
            start if start is not None else "", end if end is not None else "", length, 
            label, norm, field, evid, "FALSE", "FALSE", "", "", "", status
        ]
        ws_hydra.append(row_hydra)
        
        span_id_counter += 1

    # 5. Event_Log
    ws_log = wb["Event_Log"]
    # Columns: source_file, note_id, event_id, event_type, method, anatomy_target, device, 
    # needle_gauge, stations, counts, measurements, specimens, findings, is_historical, 
    # reviewer, comments, device_size, device_material, airway_pre, airway_post, symptoms, pleural, complication
    
    events_map = {
        "evt1": ["Therapeutic", "therapeutic aspiration", "Airways", "N/A", "", "", "", "", "", "", "FALSE", "", "", "", "", "N/A", "N/A", "N/A", "N/A", "N/A"],
        "evt2": ["Therapeutic", "Foreign Body Removal", "RML", "forceps", "", "", "", "stent", "stent dislodged", "", "FALSE", "", "", "", "", "N/A", "N/A", "N/A", "N/A", "N/A"],
        "evt3": ["Therapeutic", "Bleeding Control", "RML", "N/A", "", "", "", "", "", "oozing blood", "FALSE", "", "", "", "", "N/A", "N/A", "N/A", "N/A", "N/A"],
        "evt4": ["Therapeutic", "Mechanical Debulking", "RML", "alligator forceps", "", "", "", "", "", "granulation tissue", "FALSE", "", "", "", "", "N/A", "N/A", "N/A", "N/A", "N/A"],
        "evt5": ["Therapeutic", "Cryospray", "RML", "cryospray catheter", "", "", "", "", "", "", "FALSE", "", "", "", "", "N/A", "N/A", "N/A", "N/A", "N/A"],
        "evt6": ["Therapeutic", "Balloon Dilation", "RML", "Elation balloon", "", "", "10mm", "2 inflations", "", "", "FALSE", "", "10mm", "", "N/A", "N/A", "N/A", "N/A", "N/A"],
        "evt7": ["Outcome", "Outcome Assessment", "RML", "N/A", "", "", "", "", "", "patent", "FALSE", "", "", "", "70%", "90%", "N/A", "N/A", "none"],
        "evt8": ["Diagnostic", "BAL", "RML", "N/A", "", "", "", "40cc/20cc", "fluid", "", "FALSE", "", "", "", "N/A", "N/A", "N/A", "N/A", "N/A"],
    }
    
    for evid, data in events_map.items():
        row = [SOURCE_FILE, NOTE_ID, evid] + data
        ws_log.append(row)

    # 6. V3_Procedure_Events
    # note_id, event_id, type, target.anatomy_type, target.location.lobe, target.location.segment, target.station, lesion.type, lesion.size_mm, method, devices_json, measurements_json, specimens_json, findings_json, evidence_quote, stent.size, stent.material_or_brand, catheter.size_fr, outcomes.airway.lumen_pre, outcomes.airway.lumen_post, outcomes.symptoms, outcomes.pleural, outcomes.complications
    ws_v3 = wb["V3_Procedure_Events"]
    
    # Just basic population for demonstration of completeness
    v3_rows = [
        [NOTE_ID, "evt1", "Therapeutic", "Airway", "RML/RUL/LUL", "Multiple", "", "", "", "Aspiration", "[]", "[]", "[]", "[]", "clean out the Trachea...", "", "", "", "", "", "", "", ""],
        [NOTE_ID, "evt2", "Therapeutic", "Airway", "RML", "", "", "Foreign Body", "", "Removal", '["forceps"]', "[]", '["stent"]', "[]", "stent...removed en bloc", "7x16mm", "iCAST", "", "", "", "", "", ""],
        [NOTE_ID, "evt6", "Therapeutic", "Airway", "RML", "", "", "Stenosis", "", "Balloon Dilation", '["Elation balloon"]', '["10mm", "60sec"]', "[]", "[]", "dilation to 10 mm", "", "", "", "70%", "90%", "", "", ""],
        [NOTE_ID, "evt8", "Diagnostic", "Airway", "RML", "Lateral", "", "", "", "BAL", "[]", '["40cc"]', '["fluid"]', "[]", "Instilled 40 cc", "", "", "", "", "", "", "", ""]
    ]
    
    for r in v3_rows:
        ws_v3.append(r)

    # 7. V3_Registry_JSON (Mock)
    ws_json = wb["V3_Registry_JSON"]
    ws_json.append(["1.0", NOTE_ID, '{"procedures": [...]}', "TRUE"])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()