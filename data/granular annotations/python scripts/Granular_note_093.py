import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_093"
SOURCE_FILE = "note_093.txt"
NOTE_TEXT = """NOTE_ID:  note_093 SOURCE_FILE: note_093.txt INDICATION FOR OPERATION:  [REDACTED]is a 57 year old-year-old male who presents with airway stenosis.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31624 Dx bronchoscope/lavage (BAL)    
31625 Endobronchial Biopsy(s)
31629 TBNA single lobe   
31652 EBUS sampling 1 or 2 nodes
31630 Balloon dilation
31636 Dilate and bronchial stent initial bronchus
31640 Bronchoscopy with excision 
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
31635 Foreign body removal
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental 
effort required)
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required multiple modalities and increased effort to access and open the RML.
This resulted in >100% increased work due to Increased intensity, Time, Technical difficulty of procedure, and Physical and mental effort required.
Apply to: 31629 TBNA single lobe   
31652 EBUS sampling 1 or 2 nodes
31630 Balloon dilation
31640 Bronchoscopy with excision .
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
Flexible Hybrid (Pedatric) Bronchoscope
Linear EBUS 
ESTIMATED BLOOD LOSS:   Moderate
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
Initial Airway Inspection Findings:
  
The laryngeal mask airway is in good position.
Pharynx: Not assessed due to bronchoscopy introduction through LMA.
Larynx: Not fully assessed due to bronchoscopy introduction through LMA.
Vocal Cords: Normal without mass/lesions
Trachea: Distal 1/3 normal.
Main Carina: Sharp
Right Lung Proximal Airways: RML orifice unable to be identified on initial inspection since the airway had swallowed up the prior RML stent and this was now fused shut.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Left Lung Proximal Airways: Normal anatomic branching to segmental level.
No evidence of mass, lesions, bleeding or other endobronchial pathology.
Mucosa: Normal.
Secretions: Minimal, thin, and clear.
Successful therapeutic aspiration was performed to clean out the Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus.
VIEW OF OBLITERATED RML AND PATENT RLL FROM BRONCHUS INTERMEDIUS
 
PRIOR VIEW OF RML AND RLL FROM PRIOR BRONCHOSCOPY [REDACTED]
 
PRIOR VIEW OF RML AND RLL FROM PRIOR BRONCHOSCOPY [REDACTED] 
The RML was unable to be identified and there was suspicion that the RML stent migrated distally with overgrowth of tissue, effectively closing off the RML.
The EBUS scope was utilized to identify the RML orifice and distal stent.
Transbronchial needle aspiration was utilized to re-create/re-open the RML orifice. During this maneuver, frank pus was aspirated from the RML.
RML stent with needle within the airway:
 
Once an pinhole opening was made with above TBNA, pulmonary forceps were used to probe the RML and blunt dissect to widen the opening.
Endobronchial biopsy was performed at RML.  Lesion was successfully removed.  Samples sent for Pathology.
A 4Fr Fogarty balloon was inserted to further open the RML.  
Purulent secretions distal to the obstruction were aspirated.
The hybrid scope was used to explore the distal RML and the medial and lateral subsegments were successfully identified.
Bronchial alveolar lavage was performed at Lateral Segment of RML (RB4) and Medial Segment of RML (RB5).
Instilled 40 cc of NS, suction returned with 25 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
Balloon dilation was performed at RML Carina (RC2).
6/7/8 Elation balloon was used to perform dilation to 8 mm at the RML Carina (RC2).
Total 3 inflations with dilation time of 60 seconds each.
After above excision and dilation, the RML was finally patent and the stent could be visualized.
VIEW OF RML AFTER RECANNALIZATION
 
VIEW OF DISTAL RML AFTER RECANNALIZATION
 
The foreign body/old stent (iCAST 7mm x 16mm) was identified and grasped with the pulmonary rat tooth forceps.
This was removed en bloc with the bronchoscope.
Bleeding/oozing was noted from the RML and this was treated with multiple modalities including cold saline, TXA, epinephrine, and electrocautery/coagulation.
Endobronchial debris and hemorrhage at RML was treated with the following modalities:
Modality	Tools	Setting/Mode	Duration	Results
Electrocautery	Probe 	Forced coag	2-5sec	Ablation/Coagulation
 
In effort to prevent stent from getting swallowed up again, a longer stent was selected.
Next, an iCAST 7mm x 22mm stent was deployed into the RML however this blocked off the RLL and was subsequently removed en bloc with the bronchoscope using the rat tooth forceps.
An iCAST 7mm x 16mm stent was placed in the RML but was noted to be too proximal and was subsequently removed en bloc with the bronchoscope using the rat tooth forceps.
Another iCAST 7mm x 16mm stent was placed in the RML and seated appropriately.
Balloon dilation was performed at RML Carina (RC2) through the stent to seat the stent.
6/7/8 Elation balloon was used to perform dilation to 8 mm at the RML Carina (RC2).
Total 1 inflations with dilation time of 30 seconds.
RML stent at end of case:
 
The patient tolerated the procedure well.
There were no immediate complications.  At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
--iCAST 7mm x 16mm stent (path)
--RML BAL (cell count, micro, cyto)
IMPRESSION/PLAN: [REDACTED]is a 57 year old-year-old male who presents for bronchoscopy for evaluation of airway stenosis.
The patient's RML was closed off around his iCAST stent. The RML was reopened and the stent was removed.
A new iCAST stent (7mm x 16mm) was placed. Patient tolerated the procedure and there were no immediate complications.
--Follow-up path and BAL results
--Repeat bronchoscopy in 2 weeks for re-evaluation
--Augmentin (875mg-125mg) 1 tab by mouth twice daily for 7 days
"""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# -------------------------------------------------------------------------
# CONFIGURATION
# -------------------------------------------------------------------------
PROCEDURE_FLAGS = [
    "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
    "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy",
    "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration",
    "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
    "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation",
    "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
    "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis",
    "pleural_biopsy", "fibrinolytic_therapy"
]

# Set flags based on note content (manually derived from analysis)
FLAGS = {flag: 0 for flag in PROCEDURE_FLAGS}
FLAGS["diagnostic_bronchoscopy"] = 1
FLAGS["therapeutic_aspiration"] = 1 # Source 14
FLAGS["linear_ebus"] = 1 # Source 8, 16
FLAGS["tbna_conventional"] = 1 # Source 17 (31629 listed)
FLAGS["endobronchial_biopsy"] = 1 # Source 19
FLAGS["airway_dilation"] = 1 # Source 25, 34
FLAGS["bal"] = 1 # Source 22
FLAGS["foreign_body_removal"] = 1 # Source 28
FLAGS["thermal_ablation"] = 1 # Source 30 (Electrocautery)
FLAGS["airway_stent"] = 1 # Source 33
FLAGS["bronchoscopy_with_excision"] = 1 # 31640 - mapped to debulking usually, but flags are strict.
# Note: 'bronchoscopy_with_excision' is not in the 30 flags list, 'tumor_debulking_non_thermal' is closest but electocautery was thermal.
# We stick to the provided list.

# -------------------------------------------------------------------------
# SPAN DATA (HARDCODED EXTRACTION)
# -------------------------------------------------------------------------
# Format: (span_text, label, normalized_value, context_prefix, schema_field, event_id)
# event_id: links to Event Log.
# schema_field: typically matches JSON structure like 'target.location' or 'method'

SPANS = [
    # Event 1: Therapeutic Aspiration
    ("Therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", None, "method", "ev1"),
    ("clean out the Trachea", "PROC_ACTION", "clean airway", None, "action", "ev1"),
    ("Trachea (Distal 1/3)", "ANAT_AIRWAY", "Trachea", None, "target.anatomy", "ev1"),
    
    # Event 2: EBUS
    ("EBUS scope", "DEV_INSTRUMENT", "EBUS Scope", None, "device", "ev2"),
    ("identify the RML orifice", "PROC_ACTION", "identify airway", None, "action", "ev2"),
    
    # Event 3: TBNA (Recanalization)
    ("Transbronchial needle aspiration", "PROC_METHOD", "TBNA", None, "method", "ev3"),
    ("re-create/re-open the RML orifice", "PROC_ACTION", "recanalize", None, "action", "ev3"),
    ("RML", "ANAT_AIRWAY", "RML", "frank pus was aspirated from the ", "target.anatomy", "ev3"),
    ("frank pus", "OBS_LESION", "pus", None, "findings", "ev3"),
    
    # Event 4: Endobronchial Biopsy (Removal)
    ("Endobronchial biopsy", "PROC_METHOD", "Endobronchial Biopsy", None, "method", "ev4"),
    ("RML", "ANAT_AIRWAY", "RML", "Endobronchial biopsy was performed at ", "target.anatomy", "ev4"),
    ("Lesion", "OBS_LESION", "Lesion", None, "lesion", "ev4"),
    ("successfully removed", "PROC_ACTION", "Remove", "Lesion was ", "action", "ev4"),
    
    # Event 5: Balloon Dilation (Fogarty)
    ("4Fr Fogarty balloon", "DEV_INSTRUMENT", "Fogarty Balloon", None, "device", "ev5"),
    ("open the RML", "PROC_ACTION", "open", "inserted to further ", "action", "ev5"),
    
    # Event 6: BAL
    ("Bronchial alveolar lavage", "PROC_METHOD", "BAL", None, "method", "ev6"),
    ("Lateral Segment of RML (RB4)", "ANAT_AIRWAY", "RML Lateral", None, "target.anatomy", "ev6"),
    ("Medial Segment of RML (RB5)", "ANAT_AIRWAY", "RML Medial", None, "target.anatomy", "ev6"),
    
    # Event 7: Balloon Dilation (Elation)
    ("Balloon dilation", "PROC_METHOD", "Balloon Dilation", "Cytology.\n", "method", "ev7"),
    ("RML Carina (RC2)", "ANAT_AIRWAY", "RML Carina", "dilation was performed at ", "target.anatomy", "ev7"),
    ("6/7/8 Elation balloon", "DEV_INSTRUMENT", "Elation Balloon", None, "device", "ev7"),
    ("8 mm", "MEAS_AIRWAY_DIAM", "8 mm", "dilation to ", "measurements", "ev7"),
    ("3 inflations", "MEAS_COUNT", "3", None, "measurements", "ev7"),
    ("60 seconds", "CTX_TIME", "60s", None, "measurements", "ev7"),
    
    # Event 8: Foreign Body Removal (Old Stent)
    ("foreign body/old stent", "OBS_LESION", "Old Stent", None, "findings", "ev8"),
    ("iCAST 7mm x 16mm", "DEV_STENT", "iCAST 7x16mm", "old stent (", "device", "ev8"),
    ("pulmonary rat tooth forceps", "DEV_INSTRUMENT", "Rat Tooth Forceps", None, "device", "ev8"),
    ("removed en bloc", "PROC_ACTION", "Remove", "This was ", "action", "ev8"),
    ("RML", "ANAT_AIRWAY", "RML", "Bleeding/oozing was noted from the ", "target.anatomy", "ev8"),

    # Event 9: Electrocautery
    ("Electrocautery", "PROC_METHOD", "Electrocautery", "modality ... ", "method", "ev9"),
    ("hemorrhage", "OBS_LESION", "Hemorrhage", None, "findings", "ev9"),
    ("RML", "ANAT_AIRWAY", "RML", "debris and hemorrhage at ", "target.anatomy", "ev9"),
    ("Probe", "DEV_INSTRUMENT", "Probe", None, "device", "ev9"),
    ("Forced coag", "PROC_METHOD", "Forced Coagulation", None, "settings", "ev9"),

    # Event 10: Stent Placement (Attempt 1 - Removed)
    ("iCAST 7mm x 22mm stent", "DEV_STENT", "iCAST 7x22mm", None, "device", "ev10"),
    ("RML", "ANAT_AIRWAY", "RML", "deployed into the ", "target.anatomy", "ev10"),
    ("removed en bloc", "PROC_ACTION", "Remove", "subsequently ", "action", "ev10"),
    
    # Event 11: Stent Placement (Attempt 2 - Removed)
    ("iCAST 7mm x 16mm stent", "DEV_STENT", "iCAST 7x16mm", "Next, an ", "device", "ev11"),
    ("RML", "ANAT_AIRWAY", "RML", "placed in the ", "target.anatomy", "ev11"),
    ("too proximal", "OBS_LESION", "Malposition", None, "findings", "ev11"),
    ("removed en bloc", "PROC_ACTION", "Remove", "proximal and was subsequently ", "action", "ev11"),

    # Event 12: Stent Placement (Attempt 3 - Successful)
    ("Another iCAST 7mm x 16mm stent", "DEV_STENT", "iCAST 7x16mm", None, "device", "ev12"),
    ("placed in the RML", "PROC_ACTION", "Place", None, "action", "ev12"),
    ("seated appropriately", "OUTCOME_AIRWAY_LUMEN_POST", "Patent", None, "outcome", "ev12"),
    ("7mm", "DEV_STENT_SIZE", "7mm", "Another iCAST ", "device.size", "ev12"),
    ("iCAST", "DEV_STENT_MATERIAL", "iCAST", "Another ", "device.material", "ev12"),

    # Event 13: Balloon Dilation (Seating)
    ("Balloon dilation", "PROC_METHOD", "Balloon Dilation", "appropriately.\n", "method", "ev13"),
    ("RML Carina (RC2)", "ANAT_AIRWAY", "RML Carina", "performed at ", "target.anatomy", "ev13"),
    ("through the stent", "PROC_METHOD", "Through Stent", None, "method", "ev13"),
    ("6/7/8 Elation balloon", "DEV_INSTRUMENT", "Elation Balloon", "stent.\n", "device", "ev13"),
    ("8 mm", "MEAS_AIRWAY_DIAM", "8 mm", "perform dilation to ", "measurements", "ev13"),
    ("1 inflations", "MEAS_COUNT", "1", None, "measurements", "ev13"),
    ("30 seconds", "CTX_TIME", "30s", None, "measurements", "ev13"),

    # Global/Outcomes
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", None, "outcome", "ev12"), # Associate with main therapeutic event
    ("RML was closed off", "OUTCOME_AIRWAY_LUMEN_PRE", "Occluded", None, "outcome", "ev12"),
]

# -------------------------------------------------------------------------
# EVENT LOG GENERATION
# -------------------------------------------------------------------------
EVENTS = [
    {
        "event_id": "ev1", "event_type": "Therapeutic Aspiration", "method": "Suction", 
        "anatomy_target": "Trachea/Mainstems", "findings": "Mucus"
    },
    {
        "event_id": "ev2", "event_type": "Diagnostic Bronchoscopy", "method": "EBUS", 
        "anatomy_target": "RML", "findings": "Orifice Identification"
    },
    {
        "event_id": "ev3", "event_type": "TBNA", "method": "Needle Aspiration", 
        "anatomy_target": "RML", "findings": "Pus/Stenosis", "notes": "Used for recanalization"
    },
    {
        "event_id": "ev4", "event_type": "Endobronchial Biopsy", "method": "Forceps", 
        "anatomy_target": "RML", "findings": "Lesion removed"
    },
    {
        "event_id": "ev5", "event_type": "Balloon Dilation", "method": "Fogarty 4Fr", 
        "anatomy_target": "RML", "findings": "Purulent secretions"
    },
    {
        "event_id": "ev6", "event_type": "BAL", "method": "Lavage", 
        "anatomy_target": "RML (RB4/RB5)", "specimens": "Micro/Cyto"
    },
    {
        "event_id": "ev7", "event_type": "Balloon Dilation", "method": "Elation 6/7/8", 
        "anatomy_target": "RML Carina", "measurements": "8mm x 60s x 3"
    },
    {
        "event_id": "ev8", "event_type": "Foreign Body Removal", "method": "Rat Tooth Forceps", 
        "anatomy_target": "RML", "device": "iCAST 7x16mm (Old)"
    },
    {
        "event_id": "ev9", "event_type": "Thermal Ablation", "method": "Electrocautery", 
        "anatomy_target": "RML", "findings": "Hemorrhage treated"
    },
    {
        "event_id": "ev10", "event_type": "Airway Stent", "method": "Placement (Failed)", 
        "anatomy_target": "RML", "device": "iCAST 7x22mm", "notes": "Removed (blocked RLL)"
    },
    {
        "event_id": "ev11", "event_type": "Airway Stent", "method": "Placement (Failed)", 
        "anatomy_target": "RML", "device": "iCAST 7x16mm", "notes": "Removed (too proximal)"
    },
    {
        "event_id": "ev12", "event_type": "Airway Stent", "method": "Placement (Successful)", 
        "anatomy_target": "RML", "device": "iCAST 7x16mm", 
        "outcome_airway_lumen_pre": "Occluded", "outcome_airway_lumen_post": "Patent", "outcome_complication": "None",
        "device_size": "7mm x 16mm", "device_material": "iCAST"
    },
    {
        "event_id": "ev13", "event_type": "Balloon Dilation", "method": "Elation 6/7/8", 
        "anatomy_target": "RML Carina", "measurements": "8mm x 30s x 1", "notes": "Stent seating"
    }
]

# -------------------------------------------------------------------------
# REGISTRY JSON STRUCTURE
# -------------------------------------------------------------------------
REGISTRY_DATA = {
    "schema_version": "v3",
    "note_id": NOTE_ID,
    "no_immediate_complications": True,
    "procedures": [
        {
            "event_id": "ev12",
            "type": "Airway Stent",
            "target": {"anatomy_type": "Airway", "location": {"lobe": "RML"}},
            "method": "Stent Placement",
            "stent": {"type": "Covered", "brand": "iCAST", "size": "7mm x 16mm"},
            "outcomes": {
                "airway": {"lumen_pre": "Occluded", "lumen_post": "Patent"},
                "complications": "None"
            }
        }
    ]
}

# -------------------------------------------------------------------------
# HELPER FUNCTIONS
# -------------------------------------------------------------------------
def create_workbook():
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    return wb

def get_row(sheet):
    return sheet.max_row + 1

def hydrate_spans(spans, full_text):
    hydrated = []
    for span in spans:
        text, label, norm, ctx, schema, ev_id = span
        
        # Logic to find start/end char
        start = -1
        end = -1
        status = "ambiguous"
        
        count = full_text.count(text)
        if count == 0:
            status = "not_found"
        elif count == 1:
            start = full_text.find(text)
            end = start + len(text)
            status = "hydrated_unique"
        elif ctx:
            # Find all occurrences
            occurrences = [m.start() for m in re.finditer(re.escape(text), full_text)]
            # Check context
            for occ in occurrences:
                snippet = full_text[max(0, occ - 120):occ]
                if ctx in snippet:
                    start = occ
                    end = start + len(text)
                    status = "hydrated_context"
                    break
        
        # Fallback: if still ambiguous and no context, take first (or mark ambiguous)
        if start == -1 and count > 1:
            # Just take first for Phase 0 script generation purposes if no context provided
            start = full_text.find(text)
            end = start + len(text)
            status = "hydrated_first_fallback"

        hydrated.append({
            "text": text, "label": label, "norm": norm, "ctx": ctx,
            "schema": schema, "ev_id": ev_id, "start": start, "end": end,
            "status": status
        })
    return hydrated

# -------------------------------------------------------------------------
# MAIN GENERATION
# -------------------------------------------------------------------------
def main():
    if not os.path.exists(TEMPLATE_PATH):
        # Create a dummy template if missing (for standalone testing validity)
        wb = openpyxl.Workbook()
        for name in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            wb.create_sheet(name)
        wb.save(TEMPLATE_PATH)

    wb = create_workbook()
    
    # 1. Note_Text
    ws = wb["Note_Text"]
    ws.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # 2. Note_Index
    ws = wb["Note_Index"]
    header = ["source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "notes"] + PROCEDURE_FLAGS
    # Ensure header exists if empty
    if ws.max_row == 1:
        # Assuming template has header, if not we'd write it. 
        # But we append data.
        pass
        
    row_data = [SOURCE_FILE, NOTE_ID, "", "", "", "", "New", ""]
    for flag in PROCEDURE_FLAGS:
        row_data.append(FLAGS.get(flag, 0))
    ws.append(row_data)
    
    # 3. Span_Annotations
    ws_span = wb["Span_Annotations"]
    # Columns: source_file, note_id, span_id, section, ctx, text, match_idx, start, end, len, label, norm, schema, event, neg, hist, time, rev, comm, status
    
    # 4. Span_Hydrated
    ws_hydro = wb["Span_Hydrated"]
    
    hydrated_data = hydrate_spans(SPANS, NOTE_TEXT)
    
    for i, h in enumerate(hydrated_data):
        span_id = f"{NOTE_ID}_s{i+1:03d}"
        
        # Common Row
        # schema cols: source, note, span_id, section(blank), ctx, text, match_idx(blank), start, end, len, label, norm, schema, event, neg, hist, time, rev, comm, status
        row_base = [
            SOURCE_FILE, NOTE_ID, span_id, "", h["ctx"], h["text"], "", 
            "", "", f"=LEN(F{ws_span.max_row+1})", # Formula for len
            h["label"], h["norm"], h["schema"], h["ev_id"], 
            0, 0, "", "", "", "needs_hydration"
        ]
        ws_span.append(row_base)
        
        # Hydrated Row
        row_hydro = list(row_base)
        row_hydro[7] = h["start"] # start
        row_hydro[8] = h["end"]   # end
        row_hydro[9] = len(h["text"])
        row_hydro[19] = h["status"]
        ws_hydro.append(row_hydro)

    # 5. Event_Log
    ws_event = wb["Event_Log"]
    for ev in EVENTS:
        # source, note, id, type, method, anat, dev, gauge, station, count, meas, spec, find, hist, rev, comm, dev_sz, dev_mat, out_lum_pre, out_lum_post, out_sym, out_pleur, out_comp
        row = [
            SOURCE_FILE, NOTE_ID, ev["event_id"], ev["event_type"], ev.get("method"),
            ev.get("anatomy_target"), ev.get("device"), "", "", "", ev.get("measurements"),
            ev.get("specimens"), ev.get("findings"), 0, "", ev.get("notes"),
            ev.get("device_size"), ev.get("device_material"),
            ev.get("outcome_airway_lumen_pre"), ev.get("outcome_airway_lumen_post"),
            "", "", ev.get("outcome_complication")
        ]
        ws_event.append(row)

    # 6. V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    # Just map the main registry event (Stent)
    # note, event, type, anat_type, lobe, seg, station, les_type, les_sz, method, dev_json, meas_json, spec_json, find_json, quote, stent_sz, stent_mat, cath_sz, out_lum_pre, out_lum_post, out_sym, out_pleur, out_comp
    row_v3 = [
        NOTE_ID, "ev12", "Airway Stent", "Airway", "RML", "", "", "Stenosis", "",
        "Stent Placement", "", "", "", "", "", "7mm x 16mm", "iCAST", "",
        "Occluded", "Patent", "", "", "None"
    ]
    ws_v3.append(row_v3)

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    ws_json.append([json.dumps(REGISTRY_DATA, indent=2)])

    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    main()