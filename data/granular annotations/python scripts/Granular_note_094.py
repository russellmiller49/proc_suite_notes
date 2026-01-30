import openpyxl
from openpyxl.utils import get_column_letter
import json
import datetime
import re

# -------------------------------------------------------------------------
# INPUTS
# -------------------------------------------------------------------------
NOTE_ID = "note_094"
SOURCE_FILE = "note_094.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_094 SOURCE_FILE: note_094.txt INDICATION FOR OPERATION:  [REDACTED]is a 56 year old-year-old male who presents with bilateral lung transplant and complication of anastomosis dehiscence, ischemic lung injury, and bronchial stenosis requiring endobronchial stents.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
PROCEDURE:  
31899 Unlisted Procedure (Trach Change with Mature Tract or Procedure NOS)
31646 Therapeutic aspiration subsequent episodes
31622 Dx bronchoscope/cell washing          
31624 Dx bronchoscope/lavage (BAL)    
31640 Bronchoscopy with excision 
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure (50 MODIFIER):
This patient required a bilateral procedure today when this procedure would typically be unilateral.
Apply to: 
31640 Bronchoscopy with excision 
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy).
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
Disposable Bronchoscope
PROCEDURE IN DETAIL:
A timeout was performed (confirming the patient's name, procedure type, and procedure location).
Sedation initiated.  Mechanical ventilation initiated via patient's tracheostomy tube.
The Flexible Therapeutic Bronchoscope was advanced for airway examination.
Endobronchial topical lidocaine applied to the main carina, right carina 1, and left carina 2.
Initial Airway Examination Findings:
Trachea: Tracheostomy tube in good position.
Distal 1/3 normal.
Main Carina: Sharp
Right Lung Proximal Airways: RMSB intact with visible sutures, greatly improved from prior.
No evidence of prior RMS dehiscence  There is moderate stenosis at RMSB anastomosis site due to circumferential fibrous tissue, but able to easily traverse with therapeutic bronchoscope without touching.
RUL bronchus in good repair - prior area of full thickness erosion completely resolved;
small amount of fibrous scarring into RUL.  RB1-2 normal.  RB3 mildly stenotic but overall greatly improved.
Bronchus intermedius with tan exudative debris and granulation tissue along the medial aspect.
RML stent (AeroMini 8mm x 15mm) in good position and patent.
Small amount of tan exudative debris overlying RC2 carina minimally overlying edge of stent; minimal granulation tissue.
Mild amount of non-obstructing mucous adhered to inside of stent.  Able to traverse stent with therapeutic bronchoscope.
RB4/5 patent and normal.  RLL, RB6-10 are widely patent and appear healthy.
Left Lung Proximal Airways: LMS anastomosis intact with visible sutures.
There is mild stenosis at LMSB anastomosis site due to circumferential fibrous tissue, but able to easily traverse with therapeutic bronchoscope without touching.
LUL bronchus in good repair - prior area of full thickness erosion completely resolved.
Small amoutn of fibrous scarring in LUL.  LB1-3 patent and healthy.
Lingula stent (AeroMini 6mm x 10mm) in good position and patent.  Minimal non-obstructing mucus in stent.
Unable to traverse with therapeutic bronchoscope due to small size of stent but able to see fully patent LB4/LB5.
LLL, LB6-10 are patent and health.
Mucosa: As described above.
Secretions: Minimal, thin, and clear/white.
All secretions were suctioned to clear (therapeutic aspiration). 
Successful therapeutic aspiration was performed to clean out the distal trachea, RMSB, RUL, BI, RML stent, RLL, LMSB, LUL, lingula stent, LLL from mucus.
Endobronchial lesion of exudative debris was noted at RC2 and excised with mechanical debridement using bland alligator forceps.
This improved patency of RML stent.
Endobronchial lesion of exudative debris was noted in bronchus intermedius and excised with mechanical debridement using bland alligator forceps.
Endobronchial lesion of exudative debris was noted at LC1 and excised with mechanical debridement using bland alligator forceps.
This improved patency of lingula stent.
Endobronchial lesion of granulation tissue at bronchus intermedius, RMSB anastomosis stricture, and left carina 1 was treated with the following modalities:
Modality	Tools	Setting/Mode	Duration	Results
Cryoprobe	1.7mm cryoprobe	 	Numerous 30 seconds overlapping treatments	30 second overlapping treatments of granulation tissue at  bronchus intermedius, RMSB anastomosis stricture, LC1, and lingula takeoff.
Prior to treatment, the RML bronchus/stent was note to be 90% patent.  After treatment, the airway/stent was 100% patent.
Prior to treatment the bronchus intermedius was noted to be 90% patent.  After treatment, the airway was 90% patent.
Prior to treatment, the lingula bronchus/stent was note to be 90% patent.  After treatment, the airway/stent was 100% patent.
Therapeutic bronchoscope was removed.  Disposable bronchoscope with unused sterile channel was advanced.
Bronchial alveolar lavage was performed at RML bronchus.
Instilled 60 cc of NS, suction returned with 25 cc of NS.  Samples sent for Cell Count and Microbiology (Cultures/Viral/Fungal).
Residual secretions and saline was suctioned to clear.  Disposable bronchoscope removed.  Procedure completed.
The patient tolerated the procedure well.
There were no immediate complications.  At the conclusion of the operation, the patient transported to the recovery room in stable condition.
ESTIMATED BLOOD LOSS:   None
COMPLICATIONS:    None
SPECIMEN(S): 
RML BAL:  cell count, cultures/micro
IMPRESSION/PLAN: [REDACTED]is a 56 year old-year-old male who presents for bronchoscopy for airway evaluation and stent check.
-All airways continue to greatly improve.  Areas of mild-to-moderate stenosis as described above.
-RML stent and lingula stent in good position and patent.
Some mild tan exudative debris at edges of proximal stent causing minimal obstruction.
This debris was excised and areas of granulation tissue causing any stenosis were treated with cryoprobe.
-Repeat bronchoscopy in 3-4 weeks.  If stability in airways, then would pursue RML stent holiday (placed [REDACTED]) at that time.
Since lingula stent just placed [REDACTED], may need to remain in place longer to allow durable remodeling of airway.
"""

# -------------------------------------------------------------------------
# LOGIC & PARSING
# -------------------------------------------------------------------------

def get_procedure_flags(text):
    flags = {
        # Bronchoscopy (23)
        'diagnostic_bronchoscopy': 0, 'bal': 0, 'bronchial_wash': 0, 'brushings': 0,
        'endobronchial_biopsy': 0, 'tbna_conventional': 0, 'linear_ebus': 0, 'radial_ebus': 0,
        'navigational_bronchoscopy': 0, 'transbronchial_biopsy': 0, 'transbronchial_cryobiopsy': 0,
        'therapeutic_aspiration': 0, 'foreign_body_removal': 0, 'airway_dilation': 0,
        'airway_stent': 0, 'thermal_ablation': 0, 'tumor_debulking_non_thermal': 0,
        'cryotherapy': 0, 'blvr': 0, 'peripheral_ablation': 0, 'bronchial_thermoplasty': 0,
        'whole_lung_lavage': 0, 'rigid_bronchoscopy': 0,
        # Pleural (7)
        'thoracentesis': 0, 'chest_tube': 0, 'ipc': 0, 'medical_thoracoscopy': 0,
        'pleurodesis': 0, 'pleural_biopsy': 0, 'fibrinolytic_therapy': 0
    }
    
    txt_lower = text.lower()
    
    # Logic based on Note 094 content
    if "bronchoscopy" in txt_lower or "bronchoscope" in txt_lower:
        flags['diagnostic_bronchoscopy'] = 1
    if "bronchial alveolar lavage" in txt_lower or "bal" in txt_lower:
        flags['bal'] = 1
    if "therapeutic aspiration" in txt_lower:
        flags['therapeutic_aspiration'] = 1
    # 31622 implies washing
    if "31622" in txt_lower or "cell washing" in txt_lower:
        flags['bronchial_wash'] = 1
    # 31640 Bronchoscopy with excision -> tumor_debulking_non_thermal (mechanical debridement)
    if "excision" in txt_lower or "debridement" in txt_lower or "31640" in txt_lower:
        flags['tumor_debulking_non_thermal'] = 1
    # 31641 relief of stenosis (cryotherapy)
    if "cryoprobe" in txt_lower or "cryotherapy" in txt_lower:
        flags['cryotherapy'] = 1
        
    # Exclude stent placement (only stent check/cleaning performed, no new stent mentioned)
    
    return flags

def get_spans(text):
    spans = []
    
    def add_span(text_match, label, normalized, field, event_id, prefix=None, is_hist=False, is_neg=False, match_idx=0):
        spans.append({
            "span_text": text_match,
            "label": label,
            "normalized_value": normalized,
            "schema_field": field,
            "event_id": event_id,
            "context_prefix": prefix,
            "is_historical": is_hist,
            "is_negated": is_neg,
            "match_index": match_idx
        })

    # Event 1: Therapeutic Aspiration (General)
    add_span("Therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", "method", "ev1")
    add_span("All secretions were suctioned to clear", "PROC_ACTION", "Suction", "action", "ev1")
    add_span("distal trachea", "ANAT_AIRWAY", "Trachea", "target.anatomy_type", "ev1")
    add_span("RMSB", "ANAT_AIRWAY", "RMSB", "target.anatomy_type", "ev1", prefix="clean out the ")
    add_span("RUL", "ANAT_AIRWAY", "RUL", "target.anatomy_type", "ev1", prefix="RMSB, ")
    
    # Event 2: Debridement RML/RC2
    add_span("RC2", "ANAT_AIRWAY", "RC2", "target.anatomy_type", "ev2", prefix="debris was noted at ")
    add_span("RML stent", "DEV_STENT", "Stent", "stent", "ev2", prefix="improved patency of ")
    add_span("Endobronchial lesion of exudative debris", "OBS_LESION", "Exudative Debris", "lesion.type", "ev2", prefix="stent.\n")
    add_span("excised", "PROC_ACTION", "Excision", "method", "ev2", prefix="RC2 and ")
    add_span("mechanical debridement", "PROC_METHOD", "Mechanical Debridement", "method", "ev2", prefix="excised with ")
    add_span("bland alligator forceps", "DEV_INSTRUMENT", "Forceps", "device", "ev2", prefix="debridement using ")
    # Outcome for EV2
    add_span("90% patent", "OUTCOME_AIRWAY_LUMEN_PRE", "90", "outcomes.airway.lumen_pre", "ev2", prefix="RML bronchus/stent was note to be ")
    add_span("100% patent", "OUTCOME_AIRWAY_LUMEN_POST", "100", "outcomes.airway.lumen_post", "ev2", prefix="airway/stent was ")

    # Event 3: Debridement Bronchus Intermedius
    add_span("bronchus intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "target.anatomy_type", "ev3", prefix="noted in ")
    add_span("Endobronchial lesion of exudative debris", "OBS_LESION", "Exudative Debris", "lesion.type", "ev3", prefix="RML stent.\n")
    add_span("excised", "PROC_ACTION", "Excision", "method", "ev3", prefix="intermedius and ")
    add_span("mechanical debridement", "PROC_METHOD", "Mechanical Debridement", "method", "ev3", prefix="intermedius and excised with ")
    add_span("bland alligator forceps", "DEV_INSTRUMENT", "Forceps", "device", "ev3", prefix="intermedius and excised with mechanical debridement using ")
    # Outcome for EV3
    add_span("90% patent", "OUTCOME_AIRWAY_LUMEN_PRE", "90", "outcomes.airway.lumen_pre", "ev3", prefix="bronchus intermedius was noted to be ")
    add_span("90% patent", "OUTCOME_AIRWAY_LUMEN_POST", "90", "outcomes.airway.lumen_post", "ev3", prefix="airway was ")

    # Event 4: Debridement Lingula/LC1
    add_span("LC1", "ANAT_AIRWAY", "LC1", "target.anatomy_type", "ev4", prefix="noted at ")
    add_span("Endobronchial lesion of exudative debris", "OBS_LESION", "Exudative Debris", "lesion.type", "ev4", prefix="forceps.\n")
    add_span("excised", "PROC_ACTION", "Excision", "method", "ev4", prefix="LC1 and ")
    add_span("mechanical debridement", "PROC_METHOD", "Mechanical Debridement", "method", "ev4", prefix="LC1 and excised with ")
    add_span("bland alligator forceps", "DEV_INSTRUMENT", "Forceps", "device", "ev4", prefix="LC1 and excised with mechanical debridement using ")
    # Outcome for EV4
    add_span("90% patent", "OUTCOME_AIRWAY_LUMEN_PRE", "90", "outcomes.airway.lumen_pre", "ev4", prefix="lingula bronchus/stent was note to be ")
    add_span("100% patent", "OUTCOME_AIRWAY_LUMEN_POST", "100", "outcomes.airway.lumen_post", "ev4", prefix="lingula bronchus/stent was note to be 90% patent.  After treatment, the airway/stent was ")

    # Event 5: Cryotherapy (Overlapping sites)
    add_span("Cryoprobe", "PROC_METHOD", "Cryotherapy", "method", "ev5", prefix="Modality\tTools")
    add_span("1.7mm cryoprobe", "DEV_INSTRUMENT", "Cryoprobe", "device", "ev5")
    add_span("RMSB anastomosis stricture", "ANAT_AIRWAY", "RMSB", "target.anatomy_type", "ev5", prefix="granulation tissue at  bronchus intermedius, ")
    add_span("LC1", "ANAT_AIRWAY", "LC1", "target.anatomy_type", "ev5", prefix="RMSB anastomosis stricture, ")
    add_span("lingula takeoff", "ANAT_AIRWAY", "Lingula", "target.anatomy_type", "ev5")
    add_span("granulation tissue", "OBS_LESION", "Granulation Tissue", "lesion.type", "ev5", prefix="treatments of ")

    # Event 6: BAL RML
    add_span("Bronchial alveolar lavage", "PROC_METHOD", "BAL", "method", "ev6")
    add_span("RML bronchus", "ANAT_AIRWAY", "RML", "target.anatomy_type", "ev6", prefix="lavage was performed at ")
    add_span("Instilled 60 cc of NS", "PROC_ACTION", "Instilled 60cc NS", "action", "ev6")
    add_span("suction returned with 25 cc of NS", "PROC_ACTION", "Returned 25cc NS", "action", "ev6")

    # Stent Details (Historical/Existing context, mapped to check events or general)
    # RML Stent
    add_span("AeroMini 8mm x 15mm", "DEV_STENT", "AeroMini", "stent.brand", "ev2", is_hist=True)
    add_span("8mm", "DEV_STENT_SIZE", "8mm", "stent.size", "ev2", prefix="AeroMini ")
    add_span("15mm", "DEV_STENT_SIZE", "15mm", "stent.len", "ev2", prefix="8mm x ")
    
    # Lingula Stent
    add_span("AeroMini 6mm x 10mm", "DEV_STENT", "AeroMini", "stent.brand", "ev4", is_hist=True)
    add_span("6mm", "DEV_STENT_SIZE", "6mm", "stent.size", "ev4", prefix="AeroMini ")
    add_span("10mm", "DEV_STENT_SIZE", "10mm", "stent.len", "ev4", prefix="6mm x ")

    # Complications (Global)
    add_span("no immediate complications", "OUTCOME_COMPLICATION", "None", "outcomes.complications", "ev7")
    
    # Extra Anatomy contexts
    add_span("RMSB anastomosis", "ANAT_AIRWAY", "RMSB", "target.anatomy_type", "ev5", prefix="There is moderate stenosis at ")

    return spans

def get_event_definitions():
    return {
        "ev1": {"type": "Therapeutic Aspiration", "anatomy": "Trachea, RMSB, RUL, BI, RML, LUL, Lingula", "method": "Suction"},
        "ev2": {"type": "Mechanical Debridement", "anatomy": "RC2", "method": "Forceps", "outcome_pre": "90", "outcome_post": "100"},
        "ev3": {"type": "Mechanical Debridement", "anatomy": "Bronchus Intermedius", "method": "Forceps", "outcome_pre": "90", "outcome_post": "90"},
        "ev4": {"type": "Mechanical Debridement", "anatomy": "LC1/Lingula", "method": "Forceps", "outcome_pre": "90", "outcome_post": "100"},
        "ev5": {"type": "Cryotherapy", "anatomy": "BI, RMSB, LC1, Lingula", "method": "Cryoprobe"},
        "ev6": {"type": "BAL", "anatomy": "RML", "method": "Lavage"},
        "ev7": {"type": "General", "anatomy": "Airway", "complication": "None"}
    }

# -------------------------------------------------------------------------
# EXECUTION
# -------------------------------------------------------------------------

def hydrate_spans(spans, full_text):
    hydrated = []
    for s in spans:
        target = s['span_text']
        start = -1
        end = -1
        status = "ambiguous"
        
        # 1. Unique
        if full_text.count(target) == 1:
            start = full_text.find(target)
            end = start + len(target)
            status = "hydrated_unique"
        
        # 2. Context Prefix
        elif s.get('context_prefix'):
            prefix = s['context_prefix']
            # Find all occurrences of target
            occurrences = [m.start() for m in re.finditer(re.escape(target), full_text)]
            for occ in occurrences:
                # Look back 120 chars
                window_start = max(0, occ - 120)
                window_text = full_text[window_start:occ]
                if prefix in window_text:
                    start = occ
                    end = start + len(target)
                    status = "hydrated_prefix_window"
                    break
        
        # 3. Match Index
        elif s.get('match_index') is not None:
            idx = s['match_index']
            occurrences = [m.start() for m in re.finditer(re.escape(target), full_text)]
            if 0 <= idx < len(occurrences):
                start = occurrences[idx]
                end = start + len(target)
                status = "hydrated_match_index"

        s_copy = s.copy()
        s_copy['start_char'] = start if start != -1 else ""
        s_copy['end_char'] = end if end != -1 else ""
        s_copy['span_len'] = len(target)
        s_copy['hydration_status'] = status
        hydrated.append(s_copy)
    return hydrated

def create_workbook():
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    
    # 1. Note_Text
    ws_text = wb["Note_Text"]
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # 2. Note_Index
    ws_index = wb["Note_Index"]
    flags = get_procedure_flags(NOTE_TEXT)
    row_data = [
        SOURCE_FILE, NOTE_ID, "", "", "", "", "Pending", ""
    ]
    # Order matches the template instructions (30 flags)
    # Bronchoscopy (23)
    keys = [
        'diagnostic_bronchoscopy', 'bal', 'bronchial_wash', 'brushings',
        'endobronchial_biopsy', 'tbna_conventional', 'linear_ebus', 'radial_ebus',
        'navigational_bronchoscopy', 'transbronchial_biopsy', 'transbronchial_cryobiopsy',
        'therapeutic_aspiration', 'foreign_body_removal', 'airway_dilation',
        'airway_stent', 'thermal_ablation', 'tumor_debulking_non_thermal',
        'cryotherapy', 'blvr', 'peripheral_ablation', 'bronchial_thermoplasty',
        'whole_lung_lavage', 'rigid_bronchoscopy',
        # Pleural (7)
        'thoracentesis', 'chest_tube', 'ipc', 'medical_thoracoscopy',
        'pleurodesis', 'pleural_biopsy', 'fibrinolytic_therapy'
    ]
    for k in keys:
        row_data.append(flags[k])
    ws_index.append(row_data)
    
    # 3. Spans
    spans = get_spans(NOTE_TEXT)
    hydrated_spans = hydrate_spans(spans, NOTE_TEXT)
    
    ws_anno = wb["Span_Annotations"]
    ws_hydra = wb["Span_Hydrated"]
    
    headers = [
        "source_file", "note_id", "span_id", "section_type",
        "context_prefix", "span_text", "match_index",
        "start_char", "end_char", "span_len",
        "label", "normalized_value", "schema_field", "event_id",
        "is_negated", "is_historical", "time_anchor", "reviewer", "comments",
        "hydration_status"
    ]
    
    for idx, s in enumerate(hydrated_spans):
        # Anchor First (Raw)
        row_raw = [
            SOURCE_FILE, NOTE_ID, f"span_{idx+1}", "",
            s.get('context_prefix', ""), s['span_text'], s.get('match_index', ""),
            "", "", f"=LEN(F{idx+2})", # Formula for length
            s['label'], s['normalized_value'], s['schema_field'], s['event_id'],
            s['is_negated'], s['is_historical'], "", "", "",
            "needs_hydration"
        ]
        ws_anno.append(row_raw)
        
        # Hydrated
        row_hyd = [
            SOURCE_FILE, NOTE_ID, f"span_{idx+1}", "",
            s.get('context_prefix', ""), s['span_text'], s.get('match_index', ""),
            s['start_char'], s['end_char'], s['span_len'],
            s['label'], s['normalized_value'], s['schema_field'], s['event_id'],
            s['is_negated'], s['is_historical'], "", "", "",
            s['hydration_status']
        ]
        ws_hydra.append(row_hyd)

    # 4. Event_Log
    ws_events = wb["Event_Log"]
    ev_defs = get_event_definitions()
    for eid, data in ev_defs.items():
        row = [
            SOURCE_FILE, NOTE_ID, eid, data.get('type', ""), data.get('method', ""),
            data.get('anatomy', ""), "", "", "", "", "",
            "", "", "", "", "",
            "", "", # device size/mat
            data.get('outcome_pre', ""), data.get('outcome_post', ""),
            "", "", data.get('complication', "")
        ]
        ws_events.append(row)
        
    # 5. V3_Procedure_Events & JSON
    ws_v3 = wb["V3_Procedure_Events"]
    v3_events = []
    
    for eid, data in ev_defs.items():
        v3_row = [
            NOTE_ID, eid, data.get('type'),
            data.get('anatomy'), "", "", "", # target parts
            "", "", # lesion
            data.get('method'), "", "", "", "", "", # json fields
            "", "", "", # stent/cath
            data.get('outcome_pre', ""), data.get('outcome_post', ""),
            "", "", data.get('complication', "")
        ]
        ws_v3.append(v3_row)
        
        # Construct simplified JSON representation
        v3_events.append({
            "event_id": eid,
            "procedure_type": data.get('type'),
            "target_anatomy": data.get('anatomy'),
            "method": data.get('method'),
            "outcomes": {
                "airway_lumen_pre": data.get('outcome_pre'),
                "airway_lumen_post": data.get('outcome_post'),
                "complications": data.get('complication')
            }
        })

    registry_json = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "procedures": v3_events,
        "no_immediate_complications": True
    }
    
    ws_json = wb["V3_Registry_JSON"]
    ws_json.cell(row=1, column=1, value=json.dumps(registry_json, indent=2))

    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    create_workbook()