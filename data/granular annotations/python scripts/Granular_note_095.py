import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import json
import re
import os

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_095"
SOURCE_FILE = "note_095.txt"
PROCEDURE_DATE = "2026-01-12" # Placeholder/Current
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_095 SOURCE_FILE: note_095.txt INDICATION FOR OPERATION:  [REDACTED]is a 56 year old-year-old male who presents with airway stenosis.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31625 Endobronchial Biopsy(s)
31636 Dilate and bronchial stent initial bronchus
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
Flexible Hybrid (Pedatric) Bronchoscope
ESTIMATED BLOOD LOSS:   None
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
Initial Airway Inspection Findings:
There were significant stenosis noted at the lingula orifice. Mild stenosis at the RMS anastomosis and BI.
Successful therapeutic aspiration was performed to clean out the Right Mainstem and Left Mainstem from mucus.
Endobronchial obstruction at Right Mainstem, Bronchus Intermedius , and lingula orifice was treated with the following modalities:
Modality	Tools	Setting/Mode	Duration	Results
Electrocautery	Needle knife	Effect 4, 40W	2-3 sec radial cuts	At lingula, RMS
Endobronchial biopsy was performed at Bronchus Intermedius .
Lesion was successfully removed.  Samples sent for Pathology.
Balloon dilation was performed at Right Mainstem.
10/11/12 Elation balloon was used to perform dilation to 12 mm at the Right Mainstem.
Total 1 inflations with dilation time of 60 seconds each.
Prior to treatment, affected airway was note to be 80% patent.  After treatment, the airway was 100% patent.
Balloon dilation was performed at lingula.  6/7/8 Elation balloon was used to perform dilation to 8 mm at the LUL Lingula Carina (Lc1).
Total 1 inflations with dilation time of 60 seconds each.
The following stent (aero mini 6x10) was placed in the lingula orifice.
Prior to treatment, affected airway was note to be 20% patent.  After treatment, the airway was 100% patent.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
none
IMPRESSION/PLAN: [REDACTED]is a 56 year old-year-old male who presents for bronchoscopy for airway stenosis.
- f/u in 2 weeks for repeat bronch"""

# -------------------------------------------------------------------------
# PROCEDURE FLAGS
# -------------------------------------------------------------------------
PROC_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1, # Base
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 1,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 0,
    "airway_dilation": 1,
    "airway_stent": 1,
    "thermal_ablation": 1, # Electrocautery mentioned
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# -------------------------------------------------------------------------
# SPAN DEFINITIONS (ANCHOR-FIRST)
# -------------------------------------------------------------------------
# Each tuple: (span_text, label, normalized_value, schema_field, event_id, context_prefix)
SPANS = [
    # Event 1: Aspiration
    ("therapeutic aspiration", "PROC_METHOD", "therapeutic aspiration", "method", "evt1", "Successful"),
    ("Right Mainstem", "ANAT_AIRWAY", "RMS", "target.anatomy", "evt1", "clean out the"),
    ("Left Mainstem", "ANAT_AIRWAY", "LMS", "target.anatomy", "evt1", "Right Mainstem and"),
    
    # Event 2: Electrocautery/Needle Knife (RMS + Lingula)
    ("Electrocautery", "PROC_METHOD", "electrocautery", "method", "evt2", None),
    ("Needle knife", "DEV_INSTRUMENT", "needle knife", "device", "evt2", None),
    ("Effect 4, 40W", "MEAS_PRESS", "Effect 4, 40W", "setting", "evt2", None), # Mapping to generic measure or setting
    ("lingula", "ANAT_AIRWAY", "Lingula", "target.anatomy", "evt2", "At"),
    ("RMS", "ANAT_AIRWAY", "RMS", "target.anatomy", "evt2", "At lingula,"),
    ("radial cuts", "PROC_ACTION", "radial cuts", "findings", "evt2", None),

    # Event 3: EBBx
    ("Endobronchial biopsy", "PROC_METHOD", "endobronchial biopsy", "method", "evt3", None),
    ("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "target.anatomy", "evt3", "performed at"),
    ("Samples sent for Pathology", "PROC_ACTION", "samples sent", "specimens", "evt3", None),
    
    # Event 4: Balloon Dilation RMS
    ("Balloon dilation", "PROC_METHOD", "balloon dilation", "method", "evt4", "Pathology.\n"),
    ("Right Mainstem", "ANAT_AIRWAY", "RMS", "target.anatomy", "evt4", "performed at"),
    ("10/11/12 Elation balloon", "DEV_INSTRUMENT", "Elation balloon 10-11-12mm", "device", "evt4", None),
    ("12 mm", "MEAS_SIZE", "12 mm", "measurements", "evt4", "dilation to"),
    ("80% patent", "OUTCOME_AIRWAY_LUMEN_PRE", "80% patent", "outcomes.airway.lumen_pre", "evt4", "note to be"),
    ("100% patent", "OUTCOME_AIRWAY_LUMEN_POST", "100% patent", "outcomes.airway.lumen_post", "evt4", "airway was"),
    ("1 inflations", "MEAS_COUNT", "1", "counts", "evt4", "Total"),
    ("60 seconds", "MEAS_TIME", "60 seconds", "measurements", "evt4", "time of"),

    # Event 5: Balloon Dilation Lingula
    ("Balloon dilation", "PROC_METHOD", "balloon dilation", "method", "evt5", "patent.\n"),
    ("lingula", "ANAT_AIRWAY", "Lingula", "target.anatomy", "evt5", "performed at"),
    ("6/7/8 Elation balloon", "DEV_INSTRUMENT", "Elation balloon 6-7-8mm", "device", "evt5", None),
    ("8 mm", "MEAS_SIZE", "8 mm", "measurements", "evt5", "dilation to"),
    ("LUL Lingula Carina", "ANAT_AIRWAY", "LUL Lingula Carina", "target.anatomy", "evt5", None),
    ("Lc1", "ANAT_AIRWAY", "Lc1", "target.anatomy", "evt5", None),
    ("20% patent", "OUTCOME_AIRWAY_LUMEN_PRE", "20% patent", "outcomes.airway.lumen_pre", "evt5", "note to be"),
    ("100% patent", "OUTCOME_AIRWAY_LUMEN_POST", "100% patent", "outcomes.airway.lumen_post", "evt5", "airway was"),

    # Event 6: Stent Lingula
    ("stent", "DEV_STENT", "stent", "device", "evt6", "The following"),
    ("aero mini 6x10", "DEV_STENT_SIZE", "6x10mm", "stent.size", "evt6", None),
    ("lingula orifice", "ANAT_AIRWAY", "Lingula Orifice", "target.anatomy", "evt6", "placed in the"),

    # Complications / Global
    ("no immediate complications", "OUTCOME_COMPLICATION", "none", "outcomes.complications", "evt_global", None),
    ("Airway Inspection Findings", "CTX_TIME", "inspection", "context", "evt_global", None),
    ("significant stenosis", "OBS_LESION", "stenosis", "findings", "evt_global", "There were"),
    ("lingula orifice", "ANAT_AIRWAY", "Lingula Orifice", "target.anatomy", "evt_global", "at the"),
    ("Mild stenosis", "OBS_LESION", "stenosis", "findings", "evt_global", None),
    ("RMS anastomosis", "ANAT_AIRWAY", "RMS Anastomosis", "target.anatomy", "evt_global", None),
]

# -------------------------------------------------------------------------
# EVENT LOG DEFINITIONS
# -------------------------------------------------------------------------
EVENTS = [
    {
        "event_id": "evt1",
        "event_type": "therapeutic_aspiration",
        "method": "therapeutic aspiration",
        "anatomy_target": "RMS, LMS",
        "findings": "mucus cleaned out",
        "reviewer": "Auto",
    },
    {
        "event_id": "evt2",
        "event_type": "thermal_ablation",
        "method": "electrocautery",
        "anatomy_target": "Lingula, RMS",
        "device": "Needle knife",
        "findings": "radial cuts",
        "measurements": "Effect 4, 40W",
        "reviewer": "Auto",
    },
    {
        "event_id": "evt3",
        "event_type": "endobronchial_biopsy",
        "method": "endobronchial biopsy",
        "anatomy_target": "Bronchus Intermedius",
        "specimens": "Pathology",
        "reviewer": "Auto",
    },
    {
        "event_id": "evt4",
        "event_type": "airway_dilation",
        "method": "balloon dilation",
        "anatomy_target": "RMS",
        "device": "10/11/12 Elation balloon",
        "measurements": "12 mm, 60 seconds",
        "counts": "1 inflation",
        "outcome_airway_lumen_pre": "80% patent",
        "outcome_airway_lumen_post": "100% patent",
        "reviewer": "Auto",
    },
    {
        "event_id": "evt5",
        "event_type": "airway_dilation",
        "method": "balloon dilation",
        "anatomy_target": "Lingula (Lc1)",
        "device": "6/7/8 Elation balloon",
        "measurements": "8 mm",
        "outcome_airway_lumen_pre": "20% patent",
        "outcome_airway_lumen_post": "100% patent",
        "reviewer": "Auto",
    },
    {
        "event_id": "evt6",
        "event_type": "airway_stent",
        "method": "stent placement",
        "anatomy_target": "Lingula Orifice",
        "device": "aero mini 6x10",
        "device_size": "6x10",
        "reviewer": "Auto",
    }
]

# -------------------------------------------------------------------------
# REGISTRY JSON STRUCTURE
# -------------------------------------------------------------------------
REGISTRY_DATA = {
    "schema_version": "3.0",
    "note_id": NOTE_ID,
    "no_immediate_complications": True,
    "procedures": [
        {
            "event_id": "evt1",
            "type": "therapeutic_aspiration",
            "target": {"anatomy_type": "airway", "location": {"lobe": "rm", "segment": "mainstem"}},
            "method": "suction"
        },
        {
            "event_id": "evt2",
            "type": "thermal_ablation",
            "target": {"anatomy_type": "airway", "location": {"lobe": "lul", "segment": "lingula"}},
            "method": "electrocautery",
            "devices_json": [{"name": "Needle knife", "type": "instrument"}]
        },
        {
            "event_id": "evt3",
            "type": "endobronchial_biopsy",
            "target": {"anatomy_type": "airway", "location": {"lobe": "rm", "segment": "bi"}},
            "method": "forceps"
        },
        {
            "event_id": "evt4",
            "type": "airway_dilation",
            "target": {"anatomy_type": "airway", "location": {"lobe": "rm", "segment": "mainstem"}},
            "method": "balloon",
            "devices_json": [{"name": "10/11/12 Elation balloon", "type": "balloon"}],
            "outcomes": {"airway": {"lumen_pre": "80% patent", "lumen_post": "100% patent"}}
        },
        {
            "event_id": "evt5",
            "type": "airway_dilation",
            "target": {"anatomy_type": "airway", "location": {"lobe": "lul", "segment": "lingula"}},
            "method": "balloon",
            "devices_json": [{"name": "6/7/8 Elation balloon", "type": "balloon"}],
            "outcomes": {"airway": {"lumen_pre": "20% patent", "lumen_post": "100% patent"}}
        },
        {
            "event_id": "evt6",
            "type": "airway_stent",
            "target": {"anatomy_type": "airway", "location": {"lobe": "lul", "segment": "lingula"}},
            "method": "stent",
            "stent": {"size": "6x10", "material_or_brand": "aero mini"}
        }
    ]
}

# -------------------------------------------------------------------------
# HYDRATION LOGIC
# -------------------------------------------------------------------------
class SpanHydrator:
    def __init__(self, full_text):
        self.text = full_text

    def find_span(self, span_text, context_prefix=None, match_index=None):
        """
        Returns (start, end, status)
        """
        if not span_text:
            return None, None, "missing_text"

        matches = [m for m in re.finditer(re.escape(span_text), self.text)]
        
        if not matches:
            return None, None, "not_found"

        # 1. Exact unique match
        if len(matches) == 1:
            m = matches[0]
            return m.start(), m.end(), "hydrated_unique"

        # 2. Context match
        if context_prefix:
            # Look for context in preceding 120 chars
            candidates = []
            for m in matches:
                start = m.start()
                search_start = max(0, start - 120)
                preceding_text = self.text[search_start:start]
                if context_prefix in preceding_text:
                    candidates.append(m)
            
            if len(candidates) == 1:
                return candidates[0].start(), candidates[0].end(), "hydrated_prefix_window"
            elif len(candidates) > 1:
                # If multiple context matches, fall back to index if provided, else first
                if match_index is not None and 0 <= match_index < len(candidates):
                    return candidates[match_index].start(), candidates[match_index].end(), "hydrated_context_index"
                return candidates[0].start(), candidates[0].end(), f"ambiguous_context_count={len(candidates)}"

        # 3. Match Index
        if match_index is not None and 0 <= match_index < len(matches):
            return matches[match_index].start(), matches[match_index].end(), "hydrated_match_index"

        return None, None, f"ambiguous_count={len(matches)}"

# -------------------------------------------------------------------------
# SCRIPT EXECUTION
# -------------------------------------------------------------------------
def main():
    if not os.path.exists(TEMPLATE_PATH):
        # Create a dummy template if not found (for standalone testing), 
        # but in real use, file must exist.
        wb = Workbook()
        for sheet in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            wb.create_sheet(sheet)
        wb.save(TEMPLATE_PATH)

    wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # 1. Note_Text
    ws_text = wb["Note_Text"]
    # Check headers or append
    if ws_text.max_row == 1 and ws_text["A1"].value is None:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws_index = wb["Note_Index"]
    header = [
        "source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes"
    ] + list(PROC_FLAGS.keys())
    
    # Write header if empty
    if ws_index.max_row == 1 and ws_index["A1"].value is None:
        ws_index.append(header)
    
    row_data = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Completed", ""]
    row_data.extend(PROC_FLAGS.values())
    ws_index.append(row_data)

    # 3. Span_Annotations & 4. Span_Hydrated
    ws_anno = wb["Span_Annotations"]
    ws_hydra = wb["Span_Hydrated"]
    
    # Headers
    span_headers = [
        "source_file", "note_id", "span_id", "section_type",
        "context_prefix", "span_text", "match_index",
        "start_char", "end_char", "span_len",
        "label", "normalized_value", "schema_field", "event_id",
        "is_negated", "is_historical", "time_anchor", "reviewer", "comments",
        "hydration_status"
    ]

    if ws_anno.max_row == 1 and ws_anno["A1"].value is None:
        ws_anno.append(span_headers)
    if ws_hydra.max_row == 1 and ws_hydra["A1"].value is None:
        ws_hydra.append(span_headers)

    hydrator = SpanHydrator(NOTE_TEXT)

    for i, span_def in enumerate(SPANS):
        span_text, label, norm_val, schema_field, event_id, ctx = span_def
        span_id = f"{NOTE_ID}_s{i+1:03d}"
        
        # Annotation Row (No offsets)
        anno_row = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure",
            ctx, span_text, None, # match_index default None
            None, None, f"=LEN(F{ws_anno.max_row+1})", # Formula for len
            label, norm_val, schema_field, event_id,
            False, False, None, "Auto", "", "needs_hydration"
        ]
        ws_anno.append(anno_row)

        # Hydration
        start, end, status = hydrator.find_span(span_text, context_prefix=ctx)
        
        hydra_row = list(anno_row)
        hydra_row[7] = start
        hydra_row[8] = end
        hydra_row[9] = (end - start) if (start is not None and end is not None) else 0
        hydra_row[-1] = status
        ws_hydra.append(hydra_row)

    # 5. Event_Log
    ws_events = wb["Event_Log"]
    event_headers = [
        "source_file", "note_id", "event_id", "event_type", "method",
        "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements",
        "specimens", "findings", "is_historical", "reviewer", "comments",
        "device_size", "device_material",
        "outcome_airway_lumen_pre", "outcome_airway_lumen_post",
        "outcome_symptoms", "outcome_pleural", "outcome_complication"
    ]
    if ws_events.max_row == 1 and ws_events["A1"].value is None:
        ws_events.append(event_headers)

    for evt in EVENTS:
        row = [
            SOURCE_FILE, NOTE_ID, evt.get("event_id"), evt.get("event_type"), evt.get("method"),
            evt.get("anatomy_target"), evt.get("device"), evt.get("needle_gauge"), evt.get("stations"), evt.get("counts"), evt.get("measurements"),
            evt.get("specimens"), evt.get("findings"), False, evt.get("reviewer"), evt.get("comments"),
            evt.get("device_size"), evt.get("device_material"),
            evt.get("outcome_airway_lumen_pre"), evt.get("outcome_airway_lumen_post"),
            evt.get("outcome_symptoms"), evt.get("outcome_pleural"), evt.get("outcome_complication")
        ]
        ws_events.append(row)

    # 6. V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    v3_headers = [
        "note_id", "event_id", "type",
        "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station",
        "lesion.type", "lesion.size_mm",
        "method", "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote",
        "stent.size", "stent.material_or_brand", "catheter.size_fr",
        "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post",
        "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"
    ]
    if ws_v3.max_row == 1 and ws_v3["A1"].value is None:
        ws_v3.append(v3_headers)

    for item in REGISTRY_DATA["procedures"]:
        # Flatten simple
        devices = json.dumps(item.get("devices_json", [])) if item.get("devices_json") else ""
        outcomes = item.get("outcomes", {})
        airway = outcomes.get("airway", {})
        
        row = [
            NOTE_ID, item.get("event_id"), item.get("type"),
            item.get("target", {}).get("anatomy_type"), item.get("target", {}).get("location", {}).get("lobe"), item.get("target", {}).get("location", {}).get("segment"), None,
            None, None,
            item.get("method"), devices, "", "", "", "",
            item.get("stent", {}).get("size"), item.get("stent", {}).get("material_or_brand"), None,
            airway.get("lumen_pre"), airway.get("lumen_post"),
            outcomes.get("symptoms"), outcomes.get("pleural"), outcomes.get("complications")
        ]
        ws_v3.append(row)

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 1 and ws_json["A1"].value is None:
        ws_json.append(["schema_version", "note_id", "json_object"])
    
    ws_json.append(["3.0", NOTE_ID, json.dumps(REGISTRY_DATA, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Successfully generated {OUTPUT_PATH}")

if __name__ == "__main__":
    main()