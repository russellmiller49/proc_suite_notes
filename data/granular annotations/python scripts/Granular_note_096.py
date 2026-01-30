import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import datetime

# ==========================================
# INPUT DATA
# ==========================================
NOTE_ID = "note_096"
SOURCE_FILE = "note_096.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_096 SOURCE_FILE: note_096.txt INDICATION FOR OPERATION:  [REDACTED]is a 56 year old-year-old male who presents with respiratory failure.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: J96.90 Respiratory Failure
POSTOPERATIVE DIAGNOSIS:  J96.90 Respiratory Failure
PROCEDURE:  
31615 Visualization of windpipe (Tracheobronchoscopy through established tracheostomy incision)
31646 Therapeutic aspiration subsequent episodes
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
ESTIMATED BLOOD LOSS:   None
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
Initial Airway Inspection Findings:
The airway was inspected via the tracheostomy tube. Lidocaine was applied to the airway.
Successful therapeutic aspiration was performed to clean out the Trachea (Middle 1/3), Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus and mucus plug.
Secretions were cleared. 
The airway anatomy demonstrated improvement of the yellow necrotic tissue. It remains present, but less prominent.
The right middle lobe stent was partially occluded with mucus this was rinsed with saline and cleaned.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
None
IMPRESSION/PLAN: [REDACTED]is a 56 year old-year-old male who presents for bronchoscopy for respiratory failure.
-Follow up bronchoscopy in 2-3 weeks"""

# ==========================================
# CONFIGURATION
# ==========================================
PROCEDURE_FLAGS = [
    # Bronchoscopy (23)
    'diagnostic_bronchoscopy', 'bal', 'bronchial_wash', 'brushings', 'endobronchial_biopsy',
    'tbna_conventional', 'linear_ebus', 'radial_ebus', 'navigational_bronchoscopy',
    'transbronchial_biopsy', 'transbronchial_cryobiopsy', 'therapeutic_aspiration',
    'foreign_body_removal', 'airway_dilation', 'airway_stent', 'thermal_ablation',
    'tumor_debulking_non_thermal', 'cryotherapy', 'blvr', 'peripheral_ablation',
    'bronchial_thermoplasty', 'whole_lung_lavage', 'rigid_bronchoscopy',
    # Pleural (7)
    'thoracentesis', 'chest_tube', 'ipc', 'medical_thoracoscopy', 'pleurodesis',
    'pleural_biopsy', 'fibrinolytic_therapy'
]

# Set active flags based on note content
ACTIVE_FLAGS = {
    'diagnostic_bronchoscopy': 1, # "Visualization of windpipe", "Airway Inspection Findings"
    'therapeutic_aspiration': 1,  # "Therapeutic aspiration subsequent episodes", "Successful therapeutic aspiration"
    # Note: "right middle lobe stent" mentioned, but action is rinsing/cleaning (aspiration/maintenance), not insertion/removal.
    # Stent cleaning falls under therapeutic aspiration or general toilet. Stent flag 0.
}

# ==========================================
# SPAN EXTRACTION LOGIC
# ==========================================
# Manual definition of spans to extract
# Format: (text, label, normalized_value, schema_field, event_id)
SPANS_TO_EXTRACT = [
    # Event 1: Therapeutic Aspiration (General Airway)
    ("Therapeutic aspiration", "PROC_METHOD", "therapeutic_aspiration", "method", "evt_01"),
    ("clean out", "PROC_ACTION", "clean", "action", "evt_01"),
    ("Trachea (Middle 1/3)", "ANAT_AIRWAY", "trachea_mid", "target.anatomy_type", "evt_01"),
    ("Trachea (Distal 1/3)", "ANAT_AIRWAY", "trachea_distal", "target.anatomy_type", "evt_01"),
    ("Right Mainstem", "ANAT_AIRWAY", "rm", "target.anatomy_type", "evt_01"),
    ("Bronchus Intermedius", "ANAT_AIRWAY", "bi", "target.anatomy_type", "evt_01"),
    ("Left Mainstem", "ANAT_AIRWAY", "lm", "target.anatomy_type", "evt_01"),
    ("Carina", "ANAT_AIRWAY", "carina", "target.anatomy_type", "evt_01"),
    ("RUL Carina (RC1)", "ANAT_AIRWAY", "rul_carina", "target.anatomy_type", "evt_01"),
    ("RML Carina (RC2)", "ANAT_AIRWAY", "rml_carina", "target.anatomy_type", "evt_01"),
    ("LUL Lingula Carina (Lc1)", "ANAT_AIRWAY", "lingula_carina", "target.anatomy_type", "evt_01"),
    ("Left Carina (LC2)", "ANAT_AIRWAY", "left_carina", "target.anatomy_type", "evt_01"),
    ("mucus", "OBS_LESION", "mucus", "findings", "evt_01"),
    ("mucus plug", "OBS_LESION", "mucus_plug", "findings", "evt_01"),
    ("Secretions were cleared", "OUTCOME_AIRWAY_LUMEN_POST", "cleared", "outcomes.airway.lumen_post", "evt_01"),
    ("yellow necrotic tissue", "OBS_LESION", "necrotic_tissue", "findings", "evt_01"),
    ("improvement of the yellow necrotic tissue", "OBS_LESION", "improved", "findings", "evt_01"),

    # Event 2: Stent Management (RML)
    ("right middle lobe", "ANAT_AIRWAY", "rml", "target.location.lobe", "evt_02"),
    ("stent", "DEV_STENT", "stent", "devices", "evt_02"),
    ("partially occluded", "OBS_LESION", "occluded_partial", "findings", "evt_02"),
    ("mucus", "OBS_LESION", "mucus", "findings", "evt_02"),
    ("rinsed", "PROC_ACTION", "rinse", "method", "evt_02"),
    ("cleaned", "PROC_ACTION", "clean", "method", "evt_02"),

    # General / Complications
    ("no immediate complications", "OUTCOME_COMPLICATION", "none", "outcomes.complications", "global"),
]

# ==========================================
# HELPER FUNCTIONS
# ==========================================
def clean_text(text):
    return text.replace('\r', '').strip()

def find_offsets_anchor_first(note_text, span_text, context_prefix=None, match_index=None):
    """
    Hydrates offsets based on the rubric:
    1. unique occurrence
    2. context prefix (preceding 120 chars)
    3. match_index (nth occurrence)
    """
    # Find all occurrences
    matches = [m for m in re.finditer(re.escape(span_text), note_text)]
    
    if not matches:
        return None, None, "not_found"
    
    # 1. Unique
    if len(matches) == 1:
        return matches[0].start(), matches[0].end(), "hydrated_unique"
    
    # 2. Context Prefix
    if context_prefix:
        for m in matches:
            start = m.start()
            window_start = max(0, start - 120)
            preceding_text = note_text[window_start:start]
            if context_prefix in preceding_text:
                return m.start(), m.end(), "hydrated_prefix_window"
    
    # 3. Match Index
    if match_index is not None:
        if 0 <= match_index < len(matches):
            return matches[match_index].start(), matches[match_index].end(), "hydrated_match_index"
            
    # Fallback: ambiguous
    return None, None, f"ambiguous_count={len(matches)}"

# ==========================================
# MAIN CLASS
# ==========================================
class RegistryWorkbookGenerator:
    def __init__(self):
        self.wb = None
        self.spans_data = [] # To store span rows for hydration
        
    def load_template(self):
        try:
            self.wb = openpyxl.load_workbook(TEMPLATE_PATH)
        except FileNotFoundError:
            # Create a basic workbook if template missing (fallback for environment without file)
            self.wb = openpyxl.Workbook()
            # Create required sheets
            for sheet in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
                if sheet not in self.wb.sheetnames:
                    self.wb.create_sheet(sheet)

    def populate_note_text(self):
        ws = self.wb["Note_Text"]
        # Append row: note_id, source_file, note_text
        ws.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    def populate_note_index(self):
        ws = self.wb["Note_Index"]
        # Headers: source_file, note_id, encounter_id, procedure_date, site, reviewer, status, free_text_notes, [flags...]
        # Assuming headers exist or appending blindly if template.
        
        row_data = [
            SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "complete", ""
        ]
        
        # Append 0/1 for all flags
        for flag in PROCEDURE_FLAGS:
            row_data.append(1 if flag in ACTIVE_FLAGS else 0)
            
        ws.append(row_data)

    def populate_spans(self):
        ws_anno = self.wb["Span_Annotations"]
        ws_hydrated = self.wb["Span_Hydrated"]
        
        # Track duplicate spans for context/indexing
        # Key: span_text, Value: count encountered so far
        span_counter = {}
        
        for text, label, norm_val, field, evt_id in SPANS_TO_EXTRACT:
            # Handle duplicates for finding correct match
            # This generator logic assumes the list SPANS_TO_EXTRACT is ordered by appearance in text or logic
            # Since we can't guarantee order without scanning text, we rely on hydration logic later.
            # For this simple script, we set match_index if we know it's a repeat in the source list.
            
            # Simple heuristic: find unique text in input list
            # If text appears multiple times in SPANS_TO_EXTRACT, we need to be careful.
            # Here, "mucus" appears twice.
            
            current_count = span_counter.get(text, 0)
            match_idx = current_count # Use the count as the index for hydration logic if needed
            span_counter[text] = current_count + 1
            
            # Context logic (rudimentary)
            ctx = None
            if text == "mucus" and match_idx == 1:
                ctx = "partially occluded with" # Context for 2nd mucus
            elif text == "mucus" and match_idx == 0:
                ctx = "clean out the" # Broad context for 1st mucus

            # Create annotation row
            # source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, 
            # start_char, end_char, span_len, label, normalized_value, schema_field, event_id, 
            # is_negated, is_historical, time_anchor, reviewer, comments, hydration_status
            
            span_id = f"span_{len(self.spans_data) + 1:03d}"
            
            row = {
                "source_file": SOURCE_FILE,
                "note_id": NOTE_ID,
                "span_id": span_id,
                "section_type": "Procedure",
                "context_prefix": ctx,
                "span_text": text,
                "match_index": match_idx,
                "label": label,
                "normalized_value": norm_val,
                "schema_field": field,
                "event_id": evt_id,
                "is_negated": "FALSE",
                "is_historical": "FALSE",
                "time_anchor": "",
                "reviewer": "Auto",
                "comments": "",
                "hydration_status": "needs_hydration"
            }
            
            # Write to Span_Annotations (No offsets)
            anno_row = [
                row["source_file"], row["note_id"], row["span_id"], row["section_type"],
                row["context_prefix"], row["span_text"], row["match_index"],
                "", "", f'=LEN(INDIRECT("RC[-4]",0))', # Excel formula for len
                row["label"], row["normalized_value"], row["schema_field"], row["event_id"],
                row["is_negated"], row["is_historical"], row["time_anchor"],
                row["reviewer"], row["comments"], row["hydration_status"]
            ]
            ws_anno.append(anno_row)
            
            # Hydrate
            start, end, status = find_offsets_anchor_first(NOTE_TEXT, text, ctx, match_idx)
            row["start_char"] = start
            row["end_char"] = end
            row["hydration_status"] = status
            self.spans_data.append(row)
            
            # Write to Span_Hydrated
            hyd_row = [
                row["source_file"], row["note_id"], row["span_id"], row["section_type"],
                row["context_prefix"], row["span_text"], row["match_index"],
                start if start is not None else "", end if end is not None else "", 
                (end - start) if (start is not None and end is not None) else "",
                row["label"], row["normalized_value"], row["schema_field"], row["event_id"],
                row["is_negated"], row["is_historical"], row["time_anchor"],
                row["reviewer"], row["comments"], row["hydration_status"]
            ]
            ws_hydrated.append(hyd_row)

    def populate_event_log(self):
        ws = self.wb["Event_Log"]
        # Columns: source_file, note_id, event_id, event_type, method, anatomy_target, device, needle_gauge, stations, counts, measurements, specimens, findings, is_historical, reviewer, comments, device_size, device_material, outcome_airway_lumen_pre, outcome_airway_lumen_post, outcome_symptoms, outcome_pleural, outcome_complication
        
        events = {
            "evt_01": {
                "type": "therapeutic_aspiration",
                "method": "aspiration",
                "anatomy": "Trachea, Right Mainstem, Bronchus Intermedius, Left Mainstem, Carinas",
                "findings": "mucus, mucus plug, necrotic tissue",
                "outcome_post": "Secretions cleared, necrotic tissue improved",
            },
            "evt_02": {
                "type": "therapeutic_aspiration",
                "method": "rinse/clean",
                "anatomy": "RML",
                "device": "stent",
                "findings": "stent occluded with mucus",
                "outcome_post": "cleaned"
            }
        }
        
        for eid, data in events.items():
            row = [
                SOURCE_FILE, NOTE_ID, eid, data.get("type"), data.get("method"),
                data.get("anatomy"), data.get("device", ""), "", "", "", "",
                "", data.get("findings"), "FALSE", "Auto", "",
                "", "", # Size/Material
                "", data.get("outcome_post", ""), # Lumen Pre, Lumen Post
                "", "", # Symptoms, Pleural
                "" # Complications (global usually)
            ]
            ws.append(row)
            
        # Global complications event
        ws.append([SOURCE_FILE, NOTE_ID, "global", "complication_check", "", "", "", "", "", "", "", "", "", "FALSE", "Auto", "", "", "", "", "", "", "", "None"])

    def populate_v3_events(self):
        ws = self.wb["V3_Procedure_Events"]
        # note_id, event_id, type, target.anatomy_type, target.location.lobe, target.location.segment, target.station, lesion.type, lesion.size_mm, method, devices_json, measurements_json, specimens_json, findings_json, evidence_quote, stent.size, stent.material_or_brand, catheter.size_fr, outcomes.airway.lumen_pre, outcomes.airway.lumen_post, outcomes.symptoms, outcomes.pleural, outcomes.complications
        
        # Event 1
        ws.append([
            NOTE_ID, "evt_01", "therapeutic_aspiration", "airway", "", "", "", 
            "mucus", "", "aspiration", "[]", "[]", "[]", "['mucus', 'necrotic_tissue']", 
            "Successful therapeutic aspiration was performed...", "", "", "", "", "cleared/improved", "", "", ""
        ])
        
        # Event 2
        ws.append([
            NOTE_ID, "evt_02", "stent_maintenance", "airway", "RML", "", "", 
            "obstruction", "", "rinse", "['stent']", "[]", "[]", "['stent_occlusion']", 
            "right middle lobe stent was partially occluded... rinsed", "", "", "", "", "cleaned", "", "", ""
        ])

    def generate_json(self):
        ws = self.wb["V3_Registry_JSON"]
        
        # Construct JSON object
        registry_data = {
            "schema_version": "3.0",
            "note_id": NOTE_ID,
            "procedures": [
                {
                    "event_id": "evt_01",
                    "type": "therapeutic_aspiration",
                    "target": {"anatomy_type": "airway", "details": ["Trachea", "Mainstems", "Carinas"]},
                    "method": "aspiration",
                    "findings": ["mucus", "mucus_plug", "necrotic_tissue"],
                    "outcomes": {"airway": {"lumen_post": "cleared"}}
                },
                {
                    "event_id": "evt_02",
                    "type": "therapeutic_aspiration",
                    "target": {"anatomy_type": "airway", "location": {"lobe": "RML"}},
                    "devices": [{"type": "stent", "action": "cleaned"}],
                    "findings": ["stent_occlusion"],
                    "outcomes": {"airway": {"lumen_post": "cleaned"}}
                }
            ],
            "no_immediate_complications": True
        }
        
        json_str = json.dumps(registry_data, indent=2)
        # Write to single cell
        ws.cell(row=1, column=1, value=json_str)

    def save(self):
        self.wb.save(OUTPUT_PATH)
        print(f"Generated {OUTPUT_PATH}")

# ==========================================
# EXECUTION
# ==========================================
if __name__ == "__main__":
    generator = RegistryWorkbookGenerator()
    generator.load_template()
    generator.populate_note_text()
    generator.populate_note_index()
    generator.populate_spans()
    generator.populate_event_log()
    generator.populate_v3_events()
    generator.generate_json()
    generator.save()