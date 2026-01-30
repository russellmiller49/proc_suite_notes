import json
import re
import os
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

# =============================================================================
# INPUT DATA
# =============================================================================
NOTE_ID = "note_097"
SOURCE_FILE = "note_097.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_097 SOURCE_FILE: note_097.txt INDICATION FOR OPERATION:  [REDACTED]is a 56 year old-year-old male who presents with with bilateral lung transplant and complication of anastomosis dehiscence, ischemic lung injury, and bronchial stenosis.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
PROCEDURE:  
31646 Therapeutic aspiration subsequent episodes
31622 Dx bronchoscope/cell washing          
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT: 
Disposable Bronchoscope
PROCEDURE IN DETAIL:
A timeout was performed (confirming the patient's name, procedure type, and procedure location).  Sedation initiated.
The Disposable Bronchoscope was advanced for airway examination.  Endobronchial topical lidocaine applied to the main carina, right carina 1, and left carina 2.
Initial Airway Inspection Findings:
Tracheostomy tube in good position.
Distal trachea and main carina normal.  
RMS dehiscence continues to be healed and remains closed.
Stable fibrinous tan exudate/granulation tissue in the donor RMS, proximal RUL bronchus, BI, overlying RML take-off, and overlying RB6 take-off.
Prior area of fullness thickness erosion/ulceration along RUL bronchus anterior wall remains fully covered by fibrinous exudate/granulation tissue.  RB1-2 normal.
RB3 mildly stenotic with swirling granulation tissue/scar forming around the segmental airway take-off.
The RML with stent (Aero 8x15mm covered metallic stent) in good position and fully patent.
Able to traverse scope in the RML stent, but not into the segmental airways.
Able to see RB4-5 from stent and these segments were fully patent.  RB6 patent.
The RLL basilar segments RB7-10 appear healthy and patent.
LMS anastomosis intact with visible sutures - overlying granulation tissue causing mild to moderate stenosis.
Prior area of fullness thickness erosion/ulceration along LUL bronchus medial wall remains fully covered by fibrinous exudate/granulation tissue.
Lingula mild stentotic but fully patent.  Able to traverse scope into lingula.  LB1-5 are patent.
LLL bronchus and segments appear healthy.  LB6-10 are patent.
Successful therapeutic aspiration was performed to clean out the RMS, BI, RLL, LMS, and LLL from moderate thin mucus/secretions.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was returned to ICU in stable condition.
ESTIMATED BLOOD LOSS:   None
COMPLICATIONS:    None
SPECIMEN(S): 
None
IMPRESSION/PLAN: [REDACTED]is a 56 year old-year-old male who presents for bronchoscopy for bronchial stenosis airway evaluation.
-Plan for repeat bronchoscopy by IP in 2 weeks.
-OK to restart anticoagulation."""

# =============================================================================
# CONFIGURATION
# =============================================================================

# Procedure Flags (30 total)
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,  # 31622
    "bal": 0,
    "bronchial_wash": 1,  # 31622 / cell washing
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,  # 31646
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# Anchors: (span_text, label, normalized_value, schema_field, event_id, context_prefix)
ANCHORS = [
    # Event 1: General Procedure / Instrument
    ("Disposable Bronchoscope", "DEV_INSTRUMENT", "Disposable Bronchoscope", "method", "evt1", None),
    
    # Event 2: Right Main Stem / RUL Inspection
    ("RMS", "ANAT_AIRWAY", "RMS", "target", "evt2", "healed and remains closed"),
    ("granulation tissue", "OBS_LESION", "granulation tissue", "findings", "evt2", "Stable fibrinous tan exudate/"),
    ("dehiscence", "OBS_LESION", "dehiscence", "findings", "evt2", None),
    ("RUL bronchus", "ANAT_AIRWAY", "RUL", "target", "evt2", "proximal"),
    
    # Event 3: RB3 Stenosis
    ("RB3", "ANAT_AIRWAY", "RB3", "target", "evt3", None),
    ("mildly stenotic", "OBS_LESION", "mild stenosis", "findings", "evt3", "RB3"),
    ("granulation tissue/scar", "OBS_LESION", "granulation tissue/scar", "findings", "evt3", "swirling"),
    
    # Event 4: RML Stent Check
    ("RML", "ANAT_AIRWAY", "RML", "target", "evt4", "The"),
    ("stent", "DEV_STENT", "stent", "device", "evt4", "The RML with"),
    ("Aero 8x15mm covered metallic stent", "DEV_STENT_MATERIAL", "Aero Covered Metallic", "stent.material", "evt4", None),
    ("8x15mm", "DEV_STENT_SIZE", "8x15mm", "stent.size", "evt4", None),
    ("fully patent", "OBS_LESION", "patent", "findings", "evt4", "good position and"),
    
    # Event 5: Left Main Stem Inspection
    ("LMS anastomosis", "ANAT_AIRWAY", "LMS", "target", "evt5", None),
    ("visible sutures", "OBS_LESION", "visible sutures", "findings", "evt5", None),
    ("mild to moderate stenosis", "OBS_LESION", "mild to moderate stenosis", "findings", "evt5", "causing"),
    
    # Event 6: LUL / Lingula Inspection
    ("LUL bronchus", "ANAT_AIRWAY", "LUL", "target", "evt6", "erosion/ulceration along"),
    ("Lingula", "ANAT_AIRWAY", "Lingula", "target", "evt6", None),
    ("mild stentotic", "OBS_LESION", "mild stenosis", "findings", "evt6", "Lingula"),
    
    # Event 7: Therapeutic Aspiration
    ("therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", "method", "evt7", "Successful"),
    ("clean out", "PROC_ACTION", "clean out", "method", "evt7", None),
    ("moderate thin mucus/secretions", "OBS_LESION", "mucus/secretions", "findings", "evt7", "from"),
    ("RMS", "ANAT_AIRWAY", "RMS", "target", "evt7", "clean out the"),
    ("BI", "ANAT_AIRWAY", "BI", "target", "evt7", "RMS,"),
    ("RLL", "ANAT_AIRWAY", "RLL", "target", "evt7", "BI,"),
    ("LMS", "ANAT_AIRWAY", "LMS", "target", "evt7", "RLL,"),
    ("LLL", "ANAT_AIRWAY", "LLL", "target", "evt7", "LMS, and"),
    
    # Event 8: Outcome
    ("No immediate complications", "OUTCOME_COMPLICATION", "none", "outcomes.complications", "evt8", None),
]

# Events Data
EVENTS = [
    {
        "event_id": "evt1",
        "type": "diagnostic_bronchoscopy",
        "method": "Disposable Bronchoscope",
        "target_anatomy": "Trachea/Carina",
        "findings": "Normal"
    },
    {
        "event_id": "evt2",
        "type": "diagnostic_bronchoscopy",
        "target_anatomy": "RMS, RUL",
        "findings": "Dehiscence healed, granulation tissue present"
    },
    {
        "event_id": "evt3",
        "type": "diagnostic_bronchoscopy",
        "target_anatomy": "RB3",
        "findings": "Mildly stenotic, granulation tissue"
    },
    {
        "event_id": "evt4",
        "type": "diagnostic_bronchoscopy",
        "target_anatomy": "RML",
        "device": "Aero 8x15mm covered metallic stent",
        "findings": "Patent, good position"
    },
    {
        "event_id": "evt5",
        "type": "diagnostic_bronchoscopy",
        "target_anatomy": "LMS",
        "findings": "Visible sutures, mild-moderate stenosis"
    },
    {
        "event_id": "evt6",
        "type": "diagnostic_bronchoscopy",
        "target_anatomy": "LUL, Lingula",
        "findings": "LUL erosion covered, Lingula mild stenotic"
    },
    {
        "event_id": "evt7",
        "type": "therapeutic_aspiration",
        "target_anatomy": "RMS, BI, RLL, LMS, LLL",
        "findings": "Moderate thin mucus/secretions removed"
    },
    {
        "event_id": "evt8",
        "type": "outcome",
        "findings": "No immediate complications"
    }
]

# =============================================================================
# LOGIC
# =============================================================================

def create_workbook():
    if not os.path.exists(TEMPLATE_PATH):
        # Create a dummy template if missing (for standalone testing),
        # but in production, we assume the file exists.
        wb = openpyxl.Workbook()
        wb.create_sheet("Note_Text")
        wb.create_sheet("Note_Index")
        wb.create_sheet("Span_Annotations")
        wb.create_sheet("Span_Hydrated")
        wb.create_sheet("Event_Log")
        wb.create_sheet("V3_Procedure_Events")
        wb.create_sheet("V3_Registry_JSON")
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # 1. Note_Text
    ws_text = wb["Note_Text"]
    if ws_text.max_row == 1:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws_index = wb["Note_Index"]
    header = [
        "source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes",
        "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", "tbna_conventional",
        "linear_ebus", "radial_ebus", "navigational_bronchoscopy", "transbronchial_biopsy", "transbronchial_cryobiopsy",
        "therapeutic_aspiration", "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
        "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", "bronchial_thermoplasty",
        "whole_lung_lavage", "rigid_bronchoscopy", "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy",
        "pleurodesis", "pleural_biopsy", "fibrinolytic_therapy"
    ]
    
    # Ensure header
    if ws_index.max_row == 1 and ws_index.cell(1,1).value != "source_file":
        ws_index.append(header)
    
    row_data = [SOURCE_FILE, NOTE_ID, "", "", "", "", "Pending", ""]
    # Append flags in order
    for flag in header[8:]:
        row_data.append(PROCEDURE_FLAGS.get(flag, 0))
    ws_index.append(row_data)

    # 3. Span_Annotations (Anchor First)
    ws_span = wb["Span_Annotations"]
    if ws_span.max_row == 1:
        ws_span.append([
            "source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text",
            "match_index", "start_char", "end_char", "span_len", "label", "normalized_value",
            "schema_field", "event_id", "is_negated", "is_historical", "time_anchor", "reviewer",
            "comments", "hydration_status"
        ])
    
    span_rows = []
    
    for idx, anchor in enumerate(ANCHORS):
        span_text, label, norm_val, schema_f, evt_id, ctx = anchor
        span_id = f"{NOTE_ID}_s{idx+1:03d}"
        
        # Determine match_index automatically during hydration, but here we define intent
        # For simplicity in this script, we pass 0 for match_index if context is None,
        # otherwise we assume hydration will find it.
        
        row = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure", ctx, span_text,
            "", "", "", f"=LEN(F{ws_span.max_row + idx + 2})", label, norm_val,
            schema_f, evt_id, "FALSE", "FALSE", "", "Auto", "", "needs_hydration"
        ]
        ws_span.append(row)
        span_rows.append(row) # Store for hydration mapping

    # 4. Span_Hydrated
    ws_hyd = wb["Span_Hydrated"]
    if ws_hyd.max_row == 1:
        # Copy header from Span_Annotations
        headers = [c.value for c in ws_span[1]]
        ws_hyd.append(headers)

    hydrated_rows = []
    
    for s_row in span_rows:
        # Unpack row (indices based on header above)
        # 0:source, 1:note, 2:id, 3:sect, 4:ctx, 5:text, 6:idx, 7:start, 8:end...
        ctx_prefix = s_row[4]
        text = s_row[5]
        
        start_char = -1
        end_char = -1
        status = "ambiguous"
        
        matches = [m.start() for m in re.finditer(re.escape(text), NOTE_TEXT)]
        
        if len(matches) == 1:
            start_char = matches[0]
            status = "hydrated_unique"
        elif len(matches) > 1:
            if ctx_prefix:
                # Try to find context before
                found = False
                for m in matches:
                    snippet = NOTE_TEXT[max(0, m - 120):m]
                    if ctx_prefix in snippet:
                        start_char = m
                        status = "hydrated_prefix_window"
                        found = True
                        break
                if not found:
                    status = f"ambiguous_count={len(matches)}_ctx_not_found"
            else:
                # Default to first if no context provided (or specify strict manual index)
                # In this generator, we assume first occurrence if no context given for now
                start_char = matches[0]
                status = "hydrated_first_default"
        else:
            status = "not_found"
            
        if start_char != -1:
            end_char = start_char + len(text)
        
        # Create new row for hydrated sheet
        h_row = list(s_row)
        h_row[7] = start_char if start_char != -1 else ""
        h_row[8] = end_char if end_char != -1 else ""
        h_row[19] = status
        
        ws_hyd.append(h_row)
        hydrated_rows.append(h_row)

    # 5. Event_Log
    ws_event = wb["Event_Log"]
    if ws_event.max_row == 1:
        ws_event.append([
            "source_file", "note_id", "event_id", "event_type", "method", "anatomy_target", "device",
            "needle_gauge", "stations", "counts", "measurements", "specimens", "findings",
            "is_historical", "reviewer", "comments", "device_size", "device_material",
            "outcome_airway_lumen_pre", "outcome_airway_lumen_post", "outcome_symptoms",
            "outcome_pleural", "outcome_complication"
        ])

    for evt in EVENTS:
        row = [
            SOURCE_FILE, NOTE_ID, evt["event_id"], evt["type"], evt.get("method", ""),
            evt.get("target_anatomy", ""), evt.get("device", ""), "", "", "", "", "",
            evt.get("findings", ""), "FALSE", "Auto", "", "", "", "", "", "", "", 
            "none" if evt["type"] == "outcome" else ""
        ]
        ws_event.append(row)

    # 6. V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    if ws_v3.max_row == 1:
        ws_v3.append([
            "note_id", "event_id", "type", "target.anatomy_type", "target.location.lobe",
            "target.location.segment", "target.station", "lesion.type", "lesion.size_mm",
            "method", "devices_json", "measurements_json", "specimens_json", "findings_json",
            "evidence_quote", "stent.size", "stent.material_or_brand", "catheter.size_fr",
            "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post", "outcomes.symptoms",
            "outcomes.pleural", "outcomes.complications"
        ])

    for evt in EVENTS:
        # Simplified mapping for V3
        row = [
            NOTE_ID, evt["event_id"], evt["type"], "airway", "", "", "", "", "",
            evt.get("method", ""), "", "", "", json.dumps([evt.get("findings", "")]), "",
            "8x15mm" if "stent" in evt.get("device", "") else "",
            "Aero Covered Metallic" if "stent" in evt.get("device", "") else "",
            "", "", "", "", "", "none" if evt["type"] == "outcome" else ""
        ]
        ws_v3.append(row)

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 1:
        ws_json.append(["schema_version", "note_id", "json_output"])
    
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": EVENTS,
        "no_immediate_complications": True
    }
    
    ws_json.append(["3.0", NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    create_workbook()