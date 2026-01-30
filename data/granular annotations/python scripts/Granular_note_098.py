import openpyxl
from openpyxl.utils import get_column_letter
import json
import os
import datetime
import re

# -------------------------------------------------------------------------
# INPUTS
# -------------------------------------------------------------------------
NOTE_ID = "note_098"
SOURCE_FILE = "note_098.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_098 SOURCE_FILE: note_098.txt INDICATION FOR OPERATION:  [REDACTED]is a 56 year old-year-old male who presents with bronchial stenosis.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: Bronchial stenosis
POSTOPERATIVE DIAGNOSIS:  Bronchial stenosis
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31622 Dx bronchoscope/cell washing          
31624 Dx bronchoscope/lavage (BAL)    
31625 Endobronchial Biopsy(s)
31630 Balloon dilation
31636 Dilate and bronchial stent initial bronchus
31638 Revision of tracheal/bronchial stent     
31640 Bronchoscopy with excision 
31635 Foreign body removal
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required) Due to small right mainstem diameter difficult to use therapeutic bronchoscope and at times during the case 
the disposable bronchoscope was introduced into the mouth and into the airway along side the therapeutic bronchoscope going through the tracheostomy tube.
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required a significant amount of time and expertise to attempt dilation of stent and replace/revise stent in complicated airway with stenosis, fibrinous exudates.
This resulted in >40% increased work due to Increased intensity, Time, Technical difficulty of procedure, and Physical and mental effort required.
Apply to: 31636 Dilate and bronchial stent initial bronchus. 
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
Disposable Bronchoscope
ESTIMATED BLOOD LOSS:   Minimum
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
Initial Airway Inspection Findings:
The bronchoscope was advanced into the existing tracheostomy tube.
A total of 4 mL of 2% lidocaine was instilled onto the main carina.
The airways were examined to the subsegmental level bilaterally. 
The right anastomosis was without dehiscence but with white-yellow exudate along the anterior portion of the RMSB and extending down to the level of the RML ostium.
The left anastomosis is intact with some area of granulation tissue without significant stenosis.
There is white-yellow fibrinous plaque along the LMSB extending to the ostium of the lingular bronchus.
There is a stent in the RML extending past the ostium of the RML.
The stent was in appropriate position but the distal portion did not appear to be fully extended.
There was a moderate amount of thick, white mucus in the LLL. This was therapeutically aspirated.
Bronchial alveolar lavage was performed at Posterior-Basal Segment of LLL (LB10).
Instilled 60 cc of NS, suction returned with 15 cc of NS.  Samples sent for Cell Count and Microbiology (Cultures/Viral/Fungal).
Balloon dilation with CRE 8, 9, 10 balloon was performed  of the RML bronchus stent.
The stent was dilated x2 to 9 mm for a total of 1 minute each dilation.
The stent did not increase in diameter but appeared to be too small for the airway and migrated distally.
Using the therapeutic bronchoscope via the tracheostomy stoma and disposable bronchoscope via the mouth attempt to reposition the stent into the RML bronchus with forceps.
The foreign body/stent was grasped with forceps and removed en bloc from the RML via the tracheostomy stoma.
Examination of the stent showed that the stent did not dilate past 7 mm.
The bronchoscope was introduced into the mouth advanced along the tongue and to the level of the vocal cords.
The cuff of the tracheostomy tube was partially deflated and the bronchoscope was advanced through the vocal cords into the trachea and past the tracheostomy tube.
A jag wire was introduced into the RML. 
Under direct visualization a 8 x 15 mm Aero SEM stent was deployed in to the RML bronchus.
The stent was revised with forceps. Position was adequate and distal airways were patent.
The lingular bronchus was dilated to 8 mm with improvement in diameter was dilation
Fibrinous/yellow-white tumor was removed with forceps at the lingular ostium and LMSB.
There was no evidence of active bleeding and the bronchoscope was removed. 
The patient tolerated the procedure well.
There were no immediate complications.  At the conclusion of the operation, the was transitioned to mechanical ventilation and back to his ICU room to recover.
SPECIMEN(S): 
BAL - RML - cell count and culture
Stent - for culture
LMSB fibrinous tissue - culture and pathology
IMPRESSION/PLAN: [REDACTED]is a 56 year old-year-old male who presents for bronchoscopy for bronchial stenosis.
[ ] f/u culture and pathology results
[ ] f/u bronchoscopy in 1 week to evaluate stent and lingular bronchus
[ ] restart bivaliruding at 18:00 today"""

# -------------------------------------------------------------------------
# CONFIGURATION
# -------------------------------------------------------------------------

PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 1,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 1, # Code 31625 listed
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 1,
    "airway_dilation": 1,
    "airway_stent": 1,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 1, # Code 31640 Excision of tumor
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# -------------------------------------------------------------------------
# DATA EXTRACTION
# -------------------------------------------------------------------------

# Define Spans
# Format: (span_text, label, normalized_value, schema_field, event_id, context_prefix)
spans_data = [
    # Event 1: Therapeutic Aspiration LLL
    ("LLL", "ANAT_LUNG_LOC", "LLL", "target.location.lobe", "evt_01", "moderate amount of thick, white mucus in the "),
    ("thick, white mucus", "OBS_LESION", "mucus", "findings.description", "evt_01", "moderate amount of "),
    ("therapeutically aspirated", "PROC_ACTION", "aspiration", "method", "evt_01", "This was "),
    
    # Event 2: BAL LLL LB10
    ("Bronchial alveolar lavage", "PROC_METHOD", "BAL", "method", "evt_02", None),
    ("Posterior-Basal Segment of LLL", "ANAT_LUNG_LOC", "LLL Posterior Basal", "target.location.segment", "evt_02", None),
    ("LB10", "ANAT_LN_STATION", "LB10", "target.station", "evt_02", None),
    ("Instilled 60 cc", "MEAS_VOL", "60", "measurements.volume", "evt_02", None),
    ("returned 15 cc", "MEAS_VOL", "15", "measurements.volume", "evt_02", None),
    
    # Event 3: Balloon Dilation RML Stent (Failed)
    ("Balloon dilation", "PROC_METHOD", "balloon dilation", "method", "evt_03", "31630 "),
    ("CRE 8, 9, 10 balloon", "DEV_INSTRUMENT", "CRE Balloon 8-9-10mm", "devices.catheter", "evt_03", None),
    ("RML bronchus", "ANAT_AIRWAY", "RML", "target.location.lobe", "evt_03", "performed  of the "),
    ("dilated x2 to 9 mm", "PROC_ACTION", "dilation", "measurements.size", "evt_03", None),
    ("stent did not increase in diameter", "OUTCOME_AIRWAY_LUMEN_POST", "no change", "outcomes.airway.lumen_post", "evt_03", None),
    
    # Event 4: Foreign Body Removal (Old Stent)
    ("foreign body/stent", "OBS_LESION", "stent", "lesion.type", "evt_04", "The "),
    ("grasped with forceps", "PROC_ACTION", "grasped", "method", "evt_04", None),
    ("removed en bloc", "PROC_ACTION", "removal", "method", "evt_04", None),
    ("RML", "ANAT_AIRWAY", "RML", "target.location.lobe", "evt_04", "removed en bloc from the "),
    
    # Event 5: Stent Placement RML (New)
    ("8 x 15 mm", "DEV_STENT_SIZE", "8x15mm", "stent.size", "evt_05", "visualization a "),
    ("Aero SEM stent", "DEV_STENT", "Aero SEM", "stent.material_or_brand", "evt_05", None),
    ("deployed", "PROC_ACTION", "deployment", "method", "evt_05", "stent was "),
    ("RML bronchus", "ANAT_AIRWAY", "RML", "target.location.lobe", "evt_05", "deployed in to the "),
    
    # Event 6: Stent Revision
    ("stent", "DEV_STENT", "stent", "target.device", "evt_06", "The "),
    ("revised", "PROC_ACTION", "revision", "method", "evt_06", "The stent was "),
    ("forceps", "DEV_INSTRUMENT", "forceps", "devices.instrument", "evt_06", "revised with "),
    
    # Event 7: Dilation Lingula
    ("lingular bronchus", "ANAT_AIRWAY", "Lingula", "target.location.lobe", "evt_07", "The "),
    ("dilated to 8 mm", "PROC_ACTION", "dilation", "measurements.size", "evt_07", None),
    ("improvement in diameter", "OUTCOME_AIRWAY_LUMEN_POST", "improved", "outcomes.airway.lumen_post", "evt_07", None),
    
    # Event 8: Excision/Debulking
    ("Fibrinous/yellow-white tumor", "OBS_LESION", "tumor", "lesion.type", "evt_08", None),
    ("removed with forceps", "PROC_ACTION", "excision", "method", "evt_08", "tumor was "),
    ("lingular ostium", "ANAT_AIRWAY", "Lingular Ostium", "target.location.lobe", "evt_08", None),
    ("LMSB", "ANAT_AIRWAY", "LMSB", "target.location.lobe", "evt_08", "ostium and "),
    
    # Event 9: Complications
    ("No immediate complications", "OUTCOME_COMPLICATION", "none", "outcomes.complications", "evt_09", "There were "),
]

# Define Events
# Format: (event_id, type, target, method, device, outcome_pre, outcome_post, complications)
events_data = [
    {
        "event_id": "evt_01",
        "type": "Therapeutic Aspiration",
        "target": "LLL",
        "method": "Aspiration",
        "findings": "thick white mucus"
    },
    {
        "event_id": "evt_02",
        "type": "BAL",
        "target": "LLL Posterior Basal",
        "method": "Lavage",
        "measurements": "Instilled 60cc, Returned 15cc"
    },
    {
        "event_id": "evt_03",
        "type": "Airway Dilation",
        "target": "RML Stent",
        "method": "Balloon Dilation",
        "device": "CRE 8-9-10 balloon",
        "outcome_post": "No increase, stent migrated"
    },
    {
        "event_id": "evt_04",
        "type": "Foreign Body Removal",
        "target": "RML",
        "method": "Forceps Removal",
        "findings": "Migrated Stent"
    },
    {
        "event_id": "evt_05",
        "type": "Airway Stent",
        "target": "RML",
        "method": "Deployment",
        "device": "Aero SEM Stent",
        "size": "8x15mm"
    },
    {
        "event_id": "evt_06",
        "type": "Stent Revision",
        "target": "RML Stent",
        "method": "Forceps Revision"
    },
    {
        "event_id": "evt_07",
        "type": "Airway Dilation",
        "target": "Lingular Bronchus",
        "method": "Dilation",
        "measurements": "to 8mm",
        "outcome_post": "Improved"
    },
    {
        "event_id": "evt_08",
        "type": "Tumor Debulking",
        "target": "Lingular Ostium/LMSB",
        "method": "Forceps Excision",
        "findings": "Fibrinous tumor"
    },
    {
        "event_id": "evt_09",
        "type": "Complications",
        "outcome_complication": "None"
    }
]

# -------------------------------------------------------------------------
# HELPER FUNCTIONS
# -------------------------------------------------------------------------

def hydrate_offsets(text, span_text, context_prefix=None):
    """
    Finds start/end char offsets.
    Logic:
    1. If unique in text -> use it.
    2. If context_prefix provided -> search for prefix + span within window.
    3. Fallback -> first occurrence (or ambiguous).
    """
    count = text.count(span_text)
    if count == 0:
        return None, None, "not_found"
    
    if count == 1:
        start = text.find(span_text)
        return start, start + len(span_text), "hydrated_unique"
    
    if context_prefix:
        # Search for context + span
        # Simple moving window or regex approach
        # Normalize spaces for robust search
        # Find all occurrences of span
        starts = [m.start() for m in re.finditer(re.escape(span_text), text)]
        
        # Check preceding window for each
        for s in starts:
            window_start = max(0, s - 150)
            preceding_text = text[window_start:s]
            if context_prefix in preceding_text:
                return s, s + len(span_text), "hydrated_prefix_window"
    
    # Fallback to first if ambiguous logic not fully implemented for index
    start = text.find(span_text)
    return start, start + len(span_text), f"ambiguous_count={count}"

# -------------------------------------------------------------------------
# EXECUTION
# -------------------------------------------------------------------------

def generate_workbook():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        print(f"Error: Template {TEMPLATE_PATH} not found.")
        return

    # 1. Note_Text
    ws_text = wb["Note_Text"]
    # Check if empty, if so add header (assuming template has it, but safe to append)
    if ws_text.max_row == 1 and ws_text.cell(1,1).value is None:
        ws_text.append(["note_id", "source_file", "note_text"])
    
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws_index = wb["Note_Index"]
    # Assuming header exists.
    # Columns: source_file, note_id, encounter_id, procedure_date, site, reviewer, status, free_text_notes, [flags...]
    
    flag_values = [PROCEDURE_FLAGS.get(key, 0) for key in PROCEDURE_FLAGS]
    # Ordering of flags in template usually matches the dict keys if sorted or defined. 
    # To be safe, let's map explicitly based on the known list in prompt instructions.
    
    ordered_flags = [
        "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
        "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy",
        "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration",
        "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
        "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation",
        "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
        "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis",
        "pleural_biopsy", "fibrinolytic_therapy"
    ]
    
    row_values = [
        SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Success", ""
    ] + [PROCEDURE_FLAGS[f] for f in ordered_flags]
    
    ws_index.append(row_values)

    # 3. Span_Annotations & 4. Span_Hydrated
    ws_anno = wb["Span_Annotations"]
    ws_hydra = wb["Span_Hydrated"]
    
    # Headers should exist.
    
    for span in spans_data:
        txt, label, norm, schema, evt, ctx = span
        
        # Calculate offsets for Hydrated
        start, end, status = hydrate_offsets(NOTE_TEXT, txt, ctx)
        
        # Common Row
        # source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, 
        # start_char, end_char, span_len, label, normalized_value, schema_field, event_id, 
        # is_negated, is_historical, time_anchor, reviewer, comments, hydration_status
        
        span_len = f"=LEN(INDIRECT(\"F\" & ROW()))" # Excel formula
        
        # Anno Row (Blank offsets)
        anno_row = [
            SOURCE_FILE, NOTE_ID, "", "", ctx, txt, "", 
            "", "", span_len, label, norm, schema, evt,
            "FALSE", "FALSE", "", "", "", "needs_hydration"
        ]
        ws_anno.append(anno_row)
        
        # Hydra Row (Filled offsets)
        hydra_row = [
            SOURCE_FILE, NOTE_ID, "", "", ctx, txt, "", 
            start, end, len(txt) if start is not None else 0, label, norm, schema, evt,
            "FALSE", "FALSE", "", "", "", status
        ]
        ws_hydra.append(hydra_row)

    # 5. Event_Log
    ws_event = wb["Event_Log"]
    for e in events_data:
        # source_file, note_id, event_id, event_type, method, anatomy_target, device, needle_gauge, stations, counts, measurements, specimens, findings, is_historical, reviewer, comments, device_size, device_material, outcome_airway_lumen_pre, outcome_airway_lumen_post, outcome_symptoms, outcome_pleural, outcome_complication
        
        row = [
            SOURCE_FILE, NOTE_ID, e.get("event_id"), e.get("type"), e.get("method"),
            e.get("target"), e.get("device"), "", "", "", e.get("measurements"),
            "", e.get("findings"), "FALSE", "", "",
            e.get("size"), "", # size, material
            "", e.get("outcome_post"), # lumen pre/post
            "", "", e.get("outcome_complication") # symp, pleural, comp
        ]
        ws_event.append(row)

    # 6. V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    # note_id, event_id, type, target.anatomy_type, target.location.lobe, target.location.segment, target.station, lesion.type, lesion.size_mm, method, devices_json, measurements_json, specimens_json, findings_json, evidence_quote, stent.size, stent.material_or_brand, catheter.size_fr, outcomes.airway.lumen_pre, outcomes.airway.lumen_post, outcomes.symptoms, outcomes.pleural, outcomes.complications
    
    for e in events_data:
        row = [
            NOTE_ID, e.get("event_id"), e.get("type"),
            "Airway" if "Bronchus" in str(e.get("target")) or "RML" in str(e.get("target")) else "", # anatomy_type
            e.get("target"), "", "", # lobe, segment, station
            "", "", # lesion type, size
            e.get("method"),
            json.dumps({"device": e.get("device")}) if e.get("device") else "", # devices_json
            json.dumps({"meas": e.get("measurements")}) if e.get("measurements") else "", # meas_json
            "", "", "", # spec, find, quote
            e.get("size"), "", "", # stent size, mat, cath size
            "", e.get("outcome_post"), "", "", e.get("outcome_complication")
        ]
        ws_v3.append(row)

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 1:
        ws_json.append(["schema_version", "note_id", "json_output"])
    
    # Construct JSON
    registry_data = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "procedures": events_data,
        "no_immediate_complications": True
    }
    
    ws_json.append(["v3.0", NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()