import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os
from datetime import datetime

# =============================================================================
# INPUT CONFIGURATION
# =============================================================================
NOTE_ID = "note_099"
SOURCE_FILE = "note_099.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_099 SOURCE_FILE: note_099.txt INDICATION FOR OPERATION:  [REDACTED]is a 56 year old-year-old male who presents with bilateral lung transplant with complication of anastomosis dehiscence, ischemic lung injury, and bronchial stenosis.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
PROCEDURE:  
31899 Unlisted Procedure (Trach Change with Mature Tract or Procedure NOS)
31646 Therapeutic aspiration subsequent episodes
31622 Dx bronchoscope/cell washing          
31624 Dx bronchoscope/lavage (BAL)    
31625 Endobronchial Biopsy(s)
31630 Balloon dilation
31636 Dilate and bronchial stent initial bronchus
31637 Dilate and bronchial stent additional bronchus
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
31635 Foreign body removal
22 Substantially greater work than normal (i.e., increased intensity, 
time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required an unusual amount of clinical expertise and use of simultaneous bronchoscopes due to the severity of the patient's profound airway injury.
This resulted in >50% increased work due to Increased intensity, Time, Technical difficulty of procedure, Severity of patient's condition, and Physical and mental effort required.
Apply to:
31625 Endobronchial Biopsy(s)
31636 Dilate and bronchial stent initial bronchus
31637 Dilate and bronchial stent additional bronchus
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
31635 Foreign body removal.
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
Disposable Bronchoscope - Slim scope
Disposable Bronchoscope - Regular scope
PROCEDURE IN DETAIL:
A timeout was performed (confirming the patient's name, procedure type, and procedure location).
Sedation initiated.
The Flexible Therapeutic Bronchoscope was advanced for airway examination.
Endobronchial topical lidocaine applied to the main carina, right carina 1, and left carina 2.
Initial Airway Inspection Findings:
Tracheostomy tube in good position.
Distal trachea and main carina normal.  
RMS dehiscence continues to be healed and remains closed.
Prior metallic hemoclip seen protruding from the posterior membrane near the RMS anastomosis site is no longer visible.
Further fibrinous exudate/granulation tissue in the donor RMS, proximal RUL bronchus, BI, overlying RML take-off, and overlying RB6 take-off.
This fibrinous exudate/granulation tissue has become more bulky and partially obstructing these airways.
Previously seen area of fullness thickness erosion/ulceration along RUL bronchus anterior wall remains fully covered by fibrinous exudate/granulation tissue.
RB1-2 normal.  RB3 mildly stenotic with swirling granulation tissue/scar forming around the segmental airway take-off.
The RML take-off is completely sealed and obstructed by the fibrinous exudate/granulation tissue.  RB4-5 not visible.
RB6 mildly stenotic due to surrounding granulation tissue.  The RLL basilar segments appear healthy.
LMS anastomosis intact with visible sutures - having increasing amounts of overlying granulation tissue with increasing - causing mild to moderate stenosis.
LUL bronchus with evidence of continued healing with fibrin exudates/granulation tissue.
No longer able to see underlying mediastinum/pulmonary artery along medial aspect of LUL.
This entire area continues to look improved from prior.  However, heaped granulation tissue forming around the lingula take-off and causing moderate stenosis.
LB1-5 are patent.  LLL bronchus and segments appear healthy.  LB6-10 are patent.
Successful therapeutic aspiration was performed to clean out the RMS, BI, RLL, LMS, and LLL from thin mucus.
Attention turned back to completely obstructed RML bronchus.  RML completely sealed off - covered by fibrinous exudate and granulation tissue.
Dilation was performed at RML bronchus.  Therapeutic bronchoscope was used to perform blunt dilation to 6.2 mm at the RML bronchus.
Able to achieve patent RML airway.  Copious tan-white secretions immediately started pouring out.
Successful therapeutic aspiration was performed to clean out the RML from copious white-tan mucus. RB4-5 segmental airways also patent afterward.
Bronchial alveolar lavage was performed at RML.  Instilled 40 cc of NS, suction returned with 25 cc of NS.
Samples sent for Cell Count and Microbiology (Cultures/Viral/Fungal).
A second slim disposable bronchoscope was advanced via oral approach into the trachea and past the tracheostomy tube cuff.
This 2nd bronchoscope assisted with dual-scope visualization during the following procedures.
The following stent (iCast fully covered 7mm x 22mm stent) was placed successfully deployed/placed in the RML bronchus.
Balloon dilation was performed at RML bronchus stent.  8/9/10 Elation balloon was used to perform dilation to 8 mm within the full RML bronchus stent.
Total 1 inflations with dilation time of 10 seconds.  Achieving 100% patency of the RML airway.
The proximal portion of the stent was flared using blunt dilation with tip of bronchoscope, blunt forceps, and serial balloon dilations with th Elation balloon and Fogarty 7 balloon.
Balloon dilation was performed at proximal portion of RML bronchus stent.
8/9/10 Elation balloon was used to perform serial dilations to 8 mm, 9 mm, and 10 mm at the proximal portion of RML bronchus stent.
Total 5 inflations with dilation time of 5-10 seconds each. 
Attention then turn to the stenotic lingula airway.
The following stent (iCast fully covered 7mm x 16mm stent) was deployed in the lingula bronchus.
However, stent migrated proximally and distal edge of stent became caught on LC1, so could not be advanced back into lingula despite multiple attempts with balloon and forceps.
Foreign body removal of iCast stent overlying lingula take-off was extracted using blunt forceps.
Balloon dilation was performed at lingula bronchus.  8/9/10 Elation balloon was used to perform dilation to 8 mm within the lingular bronchus.
Total 2 inflations with dilation time of 30 seconds each.  Improved patency of lingula bronchus.
Attention then turned to the bulky partially obstructing fibrinous exudate within the RMS and BI.
The slim disposable bronchoscope was withdrawn and a regular disposable bronchoscope was advanced via oral approach into the trachea and past the tracheostomy tube cuff.
Endobronchial biopsy was performed at RMS bronchus fibrinous exudate/debris.  Portions of lesion were successfully removed.
Samples sent for tissue culture and pathology. 
Endobronchial fibrinous exudate/debris at RMS bronchus was treated with the following modalities:
-With first bronchoscope, the fibrinous exudate/debris was stretched away from the RMS wall.
-With second bronchoscope, Erbe HybridKnife electrocautery was used to cut/debride the debris.
Able to excise/debride large portions of the debris/exudate in this manner.
Modality	Tools	Setting/Mode	Duration	Results
Electrocautery	Erbe HybridKnife	Effect 2, 40 Watts	1-5 second bursts	Excised large portions of RMS debris/exudate.
Prior to treatment, affected airway was note to be 60% patent.  After treatment, the airway was 70% patent.
During the above tissue debridement, a very small laceration to RC1 was made.
Small amount of bleeding from this site easily controlled and stopped with direct bronchoscopic pressure and 0.2mg epinephrine.
Final airway exam performed.  No active bleeding present.  Residual secretions were suctioned to clear.
The patient tolerated the procedure well.
There were no immediate complications.  At the conclusion of the operation, the patient was returned to ICU in stable condition.
ESTIMATED BLOOD LOSS:   Minimum
COMPLICATIONS:    None
SPECIMEN(S): 
RML bronchus BAL - cell count, cultures
RMS bronchus exudate/debris - tissue culture, pathology.
IMPRESSION/PLAN: [REDACTED]is a 56 year old-year-old male who presents for bronchoscopy for airway dilation, stent placement, and airway debridement.
-Successful dilation of RML bronchus - achieved patency.  Copious secretions behind blockage suctioned to clear.
-RML BAL performed.
-iCast fully covered stent (7mm x 22mm) successfully placed within RML bronchus, achieving full patency.
-Attempted stenting of lingula bronchus with stent malposition and needed to be removed.
Additional stents of the same type were not available.  Balloon dilation of lingula bronchus performed.
-Debridement of bulky partially obstructing fibrinous exudate/debris in the RMS bronchus.
-Will plan for repeat bronchoscopy with airway evaluation and likely lingula stent placement early next week - likely Tues, [REDACTED] as this is the soonest desired stents will arrive."""

# =============================================================================
# FLAG DETERMINATION LOGIC
# =============================================================================
flags = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 1,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 1,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 1, # Stent removal
    "airway_dilation": 1,
    "airway_stent": 1,
    "thermal_ablation": 1, # Electrocautery
    "tumor_debulking_non_thermal": 1, # Mechanical stretching explicitly mentioned
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# =============================================================================
# SPAN DATA (ANCHOR-FIRST)
# =============================================================================
# Fields: span_text, label, normalized_value, context_prefix (optional), event_id (optional), is_negated, comments
# Note: Schema field will be inferred from label if standard, else specified.

spans_data = [
    # --- Event 1: General Inspection/Aspiration ---
    {
        "span_text": "Flexible Therapeutic Bronchoscope",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "Flexible Bronchoscope",
        "event_id": "ev1",
        "context_prefix": "Sedation initiated.\nThe "
    },
    {
        "span_text": "Therapeutic aspiration",
        "label": "PROC_METHOD",
        "normalized_value": "aspiration",
        "event_id": "ev1",
        "context_prefix": "Successful "
    },
    {
        "span_text": "RMS",
        "label": "ANAT_AIRWAY",
        "normalized_value": "RMS",
        "event_id": "ev1",
        "context_prefix": "clean out the "
    },
    {
        "span_text": "BI",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Bronchus Intermedius",
        "event_id": "ev1",
        "context_prefix": "RMS, "
    },
    {
        "span_text": "RLL",
        "label": "ANAT_LUNG_LOC",
        "normalized_value": "RLL",
        "event_id": "ev1",
        "context_prefix": "BI, "
    },
    {
        "span_text": "LMS",
        "label": "ANAT_AIRWAY",
        "normalized_value": "LMS",
        "event_id": "ev1",
        "context_prefix": "RLL, "
    },
    {
        "span_text": "LLL",
        "label": "ANAT_LUNG_LOC",
        "normalized_value": "LLL",
        "event_id": "ev1",
        "context_prefix": "LMS, and "
    },

    # --- Event 2: RML Blunt Dilation & Aspiration ---
    {
        "span_text": "Dilation",
        "label": "PROC_METHOD",
        "normalized_value": "dilation",
        "event_id": "ev2",
        "context_prefix": "completely sealed off - covered by fibrinous exudate and granulation tissue.\n"
    },
    {
        "span_text": "RML bronchus",
        "label": "ANAT_AIRWAY",
        "normalized_value": "RML bronchus",
        "event_id": "ev2",
        "context_prefix": "Dilation was performed at "
    },
    {
        "span_text": "blunt dilation",
        "label": "PROC_METHOD",
        "normalized_value": "blunt dilation",
        "event_id": "ev2",
        "context_prefix": "Therapeutic bronchoscope was used to perform "
    },
    {
        "span_text": "6.2 mm",
        "label": "MEAS_AIRWAY_DIAM",
        "normalized_value": "6.2 mm",
        "event_id": "ev2",
        "context_prefix": "blunt dilation to "
    },
    {
        "span_text": "Therapeutic aspiration",
        "label": "PROC_METHOD",
        "normalized_value": "aspiration",
        "event_id": "ev2",
        "context_prefix": "started pouring out.\nSuccessful "
    },
    {
        "span_text": "RML",
        "label": "ANAT_AIRWAY",
        "normalized_value": "RML bronchus",
        "event_id": "ev2",
        "context_prefix": "clean out the "
    },
    {
        "span_text": "copious white-tan mucus",
        "label": "OBS_ROSE",
        "normalized_value": "copious mucus",
        "event_id": "ev2"
    },

    # --- Event 3: RML BAL ---
    {
        "span_text": "Bronchial alveolar lavage",
        "label": "PROC_METHOD",
        "normalized_value": "BAL",
        "event_id": "ev3"
    },
    {
        "span_text": "RML",
        "label": "ANAT_AIRWAY",
        "normalized_value": "RML bronchus",
        "event_id": "ev3",
        "context_prefix": "lavage was performed at "
    },
    {
        "span_text": "40 cc",
        "label": "MEAS_VOL",
        "normalized_value": "40 cc",
        "event_id": "ev3",
        "schema_field": "instilled_vol"
    },
    {
        "span_text": "25 cc",
        "label": "MEAS_VOL",
        "normalized_value": "25 cc",
        "event_id": "ev3",
        "schema_field": "return_vol"
    },

    # --- Event 4: RML Stenting ---
    {
        "span_text": "iCast fully covered 7mm x 22mm stent",
        "label": "DEV_STENT",
        "normalized_value": "iCast 7x22mm",
        "event_id": "ev4"
    },
    {
        "span_text": "RML bronchus",
        "label": "ANAT_AIRWAY",
        "normalized_value": "RML bronchus",
        "event_id": "ev4",
        "context_prefix": "placed successfully deployed/placed in the "
    },
    {
        "span_text": "Balloon dilation",
        "label": "PROC_METHOD",
        "normalized_value": "balloon dilation",
        "event_id": "ev4",
        "context_prefix": "in the RML bronchus.\n"
    },
    {
        "span_text": "8/9/10 Elation balloon",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "Elation balloon 8/9/10",
        "event_id": "ev4",
        "context_prefix": "RML bronchus stent.  "
    },
    {
        "span_text": "100% patency",
        "label": "OUTCOME_AIRWAY_LUMEN_POST",
        "normalized_value": "100%",
        "event_id": "ev4",
        "context_prefix": "Achieving "
    },
    {
        "span_text": "Fogarty 7 balloon",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "Fogarty 7 balloon",
        "event_id": "ev4"
    },

    # --- Event 5: Lingula Stent Attempt (Failed) ---
    {
        "span_text": "iCast fully covered 7mm x 16mm stent",
        "label": "DEV_STENT",
        "normalized_value": "iCast 7x16mm",
        "event_id": "ev5"
    },
    {
        "span_text": "lingula bronchus",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Lingula bronchus",
        "event_id": "ev5",
        "context_prefix": "deployed in the "
    },
    {
        "span_text": "stent migrated proximally",
        "label": "OBS_ROSE",
        "normalized_value": "stent migration",
        "event_id": "ev5"
    },
    {
        "span_text": "Foreign body removal",
        "label": "PROC_METHOD",
        "normalized_value": "foreign body removal",
        "event_id": "ev5",
        "context_prefix": "balloon and forceps.\n"
    },
    {
        "span_text": "iCast stent",
        "label": "DEV_STENT",
        "normalized_value": "iCast stent",
        "event_id": "ev5",
        "context_prefix": "removal of "
    },
    {
        "span_text": "blunt forceps",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "forceps",
        "event_id": "ev5",
        "context_prefix": "extracted using "
    },

    # --- Event 6: Lingula Dilation (Rescue) ---
    {
        "span_text": "Balloon dilation",
        "label": "PROC_METHOD",
        "normalized_value": "balloon dilation",
        "event_id": "ev6",
        "context_prefix": "extracted using blunt forceps.\n"
    },
    {
        "span_text": "lingula bronchus",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Lingula bronchus",
        "event_id": "ev6",
        "context_prefix": "performed at "
    },
    {
        "span_text": "8/9/10 Elation balloon",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "Elation balloon 8/9/10",
        "event_id": "ev6",
        "context_prefix": "lingula bronchus.  "
    },
    {
        "span_text": "8 mm",
        "label": "MEAS_AIRWAY_DIAM",
        "normalized_value": "8 mm",
        "event_id": "ev6",
        "context_prefix": "perform dilation to "
    },
    {
        "span_text": "Improved patency",
        "label": "OUTCOME_AIRWAY_LUMEN_POST",
        "normalized_value": "improved",
        "event_id": "ev6",
        "context_prefix": "seconds each.  "
    },

    # --- Event 7: RMS Debridement/Biopsy ---
    {
        "span_text": "fibrinous exudate/debris",
        "label": "OBS_LESION",
        "normalized_value": "exudate/debris",
        "event_id": "ev7",
        "context_prefix": "partially obstructing "
    },
    {
        "span_text": "RMS",
        "label": "ANAT_AIRWAY",
        "normalized_value": "RMS",
        "event_id": "ev7",
        "context_prefix": "within the "
    },
    {
        "span_text": "Endobronchial biopsy",
        "label": "PROC_METHOD",
        "normalized_value": "biopsy",
        "event_id": "ev7",
        "context_prefix": "cuff.\n"
    },
    {
        "span_text": "stretched away",
        "label": "PROC_METHOD",
        "normalized_value": "mechanical dilation/stretching",
        "event_id": "ev7"
    },
    {
        "span_text": "Erbe HybridKnife electrocautery",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "Erbe HybridKnife",
        "event_id": "ev7"
    },
    {
        "span_text": "cut/debride",
        "label": "PROC_METHOD",
        "normalized_value": "electrocautery debridement",
        "event_id": "ev7"
    },
    {
        "span_text": "Effect 2, 40 Watts",
        "label": "MEAS_COUNT",
        "normalized_value": "Effect 2, 40 Watts",
        "event_id": "ev7",
        "schema_field": "setting"
    },
    {
        "span_text": "60% patent",
        "label": "OUTCOME_AIRWAY_LUMEN_PRE",
        "normalized_value": "60%",
        "event_id": "ev7",
        "context_prefix": "note to be "
    },
    {
        "span_text": "70% patent",
        "label": "OUTCOME_AIRWAY_LUMEN_POST",
        "normalized_value": "70%",
        "event_id": "ev7",
        "context_prefix": "airway was "
    },
    {
        "span_text": "very small laceration",
        "label": "OUTCOME_COMPLICATION",
        "normalized_value": "laceration",
        "event_id": "ev7"
    },
    {
        "span_text": "RC1",
        "label": "ANAT_AIRWAY",
        "normalized_value": "RC1",
        "event_id": "ev7",
        "context_prefix": "laceration to "
    },
    {
        "span_text": "0.2mg epinephrine",
        "label": "PROC_METHOD",
        "normalized_value": "epinephrine instillation",
        "event_id": "ev7"
    }
]

# =============================================================================
# HYDRATION LOGIC
# =============================================================================
def hydrate_spans(text, spans):
    hydrated = []
    for s in spans:
        target = s["span_text"]
        prefix = s.get("context_prefix", "")
        
        start_index = -1
        end_index = -1
        hydration_status = "ambiguous"
        
        # Strategy 1: Unique occurrence
        count = text.count(target)
        if count == 1:
            start_index = text.find(target)
            end_index = start_index + len(target)
            hydration_status = "hydrated_unique"
        
        # Strategy 2: Context Prefix
        elif prefix:
            # Find all occurrences of target
            occurrences = [m.start() for m in re.finditer(re.escape(target), text)]
            best_occ = -1
            
            # Check preceding text for each occurrence
            for occ in occurrences:
                # Look at window before occurrence (len(prefix) + buffer)
                window_start = max(0, occ - len(prefix) - 50)
                window_text = text[window_start:occ]
                if prefix in window_text or prefix.strip() in window_text:
                    best_occ = occ
                    break
            
            if best_occ != -1:
                start_index = best_occ
                end_index = start_index + len(target)
                hydration_status = "hydrated_prefix_window"
            else:
                hydration_status = f"ambiguous_prefix_not_found_count={count}"
        
        else:
            hydration_status = f"ambiguous_count={count}"

        # Create row
        row = {
            "source_file": SOURCE_FILE,
            "note_id": NOTE_ID,
            "span_id": f"{s.get('event_id', 'gen')}_{spans.index(s)}",
            "section_type": "Procedure",
            "context_prefix": prefix,
            "span_text": target,
            "match_index": "",
            "start_char": start_index if start_index != -1 else "",
            "end_char": end_index if end_index != -1 else "",
            "span_len": len(target),
            "label": s["label"],
            "normalized_value": s.get("normalized_value", ""),
            "schema_field": s.get("schema_field", ""),
            "event_id": s.get("event_id", ""),
            "is_negated": s.get("is_negated", ""),
            "is_historical": s.get("is_historical", ""),
            "time_anchor": "",
            "reviewer": "Auto",
            "comments": s.get("comments", ""),
            "hydration_status": hydration_status
        }
        hydrated.append(row)
    return hydrated

# =============================================================================
# EXCEL GENERATION
# =============================================================================
def generate_excel():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        # Create blank if template missing (fallback for safety)
        wb = openpyxl.Workbook()
        
    # 1. Note_Text
    ws_text = wb["Note_Text"] if "Note_Text" in wb.sheetnames else wb.create_sheet("Note_Text")
    if ws_text.max_row == 1:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws_index = wb["Note_Index"] if "Note_Index" in wb.sheetnames else wb.create_sheet("Note_Index")
    # Headers should be in template, else append
    row_meta = [SOURCE_FILE, NOTE_ID, "", "", "", "", "Ready", ""]
    # Append flags
    flag_values = [flags[k] for k in flags]
    ws_index.append(row_meta + flag_values)

    # 3. Span_Annotations (Anchor First)
    ws_span = wb["Span_Annotations"] if "Span_Annotations" in wb.sheetnames else wb.create_sheet("Span_Annotations")
    # Write only anchor logic (no offsets)
    for s in spans_data:
        row = [
            SOURCE_FILE, NOTE_ID, f"{s.get('event_id','gen')}_{spans_data.index(s)}", "Procedure",
            s.get("context_prefix", ""), s["span_text"], "", "", "", f"=LEN(F{ws_span.max_row+1})",
            s["label"], s.get("normalized_value", ""), s.get("schema_field", ""), s.get("event_id", ""),
            s.get("is_negated", ""), s.get("is_historical", ""), "", "Auto", s.get("comments", ""),
            "needs_hydration"
        ]
        ws_span.append(row)

    # 4. Span_Hydrated
    ws_hyd = wb["Span_Hydrated"] if "Span_Hydrated" in wb.sheetnames else wb.create_sheet("Span_Hydrated")
    hydrated_rows = hydrate_spans(NOTE_TEXT, spans_data)
    # Headers assumed present in template, append data
    for r in hydrated_rows:
        ws_hyd.append([
            r["source_file"], r["note_id"], r["span_id"], r["section_type"],
            r["context_prefix"], r["span_text"], r["match_index"],
            r["start_char"], r["end_char"], r["span_len"],
            r["label"], r["normalized_value"], r["schema_field"], r["event_id"],
            r["is_negated"], r["is_historical"], r["time_anchor"],
            r["reviewer"], r["comments"], r["hydration_status"]
        ])

    # 5. Event_Log (V3 structure implied by columns)
    ws_event = wb["Event_Log"] if "Event_Log" in wb.sheetnames else wb.create_sheet("Event_Log")
    
    events_summary = [
        {
            "event_id": "ev1", "type": "Therapeutic Aspiration", "method": "Aspiration", "anatomy": "RMS, BI, RLL, LMS, LLL",
            "device": "Flexible Bronchoscope", "outcome_symptoms": "", "outcome_complication": "None"
        },
        {
            "event_id": "ev2", "type": "Airway Dilation", "method": "Blunt Dilation", "anatomy": "RML Bronchus",
            "device": "Therapeutic Bronchoscope", "measurements": "6.2 mm", "findings": "Copious secretions"
        },
        {
            "event_id": "ev3", "type": "BAL", "method": "Lavage", "anatomy": "RML Bronchus",
            "specimens": "Microbiology/Cell Count", "measurements": "40cc in, 25cc out"
        },
        {
            "event_id": "ev4", "type": "Airway Stent", "method": "Stent Placement", "anatomy": "RML Bronchus",
            "device": "iCast 7x22mm", "outcome_lumen_post": "100%", "device_size": "7x22mm"
        },
        {
            "event_id": "ev5", "type": "Foreign Body Removal (Failed Stent)", "method": "Removal", "anatomy": "Lingula Bronchus",
            "device": "iCast 7x16mm", "findings": "Stent migrated/malpositioned", "outcome_complication": "Stent migration"
        },
        {
            "event_id": "ev6", "type": "Airway Dilation", "method": "Balloon Dilation", "anatomy": "Lingula Bronchus",
            "device": "Elation Balloon", "measurements": "8mm", "outcome_lumen_post": "Improved"
        },
        {
            "event_id": "ev7", "type": "Tumor Debulking/Biopsy", "method": "Electrocautery/Biopsy", "anatomy": "RMS Bronchus",
            "device": "HybridKnife", "findings": "Exudate debrided", "outcome_lumen_pre": "60%", "outcome_lumen_post": "70%", "outcome_complication": "Small laceration"
        }
    ]
    
    for ev in events_summary:
        ws_event.append([
            SOURCE_FILE, NOTE_ID, ev["event_id"], ev["type"], ev["method"],
            ev["anatomy"], ev.get("device",""), "", "", "", ev.get("measurements",""),
            ev.get("specimens",""), ev.get("findings",""), "", "Auto", "",
            ev.get("device_size",""), "",
            ev.get("outcome_lumen_pre",""), ev.get("outcome_lumen_post",""),
            ev.get("outcome_symptoms",""), "", ev.get("outcome_complication","")
        ])

    # 6. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"] if "V3_Registry_JSON" in wb.sheetnames else wb.create_sheet("V3_Registry_JSON")
    
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": False, # Complication occurred
        "procedures": [
            {
                "event_id": "ev4",
                "type": "Airway Stent",
                "target": {"anatomy_type": "Airway", "location": {"lobe": "RML"}, "segment": "Bronchus"},
                "stent": {"brand": "iCast", "size": "7x22mm"},
                "outcomes": {"airway": {"lumen_post": "100%"}}
            },
            {
                "event_id": "ev7",
                "type": "Tumor Debulking",
                "method": "Electrocautery",
                "target": {"anatomy_type": "Airway", "location": {"lobe": "Right Main Stem"}},
                "outcomes": {
                    "airway": {"lumen_pre": "60%", "lumen_post": "70%"},
                    "complications": "Small laceration"
                }
            }
        ]
    }
    
    ws_json.append([json.dumps(registry_data, indent=2)])

    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_excel()