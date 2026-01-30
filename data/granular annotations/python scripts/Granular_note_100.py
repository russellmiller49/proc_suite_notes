import openpyxl
from openpyxl.utils import get_column_letter
import datetime
import json
import re

# =============================================================================
# 1. INPUT DATA
# =============================================================================
NOTE_ID = "note_100"
SOURCE_FILE = "note_100.txt"
PROCEDURE_DATE = "2026-01-12" # inferred or current
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_100 SOURCE_FILE: note_100.txt INDICATION FOR OPERATION:  [REDACTED]is a 56 year old-year-old male who presents with respiratory failure.
PREOPERATIVE DIAGNOSIS: J96.90 Respiratory Failure
POSTOPERATIVE DIAGNOSIS:  J96.90 Respiratory Failure
PROCEDURE:  
31646 Therapeutic aspiration subsequent episodes
31622 Dx bronchoscope/cell washing          
ANESTHESIA: 
Local ONLY
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Disposable Bronchoscope
ESTIMATED BLOOD LOSS:   None
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
A timeout was performed (confirming the patient's name, procedure type, and procedure location).
Lidocaine applied to trachea, main carina, RMS, & LMS.
Initial Airway Inspection Findings:
The tracheostomy tube is in good position.
The visualized portion of the trachea is of normal caliber. The carina is sharp.
The tracheobronchial tree was examined to at least the first subsegmental level.
Airway exam notable for persistent pale plaque of tissue along the anterior wall of the RMSB/BI and RML orifice.
The previously seen metallic hemoclip near the RMS anastomosis posterior membrane is no longer present.
The RML orifice is now completely stenosed by overlying necrotic debris.  Unable to probe into the airway to achieve patency.
The LMSB anastomosis shows mild granulation tissue and similar pale soft tissue plaque along the anterior aspect of the distal LMSB.
Segmental airways on the left are widely patent. Clear secretions bilaterally.
Successful therapeutic aspiration was performed to clean out the Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, and RLL and LLL from mucus.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
None
IMPRESSION/PLAN: [REDACTED]is a 56 year old-year-old male who presents for bronchoscopy for respiratory failure and bronchial anastomotic dehiscence and stenosis.
-Plan for bronchoscopy under GA for mngmnt of RML stenosis - likely excision of overlying debris, balloon dilation, and stent placement.
Likely on [REDACTED]."""

# =============================================================================
# 2. FLAGGING LOGIC
# =============================================================================
# Initialize all flags to 0
flags = {
    # Bronchoscopy (23)
    "diagnostic_bronchoscopy": 0, "bal": 0, "bronchial_wash": 0, "brushings": 0,
    "endobronchial_biopsy": 0, "tbna_conventional": 0, "linear_ebus": 0, "radial_ebus": 0,
    "navigational_bronchoscopy": 0, "transbronchial_biopsy": 0, "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0, "foreign_body_removal": 0, "airway_dilation": 0,
    "airway_stent": 0, "thermal_ablation": 0, "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0, "blvr": 0, "peripheral_ablation": 0, "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0, "rigid_bronchoscopy": 0,
    # Pleural (7)
    "thoracentesis": 0, "chest_tube": 0, "ipc": 0, "medical_thoracoscopy": 0,
    "pleurodesis": 0, "pleural_biopsy": 0, "fibrinolytic_therapy": 0
}

# Apply logic based on text content
lower_text = NOTE_TEXT.lower()

# Diagnostic Bronchoscopy
if "dx bronchoscope" in lower_text or "bronchoscopy" in lower_text:
    flags["diagnostic_bronchoscopy"] = 1

# Bronchial Wash
if "cell washing" in lower_text or "bronchial wash" in lower_text:
    flags["bronchial_wash"] = 1

# Therapeutic Aspiration
if "therapeutic aspiration" in lower_text or "aspiration" in lower_text:
    flags["therapeutic_aspiration"] = 1

# Note: The plan mentions stent and dilation, but the procedure text confirms only aspiration/wash this time.
# "Unable to probe into the airway to achieve patency" implies failure or lack of attempt at dilation in this session beyond probing.
# Flags remain 0 for stent/dilation.

# =============================================================================
# 3. SPAN EXTRACTION (ANCHOR-FIRST)
# =============================================================================
# Structure: (span_text, label, normalized_value, event_id, context_prefix_or_None)

spans_data = []

def add_span(text, label, norm, ev_id, prefix=None, schema_field=None, is_hist=None, comments=None):
    spans_data.append({
        "span_text": text,
        "label": label,
        "normalized_value": norm,
        "event_id": ev_id,
        "context_prefix": prefix,
        "schema_field": schema_field,
        "is_historical": is_hist,
        "comments": comments
    })

# EVENT 1: General/Diagnostic & Anesthesia
ev1 = "evt_001" # General
add_span("Disposable Bronchoscope", "DEV_INSTRUMENT", "Bronchoscope", ev1)
add_span("Lidocaine", "PROC_METHOD", "Anesthesia - Local", ev1)
add_span("trachea", "ANAT_AIRWAY", "Trachea", ev1, prefix="Lidocaine applied to ")
add_span("main carina", "ANAT_AIRWAY", "Main Carina", ev1)
add_span("RMS", "ANAT_AIRWAY", "RMS", ev1, prefix="main carina, ")
add_span("LMS", "ANAT_AIRWAY", "LMS", ev1, prefix="RMS, & ")

# EVENT 2: Inspection Findings (Trachea/Carina)
ev2 = "evt_002"
add_span("tracheostomy tube", "DEV_CATHETER", "Tracheostomy Tube", ev2)
add_span("good position", "OBS_ROSE", "Good Position", ev2)
add_span("trachea", "ANAT_AIRWAY", "Trachea", ev2, prefix="visualized portion of the ")
add_span("normal caliber", "OBS_ROSE", "Normal Caliber", ev2)
add_span("The carina is sharp", "OBS_ROSE", "Sharp Carina", ev2)

# EVENT 3: RML/RMS Findings
ev3 = "evt_003"
add_span("pale plaque of tissue", "OBS_LESION", "Plaque", ev3, prefix="persistent ")
add_span("RMSB/BI", "ANAT_AIRWAY", "RMSB/BI", ev3)
add_span("RML orifice", "ANAT_AIRWAY", "RML Orifice", ev3, prefix="and ")
add_span("RML orifice", "ANAT_AIRWAY", "RML Orifice", ev3, prefix="The ")
add_span("completely stenosed", "OBS_LESION", "Stenosis - Complete", ev3)
add_span("necrotic debris", "OBS_LESION", "Necrotic Debris", ev3)
add_span("Unable to probe", "PROC_ACTION", "Probe Failure", ev3)

# EVENT 4: LMS Findings
ev4 = "evt_004"
add_span("LMSB anastomosis", "ANAT_AIRWAY", "LMSB Anastomosis", ev4)
add_span("mild granulation tissue", "OBS_LESION", "Granulation Tissue", ev4)
add_span("pale soft tissue plaque", "OBS_LESION", "Plaque", ev4)

# EVENT 5: Therapeutic Aspiration
ev5 = "evt_005"
add_span("Therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", ev5, prefix="31646 ")
add_span("Successful therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", ev5, prefix="12] ")
add_span("clean out", "PROC_ACTION", "Clean Out", ev5)
add_span("Trachea (Distal 1/3)", "ANAT_AIRWAY", "Trachea Distal", ev5)
add_span("Right Mainstem", "ANAT_AIRWAY", "RMS", ev5)
add_span("Bronchus Intermedius", "ANAT_AIRWAY", "BI", ev5, prefix="Right Mainstem, ")
add_span("Left Mainstem", "ANAT_AIRWAY", "LMS", ev5, prefix="Intermedius , ")
add_span("RLL", "ANAT_AIRWAY", "RLL", ev5)
add_span("LLL", "ANAT_AIRWAY", "LLL", ev5)
add_span("mucus", "OBS_LESION", "Mucus", ev5)

# EVENT 6: Outcomes
ev6 = "evt_006"
add_span("No immediate complications", "OUTCOME_COMPLICATION", "None", ev6)
add_span("tolerated the procedure well", "OUTCOME_SYMPTOMS", "Tolerated Well", ev6)
add_span("extubated", "PROC_ACTION", "Extubation", ev6)

# PLAN (Historical/Future - excluded from flags but noted)
ev_plan = "evt_plan"
add_span("RML stenosis", "OBS_LESION", "Stenosis", ev_plan, prefix="mngmnt of ", is_hist="Yes", comments="Future Plan Context")
add_span("balloon dilation", "PROC_METHOD", "Balloon Dilation", ev_plan, prefix="debris, ", is_hist="Yes", comments="Future Plan Context")
add_span("stent placement", "PROC_METHOD", "Airway Stent", ev_plan, prefix="dilation, and ", is_hist="Yes", comments="Future Plan Context")

# =============================================================================
# 4. HYDRATION LOGIC
# =============================================================================

def hydrate_spans(text, span_list):
    hydrated = []
    text_lower = text.lower()
    
    for s in span_list:
        phrase = s["span_text"]
        phrase_len = len(phrase)
        
        # 1. Exact Match Check
        count = text.count(phrase)
        
        start_idx = -1
        end_idx = -1
        status = "ambiguous"
        match_index = None

        if count == 0:
            status = "not_found"
        elif count == 1:
            start_idx = text.find(phrase)
            end_idx = start_idx + phrase_len
            status = "hydrated_unique"
            match_index = 1
        else:
            # Multiple matches
            if s["context_prefix"]:
                # Window search
                prefix = s["context_prefix"]
                # Find all occurrences of phrase
                matches = [m.start() for m in re.finditer(re.escape(phrase), text)]
                
                best_m = -1
                for m_start in matches:
                    # Look back 120 chars
                    window_start = max(0, m_start - 120)
                    window_text = text[window_start:m_start]
                    if prefix in window_text:
                        best_m = m_start
                        break
                
                if best_m != -1:
                    start_idx = best_m
                    end_idx = start_idx + phrase_len
                    status = "hydrated_prefix_window"
                    # Determine match index
                    matches = [m.start() for m in re.finditer(re.escape(phrase), text)]
                    match_index = matches.index(best_m) + 1
                else:
                    status = f"ambiguous_count={count}_prefix_not_found"
            else:
                 status = f"ambiguous_count={count}"
        
        row = s.copy()
        row["start_char"] = start_idx if start_idx != -1 else ""
        row["end_char"] = end_idx if end_idx != -1 else ""
        row["span_len"] = phrase_len
        row["match_index"] = match_index if match_index else ""
        row["hydration_status"] = status
        hydrated.append(row)
        
    return hydrated

hydrated_spans = hydrate_spans(NOTE_TEXT, spans_data)

# =============================================================================
# 5. GENERATE EXCEL
# =============================================================================
def generate_excel(template_path: str = TEMPLATE_PATH, output_path: str = OUTPUT_PATH) -> None:
    try:
        wb = openpyxl.load_workbook(template_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        for sheet in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            if sheet not in wb.sheetnames:
                wb.create_sheet(sheet)

    ws_text = wb["Note_Text"]
    if ws_text.max_row == 1:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    ws_index = wb["Note_Index"]
    headers_index = [
        "source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes",
        "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", "tbna_conventional",
        "linear_ebus", "radial_ebus", "navigational_bronchoscopy", "transbronchial_biopsy", "transbronchial_cryobiopsy",
        "therapeutic_aspiration", "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
        "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", "bronchial_thermoplasty",
        "whole_lung_lavage", "rigid_bronchoscopy",
        "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis", "pleural_biopsy", "fibrinolytic_therapy"
    ]
    if ws_index.max_row == 1:
        ws_index.append(headers_index)

    row_index = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Success", ""]
    for h in headers_index[8:]:
        row_index.append(flags.get(h, 0))
    ws_index.append(row_index)

    ws_span = wb["Span_Annotations"]
    headers_span = [
        "source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", "match_index",
        "start_char", "end_char", "span_len", "label", "normalized_value", "schema_field", "event_id",
        "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"
    ]
    if ws_span.max_row == 1:
        ws_span.append(headers_span)

    span_id_counter = 1
    for s in spans_data:
        row = [
            SOURCE_FILE, NOTE_ID, f"span_{span_id_counter:03d}", "", s.get("context_prefix"), s["span_text"], "",
            "", "", f"=LEN(F{ws_span.max_row+1})", s["label"], s["normalized_value"], s.get("schema_field"), s["event_id"],
            "", s.get("is_historical"), "", "", s.get("comments"), "needs_hydration"
        ]
        ws_span.append(row)
        span_id_counter += 1

    ws_hyd = wb["Span_Hydrated"]
    if ws_hyd.max_row == 1:
        ws_hyd.append(headers_span)

    span_id_counter = 1
    for s in hydrated_spans:
        row = [
            SOURCE_FILE, NOTE_ID, f"span_{span_id_counter:03d}", "", s.get("context_prefix"), s["span_text"], s.get("match_index"),
            s["start_char"], s["end_char"], s["span_len"], s["label"], s["normalized_value"], s.get("schema_field"), s["event_id"],
            "", s.get("is_historical"), "", "", s.get("comments"), s["hydration_status"]
        ]
        ws_hyd.append(row)
        span_id_counter += 1

    ws_event = wb["Event_Log"]
    headers_event = [
        "source_file", "note_id", "event_id", "event_type", "method", "anatomy_target", "device", "needle_gauge",
        "stations", "counts", "measurements", "specimens", "findings", "is_historical", "reviewer", "comments",
        "device_size", "device_material", "outcome_airway_lumen_pre", "outcome_airway_lumen_post",
        "outcome_symptoms", "outcome_pleural", "outcome_complication"
    ]
    if ws_event.max_row == 1:
        ws_event.append(headers_event)

    events = [
        {"id": "evt_001", "type": "Diagnostic Bronchoscopy", "method": "Flexible", "anatomy": "Trachea, RMS, LMS", "device": "Disposable Bronchoscope", "findings": "Lidocaine applied"},
        {"id": "evt_002", "type": "Inspection", "anatomy": "Trachea", "findings": "Tracheostomy tube good position, normal caliber, sharp carina"},
        {"id": "evt_003", "type": "Inspection", "anatomy": "RML Orifice", "findings": "Pale plaque, necrotic debris, completely stenosed, unable to probe"},
        {"id": "evt_004", "type": "Inspection", "anatomy": "LMSB", "findings": "Mild granulation tissue, pale soft tissue plaque"},
        {"id": "evt_005", "type": "Therapeutic Aspiration", "method": "Suction/Lavage", "anatomy": "Trachea, RMS, BI, LMS, RLL, LLL", "findings": "Mucus removed", "outcome_symptoms": "Successful"},
        {"id": "evt_006", "type": "Outcome", "findings": "Patient tolerated well, extubated", "outcome_complication": "None", "outcome_symptoms": "Tolerated Well"},
    ]

    for e in events:
        row = [
            SOURCE_FILE, NOTE_ID, e["id"], e["type"], e.get("method"), e.get("anatomy"), e.get("device"), "",
            "", "", "", "", e.get("findings"), "", "", "",
            "", "", "", "", e.get("outcome_symptoms"), "", e.get("outcome_complication")
        ]
        ws_event.append(row)

    ws_v3 = wb["V3_Procedure_Events"]
    headers_v3 = [
        "note_id", "event_id", "type", "target.anatomy_type", "target.location.lobe", "target.location.segment",
        "target.station", "lesion.type", "lesion.size_mm", "method", "devices_json", "measurements_json",
        "specimens_json", "findings_json", "evidence_quote", "stent.size", "stent.material_or_brand",
        "catheter.size_fr", "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post", "outcomes.symptoms",
        "outcomes.pleural", "outcomes.complications"
    ]
    if ws_v3.max_row == 1:
        ws_v3.append(headers_v3)

    v3_rows = [
        [NOTE_ID, "evt_003", "Diagnostic Bronchoscopy", "Airway", "RML", "Orifice", "", "Stenosis", "", "Inspection", "", "", "", json.dumps(["Necrotic debris", "Pale plaque"]), "RML orifice is now completely stenosed", "", "", "", "", "0% (Stenosed)", "", "", ""],
        [NOTE_ID, "evt_005", "Therapeutic Aspiration", "Airway", "Multiple", "", "", "Mucus", "", "Suction", "", "", "", json.dumps(["Mucus removed"]), "Successful therapeutic aspiration", "", "", "", "", "Patent", "Improved", "", ""]
    ]
    for r in v3_rows:
        ws_v3.append(r)

    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 1:
        ws_json.append(["schema_version", "note_id", "json_object"])

    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": [{"event_id": "evt_005", "type": "Therapeutic Aspiration", "targets": ["Trachea", "RMS", "BI", "LMS", "RLL", "LLL"], "findings": ["Mucus"]}],
    }
    ws_json.append(["3.0", NOTE_ID, json.dumps(registry_data, indent=2)])

    wb.save(output_path)
    print(f"Successfully generated {output_path}")


if __name__ == "__main__":
    generate_excel()
