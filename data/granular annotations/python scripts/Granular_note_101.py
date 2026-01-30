import sys
import os
import datetime
import re
import json
from openpyxl import Workbook, load_workbook

# =============================================================================
# INPUT CONSTANTS
# =============================================================================
NOTE_ID = "note_101"
SOURCE_FILE = "note_101.txt"
PROCEDURE_DATE = ""  # Not found in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_101 SOURCE_FILE: note_101.txt INDICATION FOR OPERATION:  [REDACTED]is a 56 year old-year-old male who presents with respiratory failure.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: J96.90 Respiratory Failure
POSTOPERATIVE DIAGNOSIS:  J96.90 Respiratory Failure
PROCEDURE:  
31646 Therapeutic aspiration subsequent episodes
31622 Dx bronchoscope/cell washing          
31624 Dx bronchoscope/lavage (BAL)    
ANESTHESIA: 
Local ONLY
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Disposable Bronchoscope
ESTIMATED BLOOD LOSS:   None
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
Initial Airway Inspection Findings:
The tracheostomy tube is in good position. The visualized portion of the trachea is of normal caliber.
The carina is sharp. The tracheobronchial tree was examined to at least the first subsegmental level.
Airway exam notable for persistent pale plaque of tissue along the anterior wall of the RMSB/BI and RML orifice.
The RMSB anastomosis continues to show a partially exposed hemoclip along the posterior membrane.
The RML orifice is stenosed to ~75% of its normal patency.
The LMSB anastomosis shows mild granulation tissue and similar pale soft tissue plaque along the anterior aspect of the distal LMSB.
Segmental airways on the left are widely patent. Clear secretions bilaterally.
Successful therapeutic aspiration was performed to clean out the Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , and Left Mainstem from mucus.
Bronchial alveolar lavage was performed at Anterior-basal Segment of RLL (RB8).
Instilled 40 cc of NS, suction returned with 10 cc of NS.  Samples sent for Cell Count and Microbiology (Cultures/Viral/Fungal).
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
- RLL BAL 
IMPRESSION/PLAN: [REDACTED]is a 56 year old-year-old male who presents for bronchoscopy for respiratory failure and bronchial anastomotic dehiscence and stenosis.
- f/u BAL studies"""

# =============================================================================
# DATA DEFINITIONS
# =============================================================================

# 1. Procedure Flags (0/1)
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 1,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 0, # Hemoclip present, but not removed
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 2. Spans (Anchor-First)
# Format: (span_text, label, normalized_value, event_id, context_prefix_or_None)
SPANS = [
    # Header/Instruments
    ("31646 Therapeutic aspiration subsequent episodes", "PROC_METHOD", "Therapeutic Aspiration", "evt_asp_1", None),
    ("31622 Dx bronchoscope", "PROC_METHOD", "Diagnostic Bronchoscopy", "evt_insp_1", None),
    ("31624 Dx bronchoscope/lavage (BAL)", "PROC_METHOD", "BAL", "evt_bal_1", None),
    ("Disposable Bronchoscope", "DEV_INSTRUMENT", "Disposable Bronchoscope", "evt_global", None),
    
    # Inspection Findings
    ("RMSB/BI and RML orifice", "ANAT_AIRWAY", "RMSB, BI, RML Orifice", "evt_insp_1", None),
    ("persistent pale plaque", "OBS_LESION", "plaque", "evt_insp_1", None),
    
    ("RMSB anastomosis", "ANAT_AIRWAY", "RMSB Anastomosis", "evt_insp_2", None),
    ("partially exposed hemoclip", "OBS_LESION", "exposed hemoclip", "evt_insp_2", None),
    
    ("RML orifice", "ANAT_AIRWAY", "RML Orifice", "evt_insp_3", None),
    ("stenosed to ~75% of its normal patency", "OUTCOME_AIRWAY_LUMEN_PRE", "75% Patency", "evt_insp_3", None),
    
    ("LMSB anastomosis", "ANAT_AIRWAY", "LMSB Anastomosis", "evt_insp_4", None),
    ("mild granulation tissue", "OBS_LESION", "granulation tissue", "evt_insp_4", None),
    ("pale soft tissue plaque", "OBS_LESION", "plaque", "evt_insp_4", None),
    
    # Therapeutic Aspiration
    ("Successful therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", "evt_asp_1", None),
    ("clean out", "PROC_ACTION", "clean out", "evt_asp_1", None),
    ("Trachea (Distal 1/3)", "ANAT_AIRWAY", "Trachea Distal 1/3", "evt_asp_1", None),
    ("Right Mainstem", "ANAT_AIRWAY", "RMSB", "evt_asp_1", None),
    ("Bronchus Intermedius", "ANAT_AIRWAY", "BI", "evt_asp_1", None),
    ("Left Mainstem", "ANAT_AIRWAY", "LMSB", "evt_asp_1", "Bronchus Intermedius , and "),
    ("mucus", "OBS_LESION", "Mucus", "evt_asp_1", "Mainstem from "),

    # BAL
    ("Bronchial alveolar lavage", "PROC_METHOD", "BAL", "evt_bal_1", "13] "),
    ("Anterior-basal Segment of RLL (RB8)", "ANAT_LUNG_LOC", "RLL RB8", "evt_bal_1", None),
    ("Instilled 40 cc", "MEAS_VOL", "40 cc", "evt_bal_1", None),
    ("returned with 10 cc", "MEAS_VOL", "10 cc", "evt_bal_1", None),
    ("Cell Count", "PROC_ACTION", "Specimen Sent", "evt_bal_1", None),
    ("Microbiology", "PROC_ACTION", "Specimen Sent", "evt_bal_1", None),

    # Complications
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "evt_global", None),
]

# 3. Events Map
EVENTS = [
    {
        "event_id": "evt_insp_1",
        "type": "Diagnostic Bronchoscopy",
        "anatomy": "RMSB/BI/RML",
        "findings": "persistent pale plaque"
    },
    {
        "event_id": "evt_insp_2",
        "type": "Diagnostic Bronchoscopy",
        "anatomy": "RMSB anastomosis",
        "findings": "partially exposed hemoclip"
    },
    {
        "event_id": "evt_insp_3",
        "type": "Diagnostic Bronchoscopy",
        "anatomy": "RML orifice",
        "outcomes": {"airway_lumen_pre": "75% Patency"}
    },
    {
        "event_id": "evt_insp_4",
        "type": "Diagnostic Bronchoscopy",
        "anatomy": "LMSB anastomosis",
        "findings": "granulation tissue, plaque"
    },
    {
        "event_id": "evt_asp_1",
        "type": "Therapeutic Aspiration",
        "anatomy": "Trachea/RMSB/BI/LMSB",
        "findings": "Mucus"
    },
    {
        "event_id": "evt_bal_1",
        "type": "BAL",
        "anatomy": "RLL RB8",
        "specimens": "Cell Count, Microbiology",
        "measurements": "Instilled 40cc, Returned 10cc"
    },
    {
        "event_id": "evt_global",
        "type": "Global",
        "outcomes": {"complications": "None"}
    }
]

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def find_offsets_anchor_first(note_text, span_text, context_prefix=None):
    """
    Hydrates offsets for a span.
    Returns: (start_char, end_char, hydration_status)
    """
    matches = [m.start() for m in re.finditer(re.escape(span_text), note_text)]
    
    if not matches:
        return None, None, "missing"
    
    if len(matches) == 1:
        start = matches[0]
        return start, start + len(span_text), "hydrated_unique"
    
    # Ambiguity resolution
    if context_prefix:
        # Check window before each match
        for m_start in matches:
            window_start = max(0, m_start - 120)
            window_text = note_text[window_start:m_start]
            if context_prefix in window_text:
                return m_start, m_start + len(span_text), "hydrated_prefix_window"
    
    # Default to first if ambiguous (simple fallback for script)
    return matches[0], matches[0] + len(span_text), f"ambiguous_count={len(matches)}"

# =============================================================================
# MAIN LOGIC
# =============================================================================

def generate_workbook():
    # 1. Create/Load Workbook
    try:
        wb = load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        # Fallback if template missing (should not happen per prompt rules, but safe to create)
        wb = Workbook()
    
    # Helper to get or create sheet
    def get_sheet(name):
        if name in wb.sheetnames:
            return wb[name]
        return wb.create_sheet(name)

    # 2. Populate Note_Text
    ws_text = get_sheet("Note_Text")
    # Check headers
    if ws_text.max_row == 1 and ws_text.cell(1,1).value is None:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 3. Populate Note_Index
    ws_index = get_sheet("Note_Index")
    # Headers logic usually pre-existing in template, but we append row
    # Metadata cols: source_file, note_id, encounter_id, procedure_date, site, reviewer, status, free_text_notes
    row_meta = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Machine_Generated", ""]
    
    # Procedure flags mapping
    flag_keys = [
        "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", 
        "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy", "transbronchial_biopsy", 
        "transbronchial_cryobiopsy", "therapeutic_aspiration", "foreign_body_removal", "airway_dilation", 
        "airway_stent", "thermal_ablation", "tumor_debulking_non_thermal", "cryotherapy", "blvr", 
        "peripheral_ablation", "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
        "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis", "pleural_biopsy", "fibrinolytic_therapy"
    ]
    
    row_flags = [PROCEDURE_FLAGS.get(k, 0) for k in flag_keys]
    ws_index.append(row_meta + row_flags)

    # 4. Populate Span_Annotations and Span_Hydrated
    ws_anno = get_sheet("Span_Annotations")
    ws_hydra = get_sheet("Span_Hydrated")
    
    # Ensure headers if new
    headers = [
        "source_file", "note_id", "span_id", "section_type", "context_prefix",
        "span_text", "match_index", "start_char", "end_char", "span_len",
        "label", "normalized_value", "schema_field", "event_id",
        "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"
    ]
    if ws_anno.max_row == 1: ws_anno.append(headers)
    if ws_hydra.max_row == 1: ws_hydra.append(headers)

    for idx, (span_text, label, norm_val, evt_id, ctx) in enumerate(SPANS):
        span_id = f"{NOTE_ID}_s{idx+1:03d}"
        
        # Calculate hydration
        start, end, status = find_offsets_anchor_first(NOTE_TEXT, span_text, ctx)
        
        # Annotation Row (No offsets)
        row_anno = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure", ctx,
            span_text, "", "", "", f"=LEN(F{ws_anno.max_row+1})",
            label, norm_val, "", evt_id,
            "FALSE", "FALSE", "", "Auto", "", "needs_hydration"
        ]
        ws_anno.append(row_anno)
        
        # Hydrated Row
        row_hydra = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure", ctx,
            span_text, "", start, end, (end-start) if start is not None else 0,
            label, norm_val, "", evt_id,
            "FALSE", "FALSE", "", "Auto", "", status
        ]
        ws_hydra.append(row_hydra)

    # 5. Populate Event_Log
    ws_event = get_sheet("Event_Log")
    # Headers: source_file, note_id, event_id, event_type, method, anatomy_target, device, needle_gauge, stations, counts, measurements, specimens, findings, is_historical, reviewer, comments, device_size, device_material, outcome_airway_lumen_pre, outcome_airway_lumen_post, outcome_symptoms, outcome_pleural, outcome_complication
    if ws_event.max_row == 1:
        ws_event.append([
            "source_file", "note_id", "event_id", "event_type", "method",
            "anatomy_target", "device", "needle_gauge", "stations", "counts",
            "measurements", "specimens", "findings", "is_historical", "reviewer",
            "comments", "device_size", "device_material",
            "outcome_airway_lumen_pre", "outcome_airway_lumen_post",
            "outcome_symptoms", "outcome_pleural", "outcome_complication"
        ])

    for evt in EVENTS:
        row = [
            SOURCE_FILE, NOTE_ID, evt["event_id"], evt["type"],
            evt.get("method", ""), evt.get("anatomy", ""),
            evt.get("device", ""), "", "", "",
            evt.get("measurements", ""), evt.get("specimens", ""),
            evt.get("findings", ""), "FALSE", "Auto", "", "", "",
            evt.get("outcomes", {}).get("airway_lumen_pre", ""),
            evt.get("outcomes", {}).get("airway_lumen_post", ""),
            evt.get("outcomes", {}).get("symptoms", ""),
            evt.get("outcomes", {}).get("pleural", ""),
            evt.get("outcomes", {}).get("complications", "")
        ]
        ws_event.append(row)

    # 6. Populate V3_Procedure_Events (Simplified for script)
    ws_v3 = get_sheet("V3_Procedure_Events")
    # Columns: note_id, event_id, type, target.anatomy_type, target.location.lobe, ...
    if ws_v3.max_row == 1:
        ws_v3.append([
            "note_id", "event_id", "type", "target.anatomy_type", "target.location.lobe",
            "target.location.segment", "target.station", "lesion.type", "lesion.size_mm",
            "method", "devices_json", "measurements_json", "specimens_json", "findings_json",
            "evidence_quote", "stent.size", "stent.material_or_brand", "catheter.size_fr",
            "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post", "outcomes.symptoms",
            "outcomes.pleural", "outcomes.complications"
        ])
    
    for evt in EVENTS:
        # Simplified mapping
        row_v3 = [
            NOTE_ID, evt["event_id"], evt["type"], "Airway" if "Bronch" in evt["type"] else "Lung",
            "", "", "", "", "", # Locations blank for simple script
            evt.get("method", ""), "", "", "", "", "", # JSONs blank
            "", "", "",
            evt.get("outcomes", {}).get("airway_lumen_pre", ""),
            evt.get("outcomes", {}).get("airway_lumen_post", ""),
            "", "", evt.get("outcomes", {}).get("complications", "")
        ]
        ws_v3.append(row_v3)

    # 7. V3_Registry_JSON
    ws_json = get_sheet("V3_Registry_JSON")
    if ws_json.max_row == 1:
        ws_json.append(["schema_version", "note_id", "json_object"])
    
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": EVENTS
    }
    ws_json.append(["3.0", NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()