import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import json
import re

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_102"
SOURCE_FILE = "note_102.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_102 SOURCE_FILE: note_102.txt INDICATION FOR OPERATION:  [REDACTED]is a 56 year old-year-old male who presents with bilateral lung transplant with complication of anastomosis dehiscence and ischemic lung injury.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
PROCEDURE:  
31899 Unlisted Procedure (Trach Change with Mature Tract or Procedure NOS)
31646 Therapeutic aspiration subsequent episodes
31624 Dx bronchoscope/lavage (BAL)    
ANESTHESIA: Local
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Disposable Bronchoscope
PROCEDURE IN DETAIL:
A timeout was performed (confirming the patient's name, procedure type, and procedure location).
The Flexible Therapeutic Bronchoscope was advanced for airway evaluation.  Endobronchial topical lidocaine applied to the main carina, RMS, BI, and LC2.
Initial Airway Inspection Findings:
Distal trachea normal.  
RMS dehiscence continues to be healed and remains closed.
Small metallic object protruding from the posterior membrane near the RMS anastomosis site - consistent with likely hemoclip.
This is slightly more prominent than our bronchoscopy 1 week prior.
Continues to have fibrinous exudate/desired granulation tissue in the donor RMS, proximal RUL bronchus, BI, overlying RML take-off, and overlying RB6 take-off.
All areas appear to be improved from previous.  There is some mildly heaped nonobstructing granulation tissue along the medial aspect of the right mainstem bronchus.
Previously seen area of fullness thickness erosion/ulceration along RUL bronchus anterior wall now appears to be fully covered by fibrinous exudate suggesting early healing.
This is greatly improved from 1 week ago.  RML take-off is narrowed, but able to traverse with disposable bronchoscope "regular" size.
RLL basilar segments appear healthy. Moderate secretions in BI and RLL which were suctioned to clear (therapeutic aspiration).
LMS anastomosis intact with visible sutures and mild stenosis. LUL bronchus with evidence of continued healing with fibrin exudates/desired granulation tissue.
No longer able to see underlying mediastinum/pulmonary artery along medial aspect of LUL.
This entire area continues to look improved from 1 week ag.
Lingula take-off is narrowed, but able to traverse with disposable bronochoscope "regular" size.  LLL bronchus and segments appear health.
Successful therapeutic aspiration was performed to clean out the trachea, right mainstem bronchus, right-sided stent, bronchus intermedius, right lower lobe bronchus, left mainstem bronchus, left upper lobe, and left lower lobe from mild mucus.
Bronchial alveolar lavage/secretions obtained from RLL was performed at RLL.
Instilled 20 cc of NS, suction returned with 10 cc of NS.  Samples sent for Microbiology (Cultures/Viral/Fungal).
Residual secretions suctioned to clear.  Disposable bronchoscope was removed and procedure completed.
The patient tolerated the procedure well.
There were no immediate complications.  Patient will remain in ICU afterward.
ESTIMATED BLOOD LOSS:   None
COMPLICATIONS:    None
SPECIMEN(S): 
BAL of RLL - cultures/micro
IMPRESSION/PLAN: [REDACTED]is a 56 year old-year-old male who presents for bronchoscopy for airway evaluation, therapeutic aspiration, and removal stent.
-Right mainstem anastomosis dehiscence remains closed.
-Prir RUL full thickness erosion/ulceration is now covered with fibrinous exudate/granulation tissue.
-Encourage ongoing aggressive secretion clearance practices.
-Recommend bronchoscopic secretion clearance every 1-2 days, but will defer to transplant team's expertise.
-Ongoing discussions with transplant team regarding timing for next IP bronchoscopy.
Replacement of stent still possible if any signs of worsening of RMS dehiscence site."""

# -------------------------------------------------------------------------
# DATA STRUCTURES
# -------------------------------------------------------------------------

PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 1,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 0, # "Small metallic object" noted, but not explicitly removed in procedure details.
    "airway_dilation": 0, # Traversed narrowed areas, no balloon/dilation mentioned.
    "airway_stent": 0, # "removal stent" in indication, but "cleaned out... right-sided stent" in text. Stent left in place.
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0,
}

# Anchors: (span_text, label, normalized, schema_field, event_id, section, context_prefix)
ANCHORS = [
    # Event 1: Bronchoscopy & Inspection
    ("Flexible Therapeutic Bronchoscope", "PROC_METHOD", "Flexible Bronchoscopy", "method", "E1", "Procedure", "The"),
    ("airway evaluation", "PROC_ACTION", "Inspection", "action", "E1", "Procedure", "advanced for"),
    ("Distal trachea", "ANAT_AIRWAY", "Trachea", "anatomy", "E1", "Findings", "Findings:\n"),
    ("normal", "OBS_LESION", "Normal", "findings", "E1", "Findings", "Distal trachea"),
    
    # Event 2: RMS / Stent / Anastomosis Findings
    ("RMS", "ANAT_AIRWAY", "Right Mainstem Bronchus", "anatomy", "E2", "Findings", "lidocaine applied to the main carina,"),
    ("dehiscence", "OBS_LESION", "Dehiscence", "lesion", "E2", "Findings", "RMS"),
    ("healed", "OBS_LESION", "Healed", "status", "E2", "Findings", "continues to be"),
    ("Small metallic object", "OBS_LESION", "Foreign Body (Hemoclip)", "lesion", "E2", "Findings", "closed.\n"),
    ("hemoclip", "OBS_LESION", "Hemoclip", "lesion", "E2", "Findings", "consistent with likely"),
    ("fibrinous exudate", "OBS_LESION", "Fibrinous Exudate", "lesion", "E2", "Findings", "have"),
    ("granulation tissue", "OBS_LESION", "Granulation Tissue", "lesion", "E2", "Findings", "desired"),
    ("right mainstem bronchus", "ANAT_AIRWAY", "Right Mainstem Bronchus", "anatomy", "E2", "Findings", "medial aspect of the"),
    ("nonobstructing granulation tissue", "OBS_LESION", "Granulation Tissue", "lesion", "E2", "Findings", "mildly heaped"),

    # Event 3: RUL Findings (Ulceration healing)
    ("RUL bronchus", "ANAT_AIRWAY", "RUL Bronchus", "anatomy", "E3", "Findings", "along"),
    ("erosion/ulceration", "OBS_LESION", "Ulceration", "lesion", "E3", "Findings", "fullness thickness"),
    ("fully covered", "OBS_LESION", "Fully Covered", "status", "E3", "Findings", "appears to be"),
    ("early healing", "OBS_LESION", "Healing", "status", "E3", "Findings", "suggesting"),

    # Event 4: RML / Lingula Narrowing
    ("RML take-off", "ANAT_AIRWAY", "RML Bronchus", "anatomy", "E4", "Findings", "ago.  "),
    ("narrowed", "OBS_LESION", "Stenosis", "lesion", "E4", "Findings", "RML take-off is"),
    ("traverse", "PROC_ACTION", "Traverse", "action", "E4", "Findings", "able to"),
    ("Lingula take-off", "ANAT_AIRWAY", "Lingula", "anatomy", "E4", "Findings", "ag.\n"),
    ("narrowed", "OBS_LESION", "Stenosis", "lesion", "E4", "Findings", "Lingula take-off is"),

    # Event 5: LMS Findings
    ("LMS anastomosis", "ANAT_AIRWAY", "Left Mainstem Bronchus", "anatomy", "E5", "Findings", "suctioned to clear (therapeutic aspiration).\n"),
    ("intact", "OBS_LESION", "Intact", "status", "E5", "Findings", "LMS anastomosis"),
    ("mild stenosis", "OBS_LESION", "Stenosis (Mild)", "lesion", "E5", "Findings", "visible sutures and"),

    # Event 6: Therapeutic Aspiration (Global)
    ("Therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", "method", "E6", "Procedure", "17] Successful"),
    ("clean out", "PROC_ACTION", "Suction", "action", "E6", "Procedure", "performed to"),
    ("trachea", "ANAT_AIRWAY", "Trachea", "anatomy", "E6", "Procedure", "clean out the"),
    ("right-sided stent", "DEV_STENT", "Airway Stent", "device", "E6", "Procedure", "bronchus,"),
    ("mild mucus", "OBS_LESION", "Mucus", "lesion", "E6", "Procedure", "lower lobe from"),
    ("Moderate secretions", "OBS_LESION", "Secretions", "lesion", "E6", "Findings", "healthy."),
    ("suctioned to clear", "PROC_ACTION", "Suction", "action", "E6", "Findings", "which were"),
    
    # Event 7: BAL RLL
    ("Bronchial alveolar lavage", "PROC_METHOD", "BAL", "method", "E7", "Procedure", "RLL from mild mucus.\n"),
    ("RLL", "ANAT_LUNG_LOC", "RLL", "anatomy", "E7", "Procedure", "performed at"),
    ("Instilled 20 cc", "MEAS_VOL", "20 cc", "volume_instilled", "E7", "Procedure", "19]"),
    ("returned with 10 cc", "MEAS_VOL", "10 cc", "volume_returned", "E7", "Procedure", "suction"),
    ("Microbiology", "PROC_ACTION", "Sample Sent", "action", "E7", "Procedure", "sent for"),

    # Event 8: Complications
    ("no immediate complications", "OUTCOME_COMPLICATION", "None", "complication", "E8", "Outcomes", "There were"),
    ("tolerated the procedure well", "OUTCOME_SYMPTOMS", "Tolerated Well", "symptoms", "E8", "Outcomes", "The patient"),
]

# -------------------------------------------------------------------------
# UTILS
# -------------------------------------------------------------------------

def create_workbook():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        # Create expected sheets if missing
        for name in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            if name not in wb.sheetnames:
                wb.create_sheet(name)
    return wb

def get_context_substring(full_text, span_text, context_prefix=None, window_size=120):
    """
    Finds the correct start/end indices for a span.
    Logic:
    1. If context_prefix is provided, search for prefix + span_text near each other.
    2. If unique, return index.
    3. If not unique and no prefix, return ambiguous (-1).
    """
    if context_prefix:
        # Search for prefix
        prefix_matches = [m.start() for m in re.finditer(re.escape(context_prefix), full_text)]
        for p_idx in prefix_matches:
            # Look for span_text in a window after prefix
            search_region = full_text[p_idx:p_idx + len(context_prefix) + window_size]
            span_match = search_region.find(span_text)
            if span_match != -1:
                final_start = p_idx + span_match
                return final_start, final_start + len(span_text), "hydrated_prefix_window"
    
    # Fallback: exact count
    matches = [m.start() for m in re.finditer(re.escape(span_text), full_text)]
    if len(matches) == 1:
        return matches[0], matches[0] + len(span_text), "hydrated_unique"
    elif len(matches) > 1:
        return -1, -1, f"ambiguous_count={len(matches)}"
    else:
        return -1, -1, "not_found"

def populate_note_text(wb):
    ws = wb["Note_Text"]
    # Clear existing data if any (keeping header)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    
    # Ensure header
    if ws.cell(1,1).value != "note_id":
        ws.append(["note_id", "source_file", "note_text"])
    
    ws.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

def populate_note_index(wb):
    ws = wb["Note_Index"]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    
    # Headers should already exist in template, but let's assume they map to:
    # A=source_file, B=note_id, C=encounter_id, D=proc_date, E=site, F=reviewer, G=status, H=notes
    # Then flags start at I...
    
    row_data = [
        SOURCE_FILE, NOTE_ID, "", "", "", "", "Pending", ""
    ]
    
    # Flags in order
    flag_keys = list(PROCEDURE_FLAGS.keys())
    for k in flag_keys:
        row_data.append(PROCEDURE_FLAGS[k])
        
    ws.append(row_data)

def populate_spans(wb):
    ws_anno = wb["Span_Annotations"]
    ws_hydra = wb["Span_Hydrated"]
    
    # Headers
    headers = ["source_file", "note_id", "span_id", "section_type", 
               "context_prefix", "span_text", "match_index", 
               "start_char", "end_char", "span_len", 
               "label", "normalized_value", "schema_field", "event_id", 
               "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"]
    
    if ws_anno.max_row == 0 or ws_anno.cell(1,1).value != "source_file":
        ws_anno.append(headers)
    if ws_hydra.max_row == 0 or ws_hydra.cell(1,1).value != "source_file":
        ws_hydra.append(headers)
        
    # Clear old
    if ws_anno.max_row > 1: ws_anno.delete_rows(2, ws_anno.max_row - 1)
    if ws_hydra.max_row > 1: ws_hydra.delete_rows(2, ws_hydra.max_row - 1)

    # Process anchors
    for idx, (txt, lbl, norm, field, eid, sec, ctx) in enumerate(ANCHORS):
        span_id = f"{NOTE_ID}_s{idx+1:03d}"
        
        # Hydrate
        start, end, status = get_context_substring(NOTE_TEXT, txt, ctx)
        length = len(txt)
        
        # Anno Row (Blank start/end)
        row_anno = [
            SOURCE_FILE, NOTE_ID, span_id, sec,
            ctx, txt, "", # match_index
            "", "", f"=LEN(F{idx+2})", # start, end, formula
            lbl, norm, field, eid,
            0, 0, "", "", "", "needs_hydration"
        ]
        ws_anno.append(row_anno)
        
        # Hydra Row (Filled)
        row_hydra = [
            SOURCE_FILE, NOTE_ID, span_id, sec,
            ctx, txt, "",
            start if start != -1 else "", end if end != -1 else "", length,
            lbl, norm, field, eid,
            0, 0, "", "", "", status
        ]
        ws_hydra.append(row_hydra)

def populate_event_log(wb):
    ws = wb["Event_Log"]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
        
    # Headers: source_file, note_id, event_id, event_type, method, anatomy_target, device, needle_gauge, stations, counts, measurements, specimens, findings, is_historical, reviewer, comments, device_size, device_material, outcome_airway_lumen_pre, outcome_airway_lumen_post, outcome_symptoms, outcome_pleural, outcome_complication
    
    events = {
        "E1": {"type": "Bronchoscopy", "method": "Flexible Bronchoscopy", "anatomy": "Airway", "findings": "Normal Trachea"},
        "E2": {"type": "Findings", "anatomy": "RMS", "findings": "Dehiscence healed; Hemoclip present; Granulation tissue"},
        "E3": {"type": "Findings", "anatomy": "RUL", "findings": "Ulceration healed/covered"},
        "E4": {"type": "Findings", "anatomy": "RML, Lingula", "findings": "Narrowed/Stenosis, able to traverse"},
        "E5": {"type": "Findings", "anatomy": "LMS", "findings": "Intact, mild stenosis"},
        "E6": {"type": "Therapeutic Aspiration", "method": "Suction", "anatomy": "Global Airway", "findings": "Mucus/Secretions cleared", "device": "Airway Stent (cleaned)"},
        "E7": {"type": "BAL", "method": "BAL", "anatomy": "RLL", "measurements": "Instilled 20cc, Return 10cc", "specimens": "Microbiology"},
        "E8": {"type": "Outcomes", "outcome_complication": "None", "outcome_symptoms": "Tolerated Well"}
    }
    
    for eid, data in events.items():
        row = [
            SOURCE_FILE, NOTE_ID, eid,
            data.get("type", ""),
            data.get("method", ""),
            data.get("anatomy", ""),
            data.get("device", ""),
            "", # needle
            "", # stations
            "", # counts
            data.get("measurements", ""),
            data.get("specimens", ""),
            data.get("findings", ""),
            0, "", "", # hist, rev, comm
            "", "", # dev_size, dev_mat
            "", "", # lumen pre/post
            data.get("outcome_symptoms", ""),
            "", # pleural
            data.get("outcome_complication", "")
        ]
        ws.append(row)

def populate_v3_registry(wb):
    ws = wb["V3_Registry_JSON"]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    
    # Construct JSON structure
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": [
            {
                "event_id": "E1",
                "type": "Bronchoscopy",
                "method": "Flexible",
                "findings": ["Normal Trachea"]
            },
            {
                "event_id": "E6",
                "type": "Therapeutic Aspiration",
                "target": {"anatomy_type": "Airway", "location": {"lobe": "Multiple"}},
                "method": "Suction",
                "findings": ["Mucus", "Secretions"]
            },
            {
                "event_id": "E7",
                "type": "BAL",
                "target": {"anatomy_type": "Lobe", "location": {"lobe": "RLL"}},
                "measurements": {"volume_instilled": "20 cc", "volume_returned": "10 cc"},
                "specimens": ["Microbiology"]
            }
        ]
    }
    
    ws.append([json.dumps(registry_data, indent=2)])

def main():
    wb = create_workbook()
    populate_note_text(wb)
    populate_note_index(wb)
    populate_spans(wb)
    populate_event_log(wb)
    populate_v3_registry(wb)
    
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    main()