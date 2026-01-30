import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# =============================================================================
# CONFIGURATION & INPUTS
# =============================================================================

NOTE_ID = "note_103"
SOURCE_FILE = "note_103.txt"
PROCEDURE_DATE = "2026-01-12" # inferred or current date for placeholder
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_103 SOURCE_FILE: note_103.txt INDICATION FOR OPERATION:  [REDACTED]is a 56 year old-year-old male who presents with bilateral lung transplant with complication of anastomosis dehiscence and ischemic lung injury.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
PROCEDURE:  
31899 Unlisted Procedure (Trach Change with Mature Tract or Procedure NOS)
31646 Therapeutic aspiration subsequent episodes
31624 Dx bronchoscope/lavage (BAL)    
ANESTHESIA: Local
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Disposable Bronchoscope
PROCEDURE IN DETAIL:
A timeout was performed (confirming the patient's name, procedure type, and procedure location).
The Flexible Therapeutic Bronchoscope was advanced for airway evaluation.  Endobronchial topical lidocaine applied to the main carina, RMS, BI, and LC2.
Initial Airway Inspection Findings:
Distal trachea normal.  
RMS dehiscence continues to be healed and remains closed.
Small metallic object protruding from the posterior membrane near the RMS anastomosis site - consistent with likely hemoclip.
This is slightly more prominent than our bronchoscopy 1 week prior.
Continues to have fibrinous exudate/desired granulation tissue in the donor RMS, proximal RUL bronchus, BI, overlying RML take-off, and overlying RB6 take-off.
All areas appear to be improved from previous.  There is some mildly heaped nonobstructing granulation tissue along the medial aspect of the right mainstem bronchus.
Previously seen area of fullness thickness erosion/ulceration along RUL bronchus anterior wall now appears to be fully covered by fibrinous exudate suggesting early healing.
This is greatly improved from 1 week ago.  RML take-off is narrowed, but able to traverse with disposable bronchoscope "regular" size.
RLL basilar segments appear healthy. Moderate secretions in BI and RLL which were suctioned to clear (therapeutic aspiration).
LMS anastomosis intact with visible sutures and mild stenosis. LUL bronchus with evidence of continued healing with fibrin exudates/desired granulation tissue.
No longer able to see underlying mediastinum/pulmonary artery along medial aspect of LUL.
This entire area continues to look improved from 1 week ag.
Lingula take-off is narrowed, but able to traverse with disposable bronochoscope "regular" size.  LLL bronchus and segments appear health.
Successful therapeutic aspiration was performed to clean out the trachea, right mainstem bronchus, right-sided stent, bronchus intermedius, right lower lobe bronchus, left mainstem bronchus, left upper lobe, and left lower lobe from mild mucus.
Bronchial alveolar lavage/secretions obtained from RLL was performed at RLL.
Instilled 20 cc of NS, suction returned with 10 cc of NS.  Samples sent for Microbiology (Cultures/Viral/Fungal).
Residual secretions suctioned to clear.  Disposable bronchoscope was removed and procedure completed.
The patient tolerated the procedure well.
There were no immediate complications.  Patient will remain in ICU afterward.
ESTIMATED BLOOD LOSS:   None
COMPLICATIONS:    None
SPECIMEN(S): 
BAL of RLL - cultures/micro
IMPRESSION/PLAN: [REDACTED]is a 56 year old-year-old male who presents for bronchoscopy for airway evaluation, therapeutic aspiration, and removal stent.
-Right mainstem anastomosis dehiscence remains closed.
-Prir RUL full thickness erosion/ulceration is now covered with fibrinous exudate/granulation tissue.
-Encourage ongoing aggressive secretion clearance practices.
-Recommend bronchoscopic secretion clearance every 1-2 days, but will defer to transplant team's expertise.
-Ongoing discussions with transplant team regarding timing for next IP bronchoscopy.
Replacement of stent still possible if any signs of worsening of RMS dehiscence site."""

# =============================================================================
# PROCEDURE FLAGS
# =============================================================================

PROC_FLAGS = {
    # Bronchoscopy (23)
    "diagnostic_bronchoscopy": 1,
    "bal": 1,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 0, # Object seen ("metallic object"), but text does not explicitly confirm removal action in body, just "removal stent" in indication which didn't happen.
    "airway_dilation": 0,
    "airway_stent": 0, # Existing stent mentioned, replacement "possible", no placement described.
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural (7)
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# =============================================================================
# SPAN DATA (ANCHOR FIRST)
# =============================================================================

# Define spans with manual curation to ensure high quality "Goldilocks" phrasing.
# Structure: (text, label, normalized_value, event_id, context_prefix)
SPANS = [
    # Event 1: Airway Inspection / Findings
    ("Flexible Therapeutic Bronchoscope", "DEV_INSTRUMENT", "flexible bronchoscope", "evt_01", None),
    ("RMS dehiscence", "OBS_LESION", "dehiscence", "evt_01", None),
    ("healed and remains closed", "OBS_LESION", "healed", "evt_01", "remains closed"),
    ("Small metallic object", "OBS_LESION", "foreign_body", "evt_01", None),
    ("likely hemoclip", "OBS_LESION", "hemoclip", "evt_01", None),
    ("fibrinous exudate/desired granulation tissue", "OBS_LESION", "granulation_tissue", "evt_01", "Continues to have"),
    ("RML take-off is narrowed", "OBS_LESION", "stenosis", "evt_01", None),
    ("able to traverse with disposable bronchoscope", "OUTCOME_AIRWAY_LUMEN_POST", "traversable", "evt_01", "narrowed, but"),
    ("LMS anastomosis intact", "OBS_LESION", "anastomosis_intact", "evt_01", None),
    ("mild stenosis", "OBS_LESION", "stenosis_mild", "evt_01", "sutures and"),
    ("LUL bronchus", "ANAT_AIRWAY", "LUL", "evt_01", "mild stenosis."),
    ("Lingula take-off is narrowed", "OBS_LESION", "stenosis", "evt_01", None),
    ("able to traverse with disposable bronochoscope", "OUTCOME_AIRWAY_LUMEN_POST", "traversable", "evt_01", "narrowed, but"),
    
    # Event 2: Therapeutic Aspiration
    ("Successful therapeutic aspiration", "PROC_METHOD", "therapeutic_aspiration", "evt_02", None),
    ("clean out the trachea", "PROC_ACTION", "suction", "evt_02", None),
    ("right mainstem bronchus", "ANAT_AIRWAY", "RMS", "evt_02", "trachea,"),
    ("right-sided stent", "DEV_STENT", "stent", "evt_02", None),
    ("bronchus intermedius", "ANAT_AIRWAY", "BI", "evt_02", "stent,"),
    ("right lower lobe bronchus", "ANAT_AIRWAY", "RLL", "evt_02", "intermedius,"),
    ("left mainstem bronchus", "ANAT_AIRWAY", "LMS", "evt_02", "lobe bronchus,"),
    ("left upper lobe", "ANAT_AIRWAY", "LUL", "evt_02", "mainstem bronchus,"),
    ("left lower lobe", "ANAT_AIRWAY", "LLL", "evt_02", "upper lobe, and"),
    ("mild mucus", "OBS_LESION", "mucus_mild", "evt_02", "lobe from"),
    ("Moderate secretions", "OBS_LESION", "secretions_moderate", "evt_02", None),
    ("suctioned to clear", "PROC_ACTION", "suction", "evt_02", "RLL which were"),
    
    # Event 3: BAL
    ("Bronchial alveolar lavage", "PROC_METHOD", "bal", "evt_03", None),
    ("RLL", "ANAT_LUNG_LOC", "RLL", "evt_03", "lavage/secretions obtained from"),
    ("Instilled 20 cc of NS", "PROC_ACTION", "instill_saline", "evt_03", None),
    ("20 cc", "MEAS_VOL", "20", "evt_03", "Instilled"),
    ("suction returned with 10 cc", "MEAS_VOL", "10", "evt_03", None),
    ("Samples sent for Microbiology", "PROC_ACTION", "specimen_microbiology", "evt_03", None),
    
    # Outcomes / General
    ("No immediate complications", "OUTCOME_COMPLICATION", "none", "evt_general", None),
    ("appear to be improved from previous", "OUTCOME_SYMPTOMS", "improved", "evt_general", "All areas")
]

# =============================================================================
# EVENT LOG DATA
# =============================================================================

EVENTS = [
    {
        "event_id": "evt_01",
        "event_type": "diagnostic_bronchoscopy",
        "method": "Inspection",
        "anatomy_target": "RMS, LMS, RML, Lingula",
        "findings": "Dehiscence (healed), metallic object, granulation tissue, stenosis (RML, Lingula)",
        "outcome_airway_lumen_post": "traversable",
        "outcome_symptoms": "improved",
        "outcome_complication": "none"
    },
    {
        "event_id": "evt_02",
        "event_type": "therapeutic_aspiration",
        "method": "Suction",
        "anatomy_target": "Trachea, RMS, BI, RLL, LMS, LUL, LLL",
        "findings": "Mild mucus, Moderate secretions",
        "counts": "",
        "outcome_complication": "none"
    },
    {
        "event_id": "evt_03",
        "event_type": "bal",
        "method": "BAL",
        "anatomy_target": "RLL",
        "measurements": "Instilled 20cc, Return 10cc",
        "specimens": "Microbiology",
        "outcome_complication": "none"
    }
]

# =============================================================================
# HELPER CLASSES & FUNCTIONS
# =============================================================================

class Phase0Generator:
    def __init__(self, template_path, output_path, note_text):
        self.template_path = template_path
        self.output_path = output_path
        self.note_text = note_text
        self.wb = None
        
    def load_workbook(self):
        if not os.path.exists(self.template_path):
            # Create a basic workbook if template doesn't exist (fallback)
            self.wb = openpyxl.Workbook()
            default_sheets = ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]
            for sheet in default_sheets:
                if sheet not in self.wb.sheetnames:
                    self.wb.create_sheet(sheet)
            if "Sheet" in self.wb.sheetnames:
                del self.wb["Sheet"]
        else:
            self.wb = openpyxl.load_workbook(self.template_path)

    def write_note_text(self):
        ws = self.wb["Note_Text"]
        # Assuming headers exist: note_id, source_file, note_text
        # Find next empty row
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=NOTE_ID)
        ws.cell(row=row, column=2, value=SOURCE_FILE)
        ws.cell(row=row, column=3, value=self.note_text)

    def write_note_index(self):
        ws = self.wb["Note_Index"]
        row = ws.max_row + 1
        
        # Metadata columns (A-H)
        ws.cell(row=row, column=1, value=SOURCE_FILE)
        ws.cell(row=row, column=2, value=NOTE_ID)
        ws.cell(row=row, column=4, value=PROCEDURE_DATE)
        ws.cell(row=row, column=7, value="Complete")
        
        # Flag columns (starts at I, index 9)
        # Note: The template usually expects specific headers. 
        # We will iterate through PROC_FLAGS keys and try to match headers if possible, 
        # or just assume the order defined in the Prompt instructions if headers are missing.
        # For robustness, we assume the order in the Prompt description:
        # Bronchoscopy (23) then Pleural (7).
        
        flag_keys = [
            "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", 
            "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy", 
            "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration", 
            "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation", 
            "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", 
            "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
            "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis", 
            "pleural_biopsy", "fibrinolytic_therapy"
        ]
        
        # Starting column for flags (usually col 9)
        col_idx = 9
        for key in flag_keys:
            ws.cell(row=row, column=col_idx, value=PROC_FLAGS.get(key, 0))
            col_idx += 1

    def hydrate_span(self, span_text, context_prefix=None):
        """
        Locates span_text in NOTE_TEXT.
        Returns: start_char, end_char, hydration_status
        """
        if not span_text:
            return None, None, "missing_text"
            
        # 1. Exact unique match
        count = self.note_text.count(span_text)
        if count == 1:
            start = self.note_text.find(span_text)
            return start, start + len(span_text), "hydrated_unique"
            
        # 2. Context prefix match
        if count > 1 and context_prefix:
            # Iterate through all occurrences
            search_start = 0
            while True:
                idx = self.note_text.find(span_text, search_start)
                if idx == -1:
                    break
                
                # Check preceding window (120 chars)
                window_start = max(0, idx - 120)
                preceding_text = self.note_text[window_start:idx]
                
                if context_prefix in preceding_text:
                    return idx, idx + len(span_text), "hydrated_prefix_window"
                
                search_start = idx + 1
                
        # 3. Fallback: First occurrence if ambiguous and no valid context found
        if count > 1:
            start = self.note_text.find(span_text)
            return start, start + len(span_text), f"ambiguous_count={count}"
            
        return None, None, "not_found"

    def write_spans(self):
        ws_anno = self.wb["Span_Annotations"]
        ws_hydra = self.wb["Span_Hydrated"]
        
        # Headers should exist. Appending rows.
        start_row_anno = ws_anno.max_row + 1
        start_row_hydra = ws_hydra.max_row + 1
        
        for i, (text, label, norm, evt_id, ctx) in enumerate(SPANS):
            span_id = f"{NOTE_ID}_s{i+1:02d}"
            
            # 1. Write to Span_Annotations (Anchor First - No Offsets)
            # Columns: source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, start, end, len...
            ws_anno.cell(row=start_row_anno+i, column=1, value=SOURCE_FILE)
            ws_anno.cell(row=start_row_anno+i, column=2, value=NOTE_ID)
            ws_anno.cell(row=start_row_anno+i, column=3, value=span_id)
            ws_anno.cell(row=start_row_anno+i, column=5, value=ctx)
            ws_anno.cell(row=start_row_anno+i, column=6, value=text)
            # Start/End left blank in Anno sheet
            ws_anno.cell(row=start_row_anno+i, column=10, value=len(text))
            ws_anno.cell(row=start_row_anno+i, column=11, value=label)
            ws_anno.cell(row=start_row_anno+i, column=12, value=norm)
            ws_anno.cell(row=start_row_anno+i, column=14, value=evt_id)
            ws_anno.cell(row=start_row_anno+i, column=20, value="needs_hydration")

            # 2. Compute Hydration
            start, end, status = self.hydrate_span(text, ctx)
            
            # 3. Write to Span_Hydrated
            ws_hydra.cell(row=start_row_hydra+i, column=1, value=SOURCE_FILE)
            ws_hydra.cell(row=start_row_hydra+i, column=2, value=NOTE_ID)
            ws_hydra.cell(row=start_row_hydra+i, column=3, value=span_id)
            ws_hydra.cell(row=start_row_hydra+i, column=5, value=ctx)
            ws_hydra.cell(row=start_row_hydra+i, column=6, value=text)
            
            if start is not None:
                ws_hydra.cell(row=start_row_hydra+i, column=8, value=start)
                ws_hydra.cell(row=start_row_hydra+i, column=9, value=end)
                
            ws_hydra.cell(row=start_row_hydra+i, column=10, value=len(text))
            ws_hydra.cell(row=start_row_hydra+i, column=11, value=label)
            ws_hydra.cell(row=start_row_hydra+i, column=12, value=norm)
            ws_hydra.cell(row=start_row_hydra+i, column=14, value=evt_id)
            ws_hydra.cell(row=start_row_hydra+i, column=20, value=status)

    def write_event_log(self):
        ws = self.wb["Event_Log"]
        row = ws.max_row + 1
        
        for evt in EVENTS:
            ws.cell(row=row, column=1, value=SOURCE_FILE)
            ws.cell(row=row, column=2, value=NOTE_ID)
            ws.cell(row=row, column=3, value=evt["event_id"])
            ws.cell(row=row, column=4, value=evt["event_type"])
            ws.cell(row=row, column=5, value=evt.get("method"))
            ws.cell(row=row, column=6, value=evt.get("anatomy_target"))
            ws.cell(row=row, column=11, value=evt.get("measurements"))
            ws.cell(row=row, column=12, value=evt.get("specimens"))
            ws.cell(row=row, column=13, value=evt.get("findings"))
            
            # Granular Outcomes
            ws.cell(row=row, column=19, value=evt.get("outcome_airway_lumen_pre"))
            ws.cell(row=row, column=20, value=evt.get("outcome_airway_lumen_post"))
            ws.cell(row=row, column=21, value=evt.get("outcome_symptoms"))
            ws.cell(row=row, column=22, value=evt.get("outcome_pleural"))
            ws.cell(row=row, column=23, value=evt.get("outcome_complication"))
            row += 1

    def write_v3_registry_json(self):
        ws = self.wb["V3_Registry_JSON"]
        
        data = {
            "schema_version": "v3.0",
            "note_id": NOTE_ID,
            "no_immediate_complications": True,
            "procedures": []
        }
        
        for evt in EVENTS:
            proc_obj = {
                "event_id": evt["event_id"],
                "type": evt["event_type"],
                "method": evt.get("method"),
                "target": {"anatomy": evt.get("anatomy_target")},
                "outcomes": {
                    "airway": {
                        "lumen_post": evt.get("outcome_airway_lumen_post")
                    },
                    "symptoms": evt.get("outcome_symptoms"),
                    "complications": evt.get("outcome_complication")
                }
            }
            data["procedures"].append(proc_obj)
            
        json_str = json.dumps(data, indent=2)
        ws.cell(row=ws.max_row + 1, column=1, value=json_str)

    def save(self):
        self.wb.save(self.output_path)
        print(f"Generated {self.output_path}")

# =============================================================================
# MAIN EXECUTION
# =============================================================================

if __name__ == "__main__":
    generator = Phase0Generator(TEMPLATE_PATH, OUTPUT_PATH, NOTE_TEXT)
    generator.load_workbook()
    generator.write_note_text()
    generator.write_note_index()
    generator.write_spans()
    generator.write_event_log()
    generator.write_v3_registry_json()
    generator.save()