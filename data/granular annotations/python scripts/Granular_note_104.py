import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os
from datetime import datetime

# ==========================================
# 1. INPUT DATA
# ==========================================
NOTE_ID = "note_104"
SOURCE_FILE = "note_104.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_104 SOURCE_FILE: note_104.txt INDICATION FOR OPERATION:  [REDACTED]is a 56 year old-year-old male who presents with respiratory failure.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: J96.90 Respiratory Failure
POSTOPERATIVE DIAGNOSIS:  J96.90 Respiratory Failure
PROCEDURE:  
31646 Therapeutic aspiration subsequent episodes
31622 Dx bronchoscope/cell washing          
31624 Dx bronchoscope/lavage (BAL)    
31630 Balloon dilation
50 Bilateral Procedures (Procedure done on both sides of the body)
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required a bilateral procedure today when this procedure would typically be unilateral.
Apply to: 31624 Dx bronchoscope/lavage (BAL)    .
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
ESTIMATED BLOOD LOSS:   Minimum
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
Initial Airway Inspection Findings:
The airway was inspected from above the tracheostomy tube. The vocal cords were normal appearing.
The trach was in good position and the balloon was inflated. Secretions were suctioned and lidocaine was applied.
The airway was inspected via the tracheostomy. The right sided stent was in good position.
There was significant necrosis/granulation tissue around the suture site and medially along the transplanted lung.
This covered the orifice of the right middle lobe. The right lower lobe was preserved.
The portion of the right upper lobe take off visible through the stent appeared intact, but there was an area around 9 o'clock which could be injured.
The left lung anastamosis site had similar medial necrosis and granulation tissue.
The lingula was patent but obscured with the necrosis. The left upper lobe anterior segment had some necrosis and the left upper lobe proper was preserved.
The left lower lobe was preserved.
Endobronchial biopsy was performed at RML Carina (RC2).  Lesion was successfully removed.
Samples sent for Pathology. 
Balloon dilation was performed at RML Carina (RC2).
8/9/10 Elation balloon was used to perform dilation to 8 mm and to 9 mm at the RML Carina (RC2).
Total 2 inflations with dilation time of 60 seconds each.
Bronchial alveolar lavage was performed at Lateral Segment of RML (RB4) and Medial Segment of RML (RB5).
Instilled 40 cc of NS, suction returned with 25 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
Successful therapeutic aspiration was performed to clean out the Trachea (Middle 1/3), Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
Right middle lobe BAL
Right middle lobe EBBX
IMPRESSION/PLAN: [REDACTED]is a 56 year old-year-old male who presents for bronchoscopy for respiratory failure and airway dilation.
-Follow up bronchoscopic lab work
-Follow up bronchoscopy later this week or early next week"""

# ==========================================
# 2. DEFINITIONS & CONFIG
# ==========================================

# Procedure Flags (0/1)
PROC_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 1,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 1, # "Endobronchial biopsy was performed"
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1, # "Successful therapeutic aspiration"
    "foreign_body_removal": 0,
    "airway_dilation": 1, # "Balloon dilation was performed"
    "airway_stent": 0, # Existing stent mentioned, but no new stent placed
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# Anchors: list of dicts
# Fields: span_text, label, normalized_value, event_id, context_prefix (opt), schema_field (opt), is_negated (opt), is_historical (opt)
ANCHORS = [
    # Event 1: EBBX
    {
        "span_text": "Endobronchial biopsy",
        "label": "PROC_METHOD",
        "normalized_value": "Endobronchial Biopsy",
        "event_id": "evt_01",
        "context_prefix": "left lower lobe was preserved.\n"
    },
    {
        "span_text": "RML Carina (RC2)",
        "label": "ANAT_AIRWAY",
        "normalized_value": "RML Carina",
        "event_id": "evt_01",
        "context_prefix": "Endobronchial biopsy was performed at "
    },
    {
        "span_text": "Lesion",
        "label": "OBS_LESION",
        "normalized_value": "Lesion",
        "event_id": "evt_01",
        "context_prefix": "(RC2).  "
    },
    {
        "span_text": "successfully removed",
        "label": "PROC_ACTION",
        "normalized_value": "Removed",
        "event_id": "evt_01",
    },

    # Event 2: Balloon Dilation
    {
        "span_text": "Balloon dilation",
        "label": "PROC_METHOD",
        "normalized_value": "Balloon Dilation",
        "event_id": "evt_02",
        "context_prefix": "Samples sent for Pathology. \n"
    },
    {
        "span_text": "RML Carina (RC2)",
        "label": "ANAT_AIRWAY",
        "normalized_value": "RML Carina",
        "event_id": "evt_02",
        "context_prefix": "Balloon dilation was performed at "
    },
    {
        "span_text": "8/9/10 Elation balloon",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "Elation Balloon 8-9-10mm",
        "event_id": "evt_02"
    },
    {
        "span_text": "8 mm",
        "label": "MEAS_AIRWAY_DIAM",
        "normalized_value": "8",
        "event_id": "evt_02",
        "schema_field": "outcomes.airway.lumen_post"
    },
    {
        "span_text": "9 mm",
        "label": "MEAS_AIRWAY_DIAM",
        "normalized_value": "9",
        "event_id": "evt_02",
        "schema_field": "outcomes.airway.lumen_post"
    },
    {
        "span_text": "2 inflations",
        "label": "MEAS_COUNT",
        "normalized_value": "2",
        "event_id": "evt_02"
    },
    {
        "span_text": "60 seconds",
        "label": "MEAS_TIME",
        "normalized_value": "60 s",
        "event_id": "evt_02"
    },

    # Event 3: BAL
    {
        "span_text": "Bronchial alveolar lavage",
        "label": "PROC_METHOD",
        "normalized_value": "Bronchoalveolar Lavage",
        "event_id": "evt_03",
        "context_prefix": "dilation time of 60 seconds each.\n"
    },
    {
        "span_text": "Lateral Segment of RML (RB4)",
        "label": "ANAT_LUNG_LOC",
        "normalized_value": "RML Lateral Segment",
        "event_id": "evt_03"
    },
    {
        "span_text": "Medial Segment of RML (RB5)",
        "label": "ANAT_LUNG_LOC",
        "normalized_value": "RML Medial Segment",
        "event_id": "evt_03"
    },
    {
        "span_text": "40 cc",
        "label": "MEAS_VOL",
        "normalized_value": "40",
        "event_id": "evt_03",
        "context_prefix": "Instilled "
    },
    {
        "span_text": "25 cc",
        "label": "MEAS_VOL",
        "normalized_value": "25",
        "event_id": "evt_03",
        "context_prefix": "returned with "
    },

    # Event 4: Therapeutic Aspiration
    {
        "span_text": "therapeutic aspiration",
        "label": "PROC_METHOD",
        "normalized_value": "Therapeutic Aspiration",
        "event_id": "evt_04",
        "context_prefix": "Successful "
    },
    {
        "span_text": "Trachea (Middle 1/3)",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Trachea Middle 1/3",
        "event_id": "evt_04"
    },
    {
        "span_text": "Trachea (Distal 1/3)",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Trachea Distal 1/3",
        "event_id": "evt_04"
    },
    {
        "span_text": "Right Mainstem",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Right Mainstem Bronchus",
        "event_id": "evt_04"
    },
    {
        "span_text": "Bronchus Intermedius",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Bronchus Intermedius",
        "event_id": "evt_04"
    },
    {
        "span_text": "Left Mainstem",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Left Mainstem Bronchus",
        "event_id": "evt_04"
    },
    {
        "span_text": "Carina",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Carina",
        "event_id": "evt_04",
        "context_prefix": "Left Mainstem, "
    },
    {
        "span_text": "mucus",
        "label": "OBS_FINDING",
        "normalized_value": "Mucus",
        "event_id": "evt_04",
        "context_prefix": "from "
    },

    # General / Outcome
    {
        "span_text": "respiratory failure",
        "label": "OBS_FINDING",
        "normalized_value": "Respiratory Failure",
        "event_id": "evt_00",
        "context_prefix": "presents with "
    },
    {
        "span_text": "necrosis/granulation tissue",
        "label": "OBS_FINDING",
        "normalized_value": "Necrosis/Granulation Tissue",
        "event_id": "evt_00",
        "context_prefix": "significant "
    },
    {
        "span_text": "suture site",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Anastomosis Site",
        "event_id": "evt_00",
        "context_prefix": "around the "
    },
    {
        "span_text": "no immediate complications",
        "label": "OUTCOME_COMPLICATION",
        "normalized_value": "None",
        "event_id": "evt_00",
        "is_negated": True
    },
    
    # Existing Stent observations (Historical/Context)
    {
        "span_text": "right sided stent",
        "label": "DEV_STENT",
        "normalized_value": "Stent (Right)",
        "event_id": "evt_hist_01",
        "is_historical": True
    },
    {
        "span_text": "good position",
        "label": "OBS_FINDING",
        "normalized_value": "Good Position",
        "event_id": "evt_hist_01",
        "context_prefix": "stent was in "
    },
]

# ==========================================
# 3. HELPER FUNCTIONS
# ==========================================

def get_row_data(ws, row):
    return [cell.value for cell in ws[row]]

def clean_text(text):
    if not text: return ""
    return text.strip().replace('\r\n', '\n')

def hydrate_span(full_text, span_data):
    """
    Finds start/end chars for a span.
    Logic:
    1. Exact count == 1 -> easy.
    2. Context prefix -> search window before occurences.
    3. Match index -> pick nth.
    4. Ambiguous.
    """
    txt = span_data['span_text']
    if not txt or txt not in full_text:
        return None, None, "not_found"
    
    matches = [m.start() for m in re.finditer(re.escape(txt), full_text)]
    if not matches:
        return None, None, "not_found"

    # Case 1: Unique
    if len(matches) == 1:
        return matches[0], matches[0] + len(txt), "hydrated_unique"
    
    # Case 2: Context Prefix
    prefix = span_data.get('context_prefix')
    if prefix:
        # Check window 120 chars before
        for m_start in matches:
            window_start = max(0, m_start - 120)
            pre_window = full_text[window_start:m_start]
            # flexible match: remove newlines/spaces for check
            if prefix.strip() in pre_window or prefix.strip().replace('\n', ' ') in pre_window.replace('\n', ' '):
                return m_start, m_start + len(txt), "hydrated_prefix_window"
    
    # Case 3: Fallback (first or ambiguous)
    # Since we can't reliably do match_index without user input in this flow,
    # we'll default to the first one but mark as ambiguous if > 1 and no prefix matched
    return matches[0], matches[0] + len(txt), f"ambiguous_count={len(matches)}"

# ==========================================
# 4. MAIN GENERATION LOGIC
# ==========================================

def generate_workbook():
    # Load template
    if not os.path.exists(TEMPLATE_PATH):
        # Fallback if template missing (should not happen in real usage)
        wb = openpyxl.Workbook()
        ws_flags = wb.create_sheet("Note_Index")
        ws_text = wb.create_sheet("Note_Text")
        ws_span = wb.create_sheet("Span_Annotations")
        ws_hydrated = wb.create_sheet("Span_Hydrated")
        ws_events = wb.create_sheet("Event_Log")
        ws_v3 = wb.create_sheet("V3_Procedure_Events")
        ws_json = wb.create_sheet("V3_Registry_JSON")
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    
    # 1. Note_Text
    if "Note_Text" not in wb.sheetnames: wb.create_sheet("Note_Text")
    ws_text = wb["Note_Text"]
    # Header: note_id, source_file, note_text
    if ws_text.max_row == 1 and ws_text["A1"].value is None:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    if "Note_Index" not in wb.sheetnames: wb.create_sheet("Note_Index")
    ws_index = wb["Note_Index"]
    
    # Headers should be pre-existing in template, but we map columns
    # A=source_file, B=note_id, C=encounter, D=date, E=site, F=reviewer, G=status, H=notes
    # I..AL = flags
    flag_keys = list(PROC_FLAGS.keys())
    
    # Construct row
    row_meta = [SOURCE_FILE, NOTE_ID, "", "", "", "", "Pending", ""]
    row_flags = [PROC_FLAGS[k] for k in flag_keys]
    ws_index.append(row_meta + row_flags)

    # 3. Span_Annotations (Anchor First)
    if "Span_Annotations" not in wb.sheetnames: wb.create_sheet("Span_Annotations")
    ws_spans = wb["Span_Annotations"]
    # Header: source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, start, end, len, label, norm_val, schema, event, neg, hist, time, rev, comm, hyd_status
    
    # 4. Span_Hydrated
    if "Span_Hydrated" not in wb.sheetnames: wb.create_sheet("Span_Hydrated")
    ws_hydrated = wb["Span_Hydrated"]
    
    span_id_counter = 1
    
    # Sort ANCHORS by occurrence in text to be tidy? Not strictly necessary.
    
    for anchor in ANCHORS:
        s_text = anchor["span_text"]
        # Hydrate
        start, end, status = hydrate_span(NOTE_TEXT, anchor)
        length = len(s_text)
        
        # Build Row Data
        span_id = f"{NOTE_ID}_span_{span_id_counter:03d}"
        
        row_base = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure", 
            anchor.get("context_prefix", ""), s_text, "", 
            "", "", f"=LEN(F{ws_spans.max_row+1})", # Formula for len
            anchor.get("label", ""), anchor.get("normalized_value", ""), 
            anchor.get("schema_field", ""), anchor.get("event_id", ""),
            anchor.get("is_negated", False), anchor.get("is_historical", False), "", "", "", "needs_hydration"
        ]
        
        ws_spans.append(row_base)
        
        # Hydrated Row
        row_hyd = list(row_base)
        row_hyd[7] = start if start is not None else ""
        row_hyd[8] = end if end is not None else ""
        row_hyd[19] = status
        ws_hydrated.append(row_hyd)
        
        span_id_counter += 1

    # 5. Event_Log
    if "Event_Log" not in wb.sheetnames: wb.create_sheet("Event_Log")
    ws_events = wb["Event_Log"]
    
    # Define Events explicitly for the Log
    events_data = [
        {
            "id": "evt_01", "type": "Procedure", "method": "Endobronchial Biopsy", 
            "anatomy": "RML Carina (RC2)", "finding": "Lesion removed",
            "specimen": "RML EBBX"
        },
        {
            "id": "evt_02", "type": "Procedure", "method": "Balloon Dilation", 
            "anatomy": "RML Carina (RC2)", "device": "Elation Balloon 8-9-10mm",
            "measurements": "8mm, 9mm", "outcome_airway_post": "9mm"
        },
        {
            "id": "evt_03", "type": "Procedure", "method": "BAL", 
            "anatomy": "RML (Lateral/Medial)", "counts": "40cc in, 25cc out",
            "specimen": "RML BAL"
        },
        {
            "id": "evt_04", "type": "Procedure", "method": "Therapeutic Aspiration", 
            "anatomy": "Trachea, Mainstems, Lobars", "finding": "Mucus"
        }
    ]
    
    for e in events_data:
        # Mapping simple dict to columns is approximate but helpful for Phase 0
        # Columns: source, note, event_id, type, method, anatomy, device, gauge, station, count, meas, spec, find, hist, rev, comm, size, mat, pre, post, symp, pleur, compl
        row = [
            SOURCE_FILE, NOTE_ID, e["id"], e["type"], e.get("method"),
            e.get("anatomy"), e.get("device"), "", "", e.get("counts"), e.get("measurements"),
            e.get("specimen"), e.get("finding"), "", "", "",
            "", "", "", e.get("outcome_airway_post"), "", "", ""
        ]
        ws_events.append(row)

    # 6. V3_Procedure_Events
    if "V3_Procedure_Events" not in wb.sheetnames: wb.create_sheet("V3_Procedure_Events")
    ws_v3 = wb["V3_Procedure_Events"]
    # Just one example row to show structure capability
    # note_id, event_id, type, target...
    
    # 7. JSON
    if "V3_Registry_JSON" not in wb.sheetnames: wb.create_sheet("V3_Registry_JSON")
    ws_json = wb["V3_Registry_JSON"]
    
    json_out = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": events_data
    }
    ws_json.append([json.dumps(json_out, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()