import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import datetime
import os

# ==========================================
# INPUT DATA
# ==========================================
NOTE_ID = "note_105"
SOURCE_FILE = "note_105.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_105 SOURCE_FILE: note_105.txt INDICATION FOR OPERATION:  [REDACTED]is a 56 year old-year-old male who presents with airway evaluation and stent change.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient/surrogate in detail.
Patient/surrogate indicated a wish to proceed with surgery and informed consent was signed.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
PROCEDURE:  
31646 Therapeutic aspiration subsequent episodes
31622 Dx bronchoscope/cell washing          
ANESTHESIA: Local ONLY
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
ESTIMATED BLOOD LOSS:   None
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
A timeout was performed (confirming the patient's name, procedure type, and procedure location).
Lidocaine instilled into tracheostomy tube.
Disposable bronchoscope advanced.  Lidocaine applied to main carina, RMS, BI, and LC2.
Initial Airway Inspection Findings:
Frothy, non obstructing white secretions in RMS, BI, RML, RLL, LMS, LLL.
Successful therapeutic aspiration was performed to clean out the distal trachea RMS, BI, RML, RLL, LMS, LLL from mucus.
Uncovered metallic stent in good position - appropriately spanning from mid RMS to proximal BI, fully covering and closing the prior area of RMS anastomosis dehiscence.
Stent is intentionally "jailing off" RUL bronchus but uncovered perforations between struts allowing for ventilation of RUL;
minimal secretions overlying this area.  Able to visualize RUL bronchus and RB1 & RB2 from the RMS bronchus, but unable to visualize RB3.
RUL bronchus appears to have evidence of healing with fibrin exudates/desired granulation tissue.
Stent is perhaps mildly undersized but no signs of migration nor significant granulation over struts of stent.
Continues to have significant ischemic and necrotic debris in the donor RMS, BI, overlying RML take-off, and overlying RB6 take-off.
Leading to moderate stenosis of RML bronchus.  RLL basilar segments appear healthy.
LMS anastomosis intact with visible sutures and mild stenosis.  LUL bronchus with evidence of healing with fibrin exudates/desired granulation tissue.
No longer able to see underlying mediastinum/pulmonary artery along medial aspect of LUL.  LLL bronchus and segments appear health.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation patient will remain in ICU in similar condition to prior.
SPECIMEN(S): 
None
IMPRESSION/PLAN: [REDACTED]is a 56 year old-year-old male who presents for bronchoscopy for stent exchange.
-14x40 B.S.
uncovered metallic stent in good position.
-No plan for bronchoscopy by IP team remainder of this week unless clinical change.
-Continue to have issues with secretion clearance.  2-3x weekly bronchoscopy clean-out may be warranted until patient becomes stronger/able to clear own secretions.
Will defer to lung transplant team's discretion/expertise.  Could be done under local anesthesia only to minimize interruptions in nutrition/PT.
-Likely bronchoscopy by IP team early next week (likely Monday).  Discussions with primary team regarding RML bronchus dilation.
-Possible stent exchange late next week or the following week - stent is mildly undersized, so will consider upsizing to 16x40 B.S.
SEMS."""

# ==========================================
# CONFIGURATION & CLASSES
# ==========================================

class Span:
    def __init__(self, span_text, label, normalized_value=None, schema_field=None, 
                 event_id=None, context_prefix=None, is_negated=False, is_historical=False):
        self.span_text = span_text
        self.label = label
        self.normalized_value = normalized_value
        self.schema_field = schema_field
        self.event_id = event_id
        self.context_prefix = context_prefix
        self.is_negated = is_negated
        self.is_historical = is_historical
        # Hydration placeholders
        self.start_char = None
        self.end_char = None
        self.hydration_status = "needs_hydration"

# Procedure Flags Mapping
PROC_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 1, # Coded 31622
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1, # Coded 31646
    "foreign_body_removal": 0,
    "airway_dilation": 0, # Planned but not done
    "airway_stent": 0, # Planned exchange deferred
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# ==========================================
# EXTRACTION LOGIC
# ==========================================

def extract_spans(text):
    spans = []
    
    # Event 1: Therapeutic Aspiration
    e1_id = "evt_01"
    spans.append(Span("Therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", "method", e1_id, context_prefix="31646 "))
    spans.append(Span("clean out", "PROC_ACTION", "Aspiration", "action", e1_id, context_prefix="performed to "))
    
    # Anatomy for Aspiration
    targets = ["distal trachea", "RMS", "BI", "RML", "RLL", "LMS", "LLL"]
    for t in targets:
        spans.append(Span(t, "ANAT_AIRWAY", t, "target.anatomy", e1_id, context_prefix="clean out the "))
        
    spans.append(Span("mucus", "OBS_LESION", "Secretions", "findings", e1_id, context_prefix="from "))
    
    # Event 2: Stent Evaluation (Observation)
    e2_id = "evt_02"
    spans.append(Span("Uncovered metallic stent", "DEV_STENT", "Uncovered Metallic Stent", "device", e2_id, context_prefix="source: 9] "))
    spans.append(Span("good position", "OBS_LESION", "Good Position", "findings", e2_id, context_prefix="stent in "))
    spans.append(Span("spanning from mid RMS to proximal BI", "ANAT_AIRWAY", "RMS to BI", "target.anatomy", e2_id))
    spans.append(Span("jailing off", "OBS_LESION", "Jailing", "findings", e2_id))
    spans.append(Span("RUL bronchus", "ANAT_AIRWAY", "RUL", "target.anatomy", e2_id, context_prefix="jailing off\" "))
    spans.append(Span("mildly undersized", "OBS_LESION", "Undersized", "findings", e2_id, context_prefix="perhaps "))

    # Event 3: Airway Inspection (Stenosis/Necrosis)
    e3_id = "evt_03"
    spans.append(Span("necrotic debris", "OBS_LESION", "Necrotic Debris", "findings", e3_id, context_prefix="ischemic and "))
    spans.append(Span("RMS", "ANAT_AIRWAY", "RMS", "target.anatomy", e3_id, context_prefix="donor "))
    spans.append(Span("BI", "ANAT_AIRWAY", "BI", "target.anatomy", e3_id, context_prefix="RMS, "))
    spans.append(Span("moderate stenosis", "OBS_LESION", "Stenosis", "findings", e3_id, context_prefix="Leading to "))
    spans.append(Span("RML bronchus", "ANAT_AIRWAY", "RML", "target.anatomy", e3_id, context_prefix="stenosis of "))
    spans.append(Span("mild stenosis", "OBS_LESION", "Stenosis", "findings", e3_id, context_prefix="sutures and "))
    spans.append(Span("LMS anastomosis", "ANAT_AIRWAY", "LMS", "target.anatomy", e3_id))
    
    # Procedure Methods / Tools
    spans.append(Span("Flexible Therapeutic Bronchoscope", "DEV_INSTRUMENT", "Flexible Bronchoscope", "device", "evt_global"))
    spans.append(Span("31622", "PROC_METHOD", "Bronchoscopy", "method", "evt_global"))
    
    # Outcomes
    e_outcome = "evt_outcome"
    spans.append(Span("tolerated the procedure well", "OUTCOME_SYMPTOMS", "Tolerated Well", "outcomes.symptoms", e_outcome))
    spans.append(Span("No immediate complications", "OUTCOME_COMPLICATION", "None", "outcomes.complications", e_outcome, is_negated=True))
    
    # Planned
    spans.append(Span("16x40 B.S.", "DEV_STENT_SIZE", "16x40", "plan", "evt_plan", context_prefix="upsizing to "))
    spans.append(Span("SEMS", "DEV_STENT_MATERIAL", "SEMS", "plan", "evt_plan", context_prefix="16x40 B.S.\n"))

    return spans

def hydrate_spans(spans, full_text):
    hydrated = []
    
    for s in spans:
        start = -1
        end = -1
        status = "ambiguous"
        
        # Strategy 1: Exact Count == 1
        count = full_text.count(s.span_text)
        if count == 1:
            start = full_text.find(s.span_text)
            status = "hydrated_unique"
        
        # Strategy 2: Context Prefix
        elif count > 1 and s.context_prefix:
            # Find all occurrences
            occurrences = [m.start() for m in re.finditer(re.escape(s.span_text), full_text)]
            for occ in occurrences:
                # Check window before
                window_start = max(0, occ - 120)
                pre_window = full_text[window_start:occ]
                # Loose check for prefix
                if s.context_prefix in pre_window:
                    start = occ
                    status = "hydrated_prefix_window"
                    break
        
        # Strategy 3: Fallback first match (if ambiguous but we need to place it)
        if start == -1 and count > 0:
             start = full_text.find(s.span_text)
             status = f"hydrated_fallback_first_of_{count}"

        if start != -1:
            end = start + len(s.span_text)
            s.start_char = start
            s.end_char = end
            s.hydration_status = status
            hydrated.append(s)
        else:
            # Append even if not found, but marked missing
            s.hydration_status = "missing"
            hydrated.append(s)
            
    return hydrated

# ==========================================
# EXCEL GENERATION
# ==========================================

def generate_workbook():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        print(f"Error: Template {TEMPLATE_PATH} not found.")
        return

    # 1. Note_Text
    ws_text = wb["Note_Text"]
    # Assuming header exists, append
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws_index = wb["Note_Index"]
    row_meta = [
        SOURCE_FILE, NOTE_ID, "", "", "", "", "Ready", ""
    ]
    # Append flags
    flag_values = [PROC_FLAGS.get(k, 0) for k in PROC_FLAGS]
    ws_index.append(row_meta + flag_values)

    # 3. Span extraction & Hydration
    raw_spans = extract_spans(NOTE_TEXT)
    hydrated_spans = hydrate_spans(raw_spans, NOTE_TEXT)

    # 4. Write Span_Annotations (Anchor First)
    ws_annot = wb["Span_Annotations"]
    for s in hydrated_spans:
        row = [
            SOURCE_FILE, NOTE_ID, f"span_{hydrated_spans.index(s)+1:03d}", "procedure_note",
            s.context_prefix, s.span_text, "", # match_index left blank
            "", "", f"=LEN(F{ws_annot.max_row+1})", # formulas for len
            s.label, s.normalized_value, s.schema_field, s.event_id,
            s.is_negated, s.is_historical, "", "", "", "needs_hydration"
        ]
        ws_annot.append(row)

    # 5. Write Span_Hydrated
    ws_hydro = wb["Span_Hydrated"]
    for s in hydrated_spans:
        row = [
            SOURCE_FILE, NOTE_ID, f"span_{hydrated_spans.index(s)+1:03d}", "procedure_note",
            s.context_prefix, s.span_text, "",
            s.start_char, s.end_char, len(s.span_text),
            s.label, s.normalized_value, s.schema_field, s.event_id,
            s.is_negated, s.is_historical, "", "", "", s.hydration_status
        ]
        ws_hydro.append(row)

    # 6. Event_Log
    ws_event = wb["Event_Log"]
    # Event 1: Aspiration
    ws_event.append([
        SOURCE_FILE, NOTE_ID, "evt_01", "Therapeutic Aspiration", "Suction",
        "RMS, BI, RML, RLL, LMS, LLL", "", "", "", "", "",
        "", "Frothy secretions", False, "", "",
        "", "",
        "", "", "Improved clearance", "", ""
    ])
    # Event 2: Stent Eval
    ws_event.append([
        SOURCE_FILE, NOTE_ID, "evt_02", "Device Evaluation", "Inspection",
        "RMS to BI", "Metallic Stent", "", "", "", "",
        "", "Patent, Jailing RUL, Undersized", False, "", "",
        "", "Metal",
        "", "", "", "", ""
    ])
    # Event 3: Stenosis Eval
    ws_event.append([
        SOURCE_FILE, NOTE_ID, "evt_03", "Inspection", "Inspection",
        "RML, LMS", "", "", "", "", "",
        "", "Moderate stenosis RML, Mild stenosis LMS", False, "", "",
        "", "",
        "", "", "", "", ""
    ])
    # Outcome
    ws_event.append([
        SOURCE_FILE, NOTE_ID, "evt_outcome", "Outcome", "",
        "", "", "", "", "", "",
        "", "No complications", False, "", "",
        "", "",
        "", "", "Tolerated well", "", "None"
    ])

    # 7. V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    
    # Event 1
    ws_v3.append([
        NOTE_ID, "evt_01", "Therapeutic Aspiration",
        "Airway", "Multi-lobar", "", "",
        "Secretions", "",
        "Suction", "", "", "", '{"finding": "frothy secretions"}', "Successful therapeutic aspiration",
        "", "", "",
        "", "", "Improved", "", ""
    ])
    
    # Event 2
    ws_v3.append([
        NOTE_ID, "evt_02", "Device Evaluation",
        "Airway", "RMS", "", "",
        "", "",
        "Inspection", '{"device_type": "stent", "status": "existing"}', "", "", '{"finding": "jailing RUL", "position": "good"}', "Uncovered metallic stent in good position",
        "", "Metallic", "",
        "", "", "", "", ""
    ])

    # 8. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": [
            {
                "event_id": "evt_01",
                "type": "Therapeutic Aspiration",
                "method": "Suction",
                "target": {"anatomy": "Airway", "location": "Multi-lobar"},
                "findings": ["Frothy secretions", "Mucus plug"],
                "evidence": "Successful therapeutic aspiration was performed to clean out the distal trachea"
            },
            {
                "event_id": "evt_02",
                "type": "Device Evaluation",
                "target": {"anatomy": "Airway", "location": "RMS"},
                "device": {"type": "Stent", "material": "Metallic", "status": "Existing"},
                "findings": ["Good position", "Undersized", "Jailing RUL"],
                "evidence": "Uncovered metallic stent in good position"
            }
        ]
    }
    ws_json.append([json.dumps(registry_data, indent=4)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Successfully generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()