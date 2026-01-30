import openpyxl
from openpyxl.utils import get_column_letter
import json
import os
import re

# -------------------------------------------------------------------------
# CONSTANTS & INPUT DATA
# -------------------------------------------------------------------------

NOTE_ID = "note_106"
SOURCE_FILE = "note_106.txt"
PROCEDURE_DATE = "2026-01-12" # Placeholder/Inferred
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_106 SOURCE_FILE: note_106.txt INDICATION FOR OPERATION:  [REDACTED]is a 56 year old-year-old male who presents with airway evaluation and stent change.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
PROCEDURE:  
31646 Therapeutic aspiration subsequent episodes
31624 Dx bronchoscope/lavage (BAL)    
31636 Dilate and bronchial stent initial bronchus
31635 Foreign body removal
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Rigid Bronchoscope
Flexible Therapeutic Bronchoscope
Flexible Hybrid (Pedatric) Bronchoscope
ESTIMATED BLOOD LOSS:   None
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
Initial Airway Inspection Findings:
Successful therapeutic aspiration was performed to clean out the Right Mainstem, Bronchus Intermedius , and Left Mainstem from mucus.
Bronchial alveolar lavage was performed at Anteromedial Segment of LLL (Lb7/8).
Instilled 40 cc of NS, suction returned with 10 cc of NS.  Samples sent for Microbiology (Cultures/Viral/Fungal).
Boston Scientific 14x40 Bare metallic stent was removed with out issue.
The following stent (Boston Scientific 14x40) was placed in the Right Mainstem via direct visualization.
Dilated to 15mm x2 with elation balloon.  The stent is in good position but slightly undersized.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
 
IMPRESSION/PLAN: [REDACTED]is a 56 year old-year-old male who presents for bronchoscopy for stent exchange.
- distal airway appears to be improving. 
- minimize bronchscopy while maximize sputum clearance and decrease interruptions to nutrition and PT
- check stent and airways once a week
- plan for a stent exchange in 2 weeks (consider 16x40 boston sems)"""

# -------------------------------------------------------------------------
# PROCEDURE FLAGS
# -------------------------------------------------------------------------

PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1, # Implied by BAL and inspection
    "bal": 1,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 1, # Header explicitly codes 31635 FBR (stent removal)
    "airway_dilation": 1,
    "airway_stent": 1,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 1,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0,
}

# -------------------------------------------------------------------------
# SPAN ANNOTATIONS (ANCHOR-FIRST)
# -------------------------------------------------------------------------

# Event IDs:
# ev1: Therapeutic Aspiration
# ev2: BAL
# ev3: Stent Removal
# ev4: Stent Placement
# ev5: Dilation
# ev6: Global/Context

SPANS = [
    # Global / Setup
    {
        "text": "Rigid Bronchoscope",
        "label": "DEV_INSTRUMENT",
        "norm": "Rigid Bronchoscope",
        "event_id": "ev6",
        "context_prefix": "INSTRUMENT :"
    },
    
    # ev1: Therapeutic Aspiration
    {
        "text": "therapeutic aspiration",
        "label": "PROC_METHOD",
        "norm": "therapeutic aspiration",
        "event_id": "ev1",
        "context_prefix": "Successful"
    },
    {
        "text": "Right Mainstem",
        "label": "ANAT_AIRWAY",
        "norm": "Right Main Stem",
        "event_id": "ev1",
        "context_prefix": "clean out the"
    },
    {
        "text": "Bronchus Intermedius",
        "label": "ANAT_AIRWAY",
        "norm": "Bronchus Intermedius",
        "event_id": "ev1"
    },
    {
        "text": "Left Mainstem",
        "label": "ANAT_AIRWAY",
        "norm": "Left Main Stem",
        "event_id": "ev1"
    },

    # ev2: BAL
    {
        "text": "Bronchial alveolar lavage",
        "label": "PROC_METHOD",
        "norm": "BAL",
        "event_id": "ev2"
    },
    {
        "text": "Anteromedial Segment of LLL",
        "label": "ANAT_LUNG_LOC",
        "norm": "LLL Anteromedial Basal",
        "event_id": "ev2"
    },
    {
        "text": "Lb7/8",
        "label": "ANAT_AIRWAY",
        "norm": "LB7+LB8",
        "event_id": "ev2"
    },
    {
        "text": "Instilled 40 cc",
        "label": "MEAS_VOL",
        "norm": "40ml",
        "event_id": "ev2"
    },
    {
        "text": "returned with 10 cc",
        "label": "MEAS_VOL",
        "norm": "10ml",
        "event_id": "ev2"
    },

    # ev3: Stent Removal (FBR)
    {
        "text": "Boston Scientific 14x40 Bare metallic stent",
        "label": "DEV_STENT",
        "norm": "Boston Scientific Bare Metal Stent 14x40",
        "event_id": "ev3"
    },
    {
        "text": "removed",
        "label": "PROC_ACTION",
        "norm": "removal",
        "event_id": "ev3",
        "context_prefix": "Bare metallic stent was"
    },

    # ev4: Stent Placement
    {
        "text": "stent (Boston Scientific 14x40)",
        "label": "DEV_STENT",
        "norm": "Boston Scientific Stent 14x40",
        "event_id": "ev4"
    },
    {
        "text": "Right Mainstem",
        "label": "ANAT_AIRWAY",
        "norm": "Right Main Stem",
        "event_id": "ev4",
        "context_prefix": "placed in the"
    },
    {
        "text": "good position",
        "label": "OUTCOME_AIRWAY_LUMEN_POST",
        "norm": "patent",
        "event_id": "ev4"
    },
    {
        "text": "slightly undersized",
        "label": "OUTCOME_AIRWAY_LUMEN_POST",
        "norm": "undersized",
        "event_id": "ev4"
    },

    # ev5: Dilation
    {
        "text": "Dilated",
        "label": "PROC_METHOD",
        "norm": "Airway Dilation",
        "event_id": "ev5"
    },
    {
        "text": "15mm",
        "label": "MEAS_AIRWAY_DIAM",
        "norm": "15mm",
        "event_id": "ev5"
    },
    {
        "text": "elation balloon",
        "label": "DEV_INSTRUMENT",
        "norm": "Elation Balloon",
        "event_id": "ev5"
    },

    # Global Outcome
    {
        "text": "No immediate complications",
        "label": "OUTCOME_COMPLICATION",
        "norm": "none",
        "event_id": "ev6"
    }
]

# -------------------------------------------------------------------------
# EVENT LOG DATA
# -------------------------------------------------------------------------

EVENT_LOG = [
    {
        "event_id": "ev1",
        "event_type": "Therapeutic Bronchoscopy",
        "method": "Therapeutic Aspiration",
        "anatomy_target": "Right Mainstem, Bronchus Intermedius, Left Mainstem",
        "findings": "Mucus cleaned out"
    },
    {
        "event_id": "ev2",
        "event_type": "Diagnostic Bronchoscopy",
        "method": "BAL",
        "anatomy_target": "LLL Anteromedial Segment",
        "counts": "Instilled 40cc, Returned 10cc",
        "specimens": "Microbiology"
    },
    {
        "event_id": "ev3",
        "event_type": "Therapeutic Bronchoscopy",
        "method": "Foreign Body Removal",
        "anatomy_target": "Airway",
        "device": "Boston Scientific 14x40 Bare metallic stent",
        "findings": "Removed without issue"
    },
    {
        "event_id": "ev4",
        "event_type": "Therapeutic Bronchoscopy",
        "method": "Airway Stent",
        "anatomy_target": "Right Mainstem",
        "device": "Boston Scientific 14x40",
        "device_size": "14x40mm",
        "outcome_airway_lumen_post": "Good position but slightly undersized"
    },
    {
        "event_id": "ev5",
        "event_type": "Therapeutic Bronchoscopy",
        "method": "Airway Dilation",
        "anatomy_target": "Right Mainstem (implied)",
        "device": "Elation Balloon",
        "measurements": "15mm x2"
    }
]

# -------------------------------------------------------------------------
# V3 REGISTRY EVENTS (NESTED)
# -------------------------------------------------------------------------

V3_EVENTS = [
    {
        "event_id": "ev1",
        "type": "Therapeutic Aspiration",
        "target": {"anatomy_type": "airway", "location": {"lobe": "multiple"}},
        "method": "suction",
        "findings_json": {"content": "mucus"}
    },
    {
        "event_id": "ev2",
        "type": "BAL",
        "target": {"anatomy_type": "lung", "location": {"lobe": "LLL", "segment": "anteromedial"}},
        "measurements_json": {"instilled": "40ml", "returned": "10ml"}
    },
    {
        "event_id": "ev3",
        "type": "Foreign Body Removal",
        "method": "forceps/rigid",
        "devices_json": {"description": "Bare metallic stent 14x40"},
        "findings_json": {"outcome": "removed without issue"}
    },
    {
        "event_id": "ev4",
        "type": "Airway Stent",
        "target": {"anatomy_type": "airway", "location": {"lobe": "RMS"}},
        "stent": {"size": "14x40", "material_or_brand": "Boston Scientific"},
        "outcomes": {"airway": {"lumen_post": "patent, undersized"}}
    },
    {
        "event_id": "ev5",
        "type": "Airway Dilation",
        "method": "balloon",
        "measurements_json": {"diameter": "15mm"},
        "devices_json": {"brand": "Elation"}
    }
]

# -------------------------------------------------------------------------
# HELPER FUNCTIONS
# -------------------------------------------------------------------------

def hydrate_spans(spans, note_text):
    hydrated = []
    for s in spans:
        span_text = s["text"]
        context = s.get("context_prefix", "")
        match_index = s.get("match_index", 0)
        
        start = -1
        status = "ambiguous"
        
        # 1. Exact unique match
        if note_text.count(span_text) == 1:
            start = note_text.find(span_text)
            status = "hydrated_unique"
        
        # 2. Context prefix
        elif context:
            candidates = [m.start() for m in re.finditer(re.escape(span_text), note_text)]
            for cand in candidates:
                window_start = max(0, cand - 120)
                preceding = note_text[window_start:cand]
                if context in preceding:
                    start = cand
                    status = "hydrated_prefix_window"
                    break
        
        # 3. Match index fallback
        elif match_index > 0:
            candidates = [m.start() for m in re.finditer(re.escape(span_text), note_text)]
            if len(candidates) >= match_index:
                start = candidates[match_index - 1]
                status = "hydrated_match_index"
        
        # Fallback if context failed but there's a unique match anyway (safety)
        if start == -1 and note_text.count(span_text) == 1:
             start = note_text.find(span_text)
             status = "hydrated_unique_fallback"

        if start != -1:
            end = start + len(span_text)
            row = s.copy()
            row["start_char"] = start
            row["end_char"] = end
            row["hydration_status"] = status
            hydrated.append(row)
        else:
            # Add with blank offsets if failed
            row = s.copy()
            row["start_char"] = ""
            row["end_char"] = ""
            row["hydration_status"] = f"ambiguous_count={note_text.count(span_text)}"
            hydrated.append(row)
            
    return hydrated

def generate_excel():
    # Load Template
    if not os.path.exists(TEMPLATE_PATH):
        # Create a dummy workbook if template is missing (for robust standalone execution)
        wb = openpyxl.Workbook()
        wb.create_sheet("Note_Text")
        wb.create_sheet("Note_Index")
        wb.create_sheet("Span_Annotations")
        wb.create_sheet("Span_Hydrated")
        wb.create_sheet("Event_Log")
        wb.create_sheet("V3_Procedure_Events")
        wb.create_sheet("V3_Registry_JSON")
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # 1. Note_Text
    ws = wb["Note_Text"]
    if ws.max_row == 1:
        ws.append(["note_id", "source_file", "note_text"])
    ws.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws = wb["Note_Index"]
    header = [
        "source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes",
        "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", "tbna_conventional",
        "linear_ebus", "radial_ebus", "navigational_bronchoscopy", "transbronchial_biopsy", "transbronchial_cryobiopsy",
        "therapeutic_aspiration", "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
        "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", "bronchial_thermoplasty",
        "whole_lung_lavage", "rigid_bronchoscopy",
        "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis", "pleural_biopsy", "fibrinolytic_therapy"
    ]
    
    # Ensure header exists
    if ws.max_row == 0:
        ws.append(header)
    
    row_data = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Pending", ""]
    # Append flags in order
    flag_keys = header[8:]
    for k in flag_keys:
        row_data.append(PROCEDURE_FLAGS.get(k, 0))
    ws.append(row_data)

    # 3. Span_Annotations
    ws = wb["Span_Annotations"]
    span_header = [
        "source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", "match_index",
        "start_char", "end_char", "span_len", "label", "normalized_value", "schema_field", "event_id",
        "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"
    ]
    if ws.max_row == 0:
        ws.append(span_header)

    hydrated_data = hydrate_spans(SPANS, NOTE_TEXT)

    for idx, s in enumerate(SPANS):
        # Anchor-first: start/end are blank
        ws.append([
            SOURCE_FILE, NOTE_ID, f"span_{idx+1}", "", s.get("context_prefix", ""), s["text"], s.get("match_index", ""),
            "", "", f"=LEN(F{ws.max_row+1})", s["label"], s.get("norm", ""), "", s.get("event_id", ""),
            "", "", "", "", "", "needs_hydration"
        ])

    # 4. Span_Hydrated
    ws = wb["Span_Hydrated"]
    if ws.max_row == 0:
        ws.append(span_header)
    
    for idx, h in enumerate(hydrated_data):
        ws.append([
            SOURCE_FILE, NOTE_ID, f"span_{idx+1}", "", h.get("context_prefix", ""), h["text"], h.get("match_index", ""),
            h["start_char"], h["end_char"], len(h["text"]), h["label"], h.get("norm", ""), "", h.get("event_id", ""),
            "", "", "", "", "", h["hydration_status"]
        ])

    # 5. Event_Log
    ws = wb["Event_Log"]
    event_header = [
        "source_file", "note_id", "event_id", "event_type", "method", "anatomy_target", "device", "needle_gauge",
        "stations", "counts", "measurements", "specimens", "findings", "is_historical", "reviewer", "comments",
        "device_size", "device_material", "outcome_airway_lumen_pre", "outcome_airway_lumen_post",
        "outcome_symptoms", "outcome_pleural", "outcome_complication"
    ]
    if ws.max_row == 0:
        ws.append(event_header)
    
    for e in EVENT_LOG:
        ws.append([
            SOURCE_FILE, NOTE_ID, e["event_id"], e["event_type"], e["method"], e.get("anatomy_target", ""),
            e.get("device", ""), "", "", e.get("counts", ""), e.get("measurements", ""), e.get("specimens", ""),
            e.get("findings", ""), "", "", "", e.get("device_size", ""), e.get("device_material", ""),
            "", e.get("outcome_airway_lumen_post", ""), "", "", e.get("outcome_complication", "")
        ])

    # 6. V3_Procedure_Events
    ws = wb["V3_Procedure_Events"]
    v3_header = [
        "note_id", "event_id", "type", "target.anatomy_type", "target.location.lobe", "target.location.segment",
        "target.station", "lesion.type", "lesion.size_mm", "method", "devices_json", "measurements_json",
        "specimens_json", "findings_json", "evidence_quote", "stent.size", "stent.material_or_brand",
        "catheter.size_fr", "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post", "outcomes.symptoms",
        "outcomes.pleural", "outcomes.complications"
    ]
    if ws.max_row == 0:
        ws.append(v3_header)
    
    for e in V3_EVENTS:
        target = e.get("target", {})
        loc = target.get("location", {})
        ws.append([
            NOTE_ID, e["event_id"], e["type"], target.get("anatomy_type", ""), loc.get("lobe", ""), loc.get("segment", ""),
            target.get("station", ""), "", "", e.get("method", ""), json.dumps(e.get("devices_json", {})),
            json.dumps(e.get("measurements_json", {})), "", json.dumps(e.get("findings_json", {})), "",
            e.get("stent", {}).get("size", ""), e.get("stent", {}).get("material_or_brand", ""),
            "", "", e.get("outcomes", {}).get("airway", {}).get("lumen_post", ""), "", "", ""
        ])

    # 7. V3_Registry_JSON
    ws = wb["V3_Registry_JSON"]
    if ws.max_row == 0:
        ws.append(["schema_version", "note_id", "json_output"])
    
    registry_obj = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "procedures": V3_EVENTS,
        "no_immediate_complications": True
    }
    ws.append(["v3.0", NOTE_ID, json.dumps(registry_obj, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_excel()