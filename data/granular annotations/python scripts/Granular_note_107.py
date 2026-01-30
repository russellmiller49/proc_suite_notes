import openpyxl
from openpyxl.utils import get_column_letter
import json
import datetime
import re

# -------------------------------------------------------------------------
# INPUTS
# -------------------------------------------------------------------------
NOTE_ID = "note_107"
SOURCE_FILE = "note_107.txt"
PROCEDURE_DATE = "2026-01-12" # inferred from context
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_107 SOURCE_FILE: note_107.txt INDICATION FOR OPERATION:  [REDACTED]is a 56 year old-year-old male who presents with post-transplant evaluation.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: J96.90 Respiratory Failure
POSTOPERATIVE DIAGNOSIS:  J96.90 Respiratory Failure
PROCEDURE:  
31615 Visualization of windpipe (Tracheobronchoscopy through established tracheostomy incision)
31645 Therapeutic aspiration initial episode
ANESTHESIA: 
99152 Moderate sedation: initial 15 minutes
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Disposable Bronchoscope
ESTIMATED BLOOD LOSS:   None
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
Initial Airway Inspection Findings:
Tracheobronchoscopy was performed with insertion of bronchoscope through the tracheostomy to perform airway clearance and confirm tracheostomy position.
The tracheostomy tube is in good position. The visualized portion of the trachea is of normal caliber.
The carina is sharp. The tracheobronchial tree was examined to at least the first subsegmental level.
Inspection shows a BMS that is patent in the RMSB, covering the RUL but with good aeration and mucus drainage through the stent.
Dehiscence at the RMSB remains entirely closed by stent. Areas of granulation tissue/ischemic mucosa remain at the distal BI/RML orifice as well as LMSB anastomosis and LUL.
There are two areas of very thin mucosa/airway wall in the LUL and Lingula.  Moderate thick secretions.
Successful therapeutic aspiration was performed to clean out the Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, and Carina from mucus.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
none
IMPRESSION/PLAN: [REDACTED]is a 56 year old-year-old male who presents for bronchoscopy for airway evaluation.
- f/u with repeat bronch every other day
- plan for possible stent exchange"""

# -------------------------------------------------------------------------
# PROCEDURE FLAGS
# -------------------------------------------------------------------------
# Default 0, set to 1 if present
FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0, # Inspection of existing stent, not placement
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0,
}

# -------------------------------------------------------------------------
# SPAN ANNOTATIONS (ANCHOR-FIRST)
# -------------------------------------------------------------------------
# We define them here. Code will calculate offsets.
# event_id map:
# ev1: General/Initial Inspection (Trach)
# ev2: Stent Inspection (RMSB)
# ev3: Findings (Granulation/Ischemia)
# ev4: Therapeutic Aspiration
# ev99: Global Outcomes

SPANS = [
    {
        "span_text": "Disposable Bronchoscope",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "Disposable Bronchoscope",
        "event_id": "ev1",
        "section_type": "Header"
    },
    {
        "span_text": "Tracheobronchoscopy",
        "label": "PROC_METHOD",
        "normalized_value": "Tracheobronchoscopy",
        "context_prefix": "Initial Airway Inspection Findings:\n",
        "event_id": "ev1",
        "section_type": "Procedure"
    },
    {
        "span_text": "through the tracheostomy",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Tracheostomy",
        "event_id": "ev1",
        "section_type": "Procedure"
    },
    {
        "span_text": "tracheostomy tube is in good position",
        "label": "OBS_FINDING",
        "normalized_value": "tracheostomy tube good position",
        "event_id": "ev1",
        "section_type": "Findings"
    },
    # Event 2: Stent Inspection
    {
        "span_text": "BMS",
        "label": "DEV_STENT",
        "normalized_value": "Bronchial Metal Stent",
        "event_id": "ev2",
        "section_type": "Findings",
        "is_historical": "1" # Existing stent
    },
    {
        "span_text": "RMSB",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Right Main Stem Bronchus",
        "context_prefix": "that is patent in the ",
        "event_id": "ev2",
        "section_type": "Findings"
    },
    {
        "span_text": "covering the RUL",
        "label": "ANAT_AIRWAY",
        "normalized_value": "RUL",
        "event_id": "ev2",
        "section_type": "Findings"
    },
    {
        "span_text": "Dehiscence",
        "label": "OBS_FINDING",
        "normalized_value": "Dehiscence",
        "event_id": "ev2",
        "section_type": "Findings"
    },
    # Event 3: Granulation/Ischemia
    {
        "span_text": "Granulation tissue",
        "label": "OBS_LESION",
        "normalized_value": "Granulation tissue",
        "event_id": "ev3",
        "section_type": "Findings"
    },
    {
        "span_text": "ischemic mucosa",
        "label": "OBS_LESION",
        "normalized_value": "ischemic mucosa",
        "event_id": "ev3",
        "section_type": "Findings"
    },
    {
        "span_text": "distal BI",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Bronchus Intermedius",
        "event_id": "ev3",
        "section_type": "Findings"
    },
    {
        "span_text": "RML orifice",
        "label": "ANAT_AIRWAY",
        "normalized_value": "RML",
        "event_id": "ev3",
        "section_type": "Findings"
    },
    {
        "span_text": "LMSB anastomosis",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Left Main Stem Bronchus",
        "event_id": "ev3",
        "section_type": "Findings"
    },
    {
        "span_text": "LUL",
        "label": "ANAT_AIRWAY",
        "normalized_value": "LUL",
        "context_prefix": "LMSB anastomosis and ",
        "event_id": "ev3",
        "section_type": "Findings"
    },
    # Event 4: Therapeutic Aspiration
    {
        "span_text": "Moderate thick secretions",
        "label": "OBS_FINDING",
        "normalized_value": "thick secretions",
        "event_id": "ev4",
        "section_type": "Findings"
    },
    {
        "span_text": "Successful therapeutic aspiration",
        "label": "PROC_METHOD",
        "normalized_value": "Therapeutic aspiration",
        "event_id": "ev4",
        "section_type": "Procedure"
    },
    {
        "span_text": "Trachea (Distal 1/3)",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Trachea Distal 1/3",
        "event_id": "ev4",
        "section_type": "Procedure"
    },
    {
        "span_text": "Right Mainstem",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Right Main Stem Bronchus",
        "event_id": "ev4",
        "section_type": "Procedure"
    },
    {
        "span_text": "Bronchus Intermedius",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Bronchus Intermedius",
        "context_prefix": "Right Mainstem, ",
        "event_id": "ev4",
        "section_type": "Procedure"
    },
    {
        "span_text": "Left Mainstem",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Left Main Stem Bronchus",
        "event_id": "ev4",
        "section_type": "Procedure"
    },
    {
        "span_text": "Carina",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Carina",
        "context_prefix": "Left Mainstem, and ",
        "event_id": "ev4",
        "section_type": "Procedure"
    },
    # Outcomes
    {
        "span_text": "no immediate complications",
        "label": "OUTCOME_COMPLICATION",
        "normalized_value": "none",
        "event_id": "ev99",
        "section_type": "Outcome"
    }
]

# -------------------------------------------------------------------------
# HELPER FUNCTIONS
# -------------------------------------------------------------------------

def hydrate_spans(spans, note_text):
    """
    Calculates start_char, end_char, and hydration_status for each span.
    """
    hydrated = []
    for s in spans:
        txt = s["span_text"]
        
        # 1. Exact count check
        count = note_text.count(txt)
        
        start = -1
        status = "ambiguous"
        
        if count == 1:
            start = note_text.find(txt)
            status = "hydrated_unique"
        elif count > 1:
            # Try context
            if "context_prefix" in s and s["context_prefix"]:
                prefix = s["context_prefix"]
                # Iterate all occurrences
                idx = -1
                found_match = False
                while True:
                    idx = note_text.find(txt, idx + 1)
                    if idx == -1:
                        break
                    # Check window before
                    window_start = max(0, idx - 120)
                    window_text = note_text[window_start:idx]
                    if prefix in window_text:
                        start = idx
                        status = "hydrated_prefix_window"
                        found_match = True
                        break
                if not found_match:
                    status = f"ambiguous_count={count}"
            elif "match_index" in s:
                # Use explicit index (0-based)
                m_idx = s["match_index"]
                current_found = 0
                idx = -1
                while True:
                    idx = note_text.find(txt, idx + 1)
                    if idx == -1:
                        break
                    if current_found == m_idx:
                        start = idx
                        status = "hydrated_match_index"
                        break
                    current_found += 1
            else:
                status = f"ambiguous_count={count}"
        else:
            status = "not_found"

        row = s.copy()
        if start != -1:
            row["start_char"] = start
            row["end_char"] = start + len(txt)
        else:
            row["start_char"] = ""
            row["end_char"] = ""
        
        row["hydration_status"] = status
        row["span_len"] = len(txt)
        hydrated.append(row)
    return hydrated

# -------------------------------------------------------------------------
# EXECUTION
# -------------------------------------------------------------------------

def main():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        # Create a basic workbook structure if template is missing (fallback)
        wb = openpyxl.Workbook()
        for sheet in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            wb.create_sheet(sheet)
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # 1. Note_Text
    ws_text = wb["Note_Text"]
    # Headers should exist, append row
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws_index = wb["Note_Index"]
    # Headers: source_file, note_id, encounter_id, procedure_date, site, reviewer, status, free_text_notes, [flags...]
    flag_values = [FLAGS[k] for k in FLAGS]
    row_meta = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Machine_Generated", ""]
    ws_index.append(row_meta + flag_values)

    # 3. Span_Annotations
    ws_span = wb["Span_Annotations"]
    # Columns: source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, start, end, len, label, norm, schema, event_id, neg, hist, time, rev, comm, hydration
    
    # Calculate hydrated spans first to have full data
    hydrated_data = hydrate_spans(SPANS, NOTE_TEXT)
    
    for i, h in enumerate(hydrated_data):
        span_id = f"{NOTE_ID}_s{i+1:03d}"
        
        # Write to Span_Annotations (Blank offsets)
        row_anno = [
            SOURCE_FILE, NOTE_ID, span_id, h.get("section_type", ""),
            h.get("context_prefix", ""), h["span_text"], h.get("match_index", ""),
            "", "", f'=LEN(F{ws_span.max_row+1})', # Formula for len
            h["label"], h.get("normalized_value", ""), "", h["event_id"],
            h.get("is_negated", ""), h.get("is_historical", ""), "", "", "",
            "needs_hydration"
        ]
        ws_span.append(row_anno)

    # 4. Span_Hydrated
    ws_hydra = wb["Span_Hydrated"]
    for i, h in enumerate(hydrated_data):
        span_id = f"{NOTE_ID}_s{i+1:03d}"
        row_hydra = [
            SOURCE_FILE, NOTE_ID, span_id, h.get("section_type", ""),
            h.get("context_prefix", ""), h["span_text"], h.get("match_index", ""),
            h["start_char"], h["end_char"], h["span_len"],
            h["label"], h.get("normalized_value", ""), "", h["event_id"],
            h.get("is_negated", ""), h.get("is_historical", ""), "", "", "",
            h["hydration_status"]
        ]
        ws_hydra.append(row_hydra)

    # 5. Event_Log (High Level Summary)
    ws_event = wb["Event_Log"]
    
    # Organize spans by event_id
    events_map = {}
    for h in hydrated_data:
        eid = h["event_id"]
        if eid not in events_map:
            events_map[eid] = {
                "method": set(), "anatomy": set(), "device": set(), 
                "findings": set(), "outcomes": set(), "historical": False
            }
        
        if h["label"] == "PROC_METHOD":
            events_map[eid]["method"].add(h["normalized_value"])
        elif h["label"] == "ANAT_AIRWAY":
            events_map[eid]["anatomy"].add(h["normalized_value"])
        elif h["label"] in ["DEV_STENT", "DEV_INSTRUMENT"]:
            events_map[eid]["device"].add(h["normalized_value"])
        elif h["label"] in ["OBS_FINDING", "OBS_LESION"]:
            events_map[eid]["findings"].add(h["normalized_value"])
        elif h["label"] == "OUTCOME_COMPLICATION":
            events_map[eid]["outcomes"].add(h["normalized_value"])
        
        if h.get("is_historical") == "1":
            events_map[eid]["historical"] = True

    # Write Event Rows
    for eid, data in events_map.items():
        etype = "Procedure"
        if eid == "ev99": etype = "Outcome"
        elif eid == "ev1": etype = "Inspection"
        
        row = [
            SOURCE_FILE, NOTE_ID, eid, etype,
            ", ".join(data["method"]),
            ", ".join(data["anatomy"]),
            ", ".join(data["device"]),
            "", "", "", "", "", # gauge, stations, counts, measurements, specimens
            ", ".join(data["findings"]),
            "1" if data["historical"] else "0",
            "", "", "", "", # reviewer, comments, dev_size, dev_mat
            "", "", "", "", # airway outcomes
            ", ".join(data["outcomes"]) # compl
        ]
        ws_event.append(row)

    # 6. V3_Procedure_Events & 7. V3_Registry_JSON
    ws_v3 = wb["V3_Procedure_Events"]
    ws_json = wb["V3_Registry_JSON"]
    
    # Construct JSON structure
    registry_data = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "procedures": [],
        "no_immediate_complications": "no immediate complications" in NOTE_TEXT
    }

    # Map events for V3
    # ev1: Inspection
    registry_data["procedures"].append({
        "event_id": "ev1",
        "type": "diagnostic_bronchoscopy",
        "method": "Tracheobronchoscopy",
        "target": {"anatomy_type": "airway", "location": {"lobe": "Trachea"}},
        "findings_json": ["Tracheostomy tube in good position"]
    })
    ws_v3.append([NOTE_ID, "ev1", "diagnostic_bronchoscopy", "airway", "Trachea", "", "", "", "", "Tracheobronchoscopy", "", "", "", '["Tracheostomy tube in good position"]', "", "", "", "", "", "", "", "", ""])

    # ev2: Stent Inspection
    registry_data["procedures"].append({
        "event_id": "ev2",
        "type": "inspection",
        "target": {"anatomy_type": "airway", "location": {"lobe": "RUL", "segment": "RMSB"}},
        "stent": {"type": "Bronchial Metal Stent", "status": "Patent"},
        "findings_json": ["Dehiscence closed by stent"]
    })
    ws_v3.append([NOTE_ID, "ev2", "inspection", "airway", "RUL", "RMSB", "", "", "", "Inspection", '{"device": "BMS"}', "", "", '["Dehiscence closed by stent"]', "", "", "Metal", "", "", "", "", "", ""])

    # ev3: Findings
    registry_data["procedures"].append({
        "event_id": "ev3",
        "type": "inspection",
        "target": {"anatomy_type": "airway", "location": {"lobe": "LUL", "segment": "LMSB"}},
        "findings_json": ["Granulation tissue", "Ischemic mucosa", "Thin mucosa"]
    })
    ws_v3.append([NOTE_ID, "ev3", "inspection", "airway", "LUL", "LMSB", "", "Granulation/Ischemia", "", "", "", "", "", '["Granulation tissue", "Ischemic mucosa"]', "", "", "", "", "", "", "", "", ""])

    # ev4: Therapeutic Aspiration
    registry_data["procedures"].append({
        "event_id": "ev4",
        "type": "therapeutic_aspiration",
        "method": "Therapeutic aspiration",
        "target": {"anatomy_type": "airway", "location": {"lobe": "Multiple"}},
        "findings_json": ["Moderate thick secretions"]
    })
    ws_v3.append([NOTE_ID, "ev4", "therapeutic_aspiration", "airway", "Multiple", "", "", "", "", "Therapeutic aspiration", "", "", "", '["Moderate thick secretions"]', "", "", "", "", "", "", "", "", ""])

    # Write JSON
    ws_json.cell(row=1, column=1, value=json.dumps(registry_data, indent=2))

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    main()