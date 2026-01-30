import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_108"
SOURCE_FILE = "note_108.txt"
PROCEDURE_DATE = "2026-01-12" # Simulated date based on context or current date
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_108 SOURCE_FILE: note_108.txt INDICATION FOR OPERATION:  [REDACTED]is a 52 year old-year-old male who presents with post-transplant evaluation.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: J96.90 Respiratory Failure
POSTOPERATIVE DIAGNOSIS:  J96.90 Respiratory Failure
PROCEDURE:  
31615 Visualization of windpipe (Tracheobronchoscopy through established tracheostomy incision)
31645 Therapeutic aspiration initial episode
ANESTHESIA: 
99152 Moderate sedation: initial 15 minutes
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Disposable Bronchoscope
ESTIMATED BLOOD LOSS:   None
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
Initial Airway Inspection Findings:
Tracheobronchoscopy was performed with insertion of bronchoscope through the tracheostomy to perform airway clearance and confirm tracheostomy position.
The tracheostomy tube is in good position. The visualized portion of the trachea is of normal caliber.
The carina is sharp. The tracheobronchial tree was examined to at least the first subsegmental level.
Inspection shows a BMS that is patent in the RMSB, partially covering the RUL but with good aeration and mucus drainage through the stent.
Dehiscence at the RMSB remains entirely closed by stent. Areas of granulation tissue/ischemic mucosa remain at the distal BI/RML orifice as well as LMSB anastomosis and LUL.
There are two areas of very thin mucosa/airway wall in the LUL and Lingula.  Moderate thick secretions.
Successful therapeutic aspiration was performed to clean out the Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, and Carina from mucus.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
none
IMPRESSION/PLAN: [REDACTED]is a 52 year old-year-old male who presents for bronchoscopy for airway evaluation.
- f/u with repeat bronch every other day"""

# -------------------------------------------------------------------------
# CONFIGURATION & DEFINITIONS
# -------------------------------------------------------------------------

# Procedure Flags (30 items)
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1, # "Tracheobronchoscopy was performed"
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1, # "Successful therapeutic aspiration was performed"
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0, # Existing stent inspected, not placed
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# Annotated Spans (Anchor-First)
# context_prefix is used when text appears multiple times.
# event_id maps to the sequential logic of the procedure.
# Events:
# E1: General Bronchoscopy/Inspection
# E2: Stent Evaluation (RMSB)
# E3: Granulation/Ischemia Evaluation (BI, LMSB, LUL)
# E4: Therapeutic Aspiration
# E5: Outcome/Complications

RAW_SPANS = [
    # Metadata/Method
    {
        "span_text": "post-transplant evaluation",
        "label": "CTX_HISTORICAL",
        "normalized_value": "post-transplant",
        "event_id": "E1",
        "context_prefix": "presents with "
    },
    {
        "span_text": "Tracheobronchoscopy",
        "label": "PROC_METHOD",
        "normalized_value": "Tracheobronchoscopy",
        "event_id": "E1",
        "context_prefix": "windpipe ("
    },
    {
        "span_text": "tracheostomy tube",
        "label": "DEV_CATHETER", # Or instrument, but tube fits best here
        "normalized_value": "tracheostomy tube",
        "event_id": "E1",
        "context_prefix": "The "
    },
    {
        "span_text": "good position",
        "label": "OBS_LESION", # Finding regarding device
        "normalized_value": "good position",
        "event_id": "E1",
        "context_prefix": "tube is in "
    },
    
    # Event 2: Stent / RMSB
    {
        "span_text": "BMS",
        "label": "DEV_STENT",
        "normalized_value": "Metallic Stent",
        "event_id": "E2",
        "context_prefix": "Inspection shows a "
    },
    {
        "span_text": "RMSB",
        "label": "ANAT_AIRWAY",
        "normalized_value": "RMSB",
        "event_id": "E2",
        "context_prefix": "patent in the "
    },
    {
        "span_text": "patent",
        "label": "OBS_LESION",
        "normalized_value": "patent",
        "event_id": "E2",
        "context_prefix": "BMS that is "
    },
    {
        "span_text": "RUL",
        "label": "ANAT_LUNG_LOC",
        "normalized_value": "RUL",
        "event_id": "E2",
        "context_prefix": "covering the "
    },
    {
        "span_text": "Dehiscence",
        "label": "OBS_LESION",
        "normalized_value": "Dehiscence",
        "event_id": "E2",
        "context_prefix": "drainage through the stent.\n"
    },
    {
        "span_text": "entirely closed",
        "label": "OBS_LESION",
        "normalized_value": "closed",
        "event_id": "E2",
        "context_prefix": "remains "
    },

    # Event 3: Granulation/Ischemia (Other airways)
    {
        "span_text": "granulation tissue",
        "label": "OBS_LESION",
        "normalized_value": "granulation tissue",
        "event_id": "E3",
        "context_prefix": "Areas of "
    },
    {
        "span_text": "ischemic mucosa",
        "label": "OBS_LESION",
        "normalized_value": "ischemia",
        "event_id": "E3",
        "context_prefix": "tissue/"
    },
    {
        "span_text": "Bronchus Intermedius",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Bronchus Intermedius",
        "event_id": "E3",
        "context_prefix": "distal " # Note text says "distal BI", need to map carefully. Wait, text says "distal BI/RML". Let's look at full text.
        # Text: "distal BI/RML" -> BI is abbreviation.
        # Let's anchor "BI"
    },
    {
        "span_text": "BI",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Bronchus Intermedius",
        "event_id": "E3",
        "context_prefix": "distal "
    },
    {
        "span_text": "LMSB",
        "label": "ANAT_AIRWAY",
        "normalized_value": "LMSB",
        "event_id": "E3",
        "context_prefix": "as well as "
    },
    {
        "span_text": "LUL",
        "label": "ANAT_LUNG_LOC",
        "normalized_value": "LUL",
        "event_id": "E3",
        "context_prefix": "anastomosis and "
    },
    {
        "span_text": "very thin mucosa",
        "label": "OBS_LESION",
        "normalized_value": "mucosal atrophy",
        "event_id": "E3",
        "context_prefix": "areas of "
    },
    {
        "span_text": "Lingula",
        "label": "ANAT_LUNG_LOC",
        "normalized_value": "Lingula",
        "event_id": "E3",
        "context_prefix": "LUL and "
    },

    # Event 4: Therapeutic Aspiration
    {
        "span_text": "Therapeutic aspiration",
        "label": "PROC_METHOD",
        "normalized_value": "Therapeutic aspiration",
        "event_id": "E4",
        "context_prefix": "31645 "
    },
    {
        "span_text": "Moderate thick secretions",
        "label": "OBS_LESION",
        "normalized_value": "secretions",
        "event_id": "E4",
        "context_prefix": "Lingula.  "
    },
    {
        "span_text": "Trachea",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Trachea",
        "event_id": "E4",
        "context_prefix": "clean out the "
    },
    {
        "span_text": "Right Mainstem",
        "label": "ANAT_AIRWAY",
        "normalized_value": "RMSB",
        "event_id": "E4",
        "context_prefix": "Distal 1/3), "
    },
    {
        "span_text": "Bronchus Intermedius",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Bronchus Intermedius",
        "event_id": "E4",
        "context_prefix": "Right Mainstem, "
    },
    {
        "span_text": "Left Mainstem",
        "label": "ANAT_AIRWAY",
        "normalized_value": "LMSB",
        "event_id": "E4",
        "context_prefix": "Intermedius , "
    },
    {
        "span_text": "Carina",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Carina",
        "event_id": "E4",
        "context_prefix": "Mainstem, and "
    },
    {
        "span_text": "mucus",
        "label": "OBS_LESION",
        "normalized_value": "mucus",
        "event_id": "E4",
        "context_prefix": "Carina from "
    },
    {
        "span_text": "Successful",
        "label": "OUTCOME_AIRWAY_LUMEN_POST", # Proxy for successful clearance
        "normalized_value": "successful",
        "event_id": "E4",
        "context_prefix": ""
    },
    
    # Event 5: Outcome
    {
        "span_text": "No immediate complications",
        "label": "OUTCOME_COMPLICATION",
        "normalized_value": "none",
        "event_id": "E5",
        "context_prefix": "well.  There were "
    }
]

# -------------------------------------------------------------------------
# GENERATOR CLASS
# -------------------------------------------------------------------------

class Phase0Generator:
    def __init__(self, note_text, template_path, output_path):
        self.note_text = note_text
        self.template_path = template_path
        self.output_path = output_path
        self.wb = None
        
    def load_workbook(self):
        if not os.path.exists(self.template_path):
            # Create a blank one for demonstration if template missing
            self.wb = openpyxl.Workbook()
            self._create_missing_sheets()
        else:
            self.wb = openpyxl.load_workbook(self.template_path)
            self._create_missing_sheets()

    def _create_missing_sheets(self):
        required = ["Note_Text", "Note_Index", "Span_Annotations", 
                    "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]
        for sheet in required:
            if sheet not in self.wb.sheetnames:
                self.wb.create_sheet(sheet)
                
    def write_note_text(self):
        ws = self.wb["Note_Text"]
        # Assuming header row 1 exists or we overwrite: note_id, source_file, note_text
        # We append to end
        next_row = ws.max_row + 1
        if next_row == 1:
            ws.append(["note_id", "source_file", "note_text"])
            next_row = 2
            
        ws.cell(row=next_row, column=1, value=NOTE_ID)
        ws.cell(row=next_row, column=2, value=SOURCE_FILE)
        ws.cell(row=next_row, column=3, value=self.note_text)

    def write_note_index(self):
        ws = self.wb["Note_Index"]
        # Header if missing
        headers = ["source_file", "note_id", "encounter_id", "procedure_date", 
                   "site", "reviewer", "status", "free_text_notes"] + list(PROCEDURE_FLAGS.keys())
        
        if ws.max_row == 1 and ws.cell(1,1).value is None:
            ws.append(headers)
            
        row_data = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "pending", ""]
        row_data += [PROCEDURE_FLAGS[k] for k in PROCEDURE_FLAGS]
        ws.append(row_data)

    def compute_offsets(self, span_text, context_prefix=None, match_index=None):
        """
        Deterministic hydration logic.
        """
        st = span_text
        txt = self.note_text
        
        if not st:
            return None, None, "error_empty"

        count = txt.count(st)
        
        if count == 0:
            return None, None, "not_found"
            
        # Case 1: Unique
        if count == 1:
            start = txt.find(st)
            end = start + len(st)
            return start, end, "hydrated_unique"
            
        # Case 2: Context Prefix
        if context_prefix:
            # Find all occurrences
            starts = [m.start() for m in re.finditer(re.escape(st), txt)]
            best_start = -1
            
            for s in starts:
                # Look at window before
                window_start = max(0, s - 120)
                pre_window = txt[window_start:s]
                if context_prefix in pre_window:
                    best_start = s
                    break
            
            if best_start != -1:
                return best_start, best_start + len(st), "hydrated_prefix_window"

        # Case 3: Match Index (Fallback if logic was implemented, but here we assume simple sequential)
        # If ambiguous and no context, we default to ambiguous
        return None, None, f"ambiguous_count={count}"

    def write_spans(self):
        ws_anno = self.wb["Span_Annotations"]
        ws_hydra = self.wb["Span_Hydrated"]
        
        # Headers
        headers = ["source_file", "note_id", "span_id", "section_type",
                   "context_prefix", "span_text", "match_index",
                   "start_char", "end_char", "span_len",
                   "label", "normalized_value", "schema_field", "event_id",
                   "is_negated", "is_historical", "time_anchor", "reviewer", "comments",
                   "hydration_status"]
                   
        if ws_anno.max_row == 1 and ws_anno.cell(1,1).value is None:
            ws_anno.append(headers)
        if ws_hydra.max_row == 1 and ws_hydra.cell(1,1).value is None:
            ws_hydra.append(headers)

        span_counter = 1
        
        for item in RAW_SPANS:
            s_text = item["span_text"]
            c_prefix = item.get("context_prefix", "")
            
            # Compute hydration
            start, end, status = self.compute_offsets(s_text, c_prefix)
            
            span_len = len(s_text)
            span_id = f"{NOTE_ID}_s{span_counter:03d}"
            
            # Common row data (blank start/end for Anno)
            row_base = [
                SOURCE_FILE, NOTE_ID, span_id, "PROCEDURE",
                c_prefix, s_text, "", # match_index left blank for now
                "", "", f"=LEN(F{ws_anno.max_row + 1})", # Formula for length
                item["label"], item["normalized_value"], "", item["event_id"],
                "FALSE", "FALSE", "", "", "",
                "needs_hydration"
            ]
            
            ws_anno.append(row_base)
            
            # Hydrated row
            row_hydra = list(row_base)
            row_hydra[7] = start if start is not None else ""
            row_hydra[8] = end if end is not None else ""
            row_hydra[9] = span_len
            row_hydra[19] = status
            
            ws_hydra.append(row_hydra)
            span_counter += 1

    def write_events(self):
        ws = self.wb["Event_Log"]
        headers = ["source_file", "note_id", "event_id", "event_type", "method",
                   "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements",
                   "specimens", "findings", "is_historical", "reviewer", "comments",
                   "device_size", "device_material",
                   "outcome_airway_lumen_pre", "outcome_airway_lumen_post",
                   "outcome_symptoms", "outcome_pleural", "outcome_complication"]
        
        if ws.max_row == 1 and ws.cell(1,1).value is None:
            ws.append(headers)
            
        # Define events based on parsed data (Simplified for script generation)
        # In a real pipeline, we'd aggregate spans. Here we map explicitly.
        
        events_data = [
            {
                "eid": "E1", "type": "Diagnostic Bronchoscopy", "method": "Tracheobronchoscopy",
                "anat": "Trachea", "dev": "Tracheostomy tube", "find": "Normal caliber, sharp carina",
                "out_comp": ""
            },
            {
                "eid": "E2", "type": "Inspection", "method": "Inspection",
                "anat": "RMSB", "dev": "BMS", "find": "Patent, Dehiscence closed",
                "out_comp": ""
            },
            {
                "eid": "E3", "type": "Inspection", "method": "Inspection",
                "anat": "LMSB, LUL, BI", "dev": "", "find": "Granulation tissue, Ischemia, Atrophy",
                "out_comp": ""
            },
            {
                "eid": "E4", "type": "Therapeutic Aspiration", "method": "Aspiration",
                "anat": "Trachea, RMSB, BI, LMSB, Carina", "dev": "", "find": "Thick secretions, Mucus",
                "out_air_post": "Successful",
                "out_comp": ""
            },
            {
                "eid": "E5", "type": "Outcome", "method": "",
                "anat": "", "dev": "", "find": "",
                "out_comp": "No immediate complications"
            }
        ]
        
        for e in events_data:
            row = [
                SOURCE_FILE, NOTE_ID, e["eid"], e["type"], e["method"],
                e["anat"], e["dev"], "", "", "", "",
                "", e["find"], "FALSE", "", "",
                "", "",
                "", e.get("out_air_post", ""),
                "", "", e["out_comp"]
            ]
            ws.append(row)

    def write_v3_json(self):
        ws = self.wb["V3_Registry_JSON"]
        
        # Construct JSON object
        registry_data = {
            "schema_version": "3.0",
            "note_id": NOTE_ID,
            "no_immediate_complications": True,
            "procedures": [
                {
                    "event_id": "E1",
                    "type": "Bronchoscopy",
                    "method": "Tracheobronchoscopy",
                    "target": {"anatomy_type": "Airway", "location": {"lobe": "Trachea"}},
                    "findings_json": ["Normal caliber"]
                },
                {
                    "event_id": "E2",
                    "type": "Inspection",
                    "target": {"anatomy_type": "Airway", "location": {"lobe": "RMSB"}},
                    "stent": {"type": "BMS", "status": "Patent"},
                    "findings_json": ["Dehiscence closed"]
                },
                {
                    "event_id": "E4",
                    "type": "Therapeutic Aspiration",
                    "target": {"anatomy_type": "Airway", "location": {"lobe": "Multiple"}},
                    "outcomes": {"airway": {"lumen_post": "Successful"}}
                }
            ]
        }
        
        json_str = json.dumps(registry_data, indent=2)
        
        if ws.max_row == 1:
            ws.cell(1, 1, "json_output")
        ws.cell(2, 1, json_str)

    def run(self):
        self.load_workbook()
        self.write_note_text()
        self.write_note_index()
        self.write_spans() # Calls hydration internally
        self.write_events()
        # V3_Procedure_Events sheet left blank structure for this condensed script
        if "V3_Procedure_Events" not in self.wb.sheetnames:
            self.wb.create_sheet("V3_Procedure_Events")
        self.wb["V3_Procedure_Events"].append(["note_id", "event_id", "type", "method", "target", "outcomes"])
        
        self.write_v3_json()
        self.wb.save(self.output_path)
        print(f"Generated {self.output_path}")

# -------------------------------------------------------------------------
# EXECUTION
# -------------------------------------------------------------------------
if __name__ == "__main__":
    generator = Phase0Generator(NOTE_TEXT, TEMPLATE_PATH, OUTPUT_PATH)
    generator.run()