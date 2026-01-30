import json
import re
import sys
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

# =============================================================================
# 1. INPUTS & CONFIGURATION
# =============================================================================
NOTE_ID = "note_109"
SOURCE_FILE = "note_109.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# Embedded note text as fallback; script prioritizes reading local file if it exists.
NOTE_TEXT = """NOTE_ID:  note_109 SOURCE_FILE: note_109.txt INDICATION FOR OPERATION:  [REDACTED]is a 56 year old-year-old male who presents with Pleural Effusion.
CONSENT : Obtained before the procedure. The nature, indications, purpose, benefits, risks, potential complications, and alternatives to Chest Ultrasound and Chest tube placement were discussed with the patient or surrogate decision-maker in detail.  Patient or surrogate deci
PREOPERATIVE DIAGNOSIS:  Pleural Effusion
POSTOPERATIVE DIAGNOSIS: Same as preoperative diagnosis - see above. 
PROCEDURE:  
76604 Ultrasound, chest (includes mediastinum), real time with image documentation
32557 Insert catheter pleura with imaging (chest tube)
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)
IP [REDACTED] CODE MOD DETAILS: Unusual Procedure:
This patient required significant clinical expertise given the difficult location of the effusion, relatively small size, and severity of patient's condition. This resulted in >50% increased work due to Increased intensity, Technical difficulty of procedure, Severity of patient's condition, and Physical and mental effort required. Apply to: 
32557 Insert catheter pleura with imaging (chest tube)
Local Anesthesia (Lidocaine)
99152 Moderate sedation: initial 15 minutes
99153 Moderate sedation: each additional 15 minutes 
99153 Moderate sedation: each additional 15 minutes 
PROCEDURE IN DETAIL: 
Time out was performed. The area was prepped and draped in the usual sterile fashion. Local anesthesia was administered. 
Under ultrasound guidance, a needle was advanced into the pleural space. Fluid was aspirated confirming position. A guidewire was advanced through the needle. The needle was removed. 
The needle was pointed in the basilar direction. The introducer needle was then removed. A dilator was then inserted over the wire with a twisting motion through the thoracic wall tissue in order to form a tract for catheter insertion. The dilator was removed and the pigtail catheter (with trochar) was advanced over the guidewire in the basilar direction. The catheter was inserted into the chest until all catheter holes were well within the chest. The guidewire and trochar were then removed.  150 mL of pleural fluid was manually aspirated under sterile conditions and sent to the lab for analysis.  The tube was then attached to the collection drain apparatus and secured in place with suture and covered. 
Fluid Volume Removed: 150 ml 
Fluid Description:  Serous
Pleural pressures:   Not measured.
Drainage method:   Pleurovac
Suction: No (to Water Seal)
ULTRASOUND-GUIDANCE / RADIOLOGICAL GUIDANCE IMAGE:
Confirming pleural fluid pocket at needle and catheter-over-needle insertion site.
 
SPECIMEN(S): 
0‌None
1‌Analyses - pH, TProt, LDH, Glucose, and Cholesterol 
1‌Cell Count
1‌Cultures - Gram Stain/Culture, Blood culture bottle bedside innoculation, AFB, and Fungal
1‌Cytology 
0‌Other:
CXR ordered: yes
ESTIMATED BLOOD LOSS:  <5 mL
COMPLICATIONS:  None.  There were no immediate complications.
IMPRESSION/PLAN: [REDACTED]is a 56 year old-year-old male who presents for Chest Ultrasound and Chest tube placement on the LEFT.  The patient tolerated the procedure well.  
DISPOSITION: Remain ICU status"""

# Try to read file, fallback to embedded
try:
    if Path(SOURCE_FILE).exists():
        with open(SOURCE_FILE, "r", encoding="utf-8") as f:
            NOTE_TEXT = f.read()
except Exception as e:
    print(f"Warning: Could not read {SOURCE_FILE}, using embedded text. Error: {e}")

# =============================================================================
# 2. EXTRACTION DATA (Hardcoded for this note)
# =============================================================================

# Procedure Flags (Note_Index)
PROC_FLAGS = {
    "thoracentesis": 0,
    "chest_tube": 1,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0,
    # Bronchoscopy defaults (all 0 for this note)
    "diagnostic_bronchoscopy": 0, "bal": 0, "bronchial_wash": 0, "brushings": 0, 
    "endobronchial_biopsy": 0, "tbna_conventional": 0, "linear_ebus": 0, "radial_ebus": 0, 
    "navigational_bronchoscopy": 0, "transbronchial_biopsy": 0, "transbronchial_cryobiopsy": 0, 
    "therapeutic_aspiration": 0, "foreign_body_removal": 0, "airway_dilation": 0, 
    "airway_stent": 0, "thermal_ablation": 0, "tumor_debulking_non_thermal": 0, 
    "cryotherapy": 0, "blvr": 0, "peripheral_ablation": 0, "bronchial_thermoplasty": 0, 
    "whole_lung_lavage": 0, "rigid_bronchoscopy": 0
}

# Spans (Anchor-First)
SPANS = [
    # Diagnosis/Indication
    {
        "span_text": "Pleural Effusion",
        "label": "OBS_LESION",
        "context_prefix": "presents with ",
        "match_index": 0,
        "normalized_value": "Pleural effusion"
    },
    # Procedure Methods
    {
        "span_text": "Chest Ultrasound",
        "label": "PROC_METHOD",
        "context_prefix": "alternatives to ",
        "match_index": 0,
        "normalized_value": "Chest Ultrasound"
    },
    {
        "span_text": "Chest tube placement",
        "label": "PROC_METHOD",
        "context_prefix": "Ultrasound and ",
        "match_index": 0,
        "normalized_value": "Chest tube placement",
        "event_id": "evt1"
    },
    # Anatomy/Laterality
    {
        "span_text": "LEFT",
        "label": "LATERALITY",
        "context_prefix": "placement on the ",
        "match_index": 0,
        "normalized_value": "Left",
        "event_id": "evt1"
    },
    {
        "span_text": "pleural fluid pocket",
        "label": "ANAT_PLEURA",
        "context_prefix": "Confirming ",
        "match_index": 0,
        "normalized_value": "Pleural space"
    },
    # Devices
    {
        "span_text": "introducer needle",
        "label": "DEV_NEEDLE",
        "context_prefix": "direction. The ",
        "match_index": 0,
        "normalized_value": "Introducer needle",
        "event_id": "evt1"
    },
    {
        "span_text": "pigtail catheter",
        "label": "DEV_CATHETER",
        "context_prefix": "removed and the ",
        "match_index": 0,
        "normalized_value": "Pigtail catheter",
        "event_id": "evt1"
    },
    # Measurements
    {
        "span_text": "150 mL",
        "label": "MEAS_VOL",
        "context_prefix": "trochar were then removed.  ",
        "match_index": 0,
        "normalized_value": "150 mL",
        "event_id": "evt1"
    },
    # Outcomes
    {
        "span_text": "No immediate complications",
        "label": "OUTCOME_COMPLICATION",
        "context_prefix": "None.  There were ",
        "match_index": 0,
        "normalized_value": "none",
        "event_id": "evt1"
    },
    {
        "span_text": "Serous",
        "label": "OBS_LESION",
        "context_prefix": "Fluid Description:  ",
        "match_index": 0,
        "normalized_value": "Serous fluid",
        "comments": "Fluid character"
    }
]

# Events (V3 Structure)
EVENTS_V3 = [
    {
        "note_id": NOTE_ID,
        "event_id": "evt1",
        "type": "chest_tube",
        "target": {
            "anatomy_type": "Pleura",
            "location": {"lobe": "Left"}
        },
        "method": "Chest tube placement",
        "devices_json": json.dumps(["pigtail catheter"]),
        "measurements_json": json.dumps({"volume_removed": "150 mL"}),
        "findings_json": json.dumps(["Pleural Effusion", "Serous fluid"]),
        "outcomes": {
            "pleural": "150 mL removed",
            "complications": "none"
        }
    }
]

# Event Log (Flat)
EVENT_LOG_ROWS = [
    {
        "event_id": "evt1",
        "event_type": "chest_tube",
        "method": "Chest tube placement",
        "anatomy_target": "Left Pleura",
        "device": "pigtail catheter",
        "measurements": "150 mL",
        "findings": "Pleural Effusion, Serous",
        "outcome_pleural": "150 mL removed",
        "outcome_complication": "none"
    }
]

# =============================================================================
# 3. HELPER FUNCTIONS
# =============================================================================

def get_hydration_result(full_text, span_text, context_prefix=None, match_index=None):
    """
    Computes start_char/end_char based on anchor logic.
    Returns: (start_char, end_char, hydration_status)
    """
    if not span_text or span_text not in full_text:
        return "", "", "not_found"

    # Find all occurrences
    matches = [m.start() for m in re.finditer(re.escape(span_text), full_text)]
    
    # 1. Unique Match
    if len(matches) == 1:
        start = matches[0]
        return start, start + len(span_text), "hydrated_unique"

    # 2. Context Prefix Match
    if context_prefix:
        valid_matches = []
        for m_start in matches:
            # Look back 120 chars
            window_start = max(0, m_start - 120)
            window = full_text[window_start:m_start]
            if context_prefix in window:
                valid_matches.append(m_start)
        
        if len(valid_matches) == 1:
            start = valid_matches[0]
            return start, start + len(span_text), "hydrated_prefix_window"
        elif len(valid_matches) > 1:
            # If multiple valid context matches, try match_index on the filtered list
            idx = int(match_index) if match_index is not None else 0
            if idx < len(valid_matches):
                start = valid_matches[idx]
                return start, start + len(span_text), "hydrated_match_index_context"

    # 3. Pure Match Index (fallback if no context or context failed)
    idx = int(match_index) if match_index is not None else 0
    if idx < len(matches):
        start = matches[idx]
        return start, start + len(span_text), "hydrated_match_index"

    return "", "", f"ambiguous_count={len(matches)}"

# =============================================================================
# 4. WORKBOOK GENERATION
# =============================================================================

def generate_workbook():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        print(f"Error: Template {TEMPLATE_PATH} not found.")
        sys.exit(1)

    # ---------------------------------------------------------
    # Sheet 1: Note_Text
    # ---------------------------------------------------------
    ws_text = wb["Note_Text"] if "Note_Text" in wb.sheetnames else wb.create_sheet("Note_Text")
    if ws_text.max_row == 1:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # ---------------------------------------------------------
    # Sheet 2: Note_Index (Flags)
    # ---------------------------------------------------------
    ws_index = wb["Note_Index"] if "Note_Index" in wb.sheetnames else wb.create_sheet("Note_Index")
    # Headers: source_file, note_id, encounter_id, procedure_date, site, reviewer, status, free_text_notes, [FLAGS...]
    
    # Construct row
    row_data = [SOURCE_FILE, NOTE_ID, "", "", "", "", "ready", ""]
    
    # Map flags to columns (assuming template order or finding them)
    # We will assume the template has headers in row 1. We'll map by name.
    header_map = {cell.value: idx for idx, cell in enumerate(ws_index[1], 1)}
    
    # Append a blank row first (or data row)
    ws_index.append(row_data)
    current_row = ws_index.max_row
    
    # Fill flags
    for flag, val in PROC_FLAGS.items():
        if flag in header_map:
            col_idx = header_map[flag]
            ws_index.cell(row=current_row, column=col_idx, value=val)

    # ---------------------------------------------------------
    # Sheet 3 & 4: Span_Annotations (Anchors) & Span_Hydrated
    # ---------------------------------------------------------
    ws_anno = wb["Span_Annotations"]
    ws_hydra = wb["Span_Hydrated"]
    
    # Clear existing data rows if any (optional, but good for fresh extraction)
    # We'll just append.

    for span in SPANS:
        # Common fields
        span_id = f"{NOTE_ID}_s{len(ws_anno['A'])}" # Simple ID generation
        span_text = span.get("span_text", "")
        ctx = span.get("context_prefix", "")
        midx = span.get("match_index", "")
        label = span.get("label", "")
        norm = span.get("normalized_value", "")
        evt_id = span.get("event_id", "")
        
        # Annotation Row (No offsets)
        anno_row = [
            SOURCE_FILE, NOTE_ID, span_id, "", # section_type
            ctx, span_text, midx,
            "", "", f"=LEN(F{ws_anno.max_row+1})", # start, end, len formula
            label, norm, "", evt_id,
            "", "", "", "", span.get("comments", ""), "needs_hydration"
        ]
        ws_anno.append(anno_row)
        
        # Hydration
        start, end, status = get_hydration_result(NOTE_TEXT, span_text, ctx, midx)
        hydra_row = [
            SOURCE_FILE, NOTE_ID, span_id, "",
            ctx, span_text, midx,
            start, end, (end - start) if isinstance(start, int) else "",
            label, norm, "", evt_id,
            "", "", "", "", span.get("comments", ""), status
        ]
        ws_hydra.append(hydra_row)

    # ---------------------------------------------------------
    # Sheet 5: Event_Log
    # ---------------------------------------------------------
    ws_event = wb["Event_Log"]
    for ev in EVENT_LOG_ROWS:
        row = [
            SOURCE_FILE, NOTE_ID, ev.get("event_id"), ev.get("event_type"), ev.get("method"),
            ev.get("anatomy_target"), ev.get("device"), "", "", "", ev.get("measurements"),
            "", ev.get("findings"), "", "", "",
            "", "", # dev size/material
            "", "", "", ev.get("outcome_pleural"), ev.get("outcome_complication")
        ]
        ws_event.append(row)

    # ---------------------------------------------------------
    # Sheet 6: V3_Procedure_Events
    # ---------------------------------------------------------
    ws_v3 = wb["V3_Procedure_Events"]
    for v3 in EVENTS_V3:
        tgt = v3.get("target", {})
        loc = tgt.get("location", {})
        out = v3.get("outcomes", {})
        
        row = [
            NOTE_ID, v3.get("event_id"), v3.get("type"),
            tgt.get("anatomy_type"), loc.get("lobe"), loc.get("segment"), "",
            "", "", # lesion type/size
            v3.get("method"), v3.get("devices_json"), v3.get("measurements_json"), "", v3.get("findings_json"), "",
            "", "", "", # stent/cath details
            "", "", "", out.get("pleural"), out.get("complications")
        ]
        ws_v3.append(row)

    # ---------------------------------------------------------
    # Sheet 7: V3_Registry_JSON
    # ---------------------------------------------------------
    ws_json = wb["V3_Registry_JSON"]
    registry_obj = {
        "schema_version": "ip_registry_v3_draft",
        "note_id": NOTE_ID,
        "source_file": SOURCE_FILE,
        "procedure_date": "",
        "procedures": EVENTS_V3,
        "no_immediate_complications": True
    }
    json_str = json.dumps(registry_obj, indent=2)
    ws_json.append([json_str])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Successfully generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()