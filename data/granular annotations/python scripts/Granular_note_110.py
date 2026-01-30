import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import datetime

# =============================================================================
# INPUT DATA
# =============================================================================
NOTE_ID = "note_110"
SOURCE_FILE = "note_110.txt"
PROCEDURE_DATE = ""
NOTE_TEXT = """NOTE_ID:  note_110 SOURCE_FILE: note_110.txt INDICATION FOR OPERATION:  [REDACTED]is a 56 year old-year-old male who presents with Pleural Effusion.
The nature, purpose, risks, benefits and alternatives to Chest Ultrasound and Thoracentesis were discussed with the patient in detail.
Patient indicated a wish to proceed with procedure and informed consent was signed.
PREOPERATIVE DIAGNOSIS:  Pleural Effusion
POSTOPERATIVE DIAGNOSIS: Same as preoperative diagnosis - see above.
PROCEDURE:  
76604 Ultrasound, chest (includes mediastinum), real time with image documentation
32555 Aspirate pleura with imaging (thoracentesis)
53 Procedure was started, but terminated before completion due to unexpected circumstances or for the patient's well being
Local ONLY
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
PROCEDURE IN DETAIL:
PATIENT POSITION: 
1‌ Supine  0‌ Sitting   
0‌ Lateral Decubitus:  0‌ Right 0‌ Left 
 
CHEST ULTRASOUND FINDINGS:  1‌ Image saved and printed 
Hemithorax:   0‌ Right  1‌ Left 
Pleural Effusion: 
Volume:       0‌ None  0‌ Minimal  1‌ Small  0‌ Moderate  0‌ Large 
Echogenicity:   1‌ Anechoic  0‌ Hypoechoic  0‌ Isoechoic  0‌ Hyperechoic 
Loculations:  1‌ None  0‌Thin  0‌ Thick 
Diaphragmatic Motion:  0‌ Normal  0‌ Diminished  0‌ Absent  
Lung: 
Lung sliding before procedure: 
  1‌ Present  0‌ Absent 
Lung sliding post procedure:   1‌ Present  0‌ Absent 
Lung consolidation/atelectasis: 0‌ Present  0‌  Absent 
Pleura:  1‌ Normal  0‌ Thick  0‌ Nodular 
Based on Ultrasound Evaluation: the small effusion was visible in the 2nd intercostal space but without an access point that would avoid the lung.
There was a small window for access identified above the clavicle. 
Insertion site prepped and draped in sterile fashion.
Thoracentesis Kit was used with 5Fr Yeuh catheter. 
ANESTHESIA:   Lidocaine 1%: __8____ ml      Other: ______ 
Entry Site: 
0‌ Right ___ Intercostal Space   1‌ Left  __1st_ Intercostal Space 
1‌ Mid-clavicular   0‌ Mid-axillary  0‌ Mid-scapular  0‌ Other: 
After advancement of finder needle into the space under real-time US guidance, there was resistance and no fluid was aspirated despite three attempts.
As such, procedure aborted 
 
COMPLICATIONS:
1‌None 0‌Bleeding-EBL: ___ ml 0‌Pneumothorax 0‌Re- Expansion Pulmonary Edema 
0‌Other: 
IMPRESSION/PLAN: [REDACTED]is a 56 year old-year-old male who presents for Chest Ultrasound and Thoracentesis LEFT.
The procedure was aborted as no fluid could be aspirated.
- consider CT-guided approach, though this also may be challenging unless the effusion grows further 
DISPOSITION: ICU"""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# =============================================================================
# CONFIGURATION
# =============================================================================

# Procedure Flags (30) - Ordered as per instructions
PROCEDURE_FLAGS_ORDER = [
    "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
    "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy",
    "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration",
    "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
    "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation",
    "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
    # Pleural
    "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis",
    "pleural_biopsy", "fibrinolytic_therapy"
]

# Manual Flags for this note
NOTE_FLAGS = {
    "thoracentesis": 1,  # Explicitly attempted
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0,
    # Bronch default 0
    "diagnostic_bronchoscopy": 0,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0
}

# Spans to Extract (Anchor-First Definition)
# Structure: (span_text, label, normalized_value, schema_field, event_id_ref, context_prefix or None)
SPANS_DATA = [
    ("Pleural Effusion", "OBS_LESION", "Pleural Effusion", "indication", "E0", "presents with "),
    ("Chest Ultrasound", "PROC_METHOD", "Ultrasound", "method", "E1", "alternatives to "),
    ("Thoracentesis", "PROC_METHOD", "Thoracentesis", "method", "E2", "Ultrasound and "),
    ("Ultrasound, chest", "PROC_METHOD", "Ultrasound", "method", "E1", None),
    ("Aspirate pleura", "PROC_METHOD", "Thoracentesis", "method", "E2", None),
    ("Left", "LATERALITY", "Left", "target.laterality", "E1", "Hemithorax:   0‌ Right  1‌ "),
    ("Small", "MEAS_SIZE", "Small", "lesion.size_qual", "E1", "Volume:       0‌ None  0‌ Minimal  1‌ "),
    ("Anechoic", "OBS_LESION", "Anechoic", "lesion.characteristics", "E1", "Echogenicity:   1‌ "),
    ("small effusion", "OBS_LESION", "Pleural Effusion", "lesion.type", "E1", "Evaluation: the "),
    ("2nd intercostal space", "ANAT_PLEURA", "2nd Intercostal Space", "target.location", "E1", None),
    ("above the clavicle", "ANAT_PLEURA", "Supraclavicular", "target.location", "E2", "identified "),
    ("Thoracentesis Kit", "DEV_INSTRUMENT", "Thoracentesis Kit", "device.type", "E2", None),
    ("5Fr", "DEV_CATHETER_SIZE", "5Fr", "catheter.size_fr", "E2", None),
    ("Yeuh catheter", "DEV_CATHETER", "Yeuh Catheter", "device.type", "E2", None),
    ("Left", "LATERALITY", "Left", "target.laterality", "E2", "Right ___ Intercostal Space   1‌ "),
    ("1st_ Intercostal Space", "ANAT_PLEURA", "1st Intercostal Space", "target.location", "E2", None),
    ("finder needle", "DEV_NEEDLE", "Finder Needle", "device.type", "E2", None),
    ("no fluid was aspirated", "OUTCOME_PLEURAL", "Dry Tap", "outcomes.pleural", "E2", None),
    ("procedure aborted", "OUTCOME_PLEURAL", "Aborted", "outcomes.pleural", "E2", "As such, "),
    ("None", "OUTCOME_COMPLICATION", "None", "outcomes.complications", "E2", "COMPLICATIONS:\n1‌"),
    ("Thoracentesis LEFT", "PROC_METHOD", "Thoracentesis", "method", "E2", None),
    ("aborted", "OUTCOME_PLEURAL", "Aborted", "outcomes.pleural", "E2", "procedure was ")
]

# Event Definitions for V3
EVENTS_DATA = [
    {
        "event_id": "E1",
        "type": "Ultrasound",
        "method": "Ultrasound",
        "target_anatomy": "Pleural Space",
        "laterality": "Left",
        "findings": "Small anechoic effusion, visible 2nd ICS, no safe access there",
    },
    {
        "event_id": "E2",
        "type": "Thoracentesis",
        "method": "Thoracentesis",
        "target_anatomy": "Pleural Space",
        "laterality": "Left",
        "location": "1st Intercostal Space (Supraclavicular)",
        "device": "5Fr Yeuh Catheter, Finder Needle",
        "outcome_pleural": "Aborted (Dry Tap)",
        "complication": "None"
    }
]

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def find_offsets_hydrated(text, span_text, context_prefix=None, match_index=None):
    """
    Finds start_char and end_char for a span based on hydration logic.
    Returns: (start_char, end_char, hydration_status)
    """
    if not span_text:
        return None, None, "missing_text"
    
    # Escape special regex chars in span_text
    escaped_span = re.escape(span_text)
    
    # Case 1: Unique occurrence
    matches = list(re.finditer(escaped_span, text))
    if len(matches) == 1:
        m = matches[0]
        return m.start(), m.end(), "hydrated_unique"
    
    # Case 2: Context Prefix
    if context_prefix:
        # We look for the prefix within a window before the match
        best_match = None
        for m in matches:
            start_search = max(0, m.start() - 150)
            window = text[start_search:m.start()]
            if context_prefix in window:
                best_match = m
                break
        if best_match:
            return best_match.start(), best_match.end(), "hydrated_prefix_window"
    
    # Case 3: Match Index
    if match_index is not None and 0 <= match_index < len(matches):
        m = matches[match_index]
        return m.start(), m.end(), "hydrated_match_index"
    
    # Fallback: Ambiguous
    if len(matches) > 1:
        return None, None, f"ambiguous_count={len(matches)}"
    
    return None, None, "not_found"

def create_workbook():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        # Create a basic workbook if template missing (fallback)
        wb = openpyxl.Workbook()
        for sheet_name in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
    return wb

# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    wb = create_workbook()
    
    # -------------------------------------------------------------------------
    # 1. Note_Text
    # -------------------------------------------------------------------------
    ws_text = wb["Note_Text"]
    # Check if header exists, if not add it
    if ws_text.max_row == 1 and ws_text.cell(1,1).value is None:
        ws_text.append(["note_id", "source_file", "note_text"])
    
    # Append row
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # -------------------------------------------------------------------------
    # 2. Note_Index
    # -------------------------------------------------------------------------
    ws_index = wb["Note_Index"]
    header = [
        "source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes"
    ] + PROCEDURE_FLAGS_ORDER
    
    # If empty, write header
    if ws_index.max_row == 1 and ws_index.cell(1,1).value is None:
        ws_index.append(header)
        
    row_data = [
        SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "extraction_ready", ""
    ]
    for flag in PROCEDURE_FLAGS_ORDER:
        row_data.append(NOTE_FLAGS.get(flag, 0))
    
    ws_index.append(row_data)
    
    # -------------------------------------------------------------------------
    # 3. Span_Annotations (Anchor First) & 4. Span_Hydrated
    # -------------------------------------------------------------------------
    ws_anno = wb["Span_Annotations"]
    ws_hydra = wb["Span_Hydrated"]
    
    # Headers
    span_header = [
        "source_file", "note_id", "span_id", "section_type",
        "context_prefix", "span_text", "match_index",
        "start_char", "end_char", "span_len",
        "label", "normalized_value", "schema_field", "event_id",
        "is_negated", "is_historical", "time_anchor", "reviewer", "comments",
        "hydration_status"
    ]
    
    if ws_anno.max_row == 1 and ws_anno.cell(1,1).value is None:
        ws_anno.append(span_header)
    if ws_hydra.max_row == 1 and ws_hydra.cell(1,1).value is None:
        ws_hydra.append(span_header)
        
    span_id_counter = 1
    
    for item in SPANS_DATA:
        span_text, label, norm_val, schema_field, event_id, ctx_prefix = item
        
        # Calculate hydration
        start, end, status = find_offsets_hydrated(NOTE_TEXT, span_text, context_prefix=ctx_prefix)
        span_len_formula = f"=LEN(F{ws_anno.max_row + 1})"
        
        # Row for Annotations (Anchor First - no offsets)
        row_anno = [
            SOURCE_FILE, NOTE_ID, f"span_{span_id_counter:03d}", "Procedure Note",
            ctx_prefix, span_text, None, # match_index left blank unless strictly needed
            None, None, span_len_formula, # start/end blank
            label, norm_val, schema_field, event_id,
            "FALSE", "FALSE", "", "", "",
            "needs_hydration"
        ]
        ws_anno.append(row_anno)
        
        # Row for Hydrated
        row_hydra = [
            SOURCE_FILE, NOTE_ID, f"span_{span_id_counter:03d}", "Procedure Note",
            ctx_prefix, span_text, None,
            start, end, (end - start) if (start is not None and end is not None) else None,
            label, norm_val, schema_field, event_id,
            "FALSE", "FALSE", "", "", "",
            status
        ]
        ws_hydra.append(row_hydra)
        
        span_id_counter += 1

    # -------------------------------------------------------------------------
    # 5. Event_Log
    # -------------------------------------------------------------------------
    ws_event = wb["Event_Log"]
    event_header = [
        "source_file", "note_id", "event_id", "event_type", "method",
        "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements",
        "specimens", "findings", "is_historical", "reviewer", "comments",
        "device_size", "device_material",
        "outcome_airway_lumen_pre", "outcome_airway_lumen_post",
        "outcome_symptoms", "outcome_pleural", "outcome_complication"
    ]
    if ws_event.max_row == 1 and ws_event.cell(1,1).value is None:
        ws_event.append(event_header)
        
    # E1: Ultrasound
    ws_event.append([
        SOURCE_FILE, NOTE_ID, "E1", "Ultrasound", "Ultrasound",
        "Pleural Space (Left)", "", "", "", "", "Small effusion",
        "", "Visible 2nd ICS, no safe access", "FALSE", "", "",
        "", "",
        "", "",
        "", "Diagnostic Only", "None"
    ])
    
    # E2: Thoracentesis
    ws_event.append([
        SOURCE_FILE, NOTE_ID, "E2", "Thoracentesis", "Thoracentesis",
        "Pleural Space (Left, 1st ICS)", "Yeuh Catheter, Finder Needle", "", "", "", "",
        "", "Aborted due to dry tap", "FALSE", "", "",
        "5Fr", "",
        "", "",
        "", "Aborted (Dry Tap)", "None"
    ])

    # -------------------------------------------------------------------------
    # 6. V3_Procedure_Events
    # -------------------------------------------------------------------------
    ws_v3 = wb["V3_Procedure_Events"]
    v3_header = [
        "note_id", "event_id", "type",
        "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station",
        "lesion.type", "lesion.size_mm",
        "method", "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote",
        "stent.size", "stent.material_or_brand", "catheter.size_fr",
        "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post",
        "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"
    ]
    if ws_v3.max_row == 1 and ws_v3.cell(1,1).value is None:
        ws_v3.append(v3_header)
        
    # E1 Row
    ws_v3.append([
        NOTE_ID, "E1", "Ultrasound",
        "Pleural Space", "Left", "", "",
        "Pleural Effusion", "Small",
        "Ultrasound", json.dumps([]), json.dumps({"volume": "Small"}), json.dumps([]), json.dumps(["Visible 2nd ICS", "No access"]), "",
        "", "", "",
        "", "",
        "", "Diagnostic", "None"
    ])
    
    # E2 Row
    ws_v3.append([
        NOTE_ID, "E2", "Thoracentesis",
        "Pleural Space", "Left", "", "1st ICS",
        "Pleural Effusion", "",
        "Thoracentesis", json.dumps(["5Fr Yeuh Catheter", "Finder Needle"]), json.dumps([]), json.dumps([]), json.dumps(["Dry Tap", "Aborted"]), "",
        "", "", "5Fr",
        "", "",
        "", "Aborted", "None"
    ])

    # -------------------------------------------------------------------------
    # 7. V3_Registry_JSON
    # -------------------------------------------------------------------------
    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 1 and ws_json.cell(1,1).value is None:
        ws_json.append(["schema_version", "note_id", "json_output", "no_immediate_complications"])
        
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": [
            {
                "event_id": "E1",
                "type": "Ultrasound",
                "target": {"anatomy": "Pleural Space", "laterality": "Left"},
                "findings": ["Small effusion", "No safe access"]
            },
            {
                "event_id": "E2",
                "type": "Thoracentesis",
                "target": {"anatomy": "Pleural Space", "laterality": "Left", "location": "1st ICS"},
                "devices": [{"type": "Catheter", "name": "Yeuh", "size": "5Fr"}, {"type": "Needle", "name": "Finder"}],
                "outcomes": {"pleural": "Aborted (Dry Tap)", "complications": "None"}
            }
        ],
        "no_immediate_complications": True
    }
    
    ws_json.append(["3.0", NOTE_ID, json.dumps(registry_data, indent=2), "True"])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    main()