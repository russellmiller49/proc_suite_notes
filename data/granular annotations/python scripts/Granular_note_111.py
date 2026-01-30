import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
import json
import re
import os

# =============================================================================
# INPUT DATA
# =============================================================================

NOTE_ID = "note_111"
SOURCE_FILE = "note_111.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# Cleaned text removing [source] tags for accurate span matching
NOTE_TEXT = """NOTE_ID:  note_111 SOURCE_FILE: note_111.txt INDICATION FOR OPERATION:  [REDACTED]is a 56 year old-year-old male who presents with Pleural Effusion.
The nature, purpose, risks, benefits and alternatives to Chest Ultrasound and Chest tube placement were discussed with the patient in detail.
Patient indicated a wish to proceed with procedure and informed consent was signed.
PREOPERATIVE DIAGNOSIS:  Pleural Effusion
POSTOPERATIVE DIAGNOSIS: Same as preoperative diagnosis - see above.
PROCEDURE:  
76604 Ultrasound, chest (includes mediastinum), real time with image documentation
32557 Insert catheter pleura with imaging (chest tube)
Local ONLY
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
PROCEDURE IN DETAIL:
PATIENT POSITION: 
1‌ Supine  0‌ Sitting   
0‌ Lateral Decubitus:  0‌ Right 0‌ Left 
CHEST ULTRASOUND FINDINGS:  1‌ Image saved and printed 
Hemithorax:   1‌ Right  0‌ Left 
Pleural Effusion: 
Volume:       0‌ None  0‌ Minimal  0‌ Small  1‌ Moderate  0‌ Large 
Echogenicity:   1‌ Anechoic  0‌ Hypoechoic  0‌ Isoechoic  0‌ Hyperechoic 
Loculations:  0‌ None  1‌Thin  1‌ Thick 
Diaphragmatic Motion:  1‌ Normal  0‌ Diminished  0‌ Absent  
Lung: 
Lung sliding before procedure:  
 1‌ Present  0‌ Absent 
Lung sliding post procedure:   0‌ Present  0‌ Absent 
Lung consolidation/atelectasis: 1‌ Present  0‌  Absent 
Pleura:  0‌ Normal  1‌ Thick  0‌ Nodular 
Insertion site prepped and draped in sterile fashion.
ANESTHESIA:   Lidocaine 1%: ___8___ ml      Other: ______ 
Entry Site: 
1‌ Right __6th_ Intercostal Space   0‌ Left  ___ Intercostal Space 
0‌ Mid-clavicular   1‌ Mid-axillary  0‌ Mid-scapular  0‌ Other: 
Size:  0‌ 6Fr  0‌ 8Fr   0‌ 12FR   1‌ 14Fr  0‌ 16Fr   0‌ 18Fr  0‌ 24Fr   0‌ 32 Fr   0‌ Other: 
Sutured: 1‌ Yes 0‌ No 
PROCEDURE FINDINGS: 
A  pigtail catheter was inserted using the Seldinger technique.
Entry into the pleural space was confirmed with the easy removal of minimal serous appearing pleural fluid and air bubbles.
A guidewire was inserted using the introducer needle pointed in the apical posterior direction. The introducer needle was then removed.
A dilator was then inserted over the wire with a twisting motion in order to form a tract for catheter insertion.
The dilator was removed and the pigtail catheter (with trochar) was advanced over the guidewire.
The catheter was inserted into the chest until all catheter holes were well within the chest.
The guidewire and trochar were then removed.  The tube was then attached to the collection drain apparatus and secured in place with suture and covered.
Fluid Removed: __70___ ml 
0‌ Serous  1‌ Serosanguinous 0‌ Bloody  0‌ Chylous 0‌ Other: 
Pleural pressures:   If not measured, please check here 1‌
Drainage device:   1‌ Pleurovac    0‌ Drainage Bag  0‌ Heimlich Valve  0‌ Pneumostat  0‌ Other: 
Suction: 0‌ No 1‌Yes, - __20_ cmH20 
 
SPECIMEN(S): 
0‌None           1‌PH               1‌ LDH                
        1‌Glucose       1‌T.
Protein    1‌Cholesterol
1‌Cell Count   0‌ ADA             0‌Triglycerides            1‌Amylase 
1‌Gram Stain/ Culture            1‌AFB                         1‌Fungal Culture 
0‌Cytology      0‌Flow Cytometry                 
              0‌Other: 
CXR ordered: 1‌ Yes 0‌ No 
 
COMPLICATIONS:
1‌None 0‌Bleeding-EBL: ___ ml 0‌Pneumothorax 0‌Re- Expansion Pulmonary Edema 
0‌Other: 
IMPRESSION/PLAN: [REDACTED]is a 56 year old-year-old male who presents for Chest Ultrasound and Chest tube placement.
The patient tolerated the procedure well.  There were no immediate complications.
- f/u studies
- f/u CXR
- keep chest tube to suction -20cm H2O
- q8h flushing per RN flushing orders 
DISPOSITION: ICU"""

# =============================================================================
# CONFIGURATION
# =============================================================================

FLAGS_BRONCHOSCOPY = [
    "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
    "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy",
    "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration",
    "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
    "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation",
    "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy"
]

FLAGS_PLEURAL = [
    "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis",
    "pleural_biopsy", "fibrinolytic_therapy"
]

# Note: Ultrasound is an event but not in the binary flag list provided in instructions.
ACTIVE_FLAGS = {
    "chest_tube": 1
}

# =============================================================================
# SPAN ANNOTATIONS & EVENTS
# =============================================================================

# Events: 
# evt_01: Chest Ultrasound (Diagnostic)
# evt_02: Chest Tube Placement (Therapeutic)

SPANS = [
    # --- Event 1: Chest Ultrasound ---
    {
        "span_text": "Ultrasound, chest",
        "label": "PROC_METHOD",
        "normalized_value": "Ultrasound",
        "event_id": "evt_01",
        "context_prefix": "PROCEDURE:  \n76604 "
    },
    {
        "span_text": "Right",
        "label": "LATERALITY",
        "normalized_value": "Right",
        "event_id": "evt_01",
        "context_prefix": "Hemithorax:   1‌ "
    },
    {
        "span_text": "Pleural Effusion",
        "label": "OBS_LESION",
        "normalized_value": "Pleural Effusion",
        "event_id": "evt_01",
        "context_prefix": "Right  0‌ Left \n"
    },
    {
        "span_text": "Moderate",
        "label": "OBS_FINDING",
        "normalized_value": "Moderate Volume",
        "event_id": "evt_01",
        "context_prefix": "0‌ Minimal  0‌ Small  1‌ "
    },
    {
        "span_text": "Anechoic",
        "label": "OBS_FINDING",
        "normalized_value": "Anechoic",
        "event_id": "evt_01",
        "context_prefix": "Echogenicity:   1‌ "
    },
    {
        "span_text": "Thin",
        "label": "OBS_FINDING",
        "normalized_value": "Thin Loculations",
        "event_id": "evt_01",
        "context_prefix": "Loculations:  0‌ None  1‌"
    },
    
    # --- Event 2: Chest Tube ---
    {
        "span_text": "Insert catheter pleura",
        "label": "PROC_ACTION",
        "normalized_value": "Insert catheter",
        "event_id": "evt_02",
        "context_prefix": "32557 "
    },
    {
        "span_text": "chest tube",
        "label": "DEV_CATHETER",
        "normalized_value": "Chest Tube",
        "event_id": "evt_02",
        "context_prefix": "with imaging ("
    },
    {
        "span_text": "Right",
        "label": "LATERALITY",
        "normalized_value": "Right",
        "event_id": "evt_02",
        "context_prefix": "Entry Site: \n1‌ "
    },
    {
        "span_text": "__6th_ Intercostal Space",
        "label": "ANAT_LUNG_LOC",
        "normalized_value": "6th Intercostal Space",
        "event_id": "evt_02",
        "context_prefix": "1‌ Right "
    },
    {
        "span_text": "Mid-axillary",
        "label": "ANAT_LUNG_LOC",
        "normalized_value": "Mid-axillary",
        "event_id": "evt_02",
        "context_prefix": "0‌ Mid-clavicular   1‌ "
    },
    {
        "span_text": "14Fr",
        "label": "DEV_CATHETER_SIZE",
        "normalized_value": "14Fr",
        "event_id": "evt_02",
        "context_prefix": "0‌ 12FR   1‌ "
    },
    {
        "span_text": "pigtail catheter",
        "label": "DEV_CATHETER",
        "normalized_value": "Pigtail Catheter",
        "event_id": "evt_02",
        "context_prefix": "A  "
    },
    {
        "span_text": "Seldinger technique",
        "label": "PROC_METHOD",
        "normalized_value": "Seldinger Technique",
        "event_id": "evt_02",
        "context_prefix": "using the "
    },
    {
        "span_text": "__70___ ml",
        "label": "MEAS_VOL",
        "normalized_value": "70 ml",
        "event_id": "evt_02",
        "context_prefix": "Fluid Removed: "
    },
    {
        "span_text": "serous appearing pleural fluid",
        "label": "OBS_FINDING",
        "normalized_value": "Serous Fluid",
        "event_id": "evt_02",
        "context_prefix": "easy removal of minimal "
    },
    {
        "span_text": "- __20_ cmH20",
        "label": "MEAS_PRESS",
        "normalized_value": "-20 cmH2O",
        "event_id": "evt_02",
        "context_prefix": "1‌Yes, "
    },
    
    # --- Outcomes (Linked to evt_02) ---
    {
        "span_text": "tolerated the procedure well",
        "label": "OUTCOME_SYMPTOMS",
        "normalized_value": "Tolerated well",
        "event_id": "evt_02",
        "context_prefix": "The patient "
    },
    {
        "span_text": "no immediate complications",
        "label": "OUTCOME_COMPLICATION",
        "normalized_value": "None",
        "event_id": "evt_02",
        "context_prefix": "There were "
    }
]

EVENTS_DATA = [
    {
        "event_id": "evt_01",
        "type": "Ultrasound",
        "target_anatomy_type": "Pleura",
        "target_location": "Right",
        "findings": ["Pleural Effusion", "Moderate Volume", "Anechoic", "Thin Loculations"],
        "evidence_quote": "CHEST ULTRASOUND FINDINGS: ... Hemithorax: 1 Right ... Pleural Effusion: Moderate ... Anechoic ... Thin"
    },
    {
        "event_id": "evt_02",
        "type": "Chest Tube",
        "target_anatomy_type": "Pleura",
        "target_location": "Right 6th Intercostal Space Mid-axillary",
        "method": "Seldinger",
        "catheter_size": "14Fr",
        "outcomes_pleural": "70 ml drained",
        "outcomes_complication": "None",
        "outcomes_symptoms": "Tolerated well",
        "evidence_quote": "32557 Insert catheter pleura with imaging (chest tube)... Size: 1 14Fr... Fluid Removed: 70 ml... no immediate complications"
    }
]

# =============================================================================
# LOGIC
# =============================================================================

def create_workbook():
    if os.path.exists(TEMPLATE_PATH):
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    else:
        wb = openpyxl.Workbook()
        # Create default sheets if template missing (fallback)
        for sheet in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            if sheet not in wb.sheetnames:
                wb.create_sheet(sheet)
    return wb

def get_offset(full_text, span_text, context_prefix=None, match_index=0):
    # Escape regex special characters in span_text
    escaped_span = re.escape(span_text)
    
    # Find all matches
    matches = [m for m in re.finditer(escaped_span, full_text)]
    
    if not matches:
        return None, None, "not_found"
        
    # Logic 1: Unique match
    if len(matches) == 1:
        return matches[0].start(), matches[0].end(), "hydrated_unique"
        
    # Logic 2: Context match
    if context_prefix:
        # Normalize whitespace in context for loose matching
        clean_ctx = re.sub(r'\s+', ' ', context_prefix).strip()
        
        for m in matches:
            start = m.start()
            # Look at preceding 120 chars
            window = full_text[max(0, start-120):start]
            clean_window = re.sub(r'\s+', ' ', window)
            
            if clean_ctx in clean_window or context_prefix in window:
                return start, m.end(), "hydrated_prefix_window"
    
    # Logic 3: Match index
    if 0 <= match_index < len(matches):
        return matches[match_index].start(), matches[match_index].end(), "hydrated_match_index"
        
    return None, None, f"ambiguous_count={len(matches)}"

def run():
    wb = create_workbook()
    
    # 1. Note_Text
    ws_text = wb["Note_Text"]
    if ws_text.max_row == 1:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # 2. Note_Index
    ws_index = wb["Note_Index"]
    header = [
        "source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes"
    ] + FLAGS_BRONCHOSCOPY + FLAGS_PLEURAL
    
    # Ensure header exists if empty
    if ws_index.max_row == 1 and ws_index.cell(1,1).value is None:
        ws_index.append(header)
        
    row_data = [SOURCE_FILE, NOTE_ID, "", "", "", "", "Pending", ""]
    # Append flags
    for flag in FLAGS_BRONCHOSCOPY + FLAGS_PLEURAL:
        row_data.append(1 if flag in ACTIVE_FLAGS else 0)
    ws_index.append(row_data)
    
    # 3. Spans
    ws_span = wb["Span_Annotations"]
    ws_hydrated = wb["Span_Hydrated"]
    
    # Headers
    span_header = [
        "source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", 
        "match_index", "start_char", "end_char", "span_len", "label", "normalized_value", 
        "schema_field", "event_id", "is_negated", "is_historical", "time_anchor", "reviewer", 
        "comments", "hydration_status"
    ]
    
    if ws_span.max_row == 1 and ws_span.cell(1,1).value is None:
        ws_span.append(span_header)
    if ws_hydrated.max_row == 1 and ws_hydrated.cell(1,1).value is None:
        ws_hydrated.append(span_header)

    hydrated_rows = []
    
    for idx, span in enumerate(SPANS):
        span_id = f"span_{idx+1:03d}"
        
        # Calculate hydration
        start, end, status = get_offset(NOTE_TEXT, span["span_text"], span.get("context_prefix"), 0)
        
        # Base row data (Anchors)
        row = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure", span.get("context_prefix", ""), span["span_text"],
            0, "", "", f"=LEN(F{ws_span.max_row+1})", span["label"], span.get("normalized_value", ""),
            "", span.get("event_id", ""), "FALSE", "FALSE", "", "", "", "needs_hydration"
        ]
        ws_span.append(row)
        
        # Hydrated row
        h_row = row[:]
        h_row[7] = start if start is not None else ""
        h_row[8] = end if end is not None else ""
        h_row[9] = (end - start) if (start is not None and end is not None) else ""
        h_row[19] = status
        ws_hydrated.append(h_row)

    # 4. V3 Procedure Events
    ws_v3 = wb["V3_Procedure_Events"]
    v3_header = [
        "note_id", "event_id", "type", 
        "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station",
        "lesion.type", "lesion.size_mm", 
        "method", "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote",
        "stent.size", "stent.material_or_brand", "catheter.size_fr",
        "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post",
        "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"
    ]
    if ws_v3.max_row == 1 and ws_v3.cell(1,1).value is None:
        ws_v3.append(v3_header)

    registry_procedures = []

    for evt in EVENTS_DATA:
        # Build JSON row
        r_row = [
            NOTE_ID, evt["event_id"], evt["type"],
            evt.get("target_anatomy_type", ""), "", "", "", # Locations
            "", "", # Lesion
            evt.get("method", ""), "", "", "", json.dumps(evt.get("findings", [])), evt.get("evidence_quote", ""),
            "", "", evt.get("catheter_size", ""), # Devices
            "", "", # Airway Outcomes
            evt.get("outcomes_symptoms", ""), evt.get("outcomes_pleural", ""), evt.get("outcomes_complication", "")
        ]
        ws_v3.append(r_row)

        # Build Registry JSON Struct
        proc_obj = {
            "event_id": evt["event_id"],
            "procedure_type": evt["type"],
            "method": evt.get("method"),
            "target_anatomy": {
                "type": evt.get("target_anatomy_type"),
                "location": evt.get("target_location")
            },
            "outcomes": {
                "complications": evt.get("outcomes_complication"),
                "pleural_drainage": evt.get("outcomes_pleural")
            }
        }
        registry_procedures.append(proc_obj)

    # 5. Event Log (Legacy)
    ws_log = wb["Event_Log"]
    if ws_log.max_row == 1 and ws_log.cell(1,1).value is None:
        log_header = [
            "source_file", "note_id", "event_id", "event_type", "method", "anatomy_target",
            "device", "needle_gauge", "stations", "counts", "measurements", "specimens", "findings",
            "is_historical", "reviewer", "comments", "device_size", "device_material",
            "outcome_airway_lumen_pre", "outcome_airway_lumen_post", "outcome_symptoms",
            "outcome_pleural", "outcome_complication"
        ]
        ws_log.append(log_header)

    for evt in EVENTS_DATA:
        l_row = [
            SOURCE_FILE, NOTE_ID, evt["event_id"], evt["type"], evt.get("method", ""), evt.get("target_location", ""),
            "Catheter" if "Tube" in evt["type"] else "", "", "", "", "", "", json.dumps(evt.get("findings", [])),
            "FALSE", "", "", evt.get("catheter_size", ""), "",
            "", "", evt.get("outcomes_symptoms", ""), evt.get("outcomes_pleural", ""), evt.get("outcomes_complication", "")
        ]
        ws_log.append(l_row)

    # 6. JSON Dump
    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 1:
        ws_json.append(["schema_version", "note_id", "json_output", "no_immediate_complications"])
    
    final_json = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": registry_procedures,
        "no_immediate_complications": True
    }
    
    ws_json.append(["3.0", NOTE_ID, json.dumps(final_json, indent=2), True])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    run()