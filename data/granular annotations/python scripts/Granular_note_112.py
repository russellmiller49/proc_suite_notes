import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import json
import re
import os
from datetime import datetime

# ==========================================
# CONFIGURATION & INPUTS
# ==========================================
NOTE_ID = "note_112"
SOURCE_FILE = "note_112.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# Cleaned text based on the provided content (stripping source tags for realistic processing)
NOTE_TEXT = """NOTE_ID:  note_112 SOURCE_FILE: note_112.txt INDICATION FOR OPERATION:  [REDACTED]is a 61 year old-year-old male who presents with bronchial dehiscence.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
PROCEDURE:  
31615 Visualization of windpipe (Tracheobronchoscopy through established tracheostomy incision)
31645 Therapeutic aspiration initial episode
31624 Dx bronchoscope/lavage (BAL)    
50 Bilateral Procedures (Procedure done on both sides of the body)
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required a bilateral procedure today when this procedure would typically be unilateral.
Apply to: 31624 Dx bronchoscope/lavage (BAL)    .
ANESTHESIA: 
ICU drip of fentanyl plus midazolam pushes 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
ESTIMATED BLOOD LOSS:   None
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
Initial Airway Inspection Findings:
The endotracheal tube is in good position. The visualized portion of the trachea is of normal caliber.
The carina is sharp. The tracheobronchial tree was examined to at least the first subsegmental level.
Inspection shows a BMS that is patent in the RMSB, partially covering the RUL but with good aeration and mucus drainage through the stent.
Dehiscence at the RMSB remains entirely closed by stent. Areas of granulation tissue/ischemic mucosa remain at the distal BI/RML orifice as well as LMSB anastomosis and LUL.
There are two areas of very thin mucosa/airway wall in the LUL and Lingula.  Moderate thick secretions.
Successful therapeutic aspiration was performed to clean out the Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, and Carina from mucus.
Bronchial alveolar lavage was performed at Lateral-basal Segment of RLL (RB9).
Instilled 40 cc of NS, suction returned with 10 cc of NS.  Samples sent for Cell Count and Microbiology (Cultures/Viral/Fungal).
Bronchial alveolar lavage was performed at Anterior Segment of LUL (LB3).
Instilled 40 cc of NS, suction returned with 10 cc of NS.  Samples sent for Cell Count and Microbiology (Cultures/Viral/Fungal).
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
- RLL BAL
- LUL BAL
IMPRESSION/PLAN: [REDACTED]is a 61 year old-year-old male who presents for bronchoscopy for bronchial anastomosis dehiscence.
Stent in good position and patent, mucus cleared with suction. Dehiscence remains closed by stent.
Will need to pay attention to areas of thinning airway in the LUL anterior and lingula.
- f/u BAL studies
- bronch for stent inspection in 1 week"""

# ==========================================
# REGISTRY FLAGS (30)
# ==========================================
FLAGS = {
    # Bronchoscopy (23)
    "diagnostic_bronchoscopy": 1, # Inspection mentioned, dx codes
    "bal": 1, # Explicitly performed
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1, # CPT 31645
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0, # Inspection only, no placement
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural (7)
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# ==========================================
# ANCHOR-FIRST SPANS
# ==========================================
# Events:
# 1. Inspection (Stent Check)
# 2. Therapeutic Aspiration
# 3. BAL RLL
# 4. BAL LUL
# 5. Global Outcomes

SPANS = [
    # Event 1: Stent Inspection (BMS in RMSB)
    {
        "span_text": "Inspection shows a BMS",
        "label": "DEV_STENT",
        "normalized_value": "Stent (BMS)",
        "event_id": "evt_01",
        "comments": "Stent check"
    },
    {
        "span_text": "RMSB",
        "context_prefix": "that is patent in the ",
        "label": "ANAT_AIRWAY",
        "normalized_value": "RMSB",
        "event_id": "evt_01"
    },
    {
        "span_text": "good aeration and mucus drainage",
        "label": "OUTCOME_AIRWAY_LUMEN_POST", # Or Finding
        "normalized_value": "patent",
        "event_id": "evt_01"
    },
    {
        "span_text": "Dehiscence at the RMSB remains entirely closed",
        "label": "OBS_LESION",
        "normalized_value": "dehiscence closed",
        "event_id": "evt_01"
    },

    # Event 2: Therapeutic Aspiration
    {
        "span_text": "Successful therapeutic aspiration",
        "label": "PROC_METHOD",
        "normalized_value": "Therapeutic Aspiration",
        "event_id": "evt_02"
    },
    {
        "span_text": "Trachea (Distal 1/3)",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Trachea Distal",
        "event_id": "evt_02"
    },
    {
        "span_text": "Right Mainstem",
        "label": "ANAT_AIRWAY",
        "normalized_value": "RMS",
        "event_id": "evt_02"
    },
    {
        "span_text": "Bronchus Intermedius",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Bronchus Intermedius",
        "event_id": "evt_02"
    },
    {
        "span_text": "Left Mainstem",
        "label": "ANAT_AIRWAY",
        "normalized_value": "LMS",
        "event_id": "evt_02"
    },
    {
        "span_text": "Carina",
        "context_prefix": "Left Mainstem, and ",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Carina",
        "event_id": "evt_02"
    },
    {
        "span_text": "mucus",
        "context_prefix": "and Carina from ",
        "label": "OBS_LESION",
        "normalized_value": "mucus",
        "event_id": "evt_02"
    },

    # Event 3: BAL RLL
    {
        "span_text": "Bronchial alveolar lavage",
        "context_prefix": "from mucus.\n", # Preceded by aspiration paragraph
        "label": "PROC_METHOD",
        "normalized_value": "BAL",
        "event_id": "evt_03"
    },
    {
        "span_text": "Lateral-basal Segment of RLL",
        "label": "ANAT_LUNG_LOC",
        "normalized_value": "RLL Lateral-basal",
        "event_id": "evt_03"
    },
    {
        "span_text": "RB9",
        "label": "ANAT_LN_STATION", # Using station field for segment code
        "normalized_value": "RB9",
        "event_id": "evt_03"
    },
    {
        "span_text": "Instilled 40 cc",
        "context_prefix": "Segment of RLL (RB9).\n",
        "label": "MEAS_VOL",
        "normalized_value": "40",
        "event_id": "evt_03"
    },
    {
        "span_text": "suction returned with 10 cc",
        "context_prefix": "Instilled 40 cc of NS, ",
        "label": "MEAS_VOL",
        "normalized_value": "10",
        "event_id": "evt_03"
    },

    # Event 4: BAL LUL
    {
        "span_text": "Bronchial alveolar lavage",
        "context_prefix": "(Cultures/Viral/Fungal).\n", # Preceded by first BAL block
        "label": "PROC_METHOD",
        "normalized_value": "BAL",
        "event_id": "evt_04"
    },
    {
        "span_text": "Anterior Segment of LUL",
        "label": "ANAT_LUNG_LOC",
        "normalized_value": "LUL Anterior",
        "event_id": "evt_04"
    },
    {
        "span_text": "LB3",
        "label": "ANAT_LN_STATION",
        "normalized_value": "LB3",
        "event_id": "evt_04"
    },
    {
        "span_text": "Instilled 40 cc",
        "context_prefix": "Segment of LUL (LB3).\n",
        "label": "MEAS_VOL",
        "normalized_value": "40",
        "event_id": "evt_04"
    },
    {
        "span_text": "suction returned with 10 cc",
        "context_prefix": "Anterior Segment of LUL (LB3).\nInstilled 40 cc of NS, ",
        "label": "MEAS_VOL",
        "normalized_value": "10",
        "event_id": "evt_04"
    },

    # Findings / Complications
    {
        "span_text": "two areas of very thin mucosa/airway wall in the LUL and Lingula",
        "label": "OBS_LESION",
        "normalized_value": "mucosal thinning",
        "event_id": "evt_05"
    },
    {
        "span_text": "no immediate complications",
        "label": "OUTCOME_COMPLICATION",
        "normalized_value": "None",
        "is_negated": True,
        "event_id": "evt_global"
    }
]

# ==========================================
# HELPER FUNCTIONS
# ==========================================
def create_workbook():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        # Create a fallback workbook structure if template missing
        wb = openpyxl.Workbook()
        for sheet in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            wb.create_sheet(sheet)
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
    return wb

def hydrate_offsets(text, spans):
    hydrated_spans = []
    
    for span in spans:
        span_text = span["span_text"]
        context = span.get("context_prefix", "")
        match_idx = span.get("match_index", None)
        
        start = -1
        end = -1
        status = "ambiguous"
        
        # 1. Exact unique match
        if text.count(span_text) == 1:
            start = text.find(span_text)
            status = "hydrated_unique"
            
        # 2. Context match
        elif context and text.count(span_text) > 1:
            # Find all occurrences
            occurrences = [m.start() for m in re.finditer(re.escape(span_text), text)]
            for occ in occurrences:
                # Look back 120 chars for context
                window_start = max(0, occ - 120)
                preceding_text = text[window_start:occ]
                # Normalize newlines for fuzzy context matching
                if context.strip() in preceding_text.replace('\n', ' '):
                    start = occ
                    status = "hydrated_prefix_window"
                    break
            # Fallback if strict string check fails, try simpler check
            if start == -1:
                 for occ in occurrences:
                     if context[-10:] in text[max(0, occ-20):occ]:
                         start = occ
                         status = "hydrated_prefix_short"
                         break
                         
        # 3. Index match (manual override)
        elif match_idx is not None:
            occurrences = [m.start() for m in re.finditer(re.escape(span_text), text)]
            if 0 <= match_idx < len(occurrences):
                start = occurrences[match_idx]
                status = "hydrated_match_index"
        
        if start != -1:
            end = start + len(span_text)
            
        h_span = span.copy()
        h_span["start_char"] = start if start != -1 else ""
        h_span["end_char"] = end if end != -1 else ""
        h_span["hydration_status"] = status if start != -1 else f"ambiguous_count={text.count(span_text)}"
        h_span["span_len"] = len(span_text)
        hydrated_spans.append(h_span)
        
    return hydrated_spans

def populate_worksheets(wb, spans, hydrated_spans):
    # 1. Note_Text
    ws_text = wb["Note_Text"]
    # Check if empty, add header
    if ws_text.max_row == 1 and ws_text.cell(1,1).value is None:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index (Flags)
    ws_index = wb["Note_Index"]
    header = [
        "source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes",
        "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", "tbna_conventional",
        "linear_ebus", "radial_ebus", "navigational_bronchoscopy", "transbronchial_biopsy", "transbronchial_cryobiopsy",
        "therapeutic_aspiration", "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
        "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", "bronchial_thermoplasty",
        "whole_lung_lavage", "rigid_bronchoscopy",
        "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis", "pleural_biopsy", "fibrinolytic_therapy"
    ]
    
    # If header missing, add it
    if ws_index.max_row == 1 and ws_index.cell(1,1).value != "source_file":
        ws_index.append(header)
        
    row_data = [SOURCE_FILE, NOTE_ID, "", "", "", "Auto", "Complete", ""]
    for h in header[8:]:
        row_data.append(FLAGS.get(h, 0))
    ws_index.append(row_data)

    # 3. Span_Annotations (Anchor First)
    ws_anno = wb["Span_Annotations"]
    anno_header = ["source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", "match_index",
                   "start_char", "end_char", "span_len", "label", "normalized_value", "schema_field", "event_id",
                   "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"]
    
    if ws_anno.max_row == 1 and ws_anno.cell(1,1).value != "source_file":
        ws_anno.append(anno_header)

    for i, span in enumerate(spans):
        row = [
            SOURCE_FILE, NOTE_ID, f"span_{i+1:03}", "", span.get("context_prefix",""), span["span_text"], span.get("match_index",""),
            "", "", f"=LEN(F{ws_anno.max_row+1})", span["label"], span.get("normalized_value",""), "", span.get("event_id",""),
            span.get("is_negated", False), False, "", "Auto", span.get("comments",""), "needs_hydration"
        ]
        ws_anno.append(row)

    # 4. Span_Hydrated
    ws_hyd = wb["Span_Hydrated"]
    if ws_hyd.max_row == 1 and ws_hyd.cell(1,1).value != "source_file":
        ws_hyd.append(anno_header) # Same header

    for i, h_span in enumerate(hydrated_spans):
        row = [
            SOURCE_FILE, NOTE_ID, f"span_{i+1:03}", "", h_span.get("context_prefix",""), h_span["span_text"], h_span.get("match_index",""),
            h_span["start_char"], h_span["end_char"], h_span["span_len"], h_span["label"], h_span.get("normalized_value",""), "", h_span.get("event_id",""),
            h_span.get("is_negated", False), False, "", "Auto", h_span.get("comments",""), h_span["hydration_status"]
        ]
        ws_hyd.append(row)

    # 5. Event_Log
    ws_event = wb["Event_Log"]
    event_header = [
        "source_file", "note_id", "event_id", "event_type", "method", "anatomy_target", "device", "needle_gauge",
        "stations", "counts", "measurements", "specimens", "findings", "is_historical", "reviewer", "comments",
        "device_size", "device_material", "outcome_airway_lumen_pre", "outcome_airway_lumen_post",
        "outcome_symptoms", "outcome_pleural", "outcome_complication"
    ]
    if ws_event.max_row == 1 and ws_event.cell(1,1).value != "source_file":
        ws_event.append(event_header)
        
    # Summarize events
    events = {}
    for s in spans:
        eid = s.get("event_id")
        if not eid: continue
        if eid not in events: events[eid] = {"method": [], "anatomy": [], "device": [], "findings": [], "outcomes": [], "measurements": []}
        
        if s["label"] == "PROC_METHOD": events[eid]["method"].append(s["normalized_value"])
        if s["label"] in ["ANAT_AIRWAY", "ANAT_LUNG_LOC"]: events[eid]["anatomy"].append(s["normalized_value"])
        if s["label"] == "DEV_STENT": events[eid]["device"].append(s["normalized_value"])
        if s["label"] == "OBS_LESION": events[eid]["findings"].append(s["normalized_value"])
        if s["label"].startswith("OUTCOME"): events[eid]["outcomes"].append(f"{s['label']}={s['normalized_value']}")
        if s["label"] == "MEAS_VOL": events[eid]["measurements"].append(f"{s['normalized_value']}cc")

    for eid, data in events.items():
        row = [
            SOURCE_FILE, NOTE_ID, eid, "Procedure", 
            ", ".join(set(data["method"])), 
            ", ".join(set(data["anatomy"])),
            ", ".join(set(data["device"])),
            "", "", "", 
            ", ".join(set(data["measurements"])),
            "",
            ", ".join(set(data["findings"])),
            False, "Auto", "", "", "",
            "", # lumen pre
            "patent" if "OUTCOME_AIRWAY_LUMEN_POST=patent" in data["outcomes"] else "",
            "", "",
            "None" if "OUTCOME_COMPLICATION=None" in data["outcomes"] else ""
        ]
        ws_event.append(row)

    # 6. V3_Procedure_Events & 7. JSON
    # Construct V3 object
    v3_events = []
    
    # Event 1: Stent Check
    v3_events.append({
        "event_id": "evt_01",
        "type": "Diagnostic",
        "target": {"anatomy_type": "Airway", "location": {"lobe": "RMSB"}},
        "method": "Inspection",
        "devices_json": json.dumps({"stent": "BMS"}),
        "findings_json": json.dumps(["Dehiscence closed"]),
        "outcomes": {"airway": {"lumen_post": "patent"}}
    })
    
    # Event 2: Aspiration
    v3_events.append({
        "event_id": "evt_02",
        "type": "Therapeutic",
        "method": "Therapeutic Aspiration",
        "target": {"anatomy_type": "Airway", "location": {"lobe": "Multiple (Trachea, RMS, BI, LMS, Carina)"}},
        "findings_json": json.dumps(["Mucus"]),
        "action": "Suction"
    })
    
    # Event 3: BAL RLL
    v3_events.append({
        "event_id": "evt_03",
        "type": "Diagnostic",
        "method": "BAL",
        "target": {"anatomy_type": "Lung", "location": {"segment": "RB9 (Lat-Basal RLL)"}},
        "measurements_json": json.dumps({"instilled": 40, "return": 10})
    })

    # Event 4: BAL LUL
    v3_events.append({
        "event_id": "evt_04",
        "type": "Diagnostic",
        "method": "BAL",
        "target": {"anatomy_type": "Lung", "location": {"segment": "LB3 (Ant LUL)"}},
        "measurements_json": json.dumps({"instilled": 40, "return": 10})
    })

    # Populate V3 Sheet
    ws_v3 = wb["V3_Procedure_Events"]
    v3_header = ["note_id", "event_id", "type", "target.anatomy_type", "target.location.lobe", "target.location.segment", 
                 "target.station", "lesion.type", "lesion.size_mm", "method", "devices_json", "measurements_json", 
                 "specimens_json", "findings_json", "evidence_quote", "stent.size", "stent.material_or_brand", 
                 "catheter.size_fr", "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post", "outcomes.symptoms", 
                 "outcomes.pleural", "outcomes.complications"]
    
    if ws_v3.max_row == 1 and ws_v3.cell(1,1).value != "note_id":
        ws_v3.append(v3_header)

    for evt in v3_events:
        tgt = evt.get("target", {})
        loc = tgt.get("location", {})
        out = evt.get("outcomes", {})
        out_aw = out.get("airway", {})
        
        row = [
            NOTE_ID, evt["event_id"], evt["type"], tgt.get("anatomy_type"), loc.get("lobe"), loc.get("segment"),
            "", "", "", evt.get("method"), evt.get("devices_json"), evt.get("measurements_json"),
            "", evt.get("findings_json"), "", "", "", "", "", 
            out_aw.get("lumen_post"), "", "", ""
        ]
        ws_v3.append(row)

    # Populate JSON Sheet
    ws_json = wb["V3_Registry_JSON"]
    ws_json.cell(1,1, "json_output")
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": v3_events,
        "no_immediate_complications": True
    }
    ws_json.cell(2,1, json.dumps(registry_data, indent=2))

# ==========================================
# MAIN EXECUTION
# ==========================================
if __name__ == "__main__":
    wb = create_workbook()
    hydrated_spans = hydrate_offsets(NOTE_TEXT, SPANS)
    populate_worksheets(wb, SPANS, hydrated_spans)
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")