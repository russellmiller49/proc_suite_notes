import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# =============================================================================
# INPUT DATA
# =============================================================================

NOTE_ID = "note_113"
SOURCE_FILE = "note_113.txt"
PROCEDURE_DATE = "2026-01-12" # Simulated date based on context or leave blank
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_113 SOURCE_FILE: note_113.txt INDICATION FOR OPERATION:  [REDACTED]is a 56 year old-year-old male who presents with airway dehiscence and erosion in bilateral donor transplanted lungs.
CONSENT : Obtained before the procedure. The nature, indications, purpose, benefits, risks, potential complications, and alternatives to bronchoscopy were discussed with the patient or surrogate decision-maker in detail.
Patient or surrogate deci
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
PROCEDURE:  
31502 Tracheotomy tube change prior to establishment of fistula tract
31899 Unlisted Procedure (Trach Change with Mature Tract or Procedure NOS)
31646 Therapeutic aspiration subsequent episodes
31622 Dx bronchoscope/cell washing          
31630 Balloon dilation
31636 Dilate and bronchial stent initial bronchus
Mod 22. Pt with full thickness right lung anastomotic dehiscence that requiring increased time, intensity due to clinical condition, severity of dehiscence.
> 70% effort was required for this procedure due to time, clinical condition and intensity of the procedure.
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Rigid Bronchoscope
Rigid Tracheoscope
Flexible Therapeutic Bronchoscope
Flexible Hybrid (Pedatric) Bronchoscope
PROCEDURE IN DETAIL:
A timeout was performed (confirming the patient's name, procedure type, and procedure location).
Anesthesia initiated.  Endotracheal lidocaine applied through tracheostomy tube.
A combination of Flexible Therapeutic Bronchoscope and Flexible Hybrid (Pedatric) Bronchoscope were sequentially advanced through patient's tracheostomy tube for the initial airway examination.
Topical lidocaine applied to main carina, RMS, and LMS.
Initial Airway Inspection Findings:
The flexible bronchoscope was introduced into the existing tracheostomy tube after induction of anesthesia.
( Upper airway was examined after lower airway inspection)
Distal trachea appears normal.
Native proximal RMS bronchus and LMS bronchus appear largely normal.
Right lung:
Exuberant amount of necrotic debris extending from the RMS anastomosis site to RC1, BI, RML, and RB6.
RMS anastomosis sutures visible underlying necrotic debris.  Multiple areas of exposed cartilage noted.
At the RMS anastomotic site, there was a large area of full-thickness dehiscence at 7 o'clock, exposing the mediastinum.
Did not attempt to traverse defect.  Area of full-thickness airway erosion along RUL bronchus anterior wall - directly adjacent to RB3 take-off.
Bronchus intermedius cartilage appears fractured with excessive "accordion-like" movement of the airway with respiration.
RC2 and RML bronchus mucosa very pale.  Right truncus basalis and RLL basilar segmental airways appear somewhat healthier.
Left lung:
Exuberant amount of heaping necrotic debris extending from the LMS anastomosis site to LC2, LUL/lingula/LC1, and proximal LLL.
LMS anastomosis sutures visible underlying necrotic debris.  Donor distal LMS bronchus cartilage and LUL bronchus cartilage appear fractured with excessive "accordion-like" movement of the airway with respiration.
Large full-thickness erosion along anterior aspect of LUL bronchus completely exposing a large vessel (presumed branch of PA).
Left truncus basalis and LLL basilar segmental airways appear somewhat healthier.
Successful therapeutic aspiration was performed to clean out the distal RMS, RUL, BI, RML, RLL, distal LMS, LUL/lingula, LLL from moderate, non-obstructing mucus.
Flexible Therapeutic Bronchoscope and Flexible Hybrid (Pedatric) Bronchoscope were withdrawn.  
A tooth or gum protector was placed.
The black rigid non-ventilating tracheoscope barrel was introduced through the mouth and advanced in the midline while keeping the alignment with the axis of the trachea and minimizing pressure to the teeth.
The vocal cords were identified and the rigid tracheoscope was advanced carefully while minimizing contact with them.
The black rigid tracheoscope was initially positioned at the subglottic space.
The tracheostomy tube cuff was deflated and trach tube ventilation stopped.
Jet ventilation was initiated through the rigid barrel and chest wall movement was confirmed.
There was a large amount of secretion build-up above the tracheostomy cuff, which was suctioned to clear.
Tracheostomy tube was removed.  Black rigid tracheoscope was advanced into the distal trachea.
However, the tracheoscope was not long enough to intubate the RMS bronchus take-off.
Therefore, the black rigid tracheoscope was withdrawn to the subglottic space.
The tracheostomy tube was inserted again under direct visualization and secured in place.  Mechanical ventilation initiated through trachoestomy tube again.
Jet ventilation was stopped and the black rigid tracheoscope was extubated with ease.
The black rigid ventilating bronchoscope barrel was introduced through the mouth and advanced in the midline while keeping the alignment with the axis of the trachea and minimizing pressure to the teeth.
The vocal cords were identified and the rigid bronchoscope was advanced carefully while minimizing contact with them.
The black rigid bronchoscope was initially positioned at the subglottic space.
The tracheostomy tube cuff was deflated and trach tube ventilation stopped.
Jet ventilation was initiated through the rigid barrel and chest wall movement was confirmed.
Tracheostomy tube was removed.
Black rigid bronchoscope was advanced into the distal trachea and then the RMS bronchus take-off.
The Flexible Therapeutic Bronchoscope and Flexible Hybrid (Pedatric) Bronchoscope were advanced through the rigid bronchoscope barrel to complete the remaining portions of the procedure.
RC2 to MC (RMS + BI):  5.3, 5.2 cm
RMS bronchus :  3.7 cm
RUL bronchus:  2 cm
BI:  2 cm
The following stent (14 x 40 UltraFlex uncovered SEMS - proximal deploying) was advanced through the rigid bronchoscope barrel under direct bronchoscopic visualization and placed in the Right Mainstem bronchus.
During deployment of the sent the patient's oxygen saturation dropped to 88% and recovered with bag ventilation once the rigid was withdrawn.
Balloon dilation was performed within the RMS bronchus uncovered stent.
12/13.5/15 Elation balloon was used to perform dilation to 12, 13.5, and 15mm mm within the RMS bronchus uncovered stent  Total 3 inflations with dilation time of 10-15 seconds each.
The stent achieved complete closure of the right anastomosis dehiscence but without excessive radial pressure on the airway.
This closure was visualized during bag ventilation with tidal volumes of 500-600 mL.
Residual secretions were suctioned to clear.
The black rigid bronchoscope was withdrawn to the subglottic space.
A new tracheostomy tube was inserted under direct visualization to replace the previously removed tracheostomy tube.
Tracheostomy tube was changed from:
Shiley 6 cuffed Trach ISO/ID size 7.5 mm To  Shiley 6 cuffed Trach ISO/ID size 7.5 mm
without issue.
Tracheostomy tube was secured in place.
Mechanical ventilation initiated through trachoestomy tube again.
Jet ventilation was stopped and the black rigid bronchoscope was extubated with ease.
There was no evidence or oral dental injury
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient remained on tracheostomy tube mechanical ventilation in the operating room and transported back to the ICU in stable condition.
ESTIMATED BLOOD LOSS:   None
COMPLICATIONS:    None
SPECIMEN(S): 
None
IMPRESSION/PLAN: [REDACTED]is a 56 year old-year-old male who presents for bronchoscopy for airway evaluation.
-Profound bilateral donor transplant lung proximal airway injury with multiple sites of full-thickness dehiscence and airway erosion exposing underlying mediastinal/hilar structures.
-Successful placement of 14 x 40 UltraFlex uncovered metallic stent in the RMS bronchus, covering and closing the area of RMS anastomosis dehiscence.
-Patient to return to ICU
-Obtain post-procedure CXR
- pt does not require stent hydration for uncovered stent
- brother was updated at bedside after the procedure
-Patient will need routine stent checks - will plan for at least weekly.
At the first sign of granulation tissue build-up, we will plan to replace the bare metallic stent to avoid irrecoverable incorporation of the stent into the airway.
"""

# =============================================================================
# HYDRATION LOGIC
# =============================================================================

class Hydrator:
    def __init__(self, full_text):
        self.full_text = full_text

    def hydrate(self, span_text, context_prefix=None, match_index=0):
        """
        Returns (start, end, hydration_status)
        """
        # 1. Exact count check
        matches = [m.start() for m in re.finditer(re.escape(span_text), self.full_text)]
        
        if not matches:
            return None, None, "not_found"

        # Unique match
        if len(matches) == 1:
            return matches[0], matches[0] + len(span_text), "hydrated_unique"

        # Context match
        if context_prefix:
            # Look for context in 120 chars before
            for start_idx in matches:
                window_start = max(0, start_idx - 120)
                window = self.full_text[window_start:start_idx]
                if context_prefix in window:
                    return start_idx, start_idx + len(span_text), "hydrated_prefix_window"
        
        # Match Index fallback
        if 0 <= match_index < len(matches):
            selected = matches[match_index]
            return selected, selected + len(span_text), "hydrated_match_index"
        
        return None, None, f"ambiguous_count={len(matches)}"

# =============================================================================
# DEFINITIONS
# =============================================================================

# Procedure Flags (0 or 1)
PROCEDURE_FLAGS = {
    # Bronchoscopy (23)
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 0,
    "airway_dilation": 1,
    "airway_stent": 1,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 1,
    
    # Pleural (7)
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# Spans to Extract
# Struct: (span_text, label, normalized_value, event_id, context_prefix)
SPANS = [
    # Event 1: Rigid Access
    ("Rigid Bronchoscope", "PROC_METHOD", "Rigid Bronchoscopy", "evt_1", None),
    ("black rigid ventilating bronchoscope barrel", "DEV_INSTRUMENT", "Rigid Scope", "evt_1", None),
    ("Jet ventilation", "PROC_METHOD", "Jet Ventilation", "evt_1", "tube ventilation stopped"),
    
    # Event 2: Therapeutic Aspiration (Suctioning)
    ("Successful therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", "evt_2", None),
    ("clean out the distal RMS, RUL, BI, RML, RLL, distal LMS", "PROC_ACTION", "Clean Airway", "evt_2", None),
    ("moderate, non-obstructing mucus", "OBS_LESION", "Mucus", "evt_2", None),

    # Event 3: Findings (Right Dehiscence)
    ("RMS anastomosis site", "ANAT_AIRWAY", "RMS Anastomosis", "evt_3", "Exuberant amount"),
    ("full-thickness dehiscence", "OBS_LESION", "Dehiscence", "evt_3", "large area of"),
    ("exposing the mediastinum", "OBS_LESION", "Exposed Mediastinum", "evt_3", None),
    ("exuberant amount of necrotic debris", "OBS_LESION", "Necrosis", "evt_3", "Right lung"),

    # Event 4: Findings (Left Dehiscence/Erosion)
    ("LMS anastomosis", "ANAT_AIRWAY", "LMS Anastomosis", "evt_4", "Left lung"),
    ("Large full-thickness erosion", "OBS_LESION", "Erosion", "evt_4", None),
    ("LUL bronchus", "ANAT_AIRWAY", "LUL", "evt_4", "anterior aspect of"),
    ("exposing a large vessel", "OBS_LESION", "Exposed Vessel", "evt_4", None),

    # Event 5: Stent Placement (RMS)
    ("The following stent", "PROC_METHOD", "Stent Placement", "evt_5", None),
    ("14 x 40 UltraFlex uncovered SEMS", "DEV_STENT", "UltraFlex Uncovered SEMS", "evt_5", None),
    ("14 x 40", "DEV_STENT_SIZE", "14x40mm", "evt_5", "The following stent"),
    ("uncovered SEMS", "DEV_STENT_MATERIAL", "Metal (Uncovered)", "evt_5", "UltraFlex"),
    ("Right Mainstem bronchus", "ANAT_AIRWAY", "RMS", "evt_5", "placed in the"),
    ("stent achieved complete closure", "OUTCOME_AIRWAY_LUMEN_POST", "Complete Closure", "evt_5", None),

    # Event 6: Balloon Dilation
    ("Balloon dilation", "PROC_METHOD", "Balloon Dilation", "evt_6", None),
    ("RMS bronchus", "ANAT_AIRWAY", "RMS", "evt_6", "within the"),
    ("12/13.5/15 Elation balloon", "DEV_INSTRUMENT", "Elation Balloon", "evt_6", None),
    ("dilation to 12, 13.5, and 15mm", "MEAS_AIRWAY_DIAM", "15mm", "evt_6", None),
    
    # Event 7: Complications/General
    ("oxygen saturation dropped to 88%", "OUTCOME_COMPLICATION", "Desaturation (Transient)", "evt_7", None),
    ("There were no immediate complications", "OUTCOME_COMPLICATION", "None", "evt_7", None)
]

# Event Log Definitions
EVENT_LOG_ROWS = [
    {
        "event_id": "evt_1",
        "event_type": "Rigid Bronchoscopy",
        "method": "Rigid Bronchoscopy",
        "device": "Rigid Scope, Jet Ventilation",
        "findings": "Vocal cords identified, trachea entered"
    },
    {
        "event_id": "evt_2",
        "event_type": "Therapeutic Aspiration",
        "method": "Suction",
        "anatomy_target": "Bilateral Airways",
        "findings": "Moderate non-obstructing mucus cleared"
    },
    {
        "event_id": "evt_3",
        "event_type": "Observation",
        "findings": "RMS Anastomosis Full-thickness dehiscence, necrosis, exposed mediastinum"
    },
    {
        "event_id": "evt_4",
        "event_type": "Observation",
        "findings": "LMS/LUL erosion, exposed PA vessel, fractured cartilage"
    },
    {
        "event_id": "evt_5",
        "event_type": "Airway Stent",
        "method": "Stent Placement",
        "anatomy_target": "RMS",
        "device": "14x40 UltraFlex Uncovered SEMS",
        "device_size": "14x40mm",
        "device_material": "Uncovered Metal",
        "outcome_airway_lumen_post": "Complete closure of dehiscence",
        "outcome_complication": "Transient desaturation (recovered)"
    },
    {
        "event_id": "evt_6",
        "event_type": "Airway Dilation",
        "method": "Balloon Dilation",
        "anatomy_target": "RMS",
        "device": "12/13.5/15 Elation Balloon",
        "measurements": "Dilated to 15mm"
    },
    {
        "event_id": "evt_7",
        "event_type": "Outcomes",
        "outcome_complication": "None immediate"
    }
]

# =============================================================================
# PROCESSING
# =============================================================================

def create_workbook():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        # Fallback if template missing (should not happen per prompt, but good for safety)
        wb = openpyxl.Workbook()
    
    # Ensure sheets exist
    required_sheets = [
        "Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", 
        "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"
    ]
    
    for name in required_sheets:
        if name not in wb.sheetnames:
            wb.create_sheet(name)
            
    return wb

def main():
    wb = create_workbook()
    hydrator = Hydrator(NOTE_TEXT)
    
    # 1. Note_Text
    ws_text = wb["Note_Text"]
    # Append if empty or new
    if ws_text.max_row == 1:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index (Flags)
    ws_index = wb["Note_Index"]
    header = [
        "source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes",
        "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", "tbna_conventional",
        "linear_ebus", "radial_ebus", "navigational_bronchoscopy", "transbronchial_biopsy", "transbronchial_cryobiopsy",
        "therapeutic_aspiration", "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
        "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", "bronchial_thermoplasty",
        "whole_lung_lavage", "rigid_bronchoscopy",
        "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis", "pleural_biopsy", "fibrinolytic_therapy"
    ]
    
    if ws_index.max_row == 1:
        ws_index.append(header)
        
    row_data = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "Auto", "Complete", ""]
    # Append flags in order
    flag_keys = header[8:]
    for k in flag_keys:
        row_data.append(PROCEDURE_FLAGS.get(k, 0))
    ws_index.append(row_data)

    # 3 & 4. Spans (Annotations & Hydrated)
    ws_anno = wb["Span_Annotations"]
    ws_hydra = wb["Span_Hydrated"]
    
    span_header = [
        "source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", 
        "match_index", "start_char", "end_char", "span_len", "label", "normalized_value", 
        "schema_field", "event_id", "is_negated", "is_historical", "time_anchor", 
        "reviewer", "comments", "hydration_status"
    ]
    
    if ws_anno.max_row == 1: ws_anno.append(span_header)
    if ws_hydra.max_row == 1: ws_hydra.append(span_header)

    hydrated_spans = []

    for i, (txt, label, norm, evt, ctx) in enumerate(SPANS):
        span_id = f"span_{i+1:03d}"
        start, end, status = hydrator.hydrate(txt, context_prefix=ctx)
        
        # Base Row
        row = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure", ctx, txt, 
            0, None, None, len(txt), label, norm, 
            "", evt, 0, 0, 0, "Auto", "", "needs_hydration"
        ]
        ws_anno.append(row)
        
        # Hydrated Row
        row_h = list(row)
        row_h[7] = start
        row_h[8] = end
        row_h[19] = status
        ws_hydra.append(row_h)
        
        if start is not None:
            hydrated_spans.append({
                "span_text": txt, "label": label, "value": norm, "event_id": evt
            })

    # 5. Event_Log
    ws_event = wb["Event_Log"]
    evt_header = [
        "source_file", "note_id", "event_id", "event_type", "method", "anatomy_target", 
        "device", "needle_gauge", "stations", "counts", "measurements", "specimens", 
        "findings", "is_historical", "reviewer", "comments", "device_size", "device_material",
        "outcome_airway_lumen_pre", "outcome_airway_lumen_post", "outcome_symptoms", 
        "outcome_pleural", "outcome_complication"
    ]
    if ws_event.max_row == 1: ws_event.append(evt_header)

    for evt in EVENT_LOG_ROWS:
        row = [
            SOURCE_FILE, NOTE_ID, evt.get("event_id"), evt.get("event_type"), evt.get("method"), evt.get("anatomy_target"),
            evt.get("device"), "", "", "", evt.get("measurements"), "",
            evt.get("findings"), 0, "Auto", "", evt.get("device_size"), evt.get("device_material"),
            "", evt.get("outcome_airway_lumen_post"), "", "", evt.get("outcome_complication")
        ]
        ws_event.append(row)

    # 6. V3_Procedure_Events (Structured)
    ws_v3 = wb["V3_Procedure_Events"]
    v3_header = [
        "note_id", "event_id", "type", 
        "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station",
        "lesion.type", "lesion.size_mm", 
        "method", "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote",
        "stent.size", "stent.material_or_brand", "catheter.size_fr",
        "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post", 
        "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"
    ]
    if ws_v3.max_row == 1: ws_v3.append(v3_header)

    # Simple mapping for V3 (Subset)
    for evt in EVENT_LOG_ROWS:
        row = [
            NOTE_ID, evt.get("event_id"), evt.get("event_type"),
            "airway" if evt.get("anatomy_target") else "", "", "", "",
            "", "",
            evt.get("method"), json.dumps([evt.get("device")]) if evt.get("device") else "", "", "", json.dumps([evt.get("findings")]), "",
            evt.get("device_size"), evt.get("device_material"), "",
            "", evt.get("outcome_airway_lumen_post"), 
            "", "", evt.get("outcome_complication")
        ]
        ws_v3.append(row)

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 1: ws_json.append(["schema_version", "note_id", "json_output"])
    
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": EVENT_LOG_ROWS,
        "no_immediate_complications": True
    }
    ws_json.append(["3.0", NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Successfully generated {OUTPUT_PATH}")

if __name__ == "__main__":
    main()