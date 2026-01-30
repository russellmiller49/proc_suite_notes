import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import datetime
import os

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_114"
SOURCE_FILE = "note_114.txt"
PROCEDURE_DATE = "2024-09-06" # Inferred from "likely bare metallic stent placement on Fri, 9/6/2024" (future) -> Note likely just before or same day, but let's leave blank if uncertain or use valid format. Given context, I'll leave blank to be safe, or use current date if needed. Let's leave blank in metadata.
NOTE_TEXT = """NOTE_ID:  note_114 SOURCE_FILE: note_114.txt INDICATION FOR OPERATION:  [REDACTED]is a 56 year old-year-old male who presents with airway dehiscence and erosion in bilateral donor transplanted lungs.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
PROCEDURE:  
31899 Unlisted Procedure (Trach Change with Mature Tract or Procedure NOS)
31646 Therapeutic aspiration subsequent episodes
31622 Dx bronchoscope/cell washing          
Mod 22 - pt with increased time, intensity due to clinical condition.
Pt on mechanical ventilation with full thickness dehiscence at right anastomosis and multiple areas of ischemia and dehiscence distally that required 40% increased effort.
ANESTHESIA: 
99152 Moderate sedation: initial 15 minutes
99153 Moderate sedation: each additional 15 minutes 
Procedure performed under moderate sedation.
The following medications were provided:
Versed             2 mg
Fentanyl          50 mcg
Physician/patient face-to-face anesthesia start time:   17:33
Physician/patient face-to-face anesthesia stop time:   18:05
Total moderate sedation time was 33 minutes.
Patient was monitored continuously one-to-one throughout the entire procedure by the attending physician while anesthesia was administered.
Sedation was administered by Rachel Uslan, RN. 
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
ESTIMATED BLOOD LOSS:   None
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
A timeout was performed (confirming the patient's name, procedure type, and procedure location).
Sedation given, as described above.  Endotracheal lidocaine applied through tracheostomy tube.
Disposable bronchoscope (regular size) was advanced through patient's tracheostomy tube for the procedure.
Topical lidocaine applied to main carina, RMS, and LMS.
Initial Airway Inspection Findings:
Distal trachea appears normal.
Native proximal RMS bronchus and LMS bronchus appear largely normal.
Right lung:
Exuberant amount of necrotic debris extending from the RMS anastomosis site to RC1, BI, RML, and RB6.
RMS anastomosis sutures visible underlying necrotic debris.  Multiple areas of exposed cartilage noted.
At the RMS anastomotic site, there was a large area of full-thickness dehiscence at 7 o'clock, exposing the mediastinum.
Did not attempt to tranverse defect.  Area of full-thickness airway erosion along RUL bronchus anterior wall - directly adjacent to RB3 take-off.
Bronchus intermedius cartilage appears fractured with excessive "accordion-like" movement of the airway with respiration.
RC2 and RML bronchus mucosa very pale.  Right truncus basalis and RLL basilar segmental airways appear somewhat healthier.
Left lung:
Exuberant amount of heaping necrotic debris extending from the LMS anastomosis site to LC2, LUL/lingula/LC1, and proximal LLL.
LMS anastomosis sutures visible underlying necrotic debris.  Donor distal LMS bronchus cartilage and LUL bronchus cartilage appear fractured with excessive "accordion-like" movement of the airway with respiration.
Large full-thickness erosion along anterior aspect of LUL bronchus completely exposing a large vessel (presumed branch of PA).
Left truncus basalis and LLL basilar segmental airways appear somewhat healthier.
Successful therapeutic aspiration was performed to clean out the distal RMS, RUL, BI, RML, RLL, distal LMS, LUL/lingula, LLL from moderate, non-obstructing mucus.
Airway measurements taken in preparation of likely stent placement:
RMS bronchus :  3.5 cm
RUL bronchus:  2 cm
BI:  2 cm
RC2 to MC (RMS + BI):  5.5, 5.2 cm
Bronchoscope removed and procedure completed.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient remained in the ICU on trach-mechanical ventilation.
The patient was in stable condition.
SPECIMEN(S): 
None
IMPRESSION/PLAN: [REDACTED]is a 56 year old-year-old male who presents for bronchoscopy for airway evaluation.
-Profound bilateral donor transplant lung proximal airway injury with multiple sites of full-thickness dehiscence and airway erosion exposing underlying mediastinal/hilar structures.
-Will prepare for likely bare metallic stent placement on Fri, 9/6/2024.
Based on measurements above, will look to have the following bare metallic stents on-hand:  14x40, 14x60, 16x40, 16x60, 14x20, 16x20."""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# -------------------------------------------------------------------------
# CONFIGURATION
# -------------------------------------------------------------------------

PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# -------------------------------------------------------------------------
# SPAN DATA GENERATION
# -------------------------------------------------------------------------

# Helper to define span
def s(text, label, norm, field, event_id, ctx=None, neg=False, hist=False, time=False):
    return {
        "text": text,
        "label": label,
        "norm": norm,
        "field": field,
        "event_id": event_id,
        "context_prefix": ctx,
        "is_negated": "TRUE" if neg else "FALSE",
        "is_historical": "TRUE" if hist else "FALSE",
        "time_anchor": "TRUE" if time else "FALSE"
    }

# Defining events:
# Ev1: General / Instruments
# Ev2: Right Lung Findings (RMS, BI, Dehiscence)
# Ev3: Left Lung Findings (LMS, Erosion)
# Ev4: Therapeutic Aspiration (Action)
# Ev5: Measurements
# Ev6: Outcomes/Complications

spans_data = [
    # Event 1: Instruments / Procedure
    s("Flexible Therapeutic Bronchoscope", "DEV_INSTRUMENT", "Flexible Bronchoscope", "device", "ev1"),
    s("Moderate sedation", "PROC_METHOD", "Moderate Sedation", "method", "ev1"),
    s("tracheostomy tube", "ANAT_AIRWAY", "Trachea", "target.anatomy_type", "ev1"),
    s("Disposable bronchoscope", "DEV_INSTRUMENT", "Disposable Bronchoscope", "device", "ev1"),

    # Event 2: Right Lung Findings
    s("RMS", "ANAT_AIRWAY", "RMS", "target.location.lobe", "ev2", ctx="applied to main carina"),
    s("necrotic debris", "OBS_LESION", "Necrosis", "findings_json", "ev2", ctx="Right lung:\nExuberant amount of"),
    s("RMS anastomosis", "ANAT_AIRWAY", "RMS Anastomosis", "target.location.lobe", "ev2", ctx="extending from the"),
    s("RC1", "ANAT_AIRWAY", "RC1", "target.location.segment", "ev2"),
    s("BI", "ANAT_AIRWAY", "Bronchus Intermedius", "target.location.lobe", "ev2", ctx="RC1,"),
    s("RML", "ANAT_AIRWAY", "RML", "target.location.lobe", "ev2", ctx="BI,"),
    s("RB6", "ANAT_AIRWAY", "RB6", "target.location.segment", "ev2"),
    s("exposed cartilage", "OBS_LESION", "Exposed Cartilage", "findings_json", "ev2"),
    s("full-thickness dehiscence", "OBS_LESION", "Dehiscence", "findings_json", "ev2", ctx="large area of"),
    s("Bronchus intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "target.location.lobe", "ev2", ctx="cartilage appears fractured"),
    s("cartilage appears fractured", "OBS_LESION", "Fractured Cartilage", "findings_json", "ev2"),
    
    # Event 3: Left Lung Findings
    s("LMS", "ANAT_AIRWAY", "LMS", "target.location.lobe", "ev3", ctx="RMS, and"),
    s("Left lung", "LATERALITY", "Left", "target.location.lobe", "ev3"),
    s("necrotic debris", "OBS_LESION", "Necrosis", "findings_json", "ev3", ctx="Left lung:\nExuberant amount of heaping"),
    s("LMS anastomosis", "ANAT_AIRWAY", "LMS Anastomosis", "target.location.lobe", "ev3"),
    s("LC2", "ANAT_AIRWAY", "LC2", "target.location.segment", "ev3"),
    s("LUL", "ANAT_AIRWAY", "LUL", "target.location.lobe", "ev3", ctx="LC2,"),
    s("lingula", "ANAT_AIRWAY", "Lingula", "target.location.lobe", "ev3"),
    s("LC1", "ANAT_AIRWAY", "LC1", "target.location.segment", "ev3"),
    s("proximal LLL", "ANAT_AIRWAY", "LLL", "target.location.lobe", "ev3"),
    s("fractured", "OBS_LESION", "Fracture", "findings_json", "ev3", ctx="LUL bronchus cartilage appear"),
    s("full-thickness erosion", "OBS_LESION", "Erosion", "findings_json", "ev3", ctx="Large"),
    s("LUL bronchus", "ANAT_AIRWAY", "LUL", "target.location.lobe", "ev3", ctx="anterior aspect of"),
    
    # Event 4: Therapeutic Aspiration
    s("Successful therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", "method", "ev4"),
    s("distal RMS", "ANAT_AIRWAY", "RMS", "target.location.lobe", "ev4", ctx="clean out the"),
    s("RUL", "ANAT_AIRWAY", "RUL", "target.location.lobe", "ev4", ctx="distal RMS,"),
    s("BI", "ANAT_AIRWAY", "Bronchus Intermedius", "target.location.lobe", "ev4", ctx="RUL,"),
    s("RML", "ANAT_AIRWAY", "RML", "target.location.lobe", "ev4", ctx="BI,"),
    s("RLL", "ANAT_AIRWAY", "RLL", "target.location.lobe", "ev4", ctx="RML,"),
    s("distal LMS", "ANAT_AIRWAY", "LMS", "target.location.lobe", "ev4", ctx="RLL,"),
    s("LUL/lingula", "ANAT_AIRWAY", "LUL", "target.location.lobe", "ev4", ctx="distal LMS,"),
    s("LLL", "ANAT_AIRWAY", "LLL", "target.location.lobe", "ev4", ctx="LUL/lingula,"),
    s("clean out", "PROC_ACTION", "Clean", "method", "ev4"),
    s("mucus", "OBS_LESION", "Mucus", "findings_json", "ev4"),
    
    # Event 5: Measurements
    s("RMS bronchus", "ANAT_AIRWAY", "RMS", "target.location.lobe", "ev5", ctx="Airway measurements taken"),
    s("3.5 cm", "MEAS_SIZE", "3.5 cm", "measurements_json", "ev5"),
    s("RUL bronchus", "ANAT_AIRWAY", "RUL", "target.location.lobe", "ev5", ctx="3.5 cm\n"),
    s("2 cm", "MEAS_SIZE", "2 cm", "measurements_json", "ev5", ctx="RUL bronchus:"),
    s("BI", "ANAT_AIRWAY", "Bronchus Intermedius", "target.location.lobe", "ev5", ctx="2 cm\n"),
    s("2 cm", "MEAS_SIZE", "2 cm", "measurements_json", "ev5", ctx="BI:"),
    s("RC2 to MC", "ANAT_AIRWAY", "RC2 to MC", "target.location.segment", "ev5"),
    s("5.5", "MEAS_SIZE", "5.5 cm", "measurements_json", "ev5"),
    s("5.2 cm", "MEAS_SIZE", "5.2 cm", "measurements_json", "ev5"),

    # Event 6: Complications / Outcome
    s("no immediate complications", "OUTCOME_COMPLICATION", "None", "outcomes.complications", "ev6", neg=True),
    s("bare metallic stent", "DEV_STENT", "Bare Metallic Stent", "stent.type", "ev7", time=True, ctx="likely"),
    s("Fri, 9/6/2024", "CTX_TIME", "2024-09-06", "time", "ev7"),
]

# -------------------------------------------------------------------------
# HYDRATION LOGIC
# -------------------------------------------------------------------------

def get_offsets(text, span_text, context_prefix=None):
    if not span_text:
        return "", "", "missing_text"
    
    # Exact count check
    count = text.count(span_text)
    
    if count == 0:
        return "", "", "not_found"
    
    if count == 1:
        start = text.find(span_text)
        end = start + len(span_text)
        return start, end, "hydrated_unique"
    
    # Context match
    if context_prefix:
        # Find all occurrences
        starts = [m.start() for m in re.finditer(re.escape(span_text), text)]
        best_start = -1
        
        for s_idx in starts:
            # Look at preceding 120 chars
            window_start = max(0, s_idx - 120)
            window = text[window_start:s_idx]
            if context_prefix in window:
                best_start = s_idx
                break
        
        if best_start != -1:
            return best_start, best_start + len(span_text), "hydrated_prefix_window"

    return "", "", f"ambiguous_count={count}"

# -------------------------------------------------------------------------
# EXCEL GENERATION
# -------------------------------------------------------------------------

def generate_workbook():
    # Load template
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        print(f"Error: Template not found at {TEMPLATE_PATH}")
        return

    # 1. Note_Text
    ws_text = wb["Note_Text"]
    # Assuming row 1 header, write to row 2
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws_index = wb["Note_Index"]
    # Columns: source_file, note_id, encounter_id, procedure_date, site, reviewer, status, free_text, [flags...]
    row_data = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Pending", ""]
    
    # Append flags in order
    flag_keys = list(PROCEDURE_FLAGS.keys())
    for k in flag_keys:
        row_data.append(PROCEDURE_FLAGS[k])
    
    ws_index.append(row_data)

    # 3. Span_Annotations & 4. Span_Hydrated
    ws_span = wb["Span_Annotations"]
    ws_hydrated = wb["Span_Hydrated"]
    
    # Headers should already exist in template, appending from row 2
    
    hydrated_rows = []
    
    for idx, sp in enumerate(spans_data):
        span_id = f"{NOTE_ID}_s{idx+1:03d}"
        
        # Calculate hydration
        start, end, status = get_offsets(NOTE_TEXT, sp["text"], sp["context_prefix"])
        span_len = len(sp["text"]) if sp["text"] else 0
        
        # Common row part
        # source, note, span_id, section(blank), context, text, match_idx(blank)
        row_base = [
            SOURCE_FILE, NOTE_ID, span_id, "", 
            sp["context_prefix"], sp["text"], ""
        ]
        
        # Annotations sheet: start/end blank
        row_anno = row_base + ["", "", f"=LEN(F{ws_span.max_row+1})"]
        # Remainder: label, norm, field, event, neg, hist, time, reviewer, comment, status
        row_meta = [
            sp["label"], sp["norm"], sp["field"], sp["event_id"],
            sp["is_negated"], sp["is_historical"], sp["time_anchor"],
            "", "", "needs_hydration"
        ]
        
        ws_span.append(row_anno + row_meta)
        
        # Hydrated sheet: start/end filled
        row_hyd = row_base + [start, end, span_len]
        row_meta_hyd = [
            sp["label"], sp["norm"], sp["field"], sp["event_id"],
            sp["is_negated"], sp["is_historical"], sp["time_anchor"],
            "", "", status
        ]
        ws_hydrated.append(row_hyd + row_meta_hyd)

    # 5. Event_Log
    ws_event = wb["Event_Log"]
    
    # Mapping events manually for V3 log
    events = {
        "ev1": {"type": "Diagnostic Bronchoscopy", "method": "Moderate Sedation", "device": "Flexible Therapeutic Bronchoscope"},
        "ev2": {"type": "Inspection", "anatomy": "RMS, BI", "findings": "Necrotic debris, Dehiscence, Exposed Cartilage"},
        "ev3": {"type": "Inspection", "anatomy": "LMS, LUL", "findings": "Necrosis, Erosion, Fractured cartilage"},
        "ev4": {"type": "Therapeutic Aspiration", "method": "Suction/Clean", "anatomy": "Bilateral Bronchial Tree", "outcome_post": "Cleaned"},
        "ev5": {"type": "Measurements", "measurements": "RMS: 3.5cm, RUL: 2cm, BI: 2cm"},
        "ev6": {"type": "Outcome", "outcome_comp": "None"}
    }
    
    # Columns: source, note, event_id, type, method, anat, dev, gauge, station, count, meas, spec, find, hist, rev, comm, dev_size, dev_mat, out_lum_pre, out_lum_post, out_sym, out_pl, out_comp
    
    # EV1
    ws_event.append([SOURCE_FILE, NOTE_ID, "ev1", "Diagnostic Bronchoscopy", "Moderate Sedation", "Trachea", "Flexible Bronchoscope", "", "", "", "", "", "", "FALSE", "", "", "", "", "", "", "", "", ""])
    
    # EV2
    ws_event.append([SOURCE_FILE, NOTE_ID, "ev2", "Inspection", "", "RMS, BI", "", "", "", "", "", "", "Necrosis, Dehiscence", "FALSE", "", "", "", "", "", "", "", "", ""])

    # EV3
    ws_event.append([SOURCE_FILE, NOTE_ID, "ev3", "Inspection", "", "LMS, LUL", "", "", "", "", "", "", "Necrosis, Erosion", "FALSE", "", "", "", "", "", "", "", "", ""])

    # EV4
    ws_event.append([SOURCE_FILE, NOTE_ID, "ev4", "Therapeutic Aspiration", "Suction", "Bilateral Airways", "", "", "", "", "", "", "Mucus", "FALSE", "", "", "", "", "", "Cleaned", "", "", ""])

    # EV5
    ws_event.append([SOURCE_FILE, NOTE_ID, "ev5", "Measurements", "", "RMS, RUL, BI", "", "", "", "", "3.5cm, 2cm, 5.5cm", "", "", "FALSE", "", "", "", "", "", "", "", "", ""])

    # EV6
    ws_event.append([SOURCE_FILE, NOTE_ID, "ev6", "Outcome", "", "", "", "", "", "", "", "", "", "FALSE", "", "", "", "", "", "", "", "", "None"])

    # 6. V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    # Headers: note_id, event_id, type, target.anat, loc.lobe, loc.seg, station, les.type, les.size, method, dev_json, meas_json, spec_json, find_json, quote, stent.size, stent.mat, cath.size, out.lum.pre, out.lum.post, out.sym, out.pl, out.comp
    
    # Ev1
    ws_v3.append([NOTE_ID, "ev1", "Diagnostic Bronchoscopy", "Airway", "Trachea", "", "", "", "", "Moderate Sedation", '["Flexible Bronchoscope"]', "", "", "", "", "", "", "", "", "", "", "", ""])
    
    # Ev2
    ws_v3.append([NOTE_ID, "ev2", "Inspection", "Airway", "Right Mainstem", "", "", "Dehiscence", "", "Inspection", "", "", "", '["Necrosis", "Dehiscence"]', "full thickness dehiscence", "", "", "", "", "", "", "", ""])

    # Ev4
    ws_v3.append([NOTE_ID, "ev4", "Therapeutic Aspiration", "Airway", "Bilateral", "", "", "Mucus", "", "Aspiration", "", "", "", '["Mucus"]', "Successful therapeutic aspiration", "", "", "", "", "Patent", "", "", ""])

    # Ev6
    ws_v3.append([NOTE_ID, "ev6", "Outcome", "", "", "", "", "", "", "", "", "", "", "", "no immediate complications", "", "", "", "", "", "", "", "None"])

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    
    registry_data = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "procedures": [
            {
                "event_id": "ev1",
                "type": "Diagnostic Bronchoscopy",
                "method": "Moderate Sedation",
                "devices": ["Flexible Bronchoscope"]
            },
            {
                "event_id": "ev2",
                "type": "Inspection",
                "location": {"lobe": "Right Mainstem"},
                "findings": ["Necrosis", "Dehiscence", "Fractured Cartilage"]
            },
            {
                "event_id": "ev4",
                "type": "Therapeutic Aspiration",
                "action": "Clean",
                "findings": ["Mucus"]
            },
            {
                "event_id": "ev5",
                "type": "Measurements",
                "measurements": [
                    {"anatomy": "RMS", "value": "3.5 cm"},
                    {"anatomy": "RUL", "value": "2 cm"},
                    {"anatomy": "BI", "value": "2 cm"}
                ]
            }
        ],
        "no_immediate_complications": True
    }
    
    json_str = json.dumps(registry_data, indent=2)
    ws_json.append([json_str])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Successfully generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()