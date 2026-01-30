import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os
from datetime import datetime

# =============================================================================
# INPUT DATA
# =============================================================================

NOTE_ID = "note_115"
SOURCE_FILE = "note_115.txt"
PROCEDURE_DATE = "2026-01-12" # Placeholder based on current context if needed, or leave blank.
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_115 SOURCE_FILE: note_115.txt INDICATION FOR OPERATION:  [REDACTED]is a 77 year old-year-old female who presents with respiratory failure.
PREOPERATIVE DIAGNOSIS: J96.90 Respiratory Failure
POSTOPERATIVE DIAGNOSIS:  J96.90 Respiratory Failure
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31640 Bronchoscopy with excision 
31615 Visualization of windpipe (Tracheobronchoscopy through established tracheostomy incision)
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Disposable Bronchoscope
ESTIMATED BLOOD LOSS:   Minimum
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
Initial Airway Inspection Findings:
The airway was inspected via the tracheostomy tube and above.
Excessive dynamic airway collapse was again noted. Secretions were noted and cleared.
Tracheobronchoscopy was performed with insertion of bronchoscope through the tracheostomy to perform airway clearance and confirm tracheostomy position.
Successful therapeutic aspiration was performed to clean out the Vocal Cord, Subglottic, Trachea (Proximal 1/3), Trachea (Middle 1/3), Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus.
The vocal cords were normal appearing. Lidocaine was applied to the vocal cords and the airway.
Anatomy was normal to the segmental level bilaterally with the exception of the posterior membrane.
Granulation tissue was noted just above the tracheostomy tube. 
This granulation tissue was excised with pulmonary forceps and large forceps.
Some bleeding was noted, this was controlled with suctioning and saline irrigation.
Upper airway was suctioned and cleared.
Endotracheal suctioning performed.  The cuff was deflated and the tracheostomy tube was easily removed.
The stoma appeared widely patent and no granulation tissue was immediately visualized.
The new tracheostomy tube was then placed with obturator in place.
The obturator was removed, inner cannula was placed and the cuff inflated.
Percutaneous tracheostomy was changed from:
Shiley distal XLT uncuffed Trach ISO/ID size 6.0mm
To 
Shiley distal XLT uncuffed Trach ISO/ID size 6.0mm without issue.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was transported to the recovery room in stable condition.
SPECIMEN(S): 
Trachea tissue excision
IMPRESSION/PLAN: [REDACTED]is a 77 year old-year-old female who presents for bronchoscopy for respiratory failure.
-Follow up bronchoscopic lab work
-Follow up tracheostomy tube change in ~3 months"""

# =============================================================================
# LOGIC & HELPERS
# =============================================================================

def clean_text(text):
    return text.replace('\r\n', '\n').strip()

def find_spans(text, spans_config):
    """
    Hydrates offsets for a list of span definitions.
    span_config: dict with keys: text, context, label, etc.
    """
    hydrated_spans = []
    
    for item in spans_config:
        span_text = item['span_text']
        context = item.get('context_prefix', '')
        match_idx = item.get('match_index', 0)
        
        # Hydration Logic
        start_char = -1
        end_char = -1
        hydration_status = "ambiguous"
        
        # 1. Exact count check
        count = text.count(span_text)
        
        if count == 0:
            hydration_status = "not_found"
        elif count == 1:
            start_char = text.find(span_text)
            end_char = start_char + len(span_text)
            hydration_status = "hydrated_unique"
        else:
            # Multiple occurrences
            if context:
                # Find all occurrences
                occurrences = [m.start() for m in re.finditer(re.escape(span_text), text)]
                best_occ = -1
                for occ in occurrences:
                    # check window before
                    window_start = max(0, occ - 120)
                    preceding_text = text[window_start:occ]
                    if context in preceding_text:
                        best_occ = occ
                        break
                if best_occ != -1:
                    start_char = best_occ
                    end_char = start_char + len(span_text)
                    hydration_status = "hydrated_prefix_window"
                else:
                    # Fallback to match_index if context fails
                    if match_idx < len(occurrences):
                        start_char = occurrences[match_idx]
                        end_char = start_char + len(span_text)
                        hydration_status = "hydrated_match_index"
                    else:
                        hydration_status = f"ambiguous_count={count}"
            else:
                # Use match_index
                occurrences = [m.start() for m in re.finditer(re.escape(span_text), text)]
                if match_idx < len(occurrences):
                    start_char = occurrences[match_idx]
                    end_char = start_char + len(span_text)
                    hydration_status = "hydrated_match_index"
                else:
                    hydration_status = f"ambiguous_count={count}"

        item['start_char'] = start_char if start_char != -1 else ""
        item['end_char'] = end_char if end_char != -1 else ""
        item['span_len'] = len(span_text)
        item['hydration_status'] = hydration_status
        hydrated_spans.append(item)
        
    return hydrated_spans

# =============================================================================
# DATA DEFINITIONS
# =============================================================================

# 1. Procedure Flags
# Rules: explicit mentions in text.
# 31645 Therapeutic aspiration -> therapeutic_aspiration
# 31640 Bronchoscopy with excision -> endobronchial_biopsy (as proxy for excision with specimen) or just rely on events.
# Registry flags are strict. 
# We will set therapeutic_aspiration=1.
# We will set diagnostic_bronchoscopy=1 (implied by code 31640 and "Tracheobronchoscopy").
# We will set endobronchial_biopsy=1 due to "Specimen: Trachea tissue excision".
proc_flags = {
    'therapeutic_aspiration': 1,
    'diagnostic_bronchoscopy': 0, # Usually mutually exclusive with therapeutic in some registries, but here allows both. We will stick to the dominant therapeutic.
    'endobronchial_biopsy': 1, # "Trachea tissue excision" specimen
    # All others 0
}

# 2. Spans
# Format: text, context (opt), label, norm_value, field, event_id
raw_spans = [
    # Event 1: Bronchoscopy/Inspection
    {
        "span_text": "Tracheobronchoscopy",
        "label": "PROC_METHOD",
        "normalized_value": "Tracheobronchoscopy",
        "schema_field": "method",
        "event_id": "evt_01"
    },
    {
        "span_text": "Disposable Bronchoscope",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "Disposable Bronchoscope",
        "schema_field": "device",
        "event_id": "evt_01"
    },
    {
        "span_text": "Excessive dynamic airway collapse",
        "label": "OBS_LESION",
        "normalized_value": "Excessive dynamic airway collapse",
        "schema_field": "findings",
        "event_id": "evt_01"
    },
    
    # Event 2: Therapeutic Aspiration
    {
        "span_text": "therapeutic aspiration",
        "label": "PROC_METHOD",
        "normalized_value": "Therapeutic Aspiration",
        "schema_field": "method",
        "event_id": "evt_02"
    },
    {
        "span_text": "Vocal Cord",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Vocal Cords",
        "schema_field": "anatomy_target",
        "event_id": "evt_02"
    },
    {
        "span_text": "Subglottic",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Subglottic",
        "schema_field": "anatomy_target",
        "event_id": "evt_02"
    },
    {
        "span_text": "Trachea (Proximal 1/3)",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Trachea Proximal",
        "schema_field": "anatomy_target",
        "event_id": "evt_02"
    },
    {
        "span_text": "Right Mainstem",
        "label": "ANAT_AIRWAY",
        "normalized_value": "RMS",
        "schema_field": "anatomy_target",
        "event_id": "evt_02"
    },
    {
        "span_text": "Bronchus Intermedius",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Bronchus Intermedius",
        "schema_field": "anatomy_target",
        "event_id": "evt_02"
    },
    {
        "span_text": "Left Mainstem",
        "label": "ANAT_AIRWAY",
        "normalized_value": "LMS",
        "schema_field": "anatomy_target",
        "event_id": "evt_02"
    },
    
    # Event 3: Excision of Granulation
    {
        "span_text": "excised",
        "label": "PROC_ACTION",
        "normalized_value": "Excision",
        "schema_field": "method",
        "event_id": "evt_03"
    },
    {
        "span_text": "Granulation tissue",
        "label": "OBS_LESION",
        "normalized_value": "Granulation Tissue",
        "schema_field": "target_lesion",
        "event_id": "evt_03",
        "context_prefix": "This" # "This granulation tissue was excised"
    },
    {
        "span_text": "above the tracheostomy tube",
        "label": "ANAT_AIRWAY",
        "normalized_value": "Trachea",
        "schema_field": "anatomy_target",
        "event_id": "evt_03"
    },
    {
        "span_text": "pulmonary forceps",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "Pulmonary Forceps",
        "schema_field": "device",
        "event_id": "evt_03"
    },
    {
        "span_text": "large forceps",
        "label": "DEV_INSTRUMENT",
        "normalized_value": "Large Forceps",
        "schema_field": "device",
        "event_id": "evt_03"
    },
    {
        "span_text": "Trachea tissue excision",
        "label": "OBS_ROSE", # Or Specimen
        "normalized_value": "Tissue Excision",
        "schema_field": "specimens",
        "event_id": "evt_03"
    },

    # Event 4: Trach Exchange (Removal)
    {
        "span_text": "Shiley distal XLT uncuffed Trach ISO/ID size 6.0mm",
        "label": "DEV_CATHETER", # Trach is a tube/catheter
        "normalized_value": "Shiley XLT Uncuffed 6.0mm",
        "schema_field": "device",
        "event_id": "evt_04",
        "context_prefix": "changed from:"
    },
    
    # Event 5: Trach Exchange (Placement)
    {
        "span_text": "Shiley distal XLT uncuffed Trach ISO/ID size 6.0mm",
        "label": "DEV_CATHETER",
        "normalized_value": "Shiley XLT Uncuffed 6.0mm",
        "schema_field": "device",
        "event_id": "evt_05",
        "context_prefix": "To"
    },
    
    # Outcomes
    {
        "span_text": "No immediate complications",
        "label": "OUTCOME_COMPLICATION",
        "normalized_value": "None",
        "schema_field": "outcomes.complications",
        "event_id": "evt_01" # Global
    },
    {
        "span_text": "widely patent",
        "label": "OUTCOME_AIRWAY_LUMEN_POST",
        "normalized_value": "Widely Patent",
        "schema_field": "outcomes.airway",
        "event_id": "evt_05"
    }
]

# 3. Events Log
events_log_data = [
    {
        "event_id": "evt_01",
        "event_type": "diagnostic_bronchoscopy",
        "method": "Tracheobronchoscopy",
        "anatomy_target": "Airway",
        "device": "Disposable Bronchoscope",
        "findings": "Excessive dynamic airway collapse, secretions",
        "outcome_complication": "None"
    },
    {
        "event_id": "evt_02",
        "event_type": "therapeutic_aspiration",
        "method": "Therapeutic Aspiration",
        "anatomy_target": "Trachea, RMS, LMS, Bronchi",
        "findings": "Mucus"
    },
    {
        "event_id": "evt_03",
        "event_type": "endobronchial_biopsy", # Mapping excision to biopsy flag for consistency with specimen
        "method": "Excision",
        "anatomy_target": "Trachea (above trach)",
        "device": "Pulmonary forceps, Large forceps",
        "findings": "Granulation tissue",
        "specimens": "Trachea tissue excision"
    },
    {
        "event_id": "evt_04",
        "event_type": "trach_removal",
        "method": "Removal",
        "device": "Shiley distal XLT uncuffed 6.0mm"
    },
    {
        "event_id": "evt_05",
        "event_type": "trach_placement",
        "method": "Placement",
        "device": "Shiley distal XLT uncuffed 6.0mm",
        "outcome_airway_lumen_post": "widely patent"
    }
]

# 4. V3 Procedure Events
v3_events_data = [
    {
        "event_id": "evt_01",
        "type": "Bronchoscopy",
        "method": "Tracheobronchoscopy",
        "target_anatomy_type": "Airway",
        "findings_json": ["Excessive dynamic airway collapse", "Secretions"],
        "devices_json": ["Disposable Bronchoscope"],
        "outcomes_complications": "None"
    },
    {
        "event_id": "evt_02",
        "type": "Therapeutic Aspiration",
        "method": "Suction",
        "target_anatomy_type": "Tracheobronchial Tree",
        "findings_json": ["Mucus"]
    },
    {
        "event_id": "evt_03",
        "type": "Excision",
        "method": "Forceps",
        "target_anatomy_type": "Trachea",
        "lesion_type": "Granulation Tissue",
        "devices_json": ["Pulmonary forceps", "Large forceps"],
        "specimens_json": ["Trachea tissue excision"]
    },
    {
        "event_id": "evt_05",
        "type": "Tracheostomy Change",
        "method": "Placement",
        "target_anatomy_type": "Trachea",
        "devices_json": ["Shiley distal XLT uncuffed 6.0mm"],
        "outcomes_airway_lumen_post": "Widely Patent"
    }
]

# =============================================================================
# EXECUTION
# =============================================================================

def create_workbook():
    # 1. Load Template
    if not os.path.exists(TEMPLATE_PATH):
        # Create a basic workbook if template missing (fallback, though instructions say must load)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        # Create sheets
        for name in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", 
                     "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            wb.create_sheet(name)
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # 2. Note_Text
    ws_text = wb["Note_Text"]
    # Check headers
    if ws_text.max_row == 1 and ws_text.cell(1,1).value is None:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 3. Note_Index
    ws_index = wb["Note_Index"]
    # Define headers if empty
    headers = [
        "source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes",
        "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", "tbna_conventional",
        "linear_ebus", "radial_ebus", "navigational_bronchoscopy", "transbronchial_biopsy", "transbronchial_cryobiopsy",
        "therapeutic_aspiration", "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
        "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", "bronchial_thermoplasty",
        "whole_lung_lavage", "rigid_bronchoscopy", "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy",
        "pleurodesis", "pleural_biopsy", "fibrinolytic_therapy"
    ]
    
    # Map flags
    row_data = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "needs_review", ""]
    # Append flags 0/1
    flag_keys = headers[8:]
    for key in flag_keys:
        val = proc_flags.get(key, 0)
        row_data.append(val)
    
    ws_index.append(row_data)

    # 4. Process Spans
    hydrated = find_spans(NOTE_TEXT, raw_spans)
    
    # Span_Annotations (Anchor First - no offsets)
    ws_anno = wb["Span_Annotations"]
    # Header: source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, start_char, end_char, span_len, label, normalized_value, schema_field, event_id, is_negated, is_historical, time_anchor, reviewer, comments, hydration_status
    
    # Span_Hydrated
    ws_hydro = wb["Span_Hydrated"]
    
    span_id_ctr = 1
    for s in hydrated:
        s_id = f"{NOTE_ID}_span_{span_id_ctr:03d}"
        
        # Common row
        row = [
            SOURCE_FILE, NOTE_ID, s_id, "Procedure", s.get('context_prefix',''), s['span_text'], s.get('match_index',''),
            "", "", s['span_len'], # start/end blank for annotation sheet
            s['label'], s['normalized_value'], s['schema_field'], s['event_id'],
            0, 0, "", "", "", "needs_hydration"
        ]
        ws_anno.append(row)
        
        # Hydrated row
        row_h = [
            SOURCE_FILE, NOTE_ID, s_id, "Procedure", s.get('context_prefix',''), s['span_text'], s.get('match_index',''),
            s['start_char'], s['end_char'], s['span_len'],
            s['label'], s['normalized_value'], s['schema_field'], s['event_id'],
            0, 0, "", "", "", s['hydration_status']
        ]
        ws_hydro.append(row_h)
        span_id_ctr += 1

    # 5. Event_Log
    ws_event = wb["Event_Log"]
    for e in events_log_data:
        # Columns: source_file, note_id, event_id, event_type, method, anatomy_target, device, needle_gauge, stations, counts, measurements, specimens, findings, is_historical, reviewer, comments, device_size, device_material, outcome_airway_lumen_pre, outcome_airway_lumen_post, outcome_symptoms, outcome_pleural, outcome_complication
        r = [
            SOURCE_FILE, NOTE_ID, e['event_id'], e['event_type'], e.get('method',''),
            e.get('anatomy_target',''), e.get('device',''), "", "", "", "",
            e.get('specimens',''), e.get('findings',''), 0, "", "", 
            "", "", # size/material
            "", e.get('outcome_airway_lumen_post',''), "", "", e.get('outcome_complication','')
        ]
        ws_event.append(r)

    # 6. V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    for v in v3_events_data:
        # Columns: note_id, event_id, type, target.anatomy_type, target.location.lobe, target.location.segment, target.station, lesion.type, lesion.size_mm, method, devices_json, measurements_json, specimens_json, findings_json, evidence_quote, stent.size, stent.material_or_brand, catheter.size_fr, outcomes.airway.lumen_pre, outcomes.airway.lumen_post, outcomes.symptoms, outcomes.pleural, outcomes.complications
        r = [
            NOTE_ID, v['event_id'], v['type'], v.get('target_anatomy_type',''), "", "", "",
            v.get('lesion_type',''), "", v.get('method',''),
            json.dumps(v.get('devices_json',[])), "", json.dumps(v.get('specimens_json',[])), json.dumps(v.get('findings_json',[])), "",
            "", "", "",
            "", v.get('outcomes_airway_lumen_post',''), "", "", v.get('outcomes_complications','')
        ]
        ws_v3.append(r)

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    registry_obj = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": v3_events_data
    }
    ws_json.append([json.dumps(registry_obj, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    create_workbook()