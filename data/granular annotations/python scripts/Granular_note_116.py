import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_116"
SOURCE_FILE = "note_116.txt"
PROCEDURE_DATE = "2026-01-12" # Placeholder based on context
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_116 SOURCE_FILE: note_116.txt INDICATION FOR OPERATION:  [REDACTED]is a 68 year old-year-old female who presents with hemoptysis.
PREOPERATIVE DIAGNOSIS: R91.8 Other nonspecific abnormal finding of lung field.
POSTOPERATIVE DIAGNOSIS:  R91.8 Other nonspecific abnormal finding of lung field.
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31622 Dx bronchoscope/cell washing          
31635 Foreign body removal
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
ESTIMATED BLOOD LOSS:   Minimum
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
Initial Airway Inspection Findings:
The airway was inspected. Lidocaine was applied to vocal cords and the airway.
The vocal cords were normal appearing.
Blood and blood clots were noted throughout the airway.
Most notably the left lower lobe and the right mainstem. 
The blood and blood clots were suctioned.
This required multiple attempts to remove the foreign bodies. 
No active bleeding was noted.
The thin blood and secretions were suctioned.
Successful therapeutic aspiration was performed to clean out the Vocal Cord, Subglottic, Trachea (Proximal 1/3), Trachea (Middle 1/3), Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus and blood.
All blood clots were removed therapeutically from the airway. 
 
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
None
IMPRESSION/PLAN: [REDACTED]is a 68 year old-year-old female who presents for bronchoscopy for hemoptysis.
-Follow up bronchoscopic lab work"""

# -------------------------------------------------------------------------
# CONFIGURATION
# -------------------------------------------------------------------------
# Procedure Flags Map (1 = Present, 0 = Absent)
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1, # "Dx bronchoscope"
    "bal": 0,
    "bronchial_wash": 1, # "cell washing" in header
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1, # "Therapeutic aspiration"
    "foreign_body_removal": 1, # "Foreign body removal" (clots)
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# -------------------------------------------------------------------------
# SPAN DEFINITIONS
# -------------------------------------------------------------------------
# Raw spans to be processed and hydrated.
# Format: (span_text, label, normalized_value, schema_field, event_id, context_prefix or None)
RAW_SPANS = [
    # Event 1: Diagnostic/Inspection
    ("Flexible Therapeutic Bronchoscope", "DEV_INSTRUMENT", "therapeutic_bronchoscope", "device", "evt_01", None),
    ("vocal cords", "ANAT_AIRWAY", "vocal_cords", "target", "evt_01", "The"),
    ("normal appearing", "OBS_LESION", "normal", "findings", "evt_01", "vocal cords were"),
    ("Blood and blood clots", "OBS_LESION", "blood_clots", "findings", "evt_01", None),
    ("throughout the airway", "ANAT_AIRWAY", "tracheobronchial_tree", "target", "evt_01", None),
    ("left lower lobe", "ANAT_LUNG_LOC", "LLL", "target", "evt_01", None),
    ("right mainstem", "ANAT_AIRWAY", "RMS", "target", "evt_01", "lobe and the"),
    
    # Event 2: Therapeutic Aspiration / Foreign Body Removal (Clots)
    ("suctioned", "PROC_ACTION", "suction", "method", "evt_02", "blood clots were"),
    ("remove the foreign bodies", "PROC_METHOD", "foreign_body_removal", "method", "evt_02", None),
    ("No active bleeding", "OBS_LESION", "bleeding_active", "findings", "evt_02", None),
    ("Successful therapeutic aspiration", "PROC_METHOD", "therapeutic_aspiration", "method", "evt_02", None),
    ("clean out", "PROC_ACTION", "clean_out", "method", "evt_02", None),
    
    # Specific anatomy cleaned
    ("Vocal Cord", "ANAT_AIRWAY", "vocal_cords", "target", "evt_02", "clean out the"),
    ("Subglottic", "ANAT_AIRWAY", "subglottic", "target", "evt_02", None),
    ("Trachea (Proximal 1/3)", "ANAT_AIRWAY", "trachea_prox", "target", "evt_02", None),
    ("Trachea (Middle 1/3)", "ANAT_AIRWAY", "trachea_mid", "target", "evt_02", None),
    ("Trachea (Distal 1/3)", "ANAT_AIRWAY", "trachea_dist", "target", "evt_02", None),
    ("Right Mainstem", "ANAT_AIRWAY", "RMS", "target", "evt_02", "3),"),
    ("Bronchus Intermedius", "ANAT_AIRWAY", "bronchus_intermedius", "target", "evt_02", None),
    ("Left Mainstem", "ANAT_AIRWAY", "LMS", "target", "evt_02", None),
    ("Carina", "ANAT_AIRWAY", "carina", "target", "evt_02", "Mainstem,"),
    ("RUL Carina (RC1)", "ANAT_AIRWAY", "RUL_carina_rc1", "target", "evt_02", None),
    ("RML Carina (RC2)", "ANAT_AIRWAY", "RML_carina_rc2", "target", "evt_02", None),
    ("LUL Lingula Carina (Lc1)", "ANAT_AIRWAY", "LUL_lingula_carina_lc1", "target", "evt_02", None),
    ("Left Carina (LC2)", "ANAT_AIRWAY", "left_carina_lc2", "target", "evt_02", None),
    
    ("mucus and blood", "OBS_LESION", "mucus_blood", "findings", "evt_02", None),
    ("All blood clots were removed", "PROC_ACTION", "remove_clots", "method", "evt_02", None),
    
    # Outcomes
    ("No immediate complications", "OUTCOME_COMPLICATION", "none", "outcomes.complications", "evt_03", None)
]

# -------------------------------------------------------------------------
# HELPER FUNCTIONS
# -------------------------------------------------------------------------
def get_offsets(full_text, span_text, context_prefix=None, match_index=0):
    """
    Calculates start_char and end_char for a span.
    Logic:
    1. If context_prefix is provided, find occurrences of span_text preceded by context within window.
    2. If no context, finds all occurrences and picks match_index-th one.
    3. Returns (start, end, status)
    """
    if not span_text:
        return "", "", "missing_text"
    
    # Escape regex special chars in span_text
    escaped_span = re.escape(span_text)
    
    if context_prefix:
        # Search for context + ... + span
        # Look for span where preceding 120 chars contain context_prefix
        # We'll iterate all matches of span_text
        matches = [m for m in re.finditer(escaped_span, full_text)]
        candidate = None
        
        for m in matches:
            start = m.start()
            # Look back 120 chars
            window_start = max(0, start - 120)
            preceding_text = full_text[window_start:start]
            if context_prefix in preceding_text:
                candidate = m
                break # Take first match with correct context
        
        if candidate:
            return candidate.start(), candidate.end(), "hydrated_prefix_window"
        else:
            return "", "", "context_not_found"
            
    else:
        # Simple match by index
        matches = [m for m in re.finditer(escaped_span, full_text)]
        if not matches:
            return "", "", "text_not_found"
        
        if match_index < len(matches):
            m = matches[match_index]
            return m.start(), m.end(), "hydrated_unique" if len(matches) == 1 else "hydrated_match_index"
        else:
            return "", "", f"index_out_of_bounds_count_{len(matches)}"

def create_workbook_if_needed():
    """Creates a new workbook if template doesn't exist (fallback) or loads template."""
    if os.path.exists(TEMPLATE_PATH):
        return openpyxl.load_workbook(TEMPLATE_PATH)
    else:
        # Fallback: create basic structure (not ideal but compliant for logic)
        wb = openpyxl.Workbook()
        for sheet in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            if sheet not in wb.sheetnames:
                wb.create_sheet(sheet)
        return wb

# -------------------------------------------------------------------------
# MAIN EXECUTION
# -------------------------------------------------------------------------
def main():
    wb = create_workbook_if_needed()
    
    # ---------------------------------------
    # 1. Note_Text
    # ---------------------------------------
    ws_text = wb["Note_Text"]
    # Append row: note_id, source_file, note_text
    # Assuming header exists or append to end
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # ---------------------------------------
    # 2. Note_Index
    # ---------------------------------------
    ws_index = wb["Note_Index"]
    # Header: source_file, note_id, encounter_id, procedure_date, site, reviewer, status, free_text_notes, [flags...]
    # We need to ensure we map flags correctly to columns.
    # Assuming standard template column order for flags starts at column I (9) or similar. 
    # To be safe, we will just construct the list of values matching the prompt spec.
    
    # Flags keys strictly ordered as per prompt requirements
    bronch_keys = [
        "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", 
        "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy", 
        "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration", 
        "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation", 
        "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", 
        "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy"
    ]
    pleural_keys = [
        "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis", 
        "pleural_biopsy", "fibrinolytic_therapy"
    ]
    
    row_data = [
        SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "ready_for_review", ""
    ]
    
    # Append flag values (0/1)
    for k in bronch_keys:
        row_data.append(PROCEDURE_FLAGS.get(k, 0))
    for k in pleural_keys:
        row_data.append(PROCEDURE_FLAGS.get(k, 0))
        
    ws_index.append(row_data)

    # ---------------------------------------
    # 3. Span_Annotations & 4. Span_Hydrated
    # ---------------------------------------
    ws_anno = wb["Span_Annotations"]
    ws_hydra = wb["Span_Hydrated"]
    
    # Process spans
    spans_to_write = []
    
    for idx, (txt, label, norm, field, evt, ctx) in enumerate(RAW_SPANS):
        span_id = f"{NOTE_ID}_span_{idx+1:03d}"
        
        # Hydration
        start, end, status = get_offsets(NOTE_TEXT, txt, ctx)
        span_len = len(txt)
        
        # Meta info
        is_negated = True if label == "OUTCOME_COMPLICATION" and "no" in txt.lower() else False
        is_historical = False
        
        # Anno Row (Anchor First - no offsets)
        anno_row = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure", ctx, txt, 0, 
            "", "", f"=LEN(F{ws_anno.max_row+1})", # Formula for len
            label, norm, field, evt, 
            is_negated, is_historical, "", "algo", "", "needs_hydration"
        ]
        ws_anno.append(anno_row)
        
        # Hydra Row (With offsets)
        hydra_row = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure", ctx, txt, 0,
            start, end, span_len,
            label, norm, field, evt,
            is_negated, is_historical, "", "algo", "", status
        ]
        ws_hydra.append(hydra_row)
        
        spans_to_write.append({
            "id": span_id, "text": txt, "label": label, "norm": norm, "evt": evt
        })

    # ---------------------------------------
    # 5. Event_Log
    # ---------------------------------------
    ws_event = wb["Event_Log"]
    # Group by evt
    events = {}
    for s in spans_to_write:
        eid = s['evt']
        if eid not in events:
            events[eid] = {"method": [], "anatomy": [], "device": [], "findings": [], "outcomes": []}
        
        if s['label'] == "PROC_METHOD": events[eid]['method'].append(s['norm'])
        elif s['label'] == "PROC_ACTION": events[eid]['method'].append(s['norm'])
        elif s['label'].startswith("ANAT"): events[eid]['anatomy'].append(s['norm'])
        elif s['label'].startswith("DEV"): events[eid]['device'].append(s['norm'])
        elif s['label'].startswith("OBS"): events[eid]['findings'].append(s['norm'])
        elif s['label'].startswith("OUTCOME"): events[eid]['outcomes'].append(s['norm'])

    # Write Event Rows
    # Columns: source_file, note_id, event_id, event_type, method, anatomy_target, device, needle_gauge, 
    # stations, counts, measurements, specimens, findings, is_historical, reviewer, comments, 
    # device_size, device_material, outcome_airway_lumen_pre, outcome_airway_lumen_post, 
    # outcome_symptoms, outcome_pleural, outcome_complication
    
    # Event 1: Diagnostic
    e1 = events.get("evt_01", {})
    ws_event.append([
        SOURCE_FILE, NOTE_ID, "evt_01", "Diagnostic Bronchoscopy", 
        ", ".join(e1.get('method', ["inspection"])), 
        ", ".join(e1.get('anatomy', [])), 
        ", ".join(e1.get('device', [])), 
        "", "", "", "", "", 
        ", ".join(e1.get('findings', [])), 
        False, "algo", "Diagnostic inspection",
        "", "", "", "", "", "", ""
    ])

    # Event 2: Therapeutic Aspiration/FBR
    e2 = events.get("evt_02", {})
    ws_event.append([
        SOURCE_FILE, NOTE_ID, "evt_02", "Therapeutic Aspiration", 
        ", ".join(e2.get('method', [])), 
        "Multiple Airway Segments (see spans)", # Summarized for brevity in log
        "", 
        "", "", "", "", "", 
        ", ".join(e2.get('findings', [])), 
        False, "algo", "Clot removal",
        "", "", "", "", "", "", ""
    ])

    # Event 3: Outcome
    e3 = events.get("evt_03", {})
    ws_event.append([
        SOURCE_FILE, NOTE_ID, "evt_03", "Outcome", 
        "", "", "", "", "", "", "", "", "", 
        False, "algo", "",
        "", "", "", "", "", "", "none"
    ])

    # ---------------------------------------
    # 6. V3_Procedure_Events
    # ---------------------------------------
    ws_v3 = wb["V3_Procedure_Events"]
    # Simplified mapping for V3
    # Row 1
    ws_v3.append([
        NOTE_ID, "evt_01", "Diagnostic Bronchoscopy", 
        "Airway", "L Lobe/R Main", "", "", 
        "Blood/Clots", "", 
        "Inspection", json.dumps(["Flexible Therapeutic Bronchoscope"]), "", "", json.dumps(["normal vocal cords", "blood"]), "Blood and blood clots were noted",
        "", "", "", 
        "", "", "", "", ""
    ])
    # Row 2
    ws_v3.append([
        NOTE_ID, "evt_02", "Therapeutic Aspiration", 
        "Airway", "Bilateral", "Multiple", "", 
        "Clots", "", 
        "Suction/FBR", "", "", "", json.dumps(["mucus", "blood"]), "Successful therapeutic aspiration",
        "", "", "", 
        "", "", "", "", ""
    ])

    # ---------------------------------------
    # 7. V3_Registry_JSON
    # ---------------------------------------
    ws_json = wb["V3_Registry_JSON"]
    
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": [
            {
                "event_id": "evt_01",
                "type": "Diagnostic Bronchoscopy",
                "findings": ["Blood", "Clots"]
            },
            {
                "event_id": "evt_02",
                "type": "Therapeutic Aspiration",
                "method": ["Suction", "Foreign Body Removal"],
                "target": "Bilateral Airways"
            }
        ],
        "no_immediate_complications": True
    }
    
    ws_json.append([json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    main()