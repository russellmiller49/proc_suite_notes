import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import json
import re
import datetime

# ==========================================
# INPUT DATA
# ==========================================
NOTE_ID = "note_117"
SOURCE_FILE = "note_117.txt"
PROCEDURE_DATE = "2026-01-12" # inferred from context or left blank/current
NOTE_TEXT = """NOTE_ID:  note_117 SOURCE_FILE: note_117.txt INDICATION FOR OPERATION:  [REDACTED]is a 70 year old-year-old female who presents with lung nodule.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: R91.1 Solitary Lung Nodule
POSTOPERATIVE DIAGNOSIS:  R91.1 Solitary Lung Nodule
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31622 Dx bronchoscope/cell washing          
31623 Dx bronchoscope/brushing    
31624 Dx bronchoscope/lavage (BAL)    
31628 TBBX single lobe     
31629 TBNA single lobe   
31626 Fiducial marker placements, single or multiple     
31627 Navigational Bronchoscopy (computer assisted)
77012 Radiology / radiologic guidance for CT guided needle placement (CIOS)
76377 3D rendering with interpretation and reporting of CT, US, Tomo modality (ION Planning Station)
31652 EBUS sampling 1 
or 2 nodes
31654 Radial EBUS for peripheral lesion
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required a Transbronchial Cryo biopsies.
This resulted in >50% increased work due to Increased intensity, Time, Technical difficulty of procedure, and Physical and mental effort required.
Apply to: 31628 TBBX single lobe
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
Linear EBUS 
Radial EBUS
Ion Robotic Bronchoscope
Disposable Bronchoscope
ESTIMATED BLOOD LOSS:   Minimum
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
Initial Airway Inspection Findings:
Normal appearing airway anatomy and mucosa bilaterally to the segmental level.
CT Chest scan was placed on separate planning station to generate 3D rendering of the pathway to target.
The navigational plan was reviewed and verified.  This was then loaded into robotic bronchoscopy platform.
Successful therapeutic aspiration was performed to clean out the Trachea (Middle 1/3), Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus and blood.
Ventilation Parameters:
Mode	RR	TV	PEEP	FiO2	Flow Rate	Pmean
VCV	 	300	12	100	15	12
Robotic navigation bronchoscopy was performed with Ion platform.  Partial registration was used.
Ion robotic catheter was used to engage the Anterior Segment of RUL (RB3).
Target lesion is about 0.6 cm in diameter.   Under navigational guidance the ion robotic catheter was advanced to 1.0 cm away from the planned target.
Radial EBUS was performed to confirm that the location of the nodule is not well visualized.
The following features were noted: aerated lung.
Cone Beam CT was performed: 3-D reconstructions were performed on an independent workstation.
Cios Spin system was used for evaluation of nodule location.  Low dose spin was performed to acquire CT imaging.
This was passed on to Ion platform system for reconstruction and nodule location.
The 3D images was interpreted on an independent workstation (Ion).
Using the newly acquired nodule location, the Ion robotic system was adjusted to the new targeted location.
I personally interpreted the cone beam CT and 3-D reconstruction.
Transbronchial needle aspiration was performed with 21G Needle through the extended working channel catheter.  Total 5 samples were collected.
Samples sent for Microbiology (Cultures/Viral/Fungal) and Cytology.
Transbronchial cryobiopsy was performed with 1.1mm cryoprobe via the extended working channel catheter.
Freeze time of 6 seconds were used.  Total 6 samples were collected.  Samples sent for Microbiology (Cultures/Viral/Fungal) and Pathology.
Fiducial marker (0.8mm x 3mm soft tissue gold CIVCO) was loaded with bone wax and placed under fluoroscopy guidance.
Prior to withdraw of the bronchoscope. 
Transbronchial brushing was performed with Protected cytology brush the extended working channel catheter.
Total 1 samples were collected.  Samples sent for Cytology.
Bronchial alveolar lavage was performed the extended working channel catheter.
Instilled 20 cc of NS, suction returned with 5 cc of NS.  Samples sent for Microbiology (Cultures/Viral/Fungal).
ROSE from ION procedure was noted to be:
Atypical cells present but no evidence of malignant neoplasm
 
Successful therapeutic aspiration was performed to clean out the Trachea (Middle 1/3), Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus and blood.
Bronchial alveolar lavage was performed at Apical Segment of RUL (RB1), Posterior Segment of RUL (RB2), and Anterior Segment of RUL (RB3).
Instilled 40 cc of NS, suction returned with 15 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
EBUS-Findings
Indications: Diagnostic and Staging
Technique:
All lymph node stations were assessed.
Only those 5 mm or greater in short axis were sampled.
Lymph node sizing was performed by EBUS and sampling by transbronchial needle aspiration was performed using 22-gauge Needle.
Lymph Nodes/Sites Inspected: 4R (lower paratracheal) node
4L (lower paratracheal) node
7 (subcarinal) node
11Rs lymph node
11Ri lymph node
11L lymph node
No immediate complications
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
Elastography provided a semi-quantitative classification (Type 1–3), which was used to guide biopsy site selection and sampling strategy.
Lymph Nodes Evaluated:
Site 1: The 11L lymph node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was not sampled: Sampling this lymph node was not clinically indicated.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 2 elastographic pattern with mixed soft and stiff regions.
Given the size and ultrasound appearance, biopsies were not taken.
Site 2: The 4L (lower paratracheal) node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was not sampled: Sampling this lymph node was not clinically indicated.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 2 elastographic pattern with mixed soft and stiff regions.
Given the size and ultrasound appearance, biopsies were not taken.
Site 3: The 7 (subcarinal) node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was not sampled: Sampling this lymph node was not clinically indicated.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 2 elastographic pattern with mixed soft and stiff regions.
Given the size and ultrasound appearance, biopsies were not taken.
Site 4: The 4R (lower paratracheal) node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was not sampled: Sampling this lymph node was not clinically indicated.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 2 elastographic pattern with mixed soft and stiff regions.
Given the size and ultrasound appearance, biopsies were not taken.
Site 5: The 11Ri lymph node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was sampled. Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 2 elastographic pattern with mixed soft and stiff regions.
Given the size and ultrasound appearance, biopsies were not taken.
Site 6:  The 11Rs lymph node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was photographed. The site was sampled.. 4 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
The target lymph node demonstrated a Type 2 elastographic pattern with mixed soft and stiff regions.
Given this heterogeneous and indeterminate appearance, TBNA was directed at representative areas to ensure comprehensive sampling and to minimize the risk of underdiagnosis.
Successful therapeutic aspiration was performed to clean out the Trachea (Middle 1/3), Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus and blood.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
Right upper lobe transbronchial needle aspiration, transbronchial cryobiopsy, transbronchial brushing, lobar bronchoalveolar lavage
EBUS-TBNA 11Rs
IMPRESSION/PLAN: [REDACTED]is a 70 year old-year-old female who presents for bronchoscopy for lung nodule.
-Follow up bronchoscopic lab work
-Follow up CXR"""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# ==========================================
# EXTRACTION LOGIC
# ==========================================

# 1. Procedure Flags
# Default all to 0
PROCEDURE_FLAGS = {
    'diagnostic_bronchoscopy': 0, 'bal': 0, 'bronchial_wash': 0, 'brushings': 0,
    'endobronchial_biopsy': 0, 'tbna_conventional': 0, 'linear_ebus': 0, 'radial_ebus': 0,
    'navigational_bronchoscopy': 0, 'transbronchial_biopsy': 0, 'transbronchial_cryobiopsy': 0,
    'therapeutic_aspiration': 0, 'foreign_body_removal': 0, 'airway_dilation': 0,
    'airway_stent': 0, 'thermal_ablation': 0, 'tumor_debulking_non_thermal': 0,
    'cryotherapy': 0, 'blvr': 0, 'peripheral_ablation': 0, 'bronchial_thermoplasty': 0,
    'whole_lung_lavage': 0, 'rigid_bronchoscopy': 0, 'thoracentesis': 0, 'chest_tube': 0,
    'ipc': 0, 'medical_thoracoscopy': 0, 'pleurodesis': 0, 'pleural_biopsy': 0,
    'fibrinolytic_therapy': 0
}

# Update based on note content
PROCEDURE_FLAGS['diagnostic_bronchoscopy'] = 1
PROCEDURE_FLAGS['bal'] = 1 # "Bronchial alveolar lavage was performed"
PROCEDURE_FLAGS['brushings'] = 1 # "Transbronchial brushing was performed"
PROCEDURE_FLAGS['tbna_conventional'] = 1 # EBUS TBNA
PROCEDURE_FLAGS['linear_ebus'] = 1 # "Linear EBUS"
PROCEDURE_FLAGS['radial_ebus'] = 1 # "Radial EBUS"
PROCEDURE_FLAGS['navigational_bronchoscopy'] = 1 # "Robotic navigation bronchoscopy"
PROCEDURE_FLAGS['transbronchial_biopsy'] = 1 # Code 31628 explicitly listed
PROCEDURE_FLAGS['transbronchial_cryobiopsy'] = 1 # "required a Transbronchial Cryo biopsies"
PROCEDURE_FLAGS['therapeutic_aspiration'] = 1 # "Successful therapeutic aspiration was performed"

# 2. Spans and Events
# Event ID mapping:
# E01: Therapeutic Aspiration (Airway Cleaning)
# E02: Navigational Bronchoscopy (Nodule)
# E03: Radial EBUS (Nodule)
# E04: TBNA (Nodule)
# E05: Cryobiopsy (Nodule)
# E06: Fiducial Placement
# E07: Brushing
# E08: BAL
# E09: EBUS 11Rs (Sampled)
# E10: EBUS 11L (Inspected)
# E11: EBUS 4L (Inspected)
# E12: EBUS 7 (Inspected)
# E13: EBUS 4R (Inspected)
# E14: EBUS 11Ri (Inspected)

SPANS = []

def add_span(text, label, event_id, normalized="", context=""):
    SPANS.append({
        "text": text,
        "label": label,
        "event_id": event_id,
        "normalized": normalized,
        "context": context
    })

# E01: Aspiration
add_span("Successful therapeutic aspiration", "PROC_METHOD", "E01")
add_span("clean out", "PROC_ACTION", "E01")
add_span("Trachea (Middle 1/3)", "ANAT_AIRWAY", "E01")
add_span("Right Mainstem", "ANAT_AIRWAY", "E01")
add_span("Bronchus Intermedius", "ANAT_AIRWAY", "E01")
add_span("Left Mainstem", "ANAT_AIRWAY", "E01")
add_span("mucus and blood", "OBS_LESION", "E01")

# E02: Navigation
add_span("Robotic navigation bronchoscopy", "PROC_METHOD", "E02")
add_span("Ion platform", "DEV_INSTRUMENT", "E02")
add_span("Ion robotic catheter", "DEV_INSTRUMENT", "E02")
add_span("Anterior Segment of RUL (RB3)", "ANAT_LUNG_LOC", "E02")
add_span("0.6 cm", "MEAS_SIZE", "E02") # Nodule size

# E03: Radial EBUS
add_span("Radial EBUS", "PROC_METHOD", "E03")
add_span("nodule is not well visualized", "OBS_LESION", "E03")

# E04: TBNA Nodule
add_span("Transbronchial needle aspiration", "PROC_METHOD", "E04", context="21G Needle")
add_span("21G Needle", "DEV_NEEDLE", "E04", normalized="21G")
add_span("Total 5 samples", "MEAS_COUNT", "E04", normalized="5")

# E05: Cryo Nodule
add_span("Transbronchial cryobiopsy", "PROC_METHOD", "E05")
add_span("1.1mm cryoprobe", "DEV_INSTRUMENT", "E05")
add_span("Total 6 samples", "MEAS_COUNT", "E05", normalized="6")

# E06: Fiducial
add_span("Fiducial marker placements", "PROC_METHOD", "E06")
add_span("Fiducial marker", "DEV_INSTRUMENT", "E06")
add_span("0.8mm x 3mm", "MEAS_SIZE", "E06")

# E07: Brushing
add_span("Transbronchial brushing", "PROC_METHOD", "E07")
add_span("Protected cytology brush", "DEV_INSTRUMENT", "E07")
add_span("Total 1 samples", "MEAS_COUNT", "E07", normalized="1")

# E08: BAL
add_span("Bronchial alveolar lavage", "PROC_METHOD", "E08", context="Apical Segment")
add_span("Apical Segment of RUL (RB1)", "ANAT_LUNG_LOC", "E08")
add_span("Posterior Segment of RUL (RB2)", "ANAT_LUNG_LOC", "E08")
add_span("Instilled 40 cc", "MEAS_VOL", "E08", normalized="40")
add_span("returned with 15 cc", "MEAS_VOL", "E08", normalized="15")

# E09: EBUS 11Rs (Sampled)
add_span("11Rs lymph node", "ANAT_LN_STATION", "E09", normalized="11Rs")
add_span("endobronchial ultrasound guided transbronchial biopsies", "PROC_METHOD", "E09", normalized="EBUS-TBNA")
add_span("heterogeneous and indeterminate appearance", "OBS_LESION", "E09")

# E10: EBUS 11L
add_span("11L lymph node", "ANAT_LN_STATION", "E10", normalized="11L")
add_span("Endobronchial ultrasound (EBUS)", "PROC_METHOD", "E10", context="11L")
add_span("Type 2 elastographic pattern", "OBS_LESION", "E10")

# E11: EBUS 4L
add_span("4L (lower paratracheal) node", "ANAT_LN_STATION", "E11", normalized="4L")
# reusing procedure span difficult if exact same text, context needed
add_span("Endobronchial ultrasound (EBUS) elastography", "PROC_METHOD", "E11", context="Site 2")
add_span("Type 2 elastographic pattern", "OBS_LESION", "E11", context="Site 2")

# E12: EBUS 7
add_span("7 (subcarinal) node", "ANAT_LN_STATION", "E12", normalized="7")
add_span("Endobronchial ultrasound (EBUS) elastography", "PROC_METHOD", "E12", context="Site 3")
add_span("Type 2 elastographic pattern", "OBS_LESION", "E12", context="Site 3")

# E13: EBUS 4R
add_span("4R (lower paratracheal) node", "ANAT_LN_STATION", "E13", normalized="4R")
add_span("Endobronchial ultrasound (EBUS) elastography", "PROC_METHOD", "E13", context="Site 4")
add_span("Type 2 elastographic pattern", "OBS_LESION", "E13", context="Site 4")

# E14: EBUS 11Ri
add_span("11Ri lymph node", "ANAT_LN_STATION", "E14", normalized="11Ri")
add_span("Endobronchial ultrasound (EBUS) elastography", "PROC_METHOD", "E14", context="Site 5")
add_span("Type 2 elastographic pattern", "OBS_LESION", "E14", context="Site 5")

# Global Outcome
add_span("No immediate complications", "OUTCOME_COMPLICATION", "E00", normalized="None")

# ==========================================
# PROCESSING HELPERS
# ==========================================

def find_offsets_with_context(full_text, span_text, context_prefix=""):
    """
    Returns (start_char, end_char, hydration_status)
    Logic:
    1. Direct count check.
    2. Context window check (preceding 120 chars).
    3. Match index (not implemented here, defaulting to first or ambiguous).
    """
    if span_text not in full_text:
        return "", "", "not_found"
    
    matches = [m.start() for m in re.finditer(re.escape(span_text), full_text)]
    
    if len(matches) == 1:
        return matches[0], matches[0] + len(span_text), "hydrated_unique"
    
    if context_prefix:
        # Check context for each match
        for m_start in matches:
            window_start = max(0, m_start - 120)
            preceding_text = full_text[window_start:m_start]
            if context_prefix in preceding_text:
                return m_start, m_start + len(span_text), "hydrated_prefix_window"
    
    # Default to first match if ambiguous
    return matches[0], matches[0] + len(span_text), f"ambiguous_count={len(matches)}"

# ==========================================
# SCRIPT EXECUTION
# ==========================================

def generate_workbook():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        print(f"Error: Template {TEMPLATE_PATH} not found.")
        return

    # 1. Note_Text
    ws_text = wb["Note_Text"]
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws_index = wb["Note_Index"]
    # Row format: source_file, note_id, encounter, date, site, reviewer, status, notes, flags...
    row_meta = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Completed", ""]
    
    # Flag mapping must match sheet columns order strictly
    # Hardcoded order based on instructions
    flag_order = [
        'diagnostic_bronchoscopy', 'bal', 'bronchial_wash', 'brushings', 'endobronchial_biopsy',
        'tbna_conventional', 'linear_ebus', 'radial_ebus', 'navigational_bronchoscopy',
        'transbronchial_biopsy', 'transbronchial_cryobiopsy', 'therapeutic_aspiration',
        'foreign_body_removal', 'airway_dilation', 'airway_stent', 'thermal_ablation',
        'tumor_debulking_non_thermal', 'cryotherapy', 'blvr', 'peripheral_ablation',
        'bronchial_thermoplasty', 'whole_lung_lavage', 'rigid_bronchoscopy',
        'thoracentesis', 'chest_tube', 'ipc', 'medical_thoracoscopy', 'pleurodesis',
        'pleural_biopsy', 'fibrinolytic_therapy'
    ]
    
    flag_values = [PROCEDURE_FLAGS[f] for f in flag_order]
    ws_index.append(row_meta + flag_values)

    # 3. Span_Annotations & 4. Span_Hydrated
    ws_annot = wb["Span_Annotations"]
    ws_hydra = wb["Span_Hydrated"]
    
    # Define Span Columns: source, note, span_id, section, context, text, match_idx, start, end, len, label, norm, schema, event, negated, historical, time, reviewer, comments, status
    
    span_id_counter = 1
    
    hydrated_rows = []
    
    for s in SPANS:
        s_text = s["text"]
        start, end, status = find_offsets_with_context(NOTE_TEXT, s_text, s.get("context", ""))
        
        row_base = [
            SOURCE_FILE, NOTE_ID, f"span_{span_id_counter:03d}", "Procedure",
            s.get("context", ""), s_text, "", # match_index left blank
            "", "", f"=LEN(F{ws_annot.max_row+1})", # start/end blank in annot
            s["label"], s["normalized"], "", s["event_id"],
            "FALSE", "FALSE", "", "", "", "needs_hydration"
        ]
        ws_annot.append(row_base)
        
        # Hydrated row
        row_hydra = [
            SOURCE_FILE, NOTE_ID, f"span_{span_id_counter:03d}", "Procedure",
            s.get("context", ""), s_text, "",
            start, end, end-start if isinstance(start, int) else "",
            s["label"], s["normalized"], "", s["event_id"],
            "FALSE", "FALSE", "", "", "", status
        ]
        ws_hydra.append(row_hydra)
        span_id_counter += 1

    # 5. Event_Log
    ws_event = wb["Event_Log"]
    # Simple summary of events identified
    events_summary = {
        "E01": {"type": "Therapeutic Aspiration", "method": "Suction", "anatomy": "Airways"},
        "E02": {"type": "Navigational Bronchoscopy", "method": "Ion", "anatomy": "RUL"},
        "E03": {"type": "Radial EBUS", "method": "Radial EBUS", "anatomy": "RUL Nodule"},
        "E04": {"type": "TBNA", "method": "Needle Aspiration", "device": "21G Needle", "anatomy": "RUL Nodule"},
        "E05": {"type": "Cryobiopsy", "method": "Cryoprobe", "device": "1.1mm probe", "anatomy": "RUL Nodule"},
        "E06": {"type": "Fiducial", "method": "Placement", "device": "Marker", "anatomy": "RUL"},
        "E07": {"type": "Brushing", "method": "Brush", "anatomy": "RUL Nodule"},
        "E08": {"type": "BAL", "method": "Lavage", "anatomy": "RUL Segments"},
        "E09": {"type": "EBUS-TBNA", "method": "TBNA", "stations": "11Rs"},
        "E10": {"type": "EBUS Inspection", "method": "EBUS", "stations": "11L"},
        "E11": {"type": "EBUS Inspection", "method": "EBUS", "stations": "4L"},
        "E12": {"type": "EBUS Inspection", "method": "EBUS", "stations": "7"},
        "E13": {"type": "EBUS Inspection", "method": "EBUS", "stations": "4R"},
        "E14": {"type": "EBUS Inspection", "method": "EBUS", "stations": "11Ri"},
    }
    
    for eid, info in events_summary.items():
        ws_event.append([
            SOURCE_FILE, NOTE_ID, eid, info["type"], info["method"], 
            info.get("anatomy", ""), info.get("device", ""), "", info.get("stations", ""),
            "", "", "", "", "FALSE", "", "", "", "", "", "", "", "", "None" if eid == "E00" else ""
        ])

    # 6. V3_Procedure_Events & 7. V3_Registry_JSON
    ws_v3 = wb["V3_Procedure_Events"]
    ws_json = wb["V3_Registry_JSON"]
    
    # Construct JSON structure
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": []
    }
    
    # Map events to V3 structure (simplified for generation)
    for eid, info in events_summary.items():
        v3_row = [
            NOTE_ID, eid, info["type"], 
            "Lung" if "RUL" in info.get("anatomy","") else "Lymph Node" if "EBUS" in info["type"] else "Airway",
            "RUL" if "RUL" in info.get("anatomy","") else "",
            "", # segment
            info.get("stations", ""),
            "Nodule" if "Nodule" in info.get("anatomy","") else "",
            "", # lesion size
            info["method"],
            "", "", "", "", "", # json fields
            "", "", "", # stent
            "", "", "", "", "None" # outcomes
        ]
        ws_v3.append(v3_row)
        
        registry_data["procedures"].append({
            "event_id": eid,
            "type": info["type"],
            "method": info["method"],
            "target": {"anatomy": info.get("anatomy", "")}
        })

    ws_json.cell(row=1, column=1, value=json.dumps(registry_data, indent=2))

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Successfully generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()