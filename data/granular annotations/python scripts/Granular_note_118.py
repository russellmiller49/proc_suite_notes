import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os
from datetime import datetime

# =============================================================================
# 1. INPUT DATA
# =============================================================================

NOTE_ID = "note_118"
SOURCE_FILE = "note_118.txt"
PROCEDURE_DATE = "2026-01-12" # inferred context or leave blank
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_118 SOURCE_FILE: note_118.txt INDICATION FOR OPERATION:  [REDACTED]is a 50 year old-year-old female who presents with complex airway narrowing due to tumor.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: J98.09 Other diseases of bronchus, not elsewhere classified
POSTOPERATIVE DIAGNOSIS:  J98.09 Other diseases of bronchus, not elsewhere classified
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31622 Dx bronchoscope/cell washing          
31624 Dx bronchoscope/lavage (BAL)    
31625 Endobronchial Biopsy(s)
31636 Dilate and bronchial stent initial bronchus
31637 Dilate and bronchial stent additional bronchus
31638 Revision of tracheal/bronchial stent     
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy)
31635 Foreign body removal
31637 - total of 3 bronchial stents were placed
22 Substantially greater work 
than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required a Pt required placement of Y stent in the right mainstem bronchus, right upper lobe and BI due to extensive tumor infiltration and .
This resulted in >100% increased work due to Increased intensity, Time, Technical difficulty of procedure, Severity of patient's condition, and Physical and mental effort required.
Apply to: 31630 Balloon dilation
31636 Dilate and bronchial stent initial bronchus
31640 Bronchoscopy with excision 
31641 Destruction of tumor OR relief of stenosis by any method other than excision (eg. laser therapy, cryotherapy).
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Rigid Bronchoscope
Flexible Therapeutic Bronchoscope
Flexible Hybrid (Pedatric) Bronchoscope
ESTIMATED BLOOD LOSS:   Minimum
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: supine
Initial Airway Inspection Findings:
After placement of iGel by anesthesia the bronchoscope was introduced into the iGel.
A total of 6 mL of lidocaine was instilled onto the vocal cords.
The bronchoscope was advanced through the vocal cord into the trachea.
The airways were examined to the subsegmental level bilaterally.
The stent in the RMSB was visualized with tumor growing through the opening for the RUL.
Successful therapeutic aspiration was performed to clean out the Trachea (Middle 1/3), Trachea (Distal 1/3), Right Mainstem, and Bronchus Intermedius  from mucus.
The stent was in appropriate position. Cryoprobe and forceps were used for endobronchial biopsies of the tumor herniating through the stent.
There was a small amount of oozing controlled with APC.  Inspection showed residual tumor obstructing the stent opening.
The bronchoscope was able to advanced along the posterior portion of the stent between the stent and airway wall.
There was visible endobronchial tumor causing mixed obstruction and blocking the RUL ostium.
Endobronchial obstruction at Right Mainstem and RUL ostium was treated with the following modalities:
Modality	Tools	Setting/Mode	Duration	Results
Electrocautery	 	 	 	 
APC	1.1 mm	Intermittent through the stent and along the outside of the stent to free the RUL	 	 
Laser	 	 	 	 
CoreCath	 	 	 	 
Cryoprobe	1.7 mm	Intermittent with removal of the tumor herniating through the stent opening	 	 
Prior to treatment, affected airway was note to be 0% patent.
After treatment, the airway was 25% patent. The narrowing was due to extrinsic compression from the posterior membrane.
The RUL was patent and a small amount of mucus was behind the obstruction.
Bronchial alveolar lavage was performed at Anterior Segment of RUL (RB3).
Instilled 40 cc of NS, suction returned with 20 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
The fenestration of the stent and the RUL did not align and the decision was made to remove the stent/foreign body.
The iGel was removed and the patient was atraumatically intubated with the black rigid bronchoscope.
The proximal portion of the stent was grasped and two attempts were made to adjust the stent to align the fenestration and the RUL ostium.
Both were unsuccessful and the proximal end of the stent was grasped and rotated.
The stent and rigid bronchoscope were removed. No residual material was left in the patient.
The tumor could not be cored and measurements were taken for Y stent - to allow for stenting of the RMSB, RUL and BI.
Under apnea the stent was deployed into the RMSB. Following deployment the stent was in the BI and required revision but pulling the stent up into the distal trachea, rotating 60 degrees.
Using the p190 scope as stylet into the RUL the stent was advanced into BI.
The RUL limb was angled downward but using forceps (rigid and flexible) the stent was revised onto the RC1 with good position.
Balloon dilation was performed at RUL.  10/11/12 Elation balloon was used to perform dilation to 10 mm at the RUL bronchus.
Total 1 inflations with dilation time of 10 seconds each. This resulted in seating of the stent in good position.
Following placement the stent was in good position and all airways were patent. 
The patient tolerated the procedure well.
There were no immediate complications.  At the conclusion of the operation, the patient was successfully extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
BAL RUL
EBBx RMSB tumor
IMPRESSION/PLAN: [REDACTED]is a 50 year old-year-old female who presents for bronchoscopy for stent evaluation.
[ ] continue stent hygeing
[ ] f/u bronchoscopy in 1-2 weeks"""

# =============================================================================
# 2. DEFINITIONS & SCHEMA
# =============================================================================

# Procedure Flags (Sheet: Note_Index)
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 1,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 1,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 1, # Stent removal
    "airway_dilation": 1,
    "airway_stent": 1,
    "thermal_ablation": 1, # APC
    "tumor_debulking_non_thermal": 0, 
    "cryotherapy": 1,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 1,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# Spans (Sheet: Span_Annotations)
# Fields: span_text, label, normalized_value, event_id, context_prefix
SPANS = [
    # Event 1: Therapeutic Aspiration
    ("Therapeutic aspiration", "PROC_METHOD", "therapeutic_aspiration", "evt_01", "Successful"),
    ("Trachea (Middle 1/3)", "ANAT_AIRWAY", "Trachea", "evt_01", "clean out the"),
    ("Right Mainstem", "ANAT_AIRWAY", "RMSB", "evt_01", "Trachea (Distal 1/3),"),
    ("Bronchus Intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "evt_01", "Right Mainstem, and"),
    ("mucus", "OBS_LESION", "secretions", "evt_01", None),

    # Event 2: Biopsy (Tumor)
    ("endobronchial biopsies", "PROC_METHOD", "endobronchial_biopsy", "evt_02", None),
    ("Cryoprobe", "DEV_INSTRUMENT", "Cryoprobe", "evt_02", "appropriate position."),
    ("forceps", "DEV_INSTRUMENT", "forceps", "evt_02", "Cryoprobe and"),
    ("tumor", "OBS_LESION", "tumor", "evt_02", "biopsies of the"),
    ("herniating through the stent", "OBS_LESION", "tumor_herniation", "evt_02", None),

    # Event 3: Tumor Destruction/Recanalization (APC + Cryo) + Outcome
    ("APC", "DEV_INSTRUMENT", "APC", "evt_03", "controlled with"),
    ("removal of the tumor", "PROC_ACTION", "debulking", "evt_03", None),
    ("0% patent", "OUTCOME_AIRWAY_LUMEN_PRE", "0", "evt_03", None),
    ("25% patent", "OUTCOME_AIRWAY_LUMEN_POST", "25", "evt_03", None),
    ("extrinsic compression", "OBS_LESION", "extrinsic_compression", "evt_03", None),

    # Event 4: BAL
    ("Bronchial alveolar lavage", "PROC_METHOD", "bal", "evt_04", None),
    ("Anterior Segment of RUL (RB3)", "ANAT_AIRWAY", "RB3", "evt_04", None),
    ("Instilled 40 cc", "MEAS_VOL", "40", "evt_04", None),
    ("returned with 20 cc", "MEAS_VOL", "20", "evt_04", None),

    # Event 5: Stent Removal (Foreign Body)
    ("remove the stent/foreign body", "PROC_ACTION", "remove", "evt_05", None),
    ("rigid bronchoscope", "DEV_INSTRUMENT", "rigid_bronchoscope", "evt_05", "intubated with the black"),
    ("proximal end of the stent", "ANAT_AIRWAY", "stent", "evt_05", None),
    ("removed", "PROC_ACTION", "remove", "evt_05", "rigid bronchoscope were"),

    # Event 6: Y Stent Placement
    ("Y stent", "DEV_STENT_MATERIAL", "Y-stent", "evt_06", "measurements were taken for"),
    ("stent was deployed", "PROC_ACTION", "deploy", "evt_06", "Under apnea the"),
    ("RMSB", "ANAT_AIRWAY", "RMSB", "evt_06", "deployed into the"),
    ("BI", "ANAT_AIRWAY", "Bronchus Intermedius", "evt_06", "stent was in the"),
    ("p190 scope", "DEV_INSTRUMENT", "p190_scope", "evt_06", None),
    ("forceps", "DEV_INSTRUMENT", "forceps", "evt_06", "but using"),
    ("RUL", "ANAT_AIRWAY", "RUL", "evt_06", "stylet into the"),

    # Event 7: Balloon Dilation
    ("Balloon dilation", "PROC_METHOD", "balloon_dilation", "evt_07", None),
    ("RUL", "ANAT_AIRWAY", "RUL", "evt_07", "performed at"),
    ("10/11/12 Elation balloon", "DEV_INSTRUMENT", "Elation_Balloon_10_11_12", "evt_07", None),
    ("dilation to 10 mm", "MEAS_AIRWAY_DIAM", "10", "evt_07", None),
    ("10 seconds", "CTX_TIME", "10 seconds", "evt_07", None),
    
    # Global/Complications
    ("no immediate complications", "OUTCOME_COMPLICATION", "none", "evt_00", None),
]

# Event Log Data (Sheet: Event_Log)
EVENTS = [
    {
        "event_id": "evt_01", "event_type": "Therapeutic Aspiration", "method": "Aspiration",
        "anatomy_target": "Trachea, RMSB, BI", "findings": "Mucus cleaned out"
    },
    {
        "event_id": "evt_02", "event_type": "Endobronchial Biopsy", "method": "Cryoprobe, Forceps",
        "anatomy_target": "RMSB (Tumor via stent)", "findings": "Tumor herniating through stent",
        "specimens": "EBBx RMSB tumor"
    },
    {
        "event_id": "evt_03", "event_type": "Tumor Destruction", "method": "APC, Cryoprobe",
        "anatomy_target": "RMSB/RUL Ostium", "outcome_airway_lumen_pre": "0%", "outcome_airway_lumen_post": "25%"
    },
    {
        "event_id": "evt_04", "event_type": "BAL", "method": "Lavage",
        "anatomy_target": "Anterior Segment RUL (RB3)", "measurements": "Instilled 40cc, Return 20cc",
        "specimens": "BAL RUL"
    },
    {
        "event_id": "evt_05", "event_type": "Foreign Body Removal", "method": "Rigid Bronchoscopy, Forceps",
        "anatomy_target": "Stent (Proximal)", "findings": "Misaligned fenestration"
    },
    {
        "event_id": "evt_06", "event_type": "Airway Stent", "method": "Deployment, Revision",
        "anatomy_target": "RMSB, RUL, BI", "device": "Y Stent", "device_material": "Y-Stent"
    },
    {
        "event_id": "evt_07", "event_type": "Airway Dilation", "method": "Balloon",
        "anatomy_target": "RUL", "device": "10/11/12 Elation balloon", "measurements": "10mm",
        "outcome_airway_lumen_post": "Patent"
    }
]

# V3 Procedure Events (Sheet: V3_Procedure_Events)
V3_EVENTS = [
    {
        "note_id": NOTE_ID, "event_id": "evt_01", "type": "Therapeutic Aspiration",
        "target.anatomy_type": "Central Airway", "method": "Suction",
        "findings_json": json.dumps(["Mucus"]), "evidence_quote": "Successful therapeutic aspiration was performed to clean out the Trachea"
    },
    {
        "note_id": NOTE_ID, "event_id": "evt_02", "type": "Endobronchial Biopsy",
        "target.anatomy_type": "Bronchus", "target.location.lobe": "RUL", "method": "Cryoprobe",
        "specimens_json": json.dumps(["EBBx"]), "evidence_quote": "Cryoprobe and forceps were used for endobronchial biopsies"
    },
    {
        "note_id": NOTE_ID, "event_id": "evt_03", "type": "Tumor Destruction",
        "target.anatomy_type": "Bronchus", "method": "APC/Cryo",
        "outcomes.airway.lumen_pre": "0", "outcomes.airway.lumen_post": "25",
        "evidence_quote": "Prior to treatment, affected airway was note to be 0% patent. After treatment, the airway was 25% patent."
    },
    {
        "note_id": NOTE_ID, "event_id": "evt_04", "type": "BAL",
        "target.anatomy_type": "Bronchus", "target.location.segment": "RB3",
        "method": "Lavage", "measurements_json": json.dumps({"instilled": "40cc", "return": "20cc"})
    },
    {
        "note_id": NOTE_ID, "event_id": "evt_05", "type": "Foreign Body Removal",
        "target.anatomy_type": "Stent", "method": "Rigid/Forceps",
        "evidence_quote": "decision was made to remove the stent/foreign body"
    },
    {
        "note_id": NOTE_ID, "event_id": "evt_06", "type": "Airway Stent",
        "target.anatomy_type": "Bronchus", "stent.material_or_brand": "Y Stent",
        "evidence_quote": "stent was deployed into the RMSB"
    },
    {
        "note_id": NOTE_ID, "event_id": "evt_07", "type": "Airway Dilation",
        "target.anatomy_type": "Bronchus", "method": "Balloon",
        "devices_json": json.dumps(["Elation Balloon"]), "measurements_json": json.dumps({"diameter": "10mm"}),
        "evidence_quote": "dilation to 10 mm at the RUL bronchus"
    }
]

# =============================================================================
# 3. HELPER FUNCTIONS
# =============================================================================

def create_workbook_from_template():
    if not os.path.exists(TEMPLATE_PATH):
        # Create a dummy workbook if template is missing (fallback for environment)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Note_Text"
        wb.create_sheet("Note_Index")
        wb.create_sheet("Span_Annotations")
        wb.create_sheet("Span_Hydrated")
        wb.create_sheet("Event_Log")
        wb.create_sheet("V3_Procedure_Events")
        wb.create_sheet("V3_Registry_JSON")
        return wb
    return openpyxl.load_workbook(TEMPLATE_PATH)

def hydrate_span(text, span_text, context_prefix=None):
    """
    Finds start/end chars for span_text in text.
    Returns: (start_char, end_char, hydration_status)
    """
    matches = list(re.finditer(re.escape(span_text), text))
    if not matches:
        return None, None, "not_found"
    
    if len(matches) == 1:
        m = matches[0]
        return m.start(), m.end(), "hydrated_unique"
    
    if context_prefix:
        # Search for context in a 150-char window before the match
        for m in matches:
            start_window = max(0, m.start() - 150)
            window_text = text[start_window:m.start()]
            if context_prefix in window_text:
                return m.start(), m.end(), "hydrated_context"
    
    # Default to first if ambiguous and no context matched (or just pick first)
    m = matches[0]
    return m.start(), m.end(), f"ambiguous_count={len(matches)}"

# =============================================================================
# 4. EXECUTION
# =============================================================================

def main():
    wb = create_workbook_from_template()
    
    # -------------------------------------------------------------------------
    # Sheet: Note_Text
    # -------------------------------------------------------------------------
    ws = wb["Note_Text"]
    # Check headers
    if ws.max_row == 1 and ws.cell(1, 1).value is None:
        ws.append(["note_id", "source_file", "note_text"])
    ws.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # -------------------------------------------------------------------------
    # Sheet: Note_Index
    # -------------------------------------------------------------------------
    ws = wb["Note_Index"]
    # Headers logic
    headers = ["source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes"]
    flag_keys = list(PROCEDURE_FLAGS.keys())
    headers.extend(flag_keys)
    
    if ws.max_row == 1 and ws.cell(1, 1).value is None:
        ws.append(headers)
    
    row_data = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "Auto", "Draft", ""]
    row_data.extend([PROCEDURE_FLAGS[k] for k in flag_keys])
    ws.append(row_data)
    
    # -------------------------------------------------------------------------
    # Sheet: Span_Annotations (Anchor First)
    # -------------------------------------------------------------------------
    ws_ann = wb["Span_Annotations"]
    ann_headers = ["source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", 
                   "match_index", "start_char", "end_char", "span_len", "label", "normalized_value", 
                   "schema_field", "event_id", "is_negated", "is_historical", "time_anchor", 
                   "reviewer", "comments", "hydration_status"]
    if ws_ann.max_row == 1 and ws_ann.cell(1, 1).value is None:
        ws_ann.append(ann_headers)
        
    # -------------------------------------------------------------------------
    # Sheet: Span_Hydrated
    # -------------------------------------------------------------------------
    ws_hyd = wb["Span_Hydrated"]
    if ws_hyd.max_row == 1 and ws_hyd.cell(1, 1).value is None:
        ws_hyd.append(ann_headers)
    
    span_counter = 1
    for span_tuple in SPANS:
        s_text, s_label, s_norm, s_evt, s_ctx = span_tuple
        
        # Hydrate
        start, end, status = hydrate_span(NOTE_TEXT, s_text, s_ctx)
        
        span_id = f"{NOTE_ID}_span_{span_counter:03d}"
        span_len = len(s_text)
        
        # Base Row (Anchor) - start/end blank
        row_base = [SOURCE_FILE, NOTE_ID, span_id, "Procedure", s_ctx, s_text, 
                    1, "", "", span_len, s_label, s_norm, 
                    "", s_evt, "FALSE", "FALSE", "", 
                    "Auto", "", "needs_hydration"]
        ws_ann.append(row_base)
        
        # Hydrated Row
        row_hyd = [SOURCE_FILE, NOTE_ID, span_id, "Procedure", s_ctx, s_text, 
                   1, start, end, span_len, s_label, s_norm, 
                   "", s_evt, "FALSE", "FALSE", "", 
                   "Auto", "", status]
        ws_hyd.append(row_hyd)
        
        span_counter += 1

    # -------------------------------------------------------------------------
    # Sheet: Event_Log
    # -------------------------------------------------------------------------
    ws_evt = wb["Event_Log"]
    evt_headers = ["source_file", "note_id", "event_id", "event_type", "method", 
                   "anatomy_target", "device", "needle_gauge", "stations", "counts", 
                   "measurements", "specimens", "findings", "is_historical", "reviewer", "comments",
                   "device_size", "device_material", "outcome_airway_lumen_pre", 
                   "outcome_airway_lumen_post", "outcome_symptoms", "outcome_pleural", "outcome_complication"]
                   
    if ws_evt.max_row == 1 and ws_evt.cell(1, 1).value is None:
        ws_evt.append(evt_headers)
        
    for evt in EVENTS:
        row = [
            SOURCE_FILE, NOTE_ID, evt.get("event_id"), evt.get("event_type"), evt.get("method"),
            evt.get("anatomy_target"), evt.get("device"), "", "", "",
            evt.get("measurements"), evt.get("specimens"), evt.get("findings"), "FALSE", "Auto", "",
            evt.get("device_size"), evt.get("device_material"), evt.get("outcome_airway_lumen_pre"),
            evt.get("outcome_airway_lumen_post"), "", "", evt.get("outcome_complication")
        ]
        ws_evt.append(row)

    # -------------------------------------------------------------------------
    # Sheet: V3_Procedure_Events
    # -------------------------------------------------------------------------
    ws_v3 = wb["V3_Procedure_Events"]
    v3_headers = ["note_id", "event_id", "type", "target.anatomy_type", "target.location.lobe", 
                  "target.location.segment", "target.station", "lesion.type", "lesion.size_mm", 
                  "method", "devices_json", "measurements_json", "specimens_json", "findings_json", 
                  "evidence_quote", "stent.size", "stent.material_or_brand", "catheter.size_fr",
                  "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post", "outcomes.symptoms", 
                  "outcomes.pleural", "outcomes.complications"]
                  
    if ws_v3.max_row == 1 and ws_v3.cell(1, 1).value is None:
        ws_v3.append(v3_headers)
        
    for v3 in V3_EVENTS:
        row = [
            v3.get("note_id"), v3.get("event_id"), v3.get("type"), v3.get("target.anatomy_type"), 
            v3.get("target.location.lobe"), v3.get("target.location.segment"), v3.get("target.station"),
            v3.get("lesion.type"), v3.get("lesion.size_mm"), v3.get("method"), v3.get("devices_json"),
            v3.get("measurements_json"), v3.get("specimens_json"), v3.get("findings_json"), 
            v3.get("evidence_quote"), v3.get("stent.size"), v3.get("stent.material_or_brand"),
            v3.get("catheter.size_fr"), v3.get("outcomes.airway.lumen_pre"), 
            v3.get("outcomes.airway.lumen_post"), v3.get("outcomes.symptoms"), 
            v3.get("outcomes.pleural"), v3.get("outcomes.complications")
        ]
        ws_v3.append(row)

    # -------------------------------------------------------------------------
    # Sheet: V3_Registry_JSON
    # -------------------------------------------------------------------------
    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 1 and ws_json.cell(1, 1).value is None:
        ws_json.append(["schema_version", "note_id", "json_output"])
    
    registry_obj = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": V3_EVENTS,
        "no_immediate_complications": True
    }
    ws_json.append(["3.0", NOTE_ID, json.dumps(registry_obj, indent=2)])
    
    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    main()