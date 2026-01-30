import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os
import datetime

# =============================================================================
# INPUT DATA & CONSTANTS
# =============================================================================

NOTE_ID = "note_119"
SOURCE_FILE = "note_119.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# The full content of the note as provided
NOTE_TEXT = """NOTE_ID:  note_119 SOURCE_FILE: note_119.txt INDICATION FOR OPERATION:  [REDACTED]is a 48 year old-year-old female who presents with lung nodule.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: R91.1 Solitary Lung Nodule
POSTOPERATIVE DIAGNOSIS:  R91.1 Solitary Lung Nodule
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31622 Dx bronchoscope/cell washing          
31623 Dx bronchoscope/brushing    
31624 Dx bronchoscope/lavage (BAL)    
31628 TBBX single lobe     
31629 TBNA single lobe   
31626 Fiducial marker placements, single or multiple     
31627 Navigational Bronchoscopy (computer assisted)
77012 Radiology / radiologic guidance for CT guided needle placement (CIOS)
76377 3D rendering with interpretation and reporting of CT, US, Tomo modality (ION Planning Station)
31652 EBUS sampling 1 
or 2 nodes
31654 Radial EBUS for peripheral lesion
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required a Transbronchial Cryo biopsies.
This resulted in >40% increased work due to Increased intensity, Time, Technical difficulty of procedure, and Physical and mental effort required.
Apply to: 31628 TBBX single lobe
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Linear EBUS 
Radial EBUS
Ion Robotic Bronchoscope
Disposable Bronchoscope
ESTIMATED BLOOD LOSS:   Minimum
COMPLICATIONS:    None
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: . 
Initial Airway Inspection Findings:
Normal appearing airway anatomy and mucosa bilaterally to the segmental level.
CT Chest scan was placed on separate planning station to generate 3D rendering of the pathway to target.
The navigational plan was reviewed and verified.  This was then loaded into robotic bronchoscopy platform.
Successful therapeutic aspiration was performed to clean out the Trachea (Middle 1/3), Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus.
Ventilation Parameters:
Mode	RR	TV	PEEP	FiO2	Flow Rate	Pmean
VCV	 	300	12	100	10	15
Robotic navigation bronchoscopy was performed with Ion platform.  Partial registration was used.
Ion robotic catheter was used to engage the Anteromedial Segment of LLL (Lb7/8).
Target lesion is about 1 cm in diameter.   Under navigational guidance the ion robotic catheter was advanced to 1.0 cm away from the planned target.
Radial EBUS was performed to confirm that the location of the nodule is not well visualized.
The following features were noted: aerated lung.
Cone Beam CT was performed: 3-D reconstructions were performed on an independent workstation.
Cios Spin system was used for evaluation of nodule location.  Low dose spin was performed to acquire CT imaging.
This was passed on to Ion platform system for reconstruction and nodule location.
The 3D images was interpreted on an independent workstation (Ion).
Using the newly acquired nodule location, the Ion robotic system was adjusted to the new targeted location.
I personally interpreted the cone beam CT and 3-D reconstruction.
Transbronchial needle aspiration was performed with 21G Needle through the extended working channel catheter.  Total 4 samples were collected.
Samples sent for Microbiology (Cultures/Viral/Fungal) and Cytology.
Transbronchial cryobiopsy was performed with 1.1mm cryoprobe via the extended working channel catheter.
Freeze time of 6 seconds were used.  Total 6 samples were collected.  Samples sent for Microbiology (Cultures/Viral/Fungal) and Pathology.
Transbronchial brushing was performed with Protected cytology brush the extended working channel catheter.  Total 1 samples were collected.
Samples sent for Microbiology (Cultures/Viral/Fungal) and Cytology.
Bronchial alveolar lavage was performed the extended working channel catheter.
Instilled 20 cc of NS, suction returned with 5 cc of NS.  Samples sent for Microbiology (Cultures/Viral/Fungal) and Cytology.
Fiducial marker (0.8mm x 3mm soft tissue gold CIVCO) was loaded with bone wax and placed under fluoroscopy guidance.
Prior to withdraw of the bronchoscope. 
ROSE from ION procedure was noted to be:
No evidence of a malignant neoplasm
Successful therapeutic aspiration was performed to clean out the Trachea (Middle 1/3), Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus and blood.
Bronchial alveolar lavage was performed at Superior Segment of LLL (LB6), Anteromedial Segment of LLL (Lb7/8), Lateral-basal Segment of LLL (LB9), and Posterior-Basal Segment of LLL (LB10).
Instilled 40 cc of NS, suction returned with 15 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
EBUS-Findings
Indications: Diagnostic and Staging
Technique:
All lymph node stations were assessed.
Only those 5 mm or greater in short axis were sampled.
Lymph node sizing was performed by EBUS and sampling by transbronchial needle aspiration was performed using 22-gauge Needle.
Lymph Nodes/Sites Inspected: 4R (lower paratracheal) node
4L (lower paratracheal) node
7 (subcarinal) node
11Rs lymph node
11Ri lymph node
11L lymph node
No immediate complications
Endobronchial ultrasound (EBUS) elastography was performed to assess lymph node stiffness and tissue characteristics.
Elastography provided a semi-quantitative classification (Type 1–3), which was used to guide biopsy site selection and sampling strategy.
Lymph Nodes Evaluated:
Site 1: The 11L lymph node was < 10 mm on CT  and Metabolic activity unknown or PET-CT scan unavailable.
The lymph node was not photographed. The site was sampled.. 4 endobronchial ultrasound guided transbronchial biopsies were performed with samples obtained.
Successful therapeutic aspiration was performed to clean out the Trachea (Middle 1/3), Trachea (Distal 1/3), Right Mainstem, Bronchus Intermedius , Left Mainstem, Carina, RUL Carina (RC1), RML Carina (RC2), LUL Lingula Carina (Lc1), and Left Carina (LC2) from mucus and blood.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was extubated in the operating room and transported to the recovery room in stable condition.
SPECIMEN(S): 
•	left lower lobe transbronchial needle aspiration, transbronchial cryobiopsy, transbronchial brushing, sub segmental bronchoalveolar lavage, lobar bronchoalveolar lavage
•	EBUS-TBNA 11L
IMPRESSION/PLAN: [REDACTED]is a 48 year old-year-old female who presents for bronchoscopy for lung nodule.
-Follow up bronchoscopic lab work
-Follow up CXR"""

# =============================================================================
# FILE UPLOAD HANDLING
# =============================================================================
# Attempt to handle file upload if script is run in an environment that puts the file in CWD
if os.path.exists(SOURCE_FILE):
    try:
        with open(SOURCE_FILE, "r", encoding="utf-8") as f:
            content = f.read()
            # If content looks plausible, use it, otherwise stick to embedded
            if len(content) > 100:
                NOTE_TEXT = content
                print(f"Loaded {SOURCE_FILE} from disk.")
    except Exception as e:
        print(f"Could not read {SOURCE_FILE}, using embedded text. Error: {e}")

# =============================================================================
# HELPER CLASSES & FUNCTIONS
# =============================================================================

class WorkbookGenerator:
    def __init__(self, template_path):
        try:
            self.wb = openpyxl.load_workbook(template_path)
        except FileNotFoundError:
            self.wb = openpyxl.Workbook() # Fallback for demo if template missing
            print("Warning: Template not found. Creating blank workbook.")
        self.note_text = NOTE_TEXT
        self.spans = []
        self.events = []
        
    def get_sheet(self, name):
        if name in self.wb.sheetnames:
            return self.wb[name]
        return self.wb.create_sheet(name)

    def add_span(self, span_text, label, normalized_value, schema_field, event_id, 
                 context_prefix="", match_index=0, is_historical=False, comments=""):
        
        # Calculate offsets
        start_char, end_char, status = self.calculate_offsets(span_text, context_prefix, match_index)
        
        span_row = {
            "source_file": SOURCE_FILE,
            "note_id": NOTE_ID,
            "span_id": f"span_{len(self.spans) + 1:03d}",
            "section_type": "Procedure", # Simplified
            "context_prefix": context_prefix,
            "span_text": span_text,
            "match_index": match_index if match_index > 0 else None,
            "start_char": start_char,
            "end_char": end_char,
            "span_len": len(span_text),
            "label": label,
            "normalized_value": normalized_value,
            "schema_field": schema_field,
            "event_id": event_id,
            "is_negated": False,
            "is_historical": is_historical,
            "time_anchor": "",
            "reviewer": "Auto",
            "comments": comments,
            "hydration_status": status
        }
        self.spans.append(span_row)

    def calculate_offsets(self, span_text, context_prefix, match_index):
        text = self.note_text
        
        if not span_text:
            return None, None, "error_empty_span"

        # Strategy 1: Context Prefix
        if context_prefix:
            # Find context first
            c_start = text.find(context_prefix)
            if c_start != -1:
                # Look for span after context (within reasonable window)
                search_start = c_start + len(context_prefix)
                s_start = text.find(span_text, search_start)
                if s_start != -1 and (s_start - search_start) < 200:
                    return s_start, s_start + len(span_text), "hydrated_prefix_window"
        
        # Strategy 2: Match Index
        count = text.count(span_text)
        if count == 0:
            return None, None, "not_found"
        
        if count == 1:
            start = text.find(span_text)
            return start, start + len(span_text), "hydrated_unique"
        
        # Multiple occurrences
        if match_index > 0 and match_index <= count:
            # Find the Nth occurrence
            current = -1
            for _ in range(match_index):
                current = text.find(span_text, current + 1)
            return current, current + len(span_text), "hydrated_match_index"
        
        # Default if ambiguous
        return None, None, f"ambiguous_count={count}"

    def write_note_text(self):
        ws = self.get_sheet("Note_Text")
        # Ensure headers if new
        if ws.max_row == 1 and ws.cell(1,1).value is None:
            ws.append(["note_id", "source_file", "note_text"])
        
        # Check if already exists to avoid duplicates
        exists = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == NOTE_ID:
                exists = True
                break
        
        if not exists:
            # Clean text for Excel (remove illegal chars)
            clean_text = "".join(c for c in self.note_text if c.isprintable() or c in ['\n', '\r', '\t'])
            ws.append([NOTE_ID, SOURCE_FILE, clean_text])

    def write_note_index(self):
        ws = self.get_sheet("Note_Index")
        headers = [
            "source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes",
            "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", 
            "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy", 
            "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration", 
            "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation", 
            "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", 
            "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
            "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis", "pleural_biopsy", "fibrinolytic_therapy"
        ]
        
        # Create headers if missing
        if ws.max_row == 1 and ws.cell(1,1).value != "source_file":
            ws.append(headers)

        # Flag Logic
        flags = {k: 0 for k in headers[8:]}
        
        flags['diagnostic_bronchoscopy'] = 1
        flags['therapeutic_aspiration'] = 1
        flags['navigational_bronchoscopy'] = 1
        flags['radial_ebus'] = 1
        flags['transbronchial_biopsy'] = 1 # Cryo and needle
        flags['transbronchial_cryobiopsy'] = 1
        flags['brushings'] = 1
        flags['bal'] = 1
        flags['linear_ebus'] = 1

        row_data = [SOURCE_FILE, NOTE_ID, "", "", "", "Auto", "Pending", "Generated by Script"]
        for h in headers[8:]:
            row_data.append(flags[h])
        
        ws.append(row_data)

    def write_spans(self):
        ws_anchor = self.get_sheet("Span_Annotations")
        ws_hydrated = self.get_sheet("Span_Hydrated")
        
        headers = [
            "source_file", "note_id", "span_id", "section_type",
            "context_prefix", "span_text", "match_index",
            "start_char", "end_char", "span_len",
            "label", "normalized_value", "schema_field", "event_id",
            "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"
        ]

        if ws_anchor.max_row == 1: ws_anchor.append(headers)
        if ws_hydrated.max_row == 1: ws_hydrated.append(headers)

        for span in self.spans:
            # Anchor row (blanks for offsets)
            row_anchor = [span[k] for k in headers]
            # Blank out offsets for anchor sheet
            row_anchor[7] = None # start
            row_anchor[8] = None # end
            row_anchor[19] = "needs_hydration"
            ws_anchor.append(row_anchor)

            # Hydrated row
            row_hydrated = [span[k] for k in headers]
            ws_hydrated.append(row_hydrated)

    def write_events(self):
        ws = self.get_sheet("Event_Log")
        headers = [
            "source_file", "note_id", "event_id", "event_type", "method",
            "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements",
            "specimens", "findings", "is_historical", "reviewer", "comments",
            "device_size", "device_material",
            "outcome_airway_lumen_pre", "outcome_airway_lumen_post",
            "outcome_symptoms", "outcome_pleural", "outcome_complication"
        ]
        if ws.max_row == 1: ws.append(headers)
        
        for evt in self.events:
            row = [
                SOURCE_FILE, NOTE_ID, evt.get('event_id'), evt.get('type'), evt.get('method'),
                evt.get('target', {}).get('anatomy_type'), json.dumps(evt.get('devices_json')), 
                evt.get('needle_gauge'), evt.get('target', {}).get('station'), 
                json.dumps(evt.get('counts')), json.dumps(evt.get('measurements_json')),
                json.dumps(evt.get('specimens_json')), json.dumps(evt.get('findings_json')),
                False, "Auto", "",
                "", "", # Device size/mat
                "", "", "", "", # Outcomes
                evt.get('outcomes', {}).get('complications')
            ]
            ws.append(row)

    def write_v3_registry(self):
        ws = self.get_sheet("V3_Registry_JSON")
        if ws.max_row == 1: ws.append(["schema_version", "note_id", "json_dump"])
        
        registry_data = {
            "schema_version": "v3.0",
            "note_id": NOTE_ID,
            "procedures": self.events,
            "no_immediate_complications": True
        }
        
        ws.append(["v3.0", NOTE_ID, json.dumps(registry_data, indent=2)])

# =============================================================================
# EXTRACTION LOGIC
# =============================================================================

def extract_data(gen):
    
    # --- Event 1: Therapeutic Aspiration (Initial) ---
    evt_id = "evt_01"
    gen.add_span("Successful therapeutic aspiration was performed", "PROC_METHOD", "therapeutic_aspiration", "method", evt_id)
    gen.add_span("clean out the Trachea", "PROC_ACTION", "clean_mucus", "action", evt_id)
    gen.add_span("mucus", "OBS_FINDING", "mucus", "findings", evt_id, context_prefix="from ")
    
    gen.events.append({
        "event_id": evt_id,
        "type": "therapeutic_aspiration",
        "method": "aspiration",
        "target": {"anatomy_type": "airway_various"},
        "findings_json": ["mucus"]
    })

    # --- Event 2: Navigation / Radial EBUS / Cone Beam ---
    evt_id = "evt_02"
    gen.add_span("Robotic navigation bronchoscopy", "PROC_METHOD", "navigational_bronchoscopy", "method", evt_id)
    gen.add_span("Ion platform", "DEV_INSTRUMENT", "Ion", "devices", evt_id)
    gen.add_span("Anteromedial Segment of LLL (Lb7/8)", "ANAT_LUNG_LOC", "LLL_anteromedial", "target", evt_id)
    gen.add_span("Radial EBUS", "PROC_METHOD", "radial_ebus", "method", evt_id, match_index=2) # 2nd mention
    gen.add_span("not well visualized", "OBS_FINDING", "not_visualized", "findings", evt_id)
    gen.add_span("aerated lung", "OBS_FINDING", "aerated_lung", "findings", evt_id)
    gen.add_span("Cios Spin", "DEV_INSTRUMENT", "Cios Spin", "devices", evt_id)
    gen.add_span("Cone Beam CT", "PROC_METHOD", "cone_beam_ct", "method", evt_id)

    gen.events.append({
        "event_id": evt_id,
        "type": "navigation",
        "method": "robotic_nav_radial_ebus",
        "target": {"anatomy_type": "lung", "location": {"lobe": "LLL", "segment": "Anteromedial"}},
        "devices_json": ["Ion", "Radial EBUS", "Cios Spin"],
        "findings_json": ["nodule not well visualized", "aerated lung"]
    })

    # --- Event 3: Transbronchial Needle Aspiration (Peripheral) ---
    evt_id = "evt_03"
    gen.add_span("Transbronchial needle aspiration", "PROC_METHOD", "tbna_peripheral", "method", evt_id, match_index=1)
    gen.add_span("21G Needle", "DEV_NEEDLE", "21G", "devices", evt_id)
    gen.add_span("Total 4 samples", "MEAS_COUNT", "4", "measurements", evt_id)
    
    gen.events.append({
        "event_id": evt_id,
        "type": "biopsy",
        "method": "transbronchial_needle_aspiration",
        "target": {"anatomy_type": "lung", "location": {"lobe": "LLL"}},
        "devices_json": ["21G Needle"],
        "measurements_json": {"sample_count": 4}
    })

    # --- Event 4: Cryobiopsy ---
    evt_id = "evt_04"
    gen.add_span("Transbronchial cryobiopsy", "PROC_METHOD", "transbronchial_cryobiopsy", "method", evt_id)
    gen.add_span("1.1mm cryoprobe", "DEV_INSTRUMENT", "cryoprobe_1.1mm", "devices", evt_id)
    gen.add_span("Freeze time of 6 seconds", "MEAS_TIME", "6s", "measurements", evt_id)
    gen.add_span("Total 6 samples", "MEAS_COUNT", "6", "measurements", evt_id)

    gen.events.append({
        "event_id": evt_id,
        "type": "biopsy",
        "method": "transbronchial_cryobiopsy",
        "target": {"anatomy_type": "lung", "location": {"lobe": "LLL"}},
        "devices_json": ["1.1mm cryoprobe"],
        "measurements_json": {"freeze_time": "6s", "sample_count": 6}
    })

    # --- Event 5: Brushing ---
    evt_id = "evt_05"
    gen.add_span("Transbronchial brushing", "PROC_METHOD", "transbronchial_brushing", "method", evt_id)
    gen.add_span("Protected cytology brush", "DEV_INSTRUMENT", "cytology_brush", "devices", evt_id)
    gen.add_span("Total 1 samples", "MEAS_COUNT", "1", "measurements", evt_id)

    gen.events.append({
        "event_id": evt_id,
        "type": "biopsy",
        "method": "transbronchial_brushing",
        "target": {"anatomy_type": "lung"},
        "devices_json": ["Protected cytology brush"],
        "measurements_json": {"sample_count": 1}
    })

    # --- Event 6: BAL (Diagnostic - Nodule) ---
    evt_id = "evt_06"
    gen.add_span("Bronchial alveolar lavage", "PROC_METHOD", "bal", "method", evt_id, match_index=1)
    gen.add_span("Instilled 20 cc", "MEAS_VOL", "20cc", "measurements", evt_id)
    gen.add_span("suction returned with 5 cc", "MEAS_VOL", "5cc", "measurements", evt_id)

    gen.events.append({
        "event_id": evt_id,
        "type": "bal",
        "method": "bronchoalveolar_lavage",
        "measurements_json": {"instilled": "20cc", "return": "5cc"}
    })

    # --- Event 7: Fiducial ---
    evt_id = "evt_07"
    gen.add_span("Fiducial marker", "PROC_METHOD", "fiducial_placement", "method", evt_id)
    gen.add_span("0.8mm x 3mm soft tissue gold CIVCO", "DEV_INSTRUMENT", "fiducial_gold_civco", "devices", evt_id)

    gen.events.append({
        "event_id": evt_id,
        "type": "fiducial_marker",
        "method": "placement",
        "devices_json": ["0.8mm x 3mm soft tissue gold CIVCO"]
    })

    # --- Event 8: Therapeutic Aspiration (Post) ---
    evt_id = "evt_08"
    # Matches the second occurrence of this long string
    gen.add_span("Successful therapeutic aspiration was performed", "PROC_METHOD", "therapeutic_aspiration", "method", evt_id, match_index=2)
    gen.add_span("mucus and blood", "OBS_FINDING", "mucus_blood", "findings", evt_id)

    gen.events.append({
        "event_id": evt_id,
        "type": "therapeutic_aspiration",
        "method": "aspiration",
        "findings_json": ["mucus", "blood"]
    })

    # --- Event 9: BAL (Lobar/Segmental) ---
    evt_id = "evt_09"
    gen.add_span("Bronchial alveolar lavage", "PROC_METHOD", "bal", "method", evt_id, match_index=2)
    gen.add_span("Superior Segment of LLL (LB6)", "ANAT_LUNG_LOC", "LLL_superior", "target", evt_id)
    gen.add_span("Instilled 40 cc", "MEAS_VOL", "40cc", "measurements", evt_id)
    gen.add_span("suction returned with 15 cc", "MEAS_VOL", "15cc", "measurements", evt_id)

    gen.events.append({
        "event_id": evt_id,
        "type": "bal",
        "method": "bronchoalveolar_lavage",
        "target": {"anatomy_type": "lung", "location": {"lobe": "LLL"}},
        "measurements_json": {"instilled": "40cc", "return": "15cc"}
    })

    # --- Event 10: EBUS-TBNA ---
    evt_id = "evt_10"
    gen.add_span("Linear EBUS", "PROC_METHOD", "linear_ebus", "method", evt_id)
    gen.add_span("11L lymph node", "ANAT_LN_STATION", "11L", "target", evt_id, context_prefix="Site 1: The ")
    gen.add_span("22-gauge Needle", "DEV_NEEDLE", "22G", "devices", evt_id)
    gen.add_span("< 10 mm", "MEAS_SIZE", "<10mm", "lesion_size", evt_id)
    gen.add_span("4 endobronchial ultrasound guided transbronchial biopsies", "MEAS_COUNT", "4", "measurements", evt_id)

    gen.events.append({
        "event_id": evt_id,
        "type": "ebus_tbna",
        "method": "linear_ebus_tbna",
        "target": {"anatomy_type": "lymph_node", "station": "11L"},
        "devices_json": ["22G Needle"],
        "measurements_json": {"sample_count": 4, "size": "<10mm"}
    })

    # --- Global Outcomes ---
    evt_id = "evt_global"
    gen.add_span("No immediate complications", "OUTCOME_COMPLICATION", "none", "outcomes", evt_id, match_index=2) # Usually appears twice in template notes, ensuring correct one
    
    gen.events.append({
        "event_id": evt_id,
        "type": "outcome",
        "outcomes": {"complications": "none"}
    })

# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    gen = WorkbookGenerator(TEMPLATE_PATH)
    
    # 1. Populate Text
    gen.write_note_text()
    
    # 2. Extract Data
    extract_data(gen)
    
    # 3. Write Sheets
    gen.write_note_index()
    gen.write_spans()
    gen.write_events()
    gen.write_v3_registry()
    
    # 4. Save
    gen.wb.save(OUTPUT_PATH)
    print(f"Successfully generated {OUTPUT_PATH}")

if __name__ == "__main__":
    main()