import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os
from datetime import datetime

# =============================================================================
# INPUT DATA & CONFIGURATION
# =============================================================================

NOTE_ID = "note_120"
SOURCE_FILE = "note_120.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_120 SOURCE_FILE: note_120.txt INDICATION FOR OPERATION:  [REDACTED]is a 74 year old-year-old female who presents with ILD.
The nature, purpose, risks, benefits and alternatives to Bronchoscopy were discussed with the patient in detail.
PREOPERATIVE DIAGNOSIS: Interstitial Lung Disease
POSTOPERATIVE DIAGNOSIS:  Interstitial Lung Disease
PROCEDURE:  
31645 Therapeutic aspiration initial episode
31622 Dx bronchoscope/cell washing          
31624 Dx bronchoscope/lavage (BAL)    
31628 TBBX single lobe     
31654 Radial EBUS for peripheral lesion
D/C 31899NFK BRONCHOSCOPY, RIGID OR FLEXIBLE, INCLUDING FLUOROSCOPIC GUIDANCE AND OCCLUSION BALLOON, WHEN PERFORMED;
WITH BRONCHIAL, ENDOBRONCHIAL, OR TRANSBRONCHIAL CRYOBIOPSY(S), SINGLE LOBE
22 Substantially greater work than normal (i.e., increased intensity, time, technical difficulty of procedure, and severity of patient's condition, physical and mental effort required)
IP [REDACTED] CODE MOD DETAILS: 
Unusual Procedure:
This patient required a Transbronchial Cryo biopsies.
This resulted in >60% increased work due to Increased intensity, Time, Technical difficulty of procedure, Severity of patient's condition, and Physical and mental effort required.
Apply to: 31628 TBBX single lobe     Pt with significant bleeding after biopsy that required blocker placement and attempt to clot blood.
She continued to bleed and required admission to the ICU. Total time with blocker inflation was 24 minutes.
ANESTHESIA: 
General Anesthesia
MONITORING : Pulse oximetry, heart rate, telemetry, and BP were continuously monitored by an independent trained observer that was present throughout the entire procedure.
INSTRUMENT : 
Flexible Therapeutic Bronchoscope
Radial EBUS
ESTIMATED BLOOD LOSS:   50 cc but continued bleeding
COMPLICATIONS:    Bleeding requiring endobronchial blocker
PROCEDURE IN DETAIL:
After the successful induction of anesthesia, a timeout was performed (confirming the patient's name, procedure type, and procedure location).
PATIENT POSITION: .supine
Initial Airway Inspection Findings:
The patient was sedated by anesthesia.
Next the bronchoscope was used to fiberoptically intubate the patient with a size 7 Ardnt blocker along side the ETT.
The vocal cords are normal instructure and function. After atraumatic intubation the loop of the endobronchial blocker was moved to the RLL lateral subsegment
Evaluation of the airway was performed.
There were no endobronchial lesions and the mucosa was normal.
Bronchial alveolar lavage was performed at Medial Segment of RML (RB5).
Instilled 40 cc of NS, suction returned with 25 cc of NS.
Samples sent for Cell Count, Microbiology (Cultures/Viral/Fungal), and Cytology.
Radial EBUS was performed to confirm there were no enlarged blood vessels in the segment of the RLL lateral segment.
Transbronchial biopsy was performed with alligator forceps at Lateral-basal Segment of RLL (RB9).  Total 1 samples were collected.
Samples sent for Pathology. There was no bleeding after this biopsy
Transbronchial biopsy was performed with 1.7 mm cryoprobe with a 4 second freeze.
The scope and the cryoprobe were removed together at Lateral-basal Segment of RLL (RB9).  Total 1 samples were collected.
Samples sent for Pathology. Immediately after removal of the bronchoscope the balloon on the blocker was inflated with 3 cc.
The balloon was left inflated for a total of 2 minutes. At the time of deflation there was continued bleeding.
A total of 1000 mg of TXA, 2 mg epi, 5 cc iced saline were given through the end of the inflated blocker.
Additional isolation time of 5 minutes, followed by additional 10 minutes followed by another 10 minutes.
The ETT is at 21 cm at the front teeth and the blocker is at 29 cm at the front teeth.
The patient tolerated the procedure well.  There were no immediate complications.
At the conclusion of the operation, the patient was transferred to the ICU for continued care.
SPECIMEN(S): 
BAL RML 
TBBX right lower lobe
cryoTBBx right lower lobe
IMPRESSION/PLAN: [REDACTED]is a 74 year old-year-old female who presents for bronchoscopy for ILD.
[ ] transfer to ICU
[ ] repeat bronchoscopy in ICU"""

# =============================================================================
# LOGIC & DEFINITIONS
# =============================================================================

# 1. FLAG LOGIC
# Based on CPT codes and text evidence
FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,      # 31622
    "bal": 1,                          # 31624
    "bronchial_wash": 0,               # 31622 usually covers wash, but BAL is specific
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 1,                  # 31654
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 1,        # 31628
    "transbronchial_cryobiopsy": 1,    # Text: "Transbronchial Cryo biopsies"
    "therapeutic_aspiration": 1,       # 31645
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,                  # Used for biopsy here, not therapy
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 2. EVENTS STRUCTURE
# We define granular events for V3_Procedure_Events and Span generation
EVENTS = [
    {
        "event_id": "evt01",
        "type": "diagnostic_bronchoscopy",
        "subtype": "airway_inspection",
        "finding": "normal mucosa",
        "device": "size 7 Ardnt blocker" # Placed initially
    },
    {
        "event_id": "evt02",
        "type": "bal",
        "target": "RML",
        "segment": "Medial Segment (RB5)",
        "method": "lavage",
        "measurements": {"instilled": "40 cc", "return": "25 cc"}
    },
    {
        "event_id": "evt03",
        "type": "radial_ebus",
        "target": "RLL",
        "segment": "Lateral Segment",
        "finding": "no enlarged blood vessels"
    },
    {
        "event_id": "evt04",
        "type": "transbronchial_biopsy",
        "target": "RLL",
        "segment": "Lateral-basal Segment (RB9)",
        "device": "alligator forceps",
        "specimen_count": "1"
    },
    {
        "event_id": "evt05",
        "type": "transbronchial_cryobiopsy",
        "target": "RLL",
        "segment": "Lateral-basal Segment (RB9)",
        "device": "1.7 mm cryoprobe",
        "method": "4 second freeze",
        "specimen_count": "1",
        "complication": "Bleeding requiring endobronchial blocker"
    },
    {
        "event_id": "evt06",
        "type": "therapeutic_aspiration", # Using this to capture bleeding management
        "subtype": "bleeding_control",
        "action": ["balloon inflated", "TXA", "epi", "iced saline"],
        "target": "RLL",
        "outcome": "continued to bleed... transfer to ICU"
    }
]

# 3. SPAN EXTRACTION LIST
# Fields: span_text, context_prefix (opt), label, norm_val, event_id, field_name
SPANS_TO_EXTRACT = [
    # Metadata / Codes
    ("31645 Therapeutic aspiration initial episode", "", "PROC_METHOD", "Therapeutic Aspiration", "evt06", "method"),
    ("31622 Dx bronchoscope/cell washing", "", "PROC_METHOD", "Diagnostic Bronchoscopy", "evt01", "method"),
    ("31624 Dx bronchoscope/lavage (BAL)", "", "PROC_METHOD", "BAL", "evt02", "method"),
    ("31628 TBBX single lobe", "", "PROC_METHOD", "Transbronchial Biopsy", "evt04", "method"),
    ("31654 Radial EBUS", "", "PROC_METHOD", "Radial EBUS", "evt03", "method"),
    ("Transbronchial Cryo biopsies", "patient required a", "PROC_METHOD", "Transbronchial Cryobiopsy", "evt05", "method"),
    
    # Event 1: Setup/Inspection
    ("size 7 Ardnt blocker", "patient with a", "DEV_CATHETER", "Arndt Blocker 7F", "evt01", "devices_json"),
    ("RLL lateral subsegment", "moved to the", "ANAT_LUNG_LOC", "RLL Lateral Segment", "evt01", "target.location"),
    
    # Event 2: BAL
    ("Bronchial alveolar lavage", "normal.\n", "PROC_METHOD", "BAL", "evt02", "method"),
    ("Medial Segment of RML (RB5)", "performed at", "ANAT_LUNG_LOC", "RML Medial Segment", "evt02", "target.location"),
    ("40 cc of NS", "Instilled", "MEAS_VOL", "40 cc", "evt02", "measurements_json"),
    ("25 cc of NS", "returned with", "MEAS_VOL", "25 cc", "evt02", "measurements_json"),
    
    # Event 3: Radial EBUS
    ("Radial EBUS", "Cytology.\n", "PROC_METHOD", "Radial EBUS", "evt03", "method"),
    ("segment of the RLL lateral segment", "vessels in the", "ANAT_LUNG_LOC", "RLL Lateral Segment", "evt03", "target.location"),
    
    # Event 4: TBBX Forceps
    ("Transbronchial biopsy", "segment.\n", "PROC_METHOD", "Transbronchial Biopsy", "evt04", "method"),
    ("alligator forceps", "performed with", "DEV_INSTRUMENT", "Alligator Forceps", "evt04", "devices_json"),
    ("Lateral-basal Segment of RLL (RB9)", "forceps at", "ANAT_LUNG_LOC", "RLL Lateral-basal Segment", "evt04", "target.location"),
    ("Total 1 samples", "RB9).  ", "MEAS_COUNT", "1", "evt04", "specimens_json"),
    
    # Event 5: Cryo TBBX
    ("1.7 mm cryoprobe", "performed with", "DEV_INSTRUMENT", "Cryoprobe 1.7mm", "evt05", "devices_json"),
    ("4 second freeze", "with a", "PROC_ACTION", "Freeze 4s", "evt05", "method"),
    ("Lateral-basal Segment of RLL (RB9)", "together at", "ANAT_LUNG_LOC", "RLL Lateral-basal Segment", "evt05", "target.location"),
    
    # Event 6: Bleeding Control / Complication
    ("balloon on the blocker was inflated", "bronchoscope the", "PROC_ACTION", "Balloon Inflation", "evt06", "method"),
    ("continued bleeding", "deflation there was", "OUTCOME_COMPLICATION", "Bleeding", "evt05", "outcomes.complications"), # Attaching complication to the Cryo event
    ("1000 mg of TXA", "Total of", "PROC_ACTION", "Instill TXA", "evt06", "method"),
    ("2 mg epi", "TXA,", "PROC_ACTION", "Instill Epinephrine", "evt06", "method"),
    ("5 cc iced saline", "epi,", "PROC_ACTION", "Instill Iced Saline", "evt06", "method"),
    ("Bleeding requiring endobronchial blocker", "COMPLICATIONS:    ", "OUTCOME_COMPLICATION", "Bleeding", "evt05", "outcomes.complications"),
]

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def find_offsets_hydrated(full_text, span_text, context_prefix="", match_index=0):
    """
    Finds start_char and end_char for a span.
    Hydration logic:
    1. If unique in text -> take it.
    2. If context_prefix provided -> search for prefix + span within window.
    3. If match_index provided -> take the Nth occurrence.
    4. Fallback -> Ambiguous.
    """
    occurrences = [m.start() for m in re.finditer(re.escape(span_text), full_text)]
    
    if not occurrences:
        return None, None, "not_found"
    
    # 1. Unique
    if len(occurrences) == 1:
        start = occurrences[0]
        return start, start + len(span_text), "hydrated_unique"
        
    # 2. Context Prefix
    if context_prefix:
        # Search for context near occurrences
        for start in occurrences:
            # Look back 150 chars
            window_start = max(0, start - 150)
            preceding_text = full_text[window_start:start]
            if context_prefix in preceding_text:
                return start, start + len(span_text), "hydrated_prefix_window"
    
    # 3. Match Index
    if 0 <= match_index < len(occurrences):
        start = occurrences[match_index]
        return start, start + len(span_text), "hydrated_match_index"
        
    return None, None, f"ambiguous_count={len(occurrences)}"

# =============================================================================
# WORKBOOK GENERATION
# =============================================================================

def create_workbook():
    # Load Template
    if not os.path.exists(TEMPLATE_PATH):
        # Create a dummy workbook if template missing (for robustness in some envs, though instructions say it must exist)
        wb = openpyxl.Workbook()
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # 1. Note_Text
    if "Note_Text" not in wb.sheetnames:
        wb.create_sheet("Note_Text")
    ws_text = wb["Note_Text"]
    # Assuming headers exist or appending blindly if empty. 
    # Template usually has headers. We append.
    if ws_text.max_row == 1 and ws_text.cell(1, 1).value is None:
         ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    if "Note_Index" not in wb.sheetnames:
        wb.create_sheet("Note_Index")
    ws_index = wb["Note_Index"]
    
    # Flag columns usually start at col 9 (I) after metadata
    # Metadata: source_file, note_id, encounter_id, procedure_date, site, reviewer, status, free_text_notes
    meta_row = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Not Started", ""]
    
    # Flags dict to list (ordered)
    # Order must match template if template is rigid, but here we append.
    # We will assume the template has specific columns. 
    # For safety, we'll map keys to standard list.
    flag_keys = [
        "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", 
        "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy", 
        "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration", 
        "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation", 
        "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", 
        "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy", 
        "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis", 
        "pleural_biopsy", "fibrinolytic_therapy"
    ]
    
    flag_values = [FLAGS.get(k, 0) for k in flag_keys]
    
    # Append to Note_Index (Row 2, assuming Row 1 is header)
    ws_index.append(meta_row + flag_values)

    # 3. Span_Annotations (Anchor-First)
    if "Span_Annotations" not in wb.sheetnames:
        wb.create_sheet("Span_Annotations")
    ws_spans = wb["Span_Annotations"]
    
    # 4. Span_Hydrated (To be filled)
    if "Span_Hydrated" not in wb.sheetnames:
        wb.create_sheet("Span_Hydrated")
    ws_hydrated = wb["Span_Hydrated"]
    # Write headers for Hydrated if empty
    hydrated_headers = [
        "source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text",
        "match_index", "start_char", "end_char", "span_len", "label", "normalized_value",
        "schema_field", "event_id", "is_negated", "is_historical", "time_anchor", "reviewer",
        "comments", "hydration_status"
    ]
    if ws_hydrated.max_row == 1 and ws_hydrated.cell(1, 1).value is None:
        ws_hydrated.append(hydrated_headers)

    span_rows_data = []
    
    for idx, item in enumerate(SPANS_TO_EXTRACT):
        span_text, context_prefix, label, norm_val, event_id, field_name = item
        span_id = f"{NOTE_ID}_span_{idx+1:03d}"
        
        # Calculate hydration
        start, end, status = find_offsets_hydrated(NOTE_TEXT, span_text, context_prefix)
        span_len = len(span_text)
        
        # Row for Span_Annotations (No offsets)
        # source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, 
        # start_char, end_char, span_len, label, normalized_value, schema_field, event_id, 
        # is_negated, is_historical, time_anchor, reviewer, comments, hydration_status
        
        # We assume section_type is blank or "Procedure"
        row_anno = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure", context_prefix, span_text, 
            "", "", "", f"=LEN(F{ws_spans.max_row+1})", label, norm_val, field_name, event_id, 
            0, 0, "", "", "", "needs_hydration"
        ]
        ws_spans.append(row_anno)
        
        # Row for Span_Hydrated (With offsets)
        row_hyd = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure", context_prefix, span_text,
            0, start if start is not None else "", end if end is not None else "", span_len,
            label, norm_val, field_name, event_id, 0, 0, "", "", "", status
        ]
        ws_hydrated.append(row_hyd)

    # 5. Event_Log
    if "Event_Log" not in wb.sheetnames:
        wb.create_sheet("Event_Log")
    ws_events = wb["Event_Log"]
    # Columns: source_file, note_id, event_id, event_type, method, anatomy_target, device, 
    # needle_gauge, stations, counts, measurements, specimens, findings, is_historical, reviewer, comments,
    # device_size, device_material, outcome_airway_lumen_pre, outcome_airway_lumen_post, outcome_symptoms, 
    # outcome_pleural, outcome_complication
    
    event_rows = []
    # Map internal EVENTS dict to Event_Log rows
    for evt in EVENTS:
        # Defaults
        evt_type = evt.get("type", "")
        method = evt.get("method", "")
        if isinstance(method, list): method = ", ".join(method)
        target = evt.get("target", "")
        if evt.get("segment"): target += f" {evt.get('segment')}"
        device = evt.get("device", "")
        measurements = json.dumps(evt.get("measurements", {})) if evt.get("measurements") else ""
        findings = evt.get("finding", "")
        outcome_comp = evt.get("complication", "")
        outcome_sym = evt.get("outcome", "")
        
        row = [
            SOURCE_FILE, NOTE_ID, evt["event_id"], evt_type, method,
            target, device, "", "", evt.get("specimen_count", ""), measurements, "", findings, 0, "", "",
            "", "", "", "", outcome_sym, "", outcome_comp
        ]
        ws_events.append(row)

    # 6. V3_Procedure_Events
    if "V3_Procedure_Events" not in wb.sheetnames:
        wb.create_sheet("V3_Procedure_Events")
    ws_v3 = wb["V3_Procedure_Events"]
    
    for evt in EVENTS:
        # Construct V3 row
        # note_id, event_id, type, target.anatomy_type, target.location.lobe, target.location.segment, target.station,
        # lesion.type, lesion.size_mm, method, devices_json, measurements_json, specimens_json, findings_json, evidence_quote,
        # stent.size, stent.material_or_brand, catheter.size_fr, outcomes.airway.lumen_pre, outcomes.airway.lumen_post,
        # outcomes.symptoms, outcomes.pleural, outcomes.complications
        
        # Parsing target
        lobe = ""
        seg = ""
        if "RLL" in evt.get("target", ""): lobe = "RLL"
        if "RML" in evt.get("target", ""): lobe = "RML"
        if evt.get("segment"): seg = evt.get("segment")
        
        devices = evt.get("device", "")
        findings_json = json.dumps({"finding": evt.get("finding")}) if evt.get("finding") else ""
        
        row_v3 = [
            NOTE_ID, evt["event_id"], evt["type"], "LUNG", lobe, seg, "",
            "", "", str(evt.get("method", "")), devices, 
            json.dumps(evt.get("measurements", {})), 
            evt.get("specimen_count", ""), findings_json, "",
            "", "", "", "", "", evt.get("outcome", ""), "", evt.get("complication", "")
        ]
        ws_v3.append(row_v3)

    # 7. V3_Registry_JSON
    if "V3_Registry_JSON" not in wb.sheetnames:
        wb.create_sheet("V3_Registry_JSON")
    ws_json = wb["V3_Registry_JSON"]
    
    registry_data = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "procedures": EVENTS,
        "no_immediate_complications": False # False because of bleeding
    }
    
    ws_json.cell(row=1, column=1, value=json.dumps(registry_data, indent=2))

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    create_workbook()