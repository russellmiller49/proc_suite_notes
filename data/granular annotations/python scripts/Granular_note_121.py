import openpyxl
from openpyxl.utils import get_column_letter
import json
import os
import re

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_121"
SOURCE_FILE = "note_121.txt"
PROCEDURE_DATE = ""
NOTE_TEXT = """NOTE_ID:  note_121 SOURCE_FILE: note_121.txt Procedure Name

Bronchoscopy with Endobronchial Balloon Dilation

Procedure Narrative

Patient name: [REDACTED]

Preoperative diagnosis: Aspirated Foreign Body on Bronchus, Acquired Bronchial Stenosis
Postoperative diagnosis: same as above

Procedures performed:
1. Flexible bronchoscopy
2. Endobronchial Balloon Dilation

Indications for the procedure: Aspirated Foreign Body on Bronchus with subsequent Acquired Bronchial Stenosis

Anesthesia: General Sedation by Anesthesia Team;
Topical Lidocaine 1%(10 mL total)

Procedure: After obtaining informed consent from the patient, a timeout was performed to identify the correct patient and procedure to be performed.
Once the patient was adequately sedated, the flexible bronchoscope was placed through the LMA. A flexible bronchoscopy was then performed.
Topical Lidocaine 1% was sequentially instilled onto the Main and Secondary Carina.
Examination of the airways revealed a normal Trachea and an unremarkable left side. There was no endobronchial nodularity or masses.
Friable mucosa was noted from the distal Bronchus Medius to the takeoff of the RLL.
There was significant granulation and scar tissue as well. The entrance to the RLL was significantly stenotic and unable to identify the opening to the Superior Segment of the RLL.
Performed serial dilation with balloon dilator and after which purulent secretions were seen coming from the RLL.
Performed BAL with 30cc of normal saline and 20cc of return of cloudy fluid.
There was also evidence of course particulates at the RLL take off that were removed and send for pathology.
RLL takeoff was still significantly stenotic and unable to pass diagnostic bronchoscope.
Used APC and alligator forceps to debride scar tissue after which was able to pass the diagnostic bronchoscope.
The remaining basilar segments of the RLL appeared patent.

Following confirmation of hemostasis and therapeutic aspiration of all endobronchial secretions, the bronchoscope was removed.
Patient tolerated the procedure well with no immediate complications.

EBL: minimal

Bronchoscopes: Diagnostic and Therapeutic Bronchoscopes

Specimens:
Bronchial wash
BAL
Cell count
Bacterial culture
Fungal culture
Surgical Pathology"""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# -------------------------------------------------------------------------
# CONFIGURATION & UTILS
# -------------------------------------------------------------------------
PROCEDURE_FLAGS = [
    "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
    "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy",
    "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration",
    "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
    "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation",
    "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
    "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis",
    "pleural_biopsy", "fibrinolytic_therapy"
]

def create_workbook_if_needed():
    if not os.path.exists(TEMPLATE_PATH):
        wb = openpyxl.Workbook()
        # Create sheets
        for name in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            if name not in wb.sheetnames:
                wb.create_sheet(name)
        # Remove default
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
            
        # Add minimal headers for logic to work (completeness depends on real template)
        wb["Note_Text"].append(["note_id", "source_file", "note_text"])
        wb["Note_Index"].append(["source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes"] + PROCEDURE_FLAGS)
        wb["Span_Annotations"].append(["source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", "match_index", "start_char", "end_char", "span_len", "label", "normalized_value", "schema_field", "event_id", "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"])
        wb["Span_Hydrated"].append(["source_file", "note_id", "span_id", "section_type", "context_prefix", "span_text", "match_index", "start_char", "end_char", "span_len", "label", "normalized_value", "schema_field", "event_id", "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"])
        
        wb.save(TEMPLATE_PATH)

def get_offsets(full_text, span_text, context_prefix=None, match_index=None):
    if not span_text:
        return None, None, "missing_text"
    
    matches = [m for m in re.finditer(re.escape(span_text), full_text)]
    if not matches:
        return None, None, "not_found"
    
    if len(matches) == 1:
        return matches[0].start(), matches[0].end(), "hydrated_unique"
    
    # Ambiguous
    if context_prefix:
        # Look for prefix in window before
        for m in matches:
            start = m.start()
            window_start = max(0, start - 120)
            window = full_text[window_start:start]
            if context_prefix in window:
                return m.start(), m.end(), "hydrated_prefix_window"
    
    if match_index is not None and 0 <= match_index < len(matches):
        return matches[match_index].start(), matches[match_index].end(), "hydrated_match_index"
    
    return None, None, f"ambiguous_count={len(matches)}"

# -------------------------------------------------------------------------
# EXTRACTION LOGIC
# -------------------------------------------------------------------------
def extract_data():
    # 1. Flag Detection
    flags = {k: 0 for k in PROCEDURE_FLAGS}
    
    # Textual triggers
    lower_text = NOTE_TEXT.lower()
    
    if "bronchoscopy" in lower_text:
        flags["diagnostic_bronchoscopy"] = 1
    if "bal" in lower_text or "bronchoalveolar lavage" in lower_text:
        flags["bal"] = 1
    if "bronchial wash" in lower_text:
        flags["bronchial_wash"] = 1
    if "foreign body" in lower_text and ("removed" in lower_text or "removal" in lower_text):
        flags["foreign_body_removal"] = 1
    if "balloon" in lower_text and ("dilation" in lower_text or "dilator" in lower_text):
        flags["airway_dilation"] = 1
    if "apc" in lower_text or "argon plasma" in lower_text:
        flags["thermal_ablation"] = 1
    if "forceps" in lower_text and "debride" in lower_text:
        flags["tumor_debulking_non_thermal"] = 1
    if "therapeutic aspiration" in lower_text:
        flags["therapeutic_aspiration"] = 1

    # 2. Span Definitions (Anchor-First)
    # Define spans with manual context if needed
    spans = []
    
    # Helper to add span
    def add_span(text, label, norm_val, field=None, event_id=None, prefix=None, idx=None, neg=False, hist=False):
        spans.append({
            "text": text, "label": label, "norm": norm_val, "field": field, 
            "event": event_id, "prefix": prefix, "idx": idx, "neg": neg, "hist": hist
        })

    # --- ANATOMY & FINDINGS ---
    # "unremarkable left side"
    add_span("left side", "LATERALITY", "Left", event_id="E0", field="target.location.lobe")
    
    # "Bronchus Medius"
    add_span("Bronchus Medius", "ANAT_AIRWAY", "Bronchus intermedius", event_id="E0")
    
    # "takeoff of the RLL"
    add_span("RLL", "ANAT_LUNG_LOC", "RLL", event_id="E0", prefix="takeoff of the")
    
    # "significant granulation and scar tissue"
    add_span("granulation", "OBS_LESION", "granulation", event_id="E1")
    add_span("scar tissue", "OBS_LESION", "scar_tissue", event_id="E1", prefix="granulation and")

    # "entrance to the RLL was significantly stenotic"
    add_span("RLL", "ANAT_LUNG_LOC", "RLL", event_id="E1", prefix="entrance to the")
    add_span("significantly stenotic", "OUTCOME_AIRWAY_LUMEN_PRE", "stenosis_significant", event_id="E1")

    # --- PROCEDURE: BALLOON DILATION (E1) ---
    add_span("serial dilation", "PROC_METHOD", "dilation", event_id="E1")
    add_span("balloon dilator", "DEV_INSTRUMENT", "balloon_dilator", event_id="E1")
    
    # Outcome of dilation or just finding? "purulent secretions were seen"
    add_span("purulent secretions", "OBS_LESION", "secretions_purulent", event_id="E1")
    
    # --- PROCEDURE: BAL (E2) ---
    add_span("Performed BAL", "PROC_METHOD", "bal", event_id="E2")
    add_span("30cc", "MEAS_VOL", "30", event_id="E2", field="measurements.volume")
    add_span("20cc", "MEAS_VOL", "20", event_id="E2", field="measurements.volume_return")
    add_span("cloudy fluid", "OBS_LESION", "fluid_cloudy", event_id="E2")
    
    # --- PROCEDURE: FOREIGN BODY / DEBRIDEMENT (E3/E4) ---
    # Note mentions "course particulates at the RLL take off that were removed" -> Foreign Body
    # Note mentions "RLL takeoff was still significantly stenotic"
    # Note mentions "Used APC and alligator forceps to debride scar tissue"
    
    # E3: Foreign Body Removal
    add_span("course particulates", "OBS_LESION", "foreign_body", event_id="E3")
    add_span("RLL take off", "ANAT_LUNG_LOC", "RLL", event_id="E3", prefix="particulates at the")
    add_span("removed", "PROC_ACTION", "removal", event_id="E3", prefix="that were")
    
    # E4: APC / Debridement (grouped as one therapeutic session on the scar)
    add_span("RLL takeoff", "ANAT_LUNG_LOC", "RLL", event_id="E4", prefix="pathology.")
    add_span("significantly stenotic", "OUTCOME_AIRWAY_LUMEN_PRE", "stenosis_significant", event_id="E4", prefix="still")
    
    add_span("APC", "PROC_METHOD", "apc", event_id="E4")
    add_span("alligator forceps", "DEV_INSTRUMENT", "alligator_forceps", event_id="E4")
    add_span("debride", "PROC_ACTION", "debridement", event_id="E4")
    add_span("scar tissue", "OBS_LESION", "scar_tissue", event_id="E4", prefix="debride")
    
    # Outcome E4
    add_span("able to pass the diagnostic bronchoscope", "OUTCOME_AIRWAY_LUMEN_POST", "patent_complete", event_id="E4")
    add_span("basilar segments of the RLL appeared patent", "OUTCOME_AIRWAY_LUMEN_POST", "patent_complete", event_id="E4")
    
    # --- PROCEDURE: THERAPEUTIC ASPIRATION (E5) ---
    add_span("therapeutic aspiration", "PROC_METHOD", "aspiration", event_id="E5")
    add_span("all endobronchial secretions", "OBS_LESION", "secretions", event_id="E5")
    
    # --- GLOBAL OUTCOME ---
    add_span("no immediate complications", "OUTCOME_COMPLICATION", "none", event_id="E99")

    # 3. Events Generation
    # We will build V3_Procedure_Events and Event_Log based on grouped event_ids
    events_map = {
        "E1": {"type": "airway_dilation", "method": "Balloon Dilation", "anatomy": "RLL"},
        "E2": {"type": "bal", "method": "BAL", "anatomy": "RLL"},
        "E3": {"type": "foreign_body_removal", "method": "Removal", "anatomy": "RLL"},
        "E4": {"type": "thermal_ablation", "method": "APC and Forceps", "anatomy": "RLL"},
        "E5": {"type": "therapeutic_aspiration", "method": "Aspiration", "anatomy": "Airway"},
    }

    return flags, spans, events_map

# -------------------------------------------------------------------------
# EXECUTION
# -------------------------------------------------------------------------
def main():
    create_workbook_if_needed()
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    
    flags, raw_spans, events_map = extract_data()
    
    # 1. Populate Note_Text
    ws_text = wb["Note_Text"]
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # 2. Populate Note_Index
    ws_index = wb["Note_Index"]
    row_idx = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "pre-label", ""]
    # Append flags in order
    for f in PROCEDURE_FLAGS:
        row_idx.append(flags.get(f, 0))
    ws_index.append(row_idx)
    
    # 3. Populate Span_Annotations & Span_Hydrated
    ws_anno = wb["Span_Annotations"]
    ws_hydra = wb["Span_Hydrated"]
    
    span_id_counter = 1
    
    # For V3 events, we need to aggregate data
    v3_event_data = {eid: {
        "note_id": NOTE_ID, "event_id": eid, "type": info["type"], 
        "anatomy": info["anatomy"], "method": info["method"],
        "devices": [], "measurements": [], "specimens": [], "findings": [],
        "outcomes": {"lumen_pre": "", "lumen_post": "", "symptoms": "", "pleural": "", "complications": ""}
    } for eid, info in events_map.items()}

    # Add global complication outcome to all relevant therapeutic events or just the last one?
    # We'll attach "no immediate complications" to E4 (main therapeutic) or E99 (global).
    # For this format, we keep E99 spans but maybe not a row in V3 unless it's a procedure.
    
    for s in raw_spans:
        span_id = f"{NOTE_ID}_s{span_id_counter:03d}"
        span_id_counter += 1
        
        # Hydration
        start, end, status = get_offsets(NOTE_TEXT, s["text"], s["prefix"], s["idx"])
        length = (end - start) if start is not None else ""
        
        row_base = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure Narrative", s["prefix"], s["text"], s["idx"],
            "", "", "", # start, end, len (blank for anno)
            s["label"], s["norm"], s["field"], s["event"], 
            s["neg"], s["hist"], "", "", "", "needs_hydration"
        ]
        ws_anno.append(row_base)
        
        row_hydra = list(row_base)
        row_hydra[7] = start
        row_hydra[8] = end
        row_hydra[9] = length
        row_hydra[19] = status
        ws_hydra.append(row_hydra)
        
        # Accumulate V3 Data
        eid = s["event"]
        if eid in v3_event_data:
            ev = v3_event_data[eid]
            if s["label"] in ["DEV_INSTRUMENT", "DEV_STENT", "DEV_CATHETER"]:
                ev["devices"].append(s["norm"])
            if s["label"] in ["MEAS_VOL", "MEAS_SIZE", "MEAS_COUNT"]:
                ev["measurements"].append(f"{s['norm']} ({s['text']})")
            if s["label"] == "OBS_LESION":
                ev["findings"].append(s["norm"])
            if s["label"] == "OUTCOME_AIRWAY_LUMEN_PRE":
                ev["outcomes"]["lumen_pre"] = s["norm"]
            if s["label"] == "OUTCOME_AIRWAY_LUMEN_POST":
                ev["outcomes"]["lumen_post"] = s["norm"]
            if s["label"] == "OUTCOME_COMPLICATION":
                ev["outcomes"]["complications"] = s["norm"]

    # 4. Populate Event_Log & V3_Procedure_Events
    ws_event = wb["Event_Log"]
    ws_v3 = wb["V3_Procedure_Events"]
    
    # Global complication span handling
    # If "no immediate complications" is E99, we'll manually push it to E4 outcomes if empty
    global_compl = "none" # inferred from text
    
    for eid, ev in v3_event_data.items():
        # Event Log Row (Best Effort)
        # source_file, note_id, event_id, event_type, method, anatomy_target, device, needle_gauge, stations, counts, measurements, specimens, findings, is_historical, reviewer, comments, device_size, device_material, outcome_airway_lumen_pre, outcome_airway_lumen_post, outcome_symptoms, outcome_pleural, outcome_complication
        
        dev_str = ", ".join(ev["devices"])
        meas_str = ", ".join(ev["measurements"])
        find_str = ", ".join(ev["findings"])
        
        # Outcomes
        lumen_pre = ev["outcomes"]["lumen_pre"]
        lumen_post = ev["outcomes"]["lumen_post"]
        compl = ev["outcomes"]["complications"] if ev["outcomes"]["complications"] else (global_compl if eid == "E4" else "")
        
        ws_event.append([
            SOURCE_FILE, NOTE_ID, eid, ev["type"], ev["method"],
            ev["anatomy"], dev_str, "", "", "", meas_str,
            "", find_str, "False", "", "",
            "", "", 
            lumen_pre, lumen_post, "", "", compl
        ])
        
        # V3 Row
        # note_id, event_id, type, target.anatomy_type, target.location.lobe, target.location.segment, target.station, lesion.type, lesion.size_mm, method, devices_json, measurements_json, specimens_json, findings_json, evidence_quote, stent.size, stent.material_or_brand, catheter.size_fr, outcomes.airway.lumen_pre, outcomes.airway.lumen_post, outcomes.symptoms, outcomes.pleural, outcomes.complications
        ws_v3.append([
            NOTE_ID, eid, ev["type"], 
            "Airway", ev["anatomy"], "", "", # anatomy details
            "", "", # lesion
            ev["method"], json.dumps(ev["devices"]), json.dumps(ev["measurements"]), "[]", json.dumps(ev["findings"]), "",
            "", "", "", # device details
            lumen_pre, lumen_post, "", "", compl
        ])

    # 5. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    registry_obj = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": list(v3_event_data.values())
    }
    ws_json.append([json.dumps(registry_obj, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    main()