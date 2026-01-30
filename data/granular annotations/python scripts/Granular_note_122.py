import openpyxl
from openpyxl.utils import get_column_letter
import os
import datetime
import json
import re

# -------------------------------------------------------------------------
# INPUTS
# -------------------------------------------------------------------------
NOTE_ID = "note_122"
SOURCE_FILE = "note_122.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_122 SOURCE_FILE: note_122.txt Patient name: [REDACTED]
Preoperative diagnosis: 1:Mediastinal mass
Postoperative diagnosis: same as above

Procedures performed:
CPT 31654 Bronchoscope with Endobronchial Ultrasound guidance for endobronchial biopsy

Indications for the procedure: Mediastinal Mass

Anesthesia: General anesthesia using a LMA.
Procedure: After obtaining informed consent from the patient, he was brought to the procedure room.
A proper timeout was performed to identify the correct patient and procedure to be performed.
General anesthesia was administered to the patient and_was orally intubated with a LMA by the anesthesia team.
the T190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway was in good position. The vocal cords were normal. The subglottic space was normal.
The trachea was of normal caliber. The carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level. Bronchial mucosa and anatomy were normal;
there are no endobronchial lesions.

The UC180F convex probe EBUS bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The mediastinal mass was located in the 4R position. Sampling by transbronchial needle aspiration was performed with the Olympus Vizishot 2 EBUS TBNA 22 gauge needle.
A total of 5 biopsies were performed. ROSE evaluation yielded pus and inflammatory cells.
All samples were sent for routine cytology. Following completion of EBUS bronchoscopy, the T180 video bronchoscope was then re-inserted and after suctioning blood and secretions there was no evidence of active bleeding and the bronchoscope was subsequently removed.
Patient tolerated the procedure well with no immediate complications. Patient was extubated and returned to recovery in good condition.
Specimens:
Cytology
Bacterial culture
Fungal culture
Acid Fast culture
Tissue culture

EBL: minimal

Bronchoscopes: Olympus Q190, UC180F EBUS"""

# -------------------------------------------------------------------------
# CONSTANTS & SCHEMA
# -------------------------------------------------------------------------
PROCEDURE_FLAGS = [
    'diagnostic_bronchoscopy', 'bal', 'bronchial_wash', 'brushings', 
    'endobronchial_biopsy', 'tbna_conventional', 'linear_ebus', 'radial_ebus', 
    'navigational_bronchoscopy', 'transbronchial_biopsy', 'transbronchial_cryobiopsy', 
    'therapeutic_aspiration', 'foreign_body_removal', 'airway_dilation', 
    'airway_stent', 'thermal_ablation', 'tumor_debulking_non_thermal', 
    'cryotherapy', 'blvr', 'peripheral_ablation', 'bronchial_thermoplasty', 
    'whole_lung_lavage', 'rigid_bronchoscopy', 
    'thoracentesis', 'chest_tube', 'ipc', 'medical_thoracoscopy', 
    'pleurodesis', 'pleural_biopsy', 'fibrinolytic_therapy'
]

# -------------------------------------------------------------------------
# HELPER FUNCTIONS
# -------------------------------------------------------------------------
def clean_text(text):
    return text.strip()

def get_row_dict(sheet, header_row=1):
    headers = [cell.value for cell in sheet[header_row]]
    return headers

def append_row(sheet, data_dict, header_map):
    row = []
    for col_name in header_map:
        row.append(data_dict.get(col_name, ""))
    sheet.append(row)

def find_offsets_anchor_first(note_text, span_text, context_prefix=None, match_index=None):
    """
    Hydrates offsets based on strict logic:
    1. count(span) == 1 -> use it.
    2. context_prefix -> search 120 chars before.
    3. match_index -> use N-th occurrence.
    """
    # Find all start indices
    occurrences = [m.start() for m in re.finditer(re.escape(span_text), note_text)]
    
    if not occurrences:
        return None, None, "not_found"
    
    if len(occurrences) == 1:
        start = occurrences[0]
        return start, start + len(span_text), "hydrated_unique"
    
    # Context Prefix Logic
    if context_prefix:
        for start in occurrences:
            # Look back 120 chars
            window_start = max(0, start - 120)
            window_text = note_text[window_start:start]
            if context_prefix in window_text:
                return start, start + len(span_text), "hydrated_prefix_window"
    
    # Match Index Logic
    if match_index is not None and 1 <= match_index <= len(occurrences):
        start = occurrences[match_index - 1]
        return start, start + len(span_text), "hydrated_match_index"
    
    return None, None, f"ambiguous_count={len(occurrences)}"

# -------------------------------------------------------------------------
# EXTRACTION LOGIC
# -------------------------------------------------------------------------
def generate_flags():
    flags = {k: 0 for k in PROCEDURE_FLAGS}
    
    # Diagnostic Bronchoscopy (T190 used, T180 used)
    flags['diagnostic_bronchoscopy'] = 1
    
    # EBUS-TBNA (UC180F used, 4R sampling)
    flags['linear_ebus'] = 1
    # Note: conventional tbna is 0 since this is EBUS-guided
    # endobronchial_biopsy usually implies forceps, here it's needle aspiration
    
    return flags

def generate_spans():
    spans = []
    
    def add_span(text, label, norm_val=None, field=None, event_id=None, prefix=None, index=None, negated=False):
        spans.append({
            "span_text": text,
            "label": label,
            "normalized_value": norm_val if norm_val else text,
            "schema_field": field,
            "event_id": event_id,
            "context_prefix": prefix,
            "match_index": index,
            "is_negated": str(negated).upper()
        })

    # Event 1: Diagnostic Exam (T190)
    add_span("T190 video bronchoscope", "DEV_INSTRUMENT", "T190", "method.device", "evt1")
    add_span("tracheobronchial tree", "ANAT_AIRWAY", "tracheobronchial tree", "target.anatomy", "evt1")
    add_span("vocal cords were normal", "OBS_LESION", "normal", "findings", "evt1")
    add_span("subglottic space was normal", "OBS_LESION", "normal", "findings", "evt1")
    add_span("trachea was of normal caliber", "OBS_LESION", "normal", "findings", "evt1")
    add_span("Bronchial mucosa and anatomy were normal", "OBS_LESION", "normal", "findings", "evt1")
    add_span("no endobronchial lesions", "OBS_LESION", "none", "findings", "evt1", negated=True)

    # Event 2: EBUS-TBNA (4R)
    add_span("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "UC180F", "method.device", "evt2")
    add_span("mediastinal mass", "OBS_LESION", "mediastinal mass", "lesion.type", "evt2")
    add_span("4R", "ANAT_LN_STATION", "4R", "target.station", "evt2")
    add_span("transbronchial needle aspiration", "PROC_METHOD", "TBNA", "method", "evt2")
    add_span("Olympus Vizishot 2 EBUS TBNA 22 gauge needle", "DEV_NEEDLE", "Olympus Vizishot 2", "method.device", "evt2")
    add_span("22 gauge", "MEAS_SIZE", "22G", "device.size", "evt2")
    add_span("5 biopsies", "MEAS_COUNT", "5", "procedure.count", "evt2")
    add_span("pus and inflammatory cells", "OBS_ROSE", "pus/inflammatory cells", "findings.rose", "evt2")
    add_span("Cytology", "PROC_ACTION", "Cytology", "specimens", "evt2")
    
    # Event 3: Clearance/Exit (T180)
    add_span("T180 video bronchoscope", "DEV_INSTRUMENT", "T180", "method.device", "evt3")
    add_span("no evidence of active bleeding", "OUTCOME_COMPLICATION", "none", "outcomes.complications", "evt3", negated=True)

    # Global
    add_span("no immediate complications", "OUTCOME_COMPLICATION", "none", "outcomes.complications", "global", negated=True)

    return spans

def generate_events():
    events = [
        {
            "event_id": "evt1",
            "type": "diagnostic_bronchoscopy",
            "method": "Bronchoscopy",
            "devices_json": json.dumps(["T190 video bronchoscope"]),
            "findings_json": json.dumps(["normal vocal cords", "normal subglottic space", "normal trachea", "no endobronchial lesions"]),
            "evidence_quote": "The T190 video bronchoscope was introduced... Bronchial mucosa and anatomy were normal"
        },
        {
            "event_id": "evt2",
            "type": "linear_ebus",
            "method": "TBNA",
            "target.station": "4R",
            "lesion.type": "mediastinal mass",
            "devices_json": json.dumps(["UC180F", "Olympus Vizishot 2 EBUS TBNA 22 gauge needle"]),
            "specimens_json": json.dumps(["Cytology", "Bacterial culture", "Fungal culture", "Acid Fast culture", "Tissue culture"]),
            "findings_json": json.dumps(["pus and inflammatory cells"]),
            "evidence_quote": "Mediastinal mass located in 4R... Sampling by transbronchial needle aspiration... 5 biopsies... ROSE yielded pus"
        },
        {
            "event_id": "evt3",
            "type": "diagnostic_bronchoscopy",
            "method": "Bronchoscopy",
            "devices_json": json.dumps(["T180 video bronchoscope"]),
            "findings_json": json.dumps(["no active bleeding"]),
            "evidence_quote": "T180 video bronchoscope was then re-inserted... no evidence of active bleeding"
        }
    ]
    return events

# -------------------------------------------------------------------------
# MAIN EXECUTION
# -------------------------------------------------------------------------
def main():
    if not os.path.exists(TEMPLATE_PATH):
        # Fallback to create a blank workbook if template missing (for robustness in sandbox)
        wb = openpyxl.Workbook()
        # Create necessary sheets
        for name in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            wb.create_sheet(name)
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # 1. Populate Note_Text
    ws_text = wb["Note_Text"]
    if ws_text.max_row == 1:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Populate Note_Index
    ws_index = wb["Note_Index"]
    # Check if header exists, else write it
    idx_headers = ["source_file", "note_id", "encounter_id", "procedure_date", "site", "reviewer", "status", "free_text_notes"] + PROCEDURE_FLAGS
    if ws_index.max_row == 0 or (ws_index.max_row == 1 and ws_index.cell(1,1).value is None):
        ws_index.append(idx_headers)
    
    flags = generate_flags()
    row_data = [SOURCE_FILE, NOTE_ID, "", "", "", "", "generated", ""] + [flags[f] for f in PROCEDURE_FLAGS]
    ws_index.append(row_data)

    # 3. Populate Span_Annotations (Anchor First)
    ws_span = wb["Span_Annotations"]
    span_headers = ["source_file", "note_id", "span_id", "section_type", 
                    "context_prefix", "span_text", "match_index", 
                    "start_char", "end_char", "span_len", 
                    "label", "normalized_value", "schema_field", "event_id", 
                    "is_negated", "is_historical", "time_anchor", "reviewer", "comments", 
                    "hydration_status"]
    
    if ws_span.max_row == 0:
        ws_span.append(span_headers)

    raw_spans = generate_spans()
    
    # 4. Populate Span_Hydrated & Write Anchor Spans
    ws_hydra = wb["Span_Hydrated"]
    if ws_hydra.max_row == 0:
        ws_hydra.append(span_headers)

    for i, s in enumerate(raw_spans):
        span_id = f"{NOTE_ID}_s{i+1:03d}"
        
        # Calculate hydration
        start, end, status = find_offsets_anchor_first(
            NOTE_TEXT, 
            s["span_text"], 
            s.get("context_prefix"), 
            s.get("match_index")
        )
        
        common_data = {
            "source_file": SOURCE_FILE,
            "note_id": NOTE_ID,
            "span_id": span_id,
            "section_type": "Procedure",
            "context_prefix": s.get("context_prefix", ""),
            "span_text": s["span_text"],
            "match_index": s.get("match_index", ""),
            "span_len": len(s["span_text"]),
            "label": s["label"],
            "normalized_value": s["normalized_value"],
            "schema_field": s.get("schema_field", ""),
            "event_id": s.get("event_id", ""),
            "is_negated": s.get("is_negated", "FALSE"),
            "is_historical": "FALSE",
            "time_anchor": "",
            "reviewer": "auto",
            "comments": "",
        }

        # Write to Span_Annotations (Blank offsets)
        row_anchor = [common_data[h] if h not in ["start_char", "end_char", "hydration_status"] else "" for h in span_headers]
        row_anchor[-1] = "needs_hydration" # Set status
        ws_span.append(row_anchor)

        # Write to Span_Hydrated (Filled offsets)
        row_hydra = [common_data[h] if h not in ["start_char", "end_char", "hydration_status"] else "" for h in span_headers]
        row_hydra[span_headers.index("start_char")] = start if start is not None else ""
        row_hydra[span_headers.index("end_char")] = end if end is not None else ""
        row_hydra[span_headers.index("hydration_status")] = status
        ws_hydra.append(row_hydra)

    # 5. Event Log (with V3 columns)
    ws_event = wb["Event_Log"]
    event_headers = ["source_file", "note_id", "event_id", "event_type", "method", 
                     "anatomy_target", "device", "needle_gauge", "stations", "counts", "measurements", 
                     "specimens", "findings", "is_historical", "reviewer", "comments",
                     "device_size", "device_material",
                     "outcome_airway_lumen_pre", "outcome_airway_lumen_post",
                     "outcome_symptoms", "outcome_pleural", "outcome_complication"]
    
    if ws_event.max_row == 0:
        ws_event.append(event_headers)

    v3_events = generate_events()
    
    for ev in v3_events:
        # Map V3 structure to flat Event Log columns
        # This is a best-effort flattening for the legacy/hybrid sheet
        flat_row = {
            "source_file": SOURCE_FILE,
            "note_id": NOTE_ID,
            "event_id": ev["event_id"],
            "event_type": ev["type"],
            "method": ev.get("method", ""),
            "anatomy_target": ev.get("target.station", "") or ev.get("target.anatomy", ""),
            "device": ev.get("devices_json", ""),
            "findings": ev.get("findings_json", ""),
            "outcome_complication": "none" if "no evidence of active bleeding" in ev.get("findings_json", "") else ""
        }
        
        row_out = [flat_row.get(h, "") for h in event_headers]
        ws_event.append(row_out)

    # 6. V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    v3_headers = ["note_id", "event_id", "type", 
                  "target.anatomy_type", "target.location.lobe", "target.location.segment", "target.station",
                  "lesion.type", "lesion.size_mm", 
                  "method", "devices_json", "measurements_json", "specimens_json", "findings_json", "evidence_quote",
                  "stent.size", "stent.material_or_brand", "catheter.size_fr",
                  "outcomes.airway.lumen_pre", "outcomes.airway.lumen_post",
                  "outcomes.symptoms", "outcomes.pleural", "outcomes.complications"]
    
    if ws_v3.max_row == 0:
        ws_v3.append(v3_headers)

    for ev in v3_events:
        row_v3 = []
        # Mapping logic
        row_v3.append(NOTE_ID)
        row_v3.append(ev["event_id"])
        row_v3.append(ev["type"])
        
        # Anatomy
        row_v3.append("lymph_node" if "target.station" in ev else "airway") # anatomy_type
        row_v3.append("") # lobe
        row_v3.append("") # segment
        row_v3.append(ev.get("target.station", ""))
        
        # Lesion
        row_v3.append(ev.get("lesion.type", ""))
        row_v3.append("") # size_mm
        
        # Method/Dev/etc
        row_v3.append(ev.get("method", ""))
        row_v3.append(ev.get("devices_json", ""))
        row_v3.append(ev.get("measurements_json", ""))
        row_v3.append(ev.get("specimens_json", ""))
        row_v3.append(ev.get("findings_json", ""))
        row_v3.append(ev.get("evidence_quote", ""))
        
        # Stent/Cath
        row_v3.append("")
        row_v3.append("")
        row_v3.append("")
        
        # Outcomes
        row_v3.append("")
        row_v3.append("")
        row_v3.append("")
        row_v3.append("")
        row_v3.append("None" if ev.get("event_id") == "evt3" else "") # complications logic for evt3

        ws_v3.append(row_v3)

    # 7. JSON Output
    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 0:
        ws_json.append(["json_object"])
    
    registry_data = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": v3_events,
        "no_immediate_complications": True
    }
    
    ws_json.append([json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Successfully generated {OUTPUT_PATH}")

if __name__ == "__main__":
    main()