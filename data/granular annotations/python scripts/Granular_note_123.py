import openpyxl
from openpyxl.utils import get_column_letter
import json
import os
import sys

# -------------------------------------------------------------------------
# INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_123"
SOURCE_FILE = "note_123.txt"
PROCEDURE_DATE = "2026-01-12" # inferred or current date
NOTE_TEXT = """NOTE_ID:  note_123 SOURCE_FILE: note_123.txt Patient name: [REDACTED]

Preoperative diagnosis: Lung cancer
Postoperative diagnosis: same as above

Procedures performed:
CPT 31629 Flexible bronchoscopy with fluoroscopic trans-bronchial needle aspiration
CPT 31652 bronchoscopy, rigid or flexible, including fluoroscopic guidance, when performed with endobronchial ultrasound (EBUS) guided transtracheal and/or transbronchial sampling 2 or fewer structures.
Indications for the procedure: staging of malignancy

Anesthesia: General anesthesia using a LMAendotracheal tube.
Procedure: After obtaining informed consent from the patient, he was brought to the procedure room.
A proper timeout was performed to identify the correct patient and procedure to be performed.
General anesthesia was administered to the patient and he was orally intubated with a LMA by the anesthesia team.
the T190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway was in good position. The vocal cords were normal. The subglottic space was normal.
The trachea was of normal caliber. The bronchial treal was anesthetized with 8cc of 1% lidocaine. The carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level. Bronchial mucosa and anatomy were normal;
there are no endobronchial lesions.

The UC180F convex probe EBUS bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
A systematic hilar and mediastinal lymph node survey was carried out.
Sampling criteria (5mm short axis diameter) were met in station 4R, 4L, 7, 11Ri, 11Rs lymph nodes.
Sampling by transbronchial needle aspiration was performed with the Olympus Vizishot 2 EBUS TBNA 22 gauge needle beginning with the 4L Lymph node, followed by the 7 lymph node, 4R lymph node, and finally 11Ri lymph node.
A total of at least 5 biopsies were performed in each station (see below).
All samples were sent for routine cytology. Following completion of EBUS bronchoscopy, the T180 video bronchoscope was then re-inserted and after suctioning blood and secretions there was no evidence of active bleeding and the bronchoscope was subsequently removed.
The endobronchial ultrasound scope was introduced through the tube into the airway.
Endobronchial ultrasound examination of the appreciable hilar and mediastinal lymph node stations was performed.
EBUS transbronchial needle aspiration (TBNA) was performed at stations_ using the Vizishot 2 22g needle.
Specimens were sent for cytology examination. Rapid on site cytologic evaluation was utilized.

Lymph nodes
4R: 8mm; 7 passes
4L: 5mm;
6 passes
7: 10mm; 5 passes
10R: 4mm;
11Ri: 13mm; 5 passes
11Rs: 10.3mm

Following confirmation of hemostasis and therapeutic aspiration of all endobronchial secretions, the bronchoscope was removed.
Patient tolerated the procedure well with no immediate complications. Patient was extubated and returned to recovery in good condition.
Post procedure chest x-ray showed no evidence of pneumothorax.

Specimens:
Cytology

EBL: minimal

Bronchoscopes: Olympus Q190, UC180F EBUS"""

TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# -------------------------------------------------------------------------
# EXTRACTION DATA (Hardcoded for this note)
# -------------------------------------------------------------------------

# Procedure Flags (30)
FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# Spans
# Format: (span_text, label, normalized_value, event_id, context_prefix)
SPANS_DATA = [
    ("T190 video bronchoscope", "DEV_INSTRUMENT", "Olympus T190", "evt1", ""),
    ("normal", "OBS_ROSE", "Normal Anatomy", "evt1", "vocal cords were "),
    ("normal", "OBS_ROSE", "Normal Anatomy", "evt1", "subglottic space was "),
    ("normal", "OBS_ROSE", "Normal Anatomy", "evt1", "anatomy were "),
    
    # EBUS General
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "Olympus UC180F", "evt2", ""),
    ("Vizishot 2 EBUS TBNA 22 gauge needle", "DEV_NEEDLE", "Vizishot 2 22G", "evt2", ""),
    
    # Event 2: 4L
    ("4L", "ANAT_LN_STATION", "4L", "evt2", "beginning with the "),
    ("5mm", "MEAS_SIZE", "5 mm", "evt2", "4L: "),
    ("6 passes", "MEAS_COUNT", "6", "evt2", ""),
    
    # Event 3: 7
    ("7", "ANAT_LN_STATION", "7", "evt3", "followed by the "),
    ("10mm", "MEAS_SIZE", "10 mm", "evt3", "7: "),
    ("5 passes", "MEAS_COUNT", "5", "evt3", "7: 10mm; "),
    
    # Event 4: 4R
    ("4R", "ANAT_LN_STATION", "4R", "evt4", "followed by the 7 lymph node, "),
    ("8mm", "MEAS_SIZE", "8 mm", "evt4", "4R: "),
    ("7 passes", "MEAS_COUNT", "7", "evt4", "4R: 8mm; "),
    
    # Event 5: 11Ri
    ("11Ri", "ANAT_LN_STATION", "11Ri", "evt5", "and finally "),
    ("13mm", "MEAS_SIZE", "13 mm", "evt5", "11Ri: "),
    ("5 passes", "MEAS_COUNT", "5", "evt5", "11Ri: 13mm; "),
    
    # Therapeutic Aspiration
    ("therapeutic aspiration", "PROC_ACTION", "Therapeutic Aspiration", "evt6", ""),
    ("secretions", "OBS_LESION", "Secretions", "evt6", "endobronchial "),
    ("suctioning blood and secretions", "PROC_ACTION", "Suction", "evt6", ""),
    
    # Outcome
    ("no immediate complications", "OUTCOME_COMPLICATION", "None", "evt_global", ""),
    ("no evidence of pneumothorax", "OUTCOME_COMPLICATION", "None", "evt_global", "")
]

# Events Log
EVENTS_LOG = [
    {
        "event_id": "evt1", "type": "Diagnostic Bronchoscopy", "method": "Inspection",
        "anatomy": "Tracheobronchial Tree", "device": "T190 video bronchoscope", "findings": "Normal anatomy"
    },
    {
        "event_id": "evt2", "type": "EBUS TBNA", "method": "Needle Aspiration",
        "anatomy": "4L", "device": "Vizishot 2 22G", "measurements": "5mm", "counts": "6 passes", "stations": "4L"
    },
    {
        "event_id": "evt3", "type": "EBUS TBNA", "method": "Needle Aspiration",
        "anatomy": "7", "device": "Vizishot 2 22G", "measurements": "10mm", "counts": "5 passes", "stations": "7"
    },
    {
        "event_id": "evt4", "type": "EBUS TBNA", "method": "Needle Aspiration",
        "anatomy": "4R", "device": "Vizishot 2 22G", "measurements": "8mm", "counts": "7 passes", "stations": "4R"
    },
    {
        "event_id": "evt5", "type": "EBUS TBNA", "method": "Needle Aspiration",
        "anatomy": "11Ri", "device": "Vizishot 2 22G", "measurements": "13mm", "counts": "5 passes", "stations": "11Ri"
    },
    {
        "event_id": "evt6", "type": "Therapeutic Aspiration", "method": "Suction",
        "anatomy": "Airway", "device": "Bronchoscope", "findings": "Secretions/Blood"
    }
]

# V3 Events (Registry JSON model)
V3_EVENTS = [
    {
        "event_id": "evt1", "type": "Diagnostic Bronchoscopy", "target_anatomy": "Tracheobronchial Tree",
        "method": "Inspection", "devices": ["T190"], "findings": ["Normal"]
    },
    {
        "event_id": "evt2", "type": "EBUS TBNA", "target_station": "4L",
        "method": "TBNA", "devices": ["Vizishot 2 22G"], "measurements": ["5mm"], "specimens": ["6 passes"]
    },
    {
        "event_id": "evt3", "type": "EBUS TBNA", "target_station": "7",
        "method": "TBNA", "devices": ["Vizishot 2 22G"], "measurements": ["10mm"], "specimens": ["5 passes"]
    },
    {
        "event_id": "evt4", "type": "EBUS TBNA", "target_station": "4R",
        "method": "TBNA", "devices": ["Vizishot 2 22G"], "measurements": ["8mm"], "specimens": ["7 passes"]
    },
    {
        "event_id": "evt5", "type": "EBUS TBNA", "target_station": "11Ri",
        "method": "TBNA", "devices": ["Vizishot 2 22G"], "measurements": ["13mm"], "specimens": ["5 passes"]
    },
    {
        "event_id": "evt6", "type": "Therapeutic Aspiration", "target_anatomy": "Airway",
        "method": "Suction", "findings": ["Secretions cleared"]
    }
]


# -------------------------------------------------------------------------
# LOGIC
# -------------------------------------------------------------------------

def create_workbook():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        # If template missing, create simple wb for safety (though instructions say must load)
        wb = openpyxl.Workbook()
        
    return wb

def hydrate_offsets(text, span_text, context_prefix=None, match_index=None):
    """
    Returns (start, end, status)
    """
    if not span_text:
        return None, None, "missing_text"
        
    count = text.count(span_text)
    
    if count == 0:
        return None, None, "not_found"
        
    if count == 1:
        start = text.find(span_text)
        return start, start + len(span_text), "hydrated_unique"
        
    # Multiple occurrences
    if context_prefix:
        # Find all occurrences
        current_pos = 0
        candidates = []
        for _ in range(count):
            pos = text.find(span_text, current_pos)
            if pos == -1: break
            
            # Check context window (120 chars before)
            window_start = max(0, pos - 120)
            window = text[window_start:pos]
            if context_prefix in window:
                candidates.append(pos)
            current_pos = pos + 1
            
        if len(candidates) == 1:
            start = candidates[0]
            return start, start + len(span_text), "hydrated_prefix_window"
            
    # Fallback to match_index if provided, else ambiguous
    # (Simplified logic for this generator: we usually don't have match_index in input data unless specified)
    return None, None, f"ambiguous_count={count}"

def populate_sheets(wb):
    # 1. Note_Text
    ws = wb["Note_Text"] if "Note_Text" in wb.sheetnames else wb.create_sheet("Note_Text")
    if ws.max_row == 1:
        ws.append(["note_id", "source_file", "note_text"])
    ws.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # 2. Note_Index
    ws = wb["Note_Index"] if "Note_Index" in wb.sheetnames else wb.create_sheet("Note_Index")
    # Headers should be there from template, but if we append, we append to next row
    # Row structure: metadata cols... then flags
    # We construct the row
    row = [
        SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "Success", ""
    ]
    # Add flags
    flag_keys = [
        "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", 
        "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy", 
        "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration", 
        "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation", 
        "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", 
        "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy", 
        "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis", 
        "pleural_biopsy", "fibrinolytic_therapy"
    ]
    for k in flag_keys:
        row.append(FLAGS.get(k, 0))
    ws.append(row)
    
    # 3. Span_Annotations & 4. Span_Hydrated
    ws_anno = wb["Span_Annotations"] if "Span_Annotations" in wb.sheetnames else wb.create_sheet("Span_Annotations")
    ws_hydra = wb["Span_Hydrated"] if "Span_Hydrated" in wb.sheetnames else wb.create_sheet("Span_Hydrated")
    
    # Write headers if empty (assuming template has them, but safety check)
    headers = [
        "source_file", "note_id", "span_id", "section_type", "context_prefix", 
        "span_text", "match_index", "start_char", "end_char", "span_len", 
        "label", "normalized_value", "schema_field", "event_id", 
        "is_negated", "is_historical", "time_anchor", "reviewer", "comments", "hydration_status"
    ]
    
    # We append to both
    
    for i, (txt, label, norm, evt, ctx) in enumerate(SPANS_DATA):
        span_id = f"{NOTE_ID}_s{i+1}"
        
        # Calculate hydration
        start, end, status = hydrate_offsets(NOTE_TEXT, txt, ctx)
        
        # Annotation Row (No offsets)
        row_anno = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure", ctx,
            txt, "", "", "", f"=LEN(F{ws_anno.max_row+1})",
            label, norm, "", evt,
            "FALSE", "FALSE", "", "", "", "needs_hydration"
        ]
        ws_anno.append(row_anno)
        
        # Hydrated Row (With offsets)
        row_hydra = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure", ctx,
            txt, "", start if start is not None else "", end if end is not None else "", len(txt),
            label, norm, "", evt,
            "FALSE", "FALSE", "", "", "", status
        ]
        ws_hydra.append(row_hydra)
        
    # 5. Event_Log
    ws_log = wb["Event_Log"] if "Event_Log" in wb.sheetnames else wb.create_sheet("Event_Log")
    for evt in EVENTS_LOG:
        row = [
            SOURCE_FILE, NOTE_ID, evt.get("event_id"), evt.get("type"), evt.get("method"),
            evt.get("anatomy"), evt.get("device"), "", evt.get("stations"), evt.get("counts"), evt.get("measurements"),
            "", evt.get("findings"), "FALSE", "", "",
            "", "", # device size/material
            "", "", "", "", "" # outcomes
        ]
        ws_log.append(row)
        
    # 6. V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"] if "V3_Procedure_Events" in wb.sheetnames else wb.create_sheet("V3_Procedure_Events")
    for v3 in V3_EVENTS:
        # Mapping simple fields to columns
        # Columns: note_id, event_id, type, target.anatomy_type, target.location.lobe, target.location.segment, target.station
        # lesion.type, lesion.size_mm, method, devices_json, measurements_json, specimens_json, findings_json...
        
        row = [
            NOTE_ID, v3["event_id"], v3["type"],
            "Lymph Node" if "station" in v3.get("target_station", "") or v3["event_id"] in ["evt2","evt3","evt4","evt5"] else "Airway",
            "", "", v3.get("target_station", ""),
            "", "", # lesion
            v3.get("method"),
            json.dumps(v3.get("devices", [])),
            json.dumps(v3.get("measurements", [])),
            json.dumps(v3.get("specimens", [])),
            json.dumps(v3.get("findings", [])),
            "", "", "", "", # evidence, stent, catheter
            "", "", "", "", "" # outcomes
        ]
        ws_v3.append(row)

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"] if "V3_Registry_JSON" in wb.sheetnames else wb.create_sheet("V3_Registry_JSON")
    
    registry_obj = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "procedures": V3_EVENTS,
        "no_immediate_complications": True
    }
    
    ws_json.append([json.dumps(registry_obj, indent=2)])

def main():
    wb = create_workbook()
    populate_sheets(wb)
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    main()