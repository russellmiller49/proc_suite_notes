import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os

# =============================================================================
# INPUT DATA
# =============================================================================
NOTE_ID = "note_126"
SOURCE_FILE = "note_126.txt"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_126 SOURCE_FILE: note_126.txt Patient name: [REDACTED]

Preoperative diagnosis: 1:Lung mass
Postoperative diagnosis: same as above

Procedures performed:
CPT + 31633 Bronchoscopy, rigid or flexible, including fluoroscopic guidance, when performed;
with transbronchial needle aspiration biopsy(s), each additional lobe
CPT 31653 bronchoscopy, rigid or flexible, including fluoroscopic guidance, when performed with endobronchial ultrasound (EBUS) guided transtracheal and/or transbronchial sampling 3 or more structures.
Indications for the procedure: high suspicion for malignancy staging of malignancy

Anesthesia: General anesthesia using a LMA

Procedure: After obtaining informed consent from the patient the patient was brought to the procedure room.
A proper timeout was performed to identify the correct patient and procedure to be performed.
General anesthesia was administered to the patient and_was orally intubated with a LMA by the anesthesia team.
the diagnostic video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway was in good position. The vocal cords were normal. The subglottic space was normal.
The trachea was of normal caliber. The carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level. Bronchial mucosa and anatomy were normal;
there are no endobronchial lesions.

The convex probe EBUS bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
A systematic hilar and mediastinal lymph node survey was carried out.
Sampling criteria (5mm short axis diameter PET avid) were met in station 11 L, 7, 3P lymph nodes.
Sampling by transbronchial needle aspiration was performed with the EBUS TBNA 21 gauge needle beginning with the 11 L lymph node, followed by the 7 lymph node, followed by the 3P lymph node.
A total of at least 5 biopsies were performed in each station.
ROSE evaluation yielded malignancy at station 7 and 3P. All samples were sent for routine cytology.
Following completion of EBUS bronchoscopy, the video bronchoscope was then re-inserted and after suctioning blood and secretions there was no evidence of active bleeding and the bronchoscope was subsequently removed.
Lymph nodes
3p: 4.0 mm; 4 passes
4R: <3mm; 
4L: 3mm; 
7: 7.6; 7passes
11Rs: 4.9 mm; 
11Ri: 3.5 mm;
11L: 5.4mm;
6of passes

Following confirmation of hemostasis and therapeutic aspiration of all endobronchial secretions, the bronchoscope was removed.
Patient tolerated the procedure well with no immediate complications. Patient was extubated and returned to recovery in good condition.
Specimens:
Cytology

EBL: 10cc

Bronchoscopes: Olympus H190, convex EBUS"""

# =============================================================================
# CONFIGURATION & UTILS
# =============================================================================

PROCEDURE_FLAGS_LIST = [
    "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy",
    "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy",
    "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration",
    "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation",
    "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation",
    "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy",
    "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis",
    "pleural_biopsy", "fibrinolytic_therapy"
]

def determine_flags(text):
    flags = {k: 0 for k in PROCEDURE_FLAGS_LIST}
    text_lower = text.lower()
    
    # Logic for Note 126
    flags["diagnostic_bronchoscopy"] = 1 # "diagnostic video bronchoscope"
    flags["linear_ebus"] = 1 # "convex probe EBUS"
    flags["therapeutic_aspiration"] = 1 # "therapeutic aspiration of all endobronchial secretions"
    
    # Explicit exclusions
    flags["tbna_conventional"] = 0 # TBNA was EBUS guided
    flags["rigid_bronchoscopy"] = 0 # "LMA", "flexible" implied by video bronchoscope/EBUS context
    
    return flags

def hydrate_span(text, span_text, context_prefix=None, match_index=None):
    """
    Finds start_char and end_char for a span.
    Returns: (start, end, status_msg)
    """
    if not span_text or span_text not in text:
        return None, None, "not_found"
    
    matches = [m.start() for m in re.finditer(re.escape(span_text), text)]
    count = len(matches)
    
    if count == 0:
        return None, None, "not_found"
    
    selected_start = None
    status = "ambiguous"
    
    # Strategy 1: Unique match
    if count == 1:
        selected_start = matches[0]
        status = "hydrated_unique"
    
    # Strategy 2: Context Prefix
    elif context_prefix:
        # Search for occurrence where context appears within 120 chars before
        for m_start in matches:
            search_window_start = max(0, m_start - 120)
            preceding_text = text[search_window_start:m_start]
            if context_prefix in preceding_text:
                selected_start = m_start
                status = "hydrated_prefix_window"
                break
    
    # Strategy 3: Match Index
    if selected_start is None and match_index is not None:
        if 0 <= match_index < count:
            selected_start = matches[match_index]
            status = "hydrated_match_index"
    
    # Fallback/Fail
    if selected_start is None:
        return None, None, f"ambiguous_count={count}"
        
    return selected_start, selected_start + len(span_text), status

# =============================================================================
# DATA EXTRACTION
# =============================================================================

# Spans Storage
# Row structure: [span_text, context_prefix, label, norm_value, schema_field, event_id]
spans_data = []

# Event IDs
EVT_INSPECT = "evt_01"
EVT_TBNA_11L = "evt_02"
EVT_TBNA_7 = "evt_03"
EVT_TBNA_3P = "evt_04"
EVT_ASPIRATE = "evt_05"
EVT_OUTCOME = "evt_06"

# 1. Inspection
spans_data.append({
    "text": "diagnostic video bronchoscope",
    "label": "DEV_INSTRUMENT",
    "norm": "Flexible Bronchoscope",
    "event_id": EVT_INSPECT
})
spans_data.append({
    "text": "Bronchial mucosa and anatomy were normal",
    "label": "OBS_ROSE", # Using broadly for observation
    "norm": "Normal",
    "event_id": EVT_INSPECT
})
spans_data.append({
    "text": "no endobronchial lesions",
    "label": "OBS_LESION",
    "norm": "None",
    "event_id": EVT_INSPECT,
    "is_negated": "TRUE"
})

# 2. EBUS Instrument
spans_data.append({
    "text": "convex probe EBUS bronchoscope",
    "label": "DEV_INSTRUMENT",
    "norm": "Linear EBUS",
    "event_id": EVT_TBNA_11L # Associated with first TBNA event generally
})

# 3. TBNA 11L
spans_data.append({
    "text": "11 L lymph node",
    "context": "beginning with the",
    "label": "ANAT_LN_STATION",
    "norm": "11L",
    "event_id": EVT_TBNA_11L
})
spans_data.append({
    "text": "EBUS TBNA 21 gauge needle",
    "label": "DEV_NEEDLE",
    "norm": "21G",
    "event_id": EVT_TBNA_11L
})
spans_data.append({
    "text": "11L: 5.4mm",
    "label": "MEAS_SIZE",
    "norm": "5.4 mm",
    "event_id": EVT_TBNA_11L
})
# Note: "6of passes" likely refers to 11L based on layout at end of note, but text says "5 biopsies... in each station".
# The footer "11L: 5.4mm;\n6of passes" suggests 6 passes for 11L.
spans_data.append({
    "text": "6of passes",
    "label": "MEAS_COUNT",
    "norm": "6",
    "event_id": EVT_TBNA_11L
})

# 4. TBNA 7
spans_data.append({
    "text": "7 lymph node",
    "context": "followed by the",
    "label": "ANAT_LN_STATION",
    "norm": "7",
    "event_id": EVT_TBNA_7
})
spans_data.append({
    "text": "7: 7.6",
    "label": "MEAS_SIZE",
    "norm": "7.6 mm",
    "event_id": EVT_TBNA_7
})
spans_data.append({
    "text": "7passes",
    "label": "MEAS_COUNT",
    "norm": "7",
    "event_id": EVT_TBNA_7
})
spans_data.append({
    "text": "yielded malignancy at station 7",
    "label": "OBS_ROSE",
    "norm": "Malignancy",
    "event_id": EVT_TBNA_7
})

# 5. TBNA 3P
spans_data.append({
    "text": "3P lymph node",
    "context": "followed by the",
    "label": "ANAT_LN_STATION",
    "norm": "3P",
    "event_id": EVT_TBNA_3P
})
spans_data.append({
    "text": "3p: 4.0 mm",
    "label": "MEAS_SIZE",
    "norm": "4.0 mm",
    "event_id": EVT_TBNA_3P
})
spans_data.append({
    "text": "4 passes",
    "label": "MEAS_COUNT",
    "norm": "4",
    "event_id": EVT_TBNA_3P
})
spans_data.append({
    "text": "yielded malignancy at station 7 and 3P",
    "label": "OBS_ROSE",
    "norm": "Malignancy",
    "event_id": EVT_TBNA_3P
})

# 6. Therapeutic Aspiration
spans_data.append({
    "text": "therapeutic aspiration of all endobronchial secretions",
    "label": "PROC_METHOD",
    "norm": "Therapeutic Aspiration",
    "event_id": EVT_ASPIRATE
})

# 7. Outcome
spans_data.append({
    "text": "no immediate complications",
    "label": "OUTCOME_COMPLICATION",
    "norm": "None",
    "event_id": EVT_OUTCOME,
    "is_negated": "TRUE"
})

# =============================================================================
# WORKBOOK GENERATION
# =============================================================================

def generate_excel():
    if not os.path.exists(TEMPLATE_PATH):
        # Create a dummy template if missing (for standalone robustness, though requirements say it exists)
        wb = openpyxl.Workbook()
        for sheet in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            wb.create_sheet(sheet)
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # 1. Note_Text
    ws_text = wb["Note_Text"]
    if ws_text.max_row == 1:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 2. Note_Index
    ws_index = wb["Note_Index"]
    # Headers implied by template, appending row
    flags = determine_flags(NOTE_TEXT)
    row_meta = [
        SOURCE_FILE, NOTE_ID, "", "", "", "", "Success", ""
    ]
    # Append flags in specific order
    flag_values = [flags[k] for k in PROCEDURE_FLAGS_LIST]
    ws_index.append(row_meta + flag_values)

    # 3. Span_Annotations & 4. Span_Hydrated
    ws_anno = wb["Span_Annotations"]
    ws_hydra = wb["Span_Hydrated"]
    
    # Headers (assuming template has them, but mapping logic below)
    # columns: source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, start, end, len, label, norm, schema, event_id...
    
    for i, span in enumerate(spans_data):
        span_id = f"{NOTE_ID}_s{i+1:02d}"
        s_text = span["text"]
        s_context = span.get("context", "")
        s_label = span["label"]
        s_norm = span.get("norm", "")
        s_event = span["event_id"]
        s_neg = span.get("is_negated", "")
        
        # Hydrate
        start, end, status = hydrate_span(NOTE_TEXT, s_text, s_context)
        
        # Base Row
        row_base = [
            SOURCE_FILE, NOTE_ID, span_id, "Procedure",
            s_context, s_text, "", # match_index
            "", "", f"=LEN(F{ws_anno.max_row+1})", # start, end, len formula
            s_label, s_norm, "", s_event,
            s_neg, "", "", "", "", # is_historical, time, reviewer, comments
            "needs_hydration"
        ]
        ws_anno.append(row_base)
        
        # Hydrated Row
        row_hydra = list(row_base)
        row_hydra[7] = start if start is not None else ""
        row_hydra[8] = end if end is not None else ""
        row_hydra[9] = end - start if start is not None else ""
        row_hydra[-1] = status
        ws_hydra.append(row_hydra)

    # 5. Event_Log
    ws_event = wb["Event_Log"]
    # Group by event_id
    event_groups = {}
    for s in spans_data:
        eid = s["event_id"]
        if eid not in event_groups:
            event_groups[eid] = []
        event_groups[eid].append(s)
        
    for eid, spans in event_groups.items():
        # Summarize event
        etype = "Diagnostic"
        method = ""
        target = ""
        station = ""
        device = ""
        counts = ""
        measurements = ""
        findings = ""
        specimens = ""
        outcome_comp = ""
        
        for s in spans:
            lbl = s["label"]
            val = s["norm"]
            txt = s["text"]
            
            if lbl == "PROC_METHOD": method = val
            if lbl == "ANAT_LN_STATION": station = val
            if lbl == "DEV_INSTRUMENT": device = val
            if lbl == "DEV_NEEDLE": device += f" {val}"
            if lbl == "MEAS_COUNT": counts = val
            if lbl == "MEAS_SIZE": measurements = val
            if lbl == "OBS_ROSE": findings = val
            if lbl == "OUTCOME_COMPLICATION": outcome_comp = val
            
            if "TBNA" in eid.upper(): etype = "TBNA"
            if "ASPIRATE" in eid.upper(): etype = "Therapeutic"
            
        row_evt = [
            SOURCE_FILE, NOTE_ID, eid, etype, method,
            target, device, "", station, counts, measurements,
            specimens, findings, "", "", "", # historical, rev, comm
            "", "", # dev size/mat
            "", "", "", "", outcome_comp # outcomes
        ]
        ws_event.append(row_evt)

    # 6. V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    
    # Construct V3 Objects
    v3_events_list = []
    
    # Event: Inspection
    v3_insp = {
        "event_id": EVT_INSPECT,
        "type": "Bronchoscopy",
        "method": "Inspection",
        "devices_json": json.dumps([{"device": "Flexible Bronchoscope"}]),
        "findings_json": json.dumps(["Normal bronchial mucosa"])
    }
    v3_events_list.append(v3_insp)
    
    # Event: TBNA 11L
    v3_11l = {
        "event_id": EVT_TBNA_11L,
        "type": "TBNA",
        "target": {"station": "11L"},
        "method": "EBUS-TBNA",
        "devices_json": json.dumps([{"device": "Linear EBUS"}, {"device": "21G Needle"}]),
        "measurements_json": json.dumps({"count": 6, "size_mm": 5.4})
    }
    v3_events_list.append(v3_11l)
    
    # Event: TBNA 7
    v3_7 = {
        "event_id": EVT_TBNA_7,
        "type": "TBNA",
        "target": {"station": "7"},
        "method": "EBUS-TBNA",
        "measurements_json": json.dumps({"count": 7, "size_mm": 7.6}),
        "findings_json": json.dumps(["Malignancy"])
    }
    v3_events_list.append(v3_7)
    
    # Event: TBNA 3P
    v3_3p = {
        "event_id": EVT_TBNA_3P,
        "type": "TBNA",
        "target": {"station": "3P"},
        "method": "EBUS-TBNA",
        "measurements_json": json.dumps({"count": 4, "size_mm": 4.0}),
        "findings_json": json.dumps(["Malignancy"])
    }
    v3_events_list.append(v3_3p)
    
    # Event: Aspiration
    v3_asp = {
        "event_id": EVT_ASPIRATE,
        "type": "Therapeutic",
        "method": "Aspiration",
        "target": {"anatomy_type": "Airway"}
    }
    v3_events_list.append(v3_asp)
    
    # Write V3 Rows
    for evt in v3_events_list:
        row_v3 = [
            NOTE_ID, evt.get("event_id"), evt.get("type"),
            "Airway" if evt.get("target", {}).get("anatomy_type") else "", # anatomy_type
            "", "", evt.get("target", {}).get("station", ""), # lobe, seg, station
            "", "", # lesion type/size
            evt.get("method", ""),
            evt.get("devices_json", ""),
            evt.get("measurements_json", ""),
            "", evt.get("findings_json", ""), "",
            "", "", "", # stent info
            "", "", "", "", # outcome airway/pleural
            "None" if evt.get("event_id") == EVT_OUTCOME else "" # comp
        ]
        ws_v3.append(row_v3)

    # 7. V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    json_obj = {
        "schema_version": "v3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True,
        "procedures": v3_events_list
    }
    ws_json.append([json.dumps(json_obj, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_excel()