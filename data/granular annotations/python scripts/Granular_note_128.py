import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
from datetime import datetime

# =============================================================================
# CONFIGURATION & INPUT DATA
# =============================================================================

NOTE_ID = "note_128"
SOURCE_FILE = "note_128.txt"
PROCEDURE_DATE = "2026-01-12" # inferred from context or left blank
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_128 SOURCE_FILE: note_128.txt Patient name:  [REDACTED]

Preoperative diagnosis: Acute hypoxemic respiratory failure; Cough; Esophageal cancer; Malignant esophageal tumor;
Pneumonia
Postoperative diagnosis: Same as above with addition of significant carcinomatosis and lymphangitic/metastatic spread of patient's malignancy causing severe and near complete obstruction primarily affecting the right lung.
Procedures performed:
31622 – Bronchoscopy, flexible, with or without fluoroscopic guidance, diagnostic
Balloon Dilation, 31630, Bronchoscopy rigid or flexible including fluoroscopic guidance when performed;
with tracheal/bronchial dilation

Indications for the procedure: Acute hypoxemic respiratory failure; Cough; Esophageal cancer; Malignant esophageal tumor;
Pneumonia

Anesthesia: General anesthesia using a #8.5endotracheal tube.

Procedure: After obtaining informed consent from the patient, the patient was brought to the procedure room.
A proper timeout was performed to identify the correct patient and procedure to be performed. General anesthesia was administered.
Rigid bronchoscopy was performed with an 11mm ventilating scope. The patient, however, desaturated quickly despite relatively high driving pressures on jet ventilation (to the 60s), requiring extubation.
The patient was bagged to bring saturations up and we intubated with the rigid scope again, but had the same thing happen.
We then intubated with an 8.5 ETT and proceeded.

The T190/P190 video bronchoscope was introduced and advanced to the tracheobronchial tree.
There was significant adenopathy surrounding the vocal cords; however, patent vocal cords were noted on exam with mild erythema of the trachea following advancement.
Left lung without any apparent superficial abnormality, other than secretions noted in the bilateral airway.
Right lung with significant swelling, edema, erythema, and nodularity noted throughout, most prominently in the right middle and right lower lobe with significant secretions aspirated and near complete obstruction and narrowing of the airways.
We dilated the lower airway with an 8-9-10mm Elation balloon to better bypass the proximal tumor before assessing the distal disease.
The dilation was temporary only, and collapse recurred.

Unfortunately, as we advanced past the proximal tumor obstruction with the P190 flexible bronchoscope, we could see tumor infiltration extending into all of the distal subsegments, making no appropriate landing point for stenting.
We decided to terminate the procedure as the airway was not salvageable.

He was then taken to the ICU intubated.
Specimens: None

EBL: minimal

Bronchoscopes: Diagnostic and therapeutic bronchoscope as well as rigid bronchoscope"""

# =============================================================================
# DATA GENERATION LOGIC
# =============================================================================

def get_procedure_flags():
    """Returns the dictionary of 30 binary procedure flags."""
    flags = {
        # Bronchoscopy (23)
        "diagnostic_bronchoscopy": 1, # "T190/P190 video bronchoscope"
        "bal": 0,
        "bronchial_wash": 0,
        "brushings": 0,
        "endobronchial_biopsy": 0,
        "tbna_conventional": 0,
        "linear_ebus": 0,
        "radial_ebus": 0,
        "navigational_bronchoscopy": 0,
        "transbronchial_biopsy": 0,
        "transbronchial_cryobiopsy": 0,
        "therapeutic_aspiration": 1, # "significant secretions aspirated"
        "foreign_body_removal": 0,
        "airway_dilation": 1, # "dilated the lower airway"
        "airway_stent": 0, # "no appropriate landing point for stenting"
        "thermal_ablation": 0,
        "tumor_debulking_non_thermal": 0,
        "cryotherapy": 0,
        "blvr": 0,
        "peripheral_ablation": 0,
        "bronchial_thermoplasty": 0,
        "whole_lung_lavage": 0,
        "rigid_bronchoscopy": 1, # "Rigid bronchoscopy was performed"
        
        # Pleural (7)
        "thoracentesis": 0,
        "chest_tube": 0,
        "ipc": 0,
        "medical_thoracoscopy": 0,
        "pleurodesis": 0,
        "pleural_biopsy": 0,
        "fibrinolytic_therapy": 0,
    }
    return flags

def get_spans():
    """
    Returns a list of dictionaries representing the anchor-first spans.
    Each dict matches the columns expected in Span_Annotations.
    """
    spans = [
        # Event 1: Rigid Bronchoscopy (Attempt 1)
        {
            "span_text": "Rigid bronchoscopy",
            "label": "PROC_METHOD",
            "event_id": "evt_01",
            "context_prefix": "administered.\n",
            "normalized_value": "rigid bronchoscopy"
        },
        {
            "span_text": "11mm ventilating scope",
            "label": "DEV_INSTRUMENT",
            "event_id": "evt_01",
            "normalized_value": "rigid bronchoscope 11mm"
        },
        {
            "span_text": "desaturated quickly",
            "label": "OUTCOME_COMPLICATION",
            "event_id": "evt_01",
            "normalized_value": "desaturation"
        },
        
        # Event 2: Rigid Bronchoscopy (Attempt 2)
        {
            "span_text": "intubated with the rigid scope again",
            "label": "PROC_METHOD",
            "event_id": "evt_02",
            "normalized_value": "rigid bronchoscopy"
        },
        
        # Event 3: Flexible Diagnostic
        {
            "span_text": "T190/P190 video bronchoscope",
            "label": "DEV_INSTRUMENT",
            "event_id": "evt_03",
            "normalized_value": "Olympus P190"
        },
        {
            "span_text": "tracheobronchial tree",
            "label": "ANAT_AIRWAY",
            "event_id": "evt_03",
            "normalized_value": "tracheobronchial tree"
        },
        
        # Event 4: Right Lung Findings / Aspiration
        {
            "span_text": "Right lung",
            "label": "ANAT_LUNG_LOC",
            "event_id": "evt_04",
            "normalized_value": "Right lung"
        },
         {
            "span_text": "right middle",
            "label": "ANAT_LUNG_LOC",
            "event_id": "evt_04",
            "normalized_value": "RML"
        },
         {
            "span_text": "right lower lobe",
            "label": "ANAT_LUNG_LOC",
            "event_id": "evt_04",
            "normalized_value": "RLL"
        },
        {
            "span_text": "near complete obstruction",
            "label": "OUTCOME_AIRWAY_LUMEN_PRE",
            "event_id": "evt_04",
            "context_prefix": "secretions aspirated and ",
            "normalized_value": "90%"
        },
        {
            "span_text": "secretions aspirated",
            "label": "PROC_ACTION",
            "event_id": "evt_04",
            "normalized_value": "therapeutic aspiration"
        },

        # Event 5: Balloon Dilation
        {
            "span_text": "dilated",
            "label": "PROC_ACTION",
            "event_id": "evt_05",
            "normalized_value": "dilation"
        },
        {
            "span_text": "lower airway",
            "label": "ANAT_AIRWAY",
            "event_id": "evt_05",
            "normalized_value": "lower airway"
        },
        {
            "span_text": "8-9-10mm Elation balloon",
            "label": "DEV_INSTRUMENT",
            "event_id": "evt_05",
            "normalized_value": "Elation balloon 8-9-10mm"
        },
         {
            "span_text": "near complete obstruction",
            "label": "OUTCOME_AIRWAY_LUMEN_PRE",
            "event_id": "evt_05",
            "context_prefix": "secretions aspirated and ", # Re-using context to link finding to dilation
            "normalized_value": "90%"
        },
        {
            "span_text": "collapse recurred",
            "label": "OUTCOME_AIRWAY_LUMEN_POST",
            "event_id": "evt_05",
            "normalized_value": "100%" # Recurred means back to obstruction
        },
        {
            "span_text": "temporary only",
            "label": "OBS_FINDING",
            "event_id": "evt_05",
            "normalized_value": "ineffective"
        },

        # Event 6: Termination / Attempted Stent
        {
            "span_text": "advanced past the proximal tumor",
            "label": "PROC_ACTION",
            "event_id": "evt_06",
            "normalized_value": "inspection"
        },
        {
            "span_text": "P190 flexible bronchoscope",
            "label": "DEV_INSTRUMENT",
            "event_id": "evt_06",
            "normalized_value": "Olympus P190"
        },
         {
            "span_text": "tumor infiltration",
            "label": "OBS_LESION",
            "event_id": "evt_06",
            "normalized_value": "tumor infiltration"
        },
        {
            "span_text": "stenting",
            "label": "PROC_METHOD",
            "event_id": "evt_06",
            "is_negated": True,
            "normalized_value": "airway stent"
        },
        {
            "span_text": "airway was not salvageable",
            "label": "OBS_FINDING",
            "event_id": "evt_06",
            "normalized_value": "not salvageable"
        }
    ]
    return spans

def get_event_log():
    """
    Returns lists of row data for the Event_Log sheet.
    """
    return [
        {
            "event_id": "evt_01",
            "event_type": "rigid_bronchoscopy",
            "method": "rigid_bronchoscopy",
            "device": "11mm ventilating scope",
            "outcome_complication": "desaturation requiring extubation",
            "comments": "Patient desaturated despite jet ventilation."
        },
        {
            "event_id": "evt_02",
            "event_type": "rigid_bronchoscopy",
            "method": "rigid_bronchoscopy",
            "comments": "Second attempt, same result (desaturation)."
        },
         {
            "event_id": "evt_03",
            "event_type": "diagnostic_bronchoscopy",
            "method": "flexible",
            "device": "T190/P190 video bronchoscope",
            "findings": "Adenopathy, patent vocal cords, mild erythema trachea."
        },
        {
            "event_id": "evt_04",
            "event_type": "therapeutic_aspiration",
            "anatomy_target": "Right lung (RML, RLL)",
            "method": "aspiration",
            "outcome_airway_lumen_pre": "near complete obstruction",
            "findings": "Significant secretions aspirated."
        },
        {
            "event_id": "evt_05",
            "event_type": "airway_dilation",
            "method": "balloon_dilation",
            "anatomy_target": "Right lower airway",
            "device": "8-9-10mm Elation balloon",
            "outcome_airway_lumen_pre": "near complete",
            "outcome_airway_lumen_post": "collapse recurred",
            "comments": "Temporary only."
        },
        {
            "event_id": "evt_06",
            "event_type": "diagnostic_bronchoscopy",
            "method": "flexible",
            "findings": "Tumor infiltration extending into all distal subsegments.",
            "comments": "Procedure terminated, airway not salvageable, no stenting possible."
        }
    ]

# =============================================================================
# HYDRATION ENGINE
# =============================================================================

def find_offsets(text, span_text, context_prefix=None, match_index=0):
    """
    Calculates start_char and end_char based on constraints.
    Returns (start, end, status_msg).
    """
    if not span_text:
        return None, None, "missing_text"
    
    # 1. Exact count check
    count = text.count(span_text)
    
    if count == 0:
        return None, None, "not_found"
    
    if count == 1:
        start = text.find(span_text)
        return start, start + len(span_text), "hydrated_unique"
    
    # 2. Context Match (if prefix provided)
    if context_prefix:
        # Find all occurrences
        starts = [m.start() for m in re.finditer(re.escape(span_text), text)]
        for s in starts:
            # Look back 120 chars
            window_start = max(0, s - 120)
            window = text[window_start:s]
            if context_prefix in window:
                return s, s + len(span_text), "hydrated_prefix_window"
        return None, None, "context_not_found"

    # 3. Match Index (fallback)
    if count > 1:
        starts = [m.start() for m in re.finditer(re.escape(span_text), text)]
        if 0 <= match_index < len(starts):
            s = starts[match_index]
            return s, s + len(span_text), "hydrated_match_index"
        else:
            return None, None, f"index_out_of_bounds_count_{count}"

    return None, None, "ambiguous_unknown"

# =============================================================================
# EXCEL GENERATION
# =============================================================================

class WorkbookGenerator:
    def __init__(self):
        self.wb = openpyxl.load_workbook(TEMPLATE_PATH)
        self.note_text = NOTE_TEXT
        
    def write_note_text(self):
        sheet = self.wb["Note_Text"]
        # Determine next empty row
        row = sheet.max_row + 1
        sheet.cell(row=row, column=1, value=NOTE_ID)
        sheet.cell(row=row, column=2, value=SOURCE_FILE)
        sheet.cell(row=row, column=3, value=self.note_text)

    def write_note_index(self):
        sheet = self.wb["Note_Index"]
        row = sheet.max_row + 1
        
        # Metadata columns
        data = [
            SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "", "complete", ""
        ]
        
        # Flags
        flags = get_procedure_flags()
        flag_order = [
            "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", 
            "endobronchial_biopsy", "tbna_conventional", "linear_ebus", "radial_ebus", 
            "navigational_bronchoscopy", "transbronchial_biopsy", "transbronchial_cryobiopsy", 
            "therapeutic_aspiration", "foreign_body_removal", "airway_dilation", 
            "airway_stent", "thermal_ablation", "tumor_debulking_non_thermal", 
            "cryotherapy", "blvr", "peripheral_ablation", "bronchial_thermoplasty", 
            "whole_lung_lavage", "rigid_bronchoscopy",
            "thoracentesis", "chest_tube", "ipc", "medical_thoracoscopy", 
            "pleurodesis", "pleural_biopsy", "fibrinolytic_therapy"
        ]
        
        for f in flag_order:
            data.append(flags.get(f, 0))
            
        for col_idx, val in enumerate(data, 1):
            sheet.cell(row=row, column=col_idx, value=val)

    def process_spans(self):
        # We need to write to both Span_Annotations (unhydrated) and Span_Hydrated (hydrated)
        sheet_anno = self.wb["Span_Annotations"]
        sheet_hydra = self.wb["Span_Hydrated"]
        
        spans = get_spans()
        
        start_row_anno = sheet_anno.max_row + 1
        start_row_hydra = sheet_hydra.max_row + 1
        
        for i, span in enumerate(spans):
            # 1. Write Annotation Row (No offsets)
            # Schema: source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, 
            # start_char, end_char, span_len, label, normalized_value, schema_field, event_id, 
            # is_negated, is_historical, time_anchor, reviewer, comments, hydration_status
            
            span_id = f"{NOTE_ID}_s{i+1:03d}"
            
            # Common Data
            col_map = {
                1: SOURCE_FILE, 2: NOTE_ID, 3: span_id, 4: "Procedure",
                5: span.get("context_prefix", ""), 6: span["span_text"], 7: span.get("match_index", ""),
                8: "", 9: "", 10: f'=LEN(F{start_row_anno+i})', # Formula
                11: span["label"], 12: span.get("normalized_value", ""), 13: "", 14: span.get("event_id", ""),
                15: span.get("is_negated", False), 16: False, 17: "", 18: "Auto", 19: "",
                20: "needs_hydration"
            }
            
            for c, val in col_map.items():
                sheet_anno.cell(row=start_row_anno+i, column=c, value=val)
                
            # 2. Calculate Hydration
            start_char, end_char, status = find_offsets(
                self.note_text, 
                span["span_text"], 
                span.get("context_prefix"), 
                span.get("match_index", 0)
            )
            
            # 3. Write Hydrated Row
            # Same columns, but fill 8, 9, 20
            col_map_h = col_map.copy()
            col_map_h[8] = start_char if start_char is not None else ""
            col_map_h[9] = end_char if end_char is not None else ""
            col_map_h[10] = len(span["span_text"])
            col_map_h[20] = status
            
            for c, val in col_map_h.items():
                sheet_hydra.cell(row=start_row_hydra+i, column=c, value=val)

    def write_event_log(self):
        sheet = self.wb["Event_Log"]
        events = get_event_log()
        row = sheet.max_row + 1
        
        for evt in events:
            # Schema: source_file, note_id, event_id, event_type, method, anatomy_target, device, 
            # needle_gauge, stations, counts, measurements, specimens, findings, is_historical, 
            # reviewer, comments, device_size, device_material, 
            # outcome_airway_lumen_pre, outcome_airway_lumen_post, outcome_symptoms, outcome_pleural, outcome_complication
            
            mapping = {
                1: SOURCE_FILE, 2: NOTE_ID, 3: evt.get("event_id"), 4: evt.get("event_type"),
                5: evt.get("method"), 6: evt.get("anatomy_target"), 7: evt.get("device"),
                13: evt.get("findings"), 16: evt.get("comments"),
                19: evt.get("outcome_airway_lumen_pre"), 20: evt.get("outcome_airway_lumen_post"),
                23: evt.get("outcome_complication")
            }
            
            for c, val in mapping.items():
                if val:
                    sheet.cell(row=row, column=c, value=val)
            row += 1

    def write_v3_procedure_events(self):
        sheet = self.wb["V3_Procedure_Events"]
        events = get_event_log()
        row = sheet.max_row + 1
        
        for evt in events:
            # Mapping simplified for this script
            sheet.cell(row=row, column=1, value=NOTE_ID)
            sheet.cell(row=row, column=2, value=evt.get("event_id"))
            sheet.cell(row=row, column=3, value=evt.get("event_type"))
            sheet.cell(row=row, column=4, value=evt.get("anatomy_target")) # Target Anatomy
            sheet.cell(row=row, column=10, value=evt.get("method"))
            
            # Outcomes
            sheet.cell(row=row, column=18, value=evt.get("outcome_airway_lumen_pre"))
            sheet.cell(row=row, column=19, value=evt.get("outcome_airway_lumen_post"))
            sheet.cell(row=row, column=22, value=evt.get("outcome_complication"))
            
            row += 1

    def write_json(self):
        sheet = self.wb["V3_Registry_JSON"]
        
        data = {
            "schema_version": "3.0",
            "note_id": NOTE_ID,
            "procedures": get_event_log(),
            "no_immediate_complications": False
        }
        
        json_str = json.dumps(data, indent=2)
        sheet.cell(row=sheet.max_row+1, column=1, value=json_str)

    def save(self):
        self.wb.save(OUTPUT_PATH)
        print(f"Generated {OUTPUT_PATH}")

# =============================================================================
# MAIN EXECUTION
# =============================================================================

if __name__ == "__main__":
    generator = WorkbookGenerator()
    generator.write_note_text()
    generator.write_note_index()
    generator.process_spans()
    generator.write_event_log()
    generator.write_v3_procedure_events()
    generator.write_json()
    generator.save()