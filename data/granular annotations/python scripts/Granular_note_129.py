import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# ==============================================================================
# 1. CONSTANTS & INPUT DATA
# ==============================================================================

NOTE_ID = "note_129"
SOURCE_FILE = "note_129.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_129 SOURCE_FILE: note_129.txt Procedure Performed:
Right-sided medical thoracoscopy (pleuroscopy) with pleural biopsies, talc pleurodesis, and chest tube insertion

Indication:
Exudative pleural effusion

Medications:
Moderate sedation per anesthesia record

Topical Anesthesia:
Lidocaine 1%, 20 mL

The procedure, including risks, benefits, and alternatives, was explained to the patient.
All questions were answered, and informed consent was obtained and documented per institutional protocol.
A history and physical examination were performed and updated in the pre-procedure assessment record.
Pertinent laboratory studies and imaging were reviewed. A procedural time-out was performed.
The patient was positioned in the left lateral decubitus position on the procedural bed, with pressure points appropriately padded.
The right-sided pleural entry site was identified using ultrasound guidance.
The patient was sterilely prepped with chlorhexidine gluconate (ChloraPrep) and draped in the usual fashion.
The entry site was infiltrated with 20 mL of 1% lidocaine.
A 1-cm skin incision was made, and blunt dissection was carried down to the pleural space.
Finger sweep revealed pleural adhesions, which were lysed bluntly.

An 8-mm disposable primary port was placed in the right fourth intercostal space at the mid-axillary line.
After air entrainment and lung deflation, an Olympus semi-rigid pleuroscope was introduced through the trocar into the pleural cavity.
Suction was applied, and approximately 1,600 mL of serosanguinous pleural fluid was removed.
The pleura appeared abnormal, with areas of isolated nodularity and plaques most consistent with pachypleuritis. No significant adhesions were present.
Multiple biopsies of abnormal posterior parietal pleura were obtained using forceps.
Additionally, a large pleural specimen was obtained using an electrocautery knife.
Approximately nine total biopsies were obtained and sent for histopathological examination. Minimal bleeding was observed.
Following biopsies, talc pleurodesis was performed using a combination of talc poudrage and talc slurry, with a total of 5 grams of talc instilled.
The primary port was then removed, and a 24-French chest tube was placed through the existing stoma, sutured in place, and connected to a pleuro-vac.
The chest tube was dressed and covered with a transparent dressing.

The patient tolerated the procedure well.
A post-procedure chest radiograph demonstrated the chest tube in appropriate position, with the proximal side hole just at the pleural wall and a trace apical pneumothorax.
Complications:
None immediate

Estimated Blood Loss:
10 mL

Post-Procedure Diagnosis:
Pleural effusion

Recommendations:

Admit to inpatient service (planned)

Chest tube to suction;
plan for removal once output is <150 mL/day

Pain management following talc pleurodesis, including narcotics as needed

Await pathology results

Inpatient Interventional Pulmonology consult team to follow"""

# ==============================================================================
# 2. CONFIGURATION (PROCEDURE FLAGS)
# ==============================================================================

PROCEDURE_FLAGS = {
    # Bronchoscopy Flags
    "diagnostic_bronchoscopy": 0,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    
    # Pleural Flags
    "thoracentesis": 0, # Performed as part of medical thoracoscopy
    "chest_tube": 1,
    "ipc": 0,
    "medical_thoracoscopy": 1,
    "pleurodesis": 1,
    "pleural_biopsy": 1,
    "fibrinolytic_therapy": 0
}

# ==============================================================================
# 3. DATA DEFINITION (SPANS)
# ==============================================================================

# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    # Event 1: Medical Thoracoscopy (Access & Inspection)
    ("Right-sided", "LATERALITY", "Right", "Procedure Performed:\n", "evt_01"),
    ("medical thoracoscopy", "PROC_METHOD", "Medical Thoracoscopy", "Right-sided ", "evt_01"),
    ("pleuroscopy", "PROC_METHOD", "Pleuroscopy", "medical thoracoscopy (", "evt_01"),
    ("pleural biopsies", "PROC_METHOD", "Pleural Biopsy", "with ", "evt_01"),
    ("talc pleurodesis", "PROC_METHOD", "Pleurodesis", "pleural biopsies, ", "evt_01"),
    ("chest tube insertion", "PROC_METHOD", "Chest Tube Insertion", "pleurodesis, and ", "evt_01"),
    ("8-mm disposable primary port", "DEV_INSTRUMENT", "Port (8mm)", "An ", "evt_01"),
    ("right fourth intercostal space", "ANAT_PLEURA", "Right 4th ICS", "placed in the ", "evt_01"),
    ("Olympus semi-rigid pleuroscope", "DEV_INSTRUMENT", "Pleuroscope (Semi-rigid)", "deflation, an ", "evt_01"),
    
    # Event 2: Drainage (Therapeutic Aspiration aspect of Thoracoscopy)
    ("1,600 mL", "MEAS_VOL", "1600", "approximately ", "evt_02"),
    ("serosanguinous pleural fluid", "ANAT_PLEURA", "Pleural Fluid", "mL of ", "evt_02"),
    
    # Event 3: Pleural Biopsies
    ("biopsies of abnormal posterior parietal pleura", "PROC_METHOD", "Pleural Biopsy", "Multiple ", "evt_03"),
    ("posterior parietal pleura", "ANAT_PLEURA", "Parietal Pleura", "abnormal ", "evt_03"),
    ("forceps", "DEV_INSTRUMENT", "Forceps", "using ", "evt_03"),
    ("electrocautery knife", "DEV_INSTRUMENT", "Electrocautery Knife", "using an ", "evt_03"),
    ("nine total biopsies", "MEAS_COUNT", "9", "Approximately ", "evt_03"),
    
    # Event 4: Pleurodesis
    ("talc pleurodesis", "PROC_METHOD", "Pleurodesis", "Following biopsies, ", "evt_04"),
    ("talc poudrage", "PROC_METHOD", "Talc Poudrage", "combination of ", "evt_04"),
    ("talc slurry", "PROC_METHOD", "Talc Slurry", "poudrage and ", "evt_04"),
    ("5 grams", "MEAS_SIZE", "5g", "total of ", "evt_04"),
    
    # Event 5: Chest Tube
    ("24-French", "DEV_CATHETER_SIZE", "24Fr", "and a ", "evt_05"),
    ("chest tube", "DEV_CATHETER", "Chest Tube", "24-French ", "evt_05"),
    ("pleuro-vac", "DEV_INSTRUMENT", "Pleuro-vac", "connected to a ", "evt_05"),
    
    # Outcomes
    ("tolerated the procedure well", "OUTCOME_SYMPTOMS", "Tolerated Well", "The patient ", "evt_06"),
    ("None immediate", "OUTCOME_COMPLICATION", "None", "Complications:\n", "evt_06"),
    ("10 mL", "MEAS_VOL", "10", "Estimated Blood Loss:\n", "evt_06"),
]

# ==============================================================================
# 4. EVENT DEFINITIONS
# ==============================================================================

EVENTS = [
    {
        "event_id": "evt_01",
        "name": "Medical Thoracoscopy Initiation",
        "method": ["Medical Thoracoscopy", "Pleuroscopy"],
        "anatomy": ["Right Pleural Space", "4th Intercostal Space"],
        "devices": ["8-mm Port", "Olympus Semi-rigid Pleuroscope"],
        "outcomes": []
    },
    {
        "event_id": "evt_02",
        "name": "Pleural Fluid Drainage",
        "method": ["Drainage"],
        "anatomy": ["Pleural Fluid"],
        "devices": ["Suction"],
        "outcomes": ["1600 mL removed"]
    },
    {
        "event_id": "evt_03",
        "name": "Pleural Biopsies",
        "method": ["Pleural Biopsy"],
        "anatomy": ["Posterior Parietal Pleura"],
        "devices": ["Forceps", "Electrocautery Knife"],
        "outcomes": ["9 biopsies taken"]
    },
    {
        "event_id": "evt_04",
        "name": "Pleurodesis",
        "method": ["Pleurodesis", "Talc Poudrage", "Talc Slurry"],
        "anatomy": ["Pleural Space"],
        "devices": [],
        "outcomes": ["5 grams talc instilled"]
    },
    {
        "event_id": "evt_05",
        "name": "Chest Tube Placement",
        "method": ["Chest Tube Insertion"],
        "anatomy": ["Pleural Space"],
        "devices": ["24Fr Chest Tube", "Pleuro-vac"],
        "outcomes": []
    }
]

# ==============================================================================
# 5. HELPER FUNCTIONS
# ==============================================================================

def clean_text(text):
    if not text:
        return ""
    return text.strip().replace('\r', '')

def hydrate_span(text, span_text, context_prefix=""):
    """
    Finds the start and end offsets of a span in the text.
    Uses context_prefix to disambiguate identical strings.
    """
    cleaned_text = text.replace('\r', '')  # Normalize newlines for indexing
    cleaned_span = span_text.strip()
    cleaned_context = context_prefix.strip()

    if not cleaned_span:
        return None, None

    # Attempt to find with context first
    if cleaned_context:
        pattern = re.escape(cleaned_context) + r"\s*" + re.escape(cleaned_span)
        match = re.search(pattern, cleaned_text, re.DOTALL)
        if match:
            # The span starts after the context
            # We need to find exactly where the span text starts within the match
            full_match = match.group(0)
            start_in_match = full_match.find(cleaned_span)
            start_index = match.start() + start_in_match
            end_index = start_index + len(cleaned_span)
            return start_index, end_index

    # Fallback to simple find (first occurrence)
    start_index = cleaned_text.find(cleaned_span)
    if start_index != -1:
        return start_index, start_index + len(cleaned_span)
    
    return None, None

# ==============================================================================
# 6. WORKBOOK GENERATION
# ==============================================================================

def generate_workbook():
    # Create workbook
    wb = openpyxl.Workbook()
    
    # --- Sheet 1: Note_Text ---
    ws_text = wb.active
    ws_text.title = "Note_Text"
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # --- Sheet 2: Note_Index ---
    ws_index = wb.create_sheet("Note_Index")
    headers_index = ["NOTE_ID", "SOURCE_FILE", "PROCEDURE_DATE"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers_index)
    
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE] + [PROCEDURE_FLAGS[k] for k in PROCEDURE_FLAGS]
    ws_index.append(row_data)
    
    # --- Sheet 3: Span_Annotations (Manual/Empty Offsets) ---
    ws_annot = wb.create_sheet("Span_Annotations")
    headers_annot = ["NOTE_ID", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "START_CHAR", "END_CHAR", "EVENT_ID"]
    ws_annot.append(headers_annot)
    
    for span in SPANS:
        # Tuple: (text, label, norm, context, event_id)
        ws_annot.append([NOTE_ID, span[0], span[1], span[2], span[3], "", "", span[4]])
        
    # --- Sheet 4: Span_Hydrated (Calculated Offsets) ---
    ws_hydrated = wb.create_sheet("Span_Hydrated")
    ws_hydrated.append(headers_annot)
    
    for span in SPANS:
        s_text, s_label, s_norm, s_context, s_ev_id = span
        start, end = hydrate_span(NOTE_TEXT, s_text, s_context)
        ws_hydrated.append([NOTE_ID, s_text, s_label, s_norm, s_context, start, end, s_ev_id])

    # --- Sheet 5: Event_Log (Flattened) ---
    ws_elog = wb.create_sheet("Event_Log")
    headers_elog = ["NOTE_ID", "EVENT_ID", "EVENT_NAME", "METHOD", "ANATOMY", "DEVICES", "OUTCOMES"]
    ws_elog.append(headers_elog)
    
    for evt in EVENTS:
        ws_elog.append([
            NOTE_ID,
            evt["event_id"],
            evt["name"],
            ", ".join(evt["method"]),
            ", ".join(evt["anatomy"]),
            ", ".join(evt["devices"]),
            ", ".join(evt["outcomes"])
        ])

    # --- Sheet 6: V3_Procedure_Events ---
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    headers_v3 = ["NOTE_ID", "EVENT_ID", "EVENT_TYPE", "EVENT_METADATA_JSON"]
    ws_v3.append(headers_v3)
    
    for evt in EVENTS:
        metadata = {
            "method": evt["method"],
            "anatomy": evt["anatomy"],
            "devices": evt["devices"],
            "outcomes": evt["outcomes"]
        }
        ws_v3.append([NOTE_ID, evt["event_id"], "PROCEDURE_EVENT", json.dumps(metadata)])

    # --- Sheet 7: V3_Registry_JSON ---
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["NOTE_ID", "FULL_JSON"])
    
    full_registry_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source_file": SOURCE_FILE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS,
        "spans": [
            {
                "text": s[0],
                "label": s[1],
                "normalized": s[2],
                "event_id": s[4]
            } for s in SPANS
        ]
    }
    
    ws_json.append([NOTE_ID, json.dumps(full_registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

# ==============================================================================
# 7. EXECUTION
# ==============================================================================

if __name__ == "__main__":
    generate_workbook()