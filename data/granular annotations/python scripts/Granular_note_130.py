import openpyxl
from openpyxl import Workbook
import re
import json
import os
import datetime

# 1. Imports & Constants
NOTE_ID = "note_130"
SOURCE_FILE = "note_130.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_130 SOURCE_FILE: note_130.txt Procedure Name: Bronchoscopy

Indications:

Known lung cancer of the right lower lobe

Malignant airway disease

Evaluation for mediastinal staging

Anesthesia:
General anesthesia with controlled mechanical ventilation.
Topical anesthesia with 2% lidocaine to the tracheobronchial tree (9 mL).
See the anesthesia record for full documentation of administered medications.
Pre-Anesthesia Assessment:

ASA Physical Status Classification: III (patient with severe systemic disease)

The procedure, including risks, benefits, and alternatives, was explained to the patient.
All questions were answered, and informed consent was obtained and documented per institutional protocol.
A history and physical examination were performed and updated in the pre-procedure assessment record.
Relevant laboratory studies and radiographic imaging were reviewed. A procedural time-out was performed prior to the intervention.
Following administration of intravenous anesthetic medications and topical anesthesia to the upper airway and tracheobronchial tree, the Q180 slim video bronchoscope was introduced through the mouth via a laryngeal mask airway and advanced into the tracheobronchial tree.
The UC180F convex probe EBUS bronchoscope was subsequently introduced through the mouth via the laryngeal mask airway and advanced into the tracheobronchial tree.
A black rigid bronchial tube (12.0–11.0 mm) was introduced through the mouth and advanced into the tracheobronchial tree.
A 0-degree, 4.0-mm rigid telescope was introduced through the rigid bronchoscope and advanced into the airway.
The T180 therapeutic video bronchoscope was then introduced through the rigid bronchoscope following removal of the telescope and advanced into the tracheobronchial tree.
Procedure Description and Findings

Larynx:
Normal appearance.

Trachea and Carina:
No evidence of significant pathology.

Left Lung:
No evidence of significant pathology.
Right Lung:
A nearly obstructing (>90% obstruction) endobronchial mass was identified proximally in the apical segment of the right upper lobe (B1).
The lesion was not traversed.

A second nearly obstructing (>90% obstruction) large endobronchial mass was identified proximally within the bronchus intermedius.
This lesion was successfully traversed, allowing visualization of the distal airway.
The bronchoscope was withdrawn and replaced with the EBUS bronchoscope to perform endobronchial ultrasound examination.
Lymph Node Assessment

Lymph node sizing and sampling were performed using endobronchial ultrasound for staging of non-small cell lung cancer.
Transbronchial needle aspiration was performed using an Olympus EBUS-TBNA needle at the following stations, with specimens sent for histopathologic examination:

Station 11L (interlobar): ROSE demonstrated adequate tissue.
Five needle passes obtained.

Station 4L (lower paratracheal): Not sampled.

Station 2L (upper paratracheal): ROSE demonstrated non-diagnostic tissue.
Five needle passes obtained.

Station 2R (upper paratracheal): Not sampled.

Station 4R (lower paratracheal): ROSE demonstrated adequate tissue.
Five needle passes obtained.

Station 7 (subcarinal): ROSE demonstrated adequate tissue. Five needle passes obtained.
Stations 11Rs and 11Ri were not sampled due to direct invasion by the primary tumor.
Therapeutic Intervention

The endobronchial tumor within the right bronchus intermedius was mechanically excised using rigid bronchoscopy and rigid suction.
Following tumor debulking, the right middle lobe bronchus was patent.
The right lower lobe bronchi were noted to be occluded at the subsegmental level due to malignant disease.
A portion of the endobronchial tumor arising from the apical segment of the right upper lobe and occluding the right upper lobe bronchus was also removed using biopsy forceps.
The anterior and posterior segmental bronchi of the right upper lobe were patent;
however, the apical segmental bronchus remained completely occluded.

Complications

No immediate complications.

Estimated Blood Loss

Less than 5 mL.
Impression

Known lung cancer of the right lower lobe

Malignant endobronchial mass in the apical segment of the right upper lobe (B1)

Malignant endobronchial mass in the bronchus intermedius

Successful lymph node sizing and sampling via EBUS-TBNA

Mechanical excision of endobronchial lesions in the bronchus intermedius and right upper lobe

Post-Procedure Diagnosis

Technically successful flexible bronchoscopy with endobronchial ultrasound-guided biopsies

The patient remained stable throughout the procedure and was transferred in good condition to the post-bronchoscopy recovery area for observation until discharge criteria were met.
Preliminary findings were discussed with the patient, and follow-up with the requesting service for final pathology results was recommended."""

# 2. Configuration (Procedure Flags)
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 1,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 1,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 1,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 3. Data Definition (Spans)
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    # Scope/Instrument Intro
    ("Q180 slim video bronchoscope", "DEV_INSTRUMENT", "Olympus Q180", "tracheobronchial tree, the", "evt_intro"),
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "Olympus UC180F", "tree. The", "evt_intro"),
    ("black rigid bronchial tube", "DEV_INSTRUMENT", "Rigid Bronchoscope", "A", "evt_intro"),
    ("12.0–11.0 mm", "DEV_CATHETER_SIZE", "12.0-11.0 mm", "black rigid bronchial tube (", "evt_intro"),
    ("0-degree, 4.0-mm rigid telescope", "DEV_INSTRUMENT", "Rigid Telescope", "A", "evt_intro"),
    ("T180 therapeutic video bronchoscope", "DEV_INSTRUMENT", "Olympus T180", "The", "evt_intro"),

    # Findings (Pre-intervention)
    ("nearly obstructing (>90% obstruction)", "OUTCOME_AIRWAY_LUMEN_PRE", ">90% Obstruction", "Right Lung: A", "evt_find_01"),
    ("apical segment of the right upper lobe (B1)", "ANAT_AIRWAY", "RUL Apical Segment (RB1)", "proximally in the", "evt_find_01"),
    ("nearly obstructing (>90% obstruction)", "OUTCOME_AIRWAY_LUMEN_PRE", ">90% Obstruction", "A second", "evt_find_02"),
    ("bronchus intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "proximally within the", "evt_find_02"),

    # EBUS-TBNA (Linear EBUS)
    ("endobronchial ultrasound", "PROC_METHOD", "Linear EBUS", "replaced with the EBUS bronchoscope to perform", "evt_ebus"),
    ("Transbronchial needle aspiration", "PROC_METHOD", "TBNA", "examination.", "evt_ebus"),
    ("Olympus EBUS-TBNA needle", "DEV_NEEDLE", "EBUS-TBNA Needle", "using an", "evt_ebus"),

    # Station 11L
    ("Station 11L", "ANAT_LN_STATION", "Station 11L", "examination:", "evt_ebus_11l"),
    ("ROSE demonstrated adequate tissue", "OBS_ROSE", "Adequate", "(interlobar):", "evt_ebus_11l"),
    ("Five needle passes", "MEAS_COUNT", "5", "obtained.", "evt_ebus_11l"),

    # Station 2L
    ("Station 2L", "ANAT_LN_STATION", "Station 2L", "Not sampled.", "evt_ebus_2l"),
    ("ROSE demonstrated non-diagnostic tissue", "OBS_ROSE", "Non-diagnostic", "(upper paratracheal):", "evt_ebus_2l"),
    ("Five needle passes", "MEAS_COUNT", "5", "tissue.", "evt_ebus_2l"),

    # Station 4R
    ("Station 4R", "ANAT_LN_STATION", "Station 4R", "Not sampled.", "evt_ebus_4r"),
    ("ROSE demonstrated adequate tissue", "OBS_ROSE", "Adequate", "(lower paratracheal):", "evt_ebus_4r"),
    ("Five needle passes", "MEAS_COUNT", "5", "tissue.", "evt_ebus_4r"),

    # Station 7
    ("Station 7", "ANAT_LN_STATION", "Station 7", "obtained.", "evt_ebus_7"),
    ("ROSE demonstrated adequate tissue", "OBS_ROSE", "Adequate", "(subcarinal):", "evt_ebus_7"),
    ("Five needle passes", "MEAS_COUNT", "5", "tissue.", "evt_ebus_7"),

    # Therapeutic: Rigid Debulking BI
    ("rigid bronchoscopy", "PROC_METHOD", "Rigid Bronchoscopy", "mechanically excised using", "evt_rigid_bi"),
    ("rigid suction", "DEV_INSTRUMENT", "Suction Catheter", "rigid bronchoscopy and", "evt_rigid_bi"),
    ("mechanically excised", "PROC_ACTION", "Debulking", "intermedius was", "evt_rigid_bi"),
    ("right bronchus intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "tumor within the", "evt_rigid_bi"),
    ("patent", "OUTCOME_AIRWAY_LUMEN_POST", "Patent", "middle lobe bronchus was", "evt_rigid_bi"), # Inferred context of successful debulking

    # Therapeutic: Forceps Debulking RUL
    ("apical segment of the right upper lobe", "ANAT_AIRWAY", "RUL Apical Segment (RB1)", "arising from the", "evt_forceps_rul"),
    ("biopsy forceps", "DEV_INSTRUMENT", "Forceps", "removed using", "evt_forceps_rul"),
    ("removed", "PROC_ACTION", "Debulking", "bronchus was also", "evt_forceps_rul"),
    ("apical segmental bronchus remained completely occluded", "OUTCOME_AIRWAY_LUMEN_POST", "Occluded", "however, the", "evt_forceps_rul"),

    # General Outcomes
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications", "evt_outcome"),
    ("patient remained stable", "OUTCOME_SYMPTOMS", "Stable", "The", "evt_outcome")
]

# 4. Event Definitions
EVENTS = [
    {
        "event_id": "evt_ebus",
        "procedure_type": "Diagnostic",
        "method": "Linear EBUS-TBNA",
        "anatomy": "Mediastinum",
        "devices": "UC180F, Olympus EBUS-TBNA Needle",
        "outcomes": "Successful sampling",
        "notes": "Stations 11L, 2L, 4R, 7 sampled with ROSE"
    },
    {
        "event_id": "evt_rigid_bi",
        "procedure_type": "Therapeutic",
        "method": "Rigid Bronchoscopy, Mechanical Debulking",
        "anatomy": "Bronchus Intermedius",
        "devices": "Black Rigid Tube, Rigid Suction",
        "outcomes": "Lesion excised, distal airway visualized, RML patent"
    },
    {
        "event_id": "evt_forceps_rul",
        "procedure_type": "Therapeutic",
        "method": "Forceps Debulking",
        "anatomy": "RUL Apical Segment (RB1)",
        "devices": "Biopsy Forceps",
        "outcomes": "Partial removal, bronchus remained occluded"
    }
]

# 5. Helper Functions
def clean_text(text):
    if not text: return ""
    return text.strip().replace('\r', '')

def hydrate_span(text, span_text, context_prefix):
    clean_full = clean_text(text)
    clean_span = clean_text(span_text)
    clean_ctx = clean_text(context_prefix)
    
    if not clean_span: return 0, 0
    
    # Try finding with context first
    search_pattern = re.escape(clean_ctx) + r"\s*" + re.escape(clean_span)
    match = re.search(search_pattern, clean_full, re.IGNORECASE)
    
    if match:
        # Return start/end of the SPAN specifically (not context)
        # match.start() is start of context. 
        # We need to find where span starts within match
        full_match_str = match.group(0)
        span_start_in_match = full_match_str.lower().rfind(clean_span.lower())
        real_start = match.start() + span_start_in_match
        return real_start, real_start + len(clean_span)
    
    # Fallback: simple find
    start = clean_full.lower().find(clean_span.lower())
    if start != -1:
        return start, start + len(clean_span)
        
    return 0, 0

# 6. Workbook Generation
def generate_workbook():
    wb = Workbook()
    
    # Sheet 1: Note_Text
    ws_text = wb.active
    ws_text.title = "Note_Text"
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # Sheet 2: Note_Index
    ws_index = wb.create_sheet("Note_Index")
    headers = ["NOTE_ID", "PROCEDURE_DATE"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers)
    row_data = [NOTE_ID, PROCEDURE_DATE] + list(PROCEDURE_FLAGS.values())
    ws_index.append(row_data)
    
    # Sheet 3: Span_Annotations
    ws_span_ann = wb.create_sheet("Span_Annotations")
    ws_span_ann.append(["NOTE_ID", "START_CHAR", "END_CHAR", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "EVENT_ID"])
    for span in SPANS:
        ws_span_ann.append([NOTE_ID, "", "", span[0], span[1], span[2], span[3], span[4]])
        
    # Sheet 4: Span_Hydrated
    ws_span_hyd = wb.create_sheet("Span_Hydrated")
    ws_span_hyd.append(["NOTE_ID", "START_CHAR", "END_CHAR", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "EVENT_ID"])
    for span in SPANS:
        s, e = hydrate_span(NOTE_TEXT, span[0], span[3])
        ws_span_hyd.append([NOTE_ID, s, e, span[0], span[1], span[2], span[3], span[4]])
        
    # Sheet 5: Event_Log
    ws_events = wb.create_sheet("Event_Log")
    ws_events.append(["NOTE_ID", "EVENT_ID", "PROCEDURE_TYPE", "METHOD", "ANATOMY", "DEVICES", "OUTCOMES", "NOTES"])
    for evt in EVENTS:
        ws_events.append([
            NOTE_ID, 
            evt["event_id"], 
            evt["procedure_type"], 
            evt["method"], 
            evt["anatomy"], 
            evt["devices"], 
            evt["outcomes"], 
            evt.get("notes", "")
        ])
        
    # Sheet 6: V3_Procedure_Events (Detailed breakdown structure)
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    ws_v3.append(["NOTE_ID", "EVENT_ID", "EVENT_JSON"])
    for evt in EVENTS:
        ws_v3.append([NOTE_ID, evt["event_id"], json.dumps(evt)])

    # Sheet 7: V3_Registry_JSON
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["NOTE_ID", "FULL_JSON"])
    full_data = {
        "metadata": {"note_id": NOTE_ID, "source": SOURCE_FILE},
        "flags": PROCEDURE_FLAGS,
        "spans": [{"text": s[0], "label": s[1], "value": s[2], "event_id": s[4]} for s in SPANS],
        "events": EVENTS
    }
    ws_json.append([NOTE_ID, json.dumps(full_data, indent=2)])
    
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()