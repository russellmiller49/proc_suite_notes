import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# -------------------------------------------------------------------------
# 1. CONSTANTS & INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_131"
SOURCE_FILE = "note_131.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_131 SOURCE_FILE: note_131.txt Procedure Name: Pleuroscopy
Indication: Pleural effusion
Anesthesia: Monitored Anesthesia Care

Pre-Anesthesia Assessment

ASA Physical Status: III – Patient with severe systemic disease

The procedure, including risks, benefits, and alternatives, was explained to the patient.
All questions were answered, and informed consent was obtained and documented per institutional protocol.
A focused history and physical examination were performed and updated in the pre-procedure assessment record.
Relevant laboratory studies and imaging were reviewed. A procedural time-out was performed prior to initiation.
Procedure Description

The patient was placed on the operating table in the lateral decubitus position with appropriate padding of pressure points.
The procedural site was identified using ultrasound guidance and was sterilely prepped with chlorhexidine gluconate (Chloraprep) and draped in the usual fashion.
Local Anesthesia

The pleural entry site was infiltrated with 19 mL of 1% lidocaine.
A 10-mm reusable primary port was placed on the right side at the 6th intercostal space along the anterior axillary line using a Veress needle technique.
A 0-degree 2.0-mm pleuroscopy telescope was introduced through the incision and advanced into the pleural space, followed by a 50-degree 7.0-mm pleuroscopy optic.
Pleuroscopy Findings

Serous pleural effusion was present; approximately 3,000 mL was suctioned.
Small white and reddish raised lesions were identified on the diaphragmatic parietal pleura.

The visceral pleura appeared normal.
The chest wall (parietal) pleura was hyperemic with areas of brownish discoloration.
All three lobes collapsed easily and re-expanded fully with suction.
Biopsy

Targeted biopsies of nodular lesions on the diaphragmatic pleura were obtained using forceps (three samples) and sent for histopathologic evaluation.
Additional biopsies were obtained from the brownish lesions on the chest wall pleura.
A 15.5-Fr PleurX catheter was placed in the pleural space over the diaphragm.
Dressing

Port sites were dressed with a transparent dressing.

Estimated Blood Loss

Minimal

Complications

None immediate

Impression

Exudative pleural effusion with predominantly normal pleura, except for small diaphragmatic and chest wall lesions.
Post-Procedure Plan

The patient will be observed post-procedure until discharge criteria are met.

Chest X-ray to be obtained post-procedure."""

# -------------------------------------------------------------------------
# 2. CONFIGURATION (PROCEDURE FLAGS)
# -------------------------------------------------------------------------
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 0,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 1,  # PleurX catheter
    "medical_thoracoscopy": 1,  # "Pleuroscopy"
    "pleurodesis": 0,
    "pleural_biopsy": 1,  # Biopsies taken
    "fibrinolytic_therapy": 0,
}

# -------------------------------------------------------------------------
# 3. DATA DEFINITION (SPANS)
# -------------------------------------------------------------------------
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    # Event 01: Medical Thoracoscopy (Access & Inspection)
    ("Pleuroscopy", "PROC_METHOD", "Medical Thoracoscopy", "Procedure Name: ", "evt_01"),
    ("right side", "LATERALITY", "Right", "port was placed on the ", "evt_01"),
    ("6th intercostal space", "ANAT_PLEURA", "6th intercostal space", "right side at the ", "evt_01"),
    ("10-mm", "MEAS_SIZE", "10 mm", "A ", "evt_01"),
    ("pleuroscopy telescope", "DEV_INSTRUMENT", "Pleuroscope", "A 0-degree 2.0-mm ", "evt_01"),
    ("pleural space", "ANAT_PLEURA", "Pleural space", "advanced into the ", "evt_01"),
    
    # Event 01: Drainage (Therapeutic aspect of thoracoscopy)
    ("3,000 mL", "MEAS_VOL", "3000 mL", "approximately ", "evt_01"),
    ("suctioned", "PROC_ACTION", "Suction", "3,000 mL was ", "evt_01"),

    # Event 02: Biopsy (Diaphragmatic)
    ("biopsies", "PROC_ACTION", "Biopsy", "Targeted ", "evt_02"),
    ("nodular lesions", "OBS_LESION", "Nodule", "biopsies of ", "evt_02"),
    ("diaphragmatic pleura", "ANAT_PLEURA", "Diaphragmatic pleura", "lesions on the ", "evt_02"),
    ("forceps", "DEV_INSTRUMENT", "Biopsy Forceps", "obtained using ", "evt_02"),
    ("three samples", "MEAS_COUNT", "3", "forceps (", "evt_02"),

    # Event 03: Biopsy (Chest Wall)
    ("biopsies", "PROC_ACTION", "Biopsy", "Additional ", "evt_03"),
    ("brownish lesions", "OBS_LESION", "Lesion", "obtained from the ", "evt_03"),
    ("chest wall pleura", "ANAT_PLEURA", "Parietal pleura", "lesions on the ", "evt_03"),

    # Event 04: IPC Placement
    ("15.5-Fr", "DEV_CATHETER_SIZE", "15.5 Fr", "A ", "evt_04"),
    ("PleurX catheter", "DEV_CATHETER", "Tunnelled Indwelling Pleural Catheter", "15.5-Fr ", "evt_04"),
    ("placed", "PROC_ACTION", "Placement", "catheter was ", "evt_04"),
    ("pleural space", "ANAT_PLEURA", "Pleural space", "catheter was placed in the ", "evt_04"),
    
    # Outcomes / Complications
    ("Minimal", "OUTCOME_COMPLICATION", "None", "Estimated Blood Loss\n\n", "evt_05"),
    ("None immediate", "OUTCOME_COMPLICATION", "None", "Complications\n\n", "evt_05"),
]

# -------------------------------------------------------------------------
# 4. EVENT DEFINITIONS
# -------------------------------------------------------------------------
EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Medical Thoracoscopy",
        "action": "Inspection/Suction",
        "anatomy": ["Pleural space", "Right hemithorax"],
        "devices": ["10-mm port", "Pleuroscope"],
        "volume": "3000 mL",
        "comments": "Access obtained, fluid suctioned."
    },
    {
        "event_id": "evt_02",
        "procedure_type": "Pleural Biopsy",
        "action": "Biopsy",
        "anatomy": ["Diaphragmatic pleura"],
        "devices": ["Forceps"],
        "findings": "Nodular lesions",
        "comments": "3 samples taken."
    },
    {
        "event_id": "evt_03",
        "procedure_type": "Pleural Biopsy",
        "action": "Biopsy",
        "anatomy": ["Chest wall pleura"],
        "devices": ["Forceps"],
        "findings": "Brownish lesions"
    },
    {
        "event_id": "evt_04",
        "procedure_type": "Indwelling Pleural Catheter",
        "action": "Placement",
        "anatomy": ["Pleural space"],
        "devices": ["15.5-Fr PleurX catheter"],
        "comments": "Placed over diaphragm."
    }
]

# -------------------------------------------------------------------------
# 5. HELPER FUNCTIONS
# -------------------------------------------------------------------------
def clean_text(text):
    if not text:
        return ""
    return text.strip().replace('\r', '')

def hydrate_span(full_text, span_text, context_prefix):
    """
    Finds the start and end indices of span_text within full_text.
    Uses context_prefix to disambiguate if the span appears multiple times.
    """
    cleaned_full = clean_text(full_text)
    cleaned_span = clean_text(span_text)
    cleaned_context = clean_text(context_prefix)
    
    if not cleaned_span:
        return 0, 0

    # Pattern: Context + any chars + Span
    # Escape special regex chars in context/span
    pattern_str = re.escape(cleaned_context) + r"(.*?)" + re.escape(cleaned_span)
    match = re.search(pattern_str, cleaned_full, re.DOTALL | re.IGNORECASE)
    
    if match:
        # The span starts after context + group(1)
        # Note: group(0) is the whole match (context + filler + span)
        # We need the index where the span actually starts.
        full_match_start = match.start()
        # Length of context + filler
        context_len = len(match.group(0)) - len(cleaned_span)
        start_index = full_match_start + context_len
        end_index = start_index + len(cleaned_span)
        return start_index, end_index
    
    # Fallback: exact match of span only (if unique or no context provided)
    start_index = cleaned_full.find(cleaned_span)
    if start_index != -1:
        return start_index, start_index + len(cleaned_span)
        
    return 0, 0

# -------------------------------------------------------------------------
# 6. WORKBOOK GENERATION
# -------------------------------------------------------------------------
def generate_workbook():
    # A. Initialize Workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
        
    # B. Sheet 1: Note_Text
    ws_text = wb.create_sheet("Note_Text")
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # C. Sheet 2: Note_Index (Metadata + Flags)
    ws_index = wb.create_sheet("Note_Index")
    headers_index = ["NOTE_ID", "SOURCE_FILE", "PROCEDURE_DATE"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers_index)
    
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE]
    for key in PROCEDURE_FLAGS:
        row_data.append(PROCEDURE_FLAGS[key])
    ws_index.append(row_data)
    
    # D. Sheet 3: Span_Annotations (Raw)
    ws_anno = wb.create_sheet("Span_Annotations")
    headers_anno = ["NOTE_ID", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "EVENT_ID", "START_CHAR", "END_CHAR"]
    ws_anno.append(headers_anno)
    
    # E. Sheet 4: Span_Hydrated (Calculated Offsets)
    ws_hydrated = wb.create_sheet("Span_Hydrated")
    ws_hydrated.append(headers_anno) # Same headers
    
    # Process Spans
    for span in SPANS:
        text_val, label, norm, context, evt_id = span
        
        # Sheet 3 Row (Blank offsets)
        ws_anno.append([NOTE_ID, text_val, label, norm, context, evt_id, "", ""])
        
        # Calculate Offsets
        start, end = hydrate_span(NOTE_TEXT, text_val, context)
        
        # Sheet 4 Row (Filled offsets)
        ws_hydrated.append([NOTE_ID, text_val, label, norm, context, evt_id, start, end])

    # F. Sheet 5: Event_Log (Flattened)
    ws_event_log = wb.create_sheet("Event_Log")
    headers_log = ["NOTE_ID", "EVENT_ID", "PROCEDURE_TYPE", "ACTION", "ANATOMY", "DEVICES", "COMMENTS"]
    ws_event_log.append(headers_log)
    
    for evt in EVENTS:
        # Join lists for CSV-like cell content
        anat_str = ", ".join(evt.get("anatomy", []))
        dev_str = ", ".join(evt.get("devices", []))
        
        ws_event_log.append([
            NOTE_ID,
            evt.get("event_id", ""),
            evt.get("procedure_type", ""),
            evt.get("action", ""),
            anat_str,
            dev_str,
            evt.get("comments", "")
        ])

    # G. Sheet 6: V3_Procedure_Events (Detailed)
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    headers_v3 = ["NOTE_ID", "EVENT_ID", "PROCEDURE_TYPE", "V3_JSON_SNIPPET"]
    ws_v3.append(headers_v3)
    
    for evt in EVENTS:
        json_snippet = json.dumps(evt)
        ws_v3.append([NOTE_ID, evt.get("event_id"), evt.get("procedure_type"), json_snippet])

    # H. Sheet 7: V3_Registry_JSON
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["NOTE_ID", "FULL_JSON_BLOB"])
    
    full_registry_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source": SOURCE_FILE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS
    }
    ws_json.append([NOTE_ID, json.dumps(full_registry_data, indent=2)])

    # I. Save
    print(f"Generating workbook at: {OUTPUT_PATH}")
    wb.save(OUTPUT_PATH)
    print("Done.")

# -------------------------------------------------------------------------
# 7. EXECUTION
# -------------------------------------------------------------------------
if __name__ == "__main__":
    generate_workbook()