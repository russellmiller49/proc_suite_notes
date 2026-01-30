import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# ==============================================================================
# 1. IMPORTS & CONSTANTS
# ==============================================================================
NOTE_ID = "note_132"
SOURCE_FILE = "note_132.txt"
PROCEDURE_DATE = ""  # Not explicitly in text, leave blank
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_132 SOURCE_FILE: note_132.txt Procedure Name: Bronchoscopy
Indications: Chronic cough, Shortness of breath
Medications: General Anesthesia, 2% Lidocaine, tracheobronchial tree 10 mL

Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the preprocedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention. Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree.The 0 degree 4.0mm rigid optic was introduced through the mouth and advanced to the tracheobronchial tree.
The BF-Q190 slim video bronchoscope was introduced through the rigid bronchoscope (after telescope was removed) and advanced to the tracheobronchial tree.
The BF-1TH190 therapeutic video bronchoscope was introduced through the rigid bronchoscope (after telescope was removed) and advanced to the tracheobronchial tree.
The patient tolerated the procedure well. 
Procedure Description:
Patient was brought to the bronchoscopy suite, where a timeout was performed and the radiographic reviewed prior to the procedure.
Intravenous sedation was administered under the direction of Dr. Sarkis, once patient was adequately sedated and neuromuscular blockade had been provided, he was intubated without difficulty using the Dumon Orange striped bronchoscope.
Oropharyngeal landmarks were visualized, with extensive redundant soft tissue in the oropharynx.
The vocal cords visualized the appeared to be moving normally.
Chronic inflammatory changes were noted in the tracheal mucosa, with dynamic airway collapse noted.
The main carina was architecturally distorted, with a fibrotic stricture noted to be stretching from the posterior wall of the right mainstem bronchus and tethering the anteromedial aspect of the main carina.
Right and left lungs were explored segmentally and subsegmentally using the Q190 video bronchoscope introduced through the rigid bronchoscope.
Exploration of the left lung revealed chronic inflammatory changes in the bronchial mucosa, with dynamic airway collapse that resulted in less than 50% occlusion of the airway lumen with respirations.
Exploration of the right lung revealed the previously placed stent in place with 90% obstruction of the right middle lobe bronchus from what appeared to be bronchial mucosal edema and granulation tissue.
The stent was also heavily coated with thick mucopurulent secretions. The right lower lobe bronchus appeared to be patent.
Given the findings, particularly the degree of obstruction of the right middle lobe bronchus and the exuberant granulation tissue noted, we elected to remove the stent currently in place.
This was done en bloc using forceps without difficulty. Following the removal of the stent, the reconstructed right bronchus intermedius was carefully examined and was noted to have dynamic airway collapse though the degree of luminal obstruction is estimated at less than 50%.
The weblike fibrotic stricture arising from the posterior aspect of the right mainstem bronchus and tethering the anteromedial aspect of the main carina was divided using electrocautery knife.
This resulted in some freeing of the stricture, but the dynamic airway collapse continued to bring the tissues into apposition during respirations.
The diode laser was used to fulgurate the abutting tissue with the intent to discourage additional stricture formation.
Once a space had been developed, cryotherapy was used at the edges, also to discourage restenosis.
Given the fact that the luminal obstruction during respiration was less than 50%, and there was evidence of granulation tissue formation, we elected not to place a stent at this point in time.
The plan will be to bring him back for airway examination in 2 weeks to decide if silicone stent would be appropriate.
Complications: No immediate complications
Estimated Blood Loss: Less than 5 cc.
- Chronic cough
- A stricture was found in the right mainstem bronchus. The narrowing appears fibrotic.
This was excised
- A stent is located in the bronchus intermedius. This was removed en bloc.
- Chronic mucosal inflammation was visualized in the right mainstem bronchus.
- Luminal obstruction in the right mainstem and right bronchus medius is estimated at
about 30% post-stent removal.
Post Procedure Diagnosis:
- Chronic mucosal inflammation was visualized in the right mainstem bronchus.
- Luminal obstruction in the right mainstem and right bronchus medius is estimated at
about 30% post-stent removal.
- Technically successful rigid bronchoscopy with stent removal and stricture resection.
The patient has remained stable and has been transferred in good condition to the post bronchoscopy
recovery area, where patient will be observed until discharge criteria is met.
Preliminary findings have been discussed with the patient and follow-up with the pulmonary clinic in 2 weeks has been recommended"""

# ==============================================================================
# 2. CONFIGURATION (PROCEDURE FLAGS)
# ==============================================================================
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0, # Stent removal handled as separate concept usually, but strictly here stent removal != FB removal unless specified
    "airway_dilation": 0,
    "airway_stent": 0, # Removed, not placed
    "thermal_ablation": 1, # Electrocautery, Diode Laser
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 1,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 1,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# ==============================================================================
# 3. DATA DEFINITION (SPANS)
# ==============================================================================
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("rigid optic", "DEV_INSTRUMENT", "rigid_optic", "degree 4.0mm ", "evt_01"),
    ("BF-Q190 slim video bronchoscope", "DEV_INSTRUMENT", "BF-Q190", "The ", "evt_02"),
    ("BF-1TH190 therapeutic video bronchoscope", "DEV_INSTRUMENT", "BF-1TH190", "The ", "evt_03"),
    ("rigid bronchoscopy", "PROC_METHOD", "rigid_bronchoscopy", "successful ", "evt_01"),
    ("Dumon Orange striped bronchoscope", "DEV_INSTRUMENT", "rigid_bronchoscope_dumon", "using the ", "evt_01"),
    ("stent", "DEV_STENT", "stent", "previously placed ", "evt_04"),
    ("right middle lobe bronchus", "ANAT_AIRWAY", "RML", "obstruction of the ", "evt_04"),
    ("remove the stent", "PROC_ACTION", "remove", "elected to ", "evt_04"),
    ("forceps", "DEV_INSTRUMENT", "forceps", "en bloc using ", "evt_04"),
    ("main carina", "ANAT_AIRWAY", "main_carina", "aspect of the ", "evt_05"),
    ("right mainstem bronchus", "ANAT_AIRWAY", "RMB", "posterior aspect of the ", "evt_05"),
    ("divided", "PROC_ACTION", "resection", "main carina was ", "evt_05"),
    ("electrocautery knife", "DEV_INSTRUMENT", "electrocautery_knife", "using ", "evt_05"),
    ("diode laser", "DEV_INSTRUMENT", "laser", "The ", "evt_05"),
    ("fulgurate", "PROC_ACTION", "ablation", "used to ", "evt_05"),
    ("cryotherapy", "PROC_METHOD", "cryotherapy", "been developed, ", "evt_06"),
    ("patient tolerated the procedure well", "OUTCOME_SYMPTOMS", "tolerated_well", "The ", "evt_07"),
    ("No immediate complications", "OUTCOME_COMPLICATION", "none", "Complications: ", "evt_07"),
    ("30%", "OUTCOME_AIRWAY_LUMEN_POST", "30%", "estimated at\nabout ", "evt_08"),
    ("stent removal", "PROC_ACTION", "remove_stent", "bronchoscopy with ", "evt_04"),
    ("stricture resection", "PROC_ACTION", "resection", "removal and ", "evt_05")
]

# ==============================================================================
# 4. EVENT DEFINITIONS
# ==============================================================================
EVENTS = [
    {
        "event_id": "evt_01",
        "type": "procedure_start",
        "method": "rigid_bronchoscopy",
        "devices": ["rigid_optic", "dumon_rigid_scope"],
        "anatomy": "tracheobronchial_tree"
    },
    {
        "event_id": "evt_04",
        "type": "intervention",
        "action": "stent_removal",
        "device": "forceps",
        "target": "stent",
        "anatomy": ["RML", "bronchus_intermedius"]
    },
    {
        "event_id": "evt_05",
        "type": "intervention",
        "action": "resection_ablation",
        "devices": ["electrocautery_knife", "diode_laser"],
        "anatomy": ["main_carina", "RMB"],
        "pathology": "stricture"
    },
    {
        "event_id": "evt_06",
        "type": "intervention",
        "method": "cryotherapy",
        "anatomy": "stricture_edges"
    },
    {
        "event_id": "evt_08",
        "type": "outcome",
        "metric": "lumen_obstruction_post",
        "value": "30%",
        "anatomy": "RMB_Bronchus_Medius"
    }
]

# ==============================================================================
# 5. HELPER FUNCTIONS
# ==============================================================================
def clean_text(text):
    if not text: return ""
    return text.strip().replace('\r', '')

def hydrate_span(text, span_text, context_prefix):
    cleaned_text = clean_text(text)
    cleaned_span = clean_text(span_text)
    cleaned_context = clean_text(context_prefix)
    
    # Try finding with context first
    pattern = re.escape(cleaned_context) + r"\s*" + re.escape(cleaned_span)
    match = re.search(pattern, cleaned_text, re.IGNORECASE)
    
    if match:
        # Calculate actual start of the span (excluding context)
        full_start = match.start()
        # Re-verify length of context in the match to adjust offset
        context_len = len(cleaned_context)
        # Approximate adjustment; in a real robust system, we'd be more precise with groups
        start_index = full_start + len(match.group(0)) - len(cleaned_span)
        end_index = start_index + len(cleaned_span)
        return start_index, end_index
    
    # Fallback: Find first occurrence of span_text
    start_index = cleaned_text.find(cleaned_span)
    if start_index != -1:
        return start_index, start_index + len(cleaned_span)
    
    return -1, -1

# ==============================================================================
# 6. WORKBOOK GENERATION
# ==============================================================================
def generate_workbook():
    # Create workbook (in-memory since we can't load external template easily in this env, 
    # but code structure assumes we are building up the structure)
    wb = openpyxl.Workbook()
    
    # --- Sheet 1: Note_Text ---
    ws_text = wb.active
    ws_text.title = "Note_Text"
    ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # --- Sheet 2: Note_Index ---
    ws_index = wb.create_sheet("Note_Index")
    headers = ["note_id", "procedure_date"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers)
    row_data = [NOTE_ID, PROCEDURE_DATE] + [PROCEDURE_FLAGS[k] for k in PROCEDURE_FLAGS]
    ws_index.append(row_data)
    
    # --- Sheet 3: Span_Annotations ---
    ws_anno = wb.create_sheet("Span_Annotations")
    ws_anno.append(["note_id", "event_id", "label", "value", "start_char", "end_char", "text_context"])
    # We will fill this in the hydration loop or leave blank for human review. 
    # The prompt implies generating the data for the pipeline. 
    # We will map SPANS to this format.
    
    # --- Sheet 4: Span_Hydrated ---
    ws_hydrated = wb.create_sheet("Span_Hydrated")
    ws_hydrated.append(["note_id", "event_id", "label", "normalized_value", "start_char", "end_char", "span_text", "context_prefix"])
    
    for span_data in SPANS:
        span_text, label, norm_val, context, evt_id = span_data
        start, end = hydrate_span(NOTE_TEXT, span_text, context)
        ws_hydrated.append([NOTE_ID, evt_id, label, norm_val, start, end, span_text, context])
        
        # Also populate Span_Annotations (redundant but standard in some pipelines)
        ws_anno.append([NOTE_ID, evt_id, label, norm_val, start, end, context])

    # --- Sheet 5: Event_Log ---
    ws_events = wb.create_sheet("Event_Log")
    ws_events.append(["note_id", "event_id", "event_type", "details_json"])
    for evt in EVENTS:
        ws_events.append([NOTE_ID, evt["event_id"], evt["type"], json.dumps(evt)])

    # --- Sheet 6: V3_Procedure_Events ---
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    ws_v3.append(["note_id", "event_hierarchy_json"])
    ws_v3.append([NOTE_ID, json.dumps(EVENTS, indent=2)])

    # --- Sheet 7: V3_Registry_JSON ---
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["note_id", "full_registry_json"])
    
    registry_data = {
        "metadata": {"note_id": NOTE_ID, "source": SOURCE_FILE},
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS
    }
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated Phase 0 Workbook: {OUTPUT_PATH}")

# ==============================================================================
# 7. EXECUTION
# ==============================================================================
if __name__ == "__main__":
    generate_workbook()