import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# =============================================================================
# 1. IMPORTS & CONSTANTS
# =============================================================================

NOTE_ID = "note_133"
SOURCE_FILE = "note_133.txt"
PROCEDURE_DATE = ""  # Not explicitly in text, leave blank or infer if filename had date
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_133 SOURCE_FILE: note_133.txt Procedure Name: Bronchoscopy
Indications: Chronic cough; shortness of breath
Medications: General anesthesia;
2% lidocaine, 10 mL instilled to the tracheobronchial tree

Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered, and informed consent was documented per institutional protocol.
A history and physical examination were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention. Following administration of intravenous medications per the anesthesia record and topical anesthesia to the upper airway and tracheobronchial tree, a 0-degree 4.0-mm rigid optic was introduced through the mouth and advanced into the tracheobronchial tree.
The BF-Q190 slim video bronchoscope was then introduced through the rigid bronchoscope after removal of the telescope and advanced into the tracheobronchial tree.
The BF-1TH190 therapeutic video bronchoscope was subsequently introduced in a similar fashion. The patient tolerated the procedure well.
Procedure Description:

The patient was brought to the bronchoscopy suite, where a time-out was performed and radiographic imaging was reviewed prior to the procedure.
Intravenous anesthesia was administered under the direction of Dr. Sarkis.
Once the patient was adequately anesthetized and neuromuscular blockade had been achieved, he was intubated without difficulty using the Dumon orange-striped rigid bronchoscope.
Oropharyngeal landmarks were visualized, revealing extensive redundant soft tissue in the oropharynx.
The vocal cords were visualized and appeared to move normally.
Chronic inflammatory changes were noted in the tracheal mucosa, with evidence of dynamic airway collapse.
The main carina was architecturally distorted, with a fibrotic stricture arising from the posterior wall of the right mainstem bronchus and tethering the anteromedial aspect of the main carina.
The right and left lungs were explored segmentally and subsegmentally using the Q190 video bronchoscope introduced through the rigid bronchoscope.
Examination of the left lung demonstrated chronic inflammatory changes of the bronchial mucosa with dynamic airway collapse resulting in less than 50% luminal obstruction during respiration.
Examination of the right lung revealed a previously placed stent in situ with approximately 90% obstruction of the right middle lobe bronchus due to bronchial mucosal edema and granulation tissue.
The stent was heavily coated with thick mucopurulent secretions. The right lower lobe bronchus appeared patent.
Given these findings—particularly the severe obstruction of the right middle lobe bronchus and the exuberant granulation tissue—we elected to remove the stent.
The stent was removed en bloc using forceps without difficulty.
Following stent removal, the reconstructed right bronchus intermedius was carefully examined and demonstrated dynamic airway collapse, with an estimated luminal obstruction of less than 50%.
The web-like fibrotic stricture arising from the posterior aspect of the right mainstem bronchus and tethering the anteromedial aspect of the main carina was divided using an electrocautery knife.
This resulted in partial release of the stricture; however, dynamic airway collapse persisted, causing tissue apposition during respiration.
A diode laser was then used to fulgurate the abutting tissue to discourage further stricture formation.
Once an adequate airway lumen had been established, cryotherapy was applied to the edges of the treated area to further reduce the risk of restenosis.
Given that the degree of luminal obstruction during respiration was less than 50% following intervention and in light of the observed granulation tissue formation, a decision was made not to place a stent at this time.
The plan is to bring the patient back for repeat airway examination in two weeks to determine whether placement of a silicone stent would be appropriate.
Complications: No immediate complications
Estimated Blood Loss: <5 mL

Findings:

Chronic cough

Fibrotic stricture of the right mainstem bronchus, excised

Stent located in the bronchus intermedius, removed en bloc

Chronic mucosal inflammation of the right mainstem bronchus

Post–stent removal luminal obstruction of the right mainstem bronchus and right bronchus intermedius estimated at approximately 30%

Post-Procedure Diagnosis:

Chronic mucosal inflammation of the right mainstem bronchus

Luminal obstruction of the right mainstem bronchus and right bronchus intermedius estimated at approximately 30% following stent removal

Technically successful rigid bronchoscopy with stent removal and stricture resection

The patient remained stable throughout the procedure and was transferred in good condition to the post-bronchoscopy recovery area, where he 
will be observed until discharge criteria are met. Preliminary findings were discussed with the patient, and follow-up in the pulmonary clinic in two weeks has been recommended."""

# =============================================================================
# 2. CONFIGURATION (PROCEDURE FLAGS)
# =============================================================================

PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0, # Stent removal covered by specific action, typically FB implies accidental
    "airway_dilation": 0,
    "airway_stent": 0, # Stent REMOVAL occurred, not placement
    "thermal_ablation": 1, # Electrocautery, Diode Laser
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 1,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 1,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# =============================================================================
# 3. DATA DEFINITION (SPANS)
# =============================================================================

# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    # Event 1: Rigid Intubation/Inspection
    ("0-degree 4.0-mm rigid optic", "DEV_INSTRUMENT", "Rigid Optic", "anesthesia to the upper airway and tracheobronchial tree, a ", "evt_01"),
    ("Dumon orange-striped rigid bronchoscope", "DEV_INSTRUMENT", "Rigid Bronchoscope", "intubated without difficulty using the ", "evt_01"),
    ("tracheal mucosa", "ANAT_AIRWAY", "Trachea", "changes were noted in the ", "evt_01"),
    ("dynamic airway collapse", "OBS_LESION", "Dynamic Airway Collapse", "tracheal mucosa, with evidence of ", "evt_01"),
    ("fibrotic stricture", "OBS_LESION", "Stricture", "distorted, with a ", "evt_01"),
    ("right mainstem bronchus", "ANAT_AIRWAY", "RMB", "posterior wall of the ", "evt_01"),
    ("BF-Q190 slim video bronchoscope", "DEV_INSTRUMENT", "BF-Q190", "The ", "evt_01"),
    ("BF-1TH190 therapeutic video bronchoscope", "DEV_INSTRUMENT", "BF-1TH190", "The ", "evt_01"),
    
    # Event 2: Stent Removal
    ("right middle lobe bronchus", "ANAT_AIRWAY", "RML", "90% obstruction of the ", "evt_02"),
    ("remove the stent", "PROC_ACTION", "Stent Removal", "we elected to ", "evt_02"),
    ("stent", "DEV_STENT", "Stent", "The ", "evt_02"),
    ("removed en bloc", "PROC_METHOD", "Removal en bloc", "stent was ", "evt_02"),
    ("forceps", "DEV_INSTRUMENT", "Forceps", "using ", "evt_02"),
    
    # Event 3: Stricture Division (Electrocautery)
    ("web-like fibrotic stricture", "OBS_LESION", "Stricture", "The ", "evt_03"),
    ("right mainstem bronchus", "ANAT_AIRWAY", "RMB", "posterior aspect of the ", "evt_03"),
    ("divided", "PROC_ACTION", "Division", "main carina was ", "evt_03"),
    ("electrocautery knife", "DEV_INSTRUMENT", "Electrocautery Knife", "using an ", "evt_03"),

    # Event 4: Fulguration (Laser)
    ("diode laser", "DEV_INSTRUMENT", "Diode Laser", "A ", "evt_04"),
    ("fulgurate", "PROC_ACTION", "Fulguration", "used to ", "evt_04"),
    ("abutting tissue", "ANAT_AIRWAY", "Airway Tissue", "the ", "evt_04"),

    # Event 5: Cryotherapy
    ("cryotherapy", "PROC_METHOD", "Cryotherapy", "established, ", "evt_05"),
    ("applied to the edges", "PROC_ACTION", "Application", "was ", "evt_05"),
    ("treated area", "ANAT_AIRWAY", "Treatment Site", "of the ", "evt_05"),

    # Outcomes
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications: ", "evt_out"),
    ("less than 50%", "OUTCOME_AIRWAY_LUMEN_POST", "<50%", "estimated luminal obstruction of ", "evt_out"),
    ("tolerated the procedure well", "OUTCOME_SYMPTOMS", "Tolerated Well", "The patient ", "evt_out")
]

# =============================================================================
# 4. EVENT DEFINITIONS
# =============================================================================

EVENTS = [
    {
        "event_id": "evt_01",
        "name": "Rigid Bronchoscopy & Inspection",
        "procedure_type": "Bronchoscopy",
        "modality": "Rigid / Flexible",
        "anatomy": ["Trachea", "RMB", "Main Carina"],
        "devices": ["Dumon Rigid Bronchoscope", "Rigid Optic", "BF-Q190", "BF-1TH190"],
        "action": "Inspection",
        "notes": "Visualization of dynamic collapse and stricture."
    },
    {
        "event_id": "evt_02",
        "name": "Stent Removal",
        "procedure_type": "Therapeutic Bronchoscopy",
        "modality": "Rigid",
        "anatomy": ["RML", "Bronchus Intermedius"],
        "devices": ["Forceps"],
        "action": "Removal",
        "notes": "Stent removed en bloc due to obstruction/granulation."
    },
    {
        "event_id": "evt_03",
        "name": "Stricture Division",
        "procedure_type": "Therapeutic Bronchoscopy",
        "modality": "Rigid",
        "anatomy": ["RMB", "Main Carina"],
        "devices": ["Electrocautery Knife"],
        "action": "Division",
        "notes": "Web-like fibrotic stricture divided."
    },
    {
        "event_id": "evt_04",
        "name": "Laser Fulguration",
        "procedure_type": "Therapeutic Bronchoscopy",
        "modality": "Rigid",
        "anatomy": ["RMB"],
        "devices": ["Diode Laser"],
        "action": "Fulguration",
        "notes": "To discourage stricture formation."
    },
    {
        "event_id": "evt_05",
        "name": "Cryotherapy",
        "procedure_type": "Therapeutic Bronchoscopy",
        "modality": "Rigid",
        "anatomy": ["RMB"],
        "devices": ["Cryoprobe"],
        "action": "Cryotherapy",
        "notes": "Applied to edges of treated area."
    }
]

# =============================================================================
# 5. HELPER FUNCTIONS
# =============================================================================

def clean_text(text):
    if not text:
        return ""
    return text.strip().replace('\r', '').replace('\n', ' ')

def hydrate_span(full_text, span_text, context_prefix):
    """
    Finds the start and end indices of span_text in full_text,
    using context_prefix to disambiguate.
    """
    clean_full = clean_text(full_text)
    clean_span = clean_text(span_text)
    clean_context = clean_text(context_prefix)
    
    # 1. Construct search pattern: context + span
    search_phrase = clean_context + clean_span
    
    try:
        # Find the context+span in the text
        match = re.search(re.escape(search_phrase), clean_full, re.IGNORECASE)
        if match:
            # The span starts after the context ends
            start_index = match.start() + len(clean_context)
            end_index = start_index + len(clean_span)
            return start_index, end_index, clean_full[start_index:end_index]
        else:
            # Fallback: Try just the span if context fails (risky for duplicates)
            match_span = re.search(re.escape(clean_span), clean_full, re.IGNORECASE)
            if match_span:
                return match_span.start(), match_span.end(), clean_full[match_span.start():match_span.end()]
            return None, None, "NOT FOUND"
            
    except Exception as e:
        return None, None, f"ERROR: {str(e)}"

# =============================================================================
# 6. WORKBOOK GENERATION
# =============================================================================

def generate_workbook():
    # A. Create/Load Workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_ws = wb.active
    wb.remove(default_ws)
    
    # B. Sheet 1: Note_Text
    ws_text = wb.create_sheet("Note_Text")
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # C. Sheet 2: Note_Index
    ws_index = wb.create_sheet("Note_Index")
    headers_index = ["NOTE_ID", "PROCEDURE_DATE", "PATIENT_ID", "PROVIDER_ID"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers_index)
    
    row_data = [NOTE_ID, PROCEDURE_DATE, "", ""] # Patient/Provider blank
    for flag in PROCEDURE_FLAGS.keys():
        row_data.append(PROCEDURE_FLAGS[flag])
    ws_index.append(row_data)
    
    # D. Sheet 3: Span_Annotations (Raw)
    ws_anno = wb.create_sheet("Span_Annotations")
    headers_anno = ["NOTE_ID", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "EVENT_ID", "START_CHAR", "END_CHAR"]
    ws_anno.append(headers_anno)
    
    # E. Sheet 4: Span_Hydrated (Calculated Offsets)
    ws_hydrated = wb.create_sheet("Span_Hydrated")
    ws_hydrated.append(headers_anno + ["EXTRACTED_TEXT_CHECK"])
    
    for span in SPANS:
        span_text, label, norm_val, context, evt_id = span
        
        # Write to Annotations (No offsets yet)
        ws_anno.append([NOTE_ID, span_text, label, norm_val, context, evt_id, "", ""])
        
        # Calculate Offsets
        start, end, extracted = hydrate_span(NOTE_TEXT, span_text, context)
        ws_hydrated.append([NOTE_ID, span_text, label, norm_val, context, evt_id, start, end, extracted])

    # F. Sheet 5: Event_Log (Flattened V3 Events)
    ws_evt_log = wb.create_sheet("Event_Log")
    headers_log = ["NOTE_ID", "EVENT_ID", "EVENT_NAME", "PROCEDURE_TYPE", "MODALITY", "ANATOMY", "DEVICES", "ACTION", "NOTES"]
    ws_evt_log.append(headers_log)
    
    for evt in EVENTS:
        ws_evt_log.append([
            NOTE_ID,
            evt["event_id"],
            evt["name"],
            evt["procedure_type"],
            evt["modality"],
            ", ".join(evt["anatomy"]),
            ", ".join(evt["devices"]),
            evt["action"],
            evt["notes"]
        ])

    # G. Sheet 6: V3_Procedure_Events (JSON-ready structure)
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    ws_v3.append(["NOTE_ID", "EVENT_ID", "EVENT_JSON"])
    
    for evt in EVENTS:
        json_str = json.dumps(evt)
        ws_v3.append([NOTE_ID, evt["event_id"], json_str])

    # H. Sheet 7: V3_Registry_JSON (Full Document)
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["NOTE_ID", "FULL_JSON_BLOB"])
    
    full_registry_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source": SOURCE_FILE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS,
        "spans": [
            {
                "text": s[0], 
                "label": s[1], 
                "normalized": s[2], 
                "event_id": s[4]
            } for s in SPANS
        ]
    }
    
    ws_json.append([NOTE_ID, json.dumps(full_registry_data, indent=2)])

    # I. Save
    wb.save(OUTPUT_PATH)
    print(f"Successfully generated {OUTPUT_PATH}")

# =============================================================================
# 7. EXECUTION
# =============================================================================

if __name__ == "__main__":
    generate_workbook()