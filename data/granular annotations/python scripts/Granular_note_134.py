import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# 1. Imports & Constants
NOTE_ID = "note_134"
SOURCE_FILE = "note_134.txt"
PROCEDURE_DATE = ""
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# Embedded note text (used directly for span matching)
NOTE_TEXT = """NOTE_ID:  note_134 SOURCE_FILE: note_134.txt Procedure Name: Bronchoscopy with Endobronchial Ultrasound
Indications: Right upper lobe mass, diagnostic
Medications: General anesthesia;
2% lidocaine, tracheobronchial tree 10 mL

Pre-Procedure

Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered, and informed consent was documented per institutional protocol.
A history and physical examination were performed and updated in the preprocedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention.

Following intravenous medications per the anesthesia record and topical anesthesia to the upper airway and tracheobronchial tree, the Q180 slim video bronchoscope was introduced through the mouth via laryngeal mask airway and advanced into the tracheobronchial tree.
The UC180F convex probe EBUS bronchoscope was subsequently introduced via the same route. The patient tolerated the procedure well.
Procedure Description

The laryngeal mask airway was in normal position. The vocal cords moved normally with respiration.
The subglottic space was normal. The trachea was of normal caliber, and the carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level.
Bronchial mucosa and anatomy were normal, with no endobronchial lesions or secretions except as outlined below.
Evidence of prior surgery was noted in the left mainstem bronchus. The bronchial stump appeared well healed.
The flexible bronchoscope was withdrawn and replaced with the EBUS bronchoscope.
A systematic hilar and mediastinal lymph node survey was performed.
Lymph Nodes and Mass Evaluation

Lymph node sizing was performed using endobronchial ultrasound.
Sampling was performed using an Olympus 22-gauge EBUS-TBNA needle and sent for routine cytology.
Station 4R (lower paratracheal): Measured 1.9 mm by EBUS and 2.4 mm by CT. PET negative.
Ultrasound characteristics included hypoechoic, heterogeneous, irregular shape with sharp margins.
This node was not biopsied due to benign ultrasound characteristics and size criteria.
Station 7 (subcarinal): Measured 5.4 mm by EBUS and 4.8 mm by CT. PET negative.
Ultrasound characteristics included hypoechoic, heterogeneous, irregular shape with sharp margins.
This node was biopsied using a 22-gauge needle with five passes. ROSE preliminary analysis indicated adequate tissue.
Station 10R (hilar): Measured 3.4 mm by EBUS and 0.1 mm by CT. PET negative.
Ultrasound characteristics included hypoechoic, heterogeneous, irregular shape with sharp margins.
This node was not biopsied due to benign ultrasound characteristics and size criteria.
Station 11Rs: Measured 7.3 mm by EBUS and 5.4 mm by CT. PET negative.
Ultrasound characteristics included hypoechoic, heterogeneous, irregular shape with sharp margins.
This node was biopsied using a 22-gauge needle with five passes. ROSE preliminary analysis indicated adequate tissue.
Right upper lobe mass: Measured 19 mm by EBUS and 22 mm by CT. PET positive.
Ultrasound characteristics included hypoechoic, heterogeneous, irregular shape with sharp margins.
This lesion was biopsied using a 22-gauge needle with eight passes. ROSE preliminary analysis indicated malignancy.
All specimens were sent to cytopathology for review.

Additional Sampling

Fluoroscopically guided transbronchial brushings were obtained from the right upper lobe and sent for routine cytology.
Two samples were obtained.

Transbronchial biopsies were performed in the right upper lobe apical segment (B1) using forceps under fluoroscopic guidance.
Five biopsy passes were performed, yielding five biopsy specimens, which were sent for histopathologic examination.

Complications

No immediate complications.
Estimated Blood Loss

Less than 5 mL.

Impression

Technically successful flexible bronchoscopy with endobronchial ultrasound-guided lymph node and mass sampling

Evidence of prior left mainstem bronchial surgery with well-healed stump

Transbronchial lung biopsies performed

Lymph node sizing and sampling performed

Fluoroscopically guided transbronchial brushings obtained

Post-Procedure Diagnosis

As above.
The patient remained stable and was transferred in good condition to the post-bronchoscopy recovery area, where he will be observed until discharge criteria are met.
Preliminary findings were discussed with the patient. Follow-up with the requesting service for final pathology results was recommended.
Plan:

Await cytology and histopathology results."""

# 2. Configuration (Procedure Flags)
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 1,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 1,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 3. Data Definition (Spans)
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("Q180 slim video bronchoscope", "DEV_INSTRUMENT", "Q180 Scope", "tracheobronchial tree, the ", "evt_00"),
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "UC180F EBUS Scope", "subsequently introduced via the same route. ", "evt_00"),
    ("left mainstem bronchus", "ANAT_AIRWAY", "Left Mainstem Bronchus", "Evidence of prior surgery was noted in the ", "evt_01"),
    ("Olympus 22-gauge EBUS-TBNA needle", "DEV_NEEDLE", "22G EBUS Needle", "Sampling was performed using an ", "evt_00"),
    
    # Station 4R
    ("Station 4R", "ANAT_LN_STATION", "4R", "Sampling was performed", "evt_02"),
    ("1.9 mm", "MEAS_SIZE", "1.9 mm", "(lower paratracheal): Measured ", "evt_02"),
    ("not biopsied", "PROC_ACTION", "No Biopsy", "This node was ", "evt_02"),

    # Station 7
    ("Station 7", "ANAT_LN_STATION", "7", "size criteria.\n", "evt_03"),
    ("5.4 mm", "MEAS_SIZE", "5.4 mm", "(subcarinal): Measured ", "evt_03"),
    ("biopsied", "PROC_ACTION", "Biopsy", "This node was ", "evt_03"),
    ("22-gauge needle", "DEV_NEEDLE", "22G Needle", "This node was biopsied using a ", "evt_03"),
    ("five passes", "MEAS_COUNT", "5", "with ", "evt_03"),
    ("adequate tissue", "OBS_ROSE", "Adequate", "ROSE preliminary analysis indicated ", "evt_03"),

    # Station 10R
    ("Station 10R", "ANAT_LN_STATION", "10R", "adequate tissue.\n", "evt_04"),
    ("3.4 mm", "MEAS_SIZE", "3.4 mm", "(hilar): Measured ", "evt_04"),
    ("not biopsied", "PROC_ACTION", "No Biopsy", "This node was ", "evt_04"),

    # Station 11Rs
    ("Station 11Rs", "ANAT_LN_STATION", "11Rs", "size criteria.\n", "evt_05"),
    ("7.3 mm", "MEAS_SIZE", "7.3 mm", "Measured ", "evt_05"),
    ("biopsied", "PROC_ACTION", "Biopsy", "This node was ", "evt_05"),
    ("22-gauge needle", "DEV_NEEDLE", "22G Needle", "This node was biopsied using a ", "evt_05"),
    ("five passes", "MEAS_COUNT", "5", "with ", "evt_05"),
    ("adequate tissue", "OBS_ROSE", "Adequate", "ROSE preliminary analysis indicated ", "evt_05"),

    # RUL Mass
    ("Right upper lobe mass", "OBS_LESION", "Mass", "adequate tissue.\n", "evt_06"),
    ("19 mm", "MEAS_SIZE", "19 mm", "Measured ", "evt_06"),
    ("biopsied", "PROC_ACTION", "Biopsy", "This lesion was ", "evt_06"),
    ("22-gauge needle", "DEV_NEEDLE", "22G Needle", "This lesion was biopsied using a ", "evt_06"),
    ("eight passes", "MEAS_COUNT", "8", "with ", "evt_06"),
    ("malignancy", "OBS_ROSE", "Malignancy", "ROSE preliminary analysis indicated ", "evt_06"),

    # Brushings
    ("transbronchial brushings", "PROC_METHOD", "Brushings", "Fluoroscopically guided ", "evt_07"),
    ("right upper lobe", "ANAT_LUNG_LOC", "RUL", "obtained from the ", "evt_07"),
    ("Two samples", "MEAS_COUNT", "2", "cytology.\n", "evt_07"),

    # TBBx
    ("Transbronchial biopsies", "PROC_METHOD", "Transbronchial Biopsy", "obtained.\n\n", "evt_08"),
    ("right upper lobe apical segment (B1)", "ANAT_LUNG_LOC", "RUL B1", "performed in the ", "evt_08"),
    ("forceps", "DEV_INSTRUMENT", "Forceps", "using ", "evt_08"),
    ("Five biopsy passes", "MEAS_COUNT", "5", "guidance.\n", "evt_08"),

    # Outcomes
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications\n\n", "evt_99"),
]

# 4. Event Definitions
EVENTS = [
    {
        "event_id": "evt_01",
        "method": "Inspection",
        "anatomy": "Left Mainstem Bronchus",
        "findings": "Prior surgery, well healed stump"
    },
    {
        "event_id": "evt_02",
        "method": "Linear EBUS",
        "anatomy": "Station 4R",
        "measurements": "1.9 mm",
        "action": "Inspection Only (No Biopsy)"
    },
    {
        "event_id": "evt_03",
        "method": "Linear EBUS-TBNA",
        "anatomy": "Station 7",
        "devices": "22G Needle",
        "measurements": "5.4 mm, 5 passes",
        "findings": "ROSE: Adequate"
    },
    {
        "event_id": "evt_04",
        "method": "Linear EBUS",
        "anatomy": "Station 10R",
        "measurements": "3.4 mm",
        "action": "Inspection Only (No Biopsy)"
    },
    {
        "event_id": "evt_05",
        "method": "Linear EBUS-TBNA",
        "anatomy": "Station 11Rs",
        "devices": "22G Needle",
        "measurements": "7.3 mm, 5 passes",
        "findings": "ROSE: Adequate"
    },
    {
        "event_id": "evt_06",
        "method": "Linear EBUS-TBNA",
        "anatomy": "Right Upper Lobe Mass",
        "devices": "22G Needle",
        "measurements": "19 mm, 8 passes",
        "findings": "ROSE: Malignancy"
    },
    {
        "event_id": "evt_07",
        "method": "Transbronchial Brushings",
        "anatomy": "Right Upper Lobe",
        "measurements": "2 samples"
    },
    {
        "event_id": "evt_08",
        "method": "Transbronchial Biopsy",
        "anatomy": "RUL Apical Segment (B1)",
        "devices": "Forceps",
        "measurements": "5 passes"
    },
    {
        "event_id": "evt_99",
        "method": "Procedure Completion",
        "outcome": "No Complications"
    }
]

# 5. Helper Functions
def clean_text(text):
    if not text: return ""
    return text.strip().replace('\r', '')

def hydrate_span(text, span_text, context_prefix):
    # Normalize text for searching
    clean_full = text.replace('\r', '')
    
    # Try exact match first with context
    search_str = context_prefix + span_text if context_prefix else span_text
    start_idx = clean_full.find(search_str)
    
    if start_idx == -1:
        # Fallback: try finding just the span_text
        start_idx = clean_full.find(span_text)
        if start_idx == -1:
            return 0, 0, "" # Not found
    else:
        # Adjust start_idx to point to span_text, skipping context
        if context_prefix:
            start_idx += len(context_prefix)

    end_idx = start_idx + len(span_text)
    
    # Extract snippet for verification (context + span + suffix)
    snippet_start = max(0, start_idx - 20)
    snippet_end = min(len(clean_full), end_idx + 20)
    snippet = clean_full[snippet_start:snippet_end]
    
    return start_idx, end_idx, snippet

# 6. Workbook Generation Function
def generate_workbook():
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # --- Sheet 1: Note_Text ---
    ws_text = wb.create_sheet("Note_Text")
    ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # --- Sheet 2: Note_Index ---
    ws_index = wb.create_sheet("Note_Index")
    headers = ["note_id", "source_file", "procedure_date"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers)
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE] + [PROCEDURE_FLAGS[k] for k in PROCEDURE_FLAGS]
    ws_index.append(row_data)

    # --- Sheet 3: Span_Annotations ---
    ws_anno = wb.create_sheet("Span_Annotations")
    ws_anno.append(["note_id", "span_text", "label", "normalized_value", "start_char", "end_char", "event_id"])
    for span in SPANS:
        # Placeholder for start/end in this sheet (manual entry view)
        ws_anno.append([NOTE_ID, span[0], span[1], span[2], "", "", span[4]])

    # --- Sheet 4: Span_Hydrated ---
    ws_hydrated = wb.create_sheet("Span_Hydrated")
    ws_hydrated.append(["note_id", "span_text", "label", "normalized_value", "context_prefix", "start_char", "end_char", "snippet", "event_id"])
    
    for span in SPANS:
        s_text, label, norm_val, context, evt_id = span
        start, end, snippet = hydrate_span(NOTE_TEXT, s_text, context)
        ws_hydrated.append([NOTE_ID, s_text, label, norm_val, context, start, end, snippet, evt_id])

    # --- Sheet 5: Event_Log ---
    ws_events = wb.create_sheet("Event_Log")
    ws_events.append(["note_id", "event_id", "method", "anatomy", "devices", "measurements", "findings", "action", "outcome"])
    for evt in EVENTS:
        ws_events.append([
            NOTE_ID,
            evt.get("event_id", ""),
            evt.get("method", ""),
            evt.get("anatomy", ""),
            evt.get("devices", ""),
            evt.get("measurements", ""),
            evt.get("findings", ""),
            evt.get("action", ""),
            evt.get("outcome", "")
        ])

    # --- Sheet 6: V3_Procedure_Events ---
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    ws_v3.append(["note_id", "event_id", "event_type", "event_details_json"])
    for evt in EVENTS:
        details = {k: v for k, v in evt.items() if k != "event_id"}
        ws_v3.append([NOTE_ID, evt.get("event_id"), "Procedure_Step", json.dumps(details)])

    # --- Sheet 7: V3_Registry_JSON ---
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["note_id", "full_registry_json"])
    
    registry_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source_file": SOURCE_FILE,
            "procedure_date": PROCEDURE_DATE
        },
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS
    }
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    if os.path.exists(OUTPUT_PATH):
        os.remove(OUTPUT_PATH)
    wb.save(OUTPUT_PATH)
    print(f"Generated Phase 0 Workbook: {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()