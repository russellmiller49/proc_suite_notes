import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# ==============================================================================
# 1. IMPORTS & CONSTANTS
# ==============================================================================
NOTE_ID = "note_135"
SOURCE_FILE = "note_135.txt"
PROCEDURE_DATE = "" 
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_135 SOURCE_FILE: note_135.txt Procedure: Bronchoscopy (rigid + flexible)
Indication: Collapsed left lung due to left mainstem airway stent occlusion (stent placed on [REDACTED]).
Anesthesia/Ventilation: General anesthesia with jet ventilation. Topical 2% lidocaine to tracheobronchial tree (2 mL).
Pre-procedure status: ASA III, ECOG 3. Standard consent/time-out performed. Labs/imaging reviewed.
Protocol: Enrolled in protocol 2010-0990 (normal saline vs 4.2% bicarbonate solution for mucus obstruction clearance in airway stent lumen).
Technique

Q180 slim video bronchoscope introduced orally to trachea.

T180 therapeutic video bronchoscope introduced through rigid bronchoscope and advanced through the tracheobronchial tree.
Findings

Larynx: Normal.

Trachea/Carina: No significant pathology.

Right lung: No significant pathology.
Left lung:

Left mainstem bronchus: 14 mm x 4 cm Microvasive stent present.
Stent lumen completely obstructed by blood clot and clear retained secretions.
After instillation of 10 cc of “agent A” (per study protocol), clot successfully removed with no significant bleeding and no residual stent lumen obstruction (stent lumen patent).
Left lower lobe: Malignant airway disease with:

>90% near-complete obstruction of segmental bronchi from malignant extrinsic compression and submucosal infiltration.
An additional completely obstructing lesion in the LLL due to malignant disease.
Stent position: Stent covers the left upper lobe takeoff.

Impression

Complete occlusion of left mainstem stent (blood clot + secretions) causing collapsed left lung → successfully cleared, stent patent post-intervention.
Severe malignant obstruction of the left lower lobe (near-complete and complete obstructions), consistent with advanced malignant airway involvement.
No right-sided, tracheal, or carinal pathology identified.

No specimens obtained.

Complications / Blood Loss

No immediate complications.

Estimated blood loss: None."""

# ==============================================================================
# 2. CONFIGURATION (PROCEDURE FLAGS)
# ==============================================================================
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1, # Clearance of stent secretions/clot
    "foreign_body_removal": 0, # Clot/mucus typically aspiration, not FB
    "airway_dilation": 0,
    "airway_stent": 0, # Stent present/cleaned, not placed
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 1,
    
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# ==============================================================================
# 3. DATA DEFINITION (SPANS)
# ==============================================================================
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("Bronchoscopy (rigid + flexible)", "PROC_METHOD", "Rigid and Flexible Bronchoscopy", "Procedure: ", "evt_01"),
    ("Q180 slim video bronchoscope", "DEV_INSTRUMENT", "Q180", "Technique\n\n", "evt_01"),
    ("T180 therapeutic video bronchoscope", "DEV_INSTRUMENT", "T180", "\n\n", "evt_01"),
    ("rigid bronchoscope", "DEV_INSTRUMENT", "Rigid Bronchoscope", "introduced through ", "evt_01"),
    ("Left mainstem bronchus", "ANAT_AIRWAY", "LMB", "Left lung:\n\n", "evt_01"),
    ("Microvasive stent", "DEV_STENT", "Microvasive", "14 mm x 4 cm ", "evt_01"),
    ("14 mm x 4 cm", "DEV_STENT_SIZE", "14mm x 4cm", "bronchus: ", "evt_01"),
    ("completely obstructed", "OUTCOME_AIRWAY_LUMEN_PRE", "100%", "Stent lumen ", "evt_01"),
    ("blood clot", "OBS_LESION", "Blood Clot", "obstructed by ", "evt_01"),
    ("10 cc", "MEAS_VOL", "10 cc", "instillation of ", "evt_01"),
    ("clot successfully removed", "PROC_ACTION", "Clot Removal", "protocol), ", "evt_01"),
    ("stent lumen patent", "OUTCOME_AIRWAY_LUMEN_POST", "Patent", "obstruction (", "evt_01"),
    ("Left lower lobe", "ANAT_AIRWAY", "LLL", "patent).\n", "evt_02"),
    ("Malignant airway disease", "OBS_LESION", "Malignant Airway Disease", "Left lower lobe: ", "evt_02"),
    (">90% near-complete obstruction", "OBS_LESION", ">90% Obstruction", "disease with:\n\n", "evt_02"),
    ("completely obstructing lesion", "OBS_LESION", "Complete Obstruction", "An additional ", "evt_02"),
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications / Blood Loss\n\n", "evt_global")
]

# ==============================================================================
# 4. EVENT DEFINITIONS
# ==============================================================================
EVENTS = [
    {
        "event_id": "evt_01",
        "type": "Therapeutic",
        "subtype": "Airway Clearance",
        "method": ["Rigid Bronchoscopy", "Therapeutic Aspiration"],
        "anatomy": ["LMB"],
        "devices": ["Rigid Bronchoscope", "Q180", "T180", "Microvasive Stent"],
        "outcomes": ["Stent Patent", "Clot Removed"],
        "notes": "Clearance of occluded stent in LMB."
    },
    {
        "event_id": "evt_02",
        "type": "Diagnostic",
        "subtype": "Inspection",
        "method": ["Flexible Bronchoscopy"],
        "anatomy": ["LLL"],
        "devices": [],
        "outcomes": ["Malignant Obstruction"],
        "notes": "Diagnostic findings of malignancy in LLL."
    }
]

# ==============================================================================
# 5. HELPER FUNCTIONS
# ==============================================================================
def clean_text(text):
    if not text:
        return ""
    return text.strip().replace('\r', '')

def hydrate_span(full_text, span_text, context_prefix):
    """
    Finds the start and end indices of span_text within full_text,
    relying on context_prefix to disambiguate.
    """
    clean_full = clean_text(full_text)
    clean_span = clean_text(span_text)
    clean_context = clean_text(context_prefix)
    
    # 1. Construct search pattern: context + span
    # We use re.escape to handle special chars in medical text
    pattern = re.escape(clean_context) + r"\s*" + re.escape(clean_span)
    
    match = re.search(pattern, clean_full, re.IGNORECASE)
    
    if match:
        # The match includes the context. We need the offset of the span part.
        # Logic: full_match_start + length_of_context -> span_start
        
        # However, re.search with \s* might make length calc tricky due to variable whitespace.
        # Safer approach: find the exact span text starting near the end of the context match.
        full_match_str = match.group(0)
        
        # We know the span ends the match.
        span_start = match.end() - len(clean_span)
        span_end = match.end()
        
        return span_start, span_end
    else:
        # Fallback: simple search without context (risky for duplicates, but better than 0)
        start_idx = clean_full.find(clean_span)
        if start_idx != -1:
            return start_idx, start_idx + len(clean_span)
        return 0, 0

# ==============================================================================
# 6. WORKBOOK GENERATION FUNCTION
# ==============================================================================
def generate_workbook():
    # 1. Load or Create Workbook
    if not os.path.exists(TEMPLATE_PATH):
        wb = openpyxl.Workbook()
        # Create expected sheets
        for sheet_name in ["Note_Text", "Note_Index", "Span_Annotations", 
                           "Span_Hydrated", "Event_Log", "V3_Procedure_Events", 
                           "V3_Registry_JSON"]:
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
        # Remove default 'Sheet' if exists
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # 2. Sheet: Note_Text
    ws_text = wb["Note_Text"]
    # Clear existing rows > 1
    ws_text.delete_rows(2, ws_text.max_row)
    # Header check
    if ws_text.cell(1, 1).value != "note_id":
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 3. Sheet: Note_Index (Flags)
    ws_index = wb["Note_Index"]
    ws_index.delete_rows(2, ws_index.max_row)
    
    # Headers: note_id, procedure_date + 30 flags
    headers = ["note_id", "procedure_date"] + list(PROCEDURE_FLAGS.keys())
    
    # Write Header if empty
    if ws_index.max_column < len(headers):
        for col_num, header in enumerate(headers, 1):
            ws_index.cell(row=1, column=col_num, value=header)
            
    # Write Data
    row_data = [NOTE_ID, PROCEDURE_DATE] + list(PROCEDURE_FLAGS.values())
    ws_index.append(row_data)

    # 4. Sheet: Span_Annotations (Raw Spans)
    ws_anno = wb["Span_Annotations"]
    ws_anno.delete_rows(2, ws_anno.max_row)
    if ws_anno.max_column < 6:
        ws_anno.append(["note_id", "span_text", "label", "normalized_value", "context_prefix", "event_id"])
    
    for span in SPANS:
        # span: (text, label, norm, context, event)
        ws_anno.append([NOTE_ID, span[0], span[1], span[2], span[3], span[4]])

    # 5. Sheet: Span_Hydrated (Calculated Offsets)
    ws_hydra = wb["Span_Hydrated"]
    ws_hydra.delete_rows(2, ws_hydra.max_row)
    if ws_hydra.max_column < 8:
        ws_hydra.append(["note_id", "start_char", "end_char", "span_text", "label", "normalized_value", "context_prefix", "event_id"])
    
    cleaned_note = clean_text(NOTE_TEXT)
    
    for span in SPANS:
        s_text, s_label, s_norm, s_context, s_event = span
        start, end = hydrate_span(cleaned_note, s_text, s_context)
        ws_hydra.append([NOTE_ID, start, end, s_text, s_label, s_norm, s_context, s_event])

    # 6. Sheet: Event_Log (Flattened)
    ws_log = wb["Event_Log"]
    ws_log.delete_rows(2, ws_log.max_row)
    if ws_log.max_column < 4:
        ws_log.append(["note_id", "event_id", "key", "value"])
        
    for evt in EVENTS:
        e_id = evt["event_id"]
        # Explode dictionary
        for k, v in evt.items():
            if k == "event_id": continue
            if isinstance(v, list):
                for item in v:
                    ws_log.append([NOTE_ID, e_id, k, item])
            else:
                ws_log.append([NOTE_ID, e_id, k, str(v)])

    # 7. Sheet: V3_Procedure_Events (Detailed)
    ws_v3 = wb["V3_Procedure_Events"]
    ws_v3.delete_rows(2, ws_v3.max_row)
    if ws_v3.max_column < 8:
        ws_v3.append(["note_id", "event_id", "procedure_type", "procedure_subtype", "method_json", "anatomy_json", "devices_json", "outcomes_json"])
    
    for evt in EVENTS:
        ws_v3.append([
            NOTE_ID,
            evt["event_id"],
            evt.get("type", ""),
            evt.get("subtype", ""),
            json.dumps(evt.get("method", [])),
            json.dumps(evt.get("anatomy", [])),
            json.dumps(evt.get("devices", [])),
            json.dumps(evt.get("outcomes", []))
        ])

    # 8. Sheet: V3_Registry_JSON (Full Payload)
    ws_json = wb["V3_Registry_JSON"]
    ws_json.delete_rows(2, ws_json.max_row)
    if ws_json.max_column < 2:
        ws_json.append(["note_id", "json_payload"])
    
    # Construct Full JSON
    full_payload = {
        "metadata": {
            "note_id": NOTE_ID,
            "source": SOURCE_FILE,
            "procedure_date": PROCEDURE_DATE
        },
        "flags": PROCEDURE_FLAGS,
        "spans": [
            {
                "text": s[0],
                "label": s[1],
                "normalized": s[2],
                "context": s[3],
                "event_id": s[4]
            } for s in SPANS
        ],
        "events": EVENTS
    }
    
    ws_json.append([NOTE_ID, json.dumps(full_payload, indent=2)])

    # 9. Save
    wb.save(OUTPUT_PATH)
    print(f"Generated Phase 0 Workbook: {OUTPUT_PATH}")

# ==============================================================================
# 7. EXECUTION BLOCK
# ==============================================================================
if __name__ == "__main__":
    generate_workbook()