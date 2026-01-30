import openpyxl
from openpyxl.styles import Font
import re
import json
import os
import datetime

# ==========================================
# 1. Imports & Constants
# ==========================================

NOTE_ID = "note_136"
SOURCE_FILE = "note_136.txt"
PROCEDURE_DATE = "" 
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_136 SOURCE_FILE: note_136.txt Procedure Performed:
Flexible bronchoscopy with endobronchial ultrasound–guided transbronchial needle aspiration

Indication:
Diagnostic evaluation and mediastinal staging

Medications:
General anesthesia
Topical anesthesia: Lidocaine 2%, 10 mL to the tracheobronchial tree

The procedure, including risks, benefits, and alternatives, was explained to the patient.
All questions were answered, and informed consent was obtained and documented per institutional protocol.
A history and physical examination were performed and updated in the pre-procedure assessment record.
Pertinent laboratory studies and imaging were reviewed. A procedural time-out was performed.
Following administration of intravenous medications per the anesthesia record and topical anesthesia to the upper airway and tracheobronchial tree, a Q180 slim video bronchoscope was introduced orally via a laryngeal mask airway and advanced into the tracheobronchial tree.
Subsequently, a UC180F convex-probe EBUS bronchoscope was introduced orally via the laryngeal mask airway and advanced into the tracheobronchial tree.
The patient tolerated the procedure well.

Airway Examination

The laryngeal mask airway was in good position.
The vocal cords moved normally with respiration. The subglottic space was normal.
The trachea was of normal caliber, and the carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level.
Bronchial mucosa and anatomy were normal, with no endobronchial lesions or significant secretions.
A bronchial stump was noted without evidence of recurrence.

Endobronchial Ultrasound Findings and Sampling

The bronchoscope was withdrawn and replaced with the EBUS bronchoscope to perform ultrasound evaluation.
A systematic mediastinal and hilar lymph node survey was conducted.
Lymph node sizing and sampling were performed for non–small cell lung cancer staging using an Olympus 22-gauge EBUS-TBNA needle.
Samples were obtained from the following stations and sent for routine cytology:

Station 4L (left lower paratracheal):
Measured 4.6 mm by EBUS and 5 mm by CT;
PET negative. On ultrasound, the lymph node appeared hypoechoic, heterogeneous, irregularly shaped, with sharp margins.
The node was biopsied using a 22-gauge needle. A total of five needle passes were obtained.
Rapid on-site evaluation indicated adequate tissue.

Station 5 (subaortic):
Measured 20.7 mm by EBUS and 24.1 mm by CT; PET positive.
This node was accessed via the esophagus and biopsy was technically challenging.
On ultrasound, the lymph node appeared hypoechoic, heterogeneous, irregularly shaped, with sharp margins.
The node was biopsied using a 22-gauge needle. A total of eight needle passes were obtained.
Rapid on-site evaluation was suspicious for malignancy.

All samples were submitted to cytopathology for review.
Complications:
None immediate

Estimated Blood Loss:
<5 mL

Post-Procedure Diagnosis:

Technically successful flexible bronchoscopy with EBUS-guided transbronchial needle aspiration

Normal endobronchial examination

Mediastinal lymph node sampling performed

Disposition:
The patient remained stable and was transferred in good condition to the bronchoscopy recovery area, where he will be observed until discharge criteria are met.
Preliminary findings were discussed with the patient. Follow-up with the requesting service for final pathology results was recommended.
Recommendations:

Await cytology results

Consider CT-guided biopsy if EBUS sampling is nondiagnostic"""

# ==========================================
# 2. Configuration (Procedure Flags)
# ==========================================

PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# ==========================================
# 3. Data Definition (Spans)
# ==========================================
# Format: (span_text, label, normalized_value, context_prefix, event_id)

SPANS = [
    ("Flexible bronchoscopy", "PROC_METHOD", "Flexible Bronchoscopy", "Procedure Performed:\n", "evt_01"),
    ("endobronchial ultrasound–guided transbronchial needle aspiration", "PROC_METHOD", "Linear EBUS", "Flexible bronchoscopy with ", "evt_01"),
    ("Q180 slim video bronchoscope", "DEV_INSTRUMENT", "Olympus Q180", "tracheobronchial tree, a ", "evt_01"),
    ("UC180F convex-probe EBUS bronchoscope", "DEV_INSTRUMENT", "Olympus UC180F", "Subsequently, a ", "evt_02"),
    ("Olympus 22-gauge EBUS-TBNA needle", "DEV_NEEDLE", "Olympus EBUS Needle", "staging using an ", "evt_02"),
    ("22-gauge", "DEV_CATHETER_SIZE", "22G", "staging using an Olympus ", "evt_02"),
    
    # Station 4L
    ("Station 4L", "ANAT_LN_STATION", "4L", "cytology:\n\n", "evt_03"),
    ("Measured 4.6 mm", "MEAS_SIZE", "4.6mm", "left lower paratracheal):\n", "evt_03"),
    ("hypoechoic", "OBS_LESION", "Hypoechoic", "lymph node appeared ", "evt_03"),
    ("heterogeneous", "OBS_LESION", "Heterogeneous", "hypoechoic, ", "evt_03"),
    ("five needle passes", "MEAS_COUNT", "5", "A total of ", "evt_03"),
    ("adequate tissue", "OBS_ROSE", "Adequate", "evaluation indicated ", "evt_03"),

    # Station 5
    ("Station 5", "ANAT_LN_STATION", "5", "adequate tissue.\n\n", "evt_04"),
    ("Measured 20.7 mm", "MEAS_SIZE", "20.7mm", "(subaortic):\n", "evt_04"),
    ("accessed via the esophagus", "PROC_METHOD", "Trans-esophageal (EUS-B)", "This node was ", "evt_04"),
    ("hypoechoic", "OBS_LESION", "Hypoechoic", "On ultrasound, the lymph node appeared ", "evt_04"),
    ("heterogeneous", "OBS_LESION", "Heterogeneous", "hypoechoic, ", "evt_04"),
    ("eight needle passes", "MEAS_COUNT", "8", "A total of ", "evt_04"),
    ("suspicious for malignancy", "OBS_ROSE", "Malignant", "evaluation was ", "evt_04"),

    # Outcomes
    ("tolerated the procedure well", "OUTCOME_SYMPTOMS", "Tolerated Well", "The patient ", "evt_05"),
    ("None immediate", "OUTCOME_COMPLICATION", "None", "Complications:\n", "evt_05")
]

# ==========================================
# 4. Event Definitions
# ==========================================

EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_step": "Diagnostic Bronchoscopy",
        "device": ["Q180 slim video bronchoscope"],
        "anatomy": ["Tracheobronchial Tree"],
        "action": "Inspection",
        "outcome": "Normal airway exam"
    },
    {
        "event_id": "evt_02",
        "procedure_step": "EBUS Survey",
        "device": ["UC180F convex-probe EBUS bronchoscope", "Olympus 22-gauge EBUS-TBNA needle"],
        "anatomy": ["Mediastinum"],
        "action": "Ultrasound Survey",
        "outcome": "Nodes identified"
    },
    {
        "event_id": "evt_03",
        "procedure_step": "EBUS-TBNA",
        "device": ["Olympus 22-gauge EBUS-TBNA needle"],
        "anatomy": ["Station 4L"],
        "action": "Aspiration",
        "measurements": {"size": "4.6 mm", "passes": 5},
        "observation": "Hypoechoic, Heterogeneous",
        "rose_status": "Adequate"
    },
    {
        "event_id": "evt_04",
        "procedure_step": "EBUS-TBNA (Trans-esophageal)",
        "device": ["Olympus 22-gauge EBUS-TBNA needle"],
        "anatomy": ["Station 5"],
        "action": "Aspiration",
        "measurements": {"size": "20.7 mm", "passes": 8},
        "observation": "Hypoechoic, Heterogeneous",
        "rose_status": "Suspicious for malignancy"
    },
    {
        "event_id": "evt_05",
        "procedure_step": "Conclusion",
        "outcome": "Tolerated well, No complications"
    }
]

# ==========================================
# 5. Helper Functions
# ==========================================

def clean_text(text):
    if not text:
        return ""
    return text.strip().replace('\r', '')

def hydrate_span(full_text, span_text, context_prefix):
    """
    Finds the start and end indices of span_text in full_text,
    verifying it follows context_prefix to ensure unique matching.
    """
    clean_full = clean_text(full_text)
    clean_span = clean_text(span_text)
    clean_context = clean_text(context_prefix)

    # Simple approach: find context, then find span immediately after
    if clean_context:
        # Escape regex special characters in context and span
        pattern = re.escape(clean_context) + r"\s*" + re.escape(clean_span)
        match = re.search(pattern, clean_full, re.IGNORECASE)
        if match:
            # We found context + span. Now identify exactly where the span part is.
            # The span is at the end of the match.
            # We can find the start of the span by subtracting span length from end? 
            # Not reliable due to whitespace normalization in regex.
            # Better: Find the starting position of the span within the matched string
            full_match_str = match.group(0)
            # Find span inside the match (should be at the end)
            span_start_in_match = full_match_str.lower().rfind(clean_span.lower())
            
            start_index = match.start() + span_start_in_match
            end_index = start_index + len(clean_span)
            return start_index, end_index, clean_full[start_index:end_index]
    
    # Fallback if no context or context calculation fails: find first occurrence
    start_index = clean_full.find(clean_span)
    if start_index != -1:
        end_index = start_index + len(clean_span)
        return start_index, end_index, clean_full[start_index:end_index]

    return None, None, None

def generate_workbook():
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ------------------------------------
    # Sheet 1: Note_Text
    # ------------------------------------
    ws_text = wb.create_sheet("Note_Text")
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, clean_text(NOTE_TEXT)])
    
    # Formatting
    ws_text.column_dimensions['C'].width = 100
    for cell in ws_text[1]:
        cell.font = Font(bold=True)

    # ------------------------------------
    # Sheet 2: Note_Index
    # ------------------------------------
    ws_index = wb.create_sheet("Note_Index")
    headers = ["note_id", "source_file", "procedure_date"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers)
    
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE] + [PROCEDURE_FLAGS[k] for k in PROCEDURE_FLAGS]
    ws_index.append(row_data)
    
    for cell in ws_index[1]:
        cell.font = Font(bold=True)

    # ------------------------------------
    # Sheet 3: Span_Annotations
    # ------------------------------------
    ws_annotations = wb.create_sheet("Span_Annotations")
    headers_ann = ["note_id", "span_text", "label", "normalized_value", "context_prefix", "start_char", "end_char", "event_id"]
    ws_annotations.append(headers_ann)
    
    for span in SPANS:
        # Unpack, leaving start/end blank for now
        txt, lbl, norm, ctx, evt = span
        ws_annotations.append([NOTE_ID, txt, lbl, norm, ctx, "", "", evt])

    for cell in ws_annotations[1]:
        cell.font = Font(bold=True)

    # ------------------------------------
    # Sheet 4: Span_Hydrated
    # ------------------------------------
    ws_hydrated = wb.create_sheet("Span_Hydrated")
    ws_hydrated.append(headers_ann)
    
    for span in SPANS:
        txt, lbl, norm, ctx, evt = span
        s, e, found_text = hydrate_span(NOTE_TEXT, txt, ctx)
        if s is not None:
            ws_hydrated.append([NOTE_ID, found_text, lbl, norm, ctx, s, e, evt])
        else:
            # Fallback if not found (shouldn't happen with correct copy-paste)
            ws_hydrated.append([NOTE_ID, txt, lbl, norm, ctx, "NOT_FOUND", "NOT_FOUND", evt])

    for cell in ws_hydrated[1]:
        cell.font = Font(bold=True)

    # ------------------------------------
    # Sheet 5: Event_Log (Flattened)
    # ------------------------------------
    ws_event_log = wb.create_sheet("Event_Log")
    headers_log = ["note_id", "event_id", "procedure_step", "device", "anatomy", "action", "outcome", "details_json"]
    ws_event_log.append(headers_log)
    
    for evt in EVENTS:
        details = {k: v for k, v in evt.items() if k not in ["event_id", "procedure_step", "device", "anatomy", "action", "outcome"]}
        row = [
            NOTE_ID,
            evt.get("event_id", ""),
            evt.get("procedure_step", ""),
            ", ".join(evt.get("device", [])) if isinstance(evt.get("device"), list) else evt.get("device", ""),
            ", ".join(evt.get("anatomy", [])) if isinstance(evt.get("anatomy"), list) else evt.get("anatomy", ""),
            evt.get("action", ""),
            evt.get("outcome", ""),
            json.dumps(details)
        ]
        ws_event_log.append(row)

    for cell in ws_event_log[1]:
        cell.font = Font(bold=True)

    # ------------------------------------
    # Sheet 6: V3_Procedure_Events
    # ------------------------------------
    # Similar to Event_Log but formatted specifically for V3 ingestion if needed
    ws_v3_events = wb.create_sheet("V3_Procedure_Events")
    ws_v3_events.append(headers_log)
    # Copying data from Event_Log logic
    for evt in EVENTS:
        details = {k: v for k, v in evt.items() if k not in ["event_id", "procedure_step", "device", "anatomy", "action", "outcome"]}
        row = [
            NOTE_ID,
            evt.get("event_id", ""),
            evt.get("procedure_step", ""),
            ", ".join(evt.get("device", [])) if isinstance(evt.get("device"), list) else evt.get("device", ""),
            ", ".join(evt.get("anatomy", [])) if isinstance(evt.get("anatomy"), list) else evt.get("anatomy", ""),
            evt.get("action", ""),
            evt.get("outcome", ""),
            json.dumps(details)
        ]
        ws_v3_events.append(row)

    for cell in ws_v3_events[1]:
        cell.font = Font(bold=True)

    # ------------------------------------
    # Sheet 7: V3_Registry_JSON
    # ------------------------------------
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["note_id", "json_data"])
    
    # Construct the Full Registry Object
    registry_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source_file": SOURCE_FILE,
            "procedure_date": PROCEDURE_DATE
        },
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS,
        "spans": [
            {"text": s[0], "label": s[1], "normalized": s[2], "context": s[3], "event_id": s[4]} 
            for s in SPANS
        ]
    }
    
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])
    ws_json.column_dimensions['B'].width = 100

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated workbook: {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()