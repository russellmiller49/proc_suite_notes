import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import re
import json
import os
import datetime

# 1. Imports & Constants
NOTE_ID = "note_137"
SOURCE_FILE = "note_137.txt"
PROCEDURE_DATE = ""
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_137 SOURCE_FILE: note_137.txt Proceduralist(s):
Procedure Name: Pleuroscopy
Indication: Pleural effusion
Anesthesia: Monitored Anesthesia Care

Pre-Anesthesia Assessment

ASA Physical Status: III – Patient with severe systemic disease

The procedure, including risks, benefits, and alternatives, was explained to the patient.
All questions were answered, and informed consent was obtained and documented per institutional protocol.
A focused history and physical examination were performed and updated in the pre-procedure assessment record.
Relevant laboratory studies and imaging were reviewed. A procedural time-out was performed prior to initiation.
Procedure Description

The patient was placed on the operating table in the lateral decubitus position with appropriate padding of all pressure points.
The procedural site was identified using ultrasound guidance and was sterilely prepped with chlorhexidine gluconate (Chloraprep) and draped in the usual fashion.
Local Anesthesia

The pleural entry site was infiltrated with 15 mL of 1% lidocaine.
A 10-mm reusable primary port was placed on the left side at the 6th intercostal space along the anterior axillary line using a Veress needle technique.
An LTF VP Endoeye thoracoscope was introduced through the incision and advanced into the pleural space, followed by a 0-degree 2.0-mm pleuroscopy telescope and then a 0-degree 7.0-mm pleuroscopy telescope.
Pleuroscopy Findings

Extensive adhesions were present throughout the left hemithorax.

Adhesions in the upper hemithorax were soft and thin and were successfully lysed using the pleuroscope.
Dense, thick adhesions were noted in the lower hemithorax and were not taken down due to concern for bleeding.
The parietal pleura was carefully inspected and demonstrated multiple tumor masses involving the posterior mid and upper parietal pleura.
These masses were exophytic, friable, and fungating.

The visceral pleura overlying the upper lobe was thickened but without discrete tumor nodules.
Biopsy

Biopsies of the parietal pleural masses in the mid pleura were obtained using forceps and sent for histopathologic evaluation.
Fourteen samples were obtained.

Careful inspection of the pleural space following biopsy confirmed complete hemostasis at all biopsy sites.
The previously placed 15.5-Fr PleurX catheter was left in place in the pleural space over the diaphragm.
The port site was closed with three interrupted 3-0 silk sutures.

Dressing

Port sites were dressed with a transparent dressing.
Estimated Blood Loss

Minimal

Complications

None immediate

Post-Procedure Diagnosis

Suspected pleural metastasis

Post-Procedure Plan

The patient will be observed post-procedure until all discharge criteria are met.
Chest X-ray to be obtained post-procedure."""

# 2. Configuration (Procedure Flags)
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 0,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 1,
    "pleurodesis": 0,
    "pleural_biopsy": 1,
    "fibrinolytic_therapy": 0
}

# 3. Data Definition (Spans)
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("Pleuroscopy", "PROC_METHOD", "medical thoracoscopy", "Procedure Name:", "evt_01"),
    ("10-mm", "MEAS_SIZE", "10 mm", "A", "evt_01"),
    ("primary port", "DEV_INSTRUMENT", "port", "reusable", "evt_01"),
    ("left side", "LATERALITY", "Left", "placed on the", "evt_01"),
    ("6th intercostal space", "ANAT_PLEURA", "6th intercostal space", "at the", "evt_01"),
    ("anterior axillary line", "ANAT_PLEURA", "anterior axillary line", "along the", "evt_01"),
    ("Veress needle", "DEV_NEEDLE", "Veress needle", "using a", "evt_01"),
    ("LTF VP Endoeye thoracoscope", "DEV_INSTRUMENT", "thoracoscope", "An", "evt_01"),
    ("2.0-mm", "MEAS_SIZE", "2.0 mm", "0-degree", "evt_01"),
    ("pleuroscopy telescope", "DEV_INSTRUMENT", "telescope", "2.0-mm", "evt_01"),
    ("7.0-mm", "MEAS_SIZE", "7.0 mm", "0-degree", "evt_01"),
    ("pleuroscopy telescope", "DEV_INSTRUMENT", "telescope", "7.0-mm", "evt_01"),
    ("Extensive adhesions", "OBS_LESION", "adhesions", "Findings\n\n", "evt_02"),
    ("left hemithorax", "ANAT_PLEURA", "left hemithorax", "throughout the", "evt_02"),
    ("lysed", "PROC_ACTION", "lysis", "were successfully", "evt_02"),
    ("pleuroscope", "DEV_INSTRUMENT", "pleuroscope", "using the", "evt_02"),
    ("Biopsies", "PROC_METHOD", "pleural biopsy", "Biopsy\n\n", "evt_03"),
    ("parietal pleural masses", "OBS_LESION", "tumor", "of the", "evt_03"),
    ("mid pleura", "ANAT_PLEURA", "mid pleura", "masses in the", "evt_03"),
    ("forceps", "DEV_INSTRUMENT", "forceps", "using", "evt_03"),
    ("Fourteen", "MEAS_COUNT", "14", "samples were obtained.", "evt_03"),
    ("15.5-Fr", "DEV_CATHETER_SIZE", "15.5 Fr", "previously placed", "evt_04"),
    ("PleurX catheter", "DEV_CATHETER", "PleurX", "15.5-Fr", "evt_04"),
    ("pleural space", "ANAT_PLEURA", "pleural space", "place in the", "evt_04"),
    ("None immediate", "OUTCOME_COMPLICATION", "None", "Complications\n\n", "evt_05")
]

# 4. Event Definitions
EVENTS = [
    {
        "event_id": "evt_01",
        "type": "medical_thoracoscopy",
        "method": "medical thoracoscopy",
        "anatomy": "Left pleural space",
        "devices": ["10-mm reusable primary port", "Veress needle", "LTF VP Endoeye thoracoscope", "2.0-mm telescope", "7.0-mm telescope"],
        "outcomes": []
    },
    {
        "event_id": "evt_02",
        "type": "adhesiolysis",
        "method": "adhesiolysis",
        "anatomy": "Left hemithorax",
        "devices": ["pleuroscope"],
        "outcomes": ["adhesions lysed"]
    },
    {
        "event_id": "evt_03",
        "type": "pleural_biopsy",
        "method": "pleural biopsy",
        "anatomy": "Parietal pleura (mid)",
        "devices": ["forceps"],
        "outcomes": ["14 samples"]
    },
    {
        "event_id": "evt_04",
        "type": "ipc_check",
        "method": "inspection",
        "anatomy": "Pleural space",
        "devices": ["15.5-Fr PleurX catheter"],
        "outcomes": []
    },
    {
        "event_id": "evt_05",
        "type": "outcome",
        "method": "",
        "anatomy": "",
        "devices": [],
        "outcomes": ["No immediate complications"]
    }
]

# 5. Helper Functions
def clean_text(text):
    if not text: return ""
    return re.sub(r'[\r\n]+', ' ', text).strip()

def hydrate_span(full_text, span_text, context_prefix):
    # Normalize texts for search to handle potential newline differences
    normalized_full = full_text.replace('\n', ' ').replace('\r', '')
    normalized_span = span_text.replace('\n', ' ').replace('\r', '')
    normalized_context = context_prefix.replace('\n', ' ').replace('\r', '')

    # Find start index
    # Attempt to find with context first
    search_phrase = f"{normalized_context}{normalized_span}"
    start_index = normalized_full.find(search_phrase)
    
    if start_index != -1:
        # Found with context, adjust start to point to span
        real_start = start_index + len(normalized_context)
        return real_start, real_start + len(normalized_span)
    else:
        # Fallback: find span without context (risky if duplicates, but per instructions we try)
        start_index = normalized_full.find(normalized_span)
        if start_index != -1:
            return start_index, start_index + len(normalized_span)
    
    return -1, -1

# 6. Workbook Generation Function
def generate_workbook():
    # Create workbook
    wb = openpyxl.Workbook()
    
    # --- Sheet 1: Note_Text ---
    ws_text = wb.active
    ws_text.title = "Note_Text"
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # --- Sheet 2: Note_Index ---
    ws_index = wb.create_sheet("Note_Index")
    headers_index = ["NOTE_ID", "PROCEDURE_DATE", "SOURCE_FILE"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers_index)
    row_data = [NOTE_ID, PROCEDURE_DATE, SOURCE_FILE] + [PROCEDURE_FLAGS[k] for k in PROCEDURE_FLAGS]
    ws_index.append(row_data)
    
    # --- Sheet 3: Span_Annotations ---
    ws_anno = wb.create_sheet("Span_Annotations")
    headers_anno = ["NOTE_ID", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "START_CHAR", "END_CHAR", "EVENT_ID"]
    ws_anno.append(headers_anno)
    
    # Data for hydration later
    hydrated_data = []

    for span in SPANS:
        s_text, label, norm, ctx, evt = span
        # Placeholder for start/end
        ws_anno.append([NOTE_ID, s_text, label, norm, ctx, "", "", evt])
        hydrated_data.append(span)

    # --- Sheet 4: Span_Hydrated ---
    ws_hydra = wb.create_sheet("Span_Hydrated")
    ws_hydra.append(headers_anno)
    
    for span in hydrated_data:
        s_text, label, norm, ctx, evt = span
        start, end = hydrate_span(NOTE_TEXT, s_text, ctx)
        ws_hydra.append([NOTE_ID, s_text, label, norm, ctx, start, end, evt])

    # --- Sheet 5: Event_Log ---
    ws_event = wb.create_sheet("Event_Log")
    ws_event.append(["NOTE_ID", "EVENT_ID", "EVENT_TYPE", "METHOD", "ANATOMY", "DEVICES", "OUTCOMES"])
    
    for evt in EVENTS:
        devices_str = ", ".join(evt["devices"])
        outcomes_str = ", ".join(evt["outcomes"])
        ws_event.append([NOTE_ID, evt["event_id"], evt["type"], evt["method"], evt["anatomy"], devices_str, outcomes_str])

    # --- Sheet 6: V3_Procedure_Events ---
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    ws_v3.append(["NOTE_ID", "EVENT_ID", "JSON_PAYLOAD"])
    
    for evt in EVENTS:
        payload = json.dumps(evt)
        ws_v3.append([NOTE_ID, evt["event_id"], payload])

    # --- Sheet 7: V3_Registry_JSON ---
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["NOTE_ID", "FULL_JSON"])
    
    registry_data = {
        "note_id": NOTE_ID,
        "source": SOURCE_FILE,
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS
    }
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    if os.path.exists(OUTPUT_PATH):
        os.remove(OUTPUT_PATH)
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

# 7. Execution Block
if __name__ == "__main__":
    generate_workbook()