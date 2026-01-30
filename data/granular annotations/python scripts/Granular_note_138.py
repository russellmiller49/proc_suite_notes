import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# 1. Imports & Constants
NOTE_ID = "note_138"
SOURCE_FILE = "note_138.txt"
PROCEDURE_DATE = ""
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_138 SOURCE_FILE: note_138.txt Procedure Name: Bronchoscopy
Indications: Lung mass
Medications: General anesthesia;
2% lidocaine instilled to the tracheobronchial tree (volume per anesthesia record)

Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered, and informed consent was documented per institutional protocol.
A history and physical examination were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention. Following administration of intravenous medications per the anesthesia record and topical anesthesia to the upper airway and tracheobronchial tree, the BF-H190 slim video bronchoscope was introduced through the mouth via a laryngeal mask airway and advanced into the tracheobronchial tree.
The UC180F convex probe EBUS bronchoscope was subsequently introduced in a similar fashion. The patient tolerated the procedure well.
Procedure Description:

The laryngeal mask airway was in good position. The vocal cords moved normally with respiration.
The subglottic space was normal. The trachea was of normal caliber, and the carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level.
Bronchial mucosa and anatomy were normal, with no endobronchial lesions and no secretions noted on the left.
A stricture was identified in the right bronchus intermedius involving the right lower lobe bronchus.
The narrowing appeared fibrotic, with near-complete occlusion of the airway lumen. The lesion was not traversed.
Findings / EBUS Evaluation:

The bronchoscope was withdrawn and replaced with the EBUS bronchoscope to perform ultrasound examination.
A systematic hilar and mediastinal lymph node survey was conducted.
Lymph node sizing was performed via endobronchial ultrasound in the setting of suspected non-small cell lung cancer.
Sampling by transbronchial needle aspiration was performed using an Olympus EBUS-TBNA 22-gauge needle targeting a left pulmonary artery mass, with specimens sent for routine cytology.
The left pulmonary artery mass measured 10.1 mm by EBUS and 11 mm by CT. PET imaging was positive.
On ultrasound, the mass was hypoechoic, heterogeneous, irregularly shaped, with sharp margins, and appeared to form part of the pulmonary artery wall.
Flow was noted within the left pulmonary artery and, to a lesser extent, the right pulmonary artery.
The mass was biopsied using a 22-gauge needle. A total of three passes were performed.
Rapid on-site evaluation preliminary assessment was nondiagnostic.

All specimens were sent to cytopathology for review.
Complications: No immediate complications
Estimated Blood Loss: <5 mL

Post-Procedure Diagnosis:

Lymph node sizing and sampling performed

Technically successful flexible bronchoscopy with endobronchial ultrasound-guided biopsies

The patient remained stable throughout the procedure and was transferred in good condition to the post-bronchoscopy recovery area, where he will be observed until discharge criteria are met.
Preliminary findings were discussed with the patient, and follow-up with the requesting service for final pathology results has been recommended.
Plan:

Await cytology results"""

# 2. Configuration (Procedure Flags)
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 1, # EBUS-TBNA performed
    "linear_ebus": 1, # UC180F used
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0, # Parenchymal forceps biopsy not performed
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 3. Data Definition (Spans)
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    # Scope 1: Diagnostic Flexible
    ("BF-H190 slim video bronchoscope", "DEV_INSTRUMENT", "BF-H190", "tracheobronchial tree, the", "evt_01"),
    ("tracheobronchial tree", "ANAT_AIRWAY", "Tracheobronchial Tree", "advanced into the", "evt_01"),
    ("stricture", "OBS_LESION", "Stricture", "A", "evt_01"),
    ("right bronchus intermedius", "ANAT_AIRWAY", "RBI", "stricture was identified in the", "evt_01"),
    ("right lower lobe bronchus", "ANAT_AIRWAY", "RLL", "involving the", "evt_01"),
    ("fibrotic", "OBS_LESION", "Fibrosis", "narrowing appeared", "evt_01"),
    ("near-complete occlusion", "OUTCOME_AIRWAY_LUMEN_PRE", "Near-complete occlusion", "fibrotic, with", "evt_01"),
    
    # Scope 2: EBUS Inspection
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "UC180F", "The", "evt_02"),
    ("systematic hilar and mediastinal lymph node survey", "PROC_ACTION", "Lymph Node Survey", "A", "evt_02"),
    ("Lymph node sizing", "PROC_ACTION", "Lymph Node Sizing", "survey was conducted.", "evt_02"),
    
    # Scope 2: TBNA
    ("transbronchial needle aspiration", "PROC_METHOD", "TBNA", "Sampling by", "evt_03"),
    ("Olympus EBUS-TBNA 22-gauge needle", "DEV_NEEDLE", "Olympus EBUS-TBNA 22G", "using an", "evt_03"),
    ("left pulmonary artery mass", "OBS_LESION", "Mass", "targeting a", "evt_03"),
    ("10.1 mm", "MEAS_SIZE", "10.1 mm", "measured", "evt_03"),
    ("hypoechoic", "OBS_LESION", "Hypoechoic", "mass was", "evt_03"),
    ("heterogeneous", "OBS_LESION", "Heterogeneous", "hypoechoic,", "evt_03"),
    ("biopsied", "PROC_ACTION", "Biopsy", "mass was", "evt_03"),
    ("22-gauge", "DEV_CATHETER_SIZE", "22G", "using a", "evt_03"),
    ("three passes", "MEAS_COUNT", "3", "A total of", "evt_03"),
    ("nondiagnostic", "OBS_ROSE", "Nondiagnostic", "assessment was", "evt_03"),
    
    # Global Outcomes
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications:", "evt_global"),
    ("patient remained stable", "OUTCOME_SYMPTOMS", "Stable", "The", "evt_global")
]

# 4. Event Definitions
EVENTS = [
    {
        "event_id": "evt_01",
        "type": "Diagnostic Bronchoscopy",
        "method": "Flexible Bronchoscopy",
        "devices": ["BF-H190"],
        "anatomy": ["Right Bronchus Intermedius", "Right Lower Lobe Bronchus"],
        "findings": ["Stricture", "Fibrosis", "Near-complete occlusion"]
    },
    {
        "event_id": "evt_02",
        "type": "EBUS Inspection",
        "method": "Linear EBUS",
        "devices": ["UC180F"],
        "anatomy": ["Hilar Lymph Nodes", "Mediastinal Lymph Nodes"],
        "findings": ["Lymph node sizing"]
    },
    {
        "event_id": "evt_03",
        "type": "TBNA",
        "method": "EBUS-TBNA",
        "devices": ["Olympus EBUS-TBNA 22-gauge needle"],
        "anatomy": ["Left Pulmonary Artery Mass"],
        "findings": ["Mass (10.1mm)", "Hypoechoic", "Heterogeneous", "ROSE: Nondiagnostic"],
        "actions": ["Biopsy (3 passes)"]
    }
]

# 5. Helper Functions
def clean_text(text):
    if not text: return ""
    return re.sub(r'[\r\n]+', ' ', text).strip()

def hydrate_span(full_text, span_text, context_prefix):
    """
    Finds the start and end offsets of span_text within full_text,
    disambiguated by context_prefix.
    """
    clean_full = clean_text(full_text)
    clean_span = clean_text(span_text)
    clean_context = clean_text(context_prefix)
    
    # Create a flexible regex pattern escaping special chars
    # We want to find context + ... + span
    # This is a simplified approach assuming relative proximity
    
    # First, find context location
    context_match = re.search(re.escape(clean_context), clean_full, re.IGNORECASE)
    if not context_match:
        return "", "", "" # Context not found
    
    context_end = context_match.end()
    
    # Search for span after context
    search_region = clean_full[context_end:]
    span_match = re.search(re.escape(clean_span), search_region, re.IGNORECASE)
    
    if not span_match:
        # Fallback: try searching strictly for span if context fail was too strict
        # But per requirements, we should rely on context.
        # Let's try finding span globally if strict context fails, 
        # but risk of ambiguity is high. 
        # For this generator, we return empty if not found with context to ensure accuracy.
        return "", "", ""
        
    start_offset = context_end + span_match.start()
    end_offset = context_end + span_match.end()
    
    return start_offset, end_offset, clean_full[start_offset:end_offset]

# 6. Workbook Generation Function
def generate_workbook():
    wb = openpyxl.Workbook()
    
    # --- Sheet 1: Note_Text ---
    ws_text = wb.active
    ws_text.title = "Note_Text"
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # --- Sheet 2: Note_Index ---
    ws_index = wb.create_sheet("Note_Index")
    headers = ["NOTE_ID", "SOURCE_FILE", "PROCEDURE_DATE"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers)
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE] + [PROCEDURE_FLAGS[k] for k in PROCEDURE_FLAGS]
    ws_index.append(row_data)
    
    # --- Sheet 3: Span_Annotations ---
    ws_annotations = wb.create_sheet("Span_Annotations")
    headers_ann = ["NOTE_ID", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "START_CHAR", "END_CHAR", "EVENT_ID"]
    ws_annotations.append(headers_ann)
    
    # --- Sheet 4: Span_Hydrated ---
    ws_hydrated = wb.create_sheet("Span_Hydrated")
    ws_hydrated.append(headers_ann)
    
    # Process Spans
    for span in SPANS:
        s_text, s_label, s_norm, s_context, s_event = span
        
        # Annotation Row (Offsets Blank)
        ann_row = [NOTE_ID, s_text, s_label, s_norm, s_context, "", "", s_event]
        ws_annotations.append(ann_row)
        
        # Hydration Logic
        start, end, found_text = hydrate_span(NOTE_TEXT, s_text, s_context)
        hyd_row = [NOTE_ID, s_text, s_label, s_norm, s_context, start, end, s_event]
        ws_hydrated.append(hyd_row)

    # --- Sheet 5: Event_Log ---
    ws_event_log = wb.create_sheet("Event_Log")
    ws_event_log.append(["NOTE_ID", "EVENT_ID", "EVENT_TYPE", "DETAILS"])
    for evt in EVENTS:
        details = json.dumps(evt)
        ws_event_log.append([NOTE_ID, evt["event_id"], evt["type"], details])

    # --- Sheet 6: V3_Procedure_Events ---
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    v3_headers = ["NOTE_ID", "EVENT_ID", "ACTION", "METHOD", "ANATOMY", "DEVICES", "FINDINGS", "COMMENTS"]
    ws_v3.append(v3_headers)
    
    for evt in EVENTS:
        action = evt.get("actions", [])
        method = evt.get("method", "")
        anatomy = evt.get("anatomy", [])
        devices = evt.get("devices", [])
        findings = evt.get("findings", [])
        
        ws_v3.append([
            NOTE_ID,
            evt["event_id"],
            ", ".join(action) if isinstance(action, list) else str(action),
            method,
            ", ".join(anatomy) if isinstance(anatomy, list) else str(anatomy),
            ", ".join(devices) if isinstance(devices, list) else str(devices),
            ", ".join(findings) if isinstance(findings, list) else str(findings),
            ""
        ])

    # --- Sheet 7: V3_Registry_JSON ---
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["NOTE_ID", "JSON_OBJECT"])
    
    registry_data = {
        "note_id": NOTE_ID,
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS
    }
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])
    
    # Save
    if not os.path.exists(OUTPUT_PATH):
        wb.save(OUTPUT_PATH)
        print(f"Workbook saved to {OUTPUT_PATH}")
    else:
        print(f"Workbook already exists at {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()