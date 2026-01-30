import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import re
import json
import os
import datetime

# -----------------------------------------------------------------------------
# 1. CONSTANTS & INPUT DATA
# -----------------------------------------------------------------------------
NOTE_ID = "note_139"
SOURCE_FILE = "note_139.txt"
PROCEDURE_DATE = ""  # Not found in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_139 SOURCE_FILE: note_139.txt Procedure Name: Bronchoscopy
Indications: Multiple pulmonary nodules
Medications: Per anesthesia record

Pre-Procedure

Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered, and informed consent was documented per institutional protocol.
A history and physical examination were performed and updated in the preprocedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention.

Following intravenous medications per the anesthesia record and topical anesthesia to the upper airway and tracheobronchial tree, the T180 therapeutic video bronchoscope was introduced through the mouth via laryngeal mask airway and advanced into the tracheobronchial tree.
The Q180 slim video bronchoscope was subsequently introduced via the same route.
Procedure Description / Findings

Post-radiation changes were noted involving the vocal cords and aryepiglottic folds.
Papillomatous lesions were identified involving the right mainstem bronchus and bronchus intermedius.
Abnormal mucosa was noted along the anterolateral aspect of the upper trachea, predominantly on the left side, with post-treatment changes extending to the mid-trachea.
No clear papillomatous growth was identified in the distal airways.
The patient is status post prior endotracheal biopsies at abnormal papillomatous lesions involving the superior segment stump, proximal bronchus intermedius, and distal bronchus intermedius.
The patient is also status post intratracheal injection of cidofovir (75 mg in 15 mL) at these sites.
Right Lung Abnormalities

A small, barely obstructing (less than 10% luminal obstruction) 8–10 mm polypoid lesion was identified proximally in the bronchus intermedius and in the superior segment of the right lower lobe (B6).
Narrow-band imaging was used for airway examination and demonstrated dotted abnormal vascular patterns in the proximal and distal bronchus intermedius.
These lesions were biopsied.

Complications

No immediate complications.

Estimated Blood Loss

Minimal.
Post-Procedure Diagnosis

Endobronchial papillomatous lesions, status post biopsy

Recommendations

Follow up in clinic in 3 months.
Attending Participation

I was present and participated throughout the entire procedure, including non-key portions."""

# -----------------------------------------------------------------------------
# 2. PROCEDURE FLAGS
# -----------------------------------------------------------------------------
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 1,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# -----------------------------------------------------------------------------
# 3. SPANS
# -----------------------------------------------------------------------------
# (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("Bronchoscopy", "PROC_METHOD", "Bronchoscopy", "Procedure Name: ", "evt_01"),
    ("T180 therapeutic video bronchoscope", "DEV_INSTRUMENT", "Bronchoscope", "tracheobronchial tree, the ", "evt_01"),
    ("Q180 slim video bronchoscope", "DEV_INSTRUMENT", "Bronchoscope", "The ", "evt_01"),
    ("Papillomatous lesions", "OBS_LESION", "Papilloma", "folds.\n", "evt_02"),
    ("right mainstem bronchus", "ANAT_AIRWAY", "Right Mainstem Bronchus", "involving the ", "evt_02"),
    ("bronchus intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "bronchus and ", "evt_02"),
    ("Abnormal mucosa", "OBS_LESION", "Abnormal Mucosa", "", "evt_03"),
    ("upper trachea", "ANAT_AIRWAY", "Trachea", "aspect of the ", "evt_03"),
    ("polypoid lesion", "OBS_LESION", "Polyp", "obstruction) 8–10 mm ", "evt_04"),
    ("8–10 mm", "MEAS_SIZE", "8-10 mm", "obstruction) ", "evt_04"),
    ("bronchus intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "proximally in the ", "evt_04"),
    ("superior segment of the right lower lobe (B6)", "ANAT_AIRWAY", "RB6", "and in the ", "evt_04"),
    ("Narrow-band imaging", "PROC_METHOD", "NBI", "", "evt_05"),
    ("biopsied", "PROC_METHOD", "Endobronchial Biopsy", "These lesions were ", "evt_06"),
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications\n\n", "evt_global")
]

# -----------------------------------------------------------------------------
# 4. EVENTS (Aggregated Data)
# -----------------------------------------------------------------------------
EVENTS = [
    {
        "event_id": "evt_01",
        "name": "Scope Introduction",
        "method": "Bronchoscopy",
        "devices": ["Bronchoscope (T180)", "Bronchoscope (Q180)"]
    },
    {
        "event_id": "evt_02",
        "name": "Findings - Papilloma",
        "anatomy": ["Right Mainstem Bronchus", "Bronchus Intermedius"],
        "findings": ["Papillomatous lesions"]
    },
    {
        "event_id": "evt_03",
        "name": "Findings - Abnormal Mucosa",
        "anatomy": ["Trachea"],
        "findings": ["Abnormal Mucosa"]
    },
    {
        "event_id": "evt_04",
        "name": "Findings - Polyp",
        "anatomy": ["Bronchus Intermedius", "RB6"],
        "findings": ["Polyp"],
        "measurements": ["8-10 mm"]
    },
    {
        "event_id": "evt_05",
        "name": "Imaging",
        "method": "NBI"
    },
    {
        "event_id": "evt_06",
        "name": "Biopsy",
        "method": "Endobronchial Biopsy",
        "target_anatomy": ["Bronchus Intermedius", "RB6"],
        "target_findings": ["Polyp", "Papilloma"]
    },
    {
        "event_id": "evt_global",
        "name": "Outcome",
        "complications": "None"
    }
]

# -----------------------------------------------------------------------------
# 5. HELPER FUNCTIONS
# -----------------------------------------------------------------------------
def clean_text(text):
    if not text: return ""
    return text.replace('\r', '').strip()

def hydrate_span(text, span_text, context_prefix):
    """
    Finds the start/end offsets of span_text within text.
    Uses context_prefix to disambiguate if needed.
    """
    clean_full_text = clean_text(text)
    clean_span = clean_text(span_text)
    clean_context = clean_text(context_prefix)

    if not clean_span:
        return None, None, None

    # Try finding with context first
    search_pattern = re.escape(clean_context) + r"\s*" + re.escape(clean_span)
    match = re.search(search_pattern, clean_full_text, re.IGNORECASE)
    
    if match:
        # found with context
        start_index = match.start() + len(match.group()) - len(clean_span) # Approximate
        # Refine exact start relative to full match
        # (This is a simplified approach; rigid strictness might fail on whitespaces)
        return start_index, start_index + len(clean_span), clean_span
    
    # Fallback: simple find (first occurrence)
    start_index = clean_full_text.lower().find(clean_span.lower())
    if start_index != -1:
        return start_index, start_index + len(clean_span), clean_span
    
    return None, None, None

# -----------------------------------------------------------------------------
# 6. WORKBOOK GENERATION
# -----------------------------------------------------------------------------
def generate_workbook():
    # A. Initialize Workbook
    if os.path.exists(TEMPLATE_PATH):
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    else:
        wb = openpyxl.Workbook()
        # Create expected sheets if missing
        expected_sheets = ["Note_Text", "Note_Index", "Span_Annotations", 
                           "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]
        for sheet_name in expected_sheets:
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
    
    # B. Sheet 1: Note_Text
    ws_text = wb["Note_Text"]
    # Clear existing
    for row in ws_text.iter_rows(min_row=2, max_col=3):
        for cell in row: cell.value = None
    
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # C. Sheet 2: Note_Index
    ws_index = wb["Note_Index"]
    # Write Header if empty
    headers = ["note_id", "source_file", "procedure_date"] + list(PROCEDURE_FLAGS.keys())
    for col_idx, header in enumerate(headers, 1):
        ws_index.cell(row=1, column=col_idx, value=header)
    
    # Write Data
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE] + list(PROCEDURE_FLAGS.values())
    ws_index.append(row_data)

    # D. Sheet 3: Span_Annotations (Manual/Defined input)
    ws_anno = wb["Span_Annotations"]
    # Header: note_id, span_text, label, normalized_value, context_prefix, event_id, start_char, end_char
    ws_anno.delete_rows(2, ws_anno.max_row) # Clear old data
    
    for span in SPANS:
        # tuple: (span_text, label, normalized_value, context_prefix, event_id)
        # We leave start/end blank here for the "human" input simulation
        ws_anno.append([NOTE_ID, span[0], span[1], span[2], span[3], span[4], "", ""])

    # E. Sheet 4: Span_Hydrated (Calculated offsets)
    ws_hydrated = wb["Span_Hydrated"]
    ws_hydrated.delete_rows(2, ws_hydrated.max_row)
    # Header: note_id, span_text, label, normalized_value, start_char, end_char, event_id
    
    for span in SPANS:
        s_text = span[0]
        label = span[1]
        norm = span[2]
        ctx = span[3]
        evt = span[4]
        
        start, end, _ = hydrate_span(NOTE_TEXT, s_text, ctx)
        
        ws_hydrated.append([NOTE_ID, s_text, label, norm, start, end, evt])

    # F. Sheet 5: Event_Log (Flattened)
    ws_elog = wb["Event_Log"]
    ws_elog.delete_rows(2, ws_elog.max_row)
    # Header: note_id, event_id, category, value
    
    for evt in EVENTS:
        eid = evt["event_id"]
        for k, v in evt.items():
            if k == "event_id": continue
            if isinstance(v, list):
                for item in v:
                    ws_elog.append([NOTE_ID, eid, k, item])
            else:
                ws_elog.append([NOTE_ID, eid, k, v])

    # G. Sheet 6: V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    ws_v3.delete_rows(2, ws_v3.max_row)
    # Header: note_id, event_id, event_type, method, anatomy, devices, findings, etc...
    
    for evt in EVENTS:
        # Basic mapping to V3 columns (simplified for this structure)
        row_map = {
            "note_id": NOTE_ID,
            "event_id": evt.get("event_id"),
            "method": json.dumps(evt.get("method", [])),
            "anatomy": json.dumps(evt.get("anatomy", [])),
            "devices": json.dumps(evt.get("devices", [])),
            "findings": json.dumps(evt.get("findings", [])),
            "measurements": json.dumps(evt.get("measurements", [])),
            "complications": json.dumps(evt.get("complications", []))
        }
        # Append as a row (order depends on existing sheet, simply appending here)
        values = list(row_map.values())
        ws_v3.append(values)

    # H. Sheet 7: V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    ws_json.delete_rows(2, ws_json.max_row)
    
    registry_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source": SOURCE_FILE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS
    }
    
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])

    # I. Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()