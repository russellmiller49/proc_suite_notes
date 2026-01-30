import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# -------------------------------------------------------------------------
# 1. Imports & Constants
# -------------------------------------------------------------------------
NOTE_ID = "note_140"
SOURCE_FILE = "note_140.txt"
PROCEDURE_DATE = ""
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_140 SOURCE_FILE: note_140.txt Procedure: Bronchoscopy with radial probe endobronchial ultrasound (RP-EBUS)
Indication: Esophageal cancer with suspected malignant airway involvement
Anesthesia/Ventilation: General anesthesia with controlled mechanical ventilation.
Topical 2% lidocaine to tracheobronchial tree (6 mL).
Pre-procedure status: ASA III. Standard consent and time-out performed.
Labs and imaging reviewed.

Technique

Q180 slim video bronchoscope introduced via tracheostomy to the tracheobronchial tree.
T180 therapeutic bronchoscope also introduced via tracheostomy.

UM-BS20-26R 20 MHz radial probe ultrasound advanced through the working channel to assess airway wall and adjacent structures.
Airway Examination

Tracheostomy stoma: Normal.

Trachea: Normal caliber.

Carina: Sharp.

Bronchial tree: Examined to at least the first subsegmental level.
Mucosa/anatomy: Normal throughout; no endobronchial lesions and no secretions.

Radial Probe Ultrasound Findings

Radial EBUS performed at:

Proximal left mainstem bronchus takeoff

Mid and distal trachea

Esophageal tumor visualized infiltrating the adventitia of the posterior wall of the left mainstem bronchus.
Minimum distance from esophageal mass to left mainstem bronchial lumen: 3.2 mm.

No intraluminal airway invasion identified on bronchoscopic inspection.
Impression

Malignant extrinsic airway involvement from esophageal cancer, affecting the adventitia of the posterior wall of the left mainstem bronchus.
No endobronchial disease and no intrinsic airway obstruction at time of examination.
Findings suggest high-risk proximity of tumor to airway, relevant for oncologic planning and future airway risk assessment.
Complications / Blood Loss

No immediate complications.

Estimated blood loss: None.

Specimens: None collected.
Post-Procedure Plan

Routine post-procedure observation until discharge criteria met.

Follow-up with primary/oncology team as previously scheduled."""

# -------------------------------------------------------------------------
# 2. Configuration (Procedure Flags)
# -------------------------------------------------------------------------
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 1,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0,
}

# -------------------------------------------------------------------------
# 3. Data Definition (Spans)
# -------------------------------------------------------------------------
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    # Event 1: General Bronchoscopy / Scopes
    ("Bronchoscopy", "PROC_METHOD", "Bronchoscopy", "Procedure: ", "evt_01"),
    ("Q180 slim video bronchoscope", "DEV_INSTRUMENT", "Q180 Scope", "Technique\n\n", "evt_01"),
    ("T180 therapeutic bronchoscope", "DEV_INSTRUMENT", "T180 Scope", "", "evt_01"),
    ("tracheobronchial tree", "ANAT_AIRWAY", "Tracheobronchial Tree", "introduced via tracheostomy to the ", "evt_01"),
    
    # Event 2: Radial EBUS
    ("radial probe endobronchial ultrasound", "PROC_METHOD", "Radial EBUS", "Bronchoscopy with ", "evt_02"),
    ("UM-BS20-26R 20 MHz radial probe ultrasound", "DEV_INSTRUMENT", "UM-BS20-26R Radial Probe", "via tracheostomy.\n\n", "evt_02"),
    ("Proximal left mainstem bronchus takeoff", "ANAT_AIRWAY", "Left Mainstem Bronchus", "Radial EBUS performed at:\n\n", "evt_02"),
    ("Mid and distal trachea", "ANAT_AIRWAY", "Trachea", "Proximal left mainstem bronchus takeoff\n\n", "evt_02"),
    ("Esophageal tumor visualized infiltrating the adventitia", "OBS_LESION", "Tumor infiltration", "distal trachea\n\n", "evt_02"),
    ("3.2 mm", "MEAS_SIZE", "3.2 mm", "mainstem bronchial lumen: ", "evt_02"),
    
    # Event 3: Outcomes
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications / Blood Loss\n\n", "evt_03"),
]

# -------------------------------------------------------------------------
# 4. Event Definitions
# -------------------------------------------------------------------------
EVENTS = [
    {
        "event_id": "evt_01",
        "name": "Diagnostic Bronchoscopy",
        "concept_id": "PROC_BRONCHOSCOPY",
        "method": "Bronchoscopy",
        "devices": ["Q180 Scope", "T180 Scope"],
        "anatomy": ["Tracheobronchial Tree"],
        "outcomes": []
    },
    {
        "event_id": "evt_02",
        "name": "Radial EBUS",
        "concept_id": "PROC_RADIAL_EBUS",
        "method": "Radial EBUS",
        "devices": ["UM-BS20-26R Radial Probe"],
        "anatomy": ["Left Mainstem Bronchus", "Trachea"],
        "outcomes": ["Tumor infiltration detected", "Distance 3.2mm"]
    },
    {
        "event_id": "evt_03",
        "name": "Procedure Outcome",
        "concept_id": "OUTCOME_GLOBAL",
        "method": "Observation",
        "devices": [],
        "anatomy": [],
        "outcomes": ["No immediate complications"]
    }
]

# -------------------------------------------------------------------------
# 5. Helper Functions
# -------------------------------------------------------------------------
def clean_text(text):
    if not text:
        return ""
    return text.strip().replace('\r', '')

def hydrate_span(text, span_text, context_prefix):
    """
    Finds the start/end offsets of span_text within text,
    relying on context_prefix to disambiguate.
    """
    clean_full_text = clean_text(text)
    clean_span = clean_text(span_text)
    clean_context = clean_text(context_prefix)

    # 1. Attempt with context
    search_pattern = re.escape(clean_context) + r"\s*" + re.escape(clean_span)
    match = re.search(search_pattern, clean_full_text, re.IGNORECASE)
    
    if match:
        # The span starts after the context
        # We need to find exactly where the span part matches in the group
        full_match_str = match.group(0)
        # Find start of span within the match
        span_start_in_match = full_match_str.lower().rfind(clean_span.lower())
        
        start_index = match.start() + span_start_in_match
        end_index = start_index + len(clean_span)
        return start_index, end_index, clean_full_text[start_index:end_index]

    # 2. Fallback: exact match without context (warning: might pick first instance)
    start_index = clean_full_text.lower().find(clean_span.lower())
    if start_index != -1:
        end_index = start_index + len(clean_span)
        return start_index, end_index, clean_full_text[start_index:end_index]

    return None, None, None

# -------------------------------------------------------------------------
# 6. Workbook Generation Function
# -------------------------------------------------------------------------
def generate_workbook():
    # A. Create/Load Workbook
    if os.path.exists(TEMPLATE_PATH):
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    else:
        wb = openpyxl.Workbook()
        # Create standard sheets if missing
        required_sheets = [
            "Note_Text", "Note_Index", "Span_Annotations", 
            "Span_Hydrated", "Event_Log", "V3_Procedure_Events", 
            "V3_Registry_JSON"
        ]
        for sheet_name in required_sheets:
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
        # Remove default sheet if exists
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # B. Sheet 1: Note_Text
    ws_text = wb["Note_Text"]
    # Clear existing
    for row in ws_text.iter_rows():
        for cell in row:
            cell.value = None
    
    # Write Header and Data
    headers_text = ["note_id", "source_file", "note_text"]
    ws_text.append(headers_text)
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # C. Sheet 2: Note_Index (Metadata + Flags)
    ws_index = wb["Note_Index"]
    for row in ws_index.iter_rows():
        for cell in row:
            cell.value = None
            
    headers_index = [
        "note_id", "source_file", "procedure_date", "patient_id"
    ] + list(PROCEDURE_FLAGS.keys())
    
    ws_index.append(headers_index)
    
    row_values = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE, "Unknown"]
    for key in PROCEDURE_FLAGS:
        row_values.append(PROCEDURE_FLAGS[key])
    ws_index.append(row_values)

    # D. Sheet 3: Span_Annotations (Raw Spans)
    ws_spans = wb["Span_Annotations"]
    for row in ws_spans.iter_rows():
        for cell in row:
            cell.value = None
            
    headers_spans = ["note_id", "span_text", "label", "normalized_value", "context_prefix", "event_id", "start_char", "end_char"]
    ws_spans.append(headers_spans)
    
    # E. Sheet 4: Span_Hydrated (Calculated Offsets)
    ws_hydrated = wb["Span_Hydrated"]
    for row in ws_hydrated.iter_rows():
        for cell in row:
            cell.value = None
            
    ws_hydrated.append(headers_spans + ["extracted_text_check"])

    # Iterate Spans, calculate offsets, populate both sheets
    for span in SPANS:
        span_text, label, norm_val, context, event_id = span
        
        # Calculate offsets
        start, end, extracted = hydrate_span(NOTE_TEXT, span_text, context)
        
        # Row for Sheet 3 (Raw)
        row_raw = [NOTE_ID, span_text, label, norm_val, context, event_id, "", ""]
        ws_spans.append(row_raw)
        
        # Row for Sheet 4 (Hydrated)
        row_hyd = [NOTE_ID, span_text, label, norm_val, context, event_id, start, end, extracted]
        ws_hydrated.append(row_hyd)

    # F. Sheet 5: Event_Log
    ws_events = wb["Event_Log"]
    for row in ws_events.iter_rows():
        for cell in row:
            cell.value = None
            
    headers_events = ["note_id", "event_id", "event_name", "concept_id", "method", "devices", "anatomy", "outcomes"]
    ws_events.append(headers_events)
    
    for evt in EVENTS:
        row_evt = [
            NOTE_ID,
            evt["event_id"],
            evt["name"],
            evt["concept_id"],
            evt["method"],
            ", ".join(evt["devices"]),
            ", ".join(evt["anatomy"]),
            ", ".join(evt["outcomes"])
        ]
        ws_events.append(row_evt)

    # G. Sheet 6: V3_Procedure_Events (Detailed JSON-like structure)
    ws_v3 = wb["V3_Procedure_Events"]
    for row in ws_v3.iter_rows():
        for cell in row:
            cell.value = None
            
    headers_v3 = ["note_id", "event_id", "event_type", "procedure_concept_id", "data_payload"]
    ws_v3.append(headers_v3)
    
    for evt in EVENTS:
        payload = {
            "method": evt["method"],
            "devices": evt["devices"],
            "anatomy": evt["anatomy"],
            "outcomes": evt["outcomes"]
        }
        row_v3 = [
            NOTE_ID,
            evt["event_id"],
            evt["name"],
            evt["concept_id"],
            json.dumps(payload)
        ]
        ws_v3.append(row_v3)

    # H. Sheet 7: V3_Registry_JSON (Full Aggregation)
    ws_json = wb["V3_Registry_JSON"]
    for row in ws_json.iter_rows():
        for cell in row:
            cell.value = None
            
    headers_json = ["note_id", "full_registry_json"]
    ws_json.append(headers_json)
    
    registry_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source_file": SOURCE_FILE,
            "procedure_date": PROCEDURE_DATE
        },
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS
    }
    
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated workbook: {OUTPUT_PATH}")

# -------------------------------------------------------------------------
# 7. Execution Block
# -------------------------------------------------------------------------
if __name__ == "__main__":
    generate_workbook()