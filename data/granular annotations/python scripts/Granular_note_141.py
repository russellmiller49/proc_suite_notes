import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
import re
import json
import os
import datetime

# -------------------------------------------------------------------------
# 1. IMPORTS & CONSTANTS
# -------------------------------------------------------------------------
NOTE_ID = "note_141"
SOURCE_FILE = "note_141.txt"
PROCEDURE_DATE = "2019-04-01" # Inferred from "Plan for repeat... on 4/1/2019" implies current is shortly before or on.
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# Text content with [source] tags removed for clean processing
NOTE_TEXT = """NOTE_ID:  note_141 SOURCE_FILE: note_141.txt Preoperative Diagnosis:
Near-complete tracheal occlusion secondary to granulation tissue and severe tracheomalacia

Postoperative Diagnosis:
Near-complete tracheal occlusion secondary to granulation tissue and severe tracheomalacia

Procedure Performed:
Rigid bronchoscopy with silicone Y-stent placement

Indication:
Near-complete tracheal obstruction requiring urgent intervention

Anesthesia:
General anesthesia

Description of Procedure

Informed consent was obtained from the patient prior to the procedure after discussion, in lay terms, of the indications, procedural details, risks, benefits, and alternatives.
The patient acknowledged understanding and agreed to proceed.

The procedure was performed in the main operating room.
Initial airway inspection was performed using an ultra-slim Olympus P190 flexible bronchoscope inserted through the patient’s customized 5.5-mm tracheostomy tube.
At the distal margin of the tracheostomy tube, approximately 2 cm proximal to the main carina, there was a ridge of granulation tissue occluding approximately 80% of the tracheostomy tube lumen.
Distal to this obstruction, there was severe tracheobronchial malacia with near-complete collapse of the distal trachea and right mainstem bronchus, and approximately 75% obstruction of the left mainstem bronchus.
The airway mucosa was diffusely erythematous and inflamed.

The flexible bronchoscope and tracheostomy tube were removed, and an attempt was made to insert a 14-mm ventilating rigid bronchoscope through the tracheostomy stoma;
this was unsuccessful due to the small size of the stoma.
The tracheostomy tube was reinserted, and rigid bronchoscopy was then attempted via the oral route.
However, the patient had severely distorted upper airway anatomy and significant edema, making visualization of the posterior pharynx and vocal cords impossible.
Attempts were made to insert progressively smaller rigid bronchoscopes, down to a 6-mm scope, but these were unsuccessful due to a fixed stricture within the tracheostomy tract.
At this point, otolaryngology was consulted intraoperatively for assistance with surgical extension of the tracheostomy stoma.
Otolaryngology also attempted visualization of the vocal cords using suspension laryngoscopy but was unsuccessful. The tracheostomy stoma was surgically extended;
however, due to extensive scar tissue and concern for injury to the anterior trachea, the stoma could not be enlarged sufficiently to accommodate a large rigid bronchoscope.
It was, however, extended enough to allow insertion of smaller rigid bronchoscopes.
A 6-mm rigid bronchoscope was inserted through the tracheostomy stoma and advanced to the main carina.
Due to the small diameter of the scope, it was not possible to introduce usable instruments or deploy a silicone stent through the rigid bronchoscope.
The decision was therefore made to deploy the smallest available silicone Y-stent (14 × 10 × 10 mm) through the tracheostomy stoma.
The bronchial limbs were compressed with rigid forceps, and the stent was advanced toward the main carina under direct visualization using a flexible bronchoscope inserted through the left bronchial limb of the stent.
After measurement, the tracheal limb was trimmed to 30 mm, the left bronchial limb to 20 mm, and the right bronchial limb to 10 mm.
Advancement and seating of the stent were extremely challenging due to the patient’s severely narrowed airways.
Multiple techniques were employed, including manipulation with rigid optical forceps, advancement using Kelly clamps by otolaryngology under direct bronchoscopic visualization, gentle balloon dilation using a CRE balloon, and eventual insertion of a slightly larger 8-mm rigid bronchoscope through the tracheostomy stoma and into the stent to help seat the bronchial limbs.
During this maneuver, a small airway tear was visualized along the lateral aspect of the right mainstem bronchus, approximately 0.5 cm proximal to the right upper lobe orifice.
This was felt to be caused by the distal edge of the right bronchial limb of the stent.
Consideration was given to placement of a metallic stent to cover the defect;
however, given concerns regarding metallic stents in benign airway disease, this was deferred.
The patient received 1 gram of cefazolin intraoperatively. Thoracic surgery was consulted by phone, and based on their experience, recommended conservative management provided there was no evidence of pneumothorax.
A portable chest radiograph demonstrated no pneumothorax or pneumomediastinum.

The stent was slightly repositioned to optimize seating at the main carina.
There was mild tenting of the distal right bronchial limb at the site of the airway tear that could not be corrected;
otherwise, the stent was well seated, with resolution of distal tracheal and mainstem bronchial obstruction and patent distal airways.
The patient’s tracheostomy tube was reinserted under direct visualization through the stoma and positioned within the tracheal limb of the stent, where it was secured.
The procedure was then concluded, and the patient was transferred to the ICU.
Postoperative Course

Upon arrival in the ICU, the patient was initially stable but subsequently developed an episode of acute hypoxia with transient PaO₂ <50%, associated with difficulty ventilating.
This resolved rapidly with manual bag ventilation. Flexible bronchoscopy performed through the tracheostomy tube demonstrated no significant stent obstruction, though foamy, blood-tinged secretions were present, consistent with acute pulmonary edema.
A chest radiograph obtained during this event showed no pneumothorax.
The patient received 20 mg of furosemide intravenously, and the condom catheter was exchanged for a Foley catheter, resulting in brisk diuresis and significant improvement in tidal volumes.
Persistently elevated ventilator pressures were felt to be related in part to the small tracheostomy tube.
Under bronchoscopic visualization, the tracheostomy tube was exchanged for a 7.0 Bivona tracheostomy tube, which was positioned midway within the stent.
This resulted in improved airway compliance and ventilation.

The patient was subsequently deemed stable.
After discussion with the patient’s parents, care was transitioned to the ICU team.
Complications:
Small lateral right mainstem bronchial tear

Recommendations

Continue positive-pressure ventilation overnight

Empiric vancomycin and piperacillin–tazobactam for 7 days to prevent infection related to airway injury

Obtain noncontrast CT chest once the patient is stable enough for transport

Plan for repeat bedside flexible bronchoscopy on or around 4/1/2019, or sooner if clinical concerns arise

Notify the Interventional Pulmonology team immediately for any airway issues

Note that the silicone stent is radiolucent and will not be visible on chest radiograph;
CT or bronchoscopy should be used for evaluation of suspected stent-related complications

Administer 3% saline nebulizers three times daily starting tonight and continuing while the stent remains in place to reduce secretion burden and prevent stent obstruction

If tracheostomy tube removal or replacement is required, this must be performed under direct bronchoscopic visualization to ensure correct positioning within the stent and to avoid kinking or obstruction

Interventional Pulmonology will continue to follow"""

# -------------------------------------------------------------------------
# 2. CONFIGURATION (PROCEDURE FLAGS)
# -------------------------------------------------------------------------
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 1, # CRE balloon mentioned
    "airway_stent": 1,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 1,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0,
}

# -------------------------------------------------------------------------
# 3. DATA DEFINITION (SPANS)
# -------------------------------------------------------------------------
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("Rigid bronchoscopy", "PROC_METHOD", "Rigid Bronchoscopy", "Procedure Performed:\n", "evt_01"),
    ("silicone Y-stent placement", "PROC_METHOD", "Stent Placement", "Rigid bronchoscopy with", "evt_01"),
    ("ultra-slim Olympus P190 flexible bronchoscope", "DEV_INSTRUMENT", "Olympus P190", "using an", "evt_02"),
    ("5.5-mm tracheostomy tube", "DEV_CATHETER", "Tracheostomy Tube", "patient’s customized", "evt_02"),
    ("granulation tissue", "OBS_LESION", "Granulation Tissue", "ridge of", "evt_02"),
    ("2 cm proximal to the main carina", "ANAT_AIRWAY", "Trachea", "approximately", "evt_02"),
    ("occluding approximately 80%", "MEAS_AIRWAY_DIAM", "80% Occlusion", "tissue", "evt_02"),
    ("severe tracheobronchial malacia", "OBS_LESION", "Tracheomalacia", "obstruction, there was", "evt_02"),
    ("near-complete collapse of the distal trachea", "OBS_LESION", "Airway Collapse", "malacia with", "evt_02"),
    ("right mainstem bronchus", "ANAT_AIRWAY", "RMB", "collapse of the distal trachea and", "evt_02"),
    ("75% obstruction", "MEAS_AIRWAY_DIAM", "75% Obstruction", "approximately", "evt_02"),
    ("left mainstem bronchus", "ANAT_AIRWAY", "LMB", "obstruction of the", "evt_02"),
    ("14-mm ventilating rigid bronchoscope", "DEV_INSTRUMENT", "Rigid Bronchoscope", "insert a", "evt_03"),
    ("unsuccessful due to the small size of the stoma", "CTX_HISTORICAL", "Failure - Stoma Size", "this was", "evt_03"),
    ("6-mm rigid bronchoscope", "DEV_INSTRUMENT", "Rigid Bronchoscope", "A", "evt_03"),
    ("silicone Y-stent", "DEV_STENT", "Y-Stent", "deploy the smallest available", "evt_04"),
    ("14 × 10 × 10 mm", "DEV_STENT_SIZE", "14x10x10mm", "Y-stent (", "evt_04"),
    ("tracheal limb was trimmed to 30 mm", "DEV_STENT_SIZE", "Tracheal 30mm", "measurement, the", "evt_04"),
    ("left bronchial limb to 20 mm", "DEV_STENT_SIZE", "Left 20mm", "30 mm, the", "evt_04"),
    ("right bronchial limb to 10 mm", "DEV_STENT_SIZE", "Right 10mm", "20 mm, and the", "evt_04"),
    ("CRE balloon", "DEV_INSTRUMENT", "Balloon", "dilation using a", "evt_04"),
    ("8-mm rigid bronchoscope", "DEV_INSTRUMENT", "Rigid Bronchoscope", "slightly larger", "evt_04"),
    ("small airway tear", "OUTCOME_COMPLICATION", "Airway Tear", "maneuver, a", "evt_04"),
    ("lateral aspect of the right mainstem bronchus", "ANAT_AIRWAY", "RMB", "along the", "evt_04"),
    ("7.0 Bivona tracheostomy tube", "DEV_CATHETER", "Tracheostomy Tube", "exchanged for a", "evt_05"),
    ("resolution of distal tracheal and mainstem bronchial obstruction", "OUTCOME_AIRWAY_LUMEN_POST", "Obstruction Resolved", "seated, with", "evt_04"),
    ("patent distal airways", "OUTCOME_AIRWAY_LUMEN_POST", "Patent Airways", "obstruction and", "evt_04"),
]

# -------------------------------------------------------------------------
# 4. EVENT DEFINITIONS
# -------------------------------------------------------------------------
EVENTS = [
    {
        "event_id": "evt_01",
        "name": "Procedure Overview",
        "concept_id": "PROC-001",
        "type": "Procedure",
        "spans": ["Rigid bronchoscopy", "silicone Y-stent placement"]
    },
    {
        "event_id": "evt_02",
        "name": "Initial Inspection",
        "concept_id": "DX-001",
        "type": "Diagnostic",
        "spans": ["ultra-slim Olympus P190 flexible bronchoscope", "5.5-mm tracheostomy tube", "granulation tissue", "severe tracheobronchial malacia", "near-complete collapse"]
    },
    {
        "event_id": "evt_03",
        "name": "Rigid Scope Attempts",
        "concept_id": "PROC-002",
        "type": "Therapeutic",
        "spans": ["14-mm ventilating rigid bronchoscope", "6-mm rigid bronchoscope", "unsuccessful due to the small size of the stoma"]
    },
    {
        "event_id": "evt_04",
        "name": "Stent Deployment & Complication",
        "concept_id": "PROC-003",
        "type": "Therapeutic",
        "spans": ["silicone Y-stent", "14 × 10 × 10 mm", "CRE balloon", "small airway tear", "resolution of distal tracheal and mainstem bronchial obstruction"]
    },
    {
        "event_id": "evt_05",
        "name": "Post-Op Management",
        "concept_id": "MGMT-001",
        "type": "Management",
        "spans": ["7.0 Bivona tracheostomy tube"]
    }
]

# -------------------------------------------------------------------------
# 5. HELPER FUNCTIONS
# -------------------------------------------------------------------------
def clean_text(text):
    if not text:
        return ""
    return text.strip().replace('\r\n', '\n').replace('\r', '\n')

def hydrate_span(text, span_text, context_prefix):
    """
    Finds the exact start/end offset of span_text in text, 
    using context_prefix to disambiguate.
    """
    # 1. Normalize line endings for search
    clean_full_text = text.replace('\r\n', '\n')
    
    # 2. Construct search pattern: prefix + whitespace + span
    # We escape regex characters to be safe
    pattern_str = re.escape(context_prefix) + r"\s*" + re.escape(span_text)
    match = re.search(pattern_str, clean_full_text, re.IGNORECASE | re.DOTALL)
    
    if match:
        # The full match includes the prefix. We want just the span part.
        full_start, full_end = match.span()
        # Calculate where the span actually starts within that match
        # It's at the end of the full match minus the length of the span_text
        # (roughly, but we need to account for whitespace between prefix and span)
        
        # Strategy: find the span_text inside the matched string
        matched_string = clean_full_text[full_start:full_end]
        # Find span_text at the end of matched_string
        span_start_in_match = matched_string.lower().rfind(span_text.lower())
        
        real_start = full_start + span_start_in_match
        real_end = real_start + len(span_text)
        return real_start, real_end, clean_full_text[real_start:real_end]
    
    return "", "", ""

def create_workbooks_if_not_exists(path):
    if not os.path.exists(path):
        wb = openpyxl.Workbook()
        wb.save(path)

# -------------------------------------------------------------------------
# 6. WORKBOOK GENERATION
# -------------------------------------------------------------------------
def generate_workbook():
    print(f"Generating Phase 0 Workbook for {NOTE_ID}...")
    
    # Load or Create Workbook
    if not os.path.exists(TEMPLATE_PATH):
        # Create a dummy workbook if template missing (for standalone testing)
        wb = openpyxl.Workbook()
        wb.create_sheet("Note_Text")
        wb.create_sheet("Note_Index")
        wb.create_sheet("Span_Annotations")
        wb.create_sheet("Span_Hydrated")
        wb.create_sheet("Event_Log")
        wb.create_sheet("V3_Procedure_Events")
        wb.create_sheet("V3_Registry_JSON")
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # ----------------------------------------
    # Sheet 1: Note_Text
    # ----------------------------------------
    ws_text = wb["Note_Text"] if "Note_Text" in wb.sheetnames else wb.create_sheet("Note_Text")
    # Clear existing
    for row in ws_text.iter_rows():
        for cell in row:
            cell.value = None
            
    # Headers
    ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # ----------------------------------------
    # Sheet 2: Note_Index
    # ----------------------------------------
    ws_index = wb["Note_Index"] if "Note_Index" in wb.sheetnames else wb.create_sheet("Note_Index")
    # Clear existing
    for row in ws_index.iter_rows():
        for cell in row:
            cell.value = None

    # Headers (Metadata + Flags)
    headers = ["note_id", "source_file", "procedure_date"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers)
    
    # Values
    row_values = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE] + list(PROCEDURE_FLAGS.values())
    ws_index.append(row_values)

    # ----------------------------------------
    # Sheet 3: Span_Annotations (Pre-hydration)
    # ----------------------------------------
    ws_anno = wb["Span_Annotations"] if "Span_Annotations" in wb.sheetnames else wb.create_sheet("Span_Annotations")
    # Clear
    for row in ws_anno.iter_rows():
        for cell in row:
            cell.value = None
            
    # Headers
    headers_anno = ["note_id", "span_text", "label", "normalized_value", "context_prefix", "start_char", "end_char", "event_id"]
    ws_anno.append(headers_anno)
    
    for item in SPANS:
        # (span_text, label, normalized_value, context_prefix, event_id)
        # Leave start/end blank
        ws_anno.append([NOTE_ID, item[0], item[1], item[2], item[3], "", "", item[4]])

    # ----------------------------------------
    # Sheet 4: Span_Hydrated (Calculated Offsets)
    # ----------------------------------------
    ws_hydra = wb["Span_Hydrated"] if "Span_Hydrated" in wb.sheetnames else wb.create_sheet("Span_Hydrated")
    # Clear
    for row in ws_hydra.iter_rows():
        for cell in row:
            cell.value = None
            
    ws_hydra.append(headers_anno) # Same headers
    
    for item in SPANS:
        span_text = item[0]
        label = item[1]
        norm_val = item[2]
        context = item[3]
        evt_id = item[4]
        
        start_idx, end_idx, found_text = hydrate_span(NOTE_TEXT, span_text, context)
        
        ws_hydra.append([NOTE_ID, found_text, label, norm_val, context, start_idx, end_idx, evt_id])

    # ----------------------------------------
    # Sheet 5: Event_Log (Flattened)
    # ----------------------------------------
    ws_log = wb["Event_Log"] if "Event_Log" in wb.sheetnames else wb.create_sheet("Event_Log")
    for row in ws_log.iter_rows():
        for cell in row:
            cell.value = None
    
    ws_log.append(["note_id", "event_id", "event_name", "concept_id", "event_type", "spans_list"])
    for evt in EVENTS:
        spans_str = "; ".join(evt["spans"])
        ws_log.append([NOTE_ID, evt["event_id"], evt["name"], evt["concept_id"], evt["type"], spans_str])

    # ----------------------------------------
    # Sheet 6: V3_Procedure_Events
    # ----------------------------------------
    ws_v3 = wb["V3_Procedure_Events"] if "V3_Procedure_Events" in wb.sheetnames else wb.create_sheet("V3_Procedure_Events")
    for row in ws_v3.iter_rows():
        for cell in row:
            cell.value = None
            
    v3_headers = ["note_id", "event_id", "method", "anatomy", "devices", "medications", "specimens", "measurements", "observations", "outcomes"]
    ws_v3.append(v3_headers)
    
    # Construct V3 rows by aggregating spans per event
    for evt in EVENTS:
        e_id = evt["event_id"]
        
        # Filter spans for this event
        evt_spans = [s for s in SPANS if s[4] == e_id]
        
        # Categorize
        method = [s[2] for s in evt_spans if s[1] in ["PROC_METHOD", "PROC_ACTION"]]
        anatomy = [s[2] for s in evt_spans if s[1] in ["ANAT_AIRWAY", "ANAT_LUNG_LOC", "ANAT_PLEURA", "LATERALITY"]]
        devices = [s[2] for s in evt_spans if s[1] in ["DEV_STENT", "DEV_CATHETER", "DEV_INSTRUMENT", "DEV_NEEDLE", "DEV_STENT_SIZE"]]
        meds = [] # Not captured in spans for this note explicitly as spans
        specimens = []
        measurements = [s[2] for s in evt_spans if s[1] in ["MEAS_SIZE", "MEAS_VOL", "MEAS_AIRWAY_DIAM"]]
        observations = [s[2] for s in evt_spans if s[1] in ["OBS_LESION", "OBS_ROSE"]]
        outcomes = [s[2] for s in evt_spans if s[1].startswith("OUTCOME")]
        
        ws_v3.append([
            NOTE_ID,
            e_id,
            json.dumps(method),
            json.dumps(anatomy),
            json.dumps(devices),
            json.dumps(meds),
            json.dumps(specimens),
            json.dumps(measurements),
            json.dumps(observations),
            json.dumps(outcomes)
        ])

    # ----------------------------------------
    # Sheet 7: V3_Registry_JSON
    # ----------------------------------------
    ws_json = wb["V3_Registry_JSON"] if "V3_Registry_JSON" in wb.sheetnames else wb.create_sheet("V3_Registry_JSON")
    for row in ws_json.iter_rows():
        for cell in row:
            cell.value = None
            
    ws_json.append(["note_id", "json_object"])
    
    # Construct Full JSON
    registry_obj = {
        "metadata": {
            "note_id": NOTE_ID,
            "source_file": SOURCE_FILE,
            "procedure_date": PROCEDURE_DATE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS,
        "spans": [
            {
                "text": s[0],
                "label": s[1],
                "normalized": s[2],
                "context": s[3],
                "event_id": s[4]
            } for s in SPANS
        ]
    }
    
    ws_json.append([NOTE_ID, json.dumps(registry_obj, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Successfully saved {OUTPUT_PATH}")

# -------------------------------------------------------------------------
# 7. EXECUTION
# -------------------------------------------------------------------------
if __name__ == "__main__":
    generate_workbook()