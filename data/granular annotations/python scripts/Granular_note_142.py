import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import re
import json
import os
import datetime

# 1. Imports & Constants
NOTE_ID = "note_142"
SOURCE_FILE = "note_142.txt"
PROCEDURE_DATE = ""
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"
NOTE_TEXT = """NOTE_ID:  note_142 SOURCE_FILE: note_142.txt Procedure Name: Bronchoscopy

Indications:

Hilar mass

Anesthesia:
General anesthesia with topical anesthesia using 2% lidocaine to the tracheobronchial tree (8 mL).
See anesthesia record for administered medications.

Pre-Procedure Assessment

The procedure, including risks, benefits, and alternatives, was explained to the patient.
All questions were answered, and informed consent was obtained and documented per institutional protocol.
A history and physical examination were performed and updated in the pre-procedure assessment record.
Relevant laboratory studies and radiographic imaging were reviewed. A procedural time-out was performed prior to the intervention.
Following administration of intravenous anesthetic medications and topical anesthesia to the upper airway and tracheobronchial tree, the Q180 slim video bronchoscope was introduced through the mouth via a laryngeal mask airway and advanced into the tracheobronchial tree.
The UC180F convex probe EBUS bronchoscope was subsequently introduced through the mouth via the laryngeal mask airway and advanced into the tracheobronchial tree.
The patient tolerated the procedure well.

Procedure Description and Findings

The laryngeal mask airway was in normal position.
The vocal cords moved normally with respiration. The subglottic space was normal.
The trachea was of normal caliber, and the carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level.
Bronchial mucosa and anatomy were normal, with no endobronchial lesions identified, except for extrinsic compression of the left upper lobe posterior segment.
No significant secretions were present.

Endobronchial Ultrasound Findings

The bronchoscope was withdrawn and replaced with the EBUS bronchoscope to perform ultrasound evaluation.
A systematic hilar and mediastinal lymph node survey was conducted, revealing visible lymph nodes.
Lymph node sizing and sampling were performed using endobronchial ultrasound with an Olympus EBUS-TBNA 22-gauge needle.
Specimens were sent for routine cytopathologic evaluation.

The left upper lobe mass measured 24.2 mm by EBUS and 24 mm by CT.
PET imaging was not available. On ultrasound, the mass appeared hypoechoic, heterogeneous, irregularly shaped, with sharp margins.
Multiple surrounding vessels were noted, limiting safe needle trajectory. A small window for safe needle access was identified, and the mass was biopsied with a single pass using a 22-gauge needle.
ROSE preliminary analysis demonstrated atypical cells. However, brisk bleeding occurred following the first needle pass, and no additional ultrasound-guided passes were performed.
An esophageal approach was attempted but did not allow adequate visualization of the mass.
The EBUS bronchoscope was withdrawn, and a therapeutic bronchoscope was introduced.
A conventional TBNA was attempted in the region of extrinsic compression in the left upper lobe posterior segment;
however, the return was bloody, and no further passes were performed.

All samples were sent to cytopathology for review.
Complications

No immediate complications.

Estimated Blood Loss

Less than 5 mL.

Impression

Normal bronchoscopic examination except for extrinsic compression of the left upper lobe posterior segment

Lymph node sizing and limited sampling performed via EBUS-TBNA

Post-Procedure Diagnosis

Technically successful flexible bronchoscopy with endobronchial ultrasound-guided biopsy

The patient remained stable throughout the procedure and was transferred in good condition to the post-bronchoscopy recovery area for observation until discharge criteria were met.
Preliminary findings were discussed with the patient. Follow-up with the requesting service for final cytology results was recommended."""

# 2. Configuration (Procedure Flags)
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 1,
    "linear_ebus": 1,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 3. Data Definition (Spans)
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("Q180 slim video bronchoscope", "DEV_INSTRUMENT", "Q180 slim video bronchoscope", "tracheobronchial tree, the ", "evt_01"),
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "UC180F convex probe EBUS bronchoscope", "The ", "evt_02"),
    ("extrinsic compression", "OBS_LESION", "extrinsic compression", "except for ", "evt_01"),
    ("left upper lobe posterior segment", "ANAT_LUNG_LOC", "Left Upper Lobe Posterior Segment", "compression of the ", "evt_01"),
    ("endobronchial ultrasound", "PROC_METHOD", "Linear EBUS", "performed using ", "evt_02"),
    ("Olympus EBUS-TBNA 22-gauge needle", "DEV_NEEDLE", "Olympus EBUS-TBNA 22G", "with an ", "evt_02"),
    ("left upper lobe mass", "OBS_LESION", "Left Upper Lobe Mass", "The ", "evt_03"),
    ("24.2 mm", "MEAS_SIZE", "24.2 mm", "measured ", "evt_03"),
    ("hypoechoic", "OBS_LESION", "hypoechoic", "appeared ", "evt_03"),
    ("heterogeneous", "OBS_LESION", "heterogeneous", "hypoechoic, ", "evt_03"),
    ("irregularly shaped", "OBS_LESION", "irregularly shaped", "heterogeneous, ", "evt_03"),
    ("sharp margins", "OBS_LESION", "sharp margins", "with ", "evt_03"),
    ("biopsied", "PROC_ACTION", "biopsy", "mass was ", "evt_03"),
    ("single pass", "PROC_ACTION", "single pass", "with a ", "evt_03"),
    ("22-gauge needle", "DEV_NEEDLE", "22G", "using a ", "evt_03"),
    ("atypical cells", "OBS_ROSE", "atypical cells", "demonstrated ", "evt_03"),
    ("brisk bleeding", "OUTCOME_COMPLICATION", "bleeding", "However, ", "evt_03"),
    ("therapeutic bronchoscope", "DEV_INSTRUMENT", "therapeutic bronchoscope", "and a ", "evt_04"),
    ("conventional TBNA", "PROC_METHOD", "TBNA", "A ", "evt_04"),
    ("left upper lobe posterior segment", "ANAT_LUNG_LOC", "Left Upper Lobe Posterior Segment", "compression in the ", "evt_04"),
    ("bloody", "OBS_ROSE", "bloody return", "return was ", "evt_04"),
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications\n\n", "evt_global"),
]

# 4. Event Definitions
EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Diagnostic Bronchoscopy",
        "description": "General inspection with Q180 slim scope.",
        "findings": ["Extrinsic compression of LUL posterior segment"],
        "devices": ["Q180 slim video bronchoscope"]
    },
    {
        "event_id": "evt_02",
        "procedure_type": "Linear EBUS",
        "description": "Lymph node survey and sampling.",
        "findings": ["Visible lymph nodes"],
        "devices": ["UC180F convex probe EBUS bronchoscope", "Olympus EBUS-TBNA 22-gauge needle"]
    },
    {
        "event_id": "evt_03",
        "procedure_type": "Linear EBUS",
        "description": "Biopsy of LUL mass.",
        "findings": ["LUL Mass 24.2mm", "Hypoechoic", "Heterogeneous", "Irregular", "Sharp margins", "ROSE: Atypical cells"],
        "devices": ["UC180F convex probe EBUS bronchoscope", "22-gauge needle"],
        "complications": ["Brisk bleeding"]
    },
    {
        "event_id": "evt_04",
        "procedure_type": "Conventional TBNA",
        "description": "Attempted TBNA of extrinsic compression site.",
        "findings": ["Bloody return"],
        "devices": ["Therapeutic bronchoscope", "TBNA needle (implied)"]
    }
]

# 5. Helper Functions
def clean_text(text):
    return text.strip().replace('\r', '')

def hydrate_span(text, span_text, context_prefix):
    cleaned_text = clean_text(text)
    cleaned_span = clean_text(span_text)
    cleaned_context = clean_text(context_prefix)
    
    if not cleaned_span:
        return 0, 0
        
    start_search = 0
    if cleaned_context:
        context_idx = cleaned_text.find(cleaned_context)
        if context_idx != -1:
            start_search = context_idx + len(cleaned_context)
            
    span_start = cleaned_text.find(cleaned_span, start_search)
    if span_start == -1:
        # Fallback without context
        span_start = cleaned_text.find(cleaned_span)
        
    if span_start == -1:
        return 0, 0
        
    return span_start, span_start + len(cleaned_span)

# 6. Workbook Generation Function
def generate_workbook():
    wb = openpyxl.Workbook()
    
    # --- Sheet 1: Note_Text ---
    ws1 = wb.active
    ws1.title = "Note_Text"
    ws1.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws1.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # --- Sheet 2: Note_Index ---
    ws2 = wb.create_sheet("Note_Index")
    headers = ["NOTE_ID", "PROCEDURE_DATE"] + list(PROCEDURE_FLAGS.keys())
    ws2.append(headers)
    row_data = [NOTE_ID, PROCEDURE_DATE] + [PROCEDURE_FLAGS[k] for k in PROCEDURE_FLAGS]
    ws2.append(row_data)
    
    # --- Sheet 3: Span_Annotations ---
    ws3 = wb.create_sheet("Span_Annotations")
    ws3.append(["NOTE_ID", "START_CHAR", "END_CHAR", "LABEL", "TEXT", "CONTEXT_PREFIX", "NORMALIZED_VALUE", "EVENT_ID"])
    for span in SPANS:
        ws3.append([NOTE_ID, "", "", span[1], span[0], span[3], span[2], span[4]])
        
    # --- Sheet 4: Span_Hydrated ---
    ws4 = wb.create_sheet("Span_Hydrated")
    ws4.append(["NOTE_ID", "START_CHAR", "END_CHAR", "LABEL", "TEXT", "CONTEXT_PREFIX", "NORMALIZED_VALUE", "EVENT_ID"])
    for span in SPANS:
        s, e = hydrate_span(NOTE_TEXT, span[0], span[3])
        ws4.append([NOTE_ID, s, e, span[1], span[0], span[3], span[2], span[4]])
        
    # --- Sheet 5: Event_Log ---
    ws5 = wb.create_sheet("Event_Log")
    ws5.append(["NOTE_ID", "EVENT_ID", "PROCEDURE_TYPE", "DESCRIPTION", "FINDINGS", "DEVICES", "COMPLICATIONS"])
    for evt in EVENTS:
        ws5.append([
            NOTE_ID,
            evt["event_id"],
            evt["procedure_type"],
            evt["description"],
            "; ".join(evt.get("findings", [])),
            "; ".join(evt.get("devices", [])),
            "; ".join(evt.get("complications", []))
        ])
        
    # --- Sheet 6: V3_Procedure_Events ---
    ws6 = wb.create_sheet("V3_Procedure_Events")
    ws6.append(["NOTE_ID", "EVENT_ID", "EVENT_TYPE", "EVENT_DETAILS_JSON"])
    for evt in EVENTS:
        ws6.append([NOTE_ID, evt["event_id"], evt["procedure_type"], json.dumps(evt)])
        
    # --- Sheet 7: V3_Registry_JSON ---
    ws7 = wb.create_sheet("V3_Registry_JSON")
    ws7.append(["NOTE_ID", "FULL_JSON"])
    registry_data = {
        "metadata": {"note_id": NOTE_ID, "source": SOURCE_FILE, "date": PROCEDURE_DATE},
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS,
        "spans": [{"text": s[0], "label": s[1], "normalized": s[2], "event_id": s[4]} for s in SPANS]
    }
    ws7.append([NOTE_ID, json.dumps(registry_data, indent=2)])
    
    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()