import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# 1. Imports & Constants
NOTE_ID = "note_143"
SOURCE_FILE = "note_143.txt"
PROCEDURE_DATE = ""  # No specific date in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_143 SOURCE_FILE: note_143.txt Procedure Name: Pleuroscopy
Indication: Pleural effusion
Anesthesia: Monitored Anesthesia Care

Pre-Anesthesia Assessment

ASA Physical Status: III – Patient with severe systemic disease

The procedure, including risks, benefits, and alternatives, was explained to the patient.
All questions were answered, and informed consent was obtained and documented per institutional protocol.
A focused history and physical examination were performed and updated in the pre-procedure assessment record.
Relevant laboratory studies and imaging were reviewed. A procedural time-out was performed prior to initiation.
Procedure Description

The patient was placed on the operating table in the lateral decubitus position with appropriate padding of all pressure points.
The procedural site was identified using ultrasound guidance and was sterilely prepped with chlorhexidine gluconate (Chloraprep) and draped in the usual fashion.
A 10.0-mm integrated pleuroscope was introduced through the incision and advanced into the pleural space.
Local Anesthesia

The pleural entry site was infiltrated with 30 mL of 1% lidocaine.
A 10-mm reusable primary port was placed on the right side at the 5th intercostal space along the mid-axillary line using a Veress needle technique.
Pleuroscopy Findings

Nodularity of the lung parenchyma, predominantly involving the lower lobe.
The pleura appeared grossly normal but demonstrated excessive adipose tissue and diffuse white fibrinous strands enveloping the pleura and lung.
Small areas of pleural studding were noted; differentiation between fat versus tumor studding was uncertain.
A mass was visualized in the posterior pleural space near the spine;
it appeared highly vascular and was not biopsied due to bleeding risk.
Approximately 2,000 mL of milky-appearing pleural fluid was removed and sent for analysis.
Biopsy

Biopsies of a pleural mucosal abnormality over the diaphragm were obtained using forceps and sent for histopathologic evaluation.
Eleven samples were obtained.

A 15.5-Fr PleurX catheter was placed in the pleural space over the diaphragm.
Dressing

Port sites were dressed with a transparent dressing.

Estimated Blood Loss

Minimal

Complications

None immediate

Post-Procedure Diagnosis

Pleural adhesions

Post-Procedure Plan

The patient will be observed post-procedure until all discharge criteria are met.
Chest X-ray to be obtained post-procedure."""

# 2. Configuration (Procedure Flags)
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 0,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 1,
    "medical_thoracoscopy": 1,
    "pleurodesis": 0,
    "pleural_biopsy": 1,
    "fibrinolytic_therapy": 0
}

# 3. Data Definition (Spans)
SPANS = [
    # Event 1: Pleuroscopy & Access
    ("Pleuroscopy", "PROC_METHOD", "Medical Thoracoscopy", "Procedure Name: ", "evt_01"),
    ("10.0-mm integrated pleuroscope", "DEV_INSTRUMENT", "Pleuroscope", "A ", "evt_01"),
    ("right side", "LATERALITY", "Right", "placed on the ", "evt_01"),
    ("5th intercostal space", "ANAT_PLEURA", "5th Intercostal Space", "at the ", "evt_01"),
    ("mid-axillary line", "ANAT_PLEURA", "Mid-Axillary Line", "along the ", "evt_01"),
    ("2,000 mL", "MEAS_PLEURAL_DRAIN", "2000", "Approximately ", "evt_01"),
    
    # Event 2: Biopsy
    ("Biopsies", "PROC_ACTION", "Biopsy", "Biopsy\n\n", "evt_02"),
    ("pleural mucosal abnormality", "OBS_LESION", "Pleural abnormality", "of a ", "evt_02"),
    ("diaphragm", "ANAT_PLEURA", "Diaphragm", "over the ", "evt_02"),
    ("forceps", "DEV_INSTRUMENT", "Forceps", "using ", "evt_02"),
    ("Eleven samples", "MEAS_COUNT", "11", "histopathologic evaluation.\n", "evt_02"),

    # Event 3: IPC Placement
    ("15.5-Fr", "DEV_CATHETER_SIZE", "15.5 French", "A ", "evt_03"),
    ("PleurX catheter", "DEV_CATHETER", "Indwelling Pleural Catheter", "Fr ", "evt_03"),
    ("placed", "PROC_ACTION", "Placement", "catheter was ", "evt_03"),
    ("pleural space", "ANAT_PLEURA", "Pleural Space", "in the ", "evt_03"),
    ("diaphragm", "ANAT_PLEURA", "Diaphragm", "over the ", "evt_03"),

    # Outcome
    ("None immediate", "OUTCOME_COMPLICATION", "None", "Complications\n\n", "evt_04")
]

# 4. Event Definitions
EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Medical Thoracoscopy",
        "action": "Inspection/Drainage",
        "laterality": "Right",
        "anatomy": ["5th Intercostal Space", "Mid-Axillary Line"],
        "devices": ["Pleuroscope"],
        "samples": [],
        "measurements": [{"type": "Volume", "value": "2000", "unit": "mL"}],
        "notes": "Pleuroscopy with fluid drainage"
    },
    {
        "event_id": "evt_02",
        "procedure_type": "Medical Thoracoscopy",
        "action": "Biopsy",
        "laterality": "Right",
        "anatomy": ["Diaphragm"],
        "devices": ["Forceps"],
        "samples": ["Pleural biopsy (11 samples)"],
        "measurements": [],
        "notes": "Biopsy of pleural mucosal abnormality"
    },
    {
        "event_id": "evt_03",
        "procedure_type": "IPC Placement",
        "action": "Placement",
        "laterality": "Right",
        "anatomy": ["Pleural Space", "Diaphragm"],
        "devices": ["15.5-Fr PleurX catheter"],
        "samples": [],
        "measurements": [],
        "notes": "PleurX catheter placement"
    },
    {
        "event_id": "evt_04",
        "procedure_type": "General",
        "action": "Outcome",
        "laterality": "",
        "anatomy": [],
        "devices": [],
        "samples": [],
        "measurements": [],
        "notes": "No immediate complications"
    }
]

# 5. Helper Functions
def clean_text(text):
    if not text:
        return ""
    return re.sub(r'[\r\n]+', ' ', text).strip()

def hydrate_span(full_text, span_text, context_prefix):
    """
    Finds the start and end indices of span_text within full_text.
    Uses context_prefix to disambiguate if needed.
    """
    # Normalize texts for searching (keep original for offsets)
    text_norm = full_text.replace('\r', '').replace('\n', ' ')
    span_norm = span_text.replace('\r', '').replace('\n', ' ')
    context_norm = context_prefix.replace('\r', '').replace('\n', ' ')
    
    # Try finding with context first
    search_str = context_norm + span_norm
    start_idx = text_norm.find(search_str)
    
    if start_idx != -1:
        # Found with context, adjust start to point to span
        real_start = start_idx + len(context_norm)
        return real_start, real_start + len(span_norm)
    
    # Fallback: search just the span
    start_idx = text_norm.find(span_norm)
    if start_idx != -1:
        return start_idx, start_idx + len(span_norm)
    
    return -1, -1

# 6. Workbook Generation Function
def generate_workbook():
    # Create new workbook
    wb = openpyxl.Workbook()
    
    # --- Sheet 1: Note_Text ---
    ws1 = wb.active
    ws1.title = "Note_Text"
    ws1.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws1.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # --- Sheet 2: Note_Index ---
    ws2 = wb.create_sheet("Note_Index")
    headers = ["NOTE_ID", "SOURCE_FILE", "PROCEDURE_DATE"] + list(PROCEDURE_FLAGS.keys())
    ws2.append(headers)
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE] + list(PROCEDURE_FLAGS.values())
    ws2.append(row_data)
    
    # --- Sheet 3: Span_Annotations ---
    ws3 = wb.create_sheet("Span_Annotations")
    ws3.append(["NOTE_ID", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "START_CHAR", "END_CHAR", "EVENT_ID"])
    for span in SPANS:
        # Initial export leaves start/end blank for manual review if needed, but we calculate in Sheet 4
        ws3.append([NOTE_ID, span[0], span[1], span[2], span[3], "", "", span[4]])

    # --- Sheet 4: Span_Hydrated ---
    ws4 = wb.create_sheet("Span_Hydrated")
    ws4.append(["NOTE_ID", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "START_CHAR", "END_CHAR", "EVENT_ID"])
    for span in SPANS:
        start, end = hydrate_span(NOTE_TEXT, span[0], span[3])
        ws4.append([NOTE_ID, span[0], span[1], span[2], span[3], start, end, span[4]])
    
    # --- Sheet 5: Event_Log ---
    ws5 = wb.create_sheet("Event_Log")
    ws5.append(["NOTE_ID", "EVENT_ID", "PROCEDURE_TYPE", "ACTION", "LATERALITY", "ANATOMY", "DEVICES", "SAMPLES", "MEASUREMENTS", "NOTES"])
    for evt in EVENTS:
        ws5.append([
            NOTE_ID,
            evt["event_id"],
            evt["procedure_type"],
            evt["action"],
            evt["laterality"],
            ", ".join(evt["anatomy"]),
            ", ".join(evt["devices"]),
            ", ".join(evt["samples"]),
            json.dumps(evt["measurements"]),
            evt["notes"]
        ])

    # --- Sheet 6: V3_Procedure_Events ---
    ws6 = wb.create_sheet("V3_Procedure_Events")
    ws6.append(["note_id", "event_id", "procedure_type", "action", "laterality", "anatomy", "devices", "samples", "measurements", "notes"])
    for evt in EVENTS:
        ws6.append([
            NOTE_ID,
            evt["event_id"],
            evt["procedure_type"],
            evt["action"],
            evt["laterality"],
            json.dumps(evt["anatomy"]),
            json.dumps(evt["devices"]),
            json.dumps(evt["samples"]),
            json.dumps(evt["measurements"]),
            evt["notes"]
        ])

    # --- Sheet 7: V3_Registry_JSON ---
    ws7 = wb.create_sheet("V3_Registry_JSON")
    ws7.append(["note_id", "full_json"])
    
    registry_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source_file": SOURCE_FILE,
            "procedure_date": PROCEDURE_DATE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS
    }
    ws7.append([NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

# 7. Execution Block
if __name__ == "__main__":
    generate_workbook()