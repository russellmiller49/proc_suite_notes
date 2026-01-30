import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# 1. Imports & Constants
NOTE_ID = "note_144"
SOURCE_FILE = "note_144.txt"
PROCEDURE_DATE = ""
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_144 SOURCE_FILE: note_144.txt Procedure Name: Bronchoscopy
Indications: Metastatic lung cancer; hemoptysis

Pre-Anesthesia Assessment:

A history and physical examination were performed.
Patient medications and allergies were reviewed. The risks and benefits of the procedure, as well as sedation options and associated risks, were discussed with the patient’s parent.
All questions were answered, and informed consent was obtained. Patient identification and the proposed procedure were verified prior to the procedure by the physician, nurse, and technician in the procedure room.
Mental status examination revealed the patient to be alert and oriented. Airway examination demonstrated a normal oropharyngeal airway.
Respiratory examination was clear to auscultation. Cardiovascular examination revealed regular rate and rhythm without murmurs, S3, or S4.
The ASA physical status classification was III (a patient with severe systemic disease).
After reviewing risks and benefits, the patient was deemed an appropriate candidate to undergo the procedure under general anesthesia.
Immediately prior to administration of medications, the patient was reassessed and found to be appropriate for sedation.
Heart rate, respiratory rate, oxygen saturation, blood pressure, adequacy of ventilation, and response to care were continuously monitored throughout the procedure.
The patient’s physical status was reassessed following completion of the procedure.
Procedure, risks, benefits, and alternatives were again explained to the patient, and informed consent was documented per institutional protocol.
A history and physical examination were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention.

Following administration of intravenous medications per the anesthesia record and topical anesthesia to the upper airway and tracheobronchial tree, the Q180 slim video bronchoscope was introduced through the rigid bronchoscope after removal of the telescope and advanced into the tracheobronchial tree bilaterally.
The T180 therapeutic video bronchoscope was subsequently introduced and advanced into the airways. The procedure was accomplished without difficulty.
Procedure Description:

Left Lung Abnormalities:
A partially obstructing airway abnormality (approximately 40% luminal obstruction) was identified at the entrance to the left lower lobe.
Tumor debulking was performed using an electrocautery snare with successful coagulation.
Upon completion, the left lower lobe entrance was fully patent.
Findings:

Right Lung Abnormalities:
A nearly obstructing airway abnormality (>90% luminal obstruction) was identified at the entrance to the right lower lobe.
Tumor destruction was performed using argon plasma coagulation at 25 watts and 0.3 L/min, in combination with an electrocautery probe and electrocautery snare.
This resulted in successful recanalization.

Upon completion of intervention, both the left lower lobe and right lower lobe entrances were 100% patent.
Residual tumor remained within the anterior and lateral basal segments of the right lower lobe;
this tissue appeared chronic and extended into the peripheral airways and was not amenable to further intervention.
The right lower lobe superior segment, medial basal segment, and posterior segment were fully patent.
Therapeutic aspiration was performed at the conclusion of the procedure to remove retained blood and secretions, with good hemostasis achieved.
Complications: No immediate complications
Estimated Blood Loss: Minimal

Additional Findings:

Metastatic renal cell carcinoma with endobronchial tumors involving the left lower lobe and right lower lobe entrances

Hemoptysis

Partially obstructing (approximately 40%) airway abnormality in the left lower lobe

Nearly obstructing (>90%) airway abnormality in the right lower lobe

No specimens collected

Post-Procedure Diagnosis:

Nearly obstructing airway abnormality of the right lower lobe, status post successful debulking

No specimens collected

Plan:

Follow up with the bronchoscopist in one week

Follow up with the referring physician as previously scheduled"""

# 2. Configuration (Procedure Flags)
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 1, # Electrocautery and APC
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 1,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 3. Data Definition (Spans)
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    # Event 1: Rigid and Scopes
    ("rigid bronchoscope", "DEV_INSTRUMENT", "Rigid Bronchoscope", "introduced through the", "evt_01"),
    ("Q180 slim video bronchoscope", "DEV_INSTRUMENT", "Olympus Q180", "tracheobronchial tree, the", "evt_01"),
    ("T180 therapeutic video bronchoscope", "DEV_INSTRUMENT", "Olympus T180", "The", "evt_01"),

    # Event 2: LLL Debulking
    ("left lower lobe", "ANAT_LUNG_LOC", "LLL", "entrance to the", "evt_02"),
    ("40%", "OUTCOME_AIRWAY_LUMEN_PRE", "40", "approximately", "evt_02"),
    ("Tumor debulking", "PROC_METHOD", "Thermal Ablation", "Left Lung Abnormalities:\n", "evt_02"),
    ("electrocautery snare", "DEV_INSTRUMENT", "Electrocautery Snare", "using an", "evt_02"),
    ("fully patent", "OUTCOME_AIRWAY_LUMEN_POST", "100", "entrance was", "evt_02"),

    # Event 3: RLL Destruction
    ("right lower lobe", "ANAT_LUNG_LOC", "RLL", "entrance to the", "evt_03"),
    (">90%", "OUTCOME_AIRWAY_LUMEN_PRE", "90", "abnormality (", "evt_03"),
    ("Tumor destruction", "PROC_METHOD", "Thermal Ablation", "Right Lung Abnormalities:\n", "evt_03"),
    ("argon plasma coagulation", "PROC_METHOD", "APC", "using", "evt_03"),
    ("electrocautery probe", "DEV_INSTRUMENT", "Electrocautery Probe", "combination with an", "evt_03"),
    ("electrocautery snare", "DEV_INSTRUMENT", "Electrocautery Snare", "and", "evt_03"),
    ("100% patent", "OUTCOME_AIRWAY_LUMEN_POST", "100", "entrances were", "evt_03"),

    # Event 4: Aspiration
    ("Therapeutic aspiration", "PROC_METHOD", "Therapeutic Aspiration", "fully patent.\n", "evt_04"),
    ("retained blood and secretions", "OBS_LESION", "Blood/Secretions", "remove", "evt_04"),

    # Complications
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications:", "evt_05"),
]

# 4. Event Definitions
EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Bronchoscopy",
        "action": "Inspection/Access",
        "anatomy": "Tracheobronchial Tree",
        "devices": ["Rigid Bronchoscope", "Q180 slim video bronchoscope", "T180 therapeutic video bronchoscope"],
        "notes": "General anesthesia, rigid access."
    },
    {
        "event_id": "evt_02",
        "procedure_type": "Bronchoscopy",
        "action": "Tumor Debulking",
        "anatomy": "Left Lower Lobe",
        "devices": ["Electrocautery Snare"],
        "outcome": "40% obstruction to Fully Patent"
    },
    {
        "event_id": "evt_03",
        "procedure_type": "Bronchoscopy",
        "action": "Tumor Destruction (APC/Cautery)",
        "anatomy": "Right Lower Lobe",
        "devices": ["APC", "Electrocautery Probe", "Electrocautery Snare"],
        "outcome": ">90% obstruction to 100% Patent"
    },
    {
        "event_id": "evt_04",
        "procedure_type": "Bronchoscopy",
        "action": "Therapeutic Aspiration",
        "anatomy": "Airways",
        "devices": ["Suction"],
        "outcome": "Good hemostasis, secretions removed"
    },
    {
        "event_id": "evt_05",
        "procedure_type": "Bronchoscopy",
        "action": "Outcome Assessment",
        "anatomy": "N/A",
        "devices": [],
        "outcome": "No immediate complications"
    }
]

# 5. Helper Functions
def clean_text(text):
    return text.replace('\r', '').strip()

def find_offsets(text, span_text, context_prefix):
    # Normalize texts for searching
    text_norm = text.replace('\r', '')
    span_norm = span_text.replace('\r', '')
    context_norm = context_prefix.replace('\r', '')

    # Attempt 1: Exact search with context
    search_str = context_norm + span_norm
    start = text_norm.find(search_str)
    
    if start != -1:
        # Start index is after the context
        actual_start = start + len(context_norm)
        actual_end = actual_start + len(span_norm)
        return actual_start, actual_end

    # Attempt 2: Search for span only (fallback)
    start = text_norm.find(span_norm)
    if start != -1:
        return start, start + len(span_norm)

    return -1, -1

def hydrate_span(full_text, span_data):
    span_text, label, norm_val, context, evt_id = span_data
    start_char, end_char = find_offsets(full_text, span_text, context)
    return {
        "event_id": evt_id,
        "label": label,
        "text": span_text,
        "start_char": start_char,
        "end_char": end_char,
        "normalized_value": norm_val,
        "context_prefix": context
    }

# 6. Workbook Generation
def generate_workbook():
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # --- Sheet 1: Note_Text ---
    ws_text = wb.create_sheet("Note_Text")
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # --- Sheet 2: Note_Index ---
    ws_index = wb.create_sheet("Note_Index")
    headers = ["NOTE_ID", "SOURCE_FILE", "PROCEDURE_DATE"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers)
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE] + [PROCEDURE_FLAGS[k] for k in PROCEDURE_FLAGS]
    ws_index.append(row_data)

    # --- Sheet 3: Span_Annotations (Blank offsets) ---
    ws_spans_raw = wb.create_sheet("Span_Annotations")
    ws_spans_raw.append(["NOTE_ID", "EVENT_ID", "LABEL", "TEXT", "START_CHAR", "END_CHAR", "NORMALIZED_VALUE", "CONTEXT_PREFIX"])
    for span in SPANS:
        # span = (text, label, norm, context, evt_id)
        ws_spans_raw.append([NOTE_ID, span[4], span[1], span[0], "", "", span[2], span[3]])

    # --- Sheet 4: Span_Hydrated (Calculated offsets) ---
    ws_spans_calc = wb.create_sheet("Span_Hydrated")
    ws_spans_calc.append(["NOTE_ID", "EVENT_ID", "LABEL", "TEXT", "START_CHAR", "END_CHAR", "NORMALIZED_VALUE", "CONTEXT_PREFIX"])
    
    for span in SPANS:
        hydrated = hydrate_span(NOTE_TEXT, span)
        ws_spans_calc.append([
            NOTE_ID, 
            hydrated["event_id"], 
            hydrated["label"], 
            hydrated["text"], 
            hydrated["start_char"], 
            hydrated["end_char"], 
            hydrated["normalized_value"], 
            hydrated["context_prefix"]
        ])

    # --- Sheet 5: Event_Log ---
    ws_events = wb.create_sheet("Event_Log")
    ws_events.append(["NOTE_ID", "EVENT_ID", "PROCEDURE_TYPE", "ACTION", "ANATOMY", "DEVICES", "OUTCOME", "NOTES"])
    for evt in EVENTS:
        ws_events.append([
            NOTE_ID,
            evt["event_id"],
            evt.get("procedure_type", ""),
            evt.get("action", ""),
            evt.get("anatomy", ""),
            ", ".join(evt.get("devices", [])),
            evt.get("outcome", ""),
            evt.get("notes", "")
        ])

    # --- Sheet 6: V3_Procedure_Events ---
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    ws_v3.append(["NOTE_ID", "EVENT_ID", "EVENT_JSON"])
    for evt in EVENTS:
        ws_v3.append([NOTE_ID, evt["event_id"], json.dumps(evt)])

    # --- Sheet 7: V3_Registry_JSON ---
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["NOTE_ID", "FULL_JSON"])
    
    registry_data = {
        "note_id": NOTE_ID,
        "source_file": SOURCE_FILE,
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS,
        "spans": [hydrate_span(NOTE_TEXT, s) for s in SPANS]
    }
    
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Workbook generated at: {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()