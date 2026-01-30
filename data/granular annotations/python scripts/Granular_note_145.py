import openpyxl
from openpyxl import Workbook
import re
import json
import os
import datetime

# ==========================================
# 1. Imports & Constants
# ==========================================

NOTE_ID = "note_145"
SOURCE_FILE = "note_145.txt"
PROCEDURE_DATE = "" 
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_145 SOURCE_FILE: note_145.txt Procedure Name: Pleuroscopy
Indications: Pleural effusion
Medications: Monitored anesthesia care

Pre-Anesthesia Assessment

ASA Physical Status: III – A patient with severe systemic disease

Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered, and informed consent was documented per institutional protocol.
A history and physical examination were performed and updated in the preprocedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention.

The patient was sterilely prepped, and pleuroscopy was performed.
Procedure Description
Positioning and Preparation

The patient was placed on the operating table in the lateral decubitus position, and all pressure points were adequately padded.
The skin was prepped with chlorhexidine gluconate (Chloraprep) and draped in the usual sterile fashion.
Local Anesthesia

The pleural entry site was identified using ultrasound guidance.
The entry site was infiltrated with 30 mL of 1% lidocaine.
Incision and Port Placement

A 10-mm reusable primary port was placed on the left side at the seventh intercostal space along the anterior axillary line using a Veress needle technique.
Pleuroscopy

A 0-degree 7.0-mm pleuroscopy telescope was introduced through the incision and advanced into the pleural space.
A 0-degree 4.0-mm pleuroscopy telescope was subsequently introduced through the same incision for further inspection.
The pleura was systematically inspected via the primary port.

Findings

There was extensive pleural studding throughout the pleural space.
No purulent collections were identified. Thick pleural adhesions were present throughout.
Biopsy and Interventions

Biopsies of pleural adhesions were obtained from the upper pleura using forceps and sent for histopathologic examination.
Five biopsy specimens were obtained. An additional five specimens were sent for microbiologic cultures.
A 15.5-French PleurX catheter was placed in the pleural space over the diaphragm.
Dressing

Port sites were dressed with a transparent dressing.

Complications

No immediate complications.

Estimated Blood Loss

Minimal.
Post-Procedure Diagnosis

Pleural metastases

Post-Procedure Care

The patient will be observed post-procedure until all discharge criteria are met.
Chest radiograph to be obtained post-procedure."""

# ==========================================
# 2. Configuration (Procedure Flags)
# ==========================================

PROCEDURE_FLAGS = {
    # Bronchoscopy Flags (All 0 for this note)
    "diagnostic_bronchoscopy": 0,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural Flags
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 1,                     # PleurX catheter
    "medical_thoracoscopy": 1,    # Explicit Procedure Name: Pleuroscopy
    "pleurodesis": 0,
    "pleural_biopsy": 1,          # Biopsies obtained
    "fibrinolytic_therapy": 0
}

# ==========================================
# 3. Data Definition (Spans)
# ==========================================

# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    # Metadata / Prep
    ("Pleuroscopy", "PROC_METHOD", "Medical Thoracoscopy", "Procedure Name: ", "evt_01"),
    ("ultrasound guidance", "PROC_METHOD", "Ultrasound Guidance", "identified using ", "evt_01"),
    ("left side", "LATERALITY", "Left", "placed on the ", "evt_01"),
    
    # Event 1: Pleuroscopy & Inspection
    ("0-degree 7.0-mm pleuroscopy telescope", "DEV_INSTRUMENT", "Pleuroscope", "A ", "evt_01"),
    ("0-degree 4.0-mm pleuroscopy telescope", "DEV_INSTRUMENT", "Pleuroscope", "source: 11] A ", "evt_01"),
    ("pleura", "ANAT_PLEURA", "Pleura", "The ", "evt_01"),
    ("extensive pleural studding", "OBS_LESION", "Tumor/Nodule", "There was ", "evt_01"),
    ("pleural space", "ANAT_PLEURA", "Pleural Space", "throughout the ", "evt_01"),
    ("pleural adhesions", "OBS_LESION", "Adhesions", "Thick ", "evt_01"),
    
    # Event 2: Biopsy
    ("Biopsies", "PROC_METHOD", "Pleural Biopsy", "Biopsy and Interventions\n\n", "evt_02"),
    ("pleural adhesions", "ANAT_PLEURA", "Pleural Adhesions", "Biopsies of ", "evt_02"),
    ("upper pleura", "ANAT_PLEURA", "Upper Pleura", "from the ", "evt_02"),
    ("forceps", "DEV_INSTRUMENT", "Forceps", "using ", "evt_02"),
    ("Five", "MEAS_COUNT", "5", "specimens were obtained.\n", "evt_02"),
    
    # Event 3: IPC Placement
    ("15.5-French", "DEV_CATHETER_SIZE", "15.5Fr", "A ", "evt_03"),
    ("PleurX catheter", "DEV_CATHETER", "Indwelling Pleural Catheter", "French ", "evt_03"),
    ("pleural space", "ANAT_PLEURA", "Pleural Space", "placed in the ", "evt_03"),
    
    # Outcomes / Diagnosis
    ("Pleural metastases", "OBS_LESION", "Metastasis", "Post-Procedure Diagnosis\n\n", "evt_04"),
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications\n\n", "evt_05")
]

# ==========================================
# 4. Event Definitions
# ==========================================

EVENTS = [
    {
        "event_id": "evt_01",
        "method": "Medical Thoracoscopy",
        "anatomy": "Pleura",
        "devices": ["Pleuroscope", "Ultrasound"],
        "outcomes": []
    },
    {
        "event_id": "evt_02",
        "method": "Pleural Biopsy",
        "anatomy": "Pleural Adhesions",
        "devices": ["Forceps"],
        "outcomes": []
    },
    {
        "event_id": "evt_03",
        "method": "Indwelling Pleural Catheter Placement",
        "anatomy": "Pleural Space",
        "devices": ["PleurX Catheter"],
        "outcomes": []
    },
    {
        "event_id": "evt_04",
        "method": "Diagnosis",
        "anatomy": "Pleura",
        "devices": [],
        "outcomes": ["Pleural metastases"]
    },
    {
        "event_id": "evt_05",
        "method": "Outcome",
        "anatomy": "",
        "devices": [],
        "outcomes": ["No immediate complications"]
    }
]

# ==========================================
# 5. Helper Functions
# ==========================================

def clean_text(text):
    if not text: return ""
    return text.strip().replace('\r', '')

def hydrate_span(full_text, span_text, context_prefix):
    """
    Finds start/end offsets of span_text within full_text, 
    using context_prefix to disambiguate.
    """
    clean_full = clean_text(full_text)
    clean_span = clean_text(span_text)
    clean_context = clean_text(context_prefix)
    
    if not clean_span: return None, None, None

    # Search pattern: Context + ... + Span
    # This is a basic implementation; robust regex might be needed for variations
    search_str = clean_context + clean_span if clean_context else clean_span
    
    # Try to find the exact sequence first
    start_index = clean_full.find(search_str)
    
    if start_index == -1 and clean_context:
        # Fallback: Find context, then find span after context
        context_idx = clean_full.find(clean_context)
        if context_idx != -1:
            # Look for span after the context
            sub_text = clean_full[context_idx + len(clean_context):]
            span_sub_idx = sub_text.find(clean_span)
            if span_sub_idx != -1:
                start_index = context_idx + len(clean_context) + span_sub_idx
    
    # If still not found, fallback to first occurrence of span (risky but necessary fallback)
    if start_index == -1:
        start_index = clean_full.find(clean_span)

    if start_index != -1:
        # Calculate start relative to just the span_text
        # If we searched with context, the 'start_index' points to start of context
        # We need start of span
        if clean_context and clean_full.find(search_str) != -1:
             real_start = start_index + len(clean_context)
        elif clean_context and context_idx != -1:
             real_start = start_index # Already calculated above logic
        else:
             real_start = start_index

        end_index = real_start + len(clean_span)
        return real_start, end_index, clean_full[real_start:end_index]
    
    return None, None, None

# ==========================================
# 6. Workbook Generation Function
# ==========================================

def generate_workbook():
    # Initialize Workbook
    wb = Workbook()
    
    # --------------------------------
    # Sheet 1: Note_Text
    # --------------------------------
    ws_text = wb.active
    ws_text.title = "Note_Text"
    ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # --------------------------------
    # Sheet 2: Note_Index
    # --------------------------------
    ws_index = wb.create_sheet("Note_Index")
    headers = ["note_id", "source_file", "procedure_date"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers)
    
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE]
    for key in PROCEDURE_FLAGS:
        row_data.append(PROCEDURE_FLAGS[key])
    ws_index.append(row_data)

    # --------------------------------
    # Sheet 3: Span_Annotations
    # --------------------------------
    ws_spans = wb.create_sheet("Span_Annotations")
    ws_spans.append(["note_id", "span_text", "label", "normalized_value", "context_prefix", "event_id", "start_char", "end_char"])
    
    for span in SPANS:
        s_text, label, norm, ctx, evt = span
        # Start/End left blank for manual review if needed, but filled in Hydrated
        ws_spans.append([NOTE_ID, s_text, label, norm, ctx, evt, "", ""])

    # --------------------------------
    # Sheet 4: Span_Hydrated
    # --------------------------------
    ws_hydrated = wb.create_sheet("Span_Hydrated")
    ws_hydrated.append(["note_id", "span_text", "label", "normalized_value", "context_prefix", "event_id", "start_char", "end_char", "verified_text"])
    
    for span in SPANS:
        s_text, label, norm, ctx, evt = span
        start, end, verified = hydrate_span(NOTE_TEXT, s_text, ctx)
        ws_hydrated.append([NOTE_ID, s_text, label, norm, ctx, evt, start, end, verified])

    # --------------------------------
    # Sheet 5: Event_Log
    # --------------------------------
    ws_events = wb.create_sheet("Event_Log")
    ws_events.append(["note_id", "event_id", "method", "anatomy", "devices", "outcomes"])
    
    for evt in EVENTS:
        dev_str = ", ".join(evt["devices"])
        out_str = ", ".join(evt["outcomes"])
        ws_events.append([NOTE_ID, evt["event_id"], evt["method"], evt["anatomy"], dev_str, out_str])

    # --------------------------------
    # Sheet 6: V3_Procedure_Events
    # --------------------------------
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    ws_v3.append(["note_id", "event_id", "event_data_json"])
    
    for evt in EVENTS:
        ws_v3.append([NOTE_ID, evt["event_id"], json.dumps(evt)])

    # --------------------------------
    # Sheet 7: V3_Registry_JSON
    # --------------------------------
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["note_id", "full_registry_json"])
    
    registry_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source": SOURCE_FILE,
            "procedure_date": PROCEDURE_DATE
        },
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS
    }
    
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

# ==========================================
# 7. Execution Block
# ==========================================

if __name__ == "__main__":
    generate_workbook()