import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# 1. Imports & Constants
NOTE_ID = "note_146"
SOURCE_FILE = "note_146.txt"
PROCEDURE_DATE = ""
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_146 SOURCE_FILE: note_146.txt Procedure: Rigid and flexible bronchoscopy with endobronchial biopsy, tumor debulking, APC ablation, and silicone Y-stent placement
Indication: Mediastinal adenopathy, diagnostic and therapeutic
Anesthesia: General anesthesia;
topical 2% lidocaine to tracheobronchial tree (10 mL)

Pre-Procedure

Standard informed consent and time-out performed.

History/physical updated; labs and imaging reviewed.
Patient tolerated induction and procedure well.

Technique

Rigid bronchoscopy: Black bronchial tube 12.0–11.0 via mouth with 0° 4.0-mm rigid telescope.
Flexible bronchoscopy: Q180 slim and T180 therapeutic bronchoscopes passed through the rigid bronchoscope.
Therapeutic tools: Electrocautery snare, XPS 3000 microdebrider, ERBE APC VIO 300D (argon plasma coagulation).

Airway Findings

Vocal cords: Normal mobility.
Subglottis: Normal.

Proximal trachea: Normal caliber.

Distal trachea / carina:

Mixed obstruction (extrinsic compression + endoluminal tumor) arising from the right lateral distal trachea, extending to involve the main carina and both mainstem bronchi over ~1 cm.
~75% luminal occlusion in the lower trachea.

Carinal mass large, friable, infiltrative, and submucosal with prominent endobronchial component.
Interventions

Electrocautery snare debulking of carinal mass → ~60% recanalization at carina.

Resected tissue sent to pathology.
Endobronchial biopsies (cup forceps) from lower trachea and carina (5 samples) sent for histopathology.
Microdebrider resection of additional lower tracheal tumor → partial recanalization to ~50% of normal.
APC ablation to residual tumor in the trachea, carina, left and right mainstem bronchi → post-ablation lumen ~70% of normal.
Airway stenting:

Silicone Y-stent (Novatec 14–10–10 mm) selected and customized (6 cm tracheal limb; 1 cm each mainstem limb).
Deployed under direct vision across trachea, carina, and both mainstem bronchi.

Post-stent lumen ~90% of normal; no airway orifices obstructed;
final position satisfactory.

Complications / Blood Loss

No immediate complications.

Estimated blood loss: < 5 cc.
Impression

Severe malignant airway obstruction at the distal trachea/carina with bilateral mainstem involvement.
Technically successful rigid bronchoscopy with diagnostic biopsies, multimodal tumor debulking/ablation, and silicone Y-stent placement resulting in substantial airway patency restoration.
Post-Procedure Plan

Observe in post-bronchoscopy recovery until discharge criteria met.

Await pathology results.
Follow up with requesting/oncology service to review final histopathology and guide definitive therapy."""

# 2. Configuration (Procedure Flags)
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 1,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 1,
    "thermal_ablation": 1,
    "tumor_debulking_non_thermal": 1,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 1,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 3. Data Definition (Spans)
SPANS = [
    ("Rigid bronchoscopy", "PROC_METHOD", "Rigid Bronchoscopy", "Technique\n\n", "evt_01"),
    ("Black bronchial tube 12.0–11.0", "DEV_INSTRUMENT", "Rigid Bronchoscope", "Rigid bronchoscopy: ", "evt_01"),
    ("Flexible bronchoscopy", "PROC_METHOD", "Flexible Bronchoscopy", "rigid telescope.\n", "evt_02"),
    ("Q180 slim", "DEV_INSTRUMENT", "Bronchoscope - Diagnostic", "Flexible bronchoscopy: ", "evt_02"),
    ("T180 therapeutic bronchoscopes", "DEV_INSTRUMENT", "Bronchoscope - Therapeutic", "Q180 slim and ", "evt_02"),
    ("Distal trachea", "ANAT_AIRWAY", "Trachea - Distal", "Proximal trachea: Normal caliber.\n\n", "evt_03"),
    ("carina", "ANAT_AIRWAY", "Carina", "Distal trachea / ", "evt_03"),
    ("Mixed obstruction", "OBS_LESION", "Tumor", "Distal trachea / carina:\n\n", "evt_03"),
    ("~75% luminal occlusion", "OUTCOME_AIRWAY_LUMEN_PRE", "75%", "over ~1 cm.\n", "evt_03"),
    ("Electrocautery snare debulking", "PROC_METHOD", "Electrocautery Snare", "Interventions\n\n", "evt_04"),
    ("Electrocautery snare", "DEV_INSTRUMENT", "Snare", "Therapeutic tools: ", "evt_04"),
    ("carinal mass", "ANAT_AIRWAY", "Carina", "debulking of ", "evt_04"),
    ("~60% recanalization", "OUTCOME_AIRWAY_LUMEN_POST", "60%", "mass → ", "evt_04"),
    ("Endobronchial biopsies", "PROC_METHOD", "Endobronchial Biopsy", "", "evt_05"),
    ("cup forceps", "DEV_INSTRUMENT", "Forceps", "Endobronchial biopsies (", "evt_05"),
    ("lower trachea", "ANAT_AIRWAY", "Trachea - Lower", "from ", "evt_05"),
    ("5 samples", "MEAS_COUNT", "5", "carina (", "evt_05"),
    ("Microdebrider resection", "PROC_METHOD", "Microdebrider", "", "evt_06"),
    ("XPS 3000 microdebrider", "DEV_INSTRUMENT", "Microdebrider", "Electrocautery snare, ", "evt_06"),
    ("lower tracheal", "ANAT_AIRWAY", "Trachea - Lower", "of additional ", "evt_06"),
    ("~50% of normal", "OUTCOME_AIRWAY_LUMEN_POST", "50%", "recanalization to ", "evt_06"),
    ("APC ablation", "PROC_METHOD", "APC", "", "evt_07"),
    ("ERBE APC VIO 300D", "DEV_INSTRUMENT", "APC Probe", "XPS 3000 microdebrider, ", "evt_07"),
    ("trachea", "ANAT_AIRWAY", "Trachea", "tumor in the ", "evt_07"),
    ("carina", "ANAT_AIRWAY", "Carina", "trachea, ", "evt_07"),
    ("left and right mainstem bronchi", "ANAT_AIRWAY", "Mainstem Bronchi (Bilateral)", "carina, ", "evt_07"),
    ("~70% of normal", "OUTCOME_AIRWAY_LUMEN_POST", "70%", "post-ablation lumen ", "evt_07"),
    ("Airway stenting", "PROC_METHOD", "Airway Stent", "", "evt_08"),
    ("Silicone Y-stent", "DEV_STENT", "Y-Stent", "Airway stenting:\n\n", "evt_08"),
    ("Novatec", "DEV_STENT_MATERIAL", "Silicone", "Silicone Y-stent (", "evt_08"),
    ("14–10–10 mm", "DEV_STENT_SIZE", "14-10-10mm", "Novatec ", "evt_08"),
    ("deployed", "PROC_ACTION", "Deployed", "limb).\n", "evt_08"),
    ("trachea", "ANAT_AIRWAY", "Trachea", "vision across ", "evt_08"),
    ("carina", "ANAT_AIRWAY", "Carina", "vision across trachea, ", "evt_08"),
    ("both mainstem bronchi", "ANAT_AIRWAY", "Mainstem Bronchi (Bilateral)", "carina, and ", "evt_08"),
    ("~90% of normal", "OUTCOME_AIRWAY_LUMEN_POST", "90%", "Post-stent lumen ", "evt_08"),
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications / Blood Loss\n\n", "evt_09"),
]

# 4. Event Definitions
EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Rigid Bronchoscopy",
        "action": "Insertion",
        "anatomy": "Mouth",
        "devices": ["Black bronchial tube 12.0–11.0"],
        "comments": "Setup"
    },
    {
        "event_id": "evt_02",
        "procedure_type": "Flexible Bronchoscopy",
        "action": "Examination",
        "anatomy": "Tracheobronchial Tree",
        "devices": ["Q180 slim", "T180 therapeutic bronchoscopes"],
        "comments": "Through rigid"
    },
    {
        "event_id": "evt_03",
        "procedure_type": "Diagnostic",
        "action": "Observation",
        "anatomy": "Distal trachea, Carina",
        "lesion": "Mixed obstruction",
        "outcomes": [{"type": "Lumen Patency Pre", "value": "~75% occlusion"}]
    },
    {
        "event_id": "evt_04",
        "procedure_type": "Tumor Debulking",
        "method": "Electrocautery Snare",
        "anatomy": "Carina",
        "outcomes": [{"type": "Lumen Patency Post", "value": "~60% recanalization"}]
    },
    {
        "event_id": "evt_05",
        "procedure_type": "Endobronchial Biopsy",
        "method": "Cup forceps",
        "anatomy": "Lower trachea, Carina",
        "samples": "5"
    },
    {
        "event_id": "evt_06",
        "procedure_type": "Tumor Debulking",
        "method": "Microdebrider",
        "anatomy": "Lower Trachea",
        "outcomes": [{"type": "Lumen Patency Post", "value": "~50% of normal"}]
    },
    {
        "event_id": "evt_07",
        "procedure_type": "Thermal Ablation",
        "method": "APC",
        "anatomy": "Trachea, Carina, Mainstem Bronchi",
        "outcomes": [{"type": "Lumen Patency Post", "value": "~70% of normal"}]
    },
    {
        "event_id": "evt_08",
        "procedure_type": "Airway Stent",
        "device": "Silicone Y-stent (Novatec 14–10–10 mm)",
        "anatomy": "Trachea, Carina, Bilateral Mainstems",
        "action": "Deployed",
        "outcomes": [{"type": "Lumen Patency Post", "value": "~90% of normal"}]
    },
    {
        "event_id": "evt_09",
        "procedure_type": "Procedure Outcome",
        "outcome": "No immediate complications"
    }
]

# 5. Helper Functions
def clean_text(text):
    return text.strip().replace('\r', '')

def hydrate_span(text, span_text, context_prefix):
    cleaned_text = clean_text(text)
    clean_span = clean_text(span_text)
    clean_context = clean_text(context_prefix)
    
    # Escape regex special characters
    escaped_span = re.escape(clean_span)
    escaped_context = re.escape(clean_context)
    
    # Try finding with context first
    pattern_with_context = f"{escaped_context}\s*({escaped_span})"
    match = re.search(pattern_with_context, cleaned_text, re.DOTALL)
    
    if match:
        start_char = match.start(1)
        end_char = match.end(1)
        return start_char, end_char
    
    # Fallback: Find without context (risk of ambiguity)
    match_fallback = re.search(escaped_span, cleaned_text, re.DOTALL)
    if match_fallback:
        return match_fallback.start(), match_fallback.end()
    
    return "", ""

# 6. Workbook Generation Function
def generate_workbook():
    wb = openpyxl.Workbook()
    
    # Sheet 1: Note_Text
    ws1 = wb.active
    ws1.title = "Note_Text"
    ws1.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws1.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # Sheet 2: Note_Index
    ws2 = wb.create_sheet("Note_Index")
    headers = ["NOTE_ID", "procedure_date"] + list(PROCEDURE_FLAGS.keys())
    ws2.append(headers)
    row_data = [NOTE_ID, PROCEDURE_DATE] + list(PROCEDURE_FLAGS.values())
    ws2.append(row_data)
    
    # Sheet 3: Span_Annotations
    ws3 = wb.create_sheet("Span_Annotations")
    ws3.append(["NOTE_ID", "start_char", "end_char", "span_text", "span_label", "span_value_normalized", "context_prefix", "event_id"])
    
    # Sheet 4: Span_Hydrated
    ws4 = wb.create_sheet("Span_Hydrated")
    ws4.append(["NOTE_ID", "start_char", "end_char", "span_text", "span_label", "span_value_normalized", "context_prefix", "event_id"])
    
    for span in SPANS:
        span_text, label, normalized, context, event_id = span
        start, end = hydrate_span(NOTE_TEXT, span_text, context)
        
        # Add to Span_Annotations (blank offsets)
        ws3.append([NOTE_ID, "", "", span_text, label, normalized, context, event_id])
        
        # Add to Span_Hydrated (calculated offsets)
        ws4.append([NOTE_ID, start, end, span_text, label, normalized, context, event_id])

    # Sheet 5: Event_Log
    ws5 = wb.create_sheet("Event_Log")
    ws5.append(["NOTE_ID", "event_id", "property", "value"])
    for event in EVENTS:
        evt_id = event["event_id"]
        for key, val in event.items():
            if key != "event_id":
                if isinstance(val, (list, dict)):
                    val = json.dumps(val)
                ws5.append([NOTE_ID, evt_id, key, val])

    # Sheet 6: V3_Procedure_Events
    ws6 = wb.create_sheet("V3_Procedure_Events")
    ws6.append(["NOTE_ID", "event_id", "procedure_type", "method", "anatomy", "devices", "outcomes"])
    for event in EVENTS:
        ws6.append([
            NOTE_ID,
            event.get("event_id", ""),
            event.get("procedure_type", ""),
            event.get("method", ""),
            event.get("anatomy", ""),
            json.dumps(event.get("devices", [])),
            json.dumps(event.get("outcomes", []))
        ])

    # Sheet 7: V3_Registry_JSON
    ws7 = wb.create_sheet("V3_Registry_JSON")
    ws7.append(["NOTE_ID", "json_object"])
    registry_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source": SOURCE_FILE,
            "procedure_flags": PROCEDURE_FLAGS
        },
        "events": EVENTS,
        "spans": [
            {
                "text": s[0],
                "label": s[1],
                "normalized": s[2],
                "context": s[3],
                "event_id": s[4]
            } for s in SPANS
        ]
    }
    ws7.append([NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

# 7. Execution Block
if __name__ == "__main__":
    generate_workbook()