import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# 1. Imports & Constants
NOTE_ID = "note_147"
SOURCE_FILE = "note_147.txt"
PROCEDURE_DATE = ""
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_147 SOURCE_FILE: note_147.txt Procedure Performed:
Left-sided pleuroscopy (medical thoracoscopy) with pleural fluid drainage, pleural biopsies, and tunneled pleural catheter placement

Indication:
Pleural effusion

Medications / Anesthesia:
Monitored anesthesia care

Pre-Anesthesia Assessment:

ASA Class III (severe systemic disease)

The procedure, including risks, benefits, and alternatives, was explained to the patient.
All questions were answered, and informed consent was obtained and documented per institutional protocol.
A history and physical examination were performed and updated in the pre-procedure assessment record.
Pertinent laboratory studies and imaging were reviewed. A procedural time-out was performed.
The patient was positioned in the lateral decubitus position on the operating table with pressure points appropriately padded.
The patient was sterilely prepped with chlorhexidine gluconate (ChloraPrep) and draped in the usual fashion.
Local Anesthesia

The pleural entry site was identified using ultrasound guidance.
The entry site was infiltrated with 20 mL of 1% lidocaine.
Procedure Description

A 10-mm reusable primary port was placed on the left side at the sixth intercostal space in the mid-axillary line using a Veress needle technique.
A 10.0-mm integrated pleuroscope was introduced through the incision and advanced into the pleural space.
The pleura was inspected via the primary port.

A total of approximately 1,700 mL of amber-colored pleural fluid was removed.
Findings

The pleura demonstrated multiple areas of visible tumor studding involving the parietal pleura, visceral pleura, and lung.
The left lower lobe did not appear fully expanded and was atelectatic.
Biopsies

Biopsies of pleural tumor studding over the diaphragm were obtained using forceps and sent for histopathological examination.
A total of 11 biopsies were obtained.

Pleural Catheter Placement

A 15.5-French PleurX catheter was placed in the pleural space over the diaphragm.
Dressing

The port site was dressed with a transparent dressing.

Complications:
None immediate

Estimated Blood Loss:
Minimal

Post-Procedure Diagnosis:
Suspected pleural metastasis

Post-Procedure Plan:

Observe post-procedure until discharge criteria are met

Obtain post-procedure chest radiograph"""

# 2. Configuration (Procedure Flags)
PROCEDURE_FLAGS = {
    # Bronchoscopy Flags (All 0)
    "diagnostic_bronchoscopy": 0,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    
    # Pleural Flags
    "thoracentesis": 0, # Part of thoracoscopy/IPC
    "chest_tube": 0,
    "ipc": 1,
    "medical_thoracoscopy": 1,
    "pleurodesis": 0,
    "pleural_biopsy": 1,
    "fibrinolytic_therapy": 0
}

# 3. Data Definition (Spans)
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("Left-sided", "LATERALITY", "Left", "Procedure Performed:\n", "evt_01"),
    ("pleuroscopy", "PROC_METHOD", "Medical Thoracoscopy", "Left-sided ", "evt_01"),
    ("medical thoracoscopy", "PROC_METHOD", "Medical Thoracoscopy", "pleuroscopy (", "evt_01"),
    ("pleural biopsies", "PROC_METHOD", "Pleural Biopsy", "fluid drainage, ", "evt_01"),
    ("tunneled pleural catheter", "DEV_CATHETER", "Tunneled Pleural Catheter", "biopsies, and ", "evt_01"),
    
    ("ultrasound guidance", "PROC_METHOD", "Ultrasound Guidance", "identified using ", "evt_02"),
    
    ("10-mm", "MEAS_SIZE", "10 mm", "A ", "evt_03"),
    ("reusable primary port", "DEV_INSTRUMENT", "Trocar/Port", "mm ", "evt_03"),
    ("left side", "LATERALITY", "Left", "placed on the ", "evt_03"),
    ("Veress needle", "DEV_NEEDLE", "Veress Needle", "using a ", "evt_03"),
    
    ("10.0-mm", "MEAS_SIZE", "10 mm", "A ", "evt_04"),
    ("integrated pleuroscope", "DEV_INSTRUMENT", "Pleuroscope", "mm ", "evt_04"),
    
    ("1,700 mL", "MEAS_PLEURAL_DRAIN", "1700 mL", "approximately ", "evt_05"),
    ("amber-colored pleural fluid", "OBS_LESION", "Pleural Effusion", "mL of ", "evt_05"),
    
    ("tumor studding", "OBS_LESION", "Tumor Studding", "visible ", "evt_06"),
    ("parietal pleura", "ANAT_PLEURA", "Parietal Pleura", "involving the ", "evt_06"),
    ("visceral pleura", "ANAT_PLEURA", "Visceral Pleura", "parietal pleura, ", "evt_06"),
    ("lung", "ANAT_LUNG_LOC", "Lung", "visceral pleura, and ", "evt_06"),
    
    ("Biopsies", "PROC_METHOD", "Pleural Biopsy", "Biopsies\n\n", "evt_07"),
    ("diaphragm", "ANAT_PLEURA", "Diaphragmatic Pleura", "over the ", "evt_07"),
    ("forceps", "DEV_INSTRUMENT", "Forceps", "obtained using ", "evt_07"),
    ("11", "MEAS_COUNT", "11", "A total of ", "evt_07"),
    
    ("15.5-French", "DEV_CATHETER_SIZE", "15.5 Fr", "A ", "evt_08"),
    ("PleurX catheter", "DEV_CATHETER", "Tunneled Pleural Catheter", "French ", "evt_08"),
    ("placed", "PROC_ACTION", "Deployment", "catheter was ", "evt_08")
]

# 4. Event Definitions
EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Medical Thoracoscopy",
        "action": "Procedure Start",
        "anatomy": "Pleura",
        "device": "N/A",
        "notes": "Procedure declaration"
    },
    {
        "event_id": "evt_02",
        "procedure_type": "Medical Thoracoscopy",
        "action": "Ultrasound Localization",
        "anatomy": "Pleura",
        "device": "Ultrasound",
        "notes": "Entry site identification"
    },
    {
        "event_id": "evt_03",
        "procedure_type": "Medical Thoracoscopy",
        "action": "Access",
        "anatomy": "Left Pleural Space",
        "device": "Veress Needle, 10-mm Port",
        "notes": "Primary port placement"
    },
    {
        "event_id": "evt_04",
        "procedure_type": "Medical Thoracoscopy",
        "action": "Inspection",
        "anatomy": "Pleura",
        "device": "10.0-mm Pleuroscope",
        "notes": "Scope introduction"
    },
    {
        "event_id": "evt_05",
        "procedure_type": "Medical Thoracoscopy",
        "action": "Drainage",
        "anatomy": "Pleural Space",
        "device": "N/A",
        "notes": "Drainage of 1700mL fluid"
    },
    {
        "event_id": "evt_06",
        "procedure_type": "Medical Thoracoscopy",
        "action": "Observation",
        "anatomy": "Parietal/Visceral Pleura, Lung",
        "device": "N/A",
        "notes": "Tumor studding observed"
    },
    {
        "event_id": "evt_07",
        "procedure_type": "Pleural Biopsy",
        "action": "Biopsy",
        "anatomy": "Diaphragmatic Pleura",
        "device": "Forceps",
        "notes": "11 biopsies taken"
    },
    {
        "event_id": "evt_08",
        "procedure_type": "Indwelling Pleural Catheter",
        "action": "Deployment",
        "anatomy": "Pleural Space",
        "device": "15.5Fr PleurX Catheter",
        "notes": "Tunneled catheter placement"
    }
]

# 5. Helper Functions
def clean_text(text):
    if not text:
        return ""
    return text.strip().replace('\r\n', '\n').replace('\r', '\n')

def hydrate_span(full_text, span_text, context_prefix):
    """
    Finds the start and end indices of span_text within full_text,
    using context_prefix to disambiguate.
    """
    clean_full = clean_text(full_text)
    clean_span = clean_text(span_text)
    clean_context = clean_text(context_prefix)
    
    # Construct search pattern: context + span
    search_str = clean_context + clean_span
    
    try:
        # Find the combined string
        start_index = clean_full.index(search_str)
        # Adjust start index to point to the span, not the context
        actual_start = start_index + len(clean_context)
        actual_end = actual_start + len(clean_span)
        return actual_start, actual_end
    except ValueError:
        # Fallback: try finding just the span (risky if duplicates exist)
        try:
            start_index = clean_full.index(clean_span)
            return start_index, start_index + len(clean_span)
        except ValueError:
            return 0, 0

def find_offsets(text, span_text, context_prefix):
    s, e = hydrate_span(text, span_text, context_prefix)
    return s, e

# 6. Workbook Generation Function
def generate_workbook():
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
        
    # --- Sheet 1: Note_Text ---
    ws_text = wb.create_sheet("Note_Text")
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, clean_text(NOTE_TEXT)])
    
    # --- Sheet 2: Note_Index ---
    ws_index = wb.create_sheet("Note_Index")
    headers_index = ["NOTE_ID", "procedure_date"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers_index)
    row_data = [NOTE_ID, PROCEDURE_DATE] + [PROCEDURE_FLAGS[k] for k in PROCEDURE_FLAGS]
    ws_index.append(row_data)
    
    # --- Sheet 3: Span_Annotations ---
    ws_anno = wb.create_sheet("Span_Annotations")
    headers_anno = ["NOTE_ID", "start_char", "end_char", "span_text", "label", "normalized_value", "context_prefix", "event_id"]
    ws_anno.append(headers_anno)
    
    # Write spans with empty start/end for now (Sheet 4 calculates them)
    for span in SPANS:
        # (span_text, label, normalized_value, context_prefix, event_id)
        ws_anno.append([NOTE_ID, "", "", span[0], span[1], span[2], span[3], span[4]])
        
    # --- Sheet 4: Span_Hydrated ---
    ws_hydrated = wb.create_sheet("Span_Hydrated")
    ws_hydrated.append(headers_anno) # Same structure
    
    cleaned_note_text = clean_text(NOTE_TEXT)
    
    for span in SPANS:
        s_text = span[0]
        label = span[1]
        norm = span[2]
        ctx = span[3]
        evt = span[4]
        
        start, end = find_offsets(cleaned_note_text, s_text, ctx)
        
        ws_hydrated.append([NOTE_ID, start, end, s_text, label, norm, ctx, evt])

    # --- Sheet 5: Event_Log ---
    ws_event_log = wb.create_sheet("Event_Log")
    headers_log = ["NOTE_ID", "event_id", "procedure_type", "action", "anatomy", "device", "notes"]
    ws_event_log.append(headers_log)
    
    for evt in EVENTS:
        ws_event_log.append([
            NOTE_ID,
            evt["event_id"],
            evt["procedure_type"],
            evt["action"],
            evt["anatomy"],
            evt["device"],
            evt["notes"]
        ])

    # --- Sheet 6: V3_Procedure_Events ---
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    headers_v3 = ["NOTE_ID", "procedure_hierarchy", "procedure_events_json"]
    ws_v3.append(headers_v3)
    
    # Create the hierarchy/JSON structure
    # This implies a simplified representation for compatibility
    proc_hierarchy = {
        "procedure_type": "Medical Thoracoscopy",
        "events": EVENTS
    }
    
    ws_v3.append([NOTE_ID, "Medical Thoracoscopy", json.dumps(proc_hierarchy, indent=2)])

    # --- Sheet 7: V3_Registry_JSON ---
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["NOTE_ID", "full_registry_json"])
    
    full_registry_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source": SOURCE_FILE,
            "flags": PROCEDURE_FLAGS
        },
        "spans": [
            {
                "text": s[0],
                "label": s[1],
                "normalized": s[2],
                "event_id": s[4]
            } for s in SPANS
        ],
        "events": EVENTS
    }
    
    ws_json.append([NOTE_ID, json.dumps(full_registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()