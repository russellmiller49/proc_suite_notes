import openpyxl
from openpyxl import Workbook
import re
import json
import os
import datetime

# ==============================================================================
# 1. IMPORTS & CONSTANTS
# ==============================================================================

NOTE_ID = "note_148"
SOURCE_FILE = "note_148.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_148 SOURCE_FILE: note_148.txt Procedure: Fiberoptic bronchoscopy
Anesthesia Type: Previously tracheostomized and sedated
Indication: Hypoxemia
Consent: Emergent
Time-Out: Performed
Pre-Procedure Diagnosis: Emergent
Post-Procedure Diagnosis: Emergent
Medications: Patient previously sedated

Procedure Description

The Olympus Q190 video bronchoscope was introduced through the tracheostomy tube and advanced into the tracheobronchial tree.
A complete airway inspection was performed to at least the first subsegmental airways.
The right-sided airways were unremarkable, with only mildly edematous mucosa and clear secretions.
The left-sided airways were similar, with the exception of purulent secretions pooling in the superior segment of the left lower lobe.
The remaining airways were patent without significant central mucus plugging.
The bronchoscope was wedged into the superior segment of the left lower lobe, and bronchoalveolar lavage was performed with 60 mL of instilled saline and 20 mL of return.
The bronchoscope was then withdrawn, and the procedure was completed.
Estimated Blood Loss: None
Complications: None
Specimens Sent: Bronchoalveolar lavage, left lower lobe
Implants: None

Follow-Up: Continue ICU-level care"""

# ==============================================================================
# 2. CONFIGURATION (PROCEDURE FLAGS)
# ==============================================================================

PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 1,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# ==============================================================================
# 3. DATA DEFINITION (SPANS)
# ==============================================================================

# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("Fiberoptic bronchoscopy", "PROC_METHOD", "Fiberoptic Bronchoscopy", "Procedure: ", "evt_01"),
    ("Olympus Q190 video bronchoscope", "DEV_INSTRUMENT", "Olympus Q190", "The ", "evt_01"),
    ("tracheobronchial tree", "ANAT_AIRWAY", "Tracheobronchial Tree", "into the ", "evt_01"),
    ("superior segment of the left lower lobe", "ANAT_LUNG_LOC", "Left Lower Lobe Superior Segment", "pooling in the ", "evt_01"),
    ("bronchoalveolar lavage", "PROC_METHOD", "BAL", "and ", "evt_02"),
    ("superior segment of the left lower lobe", "ANAT_LUNG_LOC", "Left Lower Lobe Superior Segment", "wedged into the ", "evt_02"),
    ("60 mL", "MEAS_VOL", "60", "performed with ", "evt_02"),
    ("20 mL", "MEAS_VOL", "20", "saline and ", "evt_02"),
    ("None", "OUTCOME_COMPLICATION", "None", "Complications: ", "evt_03")
]

# ==============================================================================
# 4. EVENT DEFINITIONS
# ==============================================================================

EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Bronchoscopy",
        "method": "Fiberoptic Bronchoscopy",
        "anatomy": ["Tracheobronchial Tree", "Left Lower Lobe Superior Segment"],
        "devices": ["Olympus Q190"],
        "specimens": [],
        "outcomes": [],
        "notes": "Diagnostic inspection, noted purulent secretions in LLL Superior Segment."
    },
    {
        "event_id": "evt_02",
        "procedure_type": "Bronchoscopy",
        "method": "BAL",
        "anatomy": ["Left Lower Lobe Superior Segment"],
        "devices": [],
        "specimens": ["BAL Fluid"],
        "outcomes": ["Instilled: 60 mL", "Return: 20 mL"],
        "notes": "BAL performed in superior segment of LLL."
    },
    {
        "event_id": "evt_03",
        "procedure_type": "Global Outcome",
        "method": "N/A",
        "anatomy": [],
        "devices": [],
        "specimens": [],
        "outcomes": ["Complications: None"],
        "notes": "Procedure tolerated well."
    }
]

# ==============================================================================
# 5. HELPER FUNCTIONS
# ==============================================================================

def clean_text(text):
    if not text:
        return ""
    return text.strip().replace('\r', '')

def hydrate_span(full_text, span_text, context_prefix=""):
    """
    Finds the start and end offsets of a span within the full text.
    Uses context_prefix to disambiguate identical phrases.
    """
    clean_full = clean_text(full_text)
    clean_span = clean_text(span_text)
    clean_prefix = clean_text(context_prefix)

    if not clean_span:
        return 0, 0

    # If context is provided, search for prefix + span
    search_target = clean_prefix + clean_span if clean_prefix else clean_span
    
    try:
        # Use simple string search first
        start_index = clean_full.find(search_target)
        
        if start_index == -1:
            # Fallback to loose search if exact match fails
            start_index = clean_full.find(clean_span)
        
        # Adjust start_index if we included the prefix in the search
        if clean_prefix and start_index != -1 and search_target != clean_span:
            start_index += len(clean_prefix)

        if start_index != -1:
            end_index = start_index + len(clean_span)
            return start_index, end_index
    except Exception:
        pass
    
    return 0, 0

# ==============================================================================
# 6. WORKBOOK GENERATION
# ==============================================================================

def generate_workbook():
    # 1. Load or Create Workbook
    if os.path.exists(TEMPLATE_PATH):
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    else:
        wb = Workbook()

    # 2. Sheet 1: Note_Text
    if "Note_Text" not in wb.sheetnames:
        ws_text = wb.create_sheet("Note_Text")
        ws_text.append(["note_id", "source_file", "note_text"])
    else:
        ws_text = wb["Note_Text"]
        # Clear existing data if any (keep header)
        for row in ws_text.iter_rows(min_row=2, max_col=3):
            for cell in row:
                cell.value = None
    
    # Write Note Data
    clean_note = clean_text(NOTE_TEXT)
    ws_text.cell(row=2, column=1, value=NOTE_ID)
    ws_text.cell(row=2, column=2, value=SOURCE_FILE)
    ws_text.cell(row=2, column=3, value=clean_note)

    # 3. Sheet 2: Note_Index (Metadata + Flags)
    if "Note_Index" not in wb.sheetnames:
        ws_index = wb.create_sheet("Note_Index")
        headers = ["note_id", "source_file", "procedure_date"] + list(PROCEDURE_FLAGS.keys())
        ws_index.append(headers)
    else:
        ws_index = wb["Note_Index"]
    
    # Write Flags
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE]
    for key in PROCEDURE_FLAGS:
        row_data.append(PROCEDURE_FLAGS[key])
    
    # Check if row 2 exists, otherwise append
    if ws_index.max_row >= 2:
        for col_num, val in enumerate(row_data, 1):
            ws_index.cell(row=2, column=col_num, value=val)
    else:
        ws_index.append(row_data)

    # 4. Sheet 3: Span_Annotations (Raw Spans)
    if "Span_Annotations" not in wb.sheetnames:
        ws_anno = wb.create_sheet("Span_Annotations")
        ws_anno.append(["note_id", "span_text", "label", "normalized_value", "start_char", "end_char", "event_id"])
    else:
        ws_anno = wb["Span_Annotations"]
        # Clear old data
        ws_anno.delete_rows(2, ws_anno.max_row)

    for span in SPANS:
        # span format: (text, label, norm, context, evt)
        # We leave start/end blank here as per instructions for this sheet (or pre-calc)
        # But commonly this sheet is input for review. logic asks for blank rows in step 6 instruction implies distinct step
        ws_anno.append([NOTE_ID, span[0], span[1], span[2], "", "", span[4]])

    # 5. Sheet 4: Span_Hydrated (Calculated Offsets)
    if "Span_Hydrated" not in wb.sheetnames:
        ws_hydra = wb.create_sheet("Span_Hydrated")
        ws_hydra.append(["note_id", "span_text", "label", "normalized_value", "start_char", "end_char", "event_id"])
    else:
        ws_hydra = wb["Span_Hydrated"]
        ws_hydra.delete_rows(2, ws_hydra.max_row)

    for span in SPANS:
        s_text, label, norm, context, evt_id = span
        start, end = hydrate_span(clean_note, s_text, context)
        ws_hydra.append([NOTE_ID, s_text, label, norm, start, end, evt_id])

    # 6. Sheet 5: Event_Log (Flattened)
    if "Event_Log" not in wb.sheetnames:
        ws_log = wb.create_sheet("Event_Log")
        ws_log.append(["note_id", "event_id", "method", "anatomy", "device", "specimen", "outcome", "full_json"])
    else:
        ws_log = wb["Event_Log"]
        ws_log.delete_rows(2, ws_log.max_row)

    for evt in EVENTS:
        # Flatten simple lists for display
        meth = evt.get("method", "")
        anat = ", ".join(evt.get("anatomy", []))
        dev = ", ".join(evt.get("devices", []))
        spec = ", ".join(evt.get("specimens", []))
        out = ", ".join(evt.get("outcomes", []))
        full_j = json.dumps(evt)
        ws_log.append([NOTE_ID, evt["event_id"], meth, anat, dev, spec, out, full_j])

    # 7. Sheet 6: V3_Procedure_Events (Structured)
    if "V3_Procedure_Events" not in wb.sheetnames:
        ws_v3 = wb.create_sheet("V3_Procedure_Events")
        ws_v3.append(["note_id", "event_id", "procedure_type", "method", "anatomy_json", "devices_json", "specimens_json", "outcomes_json", "notes"])
    else:
        ws_v3 = wb["V3_Procedure_Events"]
        ws_v3.delete_rows(2, ws_v3.max_row)

    for evt in EVENTS:
        ws_v3.append([
            NOTE_ID,
            evt["event_id"],
            evt["procedure_type"],
            evt["method"],
            json.dumps(evt.get("anatomy", [])),
            json.dumps(evt.get("devices", [])),
            json.dumps(evt.get("specimens", [])),
            json.dumps(evt.get("outcomes", [])),
            evt.get("notes", "")
        ])

    # 8. Sheet 7: V3_Registry_JSON (Full Payload)
    if "V3_Registry_JSON" not in wb.sheetnames:
        ws_json = wb.create_sheet("V3_Registry_JSON")
        ws_json.append(["note_id", "full_registry_json"])
    else:
        ws_json = wb["V3_Registry_JSON"]
        ws_json.delete_rows(2, ws_json.max_row)

    # Construct final JSON object
    registry_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source_file": SOURCE_FILE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS
    }
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated workbook: {OUTPUT_PATH}")

# ==============================================================================
# 7. EXECUTION BLOCK
# ==============================================================================

if __name__ == "__main__":
    generate_workbook()