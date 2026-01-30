import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# -------------------------------------------------------------------------
# 1. CONSTANTS & INPUT DATA
# -------------------------------------------------------------------------
NOTE_ID = "note_149"
SOURCE_FILE = "note_149.txt"
PROCEDURE_DATE = ""
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_149 SOURCE_FILE: note_149.txt Procedure: Fiberoptic bronchoscopy
Anesthesia Type: Previously intubated and sedated
Indication: Hypoxia and hemoptysis
Consent: Emergent
Time-Out: Performed
Pre-Procedure Diagnosis: Hemoptysis
Post-Procedure Diagnosis: Hemoptysis versus oral blood aspiration
Medications: Patient previously sedated;
topical anesthesia with 10 mL of 2% lidocaine

Procedure Description

The Olympus Q190 video bronchoscope was introduced through the previously placed endotracheal tube and advanced into the tracheobronchial tree.
Blood was noted within the endotracheal tube and throughout the airway, beginning near the laryngeal inlet.
The tip of the endotracheal tube was visualized approximately 3 cm proximal to the main carina.
Diffuse blood was present throughout the airways, with multiple early clots partially obstructing the bronchi.
These were suctioned via the bronchoscope. All left- and right-sided airways were inspected to at least the first subsegmental level and were patent following suctioning, with the exception of a small foreign body identified in the posterior segment of the right lower lobe.
This material was suctioned from the airway and appeared consistent with tissue of unclear origin.
No active bleeding was visualized within the distal airways following clot removal;
however, continued blood dripping was observed originating proximal to the tip of the endotracheal tube.
No obvious airway injury was identified. Once adequate clearance was achieved and no distal bleeding source was identified, the bronchoscope was withdrawn, and the procedure was completed.
Estimated Blood Loss: No procedurally related blood loss
Complications: None
Specimens Sent: None
Implants: None

Recommendations

Plan for repeat bronchoscopy in the morning for additional clearance of blood and clots

Notify the pulmonary service immediately for increasing airway pressures or worsening hypoxia, as these findings may necessitate urgent repeat bronchoscopy

Follow-Up: Continue ICU-level care"""

# -------------------------------------------------------------------------
# 2. CONFIGURATION (PROCEDURE FLAGS)
# -------------------------------------------------------------------------
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 1,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# -------------------------------------------------------------------------
# 3. DATA DEFINITION (SPANS)
# -------------------------------------------------------------------------
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("Fiberoptic bronchoscopy", "PROC_METHOD", "Bronchoscopy", "Procedure:", "evt_01"),
    ("Olympus Q190 video bronchoscope", "DEV_INSTRUMENT", "Bronchoscope", "The", "evt_01"),
    ("tracheobronchial tree", "ANAT_AIRWAY", "Tracheobronchial Tree", "advanced into the", "evt_01"),
    ("clots", "OBS_LESION", "Clot", "multiple early", "evt_02"),
    ("suctioned", "PROC_ACTION", "Suction", "These were", "evt_02"),
    ("patent", "OUTCOME_AIRWAY_LUMEN_POST", "Patent", "level and were", "evt_02"),
    ("foreign body", "OBS_LESION", "Foreign Body", "exception of a small", "evt_03"),
    ("posterior segment of the right lower lobe", "ANAT_LUNG_LOC", "RLL Posterior Segment", "identified in the", "evt_03"),
    ("suctioned", "PROC_ACTION", "Suction", "This material was", "evt_03"),
    ("No procedurally related blood loss", "OUTCOME_COMPLICATION", "None", "Estimated Blood Loss:", "evt_04"),
    ("None", "OUTCOME_COMPLICATION", "None", "Complications:", "evt_04"),
]

# -------------------------------------------------------------------------
# 4. EVENT DEFINITIONS
# -------------------------------------------------------------------------
EVENTS = [
    {
        "event_id": "evt_01",
        "category": "Procedure",
        "description": "Diagnostic Bronchoscopy",
        "method": ["Bronchoscopy"],
        "devices": ["Bronchoscope"],
        "anatomy": ["Tracheobronchial Tree"],
        "actions": ["Inspection"],
        "outcomes": []
    },
    {
        "event_id": "evt_02",
        "category": "Therapeutic",
        "description": "Therapeutic Aspiration of Clots",
        "method": ["Therapeutic Aspiration"],
        "devices": [],
        "anatomy": ["Bronchus"],
        "actions": ["Suction"],
        "outcomes": ["Patent"]
    },
    {
        "event_id": "evt_03",
        "category": "Therapeutic",
        "description": "Foreign Body Removal",
        "method": ["Foreign Body Removal"],
        "devices": [],
        "anatomy": ["RLL Posterior Segment"],
        "actions": ["Suction"],
        "outcomes": []
    },
    {
        "event_id": "evt_04",
        "category": "Outcome",
        "description": "Overall Procedure Outcome",
        "method": [],
        "devices": [],
        "anatomy": [],
        "actions": [],
        "outcomes": ["None"]
    }
]

# -------------------------------------------------------------------------
# 5. HELPER FUNCTIONS
# -------------------------------------------------------------------------
def clean_text(text):
    if not text:
        return ""
    return re.sub(r'[\r\n]+', ' ', text).strip()

def hydrate_span(text, span_text, context_prefix):
    # Normalize inputs
    text_norm = text.lower().replace('\r\n', '\n')
    span_norm = span_text.lower().replace('\r\n', '\n')
    context_norm = context_prefix.lower().replace('\r\n', '\n')
    
    # Try finding with context first
    search_phrase = context_norm + span_norm
    start_idx = text_norm.find(search_phrase)
    
    if start_idx != -1:
        # Adjust start by length of context to point to span start
        real_start = start_idx + len(context_norm)
        # Verify alignment with original text (case insensitive check)
        # To handle whitespace variances, we rely on the length of the span_text
        end_idx = real_start + len(span_text)
        return real_start, end_idx, text[real_start:end_idx]
    
    # Fallback: search just the span
    start_idx = text_norm.find(span_norm)
    if start_idx != -1:
        end_idx = start_idx + len(span_text)
        return start_idx, end_idx, text[start_idx:end_idx]
    
    return -1, -1, "NOT FOUND"

def find_offsets(text, span_text, context_prefix):
    return hydrate_span(text, span_text, context_prefix)

# -------------------------------------------------------------------------
# 6. WORKBOOK GENERATION
# -------------------------------------------------------------------------
def generate_workbook():
    # A. Create Workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Styles
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # --- SHEET 1: Note_Text ---
    ws1 = wb.create_sheet("Note_Text")
    headers1 = ["note_id", "source_file", "note_text"]
    ws1.append(headers1)
    ws1.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    for cell in ws1[1]:
        cell.font = bold_font
        cell.fill = header_fill

    # --- SHEET 2: Note_Index ---
    ws2 = wb.create_sheet("Note_Index")
    # Columns: Note Metadata + 30 Flags
    flag_keys = list(PROCEDURE_FLAGS.keys())
    headers2 = ["note_id", "source_file", "procedure_date"] + flag_keys
    ws2.append(headers2)
    
    # Values
    flag_values = [PROCEDURE_FLAGS[k] for k in flag_keys]
    row2 = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE] + flag_values
    ws2.append(row2)
    
    for cell in ws2[1]:
        cell.font = bold_font
        cell.fill = header_fill

    # --- SHEET 3: Span_Annotations ---
    ws3 = wb.create_sheet("Span_Annotations")
    headers3 = ["note_id", "span_text", "label", "normalized_value", "context_prefix", "start_char", "end_char", "event_id"]
    ws3.append(headers3)
    
    for span in SPANS:
        # span = (text, label, norm, context, evt_id)
        # Write rows with empty start/end for "Annotations" sheet (human readable/input)
        row = [NOTE_ID, span[0], span[1], span[2], span[3], "", "", span[4]]
        ws3.append(row)

    for cell in ws3[1]:
        cell.font = bold_font
        cell.fill = header_fill

    # --- SHEET 4: Span_Hydrated ---
    ws4 = wb.create_sheet("Span_Hydrated")
    # Same headers as 3, but populated offsets
    ws4.append(headers3)
    
    for span in SPANS:
        s_text = span[0]
        s_context = span[3]
        start, end, matched = hydrate_span(NOTE_TEXT, s_text, s_context)
        row = [NOTE_ID, s_text, span[1], span[2], s_context, start, end, span[4]]
        ws4.append(row)

    for cell in ws4[1]:
        cell.font = bold_font
        cell.fill = header_fill

    # --- SHEET 5: Event_Log ---
    ws5 = wb.create_sheet("Event_Log")
    headers5 = ["note_id", "event_id", "category", "description", "method", "devices", "anatomy", "actions", "outcomes"]
    ws5.append(headers5)
    
    for evt in EVENTS:
        row = [
            NOTE_ID,
            evt["event_id"],
            evt["category"],
            evt["description"],
            ", ".join(evt["method"]),
            ", ".join(evt["devices"]),
            ", ".join(evt["anatomy"]),
            ", ".join(evt["actions"]),
            ", ".join(evt["outcomes"])
        ]
        ws5.append(row)

    for cell in ws5[1]:
        cell.font = bold_font
        cell.fill = header_fill

    # --- SHEET 6: V3_Procedure_Events ---
    ws6 = wb.create_sheet("V3_Procedure_Events")
    headers6 = ["note_id", "event_id", "event_index", "event_category", "event_description", "event_details_json"]
    ws6.append(headers6)
    
    for idx, evt in enumerate(EVENTS):
        details = {
            "method": evt["method"],
            "devices": evt["devices"],
            "anatomy": evt["anatomy"],
            "actions": evt["actions"],
            "outcomes": evt["outcomes"]
        }
        row = [
            NOTE_ID,
            evt["event_id"],
            idx + 1,
            evt["category"],
            evt["description"],
            json.dumps(details)
        ]
        ws6.append(row)

    for cell in ws6[1]:
        cell.font = bold_font
        cell.fill = header_fill

    # --- SHEET 7: V3_Registry_JSON ---
    ws7 = wb.create_sheet("V3_Registry_JSON")
    headers7 = ["note_id", "registry_json_dump"]
    ws7.append(headers7)
    
    # Construct Full JSON object
    full_json = {
        "note_metadata": {
            "note_id": NOTE_ID,
            "source_file": SOURCE_FILE,
            "procedure_date": PROCEDURE_DATE
        },
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS,
        "spans": [
            {
                "text": s[0],
                "label": s[1],
                "normalized": s[2],
                "context": s[3],
                "event_id": s[4]
            } for s in SPANS
        ]
    }
    
    ws7.append([NOTE_ID, json.dumps(full_json, indent=2)])
    for cell in ws7[1]:
        cell.font = bold_font
        cell.fill = header_fill

    # B. Save
    wb.save(OUTPUT_PATH)
    print(f"Generated workbook: {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()