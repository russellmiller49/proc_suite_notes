import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import re
import json
import os
from datetime import datetime

# 1. Imports & Constants
NOTE_ID = "note_150"
SOURCE_FILE = "note_150.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# Cleaning the text slightly to remove tags for better NLP simulation, 
# while keeping the original structure.
NOTE_TEXT = """NOTE_ID:  note_150 SOURCE_FILE: note_150.txt Procedure Name: Bronchoscopy

Indication: Pneumonia with hypoxia

Anesthesia: Patient previously intubated and sedated

Consent: Obtained by telephone from family

Time-Out: Performed

Pre-Procedure Diagnosis: Pneumonia

Post-Procedure Diagnosis: Hypoxia

Medications: Patient previously sedated

Procedure Description

An Olympus Q190 video bronchoscope was introduced through the previously placed endotracheal tube and advanced into the tracheobronchial tree.
The tip of the endotracheal tube was confirmed to be in appropriate position, approximately 2 cm above the carina.
Thick, clear, non-purulent secretions were present predominantly within the right-sided airways and were suctioned to clearance.
A complete airway inspection was performed and was notable only for a small white raised endobronchial nodule in the distal bronchus intermedius.
Targeted biopsy of the endobronchial nodule was performed using cupped forceps.
Minimal oozing was observed following biopsy, which resolved spontaneously after a brief period.
The bronchoscope was then advanced into the right middle lobe medial segment, where bronchoalveolar lavage was performed with instillation of 120 mL of saline and return of approximately 80 mL using hand suction.
The bronchoscope was subsequently withdrawn, and the procedure was completed without complication.
Estimated Blood Loss

None

Complications

None

Specimens Sent

Bronchoalveolar lavage from the right middle lobe for standard infectious evaluation

Biopsy specimen from the bronchus intermedius nodule for culture only

Implants

None

Post-Procedure Plan

Continue ICU-level care"""

# 2. Configuration (Procedure Flags)
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1, # Inspection performed
    "bal": 1,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 1,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 3. Data Definition (Spans)
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    # Event 01: Inspection
    ("Olympus Q190 video bronchoscope", "DEV_INSTRUMENT", "Olympus Q190", "Procedure Description\n\nAn ", "evt_01"),
    ("tracheobronchial tree", "ANAT_AIRWAY", "Tracheobronchial Tree", "advanced into the ", "evt_01"),
    ("airway inspection", "PROC_METHOD", "Inspection", "A complete ", "evt_01"),
    
    # Event 02: Biopsy
    ("distal bronchus intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "nodule in the ", "evt_02"),
    ("endobronchial nodule", "OBS_LESION", "Nodule", "small white raised ", "evt_02"),
    ("Targeted biopsy", "PROC_METHOD", "Endobronchial Biopsy", "distal bronchus intermedius.\n", "evt_02"),
    ("cupped forceps", "DEV_INSTRUMENT", "Forceps", "performed using ", "evt_02"),
    
    # Event 03: BAL
    ("right middle lobe medial segment", "ANAT_LUNG_LOC", "RML Medial Segment", "advanced into the ", "evt_03"),
    ("bronchoalveolar lavage", "PROC_METHOD", "BAL", "segment, where ", "evt_03"),
    ("120 mL", "MEAS_VOL", "120", "instillation of ", "evt_03"),
    ("80 mL", "MEAS_VOL", "80", "return of approximately ", "evt_03"),

    # Event 04: Outcome
    ("None", "OUTCOME_COMPLICATION", "None", "Complications\n\n", "evt_04")
]

# 4. Event Definitions
EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Bronchoscopy",
        "action": "Inspection",
        "anatomy": ["Tracheobronchial Tree"],
        "devices": ["Olympus Q190"],
        "details": "Initial airway inspection"
    },
    {
        "event_id": "evt_02",
        "procedure_type": "Bronchoscopy",
        "action": "Endobronchial Biopsy",
        "anatomy": ["Bronchus Intermedius"],
        "devices": ["Forceps"],
        "findings": ["Nodule"],
        "details": "Targeted biopsy of nodule"
    },
    {
        "event_id": "evt_03",
        "procedure_type": "Bronchoscopy",
        "action": "BAL",
        "anatomy": ["RML Medial Segment"],
        "devices": [],
        "measurements": {"instilled": "120 mL", "return": "80 mL"},
        "details": "Lavage performed"
    },
    {
        "event_id": "evt_04",
        "procedure_type": "General",
        "action": "Outcome",
        "anatomy": [],
        "devices": [],
        "complication": "None",
        "details": "Procedure completed without complication"
    }
]

# 5. Helper Functions
def clean_text(text):
    if not text:
        return ""
    return text.strip()

def hydrate_span(full_text, span_text, context_prefix):
    """
    Finds the start and end character offsets of span_text within full_text,
    using context_prefix to disambiguate.
    """
    # Normalize line endings for consistent finding
    full_text_norm = full_text.replace('\r\n', '\n')
    context_prefix_norm = context_prefix.replace('\r\n', '\n')
    span_text_norm = span_text.replace('\r\n', '\n')
    
    # Construct a search pattern
    # We look for the context followed immediately or closely by the span
    pattern = re.escape(context_prefix_norm) + r"\s*" + re.escape(span_text_norm)
    
    match = re.search(pattern, full_text_norm)
    if match:
        # The span starts after the context in the match
        # We need to find exactly where the span part starts in the matched string
        full_match_str = match.group(0)
        start_in_match = full_match_str.rfind(span_text_norm)
        
        start_char = match.start() + start_in_match
        end_char = start_char + len(span_text_norm)
        return start_char, end_char
    else:
        # Fallback: simple search if context fails (warning: might pick wrong instance)
        start_char = full_text_norm.find(span_text_norm)
        if start_char != -1:
            return start_char, start_char + len(span_text_norm)
        return -1, -1

# 6. Workbook Generation Function
def generate_workbook():
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Removes default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # --- Sheet 1: Note_Text ---
    ws_text = wb.create_sheet("Note_Text")
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # --- Sheet 2: Note_Index ---
    ws_index = wb.create_sheet("Note_Index")
    headers_index = ["NOTE_ID", "PROCEDURE_DATE"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers_index)
    row_data = [NOTE_ID, PROCEDURE_DATE] + list(PROCEDURE_FLAGS.values())
    ws_index.append(row_data)
    
    # --- Sheet 3: Span_Annotations ---
    ws_anno = wb.create_sheet("Span_Annotations")
    headers_anno = ["NOTE_ID", "START_CHAR", "END_CHAR", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "EVENT_ID"]
    ws_anno.append(headers_anno)
    
    for span_data in SPANS:
        span_txt, label, norm_val, context, evt_id = span_data
        # Placeholders for Start/End
        ws_anno.append([NOTE_ID, "", "", span_txt, label, norm_val, context, evt_id])
        
    # --- Sheet 4: Span_Hydrated ---
    ws_hydra = wb.create_sheet("Span_Hydrated")
    ws_hydra.append(headers_anno)
    
    for span_data in SPANS:
        span_txt, label, norm_val, context, evt_id = span_data
        start, end = hydrate_span(NOTE_TEXT, span_txt, context)
        ws_hydra.append([NOTE_ID, start, end, span_txt, label, norm_val, context, evt_id])

    # --- Sheet 5: Event_Log ---
    ws_log = wb.create_sheet("Event_Log")
    headers_log = ["NOTE_ID", "EVENT_ID", "PROCEDURE_TYPE", "ACTION", "ANATOMY", "DEVICES", "DETAILS"]
    ws_log.append(headers_log)
    
    for evt in EVENTS:
        # Flatten lists for CSV-like cell content
        anat_str = ", ".join(evt.get("anatomy", []))
        dev_str = ", ".join(evt.get("devices", []))
        ws_log.append([
            NOTE_ID,
            evt["event_id"],
            evt["procedure_type"],
            evt["action"],
            anat_str,
            dev_str,
            evt.get("details", "")
        ])

    # --- Sheet 6: V3_Procedure_Events ---
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    headers_v3 = ["note_id", "event_id", "event_type", "event_metadata"]
    ws_v3.append(headers_v3)
    
    for evt in EVENTS:
        metadata = json.dumps(evt)
        ws_v3.append([NOTE_ID, evt["event_id"], evt["action"], metadata])

    # --- Sheet 7: V3_Registry_JSON ---
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["note_id", "registry_json"])
    
    # Construct Full JSON Object
    registry_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source_file": SOURCE_FILE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS,
        "spans": [
            {
                "text": s[0],
                "label": s[1],
                "normalized": s[2],
                "event_id": s[4]
            } for s in SPANS
        ]
    }
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])
    
    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated workbook: {OUTPUT_PATH}")

# 7. Execution Block
if __name__ == "__main__":
    generate_workbook()