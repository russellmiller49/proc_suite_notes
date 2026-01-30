import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# ==========================================
# 1. Imports & Constants
# ==========================================

NOTE_ID = "note_151"
SOURCE_FILE = "note_151.txt"
PROCEDURE_DATE = ""
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_151 SOURCE_FILE: note_151.txt Tube Thoracostomy Procedure Note

Indication: Hemothorax

The patient was positioned in the usual fashion and the left chest was prepped and draped in a sterile manner using chlorhexidine.
Local anesthesia was achieved with a total of 10 mL of 1% lidocaine infiltrated into the skin, subcutaneous tissue, superior aspect of the rib periosteum, and parietal pleura.
A 2-cm incision was made parallel to the rib in the left midaxillary line at the level of the fourth rib.
Blunt dissection was carried out through the subcutaneous tissue superficial and superior to the rib down to the level of the pleura.
The pleural space was entered bluntly, with immediate expression of a large volume of blood.
The parietal pleural opening was then expanded bluntly, and a finger was inserted and carefully swept in all directions to confirm intrapleural placement and to assess for adhesions.
A 34 French chest tube was then inserted using the operator’s finger as a guide.
The tube was directed inferiorly and advanced without difficulty. Blood continued to pour from the insertion site during suturing of the tube to the skin, with approximately 1 liter of dark blood lost externally, outside of the pleurovac.
The chest tube was then securely connected to a pleurovac with tape, at which time an additional approximately 1 liter of blood drained rapidly before output slowed.
A sterile occlusive dressing was applied over the insertion site.
A second, nonfunctional chest tube was removed, revealing a large fibrin clot completely occluding the lumen.
The removal site bled briskly, and three interrupted sutures were placed, after which hemostasis was achieved.
No immediate complications were noted. A post-procedure chest x-ray is pending at the time of this note.
Complications: None
Estimated Blood Loss: Approximately 2 liters"""

# ==========================================
# 2. Configuration (Procedure Flags)
# ==========================================

PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 0,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 1,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0,
}

# ==========================================
# 3. Data Definition (Spans)
# ==========================================

# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("left", "LATERALITY", "Left", "fashion and the ", "evt_01"),
    ("10 mL", "MEAS_VOL", "10", "total of ", "evt_01"),
    ("2-cm", "MEAS_SIZE", "2", "A ", "evt_01"),
    ("incision", "PROC_ACTION", "Incision", "2-cm ", "evt_01"),
    ("midaxillary line", "ANAT_PLEURA", "Midaxillary Line", "left ", "evt_01"),
    ("fourth rib", "ANAT_PLEURA", "4th Rib", "level of the ", "evt_01"),
    ("pleural space", "ANAT_PLEURA", "Pleural Space", "The ", "evt_01"),
    ("entered", "PROC_ACTION", "Entry", "was ", "evt_01"),
    ("34 French", "DEV_CATHETER_SIZE", "34Fr", "A ", "evt_01"),
    ("chest tube", "DEV_CATHETER", "Chest Tube", "A 34 French ", "evt_01"),
    ("inserted", "PROC_ACTION", "Insertion", "was then ", "evt_01"),
    ("1 liter", "MEAS_PLEURAL_DRAIN", "1000", "approximately ", "evt_01"),
    ("1 liter", "MEAS_PLEURAL_DRAIN", "1000", "additional approximately ", "evt_01"),
    ("chest tube", "DEV_CATHETER", "Chest Tube", "nonfunctional ", "evt_02"),
    ("removed", "PROC_ACTION", "Removal", "was ", "evt_02"),
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "achieved.\n", "evt_01"),
    ("2 liters", "MEAS_PLEURAL_DRAIN", "2000", "Approximately ", "evt_01")
]

# ==========================================
# 4. Event Definitions
# ==========================================

EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Tube Thoracostomy",
        "action": "Insertion",
        "device_type": "Chest Tube",
        "device_size": "34Fr",
        "anatomy": "Left Pleura",
        "laterality": "Left",
        "outcome": "Successful",
        "complication": "None"
    },
    {
        "event_id": "evt_02",
        "procedure_type": "Tube Thoracostomy",
        "action": "Removal",
        "device_type": "Chest Tube",
        "note": "Removal of nonfunctional tube",
        "anatomy": "Left Pleura",
        "laterality": "Left"
    }
]

# ==========================================
# 5. Helper Functions
# ==========================================

def clean_text(text):
    if not text:
        return ""
    return text.strip()

def hydrate_span(full_text, span_text, context_prefix):
    """
    Finds the start and end indices of span_text within full_text,
    relying on context_prefix to disambiguate.
    """
    if not span_text:
        return "", "", ""
    
    # Normalize line endings for search
    search_text = full_text.replace('\r\n', '\n')
    search_span = span_text
    search_context = context_prefix

    # Construct a regex pattern that looks for context + span
    # We escape them to handle special chars like brackets or dots
    pattern_str = re.escape(search_context) + re.escape(search_span)
    
    match = re.search(pattern_str, search_text)
    
    if match:
        # The full match includes the context. We only want the span indices.
        # Start of span = Start of match + length of context
        start_index = match.start() + len(search_context)
        end_index = start_index + len(search_span)
        return start_index, end_index, span_text
    else:
        # Fallback: Try to find just the span if context fails (warning: might be ambiguous)
        # This is a strict fail-safe; ideally context matches.
        idx = search_text.find(search_span)
        if idx != -1:
            return idx, idx + len(search_span), span_text
            
    return "", "", span_text

# ==========================================
# 6. Workbook Generation Function
# ==========================================

def generate_workbook():
    # Create workbook (in memory, pretending to load template logic)
    wb = openpyxl.Workbook()
    
    # ----------------------------------------
    # Sheet 1: Note_Text
    # ----------------------------------------
    ws_text = wb.active
    ws_text.title = "Note_Text"
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # ----------------------------------------
    # Sheet 2: Note_Index
    # ----------------------------------------
    ws_index = wb.create_sheet("Note_Index")
    headers = ["NOTE_ID", "PROCEDURE_DATE"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers)
    
    row_data = [NOTE_ID, PROCEDURE_DATE] + [PROCEDURE_FLAGS[k] for k in PROCEDURE_FLAGS]
    ws_index.append(row_data)

    # ----------------------------------------
    # Sheet 3: Span_Annotations
    # ----------------------------------------
    ws_anno = wb.create_sheet("Span_Annotations")
    # Columns: NOTE_ID, START_CHAR, END_CHAR, LABEL, TEXT, NORMALIZED_VALUE, CONTEXT_PREFIX, EVENT_ID
    ws_anno.append(["NOTE_ID", "START_CHAR", "END_CHAR", "LABEL", "TEXT", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "EVENT_ID"])
    
    for span in SPANS:
        s_text, s_label, s_norm, s_context, s_evt = span
        # Placeholder blank start/end for this sheet as per instructions (or raw dump)
        ws_anno.append([NOTE_ID, "", "", s_label, s_text, s_norm, s_context, s_evt])

    # ----------------------------------------
    # Sheet 4: Span_Hydrated
    # ----------------------------------------
    ws_hydrated = wb.create_sheet("Span_Hydrated")
    ws_hydrated.append(["NOTE_ID", "START_CHAR", "END_CHAR", "LABEL", "TEXT", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "EVENT_ID"])
    
    for span in SPANS:
        s_text, s_label, s_norm, s_context, s_evt = span
        start, end, found_text = hydrate_span(NOTE_TEXT, s_text, s_context)
        ws_hydrated.append([NOTE_ID, start, end, s_label, found_text, s_norm, s_context, s_evt])

    # ----------------------------------------
    # Sheet 5: Event_Log
    # ----------------------------------------
    ws_event_log = wb.create_sheet("Event_Log")
    ws_event_log.append(["NOTE_ID", "EVENT_ID", "KEY", "VALUE"])
    
    for evt in EVENTS:
        e_id = evt["event_id"]
        for k, v in evt.items():
            if k != "event_id":
                ws_event_log.append([NOTE_ID, e_id, k, str(v)])

    # ----------------------------------------
    # Sheet 6: V3_Procedure_Events
    # ----------------------------------------
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    ws_v3.append(["NOTE_ID", "EVENT_ID", "PROCEDURE_TYPE", "DEVICES", "ANATOMY", "OUTCOMES", "FULL_JSON"])
    
    for evt in EVENTS:
        e_id = evt["event_id"]
        proc_type = evt.get("procedure_type", "")
        
        # Construct specific subsets for columns
        dev_dict = {k:v for k,v in evt.items() if "device" in k}
        anat_dict = {k:v for k,v in evt.items() if k in ["anatomy", "laterality"]}
        out_dict = {k:v for k,v in evt.items() if k in ["outcome", "complication"]}
        
        ws_v3.append([
            NOTE_ID,
            e_id,
            proc_type,
            json.dumps(dev_dict),
            json.dumps(anat_dict),
            json.dumps(out_dict),
            json.dumps(evt)
        ])

    # ----------------------------------------
    # Sheet 7: V3_Registry_JSON
    # ----------------------------------------
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["NOTE_ID", "JSON_PAYLOAD"])
    
    # Aggregate everything into one registry object
    registry_obj = {
        "metadata": {
            "note_id": NOTE_ID,
            "source_file": SOURCE_FILE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS,
        "spans": [
            {
                "text": s[0], 
                "label": s[1], 
                "normalized": s[2], 
                "event_id": s[4]
            } for s in SPANS
        ]
    }
    
    ws_json.append([NOTE_ID, json.dumps(registry_obj, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

# ==========================================
# 7. Execution Block
# ==========================================

if __name__ == "__main__":
    generate_workbook()