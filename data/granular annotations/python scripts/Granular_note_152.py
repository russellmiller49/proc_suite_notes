import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
from datetime import datetime

# =============================================================================
# 1. CONSTANTS & INPUT DATA
# =============================================================================

NOTE_ID = "note_152"
SOURCE_FILE = "note_152.txt"
PROCEDURE_DATE = "" 
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_152 SOURCE_FILE: note_152.txt Procedure: Fiberoptic bronchoscopy
Indication: Pneumonia
Anesthesia: Patient previously intubated and sedated
Consent: Obtained from daughter
Time-Out: Performed

Pre-Procedure Diagnosis: Pneumonia
Post-Procedure Diagnosis: Pneumonia

Medications: Patient previously sedated per ICU record

Procedure Description

The Olympus Q190 video bronchoscope was introduced through the previously placed endotracheal tube and advanced into the tracheobronchial tree. The tip of the endotracheal tube was noted to be at the level of the cricoid cartilage and was advanced 5 cm to an appropriate position approximately 2 cm above the carina.

The bronchoscope was then advanced with suction off into the anterior segment of the right lower lobe and wedged into position. Purulent secretions were visualized within this segment. Bronchoalveolar lavage was performed with instillation of 120 mL of sterile saline and return of 40 mL using hand suction technique, yielding purulent material.

Following completion of lavage, a complete airway inspection was performed. The remainder of the tracheobronchial tree was normal in appearance, without additional secretions or evidence of endobronchial disease.

Specimens

Bronchoalveolar lavage from the right lower lobe sent for microbiologic analysis

Estimated Blood Loss

None.

Complications

None.

Implants

None.

Post-Procedure Plan

Continue ICU-level care."""

# =============================================================================
# 2. CONFIGURATION (PROCEDURE FLAGS)
# =============================================================================

PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 1,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0,
}

# =============================================================================
# 3. DATA DEFINITION (SPANS)
# =============================================================================
# Format: (span_text, label, normalized_value, context_prefix, event_id)

SPANS = [
    # Event 1: Bronchoscopy/Inspection
    ("Fiberoptic bronchoscopy", "PROC_METHOD", "Fiberoptic bronchoscopy", "Procedure: ", "evt_01"),
    ("Olympus Q190 video bronchoscope", "DEV_INSTRUMENT", "Olympus Q190", "The ", "evt_01"),
    ("tracheobronchial tree", "ANAT_AIRWAY", "tracheobronchial tree", "into the ", "evt_01"),
    ("Purulent secretions", "OBS_LESION", "purulent secretions", "position. ", "evt_01"),
    ("right lower lobe", "ANAT_LUNG_LOC", "RLL", "segment of the ", "evt_01"),
    
    # Event 2: BAL
    ("Bronchoalveolar lavage", "PROC_METHOD", "Bronchoalveolar lavage", "segment. ", "evt_02"),
    ("120 mL", "MEAS_VOL", "120", "instillation of ", "evt_02"),
    ("40 mL", "MEAS_VOL", "40", "return of ", "evt_02"),
    ("purulent material", "OBS_LESION", "purulent material", "yielding ", "evt_02"),
    
    # Global/Outcomes
    ("None", "OUTCOME_COMPLICATION", "None", "Complications\n\n", "evt_01"),
    ("None", "OUTCOME_COMPLICATION", "None", "Blood Loss\n\n", "evt_01"),
]

# =============================================================================
# 4. EVENT DEFINITIONS (V3 REGISTRY)
# =============================================================================

EVENTS = [
    {
        "event_id": "evt_01",
        "type": "diagnostic_bronchoscopy",
        "method": "Fiberoptic bronchoscopy",
        "target": {"anatomy": ["tracheobronchial tree", "RLL"]},
        "devices": ["Olympus Q190 video bronchoscope"],
        "outcomes": ["complications: None"],
        "measurements": [],
        "specimens": []
    },
    {
        "event_id": "evt_02",
        "type": "bal",
        "method": "Bronchoalveolar lavage",
        "target": {"anatomy": ["RLL anterior segment"]},
        "devices": [],
        "outcomes": [],
        "measurements": ["instilled: 120 mL", "return: 40 mL"],
        "specimens": ["microbiology"]
    }
]

# =============================================================================
# 5. HELPER FUNCTIONS
# =============================================================================

def clean_text(text):
    if not text: return ""
    return str(text).strip()

def hydrate_span(full_text, span_text, context_prefix=""):
    """
    Finds start_char and end_char for a span.
    Logic:
    1. Find all matches of span_text.
    2. If context_prefix is provided, filter matches where prefix immediately precedes.
    3. Return first match's offsets.
    """
    if not span_text or span_text not in full_text:
        return "", ""

    # Escape for regex
    esc_span = re.escape(span_text)
    
    # Find all iterators
    matches = list(re.finditer(esc_span, full_text))
    
    if not matches:
        return "", ""

    selected_match = None

    if context_prefix:
        # Normalize context for robust matching (strip trailing whitespace)
        clean_prefix = context_prefix.rstrip()
        
        candidates = []
        for m in matches:
            start = m.start()
            # Look back from start
            # We grab a chunk somewhat larger than prefix to allow for flexible whitespace
            lookback_len = len(clean_prefix) + 20 
            window_start = max(0, start - lookback_len)
            window_text = full_text[window_start:start]
            
            # Simple check: does window_text end with clean_prefix (ignoring some whitespace)?
            # A stricter check:
            if clean_prefix in window_text:
                candidates.append(m)
        
        if candidates:
            selected_match = candidates[0]
        else:
            # Fallback to first if context matching fails (or return empty implies ambiguous)
            selected_match = matches[0]
    else:
        selected_match = matches[0]

    return selected_match.start(), selected_match.end()

# =============================================================================
# 6. WORKBOOK GENERATION
# =============================================================================

def generate_workbook():
    # 1. Load or Create Template
    if os.path.exists(TEMPLATE_PATH):
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    else:
        wb = openpyxl.Workbook()
        # Create expected sheets if missing
        expected_sheets = ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", 
                           "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]
        for s in expected_sheets:
            if s not in wb.sheetnames:
                wb.create_sheet(s)

    # 2. Sheet: Note_Text
    if "Note_Text" not in wb.sheetnames: wb.create_sheet("Note_Text")
    ws_text = wb["Note_Text"]
    # Clear existing
    ws_text.delete_rows(1, ws_text.max_row)
    ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 3. Sheet: Note_Index
    if "Note_Index" not in wb.sheetnames: wb.create_sheet("Note_Index")
    ws_index = wb["Note_Index"]
    ws_index.delete_rows(2, ws_index.max_row) # Keep header
    
    # Construct row based on hardcoded flags
    # Header logic assumed from typical template:
    # source_file, note_id, encounter_id, procedure_date, site, reviewer, status, free_text_notes, [FLAGS...]
    
    # Flags keys order must match the template header order roughly or we rely on the list provided in prompt
    ordered_flags = [
        "diagnostic_bronchoscopy", "bal", "bronchial_wash", "brushings", "endobronchial_biopsy", 
        "tbna_conventional", "linear_ebus", "radial_ebus", "navigational_bronchoscopy", 
        "transbronchial_biopsy", "transbronchial_cryobiopsy", "therapeutic_aspiration", 
        "foreign_body_removal", "airway_dilation", "airway_stent", "thermal_ablation", 
        "tumor_debulking_non_thermal", "cryotherapy", "blvr", "peripheral_ablation", 
        "bronchial_thermoplasty", "whole_lung_lavage", "rigid_bronchoscopy", "thoracentesis", 
        "chest_tube", "ipc", "medical_thoracoscopy", "pleurodesis", "pleural_biopsy", 
        "fibrinolytic_therapy"
    ]
    
    flag_values = [PROCEDURE_FLAGS.get(k, 0) for k in ordered_flags]
    
    # Pre-flag columns
    meta_row = [SOURCE_FILE, NOTE_ID, "", PROCEDURE_DATE, "", "Auto-Gen", "done", ""]
    ws_index.append(meta_row + flag_values)

    # 4. Sheets: Span_Annotations & Span_Hydrated
    if "Span_Annotations" not in wb.sheetnames: wb.create_sheet("Span_Annotations")
    ws_anno = wb["Span_Annotations"]
    ws_anno.delete_rows(2, ws_anno.max_row)

    if "Span_Hydrated" not in wb.sheetnames: wb.create_sheet("Span_Hydrated")
    ws_hydra = wb["Span_Hydrated"]
    ws_hydra.delete_rows(2, ws_hydra.max_row)

    for i, span_data in enumerate(SPANS):
        span_text, label, norm_val, context, evt_id = span_data
        span_id = f"{NOTE_ID}_s{i+1:03d}"
        
        # Annotation Row (offsets blank)
        # source_file, note_id, span_id, section_type, context_prefix, span_text, match_index, start_char, end_char, span_len, label, normalized_value, schema_field, event_id, ...
        # Simplified mapping
        row_anno = [
            SOURCE_FILE, NOTE_ID, span_id, "", context, span_text, "", "", "", "", label, norm_val, "", evt_id
        ]
        ws_anno.append(row_anno)

        # Hydration Logic
        start_char, end_char = hydrate_span(NOTE_TEXT, span_text, context)
        status = "Ambiguous" if start_char == "" else "Hydrated"
        
        row_hydra = [
            SOURCE_FILE, NOTE_ID, span_id, "", context, span_text, "", start_char, end_char, 
            (end_char - start_char) if isinstance(end_char, int) else "",
            label, norm_val, "", evt_id, "", "", "", "", "", status
        ]
        ws_hydra.append(row_hydra)

    # 5. Sheet: Event_Log (Flat)
    if "Event_Log" not in wb.sheetnames: wb.create_sheet("Event_Log")
    ws_events = wb["Event_Log"]
    ws_events.delete_rows(2, ws_events.max_row)

    for evt in EVENTS:
        evt_id = evt["event_id"]
        # source_file, note_id, event_id, event_type, method, anatomy_target, device, ...
        flat_row = [
            SOURCE_FILE, NOTE_ID, evt_id, evt["type"], evt["method"],
            ", ".join(evt["target"].get("anatomy", [])),
            ", ".join(evt["devices"]),
            "", # needle_gauge
            "", # stations
            "", # counts
            ", ".join(evt["measurements"]),
            ", ".join(evt["specimens"]),
            "", # findings
            "FALSE", # is_historical
            "Auto", # reviewer
            "", # comments
            "", # device_size
            "", # device_material
            "", # outcomes keys...
            "",
            "",
            "",
            ", ".join(evt["outcomes"]) # outcome_complication
        ]
        ws_events.append(flat_row)

    # 6. Sheet: V3_Procedure_Events
    if "V3_Procedure_Events" not in wb.sheetnames: wb.create_sheet("V3_Procedure_Events")
    ws_v3 = wb["V3_Procedure_Events"]
    ws_v3.delete_rows(2, ws_v3.max_row)
    
    v3_procedures = []

    for evt in EVENTS:
        # V3 Schema structure is slightly different, usually JSON blobs in columns
        # note_id, event_id, type, target.anatomy, ... devices_json, measurements_json...
        
        # Build mini dicts for JSON columns
        dev_json = json.dumps(evt["devices"]) if evt["devices"] else ""
        meas_json = json.dumps(evt["measurements"]) if evt["measurements"] else ""
        spec_json = json.dumps(evt["specimens"]) if evt["specimens"] else ""
        
        row_v3 = [
            NOTE_ID, evt["event_id"], evt["type"], 
            json.dumps(evt["target"]), # target columns usually spread, here just dumping for safety/simplicity or placeholder
            "", "", "", "", "", # granular target cols
            evt["method"],
            dev_json,
            meas_json,
            spec_json,
            "", # findings_json
            "", # evidence_quote
            "", "", "", # stent/cath
            "", "", "", "", # outcomes
            json.dumps(evt["outcomes"])
        ]
        ws_v3.append(row_v3)
        
        # Construct object for final JSON
        proc_obj = {
            "event_id": evt["event_id"],
            "type": evt["type"],
            "method": evt["method"],
            "target": evt["target"],
            "devices": evt["devices"],
            "measurements": evt["measurements"],
            "specimens": evt["specimens"],
            "outcomes": evt["outcomes"]
        }
        v3_procedures.append(proc_obj)

    # 7. Sheet: V3_Registry_JSON
    if "V3_Registry_JSON" not in wb.sheetnames: wb.create_sheet("V3_Registry_JSON")
    ws_json = wb["V3_Registry_JSON"]
    ws_json.delete_rows(2, ws_json.max_row)

    final_json_obj = {
        "schema_version": "ip_registry_v3_draft",
        "note_id": NOTE_ID,
        "source_file": SOURCE_FILE,
        "procedure_date": PROCEDURE_DATE,
        "procedures": v3_procedures
    }
    
    ws_json.append(["ip_registry_v3_draft", NOTE_ID, json.dumps(final_json_obj, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()