import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import re
import json
import os
import datetime

# 1. Imports & Constants
NOTE_ID = "note_153"
SOURCE_FILE = "note_153.txt"
PROCEDURE_DATE = "2026-01-13"
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_153 SOURCE_FILE: note_153.txt Indications: Hilar adenopathy, presumed metastatic breast CA
Procedure Performed: EBUS bronchoscopy single station.
Pre-operative diagnosis: hilar adenopathy 
Post-operative diagnosis: malignant hilar adenopathy 
Medications: General Anesthesia,
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention.
Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway is in good position. The vocal cords appeared normal. The subglottic space was normal.
The trachea is of normal caliber. The carina is sharp.
The tracheobronchial tree was examined to at least the first subsegmental level. Bronchial mucosa and anatomy were normal;
there are no endobronchial lesions, and no secretions. The video bronchoscope was then removed and the UC180F convex probe EBUS bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
Ultrasound was utilized to identify and measure the radiographically suspicious station 11Ri lymph node.
Sampling by transbronchial needle aspiration was performed beginning with the Olympus EBUS-TBNA 22 gauge needle.
Rapid onsite evaluation read as malignancy. All samples were sent for routine cytology.
Following completion of EBUS bronchoscopy the video bronchoscope was re-inserted and blood was suctioned from the airway.
The bronchoscope was removed and procedure completed. 

Complications: No immediate complications
Estimated Blood Loss: 10 cc.
Post Procedure Diagnosis:
- Technically successful flexible bronchoscopy with endobronchial ultrasound-guided biopsies.
- The patient has remained stable and has been transferred in good condition to the post-surgical monitoring unit.
- Will await final pathology results"""

# 2. Configuration (Procedure Flags)
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 3. Data Definition (Spans)
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("EBUS bronchoscopy", "PROC_METHOD", "linear_ebus", "Procedure Performed: ", "evt_01"),
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190 video bronchoscope", "tracheobronchial tree, the ", "evt_01"),
    ("tracheobronchial tree", "ANAT_AIRWAY", "tracheobronchial tree", "advanced to the ", "evt_01"),
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "UC180F convex probe EBUS bronchoscope", "removed and the ", "evt_02"),
    ("station 11Ri lymph node", "ANAT_LN_STATION", "11Ri", "radiographically suspicious ", "evt_02"),
    ("transbronchial needle aspiration", "PROC_METHOD", "tbna_conventional", "Sampling by ", "evt_02"),
    ("Olympus EBUS-TBNA 22 gauge needle", "DEV_NEEDLE", "Olympus EBUS-TBNA 22 gauge needle", "beginning with the ", "evt_02"),
    ("malignancy", "OBS_ROSE", "malignancy", "evaluation read as ", "evt_02"),
    ("Technically successful", "OUTCOME_COMPLICATION", "None", "Diagnosis:\n- ", "evt_03"),
]

# 4. Event Definitions
EVENTS = [
    {
        "event_id": "evt_01",
        "type": "diagnostic_bronchoscopy",
        "description": "Initial inspection with video bronchoscope",
        "devices": ["Q190 video bronchoscope"],
        "anatomy": ["tracheobronchial tree"],
        "outcomes": []
    },
    {
        "event_id": "evt_02",
        "type": "linear_ebus",
        "description": "EBUS-TBNA of station 11Ri",
        "devices": ["UC180F convex probe EBUS bronchoscope", "Olympus EBUS-TBNA 22 gauge needle"],
        "anatomy": ["station 11Ri lymph node"],
        "outcomes": ["ROSE: malignancy"]
    },
    {
        "event_id": "evt_03",
        "type": "post_procedure",
        "description": "Post procedure status",
        "devices": [],
        "anatomy": [],
        "outcomes": ["Technically successful"]
    }
]

# 5. Helper Functions
def clean_text(text):
    if not text:
        return ""
    return re.sub(r'[\r\n]+', ' ', text).strip()

def hydrate_span(full_text, span_text, context_prefix):
    """
    Finds the start and end indices of span_text within full_text,
    using context_prefix to locate the correct occurrence.
    """
    # Normalize spaces for searching
    clean_full = full_text.replace('\r', '').replace('\n', ' ')
    clean_span = span_text.replace('\r', '').replace('\n', ' ')
    clean_context = context_prefix.replace('\r', '').replace('\n', ' ')
    
    # Construct search pattern
    # We look for context followed by span
    pattern = re.escape(clean_context) + r'\s*' + re.escape(clean_span)
    match = re.search(pattern, clean_full)
    
    if match:
        # The span starts after the context
        # We need to find the exact start/end in the original string (ignoring newline diffs is tricky)
        # Simplified approach: Find the string in the original text that matches
        
        # Regex search in original text (allowing for newlines)
        search_pattern = re.escape(context_prefix) + r'[\s\r\n]*' + re.escape(span_text)
        # Relax the pattern slightly for whitespace differences
        search_pattern = search_pattern.replace(r'\ ', r'[\s\r\n]+')
        
        orig_match = re.search(search_pattern, full_text, re.DOTALL)
        if orig_match:
            full_match_str = orig_match.group(0)
            # context length in the match
            # finding where the span text actually starts inside the match
            # A robust way: span_start = match_start + (length of match - length of span)
            # BUT, we might have variable whitespace between context and span
            
            span_start = orig_match.end() - len(span_text)
            # This assumes exact length match at end, which is usually true for exact text
            return span_start, orig_match.end()
            
    # Fallback: simple find if unique
    start = full_text.find(span_text)
    if start != -1 and full_text.count(span_text) == 1:
        return start, start + len(span_text)
        
    return -1, -1

# 6. Workbook Generation Function
def generate_workbook():
    # Create workbook
    wb = openpyxl.Workbook()
    
    # --- Sheet 1: Note_Text ---
    ws_text = wb.active
    ws_text.title = "Note_Text"
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # --- Sheet 2: Note_Index ---
    ws_index = wb.create_sheet("Note_Index")
    headers_index = ["NOTE_ID", "procedure_date"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers_index)
    row_index = [NOTE_ID, PROCEDURE_DATE] + [PROCEDURE_FLAGS[k] for k in PROCEDURE_FLAGS]
    ws_index.append(row_index)
    
    # --- Sheet 3: Span_Annotations ---
    ws_annot = wb.create_sheet("Span_Annotations")
    headers_annot = ["NOTE_ID", "start_char", "end_char", "text", "label", "normalized_value", "context_prefix", "event_id"]
    ws_annot.append(headers_annot)
    
    # We will fill this using the SPANS list, but start/end will be blank here?
    # Usually populate_phase0 expects raw data here. We will just dump what we have.
    for span in SPANS:
        ws_annot.append([NOTE_ID, "", "", span[0], span[1], span[2], span[3], span[4]])

    # --- Sheet 4: Span_Hydrated ---
    ws_hydrated = wb.create_sheet("Span_Hydrated")
    ws_hydrated.append(headers_annot)
    
    for span in SPANS:
        s_text, s_label, s_norm, s_context, s_evt = span
        start, end = hydrate_span(NOTE_TEXT, s_text, s_context)
        if start == -1:
            print(f"Warning: Could not hydrate span '{s_text}' with context '{s_context}'")
        ws_hydrated.append([NOTE_ID, start, end, s_text, s_label, s_norm, s_context, s_evt])

    # --- Sheet 5: Event_Log ---
    ws_log = wb.create_sheet("Event_Log")
    headers_log = ["NOTE_ID", "event_id", "event_type", "description", "devices", "anatomy", "outcomes"]
    ws_log.append(headers_log)
    
    for evt in EVENTS:
        ws_log.append([
            NOTE_ID,
            evt["event_id"],
            evt["type"],
            evt["description"],
            ", ".join(evt["devices"]),
            ", ".join(evt["anatomy"]),
            ", ".join(evt["outcomes"])
        ])

    # --- Sheet 6: V3_Procedure_Events ---
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    headers_v3 = ["NOTE_ID", "procedure_date", "event_id", "event_type", "event_details_json"]
    ws_v3.append(headers_v3)
    
    for evt in EVENTS:
        details = {
            "description": evt["description"],
            "devices": evt["devices"],
            "anatomy": evt["anatomy"],
            "outcomes": evt["outcomes"]
        }
        ws_v3.append([NOTE_ID, PROCEDURE_DATE, evt["event_id"], evt["type"], json.dumps(details)])

    # --- Sheet 7: V3_Registry_JSON ---
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["NOTE_ID", "json_output"])
    
    registry_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source_file": SOURCE_FILE,
            "procedure_date": PROCEDURE_DATE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS,
        "spans": [
            {
                "text": s[0],
                "label": s[1],
                "normalized_value": s[2],
                "start": hydrate_span(NOTE_TEXT, s[0], s[3])[0],
                "end": hydrate_span(NOTE_TEXT, s[0], s[3])[1]
            }
            for s in SPANS
        ]
    }
    
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Workbook generated at: {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()