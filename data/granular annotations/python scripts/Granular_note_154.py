import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# =============================================================================
# 1. IMPORTS & CONSTANTS
# =============================================================================

NOTE_ID = "note_154"
SOURCE_FILE = "note_154.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_154 SOURCE_FILE: note_154.txt Indications: right lower lobe nodule 
Medications: Propofol infusion via anesthesia assistance  
Medications: General Anesthesia,
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention.
Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway was in good position. The vocal cords were normal. The subglottic space was normal.
The trachea was of normal caliber. The carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level. Bronchial mucosa and anatomy were normal;
there are no endobronchial lesions. We then removed the diagnostic Q190 bronchoscopy and the super-dimension navigational catheter was inserted through the T190 therapeutic bronchoscope and advanced into the airway.
Using navigational map we attempted to advance the 90 degree edge catheter into the proximity of the lesion within the right lower lobe.
Confirmation of placement once at the point of interest with radial ultrasound showed a concentric view within the lesion.
Biopsies were then performed with a variety of instruments to include peripheral needle, triple needle brush and forceps, under fluoroscopic visualization.
After adequate samples were obtained the bronchoscope was removed and the procedure completed
Complications: No immediate complications
Estimated Blood Loss: Less than 5 cc.
Post Procedure Diagnosis:
- Flexible bronchoscopy with successful navigational biopsy of right lower lobe nodule.  
Recommendations
- Await final pathology"""

# =============================================================================
# 2. CONFIGURATION (PROCEDURE FLAGS)
# =============================================================================

PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 1,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 1,
    "navigational_bronchoscopy": 1,
    "transbronchial_biopsy": 1,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# =============================================================================
# 3. DATA DEFINITION (SPANS)
# =============================================================================
# Format: (span_text, label, normalized_value, context_prefix, event_id)

SPANS = [
    # Event 01: Initial Inspection (Diagnostic)
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190 Bronchoscope", "tracheobronchial tree, the", "evt_01"),
    ("laryngeal mask airway", "DEV_INSTRUMENT", "LMA", "mouth, via", "evt_01"),
    ("tracheobronchial tree", "ANAT_AIRWAY", "Tracheobronchial Tree", "advanced to the", "evt_01"),
    ("normal", "OBS_ROSE", "Normal", "Bronchial mucosa and anatomy were", "evt_01"),
    ("no endobronchial lesions", "OBS_LESION", "None", "there are", "evt_01"),

    # Event 02: Navigation & Setup
    ("super-dimension navigational catheter", "DEV_CATHETER", "SuperDimension Catheter", "bronchoscopy and the", "evt_02"),
    ("T190 therapeutic bronchoscope", "DEV_INSTRUMENT", "T190 Bronchoscope", "inserted through the", "evt_02"),
    ("90 degree edge catheter", "DEV_CATHETER", "Edge Catheter 90 Degree", "advance the", "evt_02"),
    ("right lower lobe", "ANAT_LUNG_LOC", "RLL", "lesion within the", "evt_02"),

    # Event 03: Radial EBUS
    ("radial ultrasound", "PROC_METHOD", "Radial EBUS", "interest with", "evt_03"),
    ("concentric view", "OBS_LESION", "Concentric View", "showed a", "evt_03"),

    # Event 04: Biopsy
    ("Biopsies", "PROC_METHOD", "Biopsy", "within the lesion.\n", "evt_04"),
    ("peripheral needle", "DEV_NEEDLE", "Peripheral Needle", "to include", "evt_04"),
    ("triple needle brush", "DEV_INSTRUMENT", "Triple Needle Brush", "peripheral needle,", "evt_04"),
    ("forceps", "DEV_INSTRUMENT", "Forceps", "brush and", "evt_04"),

    # Outcomes
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications:", "evt_global"),
    ("successful navigational biopsy", "OUTCOME_PLEURAL", "Successful Biopsy", "bronchoscopy with", "evt_global")
]

# =============================================================================
# 4. EVENT DEFINITIONS
# =============================================================================

EVENTS = [
    {
        "event_id": "evt_01",
        "type": "diagnostic_bronchoscopy",
        "method": "White Light",
        "anatomy": "Tracheobronchial Tree",
        "devices": ["Q190 Bronchoscope", "LMA"],
        "outcomes": ["Normal Mucosa", "No Lesions"]
    },
    {
        "event_id": "evt_02",
        "type": "navigational_bronchoscopy",
        "method": "Electromagnetic Navigation",
        "anatomy": "Right Lower Lobe",
        "devices": ["T190 Bronchoscope", "SuperDimension Catheter", "Edge Catheter 90 Degree"],
        "outcomes": ["Target Reached"]
    },
    {
        "event_id": "evt_03",
        "type": "radial_ebus",
        "method": "Radial Ultrasound",
        "anatomy": "Right Lower Lobe",
        "devices": ["Radial EBUS Probe"],
        "outcomes": ["Concentric View"]
    },
    {
        "event_id": "evt_04",
        "type": "transbronchial_biopsy",
        "method": "Biopsy",
        "anatomy": "Right Lower Lobe",
        "devices": ["Peripheral Needle", "Triple Needle Brush", "Forceps"],
        "outcomes": []
    }
]

# =============================================================================
# 5. HELPER FUNCTIONS
# =============================================================================

def clean_text(text):
    if not text:
        return ""
    return str(text).strip().replace('\r\n', '\n').replace('\r', '\n')

def find_offsets(full_text, span_text, context_prefix=""):
    """
    Locates the start and end character offsets of span_text within full_text.
    Uses context_prefix to disambiguate identical strings.
    """
    clean_full = clean_text(full_text)
    clean_span = clean_text(span_text)
    clean_context = clean_text(context_prefix)

    if not clean_span:
        return None, None

    # Construct regex pattern
    # Escape special regex chars in span and context
    esc_span = re.escape(clean_span)
    
    if clean_context:
        esc_context = re.escape(clean_context)
        # Look for context followed by potential whitespace then span
        pattern = f"({esc_context})\s*({esc_span})"
        match = re.search(pattern, clean_full, re.IGNORECASE)
        if match:
            # Group 1 is context, Group 2 is span
            # We want offsets of Group 2
            start = match.start(2)
            end = match.end(2)
            return start, end
    
    # Fallback: simple search if no context or context failed
    # Finds first occurrence
    match = re.search(esc_span, clean_full, re.IGNORECASE)
    if match:
        return match.start(), match.end()

    return None, None

def hydrate_span(full_text, span_tuple):
    """
    Takes a span tuple from SPANS and returns a dictionary with 
    calculated offsets and formatting.
    """
    text_content, label, norm, context, event_id = span_tuple
    start, end = find_offsets(full_text, text_content, context)
    
    return {
        "event_id": event_id,
        "label": label,
        "text": text_content,
        "start": start,
        "end": end,
        "normalized": norm
    }

# =============================================================================
# 6. WORKBOOK GENERATION
# =============================================================================

def generate_workbook():
    # A. Initialize Workbook
    # If template doesn't exist, create a basic one in memory (fallback)
    if os.path.exists(TEMPLATE_PATH):
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    else:
        wb = openpyxl.Workbook()
        # Create expected sheets if they don't exist
        expected_sheets = ["Note_Text", "Note_Index", "Span_Annotations", 
                           "Span_Hydrated", "Event_Log", "V3_Procedure_Events", 
                           "V3_Registry_JSON"]
        for sheet_name in expected_sheets:
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
        # Remove default sheet if present
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # B. Sheet 1: Note_Text
    ws_text = wb["Note_Text"]
    # Clear existing data (rows > 1)
    for row in ws_text.iter_rows(min_row=2, max_col=3):
        for cell in row:
            cell.value = None
    
    # Write header if missing
    if ws_text["A1"].value != "note_id":
        ws_text.append(["note_id", "source_file", "note_text"])
    
    # Write Data
    ws_text.cell(row=2, column=1, value=NOTE_ID)
    ws_text.cell(row=2, column=2, value=SOURCE_FILE)
    ws_text.cell(row=2, column=3, value=NOTE_TEXT)

    # C. Sheet 2: Note_Index
    ws_index = wb["Note_Index"]
    # Headers: note_id, source_file, procedure_date, [30 flags]
    flag_keys = list(PROCEDURE_FLAGS.keys())
    headers = ["note_id", "source_file", "procedure_date"] + flag_keys
    
    # Reset/Write Header
    ws_index.delete_rows(1, ws_index.max_row)
    ws_index.append(headers)
    
    # Write Data
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE]
    for k in flag_keys:
        row_data.append(PROCEDURE_FLAGS[k])
    ws_index.append(row_data)

    # D. Sheet 3: Span_Annotations (Pre-hydration placeholder)
    ws_annot = wb["Span_Annotations"]
    annot_headers = ["note_id", "event_id", "label", "span_text", "start_char", "end_char", "normalized_value"]
    ws_annot.delete_rows(1, ws_annot.max_row)
    ws_annot.append(annot_headers)
    
    for s in SPANS:
        # (text, label, norm, context, event_id)
        ws_annot.append([NOTE_ID, s[4], s[1], s[0], "", "", s[2]])

    # E. Sheet 4: Span_Hydrated (Calculated Offsets)
    ws_hydra = wb["Span_Hydrated"]
    ws_hydra.delete_rows(1, ws_hydra.max_row)
    ws_hydra.append(annot_headers)
    
    for s in SPANS:
        h = hydrate_span(NOTE_TEXT, s)
        ws_hydra.append([
            NOTE_ID, 
            h["event_id"], 
            h["label"], 
            h["text"], 
            h["start"], 
            h["end"], 
            h["normalized"]
        ])

    # F. Sheet 5: Event_Log (Flattened)
    ws_log = wb["Event_Log"]
    log_headers = ["note_id", "event_id", "procedure_type", "anatomy", "devices", "outcomes"]
    ws_log.delete_rows(1, ws_log.max_row)
    ws_log.append(log_headers)
    
    for evt in EVENTS:
        ws_log.append([
            NOTE_ID,
            evt["event_id"],
            evt["type"],
            evt["anatomy"],
            ", ".join(evt["devices"]),
            ", ".join(evt["outcomes"])
        ])

    # G. Sheet 6: V3_Procedure_Events (JSON logic)
    ws_v3 = wb["V3_Procedure_Events"]
    v3_headers = ["note_id", "event_order", "event_id", "event_json"]
    ws_v3.delete_rows(1, ws_v3.max_row)
    ws_v3.append(v3_headers)
    
    for idx, evt in enumerate(EVENTS):
        ws_v3.append([
            NOTE_ID,
            idx + 1,
            evt["event_id"],
            json.dumps(evt)
        ])

    # H. Sheet 7: V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    ws_json.delete_rows(1, ws_json.max_row)
    ws_json.append(["note_id", "full_registry_json"])
    
    # Construct Master JSON
    master_json = {
        "metadata": {
            "note_id": NOTE_ID,
            "source_file": SOURCE_FILE,
            "procedure_date": PROCEDURE_DATE
        },
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS
    }
    
    ws_json.append([NOTE_ID, json.dumps(master_json, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Successfully generated: {OUTPUT_PATH}")

# =============================================================================
# 7. EXECUTION BLOCK
# =============================================================================

if __name__ == "__main__":
    generate_workbook()