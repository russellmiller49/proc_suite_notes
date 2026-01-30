import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# 1. Imports & Constants
NOTE_ID = "note_155"
SOURCE_FILE = "note_155.txt"
PROCEDURE_DATE = "" 
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_155 SOURCE_FILE: note_155.txt Procedure Name: EBUS bronchoscopy, peripheral bronchoscopy
Indications: Pulmonary nodule requiring diagnosis/staging.
Medications: Propofol infusion via anesthesia assistance  
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention. 
Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway was in good position. The vocal cords appeared normal. The subglottic space was normal.
The trachea was of normal caliber. The carina was sharp.
The tracheobronchial tree was examined to at least the first sub-segmental level without endobronchial lesions visualized.
The video bronchoscope was then removed and the UC180F convex probe EBUS bronchoscope was introduced through the mouth, and advanced to the tracheobronchial tree.
A systematic hilar and mediastinal lymph node survey was carried out.
Sampling criteria (5mm short axis diameter) was met in station 11RS and 7. Sampling by transbronchial needle aspiration was performed using an Olympus EBUSTBNA 22 gauge needle.
Further details regarding nodal size and number of samples are included in the EBUS procedural sheet in AHLTA.
All samples were sent for routine cytology. Onsite path evaluation did not identify malignancy.
The bronchoscope was then removed and the T190 therapeutic video bronchoscope was inserted into the airway and based on anatomical knowledge advanced into the left upper lobe and a large sheath catheter with radial ultrasound to the area of known nodule and a concentric view of the lesion was identified with the radial EBUS.
Biopsies were then performed with a variety of instruments to include peripheral needle forceps and brush with fluoroscopic guidance through the sheath.
After adequate samples were obtained the bronchoscope was removed. ROSE did not identify malignancy on preliminary samples.
The bronchoscope was then removed and the Q190 re-inserted into the airways.
To remove secretions and blood and once confident that there was no active bleeding the bronchoscope was removed and the procedure completed.
Complications: 	
-None 
Estimated Blood Loss:  10 cc.
Recommendations:
- Admit for overnight observation given the late hour of procedural completion
- Await biopsy results 
- Discharge home once criteria met."""

# 2. Configuration (Procedure Flags)
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 1,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 1,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 1,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 3. Data Definition (Spans)
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190", "tracheobronchial tree, the ", "evt_01"),
    ("tracheobronchial tree", "ANAT_AIRWAY", "Tracheobronchial Tree", "advanced to the ", "evt_01"),
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "UC180F", "removed and the ", "evt_02"),
    ("station 11RS", "ANAT_LN_STATION", "11RS", "met in ", "evt_02"),
    ("7", "ANAT_LN_STATION", "7", "station 11RS and ", "evt_02"),
    ("Olympus EBUSTBNA 22 gauge needle", "DEV_NEEDLE", "22G", "using an ", "evt_02"),
    ("Onsite path evaluation did not identify malignancy", "OBS_ROSE", "Benign", "cytology. ", "evt_02"),
    ("T190 therapeutic video bronchoscope", "DEV_INSTRUMENT", "T190", "removed and the ", "evt_03"),
    ("left upper lobe", "ANAT_LUNG_LOC", "LUL", "advanced into the ", "evt_03"),
    ("sheath catheter", "DEV_CATHETER", "Guide Sheath", "lobe and a large ", "evt_03"),
    ("radial ultrasound", "PROC_METHOD", "Radial EBUS", "catheter with ", "evt_03"),
    ("peripheral needle", "DEV_NEEDLE", "Needle", "include ", "evt_03"),
    ("forceps", "DEV_INSTRUMENT", "Forceps", "peripheral needle ", "evt_03"),
    ("brush", "DEV_INSTRUMENT", "Brush", "forceps and ", "evt_03"),
    ("ROSE did not identify malignancy", "OBS_ROSE", "Benign", "removed. ", "evt_03"),
    ("None", "OUTCOME_COMPLICATION", "None", "Complications: \t\n-", "evt_global")
]

# 4. Event Definitions
EVENTS = [
    {
        "event_id": "evt_01",
        "type": "diagnostic_bronchoscopy",
        "method": "White Light Bronchoscopy",
        "anatomy": "Tracheobronchial Tree",
        "devices": ["Q190 video bronchoscope"],
        "outcomes": []
    },
    {
        "event_id": "evt_02",
        "type": "linear_ebus",
        "method": "Linear EBUS",
        "anatomy": "Station 11RS, Station 7",
        "devices": ["UC180F convex probe EBUS bronchoscope", "Olympus EBUSTBNA 22 gauge needle"],
        "outcomes": ["ROSE: Benign"]
    },
    {
        "event_id": "evt_03",
        "type": "radial_ebus",
        "method": "Radial EBUS, Transbronchial Biopsy, Brushings",
        "anatomy": "Left Upper Lobe",
        "devices": ["T190 therapeutic video bronchoscope", "Sheath Catheter", "Radial Ultrasound", "Peripheral Needle", "Forceps", "Brush"],
        "outcomes": ["ROSE: Benign"]
    },
    {
        "event_id": "evt_global",
        "type": "outcome",
        "method": "N/A",
        "anatomy": "N/A",
        "devices": [],
        "outcomes": ["Complications: None"]
    }
]

# 5. Helper Functions
def clean_text(text):
    if not text: return ""
    return text.strip().replace('\r', '')

def hydrate_span(full_text, span_text, context_prefix):
    cleaned_full = clean_text(full_text)
    cleaned_span = clean_text(span_text)
    cleaned_prefix = clean_text(context_prefix)
    
    if not cleaned_span: return None, None
    
    # Attempt to find with prefix
    search_pattern = re.escape(cleaned_prefix) + r"\s*" + re.escape(cleaned_span)
    match = re.search(search_pattern, cleaned_full, re.IGNORECASE)
    
    if match:
        # Calculate start relative to the prefix end (approximate for span only)
        # Better: find prefix start, then add prefix len
        start_index = match.start() + len(cleaned_prefix)
        # Adjust for potential whitespace between prefix and span in regex
        # Find exact span start within match
        span_start_in_match = match.group(0).lower().find(cleaned_span.lower())
        real_start = match.start() + span_start_in_match
        real_end = real_start + len(cleaned_span)
        return real_start, real_end
    
    # Fallback: simple find (risky for duplicates)
    start_index = cleaned_full.lower().find(cleaned_span.lower())
    if start_index != -1:
        return start_index, start_index + len(cleaned_span)
        
    return None, None

def generate_workbook():
    wb = Workbook()
    
    # --- Sheet 1: Note_Text ---
    ws_text = wb.active
    ws_text.title = "Note_Text"
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # --- Sheet 2: Note_Index ---
    ws_index = wb.create_sheet("Note_Index")
    headers = ["NOTE_ID", "SOURCE_FILE", "PROCEDURE_DATE"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers)
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE] + list(PROCEDURE_FLAGS.values())
    ws_index.append(row_data)
    
    # --- Sheet 3: Span_Annotations ---
    ws_anno = wb.create_sheet("Span_Annotations")
    ws_anno.append(["NOTE_ID", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "START_CHAR", "END_CHAR", "EVENT_ID"])
    
    for span in SPANS:
        # span = (text, label, norm, prefix, evt)
        ws_anno.append([NOTE_ID, span[0], span[1], span[2], span[3], "", "", span[4]])
        
    # --- Sheet 4: Span_Hydrated ---
    ws_hydra = wb.create_sheet("Span_Hydrated")
    ws_hydra.append(["NOTE_ID", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "START_CHAR", "END_CHAR", "EVENT_ID"])
    
    for span in SPANS:
        start, end = hydrate_span(NOTE_TEXT, span[0], span[3])
        ws_hydra.append([NOTE_ID, span[0], span[1], span[2], span[3], start, end, span[4]])
        
    # --- Sheet 5: Event_Log ---
    ws_log = wb.create_sheet("Event_Log")
    ws_log.append(["NOTE_ID", "EVENT_ID", "EVENT_TYPE", "METHOD", "ANATOMY", "DEVICES", "OUTCOMES"])
    
    for evt in EVENTS:
        ws_log.append([
            NOTE_ID,
            evt["event_id"],
            evt["type"],
            evt["method"],
            evt["anatomy"],
            ", ".join(evt["devices"]),
            ", ".join(evt["outcomes"])
        ])

    # --- Sheet 6: V3_Procedure_Events ---
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    ws_v3.append(["NOTE_ID", "EVENT_ID", "EVENT_TYPE", "METHOD", "ANATOMY", "DEVICES", "OUTCOMES_JSON", "RAW_SPAN_DATA_JSON"])
    
    for evt in EVENTS:
        # Filter spans for this event
        evt_spans = [s for s in SPANS if s[4] == evt["event_id"]]
        span_data = [{"text": s[0], "label": s[1], "norm": s[2]} for s in evt_spans]
        
        ws_v3.append([
            NOTE_ID,
            evt["event_id"],
            evt["type"],
            evt["method"],
            evt["anatomy"],
            json.dumps(evt["devices"]),
            json.dumps(evt["outcomes"]),
            json.dumps(span_data)
        ])
        
    # --- Sheet 7: V3_Registry_JSON ---
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["NOTE_ID", "JSON_PAYLOAD"])
    
    payload = {
        "metadata": {
            "note_id": NOTE_ID,
            "source_file": SOURCE_FILE,
            "procedure_date": PROCEDURE_DATE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS,
        "spans": [{"text": s[0], "label": s[1], "norm": s[2], "prefix": s[3], "event_id": s[4]} for s in SPANS]
    }
    
    ws_json.append([NOTE_ID, json.dumps(payload, indent=2)])
    
    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()