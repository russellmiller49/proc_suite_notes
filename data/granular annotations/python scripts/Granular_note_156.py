import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# 1. Imports & Constants
NOTE_ID = "note_156"
SOURCE_FILE = "note_156.txt"
PROCEDURE_DATE = ""
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"
NOTE_TEXT = """NOTE_ID:  note_156 SOURCE_FILE: note_156.txt Indications: Mediastinal adenopathy
Procedure: EBUS bronchoscopy – single station
Medications: General Anesthesia
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention.
Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway was in good position. The vocal cords appeared normal. The subglottic space was normal.
The trachea was of normal caliber. The carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level. Bronchial mucosa and anatomy were normal;
there are no endobronchial lesions, and no secretions. The video bronchoscope was then removed and the UC180F convex probe EBUS bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
A large, station 7, lymph node was identified and sampling by transbronchial needle aspiration was performed with the Olympus 22G EBUS-TBNA needles with a total of 7 passes performed.
Rapid onsite pathological evaluation showed malignancy. Samples were sent for both flow and routine cytology.
Following completion of EBUS bronchoscopy, the Q190 video bronchoscope was then re-inserted and DECAMP research bronchoscopy was performed with endobronchial forceps biopsy of the right upper lobe, right middle lobe and left upper lobe as well as endobronchial brushing of bronchus intermedius.
After suctioning blood and secretions there was no evidence of active bleeding and the bronchoscope was subsequently removed.
Complications: No immediate complications
Estimated Blood Loss: 5cc
Post Procedure Diagnosis:
- Technically successful flexible bronchoscopy with endobronchial ultrasound-guided biopsies and DECAMP research protocol.
- The patient has remained stable and has been transferred in good condition to the post-procedural monitoring unit.
- Will await final pathology results"""

# 2. Configuration (Procedure Flags)
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 1,
    "endobronchial_biopsy": 1,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 3. Data Definition (Spans)
SPANS = [
    # Event 01: Initial Inspection
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190", "tracheobronchial tree, the ", "evt_01"),
    ("tracheobronchial tree", "ANAT_AIRWAY", "Tracheobronchial Tree", "advanced to the ", "evt_01"),
    
    # Event 02: EBUS TBNA
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "UC180F", "removed and the ", "evt_02"),
    ("station 7", "ANAT_LN_STATION", "7", "A large, ", "evt_02"),
    ("transbronchial needle aspiration", "PROC_ACTION", "TBNA", "sampling by ", "evt_02"),
    ("Olympus 22G EBUS-TBNA needles", "DEV_NEEDLE", "22G", "with the ", "evt_02"),
    ("7 passes", "MEAS_COUNT", "7", "total of ", "evt_02"),
    ("malignancy", "OBS_ROSE", "Malignant", "evaluation showed ", "evt_02"),

    # Event 03: Research Biopsies/Brushing
    ("endobronchial forceps biopsy", "PROC_ACTION", "Endobronchial Biopsy", "performed with ", "evt_03"),
    ("right upper lobe", "ANAT_LUNG_LOC", "RUL", "biopsy of the ", "evt_03"),
    ("right middle lobe", "ANAT_LUNG_LOC", "RML", "right upper lobe, ", "evt_03"),
    ("left upper lobe", "ANAT_LUNG_LOC", "LUL", "middle lobe and ", "evt_03"),
    ("endobronchial brushing", "PROC_ACTION", "Bronchial Brushing", "as well as ", "evt_03"),
    ("bronchus intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "brushing of ", "evt_03"),

    # Outcomes
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications: ", "evt_04"),
]

# 4. Event Definitions
EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Bronchoscopy",
        "action": "Inspection",
        "anatomy": "Tracheobronchial Tree",
        "devices": ["Q190"],
        "outcomes": []
    },
    {
        "event_id": "evt_02",
        "procedure_type": "Bronchoscopy",
        "action": "TBNA",
        "anatomy": "Station 7",
        "devices": ["UC180F", "Olympus 22G EBUS-TBNA needles"],
        "outcomes": ["ROSE: Malignancy"]
    },
    {
        "event_id": "evt_03",
        "procedure_type": "Bronchoscopy",
        "action": "Endobronchial Biopsy/Brushing",
        "anatomy": "RUL, RML, LUL, Bronchus Intermedius",
        "devices": ["Q190", "Forceps", "Brush"],
        "outcomes": []
    },
    {
        "event_id": "evt_04",
        "procedure_type": "Outcome",
        "action": "Post-op",
        "anatomy": "",
        "devices": [],
        "outcomes": ["No immediate complications"]
    }
]

# 5. Helper Functions
def clean_text(text):
    if not text:
        return ""
    return re.sub(r'[\r\n]+', ' ', text).strip()

def hydrate_span(full_text, span_text, context_prefix):
    """
    Finds the start and end indices of span_text in full_text.
    Uses context_prefix to disambiguate if span_text appears multiple times.
    """
    cleaned_full = clean_text(full_text)
    cleaned_span = clean_text(span_text)
    cleaned_context = clean_text(context_prefix)
    
    if not cleaned_span:
        return None, None

    # Search pattern: context + optional spaces + span
    # We escape them to handle special regex chars
    pattern_str = re.escape(cleaned_context) + r"\s*" + re.escape(cleaned_span)
    match = re.search(pattern_str, cleaned_full, re.IGNORECASE)
    
    if match:
        # The match includes the context. We want just the span part.
        # Start index of the span is the end of the context match in the group?
        # A simpler way: find the context, then find the span immediately after.
        
        # Absolute start of the match
        match_start = match.start()
        
        # Calculate where the span actually begins within that match
        # It equals match_start + length_of_context_part + length_of_intervening_spaces
        # We can just look for the span_text inside the matched string to get relative offset
        matched_string = match.group(0)
        span_start_relative = matched_string.lower().rfind(cleaned_span.lower())
        
        start_index = match_start + span_start_relative
        end_index = start_index + len(cleaned_span)
        return start_index, end_index
    
    # Fallback: exact match without context (if unique)
    start_index = cleaned_full.lower().find(cleaned_span.lower())
    if start_index != -1:
        # Check if it's unique
        if cleaned_full.lower().count(cleaned_span.lower()) == 1:
            return start_index, start_index + len(cleaned_span)
            
    return None, None

def generate_workbook():
    # Create workbook
    wb = openpyxl.Workbook()
    
    # --- Sheet 1: Note_Text ---
    ws_text = wb.active
    ws_text.title = "Note_Text"
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # --- Sheet 2: Note_Index ---
    ws_index = wb.create_sheet("Note_Index")
    headers = ["NOTE_ID", "SOURCE_FILE", "PROCEDURE_DATE"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers)
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE] + [PROCEDURE_FLAGS[k] for k in PROCEDURE_FLAGS]
    ws_index.append(row_data)
    
    # --- Sheet 3: Span_Annotations ---
    ws_anno = wb.create_sheet("Span_Annotations")
    ws_anno.append(["NOTE_ID", "SPAN_TEXT", "LABEL", "VALUE", "CONTEXT_PREFIX", "START_CHAR", "END_CHAR", "EVENT_ID"])
    
    for span_data in SPANS:
        span_text, label, val, ctx, evt_id = span_data
        ws_anno.append([NOTE_ID, span_text, label, val, ctx, "", "", evt_id])
        
    # --- Sheet 4: Span_Hydrated ---
    ws_hydra = wb.create_sheet("Span_Hydrated")
    ws_hydra.append(["NOTE_ID", "SPAN_TEXT", "LABEL", "VALUE", "CONTEXT_PREFIX", "START_CHAR", "END_CHAR", "EVENT_ID"])
    
    for span_data in SPANS:
        span_text, label, val, ctx, evt_id = span_data
        start, end = hydrate_span(NOTE_TEXT, span_text, ctx)
        ws_hydra.append([NOTE_ID, span_text, label, val, ctx, start, end, evt_id])

    # --- Sheet 5: Event_Log ---
    ws_elog = wb.create_sheet("Event_Log")
    ws_elog.append(["NOTE_ID", "EVENT_ID", "PROCEDURE_TYPE", "ACTION", "ANATOMY", "DEVICES", "OUTCOMES"])
    
    for evt in EVENTS:
        ws_elog.append([
            NOTE_ID,
            evt["event_id"],
            evt["procedure_type"],
            evt["action"],
            evt["anatomy"],
            ", ".join(evt["devices"]),
            ", ".join(evt["outcomes"])
        ])

    # --- Sheet 6: V3_Procedure_Events ---
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    ws_v3.append(["NOTE_ID", "EVENT_ID", "JSON_PAYLOAD"])
    
    for evt in EVENTS:
        payload = json.dumps(evt)
        ws_v3.append([NOTE_ID, evt["event_id"], payload])
        
    # --- Sheet 7: V3_Registry_JSON ---
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["NOTE_ID", "FULL_JSON"])
    
    registry_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source": SOURCE_FILE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS
    }
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])
    
    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()