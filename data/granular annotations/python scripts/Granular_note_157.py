import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# ==============================================================================
# 1. IMPORTS & CONSTANTS
# ==============================================================================

NOTE_ID = "note_157"
SOURCE_FILE = "note_157.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_157 SOURCE_FILE: note_157.txt Indications: Left lower lobe nodule
Medications: General Anesthesia,
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention. 
Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the 7.5 ETT.
The airways were friable and significant blood was present from difficult intubation.  The trachea was of normal caliber.
The carina was sharp. On the right the patient had an anatomic variant with an accessory airway just distal to the superior segment of the right lower lobe.
The left sided airway anatomy was normal. No evidence of endobronchial disease was seen to at least the first sub-segments.
Following inspection the super-dimension navigational catheter was inserted through the therapeutic bronchoscope and advanced into the airway.
Using navigational map we were able to advance the 180 degree edge catheter into the anterior-medial segment on the left lower lobe and navigated to within 1cm of the lesion.
The navigational probe was then removed and peripheral radial probe was inserted into the catheter to confirm location.
Ultrasound visualization yielded a concentric view affirming the location.  Needle biopsies were performed with fluoroscopic guidance. ROSE was non-diagnostic.
The catheter was repositioned to sample other areas of the lesion and radial probe continued to show a concentric view of the tumor.
ROSE again was read as non-diagnostic. Tissue biopsies were then performed using forceps under fluoroscopic visualization along with brush and triple needle brush biopsies.
After samples were obtained the bronchoscope was removed and the diagnostic bronchoscope was re-inserted into the airway.
Subsequent inspection did not show evidence of active bleeding and the bronchoscope was removed from the airway.
Following completion of the procedure the patient was noted to lack audible airleak when the ETT balloon was deflated.
Given the difficult intubation, anesthesia decided to leave patient intubated and transfer to the ICU to monitor and extubate when deemed appropriate.
Complications: None
Estimated Blood Loss: Less than 10 cc.
Post Procedure Diagnosis:
- Flexible bronchoscopy with successful biopsy of left lower lobe pulmonary nodule
- Will transfer to the ICU and attempt extubation later today.
- Await final pathology"""

# ==============================================================================
# 2. CONFIGURATION (PROCEDURE FLAGS)
# ==============================================================================

PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 1,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0, # Peripheral needle used, but typically mapped to transbronchial if peripheral/navigational
    "linear_ebus": 0,
    "radial_ebus": 1,
    "navigational_bronchoscopy": 1,
    "transbronchial_biopsy": 1, # Forceps and needle used on peripheral nodule
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# ==============================================================================
# 3. DATA DEFINITION (SPANS)
# ==============================================================================

# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    # Event 01: Indications & Prep
    ("Left lower lobe nodule", "OBS_LESION", "Left lower lobe nodule", "Indications: ", "evt_01"),
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190 video bronchoscope", "tracheobronchial tree, the ", "evt_01"),
    ("introduced through the 7.5 ETT", "PROC_METHOD", "via ETT", "bronchoscope was ", "evt_01"),
    
    # Event 02: Inspection / Anatomy
    ("airways were friable", "OBS_LESION", "friable airways", "The ", "evt_02"),
    ("significant blood", "OBS_LESION", "blood in airway", "friable and ", "evt_02"),
    ("right", "LATERALITY", "Right", "On the ", "evt_02"),
    ("accessory airway", "ANAT_AIRWAY", "accessory airway", "variant with an ", "evt_02"),
    ("right lower lobe", "ANAT_LUNG_LOC", "RLL", "segment of the ", "evt_02"),
    ("Left sided airway", "ANAT_AIRWAY", "Left airway", "The ", "evt_02"),
    ("normal", "OBS_LESION", "normal", "anatomy was ", "evt_02"),

    # Event 03: Navigation & EBUS
    ("super-dimension navigational catheter", "DEV_CATHETER", "SuperDimension Navigation", "inspection the ", "evt_03"),
    ("180 degree edge catheter", "DEV_CATHETER", "Edge Catheter 180", "advance the ", "evt_03"),
    ("anterior-medial segment", "ANAT_LUNG_LOC", "Left Lower Lobe Anterior Medial Segment", "into the ", "evt_03"),
    ("left lower lobe", "ANAT_LUNG_LOC", "LLL", "segment on the ", "evt_03"),
    ("peripheral radial probe", "DEV_INSTRUMENT", "Radial EBUS Probe", "removed and ", "evt_03"),
    ("concentric view", "OBS_LESION", "concentric view", "yielded a ", "evt_03"),

    # Event 04: Biopsies (Needle)
    ("Needle biopsies", "PROC_METHOD", "Transbronchial Needle Aspiration", "affirming the location.  ", "evt_04"),
    ("fluoroscopic guidance", "PROC_METHOD", "Fluoroscopy", "performed with ", "evt_04"),
    ("ROSE was non-diagnostic", "OBS_ROSE", "Non-diagnostic", "guidance. ", "evt_04"),

    # Event 05: Biopsies (Forceps/Brush)
    ("Tissue biopsies", "PROC_METHOD", "Transbronchial Biopsy", "non-diagnostic. ", "evt_05"),
    ("forceps", "DEV_INSTRUMENT", "Forceps", "using ", "evt_05"),
    ("brush", "DEV_INSTRUMENT", "Brush", "along with ", "evt_05"),
    ("triple needle brush", "DEV_INSTRUMENT", "Triple Needle Brush", "and ", "evt_05"),
    ("ROSE again was read as non-diagnostic", "OBS_ROSE", "Non-diagnostic", "tumor.\n", "evt_05"),

    # Event 06: Post-Procedure / Complications
    ("active bleeding", "OBS_LESION", "Bleeding", "evidence of ", "evt_06"),
    ("Complications: None", "OUTCOME_COMPLICATION", "None", "\n", "evt_06"),
    ("Estimated Blood Loss: Less than 10 cc", "MEAS_VOL", "10cc", "\n", "evt_06"),
    ("Flexible bronchoscopy", "PROC_METHOD", "Flexible bronchoscopy", "Diagnosis:\n- ", "evt_06"),
    ("successful biopsy", "OUTCOME_SYMPTOMS", "Successful biopsy", "with ", "evt_06"),
]

# ==============================================================================
# 4. EVENT DEFINITIONS
# ==============================================================================

EVENTS = [
    {
        "event_id": "evt_01",
        "name": "Preparation",
        "description": "Indication review and scope introduction",
        "spans": ["Left lower lobe nodule", "Q190 video bronchoscope"]
    },
    {
        "event_id": "evt_02",
        "name": "Inspection",
        "description": "Airway inspection revealing anatomical variant and blood from intubation",
        "spans": ["accessory airway", "right lower lobe", "Left sided airway"]
    },
    {
        "event_id": "evt_03",
        "name": "Navigation & EBUS",
        "description": "Navigation to LLL and Radial EBUS confirmation",
        "spans": ["super-dimension navigational catheter", "180 degree edge catheter", "peripheral radial probe", "concentric view"]
    },
    {
        "event_id": "evt_04",
        "name": "Needle Biopsy",
        "description": "Needle biopsies with non-diagnostic ROSE",
        "spans": ["Needle biopsies", "fluoroscopic guidance", "ROSE was non-diagnostic"]
    },
    {
        "event_id": "evt_05",
        "name": "Tissue Biopsy",
        "description": "Forceps and brush biopsies",
        "spans": ["Tissue biopsies", "forceps", "brush", "triple needle brush"]
    },
    {
        "event_id": "evt_06",
        "name": "Conclusion",
        "description": "Procedure conclusion and outcome summary",
        "spans": ["Complications: None", "successful biopsy"]
    }
]

# ==============================================================================
# 5. HELPER FUNCTIONS
# ==============================================================================

def clean_text(text):
    if not text:
        return ""
    return re.sub(r'[\r\n]+', ' ', text).strip()

def hydrate_span(text, span_text, context_prefix):
    """
    Finds the start and end indices of span_text in text, 
    ensuring it follows context_prefix to disambiguate duplicates.
    """
    # Normalize inputs for search
    search_text = text.replace('\r', '').replace('\n', ' ') 
    # We maintain original text length logic by using the raw text for final indices if possible,
    # but for robust finding, we often need to normalize whitespace. 
    # For Phase 0, exact char match is preferred.
    
    # Strategy: Find context_prefix + span_text
    full_pattern = context_prefix + span_text
    
    # Simple find first
    start_index = text.find(full_pattern)
    
    if start_index == -1:
        # Try looser whitespace matching
        pattern_re = re.escape(context_prefix) + r'\s*' + re.escape(span_text)
        match = re.search(pattern_re, text, re.DOTALL)
        if match:
            # We found the combined block. Now we need to isolate the span part.
            # The span starts after the context prefix.
            # Start of match + length of context (approx)
            # This is tricky with regex. Let's rely on finding the prefix, then looking ahead.
            prefix_match = re.search(re.escape(context_prefix), text)
            if prefix_match:
                search_start = prefix_match.end()
                span_match = text.find(span_text, search_start)
                if span_match != -1:
                    return span_match, span_match + len(span_text)
        return 0, 0 # Fallback
    
    # Exact match found
    # The span actually starts after the context prefix
    actual_span_start = start_index + len(context_prefix)
    actual_span_end = actual_span_start + len(span_text)
    
    return actual_span_start, actual_span_end

# ==============================================================================
# 6. WORKBOOK GENERATION FUNCTION
# ==============================================================================

def generate_workbook():
    # 1. Load or Create Workbook
    if os.path.exists(TEMPLATE_PATH):
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    else:
        wb = openpyxl.Workbook()
        # Create expected sheets if missing
        sheets = ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]
        for s in sheets:
            if s not in wb.sheetnames:
                wb.create_sheet(s)
        # Remove default sheet if exists
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # 2. Sheet: Note_Text
    ws_text = wb["Note_Text"]
    # Clear existing
    ws_text.delete_rows(1, ws_text.max_row)
    ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 3. Sheet: Note_Index
    ws_index = wb["Note_Index"]
    ws_index.delete_rows(1, ws_index.max_row)
    
    # Headers
    headers = ["note_id", "source_file", "procedure_date"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers)
    
    # Data
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE] + list(PROCEDURE_FLAGS.values())
    ws_index.append(row_data)

    # 4. Sheet: Span_Annotations (Manual/Blank Offsets)
    ws_anno = wb["Span_Annotations"]
    ws_anno.delete_rows(1, ws_anno.max_row)
    ws_anno.append(["note_id", "span_text", "label", "normalized_value", "start_char", "end_char", "context_prefix", "event_id"])
    
    for span in SPANS:
        s_text, label, norm, ctx, evt = span
        ws_anno.append([NOTE_ID, s_text, label, norm, "", "", ctx, evt])

    # 5. Sheet: Span_Hydrated (Calculated Offsets)
    ws_hyd = wb["Span_Hydrated"]
    ws_hyd.delete_rows(1, ws_hyd.max_row)
    ws_hyd.append(["note_id", "span_text", "label", "normalized_value", "start_char", "end_char", "context_prefix", "event_id"])
    
    for span in SPANS:
        s_text, label, norm, ctx, evt = span
        start, end = hydrate_span(NOTE_TEXT, s_text, ctx)
        ws_hyd.append([NOTE_ID, s_text, label, norm, start, end, ctx, evt])

    # 6. Sheet: Event_Log (Flattened)
    ws_log = wb["Event_Log"]
    ws_log.delete_rows(1, ws_log.max_row)
    ws_log.append(["note_id", "event_id", "event_name", "description", "span_text"])
    
    for evt in EVENTS:
        for s_text in evt["spans"]:
            ws_log.append([NOTE_ID, evt["event_id"], evt["name"], evt["description"], s_text])

    # 7. Sheet: V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    ws_v3.delete_rows(1, ws_v3.max_row)
    ws_v3.append(["note_id", "event_order", "event_id", "event_type", "event_data_json"])
    
    for idx, evt in enumerate(EVENTS, 1):
        evt_data = {
            "name": evt["name"],
            "description": evt["description"],
            "spans": evt["spans"]
        }
        ws_v3.append([NOTE_ID, idx, evt["event_id"], "procedure_step", json.dumps(evt_data)])

    # 8. Sheet: V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    ws_json.delete_rows(1, ws_json.max_row)
    ws_json.append(["note_id", "full_json_object"])
    
    # Construct Full JSON
    full_obj = {
        "metadata": {
            "note_id": NOTE_ID,
            "source": SOURCE_FILE,
            "procedure_date": PROCEDURE_DATE
        },
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS,
        "spans_hydrated": [
            {
                "text": s[0],
                "label": s[1],
                "normalized": s[2],
                "start": hydrate_span(NOTE_TEXT, s[0], s[3])[0],
                "end": hydrate_span(NOTE_TEXT, s[0], s[3])[1]
            }
            for s in SPANS
        ]
    }
    
    ws_json.append([NOTE_ID, json.dumps(full_obj, indent=2)])

    # 9. Save
    wb.save(OUTPUT_PATH)
    print(f"Successfully generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()