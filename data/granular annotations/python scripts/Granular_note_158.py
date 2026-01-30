import openpyxl
from openpyxl import Workbook
import re
import json
import os
import datetime

# =============================================================================
# 1. CONFIGURATION & CONSTANTS
# =============================================================================
NOTE_ID = "note_158"
SOURCE_FILE = "note_158.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_158 SOURCE_FILE: note_158.txt Procedure Name: 
1.	radial EBUS guided bronchoscopy with TNBA and endobronchial biopsies
Indications: Pulmonary Nodule
Medications: Propofol infusion via anesthesia assistance  
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention. 
Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The arytenoids were large with a large and floppy epiglottis. The vocal cords appeared normal. The subglottic space was normal.
The trachea is of normal caliber. The carina was sharp.
On the right the patient had an anatomic variant with an accessory airway just distal to the superior segment of the right lower lobe.
The left sided airway anatomy was normal. No evidence of endobronchial disease was seen to at least the first sub-segments.
We then removed the bronchoscope and inserted the P190 slim bronchoscope into the airway.
We attempted to visualize endobronchial disease within the anterior basilar segment of the left lower lobe but could not advance the scope distally enough.
We then inserted the radial ultrasound through the working channel of the bronchoscope and an eccentric view of the kno0wn nodule was obtained.
We then performed peripheral TBNAs of the point of interest. ROSE did not show obvious malignancy.
At this point we removed the slim scope and inserted the UC180F convex probe EBUS bronchoscope through the mouth, and advanced into the tracheobronchial tree.
We attempted to wedge the scope in the lower lobe to potentially visualize the tumor with ultrasound.
We saw an unusual area of soft tissue which did not obviously look like would be expected from the nodule.
Two biopsies were performed using the 22G EBUS-TBNA needle. ROSE however showed only blood.
Finally, fluoroscopy was brought into the room and the T190 therapeutic video bronchoscope was inserted into the airway and based on anatomical knowledge advanced into the left lower lobe and a large sheath catheter with radial ultrasound to the area of known nodule and a concentric view of the lesion was identified with the radial EBUS.
Biopsies were then performed with a variety of instruments to include peripheral needle forceps and brush with fluoroscopic guidance through the sheath.
After adequate samples were obtained the bronchoscope was removed. ROSE did not identify malignancy on preliminary samples.
After suctioning blood and secretions and once we were confident that there was no active bleeding the bronchoscope was removed and the procedure completed.
Following completion of the procedure the patient was noted to have audible upper airway sounds and ventilation was difficult.
And decision was made to convert to endotracheal intubation. Initially this was attempted over the bronchoscope however edema within the structures of the middle laryngeal cavity (above the vocal cords this was unsuccessful. Anesthesia was then able to intubate with the glideslope. The bronchoscope was then inserted through the ETT to confirm positioning 2cm above the carina and to suction secretions. The patient was then transferred to the ICU. 

Complications: Unexpected admission
Estimated Blood Loss: Less than 10 cc.
Post Procedure Diagnosis:
- Flexible bronchoscopy with successful biopsy of left lower lobe pulmonary nodule
- Will transfer to the ICU and attempt extubation later 
today. 
- Await final pathology"""

# =============================================================================
# 2. PROCEDURE FLAGS
# =============================================================================
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 1,
    "endobronchial_biopsy": 1, # Title says endobronchial biopsies
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 1,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 1, # Peripheral needle/forceps
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# =============================================================================
# 3. SPANS DATA
# Format: (span_text, label, normalized_value, context_prefix, event_id)
# =============================================================================
SPANS = [
    ("radial EBUS guided bronchoscopy", "PROC_METHOD", "Radial EBUS", "Procedure Name:", "evt_00"),
    ("TNBA", "PROC_METHOD", "TBNA", "guided bronchoscopy with", "evt_00"),
    ("endobronchial biopsies", "PROC_METHOD", "Endobronchial Biopsy", "TNBA and", "evt_00"),
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190 Bronchoscope", "tracheobronchial tree, the", "evt_01"),
    ("right lower lobe", "ANAT_LUNG_LOC", "RLL", "segment of the", "evt_01"),
    ("P190 slim bronchoscope", "DEV_INSTRUMENT", "P190 Bronchoscope", "and inserted the", "evt_02"),
    ("anterior basilar segment of the left lower lobe", "ANAT_LUNG_LOC", "LLL Anterior Basilar", "within the", "evt_02"),
    ("radial ultrasound", "PROC_METHOD", "Radial EBUS", "inserted the", "evt_02"),
    ("eccentric view", "OBS_LESION", "Eccentric View", "nodule was obtained.", "evt_02"),
    ("peripheral TBNAs", "PROC_METHOD", "Transbronchial Biopsy", "performed", "evt_02"),
    ("ROSE did not show obvious malignancy", "OBS_ROSE", "Negative", "interest.", "evt_02"),
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "UC180F EBUS Scope", "inserted the", "evt_03"),
    ("lower lobe", "ANAT_LUNG_LOC", "Lower Lobe", "wedge the scope in the", "evt_03"),
    ("unusual area of soft tissue", "OBS_LESION", "Soft Tissue", "saw an", "evt_03"),
    ("22G EBUS-TBNA needle", "DEV_NEEDLE", "22G", "using the", "evt_03"),
    ("Two biopsies", "PROC_ACTION", "Biopsy", "nodule.", "evt_03"),
    ("ROSE however showed only blood", "OBS_ROSE", "Bloody", "needle.", "evt_03"),
    ("fluoroscopy", "PROC_METHOD", "Fluoroscopy", "Finally,", "evt_04"),
    ("T190 therapeutic video bronchoscope", "DEV_INSTRUMENT", "T190 Bronchoscope", "room and the", "evt_04"),
    ("left lower lobe", "ANAT_LUNG_LOC", "LLL", "advanced into the", "evt_04"),
    ("large sheath catheter", "DEV_CATHETER", "Sheath", "and a", "evt_04"),
    ("radial ultrasound", "PROC_METHOD", "Radial EBUS", "catheter with", "evt_04"),
    ("concentric view", "OBS_LESION", "Concentric View", "nodule and a", "evt_04"),
    ("peripheral needle", "DEV_INSTRUMENT", "Needle", "include", "evt_04"),
    ("forceps", "DEV_INSTRUMENT", "Forceps", "peripheral needle", "evt_04"),
    ("brush", "DEV_INSTRUMENT", "Brush", "forceps and", "evt_04"),
    ("ROSE did not identify malignancy", "OBS_ROSE", "Negative", "removed.", "evt_04"),
    ("Unexpected admission", "OUTCOME_COMPLICATION", "Unexpected Admission", "Complications:", "evt_05"),
    ("edema within the structures of the middle laryngeal cavity", "OUTCOME_COMPLICATION", "Laryngeal Edema", "however", "evt_05"),
    ("convert to endotracheal intubation", "PROC_ACTION", "Intubation", "decision was made to", "evt_05"),
]

# =============================================================================
# 4. EVENTS AGGREGATION
# =============================================================================
EVENTS = [
    {
        "event_id": "evt_01",
        "name": "Diagnostic Inspection (Q190)",
        "method": ["Bronchoscopy"],
        "anatomy": ["Trachea", "RLL"],
        "devices": ["Q190"],
        "outcomes": []
    },
    {
        "event_id": "evt_02",
        "name": "Radial EBUS Attempt (P190)",
        "method": ["Radial EBUS", "TBNA"],
        "anatomy": ["LLL Anterior Basilar"],
        "devices": ["P190", "Radial Probe"],
        "outcomes": ["ROSE Negative", "Eccentric View"]
    },
    {
        "event_id": "evt_03",
        "name": "Linear EBUS (UC180F)",
        "method": ["Linear EBUS", "TBNA"],
        "anatomy": ["Lower Lobe"],
        "devices": ["UC180F", "22G Needle"],
        "outcomes": ["ROSE Bloody"]
    },
    {
        "event_id": "evt_04",
        "name": "Therapeutic Biopsy (T190)",
        "method": ["Fluoroscopy", "Radial EBUS", "Biopsy", "Brushing"],
        "anatomy": ["LLL"],
        "devices": ["T190", "Sheath", "Needle", "Forceps", "Brush"],
        "outcomes": ["Concentric View", "ROSE Negative"]
    },
    {
        "event_id": "evt_05",
        "name": "Complications",
        "method": ["Intubation"],
        "anatomy": ["Larynx"],
        "devices": ["ETT"],
        "outcomes": ["Unexpected Admission", "Laryngeal Edema"]
    }
]

# =============================================================================
# 5. HELPER FUNCTIONS
# =============================================================================
def clean_text(text):
    if not text:
        return ""
    return re.sub(r'\s+', ' ', text).strip()

def hydrate_span(text, span_text, context_prefix=""):
    """
    Finds the start and end offsets of a span within the note text.
    Uses context_prefix to disambiguate.
    """
    # 1. Clean inputs
    c_text = clean_text(text)
    c_span = clean_text(span_text)
    c_context = clean_text(context_prefix)
    
    # 2. Construct search pattern
    if c_context:
        # Escape special regex chars
        esc_context = re.escape(c_context)
        esc_span = re.escape(c_span)
        pattern = f"({esc_context})\s*({esc_span})"
    else:
        pattern = re.escape(c_span)

    # 3. Search
    match = re.search(pattern, c_text, re.IGNORECASE)
    
    if match:
        if c_context:
            # Group 2 is the span
            start = match.start(2)
            end = match.end(2)
        else:
            start = match.start()
            end = match.end()
        
        # Verify alignment with original raw text (approximate mapping)
        # Note: Since we flattened whitespace in regex, offsets might drift slightly 
        # from raw if raw has newlines. For Phase 0, exact char match on 
        # cleaned text is usually sufficient, or we assume text is pre-cleaned.
        # We will return offsets based on the CLEANED text for consistency in this script.
        return start, end, c_text[start:end]
    
    return None, None, None

# =============================================================================
# 6. WORKBOOK GENERATION
# =============================================================================
def generate_workbook():
    wb = Workbook()
    
    # --- Sheet 1: Note_Text ---
    ws_text = wb.active
    ws_text.title = "Note_Text"
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, clean_text(NOTE_TEXT)])
    
    # --- Sheet 2: Note_Index ---
    ws_index = wb.create_sheet("Note_Index")
    headers = ["NOTE_ID", "PROCEDURE_DATE"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers)
    row_data = [NOTE_ID, PROCEDURE_DATE] + list(PROCEDURE_FLAGS.values())
    ws_index.append(row_data)
    
    # --- Sheet 3: Span_Annotations ---
    ws_anno = wb.create_sheet("Span_Annotations")
    ws_anno.append(["NOTE_ID", "SPAN_TEXT", "LABEL", "VALUE_NORMALIZED", "CONTEXT_PREFIX", "EVENT_ID", "START_CHAR", "END_CHAR"])
    for s in SPANS:
        ws_anno.append([NOTE_ID, s[0], s[1], s[2], s[3], s[4], "", ""])

    # --- Sheet 4: Span_Hydrated ---
    ws_hydra = wb.create_sheet("Span_Hydrated")
    ws_hydra.append(["NOTE_ID", "SPAN_TEXT", "LABEL", "VALUE_NORMALIZED", "CONTEXT_PREFIX", "EVENT_ID", "START_CHAR", "END_CHAR", "MATCHED_TEXT"])
    
    full_text_cleaned = clean_text(NOTE_TEXT)
    
    for s in SPANS:
        span_text, label, val_norm, ctx, evt_id = s
        start, end, matched = hydrate_span(full_text_cleaned, span_text, ctx)
        ws_hydra.append([NOTE_ID, span_text, label, val_norm, ctx, evt_id, start, end, matched])

    # --- Sheet 5: Event_Log ---
    ws_event_log = wb.create_sheet("Event_Log")
    ws_event_log.append(["NOTE_ID", "EVENT_ID", "EVENT_NAME", "METHOD", "ANATOMY", "DEVICES", "OUTCOMES"])
    for evt in EVENTS:
        ws_event_log.append([
            NOTE_ID,
            evt["event_id"],
            evt["name"],
            ", ".join(evt["method"]),
            ", ".join(evt["anatomy"]),
            ", ".join(evt["devices"]),
            ", ".join(evt["outcomes"])
        ])

    # --- Sheet 6: V3_Procedure_Events ---
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    ws_v3.append(["NOTE_ID", "EVENT_ID", "EVENT_TYPE", "EVENT_DETAILS_JSON"])
    for evt in EVENTS:
        details = {
            "method": evt["method"],
            "anatomy": evt["anatomy"],
            "devices": evt["devices"],
            "outcomes": evt["outcomes"]
        }
        ws_v3.append([NOTE_ID, evt["event_id"], "ProcedureStep", json.dumps(details)])

    # --- Sheet 7: V3_Registry_JSON ---
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["NOTE_ID", "FULL_JSON_BLOB"])
    
    registry_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source": SOURCE_FILE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS,
        "spans": [
            {"text": s[0], "label": s[1], "value": s[2], "event_id": s[4]} for s in SPANS
        ]
    }
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])

    # --- Save ---
    wb.save(OUTPUT_PATH)
    print(f"Generated Phase 0 Workbook: {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()