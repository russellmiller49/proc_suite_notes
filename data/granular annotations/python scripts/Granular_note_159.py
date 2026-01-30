import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# 1. Imports & Constants
NOTE_ID = "note_159"
SOURCE_FILE = "note_159.txt"
PROCEDURE_DATE = ""
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_159 SOURCE_FILE: note_159.txt Indications: Left upper lobe mass 
Medications: Propofol infusion via anesthesia assistance  
Medications: General Anesthesia,
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention.
Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway was in good position. The vocal cords appeared normal. The subglottic space was normal.
The trachea was of normal caliber. The carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level. Bronchial mucosa and anatomy were normal;
there are no endobronchial lesions. We then removed the diagnostic Q190 bronchoscopy and the super-dimension navigational catheter was inserted through the T190 therapeutic bronchoscope and advanced into the airway.
Using navigational map we attempted to advance the 90 degree edge catheter into the proximity of the lesion within the left upper lobe.
Confirmation of placement once at the point of interest with radial ultrasound showed a concentric view within the lesion.
Biopsies were then performed with a variety of instruments to include peripheral needle, brush, triple needle brush and forceps, under fluoroscopic visualization.
After adequate samples were obtained the bronchoscope was removed and the procedure completed
Complications: No immediate complications
Estimated Blood Loss: Less than 5 cc.
Post Procedure Diagnosis:
- Flexible bronchoscopy with successful navigational biopsy of left upper lobe nodule.  
- Await final pathology"""

# 2. Configuration (Procedure Flags)
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 1,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 1,
    "navigational_bronchoscopy": 1,
    "transbronchial_biopsy": 1,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 3. Data Definition (Spans)
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190 Bronchoscope", "tracheobronchial tree, the ", "evt_01"),
    ("laryngeal mask airway", "DEV_INSTRUMENT", "Laryngeal Mask Airway", "mouth, via ", "evt_01"),
    ("diagnostic Q190 bronchoscopy", "PROC_METHOD", "Diagnostic Bronchoscopy", "removed the ", "evt_01"),
    ("super-dimension navigational catheter", "DEV_CATHETER", "SuperDimension Catheter", "and the ", "evt_02"),
    ("T190 therapeutic bronchoscope", "DEV_INSTRUMENT", "T190 Bronchoscope", "through the ", "evt_02"),
    ("navigational map", "PROC_METHOD", "Navigational Bronchoscopy", "Using ", "evt_02"),
    ("90 degree edge catheter", "DEV_CATHETER", "Edge Catheter", "advance the ", "evt_02"),
    ("left upper lobe", "ANAT_LUNG_LOC", "Left Upper Lobe", "within the ", "evt_02"),
    ("radial ultrasound", "PROC_METHOD", "Radial EBUS", "interest with ", "evt_02"),
    ("concentric view", "OBS_LESION", "Concentric View", "showed a ", "evt_02"),
    ("Biopsies", "PROC_ACTION", "Biopsy", "lesion.\n", "evt_03"),
    ("peripheral needle", "DEV_NEEDLE", "Peripheral Needle", "include ", "evt_03"),
    ("brush", "DEV_INSTRUMENT", "Brush", "needle, ", "evt_03"),
    ("triple needle brush", "DEV_INSTRUMENT", "Triple Needle Brush", "brush, ", "evt_03"),
    ("forceps", "DEV_INSTRUMENT", "Forceps", "brush and ", "evt_03"),
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications: ", "evt_04")
]

# 4. Event Definitions
EVENTS = [
    {
        "event_id": "evt_01",
        "name": "Initial Inspection",
        "method": "Diagnostic Bronchoscopy",
        "anatomy": "Airway",
        "devices": ["Q190 Bronchoscope", "Laryngeal Mask Airway"],
        "outcomes": []
    },
    {
        "event_id": "evt_02",
        "name": "Navigation and rEBUS",
        "method": "Navigational Bronchoscopy, Radial EBUS",
        "anatomy": "Left Upper Lobe",
        "devices": ["SuperDimension Catheter", "T190 Bronchoscope", "Edge Catheter"],
        "outcomes": ["Concentric View"]
    },
    {
        "event_id": "evt_03",
        "name": "Biopsy Sampling",
        "method": "Transbronchial Biopsy, Brushings",
        "anatomy": "Left Upper Lobe",
        "devices": ["Peripheral Needle", "Brush", "Triple Needle Brush", "Forceps"],
        "outcomes": []
    },
    {
        "event_id": "evt_04",
        "name": "Procedure Completion",
        "method": "",
        "anatomy": "",
        "devices": [],
        "outcomes": ["No immediate complications"]
    }
]

# 5. Helper Functions
def clean_text(text):
    if not text:
        return ""
    return re.sub(r'[\r\n]+', ' ', text).strip()

def hydrate_span(full_text, span_text, context_prefix):
    """
    Finds the start and end offsets of a span in the full text using a context prefix.
    """
    # Normalize inputs for search
    clean_full = full_text.replace('\r', '').replace('\n', ' ')
    clean_span = span_text.replace('\r', '').replace('\n', ' ')
    clean_prefix = context_prefix.replace('\r', '').replace('\n', ' ')
    
    # Construct search pattern
    # We look for prefix + span
    # We escape the prefix and span to handle special regex characters
    search_pattern = re.escape(clean_prefix) + r"\s*" + re.escape(clean_span)
    
    match = re.search(search_pattern, clean_full, re.IGNORECASE)
    
    if match:
        # The start of the span is the end of the prefix match
        # However, re.search gives the range of the whole match (prefix + span)
        # We need to calculate where the span actually starts.
        
        # A simpler approach: find the prefix, then find the span immediately following
        prefix_match = re.search(re.escape(clean_prefix), clean_full, re.IGNORECASE)
        if prefix_match:
            start_search_idx = prefix_match.end()
            # Look for the span starting from end of prefix
            span_match = re.search(re.escape(clean_span), clean_full[start_search_idx:], re.IGNORECASE)
            if span_match:
                final_start = start_search_idx + span_match.start()
                final_end = start_search_idx + span_match.end()
                return final_start, final_end
    
    return None, None

def generate_workbook():
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # --- Sheet 1: Note_Text ---
    ws_text = wb.create_sheet("Note_Text")
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # --- Sheet 2: Note_Index ---
    ws_index = wb.create_sheet("Note_Index")
    headers_index = ["NOTE_ID", "PROCEDURE_DATE"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers_index)
    row_data = [NOTE_ID, PROCEDURE_DATE] + list(PROCEDURE_FLAGS.values())
    ws_index.append(row_data)
    
    # --- Sheet 3: Span_Annotations ---
    ws_anno = wb.create_sheet("Span_Annotations")
    headers_anno = ["NOTE_ID", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "START_CHAR", "END_CHAR", "EVENT_ID"]
    ws_anno.append(headers_anno)
    
    for span in SPANS:
        # (span_text, label, normalized_value, context_prefix, event_id)
        # Leave start/end blank for this sheet as per instruction implies raw data, 
        # but usually we want them filled if possible. 
        # The prompt instruction says: "Sheet 3... Write rows with blank start/end chars."
        ws_anno.append([NOTE_ID, span[0], span[1], span[2], span[3], "", "", span[4]])
        
    # --- Sheet 4: Span_Hydrated ---
    ws_hydra = wb.create_sheet("Span_Hydrated")
    ws_hydra.append(headers_anno)
    
    for span in SPANS:
        span_text = span[0]
        context_prefix = span[3]
        start, end = hydrate_span(NOTE_TEXT, span_text, context_prefix)
        ws_hydra.append([NOTE_ID, span_text, span[1], span[2], span[3], start, end, span[4]])
        
    # --- Sheet 5: Event_Log ---
    ws_event = wb.create_sheet("Event_Log")
    headers_event = ["NOTE_ID", "EVENT_ID", "EVENT_NAME", "METHOD", "ANATOMY", "DEVICES", "OUTCOMES"]
    ws_event.append(headers_event)
    
    for evt in EVENTS:
        dev_str = ", ".join(evt["devices"])
        out_str = ", ".join(evt["outcomes"])
        ws_event.append([
            NOTE_ID, 
            evt["event_id"], 
            evt["name"], 
            evt["method"], 
            evt["anatomy"], 
            dev_str, 
            out_str
        ])
        
    # --- Sheet 6: V3_Procedure_Events ---
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    headers_v3 = ["note_id", "event_id", "event_type", "event_metadata"]
    ws_v3.append(headers_v3)
    
    for evt in EVENTS:
        metadata = {
            "method": evt["method"],
            "anatomy": evt["anatomy"],
            "devices": evt["devices"],
            "outcomes": evt["outcomes"]
        }
        ws_v3.append([NOTE_ID, evt["event_id"], evt["name"], json.dumps(metadata)])
        
    # --- Sheet 7: V3_Registry_JSON ---
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["note_id", "json_data"])
    
    registry_data = {
        "note_id": NOTE_ID,
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS,
        "spans": [
            {
                "text": s[0],
                "label": s[1],
                "normalized": s[2],
                "event_id": s[4]
            } for s in SPANS
        ]
    }
    
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])
    
    # Save
    if not os.path.exists(OUTPUT_PATH):
        wb.save(OUTPUT_PATH)
        print(f"Workbook saved to {OUTPUT_PATH}")
    else:
        print(f"Overwriting {OUTPUT_PATH}")
        wb.save(OUTPUT_PATH)

if __name__ == "__main__":
    generate_workbook()