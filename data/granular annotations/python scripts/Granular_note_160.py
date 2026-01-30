import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# 1. Imports & Constants
NOTE_ID = "note_160"
SOURCE_FILE = "note_160.txt"
PROCEDURE_DATE = ""
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"
NOTE_TEXT = """NOTE_ID:  note_160 SOURCE_FILE: note_160.txt Procedure Name: 
1. Electromagnetic navigation bronchoscopy
Indications: Pulmonary Nodule
Medications: Propofol infusion via anesthesia assistance  
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention. 
Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The arytenoids were large with a large and floppy epiglottis. The vocal cords appeared normal. The subglottic space was normal.
Tracheomalacia was noted as we advanced the scope to the carina. The carina was sharp.
The right and left sided airway anatomy was normal. No evidence of endobronchial disease was seen to at least the first sub-segments.
We then removed the bronchoscope and inserted the EBUS bronchoscope into the airway.
We attempted to visualize the nodule with endobronchial ultrasound but did not see the nodule with ultrasound.
We then inserted the therapeutic bronchoscope for electromagnetic navigation bronchoscopy.
We were not able to get to the nodule using the navigation equipment.
After suctioning blood and secretions and once we were confident that there was no active bleeding the bronchoscope was removed and the procedure completed.
Complications: None.
Estimated Blood Loss: Less than 10 cc.  No specimens were obtained for path review.
Post Procedure Diagnosis:
- Discuss his case at cardiothoracic tumor board next Tuesday to determine best next step in workup"""

# 2. Configuration (Procedure Flags)
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 1,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 3. Data Definition (Spans)
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("Electromagnetic navigation bronchoscopy", "PROC_METHOD", "Electromagnetic navigation bronchoscopy", "Procedure Name: \n1. ", "evt_01"),
    ("Pulmonary Nodule", "CTX_HISTORICAL", "Pulmonary Nodule", "Indications: ", "evt_01"),
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190 video bronchoscope", "tracheobronchial tree, the ", "evt_02"),
    ("Tracheomalacia", "OBS_LESION", "Tracheomalacia", "space was normal.\n", "evt_02"),
    ("EBUS bronchoscope", "DEV_INSTRUMENT", "EBUS bronchoscope", "and inserted the ", "evt_03"),
    ("endobronchial ultrasound", "PROC_METHOD", "endobronchial ultrasound", "nodule with ", "evt_03"),
    ("therapeutic bronchoscope", "DEV_INSTRUMENT", "therapeutic bronchoscope", "We then inserted the ", "evt_04"),
    ("electromagnetic navigation bronchoscopy", "PROC_METHOD", "electromagnetic navigation bronchoscopy", "therapeutic bronchoscope for ", "evt_04"),
    ("navigation equipment", "DEV_INSTRUMENT", "navigation equipment", "using the ", "evt_04"),
    ("suctioning", "PROC_ACTION", "suctioning", "After ", "evt_05"),
    ("None", "OUTCOME_COMPLICATION", "None", "Complications: ", "evt_06")
]

# 4. Event Definitions
EVENTS = [
    {
        "event_id": "evt_01",
        "name": "Procedure Start",
        "type": "admin",
        "description": "Procedure initiation and indications"
    },
    {
        "event_id": "evt_02",
        "name": "Initial Inspection",
        "type": "diagnostic",
        "description": "Inspection with Q190 scope, noting Tracheomalacia"
    },
    {
        "event_id": "evt_03",
        "name": "EBUS Inspection",
        "type": "diagnostic",
        "description": "Attempted visualization with EBUS scope (Linear inferred by scope name, though Radial typical for nodules; context ambiguous so EBUS generic used)"
    },
    {
        "event_id": "evt_04",
        "name": "ENB Navigation",
        "type": "diagnostic",
        "description": "Navigation attempt with therapeutic scope"
    },
    {
        "event_id": "evt_05",
        "name": "Cleaning",
        "type": "therapeutic",
        "description": "Suctioning secretions"
    },
    {
        "event_id": "evt_06",
        "name": "Completion",
        "type": "admin",
        "description": "Procedure conclusion and complications"
    }
]

# 5. Helper Functions
def clean_text(text):
    if not text:
        return ""
    return text.strip().replace('\r\n', '\n').replace('\r', '\n')

def hydrate_span(text, span_text, context_prefix):
    """
    Finds the start and end offsets of a span in the text, using a context prefix to disambiguate.
    """
    # Normalize inputs
    clean_full_text = clean_text(text)
    clean_span = clean_text(span_text)
    clean_context = clean_text(context_prefix)

    if not clean_span:
        return None, None

    # Construct regex pattern: context + optional whitespace + span
    # We escape the strings to handle special regex characters
    pattern_str = re.escape(clean_context) + r"\s*" + re.escape(clean_span)
    
    match = re.search(pattern_str, clean_full_text, re.IGNORECASE)
    
    if match:
        # The match includes the context. We need the span's position within that match.
        full_match_start = match.start()
        full_match_text = match.group(0)
        
        # Find where the span actually starts within the match
        # This handles variable whitespace between context and span
        span_start_in_match = full_match_text.lower().rfind(clean_span.lower())
        
        real_start = full_match_start + span_start_in_match
        real_end = real_start + len(clean_span)
        
        return real_start, real_end
    else:
        # Fallback: strict find if context fails (use with caution, can be ambiguous)
        start = clean_full_text.find(clean_span)
        if start != -1:
            return start, start + len(clean_span)
        
        return 0, 0

# 6. Workbook Generation Function
def generate_workbook():
    # Create workbook
    wb = openpyxl.Workbook()
    
    # --- Sheet 1: Note_Text ---
    ws_text = wb.active
    ws_text.title = "Note_Text"
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # --- Sheet 2: Note_Index ---
    ws_index = wb.create_sheet("Note_Index")
    headers_index = ["NOTE_ID", "PROCEDURE_DATE"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers_index)
    
    row_data = [NOTE_ID, PROCEDURE_DATE]
    for key in PROCEDURE_FLAGS:
        row_data.append(PROCEDURE_FLAGS[key])
    ws_index.append(row_data)
    
    # --- Sheet 3: Span_Annotations ---
    ws_spans = wb.create_sheet("Span_Annotations")
    headers_spans = ["NOTE_ID", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "EVENT_ID", "START_CHAR", "END_CHAR"]
    ws_spans.append(headers_spans)
    
    for span in SPANS:
        # Initial write without offsets
        ws_spans.append([NOTE_ID, span[0], span[1], span[2], span[3], span[4], "", ""])

    # --- Sheet 4: Span_Hydrated ---
    ws_hydrated = wb.create_sheet("Span_Hydrated")
    ws_hydrated.append(headers_spans)
    
    for span in SPANS:
        start, end = hydrate_span(NOTE_TEXT, span[0], span[3])
        ws_hydrated.append([NOTE_ID, span[0], span[1], span[2], span[3], span[4], start, end])

    # --- Sheet 5: Event_Log ---
    ws_events = wb.create_sheet("Event_Log")
    headers_events = ["NOTE_ID", "EVENT_ID", "EVENT_NAME", "EVENT_TYPE", "DESCRIPTION"]
    ws_events.append(headers_events)
    
    for evt in EVENTS:
        ws_events.append([NOTE_ID, evt["event_id"], evt["name"], evt["type"], evt["description"]])

    # --- Sheet 6: V3_Procedure_Events ---
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    # This schema is hypothetical based on "V3" requirements; providing a generic flexible structure
    headers_v3 = ["NOTE_ID", "EVENT_ID", "METHOD", "ANATOMY", "DEVICES", "OUTCOMES", "JSON_PAYLOAD"]
    ws_v3.append(headers_v3)
    
    # Aggregate data for V3
    for evt in EVENTS:
        e_id = evt["event_id"]
        
        # Filter spans for this event
        evt_spans = [s for s in SPANS if s[4] == e_id]
        
        methods = [s[2] for s in evt_spans if s[1] == "PROC_METHOD"]
        anatomy = [s[2] for s in evt_spans if s[1].startswith("ANAT")]
        devices = [s[2] for s in evt_spans if s[1].startswith("DEV")]
        outcomes = [s[2] for s in evt_spans if s[1].startswith("OUTCOME")]
        
        payload = {
            "event_id": e_id,
            "name": evt["name"],
            "methods": methods,
            "anatomy": anatomy,
            "devices": devices,
            "outcomes": outcomes
        }
        
        ws_v3.append([
            NOTE_ID, 
            e_id, 
            ", ".join(methods), 
            ", ".join(anatomy), 
            ", ".join(devices), 
            ", ".join(outcomes), 
            json.dumps(payload)
        ])

    # --- Sheet 7: V3_Registry_JSON ---
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["NOTE_ID", "JSON_FULL"])
    
    registry_data = {
        "note_id": NOTE_ID,
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS,
        "spans": [
            {
                "text": s[0], 
                "label": s[1], 
                "value": s[2], 
                "event_id": s[4]
            } for s in SPANS
        ]
    }
    
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    if not os.path.exists(OUTPUT_PATH):
        wb.save(OUTPUT_PATH)
        print(f"Workbook saved to {OUTPUT_PATH}")
    else:
        print(f"Workbook already exists at {OUTPUT_PATH}")

# 7. Execution Block
if __name__ == "__main__":
    generate_workbook()