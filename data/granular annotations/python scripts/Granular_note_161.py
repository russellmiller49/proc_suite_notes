import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# 1. Imports & Constants
NOTE_ID = "note_161"
SOURCE_FILE = "note_161.txt"
PROCEDURE_DATE = ""
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_161 SOURCE_FILE: note_161.txt Proceduralist(s): Russell Miller MD, Pulmonologist, Jeff Biberson (Fellow)Garrett Harp, MD (Fellow)
Procedure Name: EBUS Staging Bronchoscopy.
(CPT 31652 convex probe endobronchial ultrasound sampling 2 or fewer hilar or mediastinal stations or structures).
Indications: Diagnosis and staging of presumed lung cancer
Medications: General Anesthesia
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record.
Laboratory studies and radiographs 
Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the ET tube into the tracheobronchial tree.
Tracheomalacia was noted in the trachea. The carina was sharp. The right-sided airway anatomy was normal.
The left sided airway anatomy was normal. No evidence of endobronchial disease was seen to at least the first sub-segments.
A systematic hilar and mediastinal lymph node survey was carried out.
Sampling criteria (5mm short axis diameter) were met in station 11L  (5.4mm) and 4R (5.5mm) lymph nodes.
Sampling by transbronchial needle aspiration was performed beginning with the 11L Lymph node, followed by 4R lymph nodes using an Olympus Visioshot EBUSTBNA 22 gauge needle.
ROSE showed non-diagnostic tissue in the low probability 11L lymph node and benign lymphocytes in the 4Rs.
All samples were sent for routine cytology. The Q190 video bronchoscope was then re-inserted and after suctioning blood and secretions there was no evidence of active bleeding and the bronchoscope was subsequently removed.
Complications: No immediate complications
Estimated Blood Loss: 5 cc.

Post Procedure Diagnosis:
- Technically successful flexible bronchoscopy with endobronchial ultrasound-guided biopsies.
- The patient has remained stable and has been transferred in good condition to the post-surgical monitoring unit.
- Will await final pathology results"""

# 2. Configuration (Procedure Flags)
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 3. Data Definition (Spans)
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    # Event 01: Diagnostic / Scope
    ("EBUS Staging Bronchoscopy", "PROC_METHOD", "EBUS Staging", "Procedure Name: ", "evt_01"),
    ("convex probe endobronchial ultrasound", "PROC_METHOD", "Linear EBUS", "CPT 31652 ", "evt_01"),
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190 Scope", "tracheobronchial tree, the ", "evt_01"),
    ("Tracheomalacia", "OBS_LESION", "Tracheomalacia", "source: 7] ", "evt_01"),
    ("trachea", "ANAT_AIRWAY", "Trachea", "noted in the ", "evt_01"),

    # Event 02: EBUS TBNA 11L
    ("11L", "ANAT_LN_STATION", "11L", "beginning with the ", "evt_02"),
    ("5.4mm", "MEAS_SIZE", "5.4mm", "station 11L  (", "evt_02"),
    ("transbronchial needle aspiration", "PROC_METHOD", "TBNA", "Sampling by ", "evt_02"),
    ("Olympus Visioshot EBUSTBNA 22 gauge needle", "DEV_NEEDLE", "Olympus Visioshot 22G", "using an ", "evt_02"),
    ("non-diagnostic", "OBS_ROSE", "Non-diagnostic", "ROSE showed ", "evt_02"),
    
    # Event 03: EBUS TBNA 4R
    ("4R", "ANAT_LN_STATION", "4R", "followed by ", "evt_03"),
    ("5.5mm", "MEAS_SIZE", "5.5mm", "and 4R (", "evt_03"),
    ("benign lymphocytes", "OBS_ROSE", "Benign Lymphocytes", "node and ", "evt_03"),

    # Outcomes / Global
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications: ", "evt_global"),
    ("Technically successful", "OUTCOME_SYMPTOMS", "Successful", "- ", "evt_global")
]

# 4. Event Definitions
EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Bronchoscopy",
        "method": ["Linear EBUS", "Diagnostic"],
        "anatomy": ["Trachea"],
        "devices": ["Q190 video bronchoscope"],
        "outcomes": [],
        "notes": "Tracheomalacia noted."
    },
    {
        "event_id": "evt_02",
        "procedure_type": "Bronchoscopy",
        "method": ["TBNA"],
        "anatomy": ["11L"],
        "devices": ["Olympus Visioshot 22G"],
        "outcomes": ["ROSE: Non-diagnostic"],
        "notes": "Node size 5.4mm"
    },
    {
        "event_id": "evt_03",
        "procedure_type": "Bronchoscopy",
        "method": ["TBNA"],
        "anatomy": ["4R"],
        "devices": ["Olympus Visioshot 22G"],
        "outcomes": ["ROSE: Benign Lymphocytes"],
        "notes": "Node size 5.5mm"
    }
]

# 5. Helper Functions
def clean_text(text):
    if not text:
        return ""
    return text.strip().replace('\r', '')

def hydrate_span(text, span_text, context_prefix):
    """
    Finds the start and end indices of span_text within text,
    using context_prefix to locate the correct occurrence.
    """
    cleaned_text = clean_text(text)
    cleaned_span = clean_text(span_text)
    cleaned_context = clean_text(context_prefix)

    if not cleaned_span:
        return 0, 0

    # Locate context first
    context_start = cleaned_text.find(cleaned_context)
    if context_start == -1:
        # Fallback: search for span globally if context not found
        start_index = cleaned_text.find(cleaned_span)
        if start_index == -1:
            return 0, 0 # Not found
        return start_index, start_index + len(cleaned_span)

    # Search for span AFTER the context
    search_start_idx = context_start + len(cleaned_context)
    start_index = cleaned_text.find(cleaned_span, search_start_idx)
    
    # Safety: if not found after context, look globally (fallback)
    if start_index == -1:
        start_index = cleaned_text.find(cleaned_span)
        if start_index == -1:
            return 0, 0

    return start_index, start_index + len(cleaned_span)

def find_offsets(text, span_text, context_prefix):
    start, end = hydrate_span(text, span_text, context_prefix)
    return start, end

# 6. Workbook Generation Function
def generate_workbook():
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # --- Sheet 1: Note_Text ---
    ws_text = wb.create_sheet("Note_Text")
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # --- Sheet 2: Note_Index ---
    ws_index = wb.create_sheet("Note_Index")
    headers_index = ["NOTE_ID", "PROCEDURE_DATE", "SOURCE_FILE"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers_index)
    row_index = [NOTE_ID, PROCEDURE_DATE, SOURCE_FILE] + [PROCEDURE_FLAGS[k] for k in PROCEDURE_FLAGS]
    ws_index.append(row_index)

    # --- Sheet 3: Span_Annotations ---
    ws_anno = wb.create_sheet("Span_Annotations")
    headers_anno = ["NOTE_ID", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "START_CHAR", "END_CHAR", "EVENT_ID"]
    ws_anno.append(headers_anno)
    
    # --- Sheet 4: Span_Hydrated ---
    ws_hydrated = wb.create_sheet("Span_Hydrated")
    ws_hydrated.append(headers_anno)

    for span in SPANS:
        s_text, label, norm_val, context, evt_id = span
        # For Annotation sheet, leave offsets blank
        ws_anno.append([NOTE_ID, s_text, label, norm_val, context, "", "", evt_id])
        
        # For Hydrated sheet, calculate offsets
        start, end = find_offsets(NOTE_TEXT, s_text, context)
        ws_hydrated.append([NOTE_ID, s_text, label, norm_val, context, start, end, evt_id])

    # --- Sheet 5: Event_Log ---
    ws_event_log = wb.create_sheet("Event_Log")
    headers_log = ["NOTE_ID", "EVENT_ID", "PROCEDURE_TYPE", "METHOD", "ANATOMY", "DEVICES", "OUTCOMES", "NOTES"]
    ws_event_log.append(headers_log)
    
    for evt in EVENTS:
        ws_event_log.append([
            NOTE_ID,
            evt["event_id"],
            evt["procedure_type"],
            ", ".join(evt["method"]),
            ", ".join(evt["anatomy"]),
            ", ".join(evt["devices"]),
            ", ".join(evt["outcomes"]),
            evt.get("notes", "")
        ])

    # --- Sheet 6: V3_Procedure_Events ---
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    headers_v3 = ["NOTE_ID", "EVENT_ID", "DERIVED_PROCEDURE_TYPE", "DERIVED_METHOD", "DERIVED_ANATOMY", "DERIVED_DEVICES", "DERIVED_OUTCOMES", "JSON_PAYLOAD"]
    ws_v3.append(headers_v3)

    for evt in EVENTS:
        json_payload = json.dumps(evt)
        ws_v3.append([
            NOTE_ID,
            evt["event_id"],
            evt["procedure_type"],
            json.dumps(evt["method"]),
            json.dumps(evt["anatomy"]),
            json.dumps(evt["devices"]),
            json.dumps(evt["outcomes"]),
            json_payload
        ])

    # --- Sheet 7: V3_Registry_JSON ---
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["NOTE_ID", "FULL_REGISTRY_JSON"])
    
    registry_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source_file": SOURCE_FILE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS
    }
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Workbook generated successfully: {OUTPUT_PATH}")

# 7. Execution Block
if __name__ == "__main__":
    generate_workbook()