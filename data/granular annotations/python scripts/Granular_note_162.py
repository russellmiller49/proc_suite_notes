import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# 1. Imports & Constants
NOTE_ID = "note_162"
SOURCE_FILE = "note_162.txt"
PROCEDURE_DATE = ""
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_162 SOURCE_FILE: note_162.txt Procedure Name: EBUS Bronchoscopy
Indications: lung cancer diagnosis and staging (left upper lobe tumor)
Medications: General Anesthesia
Procedure performed:
: 31653 bronchoscopy with endobronchial ultrasound (EBUS) guided transbronchial sampling > 2 structures.
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention.
 Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway was in good position. The vocal cords appeared normal. The subglottic space was normal.
The trachea was of normal caliber. The carina was sharp.
The tracheobronchial tree was examined to at least the first sub-segmental level. Bronchial mucosa and anatomy were normal;
there are no endobronchial lesions, except for in the left lower lobe in which the proximal origin was mildly extrinsically compressed.
The video bronchoscope was then removed and the UC180F convex probe EBUS bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
A systematic hilar and mediastinal lymph node survey was carried out.
Sampling criteria (5mm short axis diameter) were met in station 11Rs (6.7mm), 10R (5.7mm), 4R (9.1mm), 2R (7.1 mm), 7 (15.7mm), 4L (6.9mm),  and 11L (21.1mm)  lymph nodes.
Sampling by transbronchial needle aspiration was performed beginning with the 11Rs Lymph node followed by 7, and 4R, 2R lymph nodes using an Olympus EBUSTBNA 22 gauge needle.
ROSE showed malignant cells in the 4R and 2R station consistent with N3 disease.
We then moved to the large 11L lymph node and took 8 additional passes for molecular studies.
All samples were sent for routine cytology and a dedicated pass from the 11L was sent for flow cytometry.
The Q190 video bronchoscope was then re-inserted and after suctioning blood and secretions there was no evidence of active bleeding and the bronchoscope was subsequently removed.
Complications: No immediate complications
Estimated Blood Loss: 5 cc.

Post Procedure Recommendations:
- Transfer to post-procedure unit and home per protocol
- Will await final pathology results"""

# 2. Configuration (Procedure Flags)
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 3. Data Definition (Spans)
SPANS = [
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190", "tree, the ", "evt_01"),
    ("tracheobronchial tree", "ANAT_AIRWAY", "tracheobronchial tree", "advanced to the ", "evt_01"),
    ("left lower lobe", "ANAT_LUNG_LOC", "LLL", "except for in the ", "evt_01"),
    ("mildly extrinsically compressed", "OBS_LESION", "extrinsic compression", "proximal origin was ", "evt_01"),
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "UC180F", "removed and the ", "evt_02"),
    ("hilar and mediastinal lymph node", "ANAT_LN_STATION", "hilar/mediastinal LN", "A systematic ", "evt_02"),
    ("11Rs", "ANAT_LN_STATION", "11Rs", "met in station ", "evt_02"),
    ("10R", "ANAT_LN_STATION", "10R", "11Rs (6.7mm), ", "evt_02"),
    ("4R", "ANAT_LN_STATION", "4R", "10R (5.7mm), ", "evt_02"),
    ("2R", "ANAT_LN_STATION", "2R", "4R (9.1mm), ", "evt_02"),
    ("7", "ANAT_LN_STATION", "7", "2R (7.1 mm), ", "evt_02"),
    ("4L", "ANAT_LN_STATION", "4L", "7 (15.7mm), ", "evt_02"),
    ("11L", "ANAT_LN_STATION", "11L", "4L (6.9mm),  and ", "evt_02"),
    ("transbronchial needle aspiration", "PROC_METHOD", "Linear EBUS-TBNA", "Sampling by ", "evt_03"),
    ("11Rs", "ANAT_LN_STATION", "11Rs", "beginning with the ", "evt_03"),
    ("7", "ANAT_LN_STATION", "7", "Lymph node followed by ", "evt_03"),
    ("4R", "ANAT_LN_STATION", "4R", "followed by 7, and ", "evt_03"),
    ("2R", "ANAT_LN_STATION", "2R", "4R, ", "evt_03"),
    ("Olympus EBUSTBNA 22 gauge needle", "DEV_NEEDLE", "Olympus EBUS-TBNA 22G", "using an ", "evt_03"),
    ("malignant cells", "OBS_ROSE", "malignant", "ROSE showed ", "evt_03"),
    ("N3 disease", "OBS_LESION", "N3 disease", "consistent with ", "evt_03"),
    ("11L", "ANAT_LN_STATION", "11L", "moved to the large ", "evt_04"),
    ("8", "MEAS_COUNT", "8", "took ", "evt_04"),
    ("molecular studies", "PROC_ACTION", "molecular studies", "passes for ", "evt_04"),
    ("flow cytometry", "PROC_ACTION", "flow cytometry", "sent for ", "evt_04"),
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications: ", "evt_05")
]

# 4. Event Definitions
EVENTS = [
    {
        "event_id": "evt_01",
        "name": "Diagnostic Bronchoscopy",
        "method": "White Light Bronchoscopy",
        "devices": ["Q190 video bronchoscope", "laryngeal mask airway"],
        "anatomy": ["Tracheobronchial tree", "Left Lower Lobe"],
        "observations": ["Extrinsic compression (LLL)"],
        "outcomes": []
    },
    {
        "event_id": "evt_02",
        "name": "EBUS Survey",
        "method": "Linear EBUS",
        "devices": ["UC180F convex probe EBUS bronchoscope"],
        "anatomy": ["11Rs", "10R", "4R", "2R", "7", "4L", "11L"],
        "observations": ["Normal measurements recorded"],
        "outcomes": []
    },
    {
        "event_id": "evt_03",
        "name": "EBUS TBNA Sampling (Cytology)",
        "method": "Linear EBUS-TBNA",
        "devices": ["Olympus EBUSTBNA 22 gauge needle"],
        "anatomy": ["11Rs", "7", "4R", "2R"],
        "observations": ["ROSE: Malignant cells (4R, 2R)", "N3 disease"],
        "outcomes": []
    },
    {
        "event_id": "evt_04",
        "name": "EBUS TBNA Sampling (Molecular/Flow)",
        "method": "Linear EBUS-TBNA",
        "devices": ["Olympus EBUSTBNA 22 gauge needle"],
        "anatomy": ["11L"],
        "observations": ["8 passes", "Molecular studies", "Flow cytometry"],
        "outcomes": []
    },
    {
        "event_id": "evt_05",
        "name": "Procedure Conclusion",
        "method": "Observation",
        "devices": [],
        "anatomy": [],
        "observations": [],
        "outcomes": ["No immediate complications"]
    }
]

# 5. Helper Functions
def clean_text(text):
    return text.replace('\r', '').strip()

def hydrate_span(full_text, span_text, context_prefix):
    cleaned_full = clean_text(full_text)
    cleaned_span = clean_text(span_text)
    cleaned_context = clean_text(context_prefix)
    
    if not cleaned_span:
        return 0, 0
        
    # Attempt to find via context
    search_pattern = re.escape(cleaned_context) + r"\s*" + re.escape(cleaned_span)
    match = re.search(search_pattern, cleaned_full, re.IGNORECASE)
    
    if match:
        # The span starts after the context
        start_index = match.start() + len(cleaned_context)
        # Adjust for potential whitespace between context and span in the match
        # We need to find exactly where the span text starts within the match
        full_match_str = match.group(0)
        span_start_in_match = full_match_str.lower().find(cleaned_span.lower())
        
        real_start = match.start() + span_start_in_match
        real_end = real_start + len(cleaned_span)
        return real_start, real_end
    
    # Fallback: simple find
    start_index = cleaned_full.find(cleaned_span)
    if start_index != -1:
        return start_index, start_index + len(cleaned_span)
        
    return -1, -1

# 6. Workbook Generation Function
def generate_workbook():
    # Create or Load Template
    if not os.path.exists(TEMPLATE_PATH):
        # Create a dummy workbook if template doesn't exist for this self-contained script
        wb = openpyxl.Workbook()
        ws_names = [
            "Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", 
            "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"
        ]
        # Remove default sheet
        wb.remove(wb.active)
        for name in ws_names:
            wb.create_sheet(name)
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # --- Sheet 1: Note_Text ---
    ws_text = wb["Note_Text"]
    # Clear existing
    for row in ws_text.iter_rows():
        for cell in row:
            cell.value = None
    
    ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # --- Sheet 2: Note_Index ---
    ws_index = wb["Note_Index"]
    # Headers
    headers_index = ["note_id", "source_file", "procedure_date"] + list(PROCEDURE_FLAGS.keys())
    # Clear
    for row in ws_index.iter_rows():
        for cell in row:
            cell.value = None
    ws_index.append(headers_index)
    
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE] + list(PROCEDURE_FLAGS.values())
    ws_index.append(row_data)

    # --- Sheet 3: Span_Annotations ---
    ws_anno = wb["Span_Annotations"]
    headers_anno = ["note_id", "span_text", "label", "normalized_value", "context_prefix", "start_char", "end_char", "event_id"]
    
    # Clear
    for row in ws_anno.iter_rows():
        for cell in row:
            cell.value = None
    ws_anno.append(headers_anno)
    
    for span in SPANS:
        # span structure: (text, label, norm, context, event_id)
        ws_anno.append([NOTE_ID, span[0], span[1], span[2], span[3], "", "", span[4]])

    # --- Sheet 4: Span_Hydrated ---
    ws_hydrated = wb["Span_Hydrated"]
    headers_hyd = ["note_id", "span_text", "label", "normalized_value", "context_prefix", "start_char", "end_char", "event_id"]
    
    # Clear
    for row in ws_hydrated.iter_rows():
        for cell in row:
            cell.value = None
    ws_hydrated.append(headers_hyd)
    
    cleaned_note_text = clean_text(NOTE_TEXT)
    
    for span in SPANS:
        s_text, s_label, s_norm, s_ctx, s_evt = span
        start, end = hydrate_span(cleaned_note_text, s_text, s_ctx)
        ws_hydrated.append([NOTE_ID, s_text, s_label, s_norm, s_ctx, start, end, s_evt])

    # --- Sheet 5: Event_Log ---
    ws_events = wb["Event_Log"]
    headers_evt = ["note_id", "event_id", "event_name", "method", "devices", "anatomy", "observations", "outcomes"]
    
    # Clear
    for row in ws_events.iter_rows():
        for cell in row:
            cell.value = None
    ws_events.append(headers_evt)
    
    for evt in EVENTS:
        ws_events.append([
            NOTE_ID,
            evt["event_id"],
            evt["name"],
            evt["method"],
            ", ".join(evt["devices"]),
            ", ".join(evt["anatomy"]),
            ", ".join(evt["observations"]),
            ", ".join(evt["outcomes"])
        ])

    # --- Sheet 6: V3_Procedure_Events ---
    ws_v3 = wb["V3_Procedure_Events"]
    headers_v3 = ["note_id", "event_id", "event_type", "event_metadata"]
    
    # Clear
    for row in ws_v3.iter_rows():
        for cell in row:
            cell.value = None
    ws_v3.append(headers_v3)
    
    for evt in EVENTS:
        metadata = {
            "method": evt["method"],
            "devices": evt["devices"],
            "anatomy": evt["anatomy"],
            "observations": evt["observations"],
            "outcomes": evt["outcomes"]
        }
        ws_v3.append([NOTE_ID, evt["event_id"], evt["name"], json.dumps(metadata)])

    # --- Sheet 7: V3_Registry_JSON ---
    ws_json = wb["V3_Registry_JSON"]
    # Clear
    for row in ws_json.iter_rows():
        for cell in row:
            cell.value = None
            
    registry_data = {
        "note_id": NOTE_ID,
        "procedure_date": PROCEDURE_DATE,
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS
    }
    
    ws_json.append([json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Workbook generated at: {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()