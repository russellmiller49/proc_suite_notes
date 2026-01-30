import openpyxl
from openpyxl.utils import get_column_letter
import json
import re
import os
from datetime import datetime

# =============================================================================
# INPUT DATA
# =============================================================================

NOTE_ID = "note_163"
SOURCE_FILE = "note_163.txt"
PROCEDURE_DATE = "2026-01-13" # Placeholder
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_163 SOURCE_FILE: note_163.txt PRE-PROCEDURE DIAGNISOS: Multiple pulmonary nodule 
POST- PROCEDURE DIAGNISOS: Multiple pulmonary nodule 
PROCEDURE PERFORMED:  
Flexible bronchoscopy with electromagnetic navigation under flouroscopic and EBUS guidance with transbronchial needle aspiration, transbronchial brush and transbronchial biopsy 
CPT 31654 Bronchoscope with Endobronchial Ultrasound guidance for peripheral lesion
CPT 31629 Flexible bronchoscopy with fluoroscopic trans-bronchial needle aspiration
CPT 31628 Bronchoscopy, rigid or flexible, including fluoroscopic guidance, when performed;
with transbronchial lung biopsy(s), single lobe
CPT 31623 Bronchoscopy, rigid or flexible, including fluoroscopic guidance, when performed;
with brushing or protected brushings
CPT +31627 Bronchoscopy with computer assisted image guided navigation
INDICATIONS FOR EXAMINATION:   multiple pulmonary nodules suspicious for malignancy            
SEDATION: General Anesthesia
FINDINGS: Following intravenous medications as per the record the patient was intubated with an 8. 0 ET tube by anesthesia.
The T190 video bronchoscope was then introduced through the endotracheal tube and advanced to the tracheobronchial tree.
The trachea was of normal caliber. The carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level. Bronchial mucosa and anatomy were normal;
there are no endobronchial lesions. The super-dimension navigational catheter was inserted through the T190 therapeutic bronchoscope and advanced into the airway.
Using the navigational map created preprocedurally we advanced the 190 degree edge catheter into the proximity of the right lower lobe nodule.
Radial probe was used to attempt to confirm presence within the lesion and minor adjustments were made in positioning until a concentric US view was obtained.
Biopsies were then performed with a variety of instruments to include peripheral needle, forceps, and triple needle brush under fluoroscopic visualization.
Rapid Onsite pathological evaluation was consistent with malignancy. Airway inspection was then performed to evaluate for any evidence of active bleeding and none was seen.
The bronchoscope was removed and the procedure completed. Specimens were sent for cytology/histology assessment.
ESTIMATED BLOOD LOSS:   less than 5 cc 
COMPLICATIONS: None
IMPRESSION:  
- Successful navigational bronchoscopy localization and biopsy a right lower lobe nodule
RECOMMENDATIONS
- Transfer to post-procedural unit
- Post-procedure CXR
- D/C home once criteria met
- Await pathology"""

# =============================================================================
# CONFIGURATION & FLAGGING
# =============================================================================

# Flags 0/1
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 1,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 1, # Peripheral needle
    "linear_ebus": 0,
    "radial_ebus": 1,
    "navigational_bronchoscopy": 1,
    "transbronchial_biopsy": 1,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# =============================================================================
# SPAN DATA DEFINITION
# =============================================================================

# Structure: (text, label, norm_value, context_prefix, event_id)
# Event Mapping:
# evt01: General Inspection
# evt02: Navigation & Localization
# evt03: Biopsies (Needle, Forceps, Brush)
# evt04: Outcomes/Plan

RAW_SPANS = [
    # Header/Diagnosis
    ("Multiple pulmonary nodule", "OBS_FINDING", "Pulmonary Nodules", "PRE-PROCEDURE DIAGNISOS:", None),
    
    # Event 1: Inspection
    ("Flexible bronchoscopy", "PROC_METHOD", "Flexible Bronchoscopy", "PROCEDURE PERFORMED:", "evt01"),
    ("tracheobronchial tree", "ANAT_AIRWAY", "Tracheobronchial Tree", "advanced to the", "evt01"),
    ("trachea", "ANAT_AIRWAY", "Trachea", "The", "evt01"),
    ("normal caliber", "OBS_FINDING", "Normal", "trachea was of", "evt01"),
    ("carina", "ANAT_AIRWAY", "Carina", "caliber. The", "evt01"),
    ("sharp", "OBS_FINDING", "Normal", "carina was", "evt01"),
    ("Bronchial mucosa and anatomy were normal", "OBS_FINDING", "Normal", "subsegmental level.", "evt01"),
    ("no endobronchial lesions", "OBS_FINDING", "No Endobronchial Lesions", "there are", "evt01"),
    
    # Event 2: Navigation / Localization
    ("super-dimension navigational catheter", "DEV_CATHETER", "SuperDimension Catheter", "lesions. The", "evt02"),
    ("navigational map", "PROC_METHOD", "Navigational Bronchoscopy", "Using the", "evt02"),
    ("190 degree edge catheter", "DEV_CATHETER", "Edge Catheter", "advanced the", "evt02"),
    ("right lower lobe nodule", "ANAT_LUNG_LOC", "RLL", "proximity of the", "evt02"),
    ("Radial probe", "DEV_INSTRUMENT", "Radial EBUS Probe", "", "evt02"),
    ("concentric US view", "OBS_FINDING", "Concentric View", "positioning until a", "evt02"),

    # Event 3: Biopsy
    ("Biopsies", "PROC_METHOD", "Biopsy", "", "evt03"),
    ("peripheral needle", "DEV_NEEDLE", "Peripheral Needle", "include", "evt03"),
    ("forceps", "DEV_INSTRUMENT", "Biopsy Forceps", "peripheral needle,", "evt03"),
    ("triple needle brush", "DEV_INSTRUMENT", "Cytology Brush", "forceps, and", "evt03"),
    ("fluoroscopic visualization", "PROC_METHOD", "Fluoroscopy", "under", "evt03"),

    # Event 4: Outcomes / ROSE
    ("Rapid Onsite pathological evaluation", "PROC_ACTION", "ROSE", "", "evt04"),
    ("consistent with malignancy", "OBS_ROSE", "Malignancy", "evaluation was", "evt04"),
    ("active bleeding", "OBS_FINDING", "Bleeding", "evidence of", "evt04"),
    ("none was seen", "OUTCOME_COMPLICATION", "None", "bleeding and", "evt04"),
    ("less than 5 cc", "MEAS_VOL", "<5cc", "BLOOD LOSS:", "evt04"),
    ("None", "OUTCOME_COMPLICATION", "None", "COMPLICATIONS:", "evt04"),
    ("Successful navigational bronchoscopy localization", "OBS_FINDING", "Successful Localization", "IMPRESSION:\n -", "evt04")
]

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def clean_text(text):
    return text.replace('\r', '').strip()

def find_offsets(full_text, span_text, context_prefix=None, match_index=None):
    """
    Find start_char, end_char based on hydration logic.
    Returns: (start, end, status_msg)
    """
    if not span_text:
        return None, None, "missing_text"
    
    # 1. Exact match count
    matches = [m.start() for m in re.finditer(re.escape(span_text), full_text)]
    count = len(matches)
    
    if count == 0:
        return None, None, "not_found"
    
    if count == 1:
        return matches[0], matches[0] + len(span_text), "hydrated_unique"
    
    # 2. Context Prefix
    if context_prefix:
        # Look for occurrence where prefix is in preceding 150 chars
        for m_start in matches:
            window_start = max(0, m_start - 150)
            preceding_text = full_text[window_start:m_start]
            if context_prefix in preceding_text:
                return m_start, m_start + len(span_text), "hydrated_prefix_window"
    
    # 3. Match Index
    if match_index is not None and 0 <= match_index < count:
        chosen = matches[match_index]
        return chosen, chosen + len(span_text), "hydrated_match_index"
    
    # Fallback if no context worked but multiple exist
    return None, None, f"ambiguous_count={count}"

# =============================================================================
# WORKBOOK GENERATION
# =============================================================================

def create_workbook():
    # 1. Load Template
    if not os.path.exists(TEMPLATE_PATH):
        # Fallback for demonstration if template doesn't exist, create simple one
        wb = openpyxl.Workbook()
        ws_flags = wb.create_sheet("Note_Index")
        ws_flags.append(["source_file", "note_id", "procedure_date"] + list(PROCEDURE_FLAGS.keys()))
        ws_text = wb.create_sheet("Note_Text")
        ws_text.append(["note_id", "source_file", "note_text"])
        ws_spans = wb.create_sheet("Span_Annotations")
        ws_spans.append(["source_file","note_id","span_id","section_type","context_prefix","span_text","match_index","start_char","end_char","span_len","label","normalized_value","schema_field","event_id","is_negated","is_historical","time_anchor","reviewer","comments","hydration_status"])
        ws_hydrated = wb.create_sheet("Span_Hydrated")
        ws_hydrated.append(["source_file","note_id","span_id","section_type","context_prefix","span_text","match_index","start_char","end_char","span_len","label","normalized_value","schema_field","event_id","is_negated","is_historical","time_anchor","reviewer","comments","hydration_status"])
        wb.create_sheet("Event_Log")
        wb.create_sheet("V3_Procedure_Events")
        wb.create_sheet("V3_Registry_JSON")
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # 2. Populate Note_Text
    ws_text = wb["Note_Text"]
    # Check if empty, if so add header
    if ws_text.max_row == 1 and ws_text.cell(1,1).value is None:
        ws_text.append(["note_id", "source_file", "note_text"])
    
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 3. Populate Note_Index (Flags)
    ws_index = wb["Note_Index"]
    header = [cell.value for cell in ws_index[1]]
    
    row_data = {
        "source_file": SOURCE_FILE,
        "note_id": NOTE_ID,
        "procedure_date": PROCEDURE_DATE,
        "status": "Auto-Generated"
    }
    # Add flags
    row_data.update(PROCEDURE_FLAGS)
    
    new_row = []
    for col in header:
        new_row.append(row_data.get(col, ""))
    ws_index.append(new_row)

    # 4. Process Spans & Hydration
    ws_spans = wb["Span_Annotations"]
    ws_hydrated = wb["Span_Hydrated"]
    
    # Load events for later grouping
    events_map = {} # event_id -> list of dicts

    for i, (span_txt, label, norm, ctx, evt_id) in enumerate(RAW_SPANS):
        span_id = f"span_{i+1:03d}"
        
        # Hydrate
        start, end, status = find_offsets(NOTE_TEXT, span_txt, context_prefix=ctx)
        span_len = len(span_txt) if span_txt else 0
        
        # Common data
        row_dict = {
            "source_file": SOURCE_FILE,
            "note_id": NOTE_ID,
            "span_id": span_id,
            "span_text": span_txt,
            "context_prefix": ctx,
            "label": label,
            "normalized_value": norm,
            "event_id": evt_id,
            "start_char": start,
            "end_char": end,
            "span_len": span_len,
            "hydration_status": status
        }
        
        # Write to Span_Annotations (Blank offsets)
        ws_spans.append([
            SOURCE_FILE, NOTE_ID, span_id, "", ctx, span_txt, "", "", "", span_len,
            label, norm, "", evt_id, "FALSE", "FALSE", "", "Auto", "", "needs_hydration"
        ])

        # Write to Span_Hydrated (Filled offsets)
        ws_hydrated.append([
            SOURCE_FILE, NOTE_ID, span_id, "", ctx, span_txt, "", start if start is not None else "", end if end is not None else "", span_len,
            label, norm, "", evt_id, "FALSE", "FALSE", "", "Auto", "", status
        ])
        
        # Collect for Event Log
        if evt_id:
            if evt_id not in events_map:
                events_map[evt_id] = []
            events_map[evt_id].append(row_dict)

    # 5. Populate Event_Log & V3_Procedure_Events
    ws_events = wb["Event_Log"]
    ws_v3 = wb["V3_Procedure_Events"]
    
    # Ensure headers exist if creating new
    if ws_events.max_row == 1:
        # Add basic header if missing
        ws_events.append(["source_file", "note_id", "event_id", "event_type", "method", "anatomy", "device", "outcomes"])

    v3_registry_list = []

    for evt_id, spans in events_map.items():
        # Aggregate data for this event
        method = next((s['normalized_value'] for s in spans if s['label'] == 'PROC_METHOD'), "")
        action = next((s['normalized_value'] for s in spans if s['label'] == 'PROC_ACTION'), "")
        anatomy = [s['normalized_value'] for s in spans if s['label'] in ('ANAT_AIRWAY', 'ANAT_LUNG_LOC', 'ANAT_PLEURA')]
        devices = [s['normalized_value'] for s in spans if s['label'] in ('DEV_INSTRUMENT', 'DEV_CATHETER', 'DEV_NEEDLE', 'DEV_STENT')]
        outcomes = [s['normalized_value'] for s in spans if s['label'] in ('OUTCOME_SYMPTOMS', 'OUTCOME_COMPLICATION', 'OBS_ROSE')]
        measurements = [f"{s['span_text']} ({s['normalized_value']})" for s in spans if 'MEAS' in s['label']]
        
        # Write flat event log
        ws_events.append([
            SOURCE_FILE, NOTE_ID, evt_id, 
            "Procedure", 
            method or action, 
            ", ".join(anatomy), 
            ", ".join(devices), 
            ", ".join(outcomes)
        ])
        
        # Build V3 Object
        v3_obj = {
            "event_id": evt_id,
            "type": method,
            "method": method,
            "target": {"anatomy": anatomy},
            "devices": devices,
            "outcomes": outcomes,
            "measurements": measurements
        }
        
        ws_v3.append([NOTE_ID, evt_id, method, str(anatomy), "", "", "", "", "", str(devices), str(measurements), "", "", "", "", "", "", "", "", "", ""])
        v3_registry_list.append(v3_obj)

    # 6. JSON Export
    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 1:
        ws_json.append(["schema_version", "note_id", "json_output"])
    
    final_json = {
        "schema_version": "3.0",
        "note_id": NOTE_ID,
        "no_immediate_complications": True, # Derived from note content
        "procedures": v3_registry_list
    }
    
    ws_json.append(["3.0", NOTE_ID, json.dumps(final_json, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    create_workbook()