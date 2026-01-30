import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# ==========================================
# 1. Imports & Constants
# ==========================================

NOTE_ID = "note_164"
SOURCE_FILE = "note_164.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_164 SOURCE_FILE: note_164.txt Indications: Mediastinal adenopathy of unclear etiology
Procedure Performed: CPT 31653 EBUS bronchoscopy single station.
Pre-operative diagnosis: Mediastinal adenopathy 
Post-operative diagnosis: Mediastinal  adenopathy 
Medications: General Anesthesia,
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention.
Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway is in good position. The vocal cords appeared normal. The subglottic space was normal.
The trachea is of normal caliber. The distal right trachea was partially externally compressed and the right mainstem at the orifice was significantly externally compressed (85%) which opened up to normal in the distal right mainstem.
The left mainstem and takeoff were normal. The tracheobronchial tree was examined to at least the first subsegmental level.
Bronchial mucosa and anatomy were normal; there are no endobronchial lesions, and moderate thick secretions.
The video bronchoscope was then removed and the UC180F convex probe EBUS bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
Ultrasound was utilized to identify and measure the radiographically suspicious station paratracheal mediastinal mass (level of 4R lymph node).
The mass measured in short axis at 4.8 cm. Sampling by transbronchial needle aspiration was performed beginning with the Olympus EBUS-TBNA 19 gauge Visioshot-2 needle with a total of 6 needle passes.
Rapid onsite evaluation read as suspicious for lymphoma. All samples were sent for routine cytology or flow cytometry.
Following completion of EBUS bronchoscopy the video bronchoscope was re-inserted and blood was suctioned from the airway.
The bronchoscope was removed and procedure completed. 

Complications: No immediate complications
Estimated Blood Loss: 3 cc.
Post Procedure Diagnosis:
- Technically successful flexible bronchoscopy with endobronchial ultrasound-guided biopsies.
- The patient has remained stable and has been transferred in good condition to the post-surgical monitoring unit.
- Will await final pathology results"""

# ==========================================
# 2. Configuration (Procedure Flags)
# ==========================================

PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# ==========================================
# 3. Data Definition (Spans)
# ==========================================

# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    # Event 01: Initial Inspection
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190 Scope", "tracheobronchial tree, the ", "evt_01"),
    ("distal right trachea", "ANAT_AIRWAY", "Trachea - Distal", "normal caliber. The ", "evt_01"),
    ("externally compressed", "OBS_LESION", "External Compression", "partially ", "evt_01"),
    ("right mainstem", "ANAT_AIRWAY", "RMS", "compressed and the ", "evt_01"),
    ("significantly externally compressed (85%)", "OBS_LESION", "External Compression (85%)", "orifice was ", "evt_01"),
    
    # Event 02: EBUS TBNA
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "UC180F EBUS Scope", "removed and the ", "evt_02"),
    ("level of 4R lymph node", "ANAT_LN_STATION", "4R", "mass (", "evt_02"),
    ("4.8 cm", "MEAS_SIZE", "4.8 cm", "short axis at ", "evt_02"),
    ("transbronchial needle aspiration", "PROC_METHOD", "Linear EBUS TBNA", "Sampling by ", "evt_02"),
    ("Olympus EBUS-TBNA 19 gauge Visioshot-2 needle", "DEV_NEEDLE", "Olympus Visioshot-2 19G", "beginning with the ", "evt_02"),
    ("6 needle passes", "MEAS_COUNT", "6", "total of ", "evt_02"),
    ("suspicious for lymphoma", "OBS_ROSE", "Suspicious for lymphoma", "evaluation read as ", "evt_02"),

    # Outcomes
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications: ", "evt_03"),
]

# ==========================================
# 4. Event Definitions
# ==========================================

EVENTS = [
    {
        "event_id": "evt_01",
        "type": "diagnostic_bronchoscopy",
        "method": "Inspection",
        "devices": ["Q190 video bronchoscope"],
        "anatomy": ["Trachea", "RMS"],
        "outcomes": ["External compression (85%)"]
    },
    {
        "event_id": "evt_02",
        "type": "linear_ebus",
        "method": "TBNA",
        "devices": ["UC180F convex probe EBUS bronchoscope", "Olympus Visioshot-2 19G"],
        "anatomy": ["4R"],
        "outcomes": ["ROSE: Suspicious for lymphoma", "Size: 4.8cm"]
    },
    {
        "event_id": "evt_03",
        "type": "outcome",
        "method": "N/A",
        "devices": [],
        "anatomy": [],
        "outcomes": ["No immediate complications"]
    }
]

# ==========================================
# 5. Helper Functions
# ==========================================

def clean_text(text):
    if not text:
        return ""
    return text.strip().replace('\r\n', '\n').replace('\r', '\n')

def hydrate_span(text, span_text, context_prefix):
    cleaned_text = clean_text(text)
    cleaned_span = clean_text(span_text)
    cleaned_context = clean_text(context_prefix)
    
    if not cleaned_span:
        return None, None
        
    # Create regex pattern with context
    pattern = re.escape(cleaned_context) + r"\s*" + re.escape(cleaned_span)
    match = re.search(pattern, cleaned_text, re.IGNORECASE)
    
    if match:
        # Calculate start/end based on the full match minus the context length
        full_start = match.start()
        # Find where the span actually starts within the match
        # (This is an approximation assuming the context matches 1:1 length-wise mostly)
        # Better approach: find span inside the matched string
        matched_string = match.group(0)
        span_start_in_match = matched_string.lower().rfind(cleaned_span.lower())
        
        start_char = full_start + span_start_in_match
        end_char = start_char + len(cleaned_span)
        return start_char, end_char
    
    # Fallback: simple search without context if context fails (though less accurate)
    match_fallback = re.search(re.escape(cleaned_span), cleaned_text, re.IGNORECASE)
    if match_fallback:
        return match_fallback.start(), match_fallback.end()
        
    return 0, 0

# ==========================================
# 6. Workbook Generation Function
# ==========================================

def generate_workbook():
    # Create workbook
    wb = openpyxl.Workbook()
    
    # ---------------------------
    # Sheet 1: Note_Text
    # ---------------------------
    ws_text = wb.active
    ws_text.title = "Note_Text"
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # ---------------------------
    # Sheet 2: Note_Index
    # ---------------------------
    ws_index = wb.create_sheet("Note_Index")
    headers = ["NOTE_ID", "SOURCE_FILE", "PROCEDURE_DATE"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers)
    
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE]
    for key in PROCEDURE_FLAGS:
        row_data.append(PROCEDURE_FLAGS[key])
    ws_index.append(row_data)
    
    # ---------------------------
    # Sheet 3: Span_Annotations
    # ---------------------------
    ws_spans = wb.create_sheet("Span_Annotations")
    ws_spans.append(["NOTE_ID", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "START_CHAR", "END_CHAR", "EVENT_ID"])
    
    # ---------------------------
    # Sheet 4: Span_Hydrated
    # ---------------------------
    ws_hydrated = wb.create_sheet("Span_Hydrated")
    ws_hydrated.append(["NOTE_ID", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "START_CHAR", "END_CHAR", "EVENT_ID"])
    
    for span in SPANS:
        span_text, label, normalized, context, event_id = span
        
        # Sheet 3 (Blank offsets)
        ws_spans.append([NOTE_ID, span_text, label, normalized, context, "", "", event_id])
        
        # Sheet 4 (Calculated offsets)
        start, end = hydrate_span(NOTE_TEXT, span_text, context)
        ws_hydrated.append([NOTE_ID, span_text, label, normalized, context, start, end, event_id])

    # ---------------------------
    # Sheet 5: Event_Log
    # ---------------------------
    ws_event_log = wb.create_sheet("Event_Log")
    ws_event_log.append(["NOTE_ID", "EVENT_ID", "TYPE", "METHOD", "DEVICES", "ANATOMY", "OUTCOMES"])
    
    for evt in EVENTS:
        ws_event_log.append([
            NOTE_ID,
            evt["event_id"],
            evt["type"],
            evt["method"],
            ", ".join(evt["devices"]),
            ", ".join(evt["anatomy"]),
            ", ".join(evt["outcomes"])
        ])

    # ---------------------------
    # Sheet 6: V3_Procedure_Events
    # ---------------------------
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    ws_v3.append(["NOTE_ID", "EVENT_ID", "EVENT_JSON"])
    
    for evt in EVENTS:
        ws_v3.append([NOTE_ID, evt["event_id"], json.dumps(evt)])

    # ---------------------------
    # Sheet 7: V3_Registry_JSON
    # ---------------------------
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["NOTE_ID", "FULL_JSON"])
    
    full_data = {
        "note_metadata": {
            "note_id": NOTE_ID,
            "source_file": SOURCE_FILE,
            "procedure_date": PROCEDURE_DATE
        },
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS
    }
    
    ws_json.append([NOTE_ID, json.dumps(full_data, indent=2)])
    
    # Save
    wb.save(OUTPUT_PATH)
    print(f"Workbook generated successfully: {OUTPUT_PATH}")

# ==========================================
# 7. Execution Block
# ==========================================

if __name__ == "__main__":
    generate_workbook()