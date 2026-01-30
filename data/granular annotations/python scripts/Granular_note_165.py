import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# 1. Imports & Constants
NOTE_ID = "note_165"
SOURCE_FILE = "note_165.txt"
PROCEDURE_DATE = ""  # Not found in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_165 SOURCE_FILE: note_165.txt PRE-PROCEDURE DIAGNISOS: pulmonary nodule 
POST- PROCEDURE DIAGNISOS: pulmonary nodule 
PROCEDURE PERFORMED:  
Flexible bronchoscopy with electromagnetic navigation under flouroscopic and EBUS guidance with isocyanate green dye injection for surgical resection.
CPT 31654 Bronchoscope with Endobronchial Ultrasound guidance for peripheral lesion
CPT 31629 Flexible bronchoscopy with fluoroscopic trans-bronchial needle aspiration
CPT +31627 Bronchoscopy with computer assisted image guided navigation
INDICATIONS FOR EXAMINATION:   pulmonary nodules requiring surgical resection
SEDATION: General Anesthesia
FINDINGS: Following intravenous medications as per the record the patient was intubated with an 8. 5 ET tube by anesthesia.
The T190 video bronchoscope was then introduced through the endotracheal tube and advanced to the tracheobronchial tree.
The trachea was of normal caliber. The carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level. Bronchial mucosa and anatomy were normal;
there are no endobronchial lesions. The super-dimension navigational catheter was inserted through the T190 therapeutic bronchoscope and advanced into the airway.
Using the navigational map created preprocedurally we advanced the 180 degree edge catheter into the proximity of the right upper lobe nodule.
Radial probe was used to attempt to confirm presence within the lesion.
A super dimension needle was then inserted through the bronchoscope and 0.75 milliliters of isocyanate Green were injected just below the pleura adjacent to the nodule for planned robotic surgical resection immediately following.
The bronchoscope was removed and the procedure was turned over to cardiothoracic surgery.
ESTIMATED BLOOD LOSS:   less than 5 cc 
COMPLICATIONS: None
IMPRESSION:  
- Successful navigational bronchoscopy with ICG marking of right upper lobe peripheral pulmonary nodule"""

# 2. Configuration (Procedure Flags)
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 1,
    "navigational_bronchoscopy": 1,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 3. Data Definition (Spans)
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("Flexible bronchoscopy", "PROC_METHOD", "Flexible bronchoscopy", "PROCEDURE PERFORMED:\n", "evt_01"),
    ("electromagnetic navigation", "PROC_METHOD", "electromagnetic navigation", "Flexible bronchoscopy with ", "evt_02"),
    ("EBUS", "PROC_METHOD", "EBUS", "flouroscopic and ", "evt_03"),
    ("isocyanate green dye injection", "PROC_ACTION", "ICG injection", "guidance with ", "evt_04"),
    ("T190 video bronchoscope", "DEV_INSTRUMENT", "T190 video bronchoscope", "The ", "evt_01"),
    ("trachea", "ANAT_AIRWAY", "trachea", "The ", "evt_01"),
    ("carina", "ANAT_AIRWAY", "carina", "The ", "evt_01"),
    ("tracheobronchial tree", "ANAT_AIRWAY", "tracheobronchial tree", "The ", "evt_01"),
    ("super-dimension navigational catheter", "DEV_CATHETER", "Super-dimension navigational catheter", "The ", "evt_02"),
    ("T190 therapeutic bronchoscope", "DEV_INSTRUMENT", "T190 therapeutic bronchoscope", "through the ", "evt_02"),
    ("180 degree edge catheter", "DEV_CATHETER", "180 degree edge catheter", "advanced the ", "evt_02"),
    ("right upper lobe", "ANAT_LUNG_LOC", "RUL", "proximity of the ", "evt_02"),
    ("Radial probe", "DEV_INSTRUMENT", "Radial probe", "nodule.\n", "evt_03"),
    ("super dimension needle", "DEV_NEEDLE", "Super dimension needle", "A ", "evt_04"),
    ("0.75 milliliters", "MEAS_VOL", "0.75 ml", "bronchoscope and ", "evt_04"),
    ("isocyanate Green", "PROC_METHOD", "ICG", "milliliters of ", "evt_04"),
    ("injected", "PROC_ACTION", "injected", "Green were ", "evt_04"),
    ("Successful navigational bronchoscopy", "OUTCOME_SYMPTOMS", "Successful navigational bronchoscopy", "IMPRESSION:  \n- ", "evt_05"),
    ("None", "OUTCOME_COMPLICATION", "None", "COMPLICATIONS: ", "evt_05"),
]

# 4. Event Definitions (Aggregated for V3)
EVENTS = [
    {
        "event_id": "evt_01",
        "name": "Diagnostic Bronchoscopy",
        "method": "Flexible Bronchoscopy",
        "anatomy": "Trachea, Carina, Tracheobronchial Tree",
        "devices": "T190 video bronchoscope",
        "action": "Inspection",
        "outcome": "Normal airway, no lesions"
    },
    {
        "event_id": "evt_02",
        "name": "Navigation",
        "method": "Electromagnetic Navigation",
        "anatomy": "Right Upper Lobe",
        "devices": "Super-dimension navigational catheter, 180 degree edge catheter, T190 therapeutic bronchoscope",
        "action": "Navigation to target",
        "outcome": "Advanced to proximity"
    },
    {
        "event_id": "evt_03",
        "name": "Verification",
        "method": "Radial EBUS",
        "anatomy": "Right Upper Lobe Nodule",
        "devices": "Radial probe",
        "action": "Verification",
        "outcome": "Attempt to confirm presence"
    },
    {
        "event_id": "evt_04",
        "name": "Marking",
        "method": "ICG Injection",
        "anatomy": "Right Upper Lobe (sub-pleural)",
        "devices": "Super dimension needle",
        "action": "Injection",
        "outcome": "0.75 ml ICG injected for resection"
    },
    {
        "event_id": "evt_05",
        "name": "Procedure Outcome",
        "method": "N/A",
        "anatomy": "N/A",
        "devices": "N/A",
        "action": "Conclusion",
        "outcome": "Successful, None complications, <5cc blood loss"
    }
]

# 5. Helper Functions
def clean_text(text):
    if not text:
        return ""
    return text.strip().replace('\r\n', '\n').replace('\r', '\n')

def hydrate_span(full_text, span_text, context_prefix):
    """
    Finds the start and end offsets of a span in the full text,
    validating against a context prefix to ensure the correct occurrence is found.
    """
    cleaned_full = clean_text(full_text)
    cleaned_span = clean_text(span_text)
    cleaned_context = clean_text(context_prefix)
    
    if not cleaned_span:
        return 0, 0
        
    # Attempt to find the span preceded by the context
    # Escape regex special characters in span and context
    pattern = re.escape(cleaned_context) + r'\s*' + re.escape(cleaned_span)
    match = re.search(pattern, cleaned_full, re.IGNORECASE | re.DOTALL)
    
    if match:
        # The match includes the context, so we need to adjust start index
        full_match_start = match.start()
        full_match_end = match.end()
        
        # Calculate where the span actually starts within the match
        # We search for the span inside the matched string to get precise length offset
        span_in_match = re.search(re.escape(cleaned_span), match.group(0), re.IGNORECASE | re.DOTALL)
        
        if span_in_match:
            final_start = full_match_start + span_in_match.start()
            final_end = full_match_start + span_in_match.end()
            return final_start, final_end
            
    # Fallback: simple search if context fails (less accurate, but functional for simple docs)
    start_index = cleaned_full.find(cleaned_span)
    if start_index != -1:
        return start_index, start_index + len(cleaned_span)
        
    return 0, 0

# 6. Workbook Generation Function
def generate_workbook():
    # A. Create/Load Workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
        
    # B. Sheet 1: Note_Text
    ws_text = wb.create_sheet("Note_Text")
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # C. Sheet 2: Note_Index (Metadata + Flags)
    ws_index = wb.create_sheet("Note_Index")
    headers = ["NOTE_ID", "PROCEDURE_DATE", "SOURCE_FILE"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers)
    values = [NOTE_ID, PROCEDURE_DATE, SOURCE_FILE] + [PROCEDURE_FLAGS[k] for k in PROCEDURE_FLAGS]
    ws_index.append(values)
    
    # D. Sheet 3: Span_Annotations (Raw)
    ws_raw_spans = wb.create_sheet("Span_Annotations")
    ws_raw_spans.append(["NOTE_ID", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "EVENT_ID", "START_CHAR", "END_CHAR"])
    
    for span in SPANS:
        # unpack tuple: (span_text, label, norm_val, context, event_id)
        s_text, s_label, s_norm, s_context, s_evt = span
        ws_raw_spans.append([NOTE_ID, s_text, s_label, s_norm, s_context, s_evt, "", ""])

    # E. Sheet 4: Span_Hydrated (Calculated Offsets)
    ws_hydra = wb.create_sheet("Span_Hydrated")
    ws_hydra.append(["NOTE_ID", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "EVENT_ID", "START_CHAR", "END_CHAR"])
    
    for span in SPANS:
        s_text, s_label, s_norm, s_context, s_evt = span
        start, end = hydrate_span(NOTE_TEXT, s_text, s_context)
        ws_hydra.append([NOTE_ID, s_text, s_label, s_norm, s_context, s_evt, start, end])

    # F. Sheet 5: Event_Log (Flattened)
    ws_evt_log = wb.create_sheet("Event_Log")
    ws_evt_log.append(["NOTE_ID", "EVENT_ID", "EVENT_NAME", "METHOD", "ANATOMY", "DEVICES", "ACTION", "OUTCOME"])
    
    for evt in EVENTS:
        ws_evt_log.append([
            NOTE_ID, 
            evt["event_id"], 
            evt["name"], 
            evt["method"], 
            evt["anatomy"], 
            evt["devices"], 
            evt["action"], 
            evt["outcome"]
        ])

    # G. Sheet 6: V3_Procedure_Events (JSON-embedded)
    ws_v3_evts = wb.create_sheet("V3_Procedure_Events")
    ws_v3_evts.append(["NOTE_ID", "EVENT_ID", "EVENT_TYPE", "EVENT_DETAILS_JSON"])
    
    for evt in EVENTS:
        # Construct a simple JSON representation of the event details
        details = {
            "method": evt["method"],
            "anatomy": evt["anatomy"],
            "devices": evt["devices"],
            "action": evt["action"],
            "outcome": evt["outcome"]
        }
        ws_v3_evts.append([NOTE_ID, evt["event_id"], evt["name"], json.dumps(details)])

    # H. Sheet 7: V3_Registry_JSON (Full Document Object)
    ws_v3_json = wb.create_sheet("V3_Registry_JSON")
    ws_v3_json.append(["NOTE_ID", "JSON_OBJECT"])
    
    # Construct the master JSON object
    registry_obj = {
        "metadata": {
            "note_id": NOTE_ID,
            "source_file": SOURCE_FILE,
            "procedure_date": PROCEDURE_DATE
        },
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS,
        "annotations": [
            {
                "span_text": s[0],
                "label": s[1],
                "normalized_value": s[2],
                "event_id": s[4]
            } for s in SPANS
        ]
    }
    
    ws_v3_json.append([NOTE_ID, json.dumps(registry_obj, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated workbook: {OUTPUT_PATH}")

# 7. Execution Block
if __name__ == "__main__":
    generate_workbook()