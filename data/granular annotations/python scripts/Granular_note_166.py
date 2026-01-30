import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# -------------------------------------------------------------------------
# 1. CONSTANTS & CONFIGURATION
# -------------------------------------------------------------------------
NOTE_ID = "note_166"
SOURCE_FILE = "note_166.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_166 SOURCE_FILE: note_166.txt Indications: Mediastinal adenopathy
Medications: General Anesthesia,
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention.
Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway was in good position. The vocal cords appeared normal. The subglottic space was normal.
The trachea was of normal caliber. The carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level. Bronchial mucosa and anatomy were normal;
there are no endobronchial lesions, and no secretions. The video bronchoscope was then removed and the UC180F convex probe EBUS bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The 11Ri lymph node was identified and sampling by transbronchial needle aspiration was performed with the Olympus 22G EBUS-TBNA needles with a total of 6 passes performed.
Rapid onsite pathological evaluation showed occasional lymphocytes but malignancy was not identified.
The 11Rs lymph node was then identified and sampling by transbronchial needle aspiration was performed with the Olympus 22G EBUS-TBNA needles with a total of 5 passes performed.
Rapid onsite pathological evaluation did not identify malignancy.  Samples were sent for both flow and routine cytology.
Following completion of EBUS bronchoscopy, the Q190 video bronchoscope was then re-inserted and after suctioning blood and secretions there was no evidence of active bleeding and the bronchoscope was subsequently removed.
Complications: No immediate complications
Estimated Blood Loss: 5cc
Post Procedure Diagnosis:
- Technically successful flexible bronchoscopy with endobronchial ultrasound-guided biopsies.
- The patient has remained stable and has been transferred in good condition to the post-procedural monitoring unit.
- Will await final pathology results"""

# -------------------------------------------------------------------------
# 2. PROCEDURE FLAGS
# -------------------------------------------------------------------------
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# -------------------------------------------------------------------------
# 3. SPAN DEFINITIONS
# Format: (span_text, label, normalized_value, context_prefix, event_id)
# -------------------------------------------------------------------------
SPANS = [
    # Scope 1
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190", "tracheobronchial tree, the ", "evt_01"),
    ("laryngeal mask airway", "DEV_INSTRUMENT", "LMA", "through the mouth, via ", "evt_01"),
    ("tracheobronchial tree", "ANAT_AIRWAY", "Tracheobronchial Tree", "advanced to the ", "evt_01"),
    ("Bronchial mucosa and anatomy were normal", "OBS_LESION", "Normal", "first subsegmental level. ", "evt_01"),
    
    # Scope 2 (EBUS)
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "UC180F", "removed and the ", "evt_02"),
    
    # 11Ri
    ("11Ri lymph node", "ANAT_LN_STATION", "11Ri", "The ", "evt_03"),
    ("transbronchial needle aspiration", "PROC_METHOD", "TBNA", "sampling by ", "evt_03"),
    ("Olympus 22G EBUS-TBNA needles", "DEV_NEEDLE", "Olympus 22G", "performed with the ", "evt_03"),
    ("6 passes", "MEAS_COUNT", "6", "total of ", "evt_03"),
    ("malignancy was not identified", "OBS_ROSE", "Negative", "lymphocytes but ", "evt_03"),
    
    # 11Rs
    ("11Rs lymph node", "ANAT_LN_STATION", "11Rs", "The ", "evt_04"),
    ("transbronchial needle aspiration", "PROC_METHOD", "TBNA", "identified and sampling by ", "evt_04"),
    ("Olympus 22G EBUS-TBNA needles", "DEV_NEEDLE", "Olympus 22G", "performed with the ", "evt_04"),
    ("5 passes", "MEAS_COUNT", "5", "total of ", "evt_04"),
    ("did not identify malignancy", "OBS_ROSE", "Negative", "evaluation ", "evt_04"),
    
    # Outcomes
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications: ", "evt_glob"),
    ("5cc", "MEAS_VOL", "5", "Estimated Blood Loss: ", "evt_glob")
]

# -------------------------------------------------------------------------
# 4. EVENT DEFINITIONS (V3 Aggregation)
# -------------------------------------------------------------------------
EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Diagnostic Bronchoscopy",
        "anatomy": "Tracheobronchial Tree",
        "method": "Inspection",
        "devices": ["Q190 video bronchoscope", "LMA"],
        "outcomes": ["Normal anatomy"]
    },
    {
        "event_id": "evt_02",
        "procedure_type": "EBUS",
        "anatomy": "Airway",
        "method": "Scope Introduction",
        "devices": ["UC180F convex probe EBUS bronchoscope"],
        "outcomes": []
    },
    {
        "event_id": "evt_03",
        "procedure_type": "EBUS-TBNA",
        "anatomy": "11Ri",
        "method": "Transbronchial Needle Aspiration",
        "devices": ["Olympus 22G EBUS-TBNA needles"],
        "sampling_count": 6,
        "rose_result": "Negative for malignancy"
    },
    {
        "event_id": "evt_04",
        "procedure_type": "EBUS-TBNA",
        "anatomy": "11Rs",
        "method": "Transbronchial Needle Aspiration",
        "devices": ["Olympus 22G EBUS-TBNA needles"],
        "sampling_count": 5,
        "rose_result": "Negative for malignancy"
    },
    {
        "event_id": "evt_glob",
        "procedure_type": "Global Outcomes",
        "anatomy": "",
        "method": "",
        "devices": [],
        "complications": "No immediate complications",
        "ebl": "5cc"
    }
]

# -------------------------------------------------------------------------
# 5. HELPER FUNCTIONS
# -------------------------------------------------------------------------
def clean_text(text):
    if not text:
        return ""
    return text.replace('\r', '').strip()

def hydrate_span(full_text, span_text, context_prefix):
    """
    Finds the start and end indices of span_text within full_text,
    ensuring it follows context_prefix to handle duplicate phrases.
    """
    clean_full = clean_text(full_text)
    clean_span = clean_text(span_text)
    clean_prefix = clean_text(context_prefix)

    if not clean_span:
        return 0, 0

    # Pattern: Prefix + anything + Span
    # Escape special regex characters in the prefix and span
    escaped_prefix = re.escape(clean_prefix)
    escaped_span = re.escape(clean_span)
    
    # Allow for variable whitespace in the source text vs span text
    pattern = f"({escaped_prefix})(.*?)({escaped_span})"
    
    match = re.search(pattern, clean_full, re.DOTALL | re.IGNORECASE)
    
    if match:
        # match.group(1) is prefix, (2) is gap, (3) is span
        # Start index of span is end of prefix + length of gap
        start_index = match.start(3)
        end_index = match.end(3)
        return start_index, end_index
    else:
        # Fallback: simple find if prefix not found (or strict mode failed)
        # Only use this if you are sure duplicates aren't an issue or specific logic handles it
        idx = clean_full.find(clean_span)
        if idx != -1:
            return idx, idx + len(clean_span)
            
    return -1, -1

# -------------------------------------------------------------------------
# 6. WORKBOOK GENERATION
# -------------------------------------------------------------------------
def generate_workbook():
    # Create workbook (memory only, no template file dependence for the script to run standalone)
    wb = openpyxl.Workbook()
    
    # ---------------------------
    # Sheet 1: Note_Text
    # ---------------------------
    ws_text = wb.active
    ws_text.title = "Note_Text"
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # ---------------------------
    # Sheet 2: Note_Index
    # ---------------------------
    ws_index = wb.create_sheet("Note_Index")
    headers = ["NOTE_ID", "PROCEDURE_DATE"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers)
    row_data = [NOTE_ID, PROCEDURE_DATE] + [PROCEDURE_FLAGS[k] for k in PROCEDURE_FLAGS]
    ws_index.append(row_data)

    # ---------------------------
    # Sheet 3: Span_Annotations
    # ---------------------------
    ws_anno = wb.create_sheet("Span_Annotations")
    ws_anno.append(["NOTE_ID", "SPAN_TEXT", "LABEL", "VALUE", "CONTEXT_PREFIX", "EVENT_ID", "START_CHAR", "END_CHAR"])
    
    for sp in SPANS:
        # sp = (text, label, val, prefix, evt)
        ws_anno.append([NOTE_ID, sp[0], sp[1], sp[2], sp[3], sp[4], "", ""])

    # ---------------------------
    # Sheet 4: Span_Hydrated
    # ---------------------------
    ws_hydro = wb.create_sheet("Span_Hydrated")
    ws_hydro.append(["NOTE_ID", "SPAN_TEXT", "LABEL", "VALUE", "CONTEXT_PREFIX", "EVENT_ID", "START_CHAR", "END_CHAR"])
    
    for sp in SPANS:
        start, end = hydrate_span(NOTE_TEXT, sp[0], sp[3])
        ws_hydro.append([NOTE_ID, sp[0], sp[1], sp[2], sp[3], sp[4], start, end])

    # ---------------------------
    # Sheet 5: Event_Log
    # ---------------------------
    ws_log = wb.create_sheet("Event_Log")
    ws_log.append(["NOTE_ID", "EVENT_ID", "PROCEDURE_TYPE", "ANATOMY", "METHOD", "DEVICES", "OUTCOMES"])
    
    for evt in EVENTS:
        devs = ", ".join(evt.get("devices", []))
        outs = ", ".join(evt.get("outcomes", []))
        if "complications" in evt:
            outs += f"; Complications: {evt['complications']}"
        if "rose_result" in evt:
            outs += f"; ROSE: {evt['rose_result']}"
            
        ws_log.append([
            NOTE_ID,
            evt["event_id"],
            evt["procedure_type"],
            evt.get("anatomy", ""),
            evt.get("method", ""),
            devs,
            outs
        ])

    # ---------------------------
    # Sheet 6: V3_Procedure_Events
    # ---------------------------
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    ws_v3.append(["NOTE_ID", "EVENT_JSON"])
    for evt in EVENTS:
        ws_v3.append([NOTE_ID, json.dumps(evt)])

    # ---------------------------
    # Sheet 7: V3_Registry_JSON
    # ---------------------------
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["NOTE_ID", "FULL_JSON"])
    
    full_registry_data = {
        "note_id": NOTE_ID,
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS
    }
    ws_json.append([NOTE_ID, json.dumps(full_registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Successfully generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()