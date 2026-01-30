import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# =============================================================================
# 1. IMPORTS & CONSTANTS
# =============================================================================

NOTE_ID = "note_167"
SOURCE_FILE = "note_167.txt"
PROCEDURE_DATE = "2026-01-13" # Assumed current or extracted
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

# Cleaned text derived from the input, stripping source tags
NOTE_TEXT = """NOTE_ID:  note_167 SOURCE_FILE: note_167.txt Indications: Mediastinal adenopathy in setting of known lung cancer
Medications: General Anesthesia,
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention.
Following intravenous medications as per the record the Q190 video bronchoscope was introduced through the endotracheal tube and advanced to the tracheobronchial tree.
The trachea was of normal caliber. The carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level. Right sided bronchial mucosa and anatomy were normal;
without endobronchial lesions, and no secretions. Left proximal left mainstem was normal.
There was submucosal tumor infiltration in the distal 3 cm of the left mainstem circumferentially.
The left lower lobe was completely obstructed by endobronchial tumor.
The left upper lobe orifice was approximately 70% obstructed with tumor and the LC1 carina was infiltrated with tumor.
The video bronchoscope was then removed and the UC180F convex probe EBUS bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
A systematic hilar and mediastinal lymph node survey was carried out.
Sampling criteria (5mm short axis diameter) were met in station 7, 4L lymph nodes.
Sampling by transbronchial needle aspiration was performed with the Olympus EBUSTBNA 22 gauge needle beginning with the 7 Lymph node, followed by the 4L lymph node.
ROSE evaluation yielded benign lymphocytes in the station 7 and scant lymphocytes in station 4L.
All samples were sent for routine cytology. Following completion of EBUS bronchoscopy, the Q190 video bronchoscope was then re-inserted and after suctioning blood and secretions there was no evidence of active bleeding and the bronchoscope was subsequently removed.
Decision was made to not proceed with mediastinoscopy as the left mainstem and upper lobe disease would preclude previously considered surgical resection and treatment would not be effected by the presence of N2 disease.
Complications: No immediate complications
Estimated Blood Loss: 5cc

Post Procedure Diagnosis:
- Technically successful flexible bronchoscopy with endobronchial ultrasound-guided biopsies.
- The patient has remained stable and has been transferred in good condition to the post-surgical monitoring unit.
- Will await final pathology results
- Will present at tumor board for consideration of chemo-rads"""

# =============================================================================
# 2. CONFIGURATION (PROCEDURE FLAGS)
# =============================================================================

PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# =============================================================================
# 3. DATA DEFINITION (SPANS)
# =============================================================================

# Tuple format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190", "record the", "evt_01"),
    ("submucosal tumor infiltration", "OBS_LESION", "tumor infiltration", "There was", "evt_01"),
    ("left mainstem", "ANAT_AIRWAY", "Left Mainstem Bronchus", "cm of the", "evt_01"),
    ("Left lower lobe", "ANAT_AIRWAY", "Left Lower Lobe", "The", "evt_01"),
    ("completely obstructed", "OBS_LESION", "Obstruction - Complete", "lobe was", "evt_01"),
    ("Left upper lobe orifice", "ANAT_AIRWAY", "Left Upper Lobe", "The", "evt_01"),
    ("70% obstructed", "OBS_LESION", "Obstruction - Partial", "approximately", "evt_01"),
    ("LC1 carina", "ANAT_AIRWAY", "LC1", "and the", "evt_01"),
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "UC180F", "and the", "evt_02"),
    ("station 7", "ANAT_LN_STATION", "Station 7", "met in", "evt_02"),
    ("4L", "ANAT_LN_STATION", "Station 4L", "station 7,", "evt_03"),
    ("transbronchial needle aspiration", "PROC_METHOD", "TBNA", "Sampling by", "evt_02"),
    ("Olympus EBUSTBNA 22 gauge needle", "DEV_NEEDLE", "Olympus 22G EBUS Needle", "with the", "evt_02"),
    ("7 Lymph node", "ANAT_LN_STATION", "Station 7", "with the", "evt_02"),
    ("4L lymph node", "ANAT_LN_STATION", "Station 4L", "by the", "evt_03"),
    ("ROSE evaluation", "PROC_ACTION", "ROSE", "", "evt_02"),
    ("benign lymphocytes", "OBS_ROSE", "Benign", "yielded", "evt_02"),
    ("scant lymphocytes", "OBS_ROSE", "Scant lymphocytes", "and", "evt_03"),
]

# =============================================================================
# 4. EVENT DEFINITIONS
# =============================================================================

EVENTS = [
    {
        "event_id": "evt_01",
        "method": "Diagnostic Bronchoscopy",
        "anatomy": "Tracheobronchial Tree",
        "devices": ["Q190 video bronchoscope"],
        "outcomes": ["Tumor Infiltration Identified", "Airway Obstruction Identified"],
        "notes": "Inspection revealing submucosal infiltration in left mainstem and obstruction in LLL/LUL."
    },
    {
        "event_id": "evt_02",
        "method": "Linear EBUS TBNA",
        "anatomy": "Station 7",
        "devices": ["UC180F convex probe EBUS bronchoscope", "Olympus EBUSTBNA 22 gauge needle"],
        "outcomes": ["ROSE: Benign lymphocytes"],
        "notes": "Sampling of Station 7."
    },
    {
        "event_id": "evt_03",
        "method": "Linear EBUS TBNA",
        "anatomy": "Station 4L",
        "devices": ["UC180F convex probe EBUS bronchoscope", "Olympus EBUSTBNA 22 gauge needle"],
        "outcomes": ["ROSE: Scant lymphocytes"],
        "notes": "Sampling of Station 4L."
    }
]

# =============================================================================
# 5. HELPER FUNCTIONS
# =============================================================================

def clean_text(text):
    if not text:
        return ""
    return re.sub(r'\s+', ' ', text).strip()

def find_offsets(text, span_text, context_prefix):
    if not span_text:
        return None, None
    
    # Normalize spaces for search
    search_text = clean_text(text).lower()
    search_span = clean_text(span_text).lower()
    search_context = clean_text(context_prefix).lower()
    
    if search_context:
        # Find context first
        context_start = search_text.find(search_context)
        if context_start == -1:
            return None, None # Context not found
        
        # Search for span after context
        start_index = search_text.find(search_span, context_start + len(search_context))
    else:
        start_index = search_text.find(search_span)
        
    if start_index != -1:
        return start_index, start_index + len(search_span)
    
    return None, None

def hydrate_span(text, span_data):
    span_text, label, norm_val, context, evt_id = span_data
    start, end = find_offsets(text, span_text, context)
    
    return {
        "span_text": span_text,
        "label": label,
        "normalized_value": norm_val,
        "start_char": start,
        "end_char": end,
        "event_id": evt_id
    }

# =============================================================================
# 6. WORKBOOK GENERATION FUNCTION
# =============================================================================

def generate_workbook():
    # 1. Load or Create Template
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        # Create expected sheets if they don't exist
        expected_sheets = ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]
        for sheet in expected_sheets:
            if sheet not in wb.sheetnames:
                wb.create_sheet(sheet)
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # 2. Sheet 1: Note_Text
    ws_text = wb["Note_Text"]
    # Clear existing
    for row in ws_text.iter_rows(min_row=2, max_col=3):
        for cell in row:
            cell.value = None
    
    ws_text["A1"] = "note_id"
    ws_text["B1"] = "source_file"
    ws_text["C1"] = "note_text"
    ws_text["A2"] = NOTE_ID
    ws_text["B2"] = SOURCE_FILE
    ws_text["C2"] = NOTE_TEXT

    # 3. Sheet 2: Note_Index
    ws_index = wb["Note_Index"]
    # Headers
    headers = ["note_id", "source_file", "procedure_date"] + list(PROCEDURE_FLAGS.keys())
    for col, header in enumerate(headers, 1):
        ws_index.cell(row=1, column=col, value=header)
    
    # Values
    values = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE] + list(PROCEDURE_FLAGS.values())
    for col, value in enumerate(values, 1):
        ws_index.cell(row=2, column=col, value=value)

    # 4. Sheet 3: Span_Annotations (Raw)
    ws_anno = wb["Span_Annotations"]
    ws_anno.delete_rows(2, ws_anno.max_row)
    anno_headers = ["note_id", "span_text", "label", "normalized_value", "start_char", "end_char", "event_id"]
    for col, header in enumerate(anno_headers, 1):
        ws_anno.cell(row=1, column=col, value=header)
    
    for i, span in enumerate(SPANS, 2):
        # span = (text, label, norm, context, evt)
        ws_anno.cell(row=i, column=1, value=NOTE_ID)
        ws_anno.cell(row=i, column=2, value=span[0])
        ws_anno.cell(row=i, column=3, value=span[1])
        ws_anno.cell(row=i, column=4, value=span[2])
        ws_anno.cell(row=i, column=5, value="") # Blank for raw
        ws_anno.cell(row=i, column=6, value="") # Blank for raw
        ws_anno.cell(row=i, column=7, value=span[4])

    # 5. Sheet 4: Span_Hydrated (Calculated)
    ws_hydra = wb["Span_Hydrated"]
    ws_hydra.delete_rows(2, ws_hydra.max_row)
    for col, header in enumerate(anno_headers, 1):
        ws_hydra.cell(row=1, column=col, value=header)
        
    for i, span in enumerate(SPANS, 2):
        hydrated = hydrate_span(NOTE_TEXT, span)
        ws_hydra.cell(row=i, column=1, value=NOTE_ID)
        ws_hydra.cell(row=i, column=2, value=hydrated["span_text"])
        ws_hydra.cell(row=i, column=3, value=hydrated["label"])
        ws_hydra.cell(row=i, column=4, value=hydrated["normalized_value"])
        ws_hydra.cell(row=i, column=5, value=hydrated["start_char"])
        ws_hydra.cell(row=i, column=6, value=hydrated["end_char"])
        ws_hydra.cell(row=i, column=7, value=hydrated["event_id"])

    # 6. Sheet 5: Event_Log (Flattened)
    ws_event = wb["Event_Log"]
    ws_event.delete_rows(2, ws_event.max_row)
    event_headers = ["note_id", "event_id", "method", "anatomy", "devices", "outcomes"]
    for col, header in enumerate(event_headers, 1):
        ws_event.cell(row=1, column=col, value=header)
        
    for i, evt in enumerate(EVENTS, 2):
        ws_event.cell(row=i, column=1, value=NOTE_ID)
        ws_event.cell(row=i, column=2, value=evt["event_id"])
        ws_event.cell(row=i, column=3, value=evt["method"])
        ws_event.cell(row=i, column=4, value=evt["anatomy"])
        ws_event.cell(row=i, column=5, value=", ".join(evt["devices"]))
        ws_event.cell(row=i, column=6, value=", ".join(evt["outcomes"]))

    # 7. Sheet 6: V3_Procedure_Events (Detailed)
    ws_v3 = wb["V3_Procedure_Events"]
    ws_v3.delete_rows(2, ws_v3.max_row)
    v3_headers = ["note_id", "event_id", "event_type", "event_metadata"]
    for col, header in enumerate(v3_headers, 1):
        ws_v3.cell(row=1, column=col, value=header)
        
    for i, evt in enumerate(EVENTS, 2):
        ws_v3.cell(row=i, column=1, value=NOTE_ID)
        ws_v3.cell(row=i, column=2, value=evt["event_id"])
        ws_v3.cell(row=i, column=3, value=evt["method"])
        ws_v3.cell(row=i, column=4, value=json.dumps(evt))

    # 8. Sheet 7: V3_Registry_JSON (Full Object)
    ws_json = wb["V3_Registry_JSON"]
    ws_json.delete_rows(2, ws_json.max_row)
    ws_json["A1"] = "note_id"
    ws_json["B1"] = "registry_json"
    
    registry_obj = {
        "note_id": NOTE_ID,
        "procedure_date": PROCEDURE_DATE,
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS
    }
    
    ws_json["A2"] = NOTE_ID
    ws_json["B2"] = json.dumps(registry_obj, indent=2)

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Workbook generated: {OUTPUT_PATH}")

# =============================================================================
# 7. EXECUTION BLOCK
# =============================================================================

if __name__ == "__main__":
    generate_workbook()