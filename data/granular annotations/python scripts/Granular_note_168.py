import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import re
import json
import os
import datetime

# =============================================================================
# 1. IMPORTS & CONSTANTS
# =============================================================================

NOTE_ID = "note_168"
SOURCE_FILE = "note_168.txt"
PROCEDURE_DATE = ""  # Not specified in note text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_168 SOURCE_FILE: note_168.txt Indications: Mediastinal Adenopathy with concern for brain mets
Medications: General Anesthesia,
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention.
Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway was in good position. There was partial dynamic obstruction of the trachea of unclear etiology.
The carina was sharp. The tracheobronchial tree was examined to at least the first subsegmental level.
Bronchial mucosa and anatomy were normal; there are no endobronchial lesions, and no secretions.
The video bronchoscope was then removed and the UC180F convex probe EBUS bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
Enlarged mediastinal and hilar lymph nodes were seen in multiple stations in relatively symmetric pattern.
Mediastinal lymph node sampling was performed by transbronchial needle aspiration with the Olympus EBUSTBNA 22 gauge needle beginning with the 11L Lymph node, followed by the 7 lymph node followed by the 4L lymph node and finally the 2R lymph node.
A total of 5 biopsies were performed in each station (station 2R 7 samples were collected).
ROSE evaluation yielded multiple granulomas from the 11L, 4L and 2R stations and adequate lymphocytes from station 7. All samples were sent for routine cytology and two additional samples were taken for AFB and fungal culture from the station 2R.
Following completion of EBUS bronchoscopy, the Q190 video bronchoscope was then re-inserted and after suctioning blood there and secretions there was some slow oozing seen in the bilateral lower lobes.
2ml of 1% lidocaine with epinephrine were instilled into each lower lobe and the bleeding stopped.
Once we were confident that there was no evidence of active bleeding and the bronchoscope was subsequently removed.
Complications: No immediate complications
Estimated Blood Loss: 10ml

Post Procedure Diagnosis:
- Technically successful flexible bronchoscopy with endobronchial ultrasound-guided biopsies.
- The patient has remained stable and has been transferred in good condition to the post-surgical monitoring unit.
- Will await final pathology results"""

# =============================================================================
# 2. CONFIGURATION (PROCEDURE FLAGS)
# =============================================================================

PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# =============================================================================
# 3. DATA DEFINITION (SPANS)
# =============================================================================

# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    # Event 01: Initial Inspection (Diagnostic)
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190", "tracheobronchial tree, the", "evt_01"),
    ("trachea", "ANAT_AIRWAY", "Trachea", "obstruction of the", "evt_01"),
    ("partial dynamic obstruction", "OBS_LESION", "Dynamic Obstruction", "position. There was", "evt_01"),
    ("tracheobronchial tree", "ANAT_AIRWAY", "Tracheobronchial Tree", "advanced to the", "evt_01"),

    # Event 02: EBUS-TBNA
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "UC180F EBUS", "removed and the", "evt_02"),
    ("Mediastinal lymph node sampling", "PROC_METHOD", "TBNA", "relatively symmetric pattern.", "evt_02"),
    ("transbronchial needle aspiration", "PROC_METHOD", "TBNA", "performed by", "evt_02"),
    ("Olympus EBUSTBNA 22 gauge needle", "DEV_NEEDLE", "Olympus 22G", "with the", "evt_02"),
    ("11L", "ANAT_LN_STATION", "11L", "beginning with the", "evt_02"),
    ("7", "ANAT_LN_STATION", "7", "followed by the", "evt_02"),
    ("4L", "ANAT_LN_STATION", "4L", "followed by the", "evt_02"),
    ("2R", "ANAT_LN_STATION", "2R", "finally the", "evt_02"),
    ("5 biopsies", "MEAS_COUNT", "5", "A total of", "evt_02"),
    ("ROSE evaluation yielded multiple granulomas", "OBS_ROSE", "Granulomas", "collected).", "evt_02"),
    
    # Event 03: Management of Oozing/Therapeutic
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190", "EBUS bronchoscopy, the", "evt_03"),
    ("suctioning blood", "PROC_ACTION", "Suction", "inserted and after", "evt_03"),
    ("slow oozing", "OBS_LESION", "Bleeding", "there was some", "evt_03"),
    ("bilateral lower lobes", "ANAT_LUNG_LOC", "Bilateral Lower Lobes", "seen in the", "evt_03"),
    ("instilled", "PROC_ACTION", "Instillation", "epinephrine were", "evt_03"),
    ("2ml", "MEAS_VOL", "2ml", "lower lobes.", "evt_03"),
    ("bleeding stopped", "OUTCOME_COMPLICATION", "Bleeding stopped", "lower lobe and the", "evt_03"),
    
    # Outcomes
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications:", "evt_global"),
    ("Technically successful", "OUTCOME_SYMPTOMS", "Successful", "Diagnosis:", "evt_global")
]

# =============================================================================
# 4. EVENT DEFINITIONS
# =============================================================================

EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Diagnostic Bronchoscopy",
        "devices": ["Q190 video bronchoscope"],
        "anatomy": ["Trachea", "Tracheobronchial Tree"],
        "actions": ["Inspection"],
        "observations": ["Partial dynamic obstruction"]
    },
    {
        "event_id": "evt_02",
        "procedure_type": "Linear EBUS",
        "devices": ["UC180F convex probe EBUS bronchoscope", "Olympus EBUSTBNA 22 gauge needle"],
        "anatomy": ["11L", "7", "4L", "2R"],
        "actions": ["TBNA", "Biopsy"],
        "observations": ["Granulomas (ROSE)"]
    },
    {
        "event_id": "evt_03",
        "procedure_type": "Therapeutic Aspiration",
        "devices": ["Q190 video bronchoscope"],
        "anatomy": ["Bilateral Lower Lobes"],
        "actions": ["Suctioning", "Instillation of Lidocaine/Epi"],
        "observations": ["Slow oozing", "Bleeding stopped"]
    }
]

# =============================================================================
# 5. HELPER FUNCTIONS
# =============================================================================

def clean_text(text):
    """Normalize whitespace and remove carriage returns."""
    if not text:
        return ""
    text = re.sub(r'\r\n', '\n', text)
    text = re.sub(r'\r', '\n', text)
    return text.strip()

def find_offsets(text, span_text, context_prefix):
    """
    Locate span_text within text. Use context_prefix to disambiguate.
    Returns (start, end) indices or (None, None).
    """
    # Normalize inputs for searching
    search_text = text.lower()
    search_span = span_text.lower()
    search_context = context_prefix.lower()

    # Create a pattern that allows for variable whitespace between words
    def make_pattern(s):
        return r"\s+".join(re.escape(word) for word in s.split())

    pattern_span = make_pattern(search_span)
    
    # If context is provided, search for "context + anything + span"
    if search_context:
        pattern_context = make_pattern(search_context)
        # Look for context, followed by optional non-word chars, followed by span
        full_pattern = f"({pattern_context})(.*?)({pattern_span})"
        match = re.search(full_pattern, search_text, re.DOTALL)
        if match:
            # match.group(1) is context, group(2) is gap, group(3) is span
            # We need the start/end of group(3) relative to text
            span_start = match.start(3)
            span_end = match.end(3)
            return span_start, span_end
        else:
            # Fallback: try finding just the span if context match fails
            # (Strictly, we should fail, but this helps if context is slightly off)
            pass

    # Fallback or no context: find first occurrence of span
    match = re.search(pattern_span, search_text)
    if match:
        return match.start(), match.end()
    
    return None, None

def hydrate_span(full_text, span_tpl):
    """
    Takes a SPANS tuple, finds offsets, returns a dict for the Hydrated sheet.
    """
    span_text, label, norm_val, context, evt_id = span_tpl
    clean_note = clean_text(full_text)
    
    start_char, end_char = find_offsets(clean_note, span_text, context)
    
    return {
        "event_id": evt_id,
        "label": label,
        "text": span_text,
        "normalized_value": norm_val,
        "start_char": start_char if start_char is not None else "",
        "end_char": end_char if end_char is not None else "",
        "context_prefix": context
    }

# =============================================================================
# 6. WORKBOOK GENERATION
# =============================================================================

def generate_workbook():
    # A. Create/Load Workbook
    wb = openpyxl.Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)
    
    # Helper to setup headers
    def setup_sheet(name, headers):
        ws = wb.create_sheet(name)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        return ws

    # ---------------------------------------------------------
    # Sheet 1: Note_Text
    # ---------------------------------------------------------
    ws_text = setup_sheet("Note_Text", ["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, clean_text(NOTE_TEXT)])

    # ---------------------------------------------------------
    # Sheet 2: Note_Index
    # ---------------------------------------------------------
    # Headers: note_id, source_file, procedure_date + 30 flags
    flag_keys = list(PROCEDURE_FLAGS.keys())
    headers_index = ["note_id", "source_file", "procedure_date"] + flag_keys
    ws_index = setup_sheet("Note_Index", headers_index)
    
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE]
    for k in flag_keys:
        row_data.append(PROCEDURE_FLAGS[k])
    ws_index.append(row_data)

    # ---------------------------------------------------------
    # Sheet 3: Span_Annotations
    # ---------------------------------------------------------
    # Columns: note_id, event_id, label, text, normalized_value, start_char, end_char
    ws_anno = setup_sheet("Span_Annotations", 
                          ["note_id", "event_id", "label", "text", "normalized_value", "start_char", "end_char"])
    
    for item in SPANS:
        s_text, s_label, s_norm, s_ctx, s_evt = item
        ws_anno.append([NOTE_ID, s_evt, s_label, s_text, s_norm, "", ""])

    # ---------------------------------------------------------
    # Sheet 4: Span_Hydrated
    # ---------------------------------------------------------
    ws_hydra = setup_sheet("Span_Hydrated", 
                           ["note_id", "event_id", "label", "text", "normalized_value", "start_char", "end_char"])
    
    for item in SPANS:
        data = hydrate_span(NOTE_TEXT, item)
        ws_hydra.append([
            NOTE_ID, 
            data["event_id"], 
            data["label"], 
            data["text"], 
            data["normalized_value"], 
            data["start_char"], 
            data["end_char"]
        ])

    # ---------------------------------------------------------
    # Sheet 5: Event_Log
    # ---------------------------------------------------------
    # Flat summary of events
    ws_elog = setup_sheet("Event_Log", ["note_id", "event_id", "procedure_type", "devices", "anatomy", "actions"])
    
    for evt in EVENTS:
        ws_elog.append([
            NOTE_ID,
            evt["event_id"],
            evt["procedure_type"],
            ", ".join(evt.get("devices", [])),
            ", ".join(evt.get("anatomy", [])),
            ", ".join(evt.get("actions", []))
        ])

    # ---------------------------------------------------------
    # Sheet 6: V3_Procedure_Events
    # ---------------------------------------------------------
    # JSON-serialized columns for complex list fields
    ws_v3 = setup_sheet("V3_Procedure_Events", 
                        ["note_id", "event_id", "procedure_type", "devices_json", "anatomy_json", "actions_json", "outcomes_json"])
    
    for evt in EVENTS:
        ws_v3.append([
            NOTE_ID,
            evt["event_id"],
            evt["procedure_type"],
            json.dumps(evt.get("devices", [])),
            json.dumps(evt.get("anatomy", [])),
            json.dumps(evt.get("actions", [])),
            json.dumps(evt.get("observations", []))
        ])

    # ---------------------------------------------------------
    # Sheet 7: V3_Registry_JSON
    # ---------------------------------------------------------
    # Full JSON blob of the procedure data
    ws_json = setup_sheet("V3_Registry_JSON", ["note_id", "json_data"])
    
    registry_data = {
        "note_id": NOTE_ID,
        "source_file": SOURCE_FILE,
        "procedure_date": PROCEDURE_DATE,
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS
    }
    
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])

    # B. Save
    wb.save(OUTPUT_PATH)
    print(f"Successfully generated {OUTPUT_PATH}")

# =============================================================================
# 7. EXECUTION BLOCK
# =============================================================================

if __name__ == "__main__":
    generate_workbook()