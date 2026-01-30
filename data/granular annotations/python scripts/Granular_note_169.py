import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import re
import json
import os
import datetime

# ==========================================
# 1. Imports & Constants
# ==========================================

NOTE_ID = "note_169"
SOURCE_FILE = "note_169.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_169 SOURCE_FILE: note_169.txt PRE-PROCEDURE DIAGNISOS: LEFT UPPER LOBE PULMONARY NODULE
POST- PROCEDURE DIAGNISOS: RIGHT UPPER LOBE CAVITARY PULMONARY NODULE
PROCEDURE PERFORMED:   Flexible bronchoscopy with electromagnetic navigation under flouroscopic and EBUS guidance with transbronchial needle aspiration, Transbronchial biopsy and bronchioalveolar lavage.
MEDICATIONS:    GA
FINDINGS: Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the T190 video bronchoscope was introduced through the laryngeal airway and advanced to the tracheobronchial tree.
The trachea was of normal caliber. The carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level. Bronchial mucosa and anatomy were normal;
there are no endobronchial lesions. The super-dimension navigational catheter was inserted through the T190 therapeutic bronchoscope and advanced into the airway.
Using navigational map we advanced the 180 degree edge catheter into the proximity of the lesion right upper lobe.
Radial probe was used to attempt to confirm presence within the lesion and minor adjustments were made in positioning until a concentric US view was obtained.
Biopsies were then performed with a variety of instruments to include peripheral needle, forceps, and brush under fluoroscopic visualization.
After which a mini-BAL was then performed through the super-D catheter.
We then removed the therapeutic bronchoscope with super-D catheter and reinserted the diagnostic scope at which point repeat airway inspection was then performed and once we were satisfied that no bleeding occurred, the bronchoscope was removed and the procedure completed.
Specimens were sent for both microbiological and cytology/histology assessment.

ESTIMATED BLOOD LOSS:   None 
COMPLICATIONS:                 None

IMPRESSION:  
- S/P bronchoscopy with biopsy and lavage.
- Successful navigational localization and biopsy 
RECOMMENDATIONS
- Transfer to post-procedural unit
- Post-procedure CXR
- D/C home once criteria met
- Await pathology"""

# ==========================================
# 2. Configuration (Procedure Flags)
# ==========================================

PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 1,
    "bronchial_wash": 0,
    "brushings": 1,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 1,
    "linear_ebus": 0,
    "radial_ebus": 1,
    "navigational_bronchoscopy": 1,
    "transbronchial_biopsy": 1,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# ==========================================
# 3. Data Definition (Spans)
# ==========================================

SPANS = [
    # Header/Diagnosis
    ("RIGHT UPPER LOBE", "ANAT_LUNG_LOC", "RUL", "POST- PROCEDURE DIAGNISOS:", "evt_01"),
    
    # Procedure Header
    ("Flexible bronchoscopy", "PROC_METHOD", "diagnostic_bronchoscopy", "PROCEDURE PERFORMED:   ", "evt_01"),
    ("electromagnetic navigation", "PROC_METHOD", "navigational_bronchoscopy", "Flexible bronchoscopy with ", "evt_01"),
    ("flouroscopic", "PROC_METHOD", "fluoroscopy", "under ", "evt_01"),
    ("EBUS guidance", "PROC_METHOD", "radial_ebus", "and ", "evt_01"),
    ("transbronchial needle aspiration", "PROC_METHOD", "tbna_conventional", "with ", "evt_01"),
    ("Transbronchial biopsy", "PROC_METHOD", "transbronchial_biopsy", ", ", "evt_01"),
    ("bronchioalveolar lavage", "PROC_METHOD", "bal", "and ", "evt_01"),

    # Findings - Scope/Devices
    ("T190 video bronchoscope", "DEV_INSTRUMENT", "T190", "tree, the ", "evt_01"),
    ("super-dimension navigational catheter", "DEV_CATHETER", "super_dimension", "The ", "evt_01"),
    ("T190 therapeutic bronchoscope", "DEV_INSTRUMENT", "T190", "through the ", "evt_01"),
    
    # Navigation & Target
    ("180 degree edge catheter", "DEV_CATHETER", "edge_catheter", "advanced the ", "evt_01"),
    ("right upper lobe", "ANAT_LUNG_LOC", "RUL", "proximity of the lesion ", "evt_01"),
    
    # Verification
    ("Radial probe", "DEV_INSTRUMENT", "radial_probe", "", "evt_02"),
    ("concentric US view", "OBS_LESION", "concentric_view", "until a ", "evt_02"),

    # Biopsies
    ("peripheral needle", "DEV_NEEDLE", "peripheral_needle", "include ", "evt_03"),
    ("forceps", "DEV_INSTRUMENT", "forceps", ", ", "evt_03"),
    ("brush", "DEV_INSTRUMENT", "brush", ", and ", "evt_03"),
    
    # BAL
    ("mini-BAL", "PROC_METHOD", "bal", "After which a ", "evt_04"),
    ("super-D catheter", "DEV_CATHETER", "super_dimension", "through the ", "evt_04"),

    # Outcomes
    ("no bleeding occurred", "OUTCOME_COMPLICATION", "None", "satisfied that ", "evt_05"),
    ("None", "MEAS_VOL", "0", "ESTIMATED BLOOD LOSS:   ", "evt_05"),
    ("None", "OUTCOME_COMPLICATION", "None", "COMPLICATIONS:                 ", "evt_05"),
    ("Successful navigational localization", "OUTCOME_COMPLICATION", "Success", "- ", "evt_05")
]

# ==========================================
# 4. Event Definitions
# ==========================================

EVENTS = [
    {
        "event_id": "evt_01",
        "name": "Navigation and Inspection",
        "concept_id": "CPT-31622",
        "method": ["diagnostic_bronchoscopy", "navigational_bronchoscopy"],
        "anatomy": ["RUL"],
        "devices": ["T190 scope", "Super-dimension catheter", "180 degree edge catheter"],
        "outcomes": []
    },
    {
        "event_id": "evt_02",
        "name": "Radial EBUS Verification",
        "concept_id": "CPT-31620",
        "method": ["radial_ebus"],
        "anatomy": ["RUL"],
        "devices": ["Radial probe"],
        "outcomes": ["concentric US view obtained"]
    },
    {
        "event_id": "evt_03",
        "name": "Biopsies (TBNA, TBBX, Brush)",
        "concept_id": "CPT-31629",
        "method": ["tbna_conventional", "transbronchial_biopsy", "brushings"],
        "anatomy": ["RUL"],
        "devices": ["peripheral needle", "forceps", "brush"],
        "outcomes": ["Specimens obtained"]
    },
    {
        "event_id": "evt_04",
        "name": "Mini-BAL",
        "concept_id": "CPT-31624",
        "method": ["bal"],
        "anatomy": ["RUL"],
        "devices": ["Super-D catheter"],
        "outcomes": []
    },
    {
        "event_id": "evt_05",
        "name": "Procedure Completion",
        "concept_id": "N/A",
        "method": [],
        "anatomy": [],
        "devices": [],
        "outcomes": ["No bleeding", "No complications", "Successful localization"]
    }
]

# ==========================================
# 5. Helper Functions
# ==========================================

def clean_text(text):
    if not text:
        return ""
    return text.strip().replace('\r', '')

def hydrate_span(text, span_text, context_prefix):
    """
    Finds the start and end indices of span_text within text,
    using context_prefix to disambiguate.
    """
    cleaned_text = clean_text(text)
    cleaned_span = clean_text(span_text)
    cleaned_context = clean_text(context_prefix)
    
    if not cleaned_span:
        return 0, 0

    # Attempt to find context + span
    search_pattern = re.escape(cleaned_context) + r'\s*' + re.escape(cleaned_span)
    match = re.search(search_pattern, cleaned_text, re.IGNORECASE)
    
    if match:
        # The span starts after the context in the match
        # We need to calculate the actual start/end in the full text relative to the context match
        full_match_str = match.group(0)
        # Find exactly where the span part starts within the match
        span_start_in_match = full_match_str.lower().find(cleaned_span.lower())
        if span_start_in_match == -1:
            # Fallback if case issues or spacing issues, though regex handle most
            span_start_in_match = len(cleaned_context) 
        
        start_index = match.start() + span_start_in_match
        end_index = start_index + len(cleaned_span)
        return start_index, end_index
    
    # Fallback: Find first occurrence of span if context fails (less accurate)
    start_index = cleaned_text.lower().find(cleaned_span.lower())
    if start_index != -1:
        return start_index, start_index + len(cleaned_span)
        
    return 0, 0

# ==========================================
# 6. Workbook Generation Function
# ==========================================

def generate_workbook():
    # A. Create/Load Workbook
    if os.path.exists(TEMPLATE_PATH):
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    else:
        wb = openpyxl.Workbook()
        # Create default sheets if template doesn't exist (fallback)
        for name in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            if name not in wb.sheetnames:
                wb.create_sheet(name)

    # B. Sheet 1: Note_Text
    ws_text = wb["Note_Text"]
    # Clear existing data if any (preserving header if exists, else write header)
    if ws_text.max_row == 1 and ws_text.cell(1,1).value is None:
         ws_text.append(["note_id", "source_file", "note_text"])
    else:
        # If populated, we might overwrite row 2
        pass
    
    # Write Data
    ws_text.cell(row=2, column=1, value=NOTE_ID)
    ws_text.cell(row=2, column=2, value=SOURCE_FILE)
    ws_text.cell(row=2, column=3, value=NOTE_TEXT)

    # C. Sheet 2: Note_Index
    ws_index = wb["Note_Index"]
    # Headers
    headers = ["note_id", "source_file", "procedure_date"] + list(PROCEDURE_FLAGS.keys())
    for col_num, header in enumerate(headers, 1):
        ws_index.cell(row=1, column=col_num, value=header)
    
    # Values
    row_vals = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE] + list(PROCEDURE_FLAGS.values())
    for col_num, val in enumerate(row_vals, 1):
        ws_index.cell(row=2, column=col_num, value=val)

    # D. Sheet 3: Span_Annotations (Manual/Review layer) & Sheet 4: Span_Hydrated (Calculated)
    ws_span_ann = wb["Span_Annotations"]
    ws_span_hyd = wb["Span_Hydrated"]
    
    span_headers = ["note_id", "event_id", "span_text", "label", "normalized_value", "context_prefix", "start_char", "end_char"]
    
    # Write headers for both
    for col_num, h in enumerate(span_headers, 1):
        ws_span_ann.cell(row=1, column=col_num, value=h)
        ws_span_hyd.cell(row=1, column=col_num, value=h)

    # Iterate SPANS
    for idx, span_data in enumerate(SPANS):
        # Unpack
        text_segment, label, norm_val, context, evt_id = span_data
        
        # Hydrate (calculate offsets)
        start, end = hydrate_span(NOTE_TEXT, text_segment, context)
        
        # Write to Annotations (Chars empty usually, but we fill for reference)
        row_num = idx + 2
        ann_row = [NOTE_ID, evt_id, text_segment, label, norm_val, context, "", ""]
        hyd_row = [NOTE_ID, evt_id, text_segment, label, norm_val, context, start, end]
        
        for c, val in enumerate(ann_row, 1):
            ws_span_ann.cell(row=row_num, column=c, value=val)
            
        for c, val in enumerate(hyd_row, 1):
            ws_span_hyd.cell(row=row_num, column=c, value=val)

    # E. Sheet 5: Event_Log (Flattened)
    ws_events = wb["Event_Log"]
    evt_headers = ["note_id", "event_id", "event_name", "method", "anatomy", "devices", "outcomes"]
    for c, h in enumerate(evt_headers, 1):
        ws_events.cell(row=1, column=c, value=h)
        
    for idx, evt in enumerate(EVENTS):
        r = idx + 2
        ws_events.cell(row=r, column=1, value=NOTE_ID)
        ws_events.cell(row=r, column=2, value=evt["event_id"])
        ws_events.cell(row=r, column=3, value=evt["name"])
        ws_events.cell(row=r, column=4, value=", ".join(evt["method"]))
        ws_events.cell(row=r, column=5, value=", ".join(evt["anatomy"]))
        ws_events.cell(row=r, column=6, value=", ".join(evt["devices"]))
        ws_events.cell(row=r, column=7, value=", ".join(evt["outcomes"]))

    # F. Sheet 6: V3_Procedure_Events (JSON nested)
    ws_v3 = wb["V3_Procedure_Events"]
    v3_headers = ["note_id", "event_id", "event_type", "event_metadata_json"]
    for c, h in enumerate(v3_headers, 1):
        ws_v3.cell(row=1, column=c, value=h)
        
    for idx, evt in enumerate(EVENTS):
        r = idx + 2
        ws_v3.cell(row=r, column=1, value=NOTE_ID)
        ws_v3.cell(row=r, column=2, value=evt["event_id"])
        ws_v3.cell(row=r, column=3, value=evt["name"])
        ws_v3.cell(row=r, column=4, value=json.dumps(evt))

    # G. Sheet 7: V3_Registry_JSON (Full Payload)
    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 1 and ws_json.cell(1,1).value is None:
        ws_json.append(["note_id", "full_registry_json"])
        
    full_payload = {
        "metadata": {
            "note_id": NOTE_ID,
            "source_file": SOURCE_FILE,
            "procedure_date": PROCEDURE_DATE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS
    }
    
    ws_json.cell(row=2, column=1, value=NOTE_ID)
    ws_json.cell(row=2, column=2, value=json.dumps(full_payload, indent=2))

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

# ==========================================
# 7. Execution Block
# ==========================================

if __name__ == "__main__":
    generate_workbook()