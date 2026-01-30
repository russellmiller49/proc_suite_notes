import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# --- 1. Imports & Constants ---
NOTE_ID = "note_170"
SOURCE_FILE = "note_170.txt"
PROCEDURE_DATE = "" 
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_170 SOURCE_FILE: note_170.txt Indications: right lower lobe nodule 
Medications: Propofol infusion via anesthesia assistance  
Medications: General Anesthesia,
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention.
Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway was in good position. A polyp wasseen on the left vocal cord.
The subglottic space was normal. The trachea was of normal caliber. The carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level. Bronchial mucosa and anatomy were normal;
there are no endobronchial lesions. We then removed the diagnostic Q190 bronchoscopy and the super-dimension navigational catheter was inserted through the T190 therapeutic bronchoscope and advanced into the airway.
Using navigational map we attempted to advance the 180 degree edge catheter into the proximity of the lesion within the right lower lobe.
Confirmation of placement once at the point of interest with radial ultrasound showed a concentric view within the lesion.
Biopsies were then performed with a variety of instruments to include peripheral needle, brush, triple needle brush and forceps, under fluoroscopic visualization.
After adequate samples were obtained the bronchoscope was removed and the procedure completed
Complications: No immediate complications
Estimated Blood Loss: Less than 5 cc.
Post Procedure Diagnosis:
- Flexible bronchoscopy with successful navigational biopsy of right lower lobe nodule.
-Vocal Cord polyp
Recommendations
- Await final pathology
- ENT consultation for eval of vocal cord polyp"""

# --- 2. Configuration (Procedure Flags) ---
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 1,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 0,
    "radial_ebus": 1,
    "navigational_bronchoscopy": 1,
    "transbronchial_biopsy": 1,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# --- 3. Data Definition (Spans) ---
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190 Bronchoscope", "tracheobronchial tree, the", "evt_01"),
    ("laryngeal mask airway", "DEV_INSTRUMENT", "LMA", "mouth, via", "evt_01"),
    ("polyp", "OBS_LESION", "Polyp", "good position. A", "evt_01"),
    ("left vocal cord", "ANAT_AIRWAY", "Left Vocal Cord", "wasseen on the", "evt_01"),
    ("diagnostic Q190 bronchoscopy", "PROC_METHOD", "Diagnostic Bronchoscopy", "removed the", "evt_01"),
    
    ("super-dimension navigational catheter", "DEV_CATHETER", "SuperDimension Catheter", "bronchoscopy and the", "evt_02"),
    ("T190 therapeutic bronchoscope", "DEV_INSTRUMENT", "T190 Bronchoscope", "through the", "evt_02"),
    ("navigational map", "PROC_METHOD", "Navigation", "Using", "evt_02"),
    ("180 degree edge catheter", "DEV_CATHETER", "Edge Catheter 180", "advance the", "evt_02"),
    ("right lower lobe", "ANAT_LUNG_LOC", "RLL", "within the", "evt_02"),
    ("radial ultrasound", "PROC_METHOD", "Radial EBUS", "interest with", "evt_02"),
    ("concentric view", "OBS_LESION", "Concentric View", "showed a", "evt_02"),
    
    ("Biopsies", "PROC_ACTION", "Biopsy", "within the lesion.", "evt_03"),
    ("peripheral needle", "DEV_NEEDLE", "Peripheral Needle", "include", "evt_03"),
    ("brush", "DEV_INSTRUMENT", "Brush", "peripheral needle,", "evt_03"),
    ("triple needle brush", "DEV_INSTRUMENT", "Triple Needle Brush", "needle, brush,", "evt_03"),
    ("forceps", "DEV_INSTRUMENT", "Forceps", "needle brush and", "evt_03"),
    ("fluoroscopic visualization", "PROC_METHOD", "Fluoroscopy", "forceps, under", "evt_03"),
    
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications:", "evt_global"),
    ("Less than 5 cc", "MEAS_VOL", "5cc", "Blood Loss:", "evt_global")
]

# --- 4. Event Definitions ---
EVENTS = [
    {
        "event_id": "evt_01",
        "type": "Inspection",
        "sub_type": "Diagnostic Bronchoscopy",
        "anatomy": "Airway (Left Vocal Cord)",
        "devices": ["Q190 Bronchoscope", "LMA"],
        "outcomes": ["Polyp observed"]
    },
    {
        "event_id": "evt_02",
        "type": "Navigation/Imaging",
        "sub_type": "Navigational Bronchoscopy/REBUS",
        "anatomy": "RLL",
        "devices": ["T190 Bronchoscope", "SuperDimension Catheter", "Edge Catheter 180", "Radial EBUS"],
        "outcomes": ["Concentric view obtained"]
    },
    {
        "event_id": "evt_03",
        "type": "Biopsy",
        "sub_type": "Transbronchial Biopsy/Brush",
        "anatomy": "RLL",
        "devices": ["Peripheral Needle", "Brush", "Triple Needle Brush", "Forceps", "Fluoroscopy"],
        "outcomes": ["Samples obtained"]
    },
    {
        "event_id": "evt_global",
        "type": "Global",
        "sub_type": "Procedure Outcome",
        "anatomy": "",
        "devices": [],
        "outcomes": ["No immediate complications", "EBL < 5cc"]
    }
]

# --- 5. Helper Functions ---

def clean_text(text):
    return text.replace('\r', '').strip()

def hydrate_span(full_text, span_text, context_prefix):
    """
    Finds the start and end indices of span_text within full_text.
    Uses context_prefix to disambiguate if needed.
    """
    clean_full = clean_text(full_text)
    clean_span = clean_text(span_text)
    clean_context = clean_text(context_prefix)
    
    # Attempt to find with context first
    search_phrase = clean_context + " " + clean_span if clean_context else clean_span
    start_index = clean_full.find(search_phrase)
    
    if start_index != -1:
        # Adjust start_index to point to the start of the span, skipping context
        actual_start = start_index + len(clean_context) + (1 if clean_context else 0)
        return actual_start, actual_start + len(clean_span)
    
    # Fallback: Find first occurrence of span if context fails
    start_index = clean_full.find(clean_span)
    if start_index != -1:
        return start_index, start_index + len(clean_span)
        
    return None, None

# --- 6. Workbook Generation Function ---

def generate_workbook():
    # 1. Initialize Workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # --- Sheet 1: Note_Text ---
    ws_text = wb.create_sheet("Note_Text")
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # --- Sheet 2: Note_Index ---
    ws_index = wb.create_sheet("Note_Index")
    ws_index.append(["NOTE_ID", "PROCEDURE_DATE"] + list(PROCEDURE_FLAGS.keys()))
    ws_index.append([NOTE_ID, PROCEDURE_DATE] + list(PROCEDURE_FLAGS.values()))

    # --- Sheet 3: Span_Annotations ---
    ws_anno = wb.create_sheet("Span_Annotations")
    headers_anno = ["NOTE_ID", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "START_CHAR", "END_CHAR", "EVENT_ID"]
    ws_anno.append(headers_anno)
    
    for span in SPANS:
        # (span_text, label, normalized_value, context_prefix, event_id)
        row = [NOTE_ID, span[0], span[1], span[2], span[3], "", "", span[4]]
        ws_anno.append(row)

    # --- Sheet 4: Span_Hydrated ---
    ws_hydra = wb.create_sheet("Span_Hydrated")
    ws_hydra.append(headers_anno)
    
    cleaned_note_text = clean_text(NOTE_TEXT)
    
    for span in SPANS:
        s_text, label, norm, ctx, evt_id = span
        start, end = hydrate_span(cleaned_note_text, s_text, ctx)
        row = [NOTE_ID, s_text, label, norm, ctx, start, end, evt_id]
        ws_hydra.append(row)

    # --- Sheet 5: Event_Log ---
    ws_event_log = wb.create_sheet("Event_Log")
    ws_event_log.append(["NOTE_ID", "EVENT_ID", "EVENT_TYPE", "SUB_TYPE", "ANATOMY", "DEVICES", "OUTCOMES"])
    
    for evt in EVENTS:
        row = [
            NOTE_ID,
            evt["event_id"],
            evt["type"],
            evt["sub_type"],
            evt["anatomy"],
            ", ".join(evt["devices"]),
            ", ".join(evt["outcomes"])
        ]
        ws_event_log.append(row)

    # --- Sheet 6: V3_Procedure_Events ---
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    ws_v3.append(["NOTE_ID", "PROCEDURE_TYPE", "EVENTS_JSON"])
    
    # Group events for JSON dump
    events_json = json.dumps(EVENTS, indent=2)
    ws_v3.append([NOTE_ID, "Bronchoscopy", events_json])

    # --- Sheet 7: V3_Registry_JSON ---
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["NOTE_ID", "FULL_REGISTRY_ENTRY"])
    
    registry_entry = {
        "metadata": {
            "note_id": NOTE_ID,
            "source": SOURCE_FILE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS,
        "spans": [
            {
                "text": s[0],
                "label": s[1],
                "normalized": s[2],
                "event_id": s[4]
            } for s in SPANS
        ]
    }
    
    ws_json.append([NOTE_ID, json.dumps(registry_entry, indent=2)])

    # Save
    if os.path.exists(OUTPUT_PATH):
        os.remove(OUTPUT_PATH)
    wb.save(OUTPUT_PATH)
    print(f"Workbook generated successfully: {OUTPUT_PATH}")

# --- 7. Execution Block ---
if __name__ == "__main__":
    generate_workbook()