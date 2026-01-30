import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# -------------------------------------------------------------------------
# 1. Imports & Constants
# -------------------------------------------------------------------------
NOTE_ID = "note_171"
SOURCE_FILE = "note_171.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_171 SOURCE_FILE: note_171.txt Procedure Name: EBUS bronchoscopy
Indications: Pulmonary nodule requiring diagnosis/staging.
Medications: Propofol infusion via anesthesia assistance  
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the preprocedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention. 
Following intravenous medications as per the anesthesia record and topical anesthesia to the upper airway and tracheobronchial tree the Q190 video bronchoscope was introduced through the mouth.
The vocal cords appeared normal. The subglottic space was normal. The trachea is of normal caliber. The carina was sharp.
All left sided airways were normal without endobronchial disease to the first segmental branch.
The right upper and lower lobe airways were normal without endobronchial disease to the first segmental branch.
Within the right middle lobe a shiny vascular obstructive endobronchial tumor was visualized within the lateral segment of the right middle lobe.
The lesion was friable and bled easily with minimal manipulation. Subsequently 6 endobronchial needle biopsies of the nodule were performed.
Due to bleeding from the lesion a total of 4ml of 10mg/10ml tranexamic acid was applied topically with subsequent resolution of bleeding.
The video bronchoscope was then removed and the UC180F convex probe EBUS bronchoscope was introduced through the mouth, and advanced to the tracheobronchial tree.
A systematic hilar and mediastinal lymph node survey was carried out.
Sampling criteria (5mm short axis diameter) were met in station 4L, station 4R, station 7 and station 11Rs lymph nodes.
Sampling by transbronchial needle aspiration was performed in these lymph nodes using an Olympus EBUSTBNA 22 gauge needle beginning at the N3 (4L) lymph nodes 74R11Rs.
Further details regarding nodal size and number of samples are included in the EBUS procedural sheet in AHLTA.
All samples were sent for routine cytology. Onsite path evaluation did not identify malignancy.
We then removed the EBUS bronchoscopy and inserted the therapeutic T190 Olympus bronchoscope and advanced the scope into the right middle lobe where 5 forceps endobronchial biopsies were performed.
Bleeding was moderate and required intermittent mechanical tamponade with the tip of the bronchoscope.
After biopsies were completed we monitored for evidence of active bleeding and none was seen.
At this point the bronchoscope was removed and the procedure completed. 

Complications: 
-None 
Estimated Blood Loss:  10 cc.
Recommendations:
- Transfer to PACU
- Await biopsy results 
- Discharge home once criteria met."""

# -------------------------------------------------------------------------
# 2. Configuration (Procedure Flags)
# -------------------------------------------------------------------------
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 1,
    "tbna_conventional": 1, # Non-EBUS needle biopsy performed on tumor
    "linear_ebus": 1,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# -------------------------------------------------------------------------
# 3. Data Definition (Spans)
# -------------------------------------------------------------------------
# Tuple: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    # Scope 1: Q190 - Inspection & Initial Needle Biopsy
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Olympus Q190", "tracheobronchial tree the", "evt_01"),
    ("right middle lobe", "ANAT_LUNG_LOC", "Right Middle Lobe", "Within the", "evt_01"),
    ("lateral segment", "ANAT_LUNG_LOC", "Lateral Segment", "visualized within the", "evt_01"),
    ("shiny vascular obstructive endobronchial tumor", "OBS_LESION", "Endobronchial Tumor", "middle lobe a", "evt_01"),
    ("endobronchial needle biopsies", "PROC_METHOD", "Needle Biopsy", "Subsequently 6", "evt_01"),
    ("6", "MEAS_COUNT", "6", "Subsequently", "evt_01"),
    ("tranexamic acid", "PROC_ACTION", "Tranexamic Acid", "10mg/10ml", "evt_01"),
    
    # Scope 2: EBUS-TBNA
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "Olympus UC180F", "removed and the", "evt_02"),
    ("transbronchial needle aspiration", "PROC_METHOD", "Linear EBUS-TBNA", "Sampling by", "evt_02"),
    ("station 4L", "ANAT_LN_STATION", "Station 4L", "met in", "evt_02"),
    ("station 4R", "ANAT_LN_STATION", "Station 4R", "station 4L,", "evt_02"),
    ("station 7", "ANAT_LN_STATION", "Station 7", "station 4R,", "evt_02"),
    ("station 11Rs", "ANAT_LN_STATION", "Station 11Rs", "and", "evt_02"),
    ("Olympus EBUSTBNA 22 gauge needle", "DEV_NEEDLE", "Olympus EBUS-TBNA 22G", "using an", "evt_02"),
    
    # Scope 3: T190 - Forceps Biopsy
    ("therapeutic T190 Olympus bronchoscope", "DEV_INSTRUMENT", "Olympus T190", "inserted the", "evt_03"),
    ("right middle lobe", "ANAT_LUNG_LOC", "Right Middle Lobe", "scope into the", "evt_03"),
    ("forceps endobronchial biopsies", "PROC_METHOD", "Endobronchial Biopsy", "where 5", "evt_03"),
    ("5", "MEAS_COUNT", "5", "where", "evt_03"),
    ("mechanical tamponade", "PROC_ACTION", "Mechanical Tamponade", "intermittent", "evt_03"),
    
    # Outcomes
    ("None", "OUTCOME_COMPLICATION", "None", "Complications: \n-", "evt_04")
]

# -------------------------------------------------------------------------
# 4. Event Definitions
# -------------------------------------------------------------------------
EVENTS = [
    {
        "event_id": "evt_01",
        "name": "Initial Inspection and Needle Biopsy",
        "method": ["Diagnostic Bronchoscopy", "Needle Biopsy"],
        "anatomy": ["Right Middle Lobe", "Lateral Segment"],
        "devices": ["Olympus Q190"],
        "outcomes": []
    },
    {
        "event_id": "evt_02",
        "name": "EBUS-TBNA",
        "method": ["Linear EBUS-TBNA"],
        "anatomy": ["Station 4L", "Station 4R", "Station 7", "Station 11Rs"],
        "devices": ["Olympus UC180F", "Olympus EBUS-TBNA 22G"],
        "outcomes": ["No malignancy identified (onsite)"]
    },
    {
        "event_id": "evt_03",
        "name": "Therapeutic Scope Biopsy",
        "method": ["Endobronchial Biopsy"],
        "anatomy": ["Right Middle Lobe"],
        "devices": ["Olympus T190", "Forceps"],
        "outcomes": ["Hemostasis achieved"]
    },
    {
        "event_id": "evt_04",
        "name": "Procedure Completion",
        "method": [],
        "anatomy": [],
        "devices": [],
        "outcomes": ["Complications: None"]
    }
]

# -------------------------------------------------------------------------
# 5. Helper Functions
# -------------------------------------------------------------------------
def clean_text(text):
    if not text:
        return ""
    return re.sub(r'[\r\n]+', ' ', text).strip()

def hydrate_span(full_text, span_text, context_prefix):
    """
    Finds start/end offsets for span_text.
    Uses context_prefix to disambiguate if span_text appears multiple times.
    """
    cleaned_full = clean_text(full_text)
    cleaned_span = clean_text(span_text)
    cleaned_context = clean_text(context_prefix)

    if not cleaned_span:
        return None, None

    # Search pattern: context + optional chars + span
    # We escape them to handle special regex chars in medical text
    pattern = re.escape(cleaned_context) + r"\s*(.*?)\s*" + re.escape(cleaned_span)
    
    match = re.search(pattern, cleaned_full, re.IGNORECASE)
    if match:
        # The span is at the end of the match
        full_match_str = match.group(0)
        # Find where the span actually starts within that match
        # (It's at the end)
        start_index = match.start() + full_match_str.lower().rfind(cleaned_span.lower())
        end_index = start_index + len(cleaned_span)
        return start_index, end_index
    
    # Fallback: exact match if context fails (or is empty)
    start_index = cleaned_full.lower().find(cleaned_span.lower())
    if start_index != -1:
        return start_index, start_index + len(cleaned_span)
        
    return None, None

def find_offsets(text, span_text, context_prefix=""):
    return hydrate_span(text, span_text, context_prefix)

# -------------------------------------------------------------------------
# 6. Workbook Generation
# -------------------------------------------------------------------------
def generate_workbook():
    # A. Create/Load Workbook
    if os.path.exists(TEMPLATE_PATH):
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    else:
        wb = openpyxl.Workbook()
        # Create expected sheets if missing
        expected_sheets = ["Note_Text", "Note_Index", "Span_Annotations", 
                           "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]
        for sheet_name in expected_sheets:
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
        # Remove default 'Sheet' if exists
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # B. Sheet 1: Note_Text
    ws_text = wb["Note_Text"]
    # Clear existing
    for row in ws_text.iter_rows(min_row=2, max_col=3):
        for cell in row:
            cell.value = None
    
    ws_text["A1"] = "note_id"
    ws_text["B1"] = "source_file"
    ws_text["C1"] = "note_text"
    ws_text["A2"] = NOTE_ID
    ws_text["B2"] = SOURCE_FILE
    ws_text["C2"] = NOTE_TEXT

    # C. Sheet 2: Note_Index
    ws_index = wb["Note_Index"]
    # Headers
    headers_index = ["note_id", "source_file", "procedure_date"] + list(PROCEDURE_FLAGS.keys())
    for col_num, header in enumerate(headers_index, 1):
        ws_index.cell(row=1, column=col_num, value=header)
    
    # Values
    ws_index.cell(row=2, column=1, value=NOTE_ID)
    ws_index.cell(row=2, column=2, value=SOURCE_FILE)
    ws_index.cell(row=2, column=3, value=PROCEDURE_DATE)
    
    col_idx = 4
    for key in PROCEDURE_FLAGS:
        ws_index.cell(row=2, column=col_idx, value=PROCEDURE_FLAGS[key])
        col_idx += 1

    # D. Sheet 3: Span_Annotations
    ws_anno = wb["Span_Annotations"]
    headers_anno = ["note_id", "span_text", "label", "normalized_value", "context_prefix", "event_id", "start_char", "end_char"]
    for col_num, header in enumerate(headers_anno, 1):
        ws_anno.cell(row=1, column=col_num, value=header)

    # Write Spans (Annotated) - leave offsets blank for now
    row_num = 2
    for span in SPANS:
        s_text, s_label, s_norm, s_ctx, s_evt = span
        ws_anno.cell(row=row_num, column=1, value=NOTE_ID)
        ws_anno.cell(row=row_num, column=2, value=s_text)
        ws_anno.cell(row=row_num, column=3, value=s_label)
        ws_anno.cell(row=row_num, column=4, value=s_norm)
        ws_anno.cell(row=row_num, column=5, value=s_ctx)
        ws_anno.cell(row=row_num, column=6, value=s_evt)
        row_num += 1

    # E. Sheet 4: Span_Hydrated
    ws_hydra = wb["Span_Hydrated"]
    for col_num, header in enumerate(headers_anno, 1):
        ws_hydra.cell(row=1, column=col_num, value=header)

    row_num = 2
    for span in SPANS:
        s_text, s_label, s_norm, s_ctx, s_evt = span
        start, end = hydrate_span(NOTE_TEXT, s_text, s_ctx)
        
        ws_hydra.cell(row=row_num, column=1, value=NOTE_ID)
        ws_hydra.cell(row=row_num, column=2, value=s_text)
        ws_hydra.cell(row=row_num, column=3, value=s_label)
        ws_hydra.cell(row=row_num, column=4, value=s_norm)
        ws_hydra.cell(row=row_num, column=5, value=s_ctx)
        ws_hydra.cell(row=row_num, column=6, value=s_evt)
        ws_hydra.cell(row=row_num, column=7, value=start)
        ws_hydra.cell(row=row_num, column=8, value=end)
        row_num += 1

    # F. Sheet 5: Event_Log (Flattened)
    ws_elog = wb["Event_Log"]
    headers_elog = ["note_id", "event_id", "category", "value"]
    for col_num, header in enumerate(headers_elog, 1):
        ws_elog.cell(row=1, column=col_num, value=header)

    row_num = 2
    for evt in EVENTS:
        eid = evt["event_id"]
        # Methods
        for m in evt["method"]:
            ws_elog.cell(row=row_num, column=1, value=NOTE_ID)
            ws_elog.cell(row=row_num, column=2, value=eid)
            ws_elog.cell(row=row_num, column=3, value="Method")
            ws_elog.cell(row=row_num, column=4, value=m)
            row_num += 1
        # Anatomy
        for a in evt["anatomy"]:
            ws_elog.cell(row=row_num, column=1, value=NOTE_ID)
            ws_elog.cell(row=row_num, column=2, value=eid)
            ws_elog.cell(row=row_num, column=3, value="Anatomy")
            ws_elog.cell(row=row_num, column=4, value=a)
            row_num += 1
        # Devices
        for d in evt["devices"]:
            ws_elog.cell(row=row_num, column=1, value=NOTE_ID)
            ws_elog.cell(row=row_num, column=2, value=eid)
            ws_elog.cell(row=row_num, column=3, value="Device")
            ws_elog.cell(row=row_num, column=4, value=d)
            row_num += 1
        # Outcomes
        for o in evt["outcomes"]:
            ws_elog.cell(row=row_num, column=1, value=NOTE_ID)
            ws_elog.cell(row=row_num, column=2, value=eid)
            ws_elog.cell(row=row_num, column=3, value="Outcome")
            ws_elog.cell(row=row_num, column=4, value=o)
            row_num += 1

    # G. Sheet 6: V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    headers_v3 = ["note_id", "event_id", "event_name", "method", "anatomy", "devices", "outcomes"]
    for col_num, header in enumerate(headers_v3, 1):
        ws_v3.cell(row=1, column=col_num, value=header)
    
    row_num = 2
    for evt in EVENTS:
        ws_v3.cell(row=row_num, column=1, value=NOTE_ID)
        ws_v3.cell(row=row_num, column=2, value=evt["event_id"])
        ws_v3.cell(row=row_num, column=3, value=evt["name"])
        ws_v3.cell(row=row_num, column=4, value=json.dumps(evt["method"]))
        ws_v3.cell(row=row_num, column=5, value=json.dumps(evt["anatomy"]))
        ws_v3.cell(row=row_num, column=6, value=json.dumps(evt["devices"]))
        ws_v3.cell(row=row_num, column=7, value=json.dumps(evt["outcomes"]))
        row_num += 1

    # H. Sheet 7: V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    ws_json["A1"] = "note_id"
    ws_json["B1"] = "json_object"
    
    registry_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source_file": SOURCE_FILE,
            "procedure_date": PROCEDURE_DATE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS
    }
    
    ws_json["A2"] = NOTE_ID
    ws_json["B2"] = json.dumps(registry_data, indent=2)

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated workbook: {OUTPUT_PATH}")

# -------------------------------------------------------------------------
# 7. Execution Block
# -------------------------------------------------------------------------
if __name__ == "__main__":
    generate_workbook()