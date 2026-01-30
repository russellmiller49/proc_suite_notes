import openpyxl
from openpyxl import Workbook
import re
import json
import os
import datetime

# ==============================================================================
# 1. IMPORTS & CONSTANTS
# ==============================================================================
NOTE_ID = "note_172"
SOURCE_FILE = "note_172.txt"
PROCEDURE_DATE = ""  # Not found in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_172 SOURCE_FILE: note_172.txt Indications: Left upper lobe nodule 
Medications: Propofol infusion via anesthesia assistance  
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the preprocedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention. 
Following intravenous medications as per the anesthesia record and topical anesthesia to the upper airway and tracheobronchial tree the Q190 video bronchoscope was introduced through the mouth.
The vocal cords appeared normal. The subglottic space was normal. The trachea is of normal caliber. The carina was sharp.
All left and right sided airways were normal without endobronchial disease.
The video bronchoscope was then removed and the UC180F convex probe EBUS bronchoscope was introduced through the mouth, and advanced to the tracheobronchial tree.
A systematic hilar and mediastinal lymph node survey was carried out.
Sampling criteria (5mm short axis diameter) were met in station 7 and 11L.
Sampling by transbronchial needle aspiration was performed in these lymph nodes using an Olympus Vizishot  EBUS TBNA 19 gauge needle.
All samples were sent for routine cytology. We then tried to visualize the FDG avid station 5 lymph node seen on PET CT through the pulmonary artery but were unable to clearly visualize the node and thus could not sample from the mainstem approach (transvascular).
We then removed the EBUS bronchoscopy. At this point the patient had lost her IV access and developed laryngospasm prompting us to bronchoscopically intubate the patient with a size 8.5 ETT which was easily passed through the vocal cords and into the lower trachea before being secured in place.
The Q190 bronchscope was then advanced, using anatomical knowledge,  into the proximal left upper lobe adjacent to the station 5 lymph node and utilizing radial EBUS attempted to sample this node through an airway approach with a transbronchial needle approach.
ROSE was unrevealing and we subsequently the super-dimension navigational catheter was inserted through the therapeutic bronchoscope and advanced into the airway.
Using navigational map we advanced the 180 degree edge catheter into the proximity of the lesion within the left upper lobe.
Confirmation of placement was attempted once we were in the vicinity of the point of interest with radial ultrasound.
The lesion however could not be adequately visualized. Multiple attempts to navigate to the lesion were unfruitful and we subsequently completed the procedure without attempting to sample the left upper lobe peripheral lesion.
Complications: No immediate complications
Estimated Blood Loss: Less than 5 cc.
Post Procedure Diagnosis:
- Negative Staging EBUS on preliminary evaluation 
- Flexible bronchoscopy with unsuccessful electromagnetic navigation to the left upper nodule
- Await final pathology  
- Discharge once post-op criteria met"""

# ==============================================================================
# 2. CONFIGURATION (PROCEDURE FLAGS)
# ==============================================================================
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 1,
    "navigational_bronchoscopy": 1,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# ==============================================================================
# 3. DATA DEFINITION (SPANS)
# ==============================================================================
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    # Event 01: Inspection
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190 Bronchoscope", "tracheobronchial tree the", "evt_01"),
    ("vocal cords", "ANAT_AIRWAY", "Vocal Cords", "The", "evt_01"),
    ("subglottic space", "ANAT_AIRWAY", "Subglottic Space", "normal. The", "evt_01"),
    ("trachea", "ANAT_AIRWAY", "Trachea", "normal. The", "evt_01"),
    ("carina", "ANAT_AIRWAY", "Carina", "sharp. The", "evt_01"),
    
    # Event 02: Linear EBUS (Station 7, 11L)
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "UC180F EBUS Scope", "removed and the", "evt_02"),
    ("station 7", "ANAT_LN_STATION", "Station 7", "met in", "evt_02"),
    ("11L", "ANAT_LN_STATION", "Station 11L", "station 7 and", "evt_02"),
    ("transbronchial needle aspiration", "PROC_METHOD", "TBNA", "Sampling by", "evt_02"),
    ("Olympus Vizishot  EBUS TBNA", "DEV_NEEDLE", "Olympus Vizishot", "using an", "evt_02"),
    ("19 gauge", "DEV_CATHETER_SIZE", "19G", "EBUS TBNA", "evt_02"),
    
    # Event 03: Failed Transvascular Stn 5
    ("station 5", "ANAT_LN_STATION", "Station 5", "avid", "evt_03"),
    ("pulmonary artery", "ANAT_LUNG_LOC", "Pulmonary Artery", "through the", "evt_03"),
    ("mainstem approach (transvascular)", "PROC_METHOD", "Transvascular EBUS", "sample from the", "evt_03"),
    
    # Event 04: Complication / Intubation
    ("laryngospasm", "OUTCOME_COMPLICATION", "Laryngospasm", "developed", "evt_04"),
    ("intubate", "PROC_ACTION", "Intubation", "bronchoscopically", "evt_04"),
    ("size 8.5", "DEV_CATHETER_SIZE", "8.5", "with a", "evt_04"),
    ("ETT", "DEV_CATHETER", "Endotracheal Tube", "8.5", "evt_04"),
    
    # Event 05: Radial EBUS Stn 5 (Airway)
    ("Q190 bronchscope", "DEV_INSTRUMENT", "Q190 Bronchoscope", "The", "evt_05"),
    ("station 5", "ANAT_LN_STATION", "Station 5", "adjacent to the", "evt_05"),
    ("radial EBUS", "PROC_METHOD", "Radial EBUS", "utilizing", "evt_05"),
    ("transbronchial needle", "DEV_NEEDLE", "TBNA Needle", "with a", "evt_05"),
    
    # Event 06: Navigation LUL
    ("ROSE was unrevealing", "OBS_ROSE", "Negative", "", "evt_06"),
    ("super-dimension navigational catheter", "DEV_CATHETER", "SuperDimension Catheter", "subsequently the", "evt_06"),
    ("180 degree edge catheter", "DEV_CATHETER", "Edge Catheter 180", "advanced the", "evt_06"),
    ("left upper lobe", "ANAT_LUNG_LOC", "LUL", "within the", "evt_06"),
    ("radial ultrasound", "PROC_METHOD", "Radial EBUS", "interest with", "evt_06"),
    ("lesion", "OBS_LESION", "Lesion", "proximity of the", "evt_06"),
    ("not be adequately visualized", "OUTCOME_COMPLICATION", "Visualization Failure", "however could", "evt_06"),
    ("unsuccessful electromagnetic navigation", "OUTCOME_COMPLICATION", "Navigation Failure", "bronchoscopy with", "evt_06"),
    
    # Global
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications:", "evt_global"),
    ("Less than 5 cc", "MEAS_VOL", "<5cc", "Loss:", "evt_global"),
]

# ==============================================================================
# 4. EVENT DEFINITIONS
# ==============================================================================
EVENTS = [
    {
        "event_id": "evt_01",
        "type": "diagnostic",
        "description": "Flexible Bronchoscopy Inspection",
        "devices": ["Q190 video bronchoscope"],
        "anatomy": ["vocal cords", "trachea", "carina"],
        "outcomes": ["Normal anatomy"]
    },
    {
        "event_id": "evt_02",
        "type": "therapeutic",
        "description": "Linear EBUS TBNA Stn 7, 11L",
        "devices": ["UC180F EBUS", "Olympus Vizishot 19G"],
        "anatomy": ["Station 7", "Station 11L"],
        "outcomes": ["Samples obtained"]
    },
    {
        "event_id": "evt_03",
        "type": "diagnostic",
        "description": "Attempted Transvascular Stn 5",
        "devices": ["UC180F EBUS"],
        "anatomy": ["Station 5"],
        "outcomes": ["Visualization failure"]
    },
    {
        "event_id": "evt_04",
        "type": "complication_management",
        "description": "Intubation for Laryngospasm",
        "devices": ["8.5 ETT"],
        "anatomy": ["Trachea"],
        "outcomes": ["Laryngospasm resolved", "Airway secured"]
    },
    {
        "event_id": "evt_05",
        "type": "therapeutic",
        "description": "Radial EBUS TBNA Stn 5 (Airway approach)",
        "devices": ["Q190", "Radial EBUS", "TBNA Needle"],
        "anatomy": ["Station 5"],
        "outcomes": ["ROSE unrevealing"]
    },
    {
        "event_id": "evt_06",
        "type": "diagnostic",
        "description": "Navigational Bronchoscopy LUL",
        "devices": ["SuperDimension", "Edge Catheter"],
        "anatomy": ["Left Upper Lobe"],
        "outcomes": ["Lesion not visualized", "Procedure aborted"]
    }
]

# ==============================================================================
# 5. HELPER FUNCTIONS
# ==============================================================================
def clean_text(text):
    if not text: return ""
    return text.replace('\r', '').strip()

def hydrate_span(full_text, span_text, context_prefix):
    """
    Finds the start/end offsets of span_text within full_text.
    Uses context_prefix to disambiguate.
    """
    clean_full = clean_text(full_text)
    clean_span = clean_text(span_text)
    clean_context = clean_text(context_prefix)
    
    if not clean_span or clean_span not in clean_full:
        return "", "", ""
    
    # Regex pattern: look for context followed by span within reasonable distance
    pattern = re.escape(clean_context) + r"\s*(.*?)\s*" + re.escape(clean_span)
    
    # If context provided, try to find specific instance
    if clean_context:
        matches = list(re.finditer(pattern, clean_full, re.DOTALL))
        if matches:
            # We want the span part. The span is at the end of the match.
            # match.end() is end of span. match.start() is start of context.
            # We need start of span.
            full_match = matches[0]
            # Find the span inside the match text to calculate exact offset
            # This is slightly tricky because of whitespace normalization in regex
            # Simple approach: find clean_span starting at full_match.start()
            start_search = full_match.start()
            actual_start = clean_full.find(clean_span, start_search)
            actual_end = actual_start + len(clean_span)
            return actual_start, actual_end, clean_full[actual_start:actual_end]
            
    # Fallback: Find first occurrence if context fails or is empty
    start_index = clean_full.find(clean_span)
    if start_index != -1:
        return start_index, start_index + len(clean_span), clean_full[start_index:start_index + len(clean_span)]
    
    return "", "", ""

# ==============================================================================
# 6. WORKBOOK GENERATION
# ==============================================================================
def generate_workbook():
    # 1. Initialize Workbook
    wb = Workbook()
    
    # 2. Sheet: Note_Text
    ws_text = wb.active
    ws_text.title = "Note_Text"
    ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # 3. Sheet: Note_Index
    ws_index = wb.create_sheet("Note_Index")
    headers = ["note_id", "source_file", "procedure_date"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers)
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE] + list(PROCEDURE_FLAGS.values())
    ws_index.append(row_data)
    
    # 4. Sheet: Span_Annotations (Blank offsets for manual review/input structure)
    ws_annot = wb.create_sheet("Span_Annotations")
    ws_annot.append(["note_id", "event_id", "label", "span_text", "normalized_value", "start_char", "end_char"])
    for span in SPANS:
        # span: (text, label, norm, context, evt_id)
        ws_annot.append([NOTE_ID, span[4], span[1], span[0], span[2], "", ""])
        
    # 5. Sheet: Span_Hydrated (Calculated offsets)
    ws_hydrated = wb.create_sheet("Span_Hydrated")
    ws_hydrated.append(["note_id", "event_id", "label", "span_text", "normalized_value", "start_char", "end_char", "context_used"])
    
    for span in SPANS:
        s_text, label, norm, context, evt_id = span
        start, end, found_text = hydrate_span(NOTE_TEXT, s_text, context)
        ws_hydrated.append([NOTE_ID, evt_id, label, found_text, norm, start, end, context])

    # 6. Sheet: Event_Log
    ws_events = wb.create_sheet("Event_Log")
    ws_events.append(["note_id", "event_id", "event_type", "description", "devices", "anatomy", "outcomes"])
    for evt in EVENTS:
        ws_events.append([
            NOTE_ID,
            evt["event_id"],
            evt["type"],
            evt["description"],
            ", ".join(evt["devices"]),
            ", ".join(evt["anatomy"]),
            ", ".join(evt["outcomes"])
        ])

    # 7. Sheet: V3_Procedure_Events (Compat)
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    ws_v3.append(["note_id", "procedure_type", "event_data_json"])
    # Aggregate specific events into procedure types if needed, or just dump raw
    # For Phase 0, we typically just output what we found
    ws_v3.append([NOTE_ID, "Bronchoscopy", json.dumps(EVENTS)])

    # 8. Sheet: V3_Registry_JSON
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["note_id", "json_blob"])
    
    registry_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source": SOURCE_FILE,
            "flags": PROCEDURE_FLAGS
        },
        "spans": [
            {
                "text": s[0],
                "label": s[1],
                "normalized": s[2],
                "event_id": s[4]
            } for s in SPANS
        ],
        "events": EVENTS
    }
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Successfully generated {OUTPUT_PATH}")

# ==============================================================================
# 7. EXECUTION BLOCK
# ==============================================================================
if __name__ == "__main__":
    generate_workbook()