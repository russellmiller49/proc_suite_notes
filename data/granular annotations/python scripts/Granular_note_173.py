import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import re
import json
import os
import datetime

# -------------------------------------------------------------------------
# 1. CONSTANTS & INPUT DATA
# -------------------------------------------------------------------------

NOTE_ID = "note_173"
SOURCE_FILE = "note_173.txt"
PROCEDURE_DATE = ""
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_173 SOURCE_FILE: note_173.txt Procedure Name: Peripheral bronchoscopy with radial EBUS localization.
Indications: Pulmonary nodule requiring diagnosis
Medications: Propofol infusion via anesthesia assistance  
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention. 
Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway was in good position. The vocal cords appeared normal. The subglottic space was normal.
The trachea was of normal caliber. The carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level without endobronchial lesions visualized.
Anatomy was normal with exception of what appears to be a fused anterior segment of left upper lobe with lingula.
The video bronchoscope was then removed and the T190 Therapeutic video bronchoscope was inserted into the airway.
A sheath catheter was advanced through the working channel into the segment of suspicion.
A radial US was advanced through the sheath and concentric view of lesion was seen.
Using fluoroscopy, transbronchial needle biopsies were performed. ROSE was consistent with malignancy.
We then performed 6 bronchoscopic lung biopsies in the same area under fluoroscopic visualization with forceps.
We then removed the therapeutic scope and re0inserted the Q190 videoscope.
After cleaning of blood and debris no active bleeding and none was identified.
Fluoroscopy was then used to scan for evidence of pneumothorax which was not seen.
The bronchoscope was removed and the procedure completed. 

Complications: 	
-None 
Estimated Blood Loss:  5 cc
Recommendations:
- post-procedure CXR
- Await biopsy results 
- Discharge home once criteria met."""

# -------------------------------------------------------------------------
# 2. CONFIGURATION (PROCEDURE FLAGS)
# -------------------------------------------------------------------------

PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 1,
    "linear_ebus": 0,
    "radial_ebus": 1,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 1,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# -------------------------------------------------------------------------
# 3. DATA DEFINITION (SPANS)
# -------------------------------------------------------------------------
# Format: (span_text, label, normalized_value, context_prefix, event_id)

SPANS = [
    ("Peripheral bronchoscopy", "PROC_METHOD", "Peripheral Bronchoscopy", "Procedure Name: ", "evt_01"),
    ("radial EBUS", "PROC_METHOD", "Radial EBUS", "bronchoscopy with ", "evt_01"),
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190 Scope", "tracheobronchial tree, the ", "evt_01"),
    ("fused anterior segment of left upper lobe with lingula", "ANAT_LUNG_LOC", "Left Upper Lobe", "appears to be a ", "evt_01"),
    ("T190 Therapeutic video bronchoscope", "DEV_INSTRUMENT", "T190 Scope", "removed and the ", "evt_02"),
    ("sheath catheter", "DEV_CATHETER", "Sheath", "A ", "evt_02"),
    ("radial US", "PROC_METHOD", "Radial EBUS", "A ", "evt_02"),
    ("transbronchial needle biopsies", "PROC_METHOD", "TBNA", "Using fluoroscopy, ", "evt_02"),
    ("consistent with malignancy", "OBS_ROSE", "Malignancy", "ROSE was ", "evt_02"),
    ("6", "MEAS_COUNT", "6", "We then performed ", "evt_03"),
    ("bronchoscopic lung biopsies", "PROC_METHOD", "Transbronchial Biopsy", "performed 6 ", "evt_03"),
    ("forceps", "DEV_INSTRUMENT", "Forceps", "visualization with ", "evt_03"),
    ("None", "OUTCOME_COMPLICATION", "None", "Complications: \t\n-", "evt_99")
]

# -------------------------------------------------------------------------
# 4. EVENT DEFINITIONS (JSON Structure)
# -------------------------------------------------------------------------

EVENTS = [
    {
        "event_id": "evt_01",
        "type": "diagnostic_bronchoscopy",
        "method": ["Peripheral Bronchoscopy", "Radial EBUS"],
        "devices": ["Q190 Scope"],
        "anatomy": ["Left Upper Lobe"],
        "outcomes": []
    },
    {
        "event_id": "evt_02",
        "type": "tbna_conventional",
        "method": ["Radial EBUS", "TBNA"],
        "devices": ["T190 Scope", "Sheath"],
        "anatomy": ["Left Upper Lobe"], # Inferred from context
        "outcomes": ["ROSE: Malignancy"]
    },
    {
        "event_id": "evt_03",
        "type": "transbronchial_biopsy",
        "method": ["Transbronchial Biopsy"],
        "devices": ["Forceps"],
        "anatomy": ["Left Upper Lobe"], # Inferred from "same area"
        "outcomes": []
    },
    {
        "event_id": "evt_99",
        "type": "outcome",
        "method": [],
        "devices": [],
        "anatomy": [],
        "outcomes": ["Complication: None"]
    }
]

# -------------------------------------------------------------------------
# 5. HELPER FUNCTIONS
# -------------------------------------------------------------------------

def clean_text(text):
    if not text:
        return ""
    return text.strip().replace('\r', '')

def hydrate_span(text, span_text, context_prefix=""):
    """
    Finds the start and end offsets of a span in the full text.
    Uses context_prefix to disambiguate identical phrases.
    """
    clean_full_text = text.replace('\r', '') # Normalize
    
    # 1. Try exact match with context
    search_str = context_prefix + span_text
    start_idx = clean_full_text.find(search_str)
    
    if start_idx != -1:
        # Adjustment for the length of context
        actual_start = start_idx + len(context_prefix)
        actual_end = actual_start + len(span_text)
        return actual_start, actual_end, span_text
        
    # 2. Fallback: Find first occurrence of span_text alone
    start_idx = clean_full_text.find(span_text)
    if start_idx != -1:
        return start_idx, start_idx + len(span_text), span_text
        
    return "", "", span_text

# -------------------------------------------------------------------------
# 6. WORKBOOK GENERATION
# -------------------------------------------------------------------------

def generate_workbook():
    # A. Create Workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # B. Define Styles
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # --- Sheet 1: Note_Text ---
    ws_text = wb.create_sheet("Note_Text")
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # Apply style
    for cell in ws_text[1]:
        cell.font = bold_font
        cell.fill = header_fill

    # --- Sheet 2: Note_Index ---
    ws_index = wb.create_sheet("Note_Index")
    headers_index = ["note_id", "source_file", "procedure_date"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers_index)
    
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE] + list(PROCEDURE_FLAGS.values())
    ws_index.append(row_data)
    
    for cell in ws_index[1]:
        cell.font = bold_font
        cell.fill = header_fill

    # --- Sheet 3: Span_Annotations (Human Readable) ---
    ws_human = wb.create_sheet("Span_Annotations")
    headers_human = ["note_id", "event_id", "label", "value", "normalized_value", "start_char", "end_char"]
    ws_human.append(headers_human)
    
    # We leave start/end blank here as per instruction for this specific sheet, 
    # but populating them makes it easier to verify. 
    # The instruction says "Write rows with blank start/end chars" for Sheet 3
    # and "calculated" for Sheet 4.
    
    for span in SPANS:
        # span structure: (span_text, label, normalized_value, context_prefix, event_id)
        ws_human.append([NOTE_ID, span[4], span[1], span[0], span[2], "", ""])

    for cell in ws_human[1]:
        cell.font = bold_font
        cell.fill = header_fill

    # --- Sheet 4: Span_Hydrated (Machine Readable) ---
    ws_hydrated = wb.create_sheet("Span_Hydrated")
    ws_hydrated.append(headers_human) # Same headers
    
    for span in SPANS:
        s_text = span[0]
        context = span[3]
        start, end, _ = hydrate_span(NOTE_TEXT, s_text, context)
        
        ws_hydrated.append([NOTE_ID, span[4], span[1], s_text, span[2], start, end])

    for cell in ws_hydrated[1]:
        cell.font = bold_font
        cell.fill = header_fill

    # --- Sheet 5: Event_Log (Flattened) ---
    ws_event_log = wb.create_sheet("Event_Log")
    headers_log = ["note_id", "event_id", "category", "value"]
    ws_event_log.append(headers_log)
    
    for evt in EVENTS:
        e_id = evt["event_id"]
        # Flatten method
        for m in evt["method"]:
            ws_event_log.append([NOTE_ID, e_id, "method", m])
        # Flatten devices
        for d in evt["devices"]:
            ws_event_log.append([NOTE_ID, e_id, "device", d])
        # Flatten anatomy
        for a in evt["anatomy"]:
            ws_event_log.append([NOTE_ID, e_id, "anatomy", a])
        # Flatten outcomes
        for o in evt["outcomes"]:
            ws_event_log.append([NOTE_ID, e_id, "outcome", o])

    for cell in ws_event_log[1]:
        cell.font = bold_font
        cell.fill = header_fill

    # --- Sheet 6: V3_Procedure_Events ---
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    headers_v3 = ["note_id", "event_id", "event_type", "method_json", "devices_json", "anatomy_json", "outcomes_json"]
    ws_v3.append(headers_v3)
    
    for evt in EVENTS:
        row = [
            NOTE_ID,
            evt["event_id"],
            evt["type"],
            json.dumps(evt["method"]),
            json.dumps(evt["devices"]),
            json.dumps(evt["anatomy"]),
            json.dumps(evt["outcomes"])
        ]
        ws_v3.append(row)

    for cell in ws_v3[1]:
        cell.font = bold_font
        cell.fill = header_fill

    # --- Sheet 7: V3_Registry_JSON ---
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["note_id", "full_json_blob"])
    
    # Construct the full object
    full_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source_file": SOURCE_FILE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS
    }
    
    ws_json.append([NOTE_ID, json.dumps(full_data, indent=2)])
    
    # Resize columns for visibility
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            # Cap width
            if adjusted_width > 50: adjusted_width = 50
            ws.column_dimensions[column].width = adjusted_width

    # Save
    if os.path.exists(OUTPUT_PATH):
        os.remove(OUTPUT_PATH)
    wb.save(OUTPUT_PATH)
    print(f"Successfully generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()