import openpyxl
from openpyxl import Workbook
import re
import json
import os
import datetime

# ==========================================
# 1. Imports & Constants
# ==========================================

NOTE_ID = "note_174"
SOURCE_FILE = "note_174.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_174 SOURCE_FILE: note_174.txt Procedure Name: EBUS bronchoscopy
Indications: Pulmonary nodule requiring diagnosis/staging.
Medications: Propofol infusion via anesthesia assistance  
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention. 
Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway was in good position. The vocal cords appeared normal. The subglottic space was normal.
The trachea was of normal caliber. The carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level without endobronchial lesions visualized.
Anatomy was normal with exception of what appears to be a fused anterior segment of left upper lobe with lingula.
The video bronchoscope was then removed and the UC180F convex probe EBUS bronchoscope was introduced through the mouth, and advanced to the tracheobronchial tree.
A systematic hilar and mediastinal lymph node survey was carried out.
Sampling criteria (5mm short axis diameter) was only met in station station 4R.
Sampling by transbronchial needle aspiration was performed using an Olympus EBUSTBNA 22 gauge needle.
Further details regarding nodal size and number of samples are included in the EBUS procedural sheet in AHLTA.
All samples were sent for routine cytology. Onsite path evaluation did not identify malignancy.
The bronchoscope was then removed and the P190 ultrathin video bronchoscope was inserted into the airway and based on anatomical knowledge advanced into the left upper lobe to the area of known nodule within the anterior segment and an eccentric view of the lesion was identified with the radial EBUS.
Biopsies were then performed with a variety of instruments to include peripheral needle forceps and brush.
After adequate samples were obtained the bronchoscope was removed. ROSE did not identify malignancy on preliminary samples.
The bronchoscope was then removed and the P190 re-inserted into the airways.
DECAMP research samples were then performed with brushing within the right upper lobe, right middle lobe and transbronchial biopsies in the RUL, RML and LUL.
We ten observed for evidence of active bleeding and none was identified. The bronchoscope was removed and the procedure completed.
Complications: 	
-None 
Estimated Blood Loss:  10 cc.
Recommendations:
- Transfer to PACU
- Await biopsy results 
- Discharge home once criteria met."""

# ==========================================
# 2. Configuration (Procedure Flags)
# ==========================================

PROCEDURE_FLAGS = {
    # Bronchoscopy Flags
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 1,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 1,
    "navigational_bronchoscopy": 0, # "anatomical knowledge" used, not electromagnetic nav
    "transbronchial_biopsy": 1,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    
    # Pleural Flags
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# ==========================================
# 3. Data Definition (Spans)
# ==========================================

# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    # Event 1: Initial Inspection (Q190)
    ("EBUS bronchoscopy", "PROC_METHOD", "EBUS bronchoscopy", "Procedure Name: ", "evt_01"),
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190", "tree, the ", "evt_01"),
    ("laryngeal mask airway", "DEV_INSTRUMENT", "Laryngeal Mask Airway", "mouth, via ", "evt_01"),
    ("tracheobronchial tree", "ANAT_AIRWAY", "Tracheobronchial Tree", "advanced to the ", "evt_01"),
    ("fused anterior segment of left upper lobe", "ANAT_LUNG_LOC", "LUL Anterior Segment", "appears to be a ", "evt_01"),
    
    # Event 2: EBUS TBNA (Linear)
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "UC180F", "removed and the ", "evt_02"),
    ("hilar and mediastinal lymph node", "ANAT_LN_STATION", "Hilar/Mediastinal", "A systematic ", "evt_02"),
    ("station 4R", "ANAT_LN_STATION", "4R", "met in station ", "evt_02"),
    ("transbronchial needle aspiration", "PROC_METHOD", "EBUS-TBNA", "Sampling by ", "evt_02"),
    ("Olympus EBUSTBNA 22 gauge needle", "DEV_NEEDLE", "Olympus EBUS-TBNA 22G", "using an ", "evt_02"),
    
    # Event 3: Peripheral Nodule (Thin Scope + REBUS)
    ("P190 ultrathin video bronchoscope", "DEV_INSTRUMENT", "P190", "removed and the ", "evt_03"),
    ("left upper lobe", "ANAT_LUNG_LOC", "LUL", "advanced into the ", "evt_03"),
    ("anterior segment", "ANAT_LUNG_LOC", "Anterior Segment", "nodule within the ", "evt_03"),
    ("radial EBUS", "DEV_INSTRUMENT", "Radial EBUS", "identified with the ", "evt_03"),
    ("Biopsies", "PROC_ACTION", "Biopsy", "radial EBUS.\n", "evt_03"),
    ("peripheral needle", "DEV_NEEDLE", "Peripheral Needle", "to include ", "evt_03"),
    ("forceps", "DEV_INSTRUMENT", "Forceps", "peripheral needle ", "evt_03"),
    ("brush", "DEV_INSTRUMENT", "Brush", "forceps and ", "evt_03"),
    
    # Event 4: Research Samples (DECAMP)
    ("DECAMP research samples", "PROC_ACTION", "Research Samples", "", "evt_04"),
    ("brushing", "PROC_METHOD", "Brushing", "performed with ", "evt_04"),
    ("right upper lobe", "ANAT_LUNG_LOC", "RUL", "within the ", "evt_04"),
    ("right middle lobe", "ANAT_LUNG_LOC", "RML", "right upper lobe, ", "evt_04"),
    ("transbronchial biopsies", "PROC_METHOD", "Transbronchial Biopsy", "lobe and ", "evt_04"),
    ("RUL", "ANAT_LUNG_LOC", "RUL", "in the ", "evt_04"),
    ("RML", "ANAT_LUNG_LOC", "RML", "RUL, ", "evt_04"),
    ("LUL", "ANAT_LUNG_LOC", "LUL", "RML and ", "evt_04"),
    
    # Outcomes
    ("None", "OUTCOME_COMPLICATION", "None", "Complications: \t\n-", "evt_outcome"),
    ("10 cc", "MEAS_VOL", "10 cc", "Estimated Blood Loss:  ", "evt_outcome"),
    ("ROSE did not identify malignancy", "OBS_ROSE", "Benign", "removed. ", "evt_outcome")
]

# ==========================================
# 4. Event Definitions
# ==========================================

EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Bronchoscopy",
        "action": "Inspection",
        "devices": ["Q190 video bronchoscope", "laryngeal mask airway"],
        "anatomy": ["Tracheobronchial Tree", "LUL Anterior Segment"],
        "notes": "General inspection, fused anterior segment identified."
    },
    {
        "event_id": "evt_02",
        "procedure_type": "EBUS-TBNA",
        "action": "Needle Aspiration",
        "devices": ["UC180F", "Olympus EBUSTBNA 22 gauge needle"],
        "anatomy": ["Station 4R"],
        "notes": "Systematic survey, sampling of 4R."
    },
    {
        "event_id": "evt_03",
        "procedure_type": "Radial EBUS Bronchoscopy",
        "action": "Biopsy",
        "devices": ["P190", "Radial EBUS", "Peripheral needle", "Forceps", "Brush"],
        "anatomy": ["LUL Anterior Segment Nodule"],
        "notes": "Targeted nodule biopsy."
    },
    {
        "event_id": "evt_04",
        "procedure_type": "Research Protocol",
        "action": "Brushing/TBBx",
        "devices": ["Brush", "Forceps"],
        "anatomy": ["RUL", "RML", "LUL"],
        "notes": "DECAMP research samples."
    }
]

# ==========================================
# 5. Helper Functions
# ==========================================

def clean_text(text):
    if not text:
        return ""
    return text.strip().replace('\r', '')

def hydrate_span(full_text, span_text, context_prefix):
    """
    Finds the start and end indices of span_text in full_text,
    using context_prefix to disambiguate.
    """
    clean_full = clean_text(full_text)
    clean_span = clean_text(span_text)
    clean_prefix = clean_text(context_prefix)
    
    if not clean_span:
        return "", ""

    # Attempt to find prefix + span
    search_str = clean_prefix + clean_span
    start_index = clean_full.find(search_str)
    
    if start_index == -1:
        # Fallback: try just the span (warning: might pick first occurrence)
        start_index = clean_full.find(clean_span)
    else:
        # Adjust start_index to point to the start of the span, not the prefix
        start_index += len(clean_prefix)
        
    if start_index == -1:
        return "", ""
        
    end_index = start_index + len(clean_span)
    return start_index, end_index

# ==========================================
# 6. Workbook Generation Function
# ==========================================

def generate_workbook():
    # 1. Load or Create Template
    if not os.path.exists(TEMPLATE_PATH):
        # Create a basic dummy workbook if template is missing (for standalone safety)
        wb = Workbook()
        for name in ["Note_Text", "Note_Index", "Span_Annotations", "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"]:
            wb.create_sheet(name)
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # 2. Sheet: Note_Text
    ws_text = wb["Note_Text"]
    # Clear existing
    for row in ws_text.iter_rows(min_row=2, max_col=3):
        for cell in row:
            cell.value = None
    # Write Data
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 3. Sheet: Note_Index
    ws_index = wb["Note_Index"]
    # Assuming Row 2 is the data row for this single note
    # Header mapping assumed based on standard registry format
    # Columns 1-4: ID, Reviewer, Status, Comments (Skip or Fill Basic)
    ws_index["A2"] = NOTE_ID
    ws_index["B2"] = "Auto-Generated"
    
    # Map Procedure Flags to Columns (Hardcoded column indices for safety or dynamic mapping if headers exist)
    # Since headers vary, we will try to match headers if they exist, otherwise append kv pairs
    headers = [cell.value for cell in ws_index[1]]
    
    # Write flags
    for key, value in PROCEDURE_FLAGS.items():
        if key in headers:
            col_idx = headers.index(key) + 1
            ws_index.cell(row=2, column=col_idx, value=value)

    # 4. Sheet: Span_Annotations
    ws_anno = wb["Span_Annotations"]
    # Headers: Note_ID, Span_Text, Label, Normalized_Value, Context_Prefix, Start_Char, End_Char, Event_ID
    # Clear old data
    ws_anno.delete_rows(2, ws_anno.max_row)
    
    for span in SPANS:
        ws_anno.append([
            NOTE_ID,
            span[0], # Text
            span[1], # Label
            span[2], # Norm
            span[3], # Prefix
            "",      # Start (calc next)
            "",      # End (calc next)
            span[4]  # Event ID
        ])

    # 5. Sheet: Span_Hydrated
    ws_hydra = wb["Span_Hydrated"]
    ws_hydra.delete_rows(2, ws_hydra.max_row)
    
    for span in SPANS:
        s_start, s_end = hydrate_span(NOTE_TEXT, span[0], span[3])
        ws_hydra.append([
            NOTE_ID,
            span[0],
            span[1],
            span[2],
            span[3],
            s_start,
            s_end,
            span[4]
        ])

    # 6. Sheet: Event_Log (Flattened)
    ws_log = wb["Event_Log"]
    ws_log.delete_rows(2, ws_log.max_row)
    
    for evt in EVENTS:
        ws_log.append([
            NOTE_ID,
            evt["event_id"],
            evt["procedure_type"],
            evt["action"],
            ", ".join(evt["devices"]),
            ", ".join(evt["anatomy"]),
            evt.get("notes", "")
        ])

    # 7. Sheet: V3_Procedure_Events (Detailed)
    ws_v3 = wb["V3_Procedure_Events"]
    ws_v3.delete_rows(2, ws_v3.max_row)
    
    for evt in EVENTS:
        # Construct V3 Row: Note_ID, Event_ID, Proc_Type, Method, Devices(JSON), Anatomy(JSON), Outcomes(JSON)
        row_data = [
            NOTE_ID,
            evt["event_id"],
            evt["procedure_type"],
            evt["action"],
            json.dumps(evt["devices"]),
            json.dumps(evt["anatomy"]),
            json.dumps({"notes": evt.get("notes", "")})
        ]
        ws_v3.append(row_data)

    # 8. Sheet: V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    ws_json.delete_rows(2, ws_json.max_row)
    
    registry_object = {
        "metadata": {
            "note_id": NOTE_ID,
            "source": SOURCE_FILE,
            "extraction_date": datetime.datetime.now().isoformat()
        },
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS,
        "spans_raw": [
            {"text": s[0], "label": s[1], "norm": s[2], "event": s[4]} for s in SPANS
        ]
    }
    
    ws_json.cell(row=2, column=1, value=NOTE_ID)
    ws_json.cell(row=2, column=2, value=json.dumps(registry_object, indent=2))

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated Phase 0 Workbook: {OUTPUT_PATH}")

# ==========================================
# 7. Execution Block
# ==========================================

if __name__ == "__main__":
    generate_workbook()