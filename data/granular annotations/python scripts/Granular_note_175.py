import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import re
import json
import os
import datetime

# -------------------------------------------------------------------------
# 1. Imports & Constants
# -------------------------------------------------------------------------

NOTE_ID = "note_175"
SOURCE_FILE = "note_175.txt"
PROCEDURE_DATE = ""  # Not specified in note content
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_175 SOURCE_FILE: note_175.txt Procedure Name: 
1.	radial EBUS guided bronchoscopy with TNBA and endobronchial biopsies
2.	Bronchoscopic intubation 
Indications: Lung cancer with suspicion for recurrence 
Medications: Propofol infusion via anesthesia assistance  
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the preprocedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention. 
Following intravenous medications as per the anesthesia record and topical anesthesia to the upper airway and tracheobronchial tree the Q190 video bronchoscope was introduced through the mouth.
The vocal cords appeared normal. The subglottic space was normal. The trachea is of normal caliber. The carina was sharp.
Somewhat thick secretions were seen throughout the airways. The left sided airway anatomy was normal without evidence of endobronchial disease.
On the right the patient had an anatomic variant with an accessory airway just proximal to the superior segment of the right lower lobe.
Within that segment fibrotic endoluminal obstruction was seen but all other airways did not have visual endobronchial disease to at least the first sub-segmental level.
Due to difficulty with maintaining position of LMA (edentulous) the decision was made to convert to endotracheal intubation.
This was performed over the bronchoscope without difficulty and the ETT was secured approximately 2 cm above the main carina.
Once the ETT was connected to ventilator, The flexible bronchoscope was re-inserted through the tube and advanced to the orifice of the accessory left lower lobe bronchus and the radial ultrasound was advanced through the scope into the sub-segment and a concentric view of the lesion was visualized within the segment.
Multiple forceps biopsies were then performed. The lesion was extremely firm and samples were relatively scant.
We then utilized the super-dimension 21G peripheral needle and the 19G Olympus peripheral needle to obtain a total of 6 endobronchial needle biopsies of the lesion which were placed in formalin.
After adequate samples were obtained repeat inspection was performed and no evidence of bleeding or other complications were seen.
The bronchoscope was removed and the procedure completed.  
Complications: No immediate complications
Estimated Blood Loss: Less than 5 cc.
Post Procedure Diagnosis:
- Flexible bronchoscopy with successful biopsy of left lower lobe accessory segment fibrotic endobronchial lesion.
- Await final pathology"""

# -------------------------------------------------------------------------
# 2. Configuration (Procedure Flags)
# -------------------------------------------------------------------------

PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 1,
    "tbna_conventional": 1,  # "TNBA" and "endobronchial needle biopsies" with 21G/19G
    "linear_ebus": 0,
    "radial_ebus": 1,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# -------------------------------------------------------------------------
# 3. Data Definition (Spans)
# -------------------------------------------------------------------------
# Format: (span_text, label, normalized_value, context_prefix, event_id)

SPANS = [
    # Header
    ("radial EBUS guided bronchoscopy", "PROC_METHOD", "Radial EBUS", "Procedure Name: \n1.\t", "evt_01"),
    ("TNBA", "PROC_METHOD", "TBNA", "guided bronchoscopy with ", "evt_01"),
    ("endobronchial biopsies", "PROC_METHOD", "Endobronchial Biopsy", "and ", "evt_01"),
    ("Bronchoscopic intubation", "PROC_METHOD", "Bronchoscopic Intubation", "2.\t", "evt_02"),
    
    # Initial Inspection
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190 Bronchoscope", "tree the ", "evt_03"),
    ("mouth", "ANAT_AIRWAY", "Mouth", "introduced through the ", "evt_03"),
    ("accessory airway", "ANAT_AIRWAY", "Accessory Airway", "variant with an ", "evt_03"),
    ("superior segment of the right lower lobe", "ANAT_LUNG_LOC", "RLL Superior Segment", "proximal to the ", "evt_03"),
    ("fibrotic endoluminal obstruction", "OBS_LESION", "Endoluminal Obstruction", "segment ", "evt_03"),
    
    # Intubation Conversion
    ("LMA", "DEV_INSTRUMENT", "LMA", "position of ", "evt_04"),
    ("endotracheal intubation", "PROC_METHOD", "Endotracheal Intubation", "convert to ", "evt_04"),
    ("ETT", "DEV_CATHETER", "ETT", "difficulty and the ", "evt_04"),
    
    # REBUS / Biopsy Phase (Note: Discrepancy in laterality in note, Source 7 says Right, Source 11 says Left. Capturing as written.)
    ("flexible bronchoscope", "DEV_INSTRUMENT", "Flexible Bronchoscope", "ventilator, The ", "evt_05"),
    ("accessory left lower lobe bronchus", "ANAT_AIRWAY", "LLL Accessory Bronchus", "orifice of the ", "evt_05"),
    ("radial ultrasound", "DEV_INSTRUMENT", "Radial EBUS Probe", "and the ", "evt_05"),
    ("concentric view", "OBS_LESION", "Concentric View", "segment and a ", "evt_05"),
    
    # Forceps
    ("forceps biopsies", "PROC_METHOD", "Forceps Biopsy", "Multiple ", "evt_06"),
    
    # Needle
    ("super-dimension 21G peripheral needle", "DEV_NEEDLE", "SuperDimension 21G", "utilized the ", "evt_07"),
    ("19G Olympus peripheral needle", "DEV_NEEDLE", "Olympus 19G", "and the ", "evt_07"),
    ("endobronchial needle biopsies", "PROC_METHOD", "TBNA", "total of 6 ", "evt_07"),
    ("6", "MEAS_COUNT", "6", "total of ", "evt_07"),
    
    # Diagnosis
    ("Flexible bronchoscopy", "PROC_METHOD", "Flexible Bronchoscopy", "Diagnosis:\n- ", "evt_08"),
    ("biopsy", "PROC_METHOD", "Biopsy", "successful ", "evt_08"),
    ("left lower lobe accessory segment", "ANAT_LUNG_LOC", "LLL Accessory Segment", "biopsy of ", "evt_08"),
    ("fibrotic endobronchial lesion", "OBS_LESION", "Fibrotic Lesion", "segment ", "evt_08")
]

# -------------------------------------------------------------------------
# 4. Event Definitions
# -------------------------------------------------------------------------

EVENTS = [
    {
        "event_id": "evt_01",
        "name": "Planned Procedures",
        "concept_id": "CPT-PLAN",
        "comments": "Header info"
    },
    {
        "event_id": "evt_02",
        "name": "Planned Intubation",
        "concept_id": "CPT-INTUB",
        "comments": "Header info"
    },
    {
        "event_id": "evt_03",
        "name": "Initial Inspection",
        "concept_id": "PROC-INSPECT",
        "comments": "Inspection revealed RLL accessory airway obstruction"
    },
    {
        "event_id": "evt_04",
        "name": "Intubation",
        "concept_id": "PROC-INTUB",
        "comments": "Conversion from LMA to ETT"
    },
    {
        "event_id": "evt_05",
        "name": "Radial EBUS Visualization",
        "concept_id": "PROC-REBUS",
        "comments": "Visualization of LLL accessory bronchus lesion (Note discrepancy: RLL vs LLL in text)"
    },
    {
        "event_id": "evt_06",
        "name": "Forceps Biopsy",
        "concept_id": "PROC-BIOPSY",
        "comments": "Forceps biopsies of lesion"
    },
    {
        "event_id": "evt_07",
        "name": "Needle Biopsy",
        "concept_id": "PROC-TBNA",
        "comments": "Using 21G and 19G needles"
    },
    {
        "event_id": "evt_08",
        "name": "Post Procedure Diagnosis",
        "concept_id": "DX-FINAL",
        "comments": "Final diagnosis summary"
    }
]

# -------------------------------------------------------------------------
# 5. Helper Functions
# -------------------------------------------------------------------------

def clean_text(text):
    return text.replace('\r', '').strip()

def hydrate_span(full_text, span_text, context_prefix):
    """
    Finds the start and end indices of span_text in full_text,
    ensuring it follows context_prefix to disambiguate.
    """
    clean_full = clean_text(full_text)
    clean_span = clean_text(span_text)
    clean_context = clean_text(context_prefix)

    if not clean_span:
        return 0, 0

    # Search for context first
    context_start = clean_full.find(clean_context)
    if context_start == -1:
        # Fallback: try finding just the span
        start_index = clean_full.find(clean_span)
        if start_index == -1:
            return None, None
    else:
        # Look for span *after* context
        search_start = context_start + len(clean_context)
        start_index = clean_full.find(clean_span, search_start)
        # Verify it's reasonably close (e.g. within 50 chars) to ensure correct match
        if start_index == -1 or (start_index - search_start > 200):
            # Fallback if too far or not found
            start_index = clean_full.find(clean_span)

    if start_index == -1:
        return None, None
    
    end_index = start_index + len(clean_span)
    return start_index, end_index

# -------------------------------------------------------------------------
# 6. Workbook Generation
# -------------------------------------------------------------------------

def generate_workbook():
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # ---------------------------
    # Sheet 1: Note_Text
    # ---------------------------
    ws_text = wb.create_sheet("Note_Text")
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # ---------------------------
    # Sheet 2: Note_Index
    # ---------------------------
    ws_index = wb.create_sheet("Note_Index")
    headers_index = ["NOTE_ID", "SOURCE_FILE", "PROCEDURE_DATE"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers_index)
    
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE]
    for key in PROCEDURE_FLAGS:
        row_data.append(PROCEDURE_FLAGS[key])
    ws_index.append(row_data)

    # ---------------------------
    # Sheet 3: Span_Annotations
    # ---------------------------
    ws_anno = wb.create_sheet("Span_Annotations")
    headers_anno = ["NOTE_ID", "start_char", "end_char", "text", "label", "normalized_value", "context_prefix", "event_id"]
    ws_anno.append(headers_anno)
    
    for span in SPANS:
        # (text, label, norm, context, evt)
        ws_anno.append([NOTE_ID, "", "", span[0], span[1], span[2], span[3], span[4]])

    # ---------------------------
    # Sheet 4: Span_Hydrated
    # ---------------------------
    ws_hydrated = wb.create_sheet("Span_Hydrated")
    ws_hydrated.append(headers_anno)
    
    for span in SPANS:
        s_txt, label, norm, ctx, evt = span
        start, end = hydrate_span(NOTE_TEXT, s_txt, ctx)
        if start is None:
            start, end = 0, 0
        ws_hydrated.append([NOTE_ID, start, end, s_txt, label, norm, ctx, evt])

    # ---------------------------
    # Sheet 5: Event_Log
    # ---------------------------
    ws_event = wb.create_sheet("Event_Log")
    headers_event = ["NOTE_ID", "EVENT_ID", "EVENT_NAME", "CONCEPT_ID", "COMMENTS"]
    ws_event.append(headers_event)
    
    for evt in EVENTS:
        ws_event.append([NOTE_ID, evt["event_id"], evt["name"], evt["concept_id"], evt["comments"]])

    # ---------------------------
    # Sheet 6: V3_Procedure_Events
    # ---------------------------
    # Flattens spans into a registry-like view
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    headers_v3 = [
        "NOTE_ID", "EVENT_ID", "Procedure_Method", "Anatomy", "Devices", 
        "Observations", "Measurements", "Outcomes"
    ]
    ws_v3.append(headers_v3)

    # Group spans by event_id
    event_map = {evt["event_id"]: {"methods": [], "anat": [], "dev": [], "obs": [], "meas": [], "out": []} for evt in EVENTS}
    
    for span in SPANS:
        s_txt, label, norm, ctx, evt_id = span
        if evt_id in event_map:
            if label == "PROC_METHOD":
                event_map[evt_id]["methods"].append(norm)
            elif label.startswith("ANAT"):
                event_map[evt_id]["anat"].append(norm)
            elif label.startswith("DEV"):
                event_map[evt_id]["dev"].append(norm)
            elif label.startswith("OBS"):
                event_map[evt_id]["obs"].append(norm)
            elif label.startswith("MEAS"):
                event_map[evt_id]["meas"].append(f"{label}: {norm}")
            elif label.startswith("OUTCOME"):
                event_map[evt_id]["out"].append(norm)

    for evt in EVENTS:
        eid = evt["event_id"]
        data = event_map.get(eid, {})
        ws_v3.append([
            NOTE_ID,
            eid,
            ", ".join(data.get("methods", [])),
            ", ".join(data.get("anat", [])),
            ", ".join(data.get("dev", [])),
            ", ".join(data.get("obs", [])),
            ", ".join(data.get("meas", [])),
            ", ".join(data.get("out", []))
        ])

    # ---------------------------
    # Sheet 7: V3_Registry_JSON
    # ---------------------------
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["NOTE_ID", "JSON_PAYLOAD"])
    
    # Construct JSON payload
    payload = {
        "note_id": NOTE_ID,
        "source_file": SOURCE_FILE,
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS,
        "annotations": []
    }
    
    # Add spans to JSON
    for span in SPANS:
        s_txt, label, norm, ctx, evt_id = span
        start, end = hydrate_span(NOTE_TEXT, s_txt, ctx)
        payload["annotations"].append({
            "text": s_txt,
            "label": label,
            "normalized": norm,
            "start": start,
            "end": end,
            "event_id": evt_id
        })
        
    ws_json.append([NOTE_ID, json.dumps(payload, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

# -------------------------------------------------------------------------
# 7. Execution Block
# -------------------------------------------------------------------------
if __name__ == "__main__":
    generate_workbook()