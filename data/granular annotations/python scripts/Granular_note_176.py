import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# =============================================================================
# 1. CONSTANTS & INPUT DATA
# =============================================================================

NOTE_ID = "note_176"
SOURCE_FILE = "note_176.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_176 SOURCE_FILE: note_176.txt Indications: Mediastinal adenopathy
Procedure: EBUS bronchoscopy – single station CPT 31652
Medications: General Anesthesia
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention.
Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway was in good position. The vocal cords appeared normal. The subglottic space was normal.
The trachea was of normal caliber. The carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level. Bronchial mucosa and anatomy were normal;
there are no endobronchial lesions, and no secretions. The video bronchoscope was then removed and the UC180F convex probe EBUS bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
A large, station 4R, lymph node was identified and sampling by transbronchial needle aspiration was performed with the Olympus 19G Visioshot EBUS-TBNA needles with a total of 7 passes performed.
Initial rapid onsite evaluation showed scattered lymphocytes but no evidence of malignancy or granuloma.
We then placed the Olympus mini-forceps through the working channel of the EBUS scope and adnaved the forceps through a tract in the airway created by the EBUS needle.
We were able to visualize the forceps using the EBUS ultrasound within the lymph node.
We then closed and retracted the forceps. This was performed twice and touchpreps were made with the collected specimins which showed abundant non-caseating granulomas on rapid onsite pathological evaluation.
Following completion of EBUS bronchoscopy, the Q190 video bronchoscope was then re-inserted and after suctioning blood and secretions there was no evidence of active bleeding and the bronchoscope was subsequently removed.
Complications: No immediate complications
Estimated Blood Loss: 5cc
Post Procedure Diagnosis:
- Technically successful flexible bronchoscopy with endobronchial ultrasound-guided needle and mini-forceps.
- The patient has remained stable and has been transferred in good condition to the post-procedural monitoring unit.
- Will await final pathology results"""

# =============================================================================
# 2. CONFIGURATION (PROCEDURE FLAGS)
# =============================================================================

PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 1, # Mini-forceps used for tissue biopsy via needle tract
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# =============================================================================
# 3. DATA DEFINITION (SPANS)
# =============================================================================
# Format: (span_text, label, normalized_value, context_prefix, event_id)

SPANS = [
    # Event 01: Initial Inspection
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190 video bronchoscope", "tracheobronchial tree, the", "evt_01"),
    ("laryngeal mask airway", "ANAT_AIRWAY", "LMA", "introduced through the mouth, via", "evt_01"),
    ("trachea", "ANAT_AIRWAY", "Trachea", "The", "evt_01"),
    ("carina", "ANAT_AIRWAY", "Carina", "The", "evt_01"),
    
    # Event 02: EBUS TBNA Station 4R
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "UC180F EBUS Scope", "removed and the", "evt_02"),
    ("station 4R", "ANAT_LN_STATION", "4R", "A large,", "evt_02"),
    ("transbronchial needle aspiration", "PROC_METHOD", "TBNA", "sampling by", "evt_02"),
    ("Olympus 19G Visioshot EBUS-TBNA needles", "DEV_NEEDLE", "Olympus 19G Visioshot", "performed with the", "evt_02"),
    ("7 passes", "MEAS_COUNT", "7", "total of", "evt_02"),
    ("scattered lymphocytes", "OBS_ROSE", "Lymphocytes", "evaluation showed", "evt_02"),
    ("no evidence of malignancy", "OBS_ROSE", "Negative for malignancy", "lymphocytes but", "evt_02"),
    
    # Event 03: EBUS Mini-Forceps Biopsy
    ("Olympus mini-forceps", "DEV_INSTRUMENT", "Olympus Mini-Forceps", "We then placed the", "evt_03"),
    ("tract in the airway", "ANAT_AIRWAY", "Airway Tract", "forceps through a", "evt_03"),
    ("EBUS needle", "DEV_NEEDLE", "EBUS Needle", "created by the", "evt_03"),
    ("EBUS ultrasound", "PROC_METHOD", "EBUS", "forceps using the", "evt_03"),
    ("lymph node", "ANAT_LN_STATION", "Lymph Node", "within the", "evt_03"),
    ("twice", "MEAS_COUNT", "2", "This was performed", "evt_03"),
    ("abundant non-caseating granulomas", "OBS_ROSE", "Non-caseating granulomas", "showed", "evt_03"),

    # Outcomes / Post-Procedure
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications:", "evt_04"),
    ("successful flexible bronchoscopy", "OUTCOME_SYMPTOMS", "Success", "Technically", "evt_04")
]

# =============================================================================
# 4. EVENT DEFINITIONS
# =============================================================================

EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Bronchoscopy",
        "action": "Inspection",
        "anatomy": ["Trachea", "Carina", "Tracheobronchial Tree"],
        "devices": ["Q190 video bronchoscope", "Laryngeal Mask Airway"],
        "outcomes": ["Normal anatomy"]
    },
    {
        "event_id": "evt_02",
        "procedure_type": "Bronchoscopy",
        "action": "EBUS-TBNA",
        "anatomy": ["Station 4R"],
        "devices": ["UC180F convex probe EBUS bronchoscope", "Olympus 19G Visioshot EBUS-TBNA needles"],
        "samples": ["7 passes"],
        "observations": ["Scattered lymphocytes", "No malignancy"]
    },
    {
        "event_id": "evt_03",
        "procedure_type": "Bronchoscopy",
        "action": "Transbronchial Biopsy (Mini-Forceps)",
        "anatomy": ["Station 4R"],
        "devices": ["Olympus mini-forceps", "EBUS Needle (tract creation)"],
        "samples": ["2 passes"],
        "observations": ["Abundant non-caseating granulomas"]
    },
    {
        "event_id": "evt_04",
        "procedure_type": "Post-Procedure",
        "action": "Conclusion",
        "outcomes": ["No immediate complications", "Technically successful"]
    }
]

# =============================================================================
# 5. HELPER FUNCTIONS
# =============================================================================

def clean_text(text):
    if not text:
        return ""
    return re.sub(r'[\r\n]+', ' ', text).strip()

def hydrate_span(full_text, span_text, context_prefix):
    """
    Finds the start and end offsets of span_text in full_text.
    Uses context_prefix to disambiguate if possible.
    Returns (start, end, exact_text_found).
    """
    clean_full = clean_text(full_text)
    clean_span = clean_text(span_text)
    clean_context = clean_text(context_prefix)
    
    if not clean_span:
        return "", "", ""

    # Attempt to find with context first
    search_phrase = f"{clean_context} {clean_span}" if clean_context else clean_span
    # Simple regex escape to handle special chars in medical text
    pattern = re.escape(search_phrase)
    
    match = re.search(pattern, clean_full, re.IGNORECASE)
    
    if match:
        # If context was used, adjust start to point to the span part
        if clean_context:
            # Re-locate span within the matched string to get precise offsets
            # This is a simplification; for strict extraction, we look for the span 
            # at the end of the matched group.
            actual_start = match.start() + len(match.group()) - len(clean_span)
            actual_end = match.end()
            return actual_start, actual_end, clean_full[actual_start:actual_end]
        else:
            return match.start(), match.end(), match.group()
            
    # Fallback: search just for the span (first occurrence)
    match_fallback = re.search(re.escape(clean_span), clean_full, re.IGNORECASE)
    if match_fallback:
        return match_fallback.start(), match_fallback.end(), match_fallback.group()
        
    return "", "", ""

# =============================================================================
# 6. WORKBOOK GENERATION
# =============================================================================

def generate_workbook():
    # 1. Initialize Workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # ---------------------------------------------------------
    # Sheet 1: Note_Text
    # ---------------------------------------------------------
    ws_text = wb.create_sheet("Note_Text")
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # ---------------------------------------------------------
    # Sheet 2: Note_Index (Metadata + Flags)
    # ---------------------------------------------------------
    ws_index = wb.create_sheet("Note_Index")
    
    # Headers: Metadata columns + Procedure Flag keys
    flag_keys = list(PROCEDURE_FLAGS.keys())
    headers = ["NOTE_ID", "SOURCE_FILE", "PROCEDURE_DATE"] + flag_keys
    ws_index.append(headers)
    
    # Row Values
    flag_values = [PROCEDURE_FLAGS[k] for k in flag_keys]
    row = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE] + flag_values
    ws_index.append(row)
    
    # ---------------------------------------------------------
    # Sheet 3: Span_Annotations (Raw input for reviewing)
    # ---------------------------------------------------------
    ws_spans = wb.create_sheet("Span_Annotations")
    ws_spans.append(["NOTE_ID", "EVENT_ID", "LABEL", "VALUE_NORMALIZED", "SPAN_TEXT", "CONTEXT_PREFIX", "START_CHAR", "END_CHAR"])
    
    for span in SPANS:
        s_text, s_label, s_norm, s_context, s_evt = span
        # Write row with empty start/end for now (filled in hydrated sheet)
        ws_spans.append([NOTE_ID, s_evt, s_label, s_norm, s_text, s_context, "", ""])

    # ---------------------------------------------------------
    # Sheet 4: Span_Hydrated (Calculated Offsets)
    # ---------------------------------------------------------
    ws_hydrated = wb.create_sheet("Span_Hydrated")
    ws_hydrated.append(["NOTE_ID", "EVENT_ID", "LABEL", "VALUE_NORMALIZED", "SPAN_TEXT", "CONTEXT_PREFIX", "START_CHAR", "END_CHAR"])
    
    for span in SPANS:
        s_text, s_label, s_norm, s_context, s_evt = span
        start, end, found_text = hydrate_span(NOTE_TEXT, s_text, s_context)
        ws_hydrated.append([NOTE_ID, s_evt, s_label, s_norm, found_text if found_text else s_text, s_context, start, end])

    # ---------------------------------------------------------
    # Sheet 5: Event_Log (Flattened Event View)
    # ---------------------------------------------------------
    ws_events = wb.create_sheet("Event_Log")
    ws_events.append(["NOTE_ID", "EVENT_ID", "PROCEDURE_TYPE", "ACTION", "ANATOMY", "DEVICES", "SAMPLES", "OBSERVATIONS", "OUTCOMES"])
    
    for evt in EVENTS:
        ws_events.append([
            NOTE_ID,
            evt.get("event_id", ""),
            evt.get("procedure_type", ""),
            evt.get("action", ""),
            ", ".join(evt.get("anatomy", [])),
            ", ".join(evt.get("devices", [])),
            ", ".join(evt.get("samples", [])),
            ", ".join(evt.get("observations", [])),
            ", ".join(evt.get("outcomes", []))
        ])

    # ---------------------------------------------------------
    # Sheet 6: V3_Procedure_Events (JSON-ready structure)
    # ---------------------------------------------------------
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    ws_v3.append(["NOTE_ID", "EVENT_ID", "EVENT_JSON"])
    
    for evt in EVENTS:
        # Create a compact JSON string for the event
        evt_json = json.dumps(evt)
        ws_v3.append([NOTE_ID, evt.get("event_id", ""), evt_json])

    # ---------------------------------------------------------
    # Sheet 7: V3_Registry_JSON (Full Document JSON)
    # ---------------------------------------------------------
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["NOTE_ID", "FULL_JSON"])
    
    full_doc_structure = {
        "metadata": {
            "note_id": NOTE_ID,
            "source_file": SOURCE_FILE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS,
        "spans": [
            {
                "text": s[0],
                "label": s[1],
                "normalized": s[2],
                "event_id": s[4]
            }
            for s in SPANS
        ]
    }
    
    ws_json.append([NOTE_ID, json.dumps(full_doc_structure, indent=2)])

    # 2. Save Workbook
    if not os.path.exists(OUTPUT_PATH):
        wb.save(OUTPUT_PATH)
        print(f"Workbook generated: {OUTPUT_PATH}")
    else:
        print(f"Workbook already exists: {OUTPUT_PATH}. Skipping save to prevent overwrite.")

# =============================================================================
# 7. EXECUTION
# =============================================================================
if __name__ == "__main__":
    generate_workbook()