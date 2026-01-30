import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# -------------------------------------------------------------------------
# 1. CONSTANTS & INPUT DATA
# -------------------------------------------------------------------------

NOTE_ID = "note_177"
SOURCE_FILE = "note_177.txt"
PROCEDURE_DATE = ""
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_177 SOURCE_FILE: note_177.txt Indications: diagnosis and staging of suspected lung cancer
Medications: General Anesthesia,
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention.
Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway was in good position. The vocal cords appeared normal. The subglottic space was normal.
The trachea was of normal caliber. The carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level. Bronchial mucosa and anatomy were normal;
there are no endobronchial lesions, and no secretions. The video bronchoscope was then removed and the UC180F convex probe EBUS bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
A systematic hilar and mediastinal lymph node survey was carried out.
Sampling criteria (5mm short axis diameter) were met in station 11L, 4L, 4R, 7, 11Rs and 11Ri lymph nodes.
Sampling by transbronchial needle aspiration was performed with the Olympus EBUSTBNA 22 gauge needle beginning with the 11L Lymph node4L74R11Rs11Ri.
ROSE did not identify malignancy in any of the obtained samples. All samples were sent for routine cytology.
Following completion of EBUS bronchoscopy, the Q190 video bronchoscope was then re-inserted and after suctioning blood and secretions there was no evidence of active bleeding and the bronchoscope was subsequently removed.
Complications: No immediate complications
Estimated Blood Loss: 5 cc.

Post Procedure Diagnosis:
- Technically successful flexible bronchoscopy with endobronchial ultrasound-guided biopsies.
- The patient has remained stable and has been transferred in good condition to the post-surgical monitoring unit.
- Will await final pathology results"""

# -------------------------------------------------------------------------
# 2. CONFIGURATION (PROCEDURE FLAGS)
# -------------------------------------------------------------------------

PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# -------------------------------------------------------------------------
# 3. DATA DEFINITION (SPANS)
# -------------------------------------------------------------------------
# Format: (span_text, label, normalized_value, context_prefix, event_id)

SPANS = [
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190 Bronchoscope", "tree, the", "evt_01"),
    ("examined to at least the first subsegmental level", "PROC_ACTION", "Inspection", "tree was", "evt_01"),
    ("no endobronchial lesions", "OBS_LESION", "None", "there are", "evt_01"),
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "UC180F EBUS Bronchoscope", "removed and the", "evt_02"),
    ("transbronchial needle aspiration", "PROC_METHOD", "EBUS-TBNA", "Sampling by", "evt_02"),
    ("Olympus EBUSTBNA 22 gauge needle", "DEV_NEEDLE", "Olympus EBUS-TBNA 22G", "with the", "evt_02"),
    ("11L", "ANAT_LN_STATION", "11L", "beginning with the", "evt_02"),
    ("4L", "ANAT_LN_STATION", "4L", "Lymph node", "evt_02"),
    ("7", "ANAT_LN_STATION", "7", "4L", "evt_02"),
    ("4R", "ANAT_LN_STATION", "4R", "7", "evt_02"),
    ("11Rs", "ANAT_LN_STATION", "11Rs", "4R", "evt_02"),
    ("11Ri", "ANAT_LN_STATION", "11Ri", "11Rs", "evt_02"),
    ("ROSE did not identify malignancy", "OBS_ROSE", "No Malignancy", "12] ", "evt_02"),
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications:", "evt_03"),
    ("flexible bronchoscopy", "PROC_METHOD", "Bronchoscopy", "successful", "evt_01"),
]

# -------------------------------------------------------------------------
# 4. EVENT DEFINITIONS
# -------------------------------------------------------------------------

EVENTS = [
    {
        "event_id": "evt_01",
        "name": "Diagnostic Bronchoscopy",
        "concept_id": "CPT-31622",
        "start": "Introduction of Q190 scope",
        "end": "Removal of scope",
        "spans": [s for s in SPANS if s[4] == "evt_01"]
    },
    {
        "event_id": "evt_02",
        "name": "EBUS-TBNA",
        "concept_id": "CPT-31652",
        "start": "Introduction of UC180F scope",
        "end": "Completion of sampling",
        "spans": [s for s in SPANS if s[4] == "evt_02"]
    },
    {
        "event_id": "evt_03",
        "name": "Procedure Outcome",
        "concept_id": "OUTCOME",
        "start": "Procedure End",
        "end": "Discharge",
        "spans": [s for s in SPANS if s[4] == "evt_03"]
    }
]

# -------------------------------------------------------------------------
# 5. HELPER FUNCTIONS
# -------------------------------------------------------------------------

def clean_text(text):
    if not text:
        return ""
    return text.strip()

def hydrate_span(text, span_text, context_prefix):
    """
    Finds the start and end indices of span_text in text, 
    using context_prefix to disambiguate.
    """
    # Normalize inputs
    text_norm = text.replace('\r', '').replace('\n', ' ')
    span_norm = span_text.replace('\r', '').replace('\n', ' ')
    context_norm = context_prefix.replace('\r', '').replace('\n', ' ')
    
    # Attempt 1: Exact search with context
    search_str = (context_norm + span_norm) if context_norm else span_norm
    start_index = text_norm.find(search_str)
    
    if start_index != -1:
        # If context used, adjust start to point to the actual span
        if context_norm:
            start_index += len(context_norm)
        end_index = start_index + len(span_norm)
        return start_index, end_index, span_norm
        
    # Attempt 2: Just the span text (fallback)
    start_index = text_norm.find(span_norm)
    if start_index != -1:
        end_index = start_index + len(span_norm)
        return start_index, end_index, span_norm
        
    return -1, -1, span_text # Not found

# -------------------------------------------------------------------------
# 6. WORKBOOK GENERATION FUNCTION
# -------------------------------------------------------------------------

def generate_workbook():
    # A. Create/Load Workbook
    wb = openpyxl.Workbook()
    
    # B. Sheet 1: Note_Text
    ws_text = wb.active
    ws_text.title = "Note_Text"
    ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # C. Sheet 2: Note_Index (Metadata + Flags)
    ws_index = wb.create_sheet("Note_Index")
    headers = ["note_id", "source_file", "procedure_date"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers)
    
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE]
    for key in PROCEDURE_FLAGS:
        row_data.append(PROCEDURE_FLAGS[key])
    ws_index.append(row_data)
    
    # D. Sheet 3: Span_Annotations (Raw Spans)
    ws_annot = wb.create_sheet("Span_Annotations")
    annot_headers = ["note_id", "event_id", "label", "value", "span_text", "start_char", "end_char"]
    ws_annot.append(annot_headers)
    
    for span in SPANS:
        s_text, s_label, s_val, s_ctx, s_evt = span
        # Start/End left blank in this sheet as per convention, calculated in Hydrated
        ws_annot.append([NOTE_ID, s_evt, s_label, s_val, s_text, "", ""])

    # E. Sheet 4: Span_Hydrated (Calculated Offsets)
    ws_hydrated = wb.create_sheet("Span_Hydrated")
    ws_hydrated.append(annot_headers)
    
    for span in SPANS:
        s_text, s_label, s_val, s_ctx, s_evt = span
        start, end, found_text = hydrate_span(NOTE_TEXT, s_text, s_ctx)
        ws_hydrated.append([NOTE_ID, s_evt, s_label, s_val, found_text, start, end])

    # F. Sheet 5: Event_Log (Flattened)
    ws_events = wb.create_sheet("Event_Log")
    ws_events.append(["note_id", "event_id", "event_name", "concept_id", "start_time", "end_time"])
    
    for evt in EVENTS:
        ws_events.append([
            NOTE_ID,
            evt["event_id"],
            evt["name"],
            evt["concept_id"],
            evt.get("start", ""),
            evt.get("end", "")
        ])

    # G. Sheet 6: V3_Procedure_Events (Detailed)
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    ws_v3.append(["note_id", "event_id", "event_type", "event_details_json"])
    
    for evt in EVENTS:
        details = {
            "name": evt["name"],
            "spans": [
                {"label": s[1], "value": s[2], "text": s[0]} 
                for s in evt["spans"]
            ]
        }
        ws_v3.append([NOTE_ID, evt["event_id"], evt["name"], json.dumps(details)])

    # H. Sheet 7: V3_Registry_JSON (Full Export)
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["note_id", "json_data"])
    
    registry_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source": SOURCE_FILE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS
    }
    
    # Serialize spans within events to avoid object reference issues
    for e in registry_data["events"]:
        e["spans"] = [
            {"text": s[0], "label": s[1], "normalized": s[2]} 
            for s in e["spans"]
        ]

    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])

    # I. Save
    wb.save(OUTPUT_PATH)
    print(f"Workbook generated at: {os.path.abspath(OUTPUT_PATH)}")

# -------------------------------------------------------------------------
# 7. EXECUTION
# -------------------------------------------------------------------------
if __name__ == "__main__":
    generate_workbook()