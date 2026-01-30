import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import re
import json
import os
import datetime

# 1. Imports & Constants
NOTE_ID = "note_178"
SOURCE_FILE = "note_178.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_178 SOURCE_FILE: note_178.txt Indications: Mediastinal adenopathy
Medications: General Anesthesia,
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention.
Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway was in good position. The vocal cords appeared normal. The subglottic space was normal.
The trachea was of normal caliber. The carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level. Bronchial mucosa and anatomy were normal;
there are no endobronchial lesions, and no secretions. The video bronchoscope was then removed and the UC180F convex probe EBUS bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
Just distal to the cricoid cartilage the suspicious 2L lymph node was identified Sampling by transbronchial needle aspiration was performed with both the Olympus 19G and Olympus 22G needles with a total of 8 passes performed.
Abundant lymphocytes were seen in rapid onsite pathological evaluation. Samples were sent for both flow and routine cytology.
Following completion of EBUS bronchoscopy, the Q190 video bronchoscope was then re-inserted and after suctioning blood and secretions there was no evidence of active bleeding and the bronchoscope was subsequently removed.
Complications: No immediate complications
Estimated Blood Loss: min

Post Procedure Diagnosis:
- Technically successful flexible bronchoscopy with endobronchial ultrasound-guided biopsies.
- The patient has remained stable and has been transferred in good condition to the post-surgical monitoring unit.
- Will await final pathology results"""

# 2. Configuration (Procedure Flags)
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 3. Data Definition (Spans)
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190", "tracheobronchial tree, the ", "evt_01"),
    ("laryngeal mask airway", "ANAT_AIRWAY", "LMA", "mouth, via ", "evt_01"),
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "UC180F", "removed and the ", "evt_02"),
    ("2L", "ANAT_LN_STATION", "2L", "the suspicious ", "evt_02"),
    ("transbronchial needle aspiration", "PROC_METHOD", "Linear EBUS-TBNA", "Sampling by ", "evt_02"),
    ("Olympus 19G", "DEV_NEEDLE", "19G", "with both the ", "evt_02"),
    ("Olympus 22G", "DEV_NEEDLE", "22G", "and ", "evt_02"),
    ("total of 8 passes", "MEAS_COUNT", "8", "needles with a ", "evt_02"),
    ("Abundant lymphocytes", "OBS_ROSE", "Lymphocytes", "evaluation. ", "evt_02"),
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications: ", "evt_03"),
]

# 4. Event Definitions
EVENTS = [
    {
        "event_id": "evt_01",
        "name": "Diagnostic Inspection",
        "concept_id": "CPT-31622",
        "type": "Diagnostic",
        "location": "Airway",
        "devices": ["Q190 video bronchoscope", "LMA"],
        "outcomes": [],
        "notes": "Initial airway inspection"
    },
    {
        "event_id": "evt_02",
        "name": "EBUS-TBNA 2L",
        "concept_id": "CPT-31652",
        "type": "Therapeutic/Diagnostic",
        "location": "2L",
        "devices": ["UC180F", "Olympus 19G", "Olympus 22G"],
        "outcomes": ["ROSE: Abundant lymphocytes"],
        "notes": "Linear EBUS sampling with 8 passes"
    },
    {
        "event_id": "evt_03",
        "name": "Procedure Outcome",
        "concept_id": "OUTCOME",
        "type": "Outcome",
        "location": "Global",
        "devices": [],
        "outcomes": ["No immediate complications"],
        "notes": "Patient stable"
    }
]

# 5. Helper Functions
def clean_text(text):
    if not text:
        return ""
    return re.sub(r'[\r\n]+', ' ', text).strip()

def hydrate_span(text, span_text, context_prefix=""):
    """
    Finds the start and end indices of span_text in text.
    Uses context_prefix to disambiguate if needed.
    """
    clean_full = clean_text(text)
    clean_span = clean_text(span_text)
    clean_context = clean_text(context_prefix)
    
    # Try exact match with context
    search_pattern = re.escape(clean_context) + r"\s*" + re.escape(clean_span)
    match = re.search(search_pattern, clean_full, re.IGNORECASE)
    
    if match:
        # The span starts after the context
        start_index = match.start() + len(clean_context)
        # Adjust for potential spacing between context and span in the match
        # We find exactly where the span text starts within the matched string
        full_match_str = match.group(0)
        span_start_in_match = full_match_str.lower().rfind(clean_span.lower())
        
        final_start = match.start() + span_start_in_match
        final_end = final_start + len(clean_span)
        return final_start, final_end, clean_full[final_start:final_end]
    
    # Fallback: simple search (take first occurrence if context fails)
    start = clean_full.lower().find(clean_span.lower())
    if start != -1:
        return start, start + len(clean_span), clean_full[start:start + len(clean_span)]
    
    return -1, -1, "NOT_FOUND"

def generate_workbook():
    # Create workbook
    wb = openpyxl.Workbook()
    
    # --- Sheet 1: Note_Text ---
    ws_text = wb.active
    ws_text.title = "Note_Text"
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, clean_text(NOTE_TEXT)])
    
    # --- Sheet 2: Note_Index ---
    ws_index = wb.create_sheet("Note_Index")
    headers = ["NOTE_ID", "SOURCE_FILE", "PROCEDURE_DATE"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers)
    
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE]
    for key in PROCEDURE_FLAGS:
        row_data.append(PROCEDURE_FLAGS[key])
    ws_index.append(row_data)
    
    # --- Sheet 3: Span_Annotations ---
    ws_anno = wb.create_sheet("Span_Annotations")
    ws_anno.append(["NOTE_ID", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "EVENT_ID", "START_CHAR", "END_CHAR"])
    
    for span in SPANS:
        # Unpack span tuple (text, label, norm, context, evt)
        s_text, s_label, s_norm, s_context, s_evt = span
        ws_anno.append([NOTE_ID, s_text, s_label, s_norm, s_context, s_evt, "", ""])
        
    # --- Sheet 4: Span_Hydrated ---
    ws_hydra = wb.create_sheet("Span_Hydrated")
    ws_hydra.append(["NOTE_ID", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "EVENT_ID", "START_CHAR", "END_CHAR", "VERIFIED_TEXT"])
    
    cleaned_note_text = clean_text(NOTE_TEXT)
    
    for span in SPANS:
        s_text, s_label, s_norm, s_context, s_evt = span
        start, end, verified = hydrate_span(cleaned_note_text, s_text, s_context)
        ws_hydra.append([NOTE_ID, s_text, s_label, s_norm, s_context, s_evt, start, end, verified])

    # --- Sheet 5: Event_Log ---
    ws_event = wb.create_sheet("Event_Log")
    ws_event.append(["NOTE_ID", "EVENT_ID", "EVENT_NAME", "CONCEPT_ID", "TYPE", "LOCATION", "DEVICES", "OUTCOMES", "NOTES"])
    
    for evt in EVENTS:
        ws_event.append([
            NOTE_ID,
            evt["event_id"],
            evt["name"],
            evt["concept_id"],
            evt["type"],
            evt["location"],
            ", ".join(evt["devices"]),
            ", ".join(evt["outcomes"]),
            evt["notes"]
        ])

    # --- Sheet 6: V3_Procedure_Events ---
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    ws_v3.append(["NOTE_ID", "EVENT_ID", "V3_JSON"])
    
    for evt in EVENTS:
        # Create a simplified V3 structure
        v3_obj = {
            "event_id": evt["event_id"],
            "procedure_type": evt["type"],
            "anatomy": evt["location"],
            "devices": evt["devices"],
            "outcomes": evt["outcomes"]
        }
        ws_v3.append([NOTE_ID, evt["event_id"], json.dumps(v3_obj)])

    # --- Sheet 7: V3_Registry_JSON ---
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["NOTE_ID", "FULL_JSON"])
    
    registry_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source": SOURCE_FILE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS,
        "spans": [
            {
                "text": s[0], 
                "label": s[1], 
                "norm": s[2], 
                "context": s[3], 
                "event_id": s[4]
            } for s in SPANS
        ]
    }
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])

    # Formatting
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Save
    if not os.path.exists(OUTPUT_PATH):
        wb.save(OUTPUT_PATH)
        print(f"Generated {OUTPUT_PATH}")
    else:
        # Overwrite protection or logic could go here, but strictly we just save
        wb.save(OUTPUT_PATH)
        print(f"Overwrote {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()