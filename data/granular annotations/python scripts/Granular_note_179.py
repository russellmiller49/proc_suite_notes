import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import re
import json
import os
import datetime

# -------------------------------------------------------------------------
# 1. Imports & Constants
# -------------------------------------------------------------------------
NOTE_ID = "note_179"
SOURCE_FILE = "note_179.txt"
PROCEDURE_DATE = ""  # No specific date in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_179 SOURCE_FILE: note_179.txt Indications: Hilar cyst aspiration 
Procedure: EBUS bronchoscopy – single station
Medications: General Anesthesia
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention.
Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway initially was not well positioned but after replacement with larger LMA seated in good position.
The vocal cords appeared normal. The subglottic space was normal. The trachea was of normal caliber. The carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level.
Bronchial mucosa and anatomy were normal with the exception of partial extrinsic compression of the left upper lobe bronchus distal to the lingual and most obvious in the apical segment;
there are no endobronchial lesions. The video bronchoscope was then removed and the UC180F convex probe EBUS bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The scope was advanced into the proximal left upper lobe and a 45mm heterogeneous cyst with multiple septations was seen just distal to the left PA.
Sampling by transbronchial needle aspiration was performed with the Olympus 19G Visioshot EBUS-TBNA needle.
Material was thick which could not be grossly  aspirated through the needle with attached suction.
Samples showed thick coagulated bloody material on multiple needle passes consistent with old blood.
Rapid onsite pathological evaluation showed heme and multiple foamy macrophages.
Samples and fluid were sent for both culture and routine cytology.
Following completion of EBUS bronchoscopy, the Q190 video bronchoscope was then re-inserted and after suctioning blood and secretions there was no evidence of active bleeding and the bronchoscope was subsequently removed.
Complications: No immediate complications
Post-operative diagnosis: Hemorrhagic Hilar cyst
Estimated Blood Loss: 10cc
Recommendations:
- Transferred patient to post-procedural monitoring 
- 10 day course of Augmentin to reduced likelihood of infection of capsulated cystic space.
- Will await final pathology results
- Repeat CT in 3-4 weeks or earlier if fevers, persistent/worsening chest pain, hemoptysis or other symptoms concerning for infection or cyst rupture.
- Will likely require unroofing and resection with thoracic surgery"""

# -------------------------------------------------------------------------
# 2. Configuration (Procedure Flags)
# -------------------------------------------------------------------------
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# -------------------------------------------------------------------------
# 3. Data Definition (Spans)
# -------------------------------------------------------------------------
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("EBUS bronchoscopy", "PROC_METHOD", "Linear EBUS", "Procedure: ", "evt_01"),
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190 Scope", "tracheobronchial tree, the ", "evt_01"),
    ("laryngeal mask airway", "DEV_INSTRUMENT", "LMA", "via ", "evt_01"),
    ("partial extrinsic compression", "OBS_LESION", "Extrinsic Compression", "exception of ", "evt_01"),
    ("left upper lobe bronchus", "ANAT_AIRWAY", "LUL Bronchus", "compression of the ", "evt_01"),
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "UC180F Scope", "removed and the ", "evt_02"),
    ("proximal left upper lobe", "ANAT_AIRWAY", "LUL Proximal", "into the ", "evt_02"),
    ("45mm heterogeneous cyst", "OBS_LESION", "Cyst", "and a ", "evt_02"),
    ("transbronchial needle aspiration", "PROC_ACTION", "TBNA", "Sampling by ", "evt_02"),
    ("Olympus 19G Visioshot EBUS-TBNA needle", "DEV_NEEDLE", "Olympus 19G", "with the ", "evt_02"),
    ("thick coagulated bloody material", "OBS_ROSE", "Thick bloody material", "showed ", "evt_02"),
    ("heme and multiple foamy macrophages", "OBS_ROSE", "Heme/Macrophages", "showed ", "evt_02"),
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190 Scope", "EBUS bronchoscopy, the ", "evt_03"),
    ("suctioning", "PROC_ACTION", "Suction", "and after ", "evt_03"),
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications: ", "evt_04")
]

# -------------------------------------------------------------------------
# 4. Event Definitions (V3 Registry)
# -------------------------------------------------------------------------
EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Bronchoscopy",
        "method": "Diagnostic Bronchoscopy",
        "anatomy": "Left Upper Lobe Bronchus",
        "devices": ["Q190 video bronchoscope", "LMA"],
        "observations": "Partial extrinsic compression"
    },
    {
        "event_id": "evt_02",
        "procedure_type": "Bronchoscopy",
        "method": "Linear EBUS",
        "anatomy": "Proximal Left Upper Lobe",
        "devices": ["UC180F convex probe EBUS bronchoscope", "Olympus 19G Visioshot EBUS-TBNA needle"],
        "actions": ["Transbronchial Needle Aspiration", "Cyst Aspiration"],
        "observations": "45mm heterogeneous cyst",
        "rose_results": "Heme and multiple foamy macrophages"
    },
    {
        "event_id": "evt_03",
        "procedure_type": "Bronchoscopy",
        "method": "Toilet Bronchoscopy",
        "anatomy": "Tracheobronchial Tree",
        "devices": ["Q190 video bronchoscope"],
        "actions": ["Suctioning"]
    },
    {
        "event_id": "evt_04",
        "procedure_type": "Outcome",
        "outcome_complication": "No immediate complications"
    }
]

# -------------------------------------------------------------------------
# 5. Helper Functions
# -------------------------------------------------------------------------
def clean_text(text):
    if not text:
        return ""
    return re.sub(r'[\r\n]+', ' ', text).strip()

def hydrate_span(full_text, span_text, context_prefix):
    """
    Finds the start and end indices of span_text within full_text.
    Uses context_prefix to disambiguate repeated terms.
    """
    clean_full = clean_text(full_text).lower()
    clean_span = clean_text(span_text).lower()
    clean_context = clean_text(context_prefix).lower()
    
    # Construct search pattern with context
    search_phrase = clean_context + clean_span
    
    start_index = clean_full.find(search_phrase)
    
    if start_index == -1:
        # Fallback: try finding just the span if context fails (though risking ambiguity)
        start_index = clean_full.find(clean_span)
    
    if start_index != -1:
        # Adjust start index to skip the context
        if clean_context in clean_full[start_index:start_index+len(search_phrase)]:
             real_start = start_index + len(clean_context)
        else:
             real_start = start_index
             
        # Find exact casing in original text for length calculation if needed, 
        # but indices are usually based on the string passed.
        # We will return indices based on the passed full_text (assuming simple whitespace normalization)
        
        # To map back to original text with newlines is complex without a robust map.
        # For Phase 0, we typically normalize the text in the sheet first.
        # Here we return indices relative to the normalized string.
        return real_start, real_start + len(clean_span)
    
    return None, None

def find_offsets(text, span_text, label, norm_val, context, evt_id):
    start, end = hydrate_span(text, span_text, context)
    return {
        "event_id": evt_id,
        "label": label,
        "text": span_text,
        "normalized_value": norm_val,
        "start": start,
        "end": end,
        "context_prefix": context
    }

# -------------------------------------------------------------------------
# 6. Workbook Generation Function
# -------------------------------------------------------------------------
def generate_workbook():
    wb = openpyxl.Workbook()
    
    # 1. Sheet: Note_Text
    ws_text = wb.active
    ws_text.title = "Note_Text"
    ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, clean_text(NOTE_TEXT)])
    
    # 2. Sheet: Note_Index
    ws_index = wb.create_sheet("Note_Index")
    headers = ["note_id", "source_file", "procedure_date"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers)
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE] + [PROCEDURE_FLAGS[k] for k in PROCEDURE_FLAGS]
    ws_index.append(row_data)
    
    # 3. Sheet: Span_Annotations (Raw input for pipeline)
    ws_anno = wb.create_sheet("Span_Annotations")
    ws_anno.append(["note_id", "event_id", "label", "span_text", "normalized_value", "context_prefix", "start_char", "end_char"])
    
    # 4. Sheet: Span_Hydrated (Calculated offsets)
    ws_hydrated = wb.create_sheet("Span_Hydrated")
    ws_hydrated.append(["note_id", "event_id", "label", "span_text", "normalized_value", "start_char", "end_char", "context_prefix"])
    
    normalized_text = clean_text(NOTE_TEXT)
    
    for item in SPANS:
        span_text, label, norm_val, context, evt_id = item
        
        # Add to Annotations (Leaving offsets blank as per "raw" requirement, or pre-filled if manual)
        ws_anno.append([NOTE_ID, evt_id, label, span_text, norm_val, context, "", ""])
        
        # Calculate offsets for Hydrated
        data = find_offsets(normalized_text, span_text, label, norm_val, context, evt_id)
        ws_hydrated.append([
            NOTE_ID, 
            data["event_id"], 
            data["label"], 
            data["text"], 
            data["normalized_value"], 
            data["start"], 
            data["end"],
            data["context_prefix"]
        ])

    # 5. Sheet: Event_Log (Flattened)
    ws_event_log = wb.create_sheet("Event_Log")
    ws_event_log.append(["note_id", "event_id", "event_type", "details_json"])
    for evt in EVENTS:
        ws_event_log.append([NOTE_ID, evt["event_id"], "V3_Event", json.dumps(evt)])

    # 6. Sheet: V3_Procedure_Events (Detailed columns)
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    # Define a superset of columns based on keys present in EVENTS
    all_keys = set().union(*(d.keys() for d in EVENTS))
    headers = ["note_id"] + sorted(list(all_keys))
    ws_v3.append(headers)
    
    for evt in EVENTS:
        row = [NOTE_ID]
        for k in headers[1:]:
            val = evt.get(k, "")
            if isinstance(val, (list, dict)):
                row.append(json.dumps(val))
            else:
                row.append(val)
        ws_v3.append(row)

    # 7. Sheet: V3_Registry_JSON
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["note_id", "full_json_object"])
    
    registry_object = {
        "metadata": {
            "note_id": NOTE_ID,
            "source_file": SOURCE_FILE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS
    }
    ws_json.append([NOTE_ID, json.dumps(registry_object, indent=2)])

    # Save
    if os.path.exists(OUTPUT_PATH):
        os.remove(OUTPUT_PATH)
    wb.save(OUTPUT_PATH)
    print(f"Workbook generated: {OUTPUT_PATH}")

# -------------------------------------------------------------------------
# 7. Execution Block
# -------------------------------------------------------------------------
if __name__ == "__main__":
    generate_workbook()