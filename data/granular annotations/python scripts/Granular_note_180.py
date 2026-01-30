import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# 1. Imports & Constants
NOTE_ID = "note_180"
SOURCE_FILE = "note_180.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_180 SOURCE_FILE: note_180.txt Indications: Hilar adenopathy
Procedure: EBUS bronchoscopy – single station
Medications: General Anesthesia
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention.
Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway initially was not well positioned but after replacement with larger LMA seated in good position.
The vocal cords appeared normal. The subglottic space was normal. The trachea was of normal caliber. The carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level. Bronchial mucosa and anatomy were normal;
there are no endobronchial lesions. There were thick greenish mucus plugs throughout the lungs which were easily suctioned.
The video bronchoscope was then removed and the UC180F convex probe EBUS bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The scope was advanced into the proximal left upper lobe and a 10mm hypoechoic structure was seen in the area of the 10L lymph node.
Within the hypoechoic structure were hyperechoic bands similar to those seen in loculated pleural effusions.
Sampling by transbronchial needle aspiration was performed with the Olympus 22G EBUS-TBNA needle and serous appearing liquid was noted within the suction line and attached syringe (approximately 2cc) consistent with cyst.
Subsequent pass with an Olympus 19G EBUS-TBNA needle yielded similar results with another 3cc fluid return.
Ultrasound showed shrinkage of the cyst. Three additional needle passes were then performed.
Rapid onsite pathological evaluation showed lymphocytes within the samples. Samples and fluid were sent for both flow and routine cytology.
Following completion of EBUS bronchoscopy, the Q190 video bronchoscope was then re-inserted and after suctioning blood and secretions there was no evidence of active bleeding and the bronchoscope was subsequently removed.
Complications: No immediate complications
Post-operative diagnosis: Hilar cyst (bronchogenic vs cardiac)
Estimated Blood Loss: 5cc
Recommendations:
- Transferred patient to post-procedural monitoring unit and discharge when standard criteria are met.
- 5 day course of Augmentin to reduce likelihood of infection of residual cystic space.
- Will await final pathology results
- F/U with primary Pulmonologist Dr. Butlin once results are available."""

# 2. Configuration (Procedure Flags)
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 1, # Cyst aspiration mentioned with shrinkage
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 3. Data Definition (Spans)
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190", "tracheobronchial tree, the ", "evt_01"),
    ("laryngeal mask airway", "DEV_INSTRUMENT", "LMA", "through the mouth, via ", "evt_01"),
    ("thick greenish mucus plugs", "OBS_LESION", "mucus plugs", "lesions. There were ", "evt_01"),
    ("suctioned", "PROC_ACTION", "suction", "which were easily ", "evt_01"),
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "UC180F", "removed and the ", "evt_02"),
    ("proximal left upper lobe", "ANAT_LUNG_LOC", "LUL", "advanced into the ", "evt_02"),
    ("10mm", "MEAS_SIZE", "10mm", "upper lobe and a ", "evt_02"),
    ("hypoechoic structure", "OBS_LESION", "cyst", "and a 10mm ", "evt_02"),
    ("10L", "ANAT_LN_STATION", "10L", "area of the ", "evt_02"),
    ("Olympus 22G EBUS-TBNA needle", "DEV_NEEDLE", "22G", "performed with the ", "evt_02"),
    ("serous appearing liquid", "OBS_LESION", "serous fluid", "needle and ", "evt_02"),
    ("2cc", "MEAS_VOL", "2cc", "syringe (approximately ", "evt_02"),
    ("Olympus 19G EBUS-TBNA needle", "DEV_NEEDLE", "19G", "pass with an ", "evt_02"),
    ("3cc", "MEAS_VOL", "3cc", "results with another ", "evt_02"),
    ("shrinkage", "OUTCOME_SYMPTOMS", "cyst shrinkage", "Ultrasound showed ", "evt_02"),
    ("Three", "MEAS_COUNT", "3", "shrinkage of the cyst. ", "evt_02"),
    ("lymphocytes", "OBS_ROSE", "lymphocytes", "evaluation showed ", "evt_02"),
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications: ", "evt_03")
]

# 4. Event Definitions
EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Diagnostic Bronchoscopy",
        "action": "Inspection",
        "anatomy": "Tracheobronchial Tree",
        "devices": ["Q190 video bronchoscope", "Laryngeal Mask Airway"],
        "outcomes": ["Mucus plugs suctioned"]
    },
    {
        "event_id": "evt_02",
        "procedure_type": "EBUS-TBNA",
        "action": "Aspiration/Biopsy",
        "anatomy": "10L",
        "devices": ["UC180F", "22G Needle", "19G Needle"],
        "samples": ["ROSE: Lymphocytes"],
        "outcomes": ["Cyst shrinkage", "5cc fluid removed"]
    },
    {
        "event_id": "evt_03",
        "procedure_type": "Procedure Outcome",
        "action": "Conclusion",
        "outcomes": ["No immediate complications"]
    }
]

# 5. Helper Functions
def clean_text(text):
    if not text: return ""
    return str(text).strip().replace('\r', '').replace('\n', ' ')

def hydrate_span(text, span_text, context_prefix):
    """
    Finds the start and end indices of span_text in text, 
    ensuring it follows context_prefix.
    """
    cleaned_text = text.replace('\r', '').replace('\n', ' ')
    
    # Escape special regex characters in the span and context
    pattern_str = re.escape(context_prefix) + r"\s*(" + re.escape(span_text) + r")"
    
    match = re.search(pattern_str, cleaned_text, re.IGNORECASE)
    
    if match:
        # The group(1) is the span_text part
        start_char = match.start(1)
        end_char = match.end(1)
        return start_char, end_char
    else:
        # Fallback: simple search if context fails (though strict mode prefers context)
        fallback_idx = cleaned_text.find(span_text)
        if fallback_idx != -1:
            return fallback_idx, fallback_idx + len(span_text)
        return "", ""

# 6. Workbook Generation Function
def generate_workbook():
    # Create workbook
    wb = openpyxl.Workbook()
    
    # --- Sheet 1: Note_Text ---
    ws_text = wb.active
    ws_text.title = "Note_Text"
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # --- Sheet 2: Note_Index ---
    ws_index = wb.create_sheet("Note_Index")
    headers = ["NOTE_ID", "proc_date"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers)
    
    row_data = [NOTE_ID, PROCEDURE_DATE]
    for key in PROCEDURE_FLAGS:
        row_data.append(PROCEDURE_FLAGS[key])
    ws_index.append(row_data)
    
    # --- Sheet 3: Span_Annotations ---
    ws_annot = wb.create_sheet("Span_Annotations")
    ws_annot.append(["NOTE_ID", "start_char", "end_char", "span_text", "label", "normalized_value", "context_prefix", "event_id"])
    
    for span in SPANS:
        # span = (text, label, norm, context, event_id)
        ws_annot.append([NOTE_ID, "", "", span[0], span[1], span[2], span[3], span[4]])
        
    # --- Sheet 4: Span_Hydrated ---
    ws_hydrated = wb.create_sheet("Span_Hydrated")
    ws_hydrated.append(["NOTE_ID", "start_char", "end_char", "span_text", "label", "normalized_value", "context_prefix", "event_id"])
    
    for span in SPANS:
        s_text, label, norm, context, evt_id = span
        start, end = hydrate_span(NOTE_TEXT, s_text, context)
        ws_hydrated.append([NOTE_ID, start, end, s_text, label, norm, context, evt_id])
        
    # --- Sheet 5: Event_Log ---
    ws_event = wb.create_sheet("Event_Log")
    ws_event.append(["NOTE_ID", "event_id", "procedure_type", "action", "anatomy", "devices", "outcomes"])
    
    for evt in EVENTS:
        ws_event.append([
            NOTE_ID,
            evt.get("event_id", ""),
            evt.get("procedure_type", ""),
            evt.get("action", ""),
            evt.get("anatomy", ""),
            ", ".join(evt.get("devices", [])),
            ", ".join(evt.get("outcomes", []))
        ])
        
    # --- Sheet 6: V3_Procedure_Events ---
    # Simplified representation for the registry structure
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    ws_v3.append(["NOTE_ID", "event_id", "event_json"])
    
    for evt in EVENTS:
        ws_v3.append([NOTE_ID, evt.get("event_id"), json.dumps(evt)])

    # --- Sheet 7: V3_Registry_JSON ---
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["NOTE_ID", "full_registry_json"])
    
    registry_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source": SOURCE_FILE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS,
        "spans": [
            {
                "text": s[0],
                "label": s[1],
                "normalized": s[2],
                "event_id": s[4]
            } for s in SPANS
        ]
    }
    
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])
    
    # Save
    if os.path.exists(OUTPUT_PATH):
        os.remove(OUTPUT_PATH)
    wb.save(OUTPUT_PATH)
    print(f"Generated Phase 0 workbook: {OUTPUT_PATH}")

# 7. Execution Block
if __name__ == "__main__":
    generate_workbook()