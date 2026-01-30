import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import re
import json
import os
import datetime

# -------------------------------------------------------------------------
# 1. IMPORTS & CONSTANTS
# -------------------------------------------------------------------------

NOTE_ID = "note_181"
SOURCE_FILE = "note_181.txt"
PROCEDURE_DATE = ""  # No specific date in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_181 SOURCE_FILE: note_181.txt PRE-PROCEDURE DIAGNISOS: LEFT UPPER LOBE PULMONARY NODULE
POST- PROCEDURE DIAGNISOS: LEFT UPPER LOBE PULMONARY NODULE
PROCEDURE PERFORMED:  
Flexible bronchoscopy with electromagnetic navigation under flouroscopic and EBUS guidance with transbronchial needle aspiration, Transbronchial biopsy and bronchioalveolar lavage.
CPT 31654 Bronchoscope with Endobronchial Ultrasound guidance for peripheral lesion
CPT 31629 Flexible bronchoscopy with fluoroscopic trans-bronchial needle aspiration
CPT 31628 Bronchoscopy, rigid or flexible, including fluoroscopic guidance, when performed;
with transbronchial lung biopsy(s), single lobe
CPT +31624 Bronchoscopy, rigid or flexible, including fluoroscopic guidance, when performed;
with bronchial alveolar lavage
CPT +31627 Bronchoscopy with computer assisted image guided navigation
INDICATIONS FOR EXAMINATION:   Left upper lobe lung nodule            
MEDICATIONS:    GA
FINDINGS: Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the endotracheal tube and advanced to the tracheobronchial tree.
The trachea was of normal caliber. The carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level. Bronchial mucosa and anatomy were normal;
there are no endobronchial lesions. We then removed the diagnostic Q190 bronchoscope and the super-dimension navigational catheter was inserted through the T190 therapeutic bronchoscope and advanced into the airway.
Using navigational map we attempted to advance the 180 degree edge catheter into the proximity of the lesion within apico-posterior branch of left upper lobe.
Radial probe was used to attempt to confirm presence within the lesion.
Although we were able to navigate directly to the lesion with navigation the radial probe view was suboptimal.
Biopsy was performed initially with triple needle brush and TBNA needle.
ROSE did not reveal evidence to support that we were within the lesion.
Multiple attempts were made to manipulate the catheter and biopsies were then performed with a variety of instruments to include peripheral needle, and forceps, brush under fluoroscopic visualization.
The specimens reviewed on-site remained suboptimal.  Multiple forceps biopsies were performed within the location of the lesion and placed in cell-block.
After which a mini-BAL was then performed through the super-D catheter.
We then removed the therapeutic bronchoscope with super-D catheter and reinserted the diagnostic scope at which point repeat airway inspection was then performed and once we were satisfied that no bleeding occurred, the bronchoscope was removed and the procedure completed.
ESTIMATED BLOOD LOSS:   None 
COMPLICATIONS:                 None

IMPRESSION:  
- S/P bronchoscopy with biopsy and lavage.
- Suboptimal navigational localization 
RECOMMENDATIONS
- Transfer to post-procedural unit
- Post-procedure CXR
- D/C home once criteria met
- Await pathology"""

# -------------------------------------------------------------------------
# 2. CONFIGURATION (PROCEDURE FLAGS)
# -------------------------------------------------------------------------
# Mapped based on explicit text in the note.
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 1,  # "bronchioalveolar lavage", "mini-BAL"
    "bronchial_wash": 0,
    "brushings": 1,  # "brush"
    "endobronchial_biopsy": 0,
    "tbna_conventional": 1,  # "transbronchial needle aspiration" explicitly listed
    "linear_ebus": 0,
    "radial_ebus": 1,  # "Radial probe", "EBUS guidance for peripheral lesion"
    "navigational_bronchoscopy": 1,  # "electromagnetic navigation", "super-dimension"
    "transbronchial_biopsy": 1,  # "Transbronchial biopsy"
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# -------------------------------------------------------------------------
# 3. DATA DEFINITION (SPANS)
# -------------------------------------------------------------------------
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    # Header / Procedure Performed
    ("Flexible bronchoscopy", "PROC_METHOD", "Flexible Bronchoscopy", "PROCEDURE PERFORMED:  \n", "evt_01"),
    ("electromagnetic navigation", "PROC_METHOD", "Navigational Bronchoscopy", "Flexible bronchoscopy with ", "evt_02"),
    ("EBUS guidance", "PROC_METHOD", "Radial EBUS", "flouroscopic and ", "evt_02"),
    ("transbronchial needle aspiration", "PROC_METHOD", "TBNA", "guidance with ", "evt_03"),
    ("Transbronchial biopsy", "PROC_METHOD", "Transbronchial Biopsy", "transbronchial needle aspiration, ", "evt_04"),
    ("bronchioalveolar lavage", "PROC_METHOD", "BAL", "Transbronchial biopsy and ", "evt_05"),
    
    # Anatomy
    ("Left upper lobe", "ANAT_LUNG_LOC", "Left Upper Lobe", "INDICATIONS FOR EXAMINATION:   ", "evt_02"),
    
    # Findings - Inspection
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Olympus Q190", "tree, the ", "evt_01"),
    ("Trachea", "ANAT_AIRWAY", "Trachea", "The ", "evt_01"),
    ("carina", "ANAT_AIRWAY", "Carina", "The ", "evt_01"),
    
    # Navigation
    ("super-dimension navigational catheter", "DEV_CATHETER", "SuperDimension Catheter", "bronchoscope and the ", "evt_02"),
    ("T190 therapeutic bronchoscope", "DEV_INSTRUMENT", "Olympus T190", "through the ", "evt_02"),
    ("180 degree edge catheter", "DEV_CATHETER", "Edge Catheter 180", "advance the ", "evt_02"),
    ("apico-posterior branch", "ANAT_AIRWAY", "Apico-posterior branch", "lesion within ", "evt_02"),
    ("Radial probe", "DEV_INSTRUMENT", "Radial EBUS Probe", "left upper lobe.\n", "evt_02"),
    
    # Biopsy Round 1
    ("triple needle brush", "DEV_INSTRUMENT", "Triple Needle Brush", "initially with ", "evt_03"),
    ("TBNA needle", "DEV_NEEDLE", "TBNA Needle", "brush and ", "evt_03"),
    ("ROSE", "OBS_ROSE", "ROSE", "lesion.\n", "evt_03"),
    
    # Biopsy Round 2
    ("peripheral needle", "DEV_NEEDLE", "Peripheral Needle", "include ", "evt_04"),
    ("forceps", "DEV_INSTRUMENT", "Biopsy Forceps", "needle, and ", "evt_04"),
    ("brush", "DEV_INSTRUMENT", "Cytology Brush", "forceps, ", "evt_04"),
    
    # BAL
    ("mini-BAL", "PROC_METHOD", "BAL", "After which a ", "evt_05"),
    ("super-D catheter", "DEV_CATHETER", "SuperDimension Catheter", "through the ", "evt_05"),
    
    # Complications
    ("None", "OUTCOME_COMPLICATION", "None", "COMPLICATIONS:                 ", "evt_06")
]

# -------------------------------------------------------------------------
# 4. EVENT DEFINITIONS
# -------------------------------------------------------------------------
EVENTS = [
    {
        "event_id": "evt_01",
        "name": "Initial Inspection",
        "procedure_type": "Bronchoscopy",
        "action": "Inspection",
        "devices": ["Olympus Q190"],
        "anatomy": ["Trachea", "Carina"],
        "outcomes": []
    },
    {
        "event_id": "evt_02",
        "name": "Navigation and Radial EBUS",
        "procedure_type": "Navigational Bronchoscopy",
        "action": "Navigation",
        "devices": ["Olympus T190", "SuperDimension Catheter", "Edge Catheter 180", "Radial EBUS Probe"],
        "anatomy": ["Left Upper Lobe", "Apico-posterior branch"],
        "outcomes": []
    },
    {
        "event_id": "evt_03",
        "name": "Biopsy Round 1",
        "procedure_type": "TBNA",
        "action": "Biopsy",
        "devices": ["Triple Needle Brush", "TBNA Needle"],
        "anatomy": ["Left Upper Lobe"],
        "outcomes": ["ROSE Negative"]
    },
    {
        "event_id": "evt_04",
        "name": "Biopsy Round 2",
        "procedure_type": "Transbronchial Biopsy",
        "action": "Biopsy",
        "devices": ["Peripheral Needle", "Biopsy Forceps", "Cytology Brush"],
        "anatomy": ["Left Upper Lobe"],
        "outcomes": []
    },
    {
        "event_id": "evt_05",
        "name": "Lavage",
        "procedure_type": "BAL",
        "action": "Lavage",
        "devices": ["SuperDimension Catheter"],
        "anatomy": ["Left Upper Lobe"],
        "outcomes": []
    },
    {
        "event_id": "evt_06",
        "name": "Conclusion",
        "procedure_type": "Bronchoscopy",
        "action": "Completion",
        "devices": [],
        "anatomy": [],
        "outcomes": ["No Complications"]
    }
]

# -------------------------------------------------------------------------
# 5. HELPER FUNCTIONS
# -------------------------------------------------------------------------

def clean_text(text):
    if not text:
        return ""
    return text.strip().replace('\r', '')

def hydrate_span(text, span_text, context_prefix):
    cleaned_full = clean_text(text)
    cleaned_span = clean_text(span_text)
    cleaned_context = clean_text(context_prefix)

    if not cleaned_span:
        return 0, 0

    # Try exact match with context first
    search_pattern = re.escape(cleaned_context) + r"\s*" + re.escape(cleaned_span)
    match = re.search(search_pattern, cleaned_full, re.IGNORECASE)
    
    if match:
        # The span starts after the context
        start_index = match.start() + len(cleaned_context)
        # Adjust for potential whitespace between context and span in regex match
        original_segment = match.group(0)
        actual_span_start_in_match = original_segment.lower().rfind(cleaned_span.lower())
        final_start = match.start() + actual_span_start_in_match
        final_end = final_start + len(cleaned_span)
        return final_start, final_end

    # Fallback: Find first occurrence of span if context match fails
    start_index = cleaned_full.lower().find(cleaned_span.lower())
    if start_index != -1:
        return start_index, start_index + len(cleaned_span)
    
    return 0, 0

# -------------------------------------------------------------------------
# 6. WORKBOOK GENERATION
# -------------------------------------------------------------------------

def generate_workbook():
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_ws = wb.active
    wb.remove(default_ws)

    # ---------------------------------------------------------------------
    # Sheet 1: Note_Text
    # ---------------------------------------------------------------------
    ws_text = wb.create_sheet("Note_Text")
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # ---------------------------------------------------------------------
    # Sheet 2: Note_Index
    # ---------------------------------------------------------------------
    ws_index = wb.create_sheet("Note_Index")
    headers_index = ["NOTE_ID", "SOURCE_FILE", "PROCEDURE_DATE"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers_index)
    
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE]
    for key in PROCEDURE_FLAGS:
        row_data.append(PROCEDURE_FLAGS[key])
    ws_index.append(row_data)

    # ---------------------------------------------------------------------
    # Sheet 3: Span_Annotations (Empty offsets)
    # ---------------------------------------------------------------------
    ws_spans = wb.create_sheet("Span_Annotations")
    headers_spans = ["NOTE_ID", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "START_CHAR", "END_CHAR", "EVENT_ID"]
    ws_spans.append(headers_spans)
    
    for span in SPANS:
        # (text, label, norm, context, evt)
        ws_spans.append([NOTE_ID, span[0], span[1], span[2], span[3], "", "", span[4]])

    # ---------------------------------------------------------------------
    # Sheet 4: Span_Hydrated (Calculated offsets)
    # ---------------------------------------------------------------------
    ws_hydrated = wb.create_sheet("Span_Hydrated")
    ws_hydrated.append(headers_spans)
    
    for span in SPANS:
        s_text, s_label, s_norm, s_context, s_evt = span
        start, end = hydrate_span(NOTE_TEXT, s_text, s_context)
        ws_hydrated.append([NOTE_ID, s_text, s_label, s_norm, s_context, start, end, s_evt])

    # ---------------------------------------------------------------------
    # Sheet 5: Event_Log (Flattened)
    # ---------------------------------------------------------------------
    ws_event_log = wb.create_sheet("Event_Log")
    headers_log = ["NOTE_ID", "EVENT_ID", "EVENT_NAME", "PROCEDURE_TYPE", "ACTION", "DEVICES", "ANATOMY", "OUTCOMES"]
    ws_event_log.append(headers_log)
    
    for evt in EVENTS:
        ws_event_log.append([
            NOTE_ID,
            evt["event_id"],
            evt["name"],
            evt["procedure_type"],
            evt["action"],
            ", ".join(evt["devices"]),
            ", ".join(evt["anatomy"]),
            ", ".join(evt["outcomes"])
        ])

    # ---------------------------------------------------------------------
    # Sheet 6: V3_Procedure_Events (Nested JSON fields)
    # ---------------------------------------------------------------------
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    headers_v3 = ["note_id", "event_id", "event_type", "procedure_type", "action", "devices", "anatomy", "outcomes"]
    ws_v3.append(headers_v3)
    
    for evt in EVENTS:
        ws_v3.append([
            NOTE_ID,
            evt["event_id"],
            "procedure_step", # generic type
            evt["procedure_type"],
            evt["action"],
            json.dumps(evt["devices"]),
            json.dumps(evt["anatomy"]),
            json.dumps(evt["outcomes"])
        ])

    # ---------------------------------------------------------------------
    # Sheet 7: V3_Registry_JSON
    # ---------------------------------------------------------------------
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["note_id", "json_payload"])
    
    registry_data = {
        "note_id": NOTE_ID,
        "procedure_date": PROCEDURE_DATE,
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS
    }
    
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Workbook generated at: {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()