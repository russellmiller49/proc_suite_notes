import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# -------------------------------------------------------------------------
# 1. CONSTANTS & CONFIGURATION
# -------------------------------------------------------------------------
NOTE_ID = "note_182"
SOURCE_FILE = "note_182.txt"
PROCEDURE_DATE = ""  # Not found in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_182 SOURCE_FILE: note_182.txt PRE-PROCEDURE DIAGNISOS: RIGHT UPPER LOBE PULMONARY NODULE
POST- PROCEDURE DIAGNISOS: RIGHT UPPER LOBE PULMONARY NODULE
PROCEDURE PERFORMED:  
Flexible bronchoscopy with electromagnetic navigation under flouroscopic and EBUS guidance with transbronchial needle aspiration, Transbronchial biopsy and bronchioalveolar brush.
CPT 31654 Bronchoscope with Endobronchial Ultrasound guidance for peripheral lesion
CPT 31629 Flexible bronchoscopy with fluoroscopic trans-bronchial needle aspiration
CPT 31628 Bronchoscopy, rigid or flexible, including fluoroscopic guidance, when performed;
with transbronchial lung biopsy(s), single lobe
CPT 31623 Bronchoscopy, rigid or flexible, including fluoroscopic guidance, when performed;
with brushing or protected brushings
CPT +31624 Bronchoscopy, rigid or flexible, including fluoroscopic guidance, when performed;
with bronchial alveolar lavage
CPT +31627 Bronchoscopy with computer assisted image guided navigation

INDICATIONS FOR EXAMINATION:   Left upper lobe lung nodule            
MEDICATIONS:    GA
Pre-operative diagnosis: Right upper lobe nodule
Post-operative diagnosis: SAA
Anesthesia: General per Anesthesia

PROCEDURE: History and physical has been performed.
The risks and benefits of the procedure were discussed with the patient.
All questions were answered and informed consent was obtained. Patient identification and proposed procedure were verified by the physician, anesthesia team, nurse, and pulmonary team.
Time out was performed at 0819. The bronchoscope was introduced through the laryngeal mask airway and into the trachea.
The trachea is of normal caliber. The carina is sharp.
The tracheobronchial tree of bilateral lungs was examined to at least the first subsegmental level.
The right upper lobe was noted to have an anomalous inferior takeoff in addition to the anterior, apical, and posterior segments.
There were no endobronchial lesions and no excessive secretions.
 The Covidien superDimension probe was then advanced and registration was performed.
The right upper lobe nodule was accessed by the anomalous segment (see below), located using electromagnetic navigation and the site confirmed with radial ultrasound.
Transbronchial needle aspiration was performed under flouroscopic guidance and adequate tissue confirmed by ROSE.
Multiple insturments to include transbronchoscopic needles, forceps (tranbronchial lung biopsy) and brushes were utilized.
Five needle biopsies three forceps biopsies, and one needle brush biopsy were obtained in total.
Finally, inspection bronchoscopy was performed and hemostasis was confirmed. The procedure was accomplished without difficulty.
The patient tolerated the procedure well without immediate complications. EBL 5cc.
SPECIMEN SENT TO THE LABORATORY:
--5 needle biopsies of the right middle lobe
--1 brush biopsies of the  right middle lobe
--3 forceps biopsies

The patient tolerated the procedure well without complications.
RECOMMENDATIONS:
-D/C From CEC when discharge criteria are met
-Patient will be notified when biopsy results return"""

# -------------------------------------------------------------------------
# 2. PROCEDURE FLAGS
# -------------------------------------------------------------------------
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 1,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 1,
    "linear_ebus": 0,
    "radial_ebus": 1,
    "navigational_bronchoscopy": 1,
    "transbronchial_biopsy": 1,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# -------------------------------------------------------------------------
# 3. SPANS
# Format: (span_text, label, normalized_value, context_prefix, event_id)
# -------------------------------------------------------------------------
SPANS = [
    ("Flexible bronchoscopy", "PROC_METHOD", "Flexible Bronchoscopy", "PROCEDURE PERFORMED: \n", "evt_01"),
    ("electromagnetic navigation", "PROC_METHOD", "Navigational Bronchoscopy", "Flexible bronchoscopy with ", "evt_01"),
    ("transbronchial needle aspiration", "PROC_METHOD", "TBNA", "EBUS guidance with ", "evt_02"),
    ("Transbronchial biopsy", "PROC_METHOD", "Transbronchial Biopsy", "needle aspiration, ", "evt_03"),
    ("bronchioalveolar brush", "PROC_METHOD", "Bronchial Brushing", "Transbronchial biopsy and ", "evt_04"),
    ("Right upper lobe", "ANAT_LUNG_LOC", "Right Upper Lobe", "PRE-PROCEDURE DIAGNISOS: ", "evt_01"),
    ("Covidien superDimension probe", "DEV_INSTRUMENT", "SuperDimension Probe", "The ", "evt_01"),
    ("radial ultrasound", "PROC_METHOD", "Radial EBUS", "confirmed with ", "evt_01"),
    ("Transbronchial needle aspiration", "PROC_METHOD", "TBNA", "confirmed with radial ultrasound.\n", "evt_02"),
    ("ROSE", "OBS_ROSE", "Positive", "adequate tissue confirmed by ", "evt_02"),
    ("forceps", "DEV_INSTRUMENT", "Forceps", "transbronchoscopic needles, ", "evt_03"),
    ("brushes", "DEV_INSTRUMENT", "Cytology Brush", "biopsy) and ", "evt_04"),
    ("Five", "MEAS_COUNT", "5", "", "evt_02"),
    ("three", "MEAS_COUNT", "3", "Five needle biopsies ", "evt_03"),
    ("one", "MEAS_COUNT", "1", "forceps biopsies, and ", "evt_04"),
    ("tolerated the procedure well", "OUTCOME_SYMPTOMS", "Tolerated Well", "The patient ", "evt_05"),
    ("without immediate complications", "OUTCOME_COMPLICATION", "None", "tolerated the procedure well ", "evt_05"),
]

# -------------------------------------------------------------------------
# 4. EVENTS (Aggregated Data for V3 Sheet)
# -------------------------------------------------------------------------
EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Bronchoscopy",
        "action": "Navigation/Localization",
        "method": ["Navigational Bronchoscopy", "Radial EBUS"],
        "anatomy": "Right Upper Lobe",
        "devices": ["Covidien superDimension probe"],
        "notes": "Localization of RUL nodule"
    },
    {
        "event_id": "evt_02",
        "procedure_type": "Bronchoscopy",
        "action": "Biopsy",
        "method": ["TBNA"],
        "anatomy": "Right Upper Lobe",
        "devices": ["Transbronchial Needle"],
        "notes": "5 passes, ROSE confirmed",
        "specimens": "5 needle biopsies"
    },
    {
        "event_id": "evt_03",
        "procedure_type": "Bronchoscopy",
        "action": "Biopsy",
        "method": ["Transbronchial Biopsy"],
        "anatomy": "Right Upper Lobe",
        "devices": ["Forceps"],
        "notes": "3 biopsies",
        "specimens": "3 forceps biopsies"
    },
    {
        "event_id": "evt_04",
        "procedure_type": "Bronchoscopy",
        "action": "Biopsy",
        "method": ["Bronchial Brushing"],
        "anatomy": "Right Upper Lobe",
        "devices": ["Cytology Brush"],
        "notes": "1 biopsy",
        "specimens": "1 brush biopsy"
    },
    {
        "event_id": "evt_05",
        "procedure_type": "Outcome",
        "action": "Assessment",
        "method": [],
        "anatomy": "",
        "devices": [],
        "notes": "Tolerated well, no complications"
    }
]

# -------------------------------------------------------------------------
# 5. HELPER FUNCTIONS
# -------------------------------------------------------------------------
def clean_text(text):
    return text.strip().replace('\r', '')

def hydrate_span(text, span_text, context_prefix):
    """
    Finds the exact start/end offsets of span_text in text, 
    disambiguated by context_prefix.
    """
    cleaned_text = clean_text(text)
    cleaned_span = clean_text(span_text)
    cleaned_context = clean_text(context_prefix)
    
    # 1. Try finding context + span
    search_phrase = cleaned_context + cleaned_span
    start_index = cleaned_text.find(search_phrase)
    
    if start_index != -1:
        # found context+span. The actual span starts after the context.
        actual_start = start_index + len(cleaned_context)
        actual_end = actual_start + len(cleaned_span)
        return actual_start, actual_end, cleaned_text[actual_start:actual_end]
    
    # 2. Fallback: Find span alone (warn if ambiguous, but here just take first/best guess or none)
    # This is less precise. For this generator, we assume context is sufficient.
    start_index = cleaned_text.find(cleaned_span)
    if start_index != -1:
        return start_index, start_index + len(cleaned_span), cleaned_text[start_index:start_index+len(cleaned_span)]
    
    return None, None, None

# -------------------------------------------------------------------------
# 6. WORKBOOK GENERATOR
# -------------------------------------------------------------------------
def generate_workbook():
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
        
    # --- Sheet 1: Note_Text ---
    ws_text = wb.create_sheet("Note_Text")
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # --- Sheet 2: Note_Index ---
    ws_index = wb.create_sheet("Note_Index")
    headers_index = ["NOTE_ID", "procedure_date"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers_index)
    row_data = [NOTE_ID, PROCEDURE_DATE] + [PROCEDURE_FLAGS[k] for k in PROCEDURE_FLAGS]
    ws_index.append(row_data)
    
    # --- Sheet 3: Span_Annotations (Blank offsets) ---
    ws_anno = wb.create_sheet("Span_Annotations")
    headers_anno = ["NOTE_ID", "start_char", "end_char", "text", "label", "normalized_value", "context_prefix", "event_id"]
    ws_anno.append(headers_anno)
    for span in SPANS:
        # span = (span_text, label, normalized_value, context_prefix, event_id)
        ws_anno.append([NOTE_ID, "", "", span[0], span[1], span[2], span[3], span[4]])
        
    # --- Sheet 4: Span_Hydrated (Calculated offsets) ---
    ws_hydrated = wb.create_sheet("Span_Hydrated")
    ws_hydrated.append(headers_anno)
    for span in SPANS:
        s_start, s_end, s_text = hydrate_span(NOTE_TEXT, span[0], span[3])
        if s_start is None:
            print(f"Warning: Could not hydrate span '{span[0]}' with context '{span[3]}'")
            s_start, s_end = "", ""
        ws_hydrated.append([NOTE_ID, s_start, s_end, span[0], span[1], span[2], span[3], span[4]])
        
    # --- Sheet 5: Event_Log ---
    ws_event_log = wb.create_sheet("Event_Log")
    headers_event_log = ["NOTE_ID", "event_id", "procedure_type", "action", "method", "anatomy", "devices", "notes"]
    ws_event_log.append(headers_event_log)
    for evt in EVENTS:
        ws_event_log.append([
            NOTE_ID,
            evt["event_id"],
            evt["procedure_type"],
            evt["action"],
            ", ".join(evt["method"]),
            evt["anatomy"],
            ", ".join(evt["devices"]),
            evt.get("notes", "")
        ])

    # --- Sheet 6: V3_Procedure_Events ---
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    headers_v3 = ["NOTE_ID", "event_id", "procedure_type", "action", "method", "anatomy", "devices", "comments"]
    ws_v3.append(headers_v3)
    for evt in EVENTS:
        # V3 often requires JSON strings for lists
        ws_v3.append([
            NOTE_ID,
            evt["event_id"],
            evt["procedure_type"],
            evt["action"],
            json.dumps(evt["method"]),
            evt["anatomy"],
            json.dumps(evt["devices"]),
            evt.get("notes", "")
        ])
        
    # --- Sheet 7: V3_Registry_JSON ---
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["NOTE_ID", "json_object"])
    
    registry_data = {
        "note_id": NOTE_ID,
        "procedure_date": PROCEDURE_DATE,
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS
    }
    
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])
    
    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()