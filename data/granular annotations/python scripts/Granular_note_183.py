import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import re
import json
import os
import datetime

# 1. Imports & Constants
NOTE_ID = "note_183"
SOURCE_FILE = "note_183.txt"
PROCEDURE_DATE = ""
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_183 SOURCE_FILE: note_183.txt PRE-PROCEDURE DIAGNISOS: LEFT UPPER LOBE PULMONARY NODULE
POST- PROCEDURE DIAGNISOS: LEFT UPPER LOBE PULMONARY NODULE
PROCEDURE PERFORMED:  
Flexible bronchoscopy with electromagnetic navigation under flouroscopic and EBUS guidance with transbronchial needle aspiration, Transbronchial biopsy and bronchioalveolar lavage.
CPT 31654 Bronchoscope with Endobronchial Ultrasound guidance for peripheral lesion
CPT 31628 Bronchoscopy, rigid or flexible, including fluoroscopic guidance, when performed;
with transbronchial lung biopsy(s), single lobe
CPT +31624 Bronchoscopy, rigid or flexible, including fluoroscopic guidance, when performed;
with bronchial alveolar lavage
CPT +31627 Bronchoscopy with computer assisted image guided navigation
INDICATIONS FOR EXAMINATION:   Left upper lobe lung nodule            
MEDICATIONS:    GA
FINDINGS: Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the endotracheal tube and advanced to the tracheobronchial tree.
The trachea was of normal caliber. The carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level. Bronchial mucosa and anatomy were normal;
there are no endobronchial lesions. We then removed the diagnostic Q190 bronchoscope and the super-dimension navigational catheter was inserted through the T190 therapeutic bronchoscope and advanced into the airway.
Using navigational map we attempted to advance the 180 degree edge catheter into the proximity of the lesion within apico-posterior branch of left upper lobe.
Radial probe was used to attempt to confirm presence within the lesion.
Although we were able to navigate directly to the lesion with navigation the radial probe view was suboptimal.
Biopsy was performed initially with triple needle brush and TBNA needle.
ROSE did not reveal evidence to support that we were within the lesion.
Multiple attempts were made to manipulate the catheter and biopsies were then performed with a variety of instruments to include peripheral needle, and forceps, brush under fluoroscopic visualization.
The specimens reviewed on-site remained suboptimal.  Multiple forceps biopsies were performed within the location of the lesion and placed in cell-block.
After which a mini-BAL was then performed through the super-D catheter.
We then removed the therapeutic bronchoscope with super-D catheter and reinserted the diagnostic scope at which point repeat airway inspection was then performed and once we were satisfied that no bleeding occurred, the bronchoscope was removed and the procedure completed.
ESTIMATED BLOOD LOSS:   None 
COMPLICATIONS:                 None

IMPRESSION:  
- S/P bronchoscopy with biopsy and lavage.
- Suboptimal navigational localization 
RECOMMENDATIONS
- Transfer to post-procedural unit
- Post-procedure CXR
- D/C home once criteria met
- Await pathology"""

# 2. Configuration (Procedure Flags)
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 1,
    "bronchial_wash": 0,
    "brushings": 1,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 1,
    "linear_ebus": 0,
    "radial_ebus": 1,
    "navigational_bronchoscopy": 1,
    "transbronchial_biopsy": 1,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 3. Data Definition (Spans)
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("Flexible bronchoscopy", "PROC_METHOD", "Flexible Bronchoscopy", "PERFORMED:  \n", "evt_01"),
    ("electromagnetic navigation", "PROC_METHOD", "Electromagnetic Navigation", "bronchoscopy with ", "evt_01"),
    ("EBUS guidance", "PROC_METHOD", "Radial EBUS", "flouroscopic and ", "evt_01"),
    ("transbronchial needle aspiration", "PROC_METHOD", "TBNA", "guidance with ", "evt_02"),
    ("Transbronchial biopsy", "PROC_METHOD", "Transbronchial Biopsy", "aspiration, ", "evt_02"),
    ("bronchioalveolar lavage", "PROC_METHOD", "BAL", "biopsy and ", "evt_03"),
    ("Left upper lobe lung nodule", "ANAT_LUNG_LOC", "LUL Nodule", "EXAMINATION:   ", "evt_01"),
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Bronchoscope - Q190", "tree, the ", "evt_01"),
    ("super-dimension navigational catheter", "DEV_CATHETER", "SuperDimension Catheter", "bronchoscope and the ", "evt_01"),
    ("T190 therapeutic bronchoscope", "DEV_INSTRUMENT", "Bronchoscope - T190", "through the ", "evt_01"),
    ("180 degree edge catheter", "DEV_CATHETER", "Edge Catheter 180", "advance the ", "evt_01"),
    ("apico-posterior branch of left upper lobe", "ANAT_AIRWAY", "LUL Apico-posterior Segment", "lesion within ", "evt_01"),
    ("Radial probe", "DEV_INSTRUMENT", "Radial EBUS Probe", "left upper lobe.\n", "evt_01"),
    ("triple needle brush", "DEV_INSTRUMENT", "Triple Needle Brush", "initially with ", "evt_02"),
    ("TBNA needle", "DEV_NEEDLE", "TBNA Needle", "brush and ", "evt_02"),
    ("peripheral needle", "DEV_NEEDLE", "Peripheral Needle", "to include ", "evt_02"),
    ("forceps", "DEV_INSTRUMENT", "Biopsy Forceps", "needle, and ", "evt_02"),
    ("brush", "DEV_INSTRUMENT", "Cytology Brush", "forceps, ", "evt_02"),
    ("mini-BAL", "PROC_METHOD", "Mini-BAL", "After which a ", "evt_03"),
    ("None", "OUTCOME_COMPLICATION", "None", "COMPLICATIONS:                 ", "evt_04")
]

# 4. Event Definitions
EVENTS = [
    {
        "event_id": "evt_01",
        "name": "Navigation and Inspection",
        "procedure_type": "Bronchoscopy",
        "method": ["Flexible Bronchoscopy", "Electromagnetic Navigation", "Radial EBUS"],
        "anatomy": ["LUL", "LUL Apico-posterior Segment"],
        "devices": ["Bronchoscope - Q190", "Bronchoscope - T190", "SuperDimension Catheter", "Edge Catheter 180", "Radial EBUS Probe"],
        "outcomes": []
    },
    {
        "event_id": "evt_02",
        "name": "Biopsy",
        "procedure_type": "Bronchoscopy",
        "method": ["TBNA", "Transbronchial Biopsy"],
        "anatomy": ["LUL"],
        "devices": ["Triple Needle Brush", "TBNA Needle", "Peripheral Needle", "Biopsy Forceps", "Cytology Brush"],
        "outcomes": []
    },
    {
        "event_id": "evt_03",
        "name": "Lavage",
        "procedure_type": "Bronchoscopy",
        "method": ["BAL", "Mini-BAL"],
        "anatomy": ["LUL"],
        "devices": [],
        "outcomes": []
    },
    {
        "event_id": "evt_04",
        "name": "Outcome",
        "procedure_type": "Bronchoscopy",
        "method": [],
        "anatomy": [],
        "devices": [],
        "outcomes": ["No Complications"]
    }
]

# 5. Helper Functions
def clean_text(text):
    if not text:
        return ""
    return text.strip().replace('\r', '')

def hydrate_span(text, span_text, context_prefix):
    cleaned_text = clean_text(text)
    cleaned_span = clean_text(span_text)
    cleaned_context = clean_text(context_prefix)
    
    if not cleaned_span:
        return None, None, None

    # Try finding with context first
    search_phrase = cleaned_context + cleaned_span
    start_index = cleaned_text.find(search_phrase)
    
    if start_index != -1:
        # Found with context, adjust start to point to span
        start_index += len(cleaned_context)
        end_index = start_index + len(cleaned_span)
        # Verify
        extracted = cleaned_text[start_index:end_index]
        if extracted == cleaned_span:
            return start_index, end_index, extracted

    # Fallback: simple find (riskier for duplicates)
    start_index = cleaned_text.find(cleaned_span)
    if start_index != -1:
        end_index = start_index + len(cleaned_span)
        return start_index, end_index, cleaned_text[start_index:end_index]
    
    return None, None, None

# 6. Workbook Generation Function
def generate_workbook():
    # Create workbook and default sheets
    wb = openpyxl.Workbook()
    
    # --- Sheet 1: Note_Text ---
    ws_text = wb.active
    ws_text.title = "Note_Text"
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # --- Sheet 2: Note_Index ---
    ws_index = wb.create_sheet("Note_Index")
    headers = ["NOTE_ID", "PROCEDURE_DATE"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers)
    row_data = [NOTE_ID, PROCEDURE_DATE] + [PROCEDURE_FLAGS[k] for k in PROCEDURE_FLAGS]
    ws_index.append(row_data)
    
    # --- Sheet 3: Span_Annotations ---
    ws_spans = wb.create_sheet("Span_Annotations")
    ws_spans.append(["NOTE_ID", "START_CHAR", "END_CHAR", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "EVENT_ID"])
    
    for span in SPANS:
        # unpack tuple: (span_text, label, normalized_value, context_prefix, event_id)
        s_text, label, norm_val, context, evt_id = span
        # Write initial row with empty start/end (to be filled by hydration or manual)
        ws_spans.append([NOTE_ID, "", "", s_text, label, norm_val, context, evt_id])
        
    # --- Sheet 4: Span_Hydrated ---
    ws_hydrated = wb.create_sheet("Span_Hydrated")
    ws_hydrated.append(["NOTE_ID", "START_CHAR", "END_CHAR", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "EVENT_ID", "EXTRACTED_VERIFICATION"])
    
    for span in SPANS:
        s_text, label, norm_val, context, evt_id = span
        start, end, extracted = hydrate_span(NOTE_TEXT, s_text, context)
        ws_hydrated.append([NOTE_ID, start, end, s_text, label, norm_val, context, evt_id, extracted])

    # --- Sheet 5: Event_Log ---
    ws_event_log = wb.create_sheet("Event_Log")
    ws_event_log.append(["NOTE_ID", "EVENT_ID", "CATEGORY", "VALUE"])
    
    for evt in EVENTS:
        eid = evt["event_id"]
        # Flatten dictionary lists into rows
        if evt["method"]:
            for m in evt["method"]: ws_event_log.append([NOTE_ID, eid, "Method", m])
        if evt["anatomy"]:
            for a in evt["anatomy"]: ws_event_log.append([NOTE_ID, eid, "Anatomy", a])
        if evt["devices"]:
            for d in evt["devices"]: ws_event_log.append([NOTE_ID, eid, "Device", d])
        if evt["outcomes"]:
            for o in evt["outcomes"]: ws_event_log.append([NOTE_ID, eid, "Outcome", o])

    # --- Sheet 6: V3_Procedure_Events ---
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    ws_v3.append(["NOTE_ID", "EVENT_ID", "PROCEDURE_TYPE", "METHOD", "ANATOMY", "DEVICES", "OUTCOMES"])
    
    for evt in EVENTS:
        ws_v3.append([
            NOTE_ID,
            evt["event_id"],
            evt["procedure_type"],
            json.dumps(evt["method"]),
            json.dumps(evt["anatomy"]),
            json.dumps(evt["devices"]),
            json.dumps(evt["outcomes"])
        ])

    # --- Sheet 7: V3_Registry_JSON ---
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["NOTE_ID", "JSON_OBJECT"])
    
    registry_obj = {
        "metadata": {
            "note_id": NOTE_ID,
            "source": SOURCE_FILE,
            "procedure_date": PROCEDURE_DATE
        },
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS
    }
    
    ws_json.append([NOTE_ID, json.dumps(registry_obj, indent=2)])

    # Formatting (Optional but nice)
    for ws in [ws_index, ws_spans, ws_hydrated, ws_event_log, ws_v3]:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

# 7. Execution Block
if __name__ == "__main__":
    generate_workbook()