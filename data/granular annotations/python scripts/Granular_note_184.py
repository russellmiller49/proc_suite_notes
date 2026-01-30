import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import re
import json
import os
import datetime

# -------------------------------------------------------------------------
# 1. Constants & Input Data
# -------------------------------------------------------------------------
NOTE_ID = "note_184"
SOURCE_FILE = "note_184.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_184 SOURCE_FILE: note_184.txt Indications: Diagnosis and staging of presumed lung cancer
Medications: General Anesthesia,
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention.
Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway was in good position. The vocal cords appeared normal. The subglottic space was normal.
The trachea was of normal caliber. The carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level. Bronchial mucosa and anatomy were normal;
there are no endobronchial lesions, and no secretions. The video bronchoscope was then removed and the UC180F convex probe EBUS bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
A systematic hilar and mediastinal lymph node survey was carried out.
Sampling criteria (5mm short axis diameter) were met in station 4R, 4L, Station 5 and 11L lymph nodes.
Sampling by transbronchial needle aspiration was performed with the Olympus EBUSTBNA 22 gauge needle beginning with the 4R Lymph node, followed by the 4L.
The 4R yielded benign lymphocytes on ROSE evaluation while the 4L was non-diagnostic.
Given that the station 11L would only stage the patient as N1 disease and the station 5 lymph nodes, which were highly suspicious for malignancy based on FDG avidity and ultrasound characteristics, after careful consideration we decided to attempt biopsy via the trans-vascular route thorough the pulmonary artery a total of 7 passes were performed, Five biopsies were performed with the Boston Scientific 25 gauge EBUS needle and an additional two with Olympus EBUSTBNA 22 gauge needle.
ROSE was consistent with poorly differentiated carcinoma. All samples were sent for routine cytology.
Following completion of EBUS bronchoscopy, the Q190 video bronchoscope was then re-inserted and after suctioning blood and secretions there was no evidence of active bleeding and the bronchoscope was subsequently removed.
Complications: No immediate complications
Estimated Blood Loss: 10 cc.

Post Procedure Diagnosis:
- Technically successful flexible bronchoscopy with endobronchial ultrasound-guided biopsies.
- The patient has remained stable and has been transferred in good condition to the post-surgical monitoring unit.
- Will await final pathology results"""

# -------------------------------------------------------------------------
# 2. Configuration (Procedure Flags)
# -------------------------------------------------------------------------
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# -------------------------------------------------------------------------
# 3. Data Definition (Spans)
# -------------------------------------------------------------------------
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Olympus Q190", "tree, the ", "evt_01"),
    ("tracheobronchial tree", "ANAT_AIRWAY", "Tracheobronchial Tree", "advanced to the ", "evt_01"),
    ("Bronchial mucosa and anatomy were normal", "OBS_LESION", "Normal Anatomy", "level. ", "evt_01"),
    
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "Olympus UC180F", "removed and the ", "evt_02"),
    ("hilar and mediastinal lymph node", "ANAT_LN_STATION", "Hilar/Mediastinal", "systematic ", "evt_02"),
    
    ("transbronchial needle aspiration", "PROC_METHOD", "Linear EBUS-TBNA", "Sampling by ", "evt_03"),
    ("Olympus EBUSTBNA 22 gauge needle", "DEV_NEEDLE", "Olympus EBUS-TBNA 22G", "with the ", "evt_03"),
    ("4R Lymph node", "ANAT_LN_STATION", "4R", "beginning with the ", "evt_03"),
    ("benign lymphocytes", "OBS_ROSE", "Benign Lymphocytes", "4R yielded ", "evt_03"),
    
    ("4L", "ANAT_LN_STATION", "4L", "followed by the ", "evt_04"),
    ("non-diagnostic", "OBS_ROSE", "Non-diagnostic", "4L was ", "evt_04"),
    
    ("station 5 lymph nodes", "ANAT_LN_STATION", "5", "Given that the ", "evt_05"),
    ("trans-vascular route thorough the pulmonary artery", "PROC_METHOD", "Trans-vascular TBNA", "biopsy via the ", "evt_05"),
    ("Boston Scientific 25 gauge EBUS needle", "DEV_NEEDLE", "Boston Scientific 25G", "performed with the ", "evt_05"),
    ("Olympus EBUSTBNA 22 gauge needle", "DEV_NEEDLE", "Olympus EBUS-TBNA 22G", "two with ", "evt_05"),
    ("poorly differentiated carcinoma", "OBS_ROSE", "Poorly Differentiated Carcinoma", "consistent with ", "evt_05"),
    
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications: ", "evt_06")
]

# -------------------------------------------------------------------------
# 4. Event Definitions
# -------------------------------------------------------------------------
EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Bronchoscopy",
        "method": "Diagnostic Bronchoscopy",
        "anatomy": "Tracheobronchial Tree",
        "devices": ["Olympus Q190"],
        "outcomes": ["Normal inspection"]
    },
    {
        "event_id": "evt_02",
        "procedure_type": "Bronchoscopy",
        "method": "Linear EBUS",
        "anatomy": "Hilar and Mediastinal Lymph Nodes",
        "devices": ["Olympus UC180F"],
        "outcomes": []
    },
    {
        "event_id": "evt_03",
        "procedure_type": "Bronchoscopy",
        "method": "Linear EBUS-TBNA",
        "anatomy": "4R",
        "devices": ["Olympus EBUS-TBNA 22G"],
        "outcomes": ["ROSE: Benign lymphocytes"]
    },
    {
        "event_id": "evt_04",
        "procedure_type": "Bronchoscopy",
        "method": "Linear EBUS-TBNA",
        "anatomy": "4L",
        "devices": ["Olympus EBUS-TBNA 22G"],
        "outcomes": ["ROSE: Non-diagnostic"]
    },
    {
        "event_id": "evt_05",
        "procedure_type": "Bronchoscopy",
        "method": "Linear EBUS-TBNA",
        "anatomy": "5",
        "devices": ["Boston Scientific 25G", "Olympus EBUS-TBNA 22G"],
        "outcomes": ["ROSE: Poorly differentiated carcinoma"]
    },
    {
        "event_id": "evt_06",
        "procedure_type": "Global",
        "method": "Outcome",
        "anatomy": "NA",
        "devices": [],
        "outcomes": ["No immediate complications"]
    }
]

# -------------------------------------------------------------------------
# 5. Helper Functions
# -------------------------------------------------------------------------
def clean_text(text):
    return text.replace('\r', '').strip()

def hydrate_span(text, span_text, context_prefix):
    """
    Finds start/end offsets of span_text within text.
    Uses context_prefix to disambiguate if needed.
    """
    # Normalize inputs for searching
    text_normalized = text.replace('\r', '')
    span_normalized = span_text.replace('\r', '')
    context_normalized = context_prefix.replace('\r', '')

    # Try exact match with context first
    search_phrase = context_normalized + span_normalized
    if search_phrase in text_normalized:
        start_index = text_normalized.find(search_phrase) + len(context_normalized)
        end_index = start_index + len(span_normalized)
        return start_index, end_index

    # Fallback: find just the span (first occurrence)
    if span_normalized in text_normalized:
        start_index = text_normalized.find(span_normalized)
        end_index = start_index + len(span_normalized)
        return start_index, end_index
    
    return "", ""

# -------------------------------------------------------------------------
# 6. Workbook Generation Function
# -------------------------------------------------------------------------
def generate_workbook():
    # 1. Load or Create Template
    if not os.path.exists(TEMPLATE_PATH):
        # Create a dummy workbook if template missing (for standalone testing)
        wb = openpyxl.Workbook()
        wb.create_sheet("Note_Text")
        wb.create_sheet("Note_Index")
        wb.create_sheet("Span_Annotations")
        wb.create_sheet("Span_Hydrated")
        wb.create_sheet("Event_Log")
        wb.create_sheet("V3_Procedure_Events")
        wb.create_sheet("V3_Registry_JSON")
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # 2. Sheet: Note_Text
    ws_text = wb["Note_Text"]
    # If empty, add headers
    if ws_text.max_row == 1 and ws_text.cell(1, 1).value is None:
        ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    
    # Append the note data
    # Check if note already exists to avoid duplicates in local testing
    found = False
    for row in ws_text.iter_rows(min_row=2, values_only=True):
        if row[0] == NOTE_ID:
            found = True
            break
    if not found:
        ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 3. Sheet: Note_Index (Metadata + Flags)
    ws_index = wb["Note_Index"]
    if ws_index.max_row == 1 and ws_index.cell(1, 1).value is None:
        headers = ["NOTE_ID", "PROCEDURE_DATE"] + list(PROCEDURE_FLAGS.keys())
        ws_index.append(headers)
    
    # Write Row
    row_data = [NOTE_ID, PROCEDURE_DATE] + list(PROCEDURE_FLAGS.values())
    ws_index.append(row_data)

    # 4. Sheet: Span_Annotations (Raw Spans)
    ws_anno = wb["Span_Annotations"]
    if ws_anno.max_row == 1 and ws_anno.cell(1, 1).value is None:
        ws_anno.append(["NOTE_ID", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "START_CHAR", "END_CHAR", "EVENT_ID"])
    
    for span in SPANS:
        s_text, s_label, s_norm, s_ctx, s_evt = span
        ws_anno.append([NOTE_ID, s_text, s_label, s_norm, s_ctx, "", "", s_evt])

    # 5. Sheet: Span_Hydrated (Calculated Offsets)
    ws_hydra = wb["Span_Hydrated"]
    if ws_hydra.max_row == 1 and ws_hydra.cell(1, 1).value is None:
        ws_hydra.append(["NOTE_ID", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "START_CHAR", "END_CHAR", "EVENT_ID"])
    
    for span in SPANS:
        s_text, s_label, s_norm, s_ctx, s_evt = span
        start, end = hydrate_span(NOTE_TEXT, s_text, s_ctx)
        ws_hydra.append([NOTE_ID, s_text, s_label, s_norm, s_ctx, start, end, s_evt])

    # 6. Sheet: Event_Log (Flattened)
    ws_event = wb["Event_Log"]
    if ws_event.max_row == 1 and ws_event.cell(1, 1).value is None:
        ws_event.append(["NOTE_ID", "EVENT_ID", "METHOD", "ANATOMY", "DEVICES", "OUTCOMES"])
    
    for evt in EVENTS:
        ws_event.append([
            NOTE_ID,
            evt["event_id"],
            evt["method"],
            evt["anatomy"],
            ", ".join(evt["devices"]),
            ", ".join(evt["outcomes"])
        ])

    # 7. Sheet: V3_Procedure_Events
    ws_v3 = wb["V3_Procedure_Events"]
    if ws_v3.max_row == 1 and ws_v3.cell(1, 1).value is None:
        ws_v3.append(["NOTE_ID", "EVENT_ID", "PROCEDURE_TYPE", "METHOD", "ANATOMY", "DEVICES_JSON", "OUTCOMES_JSON", "FULL_EVENT_JSON"])
    
    for evt in EVENTS:
        # Construct simplified registry JSON
        reg_json = {
            "event_id": evt["event_id"],
            "procedure_type": evt["procedure_type"],
            "method": evt["method"],
            "anatomy": evt["anatomy"],
            "devices": evt["devices"],
            "outcomes": evt["outcomes"]
        }
        ws_v3.append([
            NOTE_ID,
            evt["event_id"],
            evt["procedure_type"],
            evt["method"],
            evt["anatomy"],
            json.dumps(evt["devices"]),
            json.dumps(evt["outcomes"]),
            json.dumps(reg_json)
        ])

    # 8. Sheet: V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    if ws_json.max_row == 1 and ws_json.cell(1, 1).value is None:
        ws_json.append(["NOTE_ID", "FULL_REGISTRY_JSON_DUMP"])
    
    # Construct total JSON
    total_json = {
        "note_id": NOTE_ID,
        "procedure_date": PROCEDURE_DATE,
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS
    }
    ws_json.append([NOTE_ID, json.dumps(total_json, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

# -------------------------------------------------------------------------
# 7. Execution Block
# -------------------------------------------------------------------------
if __name__ == "__main__":
    generate_workbook()