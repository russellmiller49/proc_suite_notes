import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# ==========================================
# 1. Imports & Constants
# ==========================================

NOTE_ID = "note_185"
SOURCE_FILE = "note_185.txt"
PROCEDURE_DATE = ""
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_185 SOURCE_FILE: note_185.txt Procedure Name: EBUS Bronchoscopy
Indications: staging of bilateral lung cancer (metastatic vs synchronous primary)
Medications: General Anesthesia
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention.
 Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway was in good position. The vocal cords appeared normal. The subglottic space was normal.
The trachea was of normal caliber. The carina was sharp.
The tracheobronchial tree was examined to at least the first sub-segmental level. Bronchial mucosa and anatomy were normal;
there are no endobronchial lesions, and no secretions. The video bronchoscope was then removed and the UC180F convex probe EBUS bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
A systematic hilar and mediastinal lymph node survey was carried out.
Sampling criteria (5mm short axis diameter) were met in station 4LR,7, and 4R lymph nodes.
Sampling by transbronchial needle aspiration was performed beginning with the 4L Lymph node followed by 7, and 4R lymph nodes using an Olympus EBUSTBNA 22 gauge needle.
The 4L lymph node was extremely difficult to access due to angle of left mainstem and after one biopsy no further attempts performed.
Further details regarding nodal size and number of samples are included in the EBUS procedural sheet which is available in EMR.
ROSE showed non-diagnostic tissue in the 4L and 4R lymph nodes and benign lymphocytes in the station 7 lymph node.
All samples were sent for routine cytology. The Q190 video bronchoscope was then re-inserted and after suctioning blood and secretions there was no evidence of active bleeding and the bronchoscope was subsequently removed.
Complications: No immediate complications
Estimated Blood Loss: 5 cc.

Post Procedure Recommendations:
- Transfer to PACU and home per protocol
- Will await final pathology results"""

# ==========================================
# 2. Configuration (Procedure Flags)
# ==========================================

PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# ==========================================
# 3. Data Definition (Spans)
# ==========================================

# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Olympus Q190", "tracheobronchial tree, the ", "evt_01"),
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "Olympus UC180F", "removed and the ", "evt_02"),
    ("transbronchial needle aspiration", "PROC_METHOD", "EBUS-TBNA", "Sampling by ", "evt_02"),
    ("4L", "ANAT_LN_STATION", "Station 4L", "beginning with the ", "evt_02"),
    ("7", "ANAT_LN_STATION", "Station 7", "followed by ", "evt_02"),
    ("4R", "ANAT_LN_STATION", "Station 4R", ", and ", "evt_02"),
    ("Olympus EBUSTBNA 22 gauge needle", "DEV_NEEDLE", "22G EBUS Needle", "using an ", "evt_02"),
    ("ROSE", "PROC_METHOD", "ROSE", "", "evt_02"),
    ("non-diagnostic", "OBS_ROSE", "Non-diagnostic", "ROSE showed ", "evt_02"),
    ("benign lymphocytes", "OBS_ROSE", "Benign", "lymph nodes and ", "evt_02"),
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications: ", "evt_global")
]

# ==========================================
# 4. Event Definitions
# ==========================================

EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Bronchoscopy",
        "action": "Inspection",
        "anatomy": ["Tracheobronchial Tree"],
        "devices": ["Olympus Q190"],
        "notes": "Initial airway inspection, normal anatomy."
    },
    {
        "event_id": "evt_02",
        "procedure_type": "EBUS-TBNA",
        "action": "Needle Aspiration",
        "anatomy": ["Station 4L", "Station 7", "Station 4R"],
        "devices": ["Olympus UC180F", "22G EBUS Needle"],
        "outcomes": {
            "ROSE": "Non-diagnostic (4L, 4R), Benign (7)"
        },
        "notes": "4L difficult access, 1 biopsy only. Others sampled per protocol."
    }
]

# ==========================================
# 5. Helper Functions
# ==========================================

def clean_text(text):
    return text.strip()

def hydrate_span(full_text, span_text, context_prefix=""):
    """
    Finds the start and end indices of span_text in full_text.
    Uses context_prefix to locate the specific occurrence if provided.
    """
    # Normalize spaces for search (optional, but helpful if line breaks differ)
    # keeping it simple for strict matching against provided text
    
    start_index = -1
    
    if context_prefix:
        # Construct regex to find prefix + span
        # Escape special characters in prefix and span
        escaped_prefix = re.escape(context_prefix)
        escaped_span = re.escape(span_text)
        pattern = f"({escaped_prefix})({escaped_span})"
        
        match = re.search(pattern, full_text)
        if match:
            # group 1 is prefix, group 2 is span
            # start of span is match.start() + length of prefix
            start_index = match.start(2)
            end_index = match.end(2)
            return start_index, end_index
    
    # Fallback or if no context provided: find first occurrence
    start_index = full_text.find(span_text)
    if start_index != -1:
        return start_index, start_index + len(span_text)
    
    return None, None

def create_dummy_template():
    """Creates a basic template if file is missing to allow script to run."""
    wb = openpyxl.Workbook()
    sheets = [
        "Note_Text", "Note_Index", "Span_Annotations", 
        "Span_Hydrated", "Event_Log", "V3_Procedure_Events", "V3_Registry_JSON"
    ]
    # Rename active sheet
    wb.active.title = sheets[0]
    for s in sheets[1:]:
        wb.create_sheet(s)
    wb.save(TEMPLATE_PATH)

# ==========================================
# 6. Workbook Generation Function
# ==========================================

def generate_workbook():
    # Check for template
    if not os.path.exists(TEMPLATE_PATH):
        print(f"Template {TEMPLATE_PATH} not found. Creating dummy.")
        create_dummy_template()

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    
    # --- Sheet 1: Note_Text ---
    ws_text = wb["Note_Text"]
    # Clear existing
    ws_text.delete_rows(2, ws_text.max_row)
    # Write header if empty (assuming template has headers, but safe to write)
    if ws_text["A1"].value is None:
        ws_text.append(["note_id", "source_file", "note_text"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # --- Sheet 2: Note_Index ---
    ws_index = wb["Note_Index"]
    ws_index.delete_rows(2, ws_index.max_row)
    
    # Prepare row: ID, Date, then 30 flags
    index_header = ["note_id", "procedure_date"] + list(PROCEDURE_FLAGS.keys())
    # Ensure header exists
    if ws_index["A1"].value is None:
        ws_index.append(index_header)
        
    row_data = [NOTE_ID, PROCEDURE_DATE] + list(PROCEDURE_FLAGS.values())
    ws_index.append(row_data)

    # --- Sheet 3: Span_Annotations ---
    ws_span = wb["Span_Annotations"]
    ws_span.delete_rows(2, ws_span.max_row)
    span_header = ["note_id", "span_text", "label", "normalized_value", "start_char", "end_char"]
    if ws_span["A1"].value is None:
        ws_span.append(span_header)

    # We will buffer hydrated data for Sheet 4
    hydrated_rows = []

    for span_tuple in SPANS:
        text_segment, label, norm, context, evt_id = span_tuple
        
        # Calculate offsets
        start, end = hydrate_span(NOTE_TEXT, text_segment, context)
        
        # Sheet 3 gets blanks for offsets (human labeling simulation)
        ws_span.append([NOTE_ID, text_segment, label, norm, "", ""])
        
        # Store for Sheet 4
        hydrated_rows.append([NOTE_ID, text_segment, label, norm, start, end, evt_id])

    # --- Sheet 4: Span_Hydrated ---
    ws_hydra = wb["Span_Hydrated"]
    ws_hydra.delete_rows(2, ws_hydra.max_row)
    hydra_header = ["note_id", "span_text", "label", "normalized_value", "start_char", "end_char", "event_id"]
    if ws_hydra["A1"].value is None:
        ws_hydra.append(hydra_header)
    
    for row in hydrated_rows:
        ws_hydra.append(row)

    # --- Sheet 5: Event_Log ---
    ws_event = wb["Event_Log"]
    ws_event.delete_rows(2, ws_event.max_row)
    event_header = ["note_id", "event_id", "procedure_type", "action", "anatomy", "devices", "outcomes"]
    if ws_event["A1"].value is None:
        ws_event.append(event_header)

    for evt in EVENTS:
        # Flatten lists for CSV-like cell
        anat_str = ", ".join(evt.get("anatomy", []))
        dev_str = ", ".join(evt.get("devices", []))
        out_str = json.dumps(evt.get("outcomes", {}))
        
        ws_event.append([
            NOTE_ID,
            evt["event_id"],
            evt.get("procedure_type", ""),
            evt.get("action", ""),
            anat_str,
            dev_str,
            out_str
        ])

    # --- Sheet 6: V3_Procedure_Events ---
    # Similar to Event_Log but strictly V3 schema if needed. 
    # For now, we replicate Event_Log structure or add specific JSON fields.
    ws_v3 = wb["V3_Procedure_Events"]
    ws_v3.delete_rows(2, ws_v3.max_row)
    if ws_v3["A1"].value is None:
        ws_v3.append(event_header) # Reuse header
    
    for evt in EVENTS:
        anat_str = ", ".join(evt.get("anatomy", []))
        dev_str = ", ".join(evt.get("devices", []))
        out_str = json.dumps(evt.get("outcomes", {}))
        ws_v3.append([
            NOTE_ID,
            evt["event_id"],
            evt.get("procedure_type", ""),
            evt.get("action", ""),
            anat_str,
            dev_str,
            out_str
        ])

    # --- Sheet 7: V3_Registry_JSON ---
    ws_json = wb["V3_Registry_JSON"]
    ws_json.delete_rows(2, ws_json.max_row)
    if ws_json["A1"].value is None:
        ws_json.append(["note_id", "json_object"])
    
    # Construct Full JSON
    registry_data = {
        "note_id": NOTE_ID,
        "procedure_date": PROCEDURE_DATE,
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS,
        "spans": [
            {
                "text": r[1],
                "label": r[2],
                "normalized": r[3],
                "start": r[4],
                "end": r[5],
                "event_id": r[6]
            }
            for r in hydrated_rows
        ]
    }
    
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Successfully generated {OUTPUT_PATH}")

# ==========================================
# 7. Execution Block
# ==========================================

if __name__ == "__main__":
    generate_workbook()