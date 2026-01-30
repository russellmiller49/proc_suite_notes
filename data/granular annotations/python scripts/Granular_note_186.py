import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import re
import json
import os
import datetime

# ==========================================
# 1. Constants & Configuration
# ==========================================
NOTE_ID = "note_186"
SOURCE_FILE = "note_186.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_186 SOURCE_FILE: note_186.txt Indications: Pulmonary nodule requiring diagnosis/staging.
Medications: Propofol infusion via anesthesia assistance  
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention. 
Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway was in good position. The vocal cords appeared normal. The subglottic space was normal.
The trachea was of normal caliber. The carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level without endobronchial lesions visualized.
Anatomy was normal with exception of fish mouth dynamic obstruction of the right middle lobe.
The bronchoscope was then removed and the P190 ultrathin video bronchoscope was inserted into the airway and advanced into the apical segment of the right upper lobe were endobronchial tumor was visualized.
The radial EBUS probe was inserted through the bronchoscope and advanced into the airway and an eccentric view of the lesion was identified.
Biopsies were then performed with a variety of instruments to include peripheral needle forceps and brush. ROSE identified malignancy.
After adequate samples were obtained, the video bronchoscope was then removed and the UC180F convex probe EBUS bronchoscope was introduced through the mouth, and advanced to the tracheobronchial tree.
A systematic hilar and mediastinal lymph node survey was carried out.
Sampling criteria (5mm short axis diameter) was only met in station station 4R and 11Rs.
Sampling by transbronchial needle aspiration was performed using an Olympus EBUSTBNA 22 gauge needle.
All samples were sent for routine cytology. Onsite path evaluation showed adequate lymphocytes in the 4R lymph node and were non-diagnostic in the 11Rs.
The bronchoscope was then removed and the P190 ultrathin video bronchoscope was inserted again into the airway and 5 more forceps biopsies were obtained in the apical segment of the right upper lobe for cell block and molecular analysis.
The bronchoscope was then removed and the Q190 re-inserted into the airways.
DECAMP research samples were then performed with brushing within the right upper lobe, right middle lobe and transbronchial biopsies in the RUL, RML and LUL.
We then observed for evidence of active bleeding and none was identified. The bronchoscope was removed and the procedure completed.
Complications: 	
-None 
Estimated Blood Loss:  5 cc.
Recommendations:
- Transfer to post-procedure unit
- Await biopsy results 
- Discharge home once criteria met.
Dr. Miller (Attending) was personally present and involved in all key phases of the procedure."""

# ==========================================
# 2. Procedure Flags
# ==========================================
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 1,
    "endobronchial_biopsy": 1,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 1,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 1,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# ==========================================
# 3. Spans
# ==========================================
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    # Event 1: Initial Inspection (Q190)
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190", "tracheobronchial tree, the ", "evt_01"),
    ("laryngeal mask airway", "DEV_INSTRUMENT", "LMA", "the mouth, via ", "evt_01"),
    ("fish mouth dynamic obstruction", "OBS_LESION", "fish mouth dynamic obstruction", "exception of ", "evt_01"),
    ("right middle lobe", "ANAT_LUNG_LOC", "RML", "obstruction of the ", "evt_01"),

    # Event 2: RUL Apical Nodule (P190 + Radial EBUS)
    ("P190 ultrathin video bronchoscope", "DEV_INSTRUMENT", "P190", "removed and the ", "evt_02"),
    ("apical segment of the right upper lobe", "ANAT_LUNG_LOC", "RUL Apical", "advanced into the ", "evt_02"),
    ("endobronchial tumor", "OBS_LESION", "Endobronchial Tumor", "upper lobe were ", "evt_02"),
    ("Radial EBUS probe", "PROC_METHOD", "Radial EBUS", "The ", "evt_02"),
    ("eccentric view", "OBS_LESION", "Eccentric View", "airway and an ", "evt_02"),
    ("Biopsies", "PROC_METHOD", "Biopsy", "identified. ", "evt_02"),
    ("peripheral needle forceps", "DEV_INSTRUMENT", "Forceps", "to include ", "evt_02"),
    ("brush", "DEV_INSTRUMENT", "Brush", "forceps and ", "evt_02"),
    ("ROSE identified malignancy", "OBS_ROSE", "Malignancy", "brush. ", "evt_02"),

    # Event 3: EBUS TBNA
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "UC180F", "removed and the ", "evt_03"),
    ("station 4R", "ANAT_LN_STATION", "4R", "only met in station ", "evt_03"),
    ("11Rs", "ANAT_LN_STATION", "11Rs", "4R and ", "evt_03"),
    ("transbronchial needle aspiration", "PROC_METHOD", "TBNA", "Sampling by ", "evt_03"),
    ("Olympus EBUSTBNA 22 gauge needle", "DEV_NEEDLE", "22G", "using an ", "evt_03"),
    ("adequate lymphocytes", "OBS_ROSE", "Adequate Lymphocytes", "evaluation showed ", "evt_03"),
    ("4R lymph node", "ANAT_LN_STATION", "4R", "lymphocytes in the ", "evt_03"),
    ("non-diagnostic", "OBS_ROSE", "Non-diagnostic", "and were ", "evt_03"),
    ("11Rs", "ANAT_LN_STATION", "11Rs", "diagnostic in the ", "evt_03"),

    # Event 4: More RUL Biopsies (P190)
    ("P190 ultrathin video bronchoscope", "DEV_INSTRUMENT", "P190", "removed and the ", "evt_04"),
    ("5", "MEAS_COUNT", "5", "airway and ", "evt_04"),
    ("forceps biopsies", "PROC_METHOD", "Biopsy", "more ", "evt_04"),
    ("apical segment of the right upper lobe", "ANAT_LUNG_LOC", "RUL Apical", "obtained in the ", "evt_04"),

    # Event 5: DECAMP Research
    ("Q190", "DEV_INSTRUMENT", "Q190", "removed and the ", "evt_05"),
    ("brushing", "PROC_METHOD", "Brushing", "performed with ", "evt_05"),
    ("right upper lobe", "ANAT_LUNG_LOC", "RUL", "within the ", "evt_05"),
    ("right middle lobe", "ANAT_LUNG_LOC", "RML", "right upper lobe, ", "evt_05"),
    ("transbronchial biopsies", "PROC_METHOD", "Transbronchial Biopsy", "RML and ", "evt_05"),
    ("RUL", "ANAT_LUNG_LOC", "RUL", "in the ", "evt_05"),
    ("RML", "ANAT_LUNG_LOC", "RML", "RUL, ", "evt_05"),
    ("LUL", "ANAT_LUNG_LOC", "LUL", "RML and ", "evt_05"),

    # Outcome
    ("None", "OUTCOME_COMPLICATION", "None", "Complications: \t\n-", "evt_out"),
]

# ==========================================
# 4. Events (V3 Structure)
# ==========================================
EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Diagnostic Bronchoscopy",
        "devices": ["Q190 video bronchoscope", "laryngeal mask airway"],
        "anatomy": ["right middle lobe"],
        "observations": ["fish mouth dynamic obstruction"]
    },
    {
        "event_id": "evt_02",
        "procedure_type": "Radial EBUS",
        "devices": ["P190 ultrathin video bronchoscope", "peripheral needle forceps", "brush"],
        "anatomy": ["apical segment of the right upper lobe"],
        "observations": ["endobronchial tumor", "eccentric view", "ROSE identified malignancy"]
    },
    {
        "event_id": "evt_03",
        "procedure_type": "Linear EBUS (TBNA)",
        "devices": ["UC180F convex probe EBUS bronchoscope", "Olympus EBUSTBNA 22 gauge needle"],
        "anatomy": ["station 4R", "11Rs"],
        "observations": ["adequate lymphocytes (4R)", "non-diagnostic (11Rs)"]
    },
    {
        "event_id": "evt_04",
        "procedure_type": "Endobronchial Biopsy",
        "devices": ["P190 ultrathin video bronchoscope", "Forceps"],
        "anatomy": ["apical segment of the right upper lobe"],
        "action": "5 biopsies for molecular"
    },
    {
        "event_id": "evt_05",
        "procedure_type": "Research Sampling (DECAMP)",
        "devices": ["Q190"],
        "anatomy": ["RUL", "RML", "LUL"],
        "action": "Brushing, Transbronchial Biopsies"
    }
]

# ==========================================
# 5. Helper Functions
# ==========================================
def clean_text(text):
    if not text: return ""
    return str(text).strip().replace('\r', '')

def hydrate_span(full_text, span_text, context_prefix):
    """
    Finds the start and end indices of span_text in full_text,
    verifying it follows context_prefix to ensure uniqueness.
    """
    clean_full = clean_text(full_text)
    clean_span = clean_text(span_text)
    clean_context = clean_text(context_prefix)
    
    # Construct regex: Escape special chars, allow whitespace flexibility
    # Look for context + whitespace + span
    pattern_str = re.escape(clean_context) + r"\s*" + re.escape(clean_span)
    match = re.search(pattern_str, clean_full, re.IGNORECASE)
    
    if match:
        # The match includes the context. We need the offset of the span part.
        # This is rough; precise char counting depends on exact whitespace in original.
        # Better approach: Find the span *within* the matched string.
        full_match_str = match.group(0)
        span_start_in_match = full_match_str.lower().rfind(clean_span.lower())
        
        real_start = match.start() + span_start_in_match
        real_end = real_start + len(clean_span)
        return real_start, real_end
    
    # Fallback: simple find if unique (risky but catch-all)
    start = clean_full.find(clean_span)
    if start != -1:
        return start, start + len(clean_span)
    
    return None, None

def generate_workbook():
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------------------
    # Sheet 1: Note_Text
    # -------------------------------------------------------------------------
    ws_text = wb.active
    ws_text.title = "Note_Text"
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # -------------------------------------------------------------------------
    # Sheet 2: Note_Index (Metadata + Flags)
    # -------------------------------------------------------------------------
    ws_index = wb.create_sheet("Note_Index")
    headers = ["NOTE_ID", "SOURCE_FILE", "PROCEDURE_DATE"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers)
    
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE]
    for key in PROCEDURE_FLAGS:
        row_data.append(PROCEDURE_FLAGS[key])
    ws_index.append(row_data)

    # -------------------------------------------------------------------------
    # Sheet 3: Span_Annotations (Raw)
    # -------------------------------------------------------------------------
    ws_spans = wb.create_sheet("Span_Annotations")
    span_headers = ["NOTE_ID", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "START_CHAR", "END_CHAR", "EVENT_ID"]
    ws_spans.append(span_headers)
    
    for span in SPANS:
        # (text, label, norm, context, evt)
        ws_spans.append([NOTE_ID, span[0], span[1], span[2], span[3], "", "", span[4]])

    # -------------------------------------------------------------------------
    # Sheet 4: Span_Hydrated (Calculated Offsets)
    # -------------------------------------------------------------------------
    ws_hydrated = wb.create_sheet("Span_Hydrated")
    ws_hydrated.append(span_headers)
    
    for span in SPANS:
        s_text, s_label, s_norm, s_context, s_evt = span
        start, end = hydrate_span(NOTE_TEXT, s_text, s_context)
        ws_hydrated.append([NOTE_ID, s_text, s_label, s_norm, s_context, start, end, s_evt])

    # -------------------------------------------------------------------------
    # Sheet 5: Event_Log (Flattened)
    # -------------------------------------------------------------------------
    ws_events = wb.create_sheet("Event_Log")
    ws_events.append(["NOTE_ID", "EVENT_ID", "PROCEDURE_TYPE", "DEVICES", "ANATOMY", "OBSERVATIONS", "ACTION"])
    
    for evt in EVENTS:
        ws_events.append([
            NOTE_ID,
            evt.get("event_id", ""),
            evt.get("procedure_type", ""),
            ", ".join(evt.get("devices", [])),
            ", ".join(evt.get("anatomy", [])),
            ", ".join(evt.get("observations", [])),
            evt.get("action", "")
        ])

    # -------------------------------------------------------------------------
    # Sheet 6: V3_Procedure_Events (JSON-ready structure)
    # -------------------------------------------------------------------------
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    ws_v3.append(["NOTE_ID", "EVENT_ID", "JSON_PAYLOAD"])
    
    for evt in EVENTS:
        payload = json.dumps(evt)
        ws_v3.append([NOTE_ID, evt.get("event_id"), payload])

    # -------------------------------------------------------------------------
    # Sheet 7: V3_Registry_JSON (Full Document)
    # -------------------------------------------------------------------------
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["NOTE_ID", "FULL_JSON"])
    
    registry_data = {
        "note_id": NOTE_ID,
        "source_file": SOURCE_FILE,
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS
    }
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Successfully generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()