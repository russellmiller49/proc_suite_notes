import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# -------------------------------------------------------------------------
# 1. Imports & Constants
# -------------------------------------------------------------------------
NOTE_ID = "note_187"
SOURCE_FILE = "note_187.txt"
PROCEDURE_DATE = ""
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_187 SOURCE_FILE: note_187.txt Procedure Name: EBUS bronchoscopy, transbronchial lung biopsy and bronchoalveolar lavage
Indications: Mediastinal adenopathy and interstitial lung disease 
Medications: Propofol infusion via anesthesia assistance  
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the preprocedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention. 
Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway is in good position. The vocal cords appeared normal. The subglottic space was normal.
The trachea is of normal caliber. The carina is sharp.
The tracheobronchial tree was examined to at least the first subsegmental level. Bronchial mucosa and anatomy were normal;
there are no endobronchial lesions, and no secretions. The video bronchoscope was then removed and the UC180F convex probe EBUS bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
Ultrasound was utilized to identify and measure the radiographically enlarged station 7, 4R and 11Rs lymph nodes.
Sampling by transbronchial needle aspiration was performed beginning with the station 4R lymph node, followed by the station 7 lymph node and finally the 4L lymph node using an Olympus EBUSTBNA 22 gauge needle.
Rapid onsite evaluation read as benign lymphoid tissue in all sampled nodes.
All samples were sent for routine cytology and flow cytometry.
Following completion of EBUS bronchoscopy the video bronchoscope was re-inserted and blood was suctioned from the airway.
Using fluoroscopy transbronchial forceps lung biopsies were performed in both the right middle lobe and right lower lobe.
Finally BAL was performed with 120cc instillation and 40cc return.
Prior to removing the bronchoscope an airway examination was performed and no evidence of active bleeding was seen.
The bronchoscope was removed and procedure completed. 

Complications: No immediate complications
Estimated Blood Loss: Less than 5 cc.
Post Procedure Diagnosis:
- Technically successful EBUS bronchoscopy, transbronchial lung biopsy and bronchoalveolar lavage 
- Will await final pathology and culture results"""

# -------------------------------------------------------------------------
# 2. Configuration (Procedure Flags)
# -------------------------------------------------------------------------
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 1,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 1,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# -------------------------------------------------------------------------
# 3. Data Definition (Spans)
# -------------------------------------------------------------------------
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    # Inspection
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190 video bronchoscope", "tree, the ", "evt_01"),
    ("trachea", "ANAT_AIRWAY", "trachea", "The ", "evt_01"),
    ("carina", "ANAT_AIRWAY", "carina", "The ", "evt_01"),
    
    # EBUS
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "UC180F convex probe EBUS bronchoscope", "removed and the ", "evt_02"),
    ("transbronchial needle aspiration", "PROC_METHOD", "transbronchial needle aspiration", "Sampling by ", "evt_02"),
    ("station 4R", "ANAT_LN_STATION", "4R", "beginning with the ", "evt_02"),
    ("station 7", "ANAT_LN_STATION", "7", "followed by the ", "evt_02"),
    ("4L", "ANAT_LN_STATION", "4L", "finally the ", "evt_02"),
    ("Olympus EBUSTBNA 22 gauge needle", "DEV_NEEDLE", "Olympus EBUSTBNA 22 gauge", "using an ", "evt_02"),
    ("benign lymphoid tissue", "OBS_ROSE", "benign lymphoid tissue", "read as ", "evt_02"),
    
    # Transbronchial Biopsy
    ("transbronchial forceps lung biopsies", "PROC_METHOD", "transbronchial forceps lung biopsies", "fluoroscopy ", "evt_03"),
    ("right middle lobe", "ANAT_LUNG_LOC", "right middle lobe", "in both the ", "evt_03"),
    ("right lower lobe", "ANAT_LUNG_LOC", "right lower lobe", "lobe and ", "evt_03"),
    
    # BAL
    ("BAL", "PROC_METHOD", "BAL", "Finally ", "evt_04"),
    ("120cc", "MEAS_VOL", "120", "with ", "evt_04"),
    ("40cc", "MEAS_VOL", "40", "instillation and ", "evt_04"),
    
    # Outcome
    ("No immediate complications", "OUTCOME_COMPLICATION", "No immediate complications", "Complications: ", "evt_05")
]

# -------------------------------------------------------------------------
# 4. Event Definitions
# -------------------------------------------------------------------------
EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Diagnostic Bronchoscopy",
        "action": "Inspection",
        "anatomy": ["trachea", "carina"],
        "devices": ["Q190 video bronchoscope"],
        "outcomes": []
    },
    {
        "event_id": "evt_02",
        "procedure_type": "Linear EBUS",
        "action": "TBNA",
        "anatomy": ["4R", "7", "4L"],
        "devices": ["UC180F convex probe EBUS bronchoscope", "Olympus EBUSTBNA 22 gauge needle"],
        "outcomes": ["benign lymphoid tissue"]
    },
    {
        "event_id": "evt_03",
        "procedure_type": "Transbronchial Biopsy",
        "action": "Biopsy",
        "anatomy": ["right middle lobe", "right lower lobe"],
        "devices": ["forceps"],
        "outcomes": []
    },
    {
        "event_id": "evt_04",
        "procedure_type": "BAL",
        "action": "Lavage",
        "anatomy": [],
        "devices": [],
        "outcomes": ["Instill: 120cc", "Return: 40cc"]
    },
    {
        "event_id": "evt_05",
        "procedure_type": "Outcome",
        "action": "Assessment",
        "anatomy": [],
        "devices": [],
        "outcomes": ["No immediate complications"]
    }
]

# -------------------------------------------------------------------------
# 5. Helper Functions
# -------------------------------------------------------------------------
def clean_text(text):
    return text.strip().replace('\r', '').replace('\n', ' ')

def hydrate_span(full_text, span_text, context_prefix=None):
    """
    Finds start/end offsets for span_text in full_text.
    Uses context_prefix to disambiguate if provided.
    """
    if not span_text or span_text not in full_text:
        return "", "", ""
        
    start_index = -1
    
    if context_prefix:
        pattern = re.escape(context_prefix) + r"\s*" + re.escape(span_text)
        match = re.search(pattern, full_text)
        if match:
            # The span starts after the prefix
            # We calculate distinct length of prefix part in match
            full_match_str = match.group(0)
            prefix_len = full_match_str.rfind(span_text)
            start_index = match.start() + prefix_len
    
    if start_index == -1:
        # Fallback to first occurrence if context failed or not provided
        start_index = full_text.find(span_text)

    if start_index == -1:
        return "", "", ""
        
    end_index = start_index + len(span_text)
    return str(start_index), str(end_index), span_text

# -------------------------------------------------------------------------
# 6. Workbook Generation Function
# -------------------------------------------------------------------------
def generate_workbook():
    # A. Create/Load Workbook
    if not os.path.exists(TEMPLATE_PATH):
        wb = openpyxl.Workbook()
        # Create minimal required sheets
        wb.create_sheet("Note_Text")
        wb.create_sheet("Note_Index")
        wb.create_sheet("Span_Annotations")
        wb.create_sheet("Span_Hydrated")
        wb.create_sheet("Event_Log")
        wb.create_sheet("V3_Procedure_Events")
        wb.create_sheet("V3_Registry_JSON")
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # B. Sheet 1: Note_Text
    ws_text = wb["Note_Text"]
    # Clear existing
    for row in ws_text.iter_rows(min_row=2, max_col=3):
        for cell in row:
            cell.value = None
    
    ws_text.cell(row=2, column=1, value=NOTE_ID)
    ws_text.cell(row=2, column=2, value=SOURCE_FILE)
    ws_text.cell(row=2, column=3, value=NOTE_TEXT)

    # C. Sheet 2: Note_Index (Metadata + Flags)
    ws_index = wb["Note_Index"]
    # Write Header if empty
    headers = ["note_id", "source_file", "procedure_date"] + list(PROCEDURE_FLAGS.keys())
    for col_idx, header in enumerate(headers, 1):
        ws_index.cell(row=1, column=col_idx, value=header)
    
    # Write Data
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE] + list(PROCEDURE_FLAGS.values())
    for col_idx, val in enumerate(row_data, 1):
        ws_index.cell(row=2, column=col_idx, value=val)

    # D. Sheet 3 & 4: Span_Annotations (Blank) & Span_Hydrated (Calculated)
    ws_anno = wb["Span_Annotations"]
    ws_hydrated = wb["Span_Hydrated"]
    
    # Headers
    span_headers = ["note_id", "start_char", "end_char", "text", "label", "value", "context", "event_id"]
    for col_idx, h in enumerate(span_headers, 1):
        ws_anno.cell(row=1, column=col_idx, value=h)
        ws_hydrated.cell(row=1, column=col_idx, value=h)

    # Write Spans
    for i, span_data in enumerate(SPANS, 2):
        span_text, label, norm_val, context, evt_id = span_data
        
        # Hydration logic
        start_c, end_c, actual_text = hydrate_span(NOTE_TEXT, span_text, context)
        
        # Write to Annotation (Blank Offsets)
        ws_anno.cell(row=i, column=1, value=NOTE_ID)
        ws_anno.cell(row=i, column=2, value="")
        ws_anno.cell(row=i, column=3, value="")
        ws_anno.cell(row=i, column=4, value=span_text)
        ws_anno.cell(row=i, column=5, value=label)
        ws_anno.cell(row=i, column=6, value=norm_val)
        ws_anno.cell(row=i, column=7, value=context)
        ws_anno.cell(row=i, column=8, value=evt_id)

        # Write to Hydrated (Calculated Offsets)
        ws_hydrated.cell(row=i, column=1, value=NOTE_ID)
        ws_hydrated.cell(row=i, column=2, value=start_c)
        ws_hydrated.cell(row=i, column=3, value=end_c)
        ws_hydrated.cell(row=i, column=4, value=actual_text)
        ws_hydrated.cell(row=i, column=5, value=label)
        ws_hydrated.cell(row=i, column=6, value=norm_val)
        ws_hydrated.cell(row=i, column=7, value=context)
        ws_hydrated.cell(row=i, column=8, value=evt_id)

    # E. Sheet 5: Event_Log (Flattened)
    ws_event = wb["Event_Log"]
    event_headers = ["note_id", "event_id", "procedure_type", "action", "anatomy", "devices", "outcomes"]
    for col_idx, h in enumerate(event_headers, 1):
        ws_event.cell(row=1, column=col_idx, value=h)

    for i, evt in enumerate(EVENTS, 2):
        ws_event.cell(row=i, column=1, value=NOTE_ID)
        ws_event.cell(row=i, column=2, value=evt["event_id"])
        ws_event.cell(row=i, column=3, value=evt["procedure_type"])
        ws_event.cell(row=i, column=4, value=evt["action"])
        ws_event.cell(row=i, column=5, value=", ".join(evt["anatomy"]))
        ws_event.cell(row=i, column=6, value=", ".join(evt["devices"]))
        ws_event.cell(row=i, column=7, value=", ".join(evt["outcomes"]))

    # F. Sheet 6: V3_Procedure_Events (JSON parts)
    ws_v3 = wb["V3_Procedure_Events"]
    v3_headers = ["note_id", "event_id", "procedure_type", "anatomy_json", "device_json", "outcome_json"]
    for col_idx, h in enumerate(v3_headers, 1):
        ws_v3.cell(row=1, column=col_idx, value=h)

    for i, evt in enumerate(EVENTS, 2):
        ws_v3.cell(row=i, column=1, value=NOTE_ID)
        ws_v3.cell(row=i, column=2, value=evt["event_id"])
        ws_v3.cell(row=i, column=3, value=evt["procedure_type"])
        ws_v3.cell(row=i, column=4, value=json.dumps(evt["anatomy"]))
        ws_v3.cell(row=i, column=5, value=json.dumps(evt["devices"]))
        ws_v3.cell(row=i, column=6, value=json.dumps(evt["outcomes"]))

    # G. Sheet 7: V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    ws_json.cell(row=1, column=1, value="note_id")
    ws_json.cell(row=1, column=2, value="full_json")
    
    registry_data = {
        "note_id": NOTE_ID,
        "source": SOURCE_FILE,
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS
    }
    
    ws_json.cell(row=2, column=1, value=NOTE_ID)
    ws_json.cell(row=2, column=2, value=json.dumps(registry_data, indent=2))

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

# -------------------------------------------------------------------------
# 7. Execution Block
# -------------------------------------------------------------------------
if __name__ == "__main__":
    generate_workbook()