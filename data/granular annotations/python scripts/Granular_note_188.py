import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# 1. Imports & Constants
NOTE_ID = "note_188"
SOURCE_FILE = "note_188.txt"
PROCEDURE_DATE = ""
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_188 SOURCE_FILE: note_188.txt Procedure Name: EBUS bronchoscopy and peripheral biopsy
Indications: Pulmonary nodule requiring diagnosis/staging.
Medications: Propofol infusion via anesthesia assistance  
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention. 
Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway was in good position. The vocal cords appeared normal. The subglottic space was normal.
The trachea was of normal caliber. The carina was sharp.
The tracheobronchial tree was examined to at least the first subsegmental level without endobronchial lesions visualized.
The video bronchoscope was then removed and the UC180F convex probe EBUS bronchoscope was introduced through the mouth, and advanced to the tracheobronchial tree.
A systematic hilar and mediastinal lymph node survey was carried out.
Sampling criteria (5mm short axis diameter) was only met in station station 7. Sampling by transbronchial needle aspiration was performed using an Olympus EBUSTBNA 22 gauge needle.
All samples were sent for routine cytology. Onsite path evaluation revealed benign lymphocytes.
The bronchoscope was then removed and the PX190 ultrathin video bronchoscope was inserted into the airway and based on anatomical knowledge advanced into the left upper lobe to the area of known nodule within the Apical-posterior segment the lesion was identified visually.
Biopsies were then performed with micro forceps.  After adequate samples were obtained the bronchoscope was removed.
ROSE was consistent with malignancy within the lesion.  The bronchoscope was then removed and the P190 re-inserted into the airways.
We then observed for evidence of active bleeding and none was identified. The bronchoscope was removed and the procedure completed.
Complications: 	
-None 
Estimated Blood Loss:  less than 5 cc.
Recommendations:
- Transfer to post-op ward
- Await biopsy results 
- Discharge home once criteria met."""

# 2. Configuration (Procedure Flags)
PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 1, # Micro forceps biopsy
    "tbna_conventional": 0,
    "linear_ebus": 1, # Convex probe used
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 3. Data Definition (Spans)
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Olympus Q190", "tracheobronchial tree, the ", "evt_01"),
    ("tracheobronchial tree", "ANAT_AIRWAY", "Tracheobronchial Tree", "advanced to the ", "evt_01"),
    ("endobronchial lesions", "OBS_LESION", "No Lesions", "level without ", "evt_01"),
    
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "Olympus UC180F", "removed and the ", "evt_02"),
    ("station 7", "ANAT_LN_STATION", "7", "met in station ", "evt_02"),
    ("transbronchial needle aspiration", "PROC_METHOD", "EBUS-TBNA", "Sampling by ", "evt_02"),
    ("Olympus EBUSTBNA 22 gauge needle", "DEV_NEEDLE", "22G", "using an ", "evt_02"),
    ("benign lymphocytes", "OBS_ROSE", "Benign", "evaluation revealed ", "evt_02"),
    
    ("PX190 ultrathin video bronchoscope", "DEV_INSTRUMENT", "Olympus PX190", "removed and the ", "evt_03"),
    ("left upper lobe", "ANAT_LUNG_LOC", "LUL", "advanced into the ", "evt_03"),
    ("Apical-posterior segment", "ANAT_LUNG_LOC", "LUL Apic-Post", "nodule within the ", "evt_03"),
    ("Biopsies", "PROC_ACTION", "Biopsy", "visualized.\n", "evt_03"),
    ("micro forceps", "DEV_INSTRUMENT", "Micro Forceps", "performed with ", "evt_03"),
    ("malignancy", "OBS_ROSE", "Malignancy", "consistent with ", "evt_03"),
    
    ("P190", "DEV_INSTRUMENT", "Olympus P190", "removed and the ", "evt_04"),
    ("active bleeding", "OUTCOME_COMPLICATION", "None", "evidence of ", "evt_04"),
    
    ("None", "OUTCOME_COMPLICATION", "None", "Complications: \t\n-", "evt_global"),
    ("less than 5 cc", "MEAS_VOL", "<5ml", "Blood Loss:  ", "evt_global")
]

# 4. Event Definitions
EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Bronchoscopy",
        "action": "Inspection",
        "anatomy": ["Tracheobronchial Tree"],
        "devices": ["Olympus Q190"],
        "notes": "Initial inspection, normal findings."
    },
    {
        "event_id": "evt_02",
        "procedure_type": "Bronchoscopy",
        "action": "EBUS-TBNA",
        "anatomy": ["Station 7"],
        "devices": ["Olympus UC180F", "22G Needle"],
        "notes": "Station 7 sampled, benign lymphocytes."
    },
    {
        "event_id": "evt_03",
        "procedure_type": "Bronchoscopy",
        "action": "Biopsy",
        "anatomy": ["LUL", "LUL Apic-Post"],
        "devices": ["Olympus PX190", "Micro Forceps"],
        "notes": "Peripheral biopsy of nodule, ROSE malignant."
    },
    {
        "event_id": "evt_04",
        "procedure_type": "Bronchoscopy",
        "action": "Inspection",
        "anatomy": ["Airway"],
        "devices": ["Olympus P190"],
        "notes": "Final check, no bleeding."
    }
]

# 5. Helper Functions
def clean_text(text):
    if not text: return ""
    return re.sub(r'[\r\n]+', ' ', text).strip()

def hydrate_span(text, span_text, context_prefix):
    # Normalize texts for searching
    n_text = text.replace('\r', '').replace('\n', ' ')
    n_span = span_text.replace('\r', '').replace('\n', ' ')
    n_context = context_prefix.replace('\r', '').replace('\n', ' ')
    
    # Attempt simple find
    start = n_text.find(n_span)
    
    # If context provided, try to find with context
    if start != -1 and n_context:
        combined = n_context + n_span
        c_start = n_text.find(combined)
        if c_start != -1:
            start = c_start + len(n_context)
    
    if start == -1:
        return 0, 0, "NOT_FOUND"
        
    end = start + len(n_span)
    return start, end, text[start:end]

# 6. Workbook Generation Function
def generate_workbook():
    # Create workbook
    wb = openpyxl.Workbook()
    
    # --- Sheet 1: Note_Text ---
    ws_text = wb.active
    ws_text.title = "Note_Text"
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # --- Sheet 2: Note_Index ---
    ws_index = wb.create_sheet("Note_Index")
    headers = ["NOTE_ID", "PROCEDURE_DATE", "MD_NAME"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers)
    row_data = [NOTE_ID, PROCEDURE_DATE, ""] + [PROCEDURE_FLAGS[k] for k in PROCEDURE_FLAGS]
    ws_index.append(row_data)
    
    # --- Sheet 3: Span_Annotations ---
    ws_raw = wb.create_sheet("Span_Annotations")
    ws_raw.append(["NOTE_ID", "SPAN_TEXT", "LABEL", "VALUE_NORMALIZED", "CONTEXT_PREFIX", "EVENT_ID", "START_CHAR", "END_CHAR"])
    for span in SPANS:
        ws_raw.append([NOTE_ID, span[0], span[1], span[2], span[3], span[4], "", ""])

    # --- Sheet 4: Span_Hydrated ---
    ws_hydra = wb.create_sheet("Span_Hydrated")
    ws_hydra.append(["NOTE_ID", "SPAN_TEXT", "LABEL", "VALUE_NORMALIZED", "CONTEXT_PREFIX", "EVENT_ID", "START_CHAR", "END_CHAR", "VERIFICATION"])
    
    for span in SPANS:
        start, end, verify = hydrate_span(NOTE_TEXT, span[0], span[3])
        ws_hydra.append([NOTE_ID, span[0], span[1], span[2], span[3], span[4], start, end, verify])

    # --- Sheet 5: Event_Log ---
    ws_event_log = wb.create_sheet("Event_Log")
    ws_event_log.append(["NOTE_ID", "EVENT_ID", "PROCEDURE_TYPE", "ACTION", "ANATOMY", "DEVICES", "NOTES"])
    for evt in EVENTS:
        ws_event_log.append([
            NOTE_ID,
            evt["event_id"],
            evt["procedure_type"],
            evt["action"],
            ", ".join(evt["anatomy"]),
            ", ".join(evt["devices"]),
            evt["notes"]
        ])

    # --- Sheet 6: V3_Procedure_Events ---
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    ws_v3.append(["note_id", "event_id", "event_type", "event_metadata", "comments"])
    for evt in EVENTS:
        meta = {
            "method": evt["action"],
            "anatomy": evt["anatomy"],
            "devices": evt["devices"]
        }
        ws_v3.append([NOTE_ID, evt["event_id"], "procedure_step", json.dumps(meta), evt["notes"]])

    # --- Sheet 7: V3_Registry_JSON ---
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["note_id", "json_data"])
    
    registry_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source_file": SOURCE_FILE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS,
        "spans": [
            {
                "text": s[0],
                "label": s[1],
                "normalized": s[2],
                "event_id": s[4]
            } for s in SPANS
        ]
    }
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    if os.path.exists(OUTPUT_PATH):
        os.remove(OUTPUT_PATH)
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

# 7. Execution Block
if __name__ == "__main__":
    generate_workbook()