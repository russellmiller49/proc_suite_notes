import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# =============================================================================
# 1. IMPORTS & CONSTANTS
# =============================================================================

NOTE_ID = "note_189"
SOURCE_FILE = "note_189.txt"
PROCEDURE_DATE = ""
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_189 SOURCE_FILE: note_189.txt Procedure Name: EBUS bronchoscopy, radial EBUS guided bronchoscopy
Indications: Pulmonary nodule requiring diagnosis/staging.
Medications: Propofol infusion via anesthesia assistance  
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the preprocedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention. 
Following intravenous medications as per the anesthesia record and topical anesthesia to the upper airway and tracheobronchial tree the Q190 video bronchoscope was introduced through the mouth.
The vocal cords appeared normal. The subglottic space was normal. The trachea is of normal caliber. The carina was sharp.
All left and right sided airways were normal without endobronchial disease to the first segmental branch.
The video bronchoscope was then removed and the UC180F convex probe EBUS bronchoscope was introduced through the mouth, and advanced to the tracheobronchial tree.
A systematic hilar and mediastinal lymph node survey was carried out.
Sampling criteria (5mm short axis diameter) were met in station 11L, station 7, 10R 11Rs lymph nodes.
Sampling by transbronchial needle aspiration was performed in these lymph nodes using an Olympus EBUSTBNA 22 gauge needle beginning at the N3 (11L) lymph nodes 711Rs10R.
Further details regarding nodal size and number of samples are included in the attached EBUS procedural sheet.
All samples were sent for routine cytology. Onsite path eval did not identify malignancy.
We then removed the EBUS bronchoscopy and the ultrathin bronchoscope was inserted and based on anatomical knowledge advanced into the right middle lobe to the area of known nodule using radial EBUS confirmed a concentric view within the lesion.
Biopsies were then performed with a variety of instruments to include peripheral needle, forceps and brush.
After adequate samples were obtained the bronchoscope was removed. While emerging from anesthesia patient had large volume emesis and possible aspiration event.
He was immediately placed on right side. He transiently became hypoxic but saturations improved relatively quickly.
Patinet was then transferred to the PACU. X-ray was performed which right greater than left patchy opacities with increased pulmonary vasculature and interstitial markings alone with a small right apical pneumothorax.
Right sided airspace opacities, although possibly representing aspiration event, are more likely related to retained saline instilled during bronchoscopy and is not unexpected.
Patient was informed of pneumothorax. He was hemodynamically stable and oxygenating well.
Repeat CXR 1 hour later showed no expansion of pneumothorax and decision was made to observe for expansion and not place chest tube unless he develops symptoms or has expansion of pneumothorax.
Complications: 
1.	Small left apical pneumothorax
2.	Post procedure emesis with possible aspiration  
Estimated Blood Loss: Less than 10 cc.
Recommendations:
- Admit to IM for overnight observation
- Repeat CXR in 4 hours and then AM if not changes
- Consider diuresis given pulmonary edema and Systolic heart failure
- Continue to hold Plavix for now.
- Would NOT treat with antibiotics 
- Await final pathology"""

# =============================================================================
# 2. CONFIGURATION (PROCEDURE FLAGS)
# =============================================================================

PROCEDURE_FLAGS = {
    # Bronchoscopy
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 1,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 1,
    "linear_ebus": 1,
    "radial_ebus": 1,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 1,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    # Pleural
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# =============================================================================
# 3. DATA DEFINITION (SPANS)
# =============================================================================

SPANS = [
    # Event 01: Initial Inspection
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190", "tracheobronchial tree the", "evt_01"),
    ("vocal cords", "ANAT_AIRWAY", "vocal cords", "The", "evt_01"),
    ("subglottic space", "ANAT_AIRWAY", "subglottic space", "The", "evt_01"),
    ("trachea", "ANAT_AIRWAY", "trachea", "The", "evt_01"),
    ("carina", "ANAT_AIRWAY", "carina", "The", "evt_01"),
    ("left and right sided airways", "ANAT_AIRWAY", "bilateral airways", "All", "evt_01"),

    # Event 02: Linear EBUS & TBNA
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "UC180F", "removed and the", "evt_02"),
    ("station 11L", "ANAT_LN_STATION", "11L", "were met in", "evt_02"),
    ("station 7", "ANAT_LN_STATION", "7", "station 11L,", "evt_02"),
    ("10R", "ANAT_LN_STATION", "10R", "station 7,", "evt_02"),
    ("11Rs", "ANAT_LN_STATION", "11Rs", "10R", "evt_02"),
    ("transbronchial needle aspiration", "PROC_METHOD", "TBNA", "Sampling by", "evt_02"),
    ("Olympus EBUSTBNA 22 gauge needle", "DEV_NEEDLE", "22G", "using an", "evt_02"),
    ("Onsite path eval did not identify malignancy", "OBS_ROSE", "benign", "cytology.", "evt_02"),

    # Event 03: Radial EBUS & Peripheral Biopsies
    ("ultrathin bronchoscope", "DEV_INSTRUMENT", "Ultrathin", "bronchoscopy and the", "evt_03"),
    ("right middle lobe", "ANAT_LUNG_LOC", "RML", "into the", "evt_03"),
    ("radial EBUS", "PROC_METHOD", "Radial EBUS", "nodule using", "evt_03"),
    ("concentric view", "OBS_LESION", "concentric", "confirmed a", "evt_03"),
    ("peripheral needle", "DEV_NEEDLE", "Needle", "to include", "evt_03"),
    ("forceps", "DEV_INSTRUMENT", "Forceps", "peripheral needle,", "evt_03"),
    ("brush", "DEV_INSTRUMENT", "Brush", "forceps and", "evt_03"),
    
    # Event 04: Complications
    ("large volume emesis", "OUTCOME_COMPLICATION", "Emesis", "patient had", "evt_04"),
    ("possible aspiration event", "OUTCOME_COMPLICATION", "Aspiration", "emesis and", "evt_04"),
    ("small right apical pneumothorax", "OUTCOME_COMPLICATION", "Pneumothorax", "alone with a", "evt_04"),
    ("Small left apical pneumothorax", "OUTCOME_COMPLICATION", "Pneumothorax", "1.", "evt_04"),
]

# =============================================================================
# 4. EVENT DEFINITIONS
# =============================================================================

EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Bronchoscopy",
        "action": "Inspection",
        "anatomy": ["vocal cords", "trachea", "carina", "bronchial tree"],
        "devices": ["Q190 video bronchoscope"],
        "description": "Routine inspection showing normal airways."
    },
    {
        "event_id": "evt_02",
        "procedure_type": "Bronchoscopy",
        "action": "TBNA",
        "anatomy": ["11L", "7", "10R", "11Rs"],
        "devices": ["UC180F convex probe EBUS bronchoscope", "Olympus EBUSTBNA 22 gauge needle"],
        "description": "Linear EBUS TBNA of mediastinal and hilar lymph nodes."
    },
    {
        "event_id": "evt_03",
        "procedure_type": "Bronchoscopy",
        "action": "Biopsy",
        "anatomy": ["Right Middle Lobe"],
        "devices": ["Ultrathin bronchoscope", "Radial EBUS", "Peripheral needle", "Forceps", "Brush"],
        "description": "Radial EBUS guided biopsy of RML nodule using needle, forceps, and brush."
    },
    {
        "event_id": "evt_04",
        "procedure_type": "Outcome",
        "action": "Complication",
        "anatomy": ["Lungs"],
        "devices": [],
        "description": "Post-procedure emesis/aspiration and small pneumothorax (conflicting laterality in note, captured both)."
    }
]

# =============================================================================
# 5. HELPER FUNCTIONS
# =============================================================================

def clean_text(text):
    if not text:
        return ""
    return re.sub(r'[\r\n]+', ' ', text).strip()

def hydrate_span(full_text, span_text, context_prefix):
    """
    Finds the exact start/end offsets of span_text within full_text,
    verifying it follows context_prefix to ensure uniqueness.
    """
    # Normalize for searching (collapse whitespace)
    search_text = re.sub(r'\s+', ' ', full_text)
    search_span = re.sub(r'\s+', ' ', span_text)
    search_context = re.sub(r'\s+', ' ', context_prefix)

    # Create regex pattern: context + whitespace + span
    pattern = re.escape(search_context) + r'\s*' + re.escape(search_span)
    
    match = re.search(pattern, search_text, re.IGNORECASE)
    if match:
        # We found the combined string. Now we need to map back to original indices.
        # This is tricky with whitespace normalization. 
        # Strategy: Find the start index of the match in the NORMALIZED string,
        # then locate the corresponding text in the original string.
        
        # Simpler approach since strict byte-offests aren't requested for DB storage yet, 
        # just for Excel verification:
        
        start_index = full_text.find(span_text) # Simple find first
        
        # If strict context needed or duplicates exist:
        if full_text.count(span_text) > 1:
            # Fallback: locate via context in original string roughly
            c_start = full_text.find(context_prefix)
            if c_start != -1:
                start_index = full_text.find(span_text, c_start)
        
        if start_index != -1:
            end_index = start_index + len(span_text)
            return start_index, end_index
            
    return "", ""

def find_offsets(text, span_text, label, value, context, event_id):
    start, end = hydrate_span(text, span_text, context)
    return {
        "text": span_text,
        "label": label,
        "normalized_value": value,
        "start_char": start,
        "end_char": end,
        "context_prefix": context,
        "event_id": event_id
    }

# =============================================================================
# 6. WORKBOOK GENERATION
# =============================================================================

def generate_workbook():
    # 1. Load or Create Template
    if not os.path.exists(TEMPLATE_PATH):
        wb = openpyxl.Workbook()
        wb.create_sheet("Note_Text")
        wb.create_sheet("Note_Index")
        wb.create_sheet("Span_Annotations")
        wb.create_sheet("Span_Hydrated")
        wb.create_sheet("Event_Log")
        wb.create_sheet("V3_Procedure_Events")
        wb.create_sheet("V3_Registry_JSON")
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
    else:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # 2. Sheet 1: Note_Text
    ws_text = wb["Note_Text"]
    # Clear existing
    for row in ws_text.iter_rows(min_row=2, max_col=3):
        for cell in row:
            cell.value = None
    
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])

    # 3. Sheet 2: Note_Index
    ws_index = wb["Note_Index"]
    # Headers should exist in template, but we'll append data
    # Row structure: Note_ID, Procedure_Date, [30 Flags...]
    
    row_data = [NOTE_ID, PROCEDURE_DATE]
    for key in PROCEDURE_FLAGS:
        row_data.append(PROCEDURE_FLAGS[key])
    
    ws_index.append(row_data)

    # 4. Sheet 3 & 4: Spans (Annotations & Hydrated)
    ws_anno = wb["Span_Annotations"]
    ws_hydra = wb["Span_Hydrated"]
    
    # Headers: Note_ID, Span_Text, Label, Normalized_Value, Context_Prefix, Event_ID, Start_Char, End_Char
    
    for span in SPANS:
        s_text, s_label, s_val, s_context, s_evt = span
        
        # Annotation Row (No offsets)
        ws_anno.append([NOTE_ID, s_text, s_label, s_val, s_context, s_evt, "", ""])
        
        # Hydrated Row (Calculated offsets)
        offsets = find_offsets(NOTE_TEXT, s_text, s_label, s_val, s_context, s_evt)
        ws_hydra.append([
            NOTE_ID, 
            offsets["text"], 
            offsets["label"], 
            offsets["normalized_value"], 
            offsets["context_prefix"], 
            offsets["event_id"], 
            offsets["start_char"], 
            offsets["end_char"]
        ])

    # 5. Sheet 5: Event_Log (Flattened)
    ws_events = wb["Event_Log"]
    for evt in EVENTS:
        # Convert lists to comma-string for flat CSV-style representation
        anat_str = ", ".join(evt["anatomy"])
        dev_str = ", ".join(evt["devices"])
        ws_events.append([
            NOTE_ID,
            evt["event_id"],
            evt["procedure_type"],
            evt["action"],
            anat_str,
            dev_str,
            evt["description"]
        ])

    # 6. Sheet 6: V3_Procedure_Events (Detailed)
    ws_v3 = wb["V3_Procedure_Events"]
    for evt in EVENTS:
        ws_v3.append([
            NOTE_ID,
            evt["event_id"],
            evt["procedure_type"],
            evt["action"],
            json.dumps(evt["anatomy"]),
            json.dumps(evt["devices"]),
            evt["description"]
        ])

    # 7. Sheet 7: V3_Registry_JSON
    ws_json = wb["V3_Registry_JSON"]
    
    # Construct the full Registry Object
    registry_obj = {
        "metadata": {
            "note_id": NOTE_ID,
            "source_file": SOURCE_FILE,
            "procedure_date": PROCEDURE_DATE
        },
        "flags": PROCEDURE_FLAGS,
        "events": EVENTS,
        "spans": [
            find_offsets(NOTE_TEXT, s[0], s[1], s[2], s[3], s[4]) for s in SPANS
        ]
    }
    
    ws_json.append([NOTE_ID, json.dumps(registry_obj, indent=2)])

    # 8. Save
    wb.save(OUTPUT_PATH)
    print(f"Workbook generated successfully: {OUTPUT_PATH}")

# =============================================================================
# 7. EXECUTION BLOCK
# =============================================================================

if __name__ == "__main__":
    generate_workbook()