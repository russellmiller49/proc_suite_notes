import re
import json
import os
import datetime
from openpyxl import Workbook, load_workbook

# 1. Imports & Constants
NOTE_ID = "note_190"
SOURCE_FILE = "note_190.txt"
PROCEDURE_DATE = ""  # Not specified in text
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_190 SOURCE_FILE: note_190.txt Indications: Mediastinal adenopathy and lung mass
Procedure Performed: EBUS bronchoscopy single station.
Pre-operative diagnosis: Lung mass 
Post-operative diagnosis: malignant mediastinal adenopathy, new onset A. fib 
Medications: General Anesthesia,
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the pre-procedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention.
Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway is in good position. The vocal cords appeared normal. The subglottic space was normal.
The trachea is of normal caliber. The carina is sharp. Blood was seen within the right main-stem and gently suctioned.
The tracheobronchial tree was examined to at least the first sub-segmental level.
Obvious sub-mucosal infiltration of tumor was seen through the bronchus intermedius extending into the right middle lobe.
The right upper lobe was obstructed from combination extrinsic compression and sub-mucosal infiltration.
Blood was seen slowly oozing from the right upper lobe. The right lower lobe bronchial mucosa and anatomy were normal;
without endobronchial lesions. The left sided airways had normal bronchial mucosa and anatomy; without endobronchial lesions.
The video bronchoscope was then removed and the UC180F convex probe EBUS bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
Ultrasound was utilized to identify and measure the 4R lymph node which was a conglomerate of nodes extending into the station 10R .
Sampling by transbronchial needle aspiration was performed  with the Olympus EBUS-TBNA 22 gauge needle.
Rapid onsite evaluation read as malignancy. All samples were sent for routine cytology.
Near the end of the procedure the patient became significantly tachycardic and hypotensive with new onset A. fib with RVR.
The EBUS bronchoscope was removed and the Q190 video bronchoscope was re-inserted and blood was suctioned from the airway.
The bronchoscope was removed and procedure completed. Cardiology was called to the bedside and decision was made for bedside cardioversion which was successful in terminating the rhythm for a short period however she subsequently returned to atrial flutter but with a more controlled rate.
After awakening the patient had difficulty with oxygen saturation for a few minutes and we reintroduced the bronchoscope through the nasal passage to evaluate for evidence of active bleeding within the airway, none was seen.
The patient was then transferred to the ICU for observation status post cardioversion in good condition.
Complications: Intraoperative atrial fibrillation with RVR
Estimated Blood Loss: 10 cc.
Post Procedure Diagnosis:
- Technically successful flexible bronchoscopy with endobronchial ultrasound-guided biopsies.
-New onset atrial fibrillation
Recommendations
- Will await final pathology results
- We’ll discuss with radiation oncology early treatment to right upper lobe given active hemoptysis."""

# 2. Configuration (Procedure Flags)
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 3. Data Definition (Spans)
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    # Event 1: Diagnostic Bronchoscopy (Inspection)
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190 Bronchoscope", "tracheobronchial tree, the", "evt_01"),
    ("tracheobronchial tree", "ANAT_AIRWAY", "Tracheobronchial Tree", "advanced to the", "evt_01"),
    ("right main-stem", "ANAT_AIRWAY", "RMS", "Blood was seen within the", "evt_01"),
    ("suctioned", "PROC_ACTION", "Suction", "right main-stem and gently", "evt_01"),
    ("sub-mucosal infiltration of tumor", "OBS_LESION", "Tumor Infiltration", "Obvious", "evt_01"),
    ("bronchus intermedius", "ANAT_AIRWAY", "Bronchus Intermedius", "seen through the", "evt_01"),
    ("right middle lobe", "ANAT_LUNG_LOC", "RML", "extending into the", "evt_01"),
    ("right upper lobe", "ANAT_LUNG_LOC", "RUL", "The", "evt_01"),
    ("obstructed", "OBS_LESION", "Obstruction", "right upper lobe was", "evt_01"),
    ("Blood", "OBS_LESION", "Blood", "infiltration. ", "evt_01"),

    # Event 2: EBUS-TBNA
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "UC180F EBUS Scope", "removed and the", "evt_02"),
    ("4R lymph node", "ANAT_LN_STATION", "4R", "identify and measure the", "evt_02"),
    ("station 10R", "ANAT_LN_STATION", "10R", "extending into the", "evt_02"),
    ("transbronchial needle aspiration", "PROC_METHOD", "TBNA", "Sampling by", "evt_02"),
    ("Olympus EBUS-TBNA 22 gauge needle", "DEV_NEEDLE", "Olympus 22G EBUS Needle", "with the", "evt_02"),
    ("malignancy", "OBS_ROSE", "Malignancy", "Rapid onsite evaluation read as", "evt_02"),

    # Event 3: Complication (A-Fib) & Management
    ("tachycardic and hypotensive", "OUTCOME_COMPLICATION", "Tachycardia and Hypotension", "patient became significantly", "evt_03"),
    ("new onset A. fib with RVR", "OUTCOME_COMPLICATION", "Atrial Fibrillation with RVR", "hypotensive with", "evt_03"),
    ("bedside cardioversion", "PROC_ACTION", "Cardioversion", "decision was made for", "evt_03"),
    ("terminating the rhythm", "OUTCOME_SYMPTOMS", "Rhythm Terminated", "successful in", "evt_03"),
    
    # Event 4: Re-look Suctioning
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190 Bronchoscope", "removed and the", "evt_04"),
    ("blood", "OBS_LESION", "Blood", "re-inserted and", "evt_04"),
    ("suctioned", "PROC_ACTION", "Suction", "and blood was", "evt_04")
]

# 4. Event Definitions (V3 Structured Data)
EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Diagnostic Bronchoscopy",
        "action": "Inspection",
        "anatomy": ["Tracheobronchial Tree", "RMS", "Bronchus Intermedius", "RML", "RUL"],
        "devices": ["Q190 Bronchoscope"],
        "observations": ["Tumor Infiltration", "Obstruction", "Blood"]
    },
    {
        "event_id": "evt_02",
        "procedure_type": "Linear EBUS",
        "action": "TBNA",
        "anatomy": ["4R", "10R"],
        "devices": ["UC180F EBUS Scope", "Olympus 22G EBUS Needle"],
        "observations": ["Malignancy (ROSE)"]
    },
    {
        "event_id": "evt_03",
        "procedure_type": "Complication Management",
        "action": "Cardioversion",
        "anatomy": [],
        "devices": [],
        "observations": ["Tachycardia", "Hypotension", "Atrial Fibrillation"]
    },
    {
        "event_id": "evt_04",
        "procedure_type": "Therapeutic Bronchoscopy",
        "action": "Suction",
        "anatomy": ["Airway"],
        "devices": ["Q190 Bronchoscope"],
        "observations": ["Blood"]
    }
]

# 5. Helper Functions
def clean_text(text):
    if not text:
        return ""
    return re.sub(r'[\r\n]+', ' ', text).strip()

def hydrate_span(text, span_text, context_prefix):
    """
    Finds the start and end offsets of a span in the text, using context to disambiguate.
    """
    # Simple normalization for finding
    norm_text = text.replace('\n', ' ').replace('\r', '')
    norm_span = span_text.replace('\n', ' ').replace('\r', '')
    norm_context = context_prefix.replace('\n', ' ').replace('\r', '')
    
    # Construct a search pattern
    # We look for context + optional whitespace + span
    # Escape special regex chars in context and span
    pattern_str = re.escape(norm_context) + r'\s*' + re.escape(norm_span)
    match = re.search(pattern_str, norm_text, re.IGNORECASE)
    
    if match:
        # The match includes the context. We need the offset of the span part.
        # Find where the span starts within the match
        full_match_str = match.group(0)
        # Find the start index of the span within the matched string (case insensitive search from right to ensure we get the span)
        span_start_in_match = full_match_str.lower().rfind(norm_span.lower())
        
        start_char = match.start() + span_start_in_match
        end_char = start_char + len(norm_span)
        return start_char, end_char
    
    # Fallback: try finding just the span if context fails (less accurate)
    match_fallback = re.search(re.escape(norm_span), norm_text, re.IGNORECASE)
    if match_fallback:
        return match_fallback.start(), match_fallback.end()
        
    return "", ""

# 6. Workbook Generation Function
def generate_workbook():
    # Create or load workbook
    wb = Workbook()
    
    # --- Sheet 1: Note_Text ---
    ws1 = wb.active
    ws1.title = "Note_Text"
    ws1.append(["note_id", "source_file", "note_text"])
    ws1.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # --- Sheet 2: Note_Index ---
    ws2 = wb.create_sheet("Note_Index")
    headers_index = ["note_id", "procedure_date"] + list(PROCEDURE_FLAGS.keys())
    row_index = [NOTE_ID, PROCEDURE_DATE] + list(PROCEDURE_FLAGS.values())
    ws2.append(headers_index)
    ws2.append(row_index)
    
    # --- Sheet 3: Span_Annotations ---
    ws3 = wb.create_sheet("Span_Annotations")
    headers_spans = ["note_id", "span_text", "label", "normalized_value", "context_prefix", "event_id", "start_char", "end_char"]
    ws3.append(headers_spans)
    
    for span_data in SPANS:
        # span_text, label, normalized, context, event_id
        row = [NOTE_ID, span_data[0], span_data[1], span_data[2], span_data[3], span_data[4], "", ""]
        ws3.append(row)
        
    # --- Sheet 4: Span_Hydrated ---
    ws4 = wb.create_sheet("Span_Hydrated")
    ws4.append(headers_spans)
    
    for span_data in SPANS:
        start, end = hydrate_span(NOTE_TEXT, span_data[0], span_data[3])
        row = [NOTE_ID, span_data[0], span_data[1], span_data[2], span_data[3], span_data[4], start, end]
        ws4.append(row)
        
    # --- Sheet 5: Event_Log ---
    ws5 = wb.create_sheet("Event_Log")
    ws5.append(["note_id", "event_id", "procedure_type", "action", "anatomy", "devices", "observations"])
    
    for evt in EVENTS:
        row = [
            NOTE_ID,
            evt["event_id"],
            evt["procedure_type"],
            evt["action"],
            ", ".join(evt["anatomy"]),
            ", ".join(evt["devices"]),
            ", ".join(evt["observations"])
        ]
        ws5.append(row)
        
    # --- Sheet 6: V3_Procedure_Events ---
    ws6 = wb.create_sheet("V3_Procedure_Events")
    ws6.append(["note_id", "event_id", "event_json"])
    
    for evt in EVENTS:
        ws6.append([NOTE_ID, evt["event_id"], json.dumps(evt)])
        
    # --- Sheet 7: V3_Registry_JSON ---
    ws7 = wb.create_sheet("V3_Registry_JSON")
    ws7.append(["note_id", "full_json"])
    
    full_registry_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source": SOURCE_FILE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS,
        "spans": [
            {
                "text": s[0],
                "label": s[1],
                "normalized": s[2],
                "event_id": s[4]
            } for s in SPANS
        ]
    }
    ws7.append([NOTE_ID, json.dumps(full_registry_data, indent=2)])
    
    # Save
    wb.save(OUTPUT_PATH)
    print(f"Workbook generated successfully: {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()