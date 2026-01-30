import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# 1. Imports & Constants
NOTE_ID = "note_191"
SOURCE_FILE = "note_191.txt"
PROCEDURE_DATE = "" 
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_191 SOURCE_FILE: note_191.txt Procedure Name: EBUS bronchoscopy
Indications: Mediastinal adenopathy 
Medications: Propofol infusion via anesthesia assistance  
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the preprocedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention. 
Following intravenous medications as per the record and topical anesthesia to the upper airway and tracheobronchial tree, the Q190 video bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
The laryngeal mask airway is in good position. The vocal cords appeared normal. The subglottic space was normal.
The trachea is of normal caliber. The carina is sharp.
The tracheobronchial tree was examined to at least the first sub-segmental level. Bronchial mucosa and anatomy were normal;
there are no endobronchial lesions, and no secretions. The video bronchoscope was then removed and the UC180F convex probe EBUS bronchoscope was introduced through the mouth, via laryngeal mask airway and advanced to the tracheobronchial tree.
Ultrasound was utilized to identify and measure the radiographically enlarged station 7, 4R and 4L lymph nodes.
Sampling by transbronchial needle aspiration using Olympus EBUSTBNA 22 gauge and 19G needles.
At least 5 needle passes were performed at each lymph node station. All samples were sent for routine cytology.
ROSE identified granulomas within the 4R lymph node. Following completion of EBUS bronchoscopy the video bronchoscope was re-inserted and blood was suctioned from the airway.
There was no evidence of active bleeding and the bronchoscope was removed and procedure completed.
Complications: No immediate complications
Estimated Blood Loss: Less than 5 cc.
Post Procedure Diagnosis:
- Technically successful EBUS bronchoscopy
- Will await final pathology results"""

# 2. Configuration (Procedure Flags)
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 0,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 0,
    "navigational_bronchoscopy": 0,
    "transbronchial_biopsy": 0,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 3. Data Definition (Spans)
# Format: (span_text, label, normalized_value, context_prefix, event_id)
SPANS = [
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190", "tree, the ", "evt_01"),
    ("tracheobronchial tree", "ANAT_AIRWAY", "Tracheobronchial tree", "to the ", "evt_01"),
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "UC180F", "removed and the ", "evt_02"),
    ("station 7", "ANAT_LN_STATION", "7", "radiographically enlarged ", "evt_02"),
    ("4R", "ANAT_LN_STATION", "4R", "station 7, ", "evt_02"),
    ("4L", "ANAT_LN_STATION", "4L", "4R and ", "evt_02"),
    ("transbronchial needle aspiration", "PROC_METHOD", "TBNA", "Sampling by ", "evt_02"),
    ("22 gauge", "DEV_NEEDLE", "22G", "Olympus EBUSTBNA ", "evt_02"),
    ("19G", "DEV_NEEDLE", "19G", "22 gauge and ", "evt_02"),
    ("5", "MEAS_COUNT", "5", "At least ", "evt_02"),
    ("granulomas", "OBS_ROSE", "Granulomas", "ROSE identified ", "evt_03"),
    ("4R", "ANAT_LN_STATION", "4R", "within the ", "evt_03"),
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications: ", "evt_04")
]

# 4. Event Definitions
EVENTS = [
    {
        "event_id": "evt_01",
        "procedure_type": "Bronchoscopy",
        "action": "Inspection",
        "anatomy": ["Tracheobronchial tree"],
        "devices": ["Q190 video bronchoscope"],
        "description": "Initial diagnostic inspection"
    },
    {
        "event_id": "evt_02",
        "procedure_type": "EBUS",
        "action": "TBNA",
        "anatomy": ["Station 7", "Station 4R", "Station 4L"],
        "devices": ["UC180F convex probe EBUS bronchoscope", "22G Needle", "19G Needle"],
        "description": "EBUS TBNA sampling"
    },
    {
        "event_id": "evt_03",
        "procedure_type": "Bronchoscopy",
        "action": "ROSE",
        "anatomy": ["Station 4R"],
        "devices": [],
        "description": "Rapid On-Site Evaluation identified granulomas"
    },
    {
        "event_id": "evt_04",
        "procedure_type": "Procedure Outcome",
        "action": "Assessment",
        "anatomy": [],
        "devices": [],
        "description": "No immediate complications"
    }
]

# 5. Helper Functions
def clean_text(text):
    if not text: return ""
    return re.sub(r'[\r\n]+', ' ', text).strip()

def find_offsets(text, span_text, context_prefix):
    if not span_text: return None, None
    # Normalize spaces for search
    text_norm = " ".join(text.split())
    span_norm = " ".join(span_text.split())
    prefix_norm = " ".join(context_prefix.split())
    
    # Construct regex pattern
    # Escape special regex chars in span and prefix
    pattern = re.escape(prefix_norm) + r"\s*" + re.escape(span_norm)
    
    match = re.search(pattern, text_norm, re.IGNORECASE)
    if match:
        # We found the sequence "prefix + span".
        # The start of the span is match.start() + len(prefix) (approx, need to adjust for whitespace)
        # However, to map back to original text with newlines is tricky. 
        # A safer approach for this Phase 0 script which populates Excel:
        # Just return the found text as-is if strictly required, but the instruction asks for
        # start/end chars. 
        
        # Simple strategy: find exact substring in original text
        # If context provided, try to find that specific instance
        
        search_start = 0
        if context_prefix:
            prefix_loc = text.find(context_prefix)
            if prefix_loc != -1:
                search_start = prefix_loc + len(context_prefix)
        
        start_index = text.find(span_text, search_start)
        if start_index != -1:
            end_index = start_index + len(span_text)
            return start_index, end_index
            
    # Fallback: simple find
    start = text.find(span_text)
    if start != -1:
        return start, start + len(span_text)
        
    return None, None

def hydrate_span(full_text, span_data):
    span_text, label, normalized, context, evt_id = span_data
    start, end = find_offsets(full_text, span_text, context)
    return {
        "text": span_text,
        "label": label,
        "normalized": normalized,
        "start": start,
        "end": end,
        "event_id": evt_id
    }

# 6. Workbook Generation
def generate_workbook():
    # Create workbook
    wb = openpyxl.Workbook()
    
    # --- Sheet 1: Note_Text ---
    ws1 = wb.active
    ws1.title = "Note_Text"
    ws1.append(["note_id", "source_file", "note_text"])
    ws1.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # --- Sheet 2: Note_Index ---
    ws2 = wb.create_sheet("Note_Index")
    headers = ["note_id", "source_file", "procedure_date"] + list(PROCEDURE_FLAGS.keys())
    ws2.append(headers)
    row_data = [NOTE_ID, SOURCE_FILE, PROCEDURE_DATE] + list(PROCEDURE_FLAGS.values())
    ws2.append(row_data)
    
    # --- Sheet 3: Span_Annotations ---
    ws3 = wb.create_sheet("Span_Annotations")
    ws3.append(["note_id", "span_text", "label", "normalized_value", "start_char", "end_char", "context_prefix", "event_id"])
    for span in SPANS:
        # span = (text, label, norm, context, evt)
        ws3.append([NOTE_ID, span[0], span[1], span[2], "", "", span[3], span[4]])
        
    # --- Sheet 4: Span_Hydrated ---
    ws4 = wb.create_sheet("Span_Hydrated")
    ws4.append(["note_id", "span_text", "label", "normalized_value", "start_char", "end_char", "event_id"])
    for span in SPANS:
        hydrated = hydrate_span(NOTE_TEXT, span)
        ws4.append([NOTE_ID, hydrated['text'], hydrated['label'], hydrated['normalized'], hydrated['start'], hydrated['end'], hydrated['event_id']])

    # --- Sheet 5: Event_Log ---
    ws5 = wb.create_sheet("Event_Log")
    ws5.append(["note_id", "event_id", "procedure_type", "action", "anatomy", "devices", "description"])
    for evt in EVENTS:
        ws5.append([
            NOTE_ID,
            evt["event_id"],
            evt["procedure_type"],
            evt["action"],
            ", ".join(evt["anatomy"]),
            ", ".join(evt["devices"]),
            evt["description"]
        ])
        
    # --- Sheet 6: V3_Procedure_Events ---
    ws6 = wb.create_sheet("V3_Procedure_Events")
    ws6.append(["note_id", "event_id", "event_json"])
    for evt in EVENTS:
        ws6.append([NOTE_ID, evt["event_id"], json.dumps(evt)])

    # --- Sheet 7: V3_Registry_JSON ---
    ws7 = wb.create_sheet("V3_Registry_JSON")
    ws7.append(["note_id", "registry_json"])
    
    registry_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source": SOURCE_FILE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS,
        "spans": [hydrate_span(NOTE_TEXT, s) for s in SPANS]
    }
    ws7.append([NOTE_ID, json.dumps(registry_data, indent=2)])

    # Save
    if not os.path.exists(TEMPLATE_PATH):
        # We don't strictly need the template file to exist to write the output, 
        # but the instructions imply using it if possible. 
        # Here we just save the constructed workbook.
        pass
        
    wb.save(OUTPUT_PATH)
    print(f"Generated {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()