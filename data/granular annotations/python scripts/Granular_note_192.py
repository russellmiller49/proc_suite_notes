import openpyxl
from openpyxl.utils import get_column_letter
import re
import json
import os
import datetime

# 1. Imports & Constants
NOTE_ID = "note_192"
SOURCE_FILE = "note_192.txt"
PROCEDURE_DATE = ""
TEMPLATE_PATH = "phase0_golden_registry_labeling_worksheet_anchor_first_therapeutic_pleural.xlsx"
OUTPUT_PATH = f"phase0_extraction_{NOTE_ID}.xlsx"

NOTE_TEXT = """NOTE_ID:  note_192 SOURCE_FILE: note_192.txt Indications: Left upper lobe mass 
Medications: Propofol infusion via anesthesia assistance  
Procedure, risks, benefits, and alternatives were explained to the patient.
All questions were answered and informed consent was documented as per institutional protocol.
A history and physical were performed and updated in the preprocedure assessment record. Laboratory studies and radiographs were reviewed.
A time-out was performed prior to the intervention. 
Following intravenous medications as per the anesthesia record and topical anesthesia to the upper airway and tracheobronchial tree the Q190 video bronchoscope was introduced through the mouth.
The vocal cords appeared normal. The subglottic space was normal. The trachea is of normal caliber. The carina was sharp.
All left and right sided airways were normal without endobronchial disease.
The video bronchoscope was then removed and the UC180F convex probe EBUS bronchoscope was introduced through the mouth, and advanced to the tracheobronchial tree.
A systematic hilar and mediastinal lymph node survey was carried out.
Sampling criteria (5mm short axis diameter) were met in station 11Ri (5.0 mm), and 11Rs(6.6 mm).
Additionally the 4L was 4.9 mm but slightly hypoechoic and decision was made to sample this node as well.
Sampling by transbronchial needle aspiration was performed in these lymph nodes using an Olympus Visioshot 2 EBUSTBNA 21 gauge needle.
All samples were sent for routine cytology. We then removed the EBUS bronchoscopy and the super-dimension navigational catheter was inserted through the therapeutic bronchoscope and advanced into the airway.
Using navigational map we advanced the 180  degree edge catheter into the proximity of the lesion within the left upper lobe.
Confirmation of placement once at the point of interest with radial ultrasound showed a concentric view.
Biopsies were then performed with a variety of instruments to include peripheral needle, triple needle brush and forceps, under fluoroscopic visualization.
After adequate samples were obtained the bronchoscope was removed. Fluoroscopic inspection was performed and no pneumothorax was visualized at the conclusion of the procedure.
Complications: No immediate complications
Estimated Blood Loss: Less than 5 cc.
Post Procedure Diagnosis:
- Flexible bronchoscopy with successful and biopsy of left upper nodule
- Negative Staging EBUS on preliminary evaluation 
- Await final pathology  
- Post-procedure xray did not show pneumothorax.
- Patient was noted to be slightly hypoxic post-procedure likely as a result of sedation and decision was made to admit patient for overnight observation."""

# 2. Configuration (Procedure Flags)
PROCEDURE_FLAGS = {
    "diagnostic_bronchoscopy": 1,
    "bal": 0,
    "bronchial_wash": 0,
    "brushings": 1,
    "endobronchial_biopsy": 0,
    "tbna_conventional": 0,
    "linear_ebus": 1,
    "radial_ebus": 1,
    "navigational_bronchoscopy": 1,
    "transbronchial_biopsy": 1,
    "transbronchial_cryobiopsy": 0,
    "therapeutic_aspiration": 0,
    "foreign_body_removal": 0,
    "airway_dilation": 0,
    "airway_stent": 0,
    "thermal_ablation": 0,
    "tumor_debulking_non_thermal": 0,
    "cryotherapy": 0,
    "blvr": 0,
    "peripheral_ablation": 0,
    "bronchial_thermoplasty": 0,
    "whole_lung_lavage": 0,
    "rigid_bronchoscopy": 0,
    "thoracentesis": 0,
    "chest_tube": 0,
    "ipc": 0,
    "medical_thoracoscopy": 0,
    "pleurodesis": 0,
    "pleural_biopsy": 0,
    "fibrinolytic_therapy": 0
}

# 3. Data Definition (Spans)
SPANS = [
    ("Q190 video bronchoscope", "DEV_INSTRUMENT", "Q190 video bronchoscope", "tracheobronchial tree the ", "evt_01"),
    ("UC180F convex probe EBUS bronchoscope", "DEV_INSTRUMENT", "UC180F convex probe EBUS", "removed and the ", "evt_02"),
    ("11Ri", "ANAT_LN_STATION", "11Ri", "met in station ", "evt_02"),
    ("11Rs", "ANAT_LN_STATION", "11Rs", "11Ri (5.0 mm), and ", "evt_02"),
    ("4L", "ANAT_LN_STATION", "4L", "Additionally the ", "evt_02"),
    ("Olympus Visioshot 2 EBUSTBNA 21 gauge needle", "DEV_NEEDLE", "Olympus Visioshot 2 EBUSTBNA 21G", "nodes using an ", "evt_02"),
    ("super-dimension navigational catheter", "DEV_CATHETER", "super-dimension navigational catheter", "bronchoscopy and the ", "evt_03"),
    ("180  degree edge catheter", "DEV_CATHETER", "180 degree edge catheter", "advanced the ", "evt_03"),
    ("left upper lobe", "ANAT_LUNG_LOC", "LUL", "within the ", "evt_03"),
    ("radial ultrasound", "PROC_METHOD", "radial ultrasound", "interest with ", "evt_03"),
    ("peripheral needle", "DEV_INSTRUMENT", "peripheral needle", "include ", "evt_04"),
    ("triple needle brush", "DEV_INSTRUMENT", "triple needle brush", "peripheral needle, ", "evt_04"),
    ("forceps", "DEV_INSTRUMENT", "forceps", "brush and ", "evt_04"),
    ("No immediate complications", "OUTCOME_COMPLICATION", "None", "Complications: ", "evt_global"),
    ("slightly hypoxic", "OUTCOME_SYMPTOMS", "hypoxia", "noted to be ", "evt_global")
]

# 4. Event Definitions
EVENTS = [
    {
        "event_id": "evt_01",
        "method": "Diagnostic Bronchoscopy",
        "anatomy": "Airways",
        "devices": ["Q190 video bronchoscope"],
        "outcomes": []
    },
    {
        "event_id": "evt_02",
        "method": "Linear EBUS",
        "anatomy": "11Ri, 11Rs, 4L",
        "devices": ["UC180F convex probe EBUS bronchoscope", "Olympus Visioshot 2 EBUSTBNA 21 gauge needle"],
        "outcomes": ["Samples sent for cytology"]
    },
    {
        "event_id": "evt_03",
        "method": "Navigational Bronchoscopy",
        "anatomy": "Left upper lobe",
        "devices": ["super-dimension navigational catheter", "180 degree edge catheter", "radial ultrasound"],
        "outcomes": ["Concentric view confirmed"]
    },
    {
        "event_id": "evt_04",
        "method": "Transbronchial Biopsy",
        "anatomy": "Left upper lobe",
        "devices": ["peripheral needle", "triple needle brush", "forceps"],
        "outcomes": ["Adequate samples obtained"]
    },
    {
        "event_id": "evt_global",
        "method": "Outcome Assessment",
        "anatomy": "",
        "devices": [],
        "outcomes": ["No immediate complications", "slightly hypoxic post-procedure"]
    }
]

# 5. Helper Functions
def clean_text(text):
    return text.replace('\r', '').strip()

def hydrate_span(full_text, span_text, context_prefix):
    """
    Finds the start/end offsets of a span in the full text, using a context prefix 
    to disambiguate multiple occurrences.
    """
    cleaned_full = clean_text(full_text)
    cleaned_span = clean_text(span_text)
    cleaned_prefix = clean_text(context_prefix)
    
    # Attempt to find the prefix + span
    search_str = cleaned_prefix + cleaned_span
    start_index = cleaned_full.find(search_str)
    
    if start_index == -1:
        # Fallback: try finding just the span (risk of ambiguity)
        start_index = cleaned_full.find(cleaned_span)
        if start_index == -1:
            return 0, 0, "NOT_FOUND"
        else:
            return start_index, start_index + len(cleaned_span), "FOUND_NO_PREFIX"
    
    # Adjust start index to point to the start of the span, excluding prefix
    actual_start = start_index + len(cleaned_prefix)
    actual_end = actual_start + len(cleaned_span)
    
    return actual_start, actual_end, "FOUND_WITH_PREFIX"

# 6. Workbook Generation Function
def generate_workbook():
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
        
    # --- Sheet 1: Note_Text ---
    ws_text = wb.create_sheet("Note_Text")
    ws_text.append(["NOTE_ID", "SOURCE_FILE", "NOTE_TEXT"])
    ws_text.append([NOTE_ID, SOURCE_FILE, NOTE_TEXT])
    
    # --- Sheet 2: Note_Index ---
    ws_index = wb.create_sheet("Note_Index")
    headers = ["NOTE_ID", "PROCEDURE_DATE"] + list(PROCEDURE_FLAGS.keys())
    ws_index.append(headers)
    row_data = [NOTE_ID, PROCEDURE_DATE] + [PROCEDURE_FLAGS[k] for k in PROCEDURE_FLAGS]
    ws_index.append(row_data)
    
    # --- Sheet 3: Span_Annotations ---
    ws_ann = wb.create_sheet("Span_Annotations")
    ws_ann.append(["NOTE_ID", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "START_CHAR", "END_CHAR", "EVENT_ID"])
    for span in SPANS:
        # (span_text, label, normalized_value, context_prefix, event_id)
        ws_ann.append([NOTE_ID, span[0], span[1], span[2], span[3], "", "", span[4]])
        
    # --- Sheet 4: Span_Hydrated ---
    ws_hyd = wb.create_sheet("Span_Hydrated")
    ws_hyd.append(["NOTE_ID", "SPAN_TEXT", "LABEL", "NORMALIZED_VALUE", "CONTEXT_PREFIX", "START_CHAR", "END_CHAR", "EVENT_ID", "STATUS"])
    for span in SPANS:
        s_text, label, norm, prefix, evt = span
        start, end, status = hydrate_span(NOTE_TEXT, s_text, prefix)
        ws_hyd.append([NOTE_ID, s_text, label, norm, prefix, start, end, evt, status])
        
    # --- Sheet 5: Event_Log ---
    ws_log = wb.create_sheet("Event_Log")
    ws_log.append(["NOTE_ID", "EVENT_ID", "METHOD", "ANATOMY", "DEVICES", "OUTCOMES"])
    for evt in EVENTS:
        ws_log.append([
            NOTE_ID,
            evt["event_id"],
            evt["method"],
            evt["anatomy"],
            ", ".join(evt["devices"]),
            ", ".join(evt["outcomes"])
        ])
        
    # --- Sheet 6: V3_Procedure_Events ---
    ws_v3 = wb.create_sheet("V3_Procedure_Events")
    ws_v3.append(["NOTE_ID", "EVENT_JSON"])
    for evt in EVENTS:
        ws_v3.append([NOTE_ID, json.dumps(evt)])
        
    # --- Sheet 7: V3_Registry_JSON ---
    ws_json = wb.create_sheet("V3_Registry_JSON")
    ws_json.append(["NOTE_ID", "FULL_JSON"])
    registry_data = {
        "metadata": {
            "note_id": NOTE_ID,
            "source": SOURCE_FILE,
            "flags": PROCEDURE_FLAGS
        },
        "events": EVENTS,
        "spans": [
            {
                "text": s[0],
                "label": s[1],
                "normalized": s[2],
                "event_id": s[4]
            } for s in SPANS
        ]
    }
    ws_json.append([NOTE_ID, json.dumps(registry_data, indent=2)])
    
    # Save
    wb.save(OUTPUT_PATH)
    print(f"Workbook generated: {OUTPUT_PATH}")

if __name__ == "__main__":
    generate_workbook()